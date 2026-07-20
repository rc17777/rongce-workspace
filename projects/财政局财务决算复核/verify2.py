import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
wb = load_workbook(r'C:\Users\scrccpa\Desktop\马尔康项目决算审核报告-三级复核结果-20260720-v2.xlsx')
for name in wb.sheetnames:
    print(f'{name}: {wb[name].max_row} rows')
ws = wb['6-待摊费用合规测算']
# 打印测算表关键列核对
for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
    if row[1]:
        print(f'{row[0]} | {row[1]} | 实际={row[2]} | 标准={row[6]} | 比={row[7]} | 差异={row[8]} | {row[10]}')
