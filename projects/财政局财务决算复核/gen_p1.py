# -*- coding: utf-8 -*-
"""马尔康项目竣工财务决算审核报告 - 三级复核结果Excel生成"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

DARK = '0A1F3F'; TEAL='1A5C6E'
P0_FILL = PatternFill('solid', fgColor='F4CCCC')
P1_FILL = PatternFill('solid', fgColor='FCE5CD')
P2_FILL = PatternFill('solid', fgColor='EFEFEF')
OK_FILL = PatternFill('solid', fgColor='D9EAD3')
HDR_FILL = PatternFill('solid', fgColor=DARK)
thin = Side(style='thin', color='BFBFBF')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
H_FONT = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
T_FONT = Font(name='微软雅黑', bold=True, size=14, color=DARK)
B_FONT = Font(name='微软雅黑', size=10)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_sheet(ws, headers, widths, title, subtitle=None):
    ncol = len(headers)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    c = ws.cell(1,1,title); c.font = T_FONT; c.alignment = Alignment(vertical='center')
    ws.row_dimensions[1].height = 26
    r = 2
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
        c = ws.cell(2,1,subtitle); c.font = Font(name='微软雅黑', size=9, color='666666'); c.alignment = Alignment(vertical='center', wrap_text=True)
        ws.row_dimensions[2].height = 30
        r = 3
    for i, h in enumerate(headers, 1):
        cell = ws.cell(r, i, h); cell.font = H_FONT; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BORDER
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(r+1, 1)
    return r+1

def put_rows(ws, start, rows, lv_col=None):
    r = start
    for row in rows:
        for i, v in enumerate(row, 1):
            cell = ws.cell(r, i, v); cell.font = B_FONT; cell.alignment = WRAP; cell.border = BORDER
        if lv_col:
            lv = str(row[lv_col-1])
            fill = None
            if lv.startswith('P0'): fill = P0_FILL
            elif lv.startswith('P1'): fill = P1_FILL
            elif lv.startswith('P2'): fill = P2_FILL
            elif lv.startswith('√') or lv.startswith('通过'): fill = OK_FILL
            if fill:
                for i in range(1, len(row)+1):
                    ws.cell(r, i).fill = fill
        r += 1
    return r

# ================= Sheet 0 总览 =================
ws = wb.active; ws.title = '0-复核总览'
r = style_sheet(ws, ['项目','内容'], [22, 110], '马尔康城市环境质量提升（房屋建筑）工程（州级）项目竣工财务决算审核报告——三级复核结果总览')
rows0 = [
 ['复核对象','《马尔康城市环境质量提升（房屋建筑）工程（州级）项目-竣工财务决算审核报告》（征求意见稿20260720）+ 附件2《基本建设项目竣工财务决算审核表》'],
 ['报告编制单位','四川融策会计师事务所有限公司'],
 ['委托方','阿坝藏族羌族自治州财政局'],
 ['项目业主','阿坝藏族羌族自治州住房和城乡建设局'],
 ['复核日期','2026年7月20日'],
 ['复核方式','AI三级复核（一级：数据勾稽机械核对；二级：内容逻辑与合规；三级：终审与出具判断），全部汇总数字已逐项重新核算，计算依据列于各表'],
 ['复核依据','《基本建设财务规则》（财政部令81号）、《基本建设项目竣工财务决算管理暂行办法》（财建〔2016〕503号）、财建〔2016〕504号、《招标投标法》第24条、报告及附件内部勾稽关系'],
 ['',''],
 ['■ 复核结论','经三级复核，共发现问题 22 项（P0级 8 项 / P1级 8 项 / P2级 6 项），另列待核实事项 11 项、验证通过项 15 项。'],
 ['终审意见','★ 不同意以现稿对外出具（含征求意见稿发送）。8项P0须全部整改、11项待核实事项须项目组逐条书面回复后，重新提交复核。'],
 ['P0核心问题','①结算审减金额笔误（560,064.34→应为5,607,920.34，差504.78万元）；②"已支付84,707,685.26元"错误（应为77,288,874.98元，与未付7,418,810.28元自相矛盾）；③结算审核单位"泰宇"与"兴凯宏"前后矛盾；④附件2评审时间与报告审核时间矛盾；⑤资产交付口径四方矛盾（报告84,707,685.26元 vs 2-4表0 vs 2-5表转出 vs 2-6表待核销）；⑥公开招标时间线不足法定20日等标期；⑦规划许可证晚于开工令（无证开工约2个月）；⑧勘察合同晚于开工3个月（未勘先建）。另：附件2残留"九寨沟县双河镇罗依片区产业发展项目"全套台账（保密红线）。'],
 ['统计','P0=8 ｜ P1=8 ｜ P2=6 ｜ 验证通过=15 ｜ 待核实=11'],
 ['免责说明','本复核为AI初审结果，供项目负责人及质量复核人参考，不能替代人工复核与签批责任。所有P0/P1项须由项目组人工逐条确认。'],
]
r = put_rows(ws, r, rows0)

# ================= Sheet 1 一级复核 =================
ws = wb.create_sheet('1-一级复核(数据勾稽)')
r = style_sheet(ws, ['序号','核对项目','报告位置','报告值','核对值 / 计算过程（逐项重新核算）','差异','结论','风险等级','处理建议'],
    [6,22,16,24,42,16,10,10,32],
    '一级复核：数据勾稽机械核对表（expected/actual/diff 可复现）',
    '核对范围：报告正文全部金额引用 + 报告内2张表格 + 附件2审核表9张报表；每一处均独立重新计算，不采用报告原有结论。')
rows1 = [
 [1,'结算审核审减金额','正文"五（一）1"','审减金额560,064.34元','送审85,148,079.38 − 审定79,540,159.04 = 5,607,920.34','5,047,856.00','错误','P0','审减金额应为5,607,920.34元；审减率6.59%与正确金额匹配（5,607,920.34/85,148,079.38=6.586%≈6.59%），证实系金额笔误，非口径问题'],
 [2,'结余资金勾稽（已支付金额）','正文"六（二）"','已支付84,707,685.26元；应付未付7,418,810.28元','应付款明细表（11家）已付合计77,288,874.98 + 未付7,418,810.28 = 84,707,685.26；正文"七3"亦载明缺口7,418,810.28元','84,707,685.26−77,288,874.98=7,418,810.28','错误','P0','"已支付"应改为77,288,874.98元（=实际到位资金）。现表述"已支付=总投资"与"应付未付7,418,810.28元"直接矛盾'],
 [3,'评审/审核时间一致性','附件2封面 vs 正文首段','封面：实际评审2026年6月24日-7月8日；报告：2026年7月14日至7月20日','封面"委托评审时间2026年7月14日"，实际评审开始（6月24日）早于委托日20天；报告审核期间与封面完全不同','前后矛盾','错误','P0','核实真实工作日期后统一口径；评审开始早于委托日20天的逻辑必须解释'],
 [4,'资产交付使用口径','正文"六（一）" vs 附件2表2-4/2-5/2-6','报告：形成交付使用资产84,707,685.26元','2-4交付使用资产表合计=0（全空）；2-5转出投资表填84,707,685.26；2-6待核销基建支出表"防洪治理"挂84,707,685.26。四方口径互斥','84,707,685.26 vs 0','矛盾','P0','明确资产属性：交付使用/转出/待核销只能居其一；重编2-4/2-5/2-6表并与正文统一；说明资产接收单位'],
 [5,'印花税及滞纳金计入建设成本','正文表格R11、附件2-3表','14,737.52元计入待摊投资','《基本建设财务规则》（财政部令81号）精神：税收滞纳金属营业外支出性质，不得计入基本建设项目建设成本','待拆分','存疑','P1','拆分印花税与滞纳金；滞纳金部分调出待摊投资，核实是否应由责任单位自行承担'],
 [6,'2-2资金表资金类别归类','附件2表2-2 R5-R16','预算下达120,000,000挂"中央财政资金-财政专项资金"；到位77,288,874.98挂"地方财政资金-财政专项资金"','本项目资金为州本级财政资金（正文P040/P096），同一笔资金下达数挂中央、到位数挂地方，口径互斥；备注栏未填预算下达文号','归类矛盾','错误','P1','统一归入地方财政资金（州本级）；补填预算下达文号；120,000,000系批复估算而非"预算下达数"，建议按实际下达数填列'],
 [7,'2-6表科目属性','附件2表2-6 R16','全部投资84,707,685.26元挂"13.防洪治理 处/1"','本项目为城市立面整治及景观提升，与防洪治理无关；若全额列待核销则交付资产为0，与正文矛盾','科目误挂','错误','P1','按项目实质重选科目或改填2-4交付使用资产表'],
 [8,'2-5转出投资表完整性','附件2表2-5 R15','合计行空白','R5已填84,707,685.26，合计行未汇总','缺合计','不完整','P2','补填合计行；如确认非转出投资则清空并改填2-4表'],
 [9,'附件2残留演算数据','附件2 Sheet2下方R50-59、Sheet8','2-2表签字区下方残留10行演算数（合计1,749,395.99）；另有名为"Sheet1"的演算表','正式交付附件中混入工作演算底稿','残留','不规范','P2','清除全部演算残留，规范工作表命名'],
 [10,'附件编号一致性','正文附件目录 vs 附件2文件名','报告目录：附件3=审核表；文件名自标"附件2：…审核表"','附件2文件名编号与报告附件目录错位','错位','不规范','P2','统一附件编号'],
]
r = put_rows(ws, r, rows1, lv_col=8)
