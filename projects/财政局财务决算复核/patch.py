import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
p = r'C:\Users\scrccpa\Desktop\马尔康项目决算审核报告-三级复核结果-20260720.xlsx'
wb = load_workbook(p)
ws = wb['0-复核总览']
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    for cell in row:
        if cell.value == '■ 复核结论':
            ws.cell(cell.row, 2).value = '经三级复核，共发现具体问题 28 项（P0级 8 项 / P1级 10 项 / P2级 10 项），另形成终审判断 8 项、待核实事项 11 项、验证通过项 15 项。'
        if cell.value == '统计':
            ws.cell(cell.row, 2).value = '一级复核10项（P0×4/P1×3/P2×3）｜二级复核18项（P0×4/P1×7/P2×7）｜三级终审判断8项｜验证通过15项｜待核实11项'
wb.save(p)
print('patched')
