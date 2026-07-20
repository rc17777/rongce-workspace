import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
wb = load_workbook(r'C:\Users\scrccpa\Desktop\马尔康项目决算审核报告-三级复核结果-20260720.xlsx')
total_p0 = total_p1 = total_p2 = 0
for name in wb.sheetnames:
    ws = wb[name]
    rows = ws.max_row
    print(f'{name}: {rows} rows')
    if name.startswith(('1-','2-','3-')):
        for row in ws.iter_rows(min_row=3, values_only=True):
            for v in row:
                if isinstance(v,str):
                    if v.startswith('P0'): total_p0 += 1; break
                    if v.startswith('P1'): total_p1 += 1; break
                    if v.startswith('P2'): total_p2 += 1; break
print(f'P0={total_p0} P1={total_p1} P2={total_p2}')
