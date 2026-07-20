# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

P = r'C:\Users\scrccpa\Desktop\马尔康项目决算审核报告-三级复核结果-20260720.xlsx'
wb = load_workbook(P)
if '6-待摊费用合规测算' in wb.sheetnames:
    del wb['6-待摊费用合规测算']

DARK='0A1F3F'
P0_FILL=PatternFill('solid',fgColor='F4CCCC'); P1_FILL=PatternFill('solid',fgColor='FCE5CD')
P2_FILL=PatternFill('solid',fgColor='EFEFEF'); OK_FILL=PatternFill('solid',fgColor='D9EAD3')
HDR_FILL=PatternFill('solid',fgColor=DARK)
thin=Side(style='thin',color='BFBFBF'); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
H_FONT=Font(name='微软雅黑',bold=True,color='FFFFFF',size=10)
T_FONT=Font(name='微软雅黑',bold=True,size=13,color=DARK)
B_FONT=Font(name='微软雅黑',size=10)
WRAP=Alignment(wrap_text=True,vertical='top')
CENTER=Alignment(horizontal='center',vertical='center',wrap_text=True)

ws = wb.create_sheet('6-待摊费用合规测算')
headers = ['序号','费用项目','实际列支(元)','收费标准/依据','测算基数(元)','测算过程(逐项重算)','标准匡算值(元)','实际/标准','差异(元)','合规性·准确性·完整性结论','风险等级']
widths = [5,15,14,22,16,44,14,10,13,40,9]
ncol=len(headers)
ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncol)
c=ws.cell(1,1,'待摊投资各项费用合规性测算表（国家/省级收费标准逐项测算：合规性·准确性·完整性）'); c.font=T_FONT; c.alignment=Alignment(vertical='center')
ws.row_dimensions[1].height=24
ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=ncol)
note=('口径说明：①发改价格〔2015〕299号已放开勘察设计/招标代理/监理等政府指导价，原标准作为匡算参考上限——低于标准=价格形成合理（节约），高于标准须说明价格形成依据；'
      '②测算基数取签约时点可得的最近口径（设计/图审签约在控制价批复前用批复估算1.2亿；监理/代理签约在后用招标控制价9,734.80万）；'
      '③川价发〔2008〕141号原文费率表本机无电子版、外网被墙，该行按通行费率区间匡算，需调取原文复核（列入待核实#12）。')
c=ws.cell(2,1,note); c.font=Font(name='微软雅黑',size=9,color='666666'); c.alignment=Alignment(wrap_text=True,vertical='center')
ws.row_dimensions[2].height=52
for i,h in enumerate(headers,1):
    cell=ws.cell(3,i,h); cell.font=H_FONT; cell.fill=HDR_FILL; cell.alignment=CENTER; cell.border=BORDER
for i,w in enumerate(widths,1):
    ws.column_dimensions[get_column_letter(i)].width=w
ws.freeze_panes='A4'

rows = [
 [1,'设计费',1950000.00,'计价格〔2002〕10号《工程勘察设计收费管理规定》(2015年放开后作参考上限)',120000000.00,
  '基价内插：304.8万+(566.8−304.8)万×(12,000−10,000)/(20,000−10,000)=357.20万；专业调整(建筑市政)1.0×复杂II级1.0×改扩建附加1.1=392.92万',
  3929200.00,'49.6%','−1,979,200.00',
  '合规性√：实际为标准价49.6%，下浮50.4%，价格形成节约合理。准确性√：合同173万+补充22万=195万与列支一致。完整性✗：补充合同+22万(+12.7%)计价依据未披露，调整批复后栋数107→1133(面积反降)的工作量测算未附。','P2'],
 [2,'监理费',1523800.00,'发改价格〔2007〕670号《建设工程监理与相关服务收费管理规定》',97347978.38,
  '基价内插：181.0万+(218.6−181.0)万×(9,734.8−8,000)/2,000=213.62万；专业1.0×复杂II级1.0×高程1.0=213.62万',
  2136200.00,'71.3%','−613,400.00',
  '合规性√：下浮28.7%，在市场惯例区间(20%-40%)。准确性√：合同价与列支一致；已付121.90万+未付30.48万=合同价。','通过'],
 [3,'招标代理费',248928.00,'计价格〔2002〕1980号《招标代理服务收费管理暂行办法》差额定率累进(工程类)',93235181.14,
  '施工：100万×1.0%+400万×0.7%+500万×0.55%+4,000万×0.35%+4,323.52万×0.2%=291,970.36；监理(152.38万)：100万×1.0%+52.38万×0.7%=13,666.60；标准合计305,636.96；合同载"按标准下浮20%"→244,509.57',
  244509.57,'101.8%','+4,418.43',
  '准确性✗：实际下浮率18.55%≠合同载明的20%，多计4,418.43元。合规性：低于标准上限但计费过程与合同条款不符，须调取代理费计算底稿复核。完整性✗：设计招标代理(博晨冠宏)费用全体系缺失(见二级复核#6)。','P1'],
 [4,'清单控制价编制费',180000.00,'川价发〔2008〕141号《工程造价咨询收费标准》(报告审核依据已引用)',97347978.38,
  '反算费率：180,000/97,347,978.38=1.85‰；该规模清单编制+控制价编制通行区间1.5‰-2.5‰→146,022~243,370',
  194696.00,'92.5%','−14,696.00',
  '合规性√(匡算区间内)。准确性√：合同200,000最终核定180,000，核减20,000有依据。完整性✗：141号文原文费率未附卷，匡算值(取中值2‰)待原文复核。','P2'],
 [5,'全过程造价咨询费',566000.00,'川价发〔2008〕141号(施工阶段全过程造价控制)',93235181.14,
  '反算费率：566,000/93,235,181.14=6.07‰；该规模全过程控制通行区间5‰-8‰→466,176~745,881',
  606029.00,'93.4%','−40,029.00',
  '合规性√(匡算区间内，竞争性磋商形成)。程序问题已列二级复核#4：合同2023-2-28签订晚于开工5个月，"全过程"名不副实。','P2'],
 [6,'可行性研究编制费',283800.00,'计价格〔1999〕1283号《建设项目前期工作咨询收费暂行规定》',120000000.00,
  '1亿-5亿档28-75万内插：28+(75−28)×(12,000−10,000)/(50,000−10,000)=30.35万；行业系数(市政0.7/建筑0.8)→21.25万~24.28万',
  227650.00,'124.7%','+56,150.00',
  '合规性：高于匡算值17%~34%——2015年后市场调节价不违规，但须说明价格形成依据(竞价过程)。准确性✗：补充协议+25,800元(+10%)计价依据未披露。','P2'],
]
