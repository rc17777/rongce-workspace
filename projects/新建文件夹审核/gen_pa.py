# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

P = r'C:\Users\scrccpa\Desktop\马尔康项目决算审核报告-三级复核结果-20260720-v2.xlsx'
wb = load_workbook(P)
if '7-过控采购审核' in wb.sheetnames:
    del wb['7-过控采购审核']
ws = wb.create_sheet('7-过控采购审核')

hdr_fill = PatternFill('solid', fgColor='0A1F3F')
hdr_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
title_font = Font(name='微软雅黑', size=13, bold=True, color='0A1F3F')
body_font = Font(name='微软雅黑', size=9)
sec_fill = PatternFill('solid', fgColor='D9E2F3')
p1_fill = PatternFill('solid', fgColor='F4CCCC')
p2_fill = PatternFill('solid', fgColor='FFF2CC')
ok_fill = PatternFill('solid', fgColor='D9EAD3')
wait_fill = PatternFill('solid', fgColor='EAD1DC')
thin = Side(style='thin', color='999999')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical='top')

ws['A1'] = '全过程造价控制采购专项审核（招标文件/评标报告/合同）——N5132112023000009'
ws['A1'].font = title_font
ws['A2'] = '依据：竞争性磋商文件、评审工作报告、合同、党组会议纪要、代理比价表（OCR自83页扫描件） | 2026-07-21'
ws['A2'].font = Font(name='微软雅黑', size=8, color='666666')

headers = ['#', '审核对象', '检查点', '事实描述', '审核发现', '依据/法规', '级别', '核实方式']
for j, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=j, value=h)
    c.fill = hdr_fill; c.font = hdr_font; c.border = border; c.alignment = wrap

rows = [
 ('S1','招标文件','采购意向公开时限','意向公开2023-01-18 → 采购公告2023-02-09，仅22天','不足30日，程序瑕疵','财库〔2020〕10号','P2','四川政府采购网公告日期核对'),
 ('S2','招标文件','代理机构选择程序','管理科2023-01-07请示写明"拟指定博晨冠宏"；三家比价报价表日期为01-10','先内定后比价，程序倒置痕迹；"随机抽取候选代理机构三个"的抽取记录未附档','采购档案P4-P7','P2','调取代理机构抽取记录'),
 ('S3','招标文件','公告期','磋商文件发出(02-10)至首次响应截止(02-20)=10日','压线满足≥10日，合规','磋商办法第十条','通过',''),
 ('S4','招标文件','资格条件与评分标准','不接受联合体；报价20+方案44+履约9+人员27=100分','无排他性，分值量化合理','磋商文件附件2','通过',''),
 ('S5','评标报告','低价澄清程序','维信成最后报价44.5万，低于其余三家(51.28/56.6/61.6万)13%~28%；评审报告"供应商澄清情况：无"','触发磋商文件低于成本价预防条款却未启动书面说明程序，程序瑕疵','磋商文件须知表第5条','P1','调评审现场记录核实'),
 ('S6','评标报告','低价未中标(L10)','维信成价格分满分20，总分92.67仅差1.05分落败；胜负由技术分决定，评委方案打分离散度大','倾向性压分疑点，需原件验证','评审计分表/详细评审表','P1','调三位评委分项打分原件逐维度核对'),
 ('S7','评标报告','专家费用','费用表仅见申道忠900元，罗玲章费用未体现','可能OCR漏行或漏发','档案P67','P2','核对专家费发放凭证'),
 ('S8','评标报告','评审组织','小组3人=专家2+采购人代表1(潘强)；专家库抽取','专家占2/3，组成合规','磋商办法第十四条','通过',''),
 ('S9','评标报告','价格分验算','44.5÷56.6×20=15.72，与评审计分表一致','公式应用正确','附件2评分标准','通过',''),
]
r = 4
for row in rows:
    for j, v in enumerate(row, 1):
        c = ws.cell(row=r, column=j, value=v)
        c.font = body_font; c.border = border; c.alignment = wrap
    lv = row[6]
    fill = {'P1': p1_fill, 'P2': p2_fill, '通过': ok_fill, '待核实': wait_fill}.get(lv)
    if fill: ws.cell(row=r, column=7).fill = fill
    r += 1
print('part A rows:', r-4)
wb.save(P)
print('saved A')
