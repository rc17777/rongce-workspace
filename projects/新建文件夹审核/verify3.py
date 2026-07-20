import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
wb = load_workbook(r'C:\Users\scrccpa\Desktop\马尔康项目决算审核报告-三级复核结果-20260720-v2.xlsx')
for name in wb.sheetnames:
    print(f'{name}: {wb[name].max_row} rows')
ws = wb['7-过控采购审核']
for row in ws.iter_rows(min_row=4, max_row=22, values_only=True):
    if row[0]:
        print(f'{row[0]} | {row[1]} | {row[2]} | {row[6]}')
ov = wb['0-复核总览']
for row in ov.iter_rows(values_only=True):
    for v in row:
        if isinstance(v, str) and ('44项' in v or '34项' in v):
            print('总览:', v)
