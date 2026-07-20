# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

P = r'C:\Users\scrccpa\Desktop\马尔康项目决算审核报告-三级复核结果-20260720-v2.xlsx'
wb = load_workbook(P)
ws = wb['7-过控采购审核']

body_font = Font(name='微软雅黑', size=9)
p1_fill = PatternFill('solid', fgColor='F4CCCC')
p2_fill = PatternFill('solid', fgColor='FFF2CC')
ok_fill = PatternFill('solid', fgColor='D9EAD3')
wait_fill = PatternFill('solid', fgColor='EAD1DC')
link_fill = PatternFill('solid', fgColor='FCE5CD')
thin = Side(style='thin', color='999999')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical='top')

rows = [
 ('S10','合同','服务范围','磋商文件项目清单及四家最后报价均含"竣工结算审计服务"，合同第三条未写入','合同实质性内容与采购文件不一致；财政另行委托泰宇/兴凯宏做结算审核——重复购买服务，或供应商未完整履约仍按70%付款(已付396,200)','政府采购法第46条','P1','核定重复付费金额；核实中正天达实际履约内容'),
 ('S11','合同','供应商名称','全套档案为"中正天达建设项目管理有限公司"(统一代码91510000099571363E)，决算报告误为"中正天达建设集团有限公司"','主体矛盾责任在决算报告编制方，报告及附件2表2-2/2-4须更正','成交通知书/合同/签到表','P1','比对信用代码后更正'),
 ('S12','合同','法律依据','合同首条引用《合同法》(2021-01-01已废止)','法条引用错误','民法典','P2',''),
 ('S13','合同','付款条款表述','"支付至结算金额的100%"','服务合同无"结算金额"概念，应为合同金额，笔误','合同第四条','P2',''),
 ('S14','合同','术语与期限','附件清单"中标通知书"应为"成交通知书"(磋商采购)；合同期限"至工作完成"与结果公告"365日历天"口径不一','表述不规范','合同第十五条/第二条','P2',''),
 ('S15','合同','金额·签约·保证金','566,000元三方一致(成交/合同/公告)；02-20成交→02-28签约8天≤30日；履约保证金3%=16,980元','合规','政府采购法第46条','通过',''),
 ('S16','勾稽','过控进场滞后','开工令2022-09-22；过控2023-01-07才上党组会、02-28才签约','"施工阶段全过程"缺位4个月，印证决算复核P1(联动，不重复计数)','党组会议纪要/开工令','P1联动','已在决算复核P1列示'),
 ('S17','勾稽','代理费口径','本项目代理费10,100元由成交供应商支付(须知表第16条)，不进待摊投资','本项目口径清楚；设计招标代理费承担方式待核实后销号','磋商文件','待核实','调设计采购文件'),
 ('S18','勾稽','节资率','预算70万 → 成交56.6万','节资19.1%','','通过',''),
 ('S19','合同','履约验收','档案未见履约验收报告，保证金16,980元退还未知','服务完成情况及验收留痕缺失','财库〔2016〕205号','待核实','向州住建局调取验收资料'),
]
r = 13
for row in rows:
    for j, v in enumerate(row, 1):
        c = ws.cell(row=r, column=j, value=v)
        c.font = body_font; c.border = border; c.alignment = wrap
    lv = row[6]
    fill = {'P1': p1_fill, 'P2': p2_fill, '通过': ok_fill, '待核实': wait_fill, 'P1联动': link_fill}.get(lv)
    if fill: ws.cell(row=r, column=7).fill = fill
    r += 1

widths = [6, 10, 16, 46, 46, 22, 9, 24]
from openpyxl.utils import get_column_letter
for j, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(j)].width = w
ws.freeze_panes = 'A4'

# 统计块
sr = r + 1
ws.cell(row=sr, column=1, value='本表统计：P1×4(S5/S6/S10/S11) P2×6(S1/S2/S7/S12/S13/S14) 通过×6 待核实×2 联动×1(S16不重复计数)').font = Font(name='微软雅黑', size=9, bold=True)

# 总览 patch
ov = wb['0-复核总览']
patched = False
for row in ov.iter_rows():
    for c in row:
        if isinstance(c.value, str) and '34项问题' in c.value:
            c.value = '累计44项问题（P0×8 / P1×16 / P2×20）＋ 验证通过23项 ＋ 待核实15项（新增sheet7过控采购审核：P1×4/P2×6/通过6/待核实2/联动1）'
            patched = True
if not patched:
    ov.cell(row=ov.max_row+2, column=1, value='累计44项问题（P0×8 / P1×16 / P2×20）＋ 验证通过23项 ＋ 待核实15项（新增sheet7过控采购审核：P1×4/P2×6/通过6/待核实2/联动1）')

try:
    wb.save(P)
    print('saved to v2')
except PermissionError:
    P3 = P.replace('-v2.xlsx', '-v3.xlsx')
    import shutil; shutil.copyfile(P, P3)
    wb2 = load_workbook(P3)
    # 已在内存中的wb直接另存
    wb.save(P3)
    print('v2 locked, saved as v3:', P3)
print('rows total:', r-4)
