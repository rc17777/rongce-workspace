# -*- coding: utf-8 -*-
"""
若尔盖校园餐审计 — 三大核心指标计算与疑点筛查工具
================================================================
用途：进点拿到数据后，一键计算餐标、筛查虚报、比对价格。
作者：融策右护卫 | 2026-07-14

三大核心指标：
  1. 县级实际供餐标准 = 实际支出 ÷ 实际人数 ÷ 实际天数
  2. 虚报天数/人数    = 上报值 − 实际值
  3. 单校实际餐标      = 单校支出 ÷ 单校人数 ÷ 天数

用法：
  # 生成数据模板（首次使用）
  python 校园餐数据比对.py --init

  # 跑全部分析（数据填好后）
  python 校园餐数据比对.py --run

  # 单独跑某一项
  python 校园餐数据比对.py --meal-standard   # 餐标计算
  python 校园餐数据比对.py --false-report     # 虚报筛查
  python 校园餐数据比对.py --price-check      # 价格比对
"""
import sys
import os
import argparse
from datetime import datetime

# Windows GBK 编码坑，强制 UTF-8
if sys.stdout.encoding != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("缺少 openpyxl，请先安装：pip install openpyxl")
    sys.exit(1)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(BASE_DIR)
DATA_DIR = os.path.join(PROJECT_DIR, "raw_data")
OUTPUT_DIR = os.path.join(PROJECT_DIR, "analysis")

# 国家基础标准（元/生/天）
STANDARD_2021_AUTUMN = 5.0   # 2021秋季起
STANDARD_BEFORE_2021 = 4.0   # 2021秋季前

# 样式
HEADER_FILL = PatternFill("solid", fgColor="0A1F3F")
HEADER_FONT = Font(name="微软雅黑", color="C5955C", bold=True, size=11)
ALERT_FILL = PatternFill("solid", fgColor="FFC7CE")   # 红：疑点
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")    # 黄：关注
OK_FILL = PatternFill("solid", fgColor="C6EFCE")      # 绿：正常
THIN = Side(style="thin", color="C5955C")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def init_templates():
    """生成三张数据录入模板"""
    os.makedirs(DATA_DIR, exist_ok=True)

    # 模板1：营养改善计划申报数据（虚报筛查用）
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "申报数据"
    headers = ["学校名称", "年度", "上报受益人数", "上报在校天数",
               "实际受益人数", "实际在校天数", "补助标准(元)",
               "备注(数据来源/口径)"]
    ws.append(headers)
    # 示例行
    ws.append(["示例：XX小学", 2024, 500, 190, 480, 185, 5.0, "系统导出vs学籍核实"])
    _style_header(ws)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['H'].width = 30
    path1 = os.path.join(DATA_DIR, "01_申报数据.xlsx")
    wb.save(path1)

    # 模板2：资金支出数据（餐标计算用）
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "支出数据"
    headers = ["学校名称", "年度", "营养餐实际支出(元)", "受益人数",
               "供餐天数", "食堂人员工资(元)", "水电煤气(元)",
               "配送费(元)", "其他非营养餐支出(元)", "备注"]
    ws.append(headers)
    ws.append(["示例：XX小学", 2024, 450000, 480, 185, 0, 0, 0, 0, "含/不含说明"])
    _style_header(ws)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['J'].width = 25
    path2 = os.path.join(DATA_DIR, "02_支出数据.xlsx")
    wb.save(path2)

    # 模板3：食材采购价格（价格比对用）
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "采购价格"
    headers = ["学校名称", "食材名称", "规格单位", "采购单价(元)",
               "采购数量", "采购金额(元)", "供应商", "同期市场价(元)",
               "采购日期", "备注"]
    ws.append(headers)
    ws.append(["示例：XX小学", "大米", "斤", 3.5, 1000, 3500,
               "XX粮油公司", 2.8, "2024-09-01", ""])
    _style_header(ws)
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['G'].width = 18
    path3 = os.path.join(DATA_DIR, "03_采购价格.xlsx")
    wb.save(path3)

    print("✅ 数据模板已生成：")
    print(f"   {path1}")
    print(f"   {path2}")
    print(f"   {path3}")
    print("\n把实际数据填进去，再跑 --run")


def _load_sheet(filename):
    path = os.path.join(DATA_DIR, filename)
    if not os.path.exists(path):
        print(f"⚠️  找不到 {path}，请先 --init 生成模板并填数据")
        return None, None
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    # 过滤示例行和空行
    rows = [r for r in rows if r[0] and not str(r[0]).startswith("示例")]
    return wb, rows


def get_standard(year):
    """按年度返回国家基础标准"""
    try:
        y = int(year)
        return STANDARD_2021_AUTUMN if y >= 2022 else STANDARD_BEFORE_2021
    except Exception:
        return STANDARD_2021_AUTUMN


def analyze_meal_standard():
    """指标1&3：计算实际餐标并标注未达标"""
    _, rows = _load_sheet("02_支出数据.xlsx")
    if rows is None:
        return []
    results = []
    for r in rows:
        (school, year, expense, people, days,
         salary, utility, delivery, other, note) = r[:10]
        expense = expense or 0
        people = people or 0
        days = days or 0
        if people == 0 or days == 0:
            continue
        # 毛餐标（不扣除）
        gross = expense / people / days
        # 净餐标（扣除非营养餐支出）
        deduct = (salary or 0) + (utility or 0) + (delivery or 0) + (other or 0)
        net_expense = expense - deduct
        net = net_expense / people / days if people and days else 0
        std = get_standard(year)
        status = "达标" if net >= std else "未达标"
        results.append({
            "school": school, "year": year,
            "gross": round(gross, 2), "net": round(net, 2),
            "standard": std, "status": status,
            "gap": round(std - net, 2) if net < std else 0,
            "deduct": deduct,
        })
    return results


def analyze_false_report():
    """指标2：虚报人数/天数筛查"""
    _, rows = _load_sheet("01_申报数据.xlsx")
    if rows is None:
        return []
    results = []
    for r in rows:
        (school, year, rep_people, rep_days,
         act_people, act_days, standard, note) = r[:8]
        rep_people = rep_people or 0
        rep_days = rep_days or 0
        act_people = act_people or 0
        act_days = act_days or 0
        standard = standard or get_standard(year)
        false_people = rep_people - act_people
        false_days = rep_days - act_days
        # 套取资金 = 上报值乘积 - 实际值乘积，按补助标准
        rep_amount = rep_people * rep_days * standard
        act_amount = act_people * act_days * standard
        embezzled = rep_amount - act_amount
        has_issue = false_people > 0 or false_days > 0
        results.append({
            "school": school, "year": year,
            "false_people": false_people, "false_days": false_days,
            "embezzled": round(embezzled, 2),
            "has_issue": has_issue,
            # 190天固定天数预警
            "day190_flag": "⚠️沿用190天" if rep_days == 190 else "",
        })
    return results


def analyze_price():
    """价格比对：采购价 vs 市场价"""
    _, rows = _load_sheet("03_采购价格.xlsx")
    if rows is None:
        return []
    results = []
    for r in rows:
        (school, item, unit, price, qty, amount,
         supplier, market_price, date, note) = r[:10]
        price = price or 0
        market_price = market_price or 0
        if market_price == 0:
            ratio = None
            overpay = None
            status = "缺市场价"
        else:
            ratio = (price - market_price) / market_price * 100
            overpay = (price - market_price) * (qty or 0)
            if ratio > 30:
                status = "价格虚高(>30%)"
            elif ratio > 10:
                status = "偏高(>10%)"
            else:
                status = "正常"
        results.append({
            "school": school, "item": item, "supplier": supplier,
            "price": price, "market_price": market_price,
            "ratio": round(ratio, 1) if ratio is not None else "-",
            "overpay": round(overpay, 2) if overpay is not None else "-",
            "status": status,
        })
    return results


def write_report(meal, false_rep, price):
    """输出综合分析报告 Excel"""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb = openpyxl.Workbook()

    # Sheet1: 餐标分析
    ws = wb.active
    ws.title = "餐标分析"
    ws.append(["学校", "年度", "毛餐标(元)", "净餐标(元)", "国家标准",
               "达标情况", "缺口(元)", "扣除非营养餐支出(元)"])
    _style_header(ws)
    for m in meal:
        ws.append([m["school"], m["year"], m["gross"], m["net"],
                   m["standard"], m["status"], m["gap"], m["deduct"]])
        row = ws.max_row
        fill = ALERT_FILL if m["status"] == "未达标" else OK_FILL
        ws.cell(row=row, column=6).fill = fill
    for col in "ABCDEFGH":
        ws.column_dimensions[col].width = 14

    # Sheet2: 虚报筛查
    ws2 = wb.create_sheet("虚报筛查")
    ws2.append(["学校", "年度", "虚报人数", "虚报天数",
                "套取资金(元)", "疑点", "190天预警"])
    _style_header(ws2)
    for f in false_rep:
        ws2.append([f["school"], f["year"], f["false_people"],
                    f["false_days"], f["embezzled"],
                    "疑点" if f["has_issue"] else "-", f["day190_flag"]])
        row = ws2.max_row
        if f["has_issue"]:
            ws2.cell(row=row, column=6).fill = ALERT_FILL
        if f["day190_flag"]:
            ws2.cell(row=row, column=7).fill = WARN_FILL
    for col in "ABCDEFG":
        ws2.column_dimensions[col].width = 14

    # Sheet3: 价格比对
    ws3 = wb.create_sheet("价格比对")
    ws3.append(["学校", "食材", "供应商", "采购价", "市场价",
                "溢价率(%)", "多支出(元)", "判定"])
    _style_header(ws3)
    for p in price:
        ws3.append([p["school"], p["item"], p["supplier"], p["price"],
                    p["market_price"], p["ratio"], p["overpay"], p["status"]])
        row = ws3.max_row
        if "虚高" in str(p["status"]):
            ws3.cell(row=row, column=8).fill = ALERT_FILL
        elif "偏高" in str(p["status"]):
            ws3.cell(row=row, column=8).fill = WARN_FILL
    for col in "ABCDEFGH":
        ws3.column_dimensions[col].width = 14

    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    out = os.path.join(OUTPUT_DIR, f"校园餐审计分析_{stamp}.xlsx")
    wb.save(out)
    return out


def print_summary(meal, false_rep, price):
    print("\n" + "=" * 50)
    print("  若尔盖校园餐审计 — 疑点汇总")
    print("=" * 50)

    unmet = [m for m in meal if m["status"] == "未达标"]
    print(f"\n【餐标】共{len(meal)}校，未达标 {len(unmet)} 校")
    for m in unmet:
        print(f"   🔴 {m['school']} {m['year']}年：净餐标{m['net']}元 "
              f"(缺口{m['gap']}元/标准{m['standard']}元)")

    issues = [f for f in false_rep if f["has_issue"]]
    print(f"\n【虚报】共{len(false_rep)}条，疑点 {len(issues)} 条")
    total_emb = sum(f["embezzled"] for f in issues)
    for f in issues:
        print(f"   🔴 {f['school']} {f['year']}年：虚报{f['false_people']}人/"
              f"{f['false_days']}天，套取{f['embezzled']:.0f}元 {f['day190_flag']}")
    if issues:
        print(f"   合计疑似套取：{total_emb:.0f}元")

    high = [p for p in price if "虚高" in str(p["status"])]
    print(f"\n【价格】共{len(price)}条，虚高(>30%) {len(high)} 条")
    for p in high:
        print(f"   🔴 {p['school']} {p['item']}：采购{p['price']}元 vs "
              f"市场{p['market_price']}元 (溢价{p['ratio']}%)")

    print("\n" + "=" * 50)
    print("⚠️ 复核铁律：以上数字为脚本计算结果，出报告前须逐条")
    print("   核对原始凭证，并在底稿注明数据来源+计算方法。")
    print("=" * 50 + "\n")


def main():
    ap = argparse.ArgumentParser(description="若尔盖校园餐审计数据比对工具")
    ap.add_argument("--init", action="store_true", help="生成数据录入模板")
    ap.add_argument("--run", action="store_true", help="跑全部分析并出报告")
    ap.add_argument("--meal-standard", action="store_true", help="仅算餐标")
    ap.add_argument("--false-report", action="store_true", help="仅筛虚报")
    ap.add_argument("--price-check", action="store_true", help="仅比价格")
    args = ap.parse_args()

    if args.init:
        init_templates()
        return

    if not any([args.run, args.meal_standard, args.false_report, args.price_check]):
        ap.print_help()
        return

    meal = analyze_meal_standard() if (args.run or args.meal_standard) else []
    false_rep = analyze_false_report() if (args.run or args.false_report) else []
    price = analyze_price() if (args.run or args.price_check) else []

    print_summary(meal, false_rep, price)

    if args.run:
        out = write_report(meal, false_rep, price)
        print(f"📊 分析报告已输出：{out}")


if __name__ == "__main__":
    main()
