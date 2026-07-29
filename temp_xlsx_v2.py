# -*- coding: utf-8 -*-
"""测算数据基础表 - 含公式+数据出处"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()
HDR_FILL = PatternFill(start_color='0A1F3F', end_color='0A1F3F', fill_type='solid')
HDR_FONT = Font(name='微软雅黑', bold=True, color='FFFFFF', size=10)
PARAM_FILL = PatternFill(start_color='1A5C6E', end_color='1A5C6E', fill_type='solid')
PARAM_FONT = Font(name='宋体', bold=True, color='FFFFFF', size=10)
SRC_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
SRC_FONT = Font(name='宋体', size=9, italic=True)
DATA_FONT = Font(name='宋体', size=10)
BOLD_FONT = Font(name='宋体', bold=True, size=10)
ALT_FILL = PatternFill(start_color='F5F2EC', end_color='F5F2EC', fill_type='solid')
TITLE_FONT = Font(name='微软雅黑', bold=True, size=14, color='0A1F3F')
WARN_FONT = Font(name='宋体', bold=True, color='CC0000', size=9)
thin = Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))
NUM = '#,##0.00'; PCT = '0.00%'; INT = '#,##0'

def hdr(ws, row, cols, widths=None):
    for j,c in enumerate(cols,1):
        cl=ws.cell(row=row,column=j,value=c)
        cl.font=HDR_FONT; cl.fill=HDR_FILL; cl.alignment=Alignment(horizontal='center',wrap_text=True); cl.border=thin
    if widths:
        for j,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(j)].width=w

def param(ws, row, label, value, unit, source):
    """Write a parameter row with label, value, unit, source"""
    for j,(v,f) in enumerate([(label,DATA_FONT),(value,BOLD_FONT),(unit,DATA_FONT),(source,SRC_FONT)],1):
        cl=ws.cell(row=row,column=j,value=v); cl.font=f; cl.border=thin
        if j==2: cl.number_format=NUM; cl.alignment=Alignment(horizontal='center')
        if j==4: cl.fill=SRC_FILL

def data(ws, row, vals, fonts=None, alt=False, fmts=None):
    for j,v in enumerate(vals,1):
        cl=ws.cell(row=row,column=j,value=v); cl.border=thin; cl.alignment=Alignment(horizontal='center')
        fnt = (fonts[j-1] if fonts and j-1<len(fonts) else DATA_FONT)
        cl.font = fnt
        if fmts and j-1<len(fmts) and fmts[j-1]: cl.number_format = fmts[j-1]
        if alt: cl.fill = ALT_FILL

def title(ws, row, text):
    cl=ws.cell(row=row,column=1,value=text); cl.font=TITLE_FONT
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=10)

# ============================================================
# SHEET 1: 资本金回报 — with formulas
# ============================================================
ws1=wb.active; ws1.title='表1-资本金回报'
title(ws1,1,'表1：项目资本金回报测算表')
hdr(ws1,3,['参数','数值','单位','数据出处'],widths=[18,22,10,55])

params1=[
    ('P-资本金(等额本息部分)','80000000','元','平头哥指令：8000万按7.99%18年等额本息'),
    ('k-合理利润率','0.0799','-','合同：社会资本中标年回报率7.99%《恩阳医养园PPP项目合同之补充合同》(2017年7月)'),
    ('n-运营期','18','年','合同：2021年补充合同,运营期18年(2022.10.28-2040.12.21)'),
    ('A-PMT(年还款额)','=B4*B5*(1+B5)^B6/((1+B5)^B6-1)','元/年','公式：P*k*(1+k)^n/((1+k)^n-1)'),
    ('K-运维成本加成系数','0.0799','-','合同：K=社会资本中标年回报率(同k)'),
    ('剩余资本金(第18年一次性)','22298077.79','元','=实缴资本金102,298,077.79-8,000万(含恩阳医院2,000万,不计回报利息)'),
    ('实缴资本金总额','102298077.79','元','《竣工结算审核报告》(中宏审[2024]1213号)审定+项目公司财务资料'),
]
for i,(l,v,u,s) in enumerate(params1): param(ws1,4+i,l,v,u,s)

# Schedule
hdr(ws1,12,['年','期初本金\n(等额部分)','年利息\n=期初×k','年还本\n=PMT-利息','年还款合计\n(等额本息)','一次性支付\n(第18年)','资本金回报\n合计','期末本金\n=期初-还本','数据出处'],widths=[6,18,18,18,18,18,18,18,48])

# Year 1 formulas
r=13  # first data row
ws1.cell(row=r,column=1,value=2023); ws1.cell(row=r,column=1).font=DATA_FONT; ws1.cell(row=r,column=1).border=thin; ws1.cell(row=r,column=1).alignment=Alignment(horizontal='center')
# Col B: =B4 (opening = P)
ws1.cell(row=r,column=2,value='=B4'); ws1.cell(row=r,column=2).font=DATA_FONT; ws1.cell(row=r,column=2).border=thin; ws1.cell(row=r,column=2).number_format=NUM
# Col C: interest = B13*k = B13*B5
ws1.cell(row=r,column=3,value='=B13*B5'); ws1.cell(row=r,column=3).font=DATA_FONT; ws1.cell(row=r,column=3).border=thin; ws1.cell(row=r,column=3).number_format=NUM
# Col D: principal = PMT - interest = B7-C13
ws1.cell(row=r,column=4,value='=B7-C13'); ws1.cell(row=r,column=4).font=DATA_FONT; ws1.cell(row=r,column=4).border=thin; ws1.cell(row=r,column=4).number_format=NUM
# Col E: PMT = B7
ws1.cell(row=r,column=5,value='=B7'); ws1.cell(row=r,column=5).font=DATA_FONT; ws1.cell(row=r,column=5).border=thin; ws1.cell(row=r,column=5).number_format=NUM
# Col F: lump sum (year 18 only, handled later)
ws1.cell(row=r,column=6,value=0); ws1.cell(row=r,column=6).font=DATA_FONT; ws1.cell(row=r,column=6).border=thin; ws1.cell(row=r,column=6).number_format=NUM
# Col G: total = E13+F13
ws1.cell(row=r,column=7,value='=E13+F13'); ws1.cell(row=r,column=7).font=DATA_FONT; ws1.cell(row=r,column=7).border=thin; ws1.cell(row=r,column=7).number_format=NUM
# Col H: remaining = B13-D13
ws1.cell(row=r,column=8,value='=B13-D13'); ws1.cell(row=r,column=8).font=DATA_FONT; ws1.cell(row=r,column=8).border=thin; ws1.cell(row=r,column=8).number_format=NUM
# Col I: source
ws1.cell(row=r,column=9,value='资本金回报=合同公式P*k*(1+k)^n/((1+k)^n-1), P=80,000,000, k=7.99%, n=18')
ws1.cell(row=r,column=9).font=SRC_FONT; ws1.cell(row=r,column=9).fill=SRC_FILL; ws1.cell(row=r,column=9).border=thin

# Years 2-18
for y in range(1,18):
    row = 13+y
    ws1.cell(row=row,column=1,value=2023+y); ws1.cell(row=row,column=1).font=DATA_FONT; ws1.cell(row=row,column=1).border=thin; ws1.cell(row=row,column=1).alignment=Alignment(horizontal='center')
    prev_row = 12+y  # previous row
    # Col B: =H(prev) 
    ws1.cell(row=row,column=2,value=f'=H{prev_row}'); ws1.cell(row=row,column=2).font=DATA_FONT; ws1.cell(row=row,column=2).border=thin; ws1.cell(row=row,column=2).number_format=NUM
    # Col C: =B(row)*B5
    ws1.cell(row=row,column=3,value=f'=B{row}*B5'); ws1.cell(row=row,column=3).font=DATA_FONT; ws1.cell(row=row,column=3).border=thin; ws1.cell(row=row,column=3).number_format=NUM
    # Col D: =B7-C(row)
    ws1.cell(row=row,column=4,value=f'=B7-C{row}'); ws1.cell(row=row,column=4).font=DATA_FONT; ws1.cell(row=row,column=4).border=thin; ws1.cell(row=row,column=4).number_format=NUM
    # Col E: =B7
    ws1.cell(row=row,column=5,value='=B7'); ws1.cell(row=row,column=5).font=DATA_FONT; ws1.cell(row=row,column=5).border=thin; ws1.cell(row=row,column=5).number_format=NUM
    # Col F: last year lump sum
    if y==17:
        ws1.cell(row=row,column=6,value='=B9'); ws1.cell(row=row,column=6).font=DATA_FONT; ws1.cell(row=row,column=6).border=thin; ws1.cell(row=row,column=6).number_format=NUM
    else:
        ws1.cell(row=row,column=6,value=0); ws1.cell(row=row,column=6).font=DATA_FONT; ws1.cell(row=row,column=6).border=thin; ws1.cell(row=row,column=6).number_format=NUM
    # Col G: =E(row)+F(row)
    ws1.cell(row=row,column=7,value=f'=E{row}+F{row}'); ws1.cell(row=row,column=7).font=DATA_FONT; ws1.cell(row=row,column=7).border=thin; ws1.cell(row=row,column=7).number_format=NUM
    # Col H: =B(row)-D(row)
    ws1.cell(row=row,column=8,value=f'=B{row}-D{row}'); ws1.cell(row=row,column=8).font=DATA_FONT; ws1.cell(row=row,column=8).border=thin; ws1.cell(row=row,column=8).number_format=NUM
    # Col I: source
    srcs = ['同上年,公式递推']*17
    srcs[16] = '最后一年: 期末本金应=0; 一次性支付=22,298,077.79(含恩阳医院2,000万,不计回报利息)'
    ws1.cell(row=row,column=9,value=srcs[y-1]); ws1.cell(row=row,column=9).font=SRC_FONT; ws1.cell(row=row,column=9).fill=SRC_FILL; ws1.cell(row=row,column=9).border=thin

# Total row
tr=31
for c in range(1,10):
    ws1.cell(row=tr,column=c).font=BOLD_FONT; ws1.cell(row=tr,column=c).border=thin; ws1.cell(row=tr,column=c).alignment=Alignment(horizontal='center')
ws1.cell(row=tr,column=1,value='合计')
ws1.cell(row=tr,column=2,value='-')
for cc,l in [(3,'C13:C30'),(4,'D13:D30'),(5,'E13:E30'),(7,'G13:G30'),(8,'H13:H30')]:
    ws1.cell(row=tr,column=cc,value=f'=SUM({l})'); ws1.cell(row=tr,column=cc).number_format=NUM
ws1.cell(row=tr,column=6,value='=B9'); ws1.cell(row=tr,column=6).number_format=NUM
ws1.cell(row=tr,column=9,value='验算: 还本合计=80,000,000; 利息合计=PMT*18-P; 期末余额=0')

# Check row
tr2=32
ws1.cell(row=tr2,column=1,value='验算'); ws1.cell(row=tr2,column=1).font=WARN_FONT
ws1.cell(row=tr2,column=2,value=f'还本合计应=80,000,000 → =D{tr}=D{tr}'); ws1.cell(row=tr2,column=2).font=WARN_FONT
ws1.cell(row=tr2,column=3,value=f'利息合计应={80000000*0.0799*18:,.0f} → =C{tr}=C{tr}'); ws1.cell(row=tr2,column=3).font=WARN_FONT
ws1.cell(row=tr2,column=8,value=f'期末余额应=0 → [查看H30]'); ws1.cell(row=tr2,column=8).font=WARN_FONT
for c in range(1,10): ws1.cell(row=tr2,column=c).border=thin

# ============================================================
# SHEET 2: 银行融资
# ============================================================
ws2=wb.create_sheet('表2-银行融资')
title(ws2,1,'表2：实际融资本息测算表')
hdr(ws2,3,['参数','数值','单位','数据出处'],widths=[18,22,10,55])
param(ws2,4,'贷款总额','382800000','元','贵阳银行成都分行,累计发放贷款38,280万元')
param(ws2,5,'银行确认18年合计','710822119.32','元','贵阳银行《借款还本付息情况报告》(一级证据)')
param(ws2,6,'原征求意见稿表2合计','712464358.22','元','川融策咨询[2025]第490号征求意见稿 表2 小计行 (OCR提取)')
param(ws2,7,'缩放系数','=B5/B6','-','=银行确认数/原表合计; 逐年值=原表逐行×本系数 (二级推导)')
param(ws2,8,'贷款合同利率变化','6.95%→6.45%→5.5%→5.0%','-','贵阳银行贷款合同+还款报告; 2024.11→6.45%, 2025.11→5.5%, 2026.7→5.0%')

hdr(ws2,10,['年','原征求意见稿\n融资本息','缩放系数','本报告\n融资本息','剩余本金','适用利率','数据出处'],widths=[6,22,14,22,22,14,55])

bo=[(31686745.84,381300000,'6.95%'),(26809671.03,381300000,'6.45%'),
    (24811756.73,381250000,'5.5%/5.0%'),(21351219.32,380250000,'5.0%'),
    (20263854.17,379250000,'5.0%'),(20265833.33,378250000,'5.0%'),
    (20162465.30,377250000,'5.0%'),(20111770.82,376250000,'5.0%'),
    (23022951.39,372250000,'5.0%'),(22871874.99,368250000,'5.0%'),
    (22617395.82,364250000,'5.0%'),(22414618.06,360250000,'5.0%'),
    (27148298.64,351250000,'5.0%'),(27728124.99,341250000,'5.0%'),
    (35070729.16,323250000,'5.0%'),(34158229.18,305250000,'5.0%'),
    (311968819.45,0,'5.0%'),(0,0,'-')]
srcs2=[
    '原表2第1行; 2023年度利率6.95%','原表2第2行; 2024年度利率6.45%(11月调整)',
    '原表2第3行; 2025年度利率5.5%(11月28日调整)/5.0%','原表2第4行; 2026年7月调整为5.0%',
    '原表2第5行','原表2第6行','原表2第7行','原表2第8行','原表2第9行',
    '原表2第10行','原表2第11行','原表2第12行','原表2第13行','原表2第14行',
    '原表2第15行','原表2第16行','原表2第17行; 2039年贷款全部结清',
    '原表2第18行; 2040年无融资还款(2039年已结清)',
]

for y,(orig,rem,rate) in enumerate(bo):
    r=11+y
    data(ws2,r,[2023+y, orig, '=B7', f'=B{r}*C{r}', rem if rem>0 else '-', rate, srcs2[y]],
         fmts=[INT,NUM,PCT,NUM,NUM,None,None], alt=(y%2==1))
    # source column italics
    ws2.cell(row=r,column=7).font=SRC_FONT; ws2.cell(row=r,column=7).fill=SRC_FILL

# totals
tr2=29
data(ws2,tr2,['合计','=SUM(B11:B28)','-','=SUM(D11:D28)','-','—','原表2合计=712,464,358.22; 银行确认=710,822,119.32; 差额=1,642,238.90 → 逐年等比缩放'],
     fonts=[BOLD_FONT]*7, fmts=[INT,NUM,None,NUM,None,None,None])
ws2.cell(row=tr2,column=7).font=SRC_FONT; ws2.cell(row=tr2,column=7).fill=SRC_FILL

ws2.cell(row=30,column=1,value='缩放方法: 逐年值=原表逐年值×710,822,119.32÷712,464,358.22≈0.997695。本所未取得银行逐年还款明细,逐年值系二级推导。'); ws2.cell(row=30,column=1).font=WARN_FONT
ws2.merge_cells('A30:G30')

# ============================================================
# SHEET 3: 运维成本
# ============================================================
ws3=wb.create_sheet('表3-运维成本')
title(ws3,1,'表3：运营维护成本明细（合同公式: 运维成本x(1+K), K=7.99%）')
hdr(ws3,3,['参数','数值','单位','数据出处'],widths=[22,22,10,55])
param(ws3,4,'K(运维成本加成系数)','0.0799','-','合同: 《补充合同》(2017年7月)公式A=...+运维成本x(1+K)+...; K=社会资本中标年回报率')
param(ws3,5,'1+K','=1+B4','-','=1+7.99%')

hdr(ws3,7,['年','底层成本(税前)','K系数(1+K)','运维成本(含K)\n=底层x(1+K)','数据出处'],widths=[6,18,12,20,70])

raw2023=5341344.03; raw2024_y2=5948956.93; raw2024_y3=431456.22
raw2024=raw2024_y2+raw2024_y3
raw2025=5713241.86; raw_fut=5710000

ops = [
    (2023,raw2023,f'={raw2023}',f'=B8*C8','第一经营年度成本(2022.10.28-2023.10.27, 12个月)。原文: 征求意见稿正文"(三)运营维护成本...第一经营年度运营成本金额为5,341,344.03元"。本所验证: {raw2023:,.0f}x1.0799={raw2023*1.0799:,.2f}, 与征求意见稿表3的5,768,117.42完全一致→证明原表已含K。本所未取得2023年度原始账套独立验证。'),
    (2024,raw2024,f'={raw2024}',f'=B9*C9','第二经营年度5,948,956.93+第三经营年度(2个月)431,456.22。原文: 征求意见稿正文。本所验证: {raw2024_y2:,.0f}x1.0799={raw2024_y2*1.0799:,.2f} + {raw2024_y3:,.0f}x1.0799={raw2024_y3*1.0799:,.2f} = {raw2024*1.0799:,.2f}, 与征求意见稿表3的6,890,208.16完全一致。沿用,未经独立验证。'),
    (2025,raw2025,f'={raw2025}',f'=B10*C10','【独立验证】2025年度税前运维成本={raw2025:,.2f} = 主营业务成本Q1-Q4(1,016,316.83+667,176.24+1,476,847.53+1,313,687.15) + 管理费用Q1-Q4(558,779.11+287,198.91+213,914.32+179,321.77)。数据来源: 项目公司2025年度运营审计XLS明细账(6册), 取4个标准季度损益结转合计值, 不含年末调整项(+1,989,099.04)。'),
]
for i,(yr,raw,raw_str,fml,src) in enumerate(ops):
    data(ws3,8+i,[yr,raw, '=B5', fml, src], fmts=[INT,NUM,PCT,NUM,None], alt=(i%2==1))
    ws3.cell(row=8+i,column=5).font=SRC_FONT; ws3.cell(row=8+i,column=5).fill=SRC_FILL

for y in range(15):
    r=11+y
    data(ws3,r,[2026+y, raw_fut, '=B5', f'=B{r}*C{r}',
                f'税前成本按2025年度验证值取整{raw_fut:,.0f}x1.0799估算(第{y+1}/15年)。2025年度标准4Q合计={raw2025:,.2f}, 取整={raw_fut:,.0f}'],
         fmts=[INT,NUM,PCT,NUM,None], alt=((11+y)%2==1))
    ws3.cell(row=r,column=5).font=SRC_FONT; ws3.cell(row=r,column=5).fill=SRC_FILL

# totals
tr3=26
data(ws3,tr3,['合计','=SUM(B8:B25)','-','=SUM(D8:D25)','验算: 税前总计x(1+K)='],
     fonts=[BOLD_FONT]*5, fmts=[INT,NUM,None,NUM,None])
ws3.cell(row=tr3,column=5,value=f'税前总计x(1+K): (5341344.03+{raw2024:,.0f}+{raw2025:,.2f}+{raw_fut:,.0f}x15)x1.0799 = {(raw2023+raw2024+raw2025+raw_fut*15)*1.0799:,.2f}')
ws3.cell(row=tr3,column=5).font=WARN_FONT

# ============================================================
# SHEET 4: 第三方收入
# ============================================================
ws4=wb.create_sheet('表4-第三方收入')
title(ws4,1,'表4：第三方收入明细（不乘K系数, 直接扣减）')
hdr(ws4,3,['年','第三方收入','数据出处'],widths=[6,22,70])
inc_data=[
    (2023,4285070.31,'第一经营年度收入(2022.10.28-2023.12.21)。原文: 征求意见稿表3。沿用,未经独立验证。'),
    (2024,6229093.45,'第二经营年度5,841,165.87+第三经营年度(2个月)387,927.58。原文: 征求意见稿正文+表3。沿用,未经独立验证。'),
    (2025,5564017.88,'【独立验证】主营业务收入Q1-Q4(1,160,282+1,210,580+1,407,482+1,182,799)=4,961,142.62 + 其他业务收入Q1-Q4(166,208+29,264+173,959+233,444)=602,875.26。数据来源: 2025年度XLS明细账4季度损益结转合计。'),
]
for i,(yr,v,src) in enumerate(inc_data):
    data(ws4,4+i,[yr,v,src],fmts=[INT,NUM,None],alt=(i%2==1))
    ws4.cell(row=4+i,column=3).font=SRC_FONT; ws4.cell(row=4+i,column=3).fill=SRC_FILL
for y in range(15):
    r=7+y
    data(ws4,r,[2026+y,5560000.00,f'按2025年度验证值取整{5560000:,.0f}/年估算(第{y+1}/15年)'],fmts=[INT,NUM,None],alt=((7+y)%2==1))
    ws4.cell(row=r,column=3).font=SRC_FONT; ws4.cell(row=r,column=3).fill=SRC_FILL
tr4=22
data(ws4,tr4,['合计','=SUM(B4:B21)',''],fonts=[BOLD_FONT]*3,fmts=[INT,NUM,None])

# ============================================================
# SHEET 5: 可用性付费汇总 (cross-sheet formulas)
# ============================================================
ws5=wb.create_sheet('表5-可用性付费汇总')
title(ws5,1,'表5：可用性付费测算汇总表（单位：元）')
hdr(ws5,3,['参数','数值','单位','数据出处'],widths=[22,22,10,55])
param(ws5,4,'可用性付费公式','A=资本金回报+融资本息+运维成本x(1+K)-第三方收入','-','《补充合同》(2017年7月): A=P*k*(1+k)^n/((1+k)^n-1)+实际融资成本+运维成本x(1+K)-第三方收入')
param(ws5,5,'K取值','0.0799(即7.99%)','-','K=社会资本中标年回报率(同k)')

hdr(ws5,7,['年','资本金回报\n=表1!G列','实际融资本息\n=表2!D列','运维成本(含K)\n=表3!D列','第三方收入\n=表4!B列','可用性付费\n=SUM(B:F)','数据出处'],widths=[6,18,18,18,18,22,55])

# Yearly rows - can't easily cross-reference sheets with formulas in openpyxl dynamically
# Put computed values with formula annotations
for y in range(18):
    r=8+y
    # compute values
    cr_temp = 80000000; A_pmt = 80000000*0.0799*(1.0799)**18/((1.0799)**18-1)
    for yy in range(y):
        i_=cr_temp*0.0799; p_=A_pmt-i_
        if yy==17: p_=cr_temp; i_=A_pmt-p_
        cr_temp-=p_
        if cr_temp<0:cr_temp=0
    i_y=cr_temp*0.0799; p_y=A_pmt-i_y
    if y==17: p_y=cr_temp; i_y=A_pmt-p_y
    cap_ret=A_pmt+(22298077.79 if y==17 else 0)
    
    bo_list=[(31686745.84,381300000),(26809671.03,381300000),(24811756.73,381250000),
             (21351219.32,380250000),(20263854.17,379250000),(20265833.33,378250000),
             (20162465.30,377250000),(20111770.82,376250000),(23022951.39,372250000),
             (22871874.99,368250000),(22617395.82,364250000),(22414618.06,360250000),
             (27148298.64,351250000),(27728124.99,341250000),(35070729.16,323250000),
             (34158229.18,305250000),(311968819.45,0),(0,0)]
    sf_=710822119.32/sum(b[0] for b in bo_list)
    bk=round(bo_list[y][0]*sf_,2)
    
    if y==0: op_v=5341344.03*1.0799
    elif y==1: op_v=(5948956.93+431456.22)*1.0799
    elif y==2: op_v=5713241.86*1.0799
    else: op_v=5710000*1.0799
    
    if y==0: inc_v=4285070.31
    elif y==1: inc_v=6229093.45
    elif y==2: inc_v=5564017.88
    else: inc_v=5560000.00
    
    avail=cap_ret+bk+op_v-inc_v
    
    src5='资本金=等额本息逐年(表1)+一次性支付(仅2040); 运维x(1+7.99%); 收入不乘K'
    if y==17: src5='第18年(2040): 等额本息+C22,298,077.79一次性; 银行已结清(0); 运维含K'
    elif y==16: src5='第17年(2039): 银行贷款全部结清(最后一次大额还款)'
    
    data(ws5,r,[2023+y,round(cap_ret,2),round(bk,2),round(op_v,2),round(inc_v,2),round(avail,2),src5],
         fmts=[INT,NUM,NUM,NUM,NUM,NUM,None],alt=(y%2==1))
    ws5.cell(row=r,column=7).font=SRC_FONT; ws5.cell(row=r,column=7).fill=SRC_FILL

# totals
tr5=26
cap_tot=A_pmt*18+22298077.79
bank_tot=sum(round(b[0]*sf_,2) for b in bo_list)
op_tot=(5341344.03+5948956.93+431456.22+5713241.86+5710000*15)*1.0799
inc_tot=4285070.31+6229093.45+5564017.88+5560000*15
gtot=cap_tot+bank_tot+op_tot-inc_tot
data(ws5,tr5,['合计',round(cap_tot,2),round(bank_tot,2),round(op_tot,2),round(inc_tot,2),round(gtot,2),
              f'可用性付费总额=资本金+银行+运维-收入={gtot:,.2f}元(约{gtot/1e8:.2f}亿元)'],
     fonts=[BOLD_FONT]*7,fmts=[INT,NUM,NUM,NUM,NUM,NUM,None])

# Grand total highlight
ws5.cell(row=28,column=1,value=f'测算结论: 18年可用性付费总额 = {gtot:,.2f}元 ≈ {gtot/1e8:.2f}亿元')
ws5.cell(row=28,column=1).font=Font(name='微软雅黑',bold=True,size=14,color='0A1F3F')
ws5.merge_cells('A28:G28')

# ============================================================
# SHEET 6: 2025审计数据
# ============================================================
ws6=wb.create_sheet('表6-2025审计明细')
title(ws6,1,'表6：2025年度运营审计数据逐季拆解（单位：元）')
hdr(ws6,3,['科目','Q1损益结转','Q2损益结转','Q3损益结转','Q4损益结转','4Q合计','年末调整项','数据出处'],widths=[16,18,18,18,18,18,18,65])

qdata=[
    ('主营业务成本',1016316.83,667176.24,1476847.53,1313687.15,1409346.55,579752.49,
     'XLS: 主营业务成本明细.xls; Q1=1,016,316.83 Q2=667,176.24 Q3=1,476,847.53 Q4=1,313,687.15; 调整项=额外两个损益结转期间'),
    ('管理费用',558779.11,287198.91,213914.32,179321.77,428178.07,169550.88,
     'XLS: 管理费用明细.xls; 含工资/社保/差旅/办公费/折旧等'),
    ('主营业务收入',-1160281.89,-1210580.37,-1407481.80,-1182798.56,-1529843.40,-1251323.21,
     'XLS: 主营业务收入明细.xls; 物业收入+护工服务收入 (贷方为收入,以负数表示)'),
    ('其他业务收入',-166207.80,-29264.39,-173959.48,-233443.59,-44851.74,-283341.09,
     'XLS: 其他业务收入明细.xls; 停车场/充电宝/场地租赁等'),
]
for i,(nm,q1,q2,q3,q4,a1,a2,src) in enumerate(qdata):
    qsum=q1+q2+q3+q4; total=qsum+a1+a2
    data(ws6,4+i,[nm,q1,q2,q3,q4,round(qsum,2),round(a1+a2,2),src],fmts=[None,NUM,NUM,NUM,NUM,NUM,NUM,None],alt=(i%2==1))
    ws6.cell(row=4+i,column=8).font=SRC_FONT; ws6.cell(row=4+i,column=8).fill=SRC_FILL

# computed rows
data(ws6,8,['运维成本(主营+管理)','','','','',round(4474027.75+1239214.11,2),round(1409346.55+579752.49+428178.07+169550.88,2),
            '本报告取4Q合计=5,713,241.86(不含调整项); 含调整项合计=8,300,069.89'],fonts=[BOLD_FONT]*8,fmts=[None]*8)
data(ws6,9,['第三方收入(主营+其他)','','','','',round(4961142.62+602875.26,2),round(1529843.40+1251323.21+44851.74+283341.09,2),
            '本报告取4Q合计=5,564,017.88(不含调整项); 含调整项合计=8,392,034.11'],fonts=[BOLD_FONT]*8,fmts=[None]*8)
ws6.cell(row=11,column=1,value='注: 年末调整项包含额外2个损益结转期间, 本报告暂按4个标准季度取数, 不含调整项。'); ws6.cell(row=11,column=1).font=WARN_FONT
ws6.merge_cells('A11:H11')

# ============================================================
# SHEET 7: 版本对照
# ============================================================
ws7=wb.create_sheet('表7-版本对照')
title(ws7,1,'表7：版本差异对照 (征求意见稿 vs v4 vs v5)')
hdr(ws7,3,['项目','原征求意见稿','v4(含K因子bug)','v5(修正版)','差异说明'],widths=[18,22,22,22,55])

v4_op=104021567.44; v4_inc=99478181.64; v4_total=891208028.23
v5_op=op_tot; v5_inc=inc_tot; v5_total=gtot
orig_cap=177955158.81; orig_bank=712464358.22; orig_op=110108501.58; orig_inc=91474163.76
orig_total=orig_cap+orig_bank+orig_op-orig_inc

comp=[
    ('资本金回报',orig_cap,round(cap_tot,2),round(cap_tot,2),'原: 8229.8万全部等额本息(含恩阳医院2000万逐年还本无利息); v4/v5: 8000万等额本息+2229.8万第18年一次付(-211万)'),
    ('银行融资本息',orig_bank,710822119.33,round(bank_tot,2),'原: 表2合计712,464,358; v4/v5: 银行确认710,822,119,等比缩放逐年(-164万)'),
    ('运维成本(含K)',orig_op,round(v4_op,2),round(v5_op,2),f'v4 BUG: 2025+未乘(1+K),少算{round(v5_op-v4_op,0):,.0f}; v5修正(+730万)'),
    ('第三方收入',orig_inc,round(v4_inc,2),round(v5_inc,2),'v4/v5: 2025审计实际556万 vs 原估算506万; 收入增加→补贴减少(+800万)'),
    ('可用性付费合计',orig_total,round(v4_total,2),round(v5_total,2),f'原-v5={round(orig_total-v5_total,0):,.0f}; v4-v5={round(v4_total-v5_total,0):,.0f}'),
]
for i,(item,a,b,c,note) in enumerate(comp):
    data(ws7,4+i,[item,a,b,c,note],fmts=[None,NUM,NUM,NUM,None],alt=(i%2==1))
    ws7.cell(row=4+i,column=5).font=SRC_FONT; ws7.cell(row=4+i,column=5).fill=SRC_FILL

# Save
outpath=r'C:\Users\scrccpa\Desktop\恩阳医养园PPP测算数据基础表v5.xlsx'
wb.save(outpath)
print(f'Saved: {outpath} ({os.path.getsize(outpath):,} bytes)')
print(f'Grand total: {gtot:,.2f} ({gtot/1e8:.2f}亿)')
