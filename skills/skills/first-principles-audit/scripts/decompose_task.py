#!/usr/bin/env python3
"""
审计任务第一性原理分解 — 六原子要素模型

参考: 田川不是四川《审计平台开发_第一性原理分析》(2026-05-18)

输入: 审计任务描述
输出: 原子要素分解 + 自动化可行性 + 工具推荐
"""
import sys, io, argparse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

GRN = PatternFill(patternType='solid', fgColor='D4EDDA')
YEL = PatternFill(patternType='solid', fgColor='FFF3CD')
RED = PatternFill(patternType='solid', fgColor='FFD7D7')
HEADER = PatternFill(patternType='solid', fgColor='1A3A6E')
H = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
N = Font(name='Microsoft YaHei', size=10)
B = Font(name='Microsoft YaHei', size=10, bold=True)
TH = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
C = Alignment(horizontal='center', vertical='center', wrap_text=True)
L = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ── 审计任务模板库 ──
TASK_TEMPLATES = {
    '招投标审计': {
        '信息收集': {
            'actions': ['收集招标公告/招标文件', '收集投标文件(全部投标人)', '收集开标记录/评审打分表', '收集中标通知书/合同'],
            'auto': '✅ 全自动',
            'tools': 'unstructured-audit-data(batch_unzip/OCR) + PDF元数据提取',
            'time_saved': '80%',
        },
        '风险评估': {
            'actions': ['节资率异常检测', '投标人数最优分析(3-4家)', '围标风险特征扫描'],
            'auto': '△ 半自动',
            'tools': 'procurement-audit-models(L1-L19) + apriori-audit',
            'time_saved': '70%',
        },
        '证据获取': {
            'actions': ['TF-IDF文本雷同检测', 'PDF元数据同源分析', 'JPEG量化表指纹比对', '工商关联查询'],
            'auto': '△ 半自动',
            'tools': 'procurement-audit-models(L3-L8) + forensic脚本',
            'time_saved': '60%',
        },
        '证据评价': {
            'actions': ['判断证据是否构成围标', '判断招标文件条款是否构成设限', '判断串标行为的性质'],
            'auto': '✗ 人工',
            'tools': '审计人员专业判断',
            'time_saved': '0%',
        },
        '结论形成': {
            'actions': ['确定是否属于围标串标', '确认招标程序是否合规', '形成审计问题定性'],
            'auto': '✗ 人工',
            'tools': 'audit_finding_processor(辅助定性+法规+整改)',
            'time_saved': '30%',
        },
        '工作记录': {
            'actions': ['生成证据索引', '编制问题汇总表', '生成审计底稿'],
            'auto': '✅ 全自动',
            'tools': 'analysis-report + audit_pipeline(综合报告)',
            'time_saved': '90%',
        },
    },
    '财务审计': {
        '信息收集': {
            'actions': ['采集科目余额表/序时账', '收集凭证扫描件', '收集合同/发票', '收集银行对账单'],
            'auto': '✅ 全自动',
            'tools': 'data-analyst-cn + OCR + RAG',
            'time_saved': '70%',
        },
        '风险评估': {
            'actions': ['Benford定律首位数测试', '异常交易检测', '关联交易扫描'],
            'auto': '△ 半自动',
            'tools': 'financial-fraud-detection + anomaly-detection',
            'time_saved': '60%',
        },
        '证据获取': {
            'actions': ['函证发送/回收(可自动化)', '抽样/详查', '访谈(需人工)'],
            'auto': '△ 半自动',
            'tools': 'Python抽样脚本 + 人工访谈',
            'time_saved': '30%',
        },
        '证据评价': {
            'actions': ['判断异常交易是否属于舞弊', '评估证据充分性'],
            'auto': '✗ 人工',
            'tools': '审计人员专业判断',
            'time_saved': '0%',
        },
        '结论形成': {
            'actions': ['确定审计意见', '形成审计问题'],
            'auto': '✗ 人工',
            'tools': 'audit_finding_processor(辅助)',
            'time_saved': '20%',
        },
        '工作记录': {
            'actions': ['生成审计底稿', '编制试算平衡表', '生成审计报告'],
            'auto': '✅ 全自动',
            'tools': 'analysis-report + Excel自动化',
            'time_saved': '85%',
        },
    },
    '绩效评价': {
        '信息收集': {
            'actions': ['采集绩效目标表', '收集项目验收报告', '收集满意度调查数据'],
            'auto': '✅ 全自动',
            'tools': 'data-analyst-cn + OCR',
            'time_saved': '70%',
        },
        '风险评估': {
            'actions': ['绩效目标完成度分析', '资金使用效率分析'],
            'auto': '△ 半自动',
            'tools': 'data-analyst-cn + anomaly-detection',
            'time_saved': '50%',
        },
        '证据获取': {
            'actions': ['绩效指标对比分析', '满意度数据统计', '现场勘查(需人工)'],
            'auto': '△ 半自动',
            'tools': 'Python统计分析 + 人工现场',
            'time_saved': '40%',
        },
        '证据评价': {
            'actions': ['判断绩效是否达标', '评价绩效管理有效性'],
            'auto': '✗ 人工',
            'tools': '绩效评价专业判断',
            'time_saved': '0%',
        },
        '结论形成': {
            'actions': ['形成绩效评价结论', '提出改进建议'],
            'auto': '✗ 人工',
            'tools': 'audit_risk_navigator(辅助)',
            'time_saved': '15%',
        },
        '工作记录': {
            'actions': ['生成绩效评价底稿', '编制绩效评分表', '生成绩效评价报告'],
            'auto': '✅ 全自动',
            'tools': 'analysis-report',
            'time_saved': '85%',
        },
    },
    '资产清查': {
        '信息收集': {
            'actions': ['采集资产卡片/台账', '采集折旧明细', '收集资产盘点表'],
            'auto': '✅ 全自动',
            'tools': 'data-analyst-cn',
            'time_saved': '75%',
        },
        '风险评估': {
            'actions': ['资产账实差异分析', '闲置资产识别', '资产报废合理性分析'],
            'auto': '△ 半自动',
            'tools': 'data-analyst-cn + anomaly-detection',
            'time_saved': '55%',
        },
        '证据获取': {
            'actions': ['资产盘点(需人工)', '资产照片采集', '权属证明核查'],
            'auto': '△ 半自动',
            'tools': 'Python盘点表生成 + 人工实地盘点',
            'time_saved': '35%',
        },
        '证据评价': {
            'actions': ['判断盘盈盘亏原因', '评价资产管理规范性'],
            'auto': '✗ 人工',
            'tools': '资产管理人员专业判断',
            'time_saved': '0%',
        },
        '结论形成': {
            'actions': ['形成清查结论', '提出处理建议'],
            'auto': '✗ 人工',
            'tools': 'audit_finding_processor(辅助)',
            'time_saved': '15%',
        },
        '工作记录': {
            'actions': ['生成资产清查底稿', '编制差异对比表', '生成清查报告'],
            'auto': '✅ 全自动',
            'tools': 'analysis-report',
            'time_saved': '85%',
        },
    },
}


def decompose(task_type: str, output: str = None):
    """第一性原理分解审计任务"""

    if task_type not in TASK_TEMPLATES:
        matches = [k for k in TASK_TEMPLATES if task_type in k]
        if matches:
            task_type = matches[0]
        else:
            print(f"未找到: {task_type}")
            print(f"可用: {', '.join(TASK_TEMPLATES.keys())}")
            return

    template = TASK_TEMPLATES[task_type]
    elements = ['信息收集', '风险评估', '证据获取', '证据评价', '结论形成', '工作记录']

    print(f"\n{'='*60}")
    print(f"🔬 第一性原理分解: {task_type}")
    print(f"{'='*60}")

    total_auto = 0
    total_semi = 0
    total_human = 0
    avg_time_saved = 0

    for elem in elements:
        info = template[elem]
        marker = '✅' if '全自动' in info['auto'] else ('△' if '半自动' in info['auto'] else '✗')
        saved = int(info['time_saved'].replace('%', ''))

        print(f"\n{marker} {elem} [{info['auto']}] — 预计节省{saved}%时间")
        for a in info['actions']:
            print(f"   • {a}")
        print(f"   工具: {info['tools']}")

        if '全自动' in info['auto']:
            total_auto += 1
        elif '半自动' in info['auto']:
            total_semi += 1
        else:
            total_human += 1
        avg_time_saved += saved

    avg_time_saved //= 6

    print(f"\n{'='*60}")
    print(f"📊 自动化分析")
    print(f"   全自动: {total_auto}/6  半自动: {total_semi}/6  人工: {total_human}/6")
    print(f"   预计整体效率提升: ~{avg_time_saved}%")
    print(f"{'='*60}")

    if output:
        wb = Workbook()
        ws = wb.active
        ws.title = '任务分解'

        ws.merge_cells('A1:F1')
        ws['A1'] = f'{task_type} — 第一性原理任务分解'
        ws['A1'].font = Font(name='Microsoft YaHei', size=14, bold=True, color='1A3A6E')

        ws.merge_cells('A2:F2')
        ws['A2'] = f'自动: {total_auto}  |  半自动: {total_semi}  |  人工: {total_human}  |  效率提升: ~{avg_time_saved}%'
        ws['A2'].font = Font(name='Microsoft YaHei', size=9, italic=True, color='888888')

        sub_h = ['原子要素', '自动化等级', '具体动作', '推荐工具', '预计节省', '融策落地']
        for c, h in enumerate(sub_h, 1):
            cl = ws.cell(row=4, column=c, value=h)
            cl.font = H; cl.fill = HEADER; cl.alignment = C; cl.border = TH

        for i, elem in enumerate(elements):
            r = i + 5
            info = template[elem]
            auto_level = info['auto']
            fill = GRN if '全自动' in auto_level else (YEL if '半自动' in auto_level else RED)

            vals = [
                elem,
                auto_level,
                '\n'.join(info['actions']),
                info['tools'],
                info['time_saved'],
                '✅ 已就绪' if '全自动' in auto_level else ('⚠️ 部分就绪' if '半自动' in auto_level else '📋 需人工作业'),
            ]
            for c, val in enumerate(vals, 1):
                cl = ws.cell(row=r, column=c, value=val)
                cl.font = N; cl.alignment = L; cl.border = TH
                if c == 2:
                    cl.fill = fill; cl.alignment = C
                if c == 6:
                    cl.fill = fill; cl.alignment = C

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 16

        wb.save(output)
        print(f"\n✅ 策略表: {output}")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='审计任务第一性原理分解')
    parser.add_argument('task', nargs='?', help='审计任务类型')
    parser.add_argument('--o', '--output', dest='output', help='输出Excel路径')
    parser.add_argument('--list', action='store_true', help='列出已支持的任务类型')
    args = parser.parse_args()

    if args.list:
        print("已支持的审计任务模板:")
        for k, v in TASK_TEMPLATES.items():
            total_auto = sum(1 for e in v.values() if '全自动' in e['auto'])
            total_semi = sum(1 for e in v.values() if '半自动' in e['auto'])
            print(f"  {k} ({total_auto}全自动/{total_semi}半自动/{6-total_auto-total_semi}人工)")
    elif args.task:
        decompose(args.task, args.output)
    else:
        print("审计任务第一性原理分解器")
        print()
        print("已加载模板:")
        for k in TASK_TEMPLATES:
            print(f"  • {k}")
        print()
        task = input("输入审计任务类型 (如: 招投标审计) > ").strip()
        if task:
            decompose(task)
