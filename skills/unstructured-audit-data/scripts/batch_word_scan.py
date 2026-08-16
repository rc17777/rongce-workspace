#!/usr/bin/env python3
"""Word文件批处理: 批量扫描限制性关键词→高亮标注→生成疑点汇总Excel"""
import sys, io, os, re, argparse
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# 招标文件设限关键词库
RESTRICTION_KEYWORDS = {
    '资格限制': ['业绩', '奖项', '专利', '规模', '资质等级', '注册资金', '注册资本',
             '特定品牌', '指定型号', '唯一授权', '独家代理', '指定供应商',
             '一级资质', '二级资质', '甲级', '乙级', '项目经理资质'],
    '地域限制': ['本地', '本市', '省内', '注册地', '纳税地', '分支机构',
             '本地化服务', '驻场', '常住', '户籍'],
    '时间限制': ['成立年限', '经营年限', '从业经验', '满.*年', '不少于.*年'],
    '规模限制': ['营业额', '营业收入', '资产总额', '净资产', '从业人员',
             '社保人数', '纳税额'],
    '其他限制': ['联合体', '不接受.*进口', '原厂商', '制造商授权',
             '厂家授权', '售后服务网点', '本地库房'],
}


def scan_word_file(filepath: str, keywords: dict) -> list:
    """扫描单个Word文件"""
    from docx import Document

    hits = []
    try:
        doc = Document(filepath)
        for i, para in enumerate(doc.paragraphs):
            text = para.text.strip()
            if not text or len(text) < 5:
                continue
            for category, kws in keywords.items():
                for kw in kws:
                    if re.search(kw, text):
                        hits.append({
                            'file': Path(filepath).name,
                            'para_no': i + 1,
                            'category': category,
                            'keyword': kw,
                            'text': text[:300]
                        })
                        break  # 一段只记一次
    except Exception as e:
        print(f"  错误处理 {filepath}: {e}")
    return hits


def batch_scan(docx_dir: str, output_xlsx: str = None) -> list:
    """批量扫描目录下所有docx文件"""
    docx_dir = Path(docx_dir)
    files = list(docx_dir.glob('*.docx'))
    print(f"找到 {len(files)} 个Word文件")

    all_hits = []
    for f in files:
        print(f"扫描: {f.name}")
        hits = scan_word_file(str(f), RESTRICTION_KEYWORDS)
        all_hits.extend(hits)
        if hits:
            cats = set(h['category'] for h in hits)
            print(f"  发现 {len(hits)} 处限制关键词: {cats}")

    # Summary by keyword
    kw_count = {}
    for h in all_hits:
        k = h['keyword']
        kw_count[k] = kw_count.get(k, 0) + 1

    print(f"\n=== 汇总 ===")
    print(f"总命中: {len(all_hits)} 处")
    for kw, cnt in sorted(kw_count.items(), key=lambda x: -x[1])[:15]:
        print(f"  {kw}: {cnt}")

    # Export to Excel
    if output_xlsx and all_hits:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = '限制关键词扫描'

        from openpyxl.styles import Font, PatternFill, Alignment
        H = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        HDR = PatternFill(patternType='solid', fgColor='1A3A6E')
        YEL = PatternFill(patternType='solid', fgColor='FFF3CD')

        headers = ['文件', '段落号', '限制类别', '关键词', '文本内容']
        for c, h in enumerate(headers, 1):
            cl = ws.cell(row=1, column=c, value=h)
            cl.font = H; cl.fill = HDR

        for i, hit in enumerate(all_hits, 2):
            for c, key in enumerate(['file', 'para_no', 'category', 'keyword', 'text'], 1):
                cl = ws.cell(row=i, column=c, value=hit[key])
                cl.fill = YEL if c == 4 else None

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 60

        wb.save(output_xlsx)
        print(f"\n疑点表已保存: {output_xlsx}")

    return all_hits


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='批量扫描Word文件限制性关键词')
    parser.add_argument('docx_dir', help='Word文件目录')
    parser.add_argument('output_xlsx', nargs='?', help='输出Excel路径')
    args = parser.parse_args()

    batch_scan(args.docx_dir, args.output_xlsx)
