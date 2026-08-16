#!/usr/bin/env python3
"""PDF批处理: 投标文件相似度计算(Simhash/TF-IDF) + 价格浮动分析"""
import sys, io, os, re, argparse, json
from pathlib import Path
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')


def extract_text_fitz(pdf_path: str) -> str:
    """使用PyMuPDF提取PDF文本"""
    import fitz
    doc = fitz.open(pdf_path)
    text = ''
    for page in doc:
        text += page.get_text()
    doc.close()
    return text


def clean_text(text: str) -> str:
    """清洗文本"""
    text = re.sub(r'\s+', ' ', text)
    text = re.sub(r'[^\u4e00-\u9fffa-zA-Z0-9 \n]', '', text)
    return text.strip()


def compute_simhash_similarity(texts: dict) -> dict:
    """Simhash海明距离矩阵"""
    try:
        from simhash import Simhash
    except ImportError:
        print("需要安装simhash: pip install simhash")
        return {}

    # Compute simhash for each text
    hashes = {}
    for name, text in texts.items():
        if len(text) > 50:
            hashes[name] = Simhash(text)

    # Pairwise comparison
    results = {}
    names = list(hashes.keys())
    for i, n1 in enumerate(names):
        for n2 in names[i + 1:]:
            dist = hashes[n1].distance(hashes[n2])
            results[f'{n1} vs {n2}'] = {
                'hamming_distance': dist,
                'risk': 'HIGH' if dist <= 3 else ('MEDIUM' if dist <= 6 else 'LOW'),
                'max_bits': hashes[n1].f
            }
    return results


def compute_tfidf_similarity(texts: dict) -> dict:
    """TF-IDF余弦相似度矩阵"""
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.metrics.pairwise import cosine_similarity
    import numpy as np

    names = list(texts.keys())
    docs = [texts[n] for n in names]
    vec = TfidfVectorizer(max_features=5000, token_pattern=r'[\u4e00-\u9fff]+')
    mat = vec.fit_transform(docs)
    sim = cosine_similarity(mat)

    results = {}
    for i, n1 in enumerate(names):
        for j, n2 in enumerate(names):
            if i >= j: continue
            score = sim[i][j]
            results[f'{n1} vs {n2}'] = {
                'cosine_similarity': round(score, 4),
                'risk': 'HIGH' if score >= 0.8 else ('MEDIUM' if score >= 0.5 else 'LOW')
            }
    return results


def batch_analyze(pdf_dir: str, output: str = None, method: str = 'tfidf',
                  max_files: int = 200, page_limit: int = 50) -> dict:
    """批量分析PDF目录"""
    pdf_dir = Path(pdf_dir)
    pdf_files = list(pdf_dir.glob('*.pdf'))[:max_files]

    print(f"找到 {len(pdf_files)} 个PDF文件")
    if len(pdf_files) < 2:
        print("需要至少2个PDF文件进行对比")
        return {}

    # Extract text
    texts = {}
    for f in pdf_files:
        print(f"提取: {f.name}")
        try:
            raw = extract_text_fitz(str(f))
            clean = clean_text(raw)
            if len(clean) > 100:
                texts[f.stem] = clean
            else:
                print(f"  文本过短({len(clean)}字符)，可能需要OCR")
        except Exception as e:
            print(f"  错误: {e}")

    print(f"\n成功提取 {len(texts)} 个文件的有效文本")

    # Compute similarity
    if method == 'simhash':
        print("使用Simhash算法...")
        results = compute_simhash_similarity(texts)
    else:
        print("使用TF-IDF算法...")
        results = compute_tfidf_similarity(texts)

    # Print results
    high_risks = {k: v for k, v in results.items() if v.get('risk') == 'HIGH'}
    print(f"\n=== 相似度分析结果 ===")
    print(f"高风险(HIGH): {len(high_risks)} 对")

    for pair, info in sorted(results.items(),
                              key=lambda x: x[1].get('cosine_similarity', x[1].get('hamming_distance', 0)),
                              reverse=method == 'tfidf'):
        if method == 'tfidf':
            val = f"余弦相似度={info['cosine_similarity']:.3f}"
        else:
            val = f"海明距离={info['hamming_distance']}"
        marker = '🚨' if info['risk'] == 'HIGH' else ('⚠️' if info['risk'] == 'MEDIUM' else '  ')
        print(f"  {marker} {pair}: {val} [{info['risk']}]")

    # Export to Excel
    if output:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active; ws.title = '相似度分析'
        H = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        HDR = PatternFill(patternType='solid', fgColor='1A3A6E')
        RED = PatternFill(patternType='solid', fgColor='FFD7D7')
        YEL = PatternFill(patternType='solid', fgColor='FFF3CD')

        headers = ['对比项', '相似度值', '风险等级']
        for c, h in enumerate(headers, 1):
            cl = ws.cell(row=1, column=c, value=h); cl.font = H; cl.fill = HDR

        for i, (pair, info) in enumerate(sorted(results.items()), 2):
            ws.cell(row=i, column=1, value=pair)
            val = info.get('cosine_similarity', info.get('hamming_distance', ''))
            ws.cell(row=i, column=2, value=val)
            risk = info['risk']
            cl = ws.cell(row=i, column=3, value=risk)
            cl.fill = RED if risk == 'HIGH' else (YEL if risk == 'MEDIUM' else None)

        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 14

        wb.save(output)
        print(f"\n结果已保存: {output}")

    return results


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='PDF投标文件相似度分析')
    parser.add_argument('pdf_dir', help='PDF文件目录')
    parser.add_argument('--method', default='tfidf', choices=['tfidf', 'simhash'],
                        help='相似度算法: tfidf(默认) 或 simhash')
    parser.add_argument('--output', '-o', help='输出Excel路径')
    parser.add_argument('--max-files', type=int, default=200, help='最大处理文件数')
    args = parser.parse_args()

    batch_analyze(args.pdf_dir, args.output, args.method, args.max_files)
