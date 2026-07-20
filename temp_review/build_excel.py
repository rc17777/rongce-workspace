# -*- coding: utf-8 -*-
"""生成复核结果Excel到桌面"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, r'C:\Users\scrccpa\.openclaw\workspace\temp_review')
from findings_data import FINDINGS, RECONCILE, SETTLE_VS_ACCEPT, NO_ACCEPT_ITEMS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

HDR_FILL = PatternFill('solid', fgColor='0A1F3F')
HDR_FONT = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
BODY_FONT = Font(name='微软雅黑', size=9)
TITLE_FONT = Font(name='微软雅黑', size=13, bold=True, color='0A1F3F')
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
THIN = Border(*[Side(style='thin', color='B0B0B0')]*4)
FILLS = {'P0': PatternFill('solid', fgColor='F8CBAD'),
         'P1': PatternFill('solid', fgColor='FFE699'),
         'P2': PatternFill('solid', fgColor='E7E6E6')}
OK_FILL = PatternFill('solid', fgColor='C6EFCE')
BAD_FILL = PatternFill('solid', fgColor='FFC7CE')

def sheet_write(ws, title, header, rows, widths, fill_col=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header))
    c = ws.cell(row=1, column=1, value=title)
    c.font = TITLE_FONT; c.alignment = Alignment(vertical='center')
    ws.row_dimensions[1].height = 24
    for j, h in enumerate(header, 1):
        cell = ws.cell(row=2, column=j, value=h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = THIN
    for i, row in enumerate(rows, 3):
        for j, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = BODY_FONT; cell.alignment = WRAP; cell.border = THIN
        if fill_col is not None:
            key = str(row[fill_col])[:2]
            if key in FILLS:
                for j in range(1, len(header)+1):
                    ws.cell(row=i, column=j).fill = FILLS[key]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = 'A3'

# ---- Sheet1 复核发现汇总 ----
ws1 = wb.active
ws1.title = '复核发现汇总'
sheet_write(ws1, '审核报告复核发现汇总表（川竞泽专审字〔2026〕第027号 · 复核日期2026-07-20）',
            ['序号','复核维度','风险等级','发现描述','位置/证据','置信度','修改建议'],
            FINDINGS, [5,16,11,60,32,7,42], fill_col=2)

# ---- Sheet2 数据勾稽核验（通过项） ----
ws2 = wb.create_sheet('数据勾稽核验')
rows2 = [[i+1]+r for i, r in enumerate(RECONCILE)]
sheet_write(ws2, '数据勾稽核验表（已通过项）',
            ['序号','核验项','报告/清单数值','证据来源','证据数值','核验结果'],
            rows2, [5,30,24,34,20,10])
for i in range(3, 3+len(rows2)):
    ws2.cell(row=i, column=6).fill = OK_FILL

# ---- Sheet3 结算vs验收比对 ----
ws3 = wb.create_sheet('结算清单vs验收比对')
rows3 = [[i+1]+r for i, r in enumerate(SETTLE_VS_ACCEPT)]
sheet_write(ws3, '结算清单与验收报告数量比对明细表（差异项）',
            ['序号','项目','结算数量','验收数量','单价(元)','差异金额(元)','方向','备注'],
            rows3, [5,26,10,10,9,13,10,24])
for i in range(3, 3+len(rows3)):
    d = ws3.cell(row=i, column=7).value
    ws3.cell(row=i, column=7).fill = BAD_FILL if d in ('多结','计算矛盾') else OK_FILL
# 汇总行
tr = 3+len(rows3)
ws3.cell(row=tr, column=2, value='合计').font = Font(name='微软雅黑', size=9, bold=True)
ws3.cell(row=tr, column=6, value='多结755 / 少结8,271 / 净少结7,516').font = Font(name='微软雅黑', size=9, bold=True)
for j in range(1, 9):
    ws3.cell(row=tr, column=j).border = THIN

# ---- Sheet4 无验收佐证项目 ----
ws4 = wb.create_sheet('无验收佐证项目')
rows4 = [[i+1, n, a] for i, (n, a) in enumerate(NO_ACCEPT_ITEMS)]
sheet_write(ws4, '结算清单中验收报告附1无对应项的项目（合计约52,400元）',
            ['序号','项目','金额(元)'], rows4, [5,40,14])
tr = 4+len(rows4) if False else 3+len(rows4)
ws4.cell(row=tr, column=2, value='合计').font = Font(name='微软雅黑', size=9, bold=True)
ws4.cell(row=tr, column=3, value=sum(a for _, a in NO_ACCEPT_ITEMS)).font = Font(name='微软雅黑', size=9, bold=True)
for j in range(1, 4):
    ws4.cell(row=tr, column=j).border = THIN

# ---- Sheet5 复核统计 ----
ws5 = wb.create_sheet('复核统计')
p0 = sum(1 for f in FINDINGS if f[2].startswith('P0'))
p1 = sum(1 for f in FINDINGS if f[2].startswith('P1'))
p2 = sum(1 for f in FINDINGS if f[2].startswith('P2'))
stats = [
    ['复核对象','制造业数字化转型促进中心深度行（四川站）活动经费审核报告（川竞泽专审字〔2026〕第027号，修改稿）'],
    ['复核日期','2026-07-20'],
    ['复核方法','审计报告AI复核15维检查法：三方交叉（报告↔合同↔发票↔结算清单↔验收报告↔底稿）+ 正文十维度'],
    ['复核材料','审核报告docx、审计业务约定书、主/分包合同及发票、银行回单、结算清单xlsx、验收报告pptx、专项审计底稿doc'],
    ['检出合计', f'{p0+p1+p2}项（P0重大矛盾 {p0}项 / P1重大缺陷 {p1}项 / P2口径格式 {p2}项）'],
    ['通过核验', f'{len(RECONCILE)}项（详见"数据勾稽核验"表）'],
    ['核心结论','金额主链条勾稽一致（195,200=186,158+9,042，发票、合同、回单三环吻合）；但履约佐证存在重大缺口：5项多结755元、软文38点位仅17家有佐证、约5.24万元项目无验收对应、底稿"不涉及金额审计"与报告矛盾。建议提交前整改P0/P1项。'],
    ['重要提示','AI复核为初审工具，所有发现需人工逐条验证后采信；主合同、约定书为纯扫描件未能读取正文，签订日期等以原件为准。'],
]
sheet_write(ws5, '复核统计与结论', ['项目','内容'], stats, [14,110])

out = r'C:\Users\scrccpa\Desktop\审核报告复核结果-川竞泽专审字2026第027号.xlsx'
wb.save(out)
print('已生成:', out, f'({os.path.getsize(out)/1024:.1f} KB)')
print(f'P0:{p0} P1:{p1} P2:{p2} 通过:{len(RECONCILE)}')
