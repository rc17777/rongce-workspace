#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
融策审计驾驶舱 — 数据管理模块
文件上传 · Excel解析 · 合同/招投标资料导入 · 对接audit-blackboard
"""
import sys, os, json, hashlib, time, io, shutil
sys.stdout.reconfigure(encoding='utf-8')

from flask import Blueprint, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename
from datetime import datetime

data_bp = Blueprint('data', __name__, template_folder='../templates')

# 上传目录
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), 'uploads')
ALLOWED_EXTENSIONS = {'pdf', 'docx', 'xlsx', 'xls', 'csv', 'jpg', 'png', 'zip'}
PROJECTS_DIR = os.path.join(os.path.dirname(__file__), 'projects')

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(PROJECTS_DIR, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ========== 页面 ==========
@data_bp.route('/data')
def data_page():
    return render_template('data_manage.html')

# ========== API：文件上传 ==========
@data_bp.route('/api/upload', methods=['POST'])
def upload_file():
    """上传合同、招投标文件、财务报表等"""
    if 'file' not in request.files:
        return jsonify({"error": "没有文件"}), 400
    
    files = request.files.getlist('file')
    category = request.form.get('category', 'other')  # contract/bidding/financial/other
    project = request.form.get('project', 'default')
    
    # 按项目分目录
    project_dir = os.path.join(UPLOAD_DIR, project, category)
    os.makedirs(project_dir, exist_ok=True)
    
    results = []
    for f in files:
        if f.filename == '':
            continue
        if f and allowed_file(f.filename):
            filename = secure_filename(f.filename)
            # 加时间戳防重名
            name, ext = os.path.splitext(filename)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            saved_name = f"{name}_{timestamp}{ext}"
            filepath = os.path.join(project_dir, saved_name)
            f.save(filepath)
            
            # 提取文本（PDF/DOCX）
            text_preview = extract_text_preview(filepath, ext)
            
            results.append({
                "original": filename,
                "saved": saved_name,
                "path": filepath,
                "size": os.path.getsize(filepath),
                "category": category,
                "preview": text_preview[:500] if text_preview else None,
            })
        else:
            results.append({"original": f.filename, "error": "不支持的文件格式"})
    
    return jsonify({"success": True, "files": results})

# ========== API：已上传文件列表 ==========
@data_bp.route('/api/files')
def list_files():
    """列出已上传的文件"""
    project = request.args.get('project', 'default')
    category = request.args.get('category', '')
    
    files = []
    base = os.path.join(UPLOAD_DIR, project)
    
    if not os.path.exists(base):
        return jsonify({"files": [], "categories": []})
    
    categories = []
    for cat in os.listdir(base):
        cat_path = os.path.join(base, cat)
        if os.path.isdir(cat_path):
            categories.append(cat)
            if not category or category == cat:
                for fn in os.listdir(cat_path):
                    fp = os.path.join(cat_path, fn)
                    if os.path.isfile(fp):
                        files.append({
                            "name": fn,
                            "category": cat,
                            "size": os.path.getsize(fp),
                            "modified": datetime.fromtimestamp(os.path.getmtime(fp)).strftime("%Y-%m-%d %H:%M"),
                            "project": project,
                        })
    
    return jsonify({"files": sorted(files, key=lambda x: x['modified'], reverse=True), "categories": sorted(categories)})

# ========== API：Excel导入 ==========
@data_bp.route('/api/import/excel', methods=['POST'])
def import_excel():
    """导入Excel财务报表并结构化"""
    if 'file' not in request.files:
        return jsonify({"error": "没有文件"}), 400
    
    f = request.files['file']
    if not f.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"error": "仅支持 .xlsx / .xls 文件"}), 400
    
    filename = secure_filename(f.filename)
    filepath = os.path.join(UPLOAD_DIR, 'temp', filename)
    os.makedirs(os.path.join(UPLOAD_DIR, 'temp'), exist_ok=True)
    f.save(filepath)
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheets_data = {}
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            headers = []
            for i, row in enumerate(ws.iter_rows(max_row=min(ws.max_row, 200), values_only=True)):
                if i == 0:
                    headers = [str(c) if c else '' for c in row]
                else:
                    rows.append([str(c) if c is not None else '' for c in row])
            sheets_data[sheet_name] = {"headers": headers, "rows": rows, "total_rows": ws.max_row}
        
        return jsonify({
            "success": True,
            "filename": filename,
            "sheets": sheets_data,
            "recommendation": analyze_excel_structure(sheets_data),
        })
    except Exception as e:
        return jsonify({"error": f"Excel解析失败: {str(e)}"}), 500

# ========== API：对接audit-blackboard ==========
@data_bp.route('/api/projects/blackboard')
def list_blackboard_projects():
    """列出audit-blackboard已有项目"""
    bb_projects = []
    bb_dir = os.path.join(os.path.dirname(__file__), '..', '..', 'audit-blackboard', 'projects')
    if os.path.exists(bb_dir):
        for name in os.listdir(bb_dir):
            proj = os.path.join(bb_dir, name)
            if os.path.isdir(proj):
                status_file = os.path.join(proj, 'status.json')
                status = {}
                if os.path.exists(status_file):
                    try:
                        with open(status_file, 'r', encoding='utf-8') as sf:
                            status = json.load(sf)
                    except:
                        pass
                
                raw_files = []
                raw_dir = os.path.join(proj, 'raw_data')
                if os.path.exists(raw_dir):
                    raw_files = [fn for fn in os.listdir(raw_dir) if os.path.isfile(os.path.join(raw_dir, fn))]
                
                bb_projects.append({
                    "name": name,
                    "path": proj,
                    "status": status,
                    "raw_files": raw_files,
                    "file_count": len(raw_files),
                })
    
    return jsonify({"projects": bb_projects})

# ========== API：删除文件 ==========
@data_bp.route('/api/files/delete', methods=['POST'])
def delete_file():
    data = request.json
    project = data.get('project', 'default')
    category = data.get('category', '')
    filename = data.get('filename', '')
    
    filepath = os.path.join(UPLOAD_DIR, project, category, filename)
    if os.path.exists(filepath):
        os.remove(filepath)
        return jsonify({"success": True})
    return jsonify({"error": "文件不存在"}), 404

# ========== 工具函数 ==========
def extract_text_preview(filepath, ext):
    """提取文件文本预览"""
    try:
        ext = ext.lower()
        if ext == '.pdf':
            try:
                import fitz  # PyMuPDF
                doc = fitz.open(filepath)
                text = ''
                for page in doc[:3]:  # 前3页
                    text += page.get_text()
                doc.close()
                return text
            except ImportError:
                return "[需要安装 PyMuPDF 以提取PDF文本]"
        
        elif ext in ['.docx']:
            try:
                from docx import Document
                doc = Document(filepath)
                text = '\n'.join([p.text for p in doc.paragraphs[:20] if p.text.strip()])
                return text
            except ImportError:
                return "[需要安装 python-docx 以提取DOCX文本]"
        
        elif ext in ['.xlsx', '.xls', '.csv']:
            return "[Excel文件 — 请使用\"导入Excel\"功能查看结构化数据]"
        
        return None
    except Exception as e:
        return f"[预览失败: {str(e)}]"

def analyze_excel_structure(sheets_data):
    """分析Excel结构，推荐识别科目表/余额表/序时账等"""
    recos = []
    for name, data in sheets_data.items():
        headers = data.get('headers', [])
        header_text = ' '.join(headers).lower()
        
        if any(k in header_text for k in ['科目', '余额', '期末', '期初']):
            recos.append(f"📊 Sheet「{name}」→ 疑似**科目余额表**")
        elif any(k in header_text for k in ['凭证', '摘要', '借方', '贷方']):
            recos.append(f"📋 Sheet「{name}」→ 疑似**序时账/记账凭证**")
        elif any(k in header_text for k in ['收入', '成本', '利润', '损益']):
            recos.append(f"💰 Sheet「{name}」→ 疑似**利润表/损益表**")
        elif any(k in header_text for k in ['合同', '租赁', '商户']):
            recos.append(f"📄 Sheet「{name}」→ 疑似**合同/租赁台账**")
        elif any(k in header_text for k in ['投标', '报价', '供应商']):
            recos.append(f"🎯 Sheet「{name}」→ 疑似**招投标数据**")
        else:
            recos.append(f"❓ Sheet「{name}」→ 未识别类型（{len(headers)}列，{data['total_rows']}行）")
    
    return recos
