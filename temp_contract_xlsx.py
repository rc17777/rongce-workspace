"""Generate contract summary Excel for 恩阳医养园PPP项目"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ===== Style definitions =====
header_font = Font(name='微软雅黑', bold=True, size=12, color='FFFFFF')
header_fill = PatternFill(start_color='0A1F3F', end_color='0A1F3F', fill_type='solid')
sub_header_fill = PatternFill(start_color='1A5C6E', end_color='1A5C6E', fill_type='solid')
sub_header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
cell_font = Font(name='微软雅黑', size=10)
bold_font = Font(name='微软雅黑', size=10, bold=True)
warn_font = Font(name='微软雅黑', size=10, bold=True, color='FF0000')
title_font = Font(name='微软雅黑', bold=True, size=14, color='0A1F3F')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='center')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
gold_fill = PatternFill(start_color='FFF2E5', end_color='FFF2E5', fill_type='solid')
light_fill = PatternFill(start_color='F5F2EC', end_color='F5F2EC', fill_type='solid')

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_row(ws, row, cols, bold=False, fill=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = bold_font if bold else cell_font
        cell.alignment = wrap_align
        cell.border = thin_border
        if fill:
            cell.fill = fill

def auto_width(ws, min_w=12, max_w=50):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = min_w
        for cell in col:
            if cell.value:
                lines = str(cell.value).split('\n')
                for line in lines:
                    # Approximate: Chinese chars count as 2
                    l = sum(2 if ord(c) > 127 else 1 for c in line)
                    max_len = max(max_len, l)
        ws.column_dimensions[letter].width = min(max_len + 4, max_w)

# =============================================
# Sheet 1: 合同体系概览
# =============================================
ws1 = wb.active
ws1.title = "合同体系概览"

data1 = [
    ["恩阳医养园PPP项目 — 合同体系概览", "", "", "", ""],
    ["", "", "", "", ""],
    ["序号", "文件名称", "签署时间", "性质", "备注"],
    ["1", "恩阳医养园PPP项目协议", "2016-12-27", "原始三方协议",
     "甲方（卫计局）、乙1方（区人民医院）、乙2方（湖南省第五工程有限公司）"],
    ["2", "恩阳医养园PPP项目合同", "2017-06-26", "正式PPP合同",
     "项目公司（巴中市恩阳区医养园项目经营管理有限公司）承继原协议全部权利义务"],
    ["3", "补充合同（一）", "2017-07", "付费公式+绩效考核初建",
     "修改可行性缺口补助公式；融资利率上限设为7.99%；绩效考核体系初建"],
    ["4", "补充合同（二）", "2018", "绩效考核全面升级+股债分离",
     "三份补充中内容最多的：绩效考核体系定版、支付系数表、股债分离约定"],
    ["5", "补充合同（三）", "2021-05", "合作期延长+付费时点调整",
     '合作期20年→23年；建设期≤2年→≤5年；首次付费时间调整为\'交付满一年后\''],
]
ws1.merge_cells('A1:E1')
ws1.cell(row=1, column=1, value=data1[0][0]).font = title_font
ws1.cell(row=1, column=1).alignment = center_align

for i, row in enumerate(data1[2:], start=3):
    for j, val in enumerate(row, start=1):
        ws1.cell(row=i, column=j, value=val)
    if i == 3:
        style_header(ws1, i, 5)
    else:
        style_row(ws1, i, 5, fill=light_fill if i % 2 == 0 else None)

auto_width(ws1)
ws1.column_dimensions['E'].width = 55

# =============================================
# Sheet 2: 合同主体
# =============================================
ws2 = wb.create_sheet("合同主体")

data2 = [
    ["恩阳医养园PPP项目 — 合同主体", "", "", "", ""],
    ["", "", "", "", ""],
    ["角色", "名称", "简称", "说明"],
    ["甲方/实施机构", "巴中市恩阳区卫生和计划生育局\n（2021年更名为巴中市恩阳区卫生健康局）", "卫计局→卫健局",
     "恩阳区人民政府授权，PPP项目实施机构"],
    ["乙1方", "巴中市恩阳区人民医院", "区人民医院", "政府出资代表方"],
    ["乙2方/社会资本", "湖南省第五工程有限公司", "湘五建",
     "公开招标选定的社会资本方，承担全部融资责任"],
    ["乙方/项目公司", "巴中市恩阳区医养园项目经营管理有限公司", "项目公司",
     "乙1方+乙2方共同组建，法定代表人：廖晓\n注册资本金在建设期内不计算建设期利息"],
    ["合同承继", "2017年PPP合同签署后，项目公司（乙方）承继原协议中\n区人民医院和湖南省第五工程有限公司的全部权利义务", "", ""],
]
ws2.merge_cells('A1:D1')
ws2.cell(row=1, column=1, value=data2[0][0]).font = title_font
ws2.cell(row=1, column=1).alignment = center_align

for i, row in enumerate(data2[2:], start=3):
    for j, val in enumerate(row, start=1):
        ws2.cell(row=i, column=j, value=val)
    if i == 3:
        style_header(ws2, i, 4)
    else:
        style_row(ws2, i, 4, bold=(i==7), fill=light_fill if i % 2 == 0 else None)

auto_width(ws2)
ws2.column_dimensions['B'].width = 45

# =============================================
# Sheet 3: 项目基本信息
# =============================================
ws3 = wb.create_sheet("项目基本信息")

data3 = [
    ["恩阳医养园PPP项目 — 项目基本信息", "", ""],
    ["", "", ""],
    ["项目", "内容", "备注"],
    ["项目名称", "恩阳医养园工程PPP项目", ""],
    ["建设地点", "巴中市恩阳区旱谷村石峡子桥侧，临H-13-1地块", "四面邻规划路"],
    ["占地面积", "约223.8亩（总用地180,458㎡，建设用地149,229㎡）", ""],
    ["一期建筑面积", "119,640.13㎡（地上63,844.16㎡ + 地下49,846.73㎡）", "含绿化、路面硬化等"],
    ["二期建筑面积", "68,974.06㎡", "科研教学楼、住院部、康复中心、救灾指挥中心等"],
    ["总投资估算", "5亿元人民币", "最终以审计部门审定为准"],
    ["运作模式", "PPP（BOT：建设-运营-移交）", "通过物有所值评价及财政承受能力论证"],
    ["融资利率基准", "4.9%（2016年五年期以上贷款基准利率）", "后调整为上限7.99%"],
    ["回报机制", "可行性缺口补助（VGF）+ 第三方收入", "社会资本不参与医疗服务盈利"],
    ["第三方收入来源", "食堂、停车场、超市、保洁等行政后勤管理收入", ""],
    ["项目公司注册地址", "恩阳区翊龙公路街11号", "电话：0827-3362218"],
]
ws3.merge_cells('A1:C1')
ws3.cell(row=1, column=1, value=data3[0][0]).font = title_font
ws3.cell(row=1, column=1).alignment = center_align

for i, row in enumerate(data3[2:], start=3):
    for j, val in enumerate(row, start=1):
        ws3.cell(row=i, column=j, value=val)
    if i == 3:
        style_header(ws3, i, 3)
    else:
        style_row(ws3, i, 3, fill=light_fill if i % 2 == 0 else None)

ws3.column_dimensions['A'].width = 18
ws3.column_dimensions['B'].width = 55
ws3.column_dimensions['C'].width = 35

# =============================================
# Sheet 4: 合作期限变更
# =============================================
ws4 = wb.create_sheet("合作期限")

data4 = [
    ["恩阳医养园PPP项目 — 合作期限（历经变更）", "", "", "", "", ""],
    ["", "", "", "", "", ""],
    ["版本", "合同依据", "合作期", "建设期", "运营期", "关键说明"],
    ["v1.0", "原PPP合同（2017）\n原PPP协议（2016）", "20年", "≤2年（24个月）",
     "18年", "建设期超2年则运营期相应减少；\n运营期遇延期不顺延"],
    ["v2.0（现行）", "2021年补充合同（三）", "23年", "≤5年（60个月）",
     "18年（保持不变）", "★核心变化：\n①建设期从2年延至5年；\n②运营期如遇延期则相应顺延；\n③原因：设计变更/三甲医院复杂性/土地/疫情"],
]
ws4.merge_cells('A1:F1')
ws4.cell(row=1, column=1, value=data4[0][0]).font = title_font
ws4.cell(row=1, column=1).alignment = center_align

for i, row in enumerate(data4[2:], start=3):
    for j, val in enumerate(row, start=1):
        ws4.cell(row=i, column=j, value=val)
    if i == 3:
        style_header(ws4, i, 6)
    else:
        style_row(ws4, i, 6)

auto_width(ws4)

# =============================================
# Sheet 5: 回报机制详解
# =============================================
ws5 = wb.create_sheet("回报机制")

data5 = [
    ["恩阳医养园PPP项目 — 回报机制详解", "", ""],
    ["", "", ""],
    ["项目", "内容", "依据/备注"],
    ["回报模式", "可行性缺口补助（VGF）+ 第三方收入", "PPP合同"],
    ["社会资本收入来源", "仅限医院行政后勤管理（食堂/停车场/超市/保洁），不参与医疗服务盈利", "PPP合同第4.1条"],
    ["政府付费公式\n（2017版）", "A = P×k×(1+k)^n / [(1+k)^n-1]\nA=运营期内各年政府运营补贴\nP=项目资本金\nk=合理利润率（中标年回报率）\nn=财政运营补贴周期", "2017补充合同"],
    ["政府付费公式\n（2018定版）", "年度政府绩效考核付费 = R1×建安工程费×70% + R2×(运营维护费+建安工程费×30%)\n\nR1=建设期绩效考核调整系数\nR2=运营期绩效考核调整系数", "2018补充合同"],
    ["融资利率", "基准：4.9%（2016年五年期以上贷款基准利率）\n上限：7.99%", "2017补充合同设定上限"],
    ["付费方式", "等额本息法，18年，一年一次", "PPP合同"],
    ['首次付费时点\n（2021调整后）', '工程竣工综合验收质量合格并交付使用满一年后30日内支付首次；\n此后每年同月支付', '2021补充合同（三）\n原为\u201c工程完成审计后三个月内\u201d'],
    ["股债分离\n（2018约定）", "融资部分（非资本金部分）按时按金额据实支付给项目公司；\n政府支付总额上限=按7.99%/18年等额本息计算的还贷总额；\n★实际融资成本低于上限时，社会资本不能赚取差价", "2018补充合同"],
    ["建设期利息", "按每月施工产值计算，纳入建安投资分年度等额支付", "PPP合同"],
    ["注册资本金利息", "建设期内不计息", "PPP合同"],
]
ws5.merge_cells('A1:C1')
ws5.cell(row=1, column=1, value=data5[0][0]).font = title_font
ws5.cell(row=1, column=1).alignment = center_align

for i, row in enumerate(data5[2:], start=3):
    for j, val in enumerate(row, start=1):
        ws5.cell(row=i, column=j, value=val)
    if i == 3:
        style_header(ws5, i, 3)
    else:
        style_row(ws5, i, 3, fill=light_fill if i % 2 == 0 else None)
        if "★" in str(ws5.cell(row=i, column=2).value or ""):
            ws5.cell(row=i, column=2).font = warn_font

ws5.column_dimensions['A'].width = 20
ws5.column_dimensions['B'].width = 65
ws5.column_dimensions['C'].width = 30

# =============================================
# Sheet 6: 绩效考核体系
# =============================================
ws6 = wb.create_sheet("绩效考核体系")

# Part A: 建设期
ws6.merge_cells('A1:F1')
ws6.cell(row=1, column=1, value="恩阳医养园PPP项目 — 绩效考核体系").font = title_font
ws6.cell(row=1, column=1).alignment = center_align

ws6.merge_cells('A3:F3')
ws6.cell(row=3, column=1, value="一、建设期绩效考核（竣工验收后一次性考核）").font = Font(name='微软雅黑', bold=True, size=11, color='1A5C6E')

build_data = [
    ["考核指标", "权重", "考核依据/标准"],
    ["工程质量", "70分",
     "①项目建设符合法律和国家行业标准，不合格须整改\n②工程达合格得满分，不达标按分项工程金额占比扣分\n③验收前甲方已使用的，视为合格得满分"],
    ["工程进度", "15分",
     "①开工日以监理开工令为准，按约定期限完工得满分\n②延期按实际拖期天数占约定总工期比重扣分\n③因政府另实施项目影响的，不扣分"],
    ["工程安全", "10分",
     "建设期内未发生《生产安全事故报告和调查处理条例》规定的较大安全责任事故得满分\n每发生一次扣1分，直至扣完"],
    ["环境保护", "5分",
     "符合国家行业环保要求，未发生《国家突发环境事件应急预案》中的较大突发环境事件得满分\n每发生一次扣1分，直至扣完"],
]
for i, row in enumerate(build_data, start=4):
    for j, val in enumerate(row, start=1):
        ws6.cell(row=i, column=j, value=val)
    if i == 4:
        style_header(ws6, i, 3)
    else:
        style_row(ws6, i, 3)

# Part B: 运营期
ws6.merge_cells('A11:F11')
ws6.cell(row=11, column=1, value="二、运营期绩效考核（每年一次，运营期报告提交后5个工作日内考核，7个工作日内完成）").font = Font(name='微软雅黑', bold=True, size=11, color='1A5C6E')

opr_data = [
    ["考核指标", "权重", "考核依据/标准"],
    ["设施维护", "60分",
     "①主体结构正常使用(30分) ②照明系统完好(10分)\n③供排水通畅(10分) ④标牌标识规范(10分)\n非不可抗力因素造成的设施不能使用不扣分"],
    ["经营项目运营", "20分",
     "①超市/停车场/食堂正常经营(15分,每项5分)\n②食品安全卫生(3分,发生较大事故不得分)\n③服务投诉(2分,每次扣0.1分)"],
    ["安全管理", "15分",
     "按规范管理运营，未发生较大安全事故得满分\n发生较大事故且责任主体为乙方，每出现一次扣3分"],
    ["维护运营资料", "5分",
     "PPP项目维护运营主要资料完善得满分\n必备运营资料每缺一项扣0.1分"],
]
for i, row in enumerate(opr_data, start=12):
    for j, val in enumerate(row, start=1):
        ws6.cell(row=i, column=j, value=val)
    if i == 12:
        style_header(ws6, i, 3)
    else:
        style_row(ws6, i, 3)

# Part C: 支付系数
ws6.merge_cells('A18:F18')
ws6.cell(row=18, column=1, value="三、考核得分与支付系数对应表").font = Font(name='微软雅黑', bold=True, size=11, color='1A5C6E')

coeff_header = ["等级", "得分区间", "建设期R1", "运营期R2", "处理方式", ""]
coeff_data = [
    ["优", "≥80分", "100%", "100%", "全额支付", ""],
    ["良", "70-79分", "95%起，每增1分+0.5%", "80%起，每增1分+2%", "按系数支付", ""],
    ["合格", "60-69分", "90%起，每增1分+0.5%", "60%起，每增1分+2%", "按系数支付", ""],
    ["不合格", "<60分", "须整改达标后重新考核再支付", "整改后按40%支付", "★建设期不合格须整改至达标", ""],
]
for i, row in enumerate([coeff_header]+coeff_data, start=19):
    for j, val in enumerate(row, start=1):
        ws6.cell(row=i, column=j, value=val)
    if i == 19:
        style_header(ws6, i, 6)
    else:
        style_row(ws6, i, 6)

ws6.column_dimensions['A'].width = 18
ws6.column_dimensions['B'].width = 30
ws6.column_dimensions['C'].width = 25
ws6.column_dimensions['D'].width = 25
ws6.column_dimensions['E'].width = 22
ws6.column_dimensions['F'].width = 15

# =============================================
# Sheet 7: 各方权利义务
# =============================================
ws7 = wb.create_sheet("各方权利义务")

ws7.merge_cells('A1:D1')
ws7.cell(row=1, column=1, value="恩阳医养园PPP项目 — 各方核心权利义务").font = title_font
ws7.cell(row=1, column=1).alignment = center_align

ws7.merge_cells('A3:D3')
ws7.cell(row=3, column=1, value="甲方（巴中市恩阳区卫生健康局）").font = Font(name='微软雅黑', bold=True, size=11, color='0A1F3F')

rights_a = [
    ["权利", "义务"],
    ["①按合同提取建设期履约保函项下款项", "①协助乙方及时获取施工许可等相关批准"],
    ["②对投资、建设、运营、维护及移交全程实时监管", "②协调临时水、电、通讯线路接通至指定地点"],
    ["③定期评估并向社会公示", "③协调城市供水、排水、燃气、供电、通信、消防等配套设施"],
    ["④委托审计机构对建设投资进行专项审计", "④负责征地拆迁工作"],
    ["⑤乙方严重违约时终止合同并收回经营权\n（转让/出租/质押经营权；重大质量安全事故；\n被注销关停；违反承诺情节严重）", "⑤及时足额支付可行性缺口补助\n★将服务费纳入跨年度财政预算并提请人大决议"],
    ["", "⑥协调项目场地周边单位关系"],
    ["", "⑦对非可归责于乙方的事项负责协调或监管"],
]
for i, row in enumerate(rights_a, start=4):
    for j, val in enumerate(row, start=1):
        ws7.cell(row=i, column=j, value=val)
    if i == 4:
        style_header(ws7, i, 4)
    else:
        style_row(ws7, i, 4, bold=(i==4))

row_a_end = 4 + len(rights_a)

ws7.merge_cells(f'A{row_a_end+1}:D{row_a_end+1}')
ws7.cell(row=row_a_end+1, column=1, value="乙方（项目公司：巴中市恩阳区医养园项目经营管理有限公司）").font = Font(name='微软雅黑', bold=True, size=11, color='0A1F3F')

rights_b = [
    ["权利", "义务"],
    ["①享有投资、建设、运营和维护项目的独占经营权", "①承担项目投资、建设、运营及维护的全部费用和风险"],
    ["②要求甲方按约支付可行性缺口补助\n（乙1方区人民医院的投资不享有回报）", "②按规划完成相应投资建设"],
    ["③可归责于第三方导致履约不能的，\n甲方酌情考虑绩效考核指标达成率", "③运营期内持续、安全、稳定提供服务"],
    ["④可引入持牌金融机构融资", "④接受建设期及运营期全过程监管并配合审计"],
    ["⑤合作期满在同等条件下享有优先续约权", "⑤及时申请建设工程所需的各种政府批准"],
    ["", "⑥配合依附于本项目的管线、杆线等维修维护"],
    ["", "⑦协助甲方申请国家专项资金"],
    ["", "★社会资本（湘五建）承担资金缺口的全部融资责任"],
]
for i, row in enumerate(rights_b, start=row_a_end+2):
    for j, val in enumerate(row, start=1):
        ws7.cell(row=i, column=j, value=val)
    if i == row_a_end+2:
        style_header(ws7, i, 4)
    else:
        style_row(ws7, i, 4, bold=(i==row_a_end+2))

ws7.column_dimensions['A'].width = 40
ws7.column_dimensions['B'].width = 40
ws7.column_dimensions['C'].width = 15
ws7.column_dimensions['D'].width = 15

# =============================================
# Sheet 8: 补充合同变更汇总
# =============================================
ws8 = wb.create_sheet("补充合同变更汇总")

data8 = [
    ["恩阳医养园PPP项目 — 三份补充合同核心变更汇总", "", "", ""],
    ["", "", "", ""],
    ["补充合同", "变更事项", "原条款", "变更后条款"],
    ["2017补充合同（一）\n（2017年7月）",
     "①付费公式调整",
     "原第35条",
     "A = P×k×(1+k)^n/[(1+k)^n-1]\n明确等额本息公式"],
    ["", "②融资利率上限", "4.9%（基准利率）", "上限7.99%\n实际融资年利率不得超过7.99%"],
    ["", "③绩效考核体系初建", "附件7/8（原版）", "建设期+运营期考核指标体系初建\n建设指标70%+运营指标30%权重"],
    ["2018补充合同（二）\n（2018年）",
     "①绩效考核全面升级", "2017年初版考核体系",
     "建设期4项指标（质量/进度/安全/环保）\n运营期4项指标（设施/经营/安全/资料）\n明确得分-支付系数对应表"],
    ["", "②付费公式定版", "2017版公式",
     "年度付费=R1×建安费×70%+R2×(运维费+建安费×30%)\n建安费70%挂钩建设考核，30%挂钩运营考核"],
    ["", "③股债分离", "无",
     "融资部分按实际合同据实支付\n总额上限=7.99%/18年等额本息\n★实际成本低于上限，社会资本不赚差价"],
    ["", "④优先效力", "无",
     "补充协议与前期协议不一致的，以补充协议为准"],
    ["2021补充合同（三）\n（2021年5月）",
     "①合作期延长", "合作期20年\n建设期≤2年（24个月）",
     "合作期23年\n建设期≤5年（60个月）"],
    ["", "②运营期顺延", "建设期超2年运营期相应减少\n运营期不顺延",
     "运营期保持18年不变\n如遇工期延长则运营期相应顺延"],
    ["", "③首次付费时点", "工程完成审计后三个月内支付首次\n年初30日内支付（2017版）",
     "工程竣工综合验收合格并交付使用\n满一年后30日内支付首次"],
    ["", "④甲方名称变更", "巴中市恩阳区卫生和计划生育局",
     "巴中市恩阳区卫生健康局\n（机构改革更名）"],
    ["", "⑤延期原因", "—",
     "设计变更、三甲医院复杂性、\n土地原因、疫情影响"],
]
ws8.merge_cells('A1:D1')
ws8.cell(row=1, column=1, value=data8[0][0]).font = title_font
ws8.cell(row=1, column=1).alignment = center_align

for i, row in enumerate(data8[2:], start=3):
    for j, val in enumerate(row, start=1):
        ws8.cell(row=i, column=j, value=val)
    if i == 3:
        style_header(ws8, i, 4)
    else:
        is_change_row = (ws8.cell(row=i, column=1).value and "补充合同" in str(ws8.cell(row=i, column=1).value))
        style_row(ws8, i, 4, bold=is_change_row, fill=gold_fill if is_change_row else (light_fill if i % 2 == 0 else None))
        if "★" in str(ws8.cell(row=i, column=4).value or ""):
            ws8.cell(row=i, column=4).font = warn_font

ws8.column_dimensions['A'].width = 22
ws8.column_dimensions['B'].width = 22
ws8.column_dimensions['C'].width = 35
ws8.column_dimensions['D'].width = 45

# =============================================
# Sheet 9: 审计关注要点
# =============================================
ws9 = wb.create_sheet("审计关注要点")

data9 = [
    ["恩阳医养园PPP项目 — 审计关注要点", "", ""],
    ["", "", ""],
    ["序号", "关注要点", "说明"],
    ["1", "总投资审计确认", "5亿为估算值，最终以审计审定为准。需核实实际工程造价是否超估算，超估算部分是否合规"],
    ["2", "融资成本是否合规", "融资利率上限7.99%为基准利率4.9%的1.63倍。需核实实际融资利率是否符合合同约定，有无超过上限"],
    ["3", "绩效考核执行情况", "建设期和运营期考核是否按2018版标准实际执行，支付金额是否与考核结果严格挂钩"],
    ["4", "股债分离执行", "融资还贷政府支付部分是否据实核查，社会资本有无赚取利差（实际融资成本低于等额本息计算值）"],
    ["5", "工期延误责任分摊", "合作期从20年延至23年，延长3年的原因是否合理，延误责任是否合理分摊给政府方和社会资本方"],
    ["6", "首次付费时点及金额", '2021年改为\u201c交付满一年后\u201d支付首次，实际是否已开始付费、付费金额是否合规、有无滞纳金'],
    ["7", "第三方收入核算", "食堂/停车场/超市/保洁收入是否真实完整纳入计算，有无少报虚报从而增加政府补助"],
    ["8", "经营权担保合规性", "项目公司有无将PPP合同权益（可行性缺口补助/保险受益权）用于非本项目融资的担保"],
    ["9", "项目公司资金到位", "注册资本金是否按期到位，建设资金是否按进度足额到位，有无抽逃资金"],
    ["10", "人大决议落实", "区人民政府是否将本项目服务费纳入跨年度财政预算并提请人大决议，决议是否在有效期内"],
]
ws9.merge_cells('A1:C1')
ws9.cell(row=1, column=1, value=data9[0][0]).font = title_font
ws9.cell(row=1, column=1).alignment = center_align

for i, row in enumerate(data9[2:], start=3):
    for j, val in enumerate(row, start=1):
        ws9.cell(row=i, column=j, value=val)
    if i == 3:
        style_header(ws9, i, 3)
    else:
        style_row(ws9, i, 3, fill=light_fill if i % 2 == 0 else None)

ws9.column_dimensions['A'].width = 8
ws9.column_dimensions['B'].width = 25
ws9.column_dimensions['C'].width = 60

# ===== Save =====
output_path = r"C:\Users\scrccpa\Desktop\恩阳医养园PPP项目合同要点提炼.xlsx"
wb.save(output_path)
print(f"Saved to: {output_path}")
print(f"Sheets: {wb.sheetnames}")
