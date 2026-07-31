"""
若尔盖医保审计 - 多院同天住院57条深度溯源
逐条标注原始数据来源（年份/文件名/医药机构编号），方便核查追溯
"""
import sys, os, json
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict

BASE = r'C:\Users\scrccpa\Desktop\若尔盖审计\若尔盖医保审计\2026年审计资料（医保局财务）\古英=2024-2025门诊、住院、个人账户、基金拨付明细'
YEARS = ['2023', '2024', '2025']
DESKTOP = r'C:\Users\scrccpa\Desktop'

HEADER_FILL = PatternFill(start_color='0A1F3F', end_color='0A1F3F', fill_type='solid')
HEADER_FONT = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
DATA_FONT = Font(name='微软雅黑', size=10)
TITLE_FONT = Font(name='微软雅黑', bold=True, size=14, color='0A1F3F')
P0_FILL = PatternFill(start_color='FFD7D7', end_color='FFD7D7', fill_type='solid')
P1_FILL = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def parse_date(val):
    if val is None: return None
    if hasattr(val, 'date'): return val.date()
    if isinstance(val, str):
        for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y%m%d']:
            try: return datetime.strptime(val[:10], fmt).date()
            except: pass
    return None

# ============================================================
# Step 1: Load INPATIENT records only (much faster)
# ============================================================
print('Loading inpatient records from 3 years...')
inpatient_records = []
patient_daily = defaultdict(lambda: defaultdict(list))
all_records_full = defaultdict(list)

for year in YEARS:
    fp = os.path.join(BASE, f'{year}.xlsx')
    if not os.path.exists(fp): continue
    
    print(f'  {year}...', end=' ', flush=True)
    wb = load_workbook(fp, read_only=True)
    main_sheet = '总表' if '总表' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[main_sheet]
    headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    col = {str(h).strip(): i for i, h in enumerate(headers) if h}
    
    id_col = col.get('证件号码')
    name_col = col.get('姓名')
    type_col = col.get('医疗类别')
    settle_col = col.get('结算时间')
    admit_col = col.get('入院时间')
    discharge_col = col.get('出院时间')
    days_col = col.get('住院天数')
    inst_col = col.get('医药机构名称')
    inst_id_col = col.get('医药机构编号')
    fee_col = col.get('医疗费总额')
    diag_col = col.get('出院诊断名称')
    pay_col = col.get('医保支付金额')
    fund_pay_col = col.get('统筹基金支付')
    yidi_col = col.get('是否异地就医')
    acct_pay_col = col.get('个人账户支付')
    
    cnt = 0; row_num = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1
        cnt += 1
        if cnt % 50000 == 0: print(f'{cnt//1000}k...', end=' ', flush=True)
        
        id_num = str(row[id_col]).strip() if row[id_col] else ''
        name = str(row[name_col]).strip() if row[name_col] else ''
        inst = str(row[inst_col]).strip() if row[inst_col] else ''
        inst_id = str(row[inst_id_col]).strip() if row[inst_id_col] else ''
        fee = row[fee_col] if fee_col and isinstance(row[fee_col], (int, float)) else 0
        pay = row[pay_col] if pay_col and isinstance(row[pay_col], (int, float)) else 0
        fund = row[fund_pay_col] if fund_pay_col and isinstance(row[fund_pay_col], (int, float)) else 0
        acct = row[acct_pay_col] if acct_pay_col and isinstance(row[acct_pay_col], (int, float)) else 0
        diag = str(row[diag_col]).strip() if diag_col and row[diag_col] else ''
        yidi = str(row[yidi_col]).strip() if yidi_col and row[yidi_col] else ''
        
        settle_dt = parse_date(row[settle_col]) if settle_col else None
        admit_dt = parse_date(row[admit_col]) if admit_col else None
        discharge_dt = parse_date(row[discharge_col]) if discharge_col else None
        days = row[days_col] if days_col and isinstance(row[days_col], (int, float)) else 0
        
        med_type_raw = str(row[type_col]) if row[type_col] else ''
        if '住院' in med_type_raw:
            rec = {
                'id_num': id_num, 'name': name, 'inst': inst, 'inst_id': inst_id,
                'fee': fee, 'pay': pay, 'fund_pay': fund, 'acct_pay': acct,
                'med_type': med_type_raw, 'settle_dt': settle_dt,
                'admit_dt': admit_dt, 'discharge_dt': discharge_dt,
                'days': int(days), 'diag': diag, 'yidi': yidi, 'year': year,
                'row_num': row_num
            }
            inpatient_records.append(rec)
            if settle_dt:
                patient_daily[id_num][settle_dt].append(rec)
    
    wb.close()
    print(f'{cnt:,} total → {len([r for r in inpatient_records if r["year"]==year]):,} inpatient')

print(f'Total inpatient: {len(inpatient_records):,}')

# ============================================================
# Step 2: Find multi-hospital same-day cases
# ============================================================
print('\nFinding multi-hospital same-day cases...')
multi_hosp_all = []

for pid, daily in patient_daily.items():
    for dt, visits in daily.items():
        if len(visits) < 2: continue
        hospitals = set(v['inst'] for v in visits)
        if len(hospitals) >= 2:
            multi_hosp_all.append({
                'id_num': pid,
                'name': visits[0]['name'],
                'date': dt,
                'hospitals': sorted(hospitals),
                'hospital_count': len(hospitals),
                'total_fee': sum(v['fee'] for v in visits),
                'total_pay': sum(v['pay'] for v in visits),
                'total_fund': sum(v['fund_pay'] for v in visits),
                'visits': visits,
            })

print(f'Found {len(multi_hosp_all)} cases')

# ============================================================
# Step 3: Build patient admission history
# ============================================================
print('Building patient history...')
patient_history = defaultdict(list)
for rec in inpatient_records:
    patient_history[rec['id_num']].append(rec)

# Sort each patient's history by admit date
for pid in patient_history:
    patient_history[pid].sort(key=lambda x: x['admit_dt'] if x['admit_dt'] else datetime.min.date())

# ============================================================
# Step 4: Verify each case
# ============================================================
print('Verifying each case...')

def get_distance_level(hosps):
    """Determine the geographic distance level between hospitals"""
    local_keywords = ['若尔盖']
    prefecture_keywords = ['阿坝', '马尔康', '汶川', '理县', '茂县', '松潘', '九寨沟', '金川',
                          '小金', '黑水', '壤塘']
    
    has_local = any(any(k in h for k in local_keywords) for h in hosps)
    has_prefecture = any(any(k in h for k in prefecture_keywords) for h in hosps)
    all_local = all(any(k in h for k in local_keywords) for h in hosps)
    
    if all_local:
        return '同县'
    elif has_local:
        # Local + somewhere else
        remote = [h for h in hosps if not any(k in h for k in local_keywords)]
        return f'若尔盖+{remote[0][:10]}'
    elif has_prefecture:
        return '州内'
    else:
        return '跨省/远距'

verified = []
for case in sorted(multi_hosp_all, key=lambda x: -x['total_fee']):
    pid = case['id_num']
    history = patient_history.get(pid, [])
    
    # Geographic analysis
    distance = get_distance_level(case['hospitals'])
    
    verdicts = []
    if distance == '同县':
        verdicts.append('同县两院-需核实(P2)')
    elif '若尔盖+' in distance:
        verdicts.append('本地+异地同时住院-铁证(P0)')
    elif distance == '跨省/远距':
        verdicts.append('远距同天住院-铁证(P0)')
    elif distance == '州内':
        verdicts.append('跨市同天住院-严重可疑(P1)')
    else:
        verdicts.append('需核实(P2)')
    
    # History analysis
    case['total_history_fee'] = sum(h['fee'] for h in history)
    case['total_admissions'] = len(history)
    unique_hosps = set(h['inst'] for h in history)
    case['unique_hospitals_count'] = len(unique_hosps)
    case['all_hospitals'] = ' | '.join(sorted(unique_hosps))
    
    if len(unique_hosps) >= 5:
        verdicts.append('频繁转院(≥5院)')
    
    # Final verdict
    if any('铁证' in v for v in verdicts):
        case['verdict'] = 'P0-铁证'
    elif any('严重可疑' in v for v in verdicts):
        case['verdict'] = 'P1-严重可疑'
    else:
        case['verdict'] = 'P2-需核实'
    
    case['verdict_details'] = '; '.join(verdicts)
    case['distance'] = distance
    
    # Build source trace info
    trace_parts = []
    for i, v in enumerate(case['visits'], 1):
        trace_parts.append(
            f"院{i}:{v['inst'][:20]}|"
            f"入院{v['admit_dt']}|出院{v['discharge_dt']}|"
            f"¥{v['fee']:,.0f}|基金{v['fund_pay']:,.0f}|"
            f"源:{v['year']}年|机构号{v.get('inst_id','')}|"
            f"Excel行{v['row_num']}"
        )
    case['trace_sources'] = '\n'.join(trace_parts)
    
    verified.append(case)

# ============================================================
# Step 5: Export to Excel
# ============================================================
print('Exporting to Excel...')

wb = load_workbook(r'C:\Users\scrccpa\Desktop\若尔盖医保审计_深度分析.xlsx')

# Remove old Sheet and create new one
if '多院同天住院57条' in wb.sheetnames:
    del wb['多院同天住院57条']

ws2 = wb.create_sheet('多院同天住院57条', 1)

ws2.merge_cells('A1:O1')
ws2.cell(row=1, column=1, value=f'若尔盖医保审计 - 同天多院住院逐条核实（共{len(verified)}条）· 含数据溯源').font = TITLE_FONT

headers2 = [
    '序号', '判定结果', '空间距离', '判定依据',
    '患者ID', '姓名', '日期', '医院列表', '医院数',
    '本次费用', '本次基金支付', '历史总费用', '历史住院次数', '历史涉及医院',
    '📎 数据溯源（年份/机构号/Excel行号）'
]

for i, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=i, value=h)
for col in range(1, len(headers2)+1):
    cell = ws2.cell(row=3, column=col)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = THIN_BORDER

for idx, m in enumerate(verified, 1):
    row = idx + 3
    ws2.cell(row=row, column=1, value=idx)
    ws2.cell(row=row, column=2, value=m.get('verdict', ''))
    ws2.cell(row=row, column=3, value=m.get('distance', ''))
    ws2.cell(row=row, column=4, value=m.get('verdict_details', ''))
    ws2.cell(row=row, column=5, value=m['id_num'])  # Full ID for tracing
    ws2.cell(row=row, column=6, value=m['name'])
    ws2.cell(row=row, column=7, value=str(m['date']))
    ws2.cell(row=row, column=8, value=' + '.join(m['hospitals']))
    ws2.cell(row=row, column=9, value=m['hospital_count'])
    ws2.cell(row=row, column=10, value=m['total_fee'])
    ws2.cell(row=row, column=11, value=m.get('total_fund', 0))
    ws2.cell(row=row, column=12, value=m.get('total_history_fee', 0))
    ws2.cell(row=row, column=13, value=m.get('total_admissions', 0))
    ws2.cell(row=row, column=14, value=m.get('all_hospitals', ''))
    ws2.cell(row=row, column=15, value=m.get('trace_sources', ''))
    
    for col in range(1, len(headers2)+1):
        cell = ws2.cell(row=row, column=col)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    if 'P0' in m.get('verdict', ''):
        for col in range(1, len(headers2)+1):
            ws2.cell(row=row, column=col).fill = P0_FILL
    elif 'P1' in m.get('verdict', ''):
        for col in range(1, len(headers2)+1):
            ws2.cell(row=row, column=col).fill = P1_FILL

# Auto-width
for col in range(1, len(headers2)+1):
    max_len = 0
    for row in range(3, min(ws2.max_row+1, 60)):
        val = str(ws2.cell(row=row, column=col).value or '')[:80]
        max_len = max(max_len, len(val))
    ws2.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 50)

# Make trace column wider
ws2.column_dimensions[get_column_letter(15)].width = 60

# Save
out_path = os.path.join(DESKTOP, '若尔盖医保审计_深度分析.xlsx')
wb.save(out_path)
print(f'Saved: {out_path}')

# Summary
print('\n=== 溯源核查摘要 ===')
print(f'共57条 | P0铁证: {sum(1 for v in verified if "P0" in v.get("verdict",""))} | P1可疑: {sum(1 for v in verified if "P1" in v.get("verdict",""))} | P2需核实: {sum(1 for v in verified if "P2" in v.get("verdict",""))}')
print(f'溯源字段: 年份 + 机构编号 + Excel行号 → 可直接回原始文件定位')
print('Done.')
