"""
若尔盖医保审计 - 数据分析模型 v1.0
模型1: 分解住院 | 模型2: 虚假住院时空碰撞 | 模型3: 定点机构异常
"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from datetime import datetime, timedelta
from collections import defaultdict
import json

BASE = r'C:\Users\scrccpa\Desktop\若尔盖审计\若尔盖医保审计\2026年审计资料（医保局财务）\古英=2024-2025门诊、住院、个人账户、基金拨付明细'
YEARS = ['2023', '2024', '2025']
OUT = r'C:\Users\scrccpa\.openclaw\workspace\audit-blackboard\projects\若尔盖医保资金审计\findings'

os.makedirs(OUT, exist_ok=True)

def parse_date(val):
    """Parse date from various formats."""
    if val is None: return None
    if hasattr(val, 'date'): return val.date()
    if isinstance(val, str):
        for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y%m%d']:
            try: return datetime.strptime(val[:10], fmt).date()
            except: pass
    return None

# ============================================================
# Phase 1: Parse all 3 years into patient-level records
# ============================================================
print('='*60)
print('Phase 1: Loading 3-year settlement data')
print('='*60)

# We'll collect: 住院 records, 门诊 records, and all records
inpatient_records = []  # (id_num, name, admit_date, discharge_date, hospital, fee, year, diag)
outpatient_records = [] # (id_num, name, visit_date, hospital, fee, year)
all_records = []        # Full records for institution analysis

for year in YEARS:
    fp = os.path.join(BASE, f'{year}.xlsx')
    if not os.path.exists(fp): continue
    
    print(f'\nLoading {year}...', end=' ', flush=True)
    wb = load_workbook(fp, read_only=True)
    # 2023 uses 'Sheet1', 2024-2025 use '总表'
    main_sheet = '总表' if '总表' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[main_sheet]
    headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    col = {str(h).strip(): i for i, h in enumerate(headers) if h}
    
    # Map columns
    id_col = col.get('证件号码')
    name_col = col.get('姓名')
    type_col = col.get('医疗类别')
    settle_col = col.get('结算时间')
    admit_col = col.get('入院时间')
    discharge_col = col.get('出院时间')
    days_col = col.get('住院天数')
    inst_col = col.get('医药机构名称')
    inst_id_col = col.get('医药机构编号')
    fee_col = col.get('医疗费总额')
    diag_col = col.get('出院诊断名称')
    yidi_col = col.get('是否异地就医')
    pay_col = col.get('医保支付金额')
    drug_col = col.get('药品费')
    treat_col = col.get('诊疗费')
    mat_col = col.get('耗材费')
    
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        count += 1
        if count % 50000 == 0: print(f'{count//1000}k...', end=' ', flush=True)
        
        id_num = str(row[id_col]).strip() if row[id_col] else ''
        name = str(row[name_col]).strip() if row[name_col] else ''
        inst = str(row[inst_col]).strip() if row[inst_col] else ''
        inst_id = str(row[inst_id_col]).strip() if row[inst_id_col] else ''
        fee = row[fee_col] if fee_col and isinstance(row[fee_col], (int, float)) else 0
        pay = row[pay_col] if pay_col and isinstance(row[pay_col], (int, float)) else 0
        diag = str(row[diag_col]).strip() if diag_col and row[diag_col] else ''
        yidi = str(row[yidi_col]).strip() if yidi_col and row[yidi_col] else ''
        
        settle_dt = parse_date(row[settle_col]) if settle_col else None
        admit_dt = parse_date(row[admit_col]) if admit_col else None
        discharge_dt = parse_date(row[discharge_col]) if discharge_col else None
        days = row[days_col] if days_col and isinstance(row[days_col], (int, float)) else 0
        
        med_type_raw = str(row[type_col]) if row[type_col] else ''
        med_type = med_type_raw.split('|')[-1].strip() if '|' in med_type_raw else med_type_raw
        
        drug_fee = row[drug_col] if drug_col and isinstance(row[drug_col], (int, float)) else 0
        treat_fee = row[treat_col] if treat_col and isinstance(row[treat_col], (int, float)) else 0
        mat_fee = row[mat_col] if mat_col and isinstance(row[mat_col], (int, float)) else 0
        
        rec = (id_num, name, inst, inst_id, fee, pay, med_type, settle_dt, 
               admit_dt, discharge_dt, int(days), diag, yidi, drug_fee, treat_fee, mat_fee, year)
        all_records.append(rec)
        
        if '住院' in med_type:
            inpatient_records.append((id_num, name, admit_dt, discharge_dt, inst, fee, pay, year, diag, days))
        elif '门诊' in med_type:
            outpatient_records.append((id_num, name, settle_dt, inst, fee, pay, year))
    
    wb.close()
    print(f'{count:,} records')

print(f'\nTotal: {len(all_records):,} records')
print(f'  住院: {len(inpatient_records):,}')
print(f'  门诊: {len(outpatient_records):,}')

# ============================================================
# Model 1: 分解住院 (Split Hospitalization)
# ============================================================
print('\n' + '='*60)
print('Model 1: 分解住院识别 (7天内同患者同院再入院)')
print('='*60)

# Group inpatient records by patient ID
patient_admissions = defaultdict(list)
for rec in inpatient_records:
    id_num, name, admit, discharge, hospital, fee, pay, year, diag, days = rec
    if admit and discharge:
        patient_admissions[id_num].append(rec)

split_hosp_findings = []
for pid, admissions in patient_admissions.items():
    if len(admissions) < 2:
        continue
    
    # Sort by admission date
    admissions.sort(key=lambda x: x[2] if x[2] else datetime.min.date())
    
    for i in range(len(admissions)):
        for j in range(i+1, len(admissions)):
            a1 = admissions[i]
            a2 = admissions[j]
            
            # Same hospital?
            if a1[4] != a2[4]:
                continue
            
            # Gap between discharge1 and admission2
            if a1[3] and a2[2]:
                gap = (a2[2] - a1[3]).days
                if 0 <= gap <= 7:
                    split_hosp_findings.append({
                        'patient_id': pid[-4:],  # last 4 digits for privacy
                        'name': a1[1],
                        'hospital': a1[4],
                        'admit1': str(a1[2]), 'discharge1': str(a1[3]),
                        'admit2': str(a2[2]), 'discharge2': str(a2[3]),
                        'gap_days': gap,
                        'fee1': a1[5], 'fee2': a2[5],
                        'total_fee': a1[5] + a2[5],
                        'year1': a1[7], 'year2': a2[7],
                        'diag1': a1[8][:30], 'diag2': a2[8][:30],
                    })

# Deduplicate by patient-hospital pair
seen = set()
unique_findings = []
for f in sorted(split_hosp_findings, key=lambda x: -x['total_fee']):
    key = (f['patient_id'], f['hospital'], f['admit1'])
    if key not in seen:
        seen.add(key)
        unique_findings.append(f)

print(f'疑似分解住院: {len(unique_findings)} 组')
total_amount = sum(f['total_fee'] for f in unique_findings)
print(f'涉及金额: ¥{total_amount:,.0f}')

if unique_findings:
    print('\nTop 10 (按金额排序):')
    print(f'{"患者":<6} {"姓名":<8} {"医院":<25} {"间隔":>4} {"金额1":>10} {"金额2":>10} {"合计":>10}')
    for f in unique_findings[:10]:
        print(f'{f["patient_id"]:<6} {f["name"]:<8} {f["hospital"][:25]:<25} {f["gap_days"]:>3}d ¥{f["fee1"]:>9,.0f} ¥{f["fee2"]:>9,.0f} ¥{f["total_fee"]:>9,.0f}')

# ============================================================
# Model 2: 虚假住院·时空碰撞
# ============================================================
print('\n' + '='*60)
print('Model 2: 虚假住院·时空碰撞')
print('='*60)

# 2a: 住院期间有门诊记录
# Build index: patient -> {date: [(type, hospital)]}
patient_daily = defaultdict(lambda: defaultdict(list))
for rec in all_records:
    id_num, name, inst, inst_id, fee, pay, med_type, settle_dt = rec[:8]
    if settle_dt:
        patient_daily[id_num][settle_dt].append((med_type, inst, fee, name))

fake_hosp1 = []
for pid, daily in patient_daily.items():
    for dt, visits in daily.items():
        has_inp = any('住院' in v[0] for v in visits)
        has_outp = any('门诊' in v[0] for v in visits)
        if has_inp and has_outp:
            inp_visits = [v for v in visits if '住院' in v[0]]
            outp_visits = [v for v in visits if '门诊' in v[0]]
            fake_hosp1.append({
                'patient_id': pid[-4:],
                'name': visits[0][3],
                'date': str(dt),
                'inpatient_at': inp_visits[0][1],
                'outpatient_at': outp_visits[0][1],
                'inp_fee': inp_visits[0][2],
                'outp_fee': outp_visits[0][2],
            })

print(f'住院期间同日有门诊: {len(fake_hosp1)} 条')

# 2b: 同天不同医院住院
same_day_multi_hosp = []
for pid, daily in patient_daily.items():
    for dt, visits in daily.items():
        inp_visits = [v for v in visits if '住院' in v[0]]
        if len(inp_visits) >= 2:
            hospitals = set(v[1] for v in inp_visits)
            if len(hospitals) >= 2:
                same_day_multi_hosp.append({
                    'patient_id': pid[-4:],
                    'name': visits[0][3],
                    'date': str(dt),
                    'hospitals': list(hospitals),
                    'total_fee': sum(v[2] for v in inp_visits),
                })

print(f'同天不同医院住院: {len(same_day_multi_hosp)} 条')
if same_day_multi_hosp:
    print('\n同天多院住院明细:')
    for f in sorted(same_day_multi_hosp, key=lambda x: -x['total_fee'])[:10]:
        print(f'  {f["name"]} | {f["date"]} | {" + ".join(f["hospitals"])} | ¥{f["total_fee"]:,.0f}')

# ============================================================
# Model 3: 定点机构异常分析
# ============================================================
print('\n' + '='*60)
print('Model 3: 定点机构异常分析')
print('='*60)

# Aggregate by institution per year
inst_stats = defaultdict(lambda: defaultdict(lambda: {
    'inpatient_count': 0, 'outpatient_count': 0, 'total_count': 0,
    'inpatient_fee': 0.0, 'outpatient_fee': 0.0, 'total_fee': 0.0,
    'total_pay': 0.0, 'yidi_count': 0,
}))

for rec in all_records:
    id_num, name, inst, inst_id, fee, pay, med_type, settle_dt = rec[:8]
    year = rec[16]
    if not inst: continue
    
    s = inst_stats[year][inst]
    s['total_count'] += 1
    s['total_fee'] += fee
    s['total_pay'] += pay
    if '住院' in med_type:
        s['inpatient_count'] += 1
        s['inpatient_fee'] += fee
    elif '门诊' in med_type or '药店' in med_type:
        s['outpatient_count'] += 1
        s['outpatient_fee'] += fee
    if rec[12] == '是':
        s['yidi_count'] += 1

print(f'\n{"机构":<30} {"年份":>4} {"总条数":>8} {"总费用":>14} {"次均":>8} {"住院":>6} {"住院%":>7} {"异地%":>6}')
print('-'*90)

all_inst_metrics = []
for year in YEARS:
    for inst, s in inst_stats[year].items():
        if s['total_count'] < 50: continue  # Skip tiny institutions
        avg_fee = s['total_fee'] / s['total_count'] if s['total_count'] else 0
        inp_ratio = s['inpatient_count'] / s['total_count'] * 100 if s['total_count'] else 0
        yidi_ratio = s['yidi_count'] / s['total_count'] * 100 if s['total_count'] else 0
        
        all_inst_metrics.append({
            'inst': inst, 'year': year,
            'count': s['total_count'], 'total_fee': s['total_fee'],
            'avg_fee': avg_fee, 'inp_ratio': inp_ratio, 'yidi_ratio': yidi_ratio,
        })

# Find outliers: avg_fee > mean + 2*std per year
for year in YEARS:
    year_metrics = [m for m in all_inst_metrics if m['year'] == year and m['count'] >= 100]
    if not year_metrics: continue
    
    avg_fees = [m['avg_fee'] for m in year_metrics]
    mean_fee = sum(avg_fees) / len(avg_fees)
    # Simple std
    variance = sum((x - mean_fee)**2 for x in avg_fees) / len(avg_fees)
    std_fee = variance ** 0.5
    threshold = mean_fee + 2 * std_fee
    
    outliers = [m for m in year_metrics if m['avg_fee'] > threshold]
    outliers.sort(key=lambda x: -x['avg_fee'])
    
    print(f'\n{year}年: 县内均值 ¥{mean_fee:,.0f} | 2σ阈值 ¥{threshold:,.0f} | 离群机构: {len(outliers)}')
    for o in outliers[:5]:
        print(f'  🚨 {o["inst"][:30]:<30} 次均 ¥{o["avg_fee"]:>8,.0f}  ({o["count"]:>6,}条)  超均值{o["avg_fee"]/mean_fee:.1f}x')

# Growth rate analysis
print(f'\n年度增长率异常（2025 vs 2023）:')
for inst in set(m['inst'] for m in all_inst_metrics):
    m23 = [m for m in all_inst_metrics if m['inst'] == inst and m['year'] == '2023']
    m25 = [m for m in all_inst_metrics if m['inst'] == inst and m['year'] == '2025']
    if m23 and m25 and m23[0]['count'] >= 100:
        growth = (m25[0]['total_fee'] - m23[0]['total_fee']) / m23[0]['total_fee'] * 100
        if abs(growth) > 50:
            icon = '📈' if growth > 0 else '📉'
            print(f'  {icon} {inst[:30]:<30} {growth:+.0f}%  (¥{m23[0]["total_fee"]:,.0f} → ¥{m25[0]["total_fee"]:,.0f})')

# ============================================================
# Save findings
# ============================================================
findings = {
    'model1_split_hospitalization': {
        'count': len(unique_findings),
        'total_amount': total_amount,
        'top10': unique_findings[:10],
    },
    'model2_fake_hospitalization': {
        'same_day_outpatient': len(fake_hosp1),
        'same_day_multi_hospital': len(same_day_multi_hosp),
        'multi_hosp_details': same_day_multi_hosp[:10],
    },
    'model3_institution_anomaly': {
        'analyzed_institutions': len(set(m['inst'] for m in all_inst_metrics)),
    }
}

out_path = os.path.join(OUT, 'data_analysis_findings.json')
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(findings, f, ensure_ascii=False, indent=2)

print(f'\n\nFindings saved to: {out_path}')
print('Analysis complete.')
