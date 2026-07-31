"""Quick stats for 2024-2025 Guying data (skip 2023)."""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

base = r'C:\Users\scrccpa\Desktop\若尔盖审计\若尔盖医保审计\2026年审计资料（医保局财务）\古英=2024-2025门诊、住院、个人账户、基金拨付明细'

for fname in ['2024.xlsx', '2025.xlsx']:
    fp = os.path.join(base, fname)
    print(f'\n{"="*60}')
    print(f'{fname} ({os.path.getsize(fp)/1024/1024:.0f}MB)')
    print(f'{"="*60}')
    
    wb = load_workbook(fp, read_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        col_map = {str(h).strip(): i for i, h in enumerate(headers) if h}
        
        fee_col = pay_col = type_col = inst_col = date_col = yidi_col = None
        for h, i in col_map.items():
            if '医疗费总额' in h: fee_col = i
            if '医保支付金额' in h: pay_col = i
            if '医疗类别' in h: type_col = i
            if '医药机构名称' in h: inst_col = i
            if '结算时间' in h: date_col = i
            if '是否异地就医' in h: yidi_col = i
        
        count = total_fee = total_pay = yidi_count = 0
        medical_types = {}
        institutions = {}
        dates_min = dates_max = None
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            count += 1
            if count % 50000 == 0: print(f'  ...{count:,}')
            
            v = row[fee_col] if fee_col is not None else None
            if v and isinstance(v, (int, float)): total_fee += v
            
            v = row[pay_col] if pay_col is not None else None
            if v and isinstance(v, (int, float)): total_pay += v
            
            v = row[type_col] if type_col is not None else None
            if v:
                vt = str(v).split('|')[-1].strip() if '|' in str(v) else str(v)
                medical_types[vt] = medical_types.get(vt, 0) + 1
            
            v = row[inst_col] if inst_col is not None else None
            if v: institutions[str(v).strip()] = institutions.get(str(v).strip(), 0) + 1
            
            v = row[date_col] if date_col is not None else None
            if v and hasattr(v, 'year'):
                if dates_min is None or v < dates_min: dates_min = v
                if dates_max is None or v > dates_max: dates_max = v
            
            v = row[yidi_col] if yidi_col is not None else None
            if v and str(v).strip() == '是': yidi_count += 1
        
        print(f'\n  Sheet: {sn} | Rows: {count:,}')
        print(f'  医疗费总额: ¥{total_fee:,.0f}')
        print(f'  医保支付总额: ¥{total_pay:,.0f}')
        print(f'  次均费用: ¥{total_fee/count:,.0f}' if count else '')
        if dates_min and dates_max:
            print(f'  日期: {dates_min.date()} ~ {dates_max.date()}')
        print(f'  异地就医: {yidi_count:,} ({yidi_count/count*100:.1f}%)' if count else '')
        
        print(f'  医疗类别:')
        for k, v in sorted(medical_types.items(), key=lambda x: -x[1])[:5]:
            print(f'    {k}: {v:,} ({v/count*100:.1f}%)')
        
        print(f'  Top 10 机构:')
        for k, v in sorted(institutions.items(), key=lambda x: -x[1])[:10]:
            print(f'    {k}: {v:,}')
    
    wb.close()

print('\nDone.')
