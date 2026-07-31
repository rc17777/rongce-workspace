"""Deep-dive v2: raw records, weekly pattern, cross-institution check"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from collections import defaultdict, Counter
from datetime import datetime, timedelta

def pdate(v):
    if not v: return None
    if hasattr(v, 'date'): return v.date()
    try: return datetime.strptime(str(v)[:10], '%Y-%m-%d').date()
    except: return None
def sf(v): return float(v) if v and isinstance(v,(int,float)) else 0.0
def ss(v): return str(v).strip() if v else ''

BASE = r'C:\Users\scrccpa\Desktop\若尔盖审计\若尔盖医保审计\2026年审计资料（医保局财务）\古英=2024-2025门诊、住院、个人账户、基金拨付明细'
INST = '若尔盖县降扎乡卫生院'
INST_LIST = [INST]

# TOP 22 frequent patients from previous analysis
TOP22 = ['2912','2920','2913','2927','2919','2917','2916','2915','2924',
         '291X','2926','2925','2922','2918','2929','2910','2914','2923',
         '2921','292X','2911','2928']

# ===== Phase 1: Full detailed records for 降扎乡 + cross-institution for top22 =====
records_jz = []  # All 降扎 records with full detail
records_top22_other = []  # Top22 patients at OTHER institutions
records_top22_pharm = []  # Top22 at pharmacies
top22_patients_full_id = {}  # full id -> last4 mapping for cross-check

for year in ['2023','2024','2025']:
    fp = f'{BASE}\\{year}.xlsx'
    wb = load_workbook(fp, read_only=True)
    sn = '总表' if '总表' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sn]
    hdrs = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    col = {str(h).strip(): i for i, h in enumerate(hdrs) if h}
    
    for r in ws.iter_rows(min_row=2, values_only=True):
        full_pid = ss(r[col['证件号码']]) if col.get('证件号码') else ''
        last4 = full_pid[-4:] if full_pid else ''
        inst = ss(r[col['医药机构名称']]) if col.get('医药机构名称') else ''
        name = ss(r[col['姓名']]) if col.get('姓名') else ''
        mt = ss(r[col['医疗类别']]) if col.get('医疗类别') else ''
        mt_clean = mt.split('|')[-1].strip() if '|' in mt else mt
        settle = pdate(r[col['结算时间']]) if col.get('结算时间') else None
        fee = sf(r[col.get('医疗费总额')])
        pay = sf(r[col.get('医保支付金额')])
        diag_idx = col.get('出院诊断名称')
        diag = ss(r[diag_idx]) if diag_idx is not None else ''

        rec = {'year':year, 'full_pid':full_pid, 'last4':last4, 'name':name,
               'type':mt_clean, 'inst':inst, 'date':settle, 'fee':fee, 'pay':pay, 'diag':diag}
        
        if inst == INST:
            records_jz.append(rec)
        elif last4 in TOP22:
            records_top22_other.append(rec)
            top22_patients_full_id[full_pid] = last4
            if '药店' in inst or '药房' in inst:
                records_top22_pharm.append(rec)
    
    wb.close()
    print(f'{year}: JZ={sum(1 for r in records_jz if r["year"]==year)}, Top22_other={sum(1 for r in records_top22_other if r["year"]==year)}')

print(f'\nJZ total: {len(records_jz)}, Top22 at other: {len(records_top22_other)}, Top22 at pharmacy: {len(records_top22_pharm)}')

# ===== Phase 2: Weekly pattern for TOP 3 patients =====
print('\n' + '='*70)
print('PHASE 2: TOP 3 患者逐条就诊明细')

# Get top 3
patient_all = defaultdict(list)
for rec in records_jz: patient_all[rec['last4']].append(rec)
top3 = sorted(patient_all.items(), key=lambda x: -len(x[1]))[:3]

for pid, recs in top3:
    recs.sort(key=lambda x: x['date'] if x['date'] else datetime.min.date())
    print(f'\n--- {pid} {recs[0]["name"]}: {len(recs)}条 ---')
    
    # Day of week distribution
    dow = Counter()
    for r in recs:
        if r['date']: dow[r['date'].weekday()] += 1
    days = ['一','二','三','四','五','六','日']
    dow_str = ' '.join(f'{days[d]}:{dow.get(d,0)}' for d in range(7))
    print(f'  星期分布: {dow_str}')
    
    # Avg fee, fee range
    fees = [r['fee'] for r in recs]
    print(f'  费用: 均¥{sum(fees)/len(fees):.0f} | 范围¥{min(fees):.0f}-¥{max(fees):.0f}')
    
    # Check fee regularity (same amount each time?)
    fee_counter = Counter(f"{f:.0f}" for f in fees)
    top_fees = fee_counter.most_common(5)
    reg = '🔴 高度规律' if len(top_fees) <= 2 else '🟡 较规律' if len(top_fees) <= 4 else '🟢 不规则'
    print(f'  费用规律: {reg} | 最常见: {" / ".join(f"¥{f}x{c}" for f,c in top_fees[:5])}')
    
    # Interval regularity
    dates = [r['date'] for r in recs if r['date']]
    intervals = [(dates[i+1]-dates[i]).days for i in range(len(dates)-1)]
    if intervals:
        int_counter = Counter(intervals)
        top_int = int_counter.most_common(5)
        avg_int = sum(intervals)/len(intervals)
        print(f'  间隔规律: 均{avg_int:.0f}天 | 最常见: {" / ".join(f"{d}天x{c}" for d,c in top_int[:5])}')
    
    # Show ALL records for 2025.12
    dec_recs = [r for r in recs if r['date'] and r['date'].month == 12 and r['date'].year == 2025]
    if dec_recs:
        print(f'  2025.12月就诊({len(dec_recs)}条):')
        for r in dec_recs:
            print(f'    {r["date"]} ({days[r["date"].weekday()]})  ¥{r["fee"]:>6,.0f}  {r["diag"][:20]}')
    
    # First date, last date
    print(f'  跨度: {dates[0]} ~ {dates[-1]}' if dates else '  无日期')

# ===== Phase 3: December 2025 spike - who are these people? =====
print('\n' + '='*70)
print('PHASE 3: 2025.12月集中刷——新老患者分析')

dec_2025 = [r for r in records_jz if r['date'] and r['date'].year == 2025 and r['date'].month == 12]
dec_patients = {}
for r in dec_2025:
    if r['last4'] not in dec_patients: dec_patients[r['last4']] = {'name':r['name'], 'dec_count':0, 'dec_fee':0, 'first_date':r['date'], 'total_before':0}
    dec_patients[r['last4']]['dec_count'] += 1
    dec_patients[r['last4']]['dec_fee'] += r['fee']

# Check if these patients existed before Dec 2025
for pid in dec_patients:
    before = [r for r in records_jz if r['last4'] == pid and r['date'] and r['date'] < datetime(2025,12,1).date()]
    dec_patients[pid]['total_before'] = len(before)
    dec_patients[pid]['first_date'] = before[0]['date'] if before else dec_patients[pid]['first_date']

new_patients = {k:v for k,v in dec_patients.items() if v['total_before'] == 0}
old_patients = {k:v for k,v in dec_patients.items() if v['total_before'] > 0}

print(f'12月就诊: {len(dec_patients)}人，{len(dec_2025)}条')
print(f'  老患者: {len(old_patients)}人（之前就有就诊记录）')
print(f'  新面孔: {len(new_patients)}人（首次出现在12月）')

if new_patients:
    print(f'\n  🚨 {len(new_patients)}个新面孔明细:')
    for pid, v in sorted(new_patients.items(), key=lambda x: -x[1]['dec_count']):
        print(f'    {pid} {v["name"]:<8} {v["dec_count"]}次 ¥{v["dec_fee"]:,.0f}  首次:{v["first_date"]}')

# ===== Phase 4: Cross-institution check for Top22 =====
print('\n' + '='*70)
print('PHASE 4: TOP22患者跨机构就诊')
print('  他们在其他地方也"高频"吗？')

other_by_patient = defaultdict(lambda: defaultdict(list))
for rec in records_top22_other:
    other_by_patient[rec['last4']][rec['inst']].append(rec)

for pid in sorted(other_by_patient.keys()):
    by_inst = other_by_patient[pid]
    # Only show if they have significant visits elsewhere (>=5)
    sig_others = {k:v for k,v in by_inst.items() if len(v) >= 5}
    if sig_others:
        name = sig_others[list(sig_others.keys())[0]][0]['name']
        total_other = sum(len(v) for v in sig_others.values())
        print(f'\n  {pid} {name}: 同时在{len(sig_others)}个其他机构就诊({total_other}条)')
        for inst, recs in sorted(sig_others.items(), key=lambda x: -len(x[1]))[:5]:
            fees = [r['fee'] for r in recs]
            print(f'    {inst[:35]:<35} {len(recs):>4}条 ¥{sum(fees):>8,.0f} 次均¥{sum(fees)/len(fees):.0f}')

# ===== Phase 5: 12.30 the DAY =====
print('\n' + '='*70)
print('PHASE 5: 2025-12-30 当日全貌')

d30 = [r for r in records_jz if r['date'] and r['date'] == datetime(2025,12,30).date()]
d29 = [r for r in records_jz if r['date'] and r['date'] == datetime(2025,12,29).date()]
d31 = [r for r in records_jz if r['date'] and r['date'] == datetime(2025,12,31).date()]

print(f'12/29: {len(d29)}条 {len(set(r["last4"] for r in d29))}人')
print(f'12/30: {len(d30)}条 {len(set(r["last4"] for r in d30))}人')
print(f'12/31: {len(d31)}条 {len(set(r["last4"] for r in d31))}人')

# Who visited on all 3 consecutive days?
p29 = set(r['last4'] for r in d29)
p30 = set(r['last4'] for r in d30)
p31 = set(r['last4'] for r in d31)
all3 = p29 & p30 & p31
if all3:
    print(f'\n🚨 连续3天(29→30→31)都来的患者: {len(all3)}人')
    for pid in all3:
        name = [r['name'] for r in d30 if r['last4']==pid][0]
        f29 = sum(r['fee'] for r in d29 if r['last4']==pid)
        f30 = sum(r['fee'] for r in d30 if r['last4']==pid)
        f31 = sum(r['fee'] for r in d31 if r['last4']==pid)
        print(f'  {pid} {name}: 29日¥{f29:,.0f} + 30日¥{f30:,.0f} + 31日¥{f31:,.0f} = ¥{f29+f30+f31:,.0f}')

print(f'\n12/30 全部就诊:')
for r in sorted(d30, key=lambda x: -x['fee']):
    print(f'  {r["last4"]} {r["name"]:<8}  ¥{r["fee"]:>6,.0f}  {r["type"]}  {r["diag"][:20]}')

print('\n✅ 降扎乡深度挖掘v2完成')
