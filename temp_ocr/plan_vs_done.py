"""Compare audit plan vs completed work"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = Workbook()

# Styles
hdr_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='0A1F3F')
done_fill = PatternFill('solid', fgColor='C8E6C9')
partial_fill = PatternFill('solid', fgColor='FFF9C4')
not_done_fill = PatternFill('solid', fgColor='FFE0E0')
na_fill = PatternFill('solid', fgColor='F5F5F5')
body_font = Font(name='Microsoft YaHei', size=10)
bold_font = Font(name='Microsoft YaHei', size=10, bold=True)
title_font = Font(name='Microsoft YaHei', size=14, bold=True, color='0A1F3F')
sub_font = Font(name='Microsoft YaHei', size=11, bold=True, color='1A5C6E')
wrap = Alignment(wrap_text=True, vertical='top')
center = Alignment(horizontal='center', vertical='top', wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def style_sheet(ws, headers, col_widths, title=None):
    sr = 1
    if title:
        ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=len(headers))
        ws.cell(row=sr, column=1, value=title).font = title_font
        sr += 1
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=sr, column=i, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = thin_border
        ws.column_dimensions[get_column_letter(i)].width = w
    return sr + 1

def write_rows(ws, sr, rows):
    for ri, row in enumerate(rows):
        r = sr + ri
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.font = body_font; c.alignment = wrap; c.border = thin_border
        status = str(row[3]) if len(row) > 3 else ''
        if '已完成' in status:
            for ci in range(1, len(row)+1):
                ws.cell(row=r, column=ci).fill = done_fill
        elif '部分完成' in status:
            for ci in range(1, len(row)+1):
                ws.cell(row=r, column=ci).fill = partial_fill
        elif '未开展' in status:
            for ci in range(1, len(row)+1):
                ws.cell(row=r, column=ci).fill = not_done_fill

# ==========================================
# Sheet 1: 实施方案对照总表
# ==========================================
ws1 = wb.active
ws1.title = '实施方案对照总表'

h1 = ['章节', '方案要求', '审计事项', '完成状态', '已完成工作', '对应发现编号', '未完成原因/下一步']
w1 = [24, 32, 28, 12, 40, 16, 36]
r1 = style_sheet(ws1, h1, w1, '医疗保障基金审计实施方案 — 完成情况对照 (' + datetime.now().strftime('%Y-%m-%d') + ')')

items = [
    # === 三(一) 基金筹集 ===
    ['三(一)基金筹集', '审查基金预算与财政补助衔接', '财政补助到位、拨付及时性', '未开展',
     '未涉及。需要财政补助拨付文件、预算批复文件、银行到账凭证。',
     '-', '需取得2024-2025年财政补助资金拨付凭证和预算文件'],

    ['三(一)基金筹集', '审查重点人群参保和资助参保', '特困/低保/重残等应保尽保', '未开展',
     '未涉及。需要参保名单与民政/残联名单比对。',
     '-', '需取得特困/低保/重残花名册+医保参保名单进行比对'],

    ['三(一)基金筹集', '审查税务征缴/退费/欠费', '应征未征/重复参保', '未开展',
     '未涉及。需要税务征缴数据和参保名单。',
     '-', '需取得税务征缴明细和退费记录'],

    # === 三(二) 基金支付 ===
    ['三(二)基金支付', '审查医保待遇审核支付', '手工报销/异地就医结算', '未开展',
     '未涉及。仅分析了门诊结算数据，未涉及手工报销和异地就医。',
     '-', '需取得手工报销明细和异地就医结算清单'],

    ['三(二)基金支付', '核查跨渠道重复报销', '医保+大病+救助重复报销', '未开展',
     '未涉及。仅分析了门诊结算单一渠道。',
     '-', '需取得大病保险和医疗救助结算明细，与医保结算交叉比对'],

    ['三(二)基金支付', '审查基金专户管理', '挤占挪用/违规出借', '未开展',
     '未涉及。需要基金专户银行对账单和支出明细。',
     '-', '需取得基金专户银行流水和支出凭证'],

    ['三(二)基金支付', '个人账户管理/死亡清退', '死亡后账户未清退', '未开展',
     '未涉及。需要死亡人口数据与个人账户消费记录比对。',
     '-', '需取得死亡人口+个人账户消费明细进行比对'],

    # === 三(三) 定点医药机构 ===
    ['三(三)定点机构', '虚假诊疗', '无诊断/虚假就诊记录', '✅ 已完成',
     'P01:降扎乡三年1860条100%无诊断\nP03:三家机构合计19,386条0诊断\nP02:卓玛泽让一人签569名患者',
     'P01 P02 P03', '已通过三年数据三次交叉验证'],

    ['三(三)定点机构', '挂床住院', '住院记录异常', '未开展',
     '未涉及。仅分析了门诊数据，未分析住院结算数据。',
     '-', '需分析住院结算数据：入院出院同日/住院天数异常/床位费与住院天数不匹配'],

    ['三(三)定点机构', '分解处方（分解住院）', '同日多次刷卡/分解处方', '✅ 已完成',
     'S03:旦真机12/25-30连续6天10次¥288\n同日多次：阿容12/18两次、旦真机12/14两次+12/28三次',
     'S03', '处方笺尚未调取(现场取证阶段)'],

    ['三(三)定点机构', '过度诊疗', '超高频低额就诊', '✅ 已完成',
     'S01:降扎乡门诊量10.3倍暴增(126→437→1297)\nS04:22人多机构轮转(最高19机构291条)\nF13:78%就诊≤¥50',
     'S01 S04', '已在门诊数据中充分分析'],

    ['三(三)定点机构', '重复收费', '同日拆分交易(金世康)', '✅ 已完成',
     'S05:金世康同日拆分交易\n11/10:126条58人100%双重\nTOP20患者同日双重天数占45-50%\n同秒两笔同一金额=系统自动拆分',
     'S05', '进销存核对尚未开展(现场取证阶段)'],

    ['三(三)定点机构', '超标准收费', '收费项目与价格对照', '未开展',
     '未涉及。需要医疗服务价格目录与结算明细中的项目单价比对。',
     '-', '需取得四川省医疗服务价格目录+结算明细中的项目清单'],

    ['三(三)定点机构', '串换药品耗材和诊疗项目', '串换(将自费项目串为医保)', '未开展',
     '未涉及。需要HIS收费明细与医保结算明细逐项比对。',
     '-', '需取得His系统收费明细与医保结算明细交叉比对'],

    ['三(三)定点机构', 'DRG/DIP高套分组', '高套诊断编码', '未开展',
     '未涉及。需要住院病案首页(含DRG分组)数据。',
     '-', '需取得住院病案首页数据+DRG分组结果'],

    # === 三(三) 定点药店 ===
    ['三(三)定点药店', '虚开串药', '药店无医生无诊断开药', '✅ 已完成',
     'S05:金世康16,440条门诊记录，无医生、无诊断、100%同日拆分\n对比降扎卫生院:次均¥76.8 vs ¥39.9',
     'S05', '已完成数据分析，现场需调取进销存'],

    ['三(三)定点药店', '刷卡套现', '个账套现/非药品刷卡', '🟡 部分完成',
     '金世康100%职工个账支出，¥126万全为个账。同日拆分模式高度可疑套现。\n但未调取进销存确认是否涉及非药品(保健品/日用品)。',
     'S05 S06', '需调取进销存+监控确认套现→下一步'],

    ['三(三)定点药店', '倒卖回流药', '药品回收再销售', '未开展',
     '未涉及。需要追踪药品批号流向。',
     '-', '需取得药品进货发票+销售记录，交叉比对批号'],

    ['三(三)定点药店', '进销存不匹配', '系统采购≠销售≠库存', '🟡 部分完成',
     'S05已识别拆分模式=系统性问题。\nS06已识别11月暴增67%=销售异常。\n但尚未取得进销存数据进行三方比对(采购-销售-库存)。',
     'S05 S06', '已列入取证清单第4步→下一步'],

    # === 三(三) 经办机构 ===
    ['三(三)经办机构', '审核把关不严', '异常数据长期未处理', '🟡 部分完成',
     'P03:三家机构19386条100%无诊断=异常数据在医保结算系统中存在2-3年未触发任何审核拦截。\n间接证明审核机制失效。',
     'P03', '需与医保经办机构确认是否有智能审核规则、为何三年未发现'],

    ['三(三)经办机构', '权限设置/内外勾结', '经办人员异常操作', '未开展',
     '未涉及。需要操作日志。',
     '-', '需取得医保核心系统操作日志+权限配置表'],

    # === 三(四) 基金运行 ===
    ['三(四)基金运行', '基金收支结余趋势分析', '可支付月数/穿底风险', '未开展',
     '未涉及。需要2024-2025年基金收支决算表和参保人数趋势数据。',
     '-', '需取得基金收支决算表、参保人数统计表'],

    ['三(四)基金运行', '药品耗材集中带量采购', '集采任务完成/结算回款', '未开展',
     '未涉及。需要集采任务指标和实际采购结算数据。',
     '-', '需取得集采任务文件+实际采购结算清单'],

    ['三(四)基金运行', '药品加成取消/价格调整', '医疗服务价格/患者负担', '未开展',
     '未涉及。仅分析了费用金额，未比对价格标准。',
     '-', '需取得医疗服务价格调整文件和实际收费项目清单'],

    # === 三(五) 以往整改 ===
    ['三(五)以往整改', '以往发现问题整改闭环', '整改是否真实/资金是否追回', '未开展',
     '未涉及。未取得以往年度审计报告、检查报告和整改台账。',
     '-', '需取得以往审计/检查/巡察问题清单和整改台账'],

    # === 二、审计范围 ===
    ['二审计范围', '延伸抽查定点零售药店', '金世康药品店', '✅ 已完成',
     '已完成金世康药品店16,440条数据的全面分析：\n- 100%个账支出(无统筹基金)\n- 同秒双笔系统性拆分\n- 11月暴涨67%\n- TOP20患者同日双重率45-50%',
     'S05 S06', '数据分析已完成，现场取证待开展'],

    ['二审计范围', '延伸抽查定点医疗机构', '降扎乡卫生院+辖曼镇卫生院', '✅ 已完成',
     '降扎乡:1734条全分析(单医/无诊断/分解处方/年底冲量)\n辖曼镇:1212条全分析(12月占40%/6.5倍增速)',
     'P01 P02 S01-S04 S07', '两家机构数据分析已充分完成'],

    # === 四、审计方法 ===
    ['四审计方法', '数据分析模型筛查', '门诊数据多维度模型', '✅ 已完成',
     '已运行以下分析模型:\n- 按机构/患者/医生/日期/费用多维度聚合\n- 交叉验证矩阵(每条≥2次独立验证)\n- 三角对比(三家机构横向比较)\n- 同日拆分检测(精确到秒)',
     '全部11条发现', '8个分析维度/3次独立读取/2年数据'],

    ['四审计方法', '重点机构延伸', '降扎+金世康+辖曼', '✅ 已完成',
     '三家机构合计19,386条数据分析完成\n降扎(1734)+金世康(16440)+辖曼(1212)',
     '全部', '数据层面已全部分析'],

    ['四审计方法', '政策制度审查/财务账表核对/业务流程穿行', '政策/财务/流程三方', '未开展',
     '未涉及。仅完成了数据分析层面。',
     '-', '需现场审计阶段完成：政策文件收集+财务凭证核对+业务流程访谈'],

    # === 五、延伸对象选择 ===
    ['五延伸选择', '基金支付金额大/增长率高/人均异常', '金世康¥126万/降扎10x/辖曼6.5x', '✅ 已完成',
     '三家机构均满足"重点延伸对象"标准：\n金世康:¥126万/两年=全县药店最大\n降扎:10.3倍增长=异常\n辖曼:6.5倍增长+12月40%=异常',
     '全部', '三家均已被识别为重点对象且已完成初步分析'],
]

write_rows(ws1, r1, items)
ws1.freeze_panes = 'A%d' % r1
ws1.auto_filter.ref = 'A%d:G%d' % (r1-1, r1-1+len(items))

# ==========================================
# Sheet 2: 统计摘要
# ==========================================
ws2 = wb.create_sheet('统计摘要')

done = sum(1 for item in items if '已完成' in str(item[3]))
partial = sum(1 for item in items if '部分完成' in str(item[3]))
not_done = sum(1 for item in items if '未开展' in str(item[3]))
total = len(items)

ws2.merge_cells('A1:D1')
ws2.cell(row=1, column=1, value='实施方案完成情况统计').font = title_font

h2 = ['状态', '数量', '占比', '说明']
w2 = [16, 10, 10, 50]
sr2 = 3
for i, (h, w) in enumerate(zip(h2, w2), 1):
    ws2.cell(row=sr2, column=i, value=h).font = hdr_font
    ws2.cell(row=sr2, column=i).fill = hdr_fill
    ws2.cell(row=sr2, column=i).alignment = center
    ws2.cell(row=sr2, column=i).border = thin_border
    ws2.column_dimensions[get_column_letter(i)].width = w

sum_data = [
    ['✅ 已完成', done, f'{done/total*100:.0f}%', '数据分析层面已完成，部分需现场取证确认'],
    ['🟡 部分完成', partial, f'{partial/total*100:.0f}%', '数据分析已做，缺少现场取证或额外数据'],
    ['❌ 未开展', not_done, f'{not_done/total*100:.0f}%', '尚未覆盖，需要额外数据或现场审计'],
    ['合计', total, '100%', ''],
]
for ri, row in enumerate(sum_data):
    r = sr2 + 1 + ri
    for ci, val in enumerate(row, 1):
        ws2.cell(row=r, column=ci, value=val).font = bold_font if ri == 3 else body_font
        ws2.cell(row=r, column=ci).border = thin_border
        ws2.cell(row=r, column=ci).alignment = wrap
    if ri == 0:
        for ci in range(1, 5): ws2.cell(row=r, column=ci).fill = done_fill
    elif ri == 1:
        for ci in range(1, 5): ws2.cell(row=r, column=ci).fill = partial_fill
    elif ri == 2:
        for ci in range(1, 5): ws2.cell(row=r, column=ci).fill = not_done_fill

# Detail breakdown
r_detail = sr2 + 6
ws2.cell(row=r_detail, column=1, value='各章节完成情况').font = sub_font

chapter_stats = [
    ('三(一) 基金筹集和财政补助', 0, 0, 3),
    ('三(二) 基金支付和待遇审核', 0, 0, 4),
    ('三(三) 定点医药机构-医疗机构', 4, 2, 4),
    ('三(三) 定点医药机构-零售药店', 1, 2, 1),
    ('三(三) 定点医药机构-经办机构', 0, 1, 1),
    ('三(四) 基金运行可持续', 0, 0, 3),
    ('三(五) 以往问题整改', 0, 0, 1),
    ('二 审计范围(机构延伸)', 2, 0, 0),
    ('四 审计方法', 2, 0, 1),
    ('五 延伸对象选择', 1, 0, 0),
]

r_detail += 1
for i, h in enumerate(['章节', '已完成', '部分完成', '未开展'], 1):
    ws2.cell(row=r_detail, column=i, value=h).font = bold_font
    ws2.cell(row=r_detail, column=i).border = thin_border
r_detail += 1

for ch, d, p, n in chapter_stats:
    ws2.cell(row=r_detail, column=1, value=ch).font = body_font
    ws2.cell(row=r_detail, column=1).border = thin_border
    ws2.cell(row=r_detail, column=2, value=d).font = body_font
    ws2.cell(row=r_detail, column=2).border = thin_border
    ws2.cell(row=r_detail, column=3, value=p).font = body_font
    ws2.cell(row=r_detail, column=3).border = thin_border
    ws2.cell(row=r_detail, column=4, value=n).font = body_font
    ws2.cell(row=r_detail, column=4).border = thin_border
    if d > 0: ws2.cell(row=r_detail, column=2).fill = done_fill
    if p > 0: ws2.cell(row=r_detail, column=3).fill = partial_fill
    if n > 0: ws2.cell(row=r_detail, column=4).fill = not_done_fill
    r_detail += 1

# ==========================================
# Sheet 3: 下一步行动计划
# ==========================================
ws3 = wb.create_sheet('下一步行动计划')
h3 = ['优先级', '行动项', '涉及方案条款', '需补充数据', '预计工作量', '预期产出']
w3 = [8, 42, 22, 36, 14, 36]
r3 = style_sheet(ws3, h3, w3, '下一步行动计划（按方案覆盖优先级排序）')

actions = [
    ['P0', '现场取证：降扎乡P01+P02(无诊断+单医)', '三(三)虚假诊疗',
     '①纸质门诊日志 ②卓玛泽让访谈 ③人员编制文件',
     '1天', '签字确认书+访谈笔录→做实P01+P02'],

    ['P0', '现场取证：金世康S05(同日拆分)', '三(三)虚开串药/刷卡套现',
     '①进销存系统数据 ②11/10监控录像 ③店主访谈',
     '1-2天', '进销存比对表+录像截图→做实S05或排除'],

    ['P0', '现场取证：降扎乡S03(分解处方)', '三(三)分解处方',
     '①旦真机等5人处方笺 ②患者面访',
     '1天', '处方对照表+面访笔录→做实或排除分解处方'],

    ['P1', '住院数据分析：挂床住院+DRG高套', '三(三)挂床住院/DRG高套',
     '①住院结算明细(2023-2025年) ②住院病案首页',
     '2-3天', '挂床住院疑点清单+DRG异常清单'],

    ['P1', '死亡人员个账清退核查', '三(二)个人账户管理',
     '①2024-2025年死亡人口数据 ②个人账户消费明细',
     '1天', '死亡后仍使用个账的疑点清单'],

    ['P1', '扩大机构分析：全县其他卫生院+药店', '二审计范围/三(三)',
     '①全县其他2-3家卫生院门诊数据 ②其他2-3家药店数据',
     '2-3天', '全县机构风险排序+完善P03(系统性无诊断)扩大验证'],

    ['P2', '基金筹集分析：财政补助+重点人群参保', '三(一)基金筹集',
     '①财政补助拨付凭证 ②特困/低保/重残花名册 ③税务征缴明细',
     '3-5天', '筹资环节疑点清单'],

    ['P2', '基金支付分析：重复报销+专户管理', '三(二)基金支付',
     '①大病保险结算明细 ②医疗救助结算明细 ③基金专户银行流水',
     '2-3天', '重复报销疑点清单+专户合规报告'],

    ['P2', '基金运行分析：收支趋势+集采', '三(四)基金运行',
     '①基金收支决算表 ②集采任务完成数据 ③参保趋势',
     '2天', '基金可持续性分析报告'],

    ['P2', '以往整改闭环核查', '三(五)以往整改',
     '①以往年度审计/检查/巡察问题清单 ②整改台账',
     '1-2天', '整改回头看报告'],

    ['P3', '经办机构审核机制审查', '三(三)经办机构',
     '①智能审核规则配置 ②操作日志 ③权限配置表',
     '2-3天', '经办管理问题清单'],

    ['P3', '串换药品耗材核查', '三(三)串换',
     '①His收费明细 ②医疗服务价格目录',
     '2天', '串换疑点清单'],
]
write_rows(ws3, r3, actions)

# ==========================================
# Sheet 4: 已发现问题与方案对应
# ==========================================
ws4 = wb.create_sheet('发现与方案对应')
h4 = ['发现编号', '等级', '发现描述', '对应方案条款', '方案要求', '完成深度', '备注']
w4 = [10, 12, 36, 22, 30, 12, 24]
r4 = style_sheet(ws4, h4, w4, '已发现问题与实施方案条款对应关系')

mapping = [
    ['P01', '🔴 确认', '降扎乡三年1860条100%无诊断', '三(三)虚假诊疗',
     '重点审查虚假诊疗', '数据充分', '可定案，需现场签字确认'],
    ['P02', '🔴 确认', '卓玛泽让一人接诊569名患者', '三(三)虚假诊疗',
     '重点审查虚假诊疗', '数据充分', '可定案，需访谈确认'],
    ['P03', '🔴 确认', '三家机构19386条均无诊断', '三(三)虚假诊疗/经办机构',
     '虚假诊疗+审核把关不严', '数据充分', '系统性问题，需扩大验证'],
    ['S01', '🟡 可疑', '降扎门诊量10.3倍暴增', '三(三)过度诊疗\n五延伸选择(增长率高)',
     '过度诊疗+增长率异常', '数据充分', '需人口数据论证合理性'],
    ['S02', '🟡 可疑', '降扎12/30单日26条冲量', '三(三)虚假诊疗',
     '年底集中刷卡', '数据充分', '需监控录像+访谈'],
    ['S03', '🟡 可疑', '旦真机6天10次分解处方', '三(三)分解处方',
     '分解处方', '数据充分', '需处方笺物理证据'],
    ['S04', '🟡 可疑', '22人多机构轮转开药', '三(三)过度诊疗\n三(二)重复报销',
     '过度诊疗+跨渠道重复报销', '数据充分', '需全口径费用汇总'],
    ['S05', '🔴 可疑', '金世康系统同秒拆分交易', '三(三)虚开串药/刷卡套现/进销存不匹配',
     '虚开串药+刷卡套现+进销存不匹配', '数据充分', '需进销存+监控物理证据'],
    ['S06', '🟡 可疑', '金世康11月暴涨67%', '三(三)虚开串药\n五延伸选择(金额大)',
     '虚开串药+支付金额大', '数据充分', '需其他药店对比'],
    ['S07', '🟡 可疑', '辖曼12月占全年40%', '三(三)虚假诊疗\n五延伸选择(增长率高)',
     '年底集中+增长率异常', '数据充分', '需公卫项目文件'],
    ['S08', '🟢 可疑', '患者互不重叠(<1%共享)', '三(三)审核把关',
     '异常数据识别', '需扩大验证', '需全县10+机构对比'],
]
write_rows(ws4, r4, mapping)

# Save
out = r'C:\Users\scrccpa\Desktop\若尔盖医保审计_实施方案对照.xlsx'
wb.save(out)
print('Saved: ' + out)
print(f'Total items: {total}, Done: {done}, Partial: {partial}, Not done: {not_done}')
