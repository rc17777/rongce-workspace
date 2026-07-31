"""Deep dive into 降扎乡卫生院 2025 data"""
import sys, os, json
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from collections import Counter, defaultdict
from datetime import datetime

print("Loading 2025.xlsx...")
data_dir = r'C:\Users\scrccpa\Desktop\若尔盖审计\若尔盖医保审计\2026年审计资料（医保局财务）\古英=2024-2025门诊、住院、个人账户、基金拨付明细'
wb = load_workbook(os.path.join(data_dir, '2025.xlsx'), read_only=True, data_only=True)
ws = wb['总表']
headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
print(f"Headers loaded: {len(headers)} cols")

# Index map
col_idx = {h: i for i, h in enumerate(headers)}

ji_zha = []
patient_records = defaultdict(list)
doctor_counter = Counter()
fee_totals = Counter()
address_counter = Counter()
township_counter = Counter()
insurance_counter = Counter()
monthly = Counter()
daily = Counter()
fee_buckets = Counter()
diag_counter = Counter()
visit_type_counter = Counter()
settle_type_counter = Counter()

print("Scanning rows...")
count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    count += 1
    if count % 50000 == 0:
        print(f"  Scanned {count} rows, found {len(ji_zha)} 降扎 records")
    
    inst = str(row[col_idx['医药机构名称']]) if row[col_idx['医药机构名称']] else ''
    if '降扎' not in inst:
        continue
    
    ji_zha.append(row)
    
    # Patient key
    cert = str(row[col_idx['证件号码']]) if row[col_idx['证件号码']] else ''
    cert4 = cert[-4:] if len(cert) >= 4 else cert
    name = str(row[col_idx['姓名']]) if row[col_idx['姓名']] else ''
    pkey = f"{name}({cert4})"
    patient_records[pkey].append(row)
    
    # Doctor
    doc = str(row[col_idx['医师姓名']]) if row[col_idx['医师姓名']] else ''
    if doc and doc != 'None':
        doctor_counter[doc] += 1
    
    # Fees
    for k in ['药品费','诊疗费','耗材费','床位费','医疗费总额','统筹基金支出','个人账户支付','个人现金支付']:
        if k in col_idx:
            v = row[col_idx[k]]
            if v and isinstance(v, (int, float)):
                fee_totals[k] += float(v)
            elif v:
                try: fee_totals[k] += float(v)
                except: pass
    
    # Fee bucket
    total_fee = row[col_idx['医疗费总额']] if '医疗费总额' in col_idx else 0
    if total_fee and isinstance(total_fee, (int, float)):
        if total_fee <= 50: fee_buckets['0-50'] += 1
        elif total_fee <= 100: fee_buckets['50-100'] += 1
        elif total_fee <= 200: fee_buckets['100-200'] += 1
        else: fee_buckets['200+'] += 1
    
    # Address
    addr = str(row[col_idx['居住地址']]) if row[col_idx['居住地址']] else ''
    if addr and addr != 'None':
        address_counter[addr] += 1
    
    twp = str(row[col_idx['乡镇(街道)']]) if row[col_idx['乡镇(街道)']] else ''
    if twp and twp != 'None':
        township_counter[twp] += 1
    
    # Insurance
    ins = str(row[col_idx['参保类型']]) if row[col_idx['参保类型']] else ''
    if ins and ins != 'None':
        insurance_counter[ins] += 1
    
    # Diagnosis
    diag = str(row[col_idx['出院诊断名称']]) if row[col_idx['出院诊断名称']] else ''
    if diag and diag != 'None' and diag.strip():
        diag_counter[diag] += 1
    
    # Visit type
    vt = str(row[col_idx['就诊方式']]) if row[col_idx['就诊方式']] else ''
    if vt and vt != 'None':
        visit_type_counter[vt] += 1
    
    st = str(row[col_idx['医疗类别']]) if row[col_idx['医疗类别']] else ''
    if st and st != 'None':
        settle_type_counter[st] += 1
    
    # Date
    dt = row[col_idx['结算时间']]
    if dt and hasattr(dt, 'strftime'):
        monthly[dt.strftime('%Y-%m')] += 1
        daily[dt.strftime('%Y-%m-%d')] += 1

wb.close()
print(f"\nTotal scanned: {count} rows")
print(f"降扎乡 records: {len(ji_zha)}")
print(f"Unique patients: {len(patient_records)}")

# === Top patients ===
patient_counts = sorted(patient_records.items(), key=lambda x: -len(x[1]))
top5 = patient_counts[:5]
top5_details = []
for pkey, recs in top5:
    detail = {'patient': pkey, 'total_visits': len(recs), 'visits': []}
    total_fee = 0
    for r in recs:
        dt = r[col_idx['结算时间']]
        dt_str = dt.strftime('%Y-%m-%d %H:%M') if dt and hasattr(dt, 'strftime') else str(dt)
        doc = str(r[col_idx['医师姓名']]) if r[col_idx['医师姓名']] else ''
        diag = str(r[col_idx['出院诊断名称']]) if r[col_idx['出院诊断名称']] else ''
        drug_fee = r[col_idx['药品费']] if '药品费' in col_idx else 0
        treat_fee = r[col_idx['诊疗费']] if '诊疗费' in col_idx else 0
        med_total = r[col_idx['医疗费总额']] if '医疗费总额' in col_idx else 0
        fund = r[col_idx['统筹基金支出']] if '统筹基金支出' in col_idx else 0
        
        med_total = float(med_total) if med_total and isinstance(med_total, (int, float)) else 0
        fund = float(fund) if fund and isinstance(fund, (int, float)) else 0
        drug_fee = float(drug_fee) if drug_fee and isinstance(drug_fee, (int, float)) else 0
        treat_fee = float(treat_fee) if treat_fee and isinstance(treat_fee, (int, float)) else 0
        
        total_fee += med_total
        detail['visits'].append({
            'time': dt_str,
            'doctor': doc,
            'diagnosis': diag,
            'drug_fee': drug_fee,
            'treat_fee': treat_fee,
            'total_fee': med_total,
            'fund_paid': fund
        })
    detail['total_fee'] = total_fee
    detail['avg_fee'] = total_fee / len(recs)
    top5_details.append(detail)

# === Doctor daily avg ===
doctor_dates = defaultdict(set)
for r in ji_zha:
    doc = str(r[col_idx['医师姓名']]) if r[col_idx['医师姓名']] else ''
    dt = r[col_idx['结算时间']]
    if doc and doc != 'None' and dt:
        doctor_dates[doc].add(dt.strftime('%Y-%m-%d') if hasattr(dt, 'strftime') else str(dt))

doctor_analysis = []
for doc, cnt in doctor_counter.most_common(15):
    dates = doctor_dates.get(doc, set())
    doctor_analysis.append({
        'name': doc,
        'total_records': cnt,
        'unique_days': len(dates),
        'avg_daily': round(cnt / len(dates), 1) if dates else 0
    })

# === Address: non-降扎 ===
non_jiangzha = []
for pkey, recs in patient_records.items():
    addrs = set()
    for r in recs:
        addr = str(r[col_idx['居住地址']]) if r[col_idx['居住地址']] else ''
        twp = str(r[col_idx['乡镇(街道)']]) if r[col_idx['乡镇(街道)']] else ''
        if addr and addr != 'None':
            addrs.add(addr)
        if twp and twp != 'None':
            addrs.add(f'[乡镇]{twp}')
    
    has_jiangzha = any('降扎' in a for a in addrs)
    if not has_jiangzha and addrs:
        non_jiangzha.append({
            'patient': pkey,
            'visits': len(recs),
            'addresses': list(addrs)
        })

# === New findings ===
new_findings = []
# F11: Doctor pattern
if doctor_analysis:
    top_doc = doctor_analysis[0]
    if top_doc['total_records'] > len(ji_zha) * 0.5:
        new_findings.append({
            'id': 'F11', 'level': 'P0',
            'finding': f"单一医生集中接诊：{top_doc['name']}接诊{top_doc['total_records']}条，占总量的{top_doc['total_records']/len(ji_zha)*100:.0f}%",
            'detail': doctor_analysis[:5]
        })

# F12: Zero diagnosis
diag_count = sum(1 for r in ji_zha if str(r[col_idx['出院诊断名称']]) in ('', 'None', '0'))
new_findings.append({
    'id': 'F12', 'level': 'P0',
    'finding': f"无诊断记录比例：{diag_count}/{len(ji_zha)}条（{diag_count/len(ji_zha)*100:.0f}%）",
    'diagnoses_found': diag_counter.most_common(10)
})

# F13: Fee structure
drug_total = fee_totals.get('药品费', 0)
total_all = fee_totals.get('医疗费总额', 1)
new_findings.append({
    'id': 'F13', 'level': 'P1',
    'finding': f"药品费占比{drug_total/total_all*100:.0f}%，诊疗费占比{fee_totals.get('诊疗费',0)/total_all*100:.0f}%，缺少诊疗服务费",
    'fee_breakdown': dict(fee_totals),
    'fee_buckets': dict(fee_buckets)
})

# F14: Address anomaly
if non_jiangzha:
    new_findings.append({
        'id': 'F14', 'level': 'P1',
        'finding': f"{len(non_jiangzha)}名患者居住地址不在降扎乡，累计就诊{sum(n['visits'] for n in non_jiangzha)}条",
        'non_jiangzha_patients': non_jiangzha[:15]
    })

# F15: Same-day patterns
same_day_patients = []
for pkey, recs in patient_records.items():
    date_counts = Counter()
    for r in recs:
        dt = r[col_idx['结算时间']]
        if dt and hasattr(dt, 'strftime'):
            date_counts[dt.strftime('%Y-%m-%d')] += 1
    multi_days = {d: c for d, c in date_counts.items() if c >= 2}
    if multi_days:
        same_day_patients.append({'patient': pkey, 'multi_days': multi_days, 'total_visits': len(recs)})

same_day_patients.sort(key=lambda x: -len(x['multi_days']))
new_findings.append({
    'id': 'F15', 'level': 'P1',
    'finding': f"{len(same_day_patients)}名患者存在同日多次就诊记录",
    'top_same_day': same_day_patients[:10]
})

# === Build output ===
result = {
    'basic_stats': {
        'total_records': len(ji_zha),
        'total_patients': len(patient_records),
        'avg_per_patient': round(len(ji_zha) / len(patient_records), 1),
        'monthly_distribution': sorted(monthly.items()),
        'top_days': sorted(daily.items(), key=lambda x: -x[1])[:20],
        'visit_types': dict(visit_type_counter),
        'settle_types': dict(settle_type_counter),
    },
    'doctor_analysis': doctor_analysis,
    'fee_analysis': {
        'totals': dict(fee_totals),
        'drug_pct': round(drug_total / total_all * 100, 1) if total_all else 0,
        'buckets': dict(fee_buckets),
        'fund_pay_rate': round(fee_totals.get('统筹基金支出', 0) / total_all * 100, 1) if total_all else 0,
    },
    'address_analysis': {
        'top_addresses': address_counter.most_common(30),
        'top_townships': township_counter.most_common(20),
        'non_jiangzha_count': len(non_jiangzha),
        'non_jiangzha_patients': non_jiangzha[:20],
    },
    'insurance_analysis': dict(insurance_counter.most_common()),
    'top5_patient_details': top5_details,
    'all_patients_ranked': [(p, len(r)) for p, r in patient_counts[:50]],
    'new_findings': new_findings,
}

out = r'C:\Users\scrccpa\.openclaw\workspace\temp_ocr\jz_deep_dive.json'
with open(out, 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2, default=str)
print(f"\nSaved: {out}")
print(f"Total records: {len(ji_zha)}, patients: {len(patient_records)}, doctors: {len(doctor_counter)}")
