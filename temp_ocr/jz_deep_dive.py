#!/usr/bin/env python3
# encoding: utf-8
import sys, os, json, traceback
from collections import defaultdict, Counter
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

FPATH = r"C:\Users\scrccpa\Desktop\若尔盖审计\若尔盖医保审计\2026年审计资料（医保局财务）\古英=2024-2025门诊、住院、个人账户、基金拨付明细\2025.xlsx"
OPATH = r"C:\Users\scrccpa\.openclaw\workspace\temp_ocr\jz_deep_dive.json"

def sf(val, d=0.0):
    if val is None: return d
    try:
        f = float(val)
        return f if f==f else d
    except: return d

def ss(val, d=""):
    if val is None: return d
    return str(val).strip()

def pd(val):
    if val is None: return None
    if isinstance(val, datetime): return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    from datetime import timedelta
    for fmt in ['%Y-%m-%d %H:%M:%S','%Y-%m-%d','%Y/%m/%d','%Y%m%d']:
        try: return datetime.strptime(s[:10], fmt).strftime('%Y-%m-%d')
        except: pass
    try:
        from datetime import timedelta
        return (datetime(1899,12,30)+timedelta(days=float(s))).strftime('%Y-%m-%d')
    except: pass
    return s[:10] if len(s)>=10 else None

def main():
    import openpyxl
    wb = openpyxl.load_workbook(FPATH, read_only=True, data_only=True)
    ws = wb['总表']
    print("Loading...")

    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue
        v = ss(row[28] if len(row)>28 else "")
        if v and '降扎' in v:
            rows.append(list(row))
        if (i+1) % 100000 == 0:
            print(f"  checked {i+1}, found {len(rows)}")
    wb.close()
    print(f"Total: {len(rows)} rows matching 降扎")

    if len(rows)==0:
        print("ZERO matches! Check col29 values...")
        wb2 = openpyxl.load_workbook(FPATH, read_only=True, data_only=True)
        ws2 = wb2['总表']
        samples=set()
        for i,row in enumerate(ws2.iter_rows(values_only=True)):
            if i==0: continue
            v=ss(row[28] if len(row)>28 else "")
            if v: samples.add(v)
            if len(samples)>=30: break
        wb2.close()
        print("Sample col29:", sorted(samples))
        sys.exit(1)

    print("Processing {} rows...".format(len(rows)))

    # 1. Basic stats
    pids=set(); pnames=set(); mcnt=defaultdict(int); dcnt=Counter()
    for r in rows:
        pid=ss(r[0] if len(r)>0 else "")
        pn=ss(r[1] if len(r)>1 else "")
        if pid: pids.add(pid)
        if pn: pnames.add(pn)
        d=pd(r[12] if len(r)>12 else "")
        if d:
            mcnt[d[:7]]+=1
            dcnt[d]+=1

    basic_stats = {
        "total_records": len(rows),
        "unique_patients": len(pids),
        "unique_names": len(pnames),
        "monthly": dict(sorted(mcnt.items())),
        "top20_dates": [{"date":d,"count":c} for d,c in dcnt.most_common(20)]
    }

    # 2. Doctor analysis
    docs = defaultdict(lambda: {"n":0,"dates":set()})
    for r in rows:
        did=ss(r[36] if len(r)>36 else "")
        dnm=ss(r[37] if len(r)>37 else "")
        k=did if did else dnm
        if not k: continue
        docs[k]["n"]+=1
        d=pd(r[12] if len(r)>12 else "")
        if d: docs[k]["dates"].add(d)

    total_doc_rec = sum(v["n"] for v in docs.values())
    doc_analysis = []
    for k,v in sorted(docs.items(), key=lambda x:-x[1]["n"]):
        nd=len(v["dates"])
        avg=round(v["n"]/nd,1) if nd>0 else 0
        pct=round(v["n"]/total_doc_rec*100,1) if total_doc_rec>0 else 0
        doc_analysis.append({
            "医师": k,
            "记录数": v["n"],
            "占比": str(pct)+"%",
            "出诊天数": nd,
            "日均接诊": avg
        })

    doc_summary = {}
    if doc_analysis:
        t=doc_analysis[0]
        doc_summary = {
            "医生数": len(doc_analysis),
            "总记录": total_doc_rec,
            "top1医生": t["医师"],
            "top1记录": t["记录数"],
            "top1占比": t["占比"],
            "一人包揽过半": t["记录数"] > total_doc_rec * 0.5
        }

    # 3. Fee analysis
    td=tt=tc=tb=tm=tp=ts=tca=0.0
    fb={"0-50":0,"50-100":0,"100-200":0,"200+":0}
    for r in rows:
        d=sf(r[43] if len(r)>43 else "")
        tr=sf(r[48] if len(r)>48 else "")
        co=sf(r[53] if len(r)>53 else "")
        be=sf(r[58] if len(r)>58 else "")
        me=sf(r[61] if len(r)>61 else "")
        po=sf(r[66] if len(r)>66 else "")
        sp=sf(r[83] if len(r)>83 else "")
        ca=sf(r[84] if len(r)>84 else "")
        td+=d; tt+=tr; tc+=co; tb+=be; tm+=me; tp+=po; ts+=sp; tca+=ca
        if me<=50: fb["0-50"]+=1
        elif me<=100: fb["50-100"]+=1
        elif me<=200: fb["100-200"]+=1
        else: fb["200+"]+=1

    tf=td+tt+tc+tb
    fee_analysis = {
        "药品费": round(td,2),
        "诊疗费": round(tt,2),
        "耗材费": round(tc,2),
        "床位费": round(tb,2),
        "药品占比": str(round(td/tf*100,1))+"%" if tf>0 else "0%",
        "诊疗占比": str(round(tt/tf*100,1))+"%" if tf>0 else "0%",
        "耗材占比": str(round(tc/tf*100,1))+"%" if tf>0 else "0%",
        "床位占比": str(round(tb/tf*100,1))+"%" if tf>0 else "0%",
        "医疗费总额": round(tm,2),
        "统筹支出": round(tp,2),
        "个人账户": round(ts,2),
        "现金支付": round(tca,2),
        "支付率": str(round(tp/tm*100,1))+"%" if tm>0 else "0%",
        "费用分布": fb
    }

    # 4. Address analysis
    pat_addr = {}
    pat_town = {}
    pat_cnt = Counter()
    pat_visits = defaultdict(list)

    for r in rows:
        pid=ss(r[0] if len(r)>0 else "")
        pn=ss(r[1] if len(r)>1 else "")
        k=pid if pid else pn
        if not k: continue
        if k not in pat_addr: pat_addr[k]=set()
        if k not in pat_town: pat_town[k]=set()
        a=ss(r[97] if len(r)>97 else "")
        t=ss(r[94] if len(r)>94 else "")
        if a: pat_addr[k].add(a)
        if t: pat_town[k].add(t)
        pat_cnt[k]+=1
        pat_visits[k].append({
            "name": pn,
            "date": pd(r[12] if len(r)>12 else ""),
            "fee": sf(r[61] if len(r)>61 else ""),
            "pool": sf(r[66] if len(r)>66 else ""),
            "drug": sf(r[43] if len(r)>43 else ""),
            "treat": sf(r[48] if len(r)>48 else ""),
            "doc": ss(r[37] if len(r)>37 else ""),
            "diag": ss(r[33] if len(r)>33 else "")
        })

    not_jz = []
    for k, towns in pat_town.items():
        if not towns: continue
        if all('降扎' not in t for t in towns):
            v=pat_cnt[k]
            if v>=5:
                not_jz.append({
                    "姓名": pat_visits[k][0]["name"],
                    "乡镇": list(towns),
                    "地址": list(pat_addr.get(k,[])),
                    "就诊次数": v
                })
    not_jz.sort(key=lambda x:-x["就诊次数"])

    n_town = sum(1 for t in pat_town.values() if t)
    n_not_jz = sum(1 for t in pat_town.values() if t and all('降扎' not in x for x in t))
    n_no_addr = sum(1 for a in pat_addr.values() if not a)

    address_analysis = {
        "患者总数": len(pat_cnt),
        "有乡镇信息": n_town,
        "非降扎乡": n_not_jz,
        "无地址": n_no_addr,
        "非降扎高频(5次+)": not_jz[:30]
    }

    # 5. Insurance type
    ic=Counter()
    for r in rows:
        i=ss(r[88] if len(r)>88 else "")
        ic[i if i else "(空)"]+=1

    insurance_analysis = {
        "分布": dict(ic.most_common()),
        "合计": sum(ic.values())
    }

    # 6. Top5 details
    top5 = pat_cnt.most_common(5)
    top5_details = []
    for pid, cnt in top5:
        vs=pat_visits[pid]
        nm=vs[0]["name"]
        dcnt2 = Counter(v["date"] for v in vs if v["date"])
        sd = {d:c for d,c in dcnt2.items() if c>=2}
        details=[]
        for v in vs:
            details.append({
                "时间": v["date"],
                "医疗费": round(v["fee"],2),
                "统筹": round(v["pool"],2),
                "药品费": round(v["drug"],2),
                "诊疗费": round(v["treat"],2),
                "医师": v["doc"],
                "诊断": v["diag"],
                "同日多次": True if v["date"] and v["date"] in sd else False
            })
        top5_details.append({
            "编号": pid,
            "姓名": nm,
            "就诊次数": cnt,
            "同日多次日期": sd,
            "逐条明细": details
        })

    # 7. Findings
    findings = []

    # Doctor concentration
    if doc_summary.get("一人包揽过半"):
        findings.append(
            "医生高度集中: %s 一人接诊 %d 条记录(占比 %s), 超过总接诊量一半, 存在一人统包风险。" %
            (doc_summary["top1医生"], doc_summary["top1记录"], doc_summary["top1占比"])
        )

    # High fee records
    high_fee=[]
    for r in rows:
        me=sf(r[61] if len(r)>61 else "")
        if me>10000:
            high_fee.append({
                "name": ss(r[1] if len(r)>1 else ""),
                "amt": me,
                "date": pd(r[12] if len(r)>12 else ""),
                "doc": ss(r[37] if len(r)>37 else ""),
                "diag": ss(r[33] if len(r)>33 else "")
            })
    if high_fee:
        high_fee.sort(key=lambda x:-x["amt"])
        findings.append(
            "高额费用: 共%d条记录医疗费超过1万元, 最高%.2f元(%s, %s, 诊断:%s)" %
            (len(high_fee), high_fee[0]["amt"], high_fee[0]["name"], high_fee[0]["date"], high_fee[0]["diag"])
        )

    # Cross-region visits
    if not_jz:
        tot_visits = sum(p["就诊次数"] for p in not_jz)
        findings.append(
            "跨区域就诊: %d名患者地址不在降扎乡且就诊>=5次, 累计%d次, 需核实是否属服务范围。" %
            (len(not_jz), tot_visits)
        )

    # Same-day multiple visits
    sd_issues=[]
    for pid,cnt in top5:
        vs=pat_visits[pid]
        dcnt2=Counter(v["date"] for v in vs if v["date"])
        sd={d:c for d,c in dcnt2.items() if c>=2}
        if sd:
            nm=vs[0]["name"]
            for d,c in sorted(sd.items()):
                sd_issues.append("%s(%s)于%s就诊%d次" % (nm, pid, d, c))
    if sd_issues:
        findings.append("同日多次就诊: " + ";".join(sd_issues[:10]))

    # Drug fee dominance
    if tf>0:
        dp=td/tf*100
        trp=tt/tf*100
        if dp>80:
            findings.append("费用结构异常: 药品费占比%.1f%%, 远高于诊疗费%.1f%%, 存在过度用药/虚开风险。" % (dp,trp))
        elif dp>60:
            findings.append("费用结构关注: 药品费占比%.1f%%, 诊疗费%.1f%%, 药品费用偏高。" % (dp,trp))

    # Payment rate
    if tm>0:
        pr=tp/tm*100
        if pr>90:
            findings.append("支付率过高: 统筹支付%.1f%%, 患者自付极低, 存在套取医保基金风险。" % pr)

    # Tiny fees
    tiny=sum(1 for r in rows if sf(r[61] if len(r)>61 else "")<10)
    if tiny>100:
        findings.append("小额记录: %d条记录(占%.1f%%)医疗费不足10元, 有刷卡套现风险。" % (tiny, tiny/len(rows)*100))

    print("Findings: %d" % len(findings))
    for f in findings:
        print("  - "+f)

    result = {
        "basic_stats": basic_stats,
        "doctor_analysis": doc_analysis,
        "doctor_summary": doc_summary,
        "fee_analysis": fee_analysis,
        "address_analysis": address_analysis,
        "insurance_analysis": insurance_analysis,
        "top5_patient_details": top5_details,
        "new_findings": findings
    }

    with open(OPATH, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print("\nSaved to", OPATH, "(", round(os.path.getsize(OPATH)/1024,1), "KB)")

if __name__=='__main__':
    try: main()
    except Exception as e:
        print("FATAL:", e)
        traceback.print_exc()
        sys.exit(1)
