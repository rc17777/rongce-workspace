"""
若尔盖医保审计 - 参保人数真实性深度核查
"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from collections import defaultdict

INSURED_DIR = r'C:\Users\scrccpa\Desktop\若尔盖审计\若尔盖医保审计\2026年审计资料（医保局财务）\2024-2025居民、职工参保名单'
BASE_DIR = r'C:\Users\scrccpa\Desktop\若尔盖审计\若尔盖医保审计\2026年审计资料（医保局财务）\古英=2024-2025门诊、住院、个人账户、基金拨付明细'

# ============================================================
# 1. Load resident insured lists
# ============================================================
print('='*60)
print('1. 城乡居民参保人数核查')
print('='*60)

resident = {}
for label, fname in [('2024', '2024年12月底城乡居民参保情况1.7.xlsx'), ('2025', '2025.年12月底61035人.xlsx')]:
    fp = os.path.join(INSURED_DIR, '城乡居民参保名单2024-2025', fname)
    if not os.path.exists(fp):
        print(f'{label}: NOT FOUND')
        continue
    wb = load_workbook(fp, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    headers = [str(h).strip() if h else '' for h in headers]
    
    id_col = name_col = gender_col = addr_col = status_col = category_col = None
    for i, h in enumerate(headers):
        if '身份证' in h or '证件' in h: id_col = i
        if '姓名' in h: name_col = i
        if '性别' in h: gender_col = i
        if '地址' in h or '户口' in h: addr_col = i
        if '身份' in h or '类别' in h: status_col = i
        if '乡镇' in h or '街道' in h: category_col = i
    
    ids = set()
    id_names = {}
    duplicates = []
    towns = defaultdict(int)
    categories = defaultdict(int)
    cnt = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        cnt += 1
        id_val = str(row[id_col]).strip().replace('\u3000','').replace(' ','') if id_col and row[id_col] else ''
        name_val = str(row[name_col]).strip() if name_col and row[name_col] else ''
        category_val = str(row[category_col]).strip() if category_col and row[category_col] else ''
        
        if not id_val: continue
        
        if id_val in ids:
            duplicates.append((id_val, name_val, id_names.get(id_val, '')))
        else:
            ids.add(id_val)
            id_names[id_val] = name_val
        
        if category_val: towns[category_val] += 1
        if status_col and row[status_col]:
            s = str(row[status_col]).strip()
            if s: categories[s] += 1
    
    resident[label] = {'ids': ids, 'count': len(ids), 'total_rows': cnt, 'duplicates': duplicates, 
                        'towns': towns, 'categories': categories}
    print(f'{label}年末: Excel{cnt}行 → 去重{len(ids)}人 | 重复{len(duplicates)}个')
    if duplicates:
        print(f'  重复身份证示例（前5）:')
        for d in duplicates[:5]:
            print(f'    {d[0]} | {d[1]} vs {d[2]}')
    
    # Top towns
    top_towns = sorted(towns.items(), key=lambda x: -x[1])[:10]
    print(f'  乡镇TOP5: {", ".join(f"{t}({c})" for t,c in top_towns[:5])}')
    wb.close()

# ============================================================
# 2. Load employee insured lists
# ============================================================
print('\n' + '='*60)
print('2. 职工参保人数核查')
print('='*60)

employee = {}
for year in ['2024', '2025']:
    fp = os.path.join(INSURED_DIR, '职工参保人员清单', f'（{year}年）单位人员缴费明细查询20260617172604304_1.xlsx')
    if year == '2025':
        fp = os.path.join(INSURED_DIR, '职工参保人员清单', f'（{year}年）单位人员缴费明细查询20260617173052937_1.xlsx')
    
    if not os.path.exists(fp):
        print(f'{year}: NOT FOUND at {fp}')
        continue
    
    print(f'Loading {year} 职工...', end=' ', flush=True)
    wb = load_workbook(fp, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    headers = [str(h).strip() if h else '' for h in headers]
    print(f'cols: {headers[:12]}')
    
    id_col = name_col = unit_col = None
    for i, h in enumerate(headers):
        if '身份证' in h or '证件' in h: id_col = i
        if '姓名' in h: name_col = i
        if '单位' in h: unit_col = i
    
    if id_col is None:
        print('  WARNING: No ID column found!')
        wb.close()
        continue
    
    ids = set()
    units = defaultdict(int)
    cnt = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        cnt += 1
        if cnt % 50000 == 0: print(f'{cnt//1000}k...', end=' ', flush=True)
        id_val = str(row[id_col]).strip().replace('\u3000','').replace(' ','') if row[id_col] else ''
        if id_val and len(id_val) >= 15:
            ids.add(id_val)
            if unit_col and row[unit_col]:
                u = str(row[unit_col]).strip()
                if u: units[u] += 1
    
    employee[year] = {'ids': ids, 'count': len(ids), 'total_rows': cnt, 'units': units}
    print(f'  去重参保人数: {len(ids)} (原始{cnt}行)')
    print(f'  涉及单位: {len(units)}个')
    top_units = sorted(units.items(), key=lambda x: -x[1])[:5]
    print(f'  单位TOP5: {", ".join(f"{u}({c})" for u,c in top_units)}')
    wb.close()

# ============================================================
# 3. Cross-reference: duplicates between resident & employee
# ============================================================
print('\n' + '='*60)
print('3. 居民-职工重复参保核查')
print('='*60)

for year in ['2024', '2025']:
    res_ids = resident.get(year, {}).get('ids', set())
    emp_ids = employee.get(year, {}).get('ids', set())
    
    if res_ids and emp_ids:
        both = res_ids & emp_ids
        print(f'{year}年: 居民{len(res_ids)} + 职工{len(emp_ids)} = 去重合并{len(res_ids | emp_ids)}')
        print(f'  重复参保(同时居民+职工): {len(both)} 人')
        if year == '2025':
            total_unique = len(res_ids | emp_ids)
            print(f'  2025年末 实际参保总人数(去重): {total_unique}')
            print(f'  医保局口径(61035居民 + 职工)  vs 实际去重 {total_unique}')

# ============================================================
# 4. Year-over-year comparison
# ============================================================
print('\n' + '='*60)
print('4. 年度变动核查')
print('='*60)

res_2024 = resident.get('2024', {}).get('ids', set())
res_2025 = resident.get('2025', {}).get('ids', set())

if res_2024 and res_2025:
    stayed = res_2024 & res_2025
    added = res_2025 - res_2024
    removed = res_2024 - res_2025
    
    print(f'2024→2025 城乡居民变动:')
    print(f'  2024年末: {len(res_2024)}')
    print(f'  2025年末: {len(res_2025)}')
    print(f'  净变动: {len(res_2025) - len(res_2024):+d}')
    print(f'  持续参保: {len(stayed)}')
    print(f'  新增参保: {len(added)} ({len(added)/len(res_2024)*100:.1f}%)')
    print(f'  退出参保: {len(removed)} ({len(removed)/len(res_2024)*100:.1f}%)')

# ============================================================
# 5. Check settlement records for "死亡" / deceased indicators
# ============================================================
print('\n' + '='*60)
print('5. 死亡人口线索核查（从结算数据反查）')
print('='*60)

death_keywords = ['死亡', '丧葬', '抚恤', '遗体', '殡葬', '尸体']
death_records = []

for year in ['2024', '2025']:
    fp = os.path.join(BASE_DIR, f'{year}.xlsx')
    if not os.path.exists(fp): continue
    wb = load_workbook(fp, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    headers = [str(h).strip() if h else '' for h in headers]
    
    id_col = name_col = diag_col = date_col = None
    for i, h in enumerate(headers):
        if '证件' in h: id_col = i
        if '姓名' in h: name_col = i
        if '诊断' in h: diag_col = i
        if '结算' in h: date_col = i
    
    found_ids = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        diag = str(row[diag_col]).strip() if diag_col and row[diag_col] else ''
        if any(kw in diag for kw in death_keywords):
            id_val = str(row[id_col]).strip() if id_col and row[id_col] else ''
            name_val = str(row[name_col]).strip() if name_col and row[name_col] else ''
            date_val = str(row[date_col])[:10] if date_col and row[date_col] else ''
            if id_val and id_val not in found_ids:
                found_ids.add(id_val)
                death_records.append({'year': year, 'id': id_val[-4:], 'name': name_val, 
                                       'date': date_val, 'diag': diag[:40]})
    wb.close()

print(f'含死亡/丧葬相关诊断的结算记录: {len(death_records)} 人')
if death_records:
    print('  示例（前10）:')
    for r in death_records[:10]:
        print(f'    {r["year"]} | {r["name"]} | {r["date"]} | {r["diag"]}')

# Check if these deceased people are still in 2025 insured list
if res_2025:
    still_insured_deceased = 0
    for r in death_records:
        # We need full ID, but we only stored last 4 digits for privacy
        pass  # Can't check with truncated IDs
    
    # Alternative: check full IDs from settlement vs insured
    full_death_ids = set()
    for year in ['2024', '2025']:
        fp = os.path.join(BASE_DIR, f'{year}.xlsx')
        if not os.path.exists(fp): continue
        wb = load_workbook(fp, read_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        headers = [str(h).strip() if h else '' for h in headers]
        
        id_col = name_col = diag_col = None
        for i, h in enumerate(headers):
            if '证件' in h: id_col = i
            if '姓名' in h: name_col = i
            if '诊断' in h: diag_col = i
        
        found = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            diag = str(row[diag_col]).strip() if diag_col and row[diag_col] else ''
            if any(kw in diag for kw in death_keywords):
                id_val = str(row[id_col]).strip().replace(' ','') if id_col and row[id_col] else ''
                if id_val and len(id_val) >= 15:
                    full_death_ids.add(id_val)
                    found += 1
        wb.close()
    
    still_insured = full_death_ids & res_2025
    print(f'\n死亡结算记录中仍在2025参保名单: {len(still_insured)} 人')
    if still_insured:
        print('  🚨 P0-死亡未注销-仍在参保!')
        for did in list(still_insured)[:10]:
            print(f'    {did[-6:]}...')

# ============================================================
# Summary
# ============================================================
print('\n' + '='*60)
print('参保核查总结')
print('='*60)
print(f'2024城乡居民: {resident.get("2024",{}).get("count",0)} 人')
print(f'2025城乡居民: {resident.get("2025",{}).get("count",0)} 人')
print(f'2024职工: {employee.get("2024",{}).get("count",0)} 人')
print(f'2025职工: {employee.get("2025",{}).get("count",0)} 人')
if res_2025 and employee.get('2025', {}).get('ids'):
    total = len(res_2025 | employee['2025']['ids'])
    print(f'2025末实际参保(去重): {total} 人')
print(f'死亡结算记录: {len(death_records)} 人')
print('Done.')
