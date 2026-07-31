"""Quick overview of 古英 settlement data 2023-2025."""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

base = r'C:\Users\scrccpa\Desktop\若尔盖审计\若尔盖医保审计\2026年审计资料（医保局财务）\古英=2024-2025门诊、住院、个人账户、基金拨付明细'

files = ['2023.xlsx', '2024.xlsx', '2025.xlsx']

for fname in files:
    fp = os.path.join(base, fname)
    if not os.path.exists(fp):
        print(f'{fname}: NOT FOUND')
        continue
    
    print(f'\n{"="*70}')
    print(f'📊 {fname} ({os.path.getsize(fp)/1024/1024:.0f}MB)')
    print(f'{"="*70}')
    
    wb = load_workbook(fp, read_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = ws.max_row or 0
        cols = ws.max_column or 0
        
        # Get headers (row 1)
        headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        
        # Find key column indices
        col_map = {}
        for i, h in enumerate(headers):
            if h:
                col_map[str(h).strip()] = i
        
        # Sample first data row for reference
        row1 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        
        print(f'\n  Sheet: {sn}')
        print(f'  Rows: {rows:,}  |  Cols: {cols}')
        
        # Find key fields
        key_fields = ['医疗类别', '结算时间', '医疗费总额', '医保支付金额', 
                      '统筹基金支出', '个人现金支付', '医药机构名称', '是否异地就医']
        for kf in key_fields:
            found = False
            for h, i in col_map.items():
                if kf in h:
                    print(f'  [{kf}] col={i}: {h}')
                    found = True
                    break
            if not found:
                print(f'  [{kf}] NOT FOUND')
        
        # Compute stats in streaming pass
        count = 0
        total_fee = 0.0
        total_pay = 0.0
        medical_types = {}
        institutions = {}
        dates_min = None
        dates_max = None
        yidi_count = 0
        
        fee_col = None
        pay_col = None
        type_col = None
        inst_col = None
        date_col = None
        yidi_col = None
        
        for h, i in col_map.items():
            if '医疗费总额' in h: fee_col = i
            if '医保支付金额' in h: pay_col = i
            if '医疗类别' in h: type_col = i
            if '医药机构名称' in h: inst_col = i
            if '结算时间' in h: date_col = i
            if '是否异地就医' in h: yidi_col = i
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            count += 1
            if count % 50000 == 0:
                print(f'    Processing... {count:,}')
            
            # Fee
            v = row[fee_col] if fee_col is not None else None
            if v and isinstance(v, (int, float)):
                total_fee += v
            
            # Pay
            v = row[pay_col] if pay_col is not None else None
            if v and isinstance(v, (int, float)):
                total_pay += v
            
            # Medical type
            v = row[type_col] if type_col is not None else None
            if v:
                vt = str(v).split('|')[-1].strip() if '|' in str(v) else str(v)
                medical_types[vt] = medical_types.get(vt, 0) + 1
            
            # Institution
            v = row[inst_col] if inst_col is not None else None
            if v:
                vt = str(v).strip()
                institutions[vt] = institutions.get(vt, 0) + 1
            
            # Date range
            v = row[date_col] if date_col is not None else None
            if v and hasattr(v, 'year'):
                if dates_min is None or v < dates_min: dates_min = v
                if dates_max is None or v > dates_max: dates_max = v
            
            # Yidi
            v = row[yidi_col] if yidi_col is not None else None
            if v and str(v).strip() == '是':
                yidi_count += 1
        
        print(f'\n  📈 汇总统计:')
        print(f'  总记录数: {count:,}')
        print(f'  医疗费总额: ¥{total_fee:,.0f}')
        print(f'  医保支付总额: ¥{total_pay:,.0f}')
        print(f'  次均费用: ¥{total_fee/count:,.0f}' if count else '')
        if dates_min and dates_max:
            print(f'  日期范围: {dates_min.date()} ~ {dates_max.date()}')
        print(f'  异地就医: {yidi_count:,} 条 ({yidi_count/count*100:.1f}%)' if count else '')
        
        print(f'\n  🏥 医疗类别分布:')
        for k, v in sorted(medical_types.items(), key=lambda x: -x[1])[:6]:
            print(f'    {k}: {v:,} ({v/count*100:.1f}%)')
        
        print(f'\n  🏨 机构分布 (Top 15):')
        for k, v in sorted(institutions.items(), key=lambda x: -x[1])[:15]:
            print(f'    {k}: {v:,}')
    
    wb.close()

print('\nDone.')
