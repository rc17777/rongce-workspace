"""Update Excel with new plan section findings"""
import sys, json
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Load existing
wb = load_workbook(r'C:\Users\scrccpa\Desktop\若尔盖医保审计_实施方案对照.xlsx')

hdr_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='0A1F3F')
done_fill = PatternFill('solid', fgColor='C8E6C9')
partial_fill = PatternFill('solid', fgColor='FFF9C4')
not_done_fill = PatternFill('solid', fgColor='FFE0E0')
body_font = Font(name='Microsoft YaHei', size=10)
bold_font = Font(name='Microsoft YaHei', size=10, bold=True)
title_font = Font(name='Microsoft YaHei', size=14, bold=True, color='0A1F3F')
wrap = Alignment(wrap_text=True, vertical='top')
center = Alignment(horizontal='center', vertical='top', wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Update Sheet 1 - add new items
ws1 = wb['实施方案对照总表']

# Find the last row
last_row = ws1.max_row
# Add new findings below
new_items = [
    ['三(二)基金支付', '核查跨渠道重复报销', '同日跨机构就医重复报销', '✅ 已完成',
     '发现1,544名患者同日跨机构就诊(同日不同机构)。典型案例:足巴(3014)同日频繁在成都三医院+圆心妙手大药房购药各自报销¥7K-24K。索郎彭措(2910):三渠道合计获得¥46.3万(统筹¥26.5万+大病¥12.9万+救助¥6.9万)，共82条记录。',
     'S09(新增)', '需进一步核实是否为合理诊治+确认是否需要退回重复报销部分'],

    ['三(二)基金支付', '核查多基金渠道使用', '同一患者使用多基金渠道', '✅ 已完成',
     '3,542名患者使用了2种及以上基金渠道(统筹+大病+救助+公务员补助)。索郎彭措¥46.3万(统筹+大病+救助)、供波扎西¥38.2万、罗洪¥35.1万。共10人超¥30万。',
     'S10(新增)', '需核实多渠道支出是否符合政策规定，是否存在"一条费用多渠道报销"'],

    ['三(二)基金支付', '异地就医结算', '异地就医费用与流向', '✅ 已完成',
     '228,829条异地就医记录(占总数54.6%)，费用¥1.78亿，统筹支出¥7,370万。主要流向:成都市三医院(8,234条)、都江堰市人民医院(7,926条)、省医院(6,579条)、郫都区(6,437条)、阿坝州林业中心医院(5,972条)。另有跨省:甘肃迭部县(1,669条药店)。',
     'S11(新增)', '异常:54.6%异地就医率偏高，需确认是否为若尔盖县医疗服务能力不足导致的正常外流，还是存在异地就医套现'],

    ['三(二)基金支付', '死亡人员账户清退', '死亡后仍有个账消费', '未开展',
     '需要死亡人口数据，暂无法执行。',
     '-', '需取得2024-2025年死亡人口名单+个账消费明细比对'],

    ['三(三)定点机构', '挂床住院(住院天数≤1)', '一日住院/分解住院', '🟡 部分完成',
     '发现644条住院天数≤1天(2.6%)，单次最高¥19,224。集中发生在省医院、成都三医院、华西等三甲机构，多为检查/日间手术类。需区分"正常日间手术"与"虚假住院"。',
     'S12(新增)', '需取得病案首页+日间手术目录确认是否为正常日间手术'],

    ['三(三)定点机构', '全县机构全量风险排序', '以金额/诊断缺失/患者规模排名', '✅ 已完成',
     '全县Top20机构已排序。新发现:①若尔盖县康乐康康宁大药房:27,908条(全县最多!)¥229万，100%无诊断 — 比金世康还大70%。②县医院:40,562条¥2,822万，78.3%无诊断。③藏医院:14,441条，91.6%无诊断。诊断缺失是全县普遍问题(75-100%)。',
     'S13(新增)', '康宁大药房应列为第三家重点延伸药店'],
]
for ri, row in enumerate(new_items):
    r = last_row + 1 + ri
    for ci, val in enumerate(row, 1):
        ws1.cell(row=r, column=ci, value=val).font = body_font
        ws1.cell(row=r, column=ci).alignment = wrap
        ws1.cell(row=r, column=ci).border = thin_border
    status = str(row[3])
    if '已完成' in status:
        for ci in range(1, len(row)+1):
            ws1.cell(row=r, column=ci).fill = done_fill
    elif '部分完成' in status:
        for ci in range(1, len(row)+1):
            ws1.cell(row=r, column=ci).fill = partial_fill
    elif '未开展' in status:
        for ci in range(1, len(row)+1):
            ws1.cell(row=r, column=ci).fill = not_done_fill

# ===== NEW Sheet: 全县概览 =====
ws_new = wb.create_sheet('全县概览')
h = ['项目', '数值', '说明']
w = [30, 24, 50]
sr = 1
for i, (hh, ww) in enumerate(zip(h, w), 1):
    ws_new.cell(row=sr, column=i, value=hh).font = hdr_font
    ws_new.cell(row=sr, column=i).fill = hdr_fill
    ws_new.cell(row=sr, column=i).alignment = center
    ws_new.cell(row=sr, column=i).border = thin_border
    ws_new.column_dimensions[get_column_letter(i)].width = ww
sr += 1

overview = [
    ('总记录数(2024-2025)', '419,378条', '2024年204,113条 + 2025年215,265条'),
    ('总费用', '¥243,390,716', '医疗费总额'),
    ('统筹基金支出', '¥104,677,620', '医保基金池支付'),
    ('大病保险支出', '¥6,171,429', ''),
    ('医疗救助支出', '¥3,032,022', ''),
    ('个人账户支出', '¥46,303,533', '职工个账'),
    ('个人现金支出', '¥77,222,248', '患者自付'),
    ('机构总数', '12,662家', '含医院/卫生院/药店/诊所'),
    ('普通住院', '22,282条 ¥1.54亿', '占总额63.3%'),
    ('普通门诊', '148,014条 ¥2,669万', '占总额11.0%'),
    ('定点药店购药', '197,406条 ¥2,020万', '占总额8.3%，统筹支出¥0'),
    ('门诊慢特病', '20,674条 ¥1,289万', '占总额5.3%'),
    ('异地就医', '228,829条 ¥1.78亿(54.6%)', '过半就诊在县外，主要流向成都/都江堰'),
    ('住院≤1天(疑似挂床)', '644条(2.6%)', '集中三甲医院，需区分日间手术'),
    ('同日跨机构就诊', '1,544人', '同日不同机构各自报销'),
    ('多基金渠道使用', '3,542人', '统等+大病+救助等多渠道'),
    ('无诊断比例(代表性机构)', '县医院78%/藏医院92%/郫都区95%', '全县普遍问题'),
]
for row in overview:
    for ci, val in enumerate(row, 1):
        ws_new.cell(row=sr, column=ci, value=val).font = body_font
        ws_new.cell(row=sr, column=ci).border = thin_border
        ws_new.cell(row=sr, column=ci).alignment = wrap
    sr += 1

# ===== NEW Sheet: 新增发现 =====
ws_new2 = wb.create_sheet('新增发现(S09-S13)')
h2 = ['编号', '等级', '对应方案条款', '发现描述', '核心数据', '取证建议']
w2 = [8, 10, 22, 40, 42, 36]
sr = 1
for i, (hh, ww) in enumerate(zip(h2, w2), 1):
    ws_new2.cell(row=sr, column=i, value=hh).font = hdr_font
    ws_new2.cell(row=sr, column=i).fill = hdr_fill
    ws_new2.cell(row=sr, column=i).alignment = center
    ws_new2.cell(row=sr, column=i).border = thin_border
    ws_new2.column_dimensions[get_column_letter(i)].width = ww
sr += 1

new_findings = [
    ['S09', '🟡 P1', '三(二)同日跨机构就诊',
     '1,544名患者同一天在不同机构各自刷医保报销',
     '足巴(3014):2024-2025年多次同日成都三医院+圆心妙手大药房，各自报销¥7K-¥24K。共计19条同日跨机构记录。典型日期:12/05同日三机构(三医院+圆心妙手+国药西南)3条¥30K。',
     '调取足巴全部就诊记录和处方，确认是否为同一疾病在同日多处重复开药报销'],

    ['S10', '🟡 P1', '三(二)多基金渠道',
     '3,542名患者使用了2+种基金渠道，10人累计超¥30万',
     '索郎彭措:统筹¥26.5万+大病¥12.9万+救助¥6.9万=¥46.3万(82条记录)。供波扎西:统筹¥22.6万+大病¥12.8万+救助¥2.8万=¥38.2万。',
     '核实多渠道路径是否符合"先统筹、后大病、再救助"的次序报销原则，排查重复报销'],

    ['S11', '🟡 P1', '三(二)异地就医',
     '54.6%就诊在县外，异地费用¥1.78亿',
     '主要流向成都(三医院8,234条/都江堰7,926条/省医院6,579条/郫都区6,437条)。跨省:甘肃迭部县1,669条(药店)。',
     '①分析异地就医人次vs本地就医人次趋势 ②筛查甘肃药店异地购药真实性 ③确认54.6%异地率是否高于同类县'],

    ['S12', '🟡 P2', '三(三)挂床住院',
     '644条住院≤1天(2.6%)，单次最高¥19,224',
     '省医院白玛泽仁¥19,224/1天;成都三医院袁翠华¥12,992/1天。集中在三甲，可能为日间手术/检查。需病案首页区分。',
     '①取得644条病案首页确认入院出院诊断 ②确认是否属于日间手术目录范围 ③非日间手术的列为挂床住院疑点'],

    ['S13', '🔴 P0', '三(三)全县风险排序',
     '康宁大药房27,908条为全县最多记录，超金世康70%',
     '全县Top药店:康宁大药房27,908条¥229万(100%无诊断,个账) > 福康大药房(？) > 金世康16,440条¥126万。Top医院:县医院40,562条¥2,822万(78%无诊断) > 藏医院14,441条(92%无诊断)。',
     '康宁大药房应列为第三家重点延伸药店，参照金世康S05取证方案(进销存+监控+访谈)'],
]
for row in new_findings:
    for ci, val in enumerate(row, 1):
        ws_new2.cell(row=sr, column=ci, value=val).font = body_font
        ws_new2.cell(row=sr, column=ci).alignment = wrap
        ws_new2.cell(row=sr, column=ci).border = thin_border
    lv = str(row[1])
    if 'P0' in lv:
        for ci in range(1, len(row)+1):
            ws_new2.cell(row=sr, column=ci).fill = PatternFill('solid', fgColor='FFE0E0')
    elif 'P1' in lv:
        for ci in range(1, len(row)+1):
            ws_new2.cell(row=sr, column=ci).fill = PatternFill('solid', fgColor='FFF9C4')
    sr += 1

# ===== Update Sheet 2 (统计摘要) =====
ws2 = wb['统计摘要']
# Recalculate with new items
ws2.cell(row=4, column=2, value=15)  # Done count
ws2.cell(row=4, column=3, value='41%')
ws2.cell(row=5, column=2, value=4)   # Partial
ws2.cell(row=5, column=3, value='11%')
ws2.cell(row=6, column=2, value=18)  # Not done
ws2.cell(row=6, column=3, value='49%')

# ===== NEW Sheet: 重复报销重点案例 =====
ws_new3 = wb.create_sheet('重复报销重点案例')
h3 = ['患者', '总基金支出', '统筹基金', '大病保险', '医疗救助', '总记录数', '同日跨机构天数', '重点机构', '疑点']
w3 = [18, 16, 14, 12, 12, 10, 14, 42, 36]
sr = 1
for i, (hh, ww) in enumerate(zip(h3, w3), 1):
    ws_new3.cell(row=sr, column=i, value=hh).font = hdr_font
    ws_new3.cell(row=sr, column=i).fill = hdr_fill
    ws_new3.cell(row=sr, column=i).alignment = center
    ws_new3.cell(row=sr, column=i).border = thin_border
    ws_new3.column_dimensions[get_column_letter(i)].width = ww
sr += 1

cases = [
    ['足巴(3014)', '¥202,264', '¥191,869', '¥34,656', '¥54,115', 19, '多次',
     '成都三医院 + 圆心妙手大药房 + 国药西南', '同日三机构报销¥30K，药店+医院同日并行'],
    ['索郎彭措(2910)', '¥463,022', '¥265,383', '¥128,995', '¥68,644', 82, '-',
     '县医院 + 藏医院 +多药店', '三渠道累计¥46万，82条记录'],
    ['供波扎西(2510)', '¥382,149', '¥226,062', '¥128,416', '¥27,671', 10, '-',
     '多机构', '大病¥12.8万居个人最高'],
    ['罗洪(0017)', '¥351,087', '¥217,667', '¥74,023', '¥59,398', 21, '-',
     '多机构', '医疗救助¥5.9万为个人最高'],
    ['李中盖(2212)', '¥273,239', '¥273,239', '¥0', '¥0', 145, '-',
     '多机构', '145条记录全走统筹(可能慢性病持续治疗)'],
    ['刘冬梅(002X)', '¥241,089', '¥241,089', '¥0', '¥0', 47, '-',
     '多机构', '¥24万全统筹'],
    ['蒙福兰(002X)', '¥230,720', '¥230,720', '¥0', '¥0', 162, '-',
     '多机构', '162条记录多机构分散,蚂蚁搬家模式'],
    ['杨军(0016)', '¥241,703', '¥215,681', '¥26,022', '¥0', 122, '-',
     '多机构', '统筹+大病双渠道，需核实'],
]
for row in cases:
    for ci, val in enumerate(row, 1):
        ws_new3.cell(row=sr, column=ci, value=val).font = body_font
        ws_new3.cell(row=sr, column=ci).alignment = wrap
        ws_new3.cell(row=sr, column=ci).border = thin_border
    sr += 1

# Save
wb.save(r'C:\Users\scrccpa\Desktop\若尔盖医保审计_实施方案对照.xlsx')
print('Updated!')
