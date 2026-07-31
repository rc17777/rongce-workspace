"""Deep dive into 金世康药品店 all records"""
import sys, os, json
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from collections import Counter, defaultdict
from datetime import datetime

DATA = r'C:\Users\scrccpa\Desktop\若尔盖审计\若尔盖医保审计\2026年审计资料（医保局财务）\古英=2024-2025门诊、住院、个人账户、基金拨付明细'

all_records = []

for year in ['2024', '2025']:
    fp = os.path.join(DATA, f'{year}.xlsx')
    print(f'Loading {year}...')
    wb = load_workbook(fp, read_only=True, data_only=True)
    
    for sn in wb.sheetnames:
        ws = wb[sn]
        headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        col = {h: i for i, h in enumerate(headers)}
        inst_col = col.get('医药机构名称', col.get('医疗机构名称', None))
        if inst_col is None:
            continue
        
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            inst = str(row[inst_col]) if row[inst_col] else ''
            if '金世康' not in inst:
                continue
            count += 1
            
            dt = row[col.get('结算时间')]
            dt_str = dt.strftime('%Y-%m-%d %H:%M:%S') if dt and hasattr(dt, 'strftime') else str(dt) if dt else ''
            date_str = dt.strftime('%Y-%m-%d') if dt and hasattr(dt, 'strftime') else ''
            hour = dt.hour if dt and hasattr(dt, 'hour') else -1
            weekday = dt.strftime('%A') if dt and hasattr(dt, 'strftime') else ''
            
            rec = {
                'year': year,
                'sheet': sn,
                'time': dt_str,
                'date': date_str,
                'hour': hour,
                'weekday': weekday,
                'name': str(row[col['姓名']]) if col.get('姓名') and row[col['姓名']] else '',
                'cert': str(row[col['证件号码']]) if col.get('证件号码') and row[col['证件号码']] else '',
                'dept': inst,
                'drug_fee': float(row[col['药品费']]) if col.get('药品费') and row[col['药品费']] and isinstance(row[col['药品费']], (int,float)) else 0,
                'treat_fee': float(row[col.get('诊疗费', -1)]) if col.get('诊疗费') and row[col.get('诊疗费')] and isinstance(row[col.get('诊疗费')], (int,float)) else 0,
                'total_fee': float(row[col['医疗费总额']]) if col.get('医疗费总额') and row[col['医疗费总额']] and isinstance(row[col['医疗费总额']], (int,float)) else 0,
                'fund_paid': float(row[col['统筹基金支出']]) if col.get('统筹基金支出') and row[col['统筹基金支出']] and isinstance(row[col['统筹基金支出']], (int,float)) else 0,
                'personal_acct': float(row[col['个人账户支付']]) if col.get('个人账户支付') and row[col['个人账户支付']] and isinstance(row[col['个人账户支付']], (int,float)) else 0,
                'personal_cash': float(row[col['个人现金支付']]) if col.get('个人现金支付') and row[col['个人现金支付']] and isinstance(row[col['个人现金支付']], (int,float)) else 0,
                'address': str(row[col['居住地址']]) if col.get('居住地址') and row[col['居住地址']] else '',
                'township': str(row[col['乡镇(街道)']]) if col.get('乡镇(街道)') and row[col['乡镇(街道)']] else '',
                'visit_type': str(row[col.get('就诊方式','')]) if col.get('就诊方式') else '',
                'med_type': str(row[col.get('医疗类别','')]) if col.get('医疗类别') else '',
                'ins_type': str(row[col['参保类型']]) if col.get('参保类型') and row[col['参保类型']] else '',
            }
            all_records.append(rec)
        
        print(f'  {sn}: {count} records')
    
    wb.close()

print(f'\nTotal: {len(all_records)} records')

# ===== ANALYSIS =====

# 1. Patient aggregation
patients = defaultdict(list)
for r in all_records:
    cert4 = r['cert'][-4:] if r['cert'] and len(r['cert']) >= 4 else r['cert']
    pk = f"{r['name']}({cert4})"
    patients[pk].append(r)

# 2. Top patients - detailed
top_patients = sorted(patients.items(), key=lambda x: -len(x[1]))[:20]
top_details = []
for pk, recs in top_patients:
    years_cnt = Counter(r['year'] for r in recs)
    monthly_cnt = Counter(r['date'][:7] for r in recs if r['date'])
    total_fee = sum(r['total_fee'] for r in recs)
    total_fund = sum(r['fund_paid'] for r in recs)
    dates = [r['date'] for r in recs if r['date']]
    day_multi = {d: c for d, c in Counter(dates).items() if c >= 2}
    
    top_details.append({
        'patient': pk,
        'visits': len(recs),
        'years': dict(years_cnt),
        'total_fee': round(total_fee, 1),
        'total_fund': round(total_fund, 1),
        'avg_fee': round(total_fee/len(recs), 1),
        'same_day_count': len(day_multi),
        'same_day_details': day_multi,
        'addresses': list(set(r['address'] for r in recs if r['address'] and r['address'].strip()))[:5],
        'townships': list(set(r['township'] for r in recs if r['township'] and r['township'].strip()))[:5],
        'sample_visits': [
            {'date': r['date'], 'time': r['time'], 'total': r['total_fee'], 'fund': r['fund_paid']}
            for r in recs[:10]
        ]
    })

# 3. Hourly distribution
hourly = Counter(r['hour'] for r in all_records if r['hour'] >= 0)

# 4. Weekday distribution
weekdays = Counter(r['weekday'] for r in all_records if r['weekday'])

# 5. Monthly trend by year
monthly_year = defaultdict(Counter)
for r in all_records:
    if r['date']:
        monthly_year[r['year']][r['date'][:7]] += 1

# 6. Fee analysis
fee_buckets = Counter()
for r in all_records:
    f = r['total_fee']
    if f <= 50: fee_buckets['0-50'] += 1
    elif f <= 100: fee_buckets['50-100'] += 1
    elif f <= 200: fee_buckets['100-200'] += 1
    elif f <= 500: fee_buckets['200-500'] += 1
    else: fee_buckets['500+'] += 1

# 7. Insurance types
ins_types = Counter(r['ins_type'] for r in all_records if r['ins_type'] and r['ins_type'] != 'None')

# 8. Address analysis
addresses = Counter(r['address'] for r in all_records if r['address'] and r['address'].strip() and r['address'] != 'None')
townships = Counter(r['township'] for r in all_records if r['township'] and r['township'].strip() and r['township'] != 'None')

# 9. Super-heavy day detail: 2025-11-10 (126 visits)
day_1110 = [r for r in all_records if r['date'] == '2025-11-10']
day_1110_detail = {
    'total': len(day_1110),
    'patients': len(set(r['cert'] for r in day_1110)),
    'hourly': dict(Counter(r['hour'] for r in day_1110)),
    'fee_range': f"{min(r['total_fee'] for r in day_1110):.1f} - {max(r['total_fee'] for r in day_1110):.1f}",
    'avg_fee': round(sum(r['total_fee'] for r in day_1110)/len(day_1110), 1),
    'total_fee': round(sum(r['total_fee'] for r in day_1110), 1),
    'total_fund': round(sum(r['fund_paid'] for r in day_1110), 1),
    'sample_records': [
        {'name': r['name'], 'cert': r['cert'][-4:], 'time': r['time'], 'fee': r['total_fee'], 'fund': r['fund_paid']}
        for r in sorted(day_1110, key=lambda x: x['time'])[:30]
    ]
}

# 10. Top 10 busiest days detail
daily_counts = Counter(r['date'] for r in all_records if r['date'])
top_days_detail = []
for date, cnt in daily_counts.most_common(15):
    day_recs = [r for r in all_records if r['date'] == date]
    top_days_detail.append({
        'date': date,
        'count': cnt,
        'patients': len(set(r['cert'] for r in day_recs)),
        'total_fee': round(sum(r['total_fee'] for r in day_recs), 1),
        'total_fund': round(sum(r['fund_paid'] for r in day_recs), 1),
        'avg_fee': round(sum(r['total_fee'] for r in day_recs)/cnt, 1),
        'weekday': day_recs[0]['weekday'] if day_recs else '',
        'hour_range': f"{min(r['hour'] for r in day_recs if r['hour']>=0)}-{max(r['hour'] for r in day_recs if r['hour']>=0)}",
        'same_day_multi': len([d for d, c in Counter(r['cert'] for r in day_recs).items() if c >= 2]),
    })

# ===== BUILD OUTPUT =====
result = {
    'total_records': len(all_records),
    'total_patients': len(patients),
    'avg_per_patient': round(len(all_records)/len(patients), 1),
    'top20_patients': top_details,
    'hourly_distribution': sorted(hourly.items()),
    'weekday_distribution': sorted(weekdays.items(), key=lambda x: -x[1]),
    'monthly_by_year': {y: sorted(c.items()) for y, c in monthly_year.items()},
    'fee_buckets': dict(fee_buckets),
    'insurance_types': dict(ins_types.most_common()),
    'top_townships': townships.most_common(30),
    'top_addresses': addresses.most_common(30),
    'day_2025_11_10': day_1110_detail,
    'top_busiest_days': top_days_detail,
    'total_fee_all': round(sum(r['total_fee'] for r in all_records), 1),
    'total_fund_all': round(sum(r['fund_paid'] for r in all_records), 1),
}

outpath = r'C:\Users\scrccpa\.openclaw\workspace\temp_ocr\jsk_deep_dive.json'
with open(outpath, 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2, default=str)

print(f'\nSaved: {outpath}')
print(f'Total fee: ¥{result["total_fee_all"]:,.0f}')
print(f'Total fund: ¥{result["total_fund_all"]:,.0f}')
print(f'Top patient: {top_details[0]["patient"]} - {top_details[0]["visits"]} visits')
