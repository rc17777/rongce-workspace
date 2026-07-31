"""Lightweight: institution anomaly deep-dive + violation cross-check"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from collections import defaultdict
from datetime import datetime

def sf(v): return float(v) if v and isinstance(v,(int,float)) else 0.0
def ss(v): return str(v).strip() if v else ''

BASE = r'C:\Users\scrccpa\Desktop\若尔盖审计\若尔盖医保审计\2026年审计资料（医保局财务）\古英=2024-2025门诊、住院、个人账户、基金拨付明细'

# === PHASE 1: Institution anomaly by month ===
print('Phase 1: Monthly patterns for suspicious institutions')
suspects = ['若尔盖县降扎乡卫生院','若尔盖县康乐康金世康药品店','若尔盖县辖曼镇卫生院','若尔盖县妇幼保健计划生育服务中心（若尔盖县妇幼保健院）']

inst_monthly = {s: defaultdict(lambda: {'count':0,'fee':0.0}) for s in suspects}

for year in ['2023','2024','2025']:
    fp = f'{BASE}\\{year}.xlsx'
    wb = load_workbook(fp, read_only=True)
    sn = '总表' if '总表' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sn]
    hdrs = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    col = {str(h).strip(): i for i, h in enumerate(hdrs) if h}
    
    for r in ws.iter_rows(min_row=2, values_only=True):
        inst = ss(r[col['医药机构名称']]) if col.get('医药机构名称') else ''
        if inst not in suspects: continue
        
        dt = r[col['结算时间']] if col.get('结算时间') else None
        if not dt or not hasattr(dt, 'month'): continue
        
        fee = sf(r[col.get('医疗费总额')])
        key = f'{year}-{dt.month:02d}'
        inst_monthly[inst][key]['count'] += 1
        inst_monthly[inst][key]['fee'] += fee
    
    wb.close()
    print(f'  {year} done')

for inst in suspects:
    data = inst_monthly[inst]
    months = sorted(data.keys())
    if len(months) < 3:
        print(f'\n{inst}: insufficient data ({len(months)} months)')
        continue
    
    print(f'\n--- {inst} ---')
    # Find peak months
    peaks = sorted(data.items(), key=lambda x: -x[1]['fee'])[:5]
    print(f'  Top 5 peak months:')
    for m, v in peaks:
        print(f'    {m}: {v["count"]:>4}条  ¥{v["fee"]:>10,.0f}')
    
    # Check if there's a single outlier month driving the growth
    by_year = defaultdict(lambda: {'count':0, 'fee':0.0})
    for m, v in data.items():
        y = m[:4]
        by_year[y]['count'] += v['count']
        by_year[y]['fee'] += v['fee']
    
    for y in sorted(by_year.keys()):
        v = by_year[y]
        avg = v['fee']/v['count'] if v['count'] else 0
        print(f'  {y}: {v["count"]:>4}条  ¥{v["fee"]:>10,.0f}  次均¥{avg:,.0f}')


# === PHASE 2: 违规追回清单交叉比对 ===
print('\n' + '='*70)
print('Phase 2: 违规追回清单内容探查')

vio_base = r'C:\Users\scrccpa\Desktop\若尔盖审计\若尔盖医保审计\2026年审计资料（医保局财务）\2024-2025违规使用医保基金清单'

for fname in ['2024年追回资金佐证材料若尔盖县.xlsx', '2025年监管追回资金佐证材料(若尔盖县).xlsx']:
    fp = f'{vio_base}\\{fname}'
    if not os.path.exists(fp):
        print(f'{fname}: NOT FOUND')
        continue
    
    print(f'\n{fname}:')
    wb = load_workbook(fp, read_only=True)
    
    # Print all sheet names
    print(f'  Sheets: {wb.sheetnames}')
    
    for sn in wb.sheetnames:
        ws = wb[sn]
        hdrs = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        hs = [str(h).strip() for h in hdrs if h]
        print(f'\n  [{sn}] {ws.max_row} rows, {len(hs)} cols')
        print(f'  Headers: {hs[:15]}')
        
        # Print first 5 data rows
        for i, r in enumerate(ws.iter_rows(min_row=2, max_row=min(8, ws.max_row or 0), values_only=True)):
            vals = [str(v)[:40] for v in r[:8] if v is not None]
            print(f'  Row {i+1}: {" | ".join(vals)}')
        
        # Count total rows by column totals
        if ws.max_row and ws.max_row > 1:
            # Try to find amount fields
            amount_cols = [i for i, h in enumerate(hs) if any(k in h for k in ['金额','费用','追回','违规','罚款'])]
            if amount_cols:
                total = 0.0
                count = 0
                for r in ws.iter_rows(min_row=2, values_only=True):
                    for ac in amount_cols:
                        v = sf(r[ac]) if ac < len(r) else 0
                        if v: total += v; count += 1
                if total > 0:
                    print(f'  金额字段({amount_cols}): 合计 ¥{total:,.0f} ({count} 条)')
    
    wb.close()

print('\nDone.')
