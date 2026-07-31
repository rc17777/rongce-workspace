"""Deep-dive: 降扎乡卫生院 all records 2023-2025"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from collections import defaultdict, Counter
from datetime import datetime

def pdate(v):
    if not v: return None
    if hasattr(v, 'date'): return v.date()
    try: return datetime.strptime(str(v)[:10], '%Y-%m-%d').date()
    except: return None
def sf(v): return float(v) if v and isinstance(v,(int,float)) else 0.0
def ss(v): return str(v).strip() if v else ''

BASE = r'C:\Users\scrccpa\Desktop\若尔盖审计\若尔盖医保审计\2026年审计资料（医保局财务）\古英=2024-2025门诊、住院、个人账户、基金拨付明细'
INST = '若尔盖县降扎乡卫生院'

records = []

for year in ['2023','2024','2025']:
    fp = f'{BASE}\\{year}.xlsx'
    wb = load_workbook(fp, read_only=True)
    sn = '总表' if '总表' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sn]
    hdrs = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    col = {str(h).strip(): i for i, h in enumerate(hdrs) if h}
    
    for r in ws.iter_rows(min_row=2, values_only=True):
        inst = ss(r[col['医药机构名称']]) if col.get('医药机构名称') else ''
        if inst != INST: continue
        
        pid = ss(r[col['证件号码']])[-4:] if col.get('证件号码') else ''
        name = ss(r[col['姓名']]) if col.get('姓名') else ''
        mt = ss(r[col['医疗类别']]) if col.get('医疗类别') else ''
        settle = pdate(r[col['结算时间']]) if col.get('结算时间') else None
        fee = sf(r[col.get('医疗费总额')])
        pay = sf(r[col.get('医保支付金额')])
        
        # Get detailed fee breakdown if available (check column exists first)
        drug_idx = col.get('药品费')
        treat_idx = col.get('诊疗费')
        mat_idx = col.get('耗材费')
        bed_idx = col.get('床位费')
        check_idx = col.get('检查费')
        other_idx = col.get('其他费')
        drug_fee = sf(r[drug_idx]) if drug_idx is not None else 0
        treat_fee = sf(r[treat_idx]) if treat_idx is not None else 0
        mat_fee = sf(r[mat_idx]) if mat_idx is not None else 0
        bed_fee = sf(r[bed_idx]) if bed_idx is not None else 0
        check_fee = sf(r[check_idx]) if check_idx is not None else 0
        other_fee = sf(r[other_idx]) if other_idx is not None else 0
        
        diag_idx = col.get('出院诊断名称')
        yidi_idx = col.get('是否异地就医')
        ins_idx = col.get('医保类型')
        
        diag = ss(r[diag_idx])[:40] if diag_idx is not None else ''
        yidi = ss(r[yidi_idx]) if yidi_idx is not None else ''
        ins_type = ss(r[ins_idx]) if ins_idx is not None else ''
        
        records.append({
            'year': year, 'pid': pid, 'name': name, 'type': mt,
            'date': settle, 'fee': fee, 'pay': pay,
            'drug': drug_fee, 'treat': treat_fee, 'mat': mat_fee,
            'bed': bed_fee, 'check': check_fee, 'other': other_fee,
            'diag': diag, 'yidi': yidi, 'ins': ins_type,
        })
    wb.close()
    print(f'{year}: {sum(1 for r in records if r["year"]==year)} records')

print(f'\nTotal: {len(records)} records for {INST}')

# === Analysis ===

# 1. By medical type
print('\n' + '='*60)
print('1. 医疗类别分布')
by_type = Counter(r['type'].split('|')[-1].strip() if '|' in r['type'] else r['type'] for r in records)
for t, c in by_type.most_common():
    fees = [r['fee'] for r in records if (r['type'].split('|')[-1].strip() if '|' in r['type'] else r['type']) == t]
    print(f'  {t}: {c:>4}条  总费用¥{sum(fees):>10,.0f}  次均¥{sum(fees)/c:,.0f}')

# 2. By month
print('\n' + '='*60)
print('2. 月度趋势')
by_month = defaultdict(lambda: {'count':0, 'fee':0.0, 'patients':set()})
for r in records:
    if r['date']:
        key = f'{r["date"].year}-{r["date"].month:02d}'
        by_month[key]['count'] += 1
        by_month[key]['fee'] += r['fee']
        by_month[key]['patients'].add(r['pid'])

prev_count = 0
for m in sorted(by_month.keys()):
    v = by_month[m]
    arrow = ''
    if prev_count > 0:
        ratio = v['count'] / prev_count
        if ratio > 2: arrow = ' 🚨'
        elif ratio > 1.5: arrow = ' ⚡'
    print(f'  {m}: {v["count"]:>4}条  ¥{v["fee"]:>10,.0f}  {len(v["patients"]):>4}人  次均¥{v["fee"]/v["count"]:,.0f}{arrow}')
    prev_count = v['count']

# 3. Top patients
print('\n' + '='*60)
print('3. 高频患者 Top 30')
patient_counts = defaultdict(lambda: {'count':0, 'fee':0.0, 'dates':[], 'name':''})
for r in records:
    p = patient_counts[r['pid']]
    p['count'] += 1
    p['fee'] += r['fee']
    if r['date']: p['dates'].append(r['date'])
    p['name'] = r['name']

for (pid, p) in sorted(patient_counts.items(), key=lambda x: -x[1]['count'])[:30]:
    dates = sorted(p['dates'])
    date_range = f'{dates[0]} ~ {dates[-1]}' if dates else 'N/A'
    avg_days_between = (dates[-1]-dates[0]).days/max(p['count']-1,1) if len(dates)>=2 else 0
    print(f'  {pid} {p["name"]:<8} {p["count"]:>4}条  ¥{p["fee"]:>10,.0f}  {date_range}  均隔{avg_days_between:.0f}天')

# 4. Top diagnoses
print('\n' + '='*60)
print('4. 诊断分布 Top 20')
diag_counts = Counter(r['diag'] for r in records if r['diag'])
for d, c in diag_counts.most_common(20):
    avg_fee = sum(r['fee'] for r in records if r['diag'] == d) / c
    print(f'  {c:>4}条  {d:<40}  次均¥{avg_fee:,.0f}')

# 5. Fee structure analysis
print('\n' + '='*60)
print('5. 费用结构分析')
total_drug = sum(r['drug'] for r in records)
total_treat = sum(r['treat'] for r in records)
total_mat = sum(r['mat'] for r in records)
total_bed = sum(r['bed'] for r in records)
total_check = sum(r['check'] for r in records)
total_other = sum(r['other'] for r in records)
total = total_drug + total_treat + total_mat + total_bed + total_check + total_other
if total > 0:
    print(f'  药品费: ¥{total_drug:>10,.0f} ({total_drug/total*100:.0f}%)')
    print(f'  诊疗费: ¥{total_treat:>10,.0f} ({total_treat/total*100:.0f}%)')
    print(f'  耗材费: ¥{total_mat:>10,.0f} ({total_mat/total*100:.0f}%)')
    print(f'  床位费: ¥{total_bed:>10,.0f} ({total_bed/total*100:.0f}%)')
    print(f'  检查费: ¥{total_check:>10,.0f} ({total_check/total*100:.0f}%)')
    print(f'  其他费: ¥{total_other:>10,.0f} ({total_other/total*100:.0f}%)')
    print(f'  合计:   ¥{total:>10,.0f}')

# 6. Insurance type distribution
print('\n' + '='*60)
print('6. 医保类型分布')
by_ins = Counter(r['ins'] for r in records if r['ins'])
for t, c in by_ins.most_common():
    print(f'  {t}: {c:>4}条')

# 7. Suspicious: same patient, same day, multiple visits
print('\n' + '='*60)
print('7. 同患者同日多次就诊')
same_day_visits = defaultdict(list)
for r in records:
    if r['date'] and r['pid']:
        same_day_visits[(r['pid'], r['date'])].append(r)

multi = {k: v for k, v in same_day_visits.items() if len(v) >= 2}
if multi:
    for (pid, dt), visits in sorted(multi.items(), key=lambda x: -sum(v['fee'] for v in x[1]))[:10]:
        total_fee = sum(v['fee'] for v in visits)
        print(f'  {pid} {visits[0]["name"]} | {dt} | {len(visits)}次 | ¥{total_fee:,.0f} | {" · ".join(v["diag"][:15] for v in visits)}')
else:
    print('  (无此模式)')

# 8. Suspicious: unusually high single-day fees
print('\n' + '='*60)
print('8. 单日单患者费用异常高 (>¥500)')
high_fee_records = [r for r in records if r['fee'] > 500]
if high_fee_records:
    for r in sorted(high_fee_records, key=lambda x: -x['fee'])[:15]:
        print(f'  {r["pid"]} {r["name"]:<8} {r["date"]} ¥{r["fee"]:>8,.0f}  {r["diag"][:30]}  {r["type"].split("|")[-1].strip() if "|" in r["type"] else r["type"]}')
else:
    print('  (无 ¥500+ 费用)')

print('\n✅ 降扎乡卫生院深挖完成')
