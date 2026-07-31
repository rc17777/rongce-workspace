"""Compare 金世康药品店 & 辖曼镇卫生院 vs 降扎乡卫生院"""
import sys, os, json
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from collections import Counter, defaultdict

DATA = r'C:\Users\scrccpa\Desktop\若尔盖审计\若尔盖医保审计\2026年审计资料（医保局财务）\古英=2024-2025门诊、住院、个人账户、基金拨付明细'

targets = {
    'jz': '降扎',       # 降扎乡卫生院
    'jsk': '金世康',     # 金世康药品店
    'xmz': '辖曼'        # 辖曼镇卫生院
}

all_data = {k: [] for k in targets}
for year in ['2023', '2024', '2025']:
    fp = os.path.join(DATA, f'{year}.xlsx')
    print(f'Loading {year}...')
    wb = load_workbook(fp, read_only=True, data_only=True)
    
    # Try sheets
    for sn in ['总表', '本地药店']:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        col = {h: i for i, h in enumerate(headers)}
        
        inst_col = col.get('医药机构名称', col.get('医疗机构名称', None))
        if inst_col is None:
            print(f'  No institution column in {sn}')
            continue
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            inst = str(row[inst_col]) if row[inst_col] else ''
            for key, kw in targets.items():
                if kw in inst:
                    rec = {
                        'year': year,
                        'sheet': sn,
                        'id': str(row[col.get('个人编号','')]) if col.get('个人编号') else '',
                        'name': str(row[col.get('姓名','')]) if col.get('姓名') else '',
                        'cert': str(row[col.get('证件号码','')]) if col.get('证件号码') else '',
                        'time': row[col.get('结算时间','')] if col.get('结算时间') else None,
                        'dept': inst,
                        'doctor': str(row[col.get('医师姓名','')]) if col.get('医师姓名') else '',
                        'diagnosis': str(row[col.get('出院诊断名称','')]) if col.get('出院诊断名称') else '',
                        'drug_fee': float(row[col['药品费']]) if col.get('药品费') and row[col['药品费']] and isinstance(row[col['药品费']], (int,float)) else 0,
                        'treat_fee': float(row[col['诊疗费']]) if col.get('诊疗费') and row[col['诊疗费']] and isinstance(row[col['诊疗费']], (int,float)) else 0,
                        'total_fee': float(row[col['医疗费总额']]) if col.get('医疗费总额') and row[col['医疗费总额']] and isinstance(row[col['医疗费总额']], (int,float)) else 0,
                        'fund_paid': float(row[col['统筹基金支出']]) if col.get('统筹基金支出') and row[col['统筹基金支出']] and isinstance(row[col['统筹基金支出']], (int,float)) else 0,
                        'visit_type': str(row[col.get('就诊方式','')]) if col.get('就诊方式') else '',
                        'med_type': str(row[col.get('医疗类别','')]) if col.get('医疗类别') else '',
                        'address': str(row[col.get('居住地址','')]) if col.get('居住地址') else '',
                        'township': str(row[col.get('乡镇(街道)','')]) if col.get('乡镇(街道)') else '',
                    }
                    all_data[key].append(rec)
        
        print(f'  {sn}: found records for targets')
    
    wb.close()

# === Analysis per institution ===
results = {}
for key, recs in all_data.items():
    if not recs:
        results[key] = {'total': 0}
        continue
    
    # Basic stats
    years = Counter(r['year'] for r in recs)
    patients = defaultdict(list)
    for r in recs:
        cert4 = r['cert'][-4:] if r['cert'] and len(r['cert']) >= 4 else r['cert']
        pk = f"{r['name']}({cert4})"
        patients[pk].append(r)
    
    doctors = Counter(r['doctor'] for r in recs if r['doctor'] and r['doctor'] != 'None')
    monthly = Counter()
    daily = Counter()
    total_fee = sum(r['total_fee'] for r in recs)
    total_fund = sum(r['fund_paid'] for r in recs)
    fee_buckets = Counter()
    diag_empty = sum(1 for r in recs if not r['diagnosis'] or r['diagnosis'] == 'None')
    
    for r in recs:
        t = r['time']
        if t and hasattr(t, 'strftime'):
            monthly[t.strftime('%Y-%m')] += 1
            daily[t.strftime('%Y-%m-%d')] += 1
        ft = r['total_fee']
        if ft <= 50: fee_buckets['0-50'] += 1
        elif ft <= 100: fee_buckets['50-100'] += 1
        elif ft <= 200: fee_buckets['100-200'] += 1
        else: fee_buckets['200+'] += 1
    
    # Top patients
    top_pts = sorted(patients.items(), key=lambda x: -len(x[1]))[:15]
    
    # Visit types
    visit_types = Counter(r['visit_type'] for r in recs if r['visit_type'] and r['visit_type'] != 'None')
    med_types = Counter(r['med_type'] for r in recs if r['med_type'] and r['med_type'] != 'None')
    
    results[key] = {
        'total': len(recs),
        'years': dict(years),
        'patients': len(patients),
        'avg_per_patient': round(len(recs)/len(patients), 1),
        'doctors': doctors.most_common(10),
        'total_fee': total_fee,
        'total_fund': total_fund,
        'avg_fee': round(total_fee/len(recs), 1) if recs else 0,
        'fund_rate': round(total_fund/total_fee*100, 1) if total_fee else 0,
        'fee_buckets': dict(fee_buckets),
        'diag_empty': diag_empty,
        'diag_empty_pct': round(diag_empty/len(recs)*100, 1),
        'monthly': sorted(monthly.items()),
        'top_days': daily.most_common(15),
        'visit_types': dict(visit_types),
        'med_types': dict(med_types),
        'top_patients': [(p, len(r)) for p, r in top_pts],
    }

# === Cross-institution patient overlap ===
# Find patients appearing in multiple institutions
patient_insts = defaultdict(set)
for key, recs in all_data.items():
    for r in recs:
        cert4 = r['cert'][-4:] if r['cert'] and len(r['cert']) >= 4 else r['cert']
        pk = f"{r['name']}({cert4})"
        patient_insts[pk].add(key)

overlap = Counter()
for pk, insts in patient_insts.items():
    overlap[frozenset(insts)] += 1

# Patients in all 3
all3 = [pk for pk, insts in patient_insts.items() if len(insts) >= 3]

# Patients shared between jz and jsk
jz_jsk = [pk for pk, insts in patient_insts.items() if 'jz' in insts and 'jsk' in insts]

# Patients shared between jz and xmz
jz_xmz = [pk for pk, insts in patient_insts.items() if 'jz' in insts and 'xmz' in insts]

out = {
    'per_institution': results,
    'overlap': {str(k): v for k, v in overlap.most_common()},
    'all3_count': len(all3),
    'all3_patients': all3[:30],
    'jz_jsk_shared': jz_jsk[:30],
    'jz_xmz_shared': jz_xmz[:30],
    'total_unique_patients': len(patient_insts),
}

outpath = r'C:\Users\scrccpa\.openclaw\workspace\temp_ocr\triangle_compare.json'
with open(outpath, 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=2, default=str)

print(f'\nSaved: {outpath}')
print(f'\n=== SUMMARY ===')
for key, r in results.items():
    name = {'jz':'降扎乡','jsk':'金世康','xmz':'辖曼镇'}[key]
    print(f'{name}: {r["total"]} records, {r["patients"]} patients, {len(r["doctors"])} doctors, avg ¥{r["avg_fee"]}, diag_empty {r["diag_empty_pct"]}%')
print(f'\nTotal unique patients across all: {len(patient_insts)}')
print(f'Patients in all 3 institutions: {len(all3)}')
print(f'Patients shared JZ+JSK: {len(jz_jsk)}')
print(f'Patients shared JZ+XMZ: {len(jz_xmz)}')
