"""Update data requirements with new file inventory"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

wb = Workbook()

# Styles
hdr_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='0A1F3F')
body_font = Font(name='Microsoft YaHei', size=10)
bold_font = Font(name='Microsoft YaHei', size=10, bold=True)
title_font = Font(name='Microsoft YaHei', size=14, bold=True, color='0A1F3F')
sub_font = Font(name='Microsoft YaHei', size=11, bold=True, color='1A5C6E')
note_font = Font(name='Microsoft YaHei', size=9, color='666666')
wrap = Alignment(wrap_text=True, vertical='top')
center = Alignment(horizontal='center', vertical='top', wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
green_fill = PatternFill('solid', fgColor='C8E6C9')
yellow_fill = PatternFill('solid', fgColor='FFF9C4')
red_fill = PatternFill('solid', fgColor='FFE0E0')
gray_fill = PatternFill('solid', fgColor='F5F5F5')

def style_sheet(ws, headers, col_widths, title=None):
    sr = 1
    if title:
        ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=len(headers))
        ws.cell(row=sr, column=1, value=title).font = title_font
        sr += 1
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.cell(row=sr, column=i, value=h).font = hdr_font
        ws.cell(row=sr, column=i).fill = hdr_fill
        ws.cell(row=sr, column=i).alignment = center
        ws.cell(row=sr, column=i).border = thin_border
        ws.column_dimensions[get_column_letter(i)].width = w
    return sr + 1

def write_rows(ws, sr, rows):
    for ri, row in enumerate(rows):
        r = sr + ri
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val).font = body_font
            ws.cell(row=r, column=ci).alignment = wrap
            ws.cell(row=r, column=ci).border = thin_border

# ==========================================
# Sheet 1: 资料资产盘点
# ==========================================
ws1 = wb.active
ws1.title = '资料资产盘点'

h1 = ['类别', '文件数', '总大小', '关键文件', '已用状态', '可利用的审计事项', '备注']
w1 = [20, 8, 10, 48, 12, 36, 30]
r1 = style_sheet(ws1, h1, w1, '若尔盖医保审计 — 现有资料资产盘点 (' + datetime.now().strftime('%Y-%m-%d %H:%M') + ')')

inventory = [
    ['门诊/住院/个账结算数据', 3, '266MB',
     '2023.xlsx(76MB) / 2024.xlsx(92MB) / 2025.xlsx(97MB)\n总表(门诊+住院)+本地药店',
     '✅ 已全量分析',
     '三(二)重复报销/住院分析/异地就医\n三(三)虚假诊疗/分解处方/虚开串药\n三(三)全县机构风险排序',
     '41.9万条记录已分析完毕，核心数据源'],

    ['收入支出明细(银行流水)', 11, '0.3MB',
     '账户历史明细 xlsx(已解压可读取)\n含9863支出户/9855收入户/4252支出户/4203收入户\n2024+2025年各自收支户明细',
     '🆕 刚发现，待分析',
     '三(二)基金专户管理(挤占挪用核查)\n三(四)基金收支趋势分析',
     '每行含交易时间/收入/支出/余额/对方账户/摘要。可直接分析基金流向'],

    ['城乡居民参保名单', 2, '23MB',
     '2024年12月底城乡居民参保情况1.7.xlsx(11.4MB)\n2025年12月底61035人.xlsx(11.2MB)\n共61,036人',
     '🆕 刚发现，待分析',
     '三(一)重点人群参保(应保尽保)\n三(一)财政补助到位\n三(二)死亡清退(含身份证号)',
     '32列含:身份证号/人员类别/特殊身份/个人缴费/财政补助/乡镇/缴费状态'],

    ['职工参保名单', 2, '122MB',
     '(2024年)单位人员缴费明细.xlsx(59MB)\n(2025年)单位人员缴费明细.xlsx(62MB)',
     '🆕 刚发现，待分析',
     '三(一)职工参保征缴\n三(二)个账分析(金世康/康宁均为职工消费)',
     '可与金世康/康宁的职工消费记录做身份确认'],

    ['违规追回清单(Excel)', 2, '48MB',
     '2024年追回资金佐证材料.xlsx(24.6MB)\n2025年监管追回资金佐证材料.xlsx(23.3MB)',
     '🆕 刚发现，待分析',
     '三(五)以往发现问题整改闭环\n确认降扎/金世康/辖曼是否曾被查处',
     '可查询三家重点机构是否有既往违规记录'],

    ['医保局稽查PDF', 11, '1755MB',
     '医保局稽查1-13.pdf(共11个文件,75-349MB/个)\n包含稽查取证材料和记录',
     '❌ 需OCR(量大)',
     '三(五)历史稽查问题\n确认重点机构既往违规证据',
     '1755MB PDF需选择性OCR(先查目录/索引,再挑重点机构相关页)'],

    ['DRG支付文件', 3, '209MB',
     '2024DRG支付文件2.pdf(95MB)\n2025DRG支付.pdf(93MB)\nDRG支付文件格式说明.pdf(21MB)',
     '❌ 需OCR(量大)',
     '三(三)DRG/DIP高套分组分析',
     '需OCR提取DRG分组和支付标准数据。格式说明PDF可能可直接提取'],

    ['预算文件', 5, '3.5MB',
     '2024年预算(若尔盖汇总).xlsx\n2025年预算(若尔盖).xlsx\n2024年基本医疗保险基金预算调整表.xls\n2025年预算调整基础数据表.xlsx',
     '🆕 刚发现，可读取',
     '三(一)预算执行分析\n三(四)基金收支预测对比',
     'xlsx可直接读取,不需要OCR'],

    ['集采/药品耗材采购', 534, '518MB',
     '张琴/2024 + 张琴/2025\n含集采任务文件/账号/报量详情/通知等',
     '❌ 文件多且分散',
     '三(四)集采任务完成核查\n集采药品配备使用情况',
     '518MB分散在534个文件,需先梳理文件结构再选重点分析'],

    ['委托支付协议', 2, '164MB',
     '集中采购协议.pdf(95MB)\n阿坝州药品集中采购中选药品货款委托预付协议.pdf(69MB)',
     '❌ 需OCR',
     '三(四)采购结算/回款情况',
     '与集采数据配合分析'],

    ['2024-2025收入支出明细(ZIP)', 7, '1MB',
     '4252支出户/9863支出户/9855收入户/4203收入户\n基本医疗保险+医疗救助两类\n2024+2025两年的收支户明细',
     '🆕 可解压读取',
     '三(二)专户管理(收入户vs支出户)\n三(四)基金收支趋势',
     '与银行流水类数据合并分析,可完整追溯基金"收→支→余"全链路'],
]

write_rows(ws1, r1, inventory)
ws1.freeze_panes = 'A%d' % r1

# Color by status
for ri in range(len(inventory)):
    r = r1 + ri
    status = inventory[ri][4]
    if '✅' in status:
        for ci in range(1, 8): ws1.cell(row=r, column=ci).fill = green_fill
    elif '🆕' in status:
        for ci in range(1, 8): ws1.cell(row=r, column=ci).fill = yellow_fill
    elif '❌' in status:
        for ci in range(1, 8): ws1.cell(row=r, column=ci).fill = red_fill

# ==========================================
# Sheet 2: 更新后资料需求(仅需外部获取)
# ==========================================
ws2 = wb.create_sheet('需外部获取的资料')

ws2.merge_cells('A1:G1')
ws2.cell(row=1, column=1, value='仅需从外部获取的资料（已在现有资料中覆盖的已移除）').font = title_font

h2 = ['序号', '方案条款', '审计事项', '所需资料', '提供单位', '用途', '已有机数据可替代?']
w2 = [6, 18, 22, 42, 20, 36, 22]
r2 = 4
for i, (h, w) in enumerate(zip(h2, w2), 1):
    ws2.cell(row=r2, column=i, value=h).font = hdr_font
    ws2.cell(row=r2, column=i).fill = hdr_fill
    ws2.cell(row=r2, column=i).alignment = center
    ws2.cell(row=r2, column=i).border = thin_border
    ws2.column_dimensions[get_column_letter(i)].width = w
r2 += 1

external_needs = [
    [1, '三(一)', '财政补助拨付核实',
     '2024-2025年县级财政配套资金预算批复和拨付记录(红头文件)',
     '县财政局', '核实县级配套是否足额、及时',
     '部分可替代:预算文件已有预算数据,参保名单已有财政补助金额字段'],

    [2, '三(一)', '重点人群资格认定',
     '特困/低保/返贫监测/重残/孤儿认定名单(含身份证号、认定时间)',
     '县民政局/残联/退役军人事务局', '与参保名单中特殊身份字段交叉验证',
     '部分可替代:参保名单已有人员身份类别字段(含普通/低保/特困等)'],

    [3, '三(一)', '税务征缴明细',
     '2024-2025年居民/职工医保征缴明细(含缴费基数、实缴金额、欠费)',
     '县税务局', '核查应征未征、重复参保缴费',
     '部分可替代:参保名单有缴费状态和金额,但缺征缴端数据'],

    [4, '三(二)', '死亡人口名单',
     '2024-2025年若尔盖县死亡人口名单(含身份证号、死亡日期)',
     '县公安局/民政局', '与结算数据+参保名单+个账消费比对,发现死亡后仍用个账',
     '无替代:必须有公安/民政的死亡数据'],

    [5, '三(二)', '大病保险理赔明细',
     '2024-2025年大病保险理赔清单(含被保险人、费用、赔付金额)',
     '大病保险承保公司(县医保局协调)', 'S09/S10已发现疑点,需理赔端原始数据交叉验证',
     '部分可替代:结算数据已有大病保险支付金额字段,但缺理赔审批流程'],

    [6, '三(二)', '医疗救助审批明细',
     '2024-2025年医疗救助审批和拨付明细(含救助对象、审批金额、拨付凭证)',
     '县医保局', 'S10已发现3,542人多渠道使用,需审批端数据核实合规性',
     '部分可替代:结算数据已有医疗救助金额字段,但缺审批记录'],

    [7, '三(三)', '降扎乡卫生院现场取证',
     '①2023-2025年纸质门诊日志\n②人员编制和排班表\n③旦真机等5人处方笺\n④12/30监控录像\n⑤内部考核制度文件',
     '降扎乡卫生院(现场)', '做实P01(无诊断)+P02(单医)+S02(年底冲量)+S03(分解处方)',
     '不可替代:必须现场取证'],

    [8, '三(三)', '金世康药品店现场取证',
     '①进销存系统数据\n②2025年11月销售小票\n③11/10监控录像\n④店主访谈',
     '金世康药品店(现场)', '做实S05(拆分交易)+S06(11月暴涨)',
     '不可替代:必须现场取证'],

    [9, '三(三)', '康宁大药房(新增重点)',
     '①进销存系统数据\n②销售明细\n③店主访谈',
     '康宁大药房(现场)', 'S13新发现:27,908条全县最大药店,需参照金世康方案取证',
     '不可替代:必须现场取证'],

    [10, '三(三)', '挂床住院核实',
     '644条住院≤1天的病案首页(入院诊断+出院诊断+手术编码)',
     '县医保局/相关医院', 'S12发现2.6%一日住院,需区分正常日间手术vs挂床',
     '不可替代:结算数据无诊断和手术编码'],

    [11, '三(三)', 'His收费明细(串换核查)',
     '县医院+藏医院His系统收费明细(含项目编码、单价、数量)',
     '县医院/藏医院', '与医保结算明细逐项比对,识别串换和超标准收费',
     '不可替代:结算数据只有结算金额,无His收费项目明细'],

    [12, '三(三)', 'DRG分组数据',
     '2024-2025年住院病案首页+DRG分组结果+费率表',
     '县医保局', '分析诊断编码一致性,识别高套分组',
     '部分可替代:DRG支付PDF(需OCR)可提取分组和支付标准'],

    [13, '三(三)', '经办机构操作日志',
     '医保核心系统操作日志+经办人员权限配置表',
     '县医保局', '识别权限不当/异常操作/内外勾结',
     '不可替代:操作日志仅在医保局系统中'],

    [14, '三(四)', '集采任务完成数据',
     '2024-2025年集采任务指标+实际采购结算清单+回款记录',
     '县医保局', '核查集采完成率/及时性/配备使用',
     '部分可替代:集采文件目录(张琴/)已有通知和报量,但缺完整数据'],

    [15, '三(五)', '以往审计/检查报告',
     '2022-2024年医保审计报告/检查决定书/巡察问题清单+整改台账',
     '县审计局/医保局', '整改闭环核查:资金是否追回/责任是否追究/同类是否反复',
     '部分可替代:违规追回清单Excel已含部分追回数据;稽查PDF需OCR'],
]

write_rows(ws2, r2, external_needs)
ws2.freeze_panes = 'A%d' % r2

# ==========================================
# Sheet 3: 下一步可立即启动的分析
# ==========================================
ws3 = wb.create_sheet('可立即启动的分析')

h3 = ['优先级', '分析任务', '对应方案', '所用数据源', '预计产出', '预计耗时']
w3 = [8, 40, 18, 36, 36, 12]
r3 = style_sheet(ws3, h3, w3, '现有资料中可立即启动的后续分析任务')

todo = [
    ['P0', '基金收支和专户管理分析(银行流水)',
     '三(二)专户管理\n三(四)收支趋势',
     '收入支出明细(已解压)\n收入户4203/9855 + 支出户4252/9863\n2024+2025年全部收支户明细',
     '基金全链路追踪报告:收入来源→支出流向→余额变动。\n识别大额异常转出/挤占挪用/违规出借',
     '2-3h'],

    ['P0', '死亡清退核查(参保名单×结算数据)',
     '三(二)死亡清退',
     '城乡居民参保名单(61,036人含身份证号)\n2024-2025结算数据\n注意:需要外部死亡名单配合',
     '①已故未停保名单(参保名单中标记为正常但有死亡线索)\n②已停保人员仍在消费个账的疑点',
     '1-2h\n(需死亡名单)'],

    ['P0', '重点人群参保核查(参保名单)',
     '三(一)重点人群参保',
     '城乡居民参保名单\n含:人员身份类别(低保/特困/孤儿等)\n个人缴费金额+财政补助金额',
     '①各类重点人群参保覆盖率\n②财政补助金额是否足额\n③个人缴费是否按规定减免',
     '1-2h'],

    ['P1', '既往违规追回分析',
     '三(五)以往整改',
     '2024年追回资金佐证材料.xlsx(24.6MB)\n2025年监管追回资金佐证材料.xlsx(23.3MB)',
     '①追回金额总额和机构分布\n②降扎/金世康/辖曼是否曾被查处\n③既往问题类型与当前发现的对照',
     '1-2h'],

    ['P1', '全县参保职工身份与金世康/康宁消费对比',
     '三(一)职工参保\n三(三)药店消费',
     '职工参保名单(2024+2025)\n金世康/康宁药店消费记录(结算数据中已有)',
     '①金世康/康宁消费患者是否均为真实职工\n②是否存在非在保人员消费\n③职工人均个账消费异常',
     '1-2h'],

    ['P2', '预算执行分析',
     '三(一)预算执行\n三(四)基金预测',
     '预算文件 xlsx(可直接读取)\n2024+2025年预算+调整表',
     '①预算vs实际收支偏差\n②预算调整合理性\n③基金结余预测vs实际',
     '1h'],

    ['P2', 'DRG支付PDF OCR(先试小样)',
     '三(三)DRG高套',
     'DRG支付文件格式说明.pdf(21MB,可先读)\n2025DRG支付.pdf(93MB)',
     '先试OCR 21MB格式说明文件,确认数据结构后再决定是否全量OCR 200MB',
     '1h(试点)'],
]

write_rows(ws3, r3, todo)

# Color
for ri in range(len(todo)):
    r = r3 + ri
    lv = todo[ri][0]
    if 'P0' in lv:
        for ci in range(1, 7): ws3.cell(row=r, column=ci).fill = red_fill
    elif 'P1' in lv:
        for ci in range(1, 7): ws3.cell(row=r, column=ci).fill = yellow_fill

# ==========================================
# Sheet 4: 总结
# ==========================================
ws4 = wb.create_sheet('总结')
ws4.merge_cells('A1:B1')
ws4.cell(row=1, column=1, value='资料资产盘点总结').font = title_font
ws4.column_dimensions['A'].width = 30
ws4.column_dimensions['B'].width = 70

summary = [
    ('目录总文件数', '580个'),
    ('目录总大小', '3,087MB (约3GB)'),
    ('', ''),
    ('✅ 已充分利用', '结算数据(2023-2025.xlsx, 266MB) — 41.9万条已全量分析'),
    ('🆕 刚发现可立即用', '收入支出银行流水(11文件,0.3MB) — 三(二)专户管理\n参保名单(4文件,145MB) — 三(一)重点人群+死亡清退\n违规追回清单(2文件,48MB) — 三(五)整改回顾\n预算文件(5文件,3.5MB) — 三(一)/三(四)预算分析'),
    ('❌ 需OCR(量大)', '医保局稽查PDF(11文件,1755MB) — 三(五)\nDRG支付PDF(3文件,209MB) — 三(三)\n委托协议PDF(2文件,164MB) — 三(四)'),
    ('❌ 需分散梳理', '集采数据(534文件,518MB)'),
    ('', ''),
    ('仍需外部获取', '①死亡人口名单(公安/民政)\n②税务征缴明细(税务局)\n③大病/救助审批明细(医保局)\n④现场取证(降扎+金世康+康宁)\n⑤His收费明细(县医院/藏医院)\n⑥操作日志(医保局经办)'),
    ('', ''),
    ('当前总完成率', '数据层面:结算数据100%分析完成\n新发现可用数据:5类文件可立即启动\n外部获取缺口:6项仍需协调'),
]

for ri, (k, v) in enumerate(summary):
    r = 3 + ri
    ws4.cell(row=r, column=1, value=k).font = bold_font if k else body_font
    ws4.cell(row=r, column=1).border = thin_border
    ws4.cell(row=r, column=2, value=v).font = body_font
    ws4.cell(row=r, column=2).border = thin_border
    ws4.cell(row=r, column=2).alignment = wrap
    if '✅' in k:
        ws4.cell(row=r, column=1).fill = green_fill
        ws4.cell(row=r, column=2).fill = green_fill
    elif '🆕' in k:
        ws4.cell(row=r, column=1).fill = yellow_fill
        ws4.cell(row=r, column=2).fill = yellow_fill
    elif '❌' in k:
        ws4.cell(row=r, column=1).fill = red_fill
        ws4.cell(row=r, column=2).fill = red_fill

# Save
out = r'C:\Users\scrccpa\Desktop\若尔盖医保审计_资料清单(更新版).xlsx'
wb.save(out)
print('Saved: ' + out)
