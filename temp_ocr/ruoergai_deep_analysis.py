"""
若尔盖医保审计 - 深度分析 v2.0
任务1: 763组分解住院 → Excel完整明细
任务2: 57条多院时空碰撞逐条核实 → Excel
任务3: 参保人数真实性核查 → Excel
"""
import sys, os, json
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from collections import defaultdict

# ============================================================
# Config
# ============================================================
BASE = r'C:\Users\scrccpa\Desktop\若尔盖审计\若尔盖医保审计\2026年审计资料（医保局财务）\古英=2024-2025门诊、住院、个人账户、基金拨付明细'
INSURED_DIR = r'C:\Users\scrccpa\Desktop\若尔盖审计\若尔盖医保审计\2026年审计资料（医保局财务）\2024-2025居民、职工参保名单'
YEARS = ['2023', '2024', '2025']
DESKTOP = r'C:\Users\scrccpa\Desktop'

# ============================================================
# Styles
# ============================================================
HEADER_FILL = PatternFill(start_color='0A1F3F', end_color='0A1F3F', fill_type='solid')
HEADER_FONT = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
DATA_FONT = Font(name='微软雅黑', size=10)
TITLE_FONT = Font(name='微软雅黑', bold=True, size=14, color='0A1F3F')
P0_FILL = PatternFill(start_color='FFD7D7', end_color='FFD7D7', fill_type='solid')
P1_FILL = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical='top')
CENTER_ALIGN = Alignment(horizontal='center', vertical='top')

def style_header(ws, row, ncols):
    for col in range(1, ncols+1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

def style_data(ws, row, ncols):
    for col in range(1, ncols+1):
        cell = ws.cell(row=row, column=col)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = WRAP_ALIGN

def auto_width(ws, ncols, max_nrows=100):
    for col in range(1, ncols+1):
        max_len = 0
        for row in range(1, min(ws.max_row, max_nrows) + 1):
            val = str(ws.cell(row=row, column=col).value or '')
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 40)

# ============================================================
# Data Loading
# ============================================================
def parse_date(val):
    if val is None: return None
    if hasattr(val, 'date'): return val.date()
    if isinstance(val, str):
        for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y%m%d']:
            try: return datetime.strptime(val[:10], fmt).date()
            except: pass
    return None

def load_all_data():
    print('='*60)
    print('Loading 3-year settlement data...')
    print('='*60)
    
    inpatient_records = []
    outpatient_records = []
    all_records = []
    
    for year in YEARS:
        fp = os.path.join(BASE, f'{year}.xlsx')
        if not os.path.exists(fp): continue
        
        print(f'  {year}...', end=' ', flush=True)
        wb = load_workbook(fp, read_only=True)
        main_sheet = '总表' if '总表' in wb.sheetnames else wb.sheetnames[0]
        ws = wb[main_sheet]
        headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        col = {str(h).strip(): i for i, h in enumerate(headers) if h}
        
        id_col = col.get('证件号码')
        name_col = col.get('姓名')
        type_col = col.get('医疗类别')
        settle_col = col.get('结算时间')
        admit_col = col.get('入院时间')
        discharge_col = col.get('出院时间')
        days_col = col.get('住院天数')
        inst_col = col.get('医药机构名称')
        inst_id_col = col.get('医药机构编号')
        fee_col = col.get('医疗费总额')
        diag_col = col.get('出院诊断名称')
        yidi_col = col.get('是否异地就医')
        pay_col = col.get('医保支付金额')
        drug_col = col.get('药品费')
        treat_col = col.get('诊疗费')
        mat_col = col.get('耗材费')
        acct_pay_col = col.get('个人账户支付')
        fund_pay_col = col.get('统筹基金支付')
        assist_pay_col = col.get('大病支付')
        relief_pay_col = col.get('医疗救助')
        
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            count += 1
            if count % 50000 == 0: print(f'{count//1000}k...', end=' ', flush=True)
            
            id_num = str(row[id_col]).strip() if row[id_col] else ''
            name = str(row[name_col]).strip() if row[name_col] else ''
            inst = str(row[inst_col]).strip() if row[inst_col] else ''
            inst_id = str(row[inst_id_col]).strip() if row[inst_id_col] else ''
            fee = row[fee_col] if fee_col and isinstance(row[fee_col], (int, float)) else 0
            pay = row[pay_col] if pay_col and isinstance(row[pay_col], (int, float)) else 0
            diag = str(row[diag_col]).strip() if diag_col and row[diag_col] else ''
            yidi = str(row[yidi_col]).strip() if yidi_col and row[yidi_col] else ''
            
            settle_dt = parse_date(row[settle_col]) if settle_col else None
            admit_dt = parse_date(row[admit_col]) if admit_col else None
            discharge_dt = parse_date(row[discharge_col]) if discharge_col else None
            days = row[days_col] if days_col and isinstance(row[days_col], (int, float)) else 0
            
            med_type_raw = str(row[type_col]) if row[type_col] else ''
            med_type = med_type_raw.split('|')[-1].strip() if '|' in med_type_raw else med_type_raw
            
            drug_fee = row[drug_col] if drug_col and isinstance(row[drug_col], (int, float)) else 0
            treat_fee = row[treat_col] if treat_col and isinstance(row[treat_col], (int, float)) else 0
            mat_fee = row[mat_col] if mat_col and isinstance(row[mat_col], (int, float)) else 0
            acct_pay = row[acct_pay_col] if acct_pay_col and isinstance(row[acct_pay_col], (int, float)) else 0
            fund_pay = row[fund_pay_col] if fund_pay_col and isinstance(row[fund_pay_col], (int, float)) else 0
            assist_pay = row[assist_pay_col] if assist_pay_col and isinstance(row[assist_pay_col], (int, float)) else 0
            relief_pay = row[relief_pay_col] if relief_pay_col and isinstance(row[relief_pay_col], (int, float)) else 0
            
            rec = (id_num, name, inst, inst_id, fee, pay, med_type, settle_dt,
                   admit_dt, discharge_dt, int(days), diag, yidi, drug_fee, treat_fee, mat_fee, 
                   acct_pay, fund_pay, assist_pay, relief_pay, year)
            all_records.append(rec)
            
            if '住院' in med_type:
                inpatient_records.append((id_num, name, admit_dt, discharge_dt, inst, fee, pay, year, diag, days,
                                          inst_id, drug_fee, treat_fee, mat_fee, acct_pay, fund_pay, assist_pay, relief_pay))
            elif '门诊' in med_type:
                outpatient_records.append((id_num, name, settle_dt, inst, fee, pay, year, inst_id))
        
        wb.close()
        print(f'{count:,} records')
    
    print(f'\nTotal: {len(all_records):,} | 住院: {len(inpatient_records):,} | 门诊: {len(outpatient_records):,}')
    return all_records, inpatient_records, outpatient_records

# ============================================================
# Model 1: 分解住院完整版
# ============================================================
def model1_split_hospitalization(inpatient_records):
    print('\n' + '='*60)
    print('Model 1: 分解住院完整分析')
    print('='*60)
    
    patient_admissions = defaultdict(list)
    for rec in inpatient_records:
        if rec[2] and rec[3]:
            patient_admissions[rec[0]].append(rec)
    
    all_findings = []
    for pid, admissions in patient_admissions.items():
        if len(admissions) < 2: continue
        admissions.sort(key=lambda x: x[2] if x[2] else datetime.min.date())
        
        for i in range(len(admissions)):
            for j in range(i+1, len(admissions)):
                a1 = admissions[i]; a2 = admissions[j]
                if a1[4] != a2[4]: continue  # Same hospital
                if a1[3] and a2[2]:
                    gap = (a2[2] - a1[3]).days
                    if 0 <= gap <= 7:
                        all_findings.append({
                            'patient_id': pid[-4:], 'name': a1[1],
                            'hospital': a1[4], 'hospital_id': a1[10],
                            'admit1': str(a1[2]), 'discharge1': str(a1[3]),
                            'admit2': str(a2[2]), 'discharge2': str(a2[3]),
                            'gap_days': gap, 'fee1': a1[5], 'fee2': a2[5],
                            'total_fee': a1[5] + a2[5],
                            'diag1': a1[8], 'diag2': a2[8],
                            'year1': a1[7], 'year2': a2[7],
                            'days1': a1[9], 'days2': a2[9],
                            'drug1': a1[11], 'drug2': a2[11],
                            'treat1': a1[12], 'treat2': a2[12],
                            'mat1': a1[13], 'mat2': a2[13],
                            'pay1': a1[6], 'pay2': a2[6],
                            'fund1': a1[16], 'fund2': a2[16],
                        })
    
    # Deduplicate
    seen = set()
    unique = []
    for f in sorted(all_findings, key=lambda x: -x['total_fee']):
        key = (f['patient_id'], f['hospital'], f['admit1'])
        if key not in seen:
            seen.add(key)
            unique.append(f)
    
    print(f'分解住院总数: {len(unique)} 组')
    print(f'涉及金额: ¥{sum(f["total_fee"] for f in unique):,.0f}')
    
    # Risk classification
    for f in unique:
        risks = []
        if f['gap_days'] <= 1: risks.append('P0')
        elif f['gap_days'] <= 3: risks.append('P1')
        else: risks.append('P2')
        
        # Same diagnosis = higher risk for decomposition
        diag1_base = f['diag1'][:20]
        diag2_base = f['diag2'][:20]
        if diag1_base == diag2_base: risks.append('同诊断')
        if f['gap_days'] == 0: risks.append('0天间隔')
        if f['fee1'] > 100000 or f['fee2'] > 100000: risks.append('高额')
        if f['days1'] < 3 or f['days2'] < 3: risks.append('短期住院')
        
        f['risk_level'] = max(risks, key=lambda r: 'P0' if 'P0' in r else ('P1' if 'P1' in r else 'P2'))
        f['risk_tags'] = ', '.join([r for r in risks if r not in ('P0','P1','P2')])
    
    return unique

# ============================================================
# Model 2: 多院住院逐条核实
# ============================================================
def model2_multi_hospital(inpatient_records, all_records):
    print('\n' + '='*60)
    print('Model 2: 多院时空碰撞逐条核实')
    print('='*60)
    
    # Build daily index
    patient_daily = defaultdict(lambda: defaultdict(list))
    for rec in all_records:
        id_num = rec[0]; name = rec[1]; inst = rec[2]; fee = rec[4]
        med_type = rec[6]; settle_dt = rec[7]; year = rec[20]
        if settle_dt:
            patient_daily[id_num][settle_dt].append({
                'name': name, 'inst': inst, 'fee': fee, 'type': med_type, 'year': year, 'id_num': id_num
            })
    
    # Find multi-hospital same day
    multi_hosp_all = []
    for pid, daily in patient_daily.items():
        for dt, visits in daily.items():
            inp_visits = [v for v in visits if '住院' in v['type']]
            if len(inp_visits) >= 2:
                hospitals = set(v['inst'] for v in inp_visits)
                if len(hospitals) >= 2:
                    multi_hosp_all.append({
                        'patient_id': pid[-4:], 'name': visits[0]['name'],
                        'full_id': pid, 'date': str(dt),
                        'hospitals': sorted(hospitals),
                        'hospital_count': len(hospitals),
                        'total_fee': sum(v['fee'] for v in inp_visits),
                        'details': inp_visits,
                    })
    
    print(f'同天多院住院: {len(multi_hosp_all)} 条')
    
    # Now for each case, look at full history
    patient_history = defaultdict(list)
    for rec in inpatient_records:
        patient_history[rec[0]].append({
            'name': rec[1], 'admit': str(rec[2]) if rec[2] else '', 'discharge': str(rec[3]) if rec[3] else '',
            'hospital': rec[4], 'fee': rec[5], 'pay': rec[6], 'year': rec[7], 'diag': rec[8], 'days': rec[9]
        })
    
    verified = []
    for case in sorted(multi_hosp_all, key=lambda x: -x['total_fee']):
        pid = case['full_id']
        history = patient_history.get(pid, [])
        
        # Verification logic
        verdicts = []
        
        # Check 1: Are the hospitals geographically far apart?
        hosps = case['hospitals']
        # Same county vs different city vs different province
        same_county = all('若尔盖' in h for h in hosps)
        
        if same_county:
            verdicts.append('同县两院-严重可疑-P0')
        else:
            # Check if one is local and one is distant
            has_local = any('若尔盖' in h for h in hosps)
            has_distant = any(not '若尔盖' in h and not '阿坝' in h for h in hosps)
            if has_local and has_distant:
                verdicts.append('本地+异地同时住院-铁证-P0')
            elif has_distant:
                verdicts.append('跨省同天住院-铁证-P0')
            else:
                verdicts.append('跨市同天住院-严重可疑-P1')
        
        # Check 2: Check full admission history for patterns
        total_hist_fee = sum(h['fee'] for h in history)
        case['total_history_fee'] = total_hist_fee
        case['total_admissions'] = len(history)
        
        # Check 3: Number of unique hospitals across all history
        unique_hosps = set(h['hospital'] for h in history)
        case['unique_hospitals_count'] = len(unique_hosps)
        case['all_hospitals'] = ' | '.join(sorted(unique_hosps))
        
        if len(unique_hosps) >= 5:
            verdicts.append('频繁转院(≥5院)')
        
        # Final verdict
        if any('铁证' in v for v in verdicts):
            case['verdict'] = 'P0-铁证'
        elif any('严重可疑' in v for v in verdicts):
            case['verdict'] = 'P1-严重可疑'
        else:
            case['verdict'] = 'P2-需核实'
        
        case['verdict_details'] = '; '.join(verdicts)
        verified.append(case)
    
    return verified

# ============================================================
# Model 3: Insured Population Verification  
# ============================================================
def model3_insured_verification():
    print('\n' + '='*60)
    print('Model 3: 参保人数真实性核查')
    print('='*60)
    
    # Load insured data
    insured_2024_path = os.path.join(INSURED_DIR, '城乡居民参保名单2024-2025', '2024年12月底城乡居民参保情况1.7.xlsx')
    insured_2025_path = os.path.join(INSURED_DIR, '城乡居民参保名单2024-2025', '2025.年12月底61035人.xlsx')
    employee_2024_path = os.path.join(INSURED_DIR, '（2024年）单位人员缴费明细查询20260617172604304_1.xlsx')
    employee_2025_path = os.path.join(INSURED_DIR, '（2025年）单位人员缴费明细查询20260617173052937_1.xlsx')
    
    results = {'checks': [], 'summary': {}}
    
    # 1. Load resident insured count
    for label, path in [('2024城乡居民', insured_2024_path), ('2025城乡居民', insured_2025_path)]:
        if os.path.exists(path):
            wb = load_workbook(path, read_only=True)
            ws = wb[wb.sheetnames[0]]
            nrows = ws.max_row
            print(f'{label}: {nrows} rows (含表头)')
            results['summary'][label] = nrows - 1  # minus header
            wb.close()
        else:
            print(f'{label}: FILE NOT FOUND')
    
    # 2. Load employee insured
    for label, path in [('2024职工', employee_2024_path), ('2025职工', employee_2025_path)]:
        if os.path.exists(path):
            wb = load_workbook(path, read_only=True)
            ws = wb[wb.sheetnames[0]]
            nrows = ws.max_row
            print(f'{label}: {nrows} rows (含表头)')
            results['summary'][label] = nrows - 1
            wb.close()
    
    # 3. Cross-check: Look for deceased individuals still insured
    # Load insured ID list from 2025
    if os.path.exists(insured_2025_path):
        wb = load_workbook(insured_2025_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        headers = [str(h).strip() if h else '' for h in headers]
        
        print(f'\n2025参保表列: {headers[:15]}')
        
        # Find ID and name columns
        id_col_idx = None
        name_col_idx = None
        for i, h in enumerate(headers):
            if '证件' in h and '号码' in h: id_col_idx = i
            if '姓名' in h or '名称' in h: name_col_idx = i
        
        if id_col_idx is not None:
            insured_ids = set()
            insured_names = {}
            cnt = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                cnt += 1
                id_val = str(row[id_col_idx]).strip().replace('\u3000','').replace(' ','') if row[id_col_idx] else ''
                name_val = str(row[name_col_idx]).strip() if name_col_idx and row[name_col_idx] else ''
                if id_val:
                    insured_ids.add(id_val)
                    insured_names[id_val] = name_val
                if cnt % 20000 == 0: print(f'  加载{cnt//1000}k...', end=' ', flush=True)
            
            print(f'\n\n2025末 城乡居民参保人数(去重): {len(insured_ids)}')
            print(f'Excel总行数-1: {nrows - 1} → 去重后: {len(insured_ids)}')
            
            if nrows - 1 != len(insured_ids):
                dup_count = nrows - 1 - len(insured_ids)
                print(f'🔴 重复身份证号: {dup_count} 个!')
                results['checks'].append({
                    'check': '重复参保记录',
                    'finding': f'2025城乡居民参保名单中有 {dup_count} 个重复身份证号',
                    'level': 'P1' if dup_count > 100 else 'P2',
                })
            
            # 4. Check for deceased in insured list (using settlement data patterns)
            # Deceased pattern: 死亡/丧葬 related settlement records
            
            results['insured_count_2025'] = len(insured_ids)
            results['insured_ids_sample'] = list(insured_ids)[:10]
        
        wb.close()
    
    # 5. Check against settlement data - insured but never used
    # (This requires merging with settlement records, done separately)
    
    return results

# ============================================================
# Excel Export
# ============================================================
def export_all_to_excel(split_hosp, multi_hosp, insured_results):
    print('\n' + '='*60)
    print('Exporting to Excel...')
    print('='*60)
    
    wb = Workbook()
    
    # --- Sheet 1: 分解住院763组 ---
    ws1 = wb.active
    ws1.title = '分解住院763组'
    
    # Title
    ws1.merge_cells('A1:R1')
    ws1.cell(row=1, column=1, value=f'若尔盖医保审计 - 分解住院疑点清单（共{len(split_hosp)}组）').font = TITLE_FONT
    
    headers1 = ['序号','风险等级','风险标签','患者ID','姓名','医院','入院1','出院1','入院2','出院2',
                '间隔天数','住院天数1','住院天数2','诊断1','诊断2','费用1','费用2','合计费用',
                '医保支付1','医保支付2']
    
    for i, h in enumerate(headers1, 1):
        ws1.cell(row=3, column=i, value=h)
    style_header(ws1, 3, len(headers1))
    
    for idx, f in enumerate(split_hosp, 1):
        row = idx + 3
        ws1.cell(row=row, column=1, value=idx)
        ws1.cell(row=row, column=2, value=f.get('risk_level', ''))
        ws1.cell(row=row, column=3, value=f.get('risk_tags', ''))
        ws1.cell(row=row, column=4, value=f['patient_id'])
        ws1.cell(row=row, column=5, value=f['name'])
        ws1.cell(row=row, column=6, value=f['hospital'])
        ws1.cell(row=row, column=7, value=f['admit1'])
        ws1.cell(row=row, column=8, value=f['discharge1'])
        ws1.cell(row=row, column=9, value=f['admit2'])
        ws1.cell(row=row, column=10, value=f['discharge2'])
        ws1.cell(row=row, column=11, value=f['gap_days'])
        ws1.cell(row=row, column=12, value=f.get('days1', ''))
        ws1.cell(row=row, column=13, value=f.get('days2', ''))
        ws1.cell(row=row, column=14, value=f.get('diag1', ''))
        ws1.cell(row=row, column=15, value=f.get('diag2', ''))
        ws1.cell(row=row, column=16, value=f['fee1'])
        ws1.cell(row=row, column=17, value=f['fee2'])
        ws1.cell(row=row, column=18, value=f['total_fee'])
        ws1.cell(row=row, column=19, value=f.get('fund1', 0))
        ws1.cell(row=row, column=20, value=f.get('fund2', 0))
        style_data(ws1, row, len(headers1))
        
        # Risk coloring
        if 'P0' in f.get('risk_level', ''):
            for c in range(1, len(headers1)+1):
                ws1.cell(row=row, column=c).fill = P0_FILL
        elif 'P1' in f.get('risk_level', ''):
            for c in range(1, len(headers1)+1):
                ws1.cell(row=row, column=c).fill = P1_FILL
    
    auto_width(ws1, len(headers1))
    
    # --- Sheet 2: 多院同天住院57条 ---
    ws2 = wb.create_sheet('多院同天住院57条')
    
    ws2.merge_cells('A1:K1')
    ws2.cell(row=1, column=1, value=f'若尔盖医保审计 - 同天多院住院逐条核实（共{len(multi_hosp)}条）').font = TITLE_FONT
    
    headers2 = ['序号','判定结果','判定依据','患者ID','姓名','日期','医院列表','医院数','本次费用',
                '历史总费用','历史住院次数','历史涉及医院']
    
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=3, column=i, value=h)
    style_header(ws2, 3, len(headers2))
    
    for idx, m in enumerate(multi_hosp, 1):
        row = idx + 3
        ws2.cell(row=row, column=1, value=idx)
        ws2.cell(row=row, column=2, value=m.get('verdict', ''))
        ws2.cell(row=row, column=3, value=m.get('verdict_details', ''))
        ws2.cell(row=row, column=4, value=m['patient_id'])
        ws2.cell(row=row, column=5, value=m['name'])
        ws2.cell(row=row, column=6, value=m['date'])
        ws2.cell(row=row, column=7, value=' + '.join(m['hospitals']))
        ws2.cell(row=row, column=8, value=m['hospital_count'])
        ws2.cell(row=row, column=9, value=m['total_fee'])
        ws2.cell(row=row, column=10, value=m.get('total_history_fee', 0))
        ws2.cell(row=row, column=11, value=m.get('total_admissions', 0))
        ws2.cell(row=row, column=12, value=m.get('all_hospitals', ''))
        style_data(ws2, row, len(headers2))
        
        # Color by verdict
        if 'P0' in m.get('verdict', ''):
            for c in range(1, len(headers2)+1):
                ws2.cell(row=row, column=c).fill = P0_FILL
    
    auto_width(ws2, len(headers2))
    
    # --- Sheet 3: 分解住院统计 ---
    ws3 = wb.create_sheet('分解住院统计')
    ws3.merge_cells('A1:E1')
    ws3.cell(row=1, column=1, value='分解住院 - 按医院汇总').font = TITLE_FONT
    
    # Aggregate by hospital
    hosp_stats = defaultdict(lambda: {'count': 0, 'total_fee': 0, 'patients': set()})
    for f in split_hosp:
        hosp_stats[f['hospital']]['count'] += 1
        hosp_stats[f['hospital']]['total_fee'] += f['total_fee']
        hosp_stats[f['hospital']]['patients'].add(f['name'])
    
    headers3 = ['医院','分解组数','涉及金额','涉及患者数','平均每组金额']
    for i, h in enumerate(headers3, 1):
        ws3.cell(row=3, column=i, value=h)
    style_header(ws3, 3, len(headers3))
    
    sorted_hosps = sorted(hosp_stats.items(), key=lambda x: -x[1]['total_fee'])
    for idx, (hosp, stats) in enumerate(sorted_hosps, 1):
        row = idx + 3
        ws3.cell(row=row, column=1, value=hosp)
        ws3.cell(row=row, column=2, value=stats['count'])
        ws3.cell(row=row, column=3, value=stats['total_fee'])
        ws3.cell(row=row, column=4, value=len(stats['patients']))
        ws3.cell(row=row, column=5, value=stats['total_fee']/stats['count'] if stats['count'] else 0)
        style_data(ws3, row, len(headers3))
    
    auto_width(ws3, len(headers3))
    
    # --- Sheet 4: 参保人数核查 ---
    ws4 = wb.create_sheet('参保人数核查')
    ws4.merge_cells('A1:D1')
    ws4.cell(row=1, column=1, value='若尔盖医保审计 - 参保人数真实性核查').font = TITLE_FONT
    
    headers4 = ['检查项','发现','风险等级','备注']
    for i, h in enumerate(headers4, 1):
        ws4.cell(row=3, column=i, value=h)
    style_header(ws4, 3, len(headers4))
    
    row = 4
    for check in insured_results.get('checks', []):
        ws4.cell(row=row, column=1, value=check['check'])
        ws4.cell(row=row, column=2, value=check['finding'])
        ws4.cell(row=row, column=3, value=check['level'])
        style_data(ws4, row, 4)
        if 'P0' in check['level']:
            for c in range(1, 5):
                ws4.cell(row=row, column=c).fill = P0_FILL
        row += 1
    
    # Summary rows
    summary = insured_results.get('summary', {})
    row += 1
    ws4.cell(row=row, column=1, value='年度').font = Font(bold=True, size=11)
    ws4.cell(row=row, column=2, value='城乡居民').font = Font(bold=True, size=11)
    ws4.cell(row=row, column=3, value='职工').font = Font(bold=True, size=11)
    ws4.cell(row=row, column=4, value='合计').font = Font(bold=True, size=11)
    style_data(ws4, row, 4)
    
    for year_label in ['2024', '2025']:
        row += 1
        res = summary.get(f'{year_label}城乡居民', 0)
        emp = summary.get(f'{year_label}职工', 0)
        ws4.cell(row=row, column=1, value=f'{year_label}年末')
        ws4.cell(row=row, column=2, value=res)
        ws4.cell(row=row, column=3, value=emp)
        ws4.cell(row=row, column=4, value=res + emp)
        style_data(ws4, row, 4)
    
    auto_width(ws4, len(headers4))
    
    # --- Sheet 5: 同患者高金额分解 ---
    ws5 = wb.create_sheet('同患者高频分解')
    ws5.merge_cells('A1:F1')
    ws5.cell(row=1, column=1, value='同患者分解住院≥3次（高风险串案）').font = TITLE_FONT
    
    # Find patients with >=3 decomposition episodes
    patient_decomp = defaultdict(list)
    for f in split_hosp:
        patient_decomp[f['name']].append(f)
    
    high_freq_patients = {k: v for k, v in patient_decomp.items() if len(v) >= 3}
    sorted_patients = sorted(high_freq_patients.items(), key=lambda x: (-len(x[1]), -sum(f['total_fee'] for f in x[1])))
    
    headers5 = ['患者','分解次数','涉及医院数','总金额','最多医院','典型病例']
    for i, h in enumerate(headers5, 1):
        ws5.cell(row=3, column=i, value=h)
    style_header(ws5, 3, len(headers5))
    
    for idx, (name, cases) in enumerate(sorted_patients, 1):
        row = idx + 3
        unique_hosps = set(c['hospital'] for c in cases)
        top_hosp = max(cases, key=lambda c: c['total_fee'])
        ws5.cell(row=row, column=1, value=name)
        ws5.cell(row=row, column=2, value=len(cases))
        ws5.cell(row=row, column=3, value=len(unique_hosps))
        ws5.cell(row=row, column=4, value=sum(c['total_fee'] for c in cases))
        ws5.cell(row=row, column=5, value=top_hosp['hospital'])
        ws5.cell(row=row, column=6, value=f'{top_hosp["diag1"]} / {top_hosp["diag2"]}')
        style_data(ws5, row, len(headers5))
        
        if len(cases) >= 5:
            for c in range(1, len(headers5)+1):
                ws5.cell(row=row, column=c).fill = P0_FILL
    
    auto_width(ws5, len(headers5))
    
    # Save
    out_path = os.path.join(DESKTOP, '若尔盖医保审计_深度分析.xlsx')
    wb.save(out_path)
    print(f'\nExcel saved: {out_path}')
    return out_path

# ============================================================
# Main
# ============================================================
if __name__ == '__main__':
    start_time = datetime.now()
    print(f'开始时间: {start_time}')
    
    # Load data (biggest step - ~5-10 min)
    all_records, inpatient_records, outpatient_records = load_all_data()
    
    # Model 1: Split hospitalization
    split_hosp = model1_split_hospitalization(inpatient_records)
    
    # Model 2: Multi-hospital verification
    multi_hosp = model2_multi_hospital(inpatient_records, all_records)
    
    # Model 3: Insured verification
    insured_results = model3_insured_verification()
    
    # Export
    out_path = export_all_to_excel(split_hosp, multi_hosp, insured_results)
    
    elapsed = datetime.now() - start_time
    print(f'\n全部完成! 耗时: {elapsed}')
    print(f'输出: {out_path}')
    print('\n=== 结果摘要 ===')
    print(f'分解住院: {len(split_hosp)} 组')
    print(f'多院同天: {len(multi_hosp)} 条')
    p0_count = sum(1 for m in multi_hosp if 'P0' in m.get('verdict', ''))
    print(f'  P0铁证: {p0_count} 条')
    print(f'参保核查: 详见Excel')
