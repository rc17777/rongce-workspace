"""Cross-check with violation list"""
import os, sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

def sf(v): return float(v) if v and isinstance(v,(int,float)) else 0.0

vio_base = r'C:\Users\scrccpa\Desktop\若尔盖审计\若尔盖医保审计\2026年审计资料（医保局财务）\2024-2025违规使用医保基金清单'

for fname in ['2024年追回资金佐证材料若尔盖县.xlsx', '2025年监管追回资金佐证材料(若尔盖县).xlsx']:
    fp = os.path.join(vio_base, fname)
    if not os.path.exists(fp):
        print(f'{fname}: NOT FOUND')
        continue
    
    print(f'\n{"="*60}')
    print(f'{fname}')
    
    wb = load_workbook(fp, read_only=True)
    print(f'Sheets: {wb.sheetnames}')
    
    for sn in wb.sheetnames:
        ws = wb[sn]
        hdrs = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        hs = [str(h).strip() for h in hdrs if h]
        rows = ws.max_row or 0
        
        print(f'\n[{sn}] {rows} rows, {len(hs)} cols')
        print(f'Headers: {hs[:15]}')
        
        # Print first 8 rows
        for i, r in enumerate(ws.iter_rows(min_row=2, max_row=min(10, rows), values_only=True)):
            vals = [str(v)[:35] for v in r[:10] if v is not None]
            print(f'  [{i+1}] {" | ".join(vals)}')
        
        # Try to compute totals if there are amount columns
        amt_cols = [i for i, h in enumerate(hs) if any(k in h for k in ['金额','费用','追回','违规','罚款','合计'])]
        if amt_cols:
            total = 0.0; cnt = 0
            for r in ws.iter_rows(min_row=2, values_only=True):
                for ac in amt_cols:
                    v = sf(r[ac]) if ac < len(r) else 0
                    if v > 0: total += v; cnt += 1
            if total > 0:
                print(f'  💰 金额列 {[hs[i] for i in amt_cols]}: 合计 ¥{total:,.0f}')
    
    wb.close()

print('\nDone.')
