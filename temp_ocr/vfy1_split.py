"""Lightweight: 分解住院诊断交叉验证"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from collections import defaultdict
from datetime import datetime

def pdate(v):
    if not v: return None
    if hasattr(v, 'date'): return v.date()
    try: return datetime.strptime(str(v)[:10], '%Y-%m-%d').date()
    except: return None

def sf(v): return float(v) if v and isinstance(v,(int,float)) else 0.0
def ss(v): return str(v).strip() if v else ''

BASE = r'C:\Users\scrccpa\Desktop\若尔盖审计\若尔盖医保审计\2026年审计资料（医保局财务）\古英=2024-2025门诊、住院、个人账户、基金拨付明细'

# Only load INPATIENT records with key fields
inp = []
for year in ['2023','2024','2025']:
    fp = f'{BASE}\\{year}.xlsx'
    wb = load_workbook(fp, read_only=True)
    sn = '总表' if '总表' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sn]
    hdrs = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    col = {str(h).strip(): i for i, h in enumerate(hdrs) if h}
    
    for r in ws.iter_rows(min_row=2, values_only=True):
        mt = ss(r[col['医疗类别']]) if col.get('医疗类别') else ''
        if '住院' not in mt: continue
        ad = pdate(r[col['入院时间']]) if col.get('入院时间') else None
        dc = pdate(r[col['出院时间']]) if col.get('出院时间') else None
        if not ad or not dc: continue
        inp.append((
            ss(r[col['证件号码']])[-4:] if col.get('证件号码') else '',
            ss(r[col['姓名']]) if col.get('姓名') else '',
            ad, dc,
            int(sf(r[col.get('住院天数')])),
            ss(r[col['医药机构名称']]) if col.get('医药机构名称') else '',
            sf(r[col.get('医疗费总额')]),
            ss(r[col['出院诊断名称']])[:30] if col.get('出院诊断名称') else '',
            year,
        ))
    wb.close()
    print(f'Loaded {year}: {sum(1 for x in inp if x[8]==year)} inpatient')

print(f'Total: {len(inp)}')

# Group by patient
peps = defaultdict(list)
for e in inp: peps[e[0]].append(e)

# Verify splits
splits = {'high':[], 'med':[], 'low':[]}
for pid, eps in peps.items():
    if len(eps)<2: continue
    eps.sort(key=lambda x: x[2])
    for i in range(len(eps)):
        for j in range(i+1, len(eps)):
            a,b = eps[i], eps[j]
            if a[5]!=b[5]: continue
            gap = (b[2]-a[3]).days
            if gap<0 or gap>7: continue
            
            same_diag = (a[7] and b[7] and a[7][:5]==b[7][:5])
            short = (a[4]<=3 and b[4]<=3)
            
            if same_diag and short:
                level = 'high'
            elif same_diag or (short and gap<=1):
                level = 'med'
            else:
                level = 'low'
            
            splits[level].append({
                'pid':pid,'name':a[1],'hosp':a[5],'gap':gap,
                'd1':a[4],'d2':b[4],'f1':a[6],'f2':b[6],'tf':a[6]+b[6],
                'diag1':a[7],'diag2':b[7],'y1':a[8],'y2':b[8]
            })

# Dedup
for lv in splits:
    seen=set(); u=[]
    for s in sorted(splits[lv], key=lambda x:-x['tf']):
        k=(s['pid'],s['hosp'],f"{s['y1']}_{s['y2']}")
        if k not in seen: seen.add(k); u.append(s)
    splits[lv]=u

print(f'\n🔴 高度可疑(同诊断+短住院): {len(splits["high"])} 组')
print(f'🟡 中度: {len(splits["med"])} 组')
print(f'🟢 低度(不同诊断/长住院): {len(splits["low"])} 组')

print(f'\n=== 高度可疑 Top 20 ===')
print(f'{"患者":<6} {"姓名":<10} {"医院":<28} {"隔":>2} {"天1":>3} {"天2":>3} {"费1":>9} {"费2":>9} {"合计":>9} {"诊断1":<20} {"诊断2":<20}')
for s in splits['high'][:20]:
    print(f'{s["pid"]:<6} {s["name"]:<10} {s["hosp"][:28]:<28} {s["gap"]:>2}d {s["d1"]:>3} {s["d2"]:>3} {s["f1"]:>9,.0f} {s["f2"]:>9,.0f} {s["tf"]:>9,.0f} {s["diag1"][:20]:<20} {s["diag2"][:20]:<20}')

# Repeat offenders
pc = defaultdict(int)
for s in splits['high']: pc[(s['pid'],s['name'])]+=1
ro = [(k,v) for k,v in pc.items() if v>=2]
if ro:
    print(f'\n🚨 多次出现:')
    for (pid,name),c in sorted(ro, key=lambda x:-x[1]):
        print(f'  {pid} {name}: {c} 组高分分解住院')

# Per-hospital high split stats
hc = defaultdict(int)
for s in splits['high']: hc[s['hosp']]+=1
print(f'\n🏥 高发医院:')
for h,c in sorted(hc.items(), key=lambda x:-x[1])[:10]:
    print(f'  {c:>4} 组  {h}')
