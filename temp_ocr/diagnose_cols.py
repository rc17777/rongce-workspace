#!/usr/bin/env python3
# encoding: utf-8
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

FPATH = r"C:\Users\scrccpa\Desktop\若尔盖审计\若尔盖医保审计\2026年审计资料（医保局财务）\古英=2024-2025门诊、住院、个人账户、基金拨付明细\2025.xlsx"

import openpyxl
wb = openpyxl.load_workbook(FPATH, read_only=True, data_only=True)
ws = wb['总表']

# Print header row
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        # Print first 50 header values with indices
        for j, h in enumerate(row[:50]):
            print(f"  Col{j}: {h}")
    break

# Also sample some data rows - look for '降扎' or institution names anywhere
print("\n--- Scanning for '降扎' in entire row ---")
count = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    for j, val in enumerate(row):
        if val and '降扎' in str(val):
            print(f"  Row {i+1}, Col{j}: {val}")
            count += 1
            if count >= 10: break
    if count >= 10: break

# Also look for 降扎 in columns near the end (might be in different col)
print("\n--- Scanning for '降扎' in all columns of first 100 rows ---")
count2 = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    for j, val in enumerate(row):
        if val and '降扎' in str(val):
            print(f"  Row {i+1}, Col{j}: {val}")
            count2 += 1
    if i >= 100: break

wb.close()
