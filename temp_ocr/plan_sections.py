"""Audit remaining plan sections using existing data"""
import sys, os, json
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from collections import Counter, defaultdict
from datetime import datetime, timedelta

DATA = r'C:\Users\scrccpa\Desktop\若尔盖审计\若尔盖医保审计\2026年审计资料（医保局财务）\古英=2024-2025门诊、住院、个人账户、基金拨付明细'

all_records = []

for year in ['2024', '2025']:
    fp = os.path.join(DATA, f'{year}.xlsx')
    print(f'Loading {year}...')
    wb = load_workbook(fp, read_only=True, data_only=True)
    
    for sn in wb.sheetnames:
        ws = wb[sn]
        headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        col = {h: i for i, h in enumerate(headers)}
        
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            count += 1
            if count % 50000 == 0:
                print(f'  {sn}: {count} rows, collected {len(all_records)}')
            
            dt = row[col.get('结算时间')]
            dt_str = dt.strftime('%Y-%m-%d %H:%M:%S') if dt and hasattr(dt, 'strftime') else ''
            
            def sf(val):
                if val is None: return 0
                if isinstance(val, (int, float)): return float(val)
                try: return float(val)
                except: return 0
            
            rec = {
                'year': year,
                'sheet': sn,
                'time': dt_str,
                'date': dt.strftime('%Y-%m-%d') if dt and hasattr(dt, 'strftime') else '',
                'name': str(row[col['姓名']]) if col.get('姓名') and row[col['姓名']] else '',
                'cert': str(row[col['证件号码']]) if col.get('证件号码') and row[col['证件号码']] else '',
                'inst': str(row[col['医药机构名称']]) if col.get('医药机构名称') and row[col['医药机构名称']] else '',
                'med_type': str(row[col.get('医疗类别','')]) if col.get('医疗类别') else '',
                'visit_type': str(row[col.get('就诊方式','')]) if col.get('就诊方式') else '',
                'ins_type': str(row[col.get('参保类型','')]) if col.get('参保类型') else '',
                'person_type': str(row[col.get('人员类别','')]) if col.get('人员类别') else '',
                'special_type': str(row[col.get('特殊人员类型','')]) if col.get('特殊人员类型') else '',
                'rescue_type': str(row[col.get('救助对象身份名称','')]) if col.get('救助对象身份名称') else '',
                'is_remote': str(row[col.get('是否异地就医','')]) if col.get('是否异地就医') else '',
                'diag': str(row[col.get('出院诊断名称','')]) if col.get('出院诊断名称') else '',
                'total_fee': sf(row[col.get('医疗费总额')]),
                'fund_paid': sf(row[col.get('统筹基金支出')]),
                'big_illness': sf(row[col.get('大病保险支付金额')]),
                'big_amount_subsidy': sf(row[col.get('大额医疗费用补助金额')]),
                'supplement': sf(row[col.get('补充医疗报销金额')]),
                'civil_service': sf(row[col.get('公务员医疗补助资金支出')]),
                'medical_rescue': sf(row[col.get('医疗救助')]),
                'gov_bottom': sf(row[col.get('政府兜底基金')]),
                'other_fund': sf(row[col.get('其它基金支付')]),
                'personal_acct': sf(row[col.get('个人账户支付')]),
                'personal_cash': sf(row[col.get('个人现金支付')]),
                'total_pay': sf(row[col.get('医保支付金额')]),
                'drug_fee': sf(row[col.get('药品费')]),
                'treat_fee': sf(row[col.get('诊疗费')]),
                'bed_fee': sf(row[col.get('床位费')]),
                'hospital_days': sf(row[col.get('住院天数')]),
                'admit_time': str(row[col.get('入院时间','')]) if col.get('入院时间') and row[col.get('入院时间')] else '',
                'discharge_time': str(row[col.get('出院时间','')]) if col.get('出院时间') and row[col.get('出院时间')] else '',
            }
            all_records.append(rec)
        print(f'  {sn}: {count} rows')
    wb.close()

print(f'\nTotal records: {len(all_records)}')

# ===== ANALYSIS =====
results = {}

# == 1. 住院数据分析 ==
print('\n=== Analyzing Inpatient Records ===')
inpatient = [r for r in all_records if '住院' in r['med_type']]
print(f'Inpatient records: {len(inpatient)}')

# 挂床住院：入院=出院同日 or 住院天数=1
same_day_discharge = [r for r in inpatient if r['hospital_days'] <= 1 and r['hospital_days'] > 0]
print(f'Same-day discharge (住院天数≤1): {len(same_day_discharge)}')

# 住院天数异常分布
hospital_days = Counter(r['hospital_days'] for r in inpatient if r['hospital_days'] > 0)
top_days = hospital_days.most_common(20)

# 住院费用分析
inp_fee_buckets = Counter()
for r in inpatient:
    f = r['total_fee']
    if f <= 500: inp_fee_buckets['0-500'] += 1
    elif f <= 2000: inp_fee_buckets['500-2000'] += 1
    elif f <= 5000: inp_fee_buckets['2000-5000'] += 1
    elif f <= 10000: inp_fee_buckets['5000-10000'] += 1
    else: inp_fee_buckets['10000+'] += 1

# Top住院机构
inp_insts = Counter(r['inst'] for r in inpatient)
# Top住院患者
inp_patients = defaultdict(list)
for r in inpatient:
    cert4 = r['cert'][-4:] if r['cert'] and len(r['cert']) >= 4 else r['cert']
    pk = f"{r['name']}({cert4})"
    inp_patients[pk].append(r)
inp_top_pts = sorted(inp_patients.items(), key=lambda x: (-len(x[1]), -sum(r['total_fee'] for r in x[1])))[:20]

results['inpatient'] = {
    'total': len(inpatient),
    'same_day_discharge': len(same_day_discharge),
    'same_day_pct': round(len(same_day_discharge)/len(inpatient)*100, 1) if inpatient else 0,
    'hospital_days_top': top_days[:15],
    'fee_buckets': dict(inp_fee_buckets),
    'top_institutions': inp_insts.most_common(15),
    'top_patients': [(p, len(r), round(sum(x['total_fee'] for x in r), 1)) for p, r in inp_top_pts],
    'total_fee': round(sum(r['total_fee'] for r in inpatient), 1),
    'total_fund': round(sum(r['fund_paid'] for r in inpatient), 1),
}

# == 2. 重复报销检测 ==
print('\n=== Analyzing Duplicate Reimbursement ===')
multi_fund_records = [r for r in all_records if r['fund_paid'] + r['big_illness'] + r['medical_rescue'] + r['civil_service'] + r['supplement'] + r['gov_bottom'] > 0]
print(f'Records with any fund payment: {len(multi_fund_records)}')

# 同患者同日多基金渠道
patient_date_multi = defaultdict(list)
for r in multi_fund_records:
    pk = r['cert'][-8:] if r['cert'] and len(r['cert']) >= 8 else r['cert']
    key = f"{pk}_{r['date']}"
    patient_date_multi[key].append(r)

# 同日多条记录
multi_rec_dates = {k: v for k, v in patient_date_multi.items() if len(v) >= 2}
print(f'Patients with 2+ records on same day: {len(multi_rec_dates)}')

# 不同机构同日结算
cross_inst_same_day = []
for key, recs in multi_rec_dates.items():
    insts = set(r['inst'] for r in recs)
    if len(insts) >= 2:
        total_funds = sum(r['fund_paid']+r['big_illness']+r['medical_rescue'] for r in recs)
        cross_inst_same_day.append({
            'key': key,
            'institutions': list(insts),
            'records': len(recs),
            'total_funds': round(total_funds, 1),
            'sample': [{'name': r['name'], 'inst': r['inst'][:20], 'fee': r['total_fee'], 'fund': r['fund_paid']} for r in recs[:5]]
        })
cross_inst_same_day.sort(key=lambda x: -x['total_funds'])

# 同患者有多基金渠道支出
patient_all_funds = defaultdict(lambda: {'fund': 0, 'big_illness': 0, 'rescue': 0, 'civil': 0, 'supp': 0, 'gov': 0, 'records': 0})
for r in all_records:
    cert4 = r['cert'][-4:] if r['cert'] and len(r['cert']) >= 4 else r['cert']
    pk = f"{r['name']}({cert4})"
    patient_all_funds[pk]['fund'] += r['fund_paid']
    patient_all_funds[pk]['big_illness'] += r['big_illness']
    patient_all_funds[pk]['rescue'] += r['medical_rescue']
    patient_all_funds[pk]['civil'] += r['civil_service']
    patient_all_funds[pk]['supp'] += r['supplement']
    patient_all_funds[pk]['gov'] += r['gov_bottom']
    patient_all_funds[pk]['records'] += 1

# Patients using 2+ fund channels
multi_channel = []
for pk, funds in patient_all_funds.items():
    channels = sum(1 for k in ['fund','big_illness','rescue','civil','supp','gov'] if funds[k] > 0)
    if channels >= 2:
        multi_channel.append({'patient': pk, 'channels': channels, **funds})
multi_channel.sort(key=lambda x: -(x['fund']+x['big_illness']+x['rescue']))

results['duplicate'] = {
    'cross_inst_same_day_count': len(cross_inst_same_day),
    'cross_inst_same_day_top': cross_inst_same_day[:20],
    'multi_channel_patients': len(multi_channel),
    'multi_channel_top': multi_channel[:20],
}

# == 3. 异地就医分析 ==
print('\n=== Analyzing Remote Medical ===')
remote = [r for r in all_records if '是' in r['is_remote']]
print(f'Remote records: {len(remote)}')
remote_insts = Counter(r['inst'] for r in remote)
remote_monthly = Counter(r['date'][:7] for r in remote if r['date'])
remote_total_fee = sum(r['total_fee'] for r in remote)
remote_total_fund = sum(r['fund_paid'] for r in remote)

results['remote'] = {
    'total': len(remote),
    'total_fee': round(remote_total_fee, 1),
    'total_fund': round(remote_total_fund, 1),
    'top_institutions': remote_insts.most_common(15),
    'monthly': sorted(remote_monthly.items()),
}

# == 4. 特殊人群分析 ==
print('\n=== Analyzing Special Populations ===')
special_persons = Counter()
rescue_persons = Counter()
for r in all_records:
    st = r['special_type']
    if st and st != 'None' and st.strip():
        special_persons[st] += 1
    rt = r['rescue_type']
    if rt and rt != 'None' and rt.strip():
        rescue_persons[rt] += 1

results['special_pop'] = {
    'special_types': special_persons.most_common(20),
    'rescue_types': rescue_persons.most_common(20),
}

# == 5. 全县机构风险排序 ==
print('\n=== Ranking All Institutions ===')
inst_stats = defaultdict(lambda: {'total': 0, 'inpatient': 0, 'outpatient': 0, 'pharmacy': 0, 
                                     'total_fee': 0, 'fund_paid': 0, 'patients': set(), 'months_active': set(),
                                     'zero_diag': 0})
for r in all_records:
    inst = r['inst']
    if not inst or inst == 'None':
        continue
    s = inst_stats[inst]
    s['total'] += 1
    if '住院' in r['med_type']:
        s['inpatient'] += 1
    elif '门诊' in r['med_type']:
        s['outpatient'] += 1
    elif '药店' in r['med_type']:
        s['pharmacy'] += 1
    s['total_fee'] += r['total_fee']
    s['fund_paid'] += r['fund_paid']
    s['patients'].add(r['cert'])
    if r['date']:
        s['months_active'].add(r['date'][:7])
    if not r['diag'] or r['diag'] == 'None' or r['diag'].strip() == '':
        s['zero_diag'] += 1

# Rank by total fee
inst_ranked = sorted(inst_stats.items(), key=lambda x: -x[1]['total_fee'])
results['institution_ranking'] = [
    {
        'name': inst,
        'total_records': s['total'],
        'inpatient': s['inpatient'],
        'outpatient': s['outpatient'],
        'pharmacy': s['pharmacy'],
        'total_fee': round(s['total_fee'], 1),
        'fund_paid': round(s['fund_paid'], 1),
        'patients': len(s['patients']),
        'months': len(s['months_active']),
        'zero_diag_pct': round(s['zero_diag']/s['total']*100, 1) if s['total'] else 0,
        'avg_fee': round(s['total_fee']/s['total'], 1) if s['total'] else 0,
    }
    for inst, s in inst_ranked[:50]
]

# == 6. 全县整体统计 ==
print('\n=== County-wide Summary ===')
total_fee_county = sum(r['total_fee'] for r in all_records)
total_fund_county = sum(r['fund_paid'] for r in all_records)
total_medical_rescue = sum(r['medical_rescue'] for r in all_records)
total_big_illness = sum(r['big_illness'] for r in all_records)
total_personal_acct = sum(r['personal_acct'] for r in all_records)
total_personal_cash = sum(r['personal_cash'] for r in all_records)

# By med_type
med_type_stats = defaultdict(lambda: {'count': 0, 'fee': 0, 'fund': 0})
for r in all_records:
    mt = r['med_type']
    med_type_stats[mt]['count'] += 1
    med_type_stats[mt]['fee'] += r['total_fee']
    med_type_stats[mt]['fund'] += r['fund_paid']

results['county_summary'] = {
    'total_records': len(all_records),
    'total_fee': round(total_fee_county, 1),
    'total_fund': round(total_fund_county, 1),
    'total_medical_rescue': round(total_medical_rescue, 1),
    'total_big_illness': round(total_big_illness, 1),
    'total_personal_acct': round(total_personal_acct, 1),
    'total_personal_cash': round(total_personal_cash, 1),
    'unique_institutions': len(inst_stats),
    'by_med_type': {k: {'count': v['count'], 'fee': round(v['fee'],1), 'fund': round(v['fund'],1)} for k, v in sorted(med_type_stats.items(), key=lambda x: -x[1]['fee'])[:10]},
}

# == 7. 挂床住院明细 ==
same_day_top = sorted(same_day_discharge, key=lambda x: -x['total_fee'])[:30]
results['suspicious_inpatient'] = [
    {
        'name': r['name'], 'cert': r['cert'][-4:],
        'inst': r['inst'], 'date': r['date'],
        'days': r['hospital_days'], 'total_fee': r['total_fee'],
        'fund_paid': r['fund_paid'], 'diag': r['diag'][:30] if r['diag'] else '',
    }
    for r in same_day_top
]

# == 8. 多基金渠道 Top ==
results['multi_fund_top'] = [
    {
        'patient': m['patient'],
        'channels': m['channels'],
        'fund': round(m['fund'], 1),
        'big_illness': round(m['big_illness'], 1),
        'rescue': round(m['rescue'], 1),
        'records': m['records'],
    }
    for m in multi_channel[:30]
]

# Save
out = r'C:\Users\scrccpa\.openclaw\workspace\temp_ocr\plan_sections.json'
with open(out, 'w', encoding='utf-8') as f:
    json.dump(results, f, ensure_ascii=False, indent=2, default=str)

print(f'\nSaved: {out}')
print(f'Inpatient: {len(inpatient)}, Same-day discharge: {len(same_day_discharge)}')
print(f'Cross-inst same day: {len(cross_inst_same_day)}')
print(f'Multi-channel patients: {len(multi_channel)}')
print(f'Remote: {len(remote)}')
print(f'Institutions: {len(inst_stats)}')
print(f'County total fee: ¥{total_fee_county:,.0f}')
