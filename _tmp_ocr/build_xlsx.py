# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8')
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("NO_OPENPYXL")
    sys.exit(1)

rows = [
    (1, "行财股", "102001", "金川县人大常委会", "周贞秀 13698177793", 10188331.19),
    (2, "行财股", "112001", "中共金川县委统一战线工作部本级", "姚佳玲 18015767817", 5972281.03),
    (3, "行财股", "113001", "金川县妇女联合会本级", "欧旨蓉 15281502442", 2804457.69),
    (4, "行财股", "115001", "金川县工商业联合会本级", "谭君 13309048777", 836669.55),
    (5, "行财股", "126001", "金川县财政局本级", "史姐 15378378808", 68561165.16),
    (6, "行财股", "127001", "金川县审计局本级", "莫小全 13037815772", 4214685.45),
    (7, "社保股", "132004", "疾控中心", "韩图华 18090244652", 11703882.22),
    (8, "行财股", "139001", "金川县委党校（事业）", "胥明浩 13309045592", 3095641.28),
    (9, "行财股", "172001", "金川县科学技术协会（行政）", "杨全香 18161491125", 1961968.90),
    (10, "资环股", "302001", "毛日乡人民政府（行政）", "龙江平 18090222269", 11353968.03),
    (11, "资环股", "308001", "撒比脚乡人民政府（行政）", "赵云阳 18090439739", 8586900.72),
    (12, "资环股", "319001", "安宁镇人民政府（行政）", "陈仙 19981463718", 12337393.40),
    (13, "企业股", "", "金川兴鸿人力资源有限责任公司", "邓拓 13568791112", None),
]

wb = Workbook()
ws = wb.active
ws.title = "监督检查收费测算"

navy = PatternFill("solid", fgColor="0A1F3F")
gold = Font(color="C5955C", bold=True, name="微软雅黑")
white = Font(color="FFFFFF", bold=True, name="微软雅黑")
thin = Side(style="thin", color="C5955C")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
right = Alignment(horizontal="right", vertical="center")

title = "2026年会计监督检查单位明细及收支审计收费测算表（依据：川发改901号文）"
ws.merge_cells("A1:H1")
ws["A1"] = title
ws["A1"].font = Font(bold=True, size=13, name="微软雅黑", color="0A1F3F")
ws["A1"].alignment = center
ws.row_dimensions[1].height = 40

headers = ["序号", "业务股室", "单位代码", "预算单位", "联系人及电话",
           "2025年1-12月收入合计(计费基数/元)", "适用费率", "测算收费(元)"]
ws.append(headers)
for c in range(1, len(headers)+1):
    cell = ws.cell(row=2, column=c)
    cell.fill = navy
    cell.font = white
    cell.alignment = center
    cell.border = border

r = 3
for row in rows:
    idx, dept, code, unit, contact, income = row
    ws.cell(row=r, column=1, value=idx).alignment = center
    ws.cell(row=r, column=2, value=dept).alignment = center
    ws.cell(row=r, column=3, value=code).alignment = center
    ws.cell(row=r, column=4, value=unit)
    ws.cell(row=r, column=5, value=contact)
    ic = ws.cell(row=r, column=6, value=(income if income is not None else "（空）"))
    if income is not None:
        ic.number_format = '#,##0.00'
    ic.alignment = right
    ws.cell(row=r, column=7, value="待定")   # 费率待川发改901确认
    ws.cell(row=r, column=8, value="待测算")
    for c in range(1, 9):
        ws.cell(row=r, column=c).border = border
    r += 1

# 合计行
ws.cell(row=r, column=1, value="合计")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
ws.cell(row=r, column=1).alignment = center
ws.cell(row=r, column=1).font = Font(bold=True, name="微软雅黑")
tot = ws.cell(row=r, column=6, value="=SUM(F3:F%d)" % (r-1))
tot.number_format = '#,##0.00'
tot.font = Font(bold=True, name="微软雅黑")
tot.alignment = right
for c in range(1, 9):
    ws.cell(row=r, column=c).border = border
    ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="F5F2EC")

widths = [6, 10, 10, 30, 22, 26, 14, 16]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

out = r"C:\Users\scrccpa\.openclaw\workspace\output\金川县会计监督检查-收支审计收费测算表.xlsx"
import os
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print("SAVED", out)
