# -*- coding: utf-8 -*-
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INF = float('inf')
# 差额定率累进档次: (档次上限元, 费率)  —— 川发改901附件 财务报表审计基准费率
TIERS = [
    (1_000_000,      0.005),   # 100万以下 5‰
    (5_000_000,      0.0015),  # 100-500万 1.5‰
    (10_000_000,     0.0008),  # 500-1000万 0.8‰
    (50_000_000,     0.0004),  # 1000-5000万 0.4‰
    (100_000_000,    0.0003),  # 5000万-1亿 0.3‰
    (500_000_000,    0.00022), # 1-5亿 0.22‰
    (1_000_000_000,  0.00015), # 5-10亿 0.15‰
    (10_000_000_000, 0.0001),  # 10-100亿 0.1‰
    (INF,            0.00006), # 100亿以上 0.06‰
]
FLOOR = 2000.0           # 100万以下 5‰ 保底2000元(针对报表审计基准费)
SS_RATE = 1.5            # 财务收支审计 = 报表审计基准 × 150%

def base_fee(b):
    """差额定率累进 报表审计基准收费"""
    fee = 0.0
    low = 0.0
    for cap, rate in TIERS:
        if b > cap:
            fee += (cap - low) * rate
            low = cap
        else:
            fee += (b - low) * rate
            break
    return max(fee, FLOOR)

# 算法自检: 文档例子 资产总额3000万 报表审计应=2.3万元
chk = base_fee(30_000_000)
print("[自检] 3000万报表审计基准 = {:,.2f} 元 (文档标称 23,000)".format(chk))
assert abs(chk - 23000) < 0.01, "算法与文档例子不符!"
print("[自检] 通过 ✓\n")

rows = [
    (1, "行财股", "102001", "金川县人大常委会", "周贞秀 13698177793", 10188331.19),
    (2, "行财股", "112001", "中共金川县委统一战线工作部本级", "姚佳玲 18015767817", 5972281.03),
    (3, "行财股", "113001", "金川县妇女联合会本级", "欧旨蓉 15281502442", 2804457.69),
    (4, "行财股", "115001", "金川县工商业联合会本级", "谭君 13309048777", 836669.55),
    (5, "行财股", "126001", "金川县财政局本级", "史姐 15378378808", 68561165.16),
    (6, "行财股", "127001", "金川县审计局本级", "莫小全 13037815772", 4214685.45),
    (7, "社保股", "132004", "疾控中心", "韩图华 18090244652", 11703882.22),
    (8, "行财股", "139001", "金川县委党校（事业）", "胥明浩 13309045592", 3095641.28),
    (9, "行财股", "172001", "金川县科学技术协会（行政）", "杨全香 18161491125", 1961968.90),
    (10, "资环股", "302001", "毛日乡人民政府（行政）", "龙江平 18090222269", 11353968.03),
    (11, "资环股", "308001", "撒比脚乡人民政府（行政）", "赵云阳 18090439739", 8586900.72),
    (12, "资环股", "319001", "安宁镇人民政府（行政）", "陈仙 19981463718", 12337393.40),
    (13, "企业股", "", "金川兴鸿人力资源有限责任公司", "邓拓 13568791112", None),
]

results = []
sum_income = sum_base = sum_ss = 0.0
print("序号 单位                         收入合计         报表审计基准   收支审计(×150%)")
for idx, dept, code, unit, contact, inc in rows:
    if inc is None:
        results.append((idx, dept, code, unit, contact, None, None, None))
        print(f"{idx:>2}  {unit[:14]:<14}  (收入空白，无法测算)")
        continue
    bf = base_fee(inc)
    ss = bf * SS_RATE
    results.append((idx, dept, code, unit, contact, inc, bf, ss))
    sum_income += inc; sum_base += bf; sum_ss += ss
    print(f"{idx:>2}  {unit[:14]:<14}  {inc:>15,.2f}  {bf:>12,.2f}  {ss:>12,.2f}")

print("-"*80)
print(f"合计(12家)  收入 {sum_income:,.2f}   报表审计基准 {sum_base:,.2f}   收支审计合计 {sum_ss:,.2f}")
print(f"收支审计合计 浮动区间(±20%): {sum_ss*0.8:,.2f}  ~  {sum_ss*1.2:,.2f}")

# ===== 生成最终 Excel =====
wb = Workbook(); ws = wb.active; ws.title = "收支审计收费测算"
navy = PatternFill("solid", fgColor="0A1F3F")
beige = PatternFill("solid", fgColor="F5F2EC")
white = Font(color="FFFFFF", bold=True, name="微软雅黑")
thin = Side(style="thin", color="C5955C")
border = Border(left=thin,right=thin,top=thin,bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
right = Alignment(horizontal="right", vertical="center")

ws.merge_cells("A1:I1")
ws["A1"] = "金川县2026年会计监督检查单位 · 财务收支审计收费测算表"
ws["A1"].font = Font(bold=True, size=13, name="微软雅黑", color="0A1F3F")
ws["A1"].alignment = center; ws.row_dimensions[1].height = 34
ws.merge_cells("A2:I2")
ws["A2"] = "依据：川发改价格〔2013〕901号《四川省会计师事务所服务收费管理办法》附件 | 财务收支审计=报表审计基准×150%，差额定率累进，计费基数=2025年1-12月收入合计"
ws["A2"].font = Font(size=9, name="微软雅黑", color="1A5C6E")
ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 30

hdr = ["序号","业务股室","单位代码","预算单位","联系人及电话","计费基数(收入合计/元)","报表审计基准费(元)","收支审计费=×150%(元)","浮动区间±20%(元)"]
ws.append(hdr)
for c in range(1,len(hdr)+1):
    cell = ws.cell(row=3, column=c); cell.fill=navy; cell.font=white; cell.alignment=center; cell.border=border

r = 4
for idx, dept, code, unit, contact, inc, bf, ss in results:
    ws.cell(row=r,column=1,value=idx).alignment=center
    ws.cell(row=r,column=2,value=dept).alignment=center
    ws.cell(row=r,column=3,value=code).alignment=center
    ws.cell(row=r,column=4,value=unit)
    ws.cell(row=r,column=5,value=contact)
    if inc is None:
        for c,txt in [(6,"（原表空白）"),(7,"—"),(8,"待补收入后测算"),(9,"—")]:
            ws.cell(row=r,column=c,value=txt).alignment=center
    else:
        a=ws.cell(row=r,column=6,value=round(inc,2)); a.number_format='#,##0.00'; a.alignment=right
        b=ws.cell(row=r,column=7,value=round(bf,2)); b.number_format='#,##0.00'; b.alignment=right
        d=ws.cell(row=r,column=8,value=round(ss,2)); d.number_format='#,##0.00'; d.alignment=right
        e=ws.cell(row=r,column=9,value="%s ~ %s"%(format(round(ss*0.8,2),',.2f'),format(round(ss*1.2,2),',.2f'))); e.alignment=center
    for c in range(1,10):
        ws.cell(row=r,column=c).border=border
    r += 1

ws.cell(row=r,column=1,value="合计")
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5)
ws.cell(row=r,column=1).alignment=center; ws.cell(row=r,column=1).font=Font(bold=True,name="微软雅黑")
for col,val in [(6,round(sum_income,2)),(7,round(sum_base,2)),(8,round(sum_ss,2))]:
    cc=ws.cell(row=r,column=col,value=val); cc.number_format='#,##0.00'; cc.font=Font(bold=True,name="微软雅黑"); cc.alignment=right
ws.cell(row=r,column=9,value="%s ~ %s"%(format(round(sum_ss*0.8,2),',.2f'),format(round(sum_ss*1.2,2),',.2f'))).alignment=center
for c in range(1,10):
    ws.cell(row=r,column=c).border=border; ws.cell(row=r,column=c).fill=beige

widths=[6,9,9,28,20,20,16,18,26]
for i,w in enumerate(widths,1):
    ws.column_dimensions[get_column_letter(i)].width=w

out = r"C:\Users\scrccpa\.openclaw\workspace\output\金川县会计监督检查-收支审计收费测算表.xlsx"
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print("\nSAVED", out)
