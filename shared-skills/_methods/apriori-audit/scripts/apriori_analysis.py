#!/usr/bin/env python3
"""
Apriori关联规则算法 — 审计异常检测

两种模式:
1. frequent  — 发现频繁结队 (不应形成关联却形成)
2. missing   — 发现缺失关联 (应形成关联却未形成)

参考: 罗鑫(沭阳县审计局)《基于Apriori算法规则核查群体骗保问题》(中国审计 2023年第16期)
"""
import sys, io, argparse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from itertools import combinations
from collections import defaultdict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

RED = PatternFill(patternType='solid', fgColor='FFD7D7')
YEL = PatternFill(patternType='solid', fgColor='FFF3CD')
HEADER = PatternFill(patternType='solid', fgColor='1A3A6E')
H = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
N = Font(name='Microsoft YaHei', size=10)
B = Font(name='Microsoft YaHei', size=10, bold=True)
TH = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
C = Alignment(horizontal='center', vertical='center', wrap_text=True)
L = Alignment(horizontal='left', vertical='center', wrap_text=True)


def load_data(filepath: str) -> dict:
    """加载事务数据: {事务ID: [项列表]}"""
    df = pd.read_excel(filepath)
    cols = df.columns.tolist()
    tid_col = cols[0] if '事务' in cols[0] or 'ID' in cols[0].upper() else cols[0]
    item_col = cols[1] if len(cols) > 1 else cols[0]

    transactions = defaultdict(set)
    for _, row in df.iterrows():
        transactions[row[tid_col]].add(str(row[item_col]))
    return dict(transactions)


def find_frequent_1_itemsets(transactions: dict, min_support: int) -> dict:
    """频繁1项集: 单个项的计数 ≥ min_support"""
    item_count = defaultdict(int)
    for items in transactions.values():
        for item in items:
            item_count[item] += 1
    return {frozenset([k]): v for k, v in item_count.items() if v >= min_support}


def apriori_gen(prev_frequent: list):
    """由k-1频繁项集生成k项集候选"""
    k = len(next(iter(prev_frequent)))
    candidates = set()
    for itemset1 in prev_frequent:
        for itemset2 in prev_frequent:
            union = itemset1 | itemset2
            if len(union) == k + 1:
                candidates.add(union)
    return candidates


def count_support(transactions: dict, candidates: set) -> dict:
    """计算候选集支持度"""
    counts = defaultdict(int)
    for items in transactions.values():
        for cand in candidates:
            if cand.issubset(items):
                counts[cand] += 1
    return counts


def apriori_frequent(transactions: dict, min_support: int, max_k: int = 5) -> list:
    """Apriori主循环: 发现所有频繁项集"""
    all_frequent = []
    freq_1 = find_frequent_1_itemsets(transactions, min_support)

    prev_frequent = set(freq_1.keys())
    while prev_frequent and len(next(iter(prev_frequent))) < max_k:
        candidates = apriori_gen(prev_frequent)
        if not candidates:
            break
        counts = count_support(transactions, candidates)
        freq_k = {k: v for k, v in counts.items() if v >= min_support}
        if not freq_k:
            break
        all_frequent.append(freq_k)
        prev_frequent = set(freq_k.keys())

    return all_frequent


def compute_association_rules(frequent_k: dict, freq_1: dict, min_confidence: float) -> list:
    """从频繁项集计算关联规则 (支持度/置信度/提升度)"""
    rules = []
    for itemset, support_count in frequent_k.items():
        if len(itemset) < 2:
            continue
        for i in range(1, len(itemset)):
            for antecedent in combinations(itemset, i):
                antecedent = frozenset(antecedent)
                consequent = itemset - antecedent
                ant_support = freq_1.get(antecedent, 0)
                if ant_support == 0:
                    continue
                confidence = support_count / ant_support
                # 提升度 = 置信度 / consequent独立概率
                total_tx = sum(1 for _ in transactions if True)
                cons_support = freq_1.get(consequent, 0)
                lift = confidence / (cons_support / total_tx) if cons_support > 0 else 0

                if confidence >= min_confidence:
                    rules.append({
                        'antecedent': ' & '.join(sorted(antecedent)),
                        'consequent': ' & '.join(sorted(consequent)),
                        'support_count': support_count,
                        'confidence': round(confidence, 4),
                        'lift': round(lift, 2),
                    })
    return sorted(rules, key=lambda x: (-x['confidence'], -x['support_count']))


def analyze_frequent(input_file: str, output_file: str,
                     min_support: int = 3, min_confidence: float = 0.6,
                     max_k: int = 5):
    """方向一: 频繁结队分析"""
    global transactions
    transactions = load_data(input_file)
    total_tx = len(transactions)

    print(f"加载: {total_tx} 条事务记录")
    print(f"参数: 最小支持度={min_support}, 最小置信度={min_confidence}, 最大项集={max_k}")

    # Apriori
    freq_1_data = find_frequent_1_itemsets(transactions, min_support)
    print(f"频繁1项集: {len(freq_1_data)} 个")

    all_freq = apriori_frequent(transactions, min_support, max_k)
    for i, fk in enumerate(all_freq):
        print(f"频繁{i+2}项集: {len(fk)} 个")
        if fk:
            top = sorted(fk.items(), key=lambda x: -x[1])[:3]
            for itemset, cnt in top:
                items_str = ', '.join(sorted(itemset))
                print(f"  {items_str}: {cnt}次")

    # Generate association rules from 2-itemsets and above
    all_rules = []
    for fk in all_freq:
        if not fk:
            continue
        if len(next(iter(fk.keys()))) >= 2:
            rules = compute_association_rules(fk, freq_1_data, min_confidence)
            all_rules.extend(rules)

    print(f"\n关联规则: {len(all_rules)} 条 (置信度≥{min_confidence})")

    # Export Excel
    wb = Workbook()

    # Sheet 1: 频繁项集
    ws1 = wb.active
    ws1.title = '频繁项集'
    ws1.merge_cells('A1:D1')
    ws1['A1'] = 'Apriori频繁项集分析 (方向一: 频繁结队检测)'
    ws1['A1'].font = Font(name='Microsoft YaHei', size=14, bold=True, color='1A3A6E')

    h1 = ['项集大小', '项集内容', '出现次数(支持度)', '关联人数']
    for c, h in enumerate(h1, 1):
        cl = ws1.cell(row=3, column=c, value=h)
        cl.font = H; cl.fill = HEADER; cl.alignment = C; cl.border = TH

    r = 4
    for k, fk in enumerate(all_freq):
        for itemset, cnt in sorted(fk.items(), key=lambda x: -x[1]):
            size = len(itemset)
            items_str = ', '.join(sorted(itemset))
            cl = ws1.cell(row=r, column=1, value=size)
            cl.font = N; cl.alignment = C; cl.border = TH
            cl.fill = RED if size >= 3 else YEL

            cl = ws1.cell(row=r, column=2, value=items_str)
            cl.font = N; cl.alignment = L; cl.border = TH

            cl = ws1.cell(row=r, column=3, value=cnt)
            cl.font = N; cl.alignment = C; cl.border = TH

            cl = ws1.cell(row=r, column=4, value=size)
            cl.font = N; cl.alignment = C; cl.border = TH
            r += 1

    ws1.column_dimensions['A'].width = 10
    ws1.column_dimensions['B'].width = 50
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 12

    # Sheet 2: 关联规则
    ws2 = wb.create_sheet('关联规则')
    ws2.merge_cells('A1:E1')
    ws2['A1'] = '关联规则 (置信度≥阈值)'
    ws2['A1'].font = Font(name='Microsoft YaHei', size=14, bold=True, color='1A3A6E')

    h2 = ['前项(条件)', '后项(结果)', '共同出现次数', '置信度', '提升度']
    for c, h in enumerate(h2, 1):
        cl = ws2.cell(row=3, column=c, value=h)
        cl.font = H; cl.fill = HEADER; cl.alignment = C; cl.border = TH

    for i, rule in enumerate(all_rules):
        r = i + 4
        vals = [rule['antecedent'], rule['consequent'], rule['support_count'],
                rule['confidence'], rule['lift']]
        for c, val in enumerate(vals, 1):
            cl = ws2.cell(row=r, column=c, value=val)
            cl.font = N; cl.alignment = C if c >= 3 else L; cl.border = TH
            if c == 4:
                cl.fill = RED if val >= 0.8 else (YEL if val >= 0.6 else None)
            if c == 5:
                cl.fill = RED if val > 3 else None

    ws2.column_dimensions['A'].width = 40
    ws2.column_dimensions['B'].width = 40
    ws2.column_dimensions['C'].width = 16
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 12

    wb.save(output_file)
    print(f"\n结果: {output_file}")
    print(f"  Sheet1: {r - 4} 个频繁项集")
    print(f"  Sheet2: {len(all_rules)} 条关联规则")


def analyze_missing(input_file: str, output_file: str,
                    min_support: float = 0.8, max_k: int = 3):
    """方向二: 缺失关联分析"""
    global transactions
    transactions = load_data(input_file)
    total_tx = len(transactions)

    min_support_count = int(total_tx * min_support)
    print(f"加载: {total_tx} 条事务")
    print(f"支持度阈值: {min_support*100:.0f}% ({min_support_count}条)")

    freq_1 = find_frequent_1_itemsets(transactions, min_support_count)
    all_freq = apriori_frequent(transactions, min_support_count, max_k)

    # Find strong rules (>80% confidence)
    strong_rules = []
    for fk in all_freq:
        if not fk:
            continue
        rules = compute_association_rules(fk, freq_1, 0.8)
        strong_rules.extend(rules)

    # Find transactions where antecedent exists but consequent is missing
    missing_cases = []
    for rule in strong_rules[:50]:  # Top 50 rules
        ant_set = set(rule['antecedent'].split(' & '))
        cons_set = set(rule['consequent'].split(' & '))
        for tid, items in transactions.items():
            if ant_set.issubset(items) and not cons_set.issubset(items):
                missing_cases.append({
                    '事务ID': tid,
                    '关联规则': f"{rule['antecedent']} → {rule['consequent']}",
                    '正常置信度': rule['confidence'],
                    '已有项': ', '.join(sorted(items)),
                    '缺失项': ', '.join(sorted(cons_set - items)),
                })

    print(f"强关联规则: {len(strong_rules)} 条")
    print(f"缺失案例: {len(missing_cases)} 个")

    # Export
    wb = Workbook()
    ws = wb.active
    ws.title = '缺失关联疑点'
    ws.merge_cells('A1:E1')
    ws['A1'] = 'Apriori缺失关联分析 (方向二: 应有却无)'
    ws['A1'].font = Font(name='Microsoft YaHei', size=14, bold=True, color='1A3A6E')

    h = ['事务ID', '关联规则', '正常置信度', '已有项', '缺失项']
    for c, hv in enumerate(h, 1):
        cl = ws.cell(row=3, column=c, value=hv)
        cl.font = H; cl.fill = HEADER; cl.alignment = C; cl.border = TH

    for i, case in enumerate(missing_cases):
        r = i + 4
        for c, key in enumerate(['事务ID', '关联规则', '正常置信度', '已有项', '缺失项'], 1):
            cl = ws.cell(row=r, column=c, value=case[key])
            cl.font = N; cl.alignment = L; cl.border = TH
            if c == 5:
                cl.fill = RED
            if c == 3:
                cl.alignment = C

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 30

    wb.save(output_file)
    print(f"结果: {output_file}")
    print(f"  {len(missing_cases)} 个缺失案例")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Apriori关联规则审计分析')
    parser.add_argument('--i', '--input', dest='input', required=True,
                        help='输入xlsx (事务ID列 + 项列)')
    parser.add_argument('--o', '--output', dest='output',
                        default='apriori_结果.xlsx', help='输出Excel')
    parser.add_argument('--mode', choices=['frequent', 'missing'],
                        default='frequent',
                        help='frequent=频繁结队检测 | missing=缺失关联检测')
    parser.add_argument('--min-support', type=float, default=3,
                        help='最小支持度(frequent模式=绝对次数, 默认3; missing模式=比例, 默认0.8)')
    parser.add_argument('--min-confidence', type=float, default=0.6,
                        help='最小置信度(默认0.6)')
    parser.add_argument('--max-k', type=int, default=5,
                        help='最大频繁项集大小(默认5)')
    args = parser.parse_args()

    if args.mode == 'frequent':
        analyze_frequent(args.input, args.output, int(args.min_support),
                          args.min_confidence, args.max_k)
    else:
        analyze_missing(args.input, args.output, args.min_support, args.max_k)
