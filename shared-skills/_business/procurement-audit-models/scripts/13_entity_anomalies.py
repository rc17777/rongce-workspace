#!/usr/bin/env python3
"""
L15-L19: 基于公共资源交易中心数据的企业/人员异常检测
参考: 郑委/周长明(温州市审计局)《工程项目招标投标常见问题及审计方法》(中国审计 2023年第6期)

检测维度:
  L15 - 陪标专业户 (从未中标但频繁投标)
  L16 - 保证金同时缴纳 (时间差<10分钟)
  L17 - 经办人一致性 (同项目不同企业负责人相同)
  L18 - 硬盘/MAC/IP相同 (同项目不同企业共享存储/网络)
  L19 - 评标专家违规 (未回避/跨项目一致/签到时序异常)
"""
import sys, io, argparse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

RED = PatternFill(patternType='solid', fgColor='FFD7D7')
YEL = PatternFill(patternType='solid', fgColor='FFF3CD')
HEADER = PatternFill(patternType='solid', fgColor='1A3A6E')
H = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
N = Font(name='Microsoft YaHei', size=10)
B = Font(name='Microsoft YaHei', size=10, bold=True)
TH = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
C = Alignment(horizontal='center', vertical='center', wrap_text=True)
L = Alignment(horizontal='left', vertical='center', wrap_text=True)


def detect_perpetual_losers(df_projects: pd.DataFrame, output_wb, pct: float = 0.05):
    """L15: 陪标专业户 — 从未中标但频繁投标"""
    print("\n--- L15: 陪标专业户检测 ---")

    # df_projects: 项目名称 | 投标单位 | 是否中标(1/0)
    bid_counts = df_projects.groupby('投标单位').agg(
        投标次数=('项目名称', 'nunique'),
        中标次数=('是否中标', 'sum')
    ).reset_index()

    losers = bid_counts[bid_counts['中标次数'] == 0].sort_values('投标次数', ascending=False)
    if len(losers) > 0:
        threshold = max(5, int(len(losers) * pct))
        suspects = losers.head(threshold)
    else:
        suspects = losers.head(0)

    print(f"  总投标人: {len(bid_counts)}, 0中标: {len(losers)}, 疑点: {len(suspects)} (前{pct*100:.0f}%)")

    ws = output_wb.create_sheet('L15-陪标专业户')
    ws.merge_cells('A1:D1')
    ws['A1'] = 'L15 陪标专业户 — 从未中标但频繁投标'
    ws['A1'].font = Font(name='Microsoft YaHei', size=14, bold=True, color='1A3A6E')

    h = ['投标单位', '投标次数', '中标次数', '风险等级']
    for c, hv in enumerate(h, 1):
        cl = ws.cell(row=3, column=c, value=hv)
        cl.font = H; cl.fill = HEADER; cl.alignment = C; cl.border = TH

    for i, (_, row) in enumerate(suspects.iterrows()):
        r = i + 4
        for c, val in enumerate([row['投标单位'], row['投标次数'], 0, 'HIGH'], 1):
            cl = ws.cell(row=r, column=c, value=val)
            cl.font = N; cl.alignment = C; cl.border = TH
            cl.fill = RED

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12

    # Also save all losers for reference
    for i, (_, row) in enumerate(losers.iterrows()):
        r = i + len(suspects) + 4
        ws.cell(row=r, column=1, value=row['投标单位'])
        ws.cell(row=r, column=2, value=row['投标次数'])

    return suspects


def detect_bond_timing(df_bonds: pd.DataFrame, output_wb, gap_min: int = 10):
    """L16: 保证金同时缴纳 — 时间差<10分钟"""
    print("\n--- L16: 保证金同时缴纳检测 ---")

    # df_bonds: 项目名称 | 投标单位 | 保证金缴纳时间(datetime)
    if '保证金缴纳时间' not in df_bonds.columns:
        print("  缺少保证金缴纳时间列，跳过")
        return pd.DataFrame()

    df_bonds['缴纳时间'] = pd.to_datetime(df_bonds['保证金缴纳时间'])

    suspicious = []
    for project, group in df_bonds.groupby('项目名称'):
        times = group.sort_values('缴纳时间')
        for i in range(len(times)):
            for j in range(i + 1, len(times)):
                gap = abs((times.iloc[j]['缴纳时间'] - times.iloc[i]['缴纳时间']).total_seconds() / 60)
                if gap < gap_min:
                    suspicious.append({
                        '项目名称': project,
                        '企业A': times.iloc[i]['投标单位'],
                        '企业B': times.iloc[j]['投标单位'],
                        '时间差(分钟)': round(gap, 1),
                        'A缴纳时间': times.iloc[i]['缴纳时间'],
                        'B缴纳时间': times.iloc[j]['缴纳时间'],
                    })

    suspects = pd.DataFrame(suspicious)
    if len(suspects) > 0:
        # Aggregate by pair
        pair_counts = suspects.groupby(['企业A', '企业B']).size().reset_index(name='次数')
        pair_counts = pair_counts[pair_counts['次数'] > 2].sort_values('次数', ascending=False)
        print(f"  疑点配对: {len(suspects)} 条, 多次出现: {len(pair_counts)} 对")
    else:
        print("  无保证金时间异常")
        pair_counts = pd.DataFrame()

    ws = output_wb.create_sheet('L16-保证金同时缴')
    ws.merge_cells('A1:E1')
    ws['A1'] = f'L16 保证金同时缴纳 — 时间差<{gap_min}分钟'
    ws['A1'].font = Font(name='Microsoft YaHei', size=14, bold=True, color='1A3A6E')

    h = ['企业A', '企业B', '同时缴纳次数', '风险等级']
    for c, hv in enumerate(h, 1):
        cl = ws.cell(row=3, column=c, value=hv)
        cl.font = H; cl.fill = HEADER; cl.alignment = C; cl.border = TH

    for i, (_, row) in enumerate(pair_counts.iterrows()):
        r = i + 4
        for c, val in enumerate([row['企业A'], row['企业B'], row['次数'], 'HIGH'], 1):
            cl = ws.cell(row=r, column=c, value=val)
            cl.font = N; cl.alignment = C; cl.border = TH
            cl.fill = RED

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 12

    return suspects


def detect_handler_duplicate(df_projects: pd.DataFrame, output_wb):
    """L17: 经办人一致性 — 同项目不同企业的项目负责人相同"""
    print("\n--- L17: 经办人一致性检测 ---")

    if '项目负责人姓名' not in df_projects.columns:
        print("  缺少项目负责人信息，跳过")
        return pd.DataFrame()

    suspicious = []
    id_col = '项目负责人身份证号码' if '项目负责人身份证号码' in df_projects.columns else None

    for project, group in df_projects.groupby('项目名称'):
        if len(group) < 2:
            continue
        names = group['项目负责人姓名'].tolist()
        if len(names) != len(set(names)):  # Has duplicates
            if id_col:
                for i in range(len(group)):
                    for j in range(i + 1, len(group)):
                        if (group.iloc[i]['项目负责人姓名'] == group.iloc[j]['项目负责人姓名'] and
                            str(group.iloc[i][id_col]) == str(group.iloc[j][id_col])):
                            suspicious.append({
                                '项目名称': project,
                                '企业A': group.iloc[i]['投标单位'],
                                '企业B': group.iloc[j]['投标单位'],
                                '负责人姓名': group.iloc[i]['项目负责人姓名'],
                                '负责人身份证': group.iloc[i][id_col],
                            })
            else:
                for i in range(len(group)):
                    for j in range(i + 1, len(group)):
                        if group.iloc[i]['项目负责人姓名'] == group.iloc[j]['项目负责人姓名']:
                            suspicious.append({
                                '项目名称': project,
                                '企业A': group.iloc[i]['投标单位'],
                                '企业B': group.iloc[j]['投标单位'],
                                '负责人姓名': group.iloc[i]['项目负责人姓名'],
                            })

    suspects = pd.DataFrame(suspicious)
    print(f"  经办人一致案例: {len(suspects)} 个")

    ws = output_wb.create_sheet('L17-经办人一致')
    ws.merge_cells('A1:E1')
    ws['A1'] = 'L17 经办人一致性 — 同项目不同企业负责人相同'
    ws['A1'].font = Font(name='Microsoft YaHei', size=14, bold=True, color='1A3A6E')

    h = ['项目名称', '企业A', '企业B', '负责人姓名'] + (['身份证号'] if id_col else []) + ['风险等级']
    for c, hv in enumerate(h, 1):
        cl = ws.cell(row=3, column=c, value=hv)
        cl.font = H; cl.fill = HEADER; cl.alignment = C; cl.border = TH

    for i, (_, row) in enumerate(suspects.iterrows()):
        r = i + 4
        vals = [row.get('项目名称', ''), row.get('企业A', ''), row.get('企业B', ''),
                row.get('负责人姓名', ''), 'HIGH']
        if id_col:
            vals.insert(4, row.get('负责人身份证', ''))
        for c, val in enumerate(vals, 1):
            cl = ws.cell(row=r, column=c, value=val)
            cl.font = N; cl.alignment = C; cl.border = TH
            cl.fill = RED

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 12

    return suspects


def detect_hardware_dup(df_hw: pd.DataFrame, output_wb):
    """L18: 硬盘/MAC/IP相同"""
    print("\n--- L18: 硬件特征码检测 ---")

    hw_fields = []
    for f in ['硬盘特征码', 'MAC地址', '公网IP地址']:
        if f in df_hw.columns:
            hw_fields.append(f)

    if not hw_fields:
        print("  无硬件特征码字段，跳过")
        return

    for field in hw_fields:
        suspicious = []
        for project, group in df_hw.groupby('项目名称'):
            vals = group[[field, '投标单位']].dropna()
            if len(vals) < 2:
                continue
            seen = {}
            for _, row in vals.iterrows():
                key = str(row[field])
                if key in seen and key != 'nan':
                    suspicious.append({
                        '项目名称': project,
                        '字段': field,
                        '匹配值': key,
                        '企业A': seen[key],
                        '企业B': row['投标单位'],
                    })
                else:
                    seen[key] = row['投标单位']

        if suspicious:
            ws = output_wb.create_sheet(f'L18-{field[:8]}')
            ws.merge_cells('A1:E1')
            ws['A1'] = f'L18 {field}相同'
            ws['A1'].font = Font(name='Microsoft YaHei', size=14, bold=True, color='1A3A6E')

            h = ['项目名称', '字段', '匹配值', '企业A', '企业B']
            for c, hv in enumerate(h, 1):
                cl = ws.cell(row=3, column=c, value=hv)
                cl.font = H; cl.fill = HEADER; cl.alignment = C; cl.border = TH

            for i, s in enumerate(suspicious):
                for c, k in enumerate(h, 1):
                    cl = ws.cell(row=i + 4, column=c, value=s[k])
                    cl.font = N; cl.alignment = C; cl.border = TH
                    cl.fill = RED

            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 16
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 22
            ws.column_dimensions['E'].width = 22

            print(f"  {field}: {len(suspicious)} 个异常")


def detect_expert_violations(df_projects: pd.DataFrame, df_experts: pd.DataFrame, df_social: pd.DataFrame,
                             output_wb):
    """L19: 评标专家违规"""
    print("\n--- L19: 评标专家违规检测 ---")

    # (a) 未按规定回避 — 专家社保在某投标单位
    if not df_experts.empty and not df_social.empty and '专家姓名' in df_projects.columns:
        suspicious_avoid = []
        for _, row in df_projects.iterrows():
            experts = str(row['专家姓名']).split(',') if pd.notna(row.get('专家姓名')) else []
            bidders = set()
            if '投标单位' in df_projects.columns:
                proj_bidders = df_projects[df_projects['项目名称'] == row['项目名称']]['投标单位'].tolist()
            for exp in experts:
                exp = exp.strip()
                exp_info = df_experts[df_experts['专家姓名'] == exp]
                if exp_info.empty:
                    continue
                exp_id = exp_info.iloc[0].get('身份证号码', '')
                social_rec = df_social[df_social['身份证号码'] == str(exp_id)]
                if not social_rec.empty:
                    employer = social_rec.iloc[0].get('缴费单位', '')
                    for bidder in proj_bidders:
                        if bidder in str(employer):
                            suspicious_avoid.append({
                                '项目名称': row['项目名称'],
                                '专家姓名': exp,
                                '身份证号': exp_id,
                                '社保单位': employer,
                                '投标单位': bidder,
                                '问题': '专家在投标单位缴纳社保',
                            })
        print(f"  专家未回避: {len(suspicious_avoid)} 个")

    # (b) 不同项目评标专家一致
    if '专家姓名' in df_projects.columns and '专家身份证号码' in df_projects.columns:
        expert_groups = df_projects.groupby(['专家姓名', '专家身份证号码'])['项目名称'].apply(set).reset_index()
        expert_groups['项目数'] = expert_groups['项目名称'].apply(len)
        same_experts = expert_groups[expert_groups['项目数'] > 1]
        print(f"  跨项目专家完全一致: {len(same_experts)} 组")

    # (c) 签到时间晚于评标时间
    if not df_experts.empty and '专家签到时间' in df_experts.columns and '评标时间' in df_experts.columns:
        df_experts['签到'] = pd.to_datetime(df_experts['专家签到时间'], errors='coerce')
        df_experts['评标'] = pd.to_datetime(df_experts['评标时间'], errors='coerce')
        late = df_experts[df_experts['签到'] > df_experts['评标']]
        print(f"  签到时序异常: {len(late)} 个")

        if len(late) > 0:
            ws = output_wb.create_sheet('L19-评标专家违规')
            ws.merge_cells('A1:E1')
            ws['A1'] = 'L19 评标专家违规检测'
            ws['A1'].font = Font(name='Microsoft YaHei', size=14, bold=True, color='1A3A6E')
            # ... simplified for now


def main():
    parser = argparse.ArgumentParser(description='L15-L19 投标实体异常检测')
    parser.add_argument('--projects', help='项目投标表 (项目名称/投标单位/是否中标)')
    parser.add_argument('--bonds', help='保证金表 (项目名称/投标单位/保证金缴纳时间)')
    parser.add_argument('--hw', help='硬盘特征码表 (项目名称/投标单位/硬盘特征码/MAC地址/公网IP)')
    parser.add_argument('--o', '--output', dest='output', default='实体异常检测.xlsx')
    args = parser.parse_args()

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    if args.projects:
        df_projects = pd.read_excel(args.projects)
        detect_perpetual_losers(df_projects, wb)
        detect_handler_duplicate(df_projects, wb)

    if args.bonds:
        df_bonds = pd.read_excel(args.bonds)
        detect_bond_timing(df_bonds, wb)

    if args.hw:
        df_hw = pd.read_excel(args.hw)
        detect_hardware_dup(df_hw, wb)

    wb.save(args.output)
    print(f"\nDone: {args.output}")


if __name__ == '__main__':
    main()
