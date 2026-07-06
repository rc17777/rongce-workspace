#!/usr/bin/env python3
"""
L13 节资率分析 + 最优围标人数检测
参考: 黄晶 王鋆《基于招标投标历史数据揭示围标行为的审计方法》(中国审计2023年第20期)

输入: 招标台账 (项目/投标人/是否中标/投标价格/招标控制价/投标单位数)
输出: 节资率疑点表 + 围标人数风险标记
"""
import sys, io, argparse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BoxWhiskerChart, Reference


def sf(c):
    return PatternFill(patternType='solid', fgColor=c)


RED = sf('FFD7D7')
YEL = sf('FFF3CD')
GRN = sf('D4EDDA')
HEADER = sf('1A3A6E')
H = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
N = Font(name='Microsoft YaHei', size=10)
B = Font(name='Microsoft YaHei', size=10, bold=True)
BR = Font(name='Microsoft YaHei', size=10, color='CC0000', bold=True)
TH = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
C = Alignment(horizontal='center', vertical='center', wrap_text=True)
L = Alignment(horizontal='left', vertical='center', wrap_text=True)


def analyze_savings_rate(df: pd.DataFrame, output: str):
    """
    节资率分析 + 最优围标人数检测

    节资率 = (招标控制价 - 中标价) / 招标控制价
    越低 → 竞争越弱 → 围标概率越高
    """
    wb = Workbook()

    # Sheet 1: 节资率分析
    ws1 = wb.active
    ws1.title = '节资率分析'

    # Compute savings rate
    df['savings_rate'] = (df['招标控制价'] - df['中标价格']) / df['招标控制价']
    df['savings_pct'] = (df['savings_rate'] * 100).round(2)

    # Box plot statistics
    q1 = df['savings_rate'].quantile(0.25)
    q3 = df['savings_rate'].quantile(0.75)
    iqr = q3 - q1
    lower_fence = q1 - 1.5 * iqr
    upper_fence = q3 + 1.5 * iqr

    mean_rate = df['savings_rate'].mean()
    median_rate = df['savings_rate'].median()

    # Flag anomalies
    df['is_outlier'] = df['savings_rate'].apply(
        lambda x: 'LOW' if x < lower_fence else ('HIGH' if x > upper_fence else 'NORMAL')
    )

    # Bidder count risk (3-4 is optimal collusion size)
    df['bidder_risk'] = df['投标单位数'].apply(
        lambda n: 'HIGH' if 3 <= n <= 4 else ('MEDIUM' if 2 <= n <= 5 else 'LOW')
    )

    # Combined risk
    def combined_risk(row):
        score = 0
        if row['is_outlier'] == 'LOW':
            score += 2
        if row['bidder_risk'] == 'HIGH':
            score += 2
        elif row['bidder_risk'] == 'MEDIUM':
            score += 1
        return 'HIGH' if score >= 3 else ('MEDIUM' if score >= 2 else 'LOW')

    df['combined_risk'] = df.apply(combined_risk, axis=1)

    # Summary stats
    ws1.merge_cells('A1:H1')
    ws1['A1'] = '招标节资率分析 — 围标风险检测'
    ws1['A1'].font = Font(name='Microsoft YaHei', size=14, bold=True, color='1A3A6E')

    ws1.merge_cells('A2:H2')
    ws1[
        'A2'] = f'统计: N={len(df)} | 均值节资率={mean_rate*100:.2f}% | 中位数={median_rate*100:.2f}% | Q1={q1*100:.2f}% Q3={q3*100:.2f}% | 下界={lower_fence*100:.2f}% [低于此值为异常]'
    ws1['A2'].font = Font(name='Microsoft YaHei', size=9, italic=True, color='888888')

    # Headers
    headers = ['项目名称', '招标控制价', '中标价格', '中标单位', '投标单位数', '节资率(%)', '节资率异常', '投标人数风险',
               '综合风险']
    for c, h in enumerate(headers, 1):
        cl = ws1.cell(row=4, column=c, value=h)
        cl.font = H; cl.fill = HEADER; cl.alignment = C; cl.border = TH

    # Sort by savings rate (lowest first = most suspicious)
    df_sorted = df.sort_values('savings_rate')

    for i, (_, row) in enumerate(df_sorted.iterrows()):
        r = i + 5
        vals = [
            row.get('项目名称', ''), row['招标控制价'], row['中标价格'],
            row.get('中标单位', ''), row['投标单位数'], row['savings_pct'],
            row['is_outlier'], row['bidder_risk'], row['combined_risk']
        ]
        for c, val in enumerate(vals, 1):
            fill = None
            if c == 7:
                fill = RED if val == 'LOW' else None
            elif c == 8:
                fill = RED if val == 'HIGH' else (YEL if val == 'MEDIUM' else None)
            elif c == 9:
                fill = RED if val == 'HIGH' else (YEL if val == 'MEDIUM' else GRN if val == 'LOW' else None)
            elif c == 6:
                fill = RED if row['is_outlier'] == 'LOW' else None

            cl = ws1.cell(row=r, column=c, value=val)
            cl.font = N; cl.alignment = C; cl.border = TH
            if fill: cl.fill = fill

    # Analysis text
    ann_row = len(df) + 6
    ws1.merge_cells(f'A{ann_row}:I{ann_row}')
    ws1[f'A{ann_row}'] = '节资率分析说明'
    ws1[f'A{ann_row}'].font = Font(name='Microsoft YaHei', size=12, bold=True, color='CC0000')

    analysis = [
        f'🔴 节资率异常低(LOW): 节资率 < {lower_fence*100:.2f}% → 竞争严重不足 → 高度围标可疑',
        f'🟡 节资率低: 节资率 < {q1*100:.2f}% (Q1) → 竞争较弱 → 中度可疑',
        f'🟢 节资率正常: 节资率在 {q1*100:.2f}% ~ {q3*100:.2f}% 之间 → 正常竞争水平',
        '',
        '🔴 最优围标人数: 3-4家投标 → 获利性和可控性最高 (天津市审计局)',
        f'   本项目有{(df["投标单位数"].isin([3,4])).sum()}个项目投标人数为3-4家',
        '',
        '综合风险 = 节资率异常(HIGH:+2) + 投标人数(HIGH:+2 MEDIUM:+1)',
        '≥3分 = HIGH | 2分 = MEDIUM | <2分 = LOW',
    ]
    for j, line in enumerate(analysis):
        ws1.merge_cells(f'A{ann_row+1+j}:I{ann_row+1+j}')
        ws1[f'A{ann_row+1+j}'] = line
        ws1[f'A{ann_row+1+j}'].font = N
        ws1[f'A{ann_row+1+j}'].alignment = L

    # Sheet 2: 投标人共现分析 (支持度/置信度)
    ws2 = wb.create_sheet('投标人共现分析')

    if '投标单位' in df.columns and '项目名称' in df.columns:
        # Cross-tabulation
        bidders = df['投标单位'].unique()
        projects = df['项目名称'].unique()
        total_projects = len(projects)

        co_occur = {}
        for i, b1 in enumerate(bidders):
            b1_projects = set(df[df['投标单位'] == b1]['项目名称'])
            b1_count = len(b1_projects)
            for b2 in bidders[i + 1:]:
                b2_projects = set(df[df['投标单位'] == b2]['项目名称'])
                co_count = len(b1_projects & b2_projects)
                if co_count > 0:
                    support = co_count / total_projects  # 支持度
                    confidence = co_count / b1_count if b1_count > 0 else 0  # 置信度(A→B)
                    confidence_rev = co_count / len(b2_projects) if len(b2_projects) > 0 else 0  # 置信度(B→A)
                    co_occur[f'{b1} ↔ {b2}'] = {
                        'co_count': co_count,
                        'total_projects': total_projects,
                        'b1_count': b1_count,
                        'b2_count': len(b2_projects),
                        'support': round(support, 4),
                        'confidence_a': round(confidence, 4),
                        'confidence_b': round(confidence_rev, 4),
                    }

        ws2.merge_cells('A1:H1')
        ws2['A1'] = '投标人共现分析 — 支持度/置信度'
        ws2['A1'].font = Font(name='Microsoft YaHei', size=14, bold=True, color='1A3A6E')
        ws2.merge_cells('A2:H2')
        ws2['A2'] = f'总项目数: {total_projects} | 总投标人: {len(bidders)} | 支持度>0.5 + 置信度>0.6 → 高度关联'
        ws2['A2'].font = Font(name='Microsoft YaHei', size=9, italic=True, color='888888')

        h2 = ['投标人A', '投标人B', '共同投标次数', '总项目数', 'A投标次数', 'B投标次数',
              '支持度(Supp)', '置信度A→B(Conf)']
        for c, h in enumerate(h2, 1):
            cl = ws2.cell(row=4, column=c, value=h)
            cl.font = H; cl.fill = HEADER; cl.alignment = C; cl.border = TH

        # Sort by combined risk: support > 0.5 AND confidence > 0.6
        high_risk = []
        for pair, data in sorted(co_occur.items(),
                                  key=lambda x: (x[1]['support'] * x[1]['confidence_a']),
                                  reverse=True):
            risk_level = 'HIGH' if data['support'] > 0.5 and max(data['confidence_a'],
                                                                 data['confidence_b']) > 0.6 else (
                'MEDIUM' if data['support'] > 0.3 else 'LOW')
            data['risk'] = risk_level
            high_risk.append((pair, data))

        for i, (pair, data) in enumerate(high_risk):
            r = i + 5
            a, b = pair.split(' ↔ ')
            vals = [a, b, data['co_count'], data['total_projects'],
                    data['b1_count'], data['b2_count'],
                    data['support'], data['confidence_a']]
            for c, val in enumerate(vals, 1):
                cl = ws2.cell(row=r, column=c, value=val)
                cl.font = N; cl.alignment = C; cl.border = TH
                if c in [7, 8]:
                    cl.fill = RED if val > 0.6 else (YEL if val > 0.3 else None)

        ann_r2 = len(high_risk) + 6
        ws2.merge_cells(f'A{ann_r2}:H{ann_r2}')
        ws2[f'A{ann_r2}'] = ('参考阈值(天津市审计局):\n'
                             '支持度>0.5: 两个投标主体在所有项目中同时出现的比例过半 → 行为高度一致\n'
                             '置信度>0.6: A投标时B也随之投标的概率>60% → 联手围标可能\n'
                             '两个数值越接近1，疑点越集中。综合节资率+支持度/置信度+投标人数(3-4)进行交叉研判。')
        ws2[f'A{ann_r2}'].font = N
        ws2[f'A{ann_r2}'].alignment = L
        ws2.row_dimensions[ann_r2].height = 70

    # Column widths
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 16
    ws1.column_dimensions['C'].width = 14
    ws1.column_dimensions['D'].width = 22
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 14
    ws1.column_dimensions['G'].width = 14
    ws1.column_dimensions['H'].width = 14
    ws1.column_dimensions['I'].width = 14

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 14
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 14
    ws2.column_dimensions['F'].width = 14
    ws2.column_dimensions['G'].width = 16
    ws2.column_dimensions['H'].width = 18

    wb.save(output)
    return df


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='节资率分析 + 投标人共现(支持度/置信度)')
    parser.add_argument('--i', '--input', dest='input', required=True,
                        help='招标台账xlsx (项目名称/投标单位/中标价格/招标控制价/投标单位数)')
    parser.add_argument('--o', '--output', dest='output', default='节资率分析.xlsx',
                        help='输出Excel路径')
    args = parser.parse_args()

    df = pd.read_excel(args.input)
    analyze_savings_rate(df, args.output)
    print(f'Done: {args.output}')
