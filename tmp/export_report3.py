import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

hf = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
hfill = PatternFill(start_color='0A1F3F', end_color='0A1F3F', fill_type='solid')
tfont = Font(name='微软雅黑', bold=True, size=14, color='0A1F3F')
sf = Font(name='微软雅黑', bold=True, size=12, color='1A5C6E')
s2 = Font(name='微软雅黑', bold=True, size=11, color='1A5C6E')
nf = Font(name='微软雅黑', size=10); bf = Font(name='微软雅黑', size=10, bold=True)
rf = Font(name='微软雅黑', size=10, bold=True, color='CC0000')
yf = Font(name='微软雅黑', size=10, bold=True, color='C5955C')
gf = Font(name='微软雅黑', size=10, color='888888')
gnf = Font(name='微软雅黑', size=10, color='228B22')

p0f = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
p1f = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
p2f = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
okf = PatternFill(start_color='D9F2D9', end_color='D9F2D9', fill_type='solid')
h2f = PatternFill(start_color='1A5C6E', end_color='1A5C6E', fill_type='solid')

tb = Border(left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))
wr = Alignment(wrap_text=True, vertical='top')
ct = Alignment(horizontal='center', vertical='center', wrap_text=True)

def sh(ws, row, n, fill=hfill):
    for c in range(1, n+1):
        cl = ws.cell(row=row, column=c)
        cl.font = hf; cl.fill = fill; cl.alignment = ct; cl.border = tb

def sr(ws, row, n, font=None, fill=None):
    for c in range(1, n+1):
        cl = ws.cell(row=row, column=c)
        cl.font = font or nf; cl.alignment = wr; cl.border = tb
        if fill: cl.fill = fill

def wc(ws, r, c, v, font=None, fill=None, num=False):
    cl = ws.cell(row=r, column=c)
    cl.value = v; cl.font = font or nf; cl.alignment = wr; cl.border = tb
    if fill: cl.fill = fill
    if num: cl.number_format = '#,##0.00'

PRJ = '2024年马尔康市日部乡片区（新建）高标准农田建设（财政转移支付）项目'

# ===== Sheet 1: Summary =====
ws1 = wb.active; ws1.title = "复核总览"
ws1.merge_cells('A1:G1'); wc(ws1, 1, 1, '工程竣工财务决算审核报告 复核意见', tfont)
ws1.merge_cells('A2:G2')
wc(ws1, 2, 1, f'{PRJ} | 川融策专审[2026]第03-12号 | 2026年4月15日 | 审计小组：陈越',
   Font(name='微软雅黑', size=10, color='666666'))

r = 4; wc(ws1, r, 1, '项目基本信息', sf); r += 1
info = [
    ('项目名称', PRJ),
    ('被审核单位', '马尔康市农业农村局（负责人：贺凌）'),
    ('委托单位', '马尔康市财政局'),
    ('报告文号', '川融策专审〔2026〕第03-12号\n⚠ 同份报告第4行另有不完整"第04-号"\n⚠ 与业务用房、周转宿舍两项目报告文号完全相同！三份报告共用同一文号'),
    ('审计期间', '2026年4月1日 - 4月15日'),
    ('概算总投资', '23,957,900.00元'),
    ('实际完成投资', '20,661,459.53元（节约13.76%）'),
    ('建安工程投资', '19,600,459.53元'),
    ('待摊投资', '1,061,000.00元（设计437,000+监理466,000+检测158,000）'),
    ('待摊占比', '5.14%（<30%，正常）'),
    ('到位资金', '23,957,900.00元（100%到位）'),
    ('应付未付', '88,700.00元'),
    ('结余资金', '3,296,440.47元'),
]
for k, v in info:
    wc(ws1, r, 1, k, bf); ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    wc(ws1, r, 2, v); r += 1

r += 1; wc(ws1, r, 1, '复核统计', sf); r += 1
for c, h in enumerate(['等级', '数量', '说明'], 1):
    wc(ws1, r, c, h, hf, hfill)
sh(ws1, r, 3)
st = [('P0',6,'文号共享(三份)、设备购置/待摊投资术语混用、设计合同先于可研(三份同病)、批复文号不一致、建安超合同未说明、报告与编制说明矛盾'),
      ('P1',3,'底稿检查、结论与问题混排、决算拖延'),
      ('P2',5,'待摊测算(基本合理)、三方勾稽表、CPA签字、三报告横向比对、应付未付明细'),
      ('合计',14,'')]
for lvl, cnt, desc in st:
    r += 1; fl = rf if 'P0' in lvl else (yf if 'P1' in lvl else bf)
    fl2 = p0f if 'P0' in lvl else (p1f if 'P1' in lvl else p2f)
    wc(ws1, r, 1, lvl, fl, fl2); wc(ws1, r, 2, cnt, None, fl2)
    wc(ws1, r, 3, desc, None, fl2)

r += 2; wc(ws1, r, 1, '综合结论', sf); r += 1
ws1.merge_cells(start_row=r, start_column=1, end_row=r+3, end_column=7)
wc(ws1, r, 1, (
    '三份报告均为融策事务所出具、同一审计小组、同一出具日期。\n'
    '最核心问题是三份报告共享同一个文号"川融策专审〔2026〕第03-12号"，说明报告编号完全依靠模板自动生成、签发时未经独立核号。\n\n'
    '另一突出问题：报告第五节(一)末尾写"设备购置1,061,000.00元"，但编制说明明确"本项目不涉及设备购置"——\n'
    '实际上1,061,000元是待摊投资(设计437,000+监理466,000+检测158,000)，模板中"设备购置"字段未替换为"待摊投资"。\n\n'
    '设计合同先于可研批复(2024.6.3 vs 2024.8.12)——三份报告共同的程序合规问题。\n\n'
    '本项目比前两份有明显进步：数据底层总体准确，待摊费用在合理区间，但模板层面的系统性问题仍未解决。'
), bf)
for c, w in zip(['A','B','C','D','E','F','G'], [22,16,30,18,18,18,18]):
    ws1.column_dimensions[c].width = w

# ===== Sheet 2: DaTan Fee Standards =====
ws2 = wb.create_sheet("待摊费用标准测算")
ws2.merge_cells('A1:H1'); wc(ws2, 1, 1, '待摊投资费用 收费标准逐笔测算', tfont)
ws2.merge_cells('A2:H2')
wc(ws2, 2, 1, '设计费:计价格[2002]10号 | 监理费:发改价格[2007]670号 | 2015年后市场调节价,标准为参考基准',
   Font(name='微软雅黑', size=9, color='888888'))

# -- Design --
r = 4; wc(ws2, r, 1, '一、设计费 437,000元（含补充合同50,000）', sf); r += 1
ws2.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
wc(ws2, r, 1, '计费额: 招标控制价审定20,312,849.99 - 暂列金584,570.18 = 19,728,279.81 = 1,972.83万\n区间1000万(38.8)-3000万(103.8), 内插: 38.8+(103.8-38.8)x972.83/2000 = 70.417万\n农业林业工程专业系数0.9(I级复杂0.85): 704,170x0.9x0.85=538,691 -> 与437,000差异18.9%, 合理',
   Font(name='微软雅黑', size=9, color='666666'))

r += 2; wc(ws2, r, 1, '测算步骤', s2); r += 1
for c, h in enumerate(['步骤','参数','取值','计算','金额(元)','依据'], 1):
    wc(ws2, r, c, h, hf, hfill)
ds = [
    ('Step1','计费额','招标控制价1,972.83万\n(扣除暂列金)','20,312,849.99-584,570.18',19728279.81,'第八条'),
    ('Step2','基价','1000-3000万内插\n38.8+65x972.83/2000','=70.417万',704170.00,'基价表'),
    ('Step3a','专业系数','农业林业工程','0.9','—','附表'),
    ('Step3b','复杂度','高标准农田I级','0.85','—','附表'),
    ('Step4','设计费','I级+农业系数','704,170x0.9x0.85',538690.50,'第七条'),
]
for step, p, desc, calc, val, basis in ds:
    r += 1; wc(ws2, r, 1, step, bf); wc(ws2, r, 2, p); wc(ws2, r, 3, desc)
    wc(ws2, r, 4, calc)
    if val: wc(ws2, r, 5, val, None, None, True)
    else: wc(ws2, r, 5, '—')
    wc(ws2, r, 6, basis)

r += 2; wc(ws2, r, 1, '多情景', s2); r += 1
for c, h in enumerate(['情景','专业系数','复杂度','测算(元)','差异率','判定'], 1):
    wc(ws2, r, c, h, hf, h2f)
b = 704170.0
for name, px, cx in [('农业0.9+I级0.85',0.9,0.85),('农业0.8+I级0.85',0.8,0.85),
                      ('农业0.9+I级0.85+改扩建1.1',0.9,0.935),('农业1.0+II级1.0',1.0,1.0)]:
    r += 1; v = b*px*cx; diff = (v-437000)/437000
    wc(ws2, r, 1, name); wc(ws2, r, 2, px); wc(ws2, r, 3, cx)
    wc(ws2, r, 4, v, None, None, True); wc(ws2, r, 5, f'{diff*100:+.1f}%')
    j = '基本合理' if abs(diff)<0.25 else '偏低'
    wc(ws2, r, 6, j, gnf if abs(diff)<0.25 else yf)
r += 1; ws2.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
wc(ws2, r, 1, '>> 设计费437,000元(含补充合同50,000)在合理范围（农业工程费率和复杂度均低于建筑工程）。',
   bf, okf)

# -- Supervision --
r += 2; wc(ws2, r, 1, '二、监理费 466,000元', sf); r += 1
ws2.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
wc(ws2, r, 1, '计费额: 施工合同 19,272,268 = 1,927.23万\n1000-3000万内插: 30.1+48x927.23/2000 = 52.35万 = 523,535\nII级+高程1.1: 523,535x1.0x1.1 = 575,889\nI级+高程1.1: 523,535x0.85x1.1 = 489,505 -> vs 466,000差异4.8%, 基本吻合',
   Font(name='微软雅黑', size=9, color='666666'))

r += 2; wc(ws2, r, 1, '测算步骤', s2); r += 1
for c, h in enumerate(['步骤','参数','取值','计算','金额(元)','依据'], 1):
    wc(ws2, r, c, h, hf, hfill)
ss = [
    ('Step1','计费额','施工合同1,927.23万','19,272,268.00',19272268.00,'670号'),
    ('Step2','基价','1000-3000万内插','523,534.80',523534.80,'基价表'),
    ('Step3a','专业','农业工程1.0','1.0','—','附表'),
    ('Step3b','复杂度','高标准农田I级','0.85','—','附表'),
    ('Step3c','高程','马尔康2600m','1.1','—','第1.0.9条'),
    ('Step4','监理费','I级+高程1.1','523,535x0.85x1.1',489505.04,''),
]
for step, p, desc, calc, val, basis in ss:
    r += 1; wc(ws2, r, 1, step, bf); wc(ws2, r, 2, p); wc(ws2, r, 3, desc)
    wc(ws2, r, 4, calc)
    if val: wc(ws2, r, 5, val, None, None, True)
    else: wc(ws2, r, 5, '—')
    wc(ws2, r, 6, basis)

r += 2; wc(ws2, r, 1, '多情景', s2); r += 1
for c, h in enumerate(['情景','复杂度','高程','测算(元)','差异率','判定'], 1):
    wc(ws2, r, c, h, hf, h2f)
b2 = 523534.80
for name, cx, alt in [('I级+高程1.1',0.85,1.1),('I级+高程1.0',0.85,1.0),
                       ('II级+高程1.1',1.0,1.1),('II级+高程1.0',1.0,1.0)]:
    r += 1; v = b2*cx*alt; diff = (v-466000)/466000
    wc(ws2, r, 1, name); wc(ws2, r, 2, cx); wc(ws2, r, 3, alt)
    wc(ws2, r, 4, v, None, None, True); wc(ws2, r, 5, f'{diff*100:+.1f}%')
    j = '基本吻合' if abs(diff)<0.06 else ('偏低' if abs(diff)<0.2 else '明显偏低')
    fj = gnf if abs(diff)<0.06 else (yf if abs(diff)<0.2 else rf)
    wc(ws2, r, 6, j, fj)
r += 1; ws2.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
wc(ws2, r, 1, '>> 监理费466,000元与I级+高程1.1测算（489,505）差异仅4.8%，基本吻合。', bf, okf)

# Detection fee note
r += 2; wc(ws2, r, 1, '三、检测费 158,000元', sf); r += 1
ws2.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
wc(ws2, r, 1, '技术检测费(四川百川四维,直接委托,2025.7.9签订合同158,000元)。检测费无统一国家收费标准，由市场定价。审核报告应注明采购方式及比价情况。',
   Font(name='微软雅黑', size=10))

r += 2; wc(ws2, r, 1, '四、综合判定', sf); r += 1
for c, h in enumerate(['费用项','合同价','测算值','差异率','判定'], 1):
    wc(ws2, r, c, h, hf, hfill)
items = [('设计费',437000.00,538690.50,'-18.9%','基本合理'),
         ('监理费',466000.00,489505.04,'-4.8%','基本吻合'),
         ('检测费',158000.00,None,'—','市场定价')]
for name, act, cal, pct, j in items:
    r += 1; wc(ws2, r, 1, name); wc(ws2, r, 2, act, None, None, True)
    if cal: wc(ws2, r, 3, cal, None, None, True)
    else: wc(ws2, r, 3, '—')
    wc(ws2, r, 4, pct); wc(ws2, r, 5, j, gnf)

for i, w in enumerate([14,18,32,32,16,18,14,14]): ws2.column_dimensions[get_column_letter(i+1)].width = w

# ===== Sheet 3: Findings =====
ws3 = wb.create_sheet("复核明细"); ws3.merge_cells('A1:H1')
wc(ws3, 1, 1, '复核发现明细（共14项）', tfont)
for c, h in enumerate(['序号','等级','维度','发现','描述','依据','位置','建议'], 1):
    wc(ws3, 3, c, h, hf)
sh(ws3, 3, 8)

finds = [
('P0','文号','报告文号与业务用房、周转宿舍两项目完全相同',
 '三份不同项目报告共用"川融策专审[2026]第03-12号"。正文第4行另有不完整"第04-号"。',
 '报告基本规范\nFP-13E','正文第4-5行\n对比三份报告','三份报告立即纠正:业务用房->第03-11号,周转宿舍->第03-12号,高标准农田->第03-13号'),

('P0','模板错误','第五节(一)末尾"设备购置"应为"待摊投资"',
 '报告写"审定建筑安装工程金额为19,600,459.53元，设备购置1,061,000.00元。"但编制说明第(五)条明确"本项目不涉及设备购置"。1,061,000元实际是待摊投资(设计437,000+监理466,000+检测158,000)。模板中"设备购置"字段未替换。',
 '审计报告准确性\nFP-10A','第五节(一)末尾\nvs 编制说明第(五)条','将"设备购置"改为"待摊投资"。全报告搜索"设备"确保无残余。'),

('P0','合规','设计合同先于可研批复——三份报告同病',
 '设计中标2024.5.28,签合同2024.6.3。可研批复2024.8.12(AZN[2024]-101号)。设计合同比批复早约2个月。这是三份报告共同的程序合规问题。',
 '《政府投资条例》第9条','第三节(一)/(二)','统一核实三份报告的设计采购前置审批情况。'),

('P0','术语','可研批复文号不一致——"AZN" vs "ANZ"',
 '正文第二节第1项:"AZN[2024]-101号";正文第四节第4段:"ANZ[2024]-101号"。编制说明同样存在两个版本。"阿坝州农业农村局"缩写应统一。',
 '报告术语规范','第二节 vs 第四节\n编制说明','确认正确文号缩写并全文统一。'),

('P0','数据','建安审定金额超合同328,191.53元未说明',
 '施工合同19,272,268.00元;结算审定19,600,459.53元;超合同328,191.53(+1.7%)。报告未说明超合同原因(设计变更?工程量增加?价差?)。',
 'FP-13F 金额勾稽\n审计报告充分性','第五节(一)vs第三节','补充说明审定超合同的原因和依据。'),

('P0','矛盾','报告"设备购置106万" vs 编制说明"不涉及设备购置"',
 '同一套报告文件内出现根本性矛盾。报告第五节称有"设备购置"1,061,000元，编制说明第(五)条称"本项目不涉及设备购置"。',
 '报告一致性','审核报告五(一) vs 编制说明(五)','统一口径：本报告应写"待摊投资1,061,000元"，删除"设备购置"表述。'),

# P1
('P1','底稿','Excel底稿需检查是否混杂其他项目数据',
 '与前两份报告同批次，建议检查附件1/2是否同样存在多项目数据混杂。',
 '职业道德守则第3号','附件1/2','逐Sheet检查。'),
('P1','结构','审计结论与问题建议混排',
 '第八节末尾嵌套"进一步加强…"建议，缺少独立的问题章节。三份报告相同缺陷。',
 '报告结构规范','第八节末尾','拆分为独立"存在的问题及建议"。'),
('P1','L-1','竣工至决算报告约8个月——未追问原因',
 '初验2025.8.15->报告2026.4.15，8个月。法定3个月内。',
 'L-1 根源追问\n财建81号令第37条','第八节','追问原因并提出整改建议。'),

# P2
('P2','完整性','缺少金额三方勾稽汇总表','合同额、审定、支付数据分散未汇总。','FP-13F','报告全文','补附三方勾稽汇总表（建安/设计/监理/检测逐行）。'),
('P2','格式','注册会计师签字缺失','报告末尾两处CPA签字处均为空白横线。三份报告相同缺陷。','审计报告签署要求','报告末尾','正式出具前完成签字。'),
('P2','FP-13E','三份报告共享6项同类问题——系统性缺陷',
 '文号共享、设计合同先于可研、结论与问题混排、CPA签字缺失、底稿混杂、待摊词语混用/粘贴错误。模板管理需要彻底整改。',
 'FP-13E 同批次一致性','三份报告对比','建议：1)建立报告文号独立核发流程 2)模板中标记化占位符(如{{设备/待摊}}) 3)同批次交叉比对程序。'),
('P2','完整性','应付未付88,700元未列明细','报告第六节只给出应付未付合计金额，未逐笔列出付款对象、金额、原因。编制说明中应付明细表也为空。','审计报告完整性','报告第六节/编制说明第二节','补充应付款明细表，列出付款对象、金额、合同依据。'),
]

for i, (lvl, dim, ttl, desc, basis, loc, fix) in enumerate(finds):
    rr = 4+i
    fl = {'P0':rf,'P1':yf,'P2':gf}[lvl]; fl2 = {'P0':p0f,'P1':p1f,'P2':p2f}[lvl]
    wc(ws3, rr, 1, i+1); wc(ws3, rr, 2, lvl, fl, fl2)
    wc(ws3, rr, 3, dim); wc(ws3, rr, 4, ttl, bf)
    wc(ws3, rr, 5, desc); wc(ws3, rr, 6, basis)
    wc(ws3, rr, 7, loc); wc(ws3, rr, 8, fix)
    sr(ws3, rr, 8, fill=fl2)

for ci, w in enumerate([6,10,16,24,56,30,26,44], 1): ws3.column_dimensions[get_column_letter(ci)].width = w
ws3.row_dimensions[3].height = 25
for rr in range(4, 18): ws3.row_dimensions[rr].height = 95

# ===== Sheet 4: Cross-check + 3-report comparison =====
ws4 = wb.create_sheet("数据核对与三报告对比")
ws4.merge_cells('A1:G1'); wc(ws4, 1, 1, '数据交叉核对 & 三份报告横向对比', tfont)

r = 3; wc(ws4, r, 1, 'A. 本报告内部核对', s2); r += 1
for c, h in enumerate(['核对项','来源A','来源B','差异','结果','备注'], 1):
    wc(ws4, r, c, h, hf); sh(ws4, r, 6)
cks = [
    ('总投资','20,661,459.53','20,661,459.53','0','一致',''),
    ('建安投资','19,600,459.53','19,600,459.53','0','一致',''),
    ('建安超合同','19,600,459.53','19,272,268.00','+328,191.53','超合同','需说明原因'),
    ('待摊投资','1,061,000','1,061,000(编制说明)','0','一致','设计437+监理466+检测158'),
    ('到位资金','23,957,900','23,957,900','0','一致','100%到位'),
    ('结余资金','3,296,440.47','3,296,440.47','0','一致',''),
    ('设计费','437,000','437,000(编制说明)','0','一致',''),
    ('监理费','466,000','466,000(编制说明)','0','一致',''),
    ('检测费','158,000','158,000(编制说明)','0','一致','直接委托'),
    ('"设备购置"表述','报告:"设备购置"','编制说明:"不涉及"','','矛盾P0','模板错误'),
    ('批复文号缩写','"AZN"','"ANZ"','','矛盾P0','应统一'),
]
for item, a, b, diff, res, note in cks:
    r += 1; wc(ws4, r, 1, item, bf); wc(ws4, r, 2, str(a)); wc(ws4, r, 3, str(b))
    wc(ws4, r, 4, str(diff))
    fl = rf if '矛盾' in res or 'P0' in res else (yf if '超合同' in res else gnf)
    wc(ws4, r, 5, res, fl); wc(ws4, r, 6, note)
    fl2 = p0f if 'P0' in res else (p1f if '超合同' in res else okf)
    sr(ws4, r, 6, fill=fl2)

r += 2; wc(ws4, r, 1, 'B. 三份报告同一问题对照表', s2); r += 1
for c, h in enumerate(['问题','业务用房(报告1)','周转宿舍(报告2)','高标准农田(报告3)','判定'], 1):
    wc(ws4, r, c, h, hf); sh(ws4, r, 5)
cross = [
    ('报告文号','03-12号','03-12号','03-12号','P0:三份相同!'),
    ('不完整文号','"第04-号"','"第04-号"','"第04-号"','P0:三份相同!'),
    ('设计合同先于可研','2023.2.25vs3.10','2023.2.25vs3.10','2024.6.3vs8.12','P0:三份同病'),
    ('施工许可证滞后','8个月','9个月','N/A(农田项目)','P0(1,2) N/A(3)'),
    ('审计结论->问题混排','有','有','有','P1:三份同病'),
    ('决算拖延','近2年','近2年','约8个月','P1:严重程度递减'),
    ('CPA签字缺失','有','有','有','P2:三份同病'),
    ('模板粘贴错误','无','有(安徽正飞/67,000)','有(设备购置/AZN)','P0(2,3)'),
    ('建安超合同','无','无','有(+328,191)','P0:仅报告3'),
    ('待摊费用测算状态','偏低','基本合理','基本合理','仅报告1需关注'),
]
for item, r1, r2, r3, judge in cross:
    r += 1; wc(ws4, r, 1, item, bf)
    wc(ws4, r, 2, r1); wc(ws4, r, 3, r2); wc(ws4, r, 4, r3)
    fl = rf if 'P0' in judge else (yf if 'P1' in judge else gf)
    wc(ws4, r, 5, judge, fl)
    fl2 = p0f if 'P0' in judge else (p1f if 'P1' in judge else p2f)
    sr(ws4, r, 5, fill=fl2)

for ci, w in enumerate([28,32,32,34,28], 1): ws4.column_dimensions[get_column_letter(ci)].width = w

# ===== Sheet 5: FP Rules =====
ws5 = wb.create_sheet("FP规则检查")
ws5.merge_cells('A1:F1'); wc(ws5, 1, 1, 'FP规则 & 三报告系统性检查', tfont)
for c, h in enumerate(['规则','检查内容','阈值','实际值','判定','说明'], 1):
    wc(ws5, 3, c, h, hf)
sh(ws5, 3, 6)
fp = [
    ('FP-13A','待摊占比','>30%黄>40%红','5.14%','正常','1,061,000/20,661,459.53'),
    ('FP-13B','建安核减率','>10%黄>15%红','3.20%','正常','结算审核减率'),
    ('FP-13E','同批次一致性','文号唯一','三报告同文号','P0','03-12号重复3次'),
    ('FP-13F','金额勾稽','三环可追溯','建安超合同+328,191','P0','需说明原因'),
    ('FP-10A','术语规范','双轨不混用',"'设备购置'vs'不涉及'",'P0','术语矛盾'),
    ('FP-10B','概算vs决算','超概>10%未批','节约13.76%','正常','未超概'),
    ('L-1','根源追问','P0须分析','无','P1','8个月拖延'),
    ('新增','批复文号一致性','全文统一','AZN vs ANZ','P0','缩写矛盾'),
    ('新增','三报告文号管理','独立核发','3份共用1个号','P0','系统性缺陷'),
]
for i, (rule, content, th, actual, res, note) in enumerate(fp):
    rr = 4+i; wc(ws5, rr, 1, rule, bf); wc(ws5, rr, 2, content)
    wc(ws5, rr, 3, th); wc(ws5, rr, 4, actual)
    if 'P0' in res: fl, fl2 = rf, p0f
    elif 'P1' in res: fl, fl2 = yf, p1f
    else: fl, fl2 = gnf, okf
    wc(ws5, rr, 5, res, fl); wc(ws5, rr, 6, note); sr(ws5, rr, 6, fill=fl2)
for ci, w in enumerate([12,26,26,28,16,44], 1): ws5.column_dimensions[get_column_letter(ci)].width = w

out = r'C:\Users\scrccpa\Desktop\2024年马尔康市日部乡片区高标准农田建设项目_复核意见.xlsx'
wb.save(out)
print(f'OK: {out}')
