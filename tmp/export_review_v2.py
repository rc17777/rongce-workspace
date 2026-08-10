import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ===== Styles =====
header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='0A1F3F', end_color='0A1F3F', fill_type='solid')
title_font = Font(name='微软雅黑', bold=True, size=14, color='0A1F3F')
sub_font = Font(name='微软雅黑', bold=True, size=12, color='1A5C6E')
sub2_font = Font(name='微软雅黑', bold=True, size=11, color='1A5C6E')
normal_font = Font(name='微软雅黑', size=10)
bold_font = Font(name='微软雅黑', size=10, bold=True)
red_font = Font(name='微软雅黑', size=10, bold=True, color='CC0000')
yellow_font = Font(name='微软雅黑', size=10, bold=True, color='C5955C')
gray_font = Font(name='微软雅黑', size=10, color='888888')
green_font = Font(name='微软雅黑', size=10, color='228B22')

p0_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
p1_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
p2_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
ok_fill = PatternFill(start_color='D9F2D9', end_color='D9F2D9', fill_type='solid')
header2_fill = PatternFill(start_color='1A5C6E', end_color='1A5C6E', fill_type='solid')

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap = Alignment(wrap_text=True, vertical='top')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_header(ws, row, ncols, fill=header_fill):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = center
        cell.border = thin_border

def style_row(ws, row, ncols, font=None, fill=None):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        if font: cell.font = font
        else: cell.font = normal_font
        cell.alignment = wrap
        cell.border = thin_border
        if fill: cell.fill = fill

def write_cell(ws, r, c, val, font=None, fill=None, num=False):
    cell = ws.cell(row=r, column=c)
    cell.value = val
    if font: cell.font = font
    else: cell.font = normal_font
    cell.alignment = wrap
    cell.border = thin_border
    if fill: cell.fill = fill
    if num: cell.number_format = '#,##0.00'

# =========================================
# Sheet 1: Review Summary
# =========================================
ws1 = wb.active
ws1.title = "复核总览"

ws1.merge_cells('A1:G1')
write_cell(ws1, 1, 1, '工程竣工财务决算审核报告 复核意见', title_font)
ws1.merge_cells('A2:G2')
write_cell(ws1, 2, 1, '阿坝州税务局业务用房维修改造项目 | 融策专审2026第03-12号 | 2026年4月15日 | 审计小组：陈越',
           Font(name='微软雅黑', size=10, color='666666'))

# Basic info block
r = 4
write_cell(ws1, r, 1, '项目基本信息', sub_font)
r += 1
info_data = [
    ('项目名称', '阿坝州税务局业务用房维修改造项目'),
    ('被审计单位', '阿坝州税务局'),
    ('委托单位', '阿坝州财政局'),
    ('报告文号', '融策专审2026第03-12号（⚠ 同报告中另出现\"第04-号\"不完整文号）'),
    ('审计小组', '陈越 / 15184481037'),
    ('审计期间', '2026年4月1日 - 4月15日'),
    ('报告出具日', '2026年4月15日'),
    ('概算总投资', '1,350,000.00元'),
    ('实际完成投资', '1,123,669.42元（节约16.77%）'),
    ('建安工程投资', '863,684.35元'),
    ('设备投资', '192,985.07元'),
    ('待摊投资', '67,000.00元（设计费40,000 + 监理费27,000）'),
    ('待摊投资占比', '5.96%（<30%阈值，正常）'),
    ('建安费核减率', '1.17%（<10%阈值，正常）'),
    ('到位资金', '1,201,500.00元'),
    ('应付未付', '0.00元'),
    ('结余资金', '77,830.58元（已由财政收回）'),
]
for k, v in info_data:
    write_cell(ws1, r, 1, k, bold_font)
    ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    write_cell(ws1, r, 2, v)
    r += 1

# Summary stats
r += 1
write_cell(ws1, r, 1, '复核结果统计', sub_font)
r += 1
for c, h in enumerate(['等级', '数量', '占比', '主要类型', '说明'], 1):
    write_cell(ws1, r, c, h, header_font, header_fill, False)
style_header(ws1, r, 5)

stats = [
    ('P0 致命', 4, '29%', '格式+合规+数据+审计程序',
     '文号矛盾、无证施工未作评价、设计合同先于可研批复、建安/设备数据互倒、待摊费用未按标准测算'),
    ('P1 重要', 4, '29%', '底稿+时效+结构+数据',
     '底稿混杂其他项目、决算拖延2年无原因分析、结论与问题混排、合同额填写错误'),
    ('P2 建议', 5, '36%', '完整性+格式+规范',
     '缺三方勾稽表、CPA签字缺失、许可证号未引、采购方式不一致、高程系数未考量'),
    ('合计', 13, '100%', '', ''),
]
for lvl, cnt, pct, tp, desc in stats:
    r += 1
    fl = red_font if 'P0' in lvl else yellow_font if 'P1' in lvl else bold_font
    fl2 = p0_fill if 'P0' in lvl else p1_fill if 'P1' in lvl else p2_fill
    write_cell(ws1, r, 1, lvl, fl, fl2)
    write_cell(ws1, r, 2, cnt, None, fl2)
    write_cell(ws1, r, 3, pct, None, fl2)
    write_cell(ws1, r, 4, tp, None, fl2)
    write_cell(ws1, r, 5, desc, None, fl2)

# Conclusion
r += 2
write_cell(ws1, r, 1, '综合结论', sub_font)
r += 1
ws1.merge_cells(start_row=r, start_column=1, end_row=r+3, end_column=7)
write_cell(ws1, r, 1, (
    '数据底层基本正确——全链路合同→审定→支付勾稽无误，合计层面各表一致。\n'
    '但格式、合规、审计程序层面问题密集：P0 4项 + P1 4项 = 8项不可直接出具。\n'
    '新增P0：两笔待摊费用（设计费、监理费）均未按国家收费标准做测算对比，审计程序执行不充分。'
), bold_font)

ws1.column_dimensions['A'].width = 22
ws1.column_dimensions['B'].width = 16
ws1.column_dimensions['C'].width = 10
for c in ['D','E','F','G']:
    ws1.column_dimensions[c].width = 18

# =========================================
# Sheet 2: DaTan Fee Standards Calculation
# =========================================
ws2 = wb.create_sheet("待摊费用标准测算")

ws2.merge_cells('A1:H1')
write_cell(ws2, 1, 1, '待摊投资费用 收费标准逐笔测算', title_font)
ws2.merge_cells('A2:H2')
write_cell(ws2, 2, 1, '设计费依据：计价格[2002]10号 | 监理费依据：发改价格[2007]670号 | 注：2015年后已放开为市场调节价，标准为参考基准', 
           Font(name='微软雅黑', size=9, color='888888'))

# ---- 设计费 ----
r = 4
write_cell(ws2, r, 1, '一、设计费 — 计价格[2002]10号《工程勘察设计收费管理规定》', sub_font)
r += 1
write_cell(ws2, r, 1, '计算公式：设计费 = 收费基价 × 专业调整系数 × 工程复杂程度调整系数 × 附加调整系数 × (1±浮动幅度)',
           Font(name='微软雅黑', size=10, italic=True, color='666666'))
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)

r += 2
write_cell(ws2, r, 1, '测算步骤', sub2_font)
r += 1
for c, h in enumerate(['步骤', '参数名称', '取值/说明', '计算过程', '金额(元)', '依据条款'], 1):
    write_cell(ws2, r, c, h, header_font, header_fill)

design_steps = [
    ('Step1', '计费额', '经批准的初步设计概算中建安费+设备费+联合试运转费\n概算135万，招标控制价134.67万，扣除暂列金4.84万', 
     '1,346,697.04 - 48,422.10', 1298274.94, '计价格[2002]10号 第八条'),
    ('Step2', '收费基价', '计费额129.83万，处于0~200万区间\n费率 = 9.0÷200 = 0.045（万元基价/万元计费额）',
     '129.83 × 0.045 = 5.8424万', 58423.50, '《工程设计收费基价表》内插法'),
    ('Step3a', '专业调整系数', '建筑、市政工程', '1.0', 1.0, '计价格[2002]10号 附表'),
    ('Step3b', '工程复杂程度', '含结构加固的维修改造，取II级', '1.0（若一般装修I级=0.85）', 1.0, '计价格[2002]10号 附表'),
    ('Step3c', '附加调整系数', '改扩建和技术改造项目，取1.1（可取值1.1~1.4）', '1.1', 1.1, '计价格[2002]10号 第1.0.12条'),
    ('Step4', '基本设计费（主情景）', 'II级1.0 + 改扩建1.1', '58,423.50 × 1.0 × 1.0 × 1.1', 64265.85, '计价格[2002]10号 第七条'),
    ('Step5', '浮动下限（-20%）', '市场调节可下浮20%', '64,265.85 × 0.8', 51412.68, '计价格[2002]10号 第九条'),
]

for i, (step, param, desc, calc, val, basis) in enumerate(design_steps):
    r += 1
    write_cell(ws2, r, 1, step, bold_font)
    write_cell(ws2, r, 2, param)
    write_cell(ws2, r, 3, desc)
    write_cell(ws2, r, 4, calc)
    write_cell(ws2, r, 5, val, None, None, isinstance(val, float))
    write_cell(ws2, r, 6, basis)
    if step == 'Step4':
        style_row(ws2, r, 6, font=bold_font, fill=p1_fill)
    elif step == 'Step5':
        style_row(ws2, r, 6, fill=p2_fill)

# 多情景对比
r += 2
write_cell(ws2, r, 1, '多情景对比', sub2_font)
r += 1
for c, h in enumerate(['情景', '复杂度', '改扩建系数', '测算金额(元)', '浮动下限(元)', '实付40,000差异(元)', '差异率', '判定'], 1):
    write_cell(ws2, r, c, h, header_font, header2_fill)

ds = [
    ('I级 + 无改扩建附加', 0.85, 1.0, 49659.98, 39727.98, 9659.98, '24.1%', '⚠ 偏离'),
    ('I级 + 改扩建1.1', 0.85, 1.1, 54625.97, 43700.78, 14625.97, '36.6%', '❌ 偏低'),
    ('II级 + 改扩建1.1（主情景）', 1.0, 1.1, 64265.85, 51412.68, 24265.85, '60.7%', '❌ 明显偏低'),
    ('II级 + 改扩建1.2', 1.0, 1.2, 70108.20, 56086.56, 30108.20, '75.3%', '❌ 明显偏低'),
    ('II级 + 改扩建1.3', 1.0, 1.3, 75950.55, 60760.44, 35950.55, '89.9%', '❌ 明显偏低'),
]
for name, cx, ex, val, lo, diff, pct, judge in ds:
    r += 1
    write_cell(ws2, r, 1, name)
    write_cell(ws2, r, 2, cx)
    write_cell(ws2, r, 3, ex)
    write_cell(ws2, r, 4, val, None, None, True)
    write_cell(ws2, r, 5, lo, None, None, True)
    write_cell(ws2, r, 6, diff, None, None, True)
    write_cell(ws2, r, 7, pct)
    fl = red_font if '明显偏低' in judge else yellow_font
    write_cell(ws2, r, 8, judge, fl)

# 设计费结论
r += 2
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
write_cell(ws2, r, 1, '▶ 设计费结论：即使最保守情景（I级+无改扩建）测算49,660元，仍比实付40,000元高24.1%。审核报告未做任何标准测算对比，审计程序不充分。',
           bold_font, ok_fill)

# ---- 监理费 ----
r += 2
write_cell(ws2, r, 1, '二、监理费 — 发改价格[2007]670号《建设工程监理与相关服务收费管理规定》', sub_font)
r += 1
write_cell(ws2, r, 1, '计算公式：监理费 = 收费基价 × 专业调整系数 × 工程复杂程度调整系数 × 高程调整系数 × (1±浮动幅度)',
           Font(name='微软雅黑', size=10, italic=True, color='666666'))
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)

r += 2
write_cell(ws2, r, 1, '测算步骤', sub2_font)
r += 1
for c, h in enumerate(['步骤', '参数名称', '取值/说明', '计算过程', '金额(元)', '依据条款'], 1):
    write_cell(ws2, r, c, h, header_font, header_fill)

supervise_steps = [
    ('Step1', '计费额', '施工合同额（建安+设备）= 106.92万元（<基价表最低档500万）',
     '1,069,200.00', 1069200.00, '发改价格[2007]670号'),
    ('Step2', '收费基价', '低于500万，按川内惯例 16.5÷500=3.3%费率计算',
     '106.92 × 3.3% = 3.5284万', 35283.60, '施工监理服务收费基价表 内插'),
    ('Step3a', '专业调整系数', '建筑、市政工程', '1.0', 1.0, '发改价格[2007]670号 附表'),
    ('Step3b', '工程复杂程度', '含结构加固，取II级', '1.0（一般装修I级=0.85）', 1.0, '发改价格[2007]670号 附表'),
    ('Step3c', '高程调整系数', '阿坝州马尔康市海拔约2600m（区间2001~3000m）', '1.1', 1.1, '发改价格[2007]670号 第1.0.9条'),
    ('Step4', '监理费（主情景）', 'II级1.0 + 高程1.1', '35,283.60 × 1.0 × 1.0 × 1.1', 38811.96, '发改价格[2007]670号'),
    ('Step5', '浮动下限（-20%）', '市场调节可下浮20%', '38,811.96 × 0.8', 31049.57, '发改价格[2007]670号'),
]
for step, param, desc, calc, val, basis in supervise_steps:
    r += 1
    write_cell(ws2, r, 1, step, bold_font)
    write_cell(ws2, r, 2, param)
    write_cell(ws2, r, 3, desc)
    write_cell(ws2, r, 4, calc)
    write_cell(ws2, r, 5, val, None, None, isinstance(val, float))
    write_cell(ws2, r, 6, basis)
    if step == 'Step4':
        style_row(ws2, r, 6, font=bold_font, fill=p1_fill)
    elif step == 'Step5':
        style_row(ws2, r, 6, fill=p2_fill)

# 监理多情景
r += 2
write_cell(ws2, r, 1, '多情景对比', sub2_font)
r += 1
for c, h in enumerate(['情景', '复杂度', '高程系数', '测算金额(元)', '浮动下限(元)', '实付27,000差异(元)', '差异率', '判定'], 1):
    write_cell(ws2, r, c, h, header_font, header2_fill)

ss = [
    ('I级 + 高程1.0（最保守）', 0.85, 1.0, 29991.06, 23992.85, 2991.06, '11.1%', '⚠ 基本吻合'),
    ('I级 + 高程1.1', 0.85, 1.1, 32990.17, 26392.13, 5990.17, '22.2%', '⚠ 偏低'),
    ('II级 + 高程1.0', 1.0, 1.0, 35283.60, 28226.88, 8283.60, '30.7%', '❌ 偏低'),
    ('II级 + 高程1.1（主情景）', 1.0, 1.1, 38811.96, 31049.57, 11811.96, '43.7%', '❌ 明显偏低'),
]
for name, cx, alt, val, lo, diff, pct, judge in ss:
    r += 1
    write_cell(ws2, r, 1, name)
    write_cell(ws2, r, 2, cx)
    write_cell(ws2, r, 3, alt)
    write_cell(ws2, r, 4, val, None, None, True)
    write_cell(ws2, r, 5, lo, None, None, True)
    write_cell(ws2, r, 6, diff, None, None, True)
    write_cell(ws2, r, 7, pct)
    fl = red_font if '明显偏低' in judge else yellow_font
    write_cell(ws2, r, 8, judge, fl)

# 监理费结论
r += 2
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
write_cell(ws2, r, 1, '▶ 监理费结论：主情景测算38,812元，实付27,000元（低于浮动下限31,050元）。最保守情景（I级+高程1.0）测算29,991元与实付基本吻合，但海拔2600m应当取高程系数1.1。审核报告未做任何标准测算，审计程序不充分。',
           bold_font, ok_fill)

# 综合
r += 2
write_cell(ws2, r, 1, '三、综合判定', sub_font)
r += 1
for c, h in enumerate(['费用项', '合同价', '测算值(主情景)', '差异', '差异率', '浮动下限', '是否在范围内', '判定'], 1):
    write_cell(ws2, r, c, h, header_font, header_fill)

summary_items = [
    ('设计费', 40000.00, 64265.85, 24265.85, '-37.8%', 51412.68, '否', '❌ 低于所有情景'),
    ('监理费', 27000.00, 38811.96, 11811.96, '-30.4%', 31049.57, '否', '⚠ 最保守情景可接受'),
]
for name, actual, calc, diff, pct, lo, in_range, judge in summary_items:
    r += 1
    write_cell(ws2, r, 1, name)
    write_cell(ws2, r, 2, actual, None, None, True)
    write_cell(ws2, r, 3, calc, None, None, True)
    write_cell(ws2, r, 4, diff, None, None, True)
    write_cell(ws2, r, 5, pct)
    write_cell(ws2, r, 6, lo, None, None, True)
    write_cell(ws2, r, 7, in_range)
    write_cell(ws2, r, 8, judge, red_font if '❌' in judge else yellow_font)

# 审计依据
r += 2
write_cell(ws2, r, 1, '四、审计依据', sub_font)
r += 1
ws2.merge_cells(start_row=r, start_column=1, end_row=r+2, end_column=8)
write_cell(ws2, r, 1, (
    '《中国注册会计师审计准则第1301号——审计证据》第六条：\n'
    '"注册会计师应当获取充分、适当的审计证据，以将审计风险降至可接受的低水平。"\n\n'
    '审核报告仅罗列设计费40,000元、监理费27,000元，未执行"按标准测算→对比→判断合理性"的审计程序。\n'
    '虽发改价格[2015]299号已将设计费、监理费放开为市场调节价（上述标准为参考基准，非强制），\n'
    '但作为审计机构，至少应完成标准测算并加以分析，而非仅抄录发包方提供的合同金额。\n\n'
    '建议补充：(1)说明采购方式及比选报价情况；(2)按标准测算；(3)对差异做合理性分析；(4)给出审计结论。'
), bold_font)

for i, w in enumerate([14, 18, 32, 32, 16, 20, 14, 16]):
    ws2.column_dimensions[get_column_letter(i+1)].width = w

# =========================================
# Sheet 3: Detailed Findings
# =========================================
ws3 = wb.create_sheet("复核明细")
ws3.merge_cells('A1:H1')
write_cell(ws3, 1, 1, '复核发现明细（共13项）', title_font)

for c, h in enumerate(['序号', '风险等级', '复核维度', '发现标题', '问题描述', '规则依据', '原文/位置', '修改建议'], 1):
    write_cell(ws3, 3, c, h, header_font)
style_header(ws3, 3, 8)

findings = [
    # P0 - 5 items
    ('P0', '格式', '报告文号双重显示且不一致',
     '同一份报告第4行显示"融策专审2026第04-号"（不完整），第5行显示"融策专审2026第03-12号"，两个不同文号并存。',
     '报告基本格式要求', '报告正文第4-5行', '确认正确文号，删除错误版本。'),
    ('P0', 'FP-13C 合规', '未批先建——施工许可证滞后约8个月',
     '实际开工2023年7月1日，施工许可证2024年3月1日，无证施工约8个月。报告仅平铺事实，未作合规性评价。',
     '《建筑法》第7条/《建筑工程施工许可管理办法》第3条/FP-13C',
     '报告第六节第6项', '在"存在的问题"章节披露此事项，说明是否已受行政处罚及补办手续。'),
    ('P0', '合规', '设计合同签订早于可研批复',
     '设计合同2023年2月25日，可研批复2023年3月10日（阿州发改行审〔2023〕79号），合同比批复早13天。',
     '《政府投资条例》第9条', '报告第三节第(二)项第3条', '核实是否有前置依据，无则应在报告中说明。'),
    ('P0', 'FP-13F 数据', '附件1决算报表数据与审核报告不一致（建安/设备互倒604.93元）',
     '附件1显示建安863,079.42、设备193,590.00；审核报告显示建安863,684.35、设备192,985.07。合计一致但分项互倒604.93元，未说明调整原因。',
     'FP-13F 金额三方勾稽', '附件1 vs 审核报告第五节', '补充建安与设备投资之间调整604.93元的依据。'),
    ('P0', '审计程序', '待摊投资未按收费标准测算——审计程序执行不充分',
     '设计费40,000元：即使最保守标准测算（I级+无改扩建）49,660元，差距24.1%。监理费27,000元：主情景测算38,812元，差距30.4%。审核报告仅罗列金额，未做"按标准测算→对比→判断合理性"的审计程序。',
     '《中国注册会计师审计准则第1301号》第6条/计价格[2002]10号/发改价格[2007]670号',
     '报告第五节/第三节第(二)项', '补充：(1)逐项按标准测算；(2)采购方式及报价情况；(3)差异合理性分析；(4)审计结论。'),

    # P1 - 4 items
    ('P1', '底稿管理', 'Excel底稿混杂其他项目数据',
     '附件1"审计记录"Sheet含"九寨沟双河镇松柏村片区产业发展项目"完整台账（773.63万），"审计记录(2)"含另一饮水项目数据。单一项目审计产品底稿混杂其他项目，涉及信息隔离风险。',
     '《中国注册会计师职业道德守则》第3号 保密', '附件1 审计记录Sheet', '拆分Excel，每个项目独立建档。'),
    ('P1', 'L-1 时效', '竣工至决算报告间隔近2年，无原因分析',
     '竣工验收2024年4月24日→决算报告2026年4月15日，间隔近2年。法定要求3个月内完成（《基本建设财务规则》第37条）。报告仅一句话带过，无原因分析、无根源追问。',
     'L-1 根源追问/财建81号令第37条', '报告第八节第(一)项', '追问拖延原因（结算拖延/管理缺位/审批积压），对症提出整改建议。'),
    ('P1', '结构', '审计结论与问题建议混排',
     '"八、审计结论"末尾嵌套"（一）加强竣工结算及竣工财务决算工作..."，问题/建议混入肯定性结论。',
     '报告结构规范', '报告第八节末尾', '拆分为独立"存在的问题及建议"章节。'),
    ('P1', '数据', '附件2审计记录合同额填写错误',
     '附件2"审计记录"Sheet施工合同额填写为1,056,669.42（审定金额），而非实际合同额1,069,200.00，导致核减额显示为0。',
     '审计底稿准确性', '附件2 审计记录Sheet', '更正送审额为实际合同额1,069,200.00。'),

    # P2 - 4 items
    ('P2', 'FP-13F 完整性', '缺少金额三方勾稽汇总表',
     '合同额、审定、支付数据分散，未汇总为一览表。',
     'FP-13F', '报告全文', '补附三方勾稽汇总表（建安/设备/设计/监理逐行）。'),
    ('P2', '格式', '注册会计师签字缺失',
     '报告末尾两处CPA签字处均为空白横线。',
     '审计报告签署要求', '报告末尾', '正式出具前完成签字。'),
    ('P2', '规范', '施工许可证编号未引用',
     '仅提及核发日期，未引证许可证编号513229202403210201。',
     '报告引用规范', '报告第六节第6项', '补充施工许可证编号。'),
    ('P2', '表述', '监理费采购方式标注不一致',
     '审核说明写"询价"，附件2审计记录写"直接委托"，口径不统一。',
     'FP-10A 双轨术语', '审核说明 vs 附件2', '统一监理费采购方式表述。'),
]

for i, (lvl, dim, title, desc, rule, source, fix) in enumerate(findings):
    r = 4 + i
    fl = {'P0': red_font, 'P1': yellow_font, 'P2': gray_font}[lvl]
    fl2 = {'P0': p0_fill, 'P1': p1_fill, 'P2': p2_fill}[lvl]
    write_cell(ws3, r, 1, i+1)
    write_cell(ws3, r, 2, lvl, fl, fl2)
    write_cell(ws3, r, 3, dim)
    write_cell(ws3, r, 4, title, bold_font)
    write_cell(ws3, r, 5, desc)
    write_cell(ws3, r, 6, rule)
    write_cell(ws3, r, 7, source)
    write_cell(ws3, r, 8, fix)
    style_row(ws3, r, 8, fill=fl2)

for i, w in enumerate([6, 10, 14, 20, 52, 32, 28, 42]):
    ws3.column_dimensions[get_column_letter(i+1)].width = w
ws3.row_dimensions[3].height = 25
for r in range(4, 4+len(findings)):
    ws3.row_dimensions[r].height = 90

# =========================================
# Sheet 4: Data Cross-Check
# =========================================
ws4 = wb.create_sheet("数据交叉核对")
ws4.merge_cells('A1:F1')
write_cell(ws4, 1, 1, '数据一致性交叉核对', title_font)

for c, h in enumerate(['核对项', '来源A（审核报告）', '来源B（附件/附表）', '差异', '结果', '备注'], 1):
    write_cell(ws4, 3, c, h, header_font)
style_header(ws4, 3, 6)

checks = [
    ('总投资', '1,123,669.42', '1,123,669.42', '0', '✅ 一致', ''),
    ('到位资金', '1,201,500.00', '1,201,500.00', '0', '✅ 一致', ''),
    ('结余资金', '77,830.58', '77,830.58', '0', '✅ 一致', '已由财政收回'),
    ('待摊投资合计', '67,000.00', '67,000.00', '0', '✅ 一致', '设计40,000+监理27,000'),
    ('建安+设备合计', '1,056,669.42', '1,056,669.42', '0', '✅ 一致', ''),
    ('建安工程分项', '863,684.35', '863,079.42（附件1）', '604.93', '❌ 不一致', 'P0-4：需补充调整说明'),
    ('设备投资分项', '192,985.07', '193,590.00（附件1）', '-604.93', '❌ 不一致', 'P0-4：与建安互倒'),
    ('应付未付', '0.00', '0.00', '0', '✅ 一致', ''),
    ('设计费 合同→审定→支付', '40,000→40,000→40,000', '40,000→40,000→40,000', '0', '✅ 一致', '三环完整'),
    ('监理费 合同→审定→支付', '27,000→27,000→27,000', '27,000→27,000→27,000', '0', '✅ 一致', '三环完整'),
    ('待摊分摊 建安(54,763.44)', '54,763.44', '54,763.44（精确计算验证）', '0', '✅ 一致', '计价格[2002]10号验证通过'),
    ('待摊分摊 设备(12,236.56)', '12,236.56', '12,236.56（精确计算验证）', '0', '✅ 一致', '发改价格[2007]670号验证通过'),
    ('设计费 vs 标准测算', '40,000', '标准测算49,660~64,266', '9,660~24,266', '⚠ 低于标准', '审计程序不充分'),
    ('监理费 vs 标准测算', '27,000', '标准测算29,991~38,812', '2,991~11,812', '⚠ 低于标准', '审计程序不充分'),
]

for i, (item, a, b, diff, result, note) in enumerate(checks):
    r = 4 + i
    write_cell(ws4, r, 1, item, bold_font)
    write_cell(ws4, r, 2, a)
    write_cell(ws4, r, 3, b)
    write_cell(ws4, r, 4, diff)
    write_cell(ws4, r, 5, result)
    write_cell(ws4, r, 6, note)
    fl = ok_fill if '一致' in result else (p1_fill if '⚠' in result else p0_fill)
    style_row(ws4, r, 6, fill=fl)

for i, w in enumerate([28, 26, 28, 20, 14, 44]):
    ws4.column_dimensions[get_column_letter(i+1)].width = w

# =========================================
# Sheet 5: FP Rules Checklist
# =========================================
ws5 = wb.create_sheet("FP规则检查")
ws5.merge_cells('A1:F1')
write_cell(ws5, 1, 1, '工程竣工财务决算审核 FP-10 / FP-13 / L-1~L-2 规则检查清单', title_font)

for c, h in enumerate(['规则ID', '检查内容', '阈值/标准', '实际值', '判定', '说明'], 1):
    write_cell(ws5, 3, c, h, header_font)
style_header(ws5, 3, 6)

fp_checks = [
    ('FP-13A', '待摊投资占比', '>30%标黄 >40%标红', '5.96%', '✅ 正常', '67,000/1,123,669.42'),
    ('FP-13B', '建安费核减率', '>10%标黄 >15%标红', '1.17%', '✅ 正常', '(1,069,200-1,056,669.42)/1,069,200'),
    ('FP-13C', '矛盾短语对检测', '互斥短语同现即标红', '施工许可证滞后8月', '❌ P0', '报告陈述事实但未作合规评价'),
    ('FP-13D', '终止项目合同链完整性', '已确定施工/监理+终止→必须说明', 'N/A', '—', '本项目非终止项目'),
    ('FP-13E', '同批次形式一致性', '文号/日期/截止日格式统一', '同一报告内两文号', '❌ P0', '04-号 vs 03-12号'),
    ('FP-13F', '金额三方勾稽完整性', '合同额→审定→支付三环可追溯', '附件1建安/设备互倒604.93', '❌ P0', '合计一致，分项不一致需补充说明'),
    ('FP-10A', '双轨术语允许', '工程/财务术语双轨不标错', '—', '✅ 正常', ''),
    ('FP-10B', '概算vs决算对比', '超概>10%且未批→P0', '节约16.77%', '✅ 正常', '未超概'),
    ('FP-10C', '签证变更合理性', '签证>30%标P1', 'N/A', '—', '维修改造，未见签证变更'),
    ('L-1', '根源追问', 'P0问题须有根源分析', '决算拖延无原因分析', '❌ P1', '竣工2年才出报告，原因未追问'),
    ('L-2', '微观→中观跃迁', '≥2同类问题应上升审视', '待摊测算缺失+决算拖延 = 同类程序缺失', '❌ P1', '应上升到审计程序充分性层面'),
    ('新增', '待摊费用标准测算', '应执行标准测算审计程序', '设计/监理费均低于标准30%+', '❌ P0', '《审计准则1301号》审计证据充分性'),
]

for i, (rule, content, threshold, actual, result, note) in enumerate(fp_checks):
    r = 4 + i
    write_cell(ws5, r, 1, rule, bold_font)
    write_cell(ws5, r, 2, content)
    write_cell(ws5, r, 3, threshold)
    write_cell(ws5, r, 4, actual)
    fl = red_font if 'P0' in result else (yellow_font if 'P1' in result else green_font if '✅' in result else None)
    write_cell(ws5, r, 5, result, fl)
    write_cell(ws5, r, 6, note)
    fl2 = p0_fill if 'P0' in result else (p1_fill if 'P1' in result else (ok_fill if '✅' in result else None))
    style_row(ws5, r, 6, fill=fl2)

for i, w in enumerate([10, 26, 28, 28, 16, 44]):
    ws5.column_dimensions[get_column_letter(i+1)].width = w

# ===== Save =====
output = r'C:\Users\scrccpa\Desktop\阿坝州税务局业务用房维修改造项目_复核意见_v2.xlsx'
wb.save(output)
print(f'Saved: {output}')
