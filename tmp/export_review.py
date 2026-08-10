import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ===== Styles =====
header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='0A1F3F', end_color='0A1F3F', fill_type='solid')
title_font = Font(name='微软雅黑', bold=True, size=14, color='0A1F3F')
sub_font = Font(name='微软雅黑', bold=True, size=11, color='1A5C6E')
normal_font = Font(name='微软雅黑', size=10)
bold_font = Font(name='微软雅黑', size=10, bold=True)
red_font = Font(name='微软雅黑', size=10, bold=True, color='CC0000')
yellow_font = Font(name='微软雅黑', size=10, bold=True, color='C5955C')
gray_font = Font(name='微软雅黑', size=10, color='888888')

p0_fill = PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid')
p1_fill = PatternFill(start_color='FFF5E0', end_color='FFF5E0', fill_type='solid')
p2_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
ok_fill = PatternFill(start_color='E0FFE0', end_color='E0FFE0', fill_type='solid')
check_fill = PatternFill(start_color='F5F2EC', end_color='F5F2EC', fill_type='solid')

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap = Alignment(wrap_text=True, vertical='top')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_header(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

def style_row(ws, row, ncols, font=normal_font, fill=None):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.alignment = wrap
        cell.border = thin_border
        if fill:
            cell.fill = fill

# =========================================
# Sheet 1: Review Summary
# =========================================
ws1 = wb.active
ws1.title = "复核总览"

ws1.merge_cells('A1:G1')
ws1.cell(row=1, column=1, value='工程竣工财务决算审核报告 复核意见').font = title_font

ws1.merge_cells('A2:G2')
ws1.cell(row=2, column=1, value='阿坝州税务局业务用房维修改造项目 | 融策专审2026第03-12号 | 2026年4月15日').font = Font(name='微软雅黑', size=10, color='666666')

# Basic info
info_data = [
    ('项目名称', '阿坝州税务局业务用房维修改造项目'),
    ('报告文号', '融策专审2026第03-12号（注意：报告中同时出现"第04-号"，见复核明细）'),
    ('审计小组', '陈越 / 15184481037'),
    ('出具日期', '2026年4月15日'),
    ('概算总投资', '1,350,000.00元'),
    ('实际完成投资', '1,123,669.42元（节约16.77%）'),
    ('建安工程投资', '863,684.35元'),
    ('设备投资', '192,985.07元'),
    ('待摊投资', '67,000.00元（设计费40,000 + 监理费27,000）'),
    ('到位资金', '1,201,500.00元'),
    ('应付未付', '0.00元'),
    ('结余资金', '77,830.58元（已收回）'),
]

ws1.cell(row=4, column=1, value='项目基本信息').font = sub_font
for i, (k, v) in enumerate(info_data):
    ws1.cell(row=5+i, column=1, value=k).font = bold_font
    ws1.merge_cells(start_row=5+i, start_column=2, end_row=5+i, end_column=5)
    ws1.cell(row=5+i, column=2, value=v).font = normal_font

# Summary stats
r = 5 + len(info_data) + 1
ws1.cell(row=r, column=1, value='复核结果统计').font = sub_font
r += 1
stats_headers = ['风险等级', '数量', '占比', '说明']
for c, h in enumerate(stats_headers, 1):
    ws1.cell(row=r, column=c, value=h)
style_header(ws1, r, 4)

stats = [
    ('P0 致命', 4, '33%', '文号矛盾、无证施工、设计合同早于可研、建安/设备数据互倒'),
    ('P1 重要', 4, '33%', '底稿混杂、决算拖延、结论与问题混排、合同额填写错误'),
    ('P2 建议', 4, '33%', '缺少勾稽表、CPA签字缺失、许可证号未引、采购方式不一致'),
    ('合计', 12, '100%', ''),
]
for i, (lvl, cnt, pct, desc) in enumerate(stats):
    r2 = r + 1 + i
    ws1.cell(row=r2, column=1, value=lvl).font = bold_font
    ws1.cell(row=r2, column=2, value=cnt).font = normal_font
    ws1.cell(row=r2, column=3, value=pct).font = normal_font
    ws1.cell(row=r2, column=4, value=desc).font = normal_font
    fill = {'P0 致命': p0_fill, 'P1 重要': p1_fill, 'P2 建议': p2_fill}.get(lvl, None)
    style_row(ws1, r2, 4, fill=fill)

ws1.column_dimensions['A'].width = 22
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 42

# =========================================
# Sheet 2: Detailed Findings
# =========================================
ws2 = wb.create_sheet("复核明细")

ws2.merge_cells('A1:H1')
ws2.cell(row=1, column=1, value='复核发现明细').font = title_font

headers2 = ['序号', '风险等级', '复核维度', '发现标题', '问题描述', '规则依据', '原文/位置', '修改建议']
for c, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=c, value=h)
style_header(ws2, 3, 8)

findings = [
    # P0
    ('P0', '格式', '报告文号双重显示且不一致',
     '同一份报告第4行显示"融策专审2026第 04- 号"（不完整），第5行显示"融策专审2026第03-12号"，两个不同文号并存。',
     '报告基本格式要求', '报告正文第4-5行', '确认正确文号，删除错误版本。'),
    ('P0', 'FP-13C 合规', '未批先建——施工许可证滞后约8个月',
     '实际开工时间2023年7月1日，施工许可证核发日期2024年3月1日，无证施工约8个月。报告仅平铺陈述事实，未作任何合规性评价。',
     '《建筑法》第7条、《建筑工程施工许可管理办法》第3条、FP-13C', '报告第六节第6项', '在"存在的问题"章节披露此事项，说明是否已受行政处罚及补办手续情况。'),
    ('P0', '合规', '设计合同签订早于可研批复',
     '设计合同签订日期2023年2月25日，可研批复日期2023年3月10日（阿州发改行审〔2023〕79号），合同比批复早13天，程序倒置。',
     '《政府投资条例》第9条', '报告第三节第(二)项第3条', '核实是否有前置依据，如无则应在报告中说明。'),
    ('P0', 'FP-13F 数据', '附件1决算报表数据与审核报告不一致',
     '附件1（单位自编报表）显示建安863,079.42、设备193,590.00；审核报告显示建安863,684.35、设备192,985.07。合计一致但建安/设备互倒604.93元。审核报告未说明调整原因。',
     'FP-13F 金额三方勾稽', '附件1 vs 审核报告第五节', '补充建安与设备投资之间调整604.93元的依据。'),

    # P1
    ('P1', '底稿管理', 'Excel底稿混杂其他项目数据',
     '附件1「审计记录」Sheet包含"九寨沟双河镇松柏村片区产业发展项目"完整台账（773.63万），"审计记录(2)"还有另一个饮水项目数据。单一项目审计产品底稿中混杂其他项目，涉及信息隔离风险。',
     '《中国注册会计师职业道德守则》第3号', '附件1审计记录Sheet', '拆分Excel，每个项目独立建档。'),
    ('P1', 'L-1 时效', '竣工至决算报告间隔近2年',
     '竣工验收2024年4月24日→决算报告2026年4月15日，间隔近2年。法定要求3个月内完成（财建81号令第37条）。报告仅一句话带过，无原因分析、无整改措施。',
     'L-1 根源追问、财建81号令第37条', '报告第八节第（一）项', '追问拖延原因（结算拖延/管理缺位/审批积压），对症提出整改建议。'),
    ('P1', '结构', '审计结论与问题建议混排',
     '"八、审计结论"末尾直接嵌套"（一）加强竣工结算及竣工财务决算工作…"，问题/建议混入肯定性结论。',
     '报告结构规范', '报告第八节末尾', '拆分为独立"存在的问题及建议"章节。'),
    ('P1', '数据', '附件2审计记录合同额填写错误',
     '附件2「审计记录」Sheet施工合同额/送审额填写为1,056,669.42（审定金额），而非实际合同额1,069,200.00，导致核减额显示为0。',
     '审计底稿准确性和完整性', '附件2审计记录Sheet', '更正送审额为实际合同额1,069,200.00。'),

    # P2
    ('P2', 'FP-13F 完整性', '缺少金额三方勾稽汇总表',
     '合同额、审定、支付数据分散在各段落，未汇总为"费用项→合同额→审定→支付→差额"一览表，不便于快速复核。',
     'FP-13F', '报告全文', '补附三方勾稽汇总表（建安/设备/设计/监理逐行）。'),
    ('P2', '格式', '注册会计师签字缺失',
     '报告末尾两处CPA签字处均为空白横线，未签署姓名。',
     '审计报告签署要求', '报告末尾', '正式出具前完成签字。'),
    ('P2', '规范', '施工许可证编号未引用',
     '报告仅提及核发日期，未引证许可证编号513229202403210201。',
     '报告引用规范', '报告第六节第6项', '补充施工许可证编号。'),
    ('P2', '表述', '监理费采购方式标注不一致',
     '审核说明写"询价"，附件2审计记录写"直接委托"，口径不统一。',
     'FP-10A 双轨术语', '审核说明 vs 附件2', '统一监理费采购方式表述。'),
]

for i, (lvl, dim, title, desc, rule, source, fix) in enumerate(findings):
    r = 4 + i
    fill_map = {'P0': p0_fill, 'P1': p1_fill, 'P2': p2_fill}
    font_map = {'P0': red_font, 'P1': yellow_font, 'P2': gray_font}
    ws2.cell(row=r, column=1, value=i+1).font = normal_font
    ws2.cell(row=r, column=2, value=lvl).font = font_map[lvl]
    ws2.cell(row=r, column=3, value=dim).font = normal_font
    ws2.cell(row=r, column=4, value=title).font = bold_font
    ws2.cell(row=r, column=5, value=desc).font = normal_font
    ws2.cell(row=r, column=6, value=rule).font = normal_font
    ws2.cell(row=r, column=7, value=source).font = normal_font
    ws2.cell(row=r, column=8, value=fix).font = normal_font
    style_row(ws2, r, 8, fill=fill_map[lvl])

for i, w in enumerate([6, 10, 14, 18, 50, 30, 30, 40]):
    ws2.column_dimensions[get_column_letter(i+1)].width = w

ws2.row_dimensions[3].height = 25
for r in range(4, 4 + len(findings)):
    ws2.row_dimensions[r].height = 80

# =========================================
# Sheet 3: Data Cross-check
# =========================================
ws3 = wb.create_sheet("数据交叉核对")

ws3.merge_cells('A1:F1')
ws3.cell(row=1, column=1, value='数据一致性交叉核对').font = title_font

headers3 = ['核对项', '来源A（审核报告）', '来源B（附件/附表）', '差异', '结果', '备注']
for c, h in enumerate(headers3, 1):
    ws3.cell(row=3, column=c, value=h)
style_header(ws3, 3, 6)

checks = [
    ('总投资', '112.37万', '112.37万', '0', '✅ 一致', ''),
    ('到位资金', '120.15万', '120.15万', '0', '✅ 一致', ''),
    ('结余资金', '7.78万', '7.78万', '0', '✅ 一致', '已由财政收回'),
    ('待摊投资合计', '6.70万', '6.70万', '0', '✅ 一致', '设计4万+监理2.7万'),
    ('建安+设备合计', '105.67万', '105.67万', '0', '✅ 一致', ''),
    ('建安工程分项', '86.37万', '86.31万（附件1）', '604.93', '❌ 不一致', 'P0-4：需补充调整说明'),
    ('设备投资分项', '19.30万', '19.36万（附件1）', '-604.93', '❌ 不一致', 'P0-4：与建安互倒'),
    ('应付未付', '0.00元', '0.00元', '0', '✅ 一致', ''),
    ('设计费 合同→审定→支付', '4万→4万→4万', '4万→4万→4万', '0', '✅ 一致', '三环完整'),
    ('监理费 合同→审定→支付', '2.7万→2.7万→2.7万', '2.7万→2.7万→2.7万', '0', '✅ 一致', '三环完整'),
    ('待摊分摊 建安', '54,763.44', '54,763.44', '0', '✅ 一致', '精确计算验证通过'),
    ('待摊分摊 设备', '12,236.56', '12,236.56', '0', '✅ 一致', '精确计算验证通过'),
]

for i, (item, src_a, src_b, diff, result, note) in enumerate(checks):
    r = 4 + i
    ws3.cell(row=r, column=1, value=item).font = bold_font
    ws3.cell(row=r, column=2, value=src_a).font = normal_font
    ws3.cell(row=r, column=3, value=src_b).font = normal_font
    ws3.cell(row=r, column=4, value=diff).font = normal_font
    ws3.cell(row=r, column=5, value=result).font = normal_font
    ws3.cell(row=r, column=6, value=note).font = normal_font
    fill = ok_fill if '一致' in result else p0_fill
    style_row(ws3, r, 6, fill=fill)

for i, w in enumerate([28, 26, 26, 14, 14, 42]):
    ws3.column_dimensions[get_column_letter(i+1)].width = w

# =========================================
# Sheet 4: DaTan Calculation
# =========================================
ws4 = wb.create_sheet("待摊投资测算")

ws4.merge_cells('A1:G1')
ws4.cell(row=1, column=1, value='待摊投资逐笔测算').font = title_font

# Section 1
ws4.cell(row=3, column=1, value='一、合同→审定→支付核对').font = sub_font
h41 = ['费用项', '供应商', '合同额', '审定金额', '已付金额', '审减', '未付']
for c, h in enumerate(h41, 1):
    ws4.cell(row=4, column=c, value=h)
style_header(ws4, 4, 7)

daitan_items = [
    ('设计费', '成都壹品为装饰工程有限责任公司', 40000, 40000, 40000, 0, 0),
    ('监理费', '中鸿莎美管理有限公司', 27000, 27000, 27000, 0, 0),
]
for i, (name, unit, c, a, p, d1, d2) in enumerate(daitan_items):
    r = 5 + i
    ws4.cell(row=r, column=1, value=name).font = normal_font
    ws4.cell(row=r, column=2, value=unit).font = normal_font
    ws4.cell(row=r, column=3, value=c).font = normal_font
    ws4.cell(row=r, column=4, value=a).font = normal_font
    ws4.cell(row=r, column=5, value=p).font = normal_font
    ws4.cell(row=r, column=6, value=d1).font = normal_font
    ws4.cell(row=r, column=7, value=d2).font = normal_font
    style_row(ws4, r, 7, fill=ok_fill)

r = 7
ws4.cell(row=r, column=1, value='合计').font = bold_font
ws4.cell(row=r, column=3, value=67000).font = bold_font
ws4.cell(row=r, column=4, value=67000).font = bold_font
ws4.cell(row=r, column=5, value=67000).font = bold_font
ws4.cell(row=r, column=6, value=0).font = bold_font
ws4.cell(row=r, column=7, value=0).font = bold_font
style_row(ws4, r, 7, fill=ok_fill)

# Section 2
r = 9
ws4.cell(row=r, column=1, value='二、待摊投资分摊计算（按建安/设备价值比例）').font = sub_font
r = 10
h42 = ['项目', '待摊前金额', '占总成本比例', '应分摊待摊', '报表分摊值', '差异', '判定']
for c, h in enumerate(h42, 1):
    ws4.cell(row=r, column=c, value=h)
style_header(ws4, r, 7)

alloc = [
    ('建安工程投资', 863684.35, '81.7365%', 54763.44, 54763.44, 0.00),
    ('设备投资', 192985.07, '18.2635%', 12236.56, 12236.56, 0.00),
]
for i, (name, val, ratio, calc, report, diff) in enumerate(alloc):
    r = 11 + i
    ws4.cell(row=r, column=1, value=name).font = normal_font
    ws4.cell(row=r, column=2, value=val).font = normal_font
    ws4.cell(row=r, column=3, value=ratio).font = normal_font
    ws4.cell(row=r, column=4, value=calc).font = normal_font
    ws4.cell(row=r, column=5, value=report).font = normal_font
    ws4.cell(row=r, column=6, value=diff).font = normal_font
    ws4.cell(row=r, column=7, value='✅ 一致').font = normal_font
    style_row(ws4, r, 7, fill=ok_fill)

r = 13
ws4.cell(row=r, column=1, value='合计').font = bold_font
ws4.cell(row=r, column=2, value=1056669.42).font = bold_font
ws4.cell(row=r, column=3, value='100%').font = bold_font
ws4.cell(row=r, column=4, value=67000.00).font = bold_font
ws4.cell(row=r, column=5, value=67000.00).font = bold_font
ws4.cell(row=r, column=6, value=0.00).font = bold_font
ws4.cell(row=r, column=7, value='✅ 一致').font = bold_font
style_row(ws4, r, 7, fill=ok_fill)

# Section 3
r = 15
ws4.cell(row=r, column=1, value='三、未涉及项目（已确认）').font = sub_font
r = 16
h43 = ['费用项', '原因说明', '确认']
for c, h in enumerate(h43, 1):
    ws4.cell(row=r, column=c, value=h)
style_header(ws4, r, 3)

non_items = [
    ('建设单位管理费', '本项目不涉及', '✅ 已确认'),
    ('工程结算审核费', '由财政局另行支付，不计入本项目待摊', '✅ 已确认'),
]
for i, (name, reason, confirm) in enumerate(non_items):
    r = 17 + i
    ws4.cell(row=r, column=1, value=name).font = normal_font
    ws4.cell(row=r, column=2, value=reason).font = normal_font
    ws4.cell(row=r, column=3, value=confirm).font = normal_font
    style_row(ws4, r, 3, fill=ok_fill)

# Section 4
r = 20
ws4.cell(row=r, column=1, value='四、结论').font = sub_font
r = 21
ws4.merge_cells(start_row=r, start_column=1, end_row=r+2, end_column=7)
ws4.cell(row=r, column=1, value=(
    '待摊投资合计67,000.00元，包含设计费40,000.00元、监理费27,000.00元。\n'
    '1. 合同→审定→支付：两笔费用均三环一致，无审减无拖欠。\n'
    '2. 分摊计算：按建安81.7365%、设备18.2635%比例，分摊结果精确验证通过。\n'
    '3. 建设单位管理费不涉及（已确认），结算审核费财政另付（已确认）。\n'
    '4. 复核结论：待摊投资核算准确，无异常。'
)).font = Font(name='微软雅黑', size=10)

for i, w in enumerate([20, 32, 18, 18, 18, 12, 12]):
    ws4.column_dimensions[get_column_letter(i+1)].width = w

# =========================================
# Sheet 5: Checklist (FP rules)
# =========================================
ws5 = wb.create_sheet("FP规则检查")

ws5.merge_cells('A1:F1')
ws5.cell(row=1, column=1, value='工程竣工财务决算审核 FP-10 / FP-13 规则检查清单').font = title_font

h5 = ['规则ID', '检查内容', '阈值', '实际值', '判定', '说明']
for c, h in enumerate(h5, 1):
    ws5.cell(row=3, column=c, value=h)
style_header(ws5, 3, 6)

fp_checks = [
    ('FP-13A', '待摊投资占比', '>30%标黄 >40%标红', '5.96%', '✅ 正常', '未触发阈值'),
    ('FP-13B', '建安费核减率', '>10%标黄 >15%标红', '1.17%', '✅ 正常', '(1,069,200-1,056,669.42)/1,069,200'),
    ('FP-13C', '矛盾短语对检测', '互斥短语同现即标红', '施工许可证滞后8月', '❌ P0', '报告中仅陈述事实未作合规评价'),
    ('FP-13D', '终止项目合同链完整性', '施工/监理单位已确定+终止→必须说明合同处理', 'N/A', '—', '本项目非终止项目'),
    ('FP-13E', '同批次形式一致性', '文号/日期/截止日格式统一', '同报告内两个文号', '❌ P0', '04-号 vs 03-12号'),
    ('FP-13F', '金额三方勾稽完整性', '合同额→审定→支付三环可追溯', '附件1建安/设备互倒604.93', '❌ P0', '合计一致，分项不一致需补充说明'),
    ('FP-10A', '双轨术语允许', '工程/财务术语双轨不标错', '—', '✅ 正常', ''),
    ('FP-10B', '概算vs决算对比', '超概>10%且未批→P0', '节约16.77%', '✅ 正常', '未超概'),
    ('FP-10C', '签证变更合理性', '签证>30%标P1', 'N/A', '—', '维修改造项目，未见签证变更'),
    ('L-1', '根源追问', 'P0问题须有根源分析', '决算拖延无原因分析', '❌ P1', '竣工2年才出报告，原因未追问'),
    ('L-2', '微观→中观跃迁', '≥2同类问题应上升审视', '仅1条问题', '—', ''),
]

for i, (rule, content, threshold, actual, result, note) in enumerate(fp_checks):
    r = 4 + i
    ws5.cell(row=r, column=1, value=rule).font = bold_font
    ws5.cell(row=r, column=2, value=content).font = normal_font
    ws5.cell(row=r, column=3, value=threshold).font = normal_font
    ws5.cell(row=r, column=4, value=actual).font = normal_font
    ws5.cell(row=r, column=5, value=result).font = normal_font
    ws5.cell(row=r, column=6, value=note).font = normal_font
    fill = p0_fill if 'P0' in result else (p1_fill if 'P1' in result else (ok_fill if '正常' in result else None))
    style_row(ws5, r, 6, fill=fill)

for i, w in enumerate([10, 24, 24, 24, 14, 40]):
    ws5.column_dimensions[get_column_letter(i+1)].width = w

# ===== Save =====
output = r'C:\Users\scrccpa\Desktop\阿坝州税务局业务用房维修改造项目_复核意见.xlsx'
wb.save(output)
print(f'Saved to: {output}')
