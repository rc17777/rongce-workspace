import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ===== Styles =====
hf = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
hfill = PatternFill(start_color='0A1F3F', end_color='0A1F3F', fill_type='solid')
tfont = Font(name='微软雅黑', bold=True, size=14, color='0A1F3F')
sf = Font(name='微软雅黑', bold=True, size=12, color='1A5C6E')
s2 = Font(name='微软雅黑', bold=True, size=11, color='1A5C6E')
nf = Font(name='微软雅黑', size=10)
bf = Font(name='微软雅黑', size=10, bold=True)
rf = Font(name='微软雅黑', size=10, bold=True, color='CC0000')
yf = Font(name='微软雅黑', size=10, bold=True, color='C5955C')
gf = Font(name='微软雅黑', size=10, color='888888')
gnf = Font(name='微软雅黑', size=10, color='228B22')

p0f = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
p1f = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
p2f = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
okf = PatternFill(start_color='D9F2D9', end_color='D9F2D9', fill_type='solid')
h2f = PatternFill(start_color='1A5C6E', end_color='1A5C6E', fill_type='solid')

tb = Border(left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))
wr = Alignment(wrap_text=True, vertical='top')
ct = Alignment(horizontal='center', vertical='center', wrap_text=True)

def sh(ws, row, n, fill=hfill):
    for c in range(1, n+1):
        cl = ws.cell(row=row, column=c)
        cl.font = hf; cl.fill = fill; cl.alignment = ct; cl.border = tb

def sr(ws, row, n, font=None, fill=None):
    for c in range(1, n+1):
        cl = ws.cell(row=row, column=c)
        cl.font = font or nf; cl.alignment = wr; cl.border = tb
        if fill: cl.fill = fill

def wc(ws, r, c, v, font=None, fill=None, num=False):
    cl = ws.cell(row=r, column=c)
    cl.value = v; cl.font = font or nf; cl.alignment = wr; cl.border = tb
    if fill: cl.fill = fill
    if num: cl.number_format = '#,##0.00'

PRJ = '马尔康市日部乡干部周转宿舍维修加固项目'
PRJ2 = '马尔康市日部乡业务用房维修加固项目'

# ============================================================
# Sheet 1: Summary
# ============================================================
ws1 = wb.active
ws1.title = "复核总览"
ws1.merge_cells('A1:G1'); wc(ws1, 1, 1, '工程竣工财务决算审核报告 复核意见', tfont)
ws1.merge_cells('A2:G2')
wc(ws1, 2, 1, f'{PRJ} | 川融策专审〔2026〕第03-12号 | 2026年4月15日 | 审计小组：陈越',
   Font(name='微软雅黑', size=10, color='666666'))

r = 4
wc(ws1, r, 1, '项目基本信息', sf); r += 1
info = [
    ('项目名称', PRJ),
    ('被审核单位', '马尔康市日部乡人民政府'),
    ('委托单位', '马尔康市财政局'),
    ('报告文号', '川融策专审〔2026〕第03-12号\n（⚠ 同份报告第4行另出现不完整"第04-号"；且与业务用房项目文号完全相同！）'),
    ('审计小组', '陈越 / 15184481037'),
    ('审计期间', '2026年4月1日 - 4月15日'),
    ('概算总投资', '2,300,000.00元'),
    ('实际完成投资', '1,965,481.06元（节约14.54%）'),
    ('建安工程投资', '1,830,881.06元'),
    ('设备投资', '0.00元（无设备购置）'),
    ('待摊投资', '134,600.00元（设计费80,000 + 监理费54,600）'),
    ('待摊占比', '6.85%（<30%阈值，正常）'),
    ('建安费核减率', '0.45%（<10%阈值，正常）'),
    ('到位资金', '2,085,300.00元'),
    ('结余资金', '119,818.94元（已由财政收回）'),
]
for k, v in info:
    wc(ws1, r, 1, k, bf)
    ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    wc(ws1, r, 2, v); r += 1

r += 1; wc(ws1, r, 1, '复核结果统计', sf); r += 1
for c, h in enumerate(['等级', '数量', '占比', '主要类型', '说明'], 1):
    wc(ws1, r, c, h, hf, hfill)
sh(ws1, r, 5)

st = [
    ('P0 致命', 7, '47%', '文号+粘贴+合规+数据',
     '报告文号与另一项目完全相同、待摊金额粘贴错误、施工单位名称粘贴错误、设计合同先于可研、施工许可证滞后9月、节约率小数异常、报告内待摊数据矛盾'),
    ('P1 重要', 3, '20%', '底稿+时效+结构',
     '底稿混杂、决算拖延2年无原因分析、结论与问题混排'),
    ('P2 建议', 5, '33%', '完整性+格式+规范',
     '缺三方勾稽表、CPA签字缺失、对比第一份报告同质问题、待摊费用例行测算'),
    ('合计', 15, '100%', '', ''),
]
for lvl, cnt, pct, tp, desc in st:
    r += 1
    fl = rf if 'P0' in lvl else (yf if 'P1' in lvl else bf)
    fl2 = p0f if 'P0' in lvl else (p1f if 'P1' in lvl else p2f)
    wc(ws1, r, 1, lvl, fl, fl2); wc(ws1, r, 2, cnt, None, fl2)
    wc(ws1, r, 3, pct, None, fl2); wc(ws1, r, 4, tp, None, fl2)
    wc(ws1, r, 5, desc, None, fl2)

r += 2; wc(ws1, r, 1, '综合结论', sf); r += 1
ws1.merge_cells(start_row=r, start_column=1, end_row=r+4, end_column=7)
wc(ws1, r, 1, (
    '本报告与第一个报告（业务用房维修加固）使用同一模板，存在大量粘贴错误，质量问题比第一份更严重。\n\n'
    '最致命问题：本报告文号"川融策专审〔2026〕第03-12号"与业务用房项目报告完全一致——两份不同项目的正式报告不能共用一个文号。\n\n'
    '第五节(二)待摊投资数据（67,000元）直接复制自业务用房报告（该书应为134,600元），同一份报告第五节开头与第五节(二)之间即存在矛盾。\n\n'
    '第四节第4段施工单位名称"安徽正飞建筑"同样复制自业务用房报告（实际为"安徽烨煌建筑工程"）。\n\n'
    '设计费80,000元、监理费54,600元按国家标准测算在合理范围内，是本报告为数不多的正常项。'
), bf)

for c, w in zip(['A','B','C','D','E','F','G'], [22,16,10,18,18,18,18]):
    ws1.column_dimensions[c].width = w

# ============================================================
# Sheet 2: DaTan Fee Standards
# ============================================================
ws2 = wb.create_sheet("待摊费用标准测算")
ws2.merge_cells('A1:H1'); wc(ws2, 1, 1, '待摊投资费用 收费标准逐笔测算', tfont)
ws2.merge_cells('A2:H2')
wc(ws2, 2, 1, '设计费：计价格[2002]10号 | 监理费：发改价格[2007]670号 | 2015年后市场调节价，标准为参考基准',
   Font(name='微软雅黑', size=9, color='888888'))

# -- Design Fee --
r = 4; wc(ws2, r, 1, '一、设计费 80,000元', sf); r += 1
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
wc(ws2, r, 1, '公式：设计费 = 收费基价 x 专业系数 x 复杂度系数 x 附加系数 x (1+/-浮动)',
   Font(name='微软雅黑', size=10, italic=True, color='666666'))

r += 2; wc(ws2, r, 1, '测算步骤', s2); r += 1
for c, h in enumerate(['步骤', '参数', '取值/说明', '计算过程', '金额(元)', '依据'], 1):
    wc(ws2, r, c, h, hf, hfill)

ds = [
    ('Step1', '计费额', '招标控制价审定1,906,976.48，扣除暂列金160,289.08\n= 1,746,687.40元 = 174.67万',
     '1,906,976.48-160,289.08', 1746687.40, '计价格[2002]10号 第八条'),
    ('Step2', '收费基价', '174.67万在0-200万区间,费率=9.0/200=0.045',
     '174.67x0.045=7.8601万', 78601.50, '基价表 内插法'),
    ('Step3a', '专业系数', '建筑市政工程', '1.0', None, '附表'),
    ('Step3b', '复杂度系数', '含结构加固维修改造', 'II级=1.0(I级=0.85)', None, '附表'),
    ('Step3c', '附加系数', '改扩建项目', '1.1(范围1.1-1.4)', None, '第1.0.12条'),
    ('Step4', '基本设计费(主情景)', 'II级+改扩建1.1', '78,601.50x1.0x1.0x1.1', 86461.65, '第七条'),
    ('Step5', '浮动下限(-20%)', '市场调节', '86,461.65x0.8', 69169.32, '第九条'),
]
for step, p, desc, calc, val, basis in ds:
    r += 1
    wc(ws2, r, 1, step, bf); wc(ws2, r, 2, p); wc(ws2, r, 3, desc)
    wc(ws2, r, 4, calc)
    if val: wc(ws2, r, 5, val, None, None, True)
    else: wc(ws2, r, 5, '—')
    wc(ws2, r, 6, basis)
    if step == 'Step4': sr(ws2, r, 6, font=bf, fill=p1f)

r += 2; wc(ws2, r, 1, '多情景', s2); r += 1
for c, h in enumerate(['情景', '复杂度', '改扩建系数', '测算(元)', '下限(元)', '差异(元)', '差异率', '判定'], 1):
    wc(ws2, r, c, h, hf, h2f)

base_d = 78601.50
for name, cx, ex in [('I级+无改扩建',0.85,1.0),('I级+改扩建1.1',0.85,1.1),
                      ('II级+改扩建1.1(主)',1.0,1.1),('II级+改扩建1.2',1.0,1.2),
                      ('II级+改扩建1.3',1.0,1.3)]:
    r += 1; v = base_d*cx*ex; lo = v*0.8; diff = v-80000
    wc(ws2, r, 1, name); wc(ws2, r, 2, cx); wc(ws2, r, 3, ex)
    wc(ws2, r, 4, v, None, None, True); wc(ws2, r, 5, lo, None, None, True)
    wc(ws2, r, 6, diff, None, None, True); wc(ws2, r, 7, f'{diff/80000*100:+.1f}%')
    j = 'OK' if abs(diff/80000)<0.15 else 'WARN'
    wc(ws2, r, 8, '基本吻合' if j=='OK' else '偏低', gnf if j=='OK' else yf)

r += 2; ws2.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
wc(ws2, r, 1, '>> 设计费80,000元在I级+改扩建(73,492)与II级+改扩建(86,462)之间，基本合理。审核报告应补充标准测算作为审计支撑。', bf, okf)

# -- Supervision Fee --
r += 2; wc(ws2, r, 1, '二、监理费 54,600元', sf); r += 1
ws2.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
wc(ws2, r, 1, '公式：监理费 = 收费基价 x 专业系数 x 复杂度系数 x 高程系数 x (1+/-浮动)',
   Font(name='微软雅黑', size=10, italic=True, color='666666'))

r += 2; wc(ws2, r, 1, '测算步骤', s2); r += 1
for c, h in enumerate(['步骤', '参数', '取值/说明', '计算过程', '金额(元)', '依据'], 1):
    wc(ws2, r, c, h, hf, hfill)

ss = [
    ('Step1', '计费额', '施工合同1,839,200元=183.92万(<500万)', '1,839,200.00', 1839200.00, '发改价格[2007]670号'),
    ('Step2', '收费基价', '低于500万, 费率3.3%', '183.92x3.3%=6.0694万', 60693.60, '基价表 内插'),
    ('Step3a', '专业系数', '建筑市政', '1.0', None, '附表'),
    ('Step3b', '复杂度系数', '含结构加固, II级', '1.0(I级=0.85)', None, '附表'),
    ('Step3c', '高程系数', '马尔康海拔约2600m', '1.1', None, '第1.0.9条'),
    ('Step4', '监理费(主情景)', 'II级+高程1.1', '60,693.60x1.0x1.0x1.1', 66762.96, ''),
    ('Step5', '浮动下限(-20%)', '市场调节', '66,762.96x0.8', 53410.37, ''),
]
for step, p, desc, calc, val, basis in ss:
    r += 1
    wc(ws2, r, 1, step, bf); wc(ws2, r, 2, p); wc(ws2, r, 3, desc)
    wc(ws2, r, 4, calc)
    if val: wc(ws2, r, 5, val, None, None, True)
    else: wc(ws2, r, 5, '—')
    wc(ws2, r, 6, basis)
    if step == 'Step4': sr(ws2, r, 6, font=bf, fill=p1f)

r += 2; wc(ws2, r, 1, '多情景', s2); r += 1
for c, h in enumerate(['情景', '复杂度', '高程系数', '测算(元)', '下限(元)', '差异(元)', '差异率', '判定'], 1):
    wc(ws2, r, c, h, hf, h2f)

base_s = 60693.60
for name, cx, alt in [('I级+高程1.0(最保守)',0.85,1.0),('I级+高程1.1',0.85,1.1),
                       ('II级+高程1.0',1.0,1.0),('II级+高程1.1(主)',1.0,1.1)]:
    r += 1; v = base_s*cx*alt; lo = v*0.8; diff = v-54600
    wc(ws2, r, 1, name); wc(ws2, r, 2, cx); wc(ws2, r, 3, alt)
    wc(ws2, r, 4, v, None, None, True); wc(ws2, r, 5, lo, None, None, True)
    wc(ws2, r, 6, diff, None, None, True); wc(ws2, r, 7, f'{diff/54600*100:+.1f}%')
    j = 'OK' if abs(diff/54600)<0.08 else ('WARN' if abs(diff/54600)<0.15 else 'FAIL')
    fj = gnf if j=='OK' else (yf if j=='WARN' else rf)
    wc(ws2, r, 8, '基本吻合' if j=='OK' else ('偏低' if j=='WARN' else '明显偏低'), fj)

r += 2; ws2.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
wc(ws2, r, 1, '>> 监理费54,600元在I级+高程1.0(51,590)与I级+高程1.1(56,749)之间，基本合理。审核报告应补充标准测算。', bf, okf)

# Summary
r += 2; wc(ws2, r, 1, '三、综合判定', sf); r += 1
for c, h in enumerate(['费用项', '合同价', '测算值(主)', '差异', '差异率', '浮动下限', '区间内?', '判定'], 1):
    wc(ws2, r, c, h, hf, hfill)
for name, act, cal, j in [('设计费',80000.00,86461.65,'基本合理'),('监理费',54600.00,66762.96,'基本合理')]:
    r += 1; diff = act-cal
    wc(ws2, r, 1, name); wc(ws2, r, 2, act, None, None, True)
    wc(ws2, r, 3, cal, None, None, True)
    wc(ws2, r, 4, diff, None, None, True); wc(ws2, r, 5, f'{diff/cal*100:+.1f}%')
    wc(ws2, r, 6, cal*0.8, None, None, True); wc(ws2, r, 7, '是')
    wc(ws2, r, 8, j, gnf)

for i, w in enumerate([14,18,32,32,16,20,14,16]): ws2.column_dimensions[get_column_letter(i+1)].width = w

# ============================================================
# Sheet 3: Detailed Findings
# ============================================================
ws3 = wb.create_sheet("复核明细"); ws3.merge_cells('A1:H1')
wc(ws3, 1, 1, '复核发现明细（共15项）', tfont)
for c, h in enumerate(['序号','等级','维度','发现标题','问题描述','依据','原文位置','建议'], 1):
    wc(ws3, 3, c, h, hf)
sh(ws3, 3, 8)

findings = [
    ('P0','文号','报告文号与另一项目完全相同',
     '本报告"川融策专审〔2026〕第03-12号"与业务用房维修加固项目报告文号完全相同。两份不同项目的正式报告不能共用同一文号。正文第4行另有不完整"第04-号"。两个项目共出现4次文号，无一正确。',
     '报告基本规范\nFP-13E 同批次一致性',
     '正文第4-5行\n对比业务用房报告',
     '为每个项目分配唯一文号。本报告建议川融策专审〔2026〕第03-11号。'),

    ('P0','粘贴错误','第五节(二)待摊投资数据复制错误——67,000应为134,600',
     '第五节开头"待摊投资134,600.00元"（正确），但第五节(二)资金支出表写"待摊投资67,000.00元"（错误）。67,000元系业务用房项目的待摊金额（设计40,000+监理27,000），直接复制粘贴未修改。本项目的正确待摊=设计80,000+监理54,600=134,600。同一报告同一节内数据自相矛盾。',
     '审计底稿准确性\nFP-13F 金额勾稽',
     '第五节(一) vs 第五节(二)\n（前者134,600 vs 后者67,000）',
     '将第五节(二)的67,000更正为134,600，金额表格同步修正。全报告统一检查粘贴错误。'),

    ('P0','粘贴错误','施工单位名称复制错误——"安徽正飞建筑"应为"安徽烨煌建筑工程"',
     '第四节第4段"根据日部乡人民政府与安徽正飞建筑科技有限公司签订的施工合同"。安徽正飞是业务用房项目的施工单位，本项目施工单位为安徽烨煌建筑工程有限公司。编制说明中同样存在此错误。',
     '审计底稿准确性',
     '审核报告第四节第4段\n编制说明第(三)节',
     '全报告搜索"安徽正飞"替换为"安徽烨煌建筑工程有限公司"。'),

    ('P0','合规','设计合同签订早于可研批复',
     '设计合同2023年2月25日，可研批复2023年3月10日（马尔发改行审[2023]78号），合同比批复早13天，程序倒置。与业务用房项目完全相同的问题。',
     '《政府投资条例》第9条',
     '报告第三节(二)第3条',
     '核实是否有前置依据，无则在报告中说明。'),

    ('P0','FP-13C','施工许可证滞后约9个月——未作合规评价',
     '开工2023年7月1日，施工许可证2024年4月1日（513229202403210101），无证施工约9个月。报告第六节仅平铺事实。',
     '《建筑法》第7条\nFP-13C',
     '报告第六节第6项',
     '在问题章节披露，说明处罚/补办情况。'),

    ('P0','数据','节约金额小数异常——"334,518.948元"',
     '第七节第1项"节约334,518.948元"，元为单位出现3位小数（即0.948分钱），明显是浮点计算未取整的结果。',
     '数据精度规范',
     '报告第七节第1项',
     '修正为334,518.95元（四舍五入到分）。'),

    ('P0','数据矛盾','报告内待摊投资金额矛盾（134,600 vs 67,000）',
     '第五节开头和第六节写待摊134,600元，第五节(二)资金支出表写67,000元。同一份报告内同一数据出现两种金额，读者无法判断哪个正确。编制说明证实正确金额为134,600。',
     'FP-13F 金额三方勾稽\n审计报告准确性',
     '第五节(一) vs 第五节(二)\n编制说明第(六)条',
     '统一修正为134,600，通篇核查。'),

    # P1
    ('P1','底稿','Excel底稿可能混杂其他项目数据',
     '与业务用房项目同批次审计，需检查附件底稿是否同样存在多项目数据混杂问题。',
     '《职业道德守则》第3号',
     '附件1/2',
     '逐Sheet检查并拆分。'),

    ('P1','L-1','竣工至决算报告间隔近2年',
     '竣工验收2024年4月24日->2026年4月15日，近2年。法定3个月内。报告仅一句话。',
     'L-1 根源追问\n财建81号令 第37条',
     '第八节末尾',
     '追问原因，对症整改。'),

    ('P1','结构','审计结论与问题建议混排',
     '第八节末尾嵌套"进一步加强…"问题建议，缺少独立章节。',
     '报告结构规范',
     '第八节末尾',
     '拆分为独立的问题及建议章节。'),

    # P2
    ('P2','FP-13F','缺少金额三方勾稽汇总表',
     '合同额、审定、支付数据分散。',
     'FP-13F', '报告全文', '补附勾稽汇总表。'),
    ('P2','格式','注册会计师签字缺失', '两处CPA签字空白。', '签署要求', '报告末尾', '正式出具前签字。'),
    ('P2','FP-13E','与业务用房报告同批次横向比对缺失',
     '两份报告同一审计小组、同一出具日期、同一委托单位。文号相同、部分段落完全相同（粘贴模板）、同类问题反复出现，审计小组未做横向交叉比对。',
     'FP-13E 同批次一致性', '两份报告对比', '对同批次报告增加横向比对程序。'),
    ('P2','规范','施工许可证编号未在正文引用', '仅提日期未引编号。', '引用规范', '第六节', '补充编号513229202403210101。'),
    ('P2','审计程序','待摊费用例行测算（建议补充）',
     '设计费80,000元和监理费54,600元按标准测算在合理范围内（P2级建议，非P0）。但审核报告同样未做标准测算审计程序。',
     '《审计准则1301号》', '第五节/第三节', '补充标准测算作为审计支撑，结论为"基本合理"即可。'),
]

for i, (lvl, dim, ttl, desc, basis, loc, fix) in enumerate(findings):
    rr = 4+i
    fl = {'P0':rf,'P1':yf,'P2':gf}[lvl]; fl2 = {'P0':p0f,'P1':p1f,'P2':p2f}[lvl]
    wc(ws3, rr, 1, i+1); wc(ws3, rr, 2, lvl, fl, fl2)
    wc(ws3, rr, 3, dim); wc(ws3, rr, 4, ttl, bf)
    wc(ws3, rr, 5, desc); wc(ws3, rr, 6, basis)
    wc(ws3, rr, 7, loc); wc(ws3, rr, 8, fix)
    sr(ws3, rr, 8, fill=fl2)

for ci, w in enumerate([6,10,16,22,56,34,30,42], 1):
    ws3.column_dimensions[get_column_letter(ci)].width = w
ws3.row_dimensions[3].height = 25
for rr in range(4, 19): ws3.row_dimensions[rr].height = 95

# ============================================================
# Sheet 4: Cross-Check + Cross-Report Comparison
# ============================================================
ws4 = wb.create_sheet("数据核对与跨报告对比")
ws4.merge_cells('A1:G1'); wc(ws4, 1, 1, '数据一致性交叉核对 & 同批次报告横向对比', tfont)

# Section A
r = 3; wc(ws4, r, 1, 'A. 本报告内部数据核对', s2); r += 1
for c, h in enumerate(['核对项','来源A','来源B','差异','结果','备注'], 1):
    wc(ws4, r, c, h, hf)
sh(ws4, r, 6)

checks = [
    ('总投资','1,965,481.06','1,965,481.06','0','一致',''),
    ('建安投资','1,830,881.06','1,830,881.06','0','一致',''),
    ('待摊投资(第五节开头)','134,600.00','134,600.00(编制说明)','0','一致',''),
    ('待摊投资(第五节二)',"67,000.00(错误!)",'134,600.00(正确)','-67,600.00','矛盾! P0','复制自业务用房报告'),
    ('到位资金','2,085,300.00','2,085,300.00','0','一致',''),
    ('结余资金','119,818.94','119,818.94','0','一致','已由财政收回'),
    ('设计合同->审定->支付','80,000->80,000->80,000','80,000->80,000->80,000','0','一致','三环完整'),
    ('监理合同->审定->支付','54,600->54,600->54,600','54,600->54,600->54,600','0','一致','三环完整'),
    ('建安+待摊合计','1,965,481.06','1,965,481.06(1,830,881.06+134,600)','0','一致',''),
]
for item, a, b, diff, res, note in checks:
    r += 1
    wc(ws4, r, 1, item, bf); wc(ws4, r, 2, a); wc(ws4, r, 3, b)
    wc(ws4, r, 4, diff)
    fl = rf if '矛盾' in res else gnf
    wc(ws4, r, 5, res, fl); wc(ws4, r, 6, note)
    fl2 = p0f if '矛盾' in res else okf
    sr(ws4, r, 6, fill=fl2)

# Section B: Cross-report comparison
r += 2; wc(ws4, r, 1, 'B. 同批次两报告横向比对', s2); r += 1
for c, h in enumerate(['比对项', f'报告1: {PRJ2}', f'报告2: {PRJ}', '问题', '判定'], 1):
    wc(ws4, r, c, h, hf)
sh(ws4, r, 5)

cross = [
    ('报告文号','川融策专审〔2026〕第03-12号','川融策专审〔2026〕第03-12号','完全相同!','P0'),
    ('不完整文号','"第04-号"','"第04-号"','完全相同!','P0'),
    ('审计小组','陈越','陈越','同一人','—'),
    ('出具日期','2026年4月15日','2026年4月15日','同一天','—'),
    ('委托单位','马尔康市财政局','马尔康市财政局','同一委托方','—'),
    ('设计合同日 vs 可研批复','2023.2.25 vs 2023.3.10','2023.2.25 vs 2023.3.10','完全相同的问题','P0(两报告均有)'),
    ('施工许可证滞后','8个月','9个月','同类问题','P0(两报告均有)'),
    ('审计结论->问题混排','相同结构','相同结构','模板问题','P1(两报告均有)'),
    ('CPA签字','空白','空白','模板问题','P2(两报告均有)'),
    ('粘贴错误','—','"安徽正飞"/67,000','仅报告2存在','P0(报告2独有)'),
    ('施工单位','安徽正飞建筑','安徽烨煌建筑','不同','—'),
    ('可研批复文号','马尔发改行审[2023]79号','马尔发改行审[2023]78号','连续编号','—'),
    ('施工许可证号','513229202403210201','513229202403210101','仅末尾不同','—'),
]
for item, r1, r2, prob, judge in cross:
    r += 1
    wc(ws4, r, 1, item, bf); wc(ws4, r, 2, r1); wc(ws4, r, 3, r2)
    wc(ws4, r, 4, prob)
    fl = rf if 'P0' in judge else (yf if 'P1' in judge else (gf if 'P2' in judge else gnf))
    wc(ws4, r, 5, judge, fl)
    fl2 = p0f if 'P0' in judge else (p1f if 'P1' in judge else (p2f if 'P2' in judge else okf))
    sr(ws4, r, 5, fill=fl2)

for ci, w in enumerate([28,36,36,30,14,22], 1): ws4.column_dimensions[get_column_letter(ci)].width = w

# ============================================================
# Sheet 5: FP Rules
# ============================================================
ws5 = wb.create_sheet("FP规则检查")
ws5.merge_cells('A1:F1'); wc(ws5, 1, 1, 'FP-10/FP-13/L规则 & 跨报告检查清单', tfont)
for c, h in enumerate(['规则ID','检查内容','阈值/标准','实际值','判定','说明'], 1):
    wc(ws5, 3, c, h, hf)
sh(ws5, 3, 6)

fp = [
    ('FP-13A','待摊占比','>30%黄>40%红','6.85%','正常','134,600/1,965,481.06'),
    ('FP-13B','建安核减率','>10%黄>15%红','0.45%','正常','(1,839,200-1,830,881)/1,839,200'),
    ('FP-13C','矛盾短语对','互斥标红','施工许可证滞后9月','P0','未作合规评价'),
    ('FP-13E','同批次一致性','文号/日期统一','两报告同文号!','P0','"03-12号"重复使用'),
    ('FP-13F','金额勾稽','三环可追溯','待摊67,000 vs 134,600矛盾','P0','第五节内自相矛盾'),
    ('FP-13F-2','施工单位名称','复制粘贴一致性','"安徽正飞"->"安徽烨煌"','P0','第四节第4段'),
    ('FP-10B','概算vs决算','超概>10%且未批','节约14.54%','正常','未超概'),
    ('L-1','根源追问','P0须有分析','决算拖延无分析','P1','同业务用房报告'),
    ('L-2','同类问题上升','>=2同类上升','两报告共享6项同类问题','P0','同批次系统性模板缺陷'),
    ('新增','跨报告文号唯一性','每份报告独立文号','与业务用房共用文号','P0','两报告不可共用文号'),
]
for i, (rule, content, th, actual, res, note) in enumerate(fp):
    rr = 4+i
    wc(ws5, rr, 1, rule, bf); wc(ws5, rr, 2, content); wc(ws5, rr, 3, th)
    wc(ws5, rr, 4, actual)
    if 'P0' in res: fl, fl2 = rf, p0f
    elif 'P1' in res: fl, fl2 = yf, p1f
    else: fl, fl2 = gnf, okf
    wc(ws5, rr, 5, res, fl); wc(ws5, rr, 6, note)
    sr(ws5, rr, 6, fill=fl2)

for ci, w in enumerate([12,26,28,28,16,44], 1): ws5.column_dimensions[get_column_letter(ci)].width = w

# Save
out = r'C:\Users\scrccpa\Desktop\马尔康市日部乡干部周转宿舍维修加固项目_复核意见.xlsx'
wb.save(out)
print(f'OK: {out}')
