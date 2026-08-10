# -*- coding: utf-8 -*-
"""给已有Excel追加"逻辑矛盾分析"Sheet"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

title_font = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
title_fill = PatternFill(start_color='0A1F3F', end_color='0A1F3F', fill_type='solid')
header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='C9252E', end_color='C9252E', fill_type='solid')
sub_fill = PatternFill(start_color='1A5C6E', end_color='1A5C6E', fill_type='solid')
data_font = Font(name='微软雅黑', size=10)
data_bold = Font(name='微软雅黑', size=10, bold=True)
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
warn_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
border = Border(left=Side('thin','CCCCCCCC'), right=Side('thin','CCCCCCCC'), top=Side('thin','CCCCCCCC'), bottom=Side('thin','CCCCCCCC'))
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_w = Alignment(horizontal='left', vertical='center', wrap_text=True)

wb = openpyxl.load_workbook(r'C:\Users\scrccpa\Desktop\郫都民政局2025绩效自评复核_复核结果.xlsx')

# 新增Sheet
ws = wb.create_sheet('逻辑矛盾分析')
ws.sheet_properties.tabColor = 'C9252E'

# 标题
ws.merge_cells('A1:H1')
c = ws.cell(row=1, column=1, value='深度逻辑分析：报告数据间的8项核心矛盾与6项文本缺陷')
c.font = title_font; c.fill = title_fill; c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 36

# Sheet 8-1: 部门级勾稽
r = 3
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
c = ws.cell(row=r, column=1, value='一、部门级数据勾稽')
c.font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
c.fill = sub_fill; c.alignment = left_w
r += 1
headers = ['核查项', '公式/来源', '计算值', '报告值', '差额', '判断']
for i, h in enumerate(headers):
    c = ws.cell(row=r, column=i+1, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = border

items = [
    ['基本支出+项目支出', '1432.49 + 10588.12', '12,020.61万', '12,023.85万', '3.24万', '⚠ 微小差异，可能是其他零星支出未列出'],
    ['项目支出执行率', '10588.12 / (14242.51-1432.49)', '82.65%', '—（未单独列出）', '—', '低于部门整体84.42%'],
    ['年初→调整预算变动', '14242.51 - 9274.21', '+4,968.30万', '+4,968.30万', '0', '✅ 增长53.6%，年中追加近5千万'],
    ['项目数一致性', '55+8+70=133 vs 报告写35', '133个', '35个（部门预算项目）', '—', '⚠ 35可能是一级项目，133是拆分后的子项目'],
    ['指标总数', '逐一统计10个项目所有指标行', '78条', '78条（报告原文）', '0', '✅'],
    ['10项目合计决算数', '逐项累加', '3,584.27万', '3,584.28万', '0.01万', '✅ 尾差可接受'],
]
for row_data in items:
    r += 1
    for i, v in enumerate(row_data):
        c = ws.cell(row=r, column=i+1, value=v)
        c.font = data_font; c.border = border
        c.alignment = center if i < 3 else left_w
    ws.row_dimensions[r].height = 22

# Sheet 8-2: 8项核心矛盾
r += 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
c = ws.cell(row=r, column=1, value='二、8项核心逻辑矛盾')
c.font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
c.fill = PatternFill(start_color='C9252E', end_color='C9252E', fill_type='solid'); c.alignment = left_w
r += 1
h2 = ['#', '矛盾名称', '涉及项目', '数据表现', '自评/复核判断差异', '严重程度', '可能的解释', '建议追问方向']
for i, h in enumerate(h2):
    c = ws.cell(row=r, column=i+1, value=h)
    c.font = header_font; c.fill = PatternFill(start_color='C9252E', end_color='C9252E', fill_type='solid'); c.alignment = center; c.border = border

contradictions = [
    [1, '未验收但数量指标满分',
     '老年认知障碍友好社区',
     '数量指标"建设3个社区"满分15分，但备注写明项目尚在实施中待验收',
     '自评: 15分(已完成3个)\n复核: 15分(维持满分)',
     '🔴 逻辑矛盾',
     '可能以"合同已签订"等同于"已完成"',
     '① 要求提供3个社区建设的佐证材料\n② 明确"完成"的定义是"签合同"还是"通过验收"'],
    [2, '质量全扣 vs 数量满分——同一项目双重标准',
     '老年认知障碍友好社区',
     '质量指标"验收合格率"→复核0分(没验收)；数量指标"建设社区数"→复核15分(也没验收)',
     '质量: 自评15→复核0(全扣) ⚔ 数量: 自评15→复核15(满分)',
     '🔴 复核标准自相矛盾',
     '复核人未统一"未验收项目如何评分"的标准',
     '① 统一标准：未验收=全部指标暂不计分 或 =产出指标按合同确认\n② 复核报告应说明统一评分规则'],
    [3, '钱没花但服务了748人',
     '政府购买居家养老服务',
     '预算132.81万→实际支出1000元(0.08%)；数量指标完成748人(目标100人)满分',
     '预算执行: 自评0→复核0.01(诚实)\n数量: 自评15→复核15(满分)\n→ 1000元 ÷ 748人 = 约1.34元/人',
     '🔴 数据异常',
     '① 可能有上年结转资金支付(报表未反映)\n② 服务商垫资未结算(产生应付账款)\n③ 服务完成数据不实',
     '① 追查748人服务费的实际资金来源\n② 核对服务合同和付款凭证\n③ 比对上年度该项目的支出数据'],
    [4, '老年大学时效数据疑似造假',
     '郫都区老年大学能力提升',
     '自评填"2025年8月7日完成"，实际竣工验收2026年1月16日，差5个月',
     '自评: 15分(按时+提前)\n复核: 11.19分(延期扣25.4%)',
     '🔴 数据不实',
     '25.4%的扣分暗示复核人判断为"延期"而非"造假"',
     '① 确认2025年8月7日这个日期从何而来\n② 如确属虚报，应定性为"自评数据不实"加处罚分'],
    [5, '超预算2.4倍，复核比自评还高分',
     '公办养老机构护理能力提升',
     '成本指标≤160万→实际386万(超2.4倍)；自评10分/15，复核10.5分/15',
     '自评: 66.7%分\n复核: 70%分 ← 反加了0.5分！',
     '🔴 评分逻辑异常',
     '复核人以"未及时更新指标值"为由豁免部分扣分',
     '① 要么承认不达标→按比例扣分\n② 要么认定指标本身不合适→该指标从总分中剔除\n③ 加分是绝对不成立的'],
    [6, '定性指标扣分标准"看心情"',
     '多个项目(见明细)',
     '11条不可考核定性指标，扣分范围0%-100%，标准不一',
     '全扣(100%): 3条(养老消费/老年大学/高龄补贴的可持续影响)\n扣30%: 5条\n不扣(0%): 2条(公办养老护理/认知障碍氛围)',
     '🔴 评分标准不统一',
     '无统一的"定性指标扣分指南"',
     '① 建议制定复核扣分标准表\n② 对同类"无量化佐证"指标统一扣分比例\n③ 或全部记为"无法评价/不适用"，不计入总分'],
    [7, '"不投诉=满意"的偷懒评价',
     '殡葬惠民/高龄补贴/认知障碍',
     '殡葬: "未收到12315投诉"=100%满意\n高龄: "未收到投诉"=100%满意\n认知障碍: 90%满意度(来源不明)',
     '自评满分 → 复核满分（未提出异议）',
     '🔶 评判标准薄弱',
     '复核人完全接受了"不投诉=满意"的逻辑',
     '① 要求补充正式的满意度调查问卷\n② 对"以不投诉代调查"的项目满意度指标不予计分'],
]

for row_data in contradictions:
    r += 1
    for i, v in enumerate(row_data):
        c = ws.cell(row=r, column=i+1, value=v)
        c.font = data_font; c.border = border
        c.alignment = left_w if i > 0 else center
    ws.row_dimensions[r].height = 80
    if '🔴' in str(row_data[5]):
        for i in range(1,9):
            ws.cell(row=r, column=i).fill = red_fill

# 第8项单独写
r += 1
contradiction_8 = [
    [8, '抽样代表性和统计解释缺失',
     '全部10个项目',
     '10项目执行率76.81% vs 部门整体84.42%；5个执行率<50%',
     '报告未讨论抽样偏差和结论外推的局限性',
     '🔶 方法论缺陷',
     '要么民政局整体问题多(抽样有代表性)，要么选择性抽查了问题多的项目(偏差)',
     '复核报告应在"局限性说明"中讨论抽样方法对结论外推性的影响'],
]
for row_data in contradiction_8:
    for i, v in enumerate(row_data):
        c = ws.cell(row=r, column=i+1, value=v)
        c.font = data_font; c.border = border
        c.alignment = left_w if i > 0 else center
    ws.row_dimensions[r].height = 80

# Sheet 8-3: 文本缺陷
r += 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
c = ws.cell(row=r, column=1, value='三、报告本身的文本逻辑缺陷')
c.font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
c.fill = sub_fill; c.alignment = left_w
r += 1
h3 = ['#', '缺陷类型', '问题描述', '影响', '修改建议', '', '', '']
for i, h in enumerate(h3):
    c = ws.cell(row=r, column=i+1, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = border

text_issues = [
    [1, '预算执行得分公式未说明',
     '复核人按"得分=权重×执行率"给预算执行打分（如执行率95.88%→得分9.59/10），但报告中未说明此公式',
     '读者看到"预算执行得分1.10"时不知道这是百分比折算，可读性差'],
    [2, '偏离度定义不清晰',
     '项目级偏离度=(自评-复核)/自评；指标级偏离度含义不完全一致（偏度率vs偏离值）',
     '两种偏离度混用容易造成误解'],
    [3, '备注栏信息密度不一致',
     '有的备注详细（如认知障碍写了一段说明），有的备注极简（如高龄补贴可持续影响全扣却只有一句话）',
     '问题越严重的指标，信息量反而越少，不符合"问题严重→说明越多"的一般规律'],
    [4, '责任归因单向化',
     '报告将问题全部归因于"绩效指标编制规范执行不到位"，但跨年项目目标设置不当、资金追加不更新等，涉及财政-部门的协同问题',
     '单方面归因可能引发被评价单位的抵触'],
    [5, '复核扣分标准未公开',
     '复核报告未附"扣分标准对照表"，如：什么情况下扣30% vs 50% vs 100%？',
     '被评价单位无法核实扣分是否公平，影响复核报告的公信力'],
    [6, '未讨论抽样方法的局限性',
     '报告说按"覆盖面广、社会关注度高、资金规模大"原则选10个项目，但未说明：\n① 这些项目是否代表了70个项目的整体状况\n② 10个项目的异常率是否适用于其他60个项目',
     '结论外推的有效性存疑'],
]
for row_data in text_issues:
    r += 1
    for i, v in enumerate(row_data):
        c = ws.cell(row=r, column=i+1, value=v)
        c.font = data_font; c.border = border
        c.alignment = left_w if i > 0 else center
    ws.row_dimensions[r].height = 60

# 列宽
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 26
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 32
ws.column_dimensions['E'].width = 28
ws.column_dimensions['F'].width = 18
ws.column_dimensions['G'].width = 28
ws.column_dimensions['H'].width = 36

ws.freeze_panes = 'A3'
wb.save(r'C:\Users\scrccpa\Desktop\郫都民政局2025绩效自评复核_复核结果.xlsx')
print('✅ 已追加"逻辑矛盾分析"Sheet到Excel')
