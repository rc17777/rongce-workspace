# -*- coding: utf-8 -*-
"""生成双清制度3个附件模板（Excel格式）"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, numbers
from openpyxl.utils import get_column_letter
import os

outdir = r'C:\Users\scrccpa\.openclaw\workspace\output'
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_font = Font(name='微软雅黑', size=11, bold=True)
body_font = Font(name='宋体', size=10)
title_font = Font(name='黑体', size=14, bold=True)
header_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_body(ws, row, cols, align='center'):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = body_font
        cell.alignment = center_align if align == 'center' else left_align
        cell.border = thin_border

# ════════════════════════════════════════
# 附件1：年度双清目标清单
# ════════════════════════════════════════
wb1 = Workbook()
ws1 = wb1.active
ws1.title = '双清目标清单'

ws1.merge_cells('A1:L1')
ws1['A1'] = '四川融策 年度双清目标清单'
ws1['A1'].font = title_font
ws1['A1'].alignment = center_align

ws1.merge_cells('A2:L2')
ws1['A2'] = '编制部门：财务部    年度：2026    编制日期：    审批人：'
ws1['A2'].font = Font(name='宋体', size=10)

# 表头
headers1 = ['序号', '项目编号', '项目名称', '客户名称', '合同金额\n（元）', '已开票金额\n（元）',
            '已回款金额\n（元）', '应收余额\n（元）', '账龄\n（天）', '责任人\n（项目经理）',
            '年度应收目标\n（元）', '备注/特殊情形']
for c, h in enumerate(headers1, 1):
    ws1.cell(row=4, column=c, value=h)
style_header(ws1, 4, len(headers1))

# 示例数据
sample1 = [
    [1, 'PS-2025-001', 'XX县绩效评价项目', 'XX县财政局', 150000, 150000, 100000, 50000, 120,
     '张三', 50000, '尾款待验收后支付'],
    [2, 'GC-2025-015', 'XX工程结算审计', 'XX住建局', 200000, 200000, 0, 200000, 90,
     '李四', 200000, '报告已提交，等待审批'],
    [3, 'PS-2026-003', 'XX街道绩效自评复核', 'XX街道办', 80000, 80000, 60000, 20000, 60,
     '王五', 20000, ''],
    [4, 'ZL-2025-008', 'XX专项债申报咨询', 'XX发改局', 120000, 120000, 120000, 0, 0,
     '张三', 0, '已全部回款'],
]
for r, row in enumerate(sample1, 5):
    for c, val in enumerate(row, 1):
        ws1.cell(row=r, column=c, value=val)
    style_body(ws1, r, len(headers1), 'left' if r > 5 else 'center')

# 汇总行
total_row = 5 + len(sample1)
ws1.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
ws1.cell(row=total_row, column=1, value='合计')
ws1.cell(row=total_row, column=5, value=550000)
ws1.cell(row=total_row, column=6, value=550000)
ws1.cell(row=total_row, column=7, value=280000)
ws1.cell(row=total_row, column=8, value=270000)
ws1.cell(row=total_row, column=11, value=270000)
for c in range(1, len(headers1)+1):
    ws1.cell(row=total_row, column=c).font = Font(name='微软雅黑', size=10, bold=True)
    ws1.cell(row=total_row, column=c).alignment = center_align
    ws1.cell(row=total_row, column=c).border = thin_border

# 列宽
widths1 = [5, 14, 22, 14, 12, 12, 12, 12, 8, 10, 14, 20]
for c, w in enumerate(widths1, 1):
    ws1.column_dimensions[get_column_letter(c)].width = w

wb1.save(os.path.join(outdir, '附件1-年度双清目标清单.xlsx'))

# ════════════════════════════════════════
# 附件2：季度双清考核结果表
# ════════════════════════════════════════
wb2 = Workbook()
ws2 = wb2.active
ws2.title = '季度考核结果'

ws2.merge_cells('A1:L1')
ws2['A1'] = '四川融策 季度双清考核结果表'
ws2['A1'].font = title_font
ws2['A1'].alignment = center_align

ws2.merge_cells('A2:L2')
ws2['A2'] = '考核期间：2026年Q__    编制日期：    编制人：    审批人：'
ws2['A2'].font = Font(name='宋体', size=10)

# Part A: 项目经理考核
ws2.merge_cells('A4:L4')
ws2['A4'] = '一、项目经理回款考核'
ws2['A4'].font = Font(name='黑体', size=12, bold=True)

headers2a = ['序号', '项目经理', '项目名称', '应收目标\n（元）', '实际回款\n（元）',
             '回款完成率', '完成率区间', '奖惩类型', '计算过程', '奖惩金额\n（元）',
             '长期挂账\n消减奖励', '合计奖惩\n（元）']
for c, h in enumerate(headers2a, 1):
    ws2.cell(row=6, column=c, value=h)
style_header(ws2, 6, len(headers2a))

sample2a = [
    [1, '张三', 'XX县绩效评价', 50000, 55000, '110%', '≥100%', '奖励',
     '(110-100)×2%×8000=1600', 1600, 1000, 2600],
    [2, '李四', 'XX工程结算审计', 200000, 170000, '85%', '75%～90%', '处罚',
     '(90-85)×1%×12000=-600', -600, 0, -600],
    [3, '王五', 'XX绩效自评复核', 20000, 19000, '95%', '90%～100%', '不奖不罚', '——', 0, 0, 0],
]
for r, row in enumerate(sample2a, 7):
    for c, val in enumerate(row, 1):
        ws2.cell(row=r, column=c, value=val)
    style_body(ws2, r, len(headers2a))

# Part B: 部门负责人考核
start_b = 7 + len(sample2a) + 2
ws2.merge_cells(start_row=start_b, start_column=1, end_row=start_b, end_column=12)
ws2.cell(row=start_b, column=1, value='二、部门负责人回款考核')
ws2.cell(row=start_b, column=1).font = Font(name='黑体', size=12, bold=True)

headers2b = ['序号', '部门', '部门负责人', '部门应收目标\n（元）', '部门实际回款\n（元）',
             '综合回款率', '完成率区间', '奖惩类型', '计算过程', '奖惩金额\n（元）',
             '超9月消减率', '合计奖惩\n（元）']
for c, h in enumerate(headers2b, 1):
    ws2.cell(row=start_b+2, column=c, value=h)
style_header(ws2, start_b+2, len(headers2b))

sample2b = [
    [1, '审计业务部', '赵六', 350000, 320000, '91.4%', '85%～95%', '不奖不罚', '——', 0, '60%', 0],
    [2, '工程咨询部', '钱七', 500000, 450000, '90%', '85%～95%', '不奖不罚', '——', 0, '75%', 0],
]
for r, row in enumerate(sample2b, start_b+3):
    for c, val in enumerate(row, 1):
        ws2.cell(row=r, column=c, value=val)
    style_body(ws2, r, len(headers2b))

# Part C: 财务部考核
start_c = start_b + 3 + len(sample2b) + 2
ws2.merge_cells(start_row=start_c, start_column=1, end_row=start_c, end_column=12)
ws2.cell(row=start_c, column=1, value='三、财务部清欠考核')
ws2.cell(row=start_c, column=1).font = Font(name='黑体', size=12, bold=True)

headers2c = ['序号', '岗位', '责任人', '清欠目标\n（元）', '实际清欠\n（元）',
             '清欠完成率', '完成率区间', '奖惩类型', '计算过程', '奖惩金额\n（元）',
             '超1年销号率', '合计奖惩\n（元）']
for c, h in enumerate(headers2c, 1):
    ws2.cell(row=start_c+2, column=c, value=h)
style_header(ws2, start_c+2, len(headers2c))

sample2c = [
    [1, '清欠岗', '孙八', 80000, 72000, '90%', '80%～100%', '不奖不罚', '——', 0, '50%', 0],
]
for r, row in enumerate(sample2c, start_c+3):
    for c, val in enumerate(row, 1):
        ws2.cell(row=r, column=c, value=val)
    style_body(ws2, r, len(headers2c))

# 汇总
sum_row = start_c + 3 + len(sample2c) + 2
ws2.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=12)
ws2.cell(row=sum_row, column=1, value='四、本季度双清考核汇总')
ws2.cell(row=sum_row, column=1).font = Font(name='黑体', size=12, bold=True)

sum_headers = ['类别', '考核人数', '奖励人数', '奖励总金额', '处罚人数', '处罚总金额', '净奖罚']
for c, h in enumerate(sum_headers, 1):
    ws2.cell(row=sum_row+2, column=c, value=h)
style_header(ws2, sum_row+2, len(sum_headers))

sum_data = [['项目经理', 3, 1, 2600, 1, 600, 2000],
            ['部门负责人', 2, 0, 0, 0, 0, 0],
            ['财务部', 1, 0, 0, 0, 0, 0],
            ['合计', 6, 1, 2600, 1, 600, 2000]]
for r, row in enumerate(sum_data, sum_row+3):
    for c, val in enumerate(row, 1):
        ws2.cell(row=r, column=c, value=val)
    style_body(ws2, r, len(sum_headers))
    if row[0] == '合计':
        for c in range(1, len(sum_headers)+1):
            ws2.cell(row=r, column=c).font = Font(name='微软雅黑', size=10, bold=True)

for c, w in enumerate([5, 10, 16, 12, 12, 10, 12, 12, 14, 12, 12, 12], 1):
    ws2.column_dimensions[get_column_letter(c)].width = w

wb2.save(os.path.join(outdir, '附件2-季度双清考核结果表.xlsx'))

# ════════════════════════════════════════
# 附件3：项目回款月报
# ════════════════════════════════════════
wb3 = Workbook()
ws3 = wb3.active
ws3.title = '回款月报'

ws3.merge_cells('A1:N1')
ws3['A1'] = '四川融策 项目回款月报'
ws3['A1'].font = title_font
ws3['A1'].alignment = center_align

ws3.merge_cells('A2:N2')
ws3['A2'] = '报告期间：2026年__月    编制部门：财务部    编制日期：    审批人：'
ws3['A2'].font = Font(name='宋体', size=10)

# Section 1: 本月回款明细
ws3.merge_cells('A4:N4')
ws3['A4'] = '一、本月回款明细'
ws3['A4'].font = Font(name='黑体', size=12, bold=True)

headers3a = ['序号', '项目编号', '项目名称', '客户名称', '合同金额\n（元）', '本月回款\n（元）',
             '累计回款\n（元）', '应收余额\n（元）', '回款完成率', '账龄\n（天）',
             '项目经理', '预警状态', '催收次数', '备注']
for c, h in enumerate(headers3a, 1):
    ws3.cell(row=6, column=c, value=h)
style_header(ws3, 6, len(headers3a))

# Color fills for warning
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

sample3a = [
    [1, 'PS-2025-001', 'XX县绩效评价', 'XX县财政局', 150000, 50000, 150000, 0, '100%', 150,
     '张三', '🟢 绿灯', 2, '本月付清尾款'],
    [2, 'GC-2025-015', 'XX工程结算审计', 'XX住建局', 200000, 0, 0, 200000, '0%', 90,
     '李四', '🟡 黄灯', 1, '等待审批'],
    [3, 'PS-2026-003', 'XX绩效自评复核', 'XX街道办', 80000, 20000, 80000, 0, '100%', 60,
     '王五', '🟢 绿灯', 1, ''],
    [4, 'GC-2025-028', 'XX全过程咨询', 'XX交通局', 300000, 0, 100000, 200000, '33%', 210,
     '张三', '🔴 红灯', 3, '已发催收函，拟启动诉讼'],
]
for r, row in enumerate(sample3a, 7):
    for c, val in enumerate(row, 1):
        ws3.cell(row=r, column=c, value=val)
    style_body(ws3, r, len(headers3a))
    # Color the warning column
    warn = row[11]
    if '红灯' in str(warn):
        ws3.cell(row=r, column=12).fill = red_fill
    elif '黄灯' in str(warn):
        ws3.cell(row=r, column=12).fill = yellow_fill
    elif '绿灯' in str(warn):
        ws3.cell(row=r, column=12).fill = green_fill

# Section 2: 应收账款账龄分析
sec2_start = 7 + len(sample3a) + 2
ws3.merge_cells(start_row=sec2_start, start_column=1, end_row=sec2_start, end_column=14)
ws3.cell(row=sec2_start, column=1, value='二、应收账款账龄分析')
ws3.cell(row=sec2_start, column=1).font = Font(name='黑体', size=12, bold=True)

headers3b = ['部门', '项目数', '应收总额\n（元）', '0-3月\n（元）', '3-6月\n（元）',
             '6-9月\n（元）', '9-12月\n（元）', '1-2年\n（元）', '2年以上\n（元）',
             '本月回款\n（元）', '本月回款率', '环比变动', '预警项目数', '备注']
for c, h in enumerate(headers3b, 1):
    ws3.cell(row=sec2_start+2, column=c, value=h)
style_header(ws3, sec2_start+2, len(headers3b))

sample3b = [
    ['审计业务部', 15, 520000, 200000, 120000, 80000, 50000, 50000, 20000, 70000, '13.5%', '+2.1%', 3, ''],
    ['工程咨询部', 8, 680000, 300000, 200000, 100000, 50000, 30000, 0, 0, '0%', '-5.3%', 2, ''],
    ['合计', 23, 1200000, 500000, 320000, 180000, 100000, 80000, 20000, 70000, '5.8%', '-0.8%', 5, ''],
]
for r, row in enumerate(sample3b, sec2_start+3):
    for c, val in enumerate(row, 1):
        ws3.cell(row=r, column=c, value=val)
    style_body(ws3, r, len(headers3b))
    if row[0] == '合计':
        for c in range(1, len(headers3b)+1):
            ws3.cell(row=r, column=c).font = Font(name='微软雅黑', size=10, bold=True)

# Section 3: 回款排名
sec3_start = sec2_start + 3 + len(sample3b) + 2
ws3.merge_cells(start_row=sec3_start, start_column=1, end_row=sec3_start, end_column=14)
ws3.cell(row=sec3_start, column=1, value='三、项目经理本月回款排名')
ws3.cell(row=sec3_start, column=1).font = Font(name='黑体', size=12, bold=True)

headers3c = ['排名', '项目经理', '负责项目数', '本月应收目标\n（元）', '本月实际回款\n（元）',
             '本月回款率', '累计回款率', '超1年挂账\n消减数', '状态']
for c, h in enumerate(headers3c, 1):
    ws3.cell(row=sec3_start+2, column=c, value=h)
style_header(ws3, sec3_start+2, len(headers3c))

sample3c = [
    [1, '王五', 4, 50000, 40000, '80%', '92%', 1, '🟢'],
    [2, '张三', 6, 120000, 60000, '50%', '78%', 0, '🟡'],
    [3, '李四', 3, 200000, 0, '0%', '45%', 0, '🔴'],
]
for r, row in enumerate(sample3c, sec3_start+3):
    for c, val in enumerate(row, 1):
        ws3.cell(row=r, column=c, value=val)
    style_body(ws3, r, len(headers3c))

widths3 = [5, 12, 16, 14, 12, 12, 12, 12, 10, 8, 10, 10, 8, 20]
for c, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(c)].width = w

wb3.save(os.path.join(outdir, '附件3-项目回款月报.xlsx'))

print('✅ 3个附件模板已生成')
for f in ['附件1-年度双清目标清单.xlsx', '附件2-季度双清考核结果表.xlsx', '附件3-项目回款月报.xlsx']:
    print(f'  {os.path.join(outdir, f)}')
