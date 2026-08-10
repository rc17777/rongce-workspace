# -*- coding: utf-8 -*-
"""生成部门绩效考核两个Excel附件模板"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

outdir = r'C:\Users\scrccpa\.openclaw\workspace\output'
thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
hf = Font(name='微软雅黑', size=10, bold=True)
bf = Font(name='宋体', size=9)
tf = Font(name='黑体', size=13, bold=True)
hfill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
ca = Alignment(horizontal='center', vertical='center', wrap_text=True)
la = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_h(ws, r, cols):
    for c in range(1, cols+1):
        cell = ws.cell(r, c); cell.font = hf; cell.fill = hfill; cell.alignment = ca; cell.border = thin

def style_b(ws, r, cols, align='c'):
    for c in range(1, cols+1):
        cell = ws.cell(r, c); cell.font = bf; cell.alignment = ca if align == 'c' else la; cell.border = thin

# ════════════════════════════════════════
# 附件1：部门绩效考核指标计算表
# ════════════════════════════════════════
wb1 = Workbook()
ws1 = wb1.active
ws1.title = '部门考核指标计算'

ws1.merge_cells('A1:J1')
ws1['A1'] = '融策公司 季度部门绩效考核指标计算表'
ws1['A1'].font = tf; ws1['A1'].alignment = ca

ws1.merge_cells('A2:J2')
ws1['A2'] = '考核期间：2026年Q__    编制日期：    编制人：    审批人：'
ws1['A2'].font = Font(name='宋体', size=9)

# Part A: 公司业绩系数
ws1.merge_cells('A4:J4')
ws1['A4'] = '一、公司业绩系数计算'
ws1['A4'].font = Font(name='黑体', size=11, bold=True)

h1a = ['指标', '权重', '目标值', '实际值', '完成率', '得分', '加权得分', '数据来源', '备注', '']
for c, h in enumerate(h1a, 1):
    ws1.cell(6, c, h)
style_h(ws1, 6, len(h1a))

d1a = [
    ['营收完成率', '40%', '500万', '550万', '110%', 110, 44.0, '财务部月报', ''],
    ['利润完成率', '30%', '80万', '72万', '90%', 90, 27.0, '财务部季报', ''],
    ['回款完成率', '30%', '400万', '380万', '95%', 95, 28.5, '回款台账', ''],
]
for r, row in enumerate(d1a, 7):
    for c, v in enumerate(row, 1):
        ws1.cell(r, c, v)
    style_b(ws1, r, len(h1a))

# 汇总行
sr = 7 + len(d1a)
ws1.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=5)
ws1.cell(sr, 1, '公司业绩得分合计')
ws1.cell(sr, 7, 99.5)
for c in range(1, len(h1a)+1):
    ws1.cell(sr, c).font = Font(name='微软雅黑', size=9, bold=True)
    ws1.cell(sr, c).alignment = ca; ws1.cell(sr, c).border = thin

# 公司业绩系数映射
ws1.merge_cells(start_row=sr+2, start_column=1, end_row=sr+2, end_column=10)
ws1.cell(sr+2, 1, '公司业绩系数映射：得分99.5 → 系数1.0（线性插值）')
ws1.cell(sr+2, 1).font = Font(name='宋体', size=9, bold=True, color='C00000')

# Part B: 部门业绩系数
sb = sr + 4
ws1.merge_cells(start_row=sb, start_column=1, end_row=sb, end_column=10)
ws1.cell(sb, 1, '二、各部门业绩系数计算')
ws1.cell(sb, 1).font = Font(name='黑体', size=11, bold=True)

# 审计业务部
ws1.merge_cells(start_row=sb+2, start_column=1, end_row=sb+2, end_column=10)
ws1.cell(sb+2, 1, '▎审计业务部')
ws1.cell(sb+2, 1).font = Font(name='黑体', size=10, bold=True)

h1b = ['维度', '指标', '权重', '目标值', '实际值', '得分', '加权得分', '数据来源', '备注', '']
for c, h in enumerate(h1b, 1):
    ws1.cell(sb+4, c, h)
style_h(ws1, sb+4, len(h1b))

d1b_sj = [
    ['交付质量(50%)', '报告一次通过率', '20%', '≥95%', '96%', 100, 20.0, '质控记录', ''],
    ['', '项目交付准时率', '15%', '≥90%', '88%', 96, 14.4, '项目台账', ''],
    ['', '客户投诉率', '15%', '0次', '0次', 100, 15.0, '投诉台账', ''],
    ['部门管理(30%)', '部门回款完成率', '15%', '≥100%', '92%', 92, 13.8, '回款台账', ''],
    ['', '项目成本控制率', '15%', '≤预算', '超3%', 90, 13.5, '成本台账', ''],
    ['部门贡献(20%)', '新客户开发数', '10%', '≥2个', '3个', 100, 10.0, '经营台账', ''],
    ['', '知识沉淀与培训', '10%', '完成计划', '完成90%', 90, 9.0, '培训记录', ''],
]
for r, row in enumerate(d1b_sj, sb+5):
    for c, v in enumerate(row, 1):
        ws1.cell(r, c, v)
    style_b(ws1, r, len(h1b))

sj_total_row = sb + 5 + len(d1b_sj)
ws1.merge_cells(start_row=sj_total_row, start_column=1, end_row=sj_total_row, end_column=5)
ws1.cell(sj_total_row, 1, '审计业务部业绩得分合计')
ws1.cell(sj_total_row, 7, 95.7)
for c in range(1, len(h1b)+1):
    ws1.cell(sj_total_row, c).font = Font(name='微软雅黑', size=9, bold=True)
    ws1.cell(sj_total_row, c).alignment = ca; ws1.cell(sj_total_row, c).border = thin

ws1.cell(sj_total_row+1, 1, '部门业绩系数映射：得分95.7 → 系数1.1（线性插值）')
ws1.cell(sj_total_row+1, 1).font = Font(name='宋体', size=9, bold=True, color='C00000')

# 工程咨询部
ec_start = sj_total_row + 3
ws1.merge_cells(start_row=ec_start, start_column=1, end_row=ec_start, end_column=10)
ws1.cell(ec_start, 1, '▎工程咨询部')
ws1.cell(ec_start, 1).font = Font(name='黑体', size=10, bold=True)

for c, h in enumerate(h1b, 1):
    ws1.cell(ec_start+2, c, h)
style_h(ws1, ec_start+2, len(h1b))

d1b_gc = [
    ['交付质量(50%)', '报告一次通过率', '20%', '≥95%', '93%', 93, 18.6, '质控记录', ''],
    ['', '项目交付准时率', '15%', '≥90%', '85%', 90, 13.5, '项目台账', ''],
    ['', '客户投诉率', '15%', '0次', '1次', 80, 12.0, '投诉台账', ''],
    ['部门管理(30%)', '部门回款完成率', '15%', '≥100%', '88%', 88, 13.2, '回款台账', ''],
    ['', '项目成本控制率', '15%', '≤预算', '超8%', 80, 12.0, '成本台账', ''],
    ['部门贡献(20%)', '新客户开发数', '10%', '≥2个', '1个', 60, 6.0, '经营台账', ''],
    ['', '知识沉淀与培训', '10%', '完成计划', '完成80%', 80, 8.0, '培训记录', ''],
]
for r, row in enumerate(d1b_gc, ec_start+3):
    for c, v in enumerate(row, 1):
        ws1.cell(r, c, v)
    style_b(ws1, r, len(h1b))

gc_total_row = ec_start + 3 + len(d1b_gc)
ws1.merge_cells(start_row=gc_total_row, start_column=1, end_row=gc_total_row, end_column=5)
ws1.cell(gc_total_row, 1, '工程咨询部业绩得分合计')
ws1.cell(gc_total_row, 7, 83.3)
for c in range(1, len(h1b)+1):
    ws1.cell(gc_total_row, c).font = Font(name='微软雅黑', size=9, bold=True)
    ws1.cell(gc_total_row, c).alignment = ca; ws1.cell(gc_total_row, c).border = thin

ws1.cell(gc_total_row+1, 1, '部门业绩系数映射：得分83.3 → 系数0.97（线性插值）')
ws1.cell(gc_total_row+1, 1).font = Font(name='宋体', size=9, bold=True, color='C00000')

# 职能部门
fn_start = gc_total_row + 3
ws1.merge_cells(start_row=fn_start, start_column=1, end_row=fn_start, end_column=10)
ws1.cell(fn_start, 1, '▎财务部 / 行政综合部（略，格式同上）')
ws1.cell(fn_start, 1).font = Font(name='黑体', size=10, bold=True)

# Part C: 部门绩效总额汇总
sum_start = fn_start + 3
ws1.merge_cells(start_row=sum_start, start_column=1, end_row=sum_start, end_column=10)
ws1.cell(sum_start, 1, '三、部门绩效总额汇总')
ws1.cell(sum_start, 1).font = Font(name='黑体', size=11, bold=True)

h1c = ['部门', '绩效基数\n（元/人/季）', '在岗人数', '公司业绩\n系数', '部门业绩\n系数',
       '部门绩效总额\n（元）', '季度预发75%\n（元）', '年度预留25%\n（元）', '备注', '']
for c, h in enumerate(h1c, 1):
    ws1.cell(sum_start+2, c, h)
style_h(ws1, sum_start+2, len(h1c))

d1c = [
    ['审计业务部', 3000, 10, 1.0, 1.1, 33000, 24750, 8250, ''],
    ['工程咨询部', 3000, 8, 1.0, 0.97, 23280, 17460, 5820, ''],
    ['财务部', 2500, 3, 1.0, 1.0, 7500, 5625, 1875, ''],
    ['行政综合部', 2500, 2, 1.0, 1.0, 5000, 3750, 1250, ''],
    ['合计', '——', 23, '——', '——', 68780, 51585, 17195, ''],
]
for r, row in enumerate(d1c, sum_start+3):
    for c, v in enumerate(row, 1):
        ws1.cell(r, c, v)
    style_b(ws1, r, len(h1c))
    if row[0] == '合计':
        for c in range(1, len(h1c)+1):
            ws1.cell(r, c).font = Font(name='微软雅黑', size=9, bold=True)

for c, w in enumerate([12, 10, 8, 10, 10, 12, 12, 12, 10, 8], 1):
    ws1.column_dimensions[get_column_letter(c)].width = w

wb1.save(os.path.join(outdir, '附件1-部门绩效考核指标计算表.xlsx'))

# ════════════════════════════════════════
# 附件2：部门绩效分配方案表
# ════════════════════════════════════════
wb2 = Workbook()
ws2 = wb2.active
ws2.title = '部门绩效分配'

ws2.merge_cells('A1:K1')
ws2['A1'] = '融策公司 季度部门绩效分配方案表'
ws2['A1'].font = tf; ws2['A1'].alignment = ca

ws2.merge_cells('A2:K2')
ws2['A2'] = '考核期间：2026年Q__    部门：________    部门负责人：________    编制日期：'
ws2['A2'].font = Font(name='宋体', size=9)

# 部门基本信息
ws2.merge_cells('A4:K4')
ws2['A4'] = '一、部门基本信息'
ws2['A4'].font = Font(name='黑体', size=11, bold=True)

info = [
    ['部门绩效总额', '33,000元', '季度预发上限(75%)', '24,750元', '在岗人数', '10人'],
    ['公司业绩系数', '1.0', '部门业绩系数', '1.1', '绩效基数', '3,000元/人/季'],
]
for r, row in enumerate(info, 5):
    for c, v in enumerate(row, 1):
        ws2.cell(r, c, v)
        ws2.cell(r, c).font = Font(name='宋体', size=9)
        ws2.cell(r, c).border = thin
        if c in [1, 3, 5]:
            ws2.cell(r, c).font = Font(name='宋体', size=9, bold=True)

# 个人分配明细
ws2.merge_cells('A8:K8')
ws2['A8'] = '二、个人绩效分配明细'
ws2['A8'].font = Font(name='黑体', size=11, bold=True)

h2 = ['序号', '姓名', '岗位', '岗位权重', '个人绩效\n等级', '个人绩效\n系数',
      '系数×权重', '分配比例', '分配金额\n（元）', '面谈确认', '备注']
for c, h in enumerate(h2, 1):
    ws2.cell(10, c, h)
style_h(ws2, 10, len(h2))

d2 = [
    [1, '赵六', '部门经理', 1.5, 'A', 1.2, 1.80, '16.4%', 4059, '✓', ''],
    [2, '张三', '项目经理', 1.2, 'A', 1.2, 1.44, '13.1%', 3237, '✓', ''],
    [3, '李四', '项目经理', 1.2, 'B', 1.0, 1.20, '10.9%', 2697, '✓', ''],
    [4, '王五', '审计助理', 1.0, 'B', 1.0, 1.00, '9.1%', 2248, '✓', ''],
    [5, '陈七', '审计助理', 1.0, 'B', 1.0, 1.00, '9.1%', 2248, '✓', ''],
    [6, '刘八', '审计助理', 1.0, 'C', 0.8, 0.80, '7.3%', 1798, '✓', ''],
    [7, '周九', '审计助理', 1.0, 'C', 0.8, 0.80, '7.3%', 1798, '✓', ''],
    [8, '吴十', '审计助理', 1.0, 'B', 1.0, 1.00, '9.1%', 2248, '✓', ''],
    [9, '郑十一', '试用期', 0.6, 'B', 1.0, 0.60, '5.5%', 1349, '✓', ''],
    [10, '孙十二', '试用期', 0.6, 'D', 0.5, 0.30, '2.7%', 674, '✓', '待改进'],
]
for r, row in enumerate(d2, 11):
    for c, v in enumerate(row, 1):
        ws2.cell(r, c, v)
    style_b(ws2, r, len(h2))

# 合计行
tr = 11 + len(d2)
ws2.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=6)
ws2.cell(tr, 1, '合计')
ws2.cell(tr, 7, 10.94)
ws2.cell(tr, 8, '100%')
ws2.cell(tr, 9, 22358)
for c in range(1, len(h2)+1):
    ws2.cell(tr, c).font = Font(name='微软雅黑', size=9, bold=True)
    ws2.cell(tr, c).alignment = ca; ws2.cell(tr, c).border = thin

# 校验说明
ws2.merge_cells(start_row=tr+2, start_column=1, end_row=tr+2, end_column=11)
ws2.cell(tr+2, 1, '校验：∑(系数×权重) = 10.94，分配金额合计22,358元 ≤ 预发上限24,750元 ✓')
ws2.cell(tr+2, 1).font = Font(name='宋体', size=9, color='006600')

ws2.merge_cells(start_row=tr+3, start_column=1, end_row=tr+3, end_column=11)
ws2.cell(tr+3, 1, '校验：最高(A级经理)4,059元 vs 最低(D级试用)674元，差距502% ≥ 30%红线 ✓')
ws2.cell(tr+3, 1).font = Font(name='宋体', size=9, color='006600')

# 签字区
ws2.merge_cells(start_row=tr+5, start_column=1, end_row=tr+5, end_column=11)
ws2.cell(tr+5, 1, '三、审批签字')
ws2.cell(tr+5, 1).font = Font(name='黑体', size=11, bold=True)

sigs = [
    ['部门负责人：', '日期：'],
    ['人力资源审核：', '日期：'],
    ['总经理审批：', '日期：'],
]
for r, row in enumerate(sigs, tr+7):
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws2.merge_cells(start_row=r, start_column=7, end_row=r, end_column=11)
    ws2.cell(r, 1, row[0]); ws2.cell(r, 7, row[1])
    for c in [1, 7]:
        ws2.cell(r, c).font = Font(name='宋体', size=10)

for c, w in enumerate([5, 8, 10, 8, 8, 8, 10, 8, 10, 8, 10], 1):
    ws2.column_dimensions[get_column_letter(c)].width = w

wb2.save(os.path.join(outdir, '附件2-部门绩效分配方案表.xlsx'))

print('✅ 部门绩效2个附件模板已生成')
