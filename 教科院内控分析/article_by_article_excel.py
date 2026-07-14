import os, sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

out = r"D:\openclaw-workspace\教科院内控分析"
output_path = os.path.join(out, "政府采购vs一般采购逐条对比.xlsx")

wb = Workbook()

# ═══════ 样式 ═══════
hdr_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
hdr2_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
title_font = Font(name='微软雅黑', bold=True, size=14, color='2F5496')
sub_font = Font(name='微软雅黑', bold=True, size=12, color='2F5496')
bold_font = Font(name='微软雅黑', bold=True, size=10)
normal_font = Font(name='微软雅黑', size=10)
red_font = Font(name='微软雅黑', size=10, color='CC0000', bold=True)
gov_fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')  # 蓝色-政府采购
gen_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')  # 绿色-一般采购
same_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
diff_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
warn_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
add_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))
wrap_align = Alignment(wrap_text=True, vertical='top', horizontal='left')
center_align = Alignment(wrap_text=True, vertical='center', horizontal='center')

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_range(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = wrap_align
            cell.font = normal_font

# ═══════ Sheet 1: 总览对比 ═══════
ws1 = wb.active
ws1.title = "总览对比"

ws1.merge_cells('A1:G1')
ws1.cell(row=1, column=1, value='政府采购管理制度 vs 一般采购管理制度 — 逐条对比分析').font = title_font
ws1.cell(row=1, column=1).alignment = center_align

ws1.merge_cells('A2:G2')
ws1.cell(row=2, column=1, value='基于V8版本（政府采购制度最后一版，一般采购制度与V10仅2行差异）').font = Font(name='微软雅黑', size=10, color='666666')
ws1.cell(row=2, column=1).alignment = center_align

headers1 = ['对比维度', '政府采购管理制度（179行/12章）', '一般采购管理制度（146行/10章）', '关系定位']
for i, h in enumerate(headers1, 1):
    ws1.cell(row=4, column=i, value=h)
style_header(ws1, 4, len(headers1))

overview = [
    ['制度定位', '上位制度：管理法定政府采购行为（集采目录以内或限额以上）', '配套制度：管理校内自主采购行为（集采目录以外且限额以下）', '互补关系\n\\n政府采购管"大额法定"\n一般采购管"小额自主"'],
    ['法律依据', '《政府采购法》\n《政府采购法实施条例》\n《政府采购需求管理办法》\n《中小企业发展管理办法》\n高新区多项配套文件', '《政府采购法》\n《政府采购法实施条例》\n《高新区财政金融局关于规范限额标准以下非政府采购项目内控制度的通知》', '上位法相同\n一般采购依据更聚焦校内'],
    ['金额边界', '货物/服务类：50万元（含）以上\n工程类：100万元（含）以上\n（执行四川省最新标准）', '货物/服务类：50万元（不含）以下\n工程类：100万元（不含）以下', '金额无缝衔接\n50万/100万为分界线'],
    ['制度体量', '179行 / 8992字符 / 12章', '146行 / 约6000字符 / 10章', '政府采购制度篇幅多33行\n内容更全面'],
    ['章节数', '12章', '10章', '政府采购多2章\n（意向公开/进口产品）'],
    ['审批层级', '外部审批为主：\n• 区财政局审批预算\n• 区教育体育局审批计划\n• 区财政局审批进口产品\n• 政府采购网公开', '内部审批为主：\n• 1万以下：校长审批\n• 1-5万：分管领导审批\n• 5万以上：校长办公会审议', '外部监管 vs 内部审批\n是两套制度的根本差异'],
    ['采购方式', '六种法定方式：\n公开招标/邀请招标/竞争性谈判\n竞争性磋商/询价/单一来源\n网上竞价（通用货物）', '两种自主方式：\n比选（5万以上公开比选）\n询价（5万以下）\n\\n议定方式（特殊情况）', '政府采购方式法定不可变\n一般采购方式自定灵活'],
    ['评审小组', '评标委员会：5人以上单数\n评审专家≥2/3\n\\n竞争性谈判/询价小组：\n3人以上单数，专家≥2/3', '比选评审小组：\n5万以上：≥7人（每处室≥1人）\n采用综合评分法\n\\n询价小组：≥3人', '政府采购：专家主导\n一般采购：校内人员主导\n\\nV10版改为委托代理机构'],
    ['供应商质疑', '法定程序：\n• 询问：3个工作日内答复\n• 质疑：7个工作日内答复\n• 投诉：向财政部门\n有明确法律后果', '校内程序：\n• 质疑：公示期/收到通知后3日内提出\n• 答复：7个工作日内\n• 投诉：向纪检小组\n无外部法定救济', '政府采购质疑有法律强制力\n一般采购质疑为内部流程'],
    ['合同签订', '中标通知书发出30日内\n\\n合同签订后7个工作日报财政备案\n\\n合同签订后2个工作日内公告', '中选通知书发出3日后10日内\n\\n合同送办公室备案', '时限要求不同\n政府采购更严格（公告+备案）'],
    ['履约验收', '组织验收小组\n\\n大型项目邀请质量检测机构\n\\n验收方签字承担法律责任\n\\n付款时限：15日/中小企业10工作日', '验收小组3人（需求处室+负责人+资产管理员）\n\\n采购人/财务不得参与验收\n\\n15个工作日内确定验收时间', '政府采购验收更严格（第三方检测/法律责任）\n一般采购强调不相容岗位分离'],
    ['档案管理', '采购文件保存≥15年\n\\n可用电子档案\n\\n资料：会议纪要/实施计划表/招投标资料/合同/验收报告/信息统计', '比选：签到表/评审表/统计表/中选通知书/验收表/评审报告/比选文件/响应文件\n\\n询价：询价函/评审报告/验收表/响应文件', '政府采购15年法定存档\n一般采购无明确年限'],
    ['信息公开', '全程公开：\n• 采购意向（活动前30日）\n• 采购公告/文件/预算\n• 中标成交结果\n• 采购合同\n\\n渠道：四川政府采购网', '公开比选：\n• 指定平台公告\n• 中选结果公示≥2工作日\n\\n邀请比选：\n• 不公开，仅向3家以上发邀请', '政府采购：全社会公开\n一般采购：选择性公开/邀请'],
    ['中小企业政策', '有明确政策要求：\n• 200万以下货物服务/400万以下工程原则上专门面向\n• 预留≥30%份额', '无专项规定', '政府采购有法定中小企业扶持义务\n一般采购无此要求'],
    ['进口产品', '有专章管理（第八章）：\n• 专家论证+行业主管部门意见\n• 报财政部门审批\n• 禁止采购目录内禁止进口产品', '无相关规定', '进口产品采购仅适用政府采购制度\n删除政府采购制度后此管控缺失'],
    ['存续状态', 'V1-V8：存在（179行，内容未变）\nV9-V10：❌ 已删除', 'V1-V10：全程存续（146行）\nV8→V10仅2行修改', '⚠️ 政府采购已删除\n一般采购仍存续\n制度体系断裂'],
]

for i, row_data in enumerate(overview):
    for j, val in enumerate(row_data):
        ws1.cell(row=5+i, column=1+j, value=val)

style_range(ws1, 5, 5+len(overview)-1, len(headers1))

# Color
for i, row_data in enumerate(overview):
    if '删除' in str(row_data[1]) or '缺失' in str(row_data[1]):
        ws1.cell(row=5+i, column=2).fill = warn_fill
    if '存续' in str(row_data[2]):
        ws1.cell(row=5+i, column=3).fill = add_fill

ws1.column_dimensions['A'].width = 16
ws1.column_dimensions['B'].width = 48
ws1.column_dimensions['C'].width = 48
ws1.column_dimensions['D'].width = 32

for i in range(len(overview)):
    ws1.row_dimensions[5+i].height = 85

# ═══════ Sheet 2: 逐章逐条对比 ═══════
ws2 = wb.create_sheet("逐章逐条对比")

ws2.merge_cells('A1:F1')
ws2.cell(row=1, column=1, value='政府采购 vs 一般采购 — 逐章逐条对比').font = sub_font
ws2.cell(row=1, column=1).alignment = center_align

ws2.merge_cells('A2:F2')
ws2.cell(row=2, column=1, value='🟦 蓝色=政府采购制度条文  🟩 绿色=一般采购制度条文  🟨 黄色=关键差异  🔴 红色=制度缺失风险').font = Font(name='微软雅黑', size=9, color='666666')
ws2.cell(row=2, column=1).alignment = center_align

headers2 = ['对比编号', '对比主题', '政府采购管理制度条文', '一般采购管理制度条文', '差异分析', '风险/备注']
for i, h in enumerate(headers2, 1):
    ws2.cell(row=4, column=i, value=h)
style_header(ws2, 4, len(headers2))

articles = [
    # 第一章 总则
    ['A1', '制度目的',
     '规范学校政府采购工作，建立规范有序的运行机制，提高预算资金使用效益',
     '深化体制改革，转变政府职能，规范单位自主采购行为，提高采购效益',
     '目的相似但定位不同：\n政府采购偏"规范"+\n一般采购偏"自主"',
     ''],
    ['A2', '适用范围定义',
     '使用财政性资金采购集中采购目录以内或采购限额标准以上的货物、工程和服务',
     '政府集中采购目录以外、采购限额标准以下的货物、服务、工程类采购项目',
     '🔑 核心分界线：\n目录内/限额上→政府采购\n目录外/限额下→一般采购\n边界清晰，互为补充',
     '⚠️ 政府采购制度删除后，"目录内+限额上"的采购失去校内制度指引'],
    ['A3', '采购目录依据',
     '执行《四川省政府集中采购目录和采购限额标准》。目录以外/限额以下的项目参照本校一般采购管理制度',
     '限额标准：50万以下货物服务、100万以下工程（按四川省最新标准调整）',
     '政府采购制度明确引用了一般采购制度，形成制度链条\n一般采购也明确了自己的边界',
     '✅ 两套制度原本设计为互补体系'],
    ['A4', '采购原则',
     '公开透明、公平竞争、公正、诚实信用',
     '（无明确原则表述）',
     '政府采购有法定四原则\n一般采购未明确原则',
     ''],
    ['A5', '内控目标',
     '全面管控与突出重点并举、分工制衡与提升效能并重、权责对等与依法惩处并行。分事行权/分岗设权/分级授权',
     '（无内控目标专条）',
     '政府采购有完整内控目标框架\n一般采购直接进入操作层面',
     '政府采购制度的内控体系建设更为规范'],

    # 管理机构
    ['B1', '决策机构',
     '校长办公会/党组织委员会\n审定内控制度、研究重大采购事项、审定预算计划',
     '校长办公会/党组织委员会\n审议内控制度和流程、审批采购项目（5万以下授权分管领导）',
     '➡️ 两套制度决策机构相同\n但政府采购无金额分层授权（全部上会），一般采购有分层（5万以下可授权）',
     ''],
    ['B2', '归口管理部门',
     '总务处（8项职责）\n拟定制度/汇总预算/确定方式/组织实施/督促归档/组织验收/纠纷调处/资料保管',
     '总务处（8项职责）\n拟定制度/审定采购文件/审核合同/组织评审验收/组建小组/归档备案/培训指导/其他',
     '职责清单高度相似\n但政府采购多了"纠纷调处"职责',
     ''],
    ['B3', '执行部门',
     '各业务处室（6项职责）\n申报预算/编制计划/确认文件/确认中标结果/合同备案/申请支付',
     '各业务处室（8项职责）\n提出需求/编制计划/执行采购/拟定比选文件/协助投诉/合同备案/五万以上过会/经费验收',
     '➡️ 一般采购执行部门职责更多，直接负责采购全流程操作\n政府采购执行部门主要是"申报+确认"',
     '一般采购对业务处室的采购操作能力要求更高'],
    ['B4', '监督部门',
     '纪检小组\n监督采购法规执行情况、参与投诉答复',
     '纪检小组\n建立监督机制、监督制度和流程执行',
     '监督职责基本一致',
     ''],
    ['B5', '临时机构',
     '（无）', 
     '评审/验收小组\n3万以上由总务处组建\n3万以下由采购处室组建',
     '➡️ 一般采购设置了临时评审/验收小组\n政府采购由外部代理机构组织评审',
     ''],

    # 预算管理
    ['C1', '预算编制',
     '专章管理（第三章）\n硬化预算约束、细化编制\n未编制预算不得采购\n\\n预留份额面向中小企业（≥30%）',
     '（无采购预算专章）\n仅在各处室职责中提及"按预算编制计划"',
     '🔑 政府采购有完整预算管理专章\n包含中小企业预留份额政策\n一般采购缺乏独立预算管控',
     '⚠️ 删除政府采购制度后：\n采购预算-计划联动机制丧失\n中小企业预留份额义务无校内制度支撑'],

    # 意向公开
    ['D1', '采购意向公开',
     '专章管理（第四章）\n公开范围/内容/时间（采购前30日）/渠道（四川政府采购网）',
     '（无）',
     '🔑 政府采购独有章节\n一般采购无此要求',
     '⚠️ 删除后意向公开失去制度依据'],

    # 需求管理
    ['E1', '需求调查',
     '专章管理（第五章）\n需求调查：咨询/论证/问卷\n调查对象≥3个\n\\n1000万以上/公益项目/技术复杂项目必须调查',
     '（无专门需求调查要求）\n仅在评审中提到参数制定需合法合规',
     '🔑 政府采购有系统需求管理\n含调查/编制/审查三环节\n一般采购缺乏需求论证机制',
     '⚠️ 删除后大额采购需求缺乏科学论证要求'],

    # 采购实施
    ['F1', '采购方式种类',
     '六种法定方式：\n公开招标/邀请招标/竞争性谈判/竞争性磋商/询价/单一来源\n\n公开招标为主要方式',
     '两种自主方式：\n比选（公开比选/邀请比选）\n询价\n\n特殊情况可议定',
     '政府采购方式法定且不可变\n一般采购方式灵活，校内自定',
     '两者适用条件完全不同\n不可互相替代'],
    ['F2', '金额分层',
     '未设置校内金额分层\n均由政府采购流程处理',
     '三级体系：\n1万以下：自行采购\n1-5万：询价/比选\n5-10万：公开比选\n10万以上：委托代理公开比选',
     '➡️ 一般采购有精细的金额分层管理\n政府采购无此设计（因为起点就是50万）',
     ''],
    ['F3', '采购方式变更',
     '严格按照高新区管理办法执行\n填写变更申请表报教育体育局审批',
     '响应不足3家→采购失败→重新组织→仍不足3家→提请校长办公会议定',
     '政府采购变更需外部审批\n一般采购变更校内决策即可',
     ''],

    # 合同管理
    ['G1', '合同签订时限',
     '中标通知书发出之日起30日内',
     '中选通知书发出之日起3日后、10日内',
     '➡️ 政府采购时限更宽松（有公告期）\n一般采购更快',
     ''],
    ['G2', '合同备案',
     '签订后7个工作日报同级财政备案\n签订后2个工作日内在指定媒体公告',
     '签订后交学校办公室一份备案',
     '➡️ 政府采购有外部备案+公告义务\n一般采购仅内部备案',
     ''],
    ['G3', '补充合同限制',
     '追加金额不得超过原合同10%',
     '（无补充合同限制条款）',
     '政府采购对追加有限制\n一般采购无此条款',
     ''],

    # 验收
    ['H1', '验收组织',
     '组织验收小组\n大型项目邀请国家认可质量检测机构\n验收方签字承担法律责任',
     '验收小组3人（需求处室+负责人+资产管理员）\n采购人和财务不得参与（不相容岗位分离）\n15个工作日内确定验收日期',
     '➡️ 政府采购验收责任更重（法律责任）\n一般采购强调不相容岗位分离（内控亮点）',
     ''],
    ['H2', '付款时限',
     '收到发票后15日内支付\n中小企业：10个工作日内支付',
     '（无明确付款时限）',
     '政府采购有法定付款时限\n一般采购无规定',
     ''],

    # 质疑投诉
    ['I1', '询问处理',
     '3个工作日内答复',
     '（无询问处理条款）',
     '政府采购有明确时限\n一般采购无此机制',
     ''],
    ['I2', '质疑处理',
     '收到书面质疑后7个工作日内书面答复供应商\n可向财政部门投诉',
     '公示期内/收到通知后3日内提出质疑\n7个工作日内答复\n可向纪检小组投诉',
     '➡️ 质疑提出时限：政府采购7日 vs 一般采购3日\n投诉渠道：财政部门（法定） vs 纪检小组（内部）',
     '⚠️ 删除政府采购制度后\n法定质疑处理机制失去校内指引'],
    ['I3', '法律后果',
     '采购人改变结果的，依法承担法律责任',
     '（无法律后果规定）',
     '政府采购有明确法律责任\n一般采购无此条款',
     ''],

    # 信息公开
    ['J1', '信息公开要求',
     '专章管理（第十一章）\n全程公开：公告/文件/预算/结果/合同\n渠道：四川政府采购网\n公共服务项目验收结果公告',
     '公开比选：指定平台公告+中选公示≥2工作日\n邀请比选：不公开',
     '🔑 政府采购全社会公开、强制\n一般采购仅公开比选公示',
     '⚠️ 删除政府采购制度后\n采购信息公开失去制度支撑'],

    # 特殊专题
    ['K1', '政府购买服务',
     '专章管理（第七章）\n需求调研论证→计划申报→信息公开→委托代理→签订合同\n合同期限可签≤3年',
     '（无）',
     '🔑 政府采购独有章节\n适用于公共服务项目采购',
     '⚠️ 删除后政府购买服务失去校内制度'],
    ['K2', '进口产品采购',
     '专章管理（第八章）\n应当采购本国产品\n确需进口：专家论证+行业主管部门意见+财政审批\n禁止采购目录内禁进产品',
     '（无）',
     '🔑 政府采购独有章节\n进口产品的内控要求极严',
     '⚠️ 删除后进口采购失去三级论证审批制度'],
    ['K3', '中小企业政策',
     '200万以下货物服务/400万以下工程→面向中小企业\n预留≥30%预算份额',
     '（无）',
     '🔑 法定义务，政府采购独有',
     '⚠️ 删除后中小企业扶持义务无校内制度支撑'],
    ['K4', '档案保存年限',
     '采购文件保存≥15年（从采购结束日起）\n可用电子档案',
     '（无明确保存年限）',
     '🔑 政府采购15年法定存档\n一般采购未规定年限',
     ''],
    ['K5', '网上竞价',
     '有专门规定：\n通用货物单项年累计≤50万\n计算机/打印设备/空调/家具等',
     '（无）',
     '🔑 政府采购独有方式\n适用于零星小额通用货物',
     ''],

    # 附则
    ['Z1', '附则',
     '解释权：总务处\n施行日期：自发布之日起',
     '解释权：总务处\n施行日期：自发布之日起执行',
     '附则内容基本一致',
     ''],
]

for i, row_data in enumerate(articles):
    for j, val in enumerate(row_data):
        ws2.cell(row=5+i, column=1+j, value=val)

style_range(ws2, 5, 5+len(articles)-1, len(headers2))

# Color columns
for i in range(len(articles)):
    # Gov procurement column
    ws2.cell(row=5+i, column=3).fill = gov_fill
    # General procurement column
    ws2.cell(row=5+i, column=4).fill = gen_fill
    # Diff column
    diff_text = str(ws2.cell(row=5+i, column=5).value)
    if '🔑' in diff_text or '➡️' in diff_text:
        ws2.cell(row=5+i, column=5).fill = diff_fill
    # Risk column
    risk_text = str(ws2.cell(row=5+i, column=6).value)
    if '⚠️' in risk_text:
        ws2.cell(row=5+i, column=6).fill = warn_fill
        ws2.cell(row=5+i, column=6).font = Font(name='微软雅黑', size=10, color='CC0000')

ws2.column_dimensions['A'].width = 10
ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 52
ws2.column_dimensions['D'].width = 52
ws2.column_dimensions['E'].width = 38
ws2.column_dimensions['F'].width = 38

for i in range(len(articles)):
    ws2.row_dimensions[5+i].height = 85

ws2.freeze_panes = 'C5'

# ═══════ Sheet 3: 政府采购独有制度 ═══════
ws3 = wb.create_sheet("政府采购独有制度")

ws3.merge_cells('A1:E1')
ws3.cell(row=1, column=1, value='政府采购管理制度独有内容 — 一般采购管理制度中无对应条款').font = sub_font
ws3.cell(row=1, column=1).alignment = center_align

ws3.merge_cells('A2:E2')
ws3.cell(row=2, column=1, value='以下11项制度/机制仅存在于政府采购管理制度中，制度删除后全部丧失校内依据').font = Font(name='微软雅黑', size=10, color='CC0000')
ws3.cell(row=2, column=1).alignment = center_align

headers3 = ['编号', '独有制度/机制', '对应章节', '制度内容摘要', '删除后影响']
for i, h in enumerate(headers3, 1):
    ws3.cell(row=4, column=i, value=h)
style_header(ws3, 4, len(headers3))

unique_items = [
    ['U1', '中小企业预留份额制度', '第三章',
     '200万以下货物服务/400万以下工程专门面向中小企业；预留≥30%预算份额',
     '🔴 法定扶持义务失去校内执行制度'],
    ['U2', '采购意向公开制度', '第四章',
     '采购活动开始前30日在四川政府采购网公开采购意向（名称/需求/预算/时间）',
     '🔴 违反财政部意向公开要求风险'],
    ['U3', '采购需求调查制度', '第五章',
     '1000万以上/公益/技术复杂项目必须进行需求调查（调查对象≥3个）；需求编制+审查',
     '🔴 大额采购缺乏科学论证机制'],
    ['U4', '政府采购方式变更审批', '第六章',
     '严格按照高新区管理办法，填写变更申请表报教育体育局审批',
     '🟡 如需变更采购方式无校内指引'],
    ['U5', '评审专家制度', '第六章',
     '评标委员会5人以上单数，专家≥2/3；专家依法独立评审，承担个人责任',
     '🟡 评审专业性和独立性可能降低'],
    ['U6', '政府购买服务管理', '第七章',
     '需求调研→计划申报→信息公开→委托代理→签订合同；合同期限可≤3年',
     '🟡 公共服务采购特殊性不被识别'],
    ['U7', '进口产品采购管理', '第八章',
     '应采购本国产品；确需进口：专家论证+行业主管部门意见+财政审批',
     '🔴 如涉及进口采购则失去三级管控'],
    ['U8', '合同外部备案+公告', '第九章',
     '签订后7日报财政备案；2日内在指定媒体公告；15年存档',
     '🟡 合规风险'],
    ['U9', '询问处理制度', '第十章',
     '供应商询问3个工作日内答复；超出代理授权范围告知供应商向采购人提出',
     '🟡 供应商权益保障机制缺失'],
    ['U10', '法定质疑投诉渠道', '第十章',
     '书面质疑7日内答复→不服可向财政部门投诉；有明确法律后果',
     '🟡 投诉可能直接升级到财政部门'],
    ['U11', '采购全程信息公开', '第十一章',
     '采购公告/文件/预算/中标结果/合同全程在四川政府采购网公开',
     '🟡 公开透明度降低'],
]

for i, row_data in enumerate(unique_items):
    for j, val in enumerate(row_data):
        ws3.cell(row=5+i, column=1+j, value=val)

style_range(ws3, 5, 5+len(unique_items)-1, len(headers3))

for i, row_data in enumerate(unique_items):
    risk = row_data[4]
    cell = ws3.cell(row=5+i, column=5)
    if '🔴' in risk:
        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        cell.font = Font(name='微软雅黑', size=10, color='CC0000', bold=True)
    elif '🟡' in risk:
        cell.fill = diff_fill

ws3.column_dimensions['A'].width = 6
ws3.column_dimensions['B'].width = 26
ws3.column_dimensions['C'].width = 12
ws3.column_dimensions['D'].width = 60
ws3.column_dimensions['E'].width = 40

for i in range(len(unique_items)):
    ws3.row_dimensions[5+i].height = 60

# ═══════ Sheet 4: 一般采购独有 ═══════
ws4 = wb.create_sheet("一般采购独有制度")

ws4.merge_cells('A1:E1')
ws4.cell(row=1, column=1, value='一般采购管理制度独有内容 — 政府采购管理制度中无对应条款').font = sub_font
ws4.cell(row=1, column=1).alignment = center_align

headers4 = ['编号', '独有制度/机制', '对应章节', '制度内容摘要', '评价']
for i, h in enumerate(headers4, 1):
    ws4.cell(row=4, column=i, value=h)
style_header(ws4, 4, len(headers4))

gen_unique = [
    ['G1', '精细金额分层管理', '第四章',
     '1万以下：自行采购\n1-5万：比选/询价（校长审批）\n5-10万：公开比选\n10万以上：委托代理公开比选',
     '✅ 亮点：金额分层清晰，审批权限明确，避免了"一刀切"'],
    ['G2', '评审/验收临时小组', '第三章',
     '3万以上由总务处组建评审验收小组\n3万以下由采购处室组建\n不同金额不同组织主体',
     '✅ 灵活高效，兼顾规模与效率'],
    ['G3', '不相容岗位分离（验收）', '第七章',
     '验收小组3人（含处室负责人+资产管理员）\n采购人和财务人员不得参与验收',
     '✅ 内控亮点：采购-验收-支付三权分离'],
    ['G4', '内部投诉渠道', '第五章',
     '质疑不满意可向纪检小组投诉\n纪检小组7个工作日内答复',
     '✅ 建立了内部救济渠道'],
    ['G5', '比选/询价标准化附件', '第十章',
     '签到表/密封检查记录/报价表/技术参数表/评审表/比选决议/验收表\n（7套标准附件模板）',
     '✅ 操作标准化程度高，有模板可依'],
    ['G6', '供应商资质通用条件', '第五章',
     '独立法人/专业能力/健全财务/3年无违法/信用良好/其他\n（6项通用资质条件）',
     '✅ 供应商准入门槛明确'],
    ['G7', '采购方式灵活变更', '第五章',
     '响应不足3家→重采→仍不足→校长办公会议定\n体现了"效率优先"的原则',
     '✅ 避免程序僵化导致采购无法推进'],
]

for i, row_data in enumerate(gen_unique):
    for j, val in enumerate(row_data):
        ws4.cell(row=5+i, column=1+j, value=val)

style_range(ws4, 5, 5+len(gen_unique)-1, len(headers4))

for i in range(len(gen_unique)):
    ws4.cell(row=5+i, column=5).fill = add_fill

ws4.column_dimensions['A'].width = 6
ws4.column_dimensions['B'].width = 26
ws4.column_dimensions['C'].width = 12
ws4.column_dimensions['D'].width = 60
ws4.column_dimensions['E'].width = 45

for i in range(len(gen_unique)):
    ws4.row_dimensions[5+i].height = 60

# ═══════ Sheet 5: V10变化 ═══════
ws5 = wb.create_sheet("V10一般采购变化")

ws5.merge_cells('A1:D1')
ws5.cell(row=1, column=1, value='一般采购管理制度 — V8→V10唯一实质性修改').font = sub_font
ws5.cell(row=1, column=1).alignment = center_align

ws5.merge_cells('A2:D2')
ws5.cell(row=2, column=1, value='仅2行修改，其余全部相同（146行/10章）').font = Font(name='微软雅黑', size=10, color='666666')
ws5.cell(row=2, column=1).alignment = center_align

headers5 = ['修改项', 'V8原文', 'V10修改后', '变化分析']
for i, h in enumerate(headers5, 1):
    ws5.cell(row=4, column=i, value=h)
style_header(ws5, 4, len(headers5))

v10_changes = [
    ['评审小组组建规则\n（第五章-比选程序-第八款）',
     '5万元以上项目，由总务处采购专员组建不少于7人的评审小组。比选评审小组成员原则上每个处室不少于1人，由总务处采购管理员、需求处室、总务处经办人、纪检小组及相关领域专家等组成。',
     '5万元以上项目，委托社会代理机构采用法定采购组织方式进行。由业务处室负责确定采购代表人选，社会代理机构从政府采购专家库随机抽取2名及以上专家，组成不少于3人的单数评审小组。',
     '🔴 重大变化：\nV8：校内7人评审小组（内部）\nV10：委托代理机构+随机抽取专家（外部）\n\n评审主体从"内部人员"转向"外部专家"\n更加专业化、独立化\n但同时意味着学校对评审过程控制力减弱'],
    ['采购方式变更审批主体\n（第五章-询价程序-第六款）',
     '…提请党政联席会/党组织委员会议定变更采购方式',
     '…提请校长办公会/党组织委员会议定变更采购方式',
     '🟡 贯穿性术语更正：\nV8此处遗漏修改（其他位置已是"校长办公会"）\nV10统一修正为"校长办公会"'],
]

for i, row_data in enumerate(v10_changes):
    for j, val in enumerate(row_data):
        ws5.cell(row=5+i, column=1+j, value=val)

style_range(ws5, 5, 5+len(v10_changes)-1, len(headers5))

ws5.column_dimensions['A'].width = 24
ws5.column_dimensions['B'].width = 55
ws5.column_dimensions['C'].width = 55
ws5.column_dimensions['D'].width = 42

ws5.row_dimensions[5].height = 160
ws5.row_dimensions[6].height = 80

# ═══════ 保存 ═══════
wb.save(output_path)
print(f"Excel saved to: {output_path}")
print("Sheets:", wb.sheetnames)
