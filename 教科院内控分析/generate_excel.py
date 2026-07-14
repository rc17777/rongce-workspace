import os, sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

out = r"D:\openclaw-workspace\教科院内控分析"
output_path = os.path.join(out, "内控制度差异性分析.xlsx")

wb = Workbook()

# ═══════════════════════ 样式定义 ═══════════════════════
header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
title_font = Font(name='微软雅黑', bold=True, size=14, color='2F5496')
subtitle_font = Font(name='微软雅黑', bold=True, size=12, color='2F5496')
bold_font = Font(name='微软雅黑', bold=True, size=10)
normal_font = Font(name='微软雅黑', size=10)
warn_font = Font(name='微软雅黑', size=10, color='FF0000')
add_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
del_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
change_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
header2_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='center', horizontal='left')
center_align = Alignment(wrap_text=True, vertical='center', horizontal='center')

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_range(ws, start_row, end_row, max_col, font=None):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = wrap_align
            if font:
                cell.font = font
            else:
                cell.font = normal_font

# ═══════════════════════ Sheet 1: 版本概览 ═══════════════════════
ws1 = wb.active
ws1.title = "版本概览"

ws1.merge_cells('A1:G1')
ws1.cell(row=1, column=1, value='成都市教科院附中内部控制制度汇编 — 版本差异性分析').font = title_font
ws1.cell(row=1, column=1).alignment = center_align

ws1.merge_cells('A2:G2')
ws1.cell(row=2, column=1, value='分析日期：2026-05-14  |  共10个版本  |  覆盖周期：2024年6月 → 至今').font = Font(name='微软雅黑', size=9, color='666666')
ws1.cell(row=2, column=1).alignment = center_align

headers1 = ['版本', '时间范围', '制度数量', '篇章数', '关键词变化', '净增行数', '变化等级']
for i, h in enumerate(headers1, 1):
    ws1.cell(row=4, column=i, value=h)
style_header(ws1, 4, len(headers1))

versions_data = [
    ['V1', '2024.06.14 - 2024.11.18', '17项', '3部分', '初始版本（基准）', '—', '—'],
    ['V2', '2024.11.08 - 2025.03.17', '17项', '3部分', '微调，无结构性变化', '+0', '⚪ 微小'],
    ['V3', '2025.03.17 - 2025.04.11', '17项', '3部分', '治理改革：「党政联席会」→「校长办公会」', '+2', '🟡 重要'],
    ['V4', '2025.04.11 - 2025.05.27', '18项', '3部分', '新增「建设项目实施管理办法」', '+40', '🟡 重要'],
    ['V5', '2025.05.27 - 2025.07.11', '21项', '4部分', '新增第四部分：3份国家内控政策文件', '+221', '🔴 重大'],
    ['V6', '2025.07.11 - 2025.10.13', '21项', '4部分', '微调，无结构性变化', '-12', '⚪ 微小'],
    ['V7', '2025.10.13 - 2025.11.05', '21项', '4部分', '教育收费制度细化（增加收费标准明细）', '-29', '🟡 重要'],
    ['V8', '2025.11.05 - 2026.01.23', '21项', '4部分', '财务重构：收支管理→财务管理+票据管理办法', '+146', '🔴 重大'],
    ['V9', '2026.01.23 - 2026.03.13', '20项', '4部分', '删除「政府采购管理制度」', '-27', '🟡 重要'],
    ['V10', '2026.03.13 - 至今', '19项', '4部分', '内部审计精简+提交清单移除', '+0', '🟡 重要'],
]

for i, row_data in enumerate(versions_data):
    for j, val in enumerate(row_data):
        ws1.cell(row=5+i, column=1+j, value=val)
style_range(ws1, 5, 5+len(versions_data)-1, len(headers1))

# Highlight important rows
for i, row_data in enumerate(versions_data):
    if row_data[6] == '🔴 重大':
        for c in range(1, len(headers1)+1):
            ws1.cell(row=5+i, column=c).fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

ws1.column_dimensions['A'].width = 8
ws1.column_dimensions['B'].width = 28
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 10
ws1.column_dimensions['E'].width = 50
ws1.column_dimensions['F'].width = 12
ws1.column_dimensions['G'].width = 12

# ═══════════════════════ Sheet 2: 制度变迁矩阵 ═══════════════════════
ws2 = wb.create_sheet("制度变迁矩阵")

ws2.merge_cells('A1:R1')
ws2.cell(row=1, column=1, value='制度变迁矩阵 — 全部制度在各版本中的存续状态').font = subtitle_font
ws2.cell(row=1, column=1).alignment = center_align

headers2 = ['序号', '制度名称', 'V1', 'V2', 'V3', 'V4', 'V5', 'V6', 'V7', 'V8', 'V9', 'V10', '首次出现', '最后存在', '状态', '所属篇章']
for i, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=i, value=h)
style_header(ws2, 3, len(headers2))

# Data: [name, v1-v10 status, first, last, state, part]
# v_status: ✅ = 存在, ❌ = 不存在
sections_data = [
    ['"三重一大"决策制度实施办法',         '✅','✅','✅','✅','✅','✅','✅','✅','✅','✅', 'V1', 'V10', '未变化', '议事决策'],
    ['党组织会议和党政联席会（行政会）议事规则', '✅','✅','❌','❌','❌','❌','❌','❌','❌','❌', 'V1', 'V2', '已删除', '议事决策'],
    ['党组织会议和校长办公会（行政会）议事规则', '❌','❌','✅','✅','✅','✅','✅','✅','✅','✅', 'V3', 'V10', '替代旧版', '议事决策'],
    ['党政联席会议提交党组织委员会讨论决定事项清单', '✅','✅','❌','❌','❌','❌','❌','❌','❌','❌', 'V1', 'V2', '已删除', '议事决策'],
    ['校长办公会议提交党组织委员会讨论决定事项清单', '❌','❌','✅','✅','✅','✅','✅','✅','✅','❌', 'V3', 'V9', '已删除', '议事决策'],
    ['会议管理制度',                        '✅','✅','✅','✅','✅','✅','✅','✅','✅','✅', 'V1', 'V10', '未变化', '议事决策'],
    ['工会会议制度',                        '✅','✅','✅','✅','✅','✅','✅','✅','✅','✅', 'V1', 'V10', '未变化', '议事决策'],
    ['教职工代表大会会议制度',               '✅','✅','✅','✅','✅','✅','✅','✅','✅','✅', 'V1', 'V10', '未变化', '议事决策'],
    ['预算管理制度',                        '✅','✅','✅','✅','✅','✅','✅','✅','✅','✅', 'V1', 'V10', '术语修订', '经济业务'],
    ['收支管理制度',                        '✅','✅','✅','✅','✅','✅','✅','❌','❌','❌', 'V1', 'V7', '被替代', '经济业务'],
    ['财务管理制度',                        '❌','❌','❌','❌','❌','❌','❌','✅','✅','✅', 'V8', 'V10', '新增替代', '经济业务'],
    ['费用报销流程及规定',                   '✅','✅','✅','✅','✅','✅','✅','✅','✅','✅', 'V1', 'V10', '未变化', '经济业务'],
    ['财务印章管理办法',                    '✅','✅','✅','✅','✅','✅','✅','✅','✅','✅', 'V1', 'V10', '未变化', '经济业务'],
    ['公务卡管理办法',                      '✅','✅','✅','✅','✅','✅','✅','✅','✅','✅', 'V1', 'V10', '未变化', '经济业务'],
    ['票据管理办法',                        '❌','❌','❌','❌','❌','❌','❌','✅','✅','✅', 'V8', 'V10', '新增', '经济业务'],
    ['工会活动及经费使用管理办法',            '✅','✅','✅','✅','✅','✅','✅','✅','✅','✅', 'V1', 'V10', '未变化', '经济业务'],
    ['教育收费管理制度',                    '✅','✅','✅','✅','✅','✅','✅','✅','✅','✅', 'V1', 'V10', '内容扩展', '经济业务'],
    ['政府采购管理制度',                    '✅','✅','✅','✅','✅','✅','✅','✅','❌','❌', 'V1', 'V8', '已删除', '经济业务'],
    ['一般采购管理制度',                    '✅','✅','✅','✅','✅','✅','✅','✅','✅','✅', 'V1', 'V10', '术语修订', '经济业务'],
    ['建设项目实施管理办法',                '❌','❌','❌','✅','✅','✅','✅','✅','✅','✅', 'V4', 'V10', '新增', '经济业务'],
    ['资产管理制度',                        '✅','✅','✅','✅','✅','✅','✅','✅','✅','✅', 'V1', 'V10', '术语修订', '经济业务'],
    ['合同管理制度',                        '✅','✅','✅','✅','✅','✅','✅','✅','✅','✅', 'V1', 'V10', '微调', '经济业务'],
    ['低值易耗品管理规定',                   '✅','✅','✅','✅','✅','✅','✅','✅','✅','✅', 'V1', 'V10', '未变化', '行政运行'],
    ['办公用物品采购入库领用制度',            '✅','✅','✅','✅','✅','✅','✅','✅','✅','✅', 'V1', 'V10', '未变化', '行政运行'],
    ['档案管理制度',                        '✅','✅','✅','✅','✅','✅','✅','✅','✅','✅', 'V1', 'V10', '未变化', '行政运行'],
    ['印章管理制度',                        '✅','✅','✅','✅','✅','✅','✅','✅','✅','✅', 'V1', 'V10', '未变化', '行政运行'],
    ['维修管理制度',                        '✅','✅','✅','✅','✅','✅','✅','✅','✅','✅', 'V1', 'V10', '未变化', '行政运行'],
    ['风险评估制度',                        '✅','✅','✅','✅','✅','✅','✅','✅','✅','✅', 'V1', 'V10', '未变化', '行政运行'],
    ['内部审计管理制度',                    '✅','✅','✅','✅','✅','✅','✅','✅','✅','✅', 'V1', 'V10', '内容精简', '行政运行'],
    ['内控关键岗位人员轮岗交流暂行办法',      '✅','✅','✅','✅','✅','✅','✅','✅','✅','✅', 'V1', 'V10', '未变化', '行政运行'],
    ['行政事业单位内部控制规范（试行）',      '❌','❌','❌','❌','✅','✅','✅','✅','✅','✅', 'V5', 'V10', '新增', '政策文件'],
    ['财政部关于全面推进行政事业单位内控建设的指导意见', '❌','❌','❌','❌','✅','✅','✅','✅','✅','✅', 'V5', 'V10', '新增', '政策文件'],
    ['行政事业单位内部控制报告管理制度（试行）', '❌','❌','❌','❌','✅','✅','✅','✅','✅','✅', 'V5', 'V10', '新增', '政策文件'],
]

for i, row_data in enumerate(sections_data):
    ws2.cell(row=4+i, column=1, value=i+1)
    for j, val in enumerate(row_data):
        ws2.cell(row=4+i, column=2+j, value=val)

style_range(ws2, 4, 4+len(sections_data)-1, len(headers2))

# Color the status cells
for i, row_data in enumerate(sections_data):
    for j in range(2, 12):  # V1-V10 columns
        cell = ws2.cell(row=4+i, column=2+j)
        cell.alignment = center_align
        if cell.value == '✅':
            cell.fill = add_fill
        elif cell.value == '❌':
            cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    # Color state column
    state_cell = ws2.cell(row=4+i, column=15)
    state = str(state_cell.value)
    if '删除' in state or '省略' in state or '精简' in state or '替代' in state:
        state_cell.fill = del_fill
        state_cell.font = Font(name='微软雅黑', size=10, color='CC0000')
    elif '新增' in state or '扩展' in state:
        state_cell.fill = add_fill
        state_cell.font = Font(name='微软雅黑', size=10, color='006100')
    elif '修订' in state:
        state_cell.fill = change_fill

ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 48
for c in ['C','D','E','F','G','H','I','J','K','L']:
    ws2.column_dimensions[c].width = 5
ws2.column_dimensions['M'].width = 10
ws2.column_dimensions['N'].width = 10
ws2.column_dimensions['O'].width = 12
ws2.column_dimensions['P'].width = 14

ws2.freeze_panes = 'C4'

# ═══════════════════════ Sheet 3: 逐版本差异详情 ═══════════════════════
ws3 = wb.create_sheet("逐版本差异详情")

ws3.merge_cells('A1:F1')
ws3.cell(row=1, column=1, value='逐版本差异详情').font = subtitle_font
ws3.cell(row=1, column=1).alignment = center_align

headers3 = ['版本变化', '时间区间', '变化类型', '变化内容', '变化行数', '影响评估']
for i, h in enumerate(headers3, 1):
    ws3.cell(row=3, column=i, value=h)
style_header(ws3, 3, len(headers3))

diff_details = [
    ['V1→V2', '2024.11 → 2025.03', '微调',
     '无结构性变化，仅8行格式修正',
     '+0行 (+8/-8)', '无实质影响'],
    
    ['V2→V3', '2025.03.17 → 2025.04.11\n⚠️ 治理改革',
     '制度变更', 
     '【删除】\n• 党组织会议和党政联席会（行政会）议事规则\n• 学校党组织委员会会议直接讨论决定事项清单\n\n【新增】\n• 党组织会议和校长办公会（行政会）议事规则\n• 校长办公会议（行政会）提交党组织委员会讨论决定事项清单\n\n【影响】预算管理/政府采购/资产管理等制度中"决策机构"表述同步从"党政联席会"变更为"校长办公会"',
     '+2行 (+66/-64)', '🟡 重要：落实中小学校党组织领导的校长负责制，涉及全部经济业务制度的决策机构表述变更'],
    
    ['V3→V4', '2025.04.11 → 2025.05.27\n⚠️ 补缺',
     '新增制度',
     '【新增】建设项目实施管理办法\n• 小型基建项目（门卫室/垃圾房/厕所等改扩建）\n• 维修改造项目（立面整治/屋面改造/环境整治等）\n• 装饰装修项目（教室/办公室/功能室/食堂/宿舍等）\n• 三级审批体系：5万以下/5万-400万/400万以上',
     '+40行 (+64/-24)', '🟡 重要：填补工程建设领域制度空白，设定了清晰的分级审批和资金管控机制'],
    
    ['V4→V5', '2025.05.27 → 2025.07.11\n⚠️ 对标国标',
     '新增篇章',
     '【新增"第四部分 政策文件"】\n1. 行政事业单位内部控制规范（试行）— 财政部2012\n2. 财政部关于全面推进行政事业单位内部控制建设的指导意见 — 2015\n3. 行政事业单位内部控制报告管理制度（试行）— 2017\n\n从"自编制度"提升为"对标国家标准"，为校内制度提供上位法依据',
     '+221行 (+222/-1)', '🔴 重大：本周期最大增量，制度合规性显著增强'],
    
    ['V5→V6', '2025.07.11 → 2025.10.13', '微调',
     '无结构性变化，轻微内容调整',
     '-12行 (+49/-61)', '无实质影响'],
    
    ['V6→V7', '2025.10.13 → 2025.11.05', '内容扩展',
     '教育收费管理制度大幅细化：从38行扩展至包含完整收费标准明细（学费/住宿费/考试费/教材费/作业本费/伙食费/校服费/教辅材料费/课后服务费/研学旅行费/医疗保险等），实现收费透明化',
     '-29行 (+18/-47)', '🟡 重要：收费管理透明化，但部分费用数据在V8版被整合'],
    
    ['V7→V8', '2025.11.05 → 2026.01.23\n⚠️ 财务重构',
     '制度替代+新增',
     '【删除】\n• 收支管理制度（164行）\n\n【新增】\n• 财务管理制度（272行）— 全面替代，覆盖：预算/决算/资金/往来/存货/固定资产/无形资产/收入/专项资金/支出/票据/公务卡/银行账户/会计档案\n• 票据管理办法（39行）— 独立制度，覆盖电子票据管理\n\n定位升级：操作型→体系型',
     '+146行 (+373/-227)', '🔴 重大：财务制度体系化重构，从收支两端扩展至全财务要素覆盖'],
    
    ['V8→V9', '2026.01.23 → 2026.03.13\n⚠️ 采购制度删除',
     '制度删除',
     '【删除】政府采购管理制度（179行，V1-V8重要制度）\n\n可能原因：\n① 学校采购已纳入上级统一采购平台\n② 整合至一般采购管理制度\n③ 响应"放管服"简化校内制度\n\n⚠️ 审计风险：需确认政府采购职能是否有明确承接制度',
     '-27行 (+139/-166)', '🟡 重要：政府采购内控可能存在制度空白，建议核实'],
    
    ['V9→V10', '2026.03.13 → 至今', '精简优化',
     '【删除】校长办公会议（行政会）提交党组织委员会讨论决定事项清单\n\n【精简】内部审计管理制度（97行→67行）：\n• 删除：审计工作抽样审计工作程序章节\n• 删除：审计工作底稿详细要求章节\n• 保留：审计范围/被审计人员权利义务/审计内容（教育收费/食堂收费/其他收费）',
     '+0行 (+7/-7)', '🟡 重要：内部审计操作标准化程度可能降低，建议关注实务中是否需要单独操作细则'],
]

for i, row_data in enumerate(diff_details):
    for j, val in enumerate(row_data):
        ws3.cell(row=4+i, column=1+j, value=val)

style_range(ws3, 4, 4+len(diff_details)-1, len(headers3))

# Row heights for readability
for i, data in enumerate(diff_details):
    if i in [1, 2, 6]:  # Detailed rows
        ws3.row_dimensions[4+i].height = 200
    elif i in [3, 7]:
        ws3.row_dimensions[4+i].height = 150
    else:
        ws3.row_dimensions[4+i].height = 80

# Color the impact column
for i, data in enumerate(diff_details):
    impact = data[5]
    cell = ws3.cell(row=4+i, column=6)
    if '重大' in impact:
        cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        cell.font = Font(name='微软雅黑', size=10, bold=True, color='CC0000')
    elif '重要' in impact:
        cell.font = Font(name='微软雅黑', size=10, color='BF8F00')

ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 22
ws3.column_dimensions['C'].width = 12
ws3.column_dimensions['D'].width = 70
ws3.column_dimensions['E'].width = 16
ws3.column_dimensions['F'].width = 40

# ═══════════════════════ Sheet 4: 核心制度内容对比 ═══════════════════════
ws4 = wb.create_sheet("核心制度内容对比")

ws4.merge_cells('A1:F1')
ws4.cell(row=1, column=1, value='核心制度内容对比（V1 初始版 vs V10 最新版）').font = subtitle_font
ws4.cell(row=1, column=1).alignment = center_align

headers4 = ['制度名称', 'V1状态', 'V10状态', 'V1行数', 'V10行数', '差异说明']
for i, h in enumerate(headers4, 1):
    ws4.cell(row=3, column=i, value=h)
style_header(ws4, 3, len(headers4))

content_compare = [
    ['"三重一大"决策制度实施办法',     '存在', '存在', '—', '—', '内容未变化'],
    ['党组织会议议事规则',            '「党政联席会」版', '「校长办公会」版', '51', '51', '名称变更，行数不变（V3变更）'],
    ['预算管理制度',                 '存在', '存在', '102', '102', '术语修订：决策机构"党政联席会"→"校长办公会"'],
    ['收支管理制度',                 '存在', '❌ 已删除', '164', '—', 'V8被财务管理制度替代'],
    ['财务管理制度',                 '❌ 不存在', '存在', '—', '272', 'V8新增，全面替代收支管理；覆盖预决算/资金/存货/固定资产/无形资产等全财务要素'],
    ['费用报销流程及规定',            '存在', '存在', '—', '—', '内容未变化'],
    ['票据管理办法',                 '❌ 不存在', '存在', '—', '39', 'V8新增；涵盖电子票据管理/财政票据种类与领用核销'],
    ['政府采购管理制度',              '存在', '❌ 已删除', '179', '—', 'V9删除；需确认是否有替代制度承接'],
    ['一般采购管理制度',              '存在', '存在', '147', '147', '术语修订：决策机构表述变更'],
    ['建设项目实施管理办法',          '❌ 不存在', '存在', '—', '38', 'V4新增；5万/60万/400万三级审批+过程控制+工程审计'],
    ['教育收费管理制度',              '存在', '存在', '38', '189', 'V7大幅扩展（增加收费明细），V8整合优化'],
    ['资产管理制度',                 '存在', '存在', '158', '158', '术语修订：决策机构"党政联席会"→"校长办公会"'],
    ['合同管理制度',                 '存在', '存在', '97', '98', '微调（+1行）'],
    ['风险评估制度',                 '存在', '存在', '96', '96', '内容未变化'],
    ['内部审计管理制度',              '存在', '存在', '97', '67', 'V10精简30行：删除审计抽样程序章节和审计底稿详细要求；保留审计范围和收费审计内容'],
]

for i, row_data in enumerate(content_compare):
    for j, val in enumerate(row_data):
        ws4.cell(row=4+i, column=1+j, value=val)

style_range(ws4, 4, 4+len(content_compare)-1, len(headers4))

for i, row_data in enumerate(content_compare):
    # Color V10 status
    v10_cell = ws4.cell(row=4+i, column=3)
    if '删除' in str(v10_cell.value):
        v10_cell.fill = del_fill
        v10_cell.font = Font(name='微软雅黑', size=10, color='CC0000')
    elif '新增' in str(v10_cell.value):
        v10_cell.fill = add_fill
    
    # Color 说明
    note_cell = ws4.cell(row=4+i, column=6)
    note = str(note_cell.value)
    if '删除' in note or '替代' in note:
        note_cell.fill = del_fill
    elif '新增' in note:
        note_cell.fill = add_fill
    elif '修订' in note or '变更' in note:
        note_cell.fill = change_fill

ws4.column_dimensions['A'].width = 30
ws4.column_dimensions['B'].width = 14
ws4.column_dimensions['C'].width = 14
ws4.column_dimensions['D'].width = 10
ws4.column_dimensions['E'].width = 10
ws4.column_dimensions['F'].width = 60

# ═══════════════════════ Sheet 5: 审计建议与风险评估 ═══════════════════════
ws5 = wb.create_sheet("审计建议与风险评估")

ws5.merge_cells('A1:E1')
ws5.cell(row=1, column=1, value='审计建议与风险评估').font = subtitle_font
ws5.cell(row=1, column=1).alignment = center_align

headers5 = ['编号', '风险/建议项', '风险等级', '关联版本', '详细说明与建议']
for i, h in enumerate(headers5, 1):
    ws5.cell(row=3, column=i, value=h)
style_header(ws5, 3, len(headers5))

risks = [
    ['R1', '政府采购制度缺失',
     '高风险',
     'V9（2026.01）',
     '政府采购管理制度（179行）自V9起被删除。学校政府采购活动需有明确的制度依据。\n\n建议：\n① 确认上级部门或统一采购平台是否已全面覆盖学校采购活动\n② 如已整合至一般采购管理制度，建议在制度汇编中加以说明\n③ 如确认存在制度空白，建议恢复或制定替代制度'],
    
    ['R2', '内部审计操作标准化降低',
     '中风险',
     'V10（2026.03）',
     '内部审计管理制度从97行精简至67行，删除了审计工作抽样审计工作程序和审计工作底稿详细要求章节。\n\n建议：\n① 评估现有审计实务是否受此精简影响\n② 如需要，可单独制定《内部审计操作规程》作为配套细则\n③ 确保审计证据收集和底稿编制仍有标准化流程'],
    
    ['R3', '校长办公会提交清单移除',
     '低风险',
     'V10（2026.03）',
     'V10删除了"校长办公会议提交党组织委员会讨论决定事项清单"。\n\n建议：确认党组织会议和校长办公会议事规则正文已完整、明确地覆盖所有需提交党组织委员会讨论决定的事项类型，避免漏项。'],
    
    ['R4', '财务制度过渡衔接',
     '关注项',
     'V8（2025.11）',
     '收支管理制度（164行）→财务管理制度（272行）的重构是积极变化，但需注意：\n① 旧制度中是否有未被新制度覆盖的特殊收支管理要求\n② 新旧制度过渡期的业务处理是否有明确指引\n③ 建议进行新旧制度逐条对照检查'],
    
    ['P1', '【积极评价】制度体系逐步健全',
     '—',
     '全程',
     '从3部分17项制度发展到4部分，制度建设趋势良好：\n✅ 2015-2017年国家内控文件纳入，合规性增强\n✅ 建设项目管理办法填补工程领域空白\n✅ 财务管理从操作型升级为体系型\n✅ 治理结构调整符合中小学校领导体制改革方向'],
    
    ['P2', '【积极评价】关键风险领域覆盖',
     '—',
     '全程',
     '以下高风险领域始终有制度覆盖：\n✅ 预算管理/资产管理/合同管理 — 全程存续\n✅ 采购管理（一般采购） — 全程存续\n✅ 费用报销/印章/公务卡 — 全程存续\n✅ 风险评估 — 全程存续（96行未变）'],
]

for i, row_data in enumerate(risks):
    for j, val in enumerate(row_data):
        ws5.cell(row=4+i, column=1+j, value=val)

style_range(ws5, 4, 4+len(risks)-1, len(headers5))

# Color risk levels
for i, row_data in enumerate(risks):
    risk_cell = ws5.cell(row=4+i, column=3)
    if risk_cell.value == '高风险':
        risk_cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        risk_cell.font = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
    elif risk_cell.value == '中风险':
        risk_cell.fill = PatternFill(start_color='FFD93D', end_color='FFD93D', fill_type='solid')
        risk_cell.font = Font(name='微软雅黑', bold=True, size=10)
    elif risk_cell.value == '低风险':
        risk_cell.fill = PatternFill(start_color='6BCB77', end_color='6BCB77', fill_type='solid')
        risk_cell.font = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
    elif risk_cell.value == '关注项':
        risk_cell.fill = PatternFill(start_color='4D96FF', end_color='4D96FF', fill_type='solid')
        risk_cell.font = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')

ws5.column_dimensions['A'].width = 6
ws5.column_dimensions['B'].width = 28
ws5.column_dimensions['C'].width = 10
ws5.column_dimensions['D'].width = 16
ws5.column_dimensions['E'].width = 75

for i in range(4):
    ws5.row_dimensions[4+i].height = 90
for i in range(4, 6):
    ws5.row_dimensions[4+i].height = 90

# ═══════════════════════ Sheet 6: 术语变更追溯 ═══════════════════════
ws6 = wb.create_sheet("术语变更追溯")

ws6.merge_cells('A1:E1')
ws6.cell(row=1, column=1, value='贯穿性术语变更追溯').font = subtitle_font
ws6.cell(row=1, column=1).alignment = center_align

headers6 = ['变更项', '旧术语（V1-V2）', '新术语（V3起）', '首次变更版本', '影响制度范围']
for i, h in enumerate(headers6, 1):
    ws6.cell(row=3, column=i, value=h)
style_header(ws6, 3, len(headers6))

term_changes = [
    ['决策机构名称', '党政联席会/党组织委员会', '校长办公会/党组织委员会', 'V3', '预算管理制度、资产管理制度、政府采购管理制度、财务管理制度（V8）等全部经济业务制度'],
    ['议事规则名称', '党组织会议和党政联席会（行政会）议事规则', '党组织会议和校长办公会（行政会）议事规则', 'V3', '第一部分议事决策规范'],
    ['提交决定清单', '党政联席会议提交党组织委员会讨论决定事项清单\n学校党组织委员会会议直接讨论决定事项清单', '校长办公会议提交党组织委员会讨论决定事项清单', 'V3', '第一部分议事决策规范'],
    ['教育体育局', '高新区教育文体局（V1）', '高新区教育体育局', '—', '三重一大制度等（上级主管部门名称更新）'],
    ['上级政策依据', '无第四部分', '增加3份国家级内控文件', 'V5', '新增第四部分政策文件'],
]

for i, row_data in enumerate(term_changes):
    for j, val in enumerate(row_data):
        ws6.cell(row=4+i, column=1+j, value=val)

style_range(ws6, 4, 4+len(term_changes)-1, len(headers6))

ws6.column_dimensions['A'].width = 18
ws6.column_dimensions['B'].width = 35
ws6.column_dimensions['C'].width = 35
ws6.column_dimensions['D'].width = 16
ws6.column_dimensions['E'].width = 60

for i in range(len(term_changes)):
    ws6.row_dimensions[4+i].height = 50

# ═══════════════════════ Sheet 7: 制度建设成熟度 ═══════════════════════
ws7 = wb.create_sheet("制度建设成熟度")

ws7.merge_cells('A1:G1')
ws7.cell(row=1, column=1, value='制度建设成熟度评估').font = subtitle_font
ws7.cell(row=1, column=1).alignment = center_align

headers7 = ['版本', '制度数量', '篇章数', '覆盖领域', '成熟度评分', '成熟度可视化', '评估']
for i, h in enumerate(headers7, 1):
    ws7.cell(row=3, column=i, value=h)
style_header(ws7, 3, len(headers7))

maturity = [
    ['V1', 17, 3, '议事决策/经济业务/行政运行', 40,
     '████░░░░░░░░', '基础框架：经济业务领域收支+采购+资产+预算已覆盖，缺建设管理和政策依据'],
    ['V2', 17, 3, '议事决策/经济业务/行政运行', 42,
     '████░░░░░░░░', '稳定运行期，微调'],
    ['V3', 17, 3, '议事决策/经济业务/行政运行', 50,
     '█████░░░░░░░', '治理结构调整：党政联席会→校长办公会，制度合规性提升'],
    ['V4', 18, 3, '议事决策/经济业务/行政运行', 58,
     '██████░░░░░░', '补齐建设管理短板'],
    ['V5', 21, 4, '议事决策/经济业务/行政运行/政策文件', 72,
     '███████░░░░░', '对标国家标准，合规性显著增强'],
    ['V6', 21, 4, '议事决策/经济业务/行政运行/政策文件', 72,
     '███████░░░░░', '稳定运行期'],
    ['V7', 21, 4, '议事决策/经济业务/行政运行/政策文件', 75,
     '████████░░░░', '收费管理透明化'],
    ['V8', 21, 4, '议事决策/经济业务/行政运行/政策文件', 88,
     '█████████░░░', '财务制度体系化重构，本周期成熟度最大提升'],
    ['V9', 20, 4, '议事决策/经济业务/行政运行/政策文件', 83,
     '████████░░░░', '采购制度删除，成熟度略有回落'],
    ['V10', 19, 4, '议事决策/经济业务/行政运行/政策文件', 80,
     '████████░░░░', '精简优化，审计操作标准化降低'],
]

for i, row_data in enumerate(maturity):
    for j, val in enumerate(row_data):
        ws7.cell(row=4+i, column=1+j, value=val)

style_range(ws7, 4, 4+len(maturity)-1, len(headers7))

# Color the score
for i, row_data in enumerate(maturity):
    score = row_data[4]
    cell = ws7.cell(row=4+i, column=5)
    if score >= 80:
        cell.fill = add_fill
    elif score >= 60:
        cell.fill = change_fill
    else:
        cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

ws7.column_dimensions['A'].width = 8
ws7.column_dimensions['B'].width = 10
ws7.column_dimensions['C'].width = 8
ws7.column_dimensions['D'].width = 42
ws7.column_dimensions['E'].width = 12
ws7.column_dimensions['F'].width = 18
ws7.column_dimensions['G'].width = 60

# ═══════════════════════ Sheet 8: 重大变化时间线 ═══════════════════════
ws8 = wb.create_sheet("重大变化时间线")

ws8.merge_cells('A1:E1')
ws8.cell(row=1, column=1, value='重大变化时间线').font = subtitle_font
ws8.cell(row=1, column=1).alignment = center_align

headers8 = ['时间节点', '触发版本', '变化性质', '核心事件', '制度建设意义']
for i, h in enumerate(headers8, 1):
    ws8.cell(row=3, column=i, value=h)
style_header(ws8, 3, len(headers8))

timeline = [
    ['2024年6月', 'V1', '🔵 起点', '初始版本建立', '3部分17项制度，构建基础内控框架'],
    ['', '', '', '— 稳定运行期 —', ''],
    ['2025年3月', 'V3', '🟡 治理改革', '「党政联席会」→「校长办公会」', '落实党组织领导的校长负责制，制度术语全面统一'],
    ['2025年4月', 'V4', '🟡 补缺', '新增「建设项目实施管理办法」', '填补工程建设领域制度空白，设三级审批体系'],
    ['2025年5月', 'V5', '🔴 对标国标', '新增第四部分：3份国家内控文件', '从自编制度提升为对标国家标准，制度合规性显著增强'],
    ['', '', '', '— 稳定运行+微调 —', ''],
    ['2025年11月', 'V8', '🔴 财务重构', '收支管理→财务管理+票据管理办法', '财务制度从操作型升级为体系型，覆盖全财务要素'],
    ['2026年1月', 'V9', '🟡 精简', '删除「政府采购管理制度」', '简化制度体系，但可能存在内控空白'],
    ['2026年3月', 'V10', '🟡 精简', '内部审计精简+提交清单移除', '去冗余提效率，但审计操作标准化有所降低'],
]

for i, row_data in enumerate(timeline):
    for j, val in enumerate(row_data):
        ws8.cell(row=4+i, column=1+j, value=val)

style_range(ws8, 4, 4+len(timeline)-1, len(headers8))

# Color by type
for i, row_data in enumerate(timeline):
    nature = row_data[2]
    cell = ws8.cell(row=4+i, column=3)
    if '🔴' in nature:
        cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    elif '🟡' in nature:
        cell.fill = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')

ws8.column_dimensions['A'].width = 14
ws8.column_dimensions['B'].width = 12
ws8.column_dimensions['C'].width = 14
ws8.column_dimensions['D'].width = 45
ws8.column_dimensions['E'].width = 55

for i in [2, 5]:
    ws8.row_dimensions[4+i].height = 25

# ═══════════════════════ 保存 ═══════════════════════
wb.save(output_path)
print(f"Excel report saved to: {output_path}")
print("Sheets:", wb.sheetnames)
