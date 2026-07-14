import os, sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

out = r"D:\openclaw-workspace\教科院内控分析"
output_path = os.path.join(out, "政府采购差异性分析.xlsx")

wb = Workbook()

# ═══════ 样式 ═══════
header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
title_font = Font(name='微软雅黑', bold=True, size=14, color='2F5496')
subtitle_font = Font(name='微软雅黑', bold=True, size=12, color='2F5496')
bold_font = Font(name='微软雅黑', bold=True, size=10)
normal_font = Font(name='微软雅黑', size=10)
red_font = Font(name='微软雅黑', size=10, color='CC0000', bold=True)
add_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
del_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
warn_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
change_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
header2_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
red_del_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))
wrap_align = Alignment(wrap_text=True, vertical='center', horizontal='left')
center_align = Alignment(wrap_text=True, vertical='center', horizontal='center')

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_range(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = wrap_align
            cell.font = normal_font

# ═══════ Sheet 1: 政府采购存续状态 ═══════
ws1 = wb.active
ws1.title = "存续状态总览"

ws1.merge_cells('A1:G1')
ws1.cell(row=1, column=1, value='政府采购管理制度 — 10版本存续状态总览').font = title_font
ws1.cell(row=1, column=1).alignment = center_align

ws1.merge_cells('A2:G2')
ws1.cell(row=2, column=1, value='结论：政府采购管理制度在V1-V8期间内容完全一致（179行/12章/8992字符），V9起整体删除').font = Font(name='微软雅黑', size=10, color='666666')
ws1.cell(row=2, column=1).alignment = center_align

headers1 = ['版本', '时间范围', '制度状态', '行数', '字符数', '章节数', '变化说明']
for i, h in enumerate(headers1, 1):
    ws1.cell(row=4, column=i, value=h)
style_header(ws1, 4, len(headers1))

proc_status = [
    ['V1', '2024.06.14-11.18', '✅ 存在（初始版）', 179, 8992, 12, '初始基准版本'],
    ['V2', '2024.11.08-2025.03.17', '✅ 存在', 179, 8992, 12, '与V1完全相同'],
    ['V3', '2025.03.17-04.11', '✅ 存在', 179, 8992, 12, '仅1处术语修改：「党政联席会」→「校长办公会」'],
    ['V4', '2025.04.11-05.27', '✅ 存在', 179, 8992, 12, '与V3完全相同'],
    ['V5', '2025.05.27-07.11', '✅ 存在', 179, 8992, 12, '与V3完全相同'],
    ['V6', '2025.07.11-10.13', '✅ 存在', 179, 8992, 12, '与V3完全相同'],
    ['V7', '2025.10.13-11.05', '✅ 存在', 179, 8992, 12, '与V3完全相同'],
    ['V8', '2025.11.05-2026.01.23', '✅ 存在（最后一版）', 179, 8992, 12, '与V3完全相同'],
    ['V9', '2026.01.23-03.13', '❌ 已删除', 0, 0, 0, '⚠️ 整体删除！一般采购未扩容吸收'],
    ['V10', '2026.03.13-至今', '❌ 已删除', 0, 0, 0, '持续缺失中'],
]

for i, row_data in enumerate(proc_status):
    for j, val in enumerate(row_data):
        ws1.cell(row=5+i, column=1+j, value=val)
style_range(ws1, 5, 14, len(headers1))

# Color rows
for i, row_data in enumerate(proc_status):
    if '已删除' in str(row_data[2]):
        for c in range(1, len(headers1)+1):
            ws1.cell(row=5+i, column=c).fill = del_fill
        ws1.cell(row=5+i, column=3).font = red_font
        ws1.cell(row=5+i, column=7).font = red_font

ws1.column_dimensions['A'].width = 8
ws1.column_dimensions['B'].width = 26
ws1.column_dimensions['C'].width = 24
ws1.column_dimensions['D'].width = 10
ws1.column_dimensions['E'].width = 10
ws1.column_dimensions['F'].width = 10
ws1.column_dimensions['G'].width = 42

# ═══════ Sheet 2: 制度章节结构 ═══════
ws2 = wb.create_sheet("制度章节结构")

ws2.merge_cells('A1:L1')
ws2.cell(row=1, column=1, value='政府采购管理制度 — 12章完整结构（V1-V8）').font = subtitle_font
ws2.cell(row=1, column=1).alignment = center_align

ws2.merge_cells('A2:L2')
ws2.cell(row=2, column=1, value='说明：V1-V8期间12章结构完整不变，仅V3修改决策机构术语1处。V9起全部删除。').font = Font(name='微软雅黑', size=10, color='666666')
ws2.cell(row=2, column=1).alignment = center_align

headers2 = ['章节序号', '章节名称', 'V1', 'V2', 'V3', 'V4', 'V5', 'V6', 'V7', 'V8', 'V9', 'V10']
for i, h in enumerate(headers2, 1):
    ws2.cell(row=4, column=i, value=h)
style_header(ws2, 4, len(headers2))

chapters = [
    ['第一章', '总 则'],
    ['第二章', '管理机构及职责'],
    ['第三章', '政府采购预算管理'],
    ['第四章', '政府采购意向公开管理'],
    ['第五章', '政府采购需求管理及采购实施计划编制'],
    ['第六章', '政府采购项目的实施'],
    ['第七章', '政府购买服务管理'],
    ['第八章', '进口产品采购管理'],
    ['第九章', '政府采购项目合同、档案及验收管理'],
    ['第十章', '询问、质疑处理'],
    ['第十一章', '政府采购信息公开'],
    ['第十二章', '附 则'],
]

for i, (num, name) in enumerate(chapters):
    ws2.cell(row=5+i, column=1, value=num)
    ws2.cell(row=5+i, column=2, value=name)
    for j in range(8):  # V1-V8
        ws2.cell(row=5+i, column=3+j, value='✅')
        ws2.cell(row=5+i, column=3+j).alignment = center_align
        ws2.cell(row=5+i, column=3+j).fill = add_fill
    for j in range(8, 10):  # V9-V10
        ws2.cell(row=5+i, column=3+j, value='❌')
        ws2.cell(row=5+i, column=3+j).alignment = center_align
        ws2.cell(row=5+i, column=3+j).fill = del_fill

# Add a note row for V3
ws2.merge_cells('A18:B18')
ws2.cell(row=18, column=1, value='V3唯一修改').font = bold_font
ws2.cell(row=18, column=1).fill = change_fill
ws2.cell(row=18, column=1).alignment = center_align

ws2.merge_cells('C18:L18')
ws2.cell(row=18, column=3, value='第二章"管理机构及职责"中：决策机构表述从"党政联席会/党组织委员会" → "校长办公会/党组织委员会"（仅1处，2个字差异）').font = Font(name='微软雅黑', size=10)
ws2.cell(row=18, column=3).fill = change_fill

# Add deletion warning
ws2.merge_cells('A20:L20')
ws2.cell(row=20, column=1, value='⚠️ V9起12章全部删除，一般采购管理制度（147行）未扩容吸收，可能依赖上级统一采购平台承接').font = red_font
ws2.cell(row=20, column=1).fill = warn_fill
ws2.cell(row=20, column=1).alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')

style_range(ws2, 5, 16, len(headers2))

ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 40
for c in 'CDEFGHIJKL':
    ws2.column_dimensions[c].width = 6

# ═══════ Sheet 3: 制度内容详解 ═══════
ws3 = wb.create_sheet("制度内容详解")

ws3.merge_cells('A1:D1')
ws3.cell(row=1, column=1, value='政府采购管理制度 — 12章内容概要（V1-V8全文）').font = subtitle_font
ws3.cell(row=1, column=1).alignment = center_align

headers3 = ['章节', '章节名称', '核心内容', '内控要点']
for i, h in enumerate(headers3, 1):
    ws3.cell(row=3, column=i, value=h)
style_header(ws3, 3, len(headers3))

content_detail = [
    ['第一章', '总 则',
     '法律依据：《政府采购法》《政府采购法实施条例》\n适用范围：使用财政性资金采购集中采购目录以内或限额标准以上的货物/工程/服务\n采购目录：执行《四川省政府集中采购目录和采购限额标准》\n原则：公开透明、公平竞争、公正、诚实信用\n内控目标：分事行权、分岗设权、分级授权',
     '目录外/限额以下项目参照《一般采购管理制度》执行\n落实主体责任、明确归口管理部门\n科学设置岗位、强化流程控制'],
    ['第二章', '管理机构及职责',
     '（一）决策机构：党政联席会→校长办公会/党组织委员会（V3修改）\n审定内控制度、研究决定重大采购事项、审定预算和计划\n（二）归口管理部门：总务处\n拟定内控制度、汇总采购预算、确定采购方式和组织形式\n（三）财务部门\n审核采购预算、复核支付、合同备案\n（四）业务部门\n提出采购需求、参与采购文件编制、履约验收',
     '四权分立：决策/归口/财务/业务\n不相容岗位分离\n重大事项集体研究+合法性审查+内部会签'],
    ['第三章', '政府采购预算管理',
     '采购预算编制：依据年度工作计划和资产配置标准\n无预算不采购：未列入预算项目不得采购\n预算调整程序\n政府采购预算与部门预算同步编制',
     '预算刚性约束\n预算-计划-采购三段联动'],
    ['第四章', '政府采购意向公开管理',
     '公开时间：采购活动开始前30日\n公开内容：项目名称、需求概况、预算金额、预计采购时间\n公开渠道：四川政府采购网\n可不公开情形：紧急采购、涉密采购',
     '透明化要求\n社会监督机制'],
    ['第五章', '政府采购需求管理及采购实施计划编制',
     '需求调查：采购前应进行充分市场调查\n需求编制：明确技术要求和商务要求\n采购实施计划：采购方式、评审方法、合同类型等\n需求审查机制',
     '需求-计划分离\n防止倾向性/排他性需求\n重大项目需求论证'],
    ['第六章', '政府采购项目的实施',
     '采购方式选择：公开招标/邀请招标/竞争性谈判/竞争性磋商/询价/单一来源\n评审专家抽取\n采购文件编制与发布\n开标评标程序\n中标成交结果确认',
     '法定采购方式适用条件\n专家独立评审\n程序合规性'],
    ['第七章', '政府购买服务管理',
     '适用范围：公共服务项目\n购买主体与承接主体资格\n购买程序：预算→计划→采购→合同→验收\n绩效评价',
     '服务类采购特殊性\n承接主体资质审查\n服务绩效管理'],
    ['第八章', '进口产品采购管理',
     '进口产品论证：专家论证+行业主管部门意见\n报批程序：财政部门审批\n采购方式：通常采用公开招标\n特殊情况处理',
     '进口产品严控\n论证-审批-采购三级管理'],
    ['第九章', '政府采购项目合同、档案及验收管理',
     '合同签订：中标通知书发出30日内\n合同内容要求\n履约验收：成立验收小组、出具验收报告\n档案管理：采购文件保存15年\n档案内容：采购预算/招标文件/投标文件/评审报告/合同/验收报告',
     '合同-验收-档案全链条\n验收小组独立运作\n15年档案保存期限'],
    ['第十章', '询问、质疑处理',
     '询问答复：3个工作日内\n质疑提出：知道权益受损之日起7个工作日内\n质疑答复：收到后7个工作日内\n投诉途径：向财政部门投诉',
     '供应商权益保障\n法定时限要求\n质疑答复程序规范'],
    ['第十一章', '政府采购信息公开',
     '公开内容：采购公告/采购文件/采购预算/中标成交结果/采购合同/验收结果\n公开渠道：四川政府采购网\n公开时限要求\n涉密信息处理',
     '全程公开透明\n法定公开时限\n信息一致性要求'],
    ['第十二章', '附 则',
     '解释权归属\n施行日期\n与国家法规冲突时以上位法为准\n未尽事宜按国家规定执行',
     '制度效力保障'],
]

for i, row_data in enumerate(content_detail):
    for j, val in enumerate(row_data):
        ws3.cell(row=4+i, column=1+j, value=val)

style_range(ws3, 4, 4+len(content_detail)-1, len(headers3))

# Zebra striping
for i in range(len(content_detail)):
    if i % 2 == 0:
        for c in range(1, len(headers3)+1):
            ws3.cell(row=4+i, column=c).fill = PatternFill(start_color='F5F8FC', end_color='F5F8FC', fill_type='solid')

ws3.column_dimensions['A'].width = 10
ws3.column_dimensions['B'].width = 28
ws3.column_dimensions['C'].width = 65
ws3.column_dimensions['D'].width = 45

for i in range(len(content_detail)):
    ws3.row_dimensions[4+i].height = 120

# ═══════ Sheet 4: V8→V9删除分析 ═══════
ws4 = wb.create_sheet("V8到V9删除分析")

ws4.merge_cells('A1:E1')
ws4.cell(row=1, column=1, value='V8→V9：政府采购制度删除影响分析').font = subtitle_font
ws4.cell(row=1, column=1).alignment = center_align

headers4 = ['分析维度', 'V8（删除前）', 'V9（删除后）', '变化', '风险评估']
for i, h in enumerate(headers4, 1):
    ws4.cell(row=3, column=i, value=h)
style_header(ws4, 3, len(headers4))

impact_data = [
    ['政府采购管理制度', '存在（179行/12章/8992字符）', '已删除', '完全移除', '🔴 高风险'],
    ['一般采购管理制度', '147行', '147行', '+0行（未扩容）', '🔴 高风险：未吸收政府采购内容'],
    ['采购预算管理', '第三章专章规定', '仅预算管理制度统一规定', '政府采购预算管理专章消失', '🟡 采购预算与部门预算衔接可能弱化'],
    ['政府采购意向公开', '第四章专章（30日前公开）', '无独立规定', '意向公开制度依据丧失', '🟡 合规风险：可能违反财政部意向公开要求'],
    ['采购需求管理', '第五章专章（需求调查/审查）', '一般采购制度有需求管理', '需求管理标准可能降低', '🟡 倾向性/排他性需求风险增加'],
    ['政府购买服务', '第七章专章', '无独立规定', '服务类采购特殊性无制度覆盖', '🟡 服务绩效管理缺失'],
    ['进口产品采购', '第八章专章（论证+审批）', '无独立规定', '进口产品管控失去校内制度依据', '🟡 如涉及进口采购则存在重大风险'],
    ['合同/档案/验收', '第九章专章（15年存档）', '一般采购制度有合同管理', '政府采购合同的特殊要求可能遗漏', '🟢 一般采购制度部分覆盖'],
    ['询问/质疑处理', '第十章专章（法定时限）', '无独立规定', '供应商质疑处理缺乏制度指引', '🟡 投诉风险增加'],
    ['采购信息公开', '第十一章专章（全程公开）', '无独立规定', '公开透明度可能降低', '🟡 合规风险'],
]

for i, row_data in enumerate(impact_data):
    for j, val in enumerate(row_data):
        ws4.cell(row=4+i, column=1+j, value=val)

style_range(ws4, 4, 4+len(impact_data)-1, len(headers4))

# Color risk
for i, row_data in enumerate(impact_data):
    risk = row_data[4]
    cell = ws4.cell(row=4+i, column=5)
    if '🔴' in risk:
        cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        cell.font = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
    elif '🟡' in risk:
        cell.fill = warn_fill
        cell.font = Font(name='微软雅黑', bold=True, size=10)
    elif '🟢' in risk:
        cell.fill = add_fill
        cell.font = Font(name='微软雅黑', bold=True, size=10)
    
    # Color V9 column
    v9_cell = ws4.cell(row=4+i, column=3)
    if '删除' in str(v9_cell.value) or '无' in str(v9_cell.value):
        v9_cell.fill = del_fill

ws4.column_dimensions['A'].width = 22
ws4.column_dimensions['B'].width = 38
ws4.column_dimensions['C'].width = 38
ws4.column_dimensions['D'].width = 28
ws4.column_dimensions['E'].width = 42

for i in range(len(impact_data)):
    ws4.row_dimensions[4+i].height = 50

# Add summary section
ws4.merge_cells('A16:E16')
ws4.cell(row=16, column=1, value='综合结论').font = subtitle_font

ws4.merge_cells('A17:E17')
ws4.cell(row=17, column=1, value='政府采购管理制度在V9被整体删除后，一般采购管理制度（147行）并未扩容吸收其内容。\n目前校内制度体系在政府采购领域存在空白，可能依赖：①上级部门统一采购平台；②财政部门直接监管；③四川省/成都高新区集中采购制度。\n建议：确认学校政府采购活动的实际制度依据，如果存在空白则应恢复或制定替代制度。').font = Font(name='微软雅黑', size=11)
ws4.cell(row=17, column=1).alignment = Alignment(wrap_text=True, vertical='top')
ws4.row_dimensions[17].height = 90

# ═══════ Sheet 5: 政府采购vs一般采购对比 ═══════
ws5 = wb.create_sheet("政府采购vs一般采购")

ws5.merge_cells('A1:F1')
ws5.cell(row=1, column=1, value='政府采购管理制度 vs 一般采购管理制度 — 制度边界对比').font = subtitle_font
ws5.cell(row=1, column=1).alignment = center_align

ws5.merge_cells('A2:F2')
ws5.cell(row=2, column=1, value='二者是互补关系：政府采购管"目录内+限额以上"，一般采购管"目录外+限额以下"。删除政府采购导致制度体系断裂。').font = Font(name='微软雅黑', size=10, color='666666')
ws5.cell(row=2, column=1).alignment = center_align

headers5 = ['对比维度', '政府采购管理制度（已删除）', '一般采购管理制度（现行）', '制度关系']
for i, h in enumerate(headers5, 1):
    ws5.cell(row=4, column=i, value=h)
style_header(ws5, 4, len(headers5))

comparison = [
    ['适用范围', '集中采购目录以内\n或采购限额标准以上\n（货物/服务50万+，工程100万+）', '集中采购目录以外\n且采购限额标准以下\n（货物/服务≤50万，工程≤100万）', '互补不重叠\n一个管"大额"\n一个管"小额"'],
    ['法律依据', '《政府采购法》\n《政府采购法实施条例》', '《政府采购法》\n《高新区财政金融局通知》', '上位法相同\n但管理要求不同'],
    ['采购方式', '公开招标/邀请招标/竞争性谈判\n竞争性磋商/询价/单一来源\n（法定六种方式）', '自主采购方式\n比选/询价/直接采购等\n（校内自定流程）', '政府采购方式法定\n一般采购方式自定'],
    ['审批层级', '需按规定报财政部门审批\n进口产品需专家论证+审批', '校内审批为主\n5万以下部门决定\n5万以上校级决定', '外部审批 vs 内部审批'],
    ['信息公开', '四川政府采购网\n全程公开（公告/文件/结果/合同）', '校内公开\n无外部公开强制要求', '公开程度差异显著'],
    ['供应商质疑', '法定时限：7日内提出\n7日内答复\n可向财政部门投诉', '校内处理\n无法定投诉途径', '救济渠道差异'],
    ['验收管理', '成立验收小组\n出具验收报告\n15年存档', '一般验收流程\n存档期限较短', '验收标准不同'],
    ['制度篇幅', '179行/12章/8992字符\n（已删除）', '147行（一直存在）', '删除后无替代'],
]

for i, row_data in enumerate(comparison):
    for j, val in enumerate(row_data):
        ws5.cell(row=5+i, column=1+j, value=val)

style_range(ws5, 5, 5+len(comparison)-1, len(headers5))

# Delete color for procurement column
for i in range(len(comparison)):
    ws5.cell(row=5+i, column=2).fill = del_fill

ws5.column_dimensions['A'].width = 16
ws5.column_dimensions['B'].width = 38
ws5.column_dimensions['C'].width = 38
ws5.column_dimensions['D'].width = 28

for i in range(len(comparison)):
    ws5.row_dimensions[5+i].height = 75

# ═══════ Sheet 6: 审计建议 ═══════
ws6 = wb.create_sheet("审计建议与风险提示")

ws6.merge_cells('A1:D1')
ws6.cell(row=1, column=1, value='政府采购制度 — 审计建议与风险提示').font = subtitle_font
ws6.cell(row=1, column=1).alignment = center_align

headers6 = ['编号', '风险/建议项', '风险等级', '详细说明与建议措施']
for i, h in enumerate(headers6, 1):
    ws6.cell(row=3, column=i, value=h)
style_header(ws6, 3, len(headers6))

audit_items = [
    ['R1', '政府采购制度整体缺失',
     '🔴 高风险',
     '发现：政府采购管理制度（179行/12章）自V9（2026.01）起被整体删除，至今未恢复。一般采购管理制度（147行）行数未变，未吸收政府采购内容。\n\n建议措施：\n1. 立即核实学校当前政府采购活动的实际制度依据（是否依赖上级统一平台、财政部门直接监管或区域集中采购）\n2. 如确认存在制度空白：恢复政府采购管理制度，或制定替代性的《政府采购操作规程》\n3. 如上级已全覆盖：在制度汇编中增加说明性条款，明确政府采购活动遵循的外部制度依据'],
    
    ['R2', '采购预算管理专章消失',
     '🟡 中风险',
     '发现：政府采购预算管理专章（第三章）随制度一并删除。该章规定了预算编制、无预算不采购、预算调整等内容。\n\n建议措施：\n1. 检查预算管理制度中是否已充分覆盖政府采购预算的特殊要求\n2. 确认"采购预算→采购计划→采购实施"的三段联动是否仍有制度保障'],
    
    ['R3', '采购信息公开合规风险',
     '🟡 中风险',
     '发现：政府采购意向公开（第四章，要求采购前30日公开）和采购信息公开（第十一章，要求全程公开于四川政府采购网）的制度依据随制度删除而丧失。\n\n建议措施：\n1. 确认学校是否仍在实际执行信息公开要求\n2. 即使依赖上级平台，校内也应保留信息公开的内部操作规范'],
    
    ['R4', '供应商权益保障机制缺失',
     '🟡 中风险',
     '发现：询问/质疑处理专章（第十章，规定7日答复等法定时限）随制度删除。\n\n建议措施：\n1. 建立供应商询问/质疑处理的操作指引\n2. 明确责任部门和答复流程\n3. 避免因处理不当引发向财政部门投诉的风险'],
    
    ['R5', '进口产品采购管控空白',
     '🟡 中风险',
     '发现：进口产品采购专章（第八章，要求专家论证+主管部门意见+财政审批）随制度删除。\n\n建议措施：\n1. 排查学校是否涉及进口产品采购\n2. 如涉及：必须建立专项管理制度\n3. 如不涉及：可在制度汇编中说明"暂不适用"'],
    
    ['R6', '政府购买服务绩效管理缺失',
     '🟢 低风险',
     '发现：政府购买服务管理专章（第七章）随制度删除，服务类采购的特殊管理要求失去制度覆盖。\n\n建议措施：\n1. 评估学校政府购买服务业务量\n2. 如业务量小：可在一般采购制度中增加服务类采购的特殊条款\n3. 如业务量大：建议单独制定管理制度'],
    
    ['P1', '采购内控框架历史健全',
     '✅ 积极评价',
     'V1-V8期间政府采购制度始终是制度汇编中篇幅最大的制度之一（179行/8992字符），内容涵盖政府采购全生命周期：\n- 预算编制 → 需求管理 → 意向公开 → 项目实施 → 政府购买服务 → 进口产品 → 合同签订 → 履约验收 → 档案管理 → 询问质疑 → 信息公开\n\n12章结构完整、权责清晰，体现了较高的制度建设水平。'],
]

for i, row_data in enumerate(audit_items):
    for j, val in enumerate(row_data):
        ws6.cell(row=4+i, column=1+j, value=val)

style_range(ws6, 4, 4+len(audit_items)-1, len(headers6))

# Color
for i, row_data in enumerate(audit_items):
    risk = row_data[2]
    cell = ws6.cell(row=4+i, column=3)
    if '🔴' in risk:
        cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        cell.font = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
    elif '🟡' in risk:
        cell.fill = warn_fill
        cell.font = Font(name='微软雅黑', bold=True, size=10)
    elif '🟢' in risk:
        cell.fill = add_fill
        cell.font = Font(name='微软雅黑', bold=True, size=10)
    elif '✅' in risk:
        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

ws6.column_dimensions['A'].width = 6
ws6.column_dimensions['B'].width = 28
ws6.column_dimensions['C'].width = 14
ws6.column_dimensions['D'].width = 90

for i in range(len(audit_items)):
    if i < 6:
        ws6.row_dimensions[4+i].height = 140
    else:
        ws6.row_dimensions[4+i].height = 130

# ═══════ 保存 ═══════
wb.save(output_path)
print(f"Excel saved to: {output_path}")
print("Sheets:", wb.sheetnames)
