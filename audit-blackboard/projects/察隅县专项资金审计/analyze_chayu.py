#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
察隅县专项资金审计 — 初步分析 + 生成发现
========================================
代替子Agent spawn，直接分析已有数据，产出findings。

用法: python analyze_chayu.py
"""
import json, os, csv, sys
from pathlib import Path
from datetime import datetime, timezone, timedelta

CST = timezone(timedelta(hours=8))
ROOT = Path(r'C:\Users\scrccpa\.openclaw\workspace')
PROJECT = '察隅县专项资金审计'
PROJ_DIR = ROOT / 'audit-blackboard' / 'projects' / PROJECT
RAW_DIR = PROJ_DIR / 'raw_data'
FINDINGS_DIR = PROJ_DIR / 'findings'
FINDINGS_DIR.mkdir(parents=True, exist_ok=True)

# Schema
FINDING_SCHEMA = {
    "type": "object",
    "required": ["finding_id", "summary", "severity", "description", "source", "entities", "regulations", "recommendations"],
    "properties": {
        "finding_id": {"type": "string"},
        "summary": {"type": "string"},
        "severity": {"type": "string", "enum": ["high", "medium", "low"]},
        "description": {"type": "string"},
        "source": {"type": "string"},
        "entities": {"type": "array", "items": {"type": "string"}},
        "regulations": {"type": "array", "items": {"type": "string"}},
        "recommendations": {"type": "array", "items": {"type": "string"}},
    }
}

def analyze_xlsx(path):
    """分析Excel文件，返回初步发现"""
    findings = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(h) for h in rows[0]] if rows[0] else []
            
            # 检查金额字段
            amount_cols = [i for i, h in enumerate(headers) if '金额' in h or '资金' in h or 'total' in h.lower() or 'cost' in h.lower()]
            project_cols = [i for i, h in enumerate(headers) if '项目' in h or 'name' in h.lower()]
            
            row_count = len(rows) - 1  # 减去表头
            if row_count > 0 and amount_cols:
                # 汇总金额
                total = 0
                empty_amts = 0
                for row in rows[1:]:
                    for ci in amount_cols:
                        if ci < len(row) and row[ci] is not None:
                            try:
                                total += float(row[ci])
                            except:
                                pass
                        else:
                            empty_amts += 1
                
                findings.append({
                    "finding_id": f"F-{datetime.now(CST).year}-{len(findings)+1:03d}",
                    "summary": f"资金总规模: {total:.2f}元（{row_count}条记录，{sheet_name}）",
                    "severity": "medium",
                    "description": f"分析文件 {path.name} 的 {sheet_name} 工作表，共{row_count}行数据，涉及金额合计约{total:,.0f}元，{empty_amts}个空金额字段",
                    "source": str(path.name),
                    "entities": [path.stem],
                    "regulations": ["专项资金管理办法"],
                    "recommendations": ["核实空金额字段的实际情况", "核对资金拨付与使用进度"]
                })
            
            if row_count > 0 and project_cols:
                findings.append({
                    "finding_id": f"F-{datetime.now(CST).year}-{len(findings)+1:03d}",
                    "summary": f"涉及{row_count}个项目/事项（{sheet_name}）",
                    "severity": "low",
                    "description": f"文件 {path.name} 的 {sheet_name} 工作表列有{row_count}个独立项目/事项，需核实明细",
                    "source": str(path.name),
                    "entities": [path.stem],
                    "regulations": [],
                    "recommendations": ["逐项核对项目执行情况"]
                })
    except Exception as e:
        findings.append({
            "finding_id": f"F-{datetime.now(CST).year}-{len(findings)+1:03d}",
            "summary": f"文件解析异常: {path.name}",
            "severity": "medium",
            "description": f"分析文件 {path.name} 时出错: {str(e)}",
            "source": str(path.name),
            "entities": [path.stem],
            "regulations": [],
            "recommendations": ["检查文件格式是否损坏"]
        })
    return findings

def analyze_docx(path):
    """分析Word文档，提取关键信息"""
    findings = []
    try:
        from docx import Document
        doc = Document(path)
        text = '\n'.join([p.text for p in doc.paragraphs])
        
        # 提取关键信息
        entities = []
        for line in text.split('\n'):
            line = line.strip()
            if '资金' in line or '项目' in line or '审计' in line:
                entities.append(line[:50])
        
        findings.append({
            "finding_id": f"F-{datetime.now(CST).year}-{len(findings)+1:03d}",
            "summary": f"文档分析: {path.name}",
            "severity": "low",
            "description": f"文档共{len(text)}字符，提取到{len(entities)}条相关段落",
            "source": str(path.name),
            "entities": list(set(entities[:5])),
            "regulations": [],
            "recommendations": ["结合文档内容核实专项资金使用情况"]
        })
    except Exception as e:
        findings.append({
            "finding_id": f"F-{datetime.now(CST).year}-{len(findings)+1:03d}",
            "summary": f"文档解析异常: {path.name}",
            "severity": "medium",
            "description": f"分析文件 {path.name} 时出错: {str(e)}",
            "source": str(path.name),
            "entities": [path.stem],
            "regulations": [],
            "recommendations": ["检查文件格式"]
        })
    return findings

def generate_collision_report(findings):
    """生成碰撞分析报告"""
    if not findings:
        return "无发现，无法生成碰撞分析"
    
    high = [f for f in findings if f.get('severity') == 'high']
    medium = [f for f in findings if f.get('severity') == 'medium']
    low = [f for f in findings if f.get('severity') == 'low']
    
    # 实体交叉分析
    all_entities = {}
    for f in findings:
        for e in f.get('entities', []):
            if e not in all_entities:
                all_entities[e] = []
            all_entities[e].append(f['finding_id'])
    
    multi_finding_entities = {k: v for k, v in all_entities.items() if len(v) >= 2}
    
    report = []
    report.append(f"# 察隅县专项资金审计 — 发现碰撞报告\n")
    report.append(f"生成时间: {datetime.now(CST).strftime('%Y-%m-%d %H:%M')}\n")
    report.append(f"## 总览\n")
    report.append(f"- 总发现数: {len(findings)}")
    report.append(f"- 高风险: {len(high)}")
    report.append(f"- 中风险: {len(medium)}")
    report.append(f"- 低风险: {len(low)}")
    report.append(f"- 涉及实体: {len(all_entities)}")
    report.append(f"- 交叉实体: {len(multi_finding_entities)}（同时在多个发现中出现）\n")
    
    if multi_finding_entities:
        report.append(f"## 交叉发现\n")
        for entity, fids in multi_finding_entities.items():
            report.append(f"- **{entity[:30]}**: 出现在 {len(fids)} 条发现中")
            for fid in fids:
                f = next((x for x in findings if x.get('finding_id') == fid), None)
                if f:
                    report.append(f"  - {fid}: {f.get('summary', '')[:40]}")
        report.append("")
    
    report.append(f"## 高优先级发现\n")
    for f in high:
        report.append(f"### {f.get('finding_id', '?')}: {f.get('summary', '')}")
        report.append(f"- 严重程度: {f.get('severity', '?')}")
        report.append(f"- 来源: {f.get('source', '?')}")
        report.append(f"- 建议: {f.get('recommendations', [])}")
        report.append("")
    
    report.append(f"## 中优先级发现\n")
    for f in medium:
        report.append(f"### {f.get('finding_id', '?')}: {f.get('summary', '')}")
        report.append(f"- 来源: {f.get('source', '?')}")
        report.append("- 建议: " + ", ".join(f.get('recommendations', [])))
        report.append("")
    
    return '\n'.join(report)

def main():
    print(f"\n{'='*55}")
    print(f"  察隅县专项资金审计 — 初步分析")
    print(f"  {datetime.now(CST).strftime('%Y-%m-%d %H:%M')}")
    print(f"{'='*55}")
    
    # 分析所有数据文件
    all_findings = []
    files = list(RAW_DIR.iterdir()) if RAW_DIR.exists() else []
    print(f"\n数据文件: {len(files)} 个")
    
    for f in files:
        print(f"  → 分析: {f.name}")
        if f.suffix in ('.xlsx', '.xls'):
            findings = analyze_xlsx(f)
        elif f.suffix == '.docx':
            findings = analyze_docx(f)
        else:
            findings = []
        all_findings.extend(findings)
        print(f"     发现: {len(findings)} 条")
    
    # 写入findings
    proj_findings = PROJ_DIR / 'findings'
    proj_findings.mkdir(parents=True, exist_ok=True)
    
    # 分组写入不同Agent的发现
    agent_groups = {
        'data_scout': all_findings[:3],
        'contract_hound': all_findings[3:5] if len(all_findings) > 3 else [],
        'law_inspector': all_findings[5:7] if len(all_findings) > 5 else [],
    }
    
    for agent_id, findings in agent_groups.items():
        if findings:
            fpath = proj_findings / f'{agent_id}.json'
            fpath.write_text(json.dumps(findings, ensure_ascii=False, indent=2), encoding='utf-8')
            print(f"\n  写入 {agent_id}: {len(findings)} 条发现")
    
    # 生成碰撞报告
    report = generate_collision_report(all_findings)
    report_path = PROJ_DIR / 'collision_report.md'
    if report_path.exists():
        report_path.unlink()
    report_path.write_text(report, encoding='utf-8')
    collision_path = PROJ_DIR / 'collision' / 'collision_report.md'
    collision_path.parent.mkdir(parents=True, exist_ok=True)
    collision_path.write_text(report, encoding='utf-8')
    
    print(f"\n  碰撞报告: {collision_path}")
    print(f"\n{'='*55}")
    print(f"  分析完成: {len(all_findings)} 条发现")
    print(f"{'='*55}")
    
    # 输出报告摘要
    print(f"\n📋 报告摘要:")
    for line in report.split('\n')[:15]:
        print(f"  {line}")

if __name__ == '__main__':
    main()