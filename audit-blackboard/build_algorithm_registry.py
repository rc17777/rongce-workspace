# -*- coding: utf-8 -*-
"""
build_algorithm_registry.py — 从《政府审计算法资产库_v4.xlsx》重建 algorithm_registry.json

用法:
    python build_algorithm_registry.py [xlsx路径]

输出:
    audit-blackboard/algorithm_registry.json （UTF-8，ensure_ascii=False）

说明:
    - 算法40要素取自"☆算法详细卡片"（42个字段）
    - 优先级/成熟度/风险机制/数据就绪度取自"☆算法资产库总览"
    - assigned_agents 为人工审核映射（见 ALGORITHM_INTEGRATION.md 映射矩阵）
"""
import sys
import os
import re
import json

sys.stdout.reconfigure(encoding="utf-8")

BASE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_XLSX = r"C:\Users\scrccpa\Desktop\算法\政府审计算法资产库_v4.xlsx"
OUT_JSON = os.path.join(BASE, "algorithm_registry.json")

# ── 人工审核的 Agent 映射（40算法 → 22Agent体系）──────────────────────────
# Agent ID = agent_specs/ 下的文件名（不含 .json）
ASSIGNED_AGENTS = {
    "PERF-OUTLIER-001":   ["data_scout", "performance_evaluator"],
    "PROC-CONCEN-001":    ["bid_hunter", "data_scout"],
    "BID-PATTERN-005":    ["bid_hunter"],
    "FUND-FRAUD-001":     ["data_scout", "performance_evaluator"],
    "HR-RF-001":          ["data_scout"],
    "HR-RF-002":          ["data_scout"],
    "REV-PREDICT-001":    ["data_scout", "budget_estimator"],
    "ENG-SAMPLE-001":     ["settlement_auditor"],
    "ENG-SCORE-001":      ["settlement_auditor", "fiscal_reviewer"],
    "CHK-RECON-001":      ["data_scout", "review_sentinel"],
    "ASSET-MATCH-001":    ["data_scout"],
    "SUPV-ANOMALY-001":   ["data_scout"],
    "RULE-MATCH-001":     ["law_inspector", "review_sentinel"],
    "SUPV-POCKET-001":    ["data_scout", "contract_hound"],
    "SUPV-TRAVEL-001":    ["data_scout"],
    "FUND-SIPHON-001":    ["data_scout", "workpaper_crafter"],
    "CHK-LOSS-001":       ["data_scout", "contract_hound"],
    "PROC-FAKE-001":      ["bid_hunter", "law_inspector"],
    "PROC-RELATED-001":   ["bid_hunter"],
    "CHK-RD-001":         ["data_scout", "law_inspector"],
    "ENG-FINAL-001":      ["settlement_auditor", "review_sentinel", "fiscal_reviewer"],
    "SUPV-WARNING-001":   ["data_scout"],
    "ENV-CHECKLIST-001":  ["law_inspector", "performance_evaluator"],
    "SOE-MIDMAN-001":     ["contract_hound", "data_scout"],
    "AGR-INSFAKE-001":    ["data_scout"],
    "FIN-SHELL-001":      ["data_scout"],
    "FIN-INSFAKE-001":    ["data_scout"],
    "ENG-RATIO-001":      ["settlement_auditor", "budget_estimator"],
    "MED-BIDRIG-001":     ["bid_hunter"],
    "ENV-RS-001":         ["data_scout", "law_inspector"],
    "BUD-CHECKLIST-001":  ["budget_estimator", "review_sentinel"],
    "SOCIAL-INS-001":     ["data_scout"],
    "SOCIAL-MAT-001":     ["data_scout"],
    "SOCIAL-WORK-001":    ["data_scout"],
    "SOCIAL-WELFARE-001": ["data_scout"],
    "BIGDATA-SERVICE-001":["data_scout"],
    "BIGDATA-SQL-001":    ["data_scout"],
    "PERF-DEVIATION-001": ["performance_evaluator"],
    "TRANSFER-TRACE-001": ["budget_estimator", "fiscal_reviewer"],
    "BOND-PENETRATE-001": ["budget_estimator", "fiscal_reviewer"],
}

# 40要素中需要单独结构化的字段（其余原样保留在 card 摘要里）
STRUCTURED_CARD_FIELDS = {
    "audit_goal":        "审计目标",
    "risk_assumption":   "风险假设",
    "scope_applicable":  "适用范围",
    "scope_excluded":    "不适用范围",
    "legal_basis":       "法规及业务依据",
    "core_fields":       "核心字段",
    "primary_keys":      "主键与关联键",
    "data_quality":      "数据质量检查",
    "formula_steps":     "计算公式/步骤",
    "parameters":        "参数与阈值",
    "threshold_basis":   "阈值依据",
    "conclusion_boundary": "结论边界",
    "manual_check":      "人工核查程序",
    "evidence":          "追加证据",
    "test_case":         "测试案例",
    "backtest":          "历史回测结果",
    "false_rate":        "误报率与漏报率",
    "visualization":     "可视化方案",
    "reuse_scenes":      "跨场景复用",
    "desensitization":   "脱敏和权限",
    "data_readiness_eval": "数据就绪度评估",
    "data_grade":        "数据质量分级",
    "validation_std":    "验证标准（目标象限）",
    "workpaper_tpl":     "底稿嵌入模板编号",
    "explain_template":  "可解释性输出模板",
    "retire_condition":  "算法退役条件",
    "review_cycle":      "定期复查周期",
    "dep_algorithms":    "前置依赖算法",
    "expected_value":    "预期审计价值",
    "history_output":    "历史产出记录",
    "risk_scoring":      "风险评分",
    "combination_rules": "多规则组合",
    "sensitivity":       "敏感性分析",
}


def parse_data_dependencies(text: str) -> list:
    """把'1. xxx；2. yyy'拆成数据依赖列表"""
    if not text:
        return []
    parts = re.split(r"[；;]\s*", text)
    out = []
    for p in parts:
        p = re.sub(r"^\s*\d+[\.、)）]\s*", "", p).strip()
        # 去掉括号内的来源说明，保留数据表名
        p = re.sub(r"（.*?）", "", p).strip()
        p = p.rstrip("。.、，,").strip()
        if p:
            out.append(p)
    return out


def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, data_only=True)

    # 1) 总览表
    overview = {}
    for row in wb["☆算法资产库总览"].iter_rows(values_only=True):
        sn = str(row[0]).strip() if row[0] else ""
        if re.match(r"^[A-Z]+-[A-Z]+-\d{3}$", sn):
            overview[sn] = {
                "biz_scene": str(row[2]).strip() if row[2] else "",
                "risk_mechanism": str(row[3]).strip() if row[3] else "",
                "family": str(row[4]).strip() if row[4] else "",
                "maturity": str(row[5]).strip() if row[5] else "",
                "priority_raw": str(row[6]).strip() if row[6] else "",
                "data_readiness": str(row[7]).strip() if row[7] else "",
            }

    # 2) 详细卡片表
    cards = {}
    cur = None
    for row in wb["☆算法详细卡片"].iter_rows(values_only=True):
        a = str(row[0]).strip() if row[0] else ""
        b = str(row[1]).strip() if row[1] else ""
        c = str(row[2]).strip() if row[2] else ""
        m = re.match(r"^算法卡[:：]\s*([A-Z0-9\-]+)", a)
        if m:
            cur = m.group(1)
            cards[cur] = {}
        elif cur and a and b:
            if a != "要素名称":  # 跳过卡片表头行
                cards[cur][a] = c

    # 总览表缺失的算法（如 ENV-CHECKLIST-001 只在卡片表）→ 从卡片推断默认值
    CARD_DEFAULTS = {
        "risk_mechanism": "合规性",
        "family": "清单对照",
        "maturity": "L3",
        "priority_raw": "高",
        "data_readiness": "B级",
    }
    missing = [sn for sn in ASSIGNED_AGENTS if sn not in cards]
    if missing:
        print("⚠ 缺失算法卡:", missing)
        sys.exit(1)
    for sn in ASSIGNED_AGENTS:
        if sn not in overview:
            overview[sn] = dict(CARD_DEFAULTS)
            print(f"ℹ {sn} 不在总览表，使用卡片推断默认值: {CARD_DEFAULTS}")
    for sn in ASSIGNED_AGENTS:
        if not overview[sn].get("biz_scene"):
            overview[sn]["biz_scene"] = cards[sn].get("适用业务场景", "")

    algorithms = {}
    for sn, agents in ASSIGNED_AGENTS.items():
        card = cards[sn]
        ov = overview[sn]
        scene = card.get("适用业务场景") or ov["biz_scene"]
        scene_list = [s.strip() for s in re.split(r"[/／、,，]", scene) if s.strip()]
        dep_text = card.get("输入数据表", "")
        algorithms[sn] = {
            "name": card.get("算法名称", ""),
            "version": card.get("版本/编制人/复核人", "v4.0"),
            "scene": scene_list,
            "biz_scene": ov["biz_scene"],
            "risk_mechanism": ov["risk_mechanism"],
            "family": ov["family"],
            "maturity": ov["maturity"],
            "priority": "P0" if ov["priority_raw"] == "高" else ("P1" if ov["priority_raw"] == "中" else ov["priority_raw"]),
            "data_readiness": ov["data_readiness"],
            "assigned_agents": agents,
            "trigger": card.get("运行与触发条件", ""),
            "data_dependencies": parse_data_dependencies(dep_text),
            "output_type": card.get("输出字段", ""),
            "card": {k: card.get(v, "") for k, v in STRUCTURED_CARD_FIELDS.items()},
        }

    # 3) agent_algorithm_map
    agent_map = {}
    for sn, agents in ASSIGNED_AGENTS.items():
        for ag in agents:
            agent_map.setdefault(ag, []).append(sn)
    for ag in agent_map:
        agent_map[ag].sort()

    registry = {
        "version": "1.0",
        "total_algorithms": len(algorithms),
        "source": "政府审计算法资产库_v4.xlsx",
        "source_path": xlsx,
        "generated_by": "build_algorithm_registry.py",
        "agent_spec_dir": "audit-blackboard/agent_specs/",
        "algorithms": algorithms,
        "agent_algorithm_map": dict(sorted(agent_map.items())),
    }

    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(registry, f, ensure_ascii=False, indent=2)

    print(f"✅ 注册表已生成: {OUT_JSON}")
    print(f"   算法数: {len(algorithms)} | Agent数: {len(agent_map)} | 分配关系数: {sum(len(v) for v in agent_map.values())}")
    for ag, lst in sorted(agent_map.items()):
        print(f"   {ag}: {len(lst)}个算法")


if __name__ == "__main__":
    main()
