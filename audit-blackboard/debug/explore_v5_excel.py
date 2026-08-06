# -*- coding: utf-8 -*-
"""Explore v5.0 Excel structure"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

path = r"C:\Users\scrccpa\Desktop\算法\政府审计算法资产库_v5.xlsx"
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
print("Sheets:", wb.sheetnames)
for ws in wb.worksheets:
    print(f"\n=== Sheet: {ws.title} | dims: {ws.dimensions} | max_row: {ws.max_row} | max_col: {ws.max_column}")
