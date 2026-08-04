# -*- coding: utf-8 -*-
"""
审盾数据帧适配器 (SDF Adapter) v1.0
将 CSV/Excel 原始数据转换为统一的 SDF 格式。
支持：标准Excel、多Sheet、中国式多行表头、CSV（多编码）

用法：
    python sdf_adapter.py --source "data.xlsx" --project pidou_2026 --label "序时账"
    python sdf_adapter.py --source "data.csv" --project pidou_2026 --label "科目余额" --encoding gbk
"""
import sys, os, json, argparse, hashlib, csv as csvmod, datetime, statistics, math
from pathlib import Path
from collections import Counter
sys.stdout.reconfigure(encoding='utf-8')

HERE = Path(__file__).parent
SCHEMA_DIR = HERE.parent / 'schemas'
SDF_SCHEMA_PATH = SCHEMA_DIR / 'sdf_schema_v1.json'

# ═══════════════════════════════════════════════════
# 字段角色自动检测（与 profile_builder 共享逻辑，独立副本）
# ═══════════════════════════════════════════════════

AMOUNT_KW = {'金额', '元', '万元', '亿元', '预算', '支出', '收入', '余额', '资金', '经费', '成本', '费用', '价', '合计', '总计'}
DATE_KW = {'日期', '时间', '年', '月', '日', 'date', 'time', 'period', '期间'}
CATEGORY_KW = {'类型', '分类', '类别', '科目', '项目', '部门', '单位', '名称', '编号', '代码', 'code'}
TEXT_KW = {'说明', '描述', '备注', '意见', '内容', '事由', '摘要', '用途', '用途说明'}
ID_KW = {'编号', '代码', 'id', '序号', '凭证号', '合同号', '发票号'}
QUANTITY_KW = {'数量', '人', '次', '个', '项', '件', '台', '辆', '㎡', '天', '小时', '张', '笔'}
RATIO_KW = {'率', '比例', '占比', '%', 'percent'}


def detect_role(col_name, sample_values):
    """根据列名和样本值推断语义角色"""
    name_lower = col_name.lower()
    for kw in ID_KW:
        if kw in col_name:
            return 'identifier', 0.9
    for kw in AMOUNT_KW:
        if kw in col_name:
            return 'amount', 0.85
    for kw in DATE_KW:
        if kw in col_name:
            return 'date', 0.85
    for kw in CATEGORY_KW:
        if kw in col_name:
            return 'category', 0.7
    for kw in RATIO_KW:
        if kw in col_name:
            return 'ratio', 0.7
    for kw in TEXT_KW:
        if kw in col_name:
            return 'text', 0.7
    for kw in QUANTITY_KW:
        if kw in col_name:
            return 'quantity', 0.7

    numeric_count = 0
    total = 0
    for v in sample_values:
        if v is None or str(v).strip() == '':
            continue
        total += 1
        try:
            float(str(v).replace(',', '').replace('，', '').replace('¥', '').replace('￥', ''))
            numeric_count += 1
        except:
            pass

    if total > 0 and numeric_count / total > 0.8:
        if numeric_count > 3:
            return 'amount', 0.5
    return 'unknown', 0.3


def infer_dtype(values):
    """推断列的数据类型"""
    non_null = [v for v in values if v is not None and str(v).strip() != '']
    if not non_null:
        return 'string'
    int_count = dec_count = date_count = bool_count = str_count = 0
    for v in non_null[:100]:
        s = str(v).strip()
        try:
            f = float(s.replace(',', '').replace('，', ''))
            if '.' in s or 'e' in s.lower():
                dec_count += 1
            elif f == int(f):
                int_count += 1
            else:
                dec_count += 1
            continue
        except:
            pass
        if s.lower() in ('true', 'false', '是', '否', 'yes', 'no'):
            bool_count += 1
            continue
        str_count += 1
    total = len(non_null[:100])
    if int_count / total > 0.6:
        return 'integer'
    if (int_count + dec_count) / total > 0.6:
        return 'decimal'
    if bool_count / total > 0.6:
        return 'boolean'
    return 'string'


def calc_stats(values):
    """计算数值列统计"""
    nums = []
    for v in values:
        if v is None or str(v).strip() == '':
            continue
        try:
            nums.append(float(str(v).replace(',', '').replace('，', '').replace('¥', '').replace('￥', '')))
        except:
            pass
    if len(nums) < 2:
        return None
    try:
        sn = sorted(nums)
        n = len(sn)
        return {
            'min': round(sn[0], 4),
            'max': round(sn[-1], 4),
            'mean': round(statistics.mean(nums), 4),
            'std': round(statistics.stdev(nums), 4) if n >= 2 else 0,
            'sum': round(sum(nums), 2),
            'p25': round(sn[n // 4], 4),
            'p50': round(sn[n // 2], 4),
            'p75': round(sn[3 * n // 4], 4)
        }
    except:
        return None


def top_values(values, n=5):
    """最高频值"""
    c = Counter(str(v) for v in values if v is not None and str(v).strip() != '')
    total = sum(c.values())
    return [{'value': v, 'count': ct, 'ratio': round(ct / total, 4)} for v, ct in c.most_common(n)]


# ═══════════════════════════════════════════════════
# 多行表头检测
# ═══════════════════════════════════════════════════

def detect_header_rows(sheet, max_check=5):
    """
    检测Excel多行表头。
    策略：从第1行开始往下找，直到某行看起来像数据而非表头。
    表头特征：非数字率高、短文本为主。
    """
    rows = list(sheet.iter_rows(min_row=1, max_row=min(max_check, sheet.max_row), values_only=True))
    header_end = 0
    for i, row in enumerate(rows):
        non_numeric = 0
        total = 0
        for cell in row:
            if cell is None:
                continue
            total += 1
            s = str(cell).strip()
            try:
                float(s.replace(',', '').replace('，', ''))
            except:
                non_numeric += 1
        if total > 0 and non_numeric / total > 0.7:
            header_end = i + 1
        else:
            break
    return max(1, header_end)


def merge_multi_headers(sheet, header_rows):
    """
    合并多行表头为单层列名。
    例如：
      行1: ['2025年度收入情况', '', '2025年度支出情况', '']
      行2: ['预算数', '实际数', '预算数', '实际数']
      → ['2025年度收入情况_预算数', '2025年度收入情况_实际数', ...]
    """
    headers = []
    for row_idx in range(1, header_rows + 1):
        row = []
        for cell in sheet[row_idx]:
            row.append(str(cell.value).strip() if cell.value else '')
        headers.append(row)
    max_cols = max(len(r) for r in headers) if headers else 0
    merged = []
    for col in range(max_cols):
        parts = []
        for r in headers:
            if col < len(r) and r[col]:
                parts.append(r[col])
        merged.append('_'.join(parts) if parts else f'Col_{col+1}')
    return merged


# ═══════════════════════════════════════════════════
# CSV Adapter
# ═══════════════════════════════════════════════════

def detect_csv_dialect(filepath):
    """自动检测CSV分隔符和编码"""
    encodings_to_try = ['utf-8', 'utf-8-sig', 'gbk', 'gb2312', 'gb18030', 'latin-1']
    for enc in encodings_to_try:
        try:
            with open(filepath, 'r', encoding=enc) as f:
                sample = f.read(4096)
            delimiter = ','
            if sample.count('\t') > sample.count(','):
                delimiter = '\t'
            elif sample.count(';') > sample.count(','):
                delimiter = ';'
            return delimiter, enc
        except:
            continue
    return ',', 'utf-8'


def adapt_csv(filepath, project_id, label):
    """CSV → SDF"""
    delimiter, encoding = detect_csv_dialect(filepath)
    with open(filepath, 'r', encoding=encoding, errors='replace') as f:
        reader = csvmod.reader(f, delimiter=delimiter)
        rows = list(reader)
    if not rows:
        raise ValueError(f'CSV文件为空: {filepath}')
    headers = rows[0]
    data_rows = rows[1:]
    return build_sdf(headers, data_rows, 'csv', filepath, project_id, label, encoding, header_rows=1)


# ═══════════════════════════════════════════════════
# Excel Adapter
# ═══════════════════════════════════════════════════

def adapt_excel(filepath, project_id, label, sheet_name=None):
    """Excel → SDF，自动处理多Sheet和多行表头"""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)

    sheets_to_process = [sheet_name] if sheet_name else wb.sheetnames
    results = []

    for sname in sheets_to_process:
        ws = wb[sname]
        header_rows = detect_header_rows(ws)
        merged_headers = merge_multi_headers(ws, header_rows)

        data_rows = []
        for row in ws.iter_rows(min_row=header_rows + 1, values_only=True):
            data_rows.append(list(row))

        sdf = build_sdf(merged_headers, data_rows, 'excel', filepath, project_id,
                        f'{label}/{sname}' if len(sheets_to_process) > 1 else label,
                        'utf-8', header_rows, sname)
        results.append(sdf)

    wb.close()
    return results


# ═══════════════════════════════════════════════════
# SDF Builder
# ═══════════════════════════════════════════════════

def build_sdf(headers, data_rows, source_type, filepath, project_id, label,
              encoding='utf-8', header_rows=1, sheet_name=None):
    """组装SDF数据帧"""
    ingestion_time = datetime.datetime.now().isoformat()

    # 计算文件哈希
    try:
        sha = hashlib.sha256()
        with open(filepath, 'rb') as f:
            while True:
                chunk = f.read(65536)
                if not chunk:
                    break
                sha.update(chunk)
        checksum = sha.hexdigest()
    except:
        checksum = 'unknown'

    # 列分析
    columns = []
    for ci, col_name in enumerate(headers):
        col_vals = [r[ci] if ci < len(r) else None for r in data_rows]
        dtype = infer_dtype(col_vals)
        role, role_conf = detect_role(str(col_name), col_vals[:20])
        null_count = sum(1 for v in col_vals if v is None or str(v).strip() == '')
        null_rate = round(null_count / max(len(col_vals), 1), 4)
        unique_rate = round(len(set(str(v) for v in col_vals if v is not None)) / max(len([v for v in col_vals if v is not None]), 1), 4)

        col_def = {
            'name': str(col_name),
            'dtype': dtype,
            'nullable': null_count > 0,
            'unique_rate': unique_rate,
            'null_rate': null_rate,
            'role': role,
            'role_confidence': role_conf,
            'sample_values': [str(v) for v in col_vals[:5] if v is not None]
        }

        if dtype in ('integer', 'decimal'):
            s = calc_stats(col_vals)
            if s:
                col_def['stats'] = s

        if role in ('category', 'text'):
            col_def['top_values'] = top_values(col_vals)

        columns.append(col_def)

    # 整体profile
    total_cells = len(data_rows) * len(headers)
    null_cells = sum(1 for r in data_rows for v in r if v is None or str(v).strip() == '')
    null_rate = round(null_cells / max(total_cells, 1), 4)

    # 重复行检测
    row_strs = [str(tuple(r)) for r in data_rows]
    dup_count = len(row_strs) - len(set(row_strs))

    sdf = {
        'sdf_version': '1.0.0',
        'source': {
            'type': source_type,
            'original_path': str(filepath),
            'ingestion_time': ingestion_time,
            'checksum': f'sha256:{checksum}',
            'header_rows': header_rows
        },
        'profile': {
            'row_count': len(data_rows),
            'col_count': len(headers),
            'encoding': encoding,
            'null_rate': round(null_rate, 4),
            'duplicate_rows': dup_count
        },
        'columns': columns,
        'data_preview': [{str(headers[i]): str(r[i]) if i < len(r) else None for i in range(min(5, len(headers)))} for r in data_rows[:5]],
        'project_ref': {
            'project_id': project_id,
            'label': label
        }
    }

    if sheet_name:
        sdf['source']['sheet_name'] = sheet_name

    return sdf


def save_sdf(sdf, output_dir, project_id, label):
    """保存SDF到JSON文件"""
    out_dir = Path(output_dir) / project_id
    out_dir.mkdir(parents=True, exist_ok=True)
    safe_label = label.replace('/', '_').replace('\\', '_')
    out_path = out_dir / f'{safe_label}_sdf.json'
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(sdf, f, ensure_ascii=False, indent=2, default=str)
    return out_path


# ═══════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════

def main():
    p = argparse.ArgumentParser(description='审盾 SDF 适配器 v1.0 — CSV/Excel → SDF')
    p.add_argument('--source', required=True, help='数据源路径 (.xlsx/.xls/.csv)')
    p.add_argument('--project', required=True, help='项目标识')
    p.add_argument('--label', required=True, help='数据集标签')
    p.add_argument('--sheet', help='指定Sheet名（默认处理全部）')
    p.add_argument('--encoding', help='CSV编码（默认自动检测）')
    p.add_argument('--output', default='.', help='输出目录')
    args = p.parse_args()

    ext = Path(args.source).suffix.lower()
    results = []

    if ext == '.csv':
        sdf = adapt_csv(args.source, args.project, args.label)
        results = [sdf]
    elif ext in ('.xlsx', '.xls'):
        results = adapt_excel(args.source, args.project, args.label, args.sheet)
    else:
        print(f'❌ 不支持的文件类型: {ext}')
        sys.exit(1)

    for sdf in results:
        sheet_label = sdf.get('source', {}).get('sheet_name', args.label)
        sdf_label = sdf.get('project_ref', {}).get('label', sheet_label)
        out_path = save_sdf(sdf, args.output, args.project, sdf_label)
        print(f'✅ SDF已保存: {out_path}')
        print(f'   {sdf["profile"]["row_count"]}行 × {sdf["profile"]["col_count"]}列')
        print(f'   空值率: {sdf["profile"]["null_rate"]:.2%}  重复行: {sdf["profile"]["duplicate_rows"]}')
        roles = Counter(c['role'] for c in sdf['columns'])
        print(f'   字段角色: {dict(roles)}')
        print()


if __name__ == '__main__':
    main()
