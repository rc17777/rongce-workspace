# -*- coding: utf-8 -*-
"""
build_algorithm_registry_v5.py — 从《政府审计算法资产库_v5.xlsx》重建 algorithm_registry.json（135算法 → 融策22Agent体系）

用法:
    python build_algorithm_registry_v5.py [xlsx路径]

输出:
    audit-blackboard/algorithm_registry.json （UTF-8，ensure_ascii=False）

映射规则（任务书）:
    - BID-*, PROC-*              → bid_hunter (招投标猎手)
    - BUDGET-*, BUD-*            → budget_estimator (预算工程师)
    - PERF-*                     → performance_evaluator (绩效评价师)
    - ENG-*                      → settlement_auditor (结算审计师)
    - SOE-*, CHK-LOSS-*          → contract_hound (合同猎犬)
    - ENV-*                      → law_inspector (法规检察官)
    - SOCIAL-*, MED-*            → data_scout (数据侦察兵) [民生类]
    - TAX-*, FIN-*               → data_scout (金融税务)
    - FUND-*, AGR-*              → data_scout (资金/农业)
    - BIGDATA-*, DATA-*, ITCOST-*→ data_scout (数据类)
    - POLICY-*, ECONRESP-*       → fiscal_reviewer (财政评审员)
    - HR-*, SUPV-*               → data_scout (人力/监管)
    - CHK-*                      → data_scout + review_sentinel
    - 其余                        → data_scout（通用数据扫描兜底）
    - 十字交叉型（ENG-FINAL等）    → 主Agent + 副Agent（多Agent协同）

    Excel 自带 "Agent映射" 列（人工设计的工作流提示）作为首要信号，
    前缀规则保证专项 Agent 必达，v4 人工映射作为继承基线。
"""
import sys
import os
import re
import json

sys.stdout.reconfigure(encoding="utf-8")

BASE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_XLSX = r"C:\Users\scrccpa\Desktop\算法\政府审计算法资产库_v5.xlsx"
OUT_JSON = os.path.join(BASE, "algorithm_registry.json")

# ── 中文 Agent 名 → agent_specs/ 下的文件名（不含 .json）────────────────────
AGENT_NAME_MAP = {
    "数据侦察兵": "data_scout",
    "招投标猎手": "bid_hunter",
    "预算工程师": "budget_estimator",
    "合同猎犬": "contract_hound",
    "法规检察官": "law_inspector",
    "底稿工匠": "workpaper_crafter",
    "报告笔杆子": "report_writer",
    "复核哨兵": "review_sentinel",
    "结算审计师": "settlement_auditor",
    "财政评审员": "fiscal_reviewer",
    "绩效评价师": "performance_evaluator",
    "评标偏离度": "expert_bias_detector",
    "会议纪要分析": "meeting_minutes_analyzer",
    "OCR预处理": "ocr_processor",
    "数据分类员": "data_classifier",
    "数据脱敏": "data_desensitizer",
    "调整分录师": "adjustment_scribe",
    "方案撰写师": "plan_writer",
}
AGENT_NAME_PATTERN = re.compile("|".join(sorted(AGENT_NAME_MAP, key=len, reverse=True)))

# 全部 18 个 Agent（含可能 0 算法的预处理类）
ALL_AGENTS = [
    "data_scout", "bid_hunter", "contract_hound", "law_inspector",
    "workpaper_crafter", "report_writer", "review_sentinel",
    "budget_estimator", "settlement_auditor", "fiscal_reviewer",
    "performance_evaluator", "expert_bias_detector",
    "meeting_minutes_analyzer", "ocr_processor", "data_classifier",
    "data_desensitizer", "adjustment_scribe", "plan_writer",
]

# ── 前缀 → 专项 Agent（startswith 匹配，最特化优先）───────────────────────────
# 注意：使用 startswith 而非 token 匹配，以正确处理 PERF2/PROC2/ENV3/SOCIAL2 等批次后缀
PREFIX_RULES = [
    ("BUDGET", "budget_estimator"),  # BUDGET-001~020
    ("PERF", "performance_evaluator"),  # PERF-OUTLIER-001, PERF2-001~004, PERF-COST-001
    ("PROC", "bid_hunter"),  # PROC-CONCEN-001, PROC2-001~004
    ("ECONRESP", "fiscal_reviewer"),  # ECONRESP-001, ECONRESP2-001~003
    ("SOCIAL", "data_scout"),  # SOCIAL-*, SOCIAL2-*, SOCIAL3-*
    ("BIGDATA", "data_scout"),  # BIGDATA-SERVICE-001, BIGDATA-SQL-001
    ("ENV", "law_inspector"),  # ENV-CHECKLIST-001, ENV-RS-001, ENV3-*, ENV4-*
    ("CHK-LOSS", "contract_hound"),  # CHK-LOSS-001（亏损穿透 → 合同猎犬）最特化
    ("CHK", "data_scout"),  # CHK-RECON-001, CHK-RD-001, CHK2-001, CHK2-002 — + review_sentinel
    ("BID", "bid_hunter"),  # BID-PATTERN-005, BID-ROTATE-001, BID-DARKMARK-001
    ("BUD", "budget_estimator"),  # BUD-CHECKLIST-001
    ("ENG", "settlement_auditor"),  # ENG-SAMPLE-001, ENG2-001 等
    ("SOE", "contract_hound"),  # SOE-MIDMAN-001, SOE3-001~003
    ("MED", "data_scout"),  # MED-BIDRIG-001, MED2-001
    ("TAX", "data_scout"),  # TAX-001, TAX-ESCAPE-001
    ("FIN", "data_scout"),  # FIN-SHELL-001, FIN2-001, FIN3-001~003
    ("FUND", "data_scout"),  # FUND-FRAUD-001, FUND-SIPHON-001, FUND2-001
    ("AGR", "data_scout"),  # AGR-INSFAKE-001, AGRI2-001, AGRI3-001~006
    ("DATA", "data_scout"),  # DATA3-001, DATA3-002
    ("ITCOST", "data_scout"),  # ITCOST-001
    ("POLICY", "fiscal_reviewer"),  # POLICY-001
    ("HR", "data_scout"),  # HR-RF-001, HR-RF-002, HR-EATEMPTY-001
    ("SUPV", "data_scout"),  # SUPV-ANOMALY-001, SUPV-POCKET-001, SUPV-TRAVEL-001
]
DEFAULT_AGENT = "data_scout"


def prefix_agent(sn: str):
    """按前缀规则返回专项 Agent；无匹配返回 None（默认 data_scout 由调用方兜底）"""
    for prefix, agent in PREFIX_RULES:
        if sn.startswith(prefix + "-") or sn.startswith(prefix):
            return agent
    return None


def parse_hint_agents(hint: str) -> list:
    """从 'Agent映射' 列解析有序 Agent ID 列表（→ 或 + 分隔，括号内角色说明忽略）"""
    if not hint:
        return []
    out = []
    for seg in re.split(r"[→>＋+]", hint):
        m = AGENT_NAME_PATTERN.search(seg)
        if m:
            aid = AGENT_NAME_MAP[m.group(0)]
            if aid not in out:
                out.append(aid)
    return out


def assign_agents(sn: str, hint: str, v4_agents: list = None) -> list:
    """
    最终分配：Excel Agent映射（人工工作流）优先 → v4 人工映射继承 → 前缀规则保底 → CHK 规则追加复核哨兵。
    去重保序，上限 3 个（主 Agent + 副 Agent）。
    """
    agents = parse_hint_agents(hint)
    for a in (v4_agents or []):
        if a and a not in agents:
            agents.append(a)
    pre = prefix_agent(sn)
    if pre and pre not in agents:
        agents.append(pre)
    if not agents:
        agents = [DEFAULT_AGENT]
    if sn.startswith("CHK") and "review_sentinel" not in agents:
        agents.append("review_sentinel")
    return agents[:3]


# ── v4.0 人工审核映射（继承基线，40算法）────────────────────────────────────
V4_ASSIGNED = {
    "PERF-OUTLIER-001": ["data_scout", "performance_evaluator"],
    "PROC-CONCEN-001": ["bid_hunter", "data_scout"],
    "BID-PATTERN-005": ["bid_hunter"],
    "FUND-FRAUD-001": ["data_scout", "performance_evaluator"],
    "HR-RF-001": ["data_scout"],
    "HR-RF-002": ["data_scout"],
    "REV-PREDICT-001": ["data_scout", "budget_estimator"],
    "ENG-SAMPLE-001": ["settlement_auditor"],
    "ENG-SCORE-001": ["settlement_auditor", "fiscal_reviewer"],
    "CHK-RECON-001": ["data_scout", "review_sentinel"],
    "ASSET-MATCH-001": ["data_scout"],
    "SUPV-ANOMALY-001": ["data_scout"],
    "RULE-MATCH-001": ["law_inspector", "review_sentinel"],
    "SUPV-POCKET-001": ["data_scout", "contract_hound"],
    "SUPV-TRAVEL-001": ["data_scout"],
    "FUND-SIPHON-001": ["data_scout", "workpaper_crafter"],
    "CHK-LOSS-001": ["data_scout", "contract_hound"],
    "PROC-FAKE-001": ["bid_hunter", "law_inspector"],
    "PROC-RELATED-001": ["bid_hunter"],
    "CHK-RD-001": ["data_scout", "law_inspector"],
    "ENG-FINAL-001": ["settlement_auditor", "review_sentinel", "fiscal_reviewer"],
    "SUPV-WARNING-001": ["data_scout"],
    "ENV-CHECKLIST-001": ["law_inspector", "performance_evaluator"],
    "SOE-MIDMAN-001": ["contract_hound", "data_scout"],
    "AGR-INSFAKE-001": ["data_scout"],
    "FIN-SHELL-001": ["data_scout"],
    "FIN-INSFAKE-001": ["data_scout"],
    "ENG-RATIO-001": ["settlement_auditor", "budget_estimator"],
    "MED-BIDRIG-001": ["bid_hunter"],
    "ENV-RS-001": ["data_scout", "law_inspector"],
    "BUD-CHECKLIST-001": ["budget_estimator", "review_sentinel"],
    "SOCIAL-INS-001": ["data_scout"],
    "SOCIAL-MAT-001": ["data_scout"],
    "SOCIAL-WORK-001": ["data_scout"],
    "SOCIAL-WELFARE-001": ["data_scout"],
    "BIGDATA-SERVICE-001": ["data_scout"],
    "BIGDATA-SQL-001": ["data_scout"],
    "PERF-DEVIATION-001": ["performance_evaluator"],
    "TRANSFER-TRACE-001": ["budget_estimator", "fiscal_reviewer"],
    "BOND-PENETRATE-001": ["budget_estimator", "fiscal_reviewer"],
}


def parse_data_deps(text: str) -> list:
    """把 '1. xxx；2. yyy' 拆成数据依赖列表"""
    if not text:
        return []
    parts = re.split(r"[；;]\s*", text)
    out = []
    for p in parts:
        p = re.sub(r"^\s*\d+[\.、)）]\s*", "", p).strip()
        p = re.sub(r"（.*?）", "", p).strip()
        p = p.rstrip("。.、，,").strip()
        if p and p not in out:
            out.append(p)
    return out


def normalize_complexity(raw: str) -> str:
    raw = str(raw or "").strip()
    if raw.startswith("L2"):
        return "L2"
    if raw.startswith("L4"):
        return "L4"
    if raw.startswith("L3"):
        return "L3-ML" if "机器学习" in raw else "L3"
    return raw or "L3"


def build():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)

    # ── 1) 总览表 ──────────────────────────────────────────────
    overview = {}
    for row in wb["☆算法资产库总览"].iter_rows(values_only=True):
        if not row or not row[1]:
            continue
        sn = str(row[1]).strip()
        if not re.match(r"^[A-Z]+[\-A-Z0-9]*-\d{3}$", sn):
            continue
        overview[sn] = {
            "seq": row[0],
            "name": str(row[2]).strip() if row[2] else "",
            "type": str(row[3]).strip() if row[3] else "",
            "scene": str(row[4]).strip() if row[4] else "",
            "risk_mechanism": str(row[5]).strip() if row[5] else "",
            "complexity_raw": str(row[6]).strip() if row[6] else "",
            "biz_line": str(row[7]).strip() if row[7] else "",
            "agent_hint": str(row[8]).strip() if row[8] else "",
            "batch": str(row[9]).strip() if row[9] else "",
            "status": str(row[10]).strip() if row[10] else "",
        }

    # ── 2) 详细卡片表 ──────────────────────────────────────────
    cards = {}
    cur = None
    for row in wb["☆算法详细卡片"].iter_rows(values_only=True):
        a = str(row[0]).strip() if row[0] else ""
        b = str(row[1]).strip() if row[1] else ""
        c = str(row[2]).strip() if row[2] else ""
        m = re.match(r"^算法卡[:：]\s*([A-Z0-9\-]+)", a)
        if m:
            cur = m.group(1)
            cards[cur] = {}
        elif cur and a and a != "要素名称":
            cards[cur][a] = c

    missing_cards = [sn for sn in overview if sn not in cards]
    if missing_cards:
        print("⚠ 缺少算法卡:", missing_cards)
        sys.exit(1)

    # ── 3) 组装 algorithms ─────────────────────────────────────
    algorithms = {}
    for sn, ov in overview.items():
        card = cards[sn]
        alg_type = ov["type"] or ("旗舰" if len(card) > 20 else "骨架")
        is_flagship = alg_type == "旗舰"

        scene_text = ov["scene"] or card.get("适用业务场景", "")
        scene_list = [s.strip() for s in re.split(r"[/／、,，]", scene_text) if s.strip()]

        agents = assign_agents(sn, ov["agent_hint"], V4_ASSIGNED.get(sn))

        # 骨架卡无 输入数据表/输出字段/运行与触发条件 → 用 证据清单/核心信号 推导
        dep_text = card.get("输入数据表") or card.get("证据清单", "")
        trigger = card.get("运行与触发条件") or (
            f"场景触发：{ov['biz_line'] or scene_list[0] if scene_list else '全场景'}；"
            f"{alg_type}算法，命中核心信号即列疑点，按人工核查程序延伸取证"
        )
        output = card.get("输出字段") or (
            f"疑点清单（核心信号命中：{card.get('核心信号', '')[:60]}…）" if card.get("核心信号") else "疑点清单"
        )

        algorithms[sn] = {
            "name": ov["name"] or card.get("算法名称", ""),
            "type": alg_type,
            "version": card.get("版本/编制人") or card.get("版本/编制人/复核人", "v5.0"),
            "scene": scene_list,
            "biz_scene": ov["scene"],
            "biz_line": ov["biz_line"],
            "risk_mechanism": ov["risk_mechanism"],
            "family": card.get("算法族") or card.get("风险机制", ov["risk_mechanism"]),
            "complexity": normalize_complexity(ov["complexity_raw"]),
            "priority": "P0" if is_flagship else "P1",
            "batch": ov["batch"],
            "status": ov["status"],
            "assigned_agents": agents,
            "agent_hint": ov["agent_hint"],
            "trigger": trigger,
            "data_dependencies": parse_data_deps(dep_text),
            "output_type": output,
            "audit_goal": card.get("审计目标", ""),
            "card": card,
        }

    # ── 4) agent_algorithm_map ─────────────────────────────────
    agent_map = {ag: [] for ag in ALL_AGENTS}
    for sn, algo in algorithms.items():
        for ag in algo["assigned_agents"]:
            agent_map.setdefault(ag, []).append(sn)
    for ag in agent_map:
        agent_map[ag].sort()

    registry = {
        "version": "5.0",
        "total_algorithms": len(algorithms),
        "source": "政府审计算法资产库_v5.xlsx",
        "source_path": xlsx,
        "generated_by": "build_algorithm_registry_v5.py",
        "generated_at": "2026-08-06",
        "agent_spec_dir": "audit-blackboard/agent_specs/",
        "algorithms": algorithms,
        "agent_algorithm_map": agent_map,
    }

    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(registry, f, ensure_ascii=False, indent=1)

    # ── 5) 统计输出 ────────────────────────────────────────────
    print(f"✅ 注册表已生成: {OUT_JSON}")
    print(f"   算法总数: {len(algorithms)}（旗舰 {sum(1 for a in algorithms.values() if a['type']=='旗舰')} / 骨架 {sum(1 for a in algorithms.values() if a['type']=='骨架')}）")
    print(f"   Agent数: {len(agent_map)} | 分配关系数: {sum(len(v) for v in agent_map.values())}")
    print("\n   Agent 算法数分布:")
    for ag in sorted(agent_map, key=lambda x: -len(agent_map[x])):
        n = len(agent_map[ag])
        bar = "█" * min(n, 60)
        print(f"     {ag:<24} {n:>3}  {bar}")

    # 汇总供文档使用
    summary = {
        "total": len(algorithms),
        "flagship": sum(1 for a in algorithms.values() if a["type"] == "旗舰"),
        "skeleton": sum(1 for a in algorithms.values() if a["type"] == "骨架"),
        "agent_counts": {ag: len(v) for ag, v in agent_map.items()},
        "complexity": {},
        "priority": {},
    }
    for a in algorithms.values():
        summary["complexity"][a["complexity"]] = summary["complexity"].get(a["complexity"], 0) + 1
        summary["priority"][a["priority"]] = summary["priority"].get(a["priority"], 0) + 1
    with open(os.path.join(BASE, "debug", "registry_summary_v5.json"), "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=1)


if __name__ == "__main__":
    build()
