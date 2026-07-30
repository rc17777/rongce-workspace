# -*- coding: utf-8 -*-
import openpyxl, re, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\scrccpa\Desktop\马尔康市党坝乡石加友村2026年特色旅居村以工代赈项目\马尔康市党坝乡石加友村2026年特色旅居村以工代赈项目.xlsx', data_only=True)
ws = None
for sn in wb.sheetnames:
    if 'F.2' in sn:
        ws = wb[sn]
        break

items = []
i = 0
while i < ws.max_row:
    # Read current row
    c1 = ws.cell(row=i+1, column=2).value  # column B = project code
    if c1 and isinstance(c1, str) and re.match(r'\(\d+\)', c1.strip()):
        # Parse code: "(1) 040101002001" -> code number, full code
        parts = c1.strip().split(')', 1)
        idx_num = parts[0].strip('(')
        full_code = parts[1].strip() if len(parts) > 1 else ''
        name = str(ws.cell(row=i+1, column=5).value or '').strip()
        qty_val = ws.cell(row=i+1, column=14).value
        qty = float(qty_val) if qty_val else 0
        
        # Search forward for subtotal and unit price rows
        labor = 0; material = 0; machine = 0; mgmt = 0; profit = 0; unit_price = 0
        for j in range(i+1, min(i+18, ws.max_row)):
            c0 = str(ws.cell(row=j+1, column=1).value or '')
            # Match "小  计" - stripped of spaces
            c0_clean = c0.replace(' ', '').replace('\u3000', '')
            if c0_clean.startswith('\u5c0f\u8ba1'):  # 小计
                labor = float(ws.cell(row=j+1, column=10).value or 0)
                material = float(ws.cell(row=j+1, column=11).value or 0)
                machine = float(ws.cell(row=j+1, column=12).value or 0)
                mgmt = float(ws.cell(row=j+1, column=13).value or 0)
                profit = float(ws.cell(row=j+1, column=14).value or 0)
            if '\u6e05\u5355\u9879\u76ee\u7efc\u5408\u5355\u4ef7' in c0_clean:  # 清单项目综合单价
                unit_price = float(ws.cell(row=j+1, column=10).value or 0)
                break
        
        items.append({
            'idx': idx_num, 'code': full_code, 'name': name, 'qty': qty,
            'unit_price': unit_price,
            'labor': labor, 'material': material, 'machine': machine,
            'mgmt': mgmt, 'profit': profit,
            'mp': mgmt + profit,
            'pct': (mgmt+profit)/unit_price*100 if unit_price else 0
        })
        i = j + 1
    else:
        i += 1

print(f"{'#':<4} {'Code':<22} {'Name':<40} {'UnitPr':>8} {'Labor':>8} {'Matl':>8} {'Mach':>8} {'Mgmt':>8} {'Profit':>8} {'M+P':>8} {'%':>6}")
print('-'*135)
total_mp = 0
for item in items:
    q = item['qty']
    mp_cost = item['mp'] * q
    total_mp += mp_cost
    name = item['name'][:38]
    print(f"{item['idx']:<4} {item['code']:<22} {name:<40} {item['unit_price']:>8.2f} {item['labor']:>8.2f} {item['material']:>8.2f} {item['machine']:>8.2f} {item['mgmt']:>8.2f} {item['profit']:>8.2f} {mp_cost:>8.0f} {item['pct']:>5.1f}%")

print(f'\n--- Summary ---')
print(f'Total M+P: {total_mp:,.2f}')
print(f'As %% of 7,170,602.52: {total_mp/7170602.52*100:.2f}%')

# Items with high M+P%
print('\n--- Items with M+P > 10% ---')
for item in items:
    if item['pct'] > 10:
        print(f"  #{item['idx']} {item['code']}: {item['name'][:35]}  M+P={item['pct']:.1f}% (UP={item['unit_price']:.2f})")
