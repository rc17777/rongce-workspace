# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ========== 创建工作簿 ==========
wb = openpyxl.Workbook()
ws = wb.active
ws.title = '串标分析最终结论'

# ========== 标题行 ==========
header = ['序号', '结论', '证据来源', '时间/日期', '风险等级']
for i, h in enumerate(header):
    cell = ws.cell(row=1, column=i+1, value=h)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.alignment = Alignment(horizontal='center')
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # 蓝底白字

# ========== 填充数据 ==========
data = [
    [1, '竞争性磋商.pdf和太长无法加载的.pdf由同一台RICOH Pro 8100S设备处理', 'PDF元数据分析', '2025-04-02 03:53:48 和 03:59:41', '高风险'],
    [2, '投标方最晚在2025年4月2日03:53:48之前接触了招标文件原件', 'PDF元数据分析', '2025-04-02 03:53:48', '高风险'],
    [3, '响应文件.pdf是扫描件，只能通过OCR提取内容', 'PDF结构分析', 'N/A', '中风险'],
    [4, '所有三个文件都包含相同的项目名称和参考号', 'OCR文本分析', 'N/A', '低风险'],
    [5, '响应文件.pdf包含完整的投标信息，包括项目概述、背景、技术方案等部分', 'OCR文本分析', 'N/A', '低风险'],
    [6, '三个文件之间的文本相似度高达0.840以上', '深度文本相似性分析', 'N/A', '高风险'],
]

for row in data:
    ws.append(row)

# ========== 样式设置 ==========
# 风险等级着色
RED_FILL = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
ORANGE_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
GREEN_FILL = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

# 添加边框
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
# 设置交替行色
for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
    if row_idx % 2 == 0:
        for cell in row:
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')  # 浅蓝背景

# 设置单元格边框
for row in ws.iter_rows(min_row=1):
    for cell in row:
        cell.border = thin_border

# 列宽自适应
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

# ========== 保存文件 ==========
out_path = r'C:\Users\15528\Desktop\四川护理职业学院2025年校级艺术团专业技能培训与迎新晚会编导服务响应文件\__串标分析最终结论.xlsx'
wb.save(out_path)
print(f'\n报告已生成: {out_path}')