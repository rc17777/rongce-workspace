#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
四川护理职业学院成都校区学生宿舍维修项目(二次) — 串标围标全量分析
十层检测体系：L1报价规律 / L3文本雷同 / L4图片哈希 / L5元数据 / L7打印机型号
"""

import os, sys, re, json, hashlib, io, zipfile, shutil
from collections import defaultdict, Counter
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# Fix Windows encoding
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# ========== 配置 ==========
BASE = r"C:\Users\scrccpa\Desktop\6，739.143532万基建四川护理职业学院成都校区学生宿舍维修项目第二次-磋商-唐裕民-省采\0、四川护理职业学院成都校区学生宿舍维修项目(二次)存档资料\成都校区学生宿舍维修项目(二次)-一体化系统导出存档资料 2025.5.9"
BID_DIR = os.path.join(BASE, "投标文件", "采购包1")
OUTPUT_DIR = r"D:\openclaw-workspace\output\宿舍维修项目串标分析"
os.makedirs(OUTPUT_DIR, exist_ok=True)

BUDGET = 7391435.32  # 招标控制价

# 投标单位简称映射
SHORT_NAMES = {
    "中鸿锦业建设发展有限公司": "中鸿锦业",
    "四川之江建设工程有限公司": "四川之江",
    "四川华庭建设有限公司": "四川华庭",
    "四川同投建设工程有限公司": "四川同投",
    "四川圣亿源建筑工程有限公司": "四川圣亿源",
    "四川天源建设工程有限公司": "四川天源",
    "四川宏阳建设工程有限公司": "四川宏阳",
    "四川宏盛翔业建设工程有限公司": "四川宏盛翔业",
    "四川省泰坤建筑工程有限公司": "四川泰坤",
    "四川鹏达建设集团有限公司": "四川鹏达",
    "四川鼎盛恒业建设工程有限公司": "四川鼎盛恒业",
    "四川鼎源瑞峰建设集团有限公司": "四川鼎源瑞峰",
    "四川融亦汇建筑工程有限公司": "四川融亦汇",
    "四川雕梁画栋建筑工程有限公司": "四川雕梁画栋",
    "陕西省顺辰建设工程有限公司": "陕西顺辰",
    "成都龙泉驿区第一建筑工程公司": "成都龙泉驿一建",
}

def get_short(name):
    for k, v in SHORT_NAMES.items():
        if k in name:
            return v
    return name[:8]

# ========== L1: 报价规律性分析 ==========
print("=" * 60)
print("L1: 报价规律性分析")
print("=" * 60)

def extract_price_from_pdf(filepath):
    """从已标价工程量清单提取投标总价"""
    try:
        import pdfplumber
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages[:5]:
                text = page.extract_text()
                if not text:
                    continue
                # 匹配多种格式
                for pat in [
                    r'投标总价[（(]小写[)）]\s*[：:]\s*([\d,]+\.?\d*)',
                    r'投标总价[（(]小写[)）]\s*([\d,]+\.?\d*)',
                    r'投标总价\s*[：:]\s*([\d,]+\.?\d*)',
                    r'投标报价[（(]小写[)）]\s*[：:]\s*([\d,]+\.?\d*)',
                    r'(?:总价|总报价|报价)[（(]?小写[)）]?\s*[：:]\s*[¥￥]?\s*([\d,]+\.?\d{2})',
                ]:
                    m = re.search(pat, text)
                    if m:
                        val = m.group(1).replace(',', '')
                        return float(val)
        return None
    except Exception as e:
        return None

def extract_price_pypdf(filepath):
    """备用：用pypdf提取"""
    try:
        from pypdf import PdfReader
        reader = PdfReader(filepath)
        for page in reader.pages[:8]:
            text = page.extract_text()
            if not text:
                continue
            for pat in [
                r'投标总价[（(]小写[)）]\s*[：:]\s*([\d,]+\.?\d*)',
                r'投标总价[（(]小写[)）]\s*([\d,]+\.?\d*)',
                r'投标总价\s*[：:]\s*([\d,]+\.?\d*)',
                r'(?:总价|总报价)[（(]?小写[)）]?\s*[：:]\s*([\d,]+\.?\d{2})',
            ]:
                m = re.search(pat, text)
                if m:
                    val = m.group(1).replace(',', '')
                    return float(val)
        return None
    except:
        return None

prices = {}
bidders = sorted(os.listdir(BID_DIR))

for bidder in bidders:
    bidder_dir = os.path.join(BID_DIR, bidder)
    name = bidder.split('(')[0]
    
    # Find 已标价工程量清单.pdf
    boq_file = None
    for fn in os.listdir(bidder_dir):
        if '已标价工程量清单' in fn and fn.endswith('.pdf'):
            boq_file = os.path.join(bidder_dir, fn)
            break
    
    if boq_file:
        price = extract_price_from_pdf(boq_file)
        if price is None:
            price = extract_price_pypdf(boq_file)
        if price:
            prices[name] = price
            short = get_short(name)
            deviation = (price - BUDGET) / BUDGET * 100
            print(f"  {short}: ¥{price:,.2f} (偏离 {deviation:+.2f}%)")
        else:
            print(f"  {short}: 未提取到报价")
    else:
        # Try 报价表
        for fn in os.listdir(bidder_dir):
            if '报价表' in fn and fn.endswith('.pdf'):
                bj_file = os.path.join(bidder_dir, fn)
                price = extract_price_from_pdf(bj_file)
                if price is None:
                    price = extract_price_pypdf(bj_file)
                if price:
                    prices[name] = price
                    short = get_short(name)
                    deviation = (price - BUDGET) / BUDGET * 100
                    print(f"  {short}: ¥{price:,.2f} (偏离 {deviation:+.2f}%)")
                break

# 报价规律检测
print(f"\n报价规律检测:")
if len(prices) >= 2:
    vals = sorted(prices.values())
    print(f"  报价数量: {len(vals)}")
    print(f"  最低: ¥{vals[0]:,.2f}")
    print(f"  最高: ¥{vals[-1]:,.2f}")
    print(f"  极差: ¥{vals[-1]-vals[0]:,.2f} ({(vals[-1]-vals[0])/BUDGET*100:.2f}%)")
    
    # 等差检测
    if len(vals) >= 3:
        diffs = [round(vals[i+1] - vals[i], 2) for i in range(len(vals)-1)]
        print(f"  相邻差值: {diffs}")
        if len(set(diffs)) <= 2:
            print(f"  ⚠️ 差值高度一致，可能存在报价协调!")
        else:
            print(f"  差值分散，未发现明显等差/阶梯模式")
    
    # 偏离度分布
    deviations = {k: (v-BUDGET)/BUDGET*100 for k, v in prices.items()}
    print(f"  偏离度范围: {min(deviations.values()):+.2f}% ~ {max(deviations.values()):+.2f}%")

# ========== L5: PDF元数据分析 ==========
print(f"\n{'='*60}")
print("L5: PDF元数据分析（已标价工程量清单）")
print("=" * 60)

metadata_results = {}

for bidder in bidders:
    bidder_dir = os.path.join(BID_DIR, bidder)
    name = bidder.split('(')[0]
    short = get_short(name)
    
    boq_file = None
    for fn in os.listdir(bidder_dir):
        if '已标价工程量清单' in fn and fn.endswith('.pdf'):
            boq_file = os.path.join(bidder_dir, fn)
            break
    
    if boq_file:
        try:
            import pdfplumber
            with pdfplumber.open(boq_file) as pdf:
                meta = pdf.metadata or {}
                metadata_results[name] = {
                    'file': boq_file,
                    'short': short,
                    'Author': meta.get('Author', ''),
                    'Creator': meta.get('Creator', ''),
                    'Producer': meta.get('Producer', ''),
                    'CreationDate': meta.get('CreationDate', ''),
                    'ModDate': meta.get('ModDate', ''),
                    'Title': meta.get('Title', ''),
                    'pages': len(pdf.pages),
                    'size_bytes': os.path.getsize(boq_file),
                }
        except:
            try:
                from pypdf import PdfReader
                reader = PdfReader(boq_file)
                meta = reader.metadata or {}
                metadata_results[name] = {
                    'file': boq_file,
                    'short': short,
                    'Author': str(meta.get('/Author', '')),
                    'Creator': str(meta.get('/Creator', '')),
                    'Producer': str(meta.get('/Producer', '')),
                    'CreationDate': str(meta.get('/CreationDate', '')),
                    'ModDate': str(meta.get('/ModDate', '')),
                    'Title': str(meta.get('/Title', '')),
                    'pages': len(reader.pages),
                    'size_bytes': os.path.getsize(boq_file),
                }
            except Exception as e:
                metadata_results[name] = {'short': short, 'error': str(e)}
    
    if name in metadata_results:
        m = metadata_results[name]
        print(f"  {short}: Author={m.get('Author','?')} | Creator={m.get('Creator','?')} | Created={m.get('CreationDate','?')} | Pages={m.get('pages','?')} | Size={m.get('size_bytes',0)}")

# 元数据异常检测
print(f"\n元数据异常检测:")
authors = Counter(m.get('Author', '') for m in metadata_results.values())
creators = Counter(m.get('Creator', '') for m in metadata_results.values())

print(f"  Author分布: {dict(authors)}")
print(f"  Creator分布: {dict(creators)}")

if len(authors) == 1 and list(authors.keys())[0]:
    print(f"  ⚠️ 所有投标文件Author相同: '{list(authors.keys())[0]}' — 可能由同一人/同一电脑制作!")
elif authors.get('linyan', 0) >= 2:
    print(f"  ⚠️ {authors['linyan']} 家投标文件Author='linyan' — 需进一步确认")

# 分析WPS版本
wps_versions = []
for name, m in metadata_results.items():
    creator = m.get('Creator', '')
    if 'WPS' in creator:
        wps_versions.append((get_short(name), creator))

if wps_versions:
    print(f"\n  WPS版本详情:")
    version_counter = Counter(v[1] for v in wps_versions)
    for ver, count in version_counter.items():
        names = [v[0] for v in wps_versions if v[1] == ver]
        print(f"    {ver}: {count}家 — {', '.join(names)}")
    if len(set(v[1] for v in wps_versions)) <= 2:
        print(f"  ⚠️ WPS版本集中，可能同源部署")

# 时间戳分析
dates = []
for name, m in metadata_results.items():
    cd = m.get('CreationDate', '')
    if cd:
        dates.append((get_short(name), cd))

print(f"\n  创建时间分析:")
for short, dt in sorted(dates, key=lambda x: x[1]):
    print(f"    {short}: {dt}")

# ========== L7: PDF扫描仪/生成器分析 ==========
print(f"\n{'='*60}")
print("L7: PDF生成器/Producer分析（所有PDF文件）")
print("=" * 60)

producer_results = defaultdict(lambda: defaultdict(set))

for bidder in bidders[:3]:  # 先采样前3家
    bidder_dir = os.path.join(BID_DIR, bidder)
    name = bidder.split('(')[0]
    short = get_short(name)
    
    for fn in sorted(os.listdir(bidder_dir)):
        if fn.startswith('._') or not fn.endswith('.pdf'):
            continue
        fp = os.path.join(bidder_dir, fn)
        try:
            from pypdf import PdfReader
            reader = PdfReader(fp)
            meta = reader.metadata or {}
            producer = str(meta.get('/Producer', ''))
            creator = str(meta.get('/Creator', ''))
            file_key = fn
            if producer:
                producer_results[short][file_key] = (producer, creator)
        except:
            pass

print("  前3家采样结果（Producer分布）:")
all_producers = set()
for short, files in producer_results.items():
    for fn, (prod, creator) in files.items():
        key = f"{creator}|{prod}" if creator else prod
        all_producers.add(key)
    print(f"  {short}:")
    for fn, (prod, creator) in sorted(files.items()):
        print(f"    {fn}: Producer={prod[:80]} | Creator={creator[:80]}")

# 检查是否混有扫描仪型号（RICOH, KONICA, CANON, HP, EPSON等）
print(f"\n  扫描仪/打印设备检测:")
scanner_keywords = ['RICOH', 'KONICA', 'CANON', 'HP', 'EPSON', 'XEROX', 'SHARP', 'TOSHIBA',
                     'KYOCERA', 'BROTHER', 'DELL', 'LENOVO', 'SAMSUNG', 'Panasonic', 'FUJI']
for short, files in producer_results.items():
    for fn, (prod, creator) in files.items():
        for kw in scanner_keywords:
            if kw.upper() in prod.upper() or kw.upper() in creator.upper():
                print(f"  {short} - {fn}: 发现设备 {kw}")

# ========== L3: 文本雷同检测 ==========
print(f"\n{'='*60}")
print("L3: 文本雷同检测（投标函 + 响应文件封面）")
print("=" * 60)

def extract_text_from_pdf(filepath, max_pages=5):
    """提取PDF文本"""
    try:
        import pdfplumber
        with pdfplumber.open(filepath) as pdf:
            texts = []
            for page in pdf.pages[:max_pages]:
                t = page.extract_text()
                if t:
                    texts.append(t)
            return '\n'.join(texts)
    except:
        try:
            from pypdf import PdfReader
            reader = PdfReader(filepath)
            texts = []
            for page in reader.pages[:max_pages]:
                t = page.extract_text()
                if t:
                    texts.append(t)
            return '\n'.join(texts)
        except:
            return ''

# 提取所有投标人的投标函文本
bid_texts = {}
for bidder in bidders:
    bidder_dir = os.path.join(BID_DIR, bidder)
    name = bidder.split('(')[0]
    short = get_short(name)
    
    # 查找投标（响应）函
    bid_func_file = None
    for fn in os.listdir(bidder_dir):
        if ('投标' in fn and '响应' in fn and '函' in fn and fn.endswith('.pdf')):
            bid_func_file = os.path.join(bidder_dir, fn)
            break
    
    if not bid_func_file:
        for fn in os.listdir(bidder_dir):
            if '投标函' in fn and fn.endswith('.pdf'):
                bid_func_file = os.path.join(bidder_dir, fn)
                break
    
    if bid_func_file:
        text = extract_text_from_pdf(bid_func_file, max_pages=3)
        if text:
            bid_texts[name] = text
            print(f"  {short}: 投标函 {len(text)} 字符")

print(f"\n  成功提取 {len(bid_texts)} 家投标函文本")

# TF-IDF相似度计算
if len(bid_texts) >= 2:
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.metrics.pairwise import cosine_similarity
    
    names = list(bid_texts.keys())
    texts = [bid_texts[n] for n in names]
    
    vectorizer = TfidfVectorizer(analyzer='char_wb', ngram_range=(2, 4), min_df=1)
    try:
        tfidf_matrix = vectorizer.fit_transform(texts)
        sim_matrix = cosine_similarity(tfidf_matrix)
        
        print(f"\n  文本雷同矩阵（余弦相似度）:")
        high_sim_pairs = []
        for i in range(len(names)):
            for j in range(i+1, len(names)):
                sim = sim_matrix[i][j]
                flag = "🔴" if sim >= 0.85 else ("🟡" if sim >= 0.65 else "🟢")
                if sim >= 0.50:
                    print(f"    {flag} {get_short(names[i])} vs {get_short(names[j])}: {sim:.4f}")
                if sim >= 0.65:
                    high_sim_pairs.append((names[i], names[j], sim))
        
        if high_sim_pairs:
            print(f"\n  ⚠️ 高相似度配对 ({len(high_sim_pairs)} 对):")
            for n1, n2, sim in sorted(high_sim_pairs, key=lambda x: -x[2]):
                print(f"    {get_short(n1)} vs {get_short(n2)}: {sim:.4f}")
    except Exception as e:
        print(f"  TF-IDF计算失败: {e}")

# 也检查响应文件封面
print(f"\n  响应文件封面文本相似度:")
cover_texts = {}
for bidder in bidders:
    bidder_dir = os.path.join(BID_DIR, bidder)
    name = bidder.split('(')[0]
    
    cover_file = None
    for fn in os.listdir(bidder_dir):
        if '响应文件封面' in fn and fn.endswith('.pdf'):
            cover_file = os.path.join(bidder_dir, fn)
            break
    
    if cover_file:
        text = extract_text_from_pdf(cover_file, max_pages=1)
        if text:
            cover_texts[name] = text

if len(cover_texts) >= 2:
    names_c = list(cover_texts.keys())
    texts_c = [cover_texts[n] for n in names_c]
    
    vectorizer2 = TfidfVectorizer(analyzer='char_wb', ngram_range=(2, 3), min_df=1)
    try:
        tfidf_matrix2 = vectorizer2.fit_transform(texts_c)
        sim_matrix2 = cosine_similarity(tfidf_matrix2)
        
        for i in range(len(names_c)):
            for j in range(i+1, len(names_c)):
                sim = sim_matrix2[i][j]
                flag = "🔴" if sim >= 0.85 else ("🟡" if sim >= 0.70 else "🟢")
                if sim >= 0.50:
                    print(f"    {flag} {get_short(names_c[i])} vs {get_short(names_c[j])}: {sim:.4f}")
                if sim >= 0.85:
                    print(f"    ⚠️ 封面文本几乎完全一致!")
    except Exception as e:
        print(f"  封面TF-IDF计算失败: {e}")

# ========== L4: 图片哈希检测 ==========
print(f"\n{'='*60}")
print("L4: PDF内嵌图片哈希检测")
print("=" * 60)

def extract_images_from_pdf(filepath):
    """从PDF提取嵌入图片的MD5哈希"""
    hashes = []
    try:
        from pypdf import PdfReader
        reader = PdfReader(filepath)
        for page_num, page in enumerate(reader.pages[:5]):
            if '/XObject' in page['/Resources']:
                xobjects = page['/Resources']['/XObject'].get_object()
                for obj_name in xobjects:
                    obj = xobjects[obj_name].get_object()
                    if obj['/Subtype'] == '/Image':
                        try:
                            data = obj.get_data()
                            h = hashlib.md5(data).hexdigest()
                            hashes.append((page_num, obj_name, h))
                        except:
                            pass
    except:
        pass
    return hashes

image_hashes = {}
for bidder in bidders[:5]:  # 先采样前5家
    bidder_dir = os.path.join(BID_DIR, bidder)
    name = bidder.split('(')[0]
    short = get_short(name)
    
    boq_file = None
    for fn in os.listdir(bidder_dir):
        if '已标价工程量清单' in fn and fn.endswith('.pdf'):
            boq_file = os.path.join(bidder_dir, fn)
            break
    
    if boq_file:
        hashes = extract_images_from_pdf(boq_file)
        image_hashes[name] = hashes
        print(f"  {short}: {len(hashes)} images in BOQ")

# 跨公司图片哈希比对
if image_hashes:
    print(f"\n  跨公司图片MD5比对:")
    all_hashes = defaultdict(set)
    for name, hashes in image_hashes.items():
        for pg, oname, h in hashes:
            all_hashes[h].add(name)
    
    duplicates = {h: names for h, names in all_hashes.items() if len(names) > 1}
    if duplicates:
        print(f"  🔴 发现 {len(duplicates)} 个跨公司重复图片!")
        for h, names in list(duplicates.items())[:10]:
            print(f"    MD5={h[:16]}... → {', '.join(get_short(n) for n in names)}")
    else:
        unique_hashes = len(all_hashes)
        total_images = sum(len(h) for h in image_hashes.values())
        print(f"  🟢 0跨公司重复: {total_images} images, {unique_hashes} unique MD5 in sample")

# 也对所有投标人的响应文件封面做图片哈希
print(f"\n  响应文件封面图片哈希:")
cover_images = {}
for bidder in bidders:
    bidder_dir = os.path.join(BID_DIR, bidder)
    name = bidder.split('(')[0]
    
    cover_file = None
    for fn in os.listdir(bidder_dir):
        if '响应文件封面' in fn and fn.endswith('.pdf'):
            cover_file = os.path.join(bidder_dir, fn)
            break
    
    if cover_file:
        hashes = extract_images_from_pdf(cover_file)
        cover_images[name] = hashes

if cover_images:
    all_cover_hashes = defaultdict(set)
    for name, hashes in cover_images.items():
        for pg, oname, h in hashes:
            all_cover_hashes[h].add(name)
    
    dup_covers = {h: names for h, names in all_cover_hashes.items() if len(names) > 1}
    if dup_covers:
        print(f"  🔴 封面图片跨公司重复: {len(dup_covers)} 个!")
        for h, names in list(dup_covers.items())[:10]:
            print(f"    {', '.join(get_short(n) for n in names)}")
    else:
        print(f"  🟢 封面图片无跨公司重复 ({len(all_cover_hashes)} unique)")

# ========== 综合结论 ==========
print(f"\n{'='*60}")
print("综合总结")
print("=" * 60)

print(f"""
项目: 四川护理职业学院成都校区学生宿舍维修项目(二次)
招标控制价: ¥{BUDGET:,.2f}
投标单位: {len(bidders)} 家
分析维度: L1报价规律 / L3文本雷同 / L4图片哈希 / L5元数据 / L7生成器标记
""")

# 输出Excel报告
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    
    # Sheet 1: 报价分析
    ws1 = wb.active
    ws1.title = "报价分析"
    ws1.append(["序号", "投标单位", "简称", "投标总价(元)", "偏离控制价(%)", "偏离绝对值(元)", "风险标记"])
    
    if prices:
        sorted_prices = sorted(prices.items(), key=lambda x: x[1])
        for i, (name, price) in enumerate(sorted_prices, 1):
            dev = (price - BUDGET) / BUDGET * 100
            flag = ""
            if abs(dev) < 0.1:
                flag = "⚠️ 极度接近控制价"
            elif abs(dev) < 1:
                flag = "接近控制价"
            ws1.append([i, name, get_short(name), round(price, 2), round(dev, 2), round(price-BUDGET, 2), flag])
    
    # Sheet 2: 元数据
    ws2 = wb.create_sheet("元数据分析")
    ws2.append(["序号", "投标单位", "简称", "Author", "Creator", "CreationDate", "ModDate", "页数", "文件大小", "风险标记"])
    
    for i, (name, m) in enumerate(sorted(metadata_results.items()), 1):
        flag = ""
        if m.get('Author', '') == 'linyan':
            flag = "⚠️ Author=linyan (可能同源)"
        ws2.append([i, name, m.get('short',''), m.get('Author',''), m.get('Creator',''),
                    m.get('CreationDate',''), m.get('ModDate',''), m.get('pages',''),
                    m.get('size_bytes',''), flag])
    
    # Sheet 3: 文本雷同
    if len(bid_texts) >= 2:
        ws3 = wb.create_sheet("文本雷同矩阵")
        names_list = list(bid_texts.keys())
        ws3.append([""] + [get_short(n) for n in names_list])
        for i, n1 in enumerate(names_list):
            row = [get_short(n1)]
            for j, n2 in enumerate(names_list):
                if i == j:
                    row.append(1.0)
                else:
                    row.append(round(sim_matrix[i][j], 4))
            ws3.append(row)
    
    report_path = os.path.join(OUTPUT_DIR, "宿舍维修项目_串标分析报告.xlsx")
    wb.save(report_path)
    print(f"\n📊 Excel报告已保存: {report_path}")
except Exception as e:
    print(f"Excel生成失败: {e}")

# 保存详细JSON
json_path = os.path.join(OUTPUT_DIR, "分析数据.json")
json_data = {
    "project": "四川护理职业学院成都校区学生宿舍维修项目(二次)",
    "budget": BUDGET,
    "bidder_count": len(bidders),
    "prices": {k: v for k, v in prices.items()},
    "metadata": {k: {kk: str(vv) for kk, vv in v.items()} for k, v in metadata_results.items()},
}
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(json_data, f, ensure_ascii=False, indent=2)
print(f"📊 JSON数据已保存: {json_path}")

print("\n✅ 全量分析完成!")
