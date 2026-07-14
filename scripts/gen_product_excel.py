#!/usr/bin/env python3
"""融策"AI+审计场景定义"咨询服务产品手册 — Excel 生成脚本"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# 样式定义
# ============================================================
DARK_BLUE = "1B3A5C"
MED_BLUE = "2E6B9E"
LIGHT_BLUE = "D6E4F0"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
RED = "C0392B"
ORANGE = "E67E22"
GREEN = "27AE60"
YELLOW_BG = "FFF8E1"

header_font = Font(name="微软雅黑", bold=True, color=WHITE, size=11)
title_font = Font(name="微软雅黑", bold=True, color=DARK_BLUE, size=14)
subtitle_font = Font(name="微软雅黑", bold=True, color=MED_BLUE, size=12)
body_font = Font(name="微软雅黑", size=10)
bold_font = Font(name="微软雅黑", bold=True, size=10)
small_font = Font(name="微软雅黑", size=9, color="666666")
red_font = Font(name="微软雅黑", bold=True, color=RED, size=10)
green_font = Font(name="微软雅黑", bold=True, color=GREEN, size=10)
orange_font = Font(name="微软雅黑", bold=True, color=ORANGE, size=10)

header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
sub_header_fill = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
light_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
gray_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
yellow_fill = PatternFill(start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type="solid")
green_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
red_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
wrap = Alignment(wrap_text=True, vertical="top")

thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

def style_data_row(ws, row, max_col, alt=False):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = body_font
        cell.alignment = left if col > 1 else center
        cell.border = thin_border
        if alt:
            cell.fill = gray_fill

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_title(ws, row, col, text, merge_end_col=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = title_font
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end_col)

# ============================================================
# Sheet 1: 产品总览
# ============================================================
ws1 = wb.active
ws1.title = "1-产品总览"
set_col_widths(ws1, [5, 18, 40, 40])

write_title(ws1, 1, 1, "融策「AI+审计场景定义」咨询服务 · 产品总览", 4)
ws1.cell(row=2, column=1, value="版本：V1.0 | 日期：2026-06-01").font = small_font

# 产品定位
ws1.merge_cells("A4:D4")
ws1.cell(row=4, column=1, value="一、产品定位").font = subtitle_font

data1 = [
    ["", "我们卖的不是", "我们卖的是"],
    ["1", "AI系统/软件", "帮审计局定义「AI在哪个环节、怎么用、用到什么程度」"],
    ["2", "技术开发服务", "审计业务语言 → AI需求语言的翻译能力"],
    ["3", "一次性项目", "让审计局逐步具备「自己定义场景」的组织能力"],
    ["", "", ""],
    ["", "一句话定位", "融策不造AI，融策帮审计局想清楚「AI在我这到底能干什么、怎么干、从哪开始」"],
]
row = 5
for i, d in enumerate(data1):
    for j, val in enumerate(d):
        ws1.cell(row=row, column=j+1, value=val)
    row += 1
    if i >= 2 and i <= 4:
        for j in range(1, 5):
            ws1.cell(row=row-1, column=j).font = bold_font if j > 1 else body_font
            ws1.cell(row=row-1, column=j).border = thin_border

# 目标客户
row += 1
ws1.cell(row=row, column=1, value="二、目标客户").font = subtitle_font
row += 1
customers = [
    ["客户类型", "特征", "需求强度"],
    ["市县审计局", "有数字化转型意识但缺乏方法论和内驱力", "🔴 高"],
    ["省级审计厅", "需要制定全省审计AI应用规划", "🟡 中"],
    ["大型国企内审部门", "预算充足、数据基础好、场景明确", "🟡 中"],
]
for i, c in enumerate(customers):
    for j, val in enumerate(c):
        ws1.cell(row=row, column=j+1, value=val)
    row += 1
style_header_row(ws1, row - len(customers), 4)

# 四层全景
row += 1
ws1.cell(row=row, column=1, value="三、四层渐进式服务全景").font = subtitle_font
row += 1
panorama = [
    ["层级", "名称", "核心问题", "周期", "定价", "产出物"],
    ["L0", "AI审计场景就绪度诊断", "你们现在在哪？", "2周", "3-5万", "《就绪度评估报告》+ 场景清单"],
    ["L1", "场景优先级筛选", "先打哪个点？", "1周", "2-3万", "《场景优先级矩阵》+ 切入点建议"],
    ["L2", "AI审计场景定义书 ⭐核心", "这个点怎么打？", "2-3周", "5-8万", "《场景定义书》(5个子模块)"],
    ["L3", "场景原型验证", "打一次试试", "2-4周", "8-15万", "《场景验证报告》+ 运行原型"],
    ["L4", "固化与赋能", "以后自己打", "2周", "3-5万", "《操作手册》+ 培训 + 移交"],
]
for i, c in enumerate(panorama):
    for j, val in enumerate(c):
        ws1.cell(row=row, column=j+1, value=val)
    row += 1
style_header_row(ws1, row - len(panorama), 6)
for r in range(row - len(panorama) + 1, row):
    if r == row - len(panorama) + 3:  # L2 row highlight
        for col in range(1, 7):
            ws1.cell(row=r, column=col).fill = yellow_fill

# 套餐
row += 1
ws1.cell(row=row, column=1, value="四、套餐定价").font = subtitle_font
row += 1
pkgs = [
    ["套餐", "包含层级", "预估总额", "建议场景数", "适用"],
    ["入门体验", "L0 + L1", "5-8万", "不限", "首次接触AI审计的审计局"],
    ["标准交付 ⭐", "L0 → L4 全流程", "21-36万", "1个场景", "有明确数字化预算的审计局"],
    ["多场景扩展", "L0+L1一次 + L2-L4 × N", "单个场景增量15-28万", "2-3个场景", "验证成功后复购"],
]
for i, c in enumerate(pkgs):
    for j, val in enumerate(c):
        ws1.cell(row=row, column=j+1, value=val)
    row += 1
style_header_row(ws1, row - len(pkgs), 5)

# 差异化
row += 1
ws1.cell(row=row, column=1, value="五、差异化优势").font = subtitle_font
row += 1
diffs = [
    ["维度", "纯AI公司", "纯咨询公司", "融策"],
    ["懂审计业务", "❌", "🟡", "✅ 核心能力"],
    ["有AI工具链", "✅", "❌", "✅ 5大AI技能体系"],
    ["能定义场景规则", "❌", "🟡", "✅ 审计经验+规则拆解"],
    ["能做原型验证", "✅", "❌", "✅ SQL+Python自主开发"],
    ["价格竞争力", "高(含技术溢价)", "中高", "中(县级可承受)"],
    ["长期关系", "项目制，交付即结束", "项目制", "可迭代复购"],
]
for i, d in enumerate(diffs):
    for j, val in enumerate(d):
        ws1.cell(row=row, column=j+1, value=val)
    row += 1
style_header_row(ws1, row - len(diffs), 4)

# ============================================================
# Sheet 2: 四层服务详情
# ============================================================
ws2 = wb.create_sheet("2-服务详情")
set_col_widths(ws2, [5, 18, 40, 40, 25])

write_title(ws2, 1, 1, "四层服务 · 详细内容", 5)

# L0
row = 3
ws2.cell(row=row, column=1, value="L0：AI审计场景就绪度诊断（2周·3-5万）").font = subtitle_font
row += 1
l0 = [
    ["步骤", "方法", "产出", "工具"],
    ["审计流程拆解", "选取2-3个核心审计类型，绘制端到端泳道图", "审计流程泳道图", "draw.io"],
    ["数据资产盘点", "梳理现有系统、数据类型、质量、可获取性", "数据资产清单+就绪度评分表", "agent-data-standard"],
    ["人员访谈", "与科长/业务骨干访谈，了解真实痛点和AI认知", "访谈纪要+痛点热力图", "结构化访谈提纲"],
    ["差距分析", "对照12项agent数据标准逐项评估", "《就绪度评估报告》", "12项检查清单"],
]
for i, c in enumerate(l0):
    for j, val in enumerate(c):
        ws2.cell(row=row, column=j+1, value=val)
    row += 1
style_header_row(ws2, row - len(l0), 5)

# L1
row += 1
ws2.cell(row=row, column=1, value="L1：场景优先级筛选（1周·2-3万）").font = subtitle_font
row += 1
l1 = [
    ["评估维度", "权重", "评分标准", "数据来源"],
    ["痛点强度", "30%", "当前环节月均投入人工(人天) / 出错率 / 延期频率", "L0访谈+流程数据"],
    ["数据就绪度", "25%", "所需数据的完整性、可获取性、结构化程度", "L0数据盘点"],
    ["见效速度", "20%", "从启动到可量化成果的预期周期(月)", "专家评估"],
    ["容错空间", "15%", "出错后果严重程度（辅助判断→自动执行的风险梯度）", "审计业务判断"],
    ["组织意愿", "10%", "业务科室的配合意愿和牵头人积极性", "L0访谈"],
]
for i, c in enumerate(l1):
    for j, val in enumerate(c):
        ws2.cell(row=row, column=j+1, value=val)
    row += 1
style_header_row(ws2, row - len(l1), 5)

# L2
row += 1
ws2.cell(row=row, column=1, value="L2：AI审计场景定义书 ⭐核心层（2-3周·5-8万）").font = subtitle_font
row += 1
l2 = [
    ["子模块", "定义内容", "示例(以涉农补贴AI初审为例)", "产出格式"],
    ["场景边界", "场景名称/业务范围/触发条件/输入输出/边界声明", "触发条件：财政拨付台账更新OR审计组下达指令", "结构化表格"],
    ["处理规则", "合规规则(政策文件→结构化)/经验规则(cot-capture→If/Then)/阈值规则(统计+判断)", "IF 身份证号∈财政供养人员库 THEN 标记'违规领取'", "YAML/Excel规则表"],
    ["例外分级(E0-E3)", "E0=AI直接判定 / E1=高置信度建议 / E2=AI标记需调查 / E3=转人工", "E0:死亡库+补贴日在死亡日后→自动标记", "分级决策表"],
    ["数据映射", "业务概念↔系统来源↔字段名↔格式↔获取方式↔就绪状态", "补贴对象姓名←一卡通系统→T_RECV.NAME(VARCHAR50)→DB直连→✅", "数据字典"],
    ["人机协同流程", "指令下发→数据处理→疑点分级→人工审阅→反馈迭代→任务结束", "详见流程图", "泳道图/流程图"],
]
for i, c in enumerate(l2):
    for j, val in enumerate(c):
        ws2.cell(row=row, column=j+1, value=val)
    row += 1
style_header_row(ws2, row - len(l2), 5)

# L3
row += 1
ws2.cell(row=row, column=1, value="L3：场景原型验证（2-4周·8-15万）").font = subtitle_font
row += 1
l3 = [
    ["步骤", "方法", "产出", "验证指标"],
    ["原型搭建", "基于定义书，用SQL+Python搭建最小可用版本", "可运行原型", "技术可运行 ☐"],
    ["历史数据回测", "用近2年审计数据跑一遍，对比原审计结果", "准确率/召回率报告", "疑点命中率≥30% / 漏报率≤10%"],
    ["真实数据试跑", "当前审计项目真实数据运行，审计人员实操验证", "试跑日志+使用反馈", "效率提升≥50%"],
    ["迭代优化", "根据反馈修改规则、调整阈值、补充例外方案", "V1.1版定义书", "人工确认耗时≤2h/百条"],
]
for i, c in enumerate(l3):
    for j, val in enumerate(c):
        ws2.cell(row=row, column=j+1, value=val)
    row += 1
style_header_row(ws2, row - len(l3), 5)

# L4
row += 1
ws2.cell(row=row, column=1, value="L4：固化与赋能（2周·3-5万）").font = subtitle_font
row += 1
l4 = [
    ["步骤", "方法", "产出"],
    ["操作SOP编写", "将全过程标准化为操作手册", "《AI辅助审计操作手册》"],
    ["提示词入库", "场景规则对应的提示词模板入库prompt-librarian体系", "YAML提示词文件"],
    ["人员培训(2天)", "理论0.5天+实操1天+案例研讨0.5天", "培训课件+考核试卷"],
    ["迭代机制建立", "明确监测人/触发人/迭代流程", "《场景迭代管理规程》"],
    ["知识移交", "全部文档/代码/数据移交，部署到审计局环境", "移交清单"],
]
for i, c in enumerate(l4):
    for j, val in enumerate(c):
        ws2.cell(row=row, column=j+1, value=val)
    row += 1
style_header_row(ws2, row - len(l4), 4)

# ============================================================
# Sheet 3: 场景定义书模板
# ============================================================
ws3 = wb.create_sheet("3-场景定义书模板")
set_col_widths(ws3, [5, 22, 35, 35, 18])

write_title(ws3, 1, 1, "场景定义书模板 (L2核心产出物)", 5)
ws3.cell(row=2, column=1, value="说明：以下为空白模板，实际交付时根据选定的审计场景逐项填写").font = small_font

row = 4
ws3.cell(row=row, column=1, value="子模块一：场景边界定义").font = subtitle_font
row += 1
bd = [
    ["序号", "定义项", "填写内容", "填写说明"],
    ["1", "场景名称", "（填写）", "精确命名，如「县级涉农专项资金拨付合规性AI初审」"],
    ["2", "所属审计类型", "（填写）", "如：专项资金审计 / 经济责任审计 / 预算执行审计"],
    ["3", "业务范围", "（填写）", "覆盖哪些资金类型 / 哪些环节"],
    ["4", "触发条件", "（填写）", "什么情况下启动AI？如：财政拨付台账更新 OR 审计组下达检查指令"],
    ["5", "输入数据", "（填写）", "AI需要哪些数据？（逐一列出）"],
    ["6", "输出结果", "（填写）", "AI产出什么？格式？"],
    ["7", "不做什么（边界声明）", "（填写）", "明确排除范围，如：不做最终定性/不做报告撰写"],
]
for i, c in enumerate(bd):
    for j, val in enumerate(c):
        ws3.cell(row=row, column=j+1, value=val)
    row += 1
style_header_row(ws3, row - len(bd), 5)

row += 1
ws3.cell(row=row, column=1, value="子模块二：处理规则定义").font = subtitle_font
row += 1
rules = [
    ["规则编号", "规则类型", "规则内容(If/Then格式)", "规则来源", "置信度"],
    ["R01", "合规规则", "IF （填写条件） THEN 标记「（填写）」", "政策文件：__________", "高/中"],
    ["R02", "合规规则", "IF （填写条件） THEN 标记「（填写）」", "政策文件：__________", "高/中"],
    ["R03", "经验规则", "IF （填写条件） THEN 标记「（填写）」", "审计专家访谈", "高/中"],
    ["R04", "经验规则", "IF （填写条件） THEN 标记「（填写）」", "审计专家访谈", "高/中"],
    ["R05", "阈值规则", "IF （填写条件） THEN 标记「（填写）」", "历史数据统计", "高/中"],
    ["R06", "阈值规则", "IF （填写条件） THEN 标记「（填写）」", "历史数据统计", "高/中"],
]
for i, c in enumerate(rules):
    for j, val in enumerate(c):
        ws3.cell(row=row, column=j+1, value=val)
    row += 1
style_header_row(ws3, row - len(rules), 5)
# Mark rule rows as needing fill
for r in range(row - len(rules) + 1, row):
    ws3.cell(row=r, column=3).fill = yellow_fill

row += 1
ws3.cell(row=row, column=1, value="子模块三：例外分级方案（最关键模块）").font = subtitle_font
row += 1
exc = [
    ["级别", "定义", "处理方式", "示例（填写）"],
    ["E0", "AI可直接判定，置信度≥95%", "自动标记，无需人工确认", "（填写本场景的E0级例外）"],
    ["E1", "AI高置信度建议，置信度80%-95%", "AI标记+推荐结论，人工一键确认", "（填写）"],
    ["E2", "AI发现异常但不推荐结论", "AI标记，人工调查后定性", "（填写）"],
    ["E3", "AI无法判断，数据缺失或超出规则范围", "直接转人工，AI不干预", "（填写）"],
    ["", "例外覆盖率目标", "上线前≥80% → 运行2周后≥90%", ""],
]
for i, c in enumerate(exc):
    for j, val in enumerate(c):
        ws3.cell(row=row, column=j+1, value=val)
    row += 1
style_header_row(ws3, row - len(exc), 5)
# Color code exception levels
colors = [green_fill, light_fill, yellow_fill, red_fill, gray_fill]
for idx in range(len(exc) - 1):
    for col in range(1, 5):
        ws3.cell(row=row - len(exc) + idx, column=col).fill = colors[idx]

row += 1
ws3.cell(row=row, column=1, value="子模块四：数据映射").font = subtitle_font
row += 1
dm = [
    ["业务概念", "系统来源", "字段名", "数据格式", "获取方式", "就绪状态"],
    ["（填写）", "（填写系统名）", "（填写表名.字段名）", "（VARCHAR/INT/DATE等）", "DB直连/API/Excel导入/函证", "✅ / ⚠️ / ❌"],
    ["（填写）", "（填写）", "（填写）", "（填写）", "（填写）", "（填写）"],
    ["（填写）", "（填写）", "（填写）", "（填写）", "（填写）", "（填写）"],
    ["（填写）", "（填写）", "（填写）", "（填写）", "（填写）", "（填写）"],
    ["（填写）", "（填写）", "（填写）", "（填写）", "（填写）", "（填写）"],
]
for i, c in enumerate(dm):
    for j, val in enumerate(c):
        ws3.cell(row=row, column=j+1, value=val)
    row += 1
style_header_row(ws3, row - len(dm), 6)

# ============================================================
# Sheet 4: 定价与套餐
# ============================================================
ws4 = wb.create_sheet("4-定价与套餐")
set_col_widths(ws4, [5, 16, 20, 18, 16, 20])

write_title(ws4, 1, 1, "定价体系与套餐方案", 6)

row = 3
ws4.cell(row=row, column=1, value="一、分层定价").font = subtitle_font
row += 1
price = [
    ["层级", "服务名称", "交付周期", "团队配置", "参考定价", "付款节点"],
    ["L0", "就绪度诊断", "2周(含1周现场)", "1人(审计骨干+AI辅助)", "3-5万", "启动30%"],
    ["L1", "场景筛选", "1周", "1人", "2-3万", "启动30%"],
    ["L2 ⭐", "场景定义书", "2-3周", "2人(审计+分析)", "5-8万", "启动40%+交付60%"],
    ["L3", "原型验证", "2-4周", "2人(审计+分析+外聘技术顾问)", "8-15万", "启动40%+里程碑40%+验收20%"],
    ["L4", "固化与赋能", "2周", "2人(含2天现场培训)", "3-5万", "启动50%+交付50%"],
]
for i, c in enumerate(price):
    for j, val in enumerate(c):
        ws4.cell(row=row, column=j+1, value=val)
    row += 1
style_header_row(ws4, row - len(price), 6)
ws4.cell(row=row - len(price) + 2, column=5).fill = yellow_fill  # L2 highlight

row += 1
ws4.cell(row=row, column=1, value="二、套餐方案").font = subtitle_font
row += 1
pkgs2 = [
    ["套餐", "包含", "预估总额", "场景数", "付款方式", "推荐场景"],
    ["入门体验", "L0+L1", "5-8万", "不限", "一次性", "首次接触AI审计、预算有限的审计局"],
    ["标准交付 ⭐", "L0→L4全流程", "21-36万", "1个", "30%-40%-30%", "有明确数字化预算、需要一个完整案例"],
    ["多场景扩展", "L0+L1一次 + L2-L4×N", "首场景21-36万 + 增量15-28万/场景", "2-3个", "按场景分签", "首场景验证成功后扩展"],
]
for i, c in enumerate(pkgs2):
    for j, val in enumerate(c):
        ws4.cell(row=row, column=j+1, value=val)
    row += 1
style_header_row(ws4, row - len(pkgs2), 6)

row += 1
ws4.cell(row=row, column=1, value="三、投入产出分析").font = subtitle_font
row += 1
io = [
    ["项目", "金额", "说明"],
    ["内部试点成本", "约3.5万", "自有时间+样板编制+客户沟通"],
    ["首年保守收入", "26-44万", "1个全流程+1个入门"],
    ["首年基准收入", "52-88万", "2个全流程+2个入门"],
    ["ROI(保守)", "7.4x - 12.6x", "投入3.5万 vs 收入26-88万"],
    ["边际成本递减", "第2个场景成本降低30%-50%", "场景定义书模板复用"],
]
for i, c in enumerate(io):
    for j, val in enumerate(c):
        ws4.cell(row=row, column=j+1, value=val)
    row += 1
style_header_row(ws4, row - len(io), 3)

# ============================================================
# Sheet 5: 验证指标
# ============================================================
ws5 = wb.create_sheet("5-验证指标")
set_col_widths(ws5, [5, 22, 20, 25, 20])

write_title(ws5, 1, 1, "场景验证指标体系", 5)

row = 3
ws5.cell(row=row, column=1, value="一、L3原型验证核心指标").font = subtitle_font
row += 1
metrics = [
    ["序号", "指标", "合格标准", "计算方法", "数据来源"],
    ["1", "疑点命中率", "≥30%", "AI标记且人工确认有问题的条数 ÷ AI标记总条数", "验证试跑日志"],
    ["2", "漏报率", "≤10%", "实际存在但AI未识别的条数 ÷ 实际存在总条数", "对比原审计结果"],
    ["3", "效率提升率", "≥50%", "(纯人工耗时 - AI辅助耗时) ÷ 纯人工耗时", "计时对比"],
    ["4", "人工确认耗时", "≤2h/百条疑点", "审阅100条AI产出疑点的人工耗时", "审计人员计时"],
    ["5", "例外覆盖率", "≥90%", "E0+E1+E2方案覆盖率", "例外触发统计"],
    ["6", "用户满意度", "≥4.0/5.0", "审计人员使用后评分", "匿名问卷"],
]
for i, c in enumerate(metrics):
    for j, val in enumerate(c):
        ws5.cell(row=row, column=j+1, value=val)
    row += 1
style_header_row(ws5, row - len(metrics), 5)

row += 1
ws5.cell(row=row, column=1, value="二、验证不通过的处置方案").font = subtitle_font
row += 1
fail = [
    ["不通过情形", "处置方案"],
    ["疑点命中率<20%", "回溯规则定义，检查规则是否过于宽泛/遗漏关键条件 → 修订定义书 → 二次验证(额外1周)"],
    ["漏报率>15%", "补充规则+降低阈值 → 但可能降低命中率，需平衡"],
    ["效率提升<30%", "检查数据准备环节耗时 → 如为数据清洗耗时，L3需追加数据工程"],
    ["连续2次验证不通过", "暂停项目，与审计局沟通：是否继续投入 or 更换场景"],
]
for i, c in enumerate(fail):
    for j, val in enumerate(c):
        ws5.cell(row=row, column=j+1, value=val)
    row += 1
style_header_row(ws5, row - len(fail), 3)

# ============================================================
# Sheet 6: 落地路线图
# ============================================================
ws6 = wb.create_sheet("6-落地路线图")
set_col_widths(ws6, [5, 18, 35, 18, 18, 20])

write_title(ws6, 1, 1, "落地实施路线图", 6)

row = 3
ws6.cell(row=row, column=1, value="一、本周快速验证（MVP）").font = subtitle_font
row += 1
mvp = [
    ["Day", "动作", "具体内容", "参与人", "产出", "完成标准"],
    ["1-2", "内部L0诊断", "选1个融策正在做的审计项目，走一遍L0诊断框架", "项目负责人+审计骨干", "内部就绪度自评", "识别出≥3个「小而痛」的场景"],
    ["3-4", "内部L2定义书", "针对选定的场景，按5个子模块模板填写场景定义书", "审计骨干+AI辅助", "内部场景定义书(初稿)", "5个子模块都有实质内容"],
    ["5", "内部评审", "评审：这份定义书能不能指导实际工作？", "融策平头哥+核心团队", "评审意见", "明确识别1-2处改进点"],
    ["6-7", "总结决策", "根据评审结果决定：是否外部推广", "融策平头哥", "GO/NO-GO决定", "形成修改后的方法论V1.1"],
]
for i, c in enumerate(mvp):
    for j, val in enumerate(c):
        ws6.cell(row=row, column=j+1, value=val)
    row += 1
style_header_row(ws6, row - len(mvp), 6)

row += 1
ws6.cell(row=row, column=1, value="二、首单落地时间线").font = subtitle_font
row += 1
timeline = [
    ["阶段", "时间", "关键动作", "里程碑", "前置条件"],
    ["内部验证", "第1周", "完成MVP验证 → 方法论V1.1", "内部GO决定", "无"],
    ["客户探询", "第2-3周", "联系2-3家合作审计局，非正式沟通AI需求", "≥1家表达兴趣", "内部验证通过"],
    ["意向沟通", "第4周", "向意向客户呈报L0+L1入门方案(5-8万)", "获得口头认可", "客户表达兴趣"],
    ["签约L0+L1", "第5-6周", "走政府购买服务流程/签合同", "首单合同签署", "预算确认"],
    ["交付L0+L1", "第6-9周", "按手册执行L0诊断+L1筛选", "交付+验收+收款", "合同签署"],
    ["升级L2-L4", "第10周起", "根据L1筛选结果，提案L2-L4", "全流程升级", "L0+L1验收通过+客户满意"],
]
for i, c in enumerate(timeline):
    for j, val in enumerate(c):
        ws6.cell(row=row, column=j+1, value=val)
    row += 1
style_header_row(ws6, row - len(timeline), 5)

row += 1
ws6.cell(row=row, column=1, value="三、关键窗口期提醒").font = subtitle_font
row += 1
windows = [
    ["窗口", "时间", "动作", "重要性"],
    ["政策窗口", "2026年5月起(已开始)", "三部门发文划19个场景，各级审计局在「知道但不知道怎么干」阶段", "🔴 当前最佳切入期"],
    ["预算窗口", "2026年9月", "各级政府编制次年预算，需在9月前完成意向沟通", "🔴 错过等一年"],
    ["竞争窗口", "2026年6-12月(~6个月)", "需在6个月内拿到2-3个案例建立壁垒", "🟡 窗口期有限"],
    ["标杆窗口", "2026年底前", "争取做1个省级或市级标杆案例，可大幅降低后续获客成本", "🟡 重要杠杆"],
]
for i, c in enumerate(windows):
    for j, val in enumerate(c):
        ws6.cell(row=row, column=j+1, value=val)
    row += 1
style_header_row(ws6, row - len(windows), 4)

# ============================================================
# Sheet 7: 风险与对策
# ============================================================
ws7 = wb.create_sheet("7-风险与对策")
set_col_widths(ws7, [5, 22, 10, 10, 45])

write_title(ws7, 1, 1, "风险评估与对策矩阵", 5)

row = 3
risks = [
    ["序号", "风险", "概率", "影响", "对策"],
    ["1", "审计局说「不需要」", "中", "高", "先做L0免费诊断(半天访谈+简易报告)，让对方看到流程盲区后再谈付费"],
    ["2", "L3验证效果不达预期", "中", "中高", "选容错空间大的场景(辅助初审)；约定「效果不达标仅收L2费用」"],
    ["3", "数据分析人才跟不上", "中", "中", "内部培养+高校兼职双轨；首单可接受效率较低；大模型辅助分析"],
    ["4", "竞品快速跟进", "低", "中", "6个月窗口期内跑出2-3个案例即为壁垒；审计经验不是AI公司短期能复制"],
    ["5", "政府预算周期错过", "中", "低", "6-8月必须做意向沟通；错过则做预算外课题/试点经费路径"],
    ["6", "内部精力不足", "高", "高", "首单控制在L0+L1轻量级(5-8万/2-3周)；不对既有审计项目造成挤压"],
    ["7", "审计局数据基础太差", "高", "中", "L0诊断即识别此风险 → 如实告知客户 → 建议先做数据治理再上AI → 可转数据治理服务"],
]
for i, c in enumerate(risks):
    for j, val in enumerate(c):
        ws7.cell(row=row, column=j+1, value=val)
    row += 1
style_header_row(ws7, row - len(risks), 5)
# Color code risk probability
for r_idx in range(len(risks)):
    prob_cell = ws7.cell(row=row - len(risks) + r_idx, column=3)
    if "高" in str(prob_cell.value):
        prob_cell.font = red_font
    elif "中" in str(prob_cell.value):
        prob_cell.font = orange_font
    else:
        prob_cell.font = green_font

# ============================================================
# Sheet 8: 团队配置
# ============================================================
ws8 = wb.create_sheet("8-团队配置")
set_col_widths(ws8, [5, 18, 15, 25, 25, 18])

write_title(ws8, 1, 1, "交付团队配置与能力要求", 6)

row = 3
team = [
    ["角色", "来源", "L0", "L1", "L2", "L3", "L4"],
    ["项目负责人(CPA/高级审计师)", "融策自有", "✅主导", "✅主导", "✅审核", "✅审核", "✅审核"],
    ["审计业务专家", "融策自有", "✅执行", "✅执行", "✅主导", "✅参与", "✅参与"],
    ["数据分析师", "融策自有/外聘", "—", "✅辅助", "✅执行", "✅主导", "—"],
    ["AI技术顾问", "合作高校/兼职", "—", "—", "—", "✅辅助", "—"],
]
for i, c in enumerate(team):
    for j, val in enumerate(c):
        ws8.cell(row=row, column=j+1, value=val)
    row += 1
style_header_row(ws8, row - len(team), 7)

row += 1
ws8.cell(row=row, column=1, value="最小可交付团队：1人完成L0+L1 / 2人完成L0-L4全流程").font = bold_font

# ============================================================
# 保存
# ============================================================
output_path = r"D:\openclaw-workspace\output\融策AI审计场景定义-产品手册.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
