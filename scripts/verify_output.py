import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\scrccpa\Desktop\健康照护师-成本构成测算-v3-2026.06.05.xlsx')

print("=== 人工费用明细 验证 ===")
ws = wb['人工费用明细']
for r in [1,2,3]:
    print(f'R{r}: {[ws.cell(row=r,c).value for c in range(1,9)]}')
for r in range(5, 25):
    c1 = ws.cell(row=r, column=1).value
    c2 = ws.cell(row=r, column=2).value
    c6 = ws.cell(row=r, column=6).value
    c7 = ws.cell(row=r, column=7).value
    c8 = ws.cell(row=r, column=8).value
    if c1 or c2:
        print(f'R{r}: seq={c1}, label={c2}, total={c6}, pp={c7}, note={c8}')

print("\n=== 成本构成总览 验证 ===")
ws3 = wb['成本构成总览']
print(f'R13 人工: {[ws3.cell(row=13,c).value for c in range(4,8)]}')
print(f'R14 SP: {[ws3.cell(row=14,c).value for c in range(3,7)]}')
print(f'R16 耗材: {[ws3.cell(row=16,c).value for c in range(4,8)]}')
print(f'R21 实际小计: {[ws3.cell(row=21,c).value for c in range(4,8)]}')
print(f'R24 全口径: {[ws3.cell(row=24,c).value for c in range(4,8)]}')
print(f'R29 差额: {[ws3.cell(row=29,c).value for c in range(4,8)]}')

print("\n=== 耗材成本明细 三级验证 ===")
ws2 = wb['耗材成本明细']
print(f'R81(隔离衣): {[ws2.cell(row=81,c).value for c in range(1,11)]}')
print(f'R110(瞳孔笔): {[ws2.cell(row=110,c).value for c in range(1,11)]}')
print(f'R117(毛巾): {[ws2.cell(row=117,c).value for c in range(1,11)]}')
print(f'R121(袜子): {[ws2.cell(row=121,c).value for c in range(1,11)]}')
print(f'R122(合计): {[ws2.cell(row=122,c).value for c in range(1,11)]}')
