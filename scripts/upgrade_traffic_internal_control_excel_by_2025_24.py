from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path

p = Path(r'outputs\交通厅内控项目\四川省交通运输厅2026年内部控制评价指标体系及监督检查模板.xlsx')
wb = load_workbook(p)

navy='0A1F3F'; teal='1A5C6E'; gold='C5955C'; warm='F5F2EC'; white='FFFFFF'; light_gold='FFF2CC'
thin=Side(style='thin', color='B7B7B7'); medium=Side(style='medium', color=gold)
border=Border(left=thin,right=thin,top=thin,bottom=thin)

def title(ws, text, last_col):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=last_col)
    c=ws.cell(1,1,text); c.fill=PatternFill('solid', fgColor=navy)
    c.font=Font(name='微软雅黑', size=15, bold=True, color=white)
    c.alignment=Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height=32

def header(ws,row,cols):
    for i,h in enumerate(cols,1):
        c=ws.cell(row,i,h); c.value=h; c.fill=PatternFill('solid', fgColor=teal)
        c.font=Font(name='微软雅黑', size=10, bold=True, color=white)
        c.alignment=Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border=Border(left=thin,right=thin,top=medium,bottom=medium)

def style(ws):
    ws.sheet_view.showGridLines=False
    for row in ws.iter_rows():
        for cell in row:
            cell.font=Font(name='微软雅黑', size=10, bold=cell.font.bold, color=cell.font.color)
            cell.alignment=Alignment(vertical='center', wrap_text=True)
            cell.border=border
    ws.page_setup.orientation='landscape'; ws.page_setup.paperSize=ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage=True; ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0

def widths(ws, mapping):
    for col,w in mapping.items(): ws.column_dimensions[col].width=w
    for r in range(1,ws.max_row+1): ws.row_dimensions[r].height=max(ws.row_dimensions[r].height or 18, 24)

# 1 Add / replace legal basis sheet
if '财会2025-24依据摘录' in wb.sheetnames:
    del wb['财会2025-24依据摘录']
ws = wb.create_sheet('财会2025-24依据摘录', 1)
cols=['条款','核心要求','对本项目的落地口径','对应工作表/成果']
title(ws,'《行政事业单位内部控制评价办法》（财会〔2025〕24号）关键依据摘录与项目映射',len(cols)); header(ws,3,cols)
rows=[
['第4条','遵循全面性、重要性、客观性、可操作性、持续性原则。','指标覆盖组织层面、业务层面、内部监督；重点关注权力集中、资金密集、资源富集领域；评分以证明材料和底稿为依据；形成评价—整改—应用闭环。','指标评分细则、问题管理台账、操作流程'],
['第5条','财政部发布年度基本指标体系和报告格式；财政部门组织实施同级部门评价监督检查。','本模板作为交通厅细化指标，应在财政年度基本指标发布后再做一次映射校准。','编制说明、评分细则'],
['第6条','各部门在基本指标体系基础上细化本部门及所属单位指标，并逐级复核所属单位评价报告，出具复核意见。','分别设置厅本级指标、厅属单位指标和复核意见/评分汇总。','厅本级指标、厅属单位指标、评分汇总表'],
['第7条','行政事业单位是内部控制评价责任主体，单位负责人对评价报告真实性、完整性负责。','问题台账和证明材料清单中明确单位负责人、责任部门、材料真实性责任。','工作机制、证明材料清单'],
['第8条','妥善保管评价文件资料，确保评价数据安全和保密。','增加档案管理、保密与独立性机制。','工作机制建议方案'],
['第9条','鼓励采用信息化手段，必要时借助外部专家力量。','采用Excel台账、数据质量检查、系统权限核验、外部第三方辅助检查。','操作流程、检查底稿'],
['第10条/第21条','评价部门或岗位应与内部控制建设牵头部门或岗位相互分离。','工作机制中明确评价组织实施岗位与内控建设牵头岗位分离，防止自建自评。','工作机制建议方案、组织层面指标'],
['第13条','单位评价内容包括组织层面、业务层面、内部监督。','三大维度完整进入指标体系。','厅属单位指标评分细则'],
['第14条','综合运用询问访谈、调查问卷、专题讨论、穿行测试、重新执行、实地查验、抽样和比较分析等方法，充分收集有效证据。','检查方法列明访谈、穿行测试、重新执行、实地查验、抽样、比较分析和数据分析。','检查底稿模板'],
['第15条','根据细化后的评价指标体系打分，形成自评得分；奖惩情况可作调整得分参考。','评分汇总表保留自评分、复核分差异，可增设奖惩调整事项。','评分汇总表'],
['第16条','评价结果划分优、良、中、差：90-100优，80-90良，60-80中，60以下差；存在内控相关违法违纪行为的，最终不得高于“中”。','已将等级公式调整为优/良/中/差，并加入违法违纪下调提示。','评分汇总表、编制说明'],
['第17条','评价全过程应详细记录，包括方法、指标内容、得分、扣分原因、评价结果、认定依据、评价人员等。','底稿模板字段覆盖全过程记录要求。','检查底稿模板'],
['第18条','单位编制评价报告并提交主管部门复核，同时提供证明材料；报告包括组织实施、程序方法、依据、得分、结果、问题及整改等。','监督检查报告框架和证明材料清单按此配置。','报告框架、证明材料清单'],
['第19条','主管部门复核发现偏差应调整得分，形成复核意见并反馈；复核意见包括自评得分、复核调整得分及原因、最终得分、评价结果、问题及整改时限。','新增复核意见要素列，评分汇总表体现自评与复核差异。','评分汇总表、报告框架'],
['第22条','部门评价至少包括所属单位三大层面整体情况、部门层面内控体系建立实施及指导监督情况、行业特点评价内容。','厅本级指标加入交通运输行业特点：交通专项资金、政府采购、合同、项目资金、资产等。','厅本级指标评分细则'],
['第28-29条','财政部门监督包括三大层面情况、组织实施和结果应用、其他要求；可查阅资料、现场核查并反馈结果。','本项目监督检查采用资料审阅+现场核查+反馈调整路径。','操作流程、报告框架'],
['第30条','不得提供虚假评价资料，负责人不得授意编制虚假报告，违规依法追责。','证明材料真实性承诺、虚假材料列重大问题。','证明材料清单、问题台账'],
['第32-34条','建立问题整改台账，跟踪整改；结果用于制度完善、预算管理、绩效管理、财会监督、考核问责、干部选任参考；与财会监督、巡视巡察、纪检监察、审计监督贯通。','问题管理台账、整改验证、结果运用机制和监督协同指标均已设置。','问题台账、工作机制、内部监督指标'],
['第38条','办法自2026年1月1日起施行。','交通厅2026年内部控制建设项目适用该办法。','编制说明'],
]
for r,row in enumerate(rows,4):
    for c,v in enumerate(row,1): ws.cell(r,c,v)
style(ws); widths(ws, {'A':18,'B':45,'C':55,'D':35})

# 2 Update 编制说明
ws = wb['编制说明']
for row in ws.iter_rows(min_row=4, max_col=6):
    if row[0].value == '编制依据':
        row[2].value = '已根据用户提供的财会〔2025〕24号原文校准：评价原则、责任主体、评价内容、评价方法、档次划分、复核意见、问题整改台账和结果应用。'
        row[5].value = '后续若财政部发布年度基本评价指标体系，可再做年度指标映射。'
    if row[0].value == '评价维度':
        row[2].value = '总分100分，评价结果按办法第16条划分为优、良、中、差；存在内控相关违法违纪行为的，最终结果不得高于“中”。'
style(ws)

# 3 Update scoring summary formulas and headers
ws = wb['评分汇总表']
# add columns if not present
headers=[ws.cell(3,c).value for c in range(1,ws.max_column+1)]
extras=['奖惩调整分','违法违纪下调','最终得分','最终档次','复核调整原因','复核意见反馈日期']
for h in extras:
    if h not in headers:
        ws.cell(3, ws.max_column+1, h)
header(ws,3,[ws.cell(3,c).value for c in range(1,ws.max_column+1)])
# find cols
hdr={ws.cell(3,c).value:c for c in range(1,ws.max_column+1)}
for r in range(4,34):
    ws.cell(r,hdr['总分'], f'=SUM(C{r}:E{r})')
    ws.cell(r,hdr['等级'], f'=IF(F{r}>=90,"优",IF(F{r}>=80,"良",IF(F{r}>=60,"中","差")))')
    ws.cell(r,hdr['奖惩调整分'], 0)
    ws.cell(r,hdr['违法违纪下调'], '否')
    final_col=get_column_letter(hdr['最终得分']); total_col=get_column_letter(hdr['总分']); adj_col=get_column_letter(hdr['奖惩调整分'])
    illegal_col=get_column_letter(hdr['违法违纪下调']); final_grade_col=get_column_letter(hdr['最终档次'])
    ws.cell(r,hdr['最终得分'], f'=MAX(0,MIN(100,{total_col}{r}+{adj_col}{r}))')
    ws.cell(r,hdr['最终档次'], f'=IF({illegal_col}{r}="是",IF({final_col}{r}>=60,"中","差"),IF({final_col}{r}>=90,"优",IF({final_col}{r}>=80,"良",IF({final_col}{r}>=60,"中","差"))))')
# validations
for dv in list(ws.data_validations.dataValidation):
    pass
dv=DataValidation(type='list', formula1='"是,否"', allow_blank=True); ws.add_data_validation(dv); dv.add(f"{get_column_letter(hdr['违法违纪下调'])}4:{get_column_letter(hdr['违法违纪下调'])}200")
style(ws); widths(ws, {'A':28,'B':16,'C':14,'D':14,'E':14,'F':10,'G':10,'H':10,'I':12,'J':30,'K':14,'L':14,'M':24,'N':12,'O':12,'P':12,'Q':12,'R':28,'S':18})

# 4 Update 工作机制 with separation and false report accountability if not already
ws=wb['1工作机制建议方案']
existing=[ws.cell(r,1).value for r in range(4,ws.max_row+1)]
add=[
['评价岗位分离机制','评价部门与建设牵头部门分离','依据办法第10条、第21条，内部控制评价部门/岗位应与内部控制建设牵头部门/岗位相互分离，防止自建自评、自评自证。','厅内部审计部门或指定评价岗位','内控建设牵头部门、厅属单位','评价启动前确认','岗位分工文件、回避声明','评价独立性不足会削弱复核结论','工作机制建议方案'],
['真实性责任机制','评价资料真实性承诺','依据办法第7条、第30条，行政事业单位是评价责任主体，单位负责人对评价报告真实性、完整性负责；不得提供虚假资料。','被评价单位负责人','评价联系人、资料提供部门','资料提交时','真实性承诺书、资料移交清单','虚假资料应列重大问题并提示追责风险','证明材料清单/问题台账'],
]
start=ws.max_row+1
for row in add:
    for c,v in enumerate(row,1): ws.cell(start,c,v)
    start+=1
style(ws)

# 5 Update report framework with复核意见
ws=wb['4监督检查报告框架']
append_rows=[
['二、总体评价','复核意见要素','按办法第19条列示自评得分、复核调整得分及调整原因、最终得分和评价结果、发现问题及整改时限。','评分汇总表、复核底稿','复核意见应逐单位反馈，避免只给总体结论。','','财会〔2025〕24号第19条'],
['三、检查发现问题','虚假评价资料风险','对明显与事实不符、证明材料伪造或负责人授意虚假报告的情形，单独列示并提示责任追究风险。','证明材料、访谈记录、底稿','此类问题应作为重大缺陷处理。','','财会〔2025〕24号第30条'],
['五、整改建议','结果应用建议','将评价结果用于制度完善、预算管理、绩效管理、财会监督、考核问责、干部选任参考，并与审计、纪检监察、巡视巡察贯通。','问题台账、结果通报','体现评价—整改—应用—优化闭环。','','财会〔2025〕24号第32-34条'],
]
start=ws.max_row+1
for row in append_rows:
    for c,v in enumerate(row,1): ws.cell(start,c,v)
    start+=1
style(ws)

# 6 Add evaluation report template sheet
if '内控评价报告要素' in wb.sheetnames:
    del wb['内控评价报告要素']
ws=wb.create_sheet('内控评价报告要素')
cols=['报告类型','章节/要素','具体内容','依据条款','取数/资料来源','备注']
title(ws,'内部控制评价报告与复核意见要素模板',len(cols)); header(ws,3,cols)
rows=[
['单位内部控制评价报告','评价工作组织实施情况','说明评价部门/岗位、人员组织、评价范围、时间安排、审批程序。','第12条、第18条','工作方案、审批记录','评价岗位应与建设牵头岗位分离'],
['单位内部控制评价报告','评价程序和方法','说明询问访谈、调查问卷、专题讨论、穿行测试、重新执行、实地查验、抽样、比较分析等方法。','第14条、第18条','检查底稿、访谈记录','方法要与证据对应'],
['单位内部控制评价报告','评价依据','列示财会〔2025〕24号、年度基本指标体系、厅细化指标、单位制度等。','第18条','依据文件清单',''],
['单位内部控制评价报告','评价得分和结果','列示组织层面、业务层面、内部监督得分，总分及优/良/中/差档次。','第15条、第16条','评分表','存在违法违纪的最终不得高于中'],
['单位内部控制评价报告','发现问题及整改情况','列示问题事实、整改措施、整改时限、责任部门和佐证资料。','第18条、第32条','问题台账、整改资料',''],
['主管部门复核意见','自评得分','列示单位报送自评得分及自评档次。','第19条','单位自评表',''],
['主管部门复核意见','复核调整得分及原因','列示复核调增/调减分值、调整依据和原因。','第19条','复核底稿、证明材料',''],
['主管部门复核意见','最终得分和评价结果','明确经复核后的最终得分和优/良/中/差档次。','第19条','评分汇总表',''],
['主管部门复核意见','发现问题及整改时限','逐项反馈复核发现问题、整改要求和完成时限。','第19条','问题台账','单位收到后补充整改措施'],
['部门内部控制评价报告','部门本级及所属单位整体情况','汇总组织层面、业务层面、内部监督整体情况。','第22条、第23条','评分汇总、单位报告',''],
['部门内部控制评价报告','部门层面体系及指导监督','说明厅对本级及所属单位内控指导、监督、复核和结果应用情况。','第22条','工作机制、复核意见',''],
['部门内部控制评价报告','行业特点评价内容','结合交通运输行业项目资金、政府采购、资产、合同、建设项目等特点评价。','第22条','厅本级指标、项目底稿',''],
]
for r,row in enumerate(rows,4):
    for c,v in enumerate(row,1): ws.cell(r,c,v)
style(ws); widths(ws, {'A':22,'B':24,'C':52,'D':18,'E':30,'F':28})

# Save new version
out = p.with_name('四川省交通运输厅2026年内部控制评价指标体系及监督检查模板_财会2025-24校准版.xlsx')
wb.save(out)
print(out.resolve())
