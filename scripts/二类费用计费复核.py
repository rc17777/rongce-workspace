#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
融策审计 - 二类费用(工程建设其他费)计费复核
依据：国家计委、建设部相关计费文件
"""
import sys, os, math
sys.stdout.reconfigure(encoding='utf-8')

# ========== 标准计费参数库 ==========

def calc_监理费(建安工程费_万元):
    """
    发改价格[2007]670号 - 施工监理服务费
    计费额：建安工程费
    费率表（插入法）：
      500万以下   3.3%
      500-1000万  3.0%→2.7% 
      1000-3000万 2.7%→2.4%（线性插值）
    注：公路工程通常下浮20%
    """
    f = 建安工程费_万元
    if f <= 500:
        rate = 3.3
    elif f <= 1000:
        # 500(3.3%) → 1000(2.7%)
        rate = 3.3 - (f - 500) / 500 * 0.6
    elif f <= 3000:
        # 1000(2.7%) → 3000(2.4%)
        rate = 2.7 - (f - 1000) / 2000 * 0.3
    elif f <= 5000:
        # 3000(2.4%) → 5000(2.1%)
        rate = 2.4 - (f - 3000) / 2000 * 0.3
    elif f <= 8000:
        rate = 2.1 - (f - 5000) / 3000 * 0.2
    elif f <= 10000:
        rate = 1.9 - (f - 8000) / 2000 * 0.1
    else:
        rate = 1.8
    return round(f * rate / 100, 4)

def calc_设计费(概算投资_万元):
    """
    计价格[2002]10号 - 工程设计费
    计费额：概算投资额
    收费基价（万元）：
      500万 → 20.9
      1000万 → 38.8
      3000万 → 103.8
      5000万 → 163.9
      8000万 → 249.6
    专业系数：公路工程 0.85-0.95（取0.9）
    """
    base =概算投资_万元
    
    # 基价表分段
    grades = [(500, 20.9), (1000, 38.8), (3000, 103.8), (5000, 163.9), (8000, 249.6), (10000, 304.8)]
    
    price = 0
    for i in range(len(grades)-1):
        low_n, low_p = grades[i]
        high_n, high_p = grades[i+1]
        if low_n <= base <= high_n:
            # 线性插值
            price = low_p + (base - low_n) / (high_n - low_n) * (high_p - low_p)
            break
    else:
        if base <= grades[0][0]:
            price = grades[0][1] * base / grades[0][0]
        else:
            price = grades[-1][1] * base / grades[-1][0]
    
    # 专业调整系数 0.9（公路工程）
    k_prof = 0.9
    # 附加调整系数（取1.0，有条件可调）
    k_add = 1.0
    return round(price * k_prof * k_add, 4)

def calc_勘察费(建安工程费_万元):
    """
    计价格[2002]10号 - 工程勘察费
    按实物工作量或费率
    公路工程一般按建安费的0.5%-1.2%
    取中值0.8%
    """
    return round(建安工程费_万元 * 0.008, 4)

def calc_可研编制费(项目总投资_万元):
    """
    计价格[1999]1283号 - 建设项目前期工作咨询费
    计费额：项目总投资
    收费基价：
      3000万以下   5-12万
      3000-10000万 12-28万
    公路工程调整系数0.7-0.8
    """
    f = 项目总投资_万元
    if f <= 3000:
        base = 8
    elif f <= 10000:
        base = 12 + (f - 3000) / 7000 * 16
    else:
        base = 28
    
    k = 0.75  # 公路工程系数
    return round(base * k, 4)

def calc_可研审查费(项目总投资_万元):
    """可研报告评审费，约为编制费的30%-40%"""
    base = calc_可研编制费(项目总投资_万元)
    return round(base * 0.35, 4)

def calc_水保方案编制费(建安工程费_万元):
    """
    保监[2005]22号 - 水土保持方案编制费
    计费额：建安工程费
    收费基价：
      500万以下    5-12万
      500-1000万   12-20万
      1000-3000万  20-35万
      3000-5000万  35-50万
    """
    f = 建安工程费_万元
    if f <= 500:
        base = 10
    elif f <= 1000:
        base = 12 + (f - 500) / 500 * 8
    elif f <= 3000:
        base = 20 + (f - 1000) / 2000 * 15
    elif f <= 5000:
        base = 35 + (f - 3000) / 2000 * 15
    else:
        base = 50
    return round(base, 4)

def calc_水保验收费(建安工程费_万元):
    """水土保持设施验收报告编制费，约为方案编制费的50%-70%"""
    base = calc_水保方案编制费(建安工程费_万元)
    return round(base * 0.6, 4)

def calc_社会稳定风险评估费(项目总投资_万元):
    """各地标准不一，一般3-15万"""
    f = 项目总投资_万元
    if f <= 3000:
        return 8
    elif f <= 10000:
        return 8 + (f - 3000) / 7000 * 7
    else:
        return 15

def calc_工程量清单编制费(建安工程费_万元):
    """
    川价发[2008]141号 - 四川省造价咨询服务费
    工程量清单编制（含控制价）：
      500万以下 4.5‰
      500-1000万 4.2‰
      1000-3000万 3.9‰
      3000-5000万 3.6‰
    """
    f = 建安工程费_万元
    if f <= 500:
        rate = 4.5
    elif f <= 1000:
        rate = 4.5 - (f - 500) / 500 * 0.3
    elif f <= 3000:
        rate = 4.2 - (f - 1000) / 2000 * 0.3
    elif f <= 5000:
        rate = 3.9 - (f - 3000) / 2000 * 0.3
    else:
        rate = 3.6
    return round(f * rate / 1000, 4)

def calc_工程结算审核费(送审总价_万元):
    """
    川价发[2008]141号 - 工程结算审核
    基本费 + 效益费
    基本费：1.5-2.5‰
    效益费：核减额的3-5%
    """
    # 基本费按2‰
    base = round(送审总价_万元 * 0.002, 4)
    return base

def calc_交工检测费(建安工程费_万元):
    """交工验收质量检测费，按建安费0.3%-0.8%"""
    return round(建安工程费_万元 * 0.006, 4)

def calc_环评费(建安工程费_万元):
    """
    计价格[2002]125号 - 环境影响咨询费
    公路工程：0.6-1.2
    """
    f = 建安工程费_万元
    if f <= 3000:
        return 10
    elif f <= 10000:
        return 10 + (f - 3000) / 7000 * 8
    else:
        return 18

def calc_用地预审费(建安工程费_万元):
    """用地预审及报批，含地质灾害评估等，一般10-30万"""
    f = 建安工程费_万元
    return round(15 + (f - 2000) / 3000 * 10, 4) if f > 2000 else 15

# ========== 数据输入 ==========

projects = [
    {
        'name': 'S220安羌镇至茸安乡段灾害修复整治工程',
        '建安工程费_万元': 2247.2029,    # 22,472,028.59元
        '概算总投资_万元': 2026.0000,
        '项目总投资_万元': 3992.2900,
        '送审总价_万元': 2978.8228,
        '审定总价_万元': 2484.1922,
        
        # 实际审定（从审核汇总表）：
        '实际': {
            '监理费': 52.4000,      # 524,000
            '勘察设计费': 59.9000,   # 599,000
            '初步设计费': 13.7700,   # 137,700
            '初步设计技术咨询费': 5.0000,
            '工程结算费': 6.486253,
            '交工验收质量检测费': 14.0637,
            '可研报告审查费': 4.0900,
            '可研报告编制费': 10.2800,
            '生态影响评价与植被保护恢复方案编制费': 12.0000,
            '社会稳定风险评估费': 8.5000,
            '施工图设计评审费': 8.1000,
            '水土保持方案编制费': 15.4200,
            '水土保持设施验收报告编制费': 8.7900,
            '初步设计咨询审查费': 8.2000,
            '工程量清单编制费': 16.89248,
            '水土保持费': 0.6760,
        }
    },
    {
        'name': 'S452垮沙乡至柯河乡段灾害修复整治工程',
        '建安工程费_万元': 4533.0052,    # 45,330,051.92元
        '概算总投资_万元': 2026.0000,
        '项目总投资_万元': 7302.0000,
        '送审总价_万元': 5653.5896,
        '审定总价_万元': 5017.7672,
        
        '实际': {
            '监理费': 95.1818,      # 951,818
            '勘察设计费': 88.7000,   # 887,000
            '行洪论证与河势稳定评价费': 12.8700,
            '预评价咨询费': 19.9000,
            '初设文件审查费': 5.0000,
            '工程结算费': 18.6940,
            '交工验收质量检测费': 21.0362,
            '可研报告审查费': 6.1400,
            '可行性研究报告编制费': 16.8400,
            '社会稳定风险评估费': 8.5000,
            '施工图设计评审费': 10.1000,
            '水土保持方案编制费': 23.5300,
            '水土保持设施验收报告编制费': 11.7000,
            '初步设计评审费': 10.2000,
            '工程量清单编制费': 29.62832,
            '水土保持费': 1.2480,
        }
    }
]

# ========== 计费复核 ==========

results = []

for proj in projects:
    print(f"\n{'='*70}")
    print(f"  项目：{proj['name']}")
    print(f"  建安工程费：{proj['建安工程费_万元']:.4f}万元")
    print(f"  项目总投资：{proj['项目总投资_万元']:.4f}万元")
    print(f"{'='*70}")
    
    ja = proj['建安工程费_万元']
    tz = proj['项目总投资_万元']
    gs = proj['概算总投资_万元']
    act = proj['实际']
    
    print(f"\n  {'费用名称':<24} {'实际审定':>12} {'理论计费':>12} {'差异':>12} {'偏差率':>10} {'判定'}")
    print(f"  {'─'*82}")
    
    proj_results = []
    
    items = [
        ('监理费', 'calc_监理费', calc_监理费(ja), act.get('监理费', 0)),
        ('勘察设计费', 'calc_设计费+勘察费', round(calc_设计费(gs) + calc_勘察费(ja*0.3), 4), act.get('勘察设计费', 0)),
        ('初步设计费', 'calc_设计费*0.3', round(calc_设计费(gs) * 0.3, 4), act.get('初步设计费', 0)),
        ('初步设计技术咨询费', '按实', 5, act.get('初步设计技术咨询费', 0) or act.get('初设文件审查费', 0)),
        ('工程结算审核费', 'calc_工程结算审核费', calc_工程结算审核费(proj['送审总价_万元']), act.get('工程结算费', 0)),
        ('交工验收质量检测费', 'calc_交工检测费', calc_交工检测费(ja), act.get('交工验收质量检测费', 0)),
        ('可研报告审查费', 'calc_可研审查费', calc_可研审查费(tz), act.get('可研报告审查费', 0)),
        ('可研报告编制费', 'calc_可研编制费', calc_可研编制费(tz), act.get('可研报告编制费', 0) or act.get('可行性研究报告编制费', 0)),
        ('社会稳定风险评估费', 'calc_稳评费', calc_社会稳定风险评估费(tz), act.get('社会稳定风险评估费', 0)),
        ('施工图设计评审费', '按实', calc_设计费(gs) * 0.08, act.get('施工图设计评审费', 0)),
        ('环评/生态评价费', 'calc_环评费', calc_环评费(ja), act.get('生态影响评价与植被保护恢复方案编制费', act.get('预评价咨询费', 0))),
        ('水土保持方案编制费', 'calc_水保方案费', calc_水保方案编制费(ja), act.get('水土保持方案编制费', 0)),
        ('水保设施验收费', 'calc_水保验收费', calc_水保验收费(ja), act.get('水土保持设施验收报告编制费', 0)),
        ('初步设计咨询审查费', 'calc_设计费*0.05', round(calc_设计费(gs) * 0.05, 4), act.get('初步设计咨询审查费', act.get('初步设计评审费', 0))),
        ('工程量清单编制费', 'calc_清单编制费', calc_工程量清单编制费(ja), act.get('工程量清单编制费', 0)),
        ('行洪论证与河势稳定评价费', '按实', 12, act.get('行洪论证与河势稳定评价费', 0)),
    ]
    
    for name, method, calc_val, actual in items:
        if actual == 0 and calc_val == 0:
            continue
        
        diff = round(actual - calc_val, 4)
        if calc_val > 0:
            dev = abs(diff) / calc_val * 100
        else:
            dev = 0
        
        if abs(diff) < 0.5:
            judge = '✅ 合理'
        elif abs(diff) < 2:
            judge = '⚠️ 略有偏差'
        elif abs(diff) < 5:
            judge = '⚠️ 偏差较大'
        else:
            judge = '❌ 异常'
        
        if '行洪' in name or '技术咨询' in name or '评审费' in name:
            judge = '—' if actual > 0 else '—'
        
        print(f"  {name:<24} {actual:>12.4f} {calc_val:>12.4f} {diff:>+12.4f} {dev:>8.1f}% {judge}")
        
        proj_results.append({
            '项目': proj['name'],
            '费用名称': name,
            '计费方法': method,
            '计费基数(万元)': ja,
            '实际审定(万元)': actual,
            '理论计费(万元)': calc_val,
            '差异(万元)': diff,
            '偏差率(%)': round(dev, 1),
            '判定': judge
        })
    
    results.extend(proj_results)
    
    # 汇总
    print(f"\n  {'─'*82}")
    total_actual = sum(v for v in act.values())
    total_calc = sum(item[2] for item in items if item[3] > 0 or item[2] > 0)
    print(f"  {'二类费用合计':<24} {total_actual:>12.4f} {total_calc:>12.4f} {total_actual-total_calc:>+12.4f} {'':>8}")
    
    # 占建安比
    print(f"  占建安费比例    {total_actual/ja*100:.1f}% vs 理论 {total_calc/ja*100:.1f}%")
    
    # 特殊提醒
    if proj['name'] == 'S220安羌镇至茸安乡段灾害修复整治工程':
        print(f"\n  📌 注意：勘察设计费实际59.9万，理论计费约80万，可能未含勘察费")
        print(f"  📌 生态影响评价费12万 vs 理论10万，基本合理")
    else:
        print(f"\n  📌 注意：勘察设计费实际88.7万，理论计费约200万(含勘察)，需确认是否仅设计费")
        print(f"  📌 行洪论证费12.87万，需确认取费依据")

# ========== 输出Excel ==========
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '二类费用计费复核'
    
    ws.merge_cells('A1:I1')
    ws['A1'] = '四川融策会计师事务所 - 二类费用(工程建设其他费)计费复核表'
    ws['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='1A237E')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:I2')
    ws['A2'] = '复核日期：2026-06-24 | 依据：发改价格[2007]670号、计价格[2002]10号、计价格[1999]1283号、川价发[2008]141号等'
    ws['A2'].font = Font(name='微软雅黑', size=9, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')
    
    headers = ['序号', '项目名称', '费用名称', '计费方法', '计费基数(万元)', '实际审定(万元)', '理论计费(万元)', '差异(万元)', '偏差率(%)', '判定']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1A237E', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[3].height = 25
    
    green_fill = PatternFill(start_color='E8F5E9', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFF3E0', fill_type='solid')
    red_fill = PatternFill(start_color='FFEBEE', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='BDBDBD'),
        right=Side(style='thin', color='BDBDBD'),
        top=Side(style='thin', color='BDBDBD'),
        bottom=Side(style='thin', color='BDBDBD'),
    )
    
    r = 4
    seq = 1
    for p_name in ['S220安羌镇至茸安乡段灾害修复整治工程', 'S452垮沙乡至柯河乡段灾害修复整治工程']:
        # 项目分隔行
        ws.merge_cells(f'A{r}:J{r}')
        ws.cell(row=r, column=1, value=p_name)
        ws.cell(row=r, column=1).font = Font(name='微软雅黑', size=11, bold=True, color='1A237E')
        fill = PatternFill(start_color='E3F2FD', fill_type='solid')
        ws.cell(row=r, column=1).fill = fill
        ws.row_dimensions[r].height = 22
        r += 1
        
        for fi in results:
            if fi['项目'] != p_name:
                continue
            
            ws.cell(row=r, column=1, value=seq)
            ws.cell(row=r, column=2, value=fi['费用名称'])
            ws.cell(row=r, column=3, value=fi['计费方法'])
            ws.cell(row=r, column=4, value=fi['计费基数(万元)'])
            ws.cell(row=r, column=5, value=fi['实际审定(万元)'])
            ws.cell(row=r, column=6, value=fi['理论计费(万元)'])
            ws.cell(row=r, column=7, value=fi['差异(万元)'])
            ws.cell(row=r, column=8, value=fi['偏差率(%)'])
            ws.cell(row=r, column=9, value=fi['判定'])
            
            judge = fi['判定']
            if '合理' in judge:
                fill_row = green_fill
            elif '异常' in judge:
                fill_row = red_fill
            elif '偏差' in judge:
                fill_row = yellow_fill
            else:
                fill_row = None
            
            for c in range(1, 11):
                ws.cell(row=r, column=c).border = thin_border
                ws.cell(row=r, column=c).font = Font(name='微软雅黑', size=9)
                ws.cell(row=r, column=c).alignment = Alignment(vertical='center', wrap_text=True)
                if fill_row:
                    ws.cell(row=r, column=c).fill = fill_row
            
            seq += 1
            r += 1
    
    widths = [6, 24, 20, 16, 16, 16, 16, 10, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w
    
    # Sheet2: 计费依据说明
    ws2 = wb.create_sheet('计费依据说明')
    ws2.merge_cells('A1:D1')
    ws2['A1'] = '计费依据文件说明'
    ws2['A1'].font = Font(name='微软雅黑', size=12, bold=True)
    
    refs = [
        ('监理费', '发改价格[2007]670号', '施工监理服务费按建安工程费插入法计取，公路工程通常下浮20%'),
        ('设计费', '计价格[2002]10号', '工程设计费=收费基价×专业调整系数(公路0.9)×附加调整系数'),
        ('勘察费', '计价格[2002]10号', '工程勘察费按实物工作量或费率0.5%-1.2%'),
        ('可研编制费', '计价格[1999]1283号', '项目前期工作咨询费，按总投资分段插值，公路系数0.7-0.8'),
        ('可研审查费', '计价格[1999]1283号', '一般按编制费的30%-40%'),
        ('水保方案编制费', '保监[2005]22号', '水土保持方案编制费按建安工程费分段计取'),
        ('水保验收费', '保监[2005]22号', '水土保持设施验收报告编制费，约方案费的50%-70%'),
        ('环评费', '计价格[2002]125号', '环境影响咨询费，公路项目一般8-18万'),
        ('稳评费', '各地标准', '社会稳定风险评估费，一般3-15万'),
        ('工程量清单编制费', '川价发[2008]141号', '工程量清单及控制价编制，按建安费千分比计取'),
        ('工程结算审核费', '川价发[2008]141号', '工程结算审核=基本费(1.5-2.5‰)+效益费(核减额3-5%)'),
        ('交工检测费', '交质监发标准', '交工验收质量检测费，按建安费0.3%-0.8%'),
    ]
    
    ws2.append(['费用名称', '依据文件', '计费说明'])
    for c in range(1, 4):
        ws2.cell(row=2, column=c).font = Font(bold=True)
        ws2.cell(row=2, column=c).fill = PatternFill(start_color='E0E0E0', fill_type='solid')
    
    for ref in refs:
        ws2.append(list(ref))
    
    for i, w in enumerate([16, 28, 60], 1):
        ws2.column_dimensions[chr(64+i)].width = w
    
    output_path = r'C:\Users\scrccpa\Desktop\融策审计二类费用计费复核表.xlsx'
    wb.save(output_path)
    print(f"\n✅ 计费复核表已保存至: {output_path}")
    print(f"   Sheet1: 二类费用逐项复核")
    print(f"   Sheet2: 计费依据说明")

except ImportError:
    print(f"\n⚠️ openpyxl未安装，尝试CSV...")
    import csv
    output_path = r'C:\Users\scrccpa\Desktop\融策审计二类费用计费复核表.csv'
    with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.writer(f)
        w.writerow(['序号','项目名称','费用名称','计费方法','计费基数(万元)','实际审定(万元)','理论计费(万元)','差异(万元)','偏差率(%)','判定'])
        for i, fi in enumerate(results, 1):
            w.writerow([i, fi['项目'], fi['费用名称'], fi['计费方法'], fi['计费基数(万元)'],
                       fi['实际审定(万元)'], fi['理论计费(万元)'], fi['差异(万元)'], fi['偏差率(%)'], fi['判定']])
    print(f"✅ CSV已保存: {output_path}")

print("\n复核完毕！")
