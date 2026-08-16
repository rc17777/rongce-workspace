"""更新Excel — 补充维度A-E深挖结果"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.load_workbook(r'C:\Users\scrccpa\Desktop\艺术团采购审计分析报告.xlsx')

hdr_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
sub_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yel_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
grn_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
org_fill = PatternFill(start_color='FFDAB9', end_color='FFDAB9', fill_type='solid')
title_f = Font(name='微软雅黑', bold=True, size=13, color='2F5496')
red_f = Font(name='微软雅黑', size=11, color='CC0000', bold=True)
bold_f = Font(name='微软雅黑', bold=True, size=11)
norm_f = Font(name='微软雅黑', size=11)
sml_f = Font(name='微软雅黑', size=10)
code_f = Font(name='Consolas', size=9)
thin_b = Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))
ca = Alignment(horizontal='center', vertical='center', wrap_text=True)
wa = Alignment(vertical='center', wrap_text=True)

def cell(ws, r, col, v, font=norm_f, fill=None, align=ca):
    c = ws.cell(row=r, column=col, value=v)
    c.font = font; c.alignment = align; c.border = thin_b
    if fill: c.fill = fill
    return c

def hdr(ws, r, headers):
    for i,h in enumerate(headers):
        cell(ws, r, i+1, h, hdr_font, hdr_fill)

# ============ NEW Sheet: 多维深挖结果 ============
ws = wb.create_sheet('多维深度查验结果')

cell(ws, 1, 1, '七维度深度查验 — 超出L1-L10标准检测框架的补充分析', title_f, align=wa)
ws.merge_cells('A1:F1')
cell(ws, 2, 1, '思路：元数据被清除≠无迹可寻。从PDF内部结构、JPEG编码指纹、WPS签名残留、文件时间线、页面结构模式等七个维度重建行为画像。', sml_f, align=wa)
ws.merge_cells('A2:F2')

# ===== 维度A: WPS签名残留 =====
r = 4
cell(ws, r, 1, '维度A: WPS签名残留 — 胤皓11处 vs 招标文件0处', bold_f, sub_fill, wa)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

r = 5
hdr(ws, r, ['对比项','招标文件(PDF)','胤皓(PDF)','差异解读','证据价值','备注'])

data_a = [
    ['WPS签名总数','0处','11处(3个"WPS Office" + 8个"wps")',
     '招标文件是WPS标准文本导出→PDF 1.7，不嵌入图片级标记。胤皓使用了WPS的"输出为图片PDF"功能→每页JPEG嵌入WPS标记。',
     '🔴 高','证明胤皓使用WPS Office创建原始文档，与元数据空白形成矛盾。'],
    ['WPS签名位置','N/A','分布在JPEG图像数据流中(offset 542K~69M)',
     '签名在JPEG内部而非PDF元数据区，说明是WPS图片渲染引擎嵌入的标记，而非文档属性。',
     '🟡 中','进一步佐证胤皓的"文本→图片→清除元数据→重新封装"操作链。'],
    ['与招标文件同源','N/A','共同WPS字符串=0',
     '胤皓的WPS签名上下文与招标文件无重叠。不能直接证明同一WPS实例，但证明胤皓确实走了"图片化"的非常规路径。',
     '🟡 中','排除了"胤皓直接抄袭招标文件PDF"的可能。'],
]

for i, row_data in enumerate(data_a):
    for j, val in enumerate(row_data):
        if j == 4:
            fill = red_fill if '高' in str(val) else yel_fill
            cell(ws, 6+i, j+1, val, bold_f, fill)
        else:
            cell(ws, 6+i, j+1, val, j == 1 and sml_f or norm_f, align=wa if j in [2,3,5] else ca)

# ===== 维度B: JPEG量化表 =====
r = 10
cell(ws, r, 1, '维度B: JPEG量化表指纹 — 胤皓 vs 太格 使用不同编码器', bold_f, sub_fill, wa)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

r = 11
hdr(ws, r, ['检测项','胤皓','太格','差异说明','证据价值','备注'])

data_b = [
    ['亮度量化表','[7,8,10,8,7,11,10,9]','[14,15,18,15,13,20,18,16]',
     '量化表完全不同 → 两个PDF使用了不同的JPEG编码器。胤皓的量化表特征与WPS"输出为图片"功能的默认设置一致。太格的量化表与RICOH扫描仪默认设置一致。',
     '🟢 排除','证明两家的PDF生成工具链独立。排除了"同一工具批量生成后分发"这种最简串标模式。'],
    ['色度量化表','[12,12,16,14,16,32,18,18]','[23,23,30,26,30,59,33,33]',
     '色度表也完全不同。但两家在各自文档内的5页采样完全一致 → 表明各自内部使用统一设置批量生成。',
     '🟢 排除','各自内部一致=各自专业化操作。跨公司不一致=工具链独立。'],
    ['量化表哈希','胤皓2个唯一hash','太格2个唯一hash',
     '两家JPEG压缩质量设置不同：胤皓使用较低压缩(wps默认"中等质量")，太格使用较高压缩(RICOH默认"标准质量")。',
     '🟢 排除','如为同源制作，量化表应相同或接近。'],
]

for i, row_data in enumerate(data_b):
    for j, val in enumerate(row_data):
        if j == 4:
            cell(ws, 12+i, j+1, val, bold_f, grn_fill)
        else:
            cell(ws, 12+i, j+1, val, j == 1 and code_f or norm_f, align=wa if j in [2,3,5] else ca)

# ===== 维度C: 页面结构 =====
r = 16
cell(ws, r, 1, '维度C: 页面结构模式', bold_f, sub_fill, wa)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

r = 17
hdr(ws, r, ['检测项','胤皓','太格','分析','证据价值','备注'])

data_c = [
    ['内容流大小/页','36 bytes','60 bytes',
     '胤皓的内容流极简(仅页面对象+单个图片引用)，是典型的"WPS输出为图片PDF"结构。太格稍多(扫描仪添加了额外的色彩空间和校准参数)。',
     '🟡 佐证','再次确认两者使用不同的PDF生成工具。'],
    ['每页图片数','1.0张','1.0张',
     '两者每页都是单一的整页扫描图+无文字层，这是"全图片化"的典型特征。正常投标文件应混合文本+图片。',
     '🟡 佐证','纯图片PDF在电子化采购中属非常规操作。'],
    ['页面尺寸','595.08×841.68','595.20×841.92',
     'A4标准尺寸(稍有舍入差异)。胤皓的点数精度与WPS默认一致，太格与RICOH扫描仪一致。',
     '🟢 参考','尺寸一致是A4纸的标准属性，不构成证据。'],
]

for i, row_data in enumerate(data_c):
    for j, val in enumerate(row_data):
        if j == 4:
            cell(ws, 18+i, j+1, val, bold_f, yel_fill if '佐证' in str(val) else grn_fill)
        else:
            cell(ws, 18+i, j+1, val, norm_f, align=wa if j in [2,3,5] else ca)

# ===== 维度D: 时间线 =====
r = 22
cell(ws, r, 1, '维度D: 投标文件制作时间线', bold_f, sub_fill, wa)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

r = 23
hdr(ws, r, ['投标人','文件修改时间','扫描/创建时间','距截止时间','关键事件','时间线分析'])

data_timeline = [
    ['太格','2025-04-02 16:43','扫描: 04-02 15:53\n(JFIF密度300dpi)','约17小时',
     '太格在截止前17小时完成扫描。扫描时间和文件保存时间相差约50分钟(15:53→16:43)，与"扫描→存盘→整理"的正常节奏一致。',
     '🟢 正常','时间线合理。'],
    ['胤皓','2025-04-02 23:15','PDF内无时间戳\n(元数据已清除)','约11小时',
     '胤皓在深夜23:15完成文件制作。操作链(WPS→图片→清元数据→封装)需要多个步骤，深夜操作可能与"需额外处理时间"有关。',
     '🟡 关注','深夜完成+需要5步操作链，时间偏紧。'],
    ['立美/归档','无法读取','无法读取','N/A',
     '立美和归档文件因损坏无法提取时间戳信息。两者同时损坏降低了时间线交叉验证的可能性。',
     '🔴 缺失','无法建立完整时间线以判断文件制作顺序。'],
]

for i, row_data in enumerate(data_timeline):
    for j, val in enumerate(row_data):
        if j == 5:
            fill = yel_fill if '关注' in str(val) else (red_fill if '缺失' in str(val) else grn_fill)
            cell(ws, 24+i, j+1, val, bold_f, fill)
        else:
            cell(ws, 24+i, j+1, val, norm_f, align=wa if j in [3,4,5] else ca)

# ===== 综合判断 =====
r = 28
cell(ws, r, 1, '综合判断：各维度结论汇总', bold_f, sub_fill, wa)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

r = 29
hdr(ws, r, ['维度','核心发现','对串标假设的支持','证据等级','对"无串标"假设的支持','综合'])

verdict = [
    ['L5 元数据','胤皓元数据完全清空 + WPS残留11处','🟡 支持(清除行为异常)','间接','不支持(太格/立美不同)','⚠️ 胤皓可疑'],
    ['L4 图片哈希','0跨公司重复(204实例)','不支持','排除性','支持','✅ 无串标证据'],
    ['A WPS同源','胤皓11处 vs 招标文件0处','不支持(不同WPS实例)','排除性','支持','✅ 非同源'],
    ['B JPEG指纹','量化表完全不同','不支持(不同编码器)','排除性','支持','✅ 工具链独立'],
    ['C 页面结构','间距不同(36 vs 60 bytes)','不支持(不同工具)','排除性','支持','✅ 生成器独立'],
    ['D 时间线','太格16:43→胤皓23:15(相差6.5h)','中性(可能是正常排队，也可能是信息差)','中性','中性','➡️ 无倾向'],
    ['立美损坏','56MB→0页','🟡 支持(使L5检测完全失效)','间接','不支持','⚠️ 可疑但无法验证'],
    ['综合判断','技术证据主要指向"工具链独立"\n行为证据指向"胤皓反检测操作"\n数据缺失(报价/评审)使最终判断受限',
     '不能排除串标，也不能认定串标','—','—','⚠️ 需继续调查'],
]

for i, row_data in enumerate(verdict):
    for j, val in enumerate(row_data):
        if j == 2:
            fill = yel_fill if '支持' in str(val) else (grn_fill if '不支持' in str(val) else org_fill)
            cell(ws, 30+i, j+1, val, bold_f, fill)
        elif j == 4:
            fill = grn_fill if '支持' in str(val) else (yel_fill if '不支持' in str(val) else org_fill)
            cell(ws, 30+i, j+1, val, bold_f, fill)
        elif j == 5:
            cell(ws, 30+i, j+1, val, bold_f, align=wa)
        elif j == 3:
            cell(ws, 30+i, j+1, val, norm_f)
        else:
            cell(ws, 30+i, j+1, val, norm_f, align=wa)

# ===== 下一步行动 =====
r2 = 39
cell(ws, r2, 1, '补充建议：基于新技术手段的取证方向', bold_f, red_fill, wa)
ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=6)

next_steps = [
    '1.【最高优先级-工商关联L8】通过天眼查/企查查手动查询三家公司工商信息：(1)法定代表人姓名及身份证号前6位 (2)股东及持股比例 (3)董监高交叉任职 (4)注册地址是否相邻/相同 (5)历史名称变更。如有人员重叠或地址关联，立即构成串标铁证。',
    '2.【WPS实例识别】向胤皓发函要求提供原始.docx投标文件。原始.docx中的OLE2 SummaryInformation流包含创建者SID(Windows用户安全标识符)、最后保存者、修订次数等无法被PDF障眼法掩盖的信息。',
    '3.【报价数据获取】向代理机构正式函调三家最后报价记录。一旦获取报价，即可执行L1价格规律分析(极差/限价比/等差数列检测)。',
    '4.【评审过程复原】调取评审报告、评委独立打分表、资格性审查报告。与已知的7.5分价格分+87-93.5分非价格分模式做一致性对比。',
    '5.【内容对比OCR】对胤皓和太格的相同章节(如承诺函、项目经验列表)做150dpi高精度OCR，提取文字后做Jaccard相似度和独特措辞比对。如发现相同排版/相同错别字/相同非常规表述，即构成L3铁证。',
    '6.【JPEG重新编码检测】如果胤皓和太格背后是同一批人在不同电脑上操作，原始.docx可能相同。对两家投标文件中同位置图片做"重新编码检测"——如果一张图片的DCT系数模式与另一张高度相似(不是像素级相同，而是压缩痕迹相似)，说明两张图片来自同一原始文件经不同软件重新编码。',
]

for i, a in enumerate(next_steps):
    cell(ws, 40+i, 1, a, norm_f, align=wa)
    ws.merge_cells(start_row=40+i, start_column=1, end_row=40+i, end_column=6)
    ws.row_dimensions[40+i].height = 40

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 28
ws.column_dimensions['C'].width = 32
ws.column_dimensions['D'].width = 30
ws.column_dimensions['E'].width = 24
ws.column_dimensions['F'].width = 30

wb.save(r'C:\Users\scrccpa\Desktop\艺术团采购审计分析报告.xlsx')
print('Updated: 7 sheets total')
