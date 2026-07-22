# -*- coding: utf-8 -*-
import json
from collections import Counter
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(r"C:\Users\scrccpa\Desktop\报告\2.部门预算项目绩效自评复核报告20260722\中国共产党成都市郫都区委员会统一战线工作部2025年部门预算项目绩效自评复核结果.xlsx")

issues = json.loads(r'''
[
["P0","报告内容逻辑","报告正文","第60段","本级项目17个误写为70个--明显照搬民政局模板","原文\"均为区委统战部本级项目70个\"。统战部应自评项目17个,写70个是照搬了民政局(35*2?)或其他模板的数字。","改为\"均为区委统战部本级项目17个\"。","报告第60段;民政局报告第59段同为\"70个\"。"],
["P0","自评复核口径","少数民族发展机动资金","附件7 R11","满意度三级指标为\"受益老年人或其家属对补贴项目的满意度\"--照搬民政局模板","少数民族发展资金满意度写\"受益老年人或其家属\",与本项目(少数民族联谊活动)风马牛不相及。","改为\"少数民族群众对民族团结活动的满意度\"等与本项目实际内容对应的表述。","附件7 R11;民政局报告满意度指标。"],
["P0","自评复核口径","寺观教堂维修及宗教教职人员补助","附件6 R7/R8","两个重量级指标复核完成值均为0","团结清真寺园林打造400\u33a1复核0;友爱竹隐寺/安德天主教堂维修2000\u33a1复核0。两项合计30分(各10+10),项目整体复核得分将被大幅拉低。","在底稿备注写明0分原因(未启动?无资料?资金未拨付?),并在报告正文点名。","附件6 R7/R8。"],
["P0","自评复核口径","10个项目全部","各附件","10个项目自评得分全部为100分,无任何差异","10个项目均为满分自评,无一项目自认不足。对比其他单位(投促自评有98/97,消防自评有27),统战部自评明显缺乏实质自检。","在报告问题分析中增加:10个复核项目自评均为100分,但复核发现多处指标未完成/不具考核性/定义不匹配,说明自评工作流于形式。","各附件合计行;表1自评得分列。"],
["P0","自评复核口径","9个项目的效益指标","附件1R11/2R15/3R12/4R11/5R10/6R11/8R10/9R13等","社会效益指标全部以\u226595%为指标值,复核均标注\"不具考核性\"","9个效益指标(除附件7/10外)全部写\u226595%\"社会效益明显/成效显著\",复核均为\"不具考核性\"。但复核后得分?自评均为满分20分。","统一切分:严重(空洞定量无佐证)\u21920分;中等(有方向无计算公式)\u2192扣60%=8分;轻微(有公式缺来源)\u2192扣30%=6分。","用户口径第3条;各附件效益指标行。"],
["P0","自评复核口径","台事侨务管理服务","附件4 R5","质量指标\"完成率\"复核标注\"指标不匹配\"","质量指标权重10分,\"完成率\"\u226595%,复核完成值\"指标不匹配\"。如认定不匹配,按口径第10条应0分,但复核得分表需确认。","如保留\"完成率\"作为质量指标,应补充具体考核内容;如确不匹配,按0分并在备注写明理由。","用户口径第10条;附件4 R5。"],
["P0","自评复核口径","统战特别费","附件5 R12","成本指标性质为\"\u2265\",与成本控制逻辑相悖","成本指标\"统战特别费\u226517.4万元\"。成本指标用\"\u2265\"意为鼓励多花钱,与严控支出逻辑相反。报告第109段已提到此问题。","改为\"\u2264\",如实际为定额预算则用\"=\";同步修正复核评分依据。","报告第109段;附件5 R12。"],
["P1","自评复核口径","附件2民主党派","附件2 R8/R11","完成值超指标值2倍以上(222%/289%),不扣分但应标注","政治引领9场完成20场、考察9场完成26场,均\u2265130%,不扣分正确。但底稿无备注标注。","在备注补充\"完成值超130%,按规则不扣分\"。","用户口径第5条。"],
["P1","自评复核口径","附件2/附件6/附件7等","多处","质量指标\"完成率\"复核值低于95%","附件1完成率88.75%、附件2完成率87.5%、附件5完成率66.67%、附件6完成率50%、附件7完成率66.67%,均低于指标值\u226595%,应按比例扣分。","已区分实际扣分情况,确认公式正确;在备注中标注实际原因(哪几项任务未完成导致)。","各附件质量指标行。"],
["P1","报告内容逻辑","报告正文","第106段","问题分析称\"社会效益均采用量化形式\",但多处效益指标实际为空洞定性","报告原文\"社会效益、可持续效益指标均采用量化形式\"。但各项目社会效益指标值全是\"\u226595%\"这样一个无测算依据的数字,形式上\"量化\"实质上是\"定性\"。","改为\"形式上定量、实质上无测算依据的空洞量化指标\"。","报告第106段。"],
["P1","报告内容逻辑","报告正文","第112段","\"42个数量指标中5个未完成\"与各项目未对照列出","报告仅给总数,未列项目名称+指标名称+计划值+完成值+未完成原因。","补表列出5个未完成指标明细。","报告第112段;底稿附件6 R7/R8等。"],
["P1","报告与底稿一致性","附件2民主党派","附件2 R7","组织民主党派协助项目招引复核完成值0","指标值3个,复核完成值0。备注需说明0分原因(无招引项目?未提供资料?)。","补充备注:为何为0,有无可核验资料。","附件2 R7。"],
["P1","报告格式","一级标题","第55/61/73/104/113段","一级标题未居中","两端对齐,字体黑体16pt。","设为居中。","Word段落。"],
["P2","报告格式","主标题","第51段","主标题居中,方正小标宋简体","已居中。","统一为2号小标宋加粗。","Word第51段。"],
["P2","报告格式","行距/数字字体","全文","需Word确认","正文仿宋三号基本OK。","统一固定值30磅/Times New Roman。","Word样式。"],
["通过","报告与底稿一致性","部门基本情况","第58段","项目分类与金额勾稽通过","28+4+17=49\u2713;937.65+86.15+721.63=1745.43。","保留。","报告第58段。"],
["通过","自评复核口径","附件10购买计算机","附件10","全部指标基本符合","数量/质量/时效/效益/成本均达标,复核一致。","保留。","附件10。"],
["通过","自评复核口径","多个项目","多处","完成值超130%已正确处理","多次完成值翻倍/数倍不扣分,方向正确。","建议统一在备注标注。","用户口径第5条。"],
["通过","报告格式","二级标题","各处","楷体GB2312加粗","符合。","保留。","Word。"],
["通过","报告格式","层级","全文","不超过三级","一/(一)/1.三级,OK。","保留。","全文。"]
]
''')

headers = ["等级","复核维度","对象","位置","复核发现","事实及判定","处理建议","核验依据"]
wb = Workbook()
ws = wb.active; ws.title = "复核问题清单"
ws.append(["中国共产党成都市郫都区委员会统一战线工作部2025年度部门预算项目绩效自评复核结果"])
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
ws.append(["复核范围:10个项目,102条指标,100%覆盖。含报告格式、底稿一致性、内容逻辑、扣分与14项口径一致性。"])
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
ws.append(headers)
for row in issues: ws.append(row)

navy="0A1F3F";teal="1A5C6E";warm="F5F2EC";thin=Side(style="thin",color="B7C3CC")
ws["A1"].font=Font(name="Microsoft YaHei",size=16,bold=True,color="FFFFFF");ws["A1"].fill=PatternFill("solid",fgColor=navy);ws["A1"].alignment=Alignment(horizontal="center",vertical="center")
ws["A2"].font=Font(name="Microsoft YaHei",size=10,italic=True,color="333333");ws["A2"].fill=PatternFill("solid",fgColor=warm);ws["A2"].alignment=Alignment(wrap_text=True,vertical="center")
for c in ws[3]: c.font=Font(name="Microsoft YaHei",bold=True,color="FFFFFF");c.fill=PatternFill("solid",fgColor=teal);c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
for row in ws.iter_rows(min_row=4,max_row=ws.max_row,min_col=1,max_col=len(headers)):
    cc={"P0":"F4CCCC","P1":"FCE4D6","P2":"FFF2CC","通过":"E2F0D9"}.get(row[0].value,"FFFFFF")
    for cell in row:
        cell.font=Font(name="Microsoft YaHei",size=10);cell.alignment=Alignment(wrap_text=True,vertical="top")
        cell.border=Border(left=thin,right=thin,top=thin,bottom=thin);cell.fill=PatternFill("solid",fgColor=cc)
for i,w in enumerate([10,20,24,30,40,70,65,45],1): ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height=28;ws.row_dimensions[2].height=35;ws.freeze_panes="A4";ws.auto_filter.ref=f"A3:H{ws.max_row}"

sm=wb.create_sheet("复核汇总")
counts=Counter(row[0] for row in issues)
sm.append(["项目","结果"])
for r in [
    ["复核材料","统战部2025年度部门预算项目绩效自评抽查复核报告.docx;复核表-盖章底稿.xlsx"],
    ["发现统计",f"P0:{counts['P0']}项; P1:{counts['P1']}项; P2:{counts['P2']}项; 通过:{counts['通过']}项"],
    ["P0结论","两处明显照搬民政局模板(第60段70个/附件7满意度老年人);寺观教堂两个大额指标0分须补原因;10个项目自评全满分自检流于形式;社会效益指标集体\"不具考核性\"需统一切分;台事侨务质量指标\"不匹配\"须0分闭环;统战特别费成本\"\u2265\"逻辑错误。"],
    ["本次特色","10个项目全部自评100分--下午五份报告中最极端的自评偏乐观案例。效益指标\"形式上定量(\u226595%)本质定性\"的问题最集中。还发现直接复制民政局报告的两处笔误。"],
    ["人工终核","盖章底稿;寺观教堂项目为什么不执行(资金被收回?民众抗拒?);满意度问卷;成本\u2265改为\u2264的政策依据;各项目完成率低于95%的原因。"],
]: sm.append(r)
for c in sm[1]:c.font=Font(name="Microsoft YaHei",bold=True,color="FFFFFF");c.fill=PatternFill("solid",fgColor=navy)
for row in sm.iter_rows():
    for c in row:c.alignment=Alignment(wrap_text=True,vertical="top");c.border=Border(left=thin,right=thin,top=thin,bottom=thin)
sm.column_dimensions["A"].width=22;sm.column_dimensions["B"].width=125;sm.freeze_panes="A2"

wb.save(OUT)
print("OK:",OUT)
print("Counts:",dict(counts))
