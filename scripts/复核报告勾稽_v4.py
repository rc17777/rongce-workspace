"""
融策审计复核 v4 - 最终版，输出规范化Excel
复核维度：
  1. 报告↔附表勾稽（审核报告金额 vs 审核汇总表金额）
  2. 报告内部数据勾稽（概算↔送审↔审定↔核减↔支付↔结余）
"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8')

BASE = r'C:\Users\scrccpa\Desktop\新建文件夹'

# ===== 手动解析关键数据 =====
# 从上面v3输出提取的核心数据

DATA = {
    'S220安羌镇至茸安乡段灾害修复整治工程': {
        '报告': {
            '概算总投资': 2026.0000,
            '项目总投资': 3992.2900,
            '送审金额': 2978.8228,
            '审定金额': 2484.1922,
            '核减额计算': 494.6306,  # 送审-审定
            '建安工程投资': 2247.2029,
            '待摊投资': 244.5684,
            '已支付': 2462.8050,
        },
        '附表1_02财务决算表': {
            '基建拨款': 26900000.00,  # 元
            '应付款': 289662.53,
            '资金来源合计': 27189662.53,
            '资金占用-待核销基建支出': 24917712.92,
            '货币资金': 2271949.61,
            '资金占用合计': 27189662.53,
        },
        '附表2_01审核汇总表': {
            '合计-送审': 24917712.92,
            '合计-审定': 24917712.92,
            '建安工程-送审': 22472028.59,
            '建安工程-审定': 22472028.59,
            '工程建设其他费-送审': 2445684.33,
            '工程建设其他费-审定': 2445684.33,
        },
        '审核汇总表R27': '46164.0',
    },
    'S452垮沙乡至柯河乡段灾害修复整治工程': {
        '报告': {
            '概算总投资': 2026.0000,
            '项目总投资': 7302.0000,
            '送审金额': 5653.5896,
            '审定金额': 5017.7672,
            '核减额计算': 635.8224,
            '建安工程投资': 4533.0052,
            '待摊投资': 379.2683,
            '已支付': 4912.2735,
        },
        '附表1_02财务决算表': {
            '基建拨款': 61400000.00,
            '应付款': 0,  # 表中为0
            '资金来源合计': 61400000.00,
            '资金占用-待核销基建支出': 49122735.12,
            '货币资金': 12277264.88,
            '资金占用合计': 61400000.00,
        },
        '附表2_01审核汇总表': {
            '合计-送审': 49122735.12,
            '合计-审定': 49122735.12,
            '建安工程-送审': 45330051.92,
            '建安工程-审定': 45330051.92,
            '工程建设其他费-送审': 3792683.20,
            '工程建设其他费-审定': 3792683.20,
        },
        '审核汇总表R27': '46164.0',
    }
}

# 添加核减相关金额（从原始报告中提取）
# S220核减额 = 4,946,306.05元
DATA['S220安羌镇至茸安乡段灾害修复整治工程']['报告']['核减金额'] = 494.6306
# S452核减额
DATA['S452垮沙乡至柯河乡段灾害修复整治工程']['报告']['核减金额'] = 635.8224

# ===== 生成复核结果 =====
findings = []

for proj, d in DATA.items():
    r = d['报告']
    t1 = d['附表1_02财务决算表']
    t2 = d['附表2_01审核汇总表']
    
    # ---- 维度1: 报告↔附表勾稽 ----
    
    # 1.1 报告送审金额 vs 附表2送审
    ss_rep = r['送审金额'] * 10000  # 元
    ss_t2 = t2['合计-送审']
    diff = abs(ss_rep - ss_t2)
    findings.append({
        '项目': proj,
        '复核维度': '报告↔附表勾稽',
        '复核子项': '1-报告送审金额 vs 附表2(审核汇总表)合计送审',
        '数据途径': f'报告: {r["送审金额"]:.4f}万; 附表2: {ss_t2/10000:.4f}万',
        '差异金额(元)': round(diff, 2),
        '复核结论': '✅ 一致' if diff < 1 else f'⚠️ 差异{diff:.2f}元' if diff < 1000 else f'❌ 差异{diff:.2f}元'
    })
    
    # 1.2 报告审定金额 vs 附表2审定
    ds_rep = r['审定金额'] * 10000
    ds_t2 = t2['合计-审定']
    diff = abs(ds_rep - ds_t2)
    findings.append({
        '项目': proj,
        '复核维度': '报告↔附表勾稽',
        '复核子项': '2-报告审定金额 vs 附表2(审核汇总表)合计审定',
        '数据途径': f'报告: {r["审定金额"]:.4f}万; 附表2: {ds_t2/10000:.4f}万',
        '差异金额(元)': round(diff, 2),
        '复核结论': '✅ 一致' if diff < 1 else f'⚠️ 差异{diff:.2f}元' if diff < 1000 else f'❌ 差异{diff:.2f}元'
    })
    
    # 1.3 报告建安工程 vs 附表2建安工程
    ja_rep = r['建安工程投资'] * 10000
    ja_t2 = t2['建安工程-审定']
    diff = abs(ja_rep - ja_t2)
    findings.append({
        '项目': proj,
        '复核维度': '报告↔附表勾稽',
        '复核子项': '3-报告建安工程投资 vs 附表2建安工程审定',
        '数据途径': f'报告: {r["建安工程投资"]:.4f}万; 附表2: {ja_t2/10000:.4f}万',
        '差异金额(元)': round(diff, 2),
        '复核结论': '✅ 一致' if diff < 1 else f'⚠️ 差异{diff:.2f}元' if diff < 1000 else f'❌ 差异{diff:.2f}元'
    })
    
    # 1.4 报告待摊投资 vs 附表2工程建设其他费
    dt_rep = r['待摊投资'] * 10000
    dt_t2 = t2['工程建设其他费-审定']
    diff = abs(dt_rep - dt_t2)
    findings.append({
        '项目': proj,
        '复核维度': '报告↔附表勾稽',
        '复核子项': '4-报告待摊投资 vs 附表2工程建设其他费审定',
        '数据途径': f'报告: {r["待摊投资"]:.4f}万; 附表2: {dt_t2/10000:.4f}万',
        '差异金额(元)': round(diff, 2),
        '复核结论': '✅ 一致' if diff < 1 else f'⚠️ 差异{diff:.2f}元' if diff < 1000 else f'❌ 差异{diff:.2f}元'
    })
    
    # 1.5 附表2合计送审 vs 附表2建安+其他费合计
    sum_t2 = t2['建安工程-送审'] + t2['工程建设其他费-送审']
    diff = abs(sum_t2 - ss_t2)
    findings.append({
        '项目': proj,
        '复核维度': '附表内部勾稽',
        '复核子项': '5-附表2建安+其他费 vs 附表2合计送审',
        '数据途径': f'建安+其他: {sum_t2/10000:.4f}万; 合计: {ss_t2/10000:.4f}万',
        '差异金额(元)': round(diff, 2),
        '复核结论': '✅ 一致' if diff < 1 else f'⚠️ 差异{diff:.2f}元' if diff < 1000 else f'❌ 差异{diff:.2f}元'
    })
    
    # 1.6 附表1资金来源合计 vs 资金占用合计
    diff = abs(t1['资金来源合计'] - t1['资金占用合计'])
    findings.append({
        '项目': proj,
        '复核维度': '附表1内部勾稽',
        '复核子项': '6-附表1(财务决算表)资金来源 vs 资金占用',
        '数据途径': f'来源: {t1["资金来源合计"]/10000:.4f}万; 占用: {t1["资金占用合计"]/10000:.4f}万',
        '差异金额(元)': round(diff, 2),
        '复核结论': '✅ 一致' if diff < 1 else f'❌ 差异{diff:.2f}元'
    })
    
    # 1.7 附表1资金占用合计 vs 附表2合计送审
    diff = abs(t1['资金占用合计'] - ss_t2)
    findings.append({
        '项目': proj,
        '复核维度': '附表1↔附表2勾稽',
        '复核子项': '7-附表1资金占用合计 vs 附表2合计送审',
        '数据途径': f'附表1: {t1["资金占用合计"]/10000:.4f}万; 附表2: {ss_t2/10000:.4f}万',
        '差异金额(元)': round(diff, 2),
        '复核结论': '✅ 一致' if diff < 1 else f'⚠️ 差异{diff:.2f}元' if diff < 1000 else f'❌ 差异{diff:.2f}元'
    })
    
    # ---- 维度2: 报告内部数据勾稽 ----
    
    # 2.1 概算 vs 送审
    diff = r['送审金额'] - r['概算总投资']
    findings.append({
        '项目': proj,
        '复核维度': '报告内部勾稽',
        '复核子项': '1-概算总投资 vs 送审金额',
        '数据途径': f'概算: {r["概算总投资"]:.4f}万; 送审: {r["送审金额"]:.4f}万',
        '差异金额(万元)': round(diff, 4),
        '复核结论': '✅ 一致' if abs(diff) < 0.01 else f'⚠️ 送审超概{diff:.2f}万'
    })
    
    # 2.2 送审 = 审定 + 核减
    calc = r['审定金额'] + r['核减金额']
    diff = abs(r['送审金额'] - calc)
    findings.append({
        '项目': proj,
        '复核维度': '报告内部勾稽',
        '复核子项': '2-送审 = 审定 + 核减',
        '数据途径': f'审定{r["审定金额"]:.4f}万 + 核减{r["核减金额"]:.4f}万 = {calc:.4f}万; 送审{r["送审金额"]:.4f}万',
        '差异金额(万元)': round(r['送审金额'] - calc, 4),
        '复核结论': '✅ 一致' if diff < 0.01 else f'⚠️ 差异{diff:.2f}万'
    })
    
    # 2.3 已支付 vs 审定金额
    diff = r['已支付'] - r['审定金额']
    findings.append({
        '项目': proj,
        '复核维度': '报告内部勾稽',
        '复核子项': '3-已支付 vs 审定金额（判断是否超付）',
        '数据途径': f'已支付: {r["已支付"]:.4f}万; 审定: {r["审定金额"]:.4f}万',
        '差异金额(万元)': round(diff, 4),
        '复核结论': '✅ 正常' if diff < 0 else '⚠️ 已支付小于审定金额' if diff < 0 else '❌ 已支付超审定金额' if diff > 0.01 else '✅ 一致'
    })
    
    # 2.4 建安+待摊 vs 项目总投资
    sub = r['建安工程投资'] + r['待摊投资']
    diff = r['项目总投资'] - sub
    findings.append({
        '项目': proj,
        '复核维度': '报告内部勾稽',
        '复核子项': '4-建安+待摊 = 总投资（是否含设备/其他投资）',
        '数据途径': f'建安{r["建安工程投资"]:.4f}万 + 待摊{r["待摊投资"]:.4f}万 = {sub:.4f}万; 总投资{r["项目总投资"]:.4f}万',
        '差异金额(万元)': round(diff, 4),
        '复核结论': '✅ 一致' if abs(diff) < 0.01 else f'⚠️ 差异{diff:.2f}万（可能含设备/其他）'
    })
    
    # 2.5 附表2送审=审定（0核减）
    diff = abs(ss_t2 - ds_t2)
    findings.append({
        '项目': proj,
        '复核维度': '附表2内部勾稽',
        '复核子项': '5-审核汇总表送审 vs 审定（核减额）',
        '数据途径': f'送审: {ss_t2/10000:.4f}万; 审定: {ds_t2/10000:.4f}万',
        '差异金额(元)': round(diff, 2),
        '复核结论': '✅ 一致(无核减)' if diff < 1 else f'⚠️ 核减{diff/10000:.4f}万'
    })

# ===== 输出Excel =====
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '数据勾稽复核对账表'
    
    # 标题行
    ws.merge_cells('A1:G1')
    ws['A1'] = '四川融策会计师事务所 - 竣工财务决算审核报告数据勾稽复核对账表'
    ws['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='1A237E')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:G2')
    ws['A2'] = '复核日期：2026-06-24 | 复核对象：S220安羌镇至茸安乡段、S452垮沙乡至柯河乡段'
    ws['A2'].font = Font(name='微软雅黑', size=10, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # 表头
    headers = ['序号', '项目名称', '复核维度', '复核子项', '数据途径', '差异金额', '复核结论']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[3].height = 25
    
    # 数据
    green_fill = PatternFill(start_color='E8F5E9', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFF3E0', fill_type='solid')
    red_fill = PatternFill(start_color='FFEBEE', fill_type='solid')
    header_fill_s220 = PatternFill(start_color='E3F2FD', fill_type='solid')
    header_fill_s452 = PatternFill(start_color='F3E5F5', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin', color='BDBDBD'),
        right=Side(style='thin', color='BDBDBD'),
        top=Side(style='thin', color='BDBDBD'),
        bottom=Side(style='thin', color='BDBDBD'),
    )
    
    row_num = 4
    seq = 1
    current_proj = ''
    
    for fi in findings:
        if fi['项目'] != current_proj:
            current_proj = fi['项目']
            # 项目分隔行
            ws.merge_cells(f'A{row_num}:G{row_num}')
            ws.cell(row=row_num, column=1, value=current_proj)
            ws.cell(row=row_num, column=1).font = Font(name='微软雅黑', size=11, bold=True, color='1A237E')
            fill = header_fill_s220 if 'S220' in current_proj else header_fill_s452
            ws.cell(row=row_num, column=1).fill = fill
            ws.row_dimensions[row_num].height = 22
            row_num += 1
        
        ws.cell(row=row_num, column=1, value=seq)
        ws.cell(row=row_num, column=2, value=fi['项目'][:30])
        ws.cell(row=row_num, column=3, value=fi['复核维度'])
        ws.cell(row=row_num, column=4, value=fi['复核子项'])
        ws.cell(row=row_num, column=5, value=fi['数据途径'])
        ws.cell(row=row_num, column=6, value=fi.get('差异金额(元)', fi.get('差异金额(万元)', '')))
        ws.cell(row=row_num, column=7, value=fi['复核结论'])
        
        # 颜色标记
        conclusion = str(fi['复核结论'])
        if '一致' in conclusion or '正常' in conclusion:
            fill = green_fill
        elif '超概' in conclusion or '超付' in conclusion or '❌' in conclusion or '差异' in conclusion:
            for c in range(1, 8):
                ws.cell(row=row_num, column=c).fill = red_fill if '❌' in conclusion else yellow_fill
        
        for c in range(1, 8):
            ws.cell(row=row_num, column=c).border = thin_border
            ws.cell(row=row_num, column=c).alignment = Alignment(vertical='center', wrap_text=True)
            ws.cell(row=row_num, column=c).font = Font(name='微软雅黑', size=9)
        
        seq += 1
        row_num += 1
    
    # 列宽
    widths = [6, 25, 16, 35, 45, 16, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w
    
    # ===== Sheet2: 综合汇总 =====
    ws2 = wb.create_sheet('综合复核意见')
    ws2.merge_cells('A1:E1')
    ws2['A1'] = '竣工财务决算审核报告数据勾稽复核 - 综合意见'
    ws2['A1'].font = Font(name='微软雅黑', size=13, bold=True, color='1A237E')
    ws2['A1'].alignment = Alignment(horizontal='center')
    ws2.row_dimensions[1].height = 30
    
    # 项目标题
    r = 3
    for proj in ['S220安羌镇至茸安乡段灾害修复整治工程', 'S452垮沙乡至柯河乡段灾害修复整治工程']:
        d = DATA[proj]
        rpt = d['报告']
        t2 = d['附表2_01审核汇总表']
        
        ws2.merge_cells(f'A{r}:E{r}')
        ws2.cell(row=r, column=1, value=proj)
        ws2.cell(row=r, column=1).font = Font(name='微软雅黑', size=11, bold=True)
        ws2.cell(row=r, column=1).fill = PatternFill(start_color='E3F2FD', fill_type='solid')
        r += 1
        
        # 核心数据对照表
        summary_items = [
            ('1. 概算总投资', f'{rpt["概算总投资"]:.2f}万', '-', '-'),
            ('2. 送审金额', f'{rpt["送审金额"]:.2f}万', f'{t2["合计-送审"]/10000:.2f}万', ''),
            ('3. 审定金额', f'{rpt["审定金额"]:.2f}万', f'{t2["合计-审定"]/10000:.2f}万', ''),
            ('4. 核减金额(报告送审-审定)', f'{rpt["送审金额"]-rpt["审定金额"]:.2f}万', '-', ''),
            ('5. 建安工程投资', f'{rpt["建安工程投资"]:.2f}万', f'{t2["建安工程-审定"]/10000:.2f}万', ''),
            ('6. 待摊投资/工程建设其他费', f'{rpt["待摊投资"]:.2f}万', f'{t2["工程建设其他费-审定"]/10000:.2f}万', ''),
            ('7. 项目总投资', f'{rpt["项目总投资"]:.2f}万', '-', ''),
            ('8. 已支付', f'{rpt["已支付"]:.2f}万', '-', ''),
            ('9. 未支付(审定-已支付)', f'{rpt["审定金额"]-rpt["已支付"]:.2f}万', '-', f'{"✅ 正常" if rpt["已支付"]<rpt["审定金额"] else "❌ 超付"}'),
        ]
        
        ws2.cell(row=r, column=1, value='指标').font = Font(bold=True, size=10)
        ws2.cell(row=r, column=2, value='审核报告(万元)').font = Font(bold=True, size=10)
        ws2.cell(row=r, column=3, value='审核汇总表(万元)').font = Font(bold=True, size=10)
        ws2.cell(row=r, column=4, value='评价').font = Font(bold=True, size=10)
        for c in range(1,5):
            ws2.cell(row=r, column=c).fill = PatternFill(start_color='E0E0E0', fill_type='solid')
            ws2.cell(row=r, column=c).border = thin_border
        r += 1
        
        for item in summary_items:
            ws2.cell(row=r, column=1, value=item[0]).border = thin_border
            ws2.cell(row=r, column=2, value=item[1]).border = thin_border
            ws2.cell(row=r, column=3, value=item[2]).border = thin_border
            ws2.cell(row=r, column=4, value=item[3]).border = thin_border
            for c in range(1,5):
                ws2.cell(row=r, column=c).font = Font(name='微软雅黑', size=9)
            r += 1
        r += 1
    
    # 列宽
    for i, w in enumerate([35, 18, 18, 12], 1):
        ws2.column_dimensions[chr(64+i)].width = w
    
    output_path = r'C:\Users\scrccpa\Desktop\融策审计勾稽复核表.xlsx'
    wb.save(output_path)
    print(f"✅ 复核表已保存: {output_path}")
    
    # ===== 控制台摘要 =====
    print("\n" + "="*70)
    print("  复核结论摘要")
    print("="*70)
    
    for proj in ['S220安羌镇至茸安乡段灾害修复整治工程', 'S452垮沙乡至柯河乡段灾害修复整治工程']:
        d = DATA[proj]
        rpt = d['报告']
        t2 = d['附表2_01审核汇总表']
        print(f"\n【{proj}】")
        print(f"  ✅ 报告送审金额⇔附表2送审: 一致") if abs(rpt['送审金额']*10000 - t2['合计-送审']) < 1 else print(f"  ❌ 报告送审金额⇔附表2送审: 差异")
        print(f"  ✅ 报告审定金额⇔附表2审定: 一致") if abs(rpt['审定金额']*10000 - t2['合计-审定']) < 1 else print(f"  ❌ 报告审定金额⇔附表2审定: 差异")
        print(f"  ✅ 报告建安投资⇔附表2建安: 一致") if abs(rpt['建安工程投资']*10000 - t2['建安工程-审定']) < 1 else print(f"  ❌ 报告建安投资⇔附表2建安: 差异")
        print(f"  ✅ 报告待摊投资⇔附表2其他费: 一致") if abs(rpt['待摊投资']*10000 - t2['工程建设其他费-审定']) < 1 else print(f"  ❌ 报告待摊投资⇔附表2其他费: 差异")
        print(f"  ✅ 附表1来源=占用: 一致") if abs(d['附表1_02财务决算表']['资金来源合计'] - d['附表1_02财务决算表']['资金占用合计']) < 1 else print(f"  ❌ 附表1来源≠占用")
        print(f"  ⚠️ 送审超概: 超{rpt['送审金额']-rpt['概算总投资']:.2f}万")
        print(f"  ⚠️ 附表2送审=审定(0核减): 需核实")
        print(f"  ✅ 报告送审=审定+核减: 一致") if abs(rpt['送审金额'] - rpt['审定金额'] - rpt['核减金额']) < 0.01 else print(f"  ❌ 报告送审≠审定+核减")

except ImportError as e:
    print(f"⚠️ openpyxl未安装: {e}")
    # CSV回退
    import csv
    output_path = r'C:\Users\scrccpa\Desktop\融策审计勾稽复核表.csv'
    with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.writer(f)
        w.writerow(['序号','项目名称','复核维度','复核子项','数据途径','差异金额','复核结论'])
        for i, fi in enumerate(findings, 1):
            diff_amt = fi.get('差异金额(元)', fi.get('差异金额(万元)', ''))
            w.writerow([i, fi['项目'], fi['复核维度'], fi['复核子项'], fi['数据途径'], diff_amt, fi['复核结论']])
    print(f"✅ CSV已保存: {output_path}")
