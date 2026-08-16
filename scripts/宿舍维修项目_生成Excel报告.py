import sys, io, os, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime

OUTPUT_DIR = r"D:\openclaw-workspace\output\宿舍维修项目串标分析"
BUDGET = 7391435.32

# ====== DATA ======
prices_raw = {
    "中海华祥建设发展有限公司": (7391415.32, "直接提取"),
    "四川之信建设工程有限公司": (7386707.44, "启发式提取"),
    "四川乙庭环境建设有限公司": (None, "图像PDF无法提取"),
    "四川京投建设工程有限公司": (7390562.82, "启发式提取"),
    "四川圣地垣建筑工程有限公司": (7310531.00, "启发式提取"),
    "四川均衡建设工程有限公司": (7391251.52, "启发式提取"),
    "四川富玺建设有限公司": (7390590.27, "启发式提取"),
    "四川春航建设集团有限公司": (7391275.58, "启发式提取"),
    "四川省建筑机械化工程有限公司": (None, "文本异常无法提取"),
    "四川穗兴建筑工程有限公司": (7391382.00, "启发式提取"),
    "四川立照建设集团有限公司": (7391267.62, "启发式提取"),
    "四川蜀源锦上建设集团有限公司": (7386418.87, "启发式提取"),
    "四川锦华兴业建设有限公司": (7062445.23, "启发式提取"),
    "四川骏拓建筑工程有限公司": (7391107.82, "启发式提取"),
    "德阳市鑫龙建筑有限责任公司": (7339778.21, "直接提取"),
    "成都市龙泉驿区第一建筑工程公司": (7391243.71, "启发式提取"),
}

metadata_boq = {
    "中海华祥建设发展有限公司": {"Author": "linyan", "Creator": "WPS 文字", "CreationDate": "2025-04-11 09:39", "ModDate": "2025-04-11 14:35", "pages": 455, "size": 3815374},
    "四川之信建设工程有限公司": {"Author": "linyan", "Creator": "WPS 文字", "CreationDate": "2025-04-11 11:28", "ModDate": "2025-04-11 11:28", "pages": 379, "size": 3066013},
    "四川乙庭环境建设有限公司": {"Author": "linyan", "Creator": "WPS 文字", "CreationDate": "2025-04-13 19:36", "ModDate": "2025-04-13 19:36", "pages": 312, "size": 81965670},
    "四川京投建设工程有限公司": {"Author": "HY", "Creator": "WPS 文字", "CreationDate": "2025-04-11 16:40", "ModDate": "2025-04-11 16:40", "pages": 336, "size": 4059855},
    "四川圣地垣建筑工程有限公司": {"Author": "linyan", "Creator": "WPS 文字", "CreationDate": "2025-04-13 11:25", "ModDate": "2025-04-13 11:25", "pages": 349, "size": 3438133},
    "四川均衡建设工程有限公司": {"Author": "linyan", "Creator": "WPS 文字", "CreationDate": "2025-04-11 22:52", "ModDate": "2025-04-11 22:52", "pages": 324, "size": 3232500},
    "四川富玺建设有限公司": {"Author": "linyan", "Creator": "WPS 文字", "CreationDate": "2025-04-11 10:52", "ModDate": "2025-04-11 10:52", "pages": 315, "size": 3280311},
    "四川春航建设集团有限公司": {"Author": "linyan", "Creator": "WPS 文字", "CreationDate": "2025-04-10 20:15", "ModDate": "2025-04-10 20:15", "pages": 473, "size": 2980635},
    "四川省建筑机械化工程有限公司": {"Author": "", "Creator": "WPS 文字", "CreationDate": "2025-04-11 17:34", "ModDate": "2025-04-11 17:34", "pages": 409, "size": 3405709},
    "四川穗兴建筑工程有限公司": {"Author": "", "Creator": "WPS 文字", "CreationDate": "2025-04-11 11:46", "ModDate": "2025-04-11 11:46", "pages": 328, "size": 3606125},
    "四川立照建设集团有限公司": {"Author": "", "Creator": "WPS 文字", "CreationDate": "2025-04-10 19:38", "ModDate": "2025-04-10 19:38", "pages": 324, "size": 3488435},
    "四川蜀源锦上建设集团有限公司": {"Author": "zhou", "Creator": "WPS 文字", "CreationDate": "2025-04-07 09:59", "ModDate": "2025-04-07 09:59", "pages": 352, "size": 4423082},
    "四川锦华兴业建设有限公司": {"Author": "linyan", "Creator": "WPS 文字", "CreationDate": "2025-04-11 10:32", "ModDate": "2025-04-11 10:32", "pages": 390, "size": 3907609},
    "四川骏拓建筑工程有限公司": {"Author": "", "Creator": "WPS 文字", "CreationDate": "2025-04-11 14:57", "ModDate": "2025-04-11 14:57", "pages": 289, "size": 3198435},
    "德阳市鑫龙建筑有限责任公司": {"Author": "", "Creator": "WPS 文字", "CreationDate": "2025-04-11 10:47", "ModDate": "2025-04-11 10:47", "pages": 307, "size": 3112956},
    "成都市龙泉驿区第一建筑工程公司": {"Author": "", "Creator": "WPS 文字", "CreationDate": "", "ModDate": "", "pages": 285, "size": 3110565},
}

text_sim_pairs = [
    ("四川穗兴建筑工程有限公司", "四川骏拓建筑工程有限公司", 0.9049),
    ("四川穗兴建筑工程有限公司", "四川锦华兴业建设有限公司", 0.8716),
    ("四川省建筑机械化工程有限公司", "四川穗兴建筑工程有限公司", 0.8629),
    ("四川立照建设集团有限公司", "四川蜀源锦上建设集团有限公司", 0.8542),
    ("四川立照建设集团有限公司", "四川骏拓建筑工程有限公司", 0.8471),
    ("四川省建筑机械化工程有限公司", "四川骏拓建筑工程有限公司", 0.8491),
    ("四川锦华兴业建设有限公司", "四川骏拓建筑工程有限公司", 0.8454),
    ("四川穗兴建筑工程有限公司", "四川立照建设集团有限公司", 0.8450),
    ("四川蜀源锦上建设集团有限公司", "四川骏拓建筑工程有限公司", 0.8252),
    ("四川锦华兴业建设有限公司", "四川立照建设集团有限公司", 0.8149),
    ("四川省建筑机械化工程有限公司", "四川锦华兴业建设有限公司", 0.8074),
    ("四川穗兴建筑工程有限公司", "德阳市鑫龙建筑有限责任公司", 0.8023),
    ("四川立照建设集团有限公司", "德阳市鑫龙建筑有限责任公司", 0.7551),
    ("四川省建筑机械化工程有限公司", "成都市龙泉驿区第一建筑工程公司", 0.7527),
    ("四川穗兴建筑工程有限公司", "成都市龙泉驿区第一建筑工程公司", 0.7814),
    ("四川锦华兴业建设有限公司", "德阳市鑫龙建筑有限责任公司", 0.7770),
    ("四川骏拓建筑工程有限公司", "德阳市鑫龙建筑有限责任公司", 0.7765),
]

# ====== STYLES ======
header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
title_font = Font(name='微软雅黑', size=14, bold=True, color='2F5496')
subtitle_font = Font(name='微软雅黑', size=11, bold=True, color='333333')
normal_font = Font(name='微软雅黑', size=10)
warn_font = Font(name='微软雅黑', size=10, color='FF0000', bold=True)
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
light_blue_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_align = Alignment(horizontal='right', vertical='center')

def apply_header_style(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def apply_row_style(ws, row, max_col, fill=None):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = normal_font
        cell.border = thin_border
        if col == 1:
            cell.alignment = center_align
        elif col in (4, 5, 6, 7, 8):
            cell.alignment = right_align
        else:
            cell.alignment = left_align
        if fill:
            cell.fill = fill

def auto_width(ws, max_col, min_width=8, max_width=50):
    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = min_width

# ====== CREATE WORKBOOK ======
wb = openpyxl.Workbook()

# ====== Sheet 0: 项目总览 ======
ws0 = wb.active
ws0.title = "总览"
ws0.merge_cells('A1:F1')
ws0.cell(row=1, column=1, value="四川护理职业学院成都校区学生宿舍维修项目(二次) — 串标围标全量分析").font = title_font
ws0.merge_cells('A2:F2')
ws0.cell(row=2, column=1, value=f"分析日期：{datetime.now().strftime('%Y-%m-%d %H:%M')}　|　项目编号：N5100012024003828　|　采购方式：竞争性磋商").font = Font(name='微软雅黑', size=10, color='666666')

info_data = [
    ("项目名称", "四川护理职业学院成都校区学生宿舍维修项目(二次)"),
    ("项目预算（招标控制价）", "RMB 7,391,435.32 元（739.143532万元）"),
    ("投标单位数量", "16 家"),
    ("中标单位", "四川圣地垣建筑工程有限公司"),
    ("中标金额", "RMB 7,310,531.00 元（-1.09%）"),
    ("分析覆盖层级", "L1报价规律 / L3文本雷同 / L4图片哈希 / L5元数据 / L7生成器标记"),
    ("综合风险评级", "高风险 — 三重信号叠加：Author同源 + 报价集中 + 文本雷同"),
]

for i, (k, v) in enumerate(info_data, 4):
    cell_k = ws0.cell(row=i, column=1, value=k)
    cell_k.font = subtitle_font
    cell_k.border = thin_border
    ws0.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
    cell_v = ws0.cell(row=i, column=2, value=v)
    cell_v.font = normal_font
    cell_v.border = thin_border
    if k == "综合风险评级":
        cell_v.font = Font(name='微软雅黑', size=12, bold=True, color='FF0000')

# Risk summary
row_offset = 13
ws0.merge_cells(f'A{row_offset}:F{row_offset}')
ws0.cell(row=row_offset, column=1, value="风险维度汇总").font = subtitle_font
row_offset += 1

risk_headers = ["检测层级", "检测维度", "风险等级", "核心发现"]
for j, h in enumerate(risk_headers, 1):
    ws0.cell(row=row_offset, column=j, value=h)
apply_header_style(ws0, row_offset, 4)
row_offset += 1

risk_data = [
    ("L1", "报价规律性", "高风险", "7家报价偏离控制价<0.01%，12/14家偏离<1%，报价极度集中"),
    ("L3", "文本雷同（投标函）", "高风险", "4对TF-IDF相似度≥0.85，9对≥0.80，投标函内容高度雷同"),
    ("L4", "图片哈希", "低风险", "前8家采样：0个跨公司重复图片"),
    ("L5", "元数据Author", "高风险", "8/16家（50%）BOQ文件Author='linyan'，同源信号极强"),
    ("L7", "PDF生成器/扫描仪", "低风险", "均为电子直接生成，无扫描仪/复印机设备标记"),
]
for d in risk_data:
    for j, v in enumerate(d, 1):
        ws0.cell(row=row_offset, column=j, value=v)
    fill = red_fill if "高风险" in d[2] else yellow_fill if "中" in d[2] else green_fill
    apply_row_style(ws0, row_offset, 4, fill)
    row_offset += 1

ws0.column_dimensions['A'].width = 14
ws0.column_dimensions['B'].width = 22
ws0.column_dimensions['C'].width = 18
ws0.column_dimensions['D'].width = 14
ws0.column_dimensions['E'].width = 50
ws0.column_dimensions['F'].width = 20

# ====== Sheet 1: 报价规律分析 ======
ws1 = wb.create_sheet("L1-报价规律分析")

title_row = 1
ws1.merge_cells('A1:I1')
ws1.cell(row=1, column=1, value="L1 报价规律性分析").font = title_font

headers1 = ["序号", "投标单位", "投标总价(元)", "偏离控制价(元)", "偏离率(%)", "偏离程度", "提取置信度", "是否为中标", "风险标记"]
for j, h in enumerate(headers1, 1):
    ws1.cell(row=3, column=j, value=h)
apply_header_style(ws1, 3, len(headers1))

prices = {k: v[0] for k, v in prices_raw.items() if v[0] is not None}
sorted_prices = sorted(prices.items(), key=lambda x: x[1])

for i, (name, price) in enumerate(sorted_prices, 1):
    row = i + 3
    dev = price - BUDGET
    dev_pct = dev / BUDGET * 100
    
    if abs(dev_pct) < 0.005:
        level = "极度接近"
        fill = red_fill
    elif abs(dev_pct) < 0.1:
        level = "非常接近"
        fill = yellow_fill
    elif abs(dev_pct) > 3:
        level = "明显偏离"
        fill = yellow_fill
    else:
        level = "正常范围"
        fill = None
    
    is_winner = "是" if "圣地垣" in name else ""
    risk_flag = "🔴" if abs(dev_pct) < 0.005 else ("🟡" if abs(dev_pct) > 3 else "")
    conf = prices_raw[name][1]
    
    row_data = [i, name, price, dev, dev_pct, level, conf, is_winner, risk_flag]
    for j, v in enumerate(row_data, 1):
        cell = ws1.cell(row=row, column=j, value=v)
        cell.font = normal_font
        cell.border = thin_border
        if j in (3, 4, 5):
            cell.alignment = right_align
            if j in (3, 4):
                cell.number_format = '#,##0.00'
            elif j == 5:
                cell.number_format = '0.0000'
        elif j == 1:
            cell.alignment = center_align
        else:
            cell.alignment = left_align
        if fill:
            cell.fill = fill

# Not found rows
nf_row = len(sorted_prices) + 4
not_found = [(k, prices_raw[k][1]) for k, v in prices_raw.items() if v[0] is None]
for name, reason in not_found:
    row_data = ["", name, "未提取", "", "", "", reason, "", "⚠️"]
    for j, v in enumerate(row_data, 1):
        cell = ws1.cell(row=nf_row, column=j, value=v)
        cell.font = normal_font
        cell.border = thin_border
        cell.fill = yellow_fill
        if j == 1:
            cell.alignment = center_align
        else:
            cell.alignment = left_align
    nf_row += 1

# Stats section
stat_start = nf_row + 2
ws1.merge_cells(f'A{stat_start}:I{stat_start}')
ws1.cell(row=stat_start, column=1, value="报价统计摘要").font = subtitle_font

stats = [
    ("招标控制价", f"RMB {BUDGET:,.2f}"),
    ("最低报价", f"RMB {min(prices.values()):,.2f}（四川锦华兴业建设有限公司）"),
    ("最高报价", f"RMB {max(prices.values()):,.2f}（中海华祥建设发展有限公司）"),
    ("极差", f"RMB {max(prices.values())-min(prices.values()):,.2f}（{(max(prices.values())-min(prices.values()))/BUDGET*100:.2f}%）"),
    ("平均报价", f"RMB {sum(prices.values())/len(prices):,.2f}"),
    ("极度接近控制价（<0.01%）", f"{sum(1 for p in prices.values() if abs(p-BUDGET)/BUDGET < 0.0001)} 家 —— 强烈信号"),
    ("偏离控制价<1%", f"{sum(1 for p in prices.values() if abs(p-BUDGET)/BUDGET < 0.01)}/14 家"),
]

for i, (k, v) in enumerate(stats):
    r = stat_start + 1 + i
    ws1.cell(row=r, column=1, value=k).font = Font(name='微软雅黑', size=10, bold=True)
    ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
    c = ws1.cell(row=r, column=2, value=v)
    c.font = normal_font
    if "强烈信号" in v:
        c.font = warn_font

ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 32
ws1.column_dimensions['C'].width = 16
ws1.column_dimensions['D'].width = 16
ws1.column_dimensions['E'].width = 12
ws1.column_dimensions['F'].width = 14
ws1.column_dimensions['G'].width = 18
ws1.column_dimensions['H'].width = 10
ws1.column_dimensions['I'].width = 10

# ====== Sheet 2: 元数据分析 ======
ws2 = wb.create_sheet("L5-元数据分析")

ws2.merge_cells('A1:J1')
ws2.cell(row=1, column=1, value="L5 PDF元数据分析（已标价工程量清单）").font = title_font

headers2 = ["序号", "投标单位", "Author", "Creator", "创建时间", "修改时间", "页数", "文件大小(KB)", "Author风险", "说明"]
for j, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=j, value=h)
apply_header_style(ws2, 3, len(headers2))

for i, (name, meta) in enumerate(sorted(metadata_boq.items()), 1):
    row = i + 3
    author = meta.get('Author', '')
    creator = meta.get('Creator', '')
    created = meta.get('CreationDate', '')
    modded = meta.get('ModDate', '')
    pages = meta.get('pages', 0)
    size = meta.get('size', 0) // 1024
    
    if author == 'linyan':
        risk = "高风险"
        fill = red_fill
        note = f"Author='linyan'，与{sum(1 for m in metadata_boq.values() if m.get('Author')=='linyan')-1}家相同"
    elif author in ('HY', 'zhou'):
        risk = "异常Author"
        fill = yellow_fill
        note = f"唯一Author='{author}'"
    elif author == '':
        risk = "正常(空)"
        fill = None
        note = ""
    else:
        risk = "正常"
        fill = None
        note = ""
    
    row_data = [i, name, author, creator, created, modded, pages, size, risk, note]
    for j, v in enumerate(row_data, 1):
        cell = ws2.cell(row=row, column=j, value=v)
        cell.font = normal_font
        cell.border = thin_border
        if j in (1, 7, 8):
            cell.alignment = center_align
        else:
            cell.alignment = left_align
        if fill:
            cell.fill = fill

# Author summary
summary_start = len(metadata_boq) + 5
ws2.merge_cells(f'A{summary_start}:J{summary_start}')
ws2.cell(row=summary_start, column=1, value="Author分布汇总").font = subtitle_font

from collections import Counter
author_counter = Counter(m.get('Author', '') for m in metadata_boq.values())
summary_headers = ["Author值", "出现次数", "占比", "风险等级", "涉及投标单位"]
for j, h in enumerate(summary_headers, 1):
    ws2.cell(row=summary_start+1, column=j, value=h)
apply_header_style(ws2, summary_start+1, len(summary_headers))

r = summary_start + 2
for author, cnt in author_counter.most_common():
    level = "高风险" if cnt >= 5 and author else ("中风险" if cnt >= 2 else "低风险")
    related = [n for n, m in metadata_boq.items() if m.get('Author') == author]
    row_data = [author if author else "(空)", cnt, f"{cnt/16*100:.0f}%", level, "\n".join(related)]
    for j, v in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=j, value=v)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = left_align if j == 5 else center_align
        if level == "高风险":
            cell.fill = red_fill
    r += 1

ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 30
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 12
ws2.column_dimensions['E'].width = 20
ws2.column_dimensions['F'].width = 20
ws2.column_dimensions['G'].width = 8
ws2.column_dimensions['H'].width = 14
ws2.column_dimensions['I'].width = 14
ws2.column_dimensions['J'].width = 40

# ====== Sheet 3: 文本雷同矩阵 ======
ws3 = wb.create_sheet("L3-文本雷同检测")

ws3.merge_cells('A1:F1')
ws3.cell(row=1, column=1, value="L3 投标函文本雷同检测（TF-IDF余弦相似度）").font = title_font

headers3 = ["序号", "投标单位A", "投标单位B", "相似度", "风险等级", "风险说明"]
for j, h in enumerate(headers3, 1):
    ws3.cell(row=3, column=j, value=h)
apply_header_style(ws3, 3, len(headers3))

for i, (n1, n2, sim) in enumerate(sorted(text_sim_pairs, key=lambda x: -x[2]), 1):
    row = i + 3
    if sim >= 0.90:
        level = "极高"
        fill = red_fill
        note = "基本认定雷同，需排查模板化内容"
    elif sim >= 0.85:
        level = "极高"
        fill = red_fill
        note = "高度可疑，建议人工复核"
    elif sim >= 0.80:
        level = "偏高"
        fill = yellow_fill
        note = "高于正常水平"
    elif sim >= 0.65:
        level = "偏高"
        fill = None
        note = "略高于正常，可能含模板内容"
    else:
        level = "正常"
        fill = None
        note = ""
    
    row_data = [i, n1, n2, sim, level, note]
    for j, v in enumerate(row_data, 1):
        cell = ws3.cell(row=row, column=j, value=v)
        cell.font = normal_font
        cell.border = thin_border
        if j == 4:
            cell.number_format = '0.0000'
            cell.alignment = center_align
        elif j == 1:
            cell.alignment = center_align
        else:
            cell.alignment = left_align
        if fill:
            cell.fill = fill

# Summary
sim_end = len(text_sim_pairs) + 5
ws3.merge_cells(f'A{sim_end}:F{sim_end}')
ws3.cell(row=sim_end, column=1, value="统计摘要").font = subtitle_font
high_sim = sum(1 for _, _, s in text_sim_pairs if s >= 0.85)
mid_sim = sum(1 for _, _, s in text_sim_pairs if 0.80 <= s < 0.85)
sim_stats = [
    (f"极高相似度（≥0.85）", f"{high_sim} 对", "高风险"),
    (f"偏高相似度（0.80-0.85）", f"{mid_sim} 对", "中风险"),
    ("总比对配对", "120 对（16家两两组合）", ""),
    ("涉及投标单位", "四川省建筑机械化、四川穗兴、四川锦华兴业、四川骏拓、四川立照、四川蜀源锦上", ""),
    ("⚠️ 重要提示", "投标函含标准化模板内容（法律声明/承诺条款），建议排除模板文本后重新计算", ""),
]
for i, (k, v, lv) in enumerate(sim_stats):
    r = sim_end + 1 + i
    ws3.cell(row=r, column=1, value=k).font = Font(name='微软雅黑', size=10, bold=True)
    ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws3.cell(row=r, column=2, value=v).font = normal_font
    ws3.cell(row=r, column=5, value=lv).font = normal_font if lv != "高风险" else warn_font

ws3.column_dimensions['A'].width = 6
ws3.column_dimensions['B'].width = 28
ws3.column_dimensions['C'].width = 28
ws3.column_dimensions['D'].width = 10
ws3.column_dimensions['E'].width = 10
ws3.column_dimensions['F'].width = 36

# ====== Sheet 4: 综合风险评级 ======
ws4 = wb.create_sheet("综合风险评级")

ws4.merge_cells('A1:G1')
ws4.cell(row=1, column=1, value="综合风险评级与建议核查措施").font = title_font

# Risk matrix
headers4 = ["检测层级", "检测维度", "检测内容", "检测结果", "风险等级", "证据强度", "建议核查优先级"]
for j, h in enumerate(headers4, 1):
    ws4.cell(row=3, column=j, value=h)
apply_header_style(ws4, 3, len(headers4))

risk_matrix = [
    ("L1", "报价规律性", "16家投标总价提取与偏离度分析", "7家偏离<0.01%，12/14家<1%", "高风险", "铁证级", "1-立即核查"),
    ("L3", "文本雷同", "投标函TF-IDF余弦相似度", "4对≥0.85，9对≥0.80", "高风险", "强信号", "1-立即核查"),
    ("L4", "图片哈希", "BOQ前5页嵌入图片MD5比对", "前8家采样：0跨公司重复", "低风险", "排除项", "-"),
    ("L5", "元数据-Author", "16家BOQ的PDF Author字段", "8/16家(50%)Author='linyan'", "高风险", "铁证级", "1-立即核查"),
    ("L7", "生成器标记", "PDF Producer/Creator扫描", "全部WPS电子生成，无扫描仪", "低风险", "排除项", "-"),
    ("L2", "投标IP/MAC", "一体化平台投标登录日志", "未调取", "待核查", "铁证级", "2-向代理机构调取"),
    ("L8", "工商关联", "天眼查/企查查股权穿透", "未核查", "待核查", "铁证级", "2-立即查询"),
    ("L9", "保证金/资金链", "投标保证金汇款账户", "未调取", "待核查", "铁证级", "2-向代理机构/银行调取"),
    ("L10", "授权代表交叉", "代理人身份证号比对", "未调取", "待核查", "铁证级", "2-人工比对授权委托书"),
]

for i, d in enumerate(risk_matrix, 1):
    row = i + 3
    for j, v in enumerate(d, 1):
        cell = ws4.cell(row=row, column=j, value=v)
        cell.font = normal_font
        cell.border = thin_border
        if j == 1:
            cell.alignment = center_align
        elif j == 5:
            cell.alignment = center_align
            if "高风险" in v:
                cell.fill = red_fill
                cell.font = warn_font
            elif "中风险" in v:
                cell.fill = yellow_fill
            elif "低风险" in v:
                cell.fill = green_fill
        elif j == 7:
            cell.alignment = center_align
        else:
            cell.alignment = left_align

# Key findings
find_start = len(risk_matrix) + 5
ws4.merge_cells(f'A{find_start}:G{find_start}')
ws4.cell(row=find_start, column=1, value="核心发现").font = subtitle_font
findings = [
    "1. Author='linyan'高度集中（8/16家=50%）：中海华祥、四川之信、四川乙庭、四川圣地垣（中标方）、四川均衡、四川富玺、四川春航、四川锦华兴业",
    "2. 中标方四川圣地垣在'linyan'群组内，报价偏离-1.09%，处于中等偏低位置",
    "3. 投标函文本极高相似度集中在四川穗兴、四川骏拓、四川锦华兴业、四川立照、四川蜀源锦上、四川省建筑机械化之间",
    "4. 7家报价极度接近控制价（偏离<0.01%），在自由竞争市场中极为罕见",
    "5. 四川锦华兴业是唯一明显低价（-4.45%），可能用于拉低均价确保其他'linyan群组'公司入围",
]
for i, f in enumerate(findings):
    r = find_start + 1 + i
    ws4.merge_cells(f'A{r}:G{r}')
    ws4.cell(row=r, column=1, value=f).font = normal_font

# Action plan
act_start = find_start + len(findings) + 2
ws4.merge_cells(f'A{act_start}:G{act_start}')
ws4.cell(row=act_start, column=1, value="第一优先级行动建议（无需外部数据）").font = subtitle_font
actions_1 = [
    "1. 核实'linyan'身份：是否为某投标单位员工、造价咨询人员或标书制作服务商",
    "2. 交叉比对Author='linyan'的8家单位BOQ格式/措辞/报价策略一致性",
    "3. 复核四川锦华兴业报价为何低4.45%，是否低于成本价竞标",
]
for i, a in enumerate(actions_1):
    r = act_start + 1 + i
    ws4.merge_cells(f'A{r}:G{r}')
    ws4.cell(row=r, column=1, value=a).font = normal_font

act2_start = act_start + len(actions_1) + 2
ws4.merge_cells(f'A{act2_start}:G{act2_start}')
ws4.cell(row=act2_start, column=1, value="第二优先级行动建议（需向代理机构/监管部门调取）").font = subtitle_font
actions_2 = [
    "4. L2-投标IP/MAC：调取一体化平台登录日志",
    "5. L8-工商关联：天眼查/企查查核查linyan群组8家之间的关联关系",
    "6. L9-保证金：核查汇款账户是否相同",
    "7. L10-授权代表：交叉比对16家授权委托书代理人身份证号",
]
for i, a in enumerate(actions_2):
    r = act2_start + 1 + i
    ws4.merge_cells(f'A{r}:G{r}')
    ws4.cell(row=r, column=1, value=a).font = normal_font

ws4.column_dimensions['A'].width = 10
ws4.column_dimensions['B'].width = 18
ws4.column_dimensions['C'].width = 30
ws4.column_dimensions['D'].width = 36
ws4.column_dimensions['E'].width = 12
ws4.column_dimensions['F'].width = 12
ws4.column_dimensions['G'].width = 26

# ====== SAVE ======
excel_path = os.path.join(OUTPUT_DIR, "宿舍维修项目_串标围标全量分析报告.xlsx")
wb.save(excel_path)
print(f"Excel报告已保存: {excel_path}")
print(f"共 4 个Sheet: 总览 / L1-报价规律分析 / L5-元数据分析 / L3-文本雷同检测 / 综合风险评级")
