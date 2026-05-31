import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ===== Colors & Styles =====
red_fill = PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid')
yellow_fill = PatternFill(start_color='FFF8E0', end_color='FFF8E0', fill_type='solid')
green_fill = PatternFill(start_color='E0FFE0', end_color='E0FFE0', fill_type='solid')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
subheader_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
subheader_font = Font(name='微软雅黑', bold=True, size=11)
normal_font = Font(name='微软雅黑', size=10)
bold_font = Font(name='微软雅黑', bold=True, size=10)
title_font = Font(name='微软雅黑', bold=True, size=14)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

wrap_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(wrap_text=True, vertical='center', horizontal='center')

def apply_style(ws, row, col, font=normal_font, fill=white_fill, align=wrap_align, border=thin_border):
    cell = ws.cell(row=row, column=col)
    cell.font = font
    cell.fill = fill
    cell.alignment = align
    cell.border = border

# =====================================================================
# Sheet 1: 审计检查清单
# =====================================================================
ws1 = wb.active
ws1.title = '审计检查清单'

# Title
ws1.merge_cells('A1:G1')
cell = ws1['A1']
cell.value = '商业物业管理审计检查清单 — 成都轨道资源经营管理有限公司'
cell.font = title_font
cell.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 35

ws1.merge_cells('A2:G2')
cell = ws1['A2']
cell.value = '编制基准：26份公司现行制度文件 + 北上广深行业实践对比 | 日期：2026年4月'
cell.font = Font(name='微软雅黑', size=9, color='666666')
cell.alignment = Alignment(horizontal='center')
ws1.row_dimensions[2].height = 22

# Headers
headers = ['序号', '审计领域', '审计要点', '审计方法', '风险等级', '制度依据', '行业对标']
col_widths = [5, 14, 32, 38, 8, 20, 25]
for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    ws1.cell(row=3, column=i, value=h)
    apply_style(ws1, 3, i, font=header_font, fill=header_fill, align=center_align)
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.row_dimensions[3].height = 25

# Data: (领域, 要点, 方法, 风险, 制度依据, 行业对标)
items = [
    # ---- 经营/物业 ----
    ('商业/物业经营管理', '经营台账完整性：台账-财务-流水三向核对', '选取3-5户商户，逐一核对台账记录、财务入账凭证和银行流水', '🔴高', 'P1A,P1B,P1G', '深圳万象城2000+SOP体系'),
    ('', '商户入驻流程合规性', '抽查近两年入驻/退租全流程档案，核验审批签字完整性', '🔴高', 'P1A,P1D', '深圳万象城严格资质审核'),
    ('', '租金收缴对账：差异率应<1%', '取全年台账vs财务vs银行流水数据，Excel VLOOKUP差异比对', '🔴高', 'P1B,P1G', '头部企业标准：差异率<1%'),
    ('', '租金调整/优惠审批', '特别关注关联商户租金优惠是否低于正常水平30%以上', '🔴高', 'P1A,P2F', '上海恒隆"选租户"策略'),
    ('', '商户装修管理', '装修押金收退记录核验；消防审核和施工时限检查', '🟡中', 'P1A,P3L', '行业通常：押金=3个月管理费'),
    ('', '租户关系管理', '满意度调查执行记录检查；不满意项闭环整改证据', '🟡中', 'P1B', '万象城/国贸满意度>95%'),
    ('', '自助设备租户管理', '核对设备台账与合同一致性；收益入账完整性检查', '🟡中', 'P1F', '行业标杆：每设备每月巡检1次'),
    ('', '物业服务品质达标', '随机抽查物业服务记录（保洁频次、绿化养护、安保巡逻）', '🟡中', 'P1B', '恒隆：15'保洁巡查；北京国贸：主动服务'),

    # ---- 物业用房 ----
    ('物业用房管理', '物业用房底数台账完整性', '【必查】对照建筑CAD图纸+现场逐间核实，重点查地下室和设备层夹层', '🔴高', 'P1D,P2F', '北京国贸83万㎡全部登记入账'),
    ('', '物业用房出租合规性', '已出租用房：检查有无审批依据、租金是否入账', '🔴高', 'P2F', '行业警示：警惕租金私账'),
    ('', '消防通道占用排查', '现场检查：是否被商户用作仓库或杂物间', '🔴高', 'P3D', '重大安全隐患'),
    ('', '"无台账"用房排查', '逐房间核对建筑图纸+现场实物，发现未登记用房', '🔴高', 'P1D,P2D', '多地审计发现"账外房"'),
    ('', '闲置用房处置方案', '闲置物业用房有无定期巡检和招租计划', '🟡中', 'P2F', '行业通常：闲置>3个月应启动招租'),

    # ---- 设施设备 ----
    ('设施设备管理', '设备台账完整性', '核查台账是否覆盖：电扶梯/空调/消防/变配电/安防五大系统', '🔴高', 'P3A', '北京中国尊：5万+传感器实时采集'),
    ('', '维保制度执行', '抽取3-5台核心设备，核对维保记录与合同约定频次一致性', '🔴高', 'P3A,P3B', '深圳万象城：设备维保KPI到点位'),
    ('', '检修计划执行偏差', '对比年度/月度检修计划vs实际执行情况', '🔴高', 'P3C', '行业标准：计划偏差<10%'),
    ('', '故障报告与根因分析', '核验重大故障的流程报告和预防措施', '🟡中', 'P3B', '上海恒隆：24h故障响应'),
    ('', '维修项目验收合规', '维修项目完工后验收资料完整性检查（施工记录+验收单）', '🟡中', 'P3I', '深圳湾万象城：数字孪生验收'),
    ('', '大修更新改造资金', '重大设备更新资金=专项审批文件检查；费用超预算补充审批', '🟡中', 'P3A, P1G', '行业标准：>10万元须专项审批'),

    # ---- 能耗管理 ----
    ('能耗管理', '水电费分摊合规性', '选取3-5户验算实际分摊金额与合同约定是否一致', '🔴高', '⚠无专项制度', '广州珠江新城"一票制"简化模式'),
    ('', '查表跟岗机制', '【必查建议】突击参与一次实际查表，核对抄表记录与系统读数', '🔴高', '⚠无专项制度', '行业警示：双人查表签字是基本内控'),
    ('', '能耗费用差异分析', '检查不同业态计价标准差异是否有制度依据', '🟡中', '⚠无专项制度', '深圳万象城：酬金制+能耗对赌'),
    ('', '能耗异常监控', '对比近3年同期能耗数据，识别异常波动', '🟡中', '⚠无专项制度', '北京：超限额惩罚电价'),
    ('', '商户分摊透明度', '商户是否收到清晰水电分摊明细；有无因分摊不清的纠纷记录', '🟡中', '⚠无专项制度', '行业标准：应每月出具结算单'),

    # ---- 停车场管理 ----
    ('停车场管理', '停车费收入完整性', '对比道闸系统抬杆记录与收费记录，识别异常放行', '🔴高', '⚠无专项制度', '深圳万象城：2017年无感停车'),
    ('', '私收费/私抬杆风险', '【必查】现场突击检查现金岗；重点时段：夜班和节假日', '🔴高', '⚠无专项制度', '停车场：资金跑冒滴漏高风险领域'),
    ('', '免费/优惠停车审批', '免费停车权限审批记录核查——"隐形损失"常被忽视', '🟡中', '⚠无专项制度', '行业通常：免费停车须总经理审批'),
    ('', '收费标准执行', '实际执行是否与物价部门备案标准一致', '🟡中', '⚠无专项制度', '四大城市均有明确定价指引'),

    # ---- 合同管理 ----
    ('合同管理', '合同台账完整性', '是否覆盖全部在管项目；有无无合同管理高风险情况', '🟡中', 'P2C', '华润万象生活：统一合同范本'),
    ('', '合同范本与备案', '关键条款：服务范围/收费/违约责任是否明确', '🟡中', 'P2C', '五大行：标准化合同模板'),
    ('', '续签/变更流程合规', '是否存在"事实合同"（到期未续签但持续服务）', '🟡中', 'P2C', '警惕：法律风险+履约漏洞'),
    ('', '合同用印管理', '用印是否经审批流程；有无预盖章或未按流程用印', '🟡中', 'P2C', '行业基本控制：印章分管'),
    ('', '关联交易与定价公允', '供应商关联关系排查；关联交易定价是否偏离市场公允价', '🟡中', 'P2C,P2D', '国企审计重点关注'),
    ('', '违约追责与档案', '违约追究记录完整性与追索档案', '🟡中', 'P2C', '行业标准：完整留痕'),

    # ---- 资产/招租 ----
    ('资产/招租管理', '经营性资产台账完整性', '涵盖：场地位置/面积/用途/权属/出租状态/租金/租期', '🔴高', 'P2D', '深圳物业协会：纳入不动产登记信息'),
    ('', '出租审批合规性', '新增出租有无资产评估；定向出租有无充分理由和专项审批', '🔴高', 'P2E,P2F', '华润万象生活：品牌组合管理'),
    ('', '租金定价合理性', '关联方租金是否偏离市场价30%以上（国有资产流失风险）', '🔴高', 'P2E,P2F', '上海恒隆：8%租金溢价'),
    ('', '空间租赁专项', '审验地下空间/商铺租赁的流程合规性', '🟡中', 'P1D', '行业标准流程'),
    ('', '资产处置合规', '评估-审批-交易-注销闭环完整性', '🟡中', 'P2E', '国有资产处置：须审计+评估+审批'),
    ('', '公共收益完整性', '广告位/场地出租/通信基站/快递柜收益是否全入账', '🔴高', 'P1C,P2D', '行业警示：公共收益漏收是常见问题'),
    ('', '招租过程透明度', '意向承租方登记-评审-定标是否留痕；围串标嫌疑排查', '🟡中', 'P2F', '国企招租：须公开比选'),

    # ---- 广告经营 ----
    ('广告经营', '广告位资源台账', '所有广告位（平面/电子屏/灯箱）数量、位置、状态登记完整性', '🟡中', 'P1C', '北京国贸：系统化管理'),
    ('', '广告出租收入完整性', '广告合同金额与收款记录一致性核查', '🔴高', 'P1C', '警惕：广告位免费赠送不入账'),
    ('', '广告内容合规', '上刊广告审核记录检查；公益广告占比是否达标', '🟡中', 'P1C', '《广告法》+地方广告管理规定'),
    ('', '到期广告位管理', '到期清刊/续签及时性；空置期合理性', '🟡中', 'P1C', '行业：到期前30天启动续签'),

    # ---- 安全管理 ----
    ('安全管理', '消防设施年检合规', '消防设施年检报告有效期核查；疏散通道畅通', '🔴高', 'P3D', '北京：月度防火检查+年度检测'),
    ('', '安全管控制度执行', '安全巡查记录完整性；隐患整改闭环率', '🔴高', 'P3E,P3F', '行业标准：隐患整改率100%'),
    ('', '应急预案演练', '突发事件预案（火灾/停电/漏水）实际演练记录和频次', '🟡中', 'P3H', '行业标准：每半年至少1次综合演练'),
    ('', '风险分级管控', '高风险点位专项管控措施是否到位', '🟡中', 'P3J', '深圳：双重预防机制全覆盖'),
    ('', '安全教育培训', '员工安全培训计划和执行记录；新员工岗前培训到位率', '🟡中', 'P3G', '行业标准：全员年度培训>8小时'),
    ('', '施工安全管控', '商户装修动火作业审批和现场监护记录', '🟡中', 'P3L', '行业标准：动火作业双人监护'),
]

row = 4
seq = 0
current_area = ''
for area, point, method, risk, basis, benchmark in items:
    if area:
        current_area = area
    risk_letter = risk[0]  # 高/中
    
    seq += 1
    ws1.cell(row=row, column=1, value=seq)
    ws1.cell(row=row, column=2, value=current_area)
    ws1.cell(row=row, column=3, value=point)
    ws1.cell(row=row, column=4, value=method)
    ws1.cell(row=row, column=5, value=risk)
    ws1.cell(row=row, column=6, value=basis)
    ws1.cell(row=row, column=7, value=benchmark)
    
    fill = red_fill if '🔴' in risk else yellow_fill
    for c in range(1, 8):
        apply_style(ws1, row, c, fill=fill)
    apply_style(ws1, row, 5, font=Font(name='微软雅黑', bold=True, size=10))  # risk bold
    
    ws1.row_dimensions[row].height = 40
    row += 1

# Freeze panes
ws1.freeze_panes = 'A4'

# Auto filter
ws1.auto_filter.ref = f'A3:G{row-1}'

# =====================================================================
# Sheet 2: 制度比对分析
# =====================================================================
ws2 = wb.create_sheet('制度比对分析')

# Title
ws2.merge_cells('A1:F1')
cell = ws2['A1']
cell.value = '全领域制度比对分析 — 成都轨道资源 vs 行业标杆'
cell.font = Font(name='微软雅黑', bold=True, size=14)
cell.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 35

headers2 = ['管理领域', '成都轨道制度覆盖', '制度评价', '行业标杆做法', '差距/建议', '优先级']
col_widths2 = [16, 20, 12, 32, 32, 8]
for i, (h, w) in enumerate(zip(headers2, col_widths2), 1):
    ws2.cell(row=2, column=i, value=h)
    apply_style(ws2, 2, i, font=header_font, fill=header_fill, align=center_align)
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.row_dimensions[2].height = 25

comparison = [
    ('商业/物业经营管理',
     '✅ P1A《商业经营管理制度》\n✅ P1B《物业经营管理制度》\n✅ P1D《空间开发制度》\n✅ P1F《自助设备租户手册》\n✅ P1G《经营管理制度总纲》',
     '★★★★★\n体系完善',
     '【华润万象生活】"一点万象"数字化平台\n2000+ SOP操作流程\n商户全生命周期管理系统',
     '制度体系完整，建议推进SOP数字化、建立商户全生命周期管理台账',
     '🟡中'),

    ('资产/招租管理',
     '✅ P2D《资产管理制度》\n✅ P2E《资产评估管理办法》\n✅ P2F《资产招商管理办法》\n✅ P1D《空间开发制度》',
     '★★★★★\n三层体系\n完整覆盖',
     '【华润万象生活】品牌组合管理系统\n招商评审-签约-进场闭环流程\n数据驱动资产管理',
     '制度体系扎实，建议增加品牌组合管理和招商过程数字化',
     '🟡中'),

    ('合同管理',
     '✅ P2C《合同管理制度》\n2024版，14页',
     '★★★★☆\n覆盖齐全',
     '【华润】统一合同范本\n【五大行】标准化模板\n合同全生命周期系统管理',
     '现有制度基本满足需求，可考虑增加合同范本库和到期自动预警',
     '🟡中'),

    ('设施设备管理',
     '✅ P3A《设施设备管理办法》\n✅ P3B《设施设备故障管理办法》\n✅ P3C《检修计划管理办法》\n✅ P3I《维修项目验收细则》',
     '★★★★★\n完整的\n四层体系',
     '【北京中国尊】5万+传感器\nIBMS智能楼宇系统\nAI预测性维护\n数字孪生验收',
     '制度体系极其完整，建议推进设备数字化（物联网传感+预测维护）',
     '🟡中'),

    ('广告经营',
     '✅ P1C《广告经营管理办法》\n2025版，31页',
     '★★★★★\n最新版，\n覆盖完整',
     '【北京国贸】广告位系统化管理\n动态定价策略\n电子屏数字化运营',
     '制度很新很完整，可考虑增加动态定价和电子化运营',
     '🟡中'),

    ('招标采购监督',
     '✅ P2A《招标采购监督委员会议事规则》\n2024版',
     '★★★★☆\n议事规则完整',
     '【行业标杆】"十大原则"采购体系\n电子招标平台全程留痕\n供应商分级管理',
     '可增加电子招标平台、供应商绩效评估制度',
     '🟡中'),

    ('能耗管理',
     '❌ 无专项制度',
     '☆☆☆☆☆\n制度空白',
     '【标杆】酬金制+能耗对赌\n【北京】冰蓄冷+惩罚电价\n【深圳能耗KPI体系】',
     '【急需建立】①能耗分摊标准 ②查表跟岗制度 ③异常监控机制 ④商户透明度规范',
     '🔴高'),

    ('停车场管理',
     '❌ 无专项制度',
     '☆☆☆☆☆\n制度空白',
     '【深圳万象城】无感停车\n【行业】道闸系统+收入台账\n【内控】双人交接班制度',
     '【急需建立】①停车收费管理 ②免费/优惠审批 ③现金岗内控 ④系统操作规范',
     '🔴高'),

    ('物业用房管理',
     '⚠ 无专项制度\n（依赖P1D+P2F+P2D衔接覆盖）',
     '★★☆☆☆\n制度缺位',
     '【北京国贸】全部登记入账\n【行业】台账-图纸-实物三一致',
     '【建议建立】物业用房专项管理制度（台账登记/出租审批/巡检巡查）',
     '🔴高'),

    ('安全管理',
     '✅ P3D《消防安全管理细则》\n✅ P3E《安全管控制度》\n✅ P3F《安全生产责任制》\n✅ P3G《安全教育培训》\n✅ P3H《应急预案》\n✅ P3J《风险分级管控》\n✅ P3K《安全绩效考核》\n✅ P3L《施工安全管理》\n共9份制度',
     '★★★★★\n极其完善的\n安全制度体系',
     '行业先进：双重预防机制\n消防安全巡查体系\n定期应急演练制度',
     '安全制度体系是公司的真正优势——9份制度覆盖全面。建议：把制度优势转化为可量化的安全运营KPI指标',
     '🟡中'),

    ('数字化/信息化管理',
     '⚠ 未纳入本次分析范围\n（制度目录中有信息化类文件）',
     '★★★☆☆',
     '【万物云】"灵石"AI平台\n能耗降低12-18%\n人均管理面积提升40%\n数字孪生系统',
     '建议进行信息化专项审计——物联网传感、系统数据安全、操作日志追溯',
     '🟡中'),

    ('财务管理/内控',
     '⚠ 未纳入本次分析范围\n（制度目录中有财务管理类文件）',
     '★★★☆☆',
     '行业标准：全面预算管理\n资产处置审批闭环\n审计-财务联动机制',
     '建议进行财务管理专项审计——预算执行/成本控制/资金管理',
     '🟡中'),
]

row2 = 3
for area, coverage, rating, benchmark, gap, priority in comparison:
    ws2.cell(row=row2, column=1, value=area)
    ws2.cell(row=row2, column=2, value=coverage)
    ws2.cell(row=row2, column=3, value=rating)
    ws2.cell(row=row2, column=4, value=benchmark)
    ws2.cell(row=row2, column=5, value=gap)
    ws2.cell(row=row2, column=6, value=priority)
    
    fill = red_fill if '空白' in rating or '制度空白' in rating or '制度缺位' in rating else (yellow_fill if '★★' in rating else white_fill)
    for c in range(1, 7):
        apply_style(ws2, row2, c, fill=fill)
    # priority column color
    if '高' in priority:
        apply_style(ws2, row2, 6, font=Font(name='微软雅黑', bold=True, size=10, color='CC0000'), fill=red_fill)
    
    ws2.row_dimensions[row2].height = 80
    row2 += 1

ws2.freeze_panes = 'A3'

# =====================================================================
# Sheet 3: 审计方法速查
# =====================================================================
ws3 = wb.create_sheet('审计方法速查')

ws3.merge_cells('A1:D1')
cell = ws3['A1']
cell.value = '审计方法速查 — 按操作类型分类'
cell.font = Font(name='微软雅黑', bold=True, size=14)
cell.alignment = Alignment(horizontal='center')
ws3.row_dimensions[1].height = 35

headers3 = ['方法类型', '方法名称', '具体操作', '适用场景']
col_widths3 = [14, 18, 40, 40]
for i, (h, w) in enumerate(zip(headers3, col_widths3), 1):
    ws3.cell(row=2, column=i, value=h)
    apply_style(ws3, 2, i, font=header_font, fill=header_fill, align=center_align)
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.row_dimensions[2].height = 25

methods = [
    ('现场核实', '台账-实物比对', '对照建筑图纸+现场逐间核实物业用房', '物业用房管理、设备台账'),
    ('现场核实', '突击跟岗', '突然检查一次抄表/收费过程', '能耗查表、停车场收费'),
    ('现场核实', '随机巡查', '不定路线巡查消防通道、设备房、空置用房', '消防通道占用、设施设备状态'),
    ('数据比对', '三向核对', '台账vs财务系统vs银行流水VLOOKUP比对', '租金收缴、停车费、广告费'),
    ('数据比对', '核验三方关联', '合同台账vs供应商库vs股权关联排查', '关联交易、围串标嫌疑'),
    ('数据比对', '凭证-合同一致性', '逐笔核对金额、周期、收款方', '合同执行、装修押金收退'),
    ('文档审查', '审批链条追溯', '从决策文档反向追溯签字审批完整性', '招租审批、租金优惠、重大维修'),
    ('文档审查', '制度更新检查', '新旧制度版本比对+传达记录', '所有制度的时效性和执行'),
    ('数据比对', '同期数据对比', '同比去年/前年同期的能耗、收入数据', '能耗异常、收入波动'),
    ('系统审计', '操作日志追溯', '检查系统登录/操作/修改日志', '财务系统、停车场系统、招商系统'),
]

row3 = 3
for mtype, mname, mop, mscope in methods:
    ws3.cell(row=row3, column=1, value=mtype)
    ws3.cell(row=row3, column=2, value=mname)
    ws3.cell(row=row3, column=3, value=mop)
    ws3.cell(row=row3, column=4, value=mscope)
    for c in range(1, 5):
        apply_style(ws3, row3, c, fill=white_fill)
    ws3.row_dimensions[row3].height = 35
    row3 += 1

ws3.freeze_panes = 'A3'

# Save
path = r'C:\Users\scrccpa\.openclaw\workspace\商业物业管理审计检查清单.xlsx'
wb.save(path)
print(f'Excel saved to: {path}')
print(f'Sheet 1: 审计检查清单 ({len(items)} items)')
print(f'Sheet 2: 制度比对分析 ({len(comparison)} areas)')
print(f'Sheet 3: 审计方法速查 ({len(methods)} methods)')
