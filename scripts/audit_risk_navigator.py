#!/usr/bin/env python3
"""
审计风险导航 — 基于审计大模型框架的智能核查方向推荐

参考: 柳絮/李欣潼《审计大模型》场景3(智能分析)
功能: 输入审计对象/范围 → 输出推荐核查方向 + 风险特征 + 分析方法
"""
import sys, io, argparse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

RED = PatternFill(patternType='solid', fgColor='FFD7D7')
YEL = PatternFill(patternType='solid', fgColor='FFF3CD')
GRN = PatternFill(patternType='solid', fgColor='D4EDDA')
HEADER = PatternFill(patternType='solid', fgColor='1A3A6E')
H = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
N = Font(name='Microsoft YaHei', size=10)
TH = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
C = Alignment(horizontal='center', vertical='center', wrap_text=True)
L = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ── 风险导航知识库 (按审计领域组织) ──
RISK_NAVIGATOR = {
    '招投标采购审计': {
        'description': '政府采购/工程招标/物资采购全流程审计',
        'directions': [
            {
                'name': '围标串标识别',
                'risk_features': [
                    '投标人报价呈等差数列/阶梯分布',
                    '多家投标人技术方案高度雷同(TF-IDF>80%)',
                    '不同投标人PDF元数据同源(Author/Creator一致)',
                    '嵌入图片哈希相同(同一台扫描仪)',
                    '保证金缴纳时间差<10分钟',
                    '投标人数3-4家(最优围标规模)',
                ],
                'methods': 'Apriori关联规则 + TF-IDF相似度 + PDF元数据交叉 + JPEG量化表指纹',
                'tools': 'procurement-audit-models L1-L19 + apriori-audit + unstructured-audit-data',
                'priority': '🔴高',
            },
            {
                'name': '招标文件合规性',
                'risk_features': [
                    '设置业绩/奖项/专利等不合理资格条件',
                    '地域限制条款(本地/本市/省内)',
                    '技术参数指向特定品牌或型号',
                    '评分标准中主观分值占比过高(>30%)',
                    '招标控制价明显高于市场价',
                ],
                'methods': 'Word关键词扫描 + 技术参数倾向性分析 + 评分标准合理性审查',
                'tools': 'unstructured-audit-data(batch_word_scan) + 手工复核',
                'priority': '🔴高',
            },
            {
                'name': '节资率异常',
                'risk_features': [
                    '节资率<Q1(下四分位数)',
                    '同类项目节资率持续下降趋势',
                    '中标价=预算上限(节资率=0)',
                ],
                'methods': '箱线图IQR异常检测 + 时间序列趋势分析',
                'tools': 'procurement-audit-models L13(savings_rate)',
                'priority': '🟡中',
            },
            {
                'name': '陪标专业户',
                'risk_features': [
                    '多次投标从未中标(前5%)',
                    '特定企业组合反复同时出现',
                    '中标集中在极少数企业(>70%)',
                ],
                'methods': '投标频次统计 + Jaccard共现系数 + 中标集中度',
                'tools': 'procurement-audit-models L11/L15',
                'priority': '🟡中',
            },
            {
                'name': '评标专家违规',
                'risk_features': [
                    '专家在投标单位缴纳社保(利益冲突)',
                    '不同项目评标专家完全一致',
                    '专家签到时间晚于评标时间',
                ],
                'methods': '专家库+社保数据交叉比对 + 时间序列校验',
                'tools': 'procurement-audit-models L19',
                'priority': '🟢低(需额外数据)',
            },
        ],
    },

    '财务收支审计': {
        'description': '预算执行/财务收支/专项资金审计',
        'directions': [
            {
                'name': '收入真实性测试',
                'risk_features': [
                    '年末/季末收入异常集中',
                    '大额整数交易占比异常',
                    '关联方交易价格偏离市场公允价',
                    '应收账款与收入增长不匹配',
                ],
                'methods': 'Benford定律首位数分布 + 时间序列异常检测 + 关联交易穿透',
                'tools': 'financial-fraud-detection + data-analyst-cn',
                'priority': '🔴高',
            },
            {
                'name': '成本完整性验证',
                'risk_features': [
                    '毛利率异常波动(与行业/历史比较)',
                    '存货跌价准备计提不足',
                    '预提费用/应付暂估金额异常',
                ],
                'methods': '毛利率趋势分析 + 存货周转率分析 + 3σ异常检测',
                'tools': 'aloudata-anomaly-detection + data-analyst-cn',
                'priority': '🔴高',
            },
            {
                'name': '专项资金合规性',
                'risk_features': [
                    '专项资金未专账核算',
                    '资金拨付进度与项目进度不匹配',
                    '大额现金支付或转入个人账户',
                ],
                'methods': '资金流向穿透分析 + 项目进度vs付款进度比对',
                'tools': 'data-analyst-cn + audit_finding_processor',
                'priority': '🟡中',
            },
            {
                'name': '费用报销规范性',
                'risk_features': [
                    '连号发票/同一商户高频消费',
                    '节假日前后集中报销',
                    '单笔金额略低于审批权限阈值',
                    '跨部门人员同时报销同类型费用',
                ],
                'methods': '发票号聚类 + 时间模式分析 + 审批阈值规避检测',
                'tools': 'apriori-audit(mode=frequent)',
                'priority': '🟡中',
            },
        ],
    },

    '资产管理审计': {
        'description': '固定资产/无形资产/存货管理审计',
        'directions': [
            {
                'name': '资产账实不符',
                'risk_features': [
                    '资产卡片信息与实物标签不一致',
                    '已报废资产仍在账面',
                    '已投入使用资产未转固',
                    '资产出租/出借未履行审批',
                ],
                'methods': '资产抽样盘点 + 资产卡片vs折旧表核对 + 使用权登记清查',
                'tools': 'data-analyst-cn + 手工盘点',
                'priority': '🔴高',
            },
            {
                'name': '资产处置不合规',
                'risk_features': [
                    '资产处置未经资产评估',
                    '处置价格明显低于评估价或市场价',
                    '报废资产去向不明',
                ],
                'methods': '处置价格vs评估价格比对 + 报废资产追踪',
                'tools': 'data-analyst-cn',
                'priority': '🟡中',
            },
        ],
    },

    '工程审计': {
        'description': '工程预算/结算/全过程跟踪审计',
        'directions': [
            {
                'name': '工程款超付',
                'risk_features': [
                    '计量支付超出实际完成工程量',
                    '预付款未按合同约定扣回',
                    '质保金提前退还',
                    '变更签证未经规范审批',
                ],
                'methods': '工程量清单vs计量支付比对 + 合同条款履约核查',
                'tools': 'data-analyst-cn',
                'priority': '🔴高',
            },
            {
                'name': '工程造价虚高',
                'risk_features': [
                    '综合单价高于同期同类项目',
                    '材料价格偏离信息价/市场价',
                    '工程量计算错误(多算/重算)',
                    '定额套用错误(高套定额)',
                ],
                'methods': '单价横向对比 + 材料价格信息库比对 + 工程量复核',
                'tools': 'data-analyst-cn + forecast-simulation(What-if)',
                'priority': '🔴高',
            },
            {
                'name': '工程进度滞后',
                'risk_features': [
                    '实际进度严重偏离计划进度',
                    '关键路径任务反复延期',
                    '进度款支付与形象进度不匹配',
                ],
                'methods': '进度偏差分析 + 挣值管理(EVM)',
                'tools': 'forecast-simulation + data-analyst-cn',
                'priority': '🟡中',
            },
        ],
    },

    '绩效评价': {
        'description': '财政支出绩效评价/部门整体绩效评价',
        'directions': [
            {
                'name': '绩效目标未完成',
                'risk_features': [
                    '产出数量/质量/时效低于目标值',
                    '成本节约率目标未实现',
                    '满意度指标不达标',
                    '绩效自评与实际不符',
                ],
                'methods': '绩效目标vs实际完成比对 + 自评报告真实性复核',
                'tools': 'data-analyst-cn',
                'priority': '🔴高',
            },
            {
                'name': '绩效指标设置不合理',
                'risk_features': [
                    '指标值明显偏低(易达成)',
                    '指标不可量化/不可考核',
                    '效益指标缺乏佐证依据',
                ],
                'methods': '指标SMART原则审核 + 同类项目指标对比',
                'tools': '手工审核',
                'priority': '🟡中',
            },
        ],
    },
}


def navigate(audit_type: str, output: str = None):
    """风险导航: 输入审计类型 → 输出核查方向矩阵"""

    if audit_type not in RISK_NAVIGATOR:
        # Fuzzy match
        matches = [k for k in RISK_NAVIGATOR if audit_type in k]
        if matches:
            audit_type = matches[0]
        else:
            print(f"未找到匹配的审计类型: {audit_type}")
            print(f"可用类型: {', '.join(RISK_NAVIGATOR.keys())}")
            return

    nav = RISK_NAVIGATOR[audit_type]
    print(f"\n📋 {audit_type}")
    print(f"   {nav['description']}")
    print(f"\n{'='*60}")

    for i, d in enumerate(nav['directions'], 1):
        marker = '🔴' if d['priority'] == '🔴高' else ('🟡' if d['priority'] == '🟡中' else '🟢')
        print(f"\n{marker} 方向{i}: {d['name']} [{d['priority']}]")
        print(f"   风险特征:")
        for f in d['risk_features']:
            print(f"     • {f}")
        print(f"   方法: {d['methods']}")
        print(f"   工具: {d['tools']}")

    if output:
        wb = Workbook()
        ws = wb.active
        ws.title = '风险导航'

        ws.merge_cells('A1:F1')
        ws['A1'] = f'{audit_type} — 智能风险导航'
        ws['A1'].font = Font(name='Microsoft YaHei', size=16, bold=True, color='1A3A6E')

        sub_h = ['序号', '核查方向', '风险特征(摘要)', '分析方法', '推荐工具', '优先级']
        for c, h in enumerate(sub_h, 1):
            cl = ws.cell(row=3, column=c, value=h)
            cl.font = H; cl.fill = HEADER; cl.alignment = C; cl.border = TH

        for i, d in enumerate(nav['directions']):
            r = i + 4
            vals = [
                i + 1,
                d['name'],
                '\n'.join(d['risk_features']),
                d['methods'],
                d['tools'],
                d['priority'],
            ]
            for c, val in enumerate(vals, 1):
                cl = ws.cell(row=r, column=c, value=val)
                cl.font = N; cl.alignment = L; cl.border = TH
                if c == 6:
                    color_map = {'🔴高': RED, '🟡中': YEL, '🟢低': GRN}
                    cl.fill = color_map.get(val, None)
                    cl.alignment = C

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 45
        ws.column_dimensions['F'].width = 12

        wb.save(output)
        print(f"\n✅ 导航表: {output}")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='审计风险导航 — 智能推荐核查方向',
        epilog='''
示例:
  python audit_risk_navigator.py "招投标采购审计"
  python audit_risk_navigator.py "财务收支审计" --o 风险导航.xlsx
  python audit_risk_navigator.py --list
        '''
    )
    parser.add_argument('type', nargs='?', help='审计类型 (如: 招投标采购审计)')
    parser.add_argument('--o', '--output', dest='output', help='输出Excel路径')
    parser.add_argument('--list', action='store_true', help='列出所有可用审计类型')
    args = parser.parse_args()

    if args.list:
        print("可用审计类型:")
        for k, v in RISK_NAVIGATOR.items():
            print(f"  {k}: {v['description']}")
    elif args.type:
        navigate(args.type, args.output)
    else:
        # Interactive
        print("💰 审计风险智能导航")
        print("可用领域:")
        for i, (k, v) in enumerate(RISK_NAVIGATOR.items(), 1):
            print(f"  {i}. {k}")
        print()
        choice = input("选择审计类型 (输入编号或名称) > ").strip()
        try:
            idx = int(choice) - 1
            keys = list(RISK_NAVIGATOR.keys())
            if 0 <= idx < len(keys):
                navigate(keys[idx], args.output)
        except ValueError:
            navigate(choice, args.output)
