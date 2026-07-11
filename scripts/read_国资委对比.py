#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""读取国资委三文对比Excel"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')

fp = r'C:\Users\scrccpa\Desktop\轨道培训\国资委2026年三文对比-1号2号15号.xlsx'

if not os.path.exists(fp):
    print(f"❌ 文件不存在: {fp}")
    # 搜索桌面所有xlsx
    desktop = r'C:\Users\scrccpa\Desktop'
    print(f"\n📂 桌面内容:")
    for f in os.listdir(desktop):
        print(f"  {f}")
    # 找轨道培训目录
    d = os.path.join(desktop, '轨道培训')
    if os.path.isdir(d):
        print(f"\n📂 轨道培训 目录内容:")
        for f in os.listdir(d):
            print(f"  {f}")
    sys.exit(1)

import openpyxl
wb = openpyxl.load_workbook(fp, data_only=True)
print(f"✅ Sheet列表: {wb.sheetnames}")

for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"\n{'#'*70}")
    print(f"# Sheet: {sn} ({ws.max_row}行 x {ws.max_column}列)")
    print(f"{'#'*70}")
    
    for r in range(1, min(ws.max_row+1, 150)):
        parts = []
        for c in range(1, min(ws.max_column+1, 25)):
            v = ws.cell(r, c).value
            if v is not None:
                txt = str(v).strip()[:200]
                parts.append(f"C{c}={txt}")
        if parts:
            print(f"  R{r:4d}: {' | '.join(parts)}")
    
    if ws.max_row > 150:
        print(f"  ... ({ws.max_row-150} more rows)")

print("\n完毕")
