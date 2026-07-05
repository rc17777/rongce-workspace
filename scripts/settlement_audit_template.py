# -*- coding: utf-8 -*-
"""结算审核底稿模板生成脚本
依据：JJYJ50-003-2026 第7章 + 13项结算顺序框架 + 4个审减重点
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

# ═══ 配色方案（融策标书模板同款） ═══
DEEP_BLUE = "0A1F3F"
TEAL = "1A5C6E"
COPPER_GOLD = "C5955C"
WARM_GRAY = "F5F2EC"
WHITE = "FFFFFF"
LIGHT_GRAY = "E8E8E8"
RED_ALERT = "C0392B"
GREEN_PASS = "27AE60"

# ═══ 样式定义 ═══
thin_border = Border(
    left=Side(style='thin', color='999999'),
    right=Side(style='thin', color='999999'),
    top=Side(style='thin', color='999999'),
    bottom=Side(style='thin', color='999999')
)

header_fill = PatternFill(start_color=DEEP_BLUE, end_color=DEEP_BLUE, fill_type='solid')
header_font = Font(name='微软雅黑', bold=True, color=WHITE, size=11)
sub_header_fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type='solid')
sub_header_font = Font(name='微软雅黑', bold=True, color=WHITE, size=10)
title_font = Font(name='微软雅黑', bold=True, color=DEEP_BLUE, size=14)
subtitle_font = Font(name='微软雅黑', bold=True, color=TEAL, size=11)
normal_font = Font(name='微软雅黑', size=10, color='333333')
small_font = Font(name='微软雅黑', size=9, color='666666')
bold_font = Font(name='微软雅黑', bold=True, size=10, color='333333')
gold_font = Font(name='微软雅黑', bold=True, color=COPPER_GOLD, size=11)
alert_font = Font(name='微软雅黑', bold=True, color=RED_ALERT, size=10)
pass_font = Font(name='微软雅黑', bold=True, color=GREEN_PASS, size=10)

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
left_top_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

warm_fill = PatternFill(start_color=WARM_GRAY, end_color=WARM_GRAY, fill_type='solid')
light_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')
white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type='solid')

def apply_border(ws, row, col, style=thin_border):
    ws.cell(row=row, column=col).border = style

def set_cell(ws, row, col, value, font=normal_font, fill=None, alignment=left_align, border=thin_border, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    return cell

def set_header_row(ws, row, headers, fill=header_fill, font=header_font):
    for i, h in enumerate(headers, 1):
        set_cell(ws, row, i, h, font=font, fill=fill, alignment=center_align)

# ═══ 创建工作簿 ═══
wb = openpyxl.Workbook()

# ═══════════════════════════════════════════
# Sheet 1: 使用说明
# ═══════════════════════════════════════════
ws0 = wb.active
ws0.title = "使用说明"
ws0.sheet_properties.tabColor = DEEP_BLUE

ws0.merge_cells('A1:H1')
set_cell(ws0, 1, 1, "工程竣工结算审核底稿模板", font=title_font, fill=None, alignment=center_align, border=None)

ws0.merge_cells('A2:H2')
set_cell(ws0, 2, 1, "依据：重庆市建设工程工程量清单计价实施细则 JJYJ50-003-2026 第7章", font=subtitle_font, fill=None, alignment=center_align, border=None)

ws0.merge_cells('A3:H3')
set_cell(ws0, 3, 1, f"版本：v1.0 | 生成日期：{datetime.now().strftime('%Y-%m-%d')} | 四川融策工程咨询有限公司", font=small_font, fill=None, alignment=center_align, border=None)

instructions = [
    ("📋 模板结构", ""),
    ("", "本模板包含5个工作表："),
    ("", "① 使用说明（本页）—— 模板结构和使用方法"),
    ("", "② 13项结算审核底稿 —— 核心工作底稿，逐项审核"),
    ("", "③ 三组易混淆事项对照 —— 清单缺陷vs变更、变更vs新增、物价vs法规"),
    ("", "④ 四个审减重点清单 —— 材差/措施/增值税/时限"),
    ("", "⑤ 措施项目调整检查 —— 四类例外逐项核实"),
    ("", ""),
    ("🔧 使用方法", ""),
    ("", "1. 在「项目信息」区域填写工程基本信息"),
    ("", "2. 按13项顺序逐项审核，在「审核结论」列选择：✓通过 / ⚠存疑 / ✗审减"),
    ("", "3. 有审减的，在「审减金额」列填写具体金额（元）"),
    ("", "4. 「存疑/审减」项必须在「审核备注/依据」列写明理由和规范依据"),
    ("", "5. 「三组对照」和「四重点」工作表作为交叉验证工具"),
    ("", "6. 审核完成后，自动汇总审减金额到合计行"),
    ("", ""),
    ("⚠ 关键法规引用", ""),
    ("", "· 7.2.1 — 13项结算顺序（铁规矩，不得打乱）"),
    ("", "· 7.2.2第2款 — 材料暂估价价差不计管理费和利润"),
    ("", "· 7.2.3第1款 — 按项计价的措施项目原则上不调（仅四类例外）"),
    ("", "· 5.2.7 — 专业工程暂估价不计入增值税计算基础"),
    ("", "· 7.1.3 — 结算申报时限按合同约定执行"),
    ("", ""),
    ("📝 备注", ""),
    ("", "· 黄色高亮单元格为必填项"),
    ("", "· 本模板适配2026年7月1日起施行的重庆新规"),
    ("", "· 其他地区可参照使用，法规依据需替换为当地版本"),
]

r = 5
for label, content in instructions:
    if label:
        set_cell(ws0, r, 1, label + content, font=bold_font if label else normal_font, fill=None, alignment=left_align, border=None)
    else:
        set_cell(ws0, r, 1, "    " + content, font=normal_font, fill=None, alignment=left_align, border=None)
    r += 1

ws0.column_dimensions['A'].width = 100

# ═══════════════════════════════════════════
# Sheet 2: 13项结算审核底稿（主表）
# ═══════════════════════════════════════════
ws1 = wb.create_sheet("13项结算审核底稿")
ws1.sheet_properties.tabColor = TEAL

# -- 项目信息区 --
project_info = [
    ("工程名称", "", "合同编号", ""),
    ("发包人", "", "承包人", ""),
    ("合同金额（元）", "", "送审金额（元）", ""),
    ("审核人", "", "审核日期", ""),
    ("编制人", "", "编制日期", ""),
]

ws1.merge_cells('A1:H1')
set_cell(ws1, 1, 1, "工程竣工结算审核底稿", font=title_font, fill=None, alignment=center_align, border=None)
ws1.merge_cells('A2:H2')
set_cell(ws1, 2, 1, "依据：JJYJ50-003-2026 第7章 | 13项结算顺序逐项审核", font=subtitle_font, fill=None, alignment=center_align, border=None)

r = 3
for label1, val1, label2, val2 in project_info:
    set_cell(ws1, r, 1, label1, font=bold_font, fill=warm_fill, alignment=center_align)
    ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    set_cell(ws1, r, 2, val1, font=normal_font, fill=white_fill, alignment=left_align)
    set_cell(ws1, r, 4, label2, font=bold_font, fill=warm_fill, alignment=center_align)
    ws1.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    set_cell(ws1, r, 5, val2, font=normal_font, fill=white_fill, alignment=left_align)
    r += 1

# -- 主表表头 --
r += 1
headers = ["序号", "结算项目", "含义/内容", "审核要点", "送审金额(元)", "审定金额(元)", "审减金额(元)", "审核结论", "审核依据/备注"]
set_header_row(ws1, r, headers)
ws1.row_dimensions[r].height = 30

# -- 13项数据 --
items = [
    (1, "合同清单单价及总价", "原合同范围内价格，合同单价×实际完成量", "量与单价是否合规", None),
    (2, "清单缺陷调整", "清单本身错误的调整（发包人责任）", "⚠区分是缺陷还是变更", None),
    (3, "暂列金额调整", "暂列金已用部分和结余", "做没做、该不该用", None),
    (4, "暂估价调整\n（材料/专业工程）", "材料税前调差，专业工程按结算", "材差不计管理费利润", "7.2.2第2款"),
    (5, "总承包服务费调整", "仅适用总承包合同", "服务内容是否履约", None),
    (6, "计日工调整", "签证确认+合同单价", "签证是否闭环", None),
    (7, "物价变化调整", "材料价格波动调差", "基准价和调差公式", None),
    (8, "法律法规变化调整", "税率变化等政策行为", "政策变动时间节点", None),
    (9, "工程变更增减", "设计变更主动改变", "变更是否签认", None),
    (10, "新增工程", "合同外新增项目", "新增是否属于索赔范畴", None),
    (11, "工程索赔", "延误/费用索赔", "证据链是否完整", None),
    (12, "违约金", "逾期罚款等", "是否触发违约条款", None),
    (13, "其他价款", "兜底项目", "合理性和必要性", None),
]

data_start = r + 1
for i, (seq, name, meaning, check_point, ref) in enumerate(items):
    row = data_start + i
    ws1.row_dimensions[row].height = 45
    row_fill = warm_fill if i % 2 == 0 else white_fill
    
    set_cell(ws1, row, 1, seq, font=bold_font, fill=row_fill, alignment=center_align)
    set_cell(ws1, row, 2, name, font=bold_font, fill=row_fill, alignment=left_align)
    set_cell(ws1, row, 3, meaning, font=normal_font, fill=row_fill, alignment=left_align)
    set_cell(ws1, row, 4, check_point, font=normal_font if not check_point.startswith("⚠") else alert_font, fill=row_fill, alignment=left_align)
    # 金额列（白色/浅灰可编辑区）
    for col in [5, 6, 7]:
        set_cell(ws1, row, col, "", font=normal_font, fill=white_fill, alignment=center_align, number_format='#,##0.00')
    # 审核结论（下拉）
    set_cell(ws1, row, 8, "", font=normal_font, fill=white_fill, alignment=center_align)
    ref_text = ref if ref else ""
    set_cell(ws1, row, 9, ref_text, font=small_font, fill=row_fill, alignment=left_align)

# 合计行
sum_row = data_start + 13
ws1.row_dimensions[sum_row].height = 28
for col in range(1, 10):
    sum_fill = PatternFill(start_color=COPPER_GOLD, end_color=COPPER_GOLD, fill_type='solid')
    if col <= 4:
        set_cell(ws1, sum_row, col, "合  计" if col == 2 else "", font=Font(name='微软雅黑', bold=True, color=WHITE, size=11), fill=sum_fill, alignment=center_align)
    elif col == 8:
        set_cell(ws1, sum_row, col, "审减合计 →", font=Font(name='微软雅黑', bold=True, color=WHITE, size=11), fill=sum_fill, alignment=center_align)
    else:
        set_cell(ws1, sum_row, col, "", font=Font(name='微软雅黑', bold=True, color=WHITE, size=11), fill=sum_fill, alignment=center_align, number_format='#,##0.00')

# 审减合计公式
ws1.cell(row=sum_row, column=5).value = f'=SUM(E{data_start}:E{data_start+12})'
ws1.cell(row=sum_row, column=6).value = f'=SUM(F{data_start}:F{data_start+12})'
ws1.cell(row=sum_row, column=7).value = f'=SUM(G{data_start}:G{data_start+12})'
for col in [5,6,7]:
    set_cell(ws1, sum_row, col, ws1.cell(row=sum_row, column=col).value, 
             font=Font(name='微软雅黑', bold=True, color=WHITE, size=11), fill=sum_fill, alignment=center_align, number_format='#,##0.00')

# -- 数据验证：审核结论下拉 --
dv_conclusion = DataValidation(type="list", formula1='"✓通过,⚠存疑,✗审减"', allow_blank=True)
dv_conclusion.error = "请选择：✓通过 / ⚠存疑 / ✗审减"
dv_conclusion.errorTitle = "无效输入"
ws1.add_data_validation(dv_conclusion)
for i in range(13):
    dv_conclusion.add(ws1.cell(row=data_start + i, column=8))

# -- 列宽 --
col_widths = {1:6, 2:22, 3:30, 4:28, 5:16, 6:16, 7:16, 8:12, 9:28}
for col, w in col_widths.items():
    ws1.column_dimensions[get_column_letter(col)].width = w

# 冻结窗格（项目信息+表头）
ws1.freeze_panes = f'A{data_start}'

# ═══════════════════════════════════════════
# Sheet 3: 三组易混淆事项对照
# ═══════════════════════════════════════════
ws2 = wb.create_sheet("三组易混淆事项对照")
ws2.sheet_properties.tabColor = "8E44AD"

ws2.merge_cells('A1:H1')
set_cell(ws2, 1, 1, "三组最容易混淆的结算事项对照表", font=title_font, fill=None, alignment=center_align, border=None)

confusion_groups = [
    {
        "group": "第1组：清单缺陷 vs 工程变更",
        "item_a": "清单缺陷（第2项）",
        "item_b": "工程变更（第9项）",
        "desc_a": "清单本身编错了——项目特征描述少了、工程量算错了",
        "desc_b": "图纸或要求主动改变——C30混凝土改成C40",
        "responsibility_a": "发包人责任，调整价款由发包人承担",
        "responsibility_b": "双方协商调整",
        "trap": "⚠ 承包人常把清单缺陷包装成工程变更——变更的计价弹性更大",
        "check_method": "仔细比对变更通知单和原始清单，判断是「编错了」还是「主动改了」"
    },
    {
        "group": "第2组：工程变更 vs 新增工程",
        "item_a": "工程变更（第9项）",
        "item_b": "新增工程（第10项）",
        "desc_a": "在原有项目基础上改——墙厚200mm改成250mm，综合单价调一下",
        "desc_b": "清单里完全没有的项目——原来没设计通风，后来加了一套",
        "responsibility_a": "调整综合单价即可",
        "responsibility_b": "需重新确定综合单价",
        "trap": "⚠ 区分不清，调价方式全错——变更调价≠新增定价",
        "check_method": "核对原清单是否存在该分项工程，是否属于合同范围内的数量/规格调整"
    },
    {
        "group": "第3组：物价变化 vs 法律法规变化",
        "item_a": "物价变化（第7项）",
        "item_b": "法律法规变化（第8项）",
        "desc_a": "市场行为——水泥涨了、钢筋跌了",
        "desc_b": "政策行为——增值税率从11%到9%、城建税调整",
        "responsibility_a": "按合同约定的调差公式走",
        "responsibility_b": "看政府文件（税率/费率调整文件）",
        "trap": "⚠ 性质不同，处理依据不同——物价看合同，法律看政府文件",
        "check_method": "区分价格波动原因：市场供需波动→第7项；政策文件发布→第8项"
    },
]

r = 3
for g in confusion_groups:
    # 组标题
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    set_cell(ws2, r, 1, g["group"], font=Font(name='微软雅黑', bold=True, color=WHITE, size=12), fill=PatternFill(start_color=TEAL, end_color=TEAL, fill_type='solid'), alignment=center_align)
    r += 1
    
    # 表头
    set_header_row(ws2, r, ["维度", g["item_a"], g["item_b"]])
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=1)
    r += 1
    
    rows_data = [
        ("含义/内容", g["desc_a"], g["desc_b"]),
        ("责任/处理方式", g["responsibility_a"], g["responsibility_b"]),
        ("常见陷阱", g["trap"], ""),
        ("审核方法", g["check_method"], ""),
    ]
    
    for label, val_a, val_b in rows_data:
        ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws2.merge_cells(start_row=r, start_column=5, end_row=r, end_column=8)
        is_trap = "陷阱" in label
        row_h = 40 if is_trap else 28
        
        set_cell(ws2, r, 1, label, font=bold_font, fill=warm_fill, alignment=center_align)
        set_cell(ws2, r, 2, val_a, font=alert_font if is_trap else normal_font, fill=white_fill, alignment=left_align)
        set_cell(ws2, r, 5, val_b, font=normal_font, fill=white_fill, alignment=left_align)
        ws2.row_dimensions[r].height = row_h
        r += 1
    
    r += 1  # 空行分隔

# 列宽
for col, w in {1:16, 2:32, 5:32}.items():
    ws2.column_dimensions[get_column_letter(col)].width = w

# ═══════════════════════════════════════════
# Sheet 4: 四个审减重点清单
# ═══════════════════════════════════════════
ws3 = wb.create_sheet("四个审减重点清单")
ws3.sheet_properties.tabColor = RED_ALERT

ws3.merge_cells('A1:I1')
set_cell(ws3, 1, 1, "四个审减重点检查清单", font=title_font, fill=None, alignment=center_align, border=None)

focus_items = [
    {
        "title": "重点一：材差不计管理费和利润",
        "ref": "7.2.2第2款",
        "content": "材料暂估价应按经招标确定或发承包双方共同确认的材料税前价格调整合同价款，价差调整变化部分不计管理费及利润。",
        "explain": "材料暂估价1000→1200元/吨，差价200元直接计入结算，但管理费和利润只按原来1000元计算。施工单位按调整后总价重算管理费利润的部分→审减净空间。",
        "checklist": [
            "材料暂估价部分是否有价差调整？",
            "价差是否单独列示（与原价分开）？",
            "管理费和利润是否只按原价计算？",
            "有无按调整后总价重算管理费利润的情况？",
        ]
    },
    {
        "title": "重点二：措施项目——能调的和不能调的",
        "ref": "7.2.3第1款",
        "content": '除合同另有约定，以及因工程变更、新增工程、工程索赔、暂列金额等引起的措施项目费用调整外，以\u201c项\u201d为计量单位的措施项目应依据合同单价计算，不做调整。',
        "explain": "按项计价的措施项目原则上不调。例外仅四种：①工程变更 ②新增工程 ③工程索赔 ④暂列金额。",
        "checklist": [
            "这笔措施费调整属于四类例外吗？（变更/新增/索赔/暂列金）",
            "有对应的工程变更单/签证单/索赔报告支持吗？",
            "调整金额的计算方法对吗？",
            "非四类例外的按项计价措施项目是否被调整了？→ 审减！",
        ]
    },
    {
        "title": "重点三：暂估价和增值税的计算基础",
        "ref": "5.2.7",
        "content": "增值税应以分部分项工程项目清单、措施项目清单、其他项目清单（专业工程暂估价除外）的合计金额作为计算基础。",
        "explain": "专业工程暂估价要从增值税计算基础里扣除。未扣除=增值税算多了。",
        "checklist": [
            "增值税计算基础中，是否扣除了专业工程暂估价？",
            "如有材料暂估价，增值税计算是否也做了对应调整？",
            "税率是否正确（按合同签订时的有效税率）？",
        ]
    },
    {
        "title": "重点四：结算申报时限——超了可能不认",
        "ref": "7.1.3",
        "content": "发承包双方在办理工程竣工结算过程中，应在合同约定的节点及相关规定时限内完成相关合同价款调整的申报及核对。",
        "explain": "结算不仅是算数题，还有程序问题。超时申报的，发包人可依法拒绝。",
        "checklist": [
            "合同约定的结算申报时限是多久？",
            "承包人是否在规定时限内完成了申报？",
            "如有超时申报，是否已取得发包人书面同意？",
            "无书面同意的超时申报→ 发包人可依法拒绝。",
        ]
    },
]

r = 3
for f in focus_items:
    # 标题行
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    set_cell(ws3, r, 1, f"{f['title']}  【依据：{f['ref']}】", 
             font=Font(name='微软雅黑', bold=True, color=WHITE, size=11), 
             fill=PatternFill(start_color=DEEP_BLUE, end_color=DEEP_BLUE, fill_type='solid'),
             alignment=left_align)
    r += 1
    
    # 规范原文
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    set_cell(ws3, r, 1, f"📜 规范原文：{f['content']}", font=small_font, fill=warm_fill, alignment=left_align)
    ws3.row_dimensions[r].height = 30
    r += 1
    
    # 解读
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    set_cell(ws3, r, 1, f"💡 实务解读：{f['explain']}", font=normal_font, fill=white_fill, alignment=left_align)
    ws3.row_dimensions[r].height = 35
    r += 1
    
    # 检查清单
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    set_cell(ws3, r, 1, "🔍 审核检查项：", font=bold_font, fill=PatternFill(start_color=COPPER_GOLD, end_color=COPPER_GOLD, fill_type='solid'), alignment=left_align)
    r += 1
    
    for j, item in enumerate(f['checklist']):
        is_last = (j == len(f['checklist']) - 1)
        check_label = "☐"  # 未选中的checkbox
        ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
        set_cell(ws3, r, 1, check_label, font=Font(name='微软雅黑', size=14, color=COPPER_GOLD), fill=white_fill, alignment=center_align)
        set_cell(ws3, r, 2, item, font=alert_font if "审减" in item else normal_font, fill=white_fill, alignment=left_align)
        ws3.row_dimensions[r].height = 22
        r += 1
    
    r += 1  # 空行

ws3.column_dimensions['A'].width = 6
ws3.column_dimensions['B'].width = 100

# ═══════════════════════════════════════════
# Sheet 5: 措施项目调整检查表
# ═══════════════════════════════════════════
ws4 = wb.create_sheet("措施项目调整检查")
ws4.sheet_properties.tabColor = "2980B9"

ws4.merge_cells('A1:H1')
set_cell(ws4, 1, 1, "措施项目费用调整检查表", font=title_font, fill=None, alignment=center_align, border=None)
ws4.merge_cells('A2:H2')
set_cell(ws4, 2, 1, "依据：7.2.3第1款 | 按项计价的措施项目原则上不调，仅四类例外", font=subtitle_font, fill=None, alignment=center_align, border=None)

r = 4
set_header_row(ws4, r, ["序号", "措施项目名称", "计量方式", "合同单价(元)", "申报调整金额(元)", "调整事由", "是否四类例外", "审核结论", "审核说明"])
ws4.row_dimensions[r].height = 28

# 示例数据行（空白模板）
exception_rules = [
    ("按项", "原则上不调（7.2.3第1款）"),
    ("按项", "原则上不调（7.2.3第1款）"),
    ("按项", "原则上不调（7.2.3第1款）"),
    ("按项", "原则上不调（7.2.3第1款）"),
    ("按项", "原则上不调（7.2.3第1款）"),
    ("按量", "按实计量，可调"),
    ("按量", "按实计量，可调"),
]

data_start2 = r + 1
for i in range(15):  # 预置15行
    row = data_start2 + i
    ws4.row_dimensions[row].height = 25
    row_fill = warm_fill if i % 2 == 0 else white_fill
    
    set_cell(ws4, row, 1, i+1, font=normal_font, fill=row_fill, alignment=center_align)
    
    if i < len(exception_rules):
        measure_type, rule = exception_rules[i]
        set_cell(ws4, row, 2, "", font=normal_font, fill=white_fill, alignment=left_align)
        set_cell(ws4, row, 3, measure_type, font=bold_font, fill=row_fill, alignment=center_align)
        set_cell(ws4, row, 4, "", font=normal_font, fill=white_fill, alignment=center_align, number_format='#,##0.00')
        set_cell(ws4, row, 5, "", font=normal_font, fill=white_fill, alignment=center_align, number_format='#,##0.00')
        set_cell(ws4, row, 6, "", font=normal_font, fill=white_fill, alignment=left_align)
        rule_cell = set_cell(ws4, row, 7, rule, 
                             font=alert_font if "不调" in rule else pass_font, 
                             fill=white_fill, alignment=center_align)
    else:
        for col in range(2, 8):
            set_cell(ws4, row, col, "", font=normal_font, fill=white_fill if col not in [3] else row_fill, alignment=center_align if col != 6 else left_align)
        set_cell(ws4, row, 3, "", font=normal_font, fill=row_fill, alignment=center_align)
    
    set_cell(ws4, row, 8, "", font=normal_font, fill=white_fill, alignment=center_align)
    set_cell(ws4, row, 9, "", font=normal_font, fill=white_fill, alignment=left_align)

# 四类例外提示
r_note = data_start2 + 16
ws4.merge_cells(start_row=r_note, start_column=1, end_row=r_note, end_column=9)
set_cell(ws4, r_note, 1, "📌 四类可调例外：①工程变更  ②新增工程  ③工程索赔  ④暂列金额  |  其他按项计价措施项目原则上不做调整", 
         font=Font(name='微软雅黑', bold=True, color=RED_ALERT, size=10), fill=warm_fill, alignment=left_align)

# 二审机制
r_note2 = r_note + 2
ws4.merge_cells(start_row=r_note2, start_column=1, end_row=r_note2, end_column=9)
set_cell(ws4, r_note2, 1, "🔍 三段审核法：①这笔调整属于四类例外吗？→ ②有变更单/签证单/索赔报告支撑吗？→ ③调整金额计算正确吗？三个问题只要一个答不上→审减", 
         font=normal_font, fill=white_fill, alignment=left_align)

# 列宽
col_widths2 = {1:6, 2:28, 3:10, 4:14, 5:16, 6:22, 7:22, 8:12, 9:24}
for col, w in col_widths2.items():
    ws4.column_dimensions[get_column_letter(col)].width = w

ws4.freeze_panes = f'A{data_start2}'

# ═══ 保存 ═══
output_path = r"C:\Users\scrccpa\.openclaw\workspace\output\结算审核底稿模板_v1.0.xlsx"
wb.save(output_path)
print(f"✅ 模板已保存：{output_path}")
print(f"工作表：{wb.sheetnames}")
