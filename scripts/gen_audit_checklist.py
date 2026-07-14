# -*- coding: utf-8 -*-
"""生成研究型审计检查清单 Excel 版"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

wb = openpyxl.Workbook()

# ====== 样式定义 ======
header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='0A1F3F', end_color='0A1F3F', fill_type='solid')
sub_header_fill = PatternFill(start_color='1A5C6E', end_color='1A5C6E', fill_type='solid')
gold_fill = PatternFill(start_color='C5955C', end_color='C5955C', fill_type='solid')
warm_fill = PatternFill(start_color='F5F2EC', end_color='F5F2EC', fill_type='solid')
white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
green_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
red_font = Font(name='微软雅黑', size=10, color='CC0000', bold=True)
normal_font = Font(name='微软雅黑', size=10)
bold_font = Font(name='微软雅黑', size=10, bold=True)
title_font = Font(name='微软雅黑', size=16, bold=True, color='0A1F3F')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_header_row(ws, row, cols, fill=header_fill):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = center_align
        cell.border = thin_border

def style_data_cell(ws, row, col, font=normal_font):
    cell = ws.cell(row=row, column=col)
    cell.font = font
    cell.alignment = wrap_align
    cell.border = thin_border
    return cell

def add_sheet_title(ws, title, subtitle=''):
    ws.merge_cells('A1:H1')
    ws['A1'] = title
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40
    if subtitle:
        ws.merge_cells('A2:H2')
        ws['A2'] = subtitle
        ws['A2'].font = Font(name='微软雅黑', size=10, color='666666')
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 25
        start_row = 3
    else:
        start_row = 2
    return start_row

# ==========================================
# Sheet 1: 立项阶段
# ==========================================
ws1 = wb.active
ws1.title = '1-立项阶段'
sr = add_sheet_title(ws1, '研究型审计检查清单 — 立项阶段', '常规审计是"进场找问题"，研究型审计是"研究定方向、方向找重点、重点挖问题"')

headers = ['编号', '检查项', '检查细项', '完成标准', '状态', '执行人', '完成日期', '备注/发现问题']
for c, h in enumerate(headers, 1):
    ws1.cell(row=sr, column=c, value=h)
style_header_row(ws1, sr, len(headers))
ws1.row_dimensions[sr].height = 30

items = [
    # 政策层面研究
    ('P1', '政策层面研究\n（定高度）', '', '', '', '', '', ''),
    ('P1.1', '', '梳理近三年国资委监管新规', '列出文件名称+文号+与本项目相关条款', '', '', '', ''),
    ('P1.2', '', '梳理近三年财政部门监管新规', '列出文件名称+文号+与本项目相关条款', '', '', '', ''),
    ('P1.3', '', '梳理行业主管部门最新政策', '列出文件名称+文号+与本项目相关条款', '', '', '', ''),
    ('P1.4', '', '梳理近期专项整治重点', '列出整治领域+时间+重点内容', '', '', '', ''),
    ('P1.5', '', '梳理巡视巡察高频问题', '列出近三年巡视报告中的共性/高频问题', '', '', '', ''),
    ('P1.6', '', '输出政策沿革脉络表', '明确本项目的审计标准来源和判断依据', '', '', '', ''),
    
    # 行业层面研究
    ('P2', '行业层面研究\n（定深度）', '', '', '', '', '', ''),
    ('P2.1', '', '识别行业固有风险', '该行业最容易出问题的3-5个领域', '', '', '', ''),
    ('P2.2', '', '了解行业通行计价规则', '定价/结算/定额/费率等行业惯例', '', '', '', ''),
    ('P2.3', '', '了解行业投资退出惯例', '如有投融资业务，了解退出机制和估值方法', '', '', '', ''),
    ('P2.4', '', '了解行业财税合规差异', '该行业的特殊税务处理/财政补贴政策', '', '', '', ''),
    ('P2.5', '', '了解行业项目盈亏周期规律', '同类项目的正常利润率/回款周期/亏损预警线', '', '', '', ''),
    ('P2.6', '', '输出行业对标分析框架', '用行业标准对照被审计单位，找出个性化短板', '', '', '', ''),
    
    # 企业层面研究
    ('P3', '企业层面研究\n（定精度）', '', '', '', '', '', ''),
    ('P3.1', '', '梳理历史审计问题清单', '近三年审计报告问题汇总+整改状态', '', '', '', ''),
    ('P3.2', '', '梳理整改短板', '哪些问题反复出现/屡审屡犯', '', '', '', ''),
    ('P3.3', '', '梳理业务薄弱点', '哪些业务部门/环节投诉多/差错多/效率低', '', '', '', ''),
    ('P3.4', '', '梳理重大投融资项目', '正在进行的重大投资/融资/建设项目清单', '', '', '', ''),
    ('P3.5', '', '梳理在建工程', '在建工程的规模/进度/资金使用情况', '', '', '', ''),
    ('P3.6', '', '输出风险全景画像', '一份A4纸以内的风险综合分析', '', '', '', ''),
    
    # 研究课题预埋
    ('P4', '研究课题预埋', '', '', '', '', '', ''),
    ('P4.1', '', '确定1-2个研究子课题', '课题命名格式:[领域]的[风险类型]研究', '', '', '', ''),
    ('P4.2', '', '明确研究课题的审计目标', '该课题要回答什么核心问题', '', '', '', ''),
    ('P4.3', '', '明确研究课题的数据需求', '哪些数据/资料是验证该课题必需的', '', '', '', ''),
]

for i, item in enumerate(items):
    row = sr + 1 + i
    for c, val in enumerate(item, 1):
        cell = style_data_cell(ws1, row, c)
        cell.value = val
    
    # 分类标题行
    if item[0] in ('P1', 'P2', 'P3', 'P4'):
        ws1.cell(row=row, column=1).font = bold_font
        ws1.cell(row=row, column=2).font = bold_font
        for c in range(1, 9):
            ws1.cell(row=row, column=c).fill = sub_header_fill
            ws1.cell(row=row, column=c).font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
        ws1.row_dimensions[row].height = 30
    else:
        ws1.row_dimensions[row].height = 45

# 状态列数据验证
dv1 = DataValidation(type='list', formula1='"未开始,进行中,已完成,不适用"', allow_blank=True)
dv1.error = '请选择有效状态'
ws1.add_data_validation(dv1)
for row in range(sr+1, sr+1+len(items)):
    dv1.add(ws1.cell(row=row, column=5))

# 列宽
col_widths_1 = [8, 20, 40, 40, 10, 12, 14, 30]
for c, w in enumerate(col_widths_1, 1):
    ws1.column_dimensions[get_column_letter(c)].width = w

# ==========================================
# Sheet 2: 方案编制
# ==========================================
ws2 = wb.create_sheet('2-方案编制')
sr = add_sheet_title(ws2, '研究型审计检查清单 — 方案编制阶段', '"一项目、一研究、一课题"——审计方案=研究课题实施计划书')

for c, h in enumerate(headers, 1):
    ws2.cell(row=sr, column=c, value=h)
style_header_row(ws2, sr, len(headers))
ws2.row_dimensions[sr].height = 30

items2 = [
    ('S1', '审计目标分层', '', '', '', '', '', ''),
    ('S1.1', '', '基础目标：核实账务合规', '明确哪些账务科目/业务流程需要合规核实', '', '', '', ''),
    ('S1.2', '', '基础目标：排查违规问题', '列出已知风险点和历史违规高发领域', '', '', '', ''),
    ('S1.3', '', '基础目标：确认整改落实', '对照上期审计整改清单逐项确认', '', '', '', ''),
    ('S1.4', '', '研究目标：梳理业务堵点', '识别业务流程中效率低下/卡顿的环节', '', '', '', ''),
    ('S1.5', '', '研究目标：剖析制度缺陷', '找出制度设计本身的问题（而非执行问题）', '', '', '', ''),
    ('S1.6', '', '研究目标：研判行业风险', '基于行业对标，预判未来发展中的风险', '', '', '', ''),
    ('S1.7', '', '研究目标：提出长效治理建议', '输出可落地的制度优化/流程再造方案', '', '', '', ''),
    
    ('S2', '审计重点前置', '', '', '', '', '', ''),
    ('S2.1', '', '工程计价与造价管控', '施工工艺/计价规则/定额套用/招标条款漏洞', '', '', '', ''),
    ('S2.2', '', '投融资与基金运作', '底层资产/退出机制/估值合理性/关联交易', '', '', '', ''),
    ('S2.3', '', '招投标合规管理', '围标串标信号/IP地址/保证金/报价规律', '', '', '', ''),
    ('S2.4', '', '资金预算与内控机制', '预算编制/执行/调整/决算全链条', '', '', '', ''),
    ('S2.5', '', '海外/跨境财税合规', '跨境财税差异/属地合规/境外采购风险', '', '', '', ''),
    ('S2.6', '', '根据本项目实际确定2-3个重点', '从上述领域中选取最匹配的', '', '', '', ''),
    
    ('S3', '审计程序定制化', '', '', '', '', '', ''),
    ('S3.1', '', '不使用通用模板程序', '每个程序的步骤/方法/标准都针对本项目', '', '', '', ''),
    ('S3.2', '', '每个重点领域设计专项检查程序', '不是泛泛的"抽查凭证"，是具体的分析模型', '', '', '', ''),
    ('S3.3', '', '明确数据需求和来源', '每个程序需要什么数据、从哪获取、什么格式', '', '', '', ''),
]

for i, item in enumerate(items2):
    row = sr + 1 + i
    for c, val in enumerate(item, 1):
        cell = style_data_cell(ws2, row, c)
        cell.value = val
    if item[0] in ('S1', 'S2', 'S3'):
        for c in range(1, 9):
            ws2.cell(row=row, column=c).fill = sub_header_fill
            ws2.cell(row=row, column=c).font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
        ws2.row_dimensions[row].height = 30
    else:
        ws2.row_dimensions[row].height = 45

dv2 = DataValidation(type='list', formula1='"未开始,进行中,已完成,不适用"', allow_blank=True)
ws2.add_data_validation(dv2)
for row in range(sr+1, sr+1+len(items2)):
    dv2.add(ws2.cell(row=row, column=5))

for c, w in enumerate(col_widths_1, 1):
    ws2.column_dimensions[get_column_letter(c)].width = w

# ==========================================
# Sheet 3: 现场实施
# ==========================================
ws3 = wb.create_sheet('3-现场实施')
sr = add_sheet_title(ws3, '研究型审计检查清单 — 现场实施阶段', '浅表问题看行为，深层问题看机制，顶级审计看体系')

for c, h in enumerate(headers, 1):
    ws3.cell(row=sr, column=c, value=h)
style_header_row(ws3, sr, len(headers))
ws3.row_dimensions[sr].height = 30

items3 = [
    ('F1', '数据全量比对研究', '', '', '', '', '', ''),
    ('F1.1', '', '获取全量数据（非抽样）', '确认数据覆盖全部时间范围和业务范围', '', '', '', ''),
    ('F1.2', '', '执行大数据比对', '同期对比/与上年对比/与预算对比/与行业均值对比', '', '', '', ''),
    ('F1.3', '', '执行周期比对', '月度/季度/年度波动分析，识别异常波动区间', '', '', '', ''),
    ('F1.4', '', '执行同业比对', '与同类单位/同规模单位的核心指标对比', '', '', '', ''),
    ('F1.5', '', '发现异常规律（非单笔违规）', '输出：一类违规/一种乱象/一个机制漏洞', '', '', '', ''),
    
    ('F2', '业务全链条溯源', '', '', '', '', '', ''),
    ('F2.1', '', '决策环节', '决策依据是否充分/程序是否合规/审批是否完整', '', '', '', ''),
    ('F2.2', '', '立项环节', '立项文件/可研报告/批复文件', '', '', '', ''),
    ('F2.3', '', '招投标环节', '招标方式/评标过程/中标结果/合同签订', '', '', '', ''),
    ('F2.4', '', '实施环节', '进度管理/质量管理/变更管理/资金拨付', '', '', '', ''),
    ('F2.5', '', '结算环节', '结算资料/工程量确认/价格审核/支付凭证', '', '', '', ''),
    ('F2.6', '', '归档环节', '档案完整性/合规性/可追溯性', '', '', '', ''),
    ('F2.7', '', '退出环节（如有）', '项目验收/资产移交/绩效评价/后评估', '', '', '', ''),
    ('F2.8', '', '定位风险源头节点', '找到问题第一次出现的环节，而非末端表象', '', '', '', ''),
    
    ('F3', '问题穿透归因\n（四层升级）', '', '', '', '', '', ''),
    ('F3.1', '', '第一层：行为层面', '谁、在什么时间、做了什么、产生了什么结果', '', '', '', ''),
    ('F3.2', '', '第二层：制度层面', '为什么制度允许这样做？制度缺失还是制度执行不力？', '', '', '', ''),
    ('F3.3', '', '第三层：机制层面', '为什么制度会缺失/失效？管控体系/激励机制的深层缺陷', '', '', '', ''),
    ('F3.4', '', '第四层：体系层面', '行业监管盲区/治理结构的系统性缺陷/体制深层症结', '', '', '', ''),
    ('F3.5', '', '⚠️ 至少穿透到第二层', '每发现一个问题，必须追问至少一次"为什么"', '', '', '', ''),
    
    ('F4', '按业务类型的专项\n穿透（如适用）', '', '', '', '', '', ''),
    ('F4.1', '', '工程审计专项', '施工工艺→计价规则→定额套用逻辑→招投标条款漏洞', '', '', '', ''),
    ('F4.2', '', '投融资审计专项', '底层资产→退出机制→估值合理性→关联交易隐匿风险', '', '', '', ''),
    ('F4.3', '', '海外项目审计专项', '跨境财税差异→属地合规漏洞→境外采购隐性风险', '', '', '', ''),
    ('F4.4', '', '预算执行审计专项', '预算编制→指标下达→执行进度→调整审批→决算比对', '', '', '', ''),
]

for i, item in enumerate(items3):
    row = sr + 1 + i
    for c, val in enumerate(item, 1):
        cell = style_data_cell(ws3, row, c)
        cell.value = val
    if item[0] in ('F1', 'F2', 'F3', 'F4'):
        for c in range(1, 9):
            ws3.cell(row=row, column=c).fill = sub_header_fill
            ws3.cell(row=row, column=c).font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
        ws3.row_dimensions[row].height = 30
    elif item[0] == 'F3.5':
        for c in range(1, 9):
            ws3.cell(row=row, column=c).fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
        ws3.cell(row=row, column=3).font = red_font
        ws3.row_dimensions[row].height = 45
    else:
        ws3.row_dimensions[row].height = 45

dv3 = DataValidation(type='list', formula1='"未开始,进行中,已完成,不适用"', allow_blank=True)
ws3.add_data_validation(dv3)
for row in range(sr+1, sr+1+len(items3)):
    dv3.add(ws3.cell(row=row, column=5))

for c, w in enumerate(col_widths_1, 1):
    ws3.column_dimensions[get_column_letter(c)].width = w

# ==========================================
# Sheet 4: 中期升华+报告撰写
# ==========================================
ws4 = wb.create_sheet('4-中期升华与报告')
sr = add_sheet_title(ws4, '研究型审计检查清单 — 中期升华与报告撰写', '把碎片化问题上升为体系性课题；把审计发现转化为治理建议')

for c, h in enumerate(headers, 1):
    ws4.cell(row=sr, column=c, value=h)
style_header_row(ws4, sr, len(headers))
ws4.row_dimensions[sr].height = 30

items4 = [
    ('M1', '中期升华：碎片→课题', '', '', '', '', '', ''),
    ('M1.1', '', '将零散问题按类别汇总', '按问题性质/领域/频次分组归类', '', '', '', ''),
    ('M1.2', '', '每类问题提炼一个研究课题', '课题命名：[领域]的[风险类型]研究', '', '', '', ''),
    ('M1.3', '', '课题转化示例参考', '零星报销不规范→"费用管控精细化体系漏洞研究"', '', '', '', ''),
    ('M1.4', '', '', '零散结算偏差→"造价管控长效机制缺陷研究"', '', '', '', ''),
    ('M1.5', '', '', '个别基金运作不规范→"基金募投管退风控体系研究"', '', '', '', ''),
    ('M1.6', '', '形成课题组会讨论材料', '至少形成1页A4纸的课题组会汇报提纲', '', '', '', ''),
    
    ('M2', '报告撰写：六段式结构', '', '', '', '', '', ''),
    ('M2.1', '', '① 现状研判', '不罗列数字，揭示趋势和结构特征', '', '', '', ''),
    ('M2.2', '', '② 问题归集', '不堆砌清单，按机制分类呈现', '', '', '', ''),
    ('M2.3', '', '③ 根源剖析', '重点回答"为什么反复出现/屡审屡犯/机制哪里缺失"', '', '', '', ''),
    ('M2.4', '', '④ 风险预判', '不整改将产生什么后果？影响范围和严重程度', '', '', '', ''),
    ('M2.5', '', '⑤ 对策体系', '可落地/可执行/可考核的制度优化方案', '', '', '', ''),
    ('M2.6', '', '⑥ 机制建设', '流程整改路径/风控完善体系/长效治理框架', '', '', '', ''),
    
    ('M3', '报告质量自检', '', '', '', '', '', ''),
    ('M3.1', '', '站位是否更高？', '立足企业治理/国资监管/高质量发展，非单纯财务合规', '', '', '', ''),
    ('M3.2', '', '剖析是否更深？', '机制层面而非行为层面', '', '', '', ''),
    ('M3.3', '', '对策是否更实？', '摒弃"加强管理、提高意识"等空话，输出可考核的措施', '', '', '', ''),
    ('M3.4', '', '是否有"问责价值"又有"治理价值"？', '不仅说谁做错了，更说系统怎么改进', '', '', '', ''),
    ('M3.5', '', '计算依据是否可追溯？', '每条结论附：数据来源+计算方法+交叉验证结果', '', '', '', ''),
]

for i, item in enumerate(items4):
    row = sr + 1 + i
    for c, val in enumerate(item, 1):
        cell = style_data_cell(ws4, row, c)
        cell.value = val
    if item[0] in ('M1', 'M2', 'M3'):
        for c in range(1, 9):
            ws4.cell(row=row, column=c).fill = sub_header_fill
            ws4.cell(row=row, column=c).font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
        ws4.row_dimensions[row].height = 30
    elif item[0] in ('M1.3', 'M1.4', 'M1.5'):
        ws4.row_dimensions[row].height = 35
    else:
        ws4.row_dimensions[row].height = 45

dv4 = DataValidation(type='list', formula1='"未开始,进行中,已完成,不适用"', allow_blank=True)
ws4.add_data_validation(dv4)
for row in range(sr+1, sr+1+len(items4)):
    dv4.add(ws4.cell(row=row, column=5))

for c, w in enumerate(col_widths_1, 1):
    ws4.column_dimensions[get_column_letter(c)].width = w

# ==========================================
# Sheet 5: 收尾归档
# ==========================================
ws5 = wb.create_sheet('5-收尾归档')
sr = add_sheet_title(ws5, '研究型审计检查清单 — 收尾归档阶段', '普通审计做完即止；研究型审计：一次审计、一次升级、一次赋能、一次沉淀')

for c, h in enumerate(headers, 1):
    ws5.cell(row=sr, column=c, value=h)
style_header_row(ws5, sr, len(headers))
ws5.row_dimensions[sr].height = 30

items5 = [
    ('C1', '问题底稿标准化', '', '', '', '', '', ''),
    ('C1.1', '', '取证完整', '每项问题有对应的原始证据/凭证/合同/文件', '', '', '', ''),
    ('C1.2', '', '定性准确', '问题定性引用的法规条款准确、版本有效', '', '', '', ''),
    ('C1.3', '', '依据充分', '定性依据+判断逻辑+数据支撑完整可追溯', '', '', '', ''),
    ('C1.4', '', '溯源链条完整', '从问题现象→直接原因→制度原因→机制原因', '', '', '', ''),
    ('C1.5', '', '逻辑闭环', '问题→证据→判断→结论→建议，链条无断点', '', '', '', ''),
    ('C1.6', '', '经得起巡视/复查/抽查', '任何第三方能仅凭底稿理解问题全貌', '', '', '', ''),
    
    ('C2', '研究成果固化', '', '', '', '', '', ''),
    ('C2.1', '', '整理研究思路笔记', '本项目的研究方法/分析逻辑/关键发现', '', '', '', ''),
    ('C2.2', '', '整理风险模型', '本项目使用的风险识别/评估模型', '', '', '', ''),
    ('C2.3', '', '整理核查方法', '本项目最有效的核查手段/技巧', '', '', '', ''),
    ('C2.4', '', '整理对标结论', '行业对标分析方法+核心发现', '', '', '', ''),
    ('C2.5', '', '输出工作指引', '可供后续同类项目直接使用的方法指引', '', '', '', ''),
    ('C2.6', '', '输出课题材料', '将研究子课题形成正式的课题报告/论文/案例', '', '', '', ''),
    
    ('C3', '案例库沉淀', '', '', '', '', '', ''),
    ('C3.1', '', '典型问题入库', '标注：问题类型+行业+严重程度+处理方式', '', '', '', ''),
    ('C3.2', '', '新型风险入库', '标注：首次发现的风险类型+信号特征+防范建议', '', '', '', ''),
    ('C3.3', '', '行业乱象入库', '标注：行业+乱象描述+根因+监管动态', '', '', '', ''),
    ('C3.4', '', '案例格式标准化', '标题/背景/问题/方法/发现/建议/启示 七要素', '', '', '', ''),
    
    ('C4', '机制优化落地', '', '', '', '', '', ''),
    ('C4.1', '', '推动制度修订', '针对发现的制度缺陷，提出具体修订条款', '', '', '', ''),
    ('C4.2', '', '推动流程再造', '针对流程堵点，提出优化后的流程图', '', '', '', ''),
    ('C4.3', '', '推动风控补位', '针对控制缺失，提出新增控制节点和措施', '', '', '', ''),
    ('C4.4', '', '跟踪整改闭环', '确认整改措施落地→验证整改效果→闭环', '', '', '', ''),
]

for i, item in enumerate(items5):
    row = sr + 1 + i
    for c, val in enumerate(item, 1):
        cell = style_data_cell(ws5, row, c)
        cell.value = val
    if item[0] in ('C1', 'C2', 'C3', 'C4'):
        for c in range(1, 9):
            ws5.cell(row=row, column=c).fill = sub_header_fill
            ws5.cell(row=row, column=c).font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
        ws5.row_dimensions[row].height = 30
    else:
        ws5.row_dimensions[row].height = 45

dv5 = DataValidation(type='list', formula1='"未开始,进行中,已完成,不适用"', allow_blank=True)
ws5.add_data_validation(dv5)
for row in range(sr+1, sr+1+len(items5)):
    dv5.add(ws5.cell(row=row, column=5))

for c, w in enumerate(col_widths_1, 1):
    ws5.column_dimensions[get_column_letter(c)].width = w

# ==========================================
# Sheet 6: 数据资产化三闭环
# ==========================================
ws6 = wb.create_sheet('6-数据资产化闭环')
sr = add_sheet_title(ws6, '审计数据资产化 — 三个最小闭环', '归档是把项目结束，资产化是让项目继续产生价值')

headers6 = ['闭环', '步骤', '具体操作', '输入', '输出', '负责人', '频次', '当前状态']
for c, h in enumerate(headers6, 1):
    ws6.cell(row=sr, column=c, value=h)
style_header_row(ws6, sr, len(headers6))
ws6.row_dimensions[sr].height = 30

items6 = [
    ('闭环一\n复核意见→模板优化', 'Step 1', '结构化复核意见：每次审底稿打标签', '复核意见文本', '标签化复核记录\n（金额不一致/依据不充分/说明不清楚/勾稽未检查/结论太虚/格式问题）', '', '每次复核时', ''),
    ('', 'Step 2', '季度统计：哪些底稿最常被打回、哪类问题最多', '标签化复核记录', '统计报表\n（底稿类型×问题类型×频次）', '', '每季度', ''),
    ('', 'Step 3', '反向优化模板：增加校验规则/填写提示/红色标注', '统计报表', '优化后的底稿模板', '', '每季度', ''),
    ('', 'Step 4', '效果验证：对比优化前后的打回率', '优化前后的复核记录', '效果评估报告', '', '半年', ''),
    
    ('闭环二\n调整分录→风险提示', 'Step 1', '每项目标注调整分录：类型+科目+行业+金额', '调整分录', '结构化调整记录表', '', '每项目结束时', ''),
    ('', 'Step 2', '跨项目规律挖掘：某行业/某客户的典型调整类型', '累计20+项目调整记录', '行业-调整类型关联图谱', '', '每半年', ''),
    ('', 'Step 3', '新项目启动时自动提醒：同类项目历史高频调整', '行业-调整类型关联图谱', '项目启动风险提示卡', '', '新项目启动时', ''),
    
    ('闭环三\n资料清单→缺失提醒', 'Step 1', '记录每次资料缺失：客户+资料名+缺失原因+延迟天数', '项目资料清单+实际获取记录', '资料缺失历史记录表', '', '每项目结束时', ''),
    ('', 'Step 2', '生成客户专属动态资料清单', '资料缺失历史记录表', '客户资料清单（含历史缺失标注）', '', '每年更新', ''),
    ('', 'Step 3', '资料格式问题标注：经常出错的资料附模板要求', '格式问题记录', '客户资料清单（含格式说明+模板）', '', '每年更新', ''),
]

for i, item in enumerate(items6):
    row = sr + 1 + i
    for c, val in enumerate(item, 1):
        cell = style_data_cell(ws6, row, c)
        cell.value = val
    if item[0] and '闭环' in str(item[0]):
        for c in range(1, 9):
            ws6.cell(row=row, column=c).fill = sub_header_fill
            ws6.cell(row=row, column=c).font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
        ws6.row_dimensions[row].height = 35
    else:
        ws6.row_dimensions[row].height = 55

dv6 = DataValidation(type='list', formula1='"未启动,进行中,已建立,待优化"', allow_blank=True)
ws6.add_data_validation(dv6)
for row in range(sr+1, sr+1+len(items6)):
    dv6.add(ws6.cell(row=row, column=8))

col_w6 = [22, 10, 38, 28, 28, 12, 14, 12]
for c, w in enumerate(col_w6, 1):
    ws6.column_dimensions[get_column_letter(c)].width = w

# ==========================================
# Sheet 7: 使用说明
# ==========================================
ws7 = wb.create_sheet('使用说明')
sr = add_sheet_title(ws7, '研究型审计检查清单 — 使用说明')

instructions = [
    ('📋 总体说明', ''),
    ('', '本检查清单基于微信公众号「蛎翁」《研究型审计完整工作手册》提炼，适配融策会计师事务所政府审计业务场景。'),
    ('', ''),
    ('🎯 使用方式', ''),
    ('', '1. 项目启动时，打印Sheet 1-2（立项+方案），在项目启动会上对照填写'),
    ('', '2. 现场实施中，逐日/逐周对照Sheet 3（现场实施），勾选完成状态'),
    ('', '3. 报告撰写前，对照Sheet 4（中期升华+报告），确保每个报告质量自检项达标'),
    ('', '4. 项目结束时，对照Sheet 5（收尾归档），逐项闭环'),
    ('', '5. Sheet 6（数据资产化闭环）作为事务所级别的持续性工作，由质控负责人跟踪'),
    ('', ''),
    ('🔄 版本迭代', ''),
    ('', '首次使用后，记录以下信息：'),
    ('', '• 哪些检查项确实提升了报告质量？→ 保留并强化'),
    ('', '• 哪些检查项在当前项目类型中不适用？→ 标注"不适用"并记录原因'),
    ('', '• 哪些检查项需要调整标准？→ 在备注栏写修改建议'),
    ('', '做完3个不同类型的项目后，汇总反馈 → 发布融策定制版v2.0'),
    ('', ''),
    ('📊 状态定义', ''),
    ('', '未开始 = 尚未执行   |   进行中 = 正在执行   |   已完成 = 已完成并确认   |   不适用 = 本项目不需要'),
    ('', ''),
    ('⚠️ 关键提醒', ''),
    ('', '• F3.5（归因至少到第二层）是研究型审计与常规审计的核心分水岭'),
    ('', '• M3.5（计算依据可追溯）是平头哥要求的铁律——每条结论必须附计算方法'),
    ('', '• Sheet 6 闭环一（复核意见→模板优化）是三个闭环中最容易启动的，建议优先启动'),
]

for i, (col1, col2) in enumerate(instructions):
    row = sr + 1 + i
    cell1 = ws7.cell(row=row, column=1, value=col1)
    cell2 = ws7.cell(row=row, column=2, value=col2)
    if col1 and not col1.startswith(' '):
        cell1.font = Font(name='微软雅黑', size=11, bold=True, color='0A1F3F')
    else:
        cell1.font = normal_font
    cell2.font = normal_font
    cell1.alignment = Alignment(vertical='top')
    cell2.alignment = Alignment(wrap_text=True, vertical='top')
    ws7.row_dimensions[row].height = 22 if col1 else 18

ws7.column_dimensions['A'].width = 30
ws7.column_dimensions['B'].width = 80

# ====== 保存 ======
outdir = os.path.expanduser(r'~\.openclaw\workspace\output')
os.makedirs(outdir, exist_ok=True)
outpath = os.path.join(outdir, '研究型审计检查清单_融策版.xlsx')
wb.save(outpath)
print(f'Saved: {outpath}')
print(f'Sheets: {wb.sheetnames}')
