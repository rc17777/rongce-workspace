"""
融策企业导出样式标准 (Rongce Export Style Standard)
=================================================
适用: Excel / Word / PPT 所有对外及内部分析报告
版本: v1.0
日期: 2026-05-30

使用方式:
    from rongce_style import RC, apply_excel_style
    # 然后在脚本中引用 RC.PRIMARY, FONT_TITLE 等常量
"""

# ===========================
# 一、企业配色体系
# ===========================
class RC:
    """融策企业色板 —— 所有导出文档统一使用"""
    # 主色系
    PRIMARY    = "1F4E79"  # 深蓝（标题、表头、强调色）
    SECONDARY  = "2E75B6"  # 中蓝（副标题、超链接）
    ACCENT     = "D4A843"  # 金色（点缀、装饰线、高亮）
    
    # 背景色
    BG_HEADER  = "1F4E79"  # 表头背景（白字）
    BG_LIGHT   = "F2F7FB"  # 浅蓝（交替行底色）
    BG_WHITE   = "FFFFFF"  # 白色底
    BG_RED     = "FFF0F0"  # 红色警示底
    BG_YELLOW  = "FFF8E1"  # 黄色关注底
    BG_GREEN   = "E8F5E9"  # 绿色正常底
    
    # 文字色
    TEXT_DARK  = "1A1A1A"  # 正文
    TEXT_MED   = "555555"  # 次级文字
    TEXT_LIGHT = "999999"  # 弱化文字
    
    # 功能色
    RED        = "C0392B"  # 🔴 严重/铁证
    ORANGE     = "E67E22"  # 🟡 警告/强信号
    GREEN      = "27AE60"  # 🟢 正常/通过
    BLUE_INFO  = "2980B9"  # ℹ️ 信息
    
    # 边框
    LINE       = "D0D0D0"  # 表格边框
    LINE_DARK  = "999999"  # 强调边框
    
    # ====== 语义配色（审计专用） ======
    FUND_GREEN    = "27AE60"  # 资金绿
    RISK_RED      = "E74C3C"  # 风险红
    REMEDIATION   = "E67E22"  # 整改橙
    REGULATION    = "2980B9"  # 制度蓝
    PROCESS       = "8E44AD"  # 流程紫
    ORG_GRAY      = "7F8C8D"  # 组织灰
    DECISION_GOLD = "F39C12"  # 决策金
    EVIDENCE_CYAN = "1ABC9C"  # 证据青

# ===========================
# 二、Logo信息
# ===========================
LOGO_PATH = r"D:\openclaw-workspace\projects\data-analysis-agent\static\Images\rongce-logo.png"
LOGO_SIDEBAR_PATH = r"D:\openclaw-workspace\projects\data-analysis-agent\static\Images\rongce-logo-sidebar.png"

# 页脚信息
FOOTER_TEXT = "四川融策会计师事务所 / 四川融策工程咨询有限公司  |  保密分析材料  |  仅供内部使用"
COMPANY_SHORT = "四川融策会计师事务所"

# ===========================
# 三、Excel样式规范
# ===========================
# 字体
EXCEL = {
    "font_title":    {"name": "微软雅黑", "size": 16, "bold": True,  "color": RC.PRIMARY},
    "font_subtitle": {"name": "微软雅黑", "size": 11, "bold": False, "color": RC.TEXT_MED},
    "font_section":  {"name": "微软雅黑", "size": 12, "bold": True,  "color": RC.PRIMARY},
    "font_header":   {"name": "微软雅黑", "size": 10, "bold": True,  "color": "FFFFFF"},
    "font_body":     {"name": "微软雅黑", "size": 10, "bold": False, "color": RC.TEXT_DARK},
    "font_bold":     {"name": "微软雅黑", "size": 10, "bold": True,  "color": RC.TEXT_DARK},
    "font_small":    {"name": "微软雅黑", "size": 9,  "bold": False, "color": RC.TEXT_MED},
}

# ===========================
# 四、Word样式规范（python-docx）
# ===========================
WORD = {
    "heading1": {"font_name": "微软雅黑", "font_size": 18, "bold": True,  "color": RC.PRIMARY},
    "heading2": {"font_name": "微软雅黑", "font_size": 14, "bold": True,  "color": RC.PRIMARY},
    "heading3": {"font_name": "微软雅黑", "font_size": 12, "bold": True,  "color": RC.SECONDARY},
    "body":     {"font_name": "微软雅黑", "font_size": 11, "bold": False, "color": RC.TEXT_DARK},
    "table_header": {"font_name": "微软雅黑", "font_size": 10, "bold": True, "color": "FFFFFF",
                     "bg": RC.BG_HEADER},
    "table_body":   {"font_name": "微软雅黑", "font_size": 10, "bold": False, "color": RC.TEXT_DARK},
    "footer":  {"font_name": "微软雅黑", "font_size": 9,  "bold": False, "color": RC.TEXT_LIGHT},
    "page_margin": {"top": 2.5, "bottom": 2.5, "left": 3.0, "right": 2.5},  # cm
}

# ===========================
# 五、PPT样式规范（python-pptx）
# ===========================
PPT = {
    "title_color":       RC.PRIMARY,
    "subtitle_color":    RC.SECONDARY,
    "body_color":        RC.TEXT_DARK,
    "accent_color":      RC.ACCENT,
    "bg_color":          "FFFFFF",
    "slide_width_cm":    33.867,  # 16:9
    "slide_height_cm":   19.05,
    "font_title":        {"name": "微软雅黑", "size": 32, "bold": True},
    "font_subtitle":     {"name": "微软雅黑", "size": 18, "bold": False},
    "font_section":      {"name": "微软雅黑", "size": 24, "bold": True},
    "font_body":         {"name": "微软雅黑", "size": 16, "bold": False},
    "font_small":        {"name": "微软雅黑", "size": 12, "bold": False},
}

# ===========================
# 六、通用规范
# ===========================
# 网格线：专业报告一律隐藏Excel网格线（sheet_view.showGridLines = False）
# 交替行色：数据表使用交替行底色提高可读性
# 列宽：中文按双字节估算（ord>127的字符计2宽度）
# 语言：中文报告用「微软雅黑」，数字/代码用 Consolas
# Logo：封面和每页左上角放置公司Logo
# 页脚：每页底部署名「四川融策会计师事务所 · 四川融策工程咨询有限公司」


# ===========================
# 七、便捷函数（Excel）
# ===========================
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def rc_font(name="微软雅黑", size=10, bold=False, color=RC.TEXT_DARK):
    return Font(name=name, size=size, bold=bold, color=color)

def rc_fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def rc_border(color=RC.LINE):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def rc_alignment(horizontal="left", vertical="center", wrap=True):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)

# 快捷样式
FONT_TITLE   = rc_font(size=16, bold=True, color=RC.PRIMARY)
FONT_SECTION = rc_font(size=12, bold=True, color=RC.PRIMARY)
FONT_HEADER  = rc_font(size=10, bold=True, color="FFFFFF")
FONT_BODY    = rc_font(size=10)
FONT_BOLD    = rc_font(size=10, bold=True)
FONT_SMALL   = rc_font(size=9, color=RC.TEXT_MED)

FILL_HEADER  = rc_fill(RC.BG_HEADER)
FILL_ODD     = rc_fill(RC.BG_LIGHT)
FILL_EVEN    = rc_fill(RC.BG_WHITE)
FILL_RED     = rc_fill(RC.BG_RED)
FILL_YELLOW  = rc_fill(RC.BG_YELLOW)
FILL_GREEN   = rc_fill(RC.BG_GREEN)

BORDER_THIN  = rc_border(RC.LINE)
ALIGN_LEFT   = rc_alignment("left")
ALIGN_CENTER = rc_alignment("center")


def apply_excel_sheet_style(ws, logo=True, hide_gridlines=True):
    """一键应用融策Excel样式"""
    from openpyxl.drawing.image import Image as XLImage
    import os
    
    if hide_gridlines:
        ws.sheet_view.showGridLines = False
    
    # Logo v2: 仅封面放置，80-100px宽，居中或右上角，避免突兀
    if logo and os.path.exists(LOGO_PATH):
        img = XLImage(LOGO_PATH)
        img.width = 100
        ws.add_image(img, "A1")  # 子类可覆盖位置
    
    # 所有已用单元格加默认字体
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if not cell.font or cell.font.name == 'Calibri':
                cell.font = Font(name="微软雅黑", size=10, color=RC.TEXT_DARK)


def write_excel_header(ws, row, headers, start_col=1):
    """写标准表头行（深蓝底白字）"""
    for i, h in enumerate(headers, start_col):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN


def write_excel_row(ws, row, values, start_col=1, row_type="normal"):
    """写标准数据行（交替色）"""
    fill_map = {
        "normal": FILL_ODD if row % 2 == 0 else FILL_EVEN,
        "red":   FILL_RED,
        "yellow": FILL_YELLOW,
        "green": FILL_GREEN,
    }
    fill = fill_map.get(row_type, fill_map["normal"])
    
    for i, val in enumerate(values, start_col):
        cell = ws.cell(row=row, column=i, value=val)
        cell.font = FONT_BODY
        cell.fill = fill
        cell.alignment = ALIGN_LEFT
        cell.border = BORDER_THIN


def auto_width(ws, min_w=8, max_w=55):
    """智能列宽：中文字符计2宽度"""
    from openpyxl.utils import get_column_letter
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value:
                for line in str(cell.value).split('\n'):
                    w = sum(2 if ord(c) > 127 else 1 for c in line)
                    max_len = max(max_len, w)
        ws.column_dimensions[col_letter].width = max(min_w, min(max_len + 4, max_w))


def add_footer(ws, row, max_col, text=None):
    """添加页脚"""
    if text is None:
        text = FOOTER_TEXT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = FONT_SMALL
    cell.alignment = Alignment(horizontal="center", vertical="center")


# ===========================
# 八、通俗解读框（写给非技术读者的解释）
# ===========================
def write_explain_box(ws, row, col, end_col, title, lines,
                      title_font=None, body_font=None,
                      title_fill=None, body_fill=None):
    """
    在表格下方插入「通俗解读」框
    - title: 解释框标题（如 "📖 通俗解读：什么是文件来源追溯？"）
    - lines: 解释文本行列表
    - 返回下一个可用行号
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    tf = title_font or Font(name="微软雅黑", size=10.5, bold=True, color="2E75B6")
    bf = body_font or Font(name="微软雅黑", size=10, color="777777")
    tfill = title_fill or PatternFill(start_color="E9F0F7", end_color="E9F0F7", fill_type="solid")
    bfill = body_fill or PatternFill(start_color="F8F9FB", end_color="F8F9FB", fill_type="solid")

    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=col, value=f"📖 {title}")
    c.font = tf; c.fill = tfill
    c.alignment = Alignment(horizontal="left", vertical="center")
    row += 1
    for line in lines:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
        c = ws.cell(row=row, column=col, value=line)
        c.font = bf; c.fill = bfill
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        row += 1
    return row + 1


print("✅ 融策企业导出样式标准 v1.1 已加载")
print(f"   配色方案: {len([v for v in dir(RC) if not v.startswith('_')])} 个色值")
print(f"   Logo: {LOGO_PATH}")
print(f"   新增: write_explain_box() 通俗解读框")
