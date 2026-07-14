#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成审计报告复核结果Excel"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ===== 配色 =====
DARK_BLUE = "000A1F3F"
TEAL = "001A5C6E"
GOLD = "00C5955C"
GOLD_LIGHT = "00E8D5B5"
WARM_WHITE = "00FAFAF8"
ALT_ROW = "00F5F2EC"
RED_BG = "00FCE4E4"
YELLOW_BG = "00FFF8E1"

hdr_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
teal_fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
gold_fill = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
alt_fill = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid")
red_fill = PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid")
yellow_fill = PatternFill(start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type="solid")
pass_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

hdr_font = Font(name='微软雅黑', bold=True, color="FFFFFF", size=11)
hdr_font_gold = Font(name='微软雅黑', bold=True, color=GOLD, size=11)
body_font = Font(name='微软雅黑', size=10, color="2D2D2D")
body_font_bold = Font(name='微软雅黑', size=10, color="2D2D2D", bold=True)
red_font = Font(name='微软雅黑', size=10, color="C62828", bold=True)
gold_font = Font(name='微软雅黑', size=10, color="A67C00", bold=True)
pass_font = Font(name='微软雅黑', size=10, color="2E7D32")
title_font = Font(name='微软雅黑', size=14, bold=True, color=DARK_BLUE)
section_font = Font(name='微软雅黑', size=12, bold=True, color=TEAL)

thin_border = Border(
    left=Side(style='thin', color='D0C8B8'),
    right=Side(style='thin', color='D0C8B8'),
    top=Side(style='thin', color='D0C8B8'),
    bottom=Side(style='thin', color='D0C8B8')
)

def set_cell(ws, row, col, value, font=body_font, fill=None, align=None, border=thin_border):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = align or Alignment(vertical='center', wrap_text=True)
    cell.border = border
    return cell

def set_title_row(ws, row, values, fill=hdr_fill, font=hdr_font):
    for i, v in enumerate(values, 1):
        set_cell(ws, row, i, v, font=font, fill=fill, 
                 align=Alignment(horizontal='center', vertical='center', wrap_text=True))

# ================================================================
# Sheet 1: 复核总览
# ================================================================
ws1 = wb.active
ws1.title = '复核总览'
ws1.sheet_properties.tabColor = DARK_BLUE

# 标题
ws1.merge_cells('A1:F1')
set_cell(ws1, 1, 1, '审计报告AI复核结果', font=title_font, border=Border())
ws1.row_dimensions[1].height = 35

ws1.merge_cells('A2:F2')
set_cell(ws1, 2, 1, '报告名称：四川食在攒劲餐饮服务有限公司 2024年7月—2025年12月 财务收支专项审计报告', 
          font=Font(name='微软雅黑', size=10, color='8B9DAF'), border=Border())
ws1.row_dimensions[2].height = 22

# 统计卡片
stats = [
    ('P0 实质性矛盾', 0, '无', '00E8F5E9'),
    ('P1 重大遗漏/错误', 8, '8处问题需修正', '00FCE4E4'),
    ('P2 口径/表达问题', 4, '4处建议优化', '00FFF8E1'),
    ('OK 数字验证通过', '14/14', '全部一致', '00E8F5E9'),
]

set_cell(ws1, 4, 1, '统计维度', font=hdr_font, fill=hdr_fill, align=Alignment(horizontal='center', vertical='center'))
ws1.merge_cells('A4:B4')
set_cell(ws1, 4, 3, '数值', font=hdr_font, fill=hdr_fill, align=Alignment(horizontal='center', vertical='center'))
set_cell(ws1, 4, 4, '说明', font=hdr_font, fill=hdr_fill, align=Alignment(horizontal='center', vertical='center'))
ws1.merge_cells('C4:D4')

for i, (name, val, desc, color) in enumerate(stats, 5):
    bg = PatternFill(start_color=color, end_color=color, fill_type='solid')
    set_cell(ws1, i, 1, name, font=body_font_bold, fill=bg, align=Alignment(vertical='center'))
    ws1.merge_cells(f'A{i}:B{i}')
    set_cell(ws1, i, 3, str(val), font=Font(name='微软雅黑', size=12, bold=True, color=DARK_BLUE), fill=bg, 
             align=Alignment(horizontal='center', vertical='center'))
    set_cell(ws1, i, 4, desc, font=body_font, fill=bg, align=Alignment(vertical='center'))
    ws1.merge_cells(f'C{i}:D{i}')
    ws1.row_dimensions[i].height = 28

# 各维度检查结果摘要
set_cell(ws1, 10, 1, '维度检查摘要', font=section_font, border=Border())
ws1.merge_cells('A10:D10')
ws1.row_dimensions[10].height = 28

set_title_row(ws1, 11, ['检查维度', '状态', '说明'])

dimensions = [
    ('① 逻辑一致性', '⚠️', 'P145/P66/P87三处笔误影响逻辑连贯性'),
    ('② 问题定性精确度', '✅', '问题定性有制度依据，定性准确'),
    ('③ 整改建议靶向性', '✅', '每条建议都有责任主体和方向'),
    ('④ 证据链完整性', '⚠️', '收入无附件的问题已指出，建议补充原始凭证'),
    ('⑤ 风险后果推演', '✅', '负债结构风险/毛利率分析充分'),
    ('⑥ 审计目标覆盖度', '⚠️', '纳税审查只到2025年1月，未覆盖完整期间'),
    ('⑦ 受众适配度', '✅', '表述专业，适合管理层阅读'),
    ('⑧ 摘要可读性', 'N/A', '报告无独立摘要段落'),
    ('⑨ 跨项目一致性', 'N/A', '仅一份报告，无法比对'),
    ('⑩ 措辞情绪化', '✅', '措辞客观中性'),
    ('⑪ 报告↔附表数据一致', '✅', '14项数字全部勾稽一致'),
]

for i, (dim, status, note) in enumerate(dimensions, 12):
    bg = alt_fill if i % 2 == 0 else None
    set_cell(ws1, i, 1, dim, font=body_font, fill=bg, align=Alignment(vertical='center'))
    set_cell(ws1, i, 2, status, font=body_font, fill=bg, align=Alignment(horizontal='center', vertical='center'))
    ws1.merge_cells(f'C{i}:D{i}')
    set_cell(ws1, i, 3, note, font=body_font, fill=bg, align=Alignment(vertical='center'))

ws1.column_dimensions['A'].width = 28
ws1.column_dimensions['B'].width = 8
ws1.column_dimensions['C'].width = 50
ws1.column_dimensions['D'].width = 15

# ================================================================
# Sheet 2: 问题清单
# ================================================================
ws2 = wb.create_sheet('问题清单')
ws2.sheet_properties.tabColor = '00FCE4E4'

set_title_row(ws2, 1, ['等级', '序号', '段落位置', '问题类型', '问题描述', '原文摘录', '修改建议'])
ws2.row_dimensions[1].height = 30

p1_items = [
    ('P1', 1, '三、（五）1.收入情况', '数字缺失', '2025年全年营业收入数字缺失', 
     '"2025年全年累计营业收入,,312.14元"',
     '补充为：10,814,312.14元'),
    ('P1', 2, '一、（四）审计程序', '格式错误', '金额格式错误（逗号/小数点位置）', 
     '"金额累计197,764,88元"',
     '改为：197,764.88元'),
    ('P1', 3, '三、（二）在职职工及聘用人员', '数据空白', '在职职工人数、未参保职工人数均为空', 
     '"公司共有在职职工人" / "未参保职工人"',
     '补充具体人数，如无法获取注明"因公司提供数据不全无法统计"'),
    ('P1', 4, '三、（二）在职职工及聘用人员', '期间错误', '劳务费累计期间表述错误', 
     '"2024年7—1月累计750,965.68元"',
     '改为：2024年7—12月'),
    ('P1', 5, '一、（四）审计程序', '审查范围过窄', '税费审查仅覆盖至2025年1月，审计期间至12月', 
     '"复核2024年7月—2025年1月主要税费缴纳凭证"',
     '改为：2024年7月—2025年12月（覆盖完整期间）'),
    ('P1', 6, '六、审计发现', '编号缺失', '审计发现问题4~8均缺少编号前缀', 
     '问题3"采购原材料未进行存货会计处理"后，直接出现火车票丢失段、科目使用不恰当等，均无编号标题',
     '统一编号为：4.部分凭证原始凭证缺失；5.会计科目使用不恰当；6.收入成本确认无依据；7.记账凭证附件不全；8.跨期入账'),
    ('P1', 7, '六、审计发现（收入无依据段）', '空号括号', '编号括号内为空', 
     '"（2024年12月192号记账凭证"——无编号',
     '补编号：如"（6）"'),
    ('P1', 8, '六、（二）结论性事项', '章节归类', '利润分配方案归入"审计发现及结论性事项"不当', 
     '"（二）结论性事项"下为利润分配详细计算（兜底利润、公积金、超额分配等）',
     '建议单列为独立章节"七、利润分配情况"'),
]

p2_items = [
    ('P2', 9, '三、（五）1.收入总体分析', '口径冲突', '整体毛利率13.01%与2025年毛利率17.03%未说明口径差异', 
     '前文"主营业务整体毛利率13.01%" vs 后文"毛利率由约0.50%提升至约17.03%"',
     '在后一处注明"2025年单年毛利率"即可消除歧义'),
    ('P2', 10, '一、（二）审计范围', '分类不当', '审计范围分类维度不统一：报表科目vs费用类型混排', 
     '"1.收入与成本；2.费用支出（含租赁费、市场推广费等）；3.资产与负债"',
     '统一为：(1)财务状况类；(2)经营成果类；(3)税务情况'),
    ('P2', 11, '一、（三）审计方法-函证', '表述不清', '函证对象未明确，餐饮散客发函可行性存疑', 
     '"对主要客户和供应商发函，验证应收应付款的真实性"',
     '明确函证对象：对大客户/长期合作方发函，散客以POS流水替代'),
    ('P2', 12, '三、（四）负债-应交税费表', '缺失说明', '应交税费明细无印花税，未说明原因', 
     '增值税18,922.96、个税188.93、企税84,865.86、城建税333.82、教育费附加143.06等，无印花税',
     '如本期间未发生印花税，应在报告中加以说明'),
]

# P1表头
set_cell(ws2, 3, 1, '【P1】重大遗漏/错误', font=section_font, fill=red_fill, border=Border())
ws2.merge_cells('A3:G3')
ws2.row_dimensions[3].height = 28

row = 4
for item in p1_items:
    bg = alt_fill if row % 2 == 0 else None
    set_cell(ws2, row, 1, item[0], font=red_font, fill=bg, align=Alignment(horizontal='center', vertical='center'))
    set_cell(ws2, row, 2, item[1], font=body_font, fill=bg, align=Alignment(horizontal='center', vertical='center'))
    set_cell(ws2, row, 3, item[2], font=body_font, fill=bg, align=Alignment(horizontal='center', vertical='center'))
    set_cell(ws2, row, 4, item[3], font=body_font_bold, fill=bg)
    set_cell(ws2, row, 5, item[4], font=body_font, fill=bg)
    set_cell(ws2, row, 6, item[5], font=Font(name='微软雅黑', size=9, color='666666'), fill=bg)
    set_cell(ws2, row, 7, item[6], font=body_font, fill=bg)
    ws2.row_dimensions[row].height = 55
    row += 1

# P2表头
row += 1
set_cell(ws2, row, 1, '【P2】口径/表达问题', font=section_font, fill=yellow_fill, border=Border())
ws2.merge_cells(f'A{row}:G{row}')
ws2.row_dimensions[row].height = 28
row += 1

for item in p2_items:
    bg = alt_fill if row % 2 == 0 else None
    set_cell(ws2, row, 1, item[0], font=gold_font, fill=bg, align=Alignment(horizontal='center', vertical='center'))
    set_cell(ws2, row, 2, item[1], font=body_font, fill=bg, align=Alignment(horizontal='center', vertical='center'))
    set_cell(ws2, row, 3, item[2], font=body_font, fill=bg, align=Alignment(horizontal='center', vertical='center'))
    set_cell(ws2, row, 4, item[3], font=body_font_bold, fill=bg)
    set_cell(ws2, row, 5, item[4], font=body_font, fill=bg)
    set_cell(ws2, row, 6, item[5], font=Font(name='微软雅黑', size=9, color='666666'), fill=bg)
    set_cell(ws2, row, 7, item[6], font=body_font, fill=bg)
    ws2.row_dimensions[row].height = 55
    row += 1

ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 6
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 18
ws2.column_dimensions['E'].width = 40
ws2.column_dimensions['F'].width = 45
ws2.column_dimensions['G'].width = 45

# ================================================================
# Sheet 3: 数字验证通过项
# ================================================================
ws3 = wb.create_sheet('数字验证通过')
ws3.sheet_properties.tabColor = 'E8F5E9'

set_title_row(ws3, 1, ['序号', '验证项', '报告值', '验证值', '差异', '结论'])
ws3.row_dimensions[1].height = 30

verify_items = [
    (1, '资产=流动资产+非流动资产', '7,372,623.71', '7,372,623.71', 0, '一致'),
    (2, '资产-负债=所有者权益', '1,523,649.89', '1,523,649.89', 0, '一致'),
    (3, '货币资金明细求和', '3,854,813.60', '3,854,813.60', 0, '一致'),
    (4, '应收账款明细求和', '271,059.90', '271,059.90', 0, '一致'),
    (5, '其他应收款明细求和', '537,431.38', '537,431.38', 0, '一致'),
    (6, '固定资产原值-折旧=净值', '829,351.91', '829,351.91', 0, '一致'),
    (7, '长期待摊原值-摊销=余额', '1,879,966.92', '1,879,966.92', 0, '一致'),
    (8, '应交税费明细求和', '104,550.01', '104,550.01', 0, '一致'),
    (9, '其他应付款明细求和', '5,403,076.93', '5,403,076.93', 0, '一致'),
    (10, '2025年营业利润计算', '1,720,772.12', '1,720,772.12', 0, '一致'),
    (11, '2025年利润总额计算', '1,723,338.07', '1,723,338.07', 0, '一致'),
    (12, '2025年净利润计算', '1,566,799.92', '1,566,799.92', 0, '一致'),
    (13, '审计期间净利润合计', '1,523,649.89', '1,523,649.89', 0, '一致'),
    (14, '利润分配-兜底利润(6%)', '709,790.39', '709,790.39', 0, '一致'),
    (15, '利润分配-法定公积金(10%)', '152,364.99', '152,364.99', 0, '一致'),
    (16, '利润分配-超额利润', '557,425.40', '557,425.40', 0, '一致'),
]

for i, (seq, name, report_val, verify_val, diff, result) in enumerate(verify_items, 2):
    bg = alt_fill if i % 2 == 0 else None
    set_cell(ws3, i, 1, seq, font=body_font, fill=bg, align=Alignment(horizontal='center', vertical='center'))
    set_cell(ws3, i, 2, name, font=body_font, fill=bg, align=Alignment(vertical='center'))
    set_cell(ws3, i, 3, report_val, font=body_font, fill=bg, align=Alignment(horizontal='right', vertical='center'))
    set_cell(ws3, i, 4, verify_val, font=body_font, fill=bg, align=Alignment(horizontal='right', vertical='center'))
    set_cell(ws3, i, 5, str(diff), font=body_font, fill=bg, align=Alignment(horizontal='center', vertical='center'))
    set_cell(ws3, i, 6, result, font=Font(name='微软雅黑', size=10, color='2E7D32', bold=True), fill=pass_fill,
             align=Alignment(horizontal='center', vertical='center'))
    ws3.row_dimensions[i].height = 24

ws3.column_dimensions['A'].width = 8
ws3.column_dimensions['B'].width = 42
ws3.column_dimensions['C'].width = 20
ws3.column_dimensions['D'].width = 20
ws3.column_dimensions['E'].width = 12
ws3.column_dimensions['F'].width = 10

# ===== 保存 =====
output = r'D:\openclaw-workspace\temp\审计报告复核结果-食在攒劲.xlsx'
wb.save(output)
print(f'复核结果Excel已生成: {output}')
