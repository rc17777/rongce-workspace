"""Generate comprehensive audit analysis Excel report"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============ Styles ============
header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
sub_header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
title_font = Font(name='微软雅黑', bold=True, size=14, color='2F5496')
bold_font = Font(name='微软雅黑', bold=True, size=11)
normal_font = Font(name='微软雅黑', size=11)
small_font = Font(name='微软雅黑', size=10, color='666666')
red_font = Font(name='微软雅黑', size=11, color='FF0000', bold=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
wrap_align = Alignment(vertical='center', wrap_text=True)

def apply_style(ws, row, col, value, font=normal_font, fill=None, alignment=center_align, border=thin_border):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if fill: cell.fill = fill
    cell.alignment = alignment
    cell.border = border
    return cell

def write_header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        apply_style(ws, row, start_col+i, h, font=header_font, fill=header_fill)

# ============ Sheet 1: 综合结论 ============
ws1 = wb.active
ws1.title = '综合审计结论'

# Title
apply_style(ws1, 1, 1, '2024年多功能急救实训室建设项目 招投标审计分析报告', font=title_font, fill=None, border=None)
apply_style(ws1, 2, 1, '项目编号: N5100012025000628 | 采购人: 四川护理职业学院 | 代理机构: 联投项目管理(集团)有限公司', font=small_font, fill=None, border=None)
apply_style(ws1, 3, 1, f'分析日期: 2026-05-27 | 预算: 1,751,000元 | 最高限价: 1,598,400元', font=small_font, fill=None, border=None)

# Project info
info_data = [
    ('项目名称', '2024年多功能急救实训室建设项目'),
    ('采购编号', 'N5100012025000628'),
    ('采购方式', '公开招标（综合评分法，电子化采购）'),
    ('采购预算', '1,751,000.00元'),
    ('最高限价', '1,598,400.00元'),
    ('中标供应商', '成都易可天地科技有限公司'),
    ('中标金额', '1,566,000.00元'),
    ('代理服务费', '16,980元（1.698万元）'),
    ('评审日期', '2025年5月26日'),
    ('评审专家', '杨京儒（采购人代表）、罗征洪、郑雁、王文霞、肖晓辉'),
    ('采购人联系人', '梅老师 028-63955482'),
    ('监督投诉', '四川省财政厅 028-86723581/028-86723539/028-86723553'),
]
for i, (k, v) in enumerate(info_data):
    apply_style(ws1, 5+i, 1, k, font=bold_font, fill=sub_header_fill, alignment=wrap_align)
    apply_style(ws1, 5+i, 2, v, font=normal_font, alignment=wrap_align)

# Conclusion
r = 18
apply_style(ws1, r, 1, '一、招标文件合规性', font=bold_font, fill=green_fill, alignment=wrap_align)
apply_style(ws1, r, 2, '合规 — 编制规范，未发现违反政府采购法律法规的条款。', font=normal_font, fill=green_fill, alignment=wrap_align)

r = 19
apply_style(ws1, r, 1, '二、串标围标风险', font=bold_font, fill=yellow_fill, alignment=wrap_align)
apply_style(ws1, r, 2, '存在疑点但非铁证 — 报价集中度异常(1.21%)但品牌/厂家完全独立(12家零交叉)', font=normal_font, fill=yellow_fill, alignment=wrap_align)

r = 20
apply_style(ws1, r, 1, '三、核心风险信号', font=bold_font, fill=red_fill, alignment=wrap_align)
risks = [
    '1. 报价极差仅1.21%（校服项目7.8%），离散度异常',
    '2. 所有38项分项均紧贴限价97-99.9%，不同品牌独立定价不应如此一致',
    '3. 三家全部放弃中小企业10%价格扣除',
    '4. 评审得分4/5评委完全一致（到小数点2位），唯一不同的评委恰好各少2分'
]
for i, risk in enumerate(risks):
    apply_style(ws1, 21+i, 1, risk, font=normal_font, alignment=wrap_align)
    ws1.merge_cells(start_row=21+i, start_column=1, end_row=21+i, end_column=2)

r = 26
apply_style(ws1, r, 1, '四、排除的信号（正常）', font=bold_font, fill=green_fill, alignment=wrap_align)
goods = [
    '1. 三家核心产品品牌/厂家完全不同（潮天汇/天堰/诚恩），12家独立厂家零交叉',
    '2. 词级TF-IDF文本相似度仅3-8%，无实质性文本复制',
    '3. 非等差数列，排除简单协同报价',
    '4. 注册地不同省市（成都2家+江西1家），品牌产地五省市各异'
]
for i, g in enumerate(goods):
    apply_style(ws1, 27+i, 1, g, font=normal_font, alignment=wrap_align)
    ws1.merge_cells(start_row=27+i, start_column=1, end_row=27+i, end_column=2)

r = 32
apply_style(ws1, r, 1, '五、后续审计建议（按优先级）', font=bold_font, fill=sub_header_fill, alignment=wrap_align)
suggestions = [
    ('P0-紧急', '向代理机构调取电子交易系统投标IP记录 — L2决定性证据'),
    ('P1-重要', '天眼查/企查查查询三家工商关联关系 — L8股东/高管交叉'),
    ('P2-建议', '获取原始.docx投标文件提取元数据 — L5同源创建/WPS版本'),
    ('P3-建议', '核实三家是否确实不属于小微企业 — 验证中小企业声明真实性'),
    ('P4-例行', '中标后验收环节抽查设备品牌和型号 — 防止以次充好'),
]
for i, (pri, sug) in enumerate(suggestions):
    fill = red_fill if 'P0' in pri else (yellow_fill if 'P1' in pri else None)
    apply_style(ws1, 33+i, 1, pri, font=bold_font, fill=fill, alignment=center_align)
    apply_style(ws1, 33+i, 2, sug, font=normal_font, fill=fill, alignment=wrap_align)

ws1.column_dimensions['A'].width = 22
ws1.column_dimensions['B'].width = 85

# ============ Sheet 2: 评审得分明细 ============
ws2 = wb.create_sheet('评审得分明细')

apply_style(ws2, 1, 1, '评审得分汇总表（数据来源：归档资料 P106 评分汇总表 + P104 价格评审表）', font=title_font, fill=None, border=None)
apply_style(ws2, 2, 1, '评审日期: 2025-05-26 | 评审专家: 杨京儒(采购人代表)、罗征洪、郑雁、王文霞、肖晓辉', font=small_font, fill=None, border=None)

# Part A: Individual judge scores
r = 4
apply_style(ws2, r, 1, 'A. 各评委打分明细（满分100分）', font=bold_font, fill=sub_header_fill, alignment=wrap_align)
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)

r = 5
headers = ['排名', '供应商名称', '郑雁', '杨京儒', '罗征洪', '王文霞', '肖晓辉', '总分合计', '平均得分', '得分极差']
write_header_row(ws2, r, headers)

score_data = [
    [1, '成都易可天地科技有限公司', 93.5, 93.5, 92.5, 93.5, 93.5, 466.50, 93.30, 1.0],
    [2, '江西正好医疗器械有限公司', 69.18, 69.18, 67.18, 69.18, 69.18, 343.90, 68.78, 2.0],
    [3, '四川省好医助医疗器械有限公司', 65.97, 65.97, 63.97, 65.97, 65.97, 327.85, 65.57, 2.0],
]

for i, row_data in enumerate(score_data):
    for j, val in enumerate(row_data):
        f = green_fill if i == 0 else None
        font_style = bold_font if j == 7 else normal_font
        if j == 1:
            apply_style(ws2, 6+i, j+1, val, font=font_style, fill=f, alignment=wrap_align)
        else:
            apply_style(ws2, 6+i, j+1, val, font=font_style, fill=f, alignment=center_align)
        if j >= 7 or (j >= 2 and j <= 6):
            ws2.cell(row=6+i, column=j+1).number_format = '0.00'

# Highlight: 4/5 judges gave identical scores
r = 10
apply_style(ws2, r, 1, '⚠️ 得分一致性异常', font=red_font, fill=red_fill, alignment=wrap_align)
notes = [
    '• 郑雁、杨京儒、王文霞、肖晓辉 4位评委对各投标人评分完全一致（精确到0.01分）',
    '• 唯一不同的评委罗征洪，恰好给每个投标人各少2分（92.5/67.18/63.97 vs 93.5/69.18/65.97）',
    '• 4/5评委对含10分主观分的评分产生完全相同的总分值，概率极低',
    '• 提示：评分存在"协调一致"或"预填"的可能'
]
for i, note in enumerate(notes):
    apply_style(ws2, 11+i, 1, note, font=normal_font, alignment=wrap_align)
    ws2.merge_cells(start_row=11+i, start_column=1, end_row=11+i, end_column=10)

# Part B: Price scores
r = 17
apply_style(ws2, r, 1, 'B. 价格评分明细（满分30分）', font=bold_font, fill=sub_header_fill, alignment=wrap_align)
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)

r = 18
price_headers = ['排名', '供应商名称', '投标报价', '评标报价', '占限价', '报价得分', '价格分差距', '得分计算']
write_header_row(ws2, r, price_headers)

price_data = [
    [1, '成都易可天地科技有限公司', 1566000, 1566000, '98.0%', 30.00, '基准', '30*(1566000/1566000)'],
    [2, '江西正好医疗器械有限公司', 1574022, 1574022, '98.5%', 29.85, -0.15, '30*(1566000/1574022)'],
    [3, '四川省好医助医疗器械有限公司', 1585000, 1585000, '99.2%', 29.64, -0.36, '30*(1566000/1585000)'],
]

for i, row_data in enumerate(price_data):
    for j, val in enumerate(row_data):
        f = green_fill if i == 0 else None
        apply_style(ws2, 19+i, j+1, val, font=normal_font, fill=f, 
                    alignment=center_align if j != 7 else wrap_align)
        if j == 2 or j == 3:
            ws2.cell(row=19+i, column=j+1).number_format = '#,##0'
        if j == 5:
            ws2.cell(row=19+i, column=j+1).number_format = '0.00'

# Part C: Non-price score analysis
r = 24
apply_style(ws2, r, 1, 'C. 非价格得分分析（满分70分 = 技术45+演示12+服务10+业绩2+环境1）', font=bold_font, fill=sub_header_fill, alignment=wrap_align)
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)

r = 25
np_headers = ['排名', '供应商名称', '总得分', '价格得分', '非价格得分', '非价格满分', '非价格得分率', '与第1名差距']
write_header_row(ws2, r, np_headers)

np_data = [
    [1, '成都易可天地科技有限公司', 93.30, 30.00, 63.30, 70, '90.4%', '—'],
    [2, '江西正好医疗器械有限公司', 68.78, 29.85, 38.93, 70, '55.6%', '-24.37'],
    [3, '四川省好医助医疗器械有限公司', 65.57, 29.64, 35.93, 70, '51.3%', '-27.37'],
]

for i, row_data in enumerate(np_data):
    for j, val in enumerate(row_data):
        f = green_fill if i == 0 else None
        font_style = red_font if (j == 7 and i > 0) else normal_font
        apply_style(ws2, 26+i, j+1, val, font=font_style, fill=f, alignment=center_align)
        if j in [2, 3, 4]:
            ws2.cell(row=26+i, column=j+1).number_format = '0.00'

r = 30
apply_style(ws2, r, 1, '⚠️ 非价格得分断崖式差距', font=red_font, fill=red_fill, alignment=wrap_align)
gap_notes = [
    '• 第1名非价格得分63.30 vs 第2名38.93 = 差距24.37分（满分70分）',
    '• 第2名和第3名非价格得分仅差3分（38.93 vs 35.93），差异不大',
    '• 排除价格因素后，第1名的技术/演示/服务得分是第2名的1.6倍',
    '• 提示：技术参数满足度可能存在"量身定制"或第2/3名"陪标"嫌疑'
]
for i, note in enumerate(gap_notes):
    apply_style(ws2, 31+i, 1, note, font=normal_font, alignment=wrap_align)
    ws2.merge_cells(start_row=31+i, start_column=1, end_row=31+i, end_column=10)

ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 30
for c in range(3, 11):
    ws2.column_dimensions[get_column_letter(c)].width = 14

# ============ Sheet 3: 投标报价对比 ============
ws3 = wb.create_sheet('投标报价与品牌对比')

apply_style(ws3, 1, 1, '分项报价与品牌/厂家对比（核心代表产品）', font=title_font, fill=None, border=None)
apply_style(ws3, 2, 1, '数据来源：投标报价表 PDF 表格提取', font=small_font, fill=None, border=None)

items = [
    ('基于现代物联网技术的新型高效急救训练系统（3套）', 240000, [
        ('四川省好医助医疗器械有限公司', 239700, '潮天汇 CTH-C23', '广州潮天汇科技有限公司', '广州市黄埔区'),
        ('江西正好医疗器械有限公司', 238200, '上海诚恩 CE/154', '上海诚恩医学科技有限公司', '上海市奉贤区'),
        ('成都易可天地科技有限公司', 234000, '天堰科技 TY9049.10', '天津天堰科技股份有限公司', '天津市滨海新区'),
    ]),
    ('心肺复苏模拟人系统/学生机（20套）', 400000, [
        ('四川省好医助医疗器械有限公司', 392000, '潮天汇 CTH-452', '广州潮天汇科技有限公司', '广州市黄埔区'),
        ('成都易可天地科技有限公司', 396000, '天堰科技 TY9019', '天津天堰科技股份有限公司', '天津市滨海新区'),
        ('江西正好医疗器械有限公司', 396000, '上海诚恩 CE/201', '上海诚恩医学科技有限公司', '上海市奉贤区'),
    ]),
    ('可视化喉镜（1台）', 16800, [
        ('四川省好医助医疗器械有限公司', 16700, '斯美特 SM', '泰兴市斯美特医疗器械有限公司', '江苏省泰兴市'),
        ('成都易可天地科技有限公司', 16500, '辉春 SPMHP-001', '泰州市辉春医疗器械有限公司', '江苏省泰州市高港区'),
        ('江西正好医疗器械有限公司', 16464, '江苏永乐 YL01-IV', '江苏永乐医疗科技有限公司', '江苏省泰兴市'),
    ]),
    ('智慧面板（2套）', 7000, [
        ('四川省好医助医疗器械有限公司', 6980, '海普迪 TTP100', '成都海普迪科技有限公司', '成都市武侯区'),
        ('江西正好医疗器械有限公司', 6860, '广州力扑 TH10C', '广州力扑智能科技有限公司', '广州市天河区'),
        ('成都易可天地科技有限公司', 6800, '锐捷 RG-SEPanel300', '锐捷网络股份有限公司', '福州市仓山区'),
    ]),
]

r = 4
for item_name, limit, bidders in items:
    apply_style(ws3, r, 1, f'{item_name} [限价: {limit:,}元]', font=bold_font, fill=sub_header_fill, alignment=wrap_align)
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1
    
    headers = ['投标人', '报价(元)', '占限价', '品牌/型号', '生产厂家', '产地', '是否限价品']
    write_header_row(ws3, r, headers)
    r += 1
    
    for bidder, price, brand, mfg, origin in bidders:
        is_limit = '是' if price == limit else '否'
        pct = price / limit * 100
        apply_style(ws3, r, 1, bidder, font=normal_font, alignment=wrap_align)
        apply_style(ws3, r, 2, price, font=normal_font, alignment=center_align)
        ws3.cell(row=r, column=2).number_format = '#,##0'
        apply_style(ws3, r, 3, f'{pct:.1f}%', font=normal_font, alignment=center_align)
        apply_style(ws3, r, 4, brand, font=normal_font, alignment=center_align)
        apply_style(ws3, r, 5, mfg, font=normal_font, alignment=center_align)
        apply_style(ws3, r, 6, origin, font=normal_font, alignment=center_align)
        fill_c = red_fill if pct >= 99 else (yellow_fill if pct >= 97 else green_fill)
        apply_style(ws3, r, 7, is_limit, font=normal_font, fill=fill_c, alignment=center_align)
        r += 1
    r += 1

# Key findings
apply_style(ws3, r, 1, '关键发现', font=bold_font, fill=sub_header_fill, alignment=wrap_align)
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
findings = [
    '✅ 三家核心产品使用完全不同的品牌和制造商（潮天汇/天堰/诚恩）',
    '✅ 12个独立厂家，0交叉重叠 — 排除"同一品牌围标"模式',
    '🟡 可视化喉镜三家均选江苏泰州/泰兴厂家（产业聚集区，不同厂家）',
    '🔴 所有品目报价均在限价97%-99.9%区间，紧贴上限',
    '🔴 好医助急救训练系统报价239,700（占限价99.9%），仅有300元折扣空间',
    '✅ 好医助选成都海普迪（智慧面板）、易可天地选成都欣荣泰联（机柜）— 本地供应链'
]
for i, f_text in enumerate(findings):
    apply_style(ws3, r+1+i, 1, f_text, font=normal_font, alignment=wrap_align)
    ws3.merge_cells(start_row=r+1+i, start_column=1, end_row=r+1+i, end_column=7)

ws3.column_dimensions['A'].width = 28
ws3.column_dimensions['B'].width = 14
ws3.column_dimensions['C'].width = 10
ws3.column_dimensions['D'].width = 22
ws3.column_dimensions['E'].width = 28
ws3.column_dimensions['F'].width = 18
ws3.column_dimensions['G'].width = 12

# ============ Sheet 4: 十层检测矩阵 ============
ws4 = wb.create_sheet('十层检测矩阵')

apply_style(ws4, 1, 1, '串标围标十层检测体系 — 本项目全量结果', font=title_font, fill=None, border=None)
apply_style(ws4, 2, 1, '参考框架: procurement-audit-models 十层检测体系 | 对比基准: 校服采购项目 TQ-CG-(2025)093', font=small_font, fill=None, border=None)

r = 4
headers = ['层级', '检测维度', '数据源', '本项目结果', '风险等级', '校服项目对比', '详细说明']
write_header_row(ws4, r, headers)

detection_data = [
    ['L1', '报价规律性', '投标总价+分项报价', '极差1.21%，全部紧贴限价97-99.9%', '🔴 高',
     '校服项目极差7.8%', '报价离散度显著低于正常竞争水平。对比校服采购项目（5家极差7.8%），本项目3家仅1.21%的极差极度异常。分项报价均紧贴天花板。'],
    ['L2', '投标IP/MAC', '电子交易系统后台', '❓ 缺数据', '— 待查',
     'N/A', '最关键的缺失证据。需向代理机构（联投项目管理公司）调取。同一IP即铁证。'],
    ['L3', 'TF-IDF文本雷同', '投标文件全文', '词级3-8%（正常）/ char级89-94%（模板）', '🟢 低',
     '校服最高36.5%', '词级分析正常，char级高相似均对应标准承诺函/声明函等模板文本。精确段落匹配显示高相似段落全为政府采购标准模板。'],
    ['L4', '图片/资源哈希', '.docx word/media/', '❓ 仅有PDF，无法解压.docx', '— 缺数据',
     '校服1095张0重复', '原始投标为.docx文件，但仅有PDF扫描版。无法提取嵌入图片做哈希比对。'],
    ['L5', '元数据交叉', '.docx core.xml', '❓ 仅有PDF，无原始.docx', '— 缺数据',
     '校服4/4 WPS版本一致', '无法提取作者/创建时间/最后修改者/WPS版本GUID。需从代理机构或投标人获取原始.docx文件。'],
    ['L6', '文档结构/样式', 'styles.xml', '❓ 仅有PDF', '— 缺数据',
     'N/A', '无法比对。'],
    ['L7', '打印机/扫描仪', 'PDF Producer/Creator', '❓ 缺数据', '— 缺数据',
     '校服已查不同设备', '未提取PDF元数据字段。'],
    ['L8', '工商关联穿透', '天眼查/企查查', '❓ 待查（网络超时）', '— 待查',
     'N/A', '需查询三家公司的股东/高管/对外投资交叉关系。'],
    ['L9', '保证金/资金链', '银行汇款凭证', 'N/A — 本项目不收取投标保证金', '— N/A',
     '校服N/A', '招标文件明确不收取投标保证金。'],
    ['L10', '代理人/IP/签到', '开标签到表/授权委托书', '❓ 归档资料扫描OCR未完成', '— 缺数据',
     '校服缺数据', '归档资料P62为开标记录表（扫描件），OCR质量不足以提取完整表格数据。'],
    ['追加', '品牌/厂家归属', '投标报价表', '12家独立厂家，0交叉重叠', '🟢 低',
     'N/A', '三家核心产品使用完全不同品牌和制造商。有效排除"同一品牌围标"模式。'],
    ['追加', '中小企业声明', '中小企业声明函', '三家均标注"不响应此评审点内容"', '🟡 关注',
     'N/A', '全部放弃10%价格扣除。若确有符合小微标准者，主动放弃不符合商业逻辑。需核实企业实际规模。'],
    ['追加', '评审得分一致性', '评分汇总表', '4/5评委得分完全相同（小数点2位精度）', '🔴 高',
     'N/A', '含10分主观分的评审中，4位评委对3家投标人的总分完全一致（精确到0.01分），唯一不同的评委恰好各少2分。概率极低。'],
]

for i, row_data in enumerate(detection_data):
    for j, val in enumerate(row_data):
        if j == 4:  # Risk level
            fill = red_fill if '高' in val else (yellow_fill if '关注' in val or '待查' in val else (green_fill if '低' in val else None))
            apply_style(ws4, 5+i, j+1, val, font=bold_font, fill=fill, alignment=center_align)
        elif j == 6:  # Detail
            apply_style(ws4, 5+i, j+1, val, font=normal_font, alignment=wrap_align)
        else:
            apply_style(ws4, 5+i, j+1, val, font=normal_font, alignment=center_align if j != 3 else wrap_align)

ws4.column_dimensions['A'].width = 8
ws4.column_dimensions['B'].width = 18
ws4.column_dimensions['C'].width = 20
ws4.column_dimensions['D'].width = 32
ws4.column_dimensions['E'].width = 12
ws4.column_dimensions['F'].width = 20
ws4.column_dimensions['G'].width = 55

# ============ Sheet 5: 招标文件合规 ============
ws5 = wb.create_sheet('招标文件合规审查')

apply_style(ws5, 1, 1, '招标文件合规性审查', font=title_font, fill=None, border=None)

r = 3
headers = ['审查事项', '审查内容', '审查结果', '法律依据', '评价']
write_header_row(ws5, r, headers)

compliance_data = [
    ['资格条件', '是否设置排他性/歧视性条款', '未发现排他性条款，仅要求政府采购法第22条基本条件', '政府采购法第22条', '✅ 合规'],
    ['采购方式', '公开招标程序合规性', '四川省政府采购一体化平台电子化采购', '政府采购法第26条', '✅ 合规'],
    ['评审办法', '综合评分法分值设置', '客观分90+主观分10+价格30=100分，客观化程度高', '财库[2007]2号', '✅ 合规'],
    ['价格分值', '价格分占比≥30%', '价格分30分(30%)，符合货物类≥30%要求', '政府采购货物和服务招标投标管理办法第55条', '✅ 合规'],
    ['中小企业', '是否落实中小企业政策', '非专门面向但给予10%价格扣除', '财库[2020]46号', '✅ 合规'],
    ['环境标志', '优先采购环保产品', '3种机柜要求优先采购环境标志产品', '财库[2019]9号', '✅ 合规'],
    ['进口产品', '是否限制进口', '明确不允许采购进口产品', '政府采购进口产品管理办法', '✅ 合规'],
    ['投标保证金', '保证金设置是否合理', '不收取投标保证金', '减负政策', '✅ 合规'],
    ['履约保证金', '履约保证金比例', '合同金额的5%', '≤10%即合规', '✅ 合规但偏高'],
    ['现场演示', '是否设置不合理演示要求', '8项需现场演示(12分)，对跨省投标人存在隐性门槛', '—', '🟡 关注'],
    ['技术参数', '是否指向特定品牌', '技术参数为功能性描述，未发现指向特定品牌的排他性参数', '政府采购法实施条例第20条', '✅ 合规'],
    ['合同条款', '付款/验收/质保条款', '验收合格后付款，3年质保', '—', '✅ 合规'],
]

for i, row_data in enumerate(compliance_data):
    for j, val in enumerate(row_data):
        fill = yellow_fill if '关注' in str(val) else None
        apply_style(ws5, 4+i, j+1, val, font=normal_font, fill=fill, 
                    alignment=wrap_align if j in [1,2,3] else center_align)

ws5.column_dimensions['A'].width = 14
ws5.column_dimensions['B'].width = 30
ws5.column_dimensions['C'].width = 42
ws5.column_dimensions['D'].width = 28
ws5.column_dimensions['E'].width = 14

# ============ Sheet 6: TF-IDF文本雷同 ============
ws6 = wb.create_sheet('TF-IDF文本雷同分析')

apply_style(ws6, 1, 1, 'TF-IDF 文本相似度分析', font=title_font, fill=None, border=None)

r = 3
headers = ['对比对', 'char-wb全文', 'char-wb技术应答', '词级全文', '词级技术应答', '判定']
write_header_row(ws6, r, headers)

tfidf_data = [
    ['好医助 vs 易可天地', '89.9%', '54.2%', '8.0%', '—', '🟢 正常（高char相似来自模板文本）'],
    ['好医助 vs 江西正好', '93.8%', '41.7%', '8.0%', '—', '🟢 正常'],
    ['易可天地 vs 江西正好', '91.0%', '46.6%', '3.5%', '—', '🟢 正常'],
]

for i, row_data in enumerate(tfidf_data):
    for j, val in enumerate(row_data):
        apply_style(ws6, 4+i, j+1, val, font=normal_font, 
                    alignment=wrap_align if j == 5 else center_align)

r = 8
apply_style(ws6, r, 1, '说明', font=bold_font, fill=sub_header_fill, alignment=wrap_align)
ws6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

notes = [
    'char-wb: 字符级word-boundary分析，适合检测逐字复制',
    '词级: jieba分词后TF-IDF向量化+余弦相似度，排除停用词',
    'char级89-94%高相似经精确定位分析，全部对应政府采购标准承诺函/声明函/法定代表人授权书等模板化文档',
    '精确段落匹配: 相似度=1.000的段落均为标准模板文本，无自行撰写内容的异常匹配',
    '阈值参考: ≥80%=可疑，≥90%=高度可疑。但需排除模板文本干扰。',
    '技术应答char-wb 41-54%处于正常范围（不同品牌产品的技术参数描述自然有一定差异）',
]
for i, note in enumerate(notes):
    apply_style(ws6, 9+i, 1, note, font=normal_font, alignment=wrap_align)
    ws6.merge_cells(start_row=9+i, start_column=1, end_row=9+i, end_column=6)

ws6.column_dimensions['A'].width = 22
for c in range(2, 7):
    ws6.column_dimensions[get_column_letter(c)].width = 22

# ============ Sheet 7: 中小企业声明 ============
ws7 = wb.create_sheet('中小企业声明分析')

apply_style(ws7, 1, 1, '中小企业声明函分析', font=title_font, fill=None, border=None)

r = 3
headers = ['投标人', '声明内容', '声明日期', '10%扣除', '问题评估']
write_header_row(ws7, r, headers)

sme_data = [
    ['四川省好医助医疗器械有限公司', '不响应此评审点内容', '2025-05-23', '放弃', '🟡 需核实：若为小微企业，主动放弃10%竞争优势不符合商业逻辑'],
    ['成都易可天地科技有限公司', '不响应此评审点内容', '2025-05-22', '放弃', '🟡 需核实：声明日期（5/22）早于另两家（5/23）'],
    ['江西正好医疗器械有限公司', '不响应此评审点内容', '2025-05-23', '放弃', '🟡 需核实：与好医助同日提交声明'],
]

for i, row_data in enumerate(sme_data):
    for j, val in enumerate(row_data):
        f = yellow_fill
        apply_style(ws7, 4+i, j+1, val, font=normal_font, fill=f, 
                    alignment=wrap_align if j in [1, 4] else center_align)

r = 8
apply_style(ws7, r, 1, '关键发现', font=bold_font, fill=sub_header_fill, alignment=wrap_align)
ws7.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

findings = [
    '1. 三家投标人均在中小企业声明函中标注"不响应此评审点内容"',
    '2. 即三家均不申请小微企业评审优惠（10%价格扣除）',
    '3. 如果三家均为非小微企业则属正常',
    '4. 但若任何一家实际符合小微标准却主动放弃，则存在以下疑问：',
    '   - 主动放弃10%价格竞争优势不符合商业逻辑',
    '   - 可能为配合其他投标人（确保价格分排序不变）',
    '5. 建议：通过国家企业信用信息公示系统或中小企业认定平台核实三家的实际企业规模',
    '6. 好医助与江西正好声明日期相同（5/23），易可天地提前一天（5/22）'
]
for i, f_text in enumerate(findings):
    apply_style(ws7, 9+i, 1, f_text, font=normal_font, alignment=wrap_align)
    ws7.merge_cells(start_row=9+i, start_column=1, end_row=9+i, end_column=5)

ws7.column_dimensions['A'].width = 30
ws7.column_dimensions['B'].width = 22
ws7.column_dimensions['C'].width = 14
ws7.column_dimensions['D'].width = 10
ws7.column_dimensions['E'].width = 55

# Save
out_path = r'D:\openclaw-workspace\output\急救实训室_extracted\招投标审计分析报告.xlsx'
wb.save(out_path)
print(f'Saved: {out_path}')
