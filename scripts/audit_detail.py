import openpyxl
PATH = r'C:\Users\scrccpa\Desktop\融策审计过程记录系统=项目经理版(6).xlsx'
wb = openpyxl.load_workbook(PATH, data_only=True)

with open(r'D:\openclaw-workspace\output\audit_tail.txt', 'w', encoding='utf-8') as f:
    # Sheet 5: 问题与证据
    ws5 = wb['5-问题与证据']
    f.write("=== 5-问题与证据 ===\n")
    for row in ws5.iter_rows(min_row=1, values_only=True):
        vals = [str(v)[:200] if v else '' for v in row]
        if any(v for v in vals):
            f.write('  ' + ' | '.join(vals) + '\n')
    
    # Sheet 1: 项目信息  
    ws1 = wb['1-项目信息']
    f.write("\n=== 1-项目信息 ===\n")
    for row in ws1.iter_rows(min_row=1, values_only=True):
        vals = [str(v)[:200] if v else '' for v in row]
        if any(v for v in vals):
            f.write('  ' + ' | '.join(vals) + '\n')
            
    # Sheet 4: 访谈记录
    ws4 = wb['4-访谈记录']
    f.write("\n=== 4-访谈记录 ===\n")
    for row in ws4.iter_rows(min_row=1, values_only=True):
        vals = [str(v)[:200] if v else '' for v in row]
        if any(v for v in vals):
            f.write('  ' + ' | '.join(vals) + '\n')
            
    # Sheet 6: 复核记录
    ws6 = wb['6-复核记录']
    f.write("\n=== 6-复核记录 ===\n")
    for row in ws6.iter_rows(min_row=1, values_only=True):
        vals = [str(v)[:200] if v else '' for v in row]
        if any(v for v in vals):
            f.write('  ' + ' | '.join(vals) + '\n')

    # Last 50 records from Sheet 2
    ws2 = wb['2-审计过程']
    f.write("\n=== 2-审计过程 最后50条记录 ===\n")
    all_rows = []
    for row in ws2.iter_rows(min_row=3, values_only=True):
        vals = [str(v) if v else '' for v in row]
        all_rows.append(vals)
    
    for r in all_rows[-50:]:
        f.write(f"[{r[0]}] {r[1]} | {r[3]} | {r[2]} | 发现={r[6][:80] if r[6] else '无'}\n")

print("完成")
