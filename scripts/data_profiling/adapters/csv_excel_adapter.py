# -*- coding: utf-8 -*-
"""
CSV/Excel → SDF 桥接适配器 (已有 profile_builder 的SDF化补充)
流程: 编码检测(CSV) / 多Sheet遍历(Excel) → 表头定位 → 行读取 → SDF

用法: python csv_excel_adapter.py --source "data.xlsx" --out sdf.json [--sheet 名称]
"""
import sys, os, json, csv, argparse
from pathlib import Path
sys.stdout.reconfigure(encoding='utf-8')

HERE = Path(__file__).parent
sys.path.insert(0, str(HERE))
from common import build_sdf, save_sdf, print_sdf_summary, sha256_file


def read_csv_sdf(source, sheet=None, label='csv'):
    path = Path(source)
    raw = path.read_bytes()
    for enc in ('utf-8-sig', 'utf-8', 'gbk', 'gb18030', 'latin-1'):
        try:
            text = raw.decode(enc)
            encoding = enc
            break
        except (UnicodeDecodeError, ValueError):
            continue
    else:
        text = raw.decode('utf-8', errors='replace')
        encoding = 'utf-8(replace)'

    # 分隔符识别: 逗号/分号/Tab
    sample = text[:4000]
    delim = ';' if sample.count(';') > sample.count(',') and sample.count(';') > 0 else (',' if ',' in sample else '\t')
    if '\t' in sample and sample.count('\t') > sample.count(','):
        delim = '\t'

    reader = csv.reader(text.splitlines(), delimiter=delim)
    rows = [r for r in reader if any(str(c).strip() for c in r)]
    if not rows:
        raise ValueError(f'CSV为空: {source}')

    header = [str(c).strip() if c else f'列{i+1}' for i, c in enumerate(rows[0])]
    ncols = len(header)
    data_rows = []
    for r in rows[1:]:
        if not any(str(c).strip() for c in r):
            continue
        data_rows.append({header[i]: (r[i] if i < len(r) else None) for i in range(ncols)})

    sdf = build_sdf('csv', os.path.basename(source), data_rows,
                    checksum=sha256_file(source),
                    extra_source={'encoding': encoding, 'delimiter': delim})
    return sdf


def read_excel_sdf(source, sheet=None, label='excel'):
    import openpyxl
    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    sheets = [sheet] if sheet else wb.sheetnames
    all_rows = []
    for sn in sheets:
        if sn not in wb.sheetnames:
            raise ValueError(f'Sheet不存在: {sn} (可选: {wb.sheetnames})')
        ws = wb[sn]
        data = list(ws.iter_rows(values_only=True))
        if not data:
            continue
        # 多行表头检测: 跳过纯标题行(单单元格或含'说明/用法')
        header_idx = 0
        while header_idx < min(3, len(data)):
            first = data[header_idx]
            non_empty = [c for c in first if c]
            if len(non_empty) == 1 and len(first) > 3:
                header_idx += 1
            elif any('说明' in str(c) or '用法' in str(c) for c in non_empty):
                header_idx += 1
            else:
                break
        header = [str(c) if c else f'列{i+1}' for i, c in enumerate(data[header_idx])]
        ncols = len(header)
        for row in data[header_idx + 1:]:
            if not any(c is not None for c in row):
                continue
            all_rows.append({header[i]: (row[i] if i < len(row) else None) for i in range(ncols)})
    wb.close()
    if not all_rows:
        raise ValueError(f'Excel无数据: {source}')
    sdf = build_sdf('excel', os.path.basename(source), all_rows,
                    checksum=sha256_file(source),
                    extra_source={'sheets': sheets})
    return sdf


def convert(source, out_path=None, sheet=None, label='data'):
    ext = Path(source).suffix.lower()
    if ext in ('.csv', '.tsv', '.txt'):
        sdf = read_csv_sdf(source, sheet, label)
    elif ext in ('.xlsx', '.xlsm', '.xls'):
        sdf = read_excel_sdf(source, sheet, label)
    else:
        raise ValueError(f'不支持的文件类型: {ext}')
    if out_path:
        save_sdf(sdf, out_path)
    return sdf


if __name__ == '__main__':
    p = argparse.ArgumentParser(description='CSV/Excel → SDF 桥接适配器')
    p.add_argument('--source', required=True, help='CSV/Excel文件路径')
    p.add_argument('--out', help='SDF输出路径')
    p.add_argument('--sheet', help='Excel Sheet名(默认全部)')
    args = p.parse_args()

    sdf = convert(args.source, args.out, args.sheet)
    print_sdf_summary(sdf)
