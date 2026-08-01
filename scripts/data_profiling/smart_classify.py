# -*- coding: utf-8 -*-
"""
智能分类三步法 — Skill 9 (P1)
预打标(归组抽样)→构建分类体系(审计师确认)→全量分类(低置信复判)
用法: python smart_classify.py --source "data.xlsx" --col "报销事由" --project "XX项目"
"""
import sys, os, json, re, argparse, datetime, hashlib, csv
from pathlib import Path
from collections import Counter
sys.stdout.reconfigure(encoding='utf-8')

HERE = Path(__file__).parent

# ─── Step 1: 预打标 (Pre-tagging) ──────────────────────
def normalize_text(text):
    """文本归一化：去空格、去标点、全角转半角"""
    if not text:
        return ''
    text = str(text).strip()
    text = re.sub(r'[\s\u3000]+', '', text)
    text = re.sub(r'[，。；：、！？（）《》【】"\'…—\.,;:!?()\[\]{}""'']', '', text)
    return text

def text_fingerprint(text):
    """简单文本指纹用于归组"""
    return hashlib.md5(text.encode('utf-8')).hexdigest()[:8]

def pre_tag(rows, col_name, sample_size=200):
    """
    Step 1: 归组相似记录 + 抽取代表性样本 + 生成候选标签
    返回: groups, samples, candidate_labels
    """
    texts = [str(r.get(col_name, '')) for r in rows if str(r.get(col_name, '')).strip()]
    total = len(texts)
    
    # 1a: 按长度分组（初步归组）
    short = [t for t in texts if len(t) <= 10]      # 极短 → 可能是代码
    medium = [t for t in texts if 10 < len(t) <= 50] # 中 → 可能是摘要
    long = [t for t in texts if len(t) > 50]          # 长 → 原文
    
    groups = {
        'short': {'count': len(short), 'sample': short[:50]},
        'medium': {'count': len(medium), 'sample': medium[:50]},
        'long': {'count': len(long), 'sample': long[:50]},
    }
    
    # 1b: 关键词聚类（简单TF）
    def extract_keywords(texts_list, top_n=30):
        """提取高频特征词（2-4字）"""
        word_freq = Counter()
        for t in texts_list[:sample_size]:
            # 简单按2-4字分词
            for wlen in [2, 3, 4]:
                for i in range(len(t) - wlen + 1):
                    w = t[i:i+wlen]
                    if re.match(r'^[\u4e00-\u9fff]+$', w):
                        word_freq[w] += 1
        # 过滤停用词
        stopwords = {'有限公司', '有限责任', '成都市', '四川省', '相关的', '进行了', '进行了对', '根据规定', '按照要求'}
        return [(w, c) for w, c in word_freq.most_common(top_n) 
                if w not in stopwords and c >= max(word_freq.values()) * 0.02]
    
    keywords = extract_keywords(texts)
    
    # 1c: 候选标签（从样本中提取，供 AI 归纳）
    sample = texts[:sample_size]
    
    # 1d: 去重后代表性样本
    seen = set()
    deduped = []
    for t in texts:
        fp = text_fingerprint(t)
        if fp not in seen:
            seen.add(fp)
            deduped.append(t)
    
    result = {
        'step': 'pretag',
        'total_records': total,
        'unique_records': len(deduped),
        'duplicate_rate': round((1 - len(deduped) / total) * 100, 1) if total else 0,
        'length_groups': groups,
        'top_keywords': keywords,
        'samples': deduped[:sample_size],
        'samples_for_ai': deduped[:min(sample_size, 100)],  # 给 AI 看的代表性样本
        'timestamp': datetime.datetime.now().isoformat()
    }
    return result

# ─── Step 2: 构建分类体系模板 ──────────────────────────
def build_classification_template(pretag_result, existing_system=None):
    """
    Step 2: 基于预打标结果生成分类体系草案
    产出: classification_system (可 JSON 化供审计师审阅)
    """
    keywords = pretag_result['top_keywords']
    samples = pretag_result['samples']
    
    # 按关键词聚成候选类别
    categories = []
    used_kws = set()
    
    for kw, freq in keywords[:15]:
        # 在每个类别名下找包含该关键词的样本
        cat_name = kw
        matching = [s for s in samples if kw in s][:5]
        if len(matching) >= 2:
            categories.append({
                'code': f'CAT{len(categories)+1:02d}',
                'name': cat_name,
                'description': f'（待确认）包含关键词"{kw}"的记录',
                'keyword_triggers': [kw],
                'sample_records': matching,
                'estimated_coverage': f'{freq}/{pretag_result["total_records"]}',
                'confidence': 'medium',
                'notes': ''
            })
            used_kws.add(kw)
    
    # 添加"其他"类别兜底
    categories.append({
        'code': f'CAT{len(categories)+1:02d}',
        'name': '其他',
        'description': '无法归入上述类别的记录',
        'keyword_triggers': [],
        'sample_records': [],
        'estimated_coverage': '',
        'confidence': 'low',
        'notes': '需审计师确认，或从该类别拆分出新类别'
    })
    
    system = {
        'step': 'build_system',
        'version': 'draft',
        'created_at': datetime.datetime.now().isoformat(),
        'classification_column': '',
        'categories': categories,
        'coverage_metrics': {
            'total_categories': len(categories),
            'duplicate_rate': pretag_result['duplicate_rate'],
            'unique_records': pretag_result['unique_records']
        },
        'review_checklist': [
            '□ 各类别定义是否互斥且穷尽（MECE）？',
            '□ 类别命名是否符合审计业务语境？',
            '□ "其他"类别占比是否 ≤10%？',
            '□ 是否有应拆分、应合并的类别？',
            '□ 分类口径是否与数据档案中的定义一致？',
        ],
        'confirmed': False,
        'confirmed_by': '',
        'confirmed_date': ''
    }
    
    return system

# ─── Step 3: 全量分类 ──────────────────────────────────
def batch_classify(rows, col_name, classification_system):
    """
    Step 3: 按确认的分类体系逐条分类
    策略: 关键词匹配(高置信度) + 待 AI 复判(低置信度)
    返回: 分类结果列表
    """
    categories = classification_system.get('categories', [])
    results = []
    low_confidence = []
    
    for i, row in enumerate(rows):
        text = str(row.get(col_name, '')).strip()
        if not text:
            results.append({'row': i, 'text': text, 'category': '空值', 'confidence': 'certain'})
            continue
        
        matched = None
        for cat in categories:
            for kw in cat.get('keyword_triggers', []):
                if kw in text:
                    if matched is None:
                        matched = cat
                    # 如果已经匹配了一个，且这个更具体（关键词更长）
                    elif len(kw) > len(matched['keyword_triggers'][0] if matched.get('keyword_triggers') else ''):
                        matched = cat
        
        if matched:
            results.append({
                'row': i,
                'text': text,
                'category': matched['name'],
                'category_code': matched['code'],
                'confidence': matched.get('confidence', 'medium'),
                'match_type': 'keyword'
            })
            
            if matched.get('confidence') == 'low':
                low_confidence.append(results[-1])
        else:
            # 未匹配 → 默认归"其他"
            results.append({
                'row': i,
                'text': text,
                'category': '其他',
                'category_code': [c['code'] for c in categories if c['name'] == '其他'][0] if any(c['name'] == '其他' for c in categories) else 'CAT99',
                'confidence': 'low',
                'match_type': 'fallback'
            })
            low_confidence.append(results[-1])
    
    # 统计
    cat_counts = Counter(r['category'] for r in results)
    
    return {
        'step': 'classify',
        'total_rows': len(rows),
        'category_distribution': dict(cat_counts.most_common()),
        'low_confidence_count': len(low_confidence),
        'low_confidence_rate': round(len(low_confidence) / max(len(rows), 1) * 100, 1),
        'low_confidence_samples': low_confidence[:50],
        'detailed_results': results,
        'timestamp': datetime.datetime.now().isoformat()
    }

# ─── 评估分类质量 ──────────────────────────────────────
def evaluate_classification(classify_result):
    """评估分类质量并给出建议"""
    dist = classify_result['category_distribution']
    total = classify_result['total_rows']
    issues = []
    
    # 检查"其他"占比
    other_pct = dist.get('其他', 0) / total * 100 if total else 0
    if other_pct > 20:
        issues.append({
            'severity': 'high',
            'issue': f'"其他"类别占比 {other_pct:.1f}%（>20%）',
            'suggestion': '建议从"其他"中识别高频模式，拆分出 2-3 个新类别'
        })
    elif other_pct > 10:
        issues.append({
            'severity': 'medium',
            'issue': f'"其他"类别占比 {other_pct:.1f}%（>10%）',
            'suggestion': '考虑复核"其他"中的记录，确认是否有遗漏类别'
        })
    
    # 检查空值
    empty_pct = dist.get('空值', 0) / total * 100 if total else 0
    if empty_pct > 10:
        issues.append({
            'severity': 'high',
            'issue': f'空值占比 {empty_pct:.1f}%',
            'suggestion': '检查原始数据，确认空值是原始缺失还是导出问题'
        })
    
    # 低置信率
    lc_rate = classify_result.get('low_confidence_rate', 0)
    if lc_rate > 30:
        issues.append({
            'severity': 'high',
            'issue': f'低置信分类占比 {lc_rate:.1f}%',
            'suggestion': '建议审计师复核低置信样本，调整分类体系关键词后重新分类'
        })
    
    return issues

# ─── CLI / 主入口 ───────────────────────────────────────
def run_full_pipeline(rows, col_name, out_dir=None, label=''):
    """完整三步流水线: 预打标 → 构建体系 → 全量分类 → 评估"""
    print(f'\n{"="*60}')
    print(f'  🏷️ 智能分类三步法 — {label}')
    print(f'  列: {col_name} | 记录: {len(rows)} 行')
    print(f'{"="*60}')
    
    # Step 1
    print(f'\n📊 Step 1: 预打标 (归组+抽样)...')
    pretag = pre_tag(rows, col_name)
    print(f'  总记录: {pretag["total_records"]} | 去重后: {pretag["unique_records"]} | 重复率: {pretag["duplicate_rate"]}%')
    print(f'  关键词: {", ".join(kw for kw, _ in pretag["top_keywords"][:10])}')
    
    # Step 2
    print(f'\n📐 Step 2: 构建分类体系...')
    system = build_classification_template(pretag)
    print(f'  候选类别: {len(system["categories"])} 个')
    for cat in system['categories']:
        print(f'    {cat["code"]}: {cat["name"]} ({cat["estimated_coverage"]})')
    
    print(f'\n  ⚠️ 分类体系为 AI 预建草案，请审计师审阅后确认。')
    print(f'  确认清单:')
    for item in system['review_checklist']:
        print(f'    {item}')
    
    # Step 3
    print(f'\n📋 Step 3: 全量分类...')
    classify = batch_classify(rows, col_name, system)
    print(f'  分类完成: {classify["total_rows"]} 行 → {len(classify["category_distribution"])} 类')
    print(f'  分布:')
    for cat, cnt in classify['category_distribution'].items():
        pct = cnt / classify['total_rows'] * 100
        print(f'    {cat:<20} {cnt:>5} ({pct:5.1f}%)')
    
    # 评估
    issues = evaluate_classification(classify)
    if issues:
        print(f'\n  ⚠️ 质量问题:')
        for iss in issues:
            print(f'    [{iss["severity"]}] {iss["issue"]}')
            print(f'    → {iss["suggestion"]}')
    else:
        print(f'\n  ✅ 分类质量达标')
    
    # 保存
    if out_dir:
        out_path = Path(out_dir)
        out_path.mkdir(parents=True, exist_ok=True)
        pipeline = {
            'label': label,
            'column': col_name,
            'run_at': datetime.datetime.now().isoformat(),
            'pretag': pretag,
            'classification_system': system,
            'classify': classify,
            'evaluation': issues
        }
        # 保存不含详情的主报告（详细记录太大）
        report = {k: v for k, v in pipeline.items() if k != 'classify'}
        report['classify_summary'] = {
            'total': classify['total_rows'],
            'distribution': classify['category_distribution'],
            'low_confidence_count': classify['low_confidence_count'],
            'low_confidence_rate': classify['low_confidence_rate']
        }
        with open(out_path / f'{label}_classify_report.json', 'w', encoding='utf-8') as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
        # 低置信记录单独存
        if classify['low_confidence_samples']:
            with open(out_path / f'{label}_low_confidence.json', 'w', encoding='utf-8') as f:
                json.dump(classify['low_confidence_samples'], f, ensure_ascii=False, indent=2)
        # 分类明细存 CSV
        csv_path = out_path / f'{label}_classified.csv'
        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(['行号', '原文', '类别', '类别代码', '置信度', '匹配方式'])
            for r in classify['detailed_results']:
                writer.writerow([r['row'], r['text'], r['category'], r.get('category_code', ''),
                                r['confidence'], r.get('match_type', '')])
        
        print(f'\n  💾 已保存到: {out_path}')
    
    return pretag, system, classify, issues

def load_from_excel(source, columns=None):
    """从 Excel 加载数据"""
    import openpyxl
    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    rows = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        data = list(ws.iter_rows(values_only=True))
        if not data:
            continue
        headers = [str(c) if c else f'col{i}' for i, c in enumerate(data[0])]
        for row in data[1:]:
            r = {}
            for i, v in enumerate(row):
                if i < len(headers):
                    r[headers[i]] = v
            rows.append(r)
        break  # 只取第一个 sheet
    wb.close()
    return rows

if __name__ == '__main__':
    p = argparse.ArgumentParser(description='智能分类三步法 v0.1 (Skill 9)')
    p.add_argument('--source', required=True, help='Excel 数据源')
    p.add_argument('--col', required=True, help='要分类的列名')
    p.add_argument('--project', default='default', help='项目标识')
    p.add_argument('--label', default='classify', help='分类标签')
    p.add_argument('--out', help='输出目录 (默认 profiles/{project}/classify/)')
    
    args = p.parse_args()
    
    rows = load_from_excel(args.source)
    if not rows:
        print(f'❌ 未找到数据: {args.source}')
        sys.exit(1)
    
    out_dir = args.out or str(HERE / 'profiles' / args.project / 'classify')
    run_full_pipeline(rows, args.col, out_dir, args.label)
