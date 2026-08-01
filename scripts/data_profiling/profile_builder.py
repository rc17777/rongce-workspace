# -*- coding: utf-8 -*-
"""
数据理解档案构建器 v0.2
从 Excel/CSV/数据字典构建审盾数据口径档案
用法: python profile_builder.py --source "data.xlsx" --project "pidou_2026" --label "运行经费"
"""
import sys, os, json, argparse, csv, datetime
from pathlib import Path
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
PROFILES_DIR = HERE / 'profiles'
TEMPLATE = HERE / 'profile_template_v0.1.json'

# ─── Field Role Detection ───────────────────────────────
AMOUNT_KEYWORDS = ['金额', '元', '万元', '亿元', '金额(万元', '预算', '支出', '收入', '余额', '资金', '经费', '成本', '费用', '价']
QUANTITY_KEYWORDS = ['数量', '人', '次', '个', '项', '件', '台', '辆', '㎡', 'm²', '天', '小时', '张', '笔']
DATE_KEYWORDS = ['日期', '时间', '年', '月', '日', 'date', 'time', 'period']
CATEGORY_KEYWORDS = ['类型', '分类', '类别', '科目', '项目', '部门', '单位', '名称', '编号']
TEXT_KEYWORDS = ['说明', '描述', '备注', '意见', '内容', '事由', '摘要']
IDENTIFIER_KEYWORDS = ['编号', '代码', 'id', 'ID', '序号', '编码']

def detect_field_role(name, values, col_letter):
    """自动推断字段角色"""
    name_lower = name.lower()
    
    # 名称匹配优先
    for kw in IDENTIFIER_KEYWORDS:
        if kw in name:
            return 'identifier'
    for kw in AMOUNT_KEYWORDS:
        if kw in name:
            return 'amount'
    for kw in DATE_KEYWORDS:
        if kw in name:
            return 'date'
    for kw in CATEGORY_KEYWORDS:
        if kw in name:
            return 'category'
    for kw in TEXT_KEYWORDS:
        if kw in name:
            return 'text'
    for kw in QUANTITY_KEYWORDS:
        if kw in name:
            return 'quantity'
    
    # 值推断
    numeric = 0
    for v in values:
        if v is None or str(v).strip() == '':
            continue
        try:
            float(str(v).replace(',', '').replace('，', ''))
            numeric += 1
        except:
            pass
    if numeric / max(len(values), 1) > 0.8 and numeric > 3:
        return 'amount'  # >80%可转数字→金额
    
    return 'unknown'

def guess_unit(name, values):
    """推测单位"""
    if '亿元' in name:
        return '亿元'
    if '万元' in name:
        return '万元'
    if '元' in name:
        return '元'
    if '%' in name or '率' in name or '比例' in name:
        return '%'
    if '人' in name:
        return '人'
    if '次' in name:
        return '次'
    if '个' in name:
        return '个'
    # 值范围推断
    try:
        nums = [float(str(v).replace(',', '')) for v in values if v is not None and str(v).strip()]
        if nums:
            avg = sum(nums) / len(nums)
            if avg > 1e6:
                return '万元'
            elif avg > 1e3:
                return '元'
    except:
        pass
    return ''

def profile_excel(source, merge_sheets=False, **kwargs):
    """从Excel构建数据档案
    
    Args:
        source: Excel文件路径
        merge_sheets: 是否合并多Sheet（当多个Sheet结构相似时）
    """
    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    sources = []
    all_sheet_fields = {}  # 收集所有sheet的字段信息
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        
        # 检测多行头：如果第一行像标题（含"说明""用法"或只有一个非空单元格），跳过
        header_row_idx = 0
        first_row = rows[0]
        non_empty = [c for c in first_row if c]
        if len(non_empty) == 1 and len(first_row) > 3:
            # 单单元格标题行，跳过
            header_row_idx = 1
        elif any('说明' in str(c) or '用法' in str(c) for c in first_row if c):
            # 含说明/用法关键词，跳过
            header_row_idx = 1
        
        header = [str(c) if c else '' for c in rows[header_row_idx]]
        ncols = len(header)
        nrows = len(rows) - 1 - header_row_idx  # exclude header
        
        # 每列提取样本+统计
        col_values = {i: [] for i in range(ncols)}
        for row in rows[1+header_row_idx:]:
            for i, v in enumerate(row):
                if i < ncols:
                    col_values[i].append(str(v) if v is not None else None)
        
        sheet_fields = []
        for i, col_name in enumerate(header):
            vals = [v for v in col_values[i] if v is not None]
            role = detect_field_role(col_name, vals[:100], get_column_letter(i+1))
            unit = guess_unit(col_name, vals[:100])
            unique = len(set(vals[:1000])) if vals else 0
            null_pct = round((len(col_values[i]) - len(vals)) / max(len(col_values[i]), 1) * 100, 1)
            
            sheet_fields.append({
                "name": col_name,
                "col": get_column_letter(i+1),
                "role": role,
                "data_type": "number" if role in ('amount', 'quantity') else "string",
                "unit": unit if role in ('amount', 'quantity') else "",
                "nullable_pct": null_pct,
                "unique_values": unique,
                "sample": [str(v)[:50] for v in vals[:5]],
                "notes": ""
            })
        
        sources.append({
            "file_name": os.path.basename(source),
            "file_path": source,
            "format": "excel",
            "sheet_or_section": sheet_name,
            "row_count": nrows,
            "col_count": ncols
        })
        
        all_sheet_fields[sheet_name] = sheet_fields
    
    wb.close()
    
    # 合并多Sheet（如果启用）
    if merge_sheets and len(all_sheet_fields) > 1:
        # 找出所有sheet共有的字段
        common_fields = None
        for sheet_name, sheet_fields in all_sheet_fields.items():
            field_names = set(f['name'] for f in sheet_fields)
            if common_fields is None:
                common_fields = field_names
            else:
                common_fields = common_fields & field_names
        
        if common_fields and len(common_fields) > 3:
            # 有足够多共有字段，合并
            merged = {}
            for sheet_name, sheet_fields in all_sheet_fields.items():
                for f in sheet_fields:
                    if f['name'] in common_fields:
                        if f['name'] not in merged:
                            merged[f['name']] = f.copy()
                            merged[f['name']]['notes'] = f"来源: {sheet_name}"
                        else:
                            # 合并样本
                            existing = merged[f['name']]
                            existing['sample'] = list(set(existing['sample'] + f['sample']))[:5]
                            existing['unique_values'] = max(existing['unique_values'], f['unique_values'])
            fields = list(merged.values())
        else:
            # 没有足够共有字段，使用第一个sheet
            first_sheet = list(all_sheet_fields.values())[0]
            fields = first_sheet
    else:
        # 不合并，使用所有sheet的字段（去重）
        seen = set()
        fields = []
        for sheet_fields in all_sheet_fields.values():
            for f in sheet_fields:
                if f['name'] not in seen:
                    fields.append(f)
                    seen.add(f['name'])
    
    # 提取金额字段作为key_calibers候选
    amount_fields = [f for f in fields if f['role'] == 'amount']
    calibers = []
    for f in amount_fields:
        calibers.append({
            "caliber_id": f"auto_{f['col']}",
            "name": f['name'],
            "definition": f"（AI推断）字段'{f['name']}'的汇总值，需确认统计口径和计算规则",
            "formula": "",
            "unit": f['unit'],
            "confirmed": False,
            "confirmed_by": "",
            "confirmed_date": "",
            "source_field": f['name'],
            "business_context": "",
            "common_misunderstandings": [],
            "cross_references": []
        })
    
    return sources, fields, calibers

def build_profile(source, project_name, project_slug, client, year, audit_type, label, confirmed_by='', merge_sheets=False):
    """构建完整数据档案"""
    profile_id = f"{project_slug}/{label}"
    now = datetime.datetime.now().isoformat()
    
    sources, fields, calibers = profile_excel(source, merge_sheets=merge_sheets)
    
    profile = {
        "profile_id": profile_id,
        "project": {
            "name": project_name,
            "slug": project_slug,
            "client": client,
            "year": year,
            "audit_type": audit_type
        },
        "created_at": now,
        "last_confirmed_at": now if confirmed_by else "",
        "confirmed_by": confirmed_by,
        "data_sources": sources,
        "table_profile": {
            "business_object": "",
            "time_granularity": "",
            "entity_scope": "",
            "fields": fields,
            "key_relationships": []
        },
        "key_calibers": calibers,
        "quality_notes": {
            "completeness_issues": [],
            "consistency_issues": [],
            "outliers_noted": [],
            "known_biases": []
        },
        "change_log": [],
        "_ai_confidence": {
            "auto_inferred_fields": [f['name'] for f in fields if f['role'] != 'unknown'],
            "needs_human_review": [f['name'] for f in fields if f['role'] == 'unknown']
        }
    }
    
    # Save
    out_dir = PROFILES_DIR / project_slug
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{label}_profile.json"
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(profile, f, ensure_ascii=False, indent=2)
    
    print(f'✅ 数据档案已生成: {out_path}')
    print(f'   表: {len(sources)} 个 | 字段: {len(fields)} 个 | 候选口径: {len(calibers)} 个')
    if not confirmed_by:
        print(f'   ⚠️ 未确认 — {len([f for f in fields if f["role"]=="unknown"])} 个字段角色未识别，请人工标注并运行 --confirm')
    
    return profile, out_path

# ─── CLI ────────────────────────────────────────────────
if __name__ == '__main__':
    p = argparse.ArgumentParser(description='数据理解档案构建器 v0.2')
    p.add_argument('--source', required=True, help='Excel/CSV 数据源路径')
    p.add_argument('--project-name', default='', help='项目名称')
    p.add_argument('--project', required=True, help='项目标识 slug (如 pidou_2026)')
    p.add_argument('--client', default='', help='委托方')
    p.add_argument('--year', type=int, default=2026, help='年度')
    p.add_argument('--type', default='绩效评价', dest='audit_type', help='审计类型')
    p.add_argument('--label', required=True, help='数据集标签 (如 运行经费)')
    p.add_argument('--confirmed-by', default='', help='确认人')
    p.add_argument('--merge-sheets', action='store_true', help='合并多Sheet（当结构相似时）')
    
    args = p.parse_args()
    build_profile(args.source, args.project_name or args.project, args.project,
                  args.client, args.year, args.audit_type, args.label, args.confirmed_by,
                  merge_sheets=args.merge_sheets)
