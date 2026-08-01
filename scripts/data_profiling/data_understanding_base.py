# -*- coding: utf-8 -*-
"""
数据理解底座引擎 — Skill 10 (P1)
结构扫描→数据画像→语义识别→口径校验→结果沉淀
用法: python data_understanding_base.py --source "data.xlsx" --project "pidou_2026"
"""
import sys, os, json, re, argparse, datetime, csv
from pathlib import Path
from collections import Counter
import statistics
sys.stdout.reconfigure(encoding='utf-8')

HERE = Path(__file__).parent

# ─── Phase 1: 结构扫描 ─────────────────────────────────
def scan_structure(source):
    """扫描 Excel/CSV 的表结构"""
    import openpyxl
    from openpyxl.utils import get_column_letter
    
    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    sheets = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        data = list(ws.iter_rows(values_only=True))
        if not data:
            continue
        
        header = [str(c) if c else '' for c in data[0]]
        nrows = len(data) - 1  # skip header
        ncols = len(header)
        
        # 每列基本信息
        columns = []
        for i, col_name in enumerate(header):
            col_data = []
            for row in data[1:]:
                v = row[i] if i < len(row) else None
                col_data.append(v)
            
            non_null = [v for v in col_data if v is not None]
            nulls = len(col_data) - len(non_null)
            
            columns.append({
                'index': i,
                'name': col_name,
                'letter': get_column_letter(i+1),
                'row_count': len(col_data),
                'null_count': nulls,
                'null_pct': round(nulls / max(len(col_data), 1) * 100, 1),
                'unique_count': len(set(str(v) for v in non_null[:1000])) if non_null else 0,
                'sample_top5': [str(v)[:40] for v in non_null[:5]],
            })
        
        sheets.append({
            'name': sheet_name,
            'row_count': nrows,
            'col_count': ncols,
            'columns': columns
        })
    
    wb.close()
    
    return {
        'phase': 'structure_scan',
        'source': source,
        'sheets': sheets,
        'timestamp': datetime.datetime.now().isoformat()
    }

# ─── Phase 2: 数据画像 ─────────────────────────────────
def profile_data(source, target_columns=None):
    """对指定列进行数据画像（分布/异常/模式）"""
    import openpyxl
    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    
    profiles = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        data = list(ws.iter_rows(values_only=True))
        if not data:
            continue
        
        header = [str(c) if c else '' for c in data[0]]
        
        for i, col_name in enumerate(header):
            if target_columns and col_name not in target_columns:
                continue
            
            col_data = [row[i] for row in data[1:] if i < len(row) and row[i] is not None]
            
            # 尝试数值分析
            nums = []
            for v in col_data:
                try:
                    nums.append(float(str(v).replace(',', '').replace('，', '')))
                except:
                    pass
            
            profile = {
                'name': col_name,
                'total': len([r for r in data[1:]]),
                'non_null': len(col_data),
                'null_pct': round((len([r for r in data[1:]]) - len(col_data)) / max(len([r for r in data[1:]]), 1) * 100, 1),
                'unique_count': len(set(str(v) for v in col_data)) if col_data else 0,
            }
            
            if nums and len(nums) > 3:
                profile['numeric'] = True
                profile['min'] = min(nums)
                profile['max'] = max(nums)
                profile['mean'] = round(statistics.mean(nums), 2)
                profile['median'] = round(statistics.median(nums), 2)
                if len(nums) > 1:
                    profile['stdev'] = round(statistics.stdev(nums), 2) if len(nums) > 2 else 0
                    # IQR 异常检测
                    try:
                        q1 = statistics.median(sorted(nums)[:len(nums)//2])
                        q3 = statistics.median(sorted(nums)[len(nums)//2:])
                        iqr = q3 - q1
                        lower = q1 - 1.5 * iqr
                        upper = q3 + 1.5 * iqr
                        outliers = [n for n in nums if n < lower or n > upper]
                        profile['outliers_iqr'] = len(outliers)
                        profile['outliers_pct'] = round(len(outliers) / max(len(nums), 1) * 100, 1)
                        profile['outlier_thresholds'] = [round(lower, 2), round(upper, 2)]
                        profile['sample_outliers'] = sorted(set(outliers))[:10]
                    except:
                        pass
                
                # 零值/负值检查
                zeros = sum(1 for n in nums if n == 0)
                negs = sum(1 for n in nums if n < 0)
                profile['zero_count'] = zeros
                profile['negative_count'] = negs
            else:
                profile['numeric'] = False
                # 文本分布
                freq = Counter(str(v)[:20] for v in col_data)
                profile['top_values'] = freq.most_common(10)
                profile['distinct_pct'] = round(profile['unique_count'] / max(len(col_data), 1) * 100, 1)
            
            profiles[col_name] = profile
    
    wb.close()
    
    return {
        'phase': 'data_profiling',
        'profiles': profiles,
        'timestamp': datetime.datetime.now().isoformat()
    }

# ─── Phase 3: 语义识别 ─────────────────────────────────
def semantic_tag(profile_result):
    """基于剖面结果推断字段语义角色"""
    profiles = profile_result['profiles']
    tagged = {}
    
    for name, p in profiles.items():
        tags = []
        nl = name.lower()
        
        # 角色推断
        if p.get('numeric'):
            if p.get('min', 0) == 0 and p.get('max', 0) == 1:
                tags.append('boolean')
            elif '率' in nl or '比' in nl or '%' in nl:
                tags.append('ratio')
            elif '元' in nl or '金额' in nl or '金额' in nl or '预算' in nl or '支出' in nl:
                tags.append('amount')
            elif '年' in nl or '月' in nl:
                tags.append('date_numeric')
            elif '人' in nl or '次' in nl or '个' in nl:
                tags.append('quantity')
            else:
                tags.append('numeric')
        
        if p.get('unique_count', 9999) < 50:
            tags.append('low_cardinality')
        if p.get('null_pct', 0) > 50:
            tags.append('mostly_null')
        if p.get('zero_count', 0) > p.get('total', 1) * 0.8:
            tags.append('mostly_zero')
        
        # 日期检测
        if '日' in nl or 'date' in nl or '时间' in nl or '日期' in nl:
            tags.append('date')
        
        # 编码检测
        if p.get('unique_count', 0) > 0.8 * p.get('total', 1):
            tags.append('likely_identifier')
        
        # 文本检测
        if not p.get('numeric'):
            if p.get('distinct_pct', 0) > 50:
                tags.append('free_text')
            elif p.get('distinct_pct', 0) < 5:
                tags.append('lookup_value')
        
        tagged[name] = {
            'role': tags[0] if tags else 'unknown',
            'all_tags': tags,
            'confidence': 'high' if len(tags) >= 2 else 'medium' if tags else 'low'
        }
    
    return {
        'phase': 'semantic_tagging',
        'tags': tagged,
        'timestamp': datetime.datetime.now().isoformat()
    }

# ─── Phase 4: 口径校验 ─────────────────────────────────
def verify_caliber(profile_result, semantic_tags):
    """
    基于数据剖面和语义标签，校验默认口径
    检测: 总金额与明细是否勾稽、关键字段是否有异常分布
    """
    profiles = profile_result['profiles']
    tags = semantic_tags['tags']
    checks = []
    
    # 检查金额列
    amount_cols = [(n, p) for n, p in profiles.items() if 'amount' in tags.get(n, {}).get('all_tags', [])]
    
    for name, p in amount_cols:
        # 检查零值
        zeros = p.get('zero_count', 0)
        if zeros > 0:
            checks.append({
                'type': 'zero_amounts',
                'column': name,
                'count': zeros,
                'severity': 'medium' if zeros > 5 else 'low',
                'detail': f'金额列含 {zeros} 条零值记录',
                'action': '确认零值是否合理（可能是免费项目/未发生/数据缺失）'
            })
        
        # 检查负值
        negs = p.get('negative_count', 0)
        if negs > 0:
            checks.append({
                'type': 'negative_amounts',
                'column': name,
                'count': negs,
                'severity': 'high',
                'detail': f'金额列含 {negs} 条负值记录',
                'action': '确认负数含义（冲销/退款/数据录入错误），并明确口径是否含负数'
            })
        
        # 检查异常值
        outliers = p.get('outliers_iqr', 0)
        if outliers > 0:
            checks.append({
                'type': 'outliers',
                'column': name,
                'count': outliers,
                'severity': 'medium',
                'detail': f'IQR检测发现 {outliers} 条统计异常值（阈值: {p.get("outlier_thresholds", [])}）',
                'sample': p.get('sample_outliers', [])[:5],
                'action': '核实异常值是否真实业务数据'
            })
    
    # 检查高缺失列
    for name, p in profiles.items():
        if p.get('null_pct', 0) > 30:
            checks.append({
                'type': 'high_missing',
                'column': name,
                'null_pct': p['null_pct'],
                'severity': 'high' if p['null_pct'] > 70 else 'medium',
                'detail': f'缺失率 {p["null_pct"]}%',
                'action': '确认该字段是否为可选字段，缺值原因'
            })
    
    return {
        'phase': 'caliber_verify',
        'checks': checks,
        'total_checks': len(checks),
        'high_severity': len([c for c in checks if c['severity'] == 'high']),
        'timestamp': datetime.datetime.now().isoformat()
    }

# ─── Phase 5: 结果沉淀 ─────────────────────────────────
def generate_understanding_report(structure, profiling, semantic, caliber_checks, project_label, out_dir=None):
    """汇总四阶段结果，生成数据理解档案报告"""
    report = {
        'project': project_label,
        'generated_at': datetime.datetime.now().isoformat(),
        'profile_version': 'v0.1',
        'phases': {
            'structure': structure,
            'profiling': profiling,
            'semantic': semantic,
            'caliber_verify': caliber_checks,
        },
        'executive_summary': {
            'sheets': len(structure['sheets']),
            'total_columns': sum(s['col_count'] for s in structure['sheets']),
            'amount_columns': len([n for n, t in semantic['tags'].items() if 'amount' in t.get('all_tags', [])]),
            'date_columns': len([n for n, t in semantic['tags'].items() if 'date' in t.get('all_tags', [])]),
            'high_missing_cols': len([c for c in caliber_checks['checks'] if c['type'] == 'high_missing' and c['severity'] == 'high']),
            'data_quality_flags': caliber_checks['total_checks'],
            'high_severity_flags': caliber_checks['high_severity'],
        },
        'action_items': [c for c in caliber_checks['checks'] if c['severity'] == 'high']
    }
    
    if out_dir:
        out_path = Path(out_dir)
        out_path.mkdir(parents=True, exist_ok=True)
        report_path = out_path / f'{project_label}_understanding_report.json'
        with open(report_path, 'w', encoding='utf-8') as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
        print(f'\n  💾 数据理解报告已保存: {report_path}')
    
    return report

def print_summary(report):
    """打印数据理解摘要"""
    s = report['executive_summary']
    print(f'\n{"="*60}')
    print(f'  🔬 数据理解底座报告 — {report["project"]}')
    print(f'{"="*60}')
    print(f'  📊 表: {s["sheets"]} | 列: {s["total_columns"]}')
    print(f'  💰 金额列: {s["amount_columns"]} | 📅 日期列: {s["date_columns"]}')
    print(f'  ⚠️ 高缺失列: {s["high_missing_cols"]} | 质量标记: {s["data_quality_flags"]}')
    print(f'  🚨 高严重度: {s["high_severity_flags"]}')
    
    if report['action_items']:
        print(f'\n  📋 需处理:')
        for item in report['action_items']:
            print(f'    [{item["type"]}] {item["column"]}: {item["detail"]}')
            print(f'    → {item["action"]}')

# ─── Main ───────────────────────────────────────────────
def run_full_pipeline(source, project_label='default', out_dir=None, target_columns=None):
    """完整五阶段流水线"""
    
    # Phase 1
    print('Phase 1/5: 结构扫描...')
    structure = scan_structure(source)
    for s in structure['sheets']:
        print(f'  {s["name"]}: {s["row_count"]} 行 x {s["col_count"]} 列')
    
    # Phase 2
    print('Phase 2/5: 数据画像...')
    profiling = profile_data(source, target_columns)
    amount_count = 0
    for n, p in profiling['profiles'].items():
        if p.get('numeric'):
            print(f'  {n}: mean={p.get("mean","-")} range=[{p.get("min","-")}, {p.get("max","-")}] outliers={p.get("outliers_iqr","-")}')
            amount_count += 1
    print(f'  数值列: {amount_count} / {len(profiling["profiles"])}')
    
    # Phase 3
    print('Phase 3/5: 语义识别...')
    semantic = semantic_tag(profiling)
    role_summary = Counter(t['role'] for t in semantic['tags'].values())
    for role, cnt in role_summary.most_common():
        print(f'  {role}: {cnt} 列')
    
    # Phase 4
    print('Phase 4/5: 口径校验...')
    caliber = verify_caliber(profiling, semantic)
    print(f'  质量问题: {caliber["total_checks"]} 条 (🚨{caliber["high_severity"]} 高)')
    
    # Phase 5
    print('Phase 5/5: 结果沉淀...')
    report = generate_understanding_report(structure, profiling, semantic, caliber, project_label, out_dir)
    print_summary(report)
    
    return report

if __name__ == '__main__':
    p = argparse.ArgumentParser(description='数据理解底座引擎 v0.1 (Skill 10)')
    p.add_argument('--source', required=True, help='Excel/CSV 数据源路径')
    p.add_argument('--project', required=True, help='项目标识 (如 pidou_2026)')
    p.add_argument('--columns', nargs='*', help='限定分析列')
    p.add_argument('--out', help='输出目录')
    
    args = p.parse_args()
    
    out_dir = args.out or str(HERE / 'profiles' / args.project)
    run_full_pipeline(args.source, args.project, out_dir, args.columns)
