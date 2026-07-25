#!/usr/bin/env python3
"""Generate 审盾 v2.0 Excel workbook with all sheets."""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

wb = Workbook()
DARK = "0A1F3F"
TEAL = "1A5C6E"
GOLD = "C5955C"
RED = "CC0000"
GREEN = "228B22"
GREY = "666666"
LIGHT_GREY = "F5F2EC"

hdr_fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
hdr_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
title_font = Font(name="微软雅黑", bold=True, size=14, color=DARK)
section_font = Font(name="微软雅黑", bold=True, size=12, color=TEAL)
normal = Font(name="微软雅黑", size=10)
bold_n = Font(name="微软雅黑", size=10, bold=True)
red_f = Font(name="微软雅黑", size=10, color=RED)
green_f = Font(name="微软雅黑", size=10, color=GREEN)
gold_bold = Font(name="微软雅黑", size=10, bold=True, color=GOLD)
grey_f = Font(name="微软雅黑", size=10, color=GREY)
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)
wrap_align = Alignment(wrap_text=True, vertical='top')

def set_cell(ws, cell, value, font=None, fill=None, alignment=None, border=None):
    c = ws[cell]
    c.value = value
    if font: c.font = font
    if fill: c.fill = fill
    if alignment: c.alignment = alignment
    if border: c.border = border
    return c

def merge_and_set(ws, range_str, value, font=None, fill=None):
    ws.merge_cells(range_str)
    set_cell(ws, range_str.split(':')[0], value, font, fill, wrap_align)

def header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        set_cell(ws, f"{get_column_letter(start_col+i)}{row}", h, hdr_font, hdr_fill, wrap_align, thin_border)

# ==================== Sheet 1: 总览 ====================
ws = wb.active
ws.title = "总览"
ws.sheet_properties.tabColor = DARK

merge_and_set(ws, "A1:H1", "融策·审盾 三位一体AI审计中台建设方案 v2.0", title_font)
merge_and_set(ws, "A2:H2", "修订：2026-07-21 | 四模型联合评审 + 苏格拉底追问后修订 | 负责人：融策平头哥", grey_f)

set_cell(ws, "A4", "▎核心定位", section_font)
merge_and_set(ws, "A5:H5", "融策·审盾 —— 用AI武装的政府财政资金安全与绩效智能守门人", Font(name="微软雅黑", size=11))
merge_and_set(ws, "A6:H6", "一期唯一用户：审计项目经理（B类用户——报告复核） | 实战场景：若尔盖审计局 校园餐+医保资金审计", grey_f)

r = 8
set_cell(ws, f"A{r}", "▎v1.0 → v2.0 关键变更（四模型联合评审驱动）", section_font)
r += 1
header_row(ws, r, ["修正点", "v1.0", "v2.0", "出发来源"])
changes = [
    ("目标用户", "模糊（三类人）", "单一用户：审计项目经理（B）", "苏格拉底Q3"),
    ("技术路线", "微调13B-32B模型", "RAG + Agentic Workflow为主", "Gemini + GPT-5.5力挺"),
    ("硬件路线", "RTX 5090 ×1（Day 1）", "现有工作站+API，GPU推迟至Day90", "Luna + GPT-5.5"),
    ("验证数据", "融策内部报告", "10份不同客户/年份/格式", "苏格拉底Q9"),
    ("验证节奏", "30天出盲测", "30天规则引擎 + 60天语义复核", "Claude + GPT-5.5 + Luna"),
    ("量化标准", "无", "检出率≥80% 误报≤20% 采纳率≥40%", "Claude + GPT-5.5"),
    ("止损机制", "无", "Day180不达标→不进入Phase 2", "Claude + GPT-5.5"),
    ("SaaS化", "多租户SaaS", "三级路径：溢价→联合体→私有化部署", "Claude + GPT-5.5 + Luna"),
    ("数据飞轮", "Phase 2才建标注平台", "Day 1启动Label Studio", "四模型全票"),
    ("团队配置", "Phase1招1人AI工程师", "Day1-90平头哥亲自+OpenClaw", "苏格拉底Q5"),
    ("业务线扩展", "Phase2做5条线", "Phase2深度打磨3条线", "Luna + GPT-5.5"),
    ("三年总投入", "¥720万", "¥325-431万（↓40-55%）", "Gemini + Claude"),
]
for a, b, c, d in changes:
    r += 1
    set_cell(ws, f"A{r}", a, normal, None, wrap_align, thin_border)
    set_cell(ws, f"B{r}", b, grey_f, None, wrap_align, thin_border)
    set_cell(ws, f"C{r}", c, Font(name="微软雅黑", size=10, color=TEAL), None, wrap_align, thin_border)
    set_cell(ws, f"D{r}", d, normal, None, wrap_align, thin_border)

r += 2
set_cell(ws, f"A{r}", "▎三年投入速查", section_font)
r += 1
header_row(ws, r, ["阶段", "时间", "一次性投入", "年运营成本", "累计投入"])
budgets = [
    ("第一阶段·验证闭环", "月1-6", "¥0-3万", "¥15-18万", "¥15-21万"),
    ("第二阶段·规模化打磨", "月7-18", "¥15-25万", "¥105-115万", "¥120-140万"),
    ("第三阶段·行业壁垒", "月19-36", "¥40-75万", "¥150-195万", "¥190-270万"),
    ("三年合计", "36个月", "¥55-103万", "¥270-328万", "¥325-431万"),
]
for a, b, c, d, e in budgets:
    r += 1
    is_total = "合计" in a
    fnt = bold_n if is_total else normal
    set_cell(ws, f"A{r}", a, fnt, None, wrap_align, thin_border)
    set_cell(ws, f"B{r}", b, normal, None, wrap_align, thin_border)
    set_cell(ws, f"C{r}", c, normal, None, wrap_align, thin_border)
    set_cell(ws, f"D{r}", d, normal, None, wrap_align, thin_border)
    set_cell(ws, f"E{r}", e, gold_bold if is_total else normal, None, wrap_align, thin_border)

ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 38
ws.column_dimensions['D'].width = 26
ws.column_dimensions['E'].width = 16

# ==================== Sheet 2: 一阶段 ====================
ws = wb.create_sheet("一阶段-验证闭环")
ws.sheet_properties.tabColor = TEAL

merge_and_set(ws, "A1:F1", "第一阶段：验证闭环（2026年7月-12月）", title_font)
merge_and_set(ws, "A2:F2", "唯一目标：证明AI复核在政府审计报告场景中的可行性与业务价值", normal)
merge_and_set(ws, "A4:F4", "❌ 本阶段不做（写死在墙上）：不做投标方案 | 不做质控汇总 | 不做其他业务线Agent | 不做SaaS | 不做微调 | 不买新GPU（Day90决策）", red_f)

r = 6
set_cell(ws, f"A{r}", "阶段1A：零阶段·准备工作（Day 1-14）¥0投入", section_font)
r += 1
header_row(ws, r, ["天数", "任务", "产出物"])
tasks_1a = [
    ("Day 1-2", "收集校园餐+医保全部原始资料；合规三问调研", "📁 原始资料包 + 📝 合规调研笔记"),
    ("Day 3-4", "手工标注第1份校园餐报告，记录所有人工检查点", "📝 人工复核Checklist v1.0"),
    ("Day 5-6", "同一份报告丢OpenClaw跑AI复核→逐条对比人工", "🔀 第1份人机对比表"),
    ("Day 7-8", "换医保资金报告，重复对比", "🔀 第2份人机对比表"),
    ("Day 9-10", "收集第3-5份不同来源/年份报告，跑AI复核", "📊 5份原始AI复核结果"),
    ("Day 11-12", "整理首周发现：采纳率+失效模式+脏数据清单", "📄 第1周验证周报"),
    ("Day 13", "定稿量化通过标准（书面确认）", "✅ 标准文档"),
    ("Day 14", "输出《政府审计AI辅助能力白皮书》初稿", "📘 白皮书v1.0"),
]
for a, b, c in tasks_1a:
    r += 1
    set_cell(ws, f"A{r}", a, bold_n, None, wrap_align, thin_border)
    set_cell(ws, f"B{r}", b, normal, None, wrap_align, thin_border)
    set_cell(ws, f"C{r}", c, normal, None, wrap_align, thin_border)

r += 2
set_cell(ws, f"A{r}", "阶段1B：规则引擎验证（Day 15-60）¥0-1万（API费）", section_font)
r += 1
header_row(ws, r, ["维度", "内容", "说明"])
p1b = [
    ("验证范围", "格式检查+合计校验+法规引用完整性+附表交叉比对", "只做规则层面，不碰语义判断"),
    ("测试报告", "10份（5份校园餐+5份医保），≥3个不同客户来源", "必须不同年份、不同格式"),
    ("通过标准", "规则引擎准确率≥95%，误报率≤10%", "未达标→延长至Day 90重测"),
    ("数据积累", "每条人工修正diff存入Label Studio", "每天至少攒10条标注"),
    ("Prompt管理", "建立评估表：每次改prompt记录预期vs实际结果", "防止调参死循环"),
]
for a, b, c in p1b:
    r += 1
    set_cell(ws, f"A{r}", a, bold_n, None, wrap_align, thin_border)
    set_cell(ws, f"B{r}", b, normal, None, wrap_align, thin_border)
    set_cell(ws, f"C{r}", c, normal, None, wrap_align, thin_border)

r += 2
set_cell(ws, f"A{r}", "阶段1C：语义复核验证（Day 61-90）¥1-3万", section_font)
r += 1
p1c = [
    ("验证范围", "语义层面：结论与证据链一致性+法规适用准确性+指标合理性"),
    ("方法", "RAG增强：每条AI意见附引用源chunk（溯源链=信任基础）"),
    ("通过标准", "①关键错误检出率≥人工80% ②误报率≤20% ③采纳率≥40%（三指标同时达标）"),
    ("盲测设计", "3份未训练报告，AI vs 平头哥独立复核，盲测对比"),
    ("API容灾", "确认Qwen/豆包API可用，完成同一报告跨模型对比"),
    ("Day90决策门", "达标→招第1个AI工程师 | 部分达标→延长30天 | 严重不达标→降级"),
]
for a, b in p1c:
    r += 1
    set_cell(ws, f"A{r}", a, bold_n, None, wrap_align, thin_border)
    ws.merge_cells(f"B{r}:F{r}")
    set_cell(ws, f"B{r}", b, normal, None, wrap_align, thin_border)

r += 2
set_cell(ws, f"A{r}", "阶段1D：工程师入场+实战交付（Day 91-180）¥10-15万", section_font)
r += 1
p1d = [
    ("招人画像", "Python+LangChain+向量数据库+FastAPI，月薪1.5-2万(成都)，不要求微调/K8s/前端"),
    ("固化为Web界面", "审计师可自行上传报告→收到AI复核结果（简陋版，能用就行）"),
    ("扩至3条业务线", "校园餐+医保资金+绩效评价"),
    ("标注数据≥1000条", "为未来微调打底"),
    ("若尔盖项目AI辅助交付", "审计师自愿使用，不强推——内部口碑种子"),
    ("Day180决策", "达标(3线效率+30%,≥10人周用≥3次)→Phase2 | 不达标→维持1人+API低成本运作"),
]
for a, b in p1d:
    r += 1
    set_cell(ws, f"A{r}", a, bold_n, None, wrap_align, thin_border)
    ws.merge_cells(f"B{r}:F{r}")
    set_cell(ws, f"B{r}", b, normal, None, wrap_align, thin_border)

ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 35

# ==================== Sheet 3: 二阶段 ====================
ws = wb.create_sheet("二阶段-规模化打磨")
ws.sheet_properties.tabColor = "4472C4"

merge_and_set(ws, "A1:E1", "第二阶段：规模化打磨（2027年1月-12月）", title_font)
merge_and_set(ws, "A2:E2", "进入条件：Phase 1 Day 180三指标全部达标 + 3条业务线验证通过", Font(name="微软雅黑", size=10, color=RED))
merge_and_set(ws, "A3:E3", "目标：AI复核从'能跑'变成'离不开'，核心团队日常使用，对外形成品牌认知", normal)

set_cell(ws, "A5", "▎阶段2A：基础设施补强（月1-3）", section_font)
r = 6
header_row(ws, r, ["配置", "说明", "投入"])
items_2a = [
    ("GPU工作站×1", "RTX 4090 24GB或等5090，本地推理敏感底稿", "¥5-7万"),
    ("Qdrant+自动更新", "每周增量重建索引，法规更新自动触发", "¥0（自研）"),
    ("AI合规框架", "输出分级(L1/L2/L3)+底稿AI水印+溯源链", "¥1-2万"),
    ("双路由API", "DeepSeek主+Qwen/豆包备，自动切换", "¥0.5万"),
    ("20TB NAS备份", "项目数据统一存储", "¥1-2万"),
]
for a, b, c in items_2a:
    r += 1
    set_cell(ws, f"A{r}", a, bold_n, None, wrap_align, thin_border)
    set_cell(ws, f"B{r}", b, normal, None, wrap_align, thin_border)
    set_cell(ws, f"C{r}", c, normal, None, wrap_align, thin_border)

r += 2
set_cell(ws, f"A{r}", "▎阶段2B：3条业务线深度打磨（月4-9）", section_font)
r += 1
header_row(ws, r, ["业务线", "核心功能", "验收标准"])
lines_2b = [
    ("绩效评价报告复核", "15维全量检查+语义推理", "采纳率≥50%，误报率≤15%"),
    ("专项资金审计复核", "校园餐/医保等专项检查", "采纳率≥45%，误报率≤20%"),
    ("经责审计底稿辅助", "底稿自动生成+交叉复核", "底稿初稿生成时间缩减≥40%"),
]
for a, b, c in lines_2b:
    r += 1
    set_cell(ws, f"A{r}", a, bold_n, None, wrap_align, thin_border)
    set_cell(ws, f"B{r}", b, normal, None, wrap_align, thin_border)
    set_cell(ws, f"C{r}", c, normal, None, wrap_align, thin_border)

r += 2
set_cell(ws, f"A{r}", "▎阶段2C：品牌化+对外试水（月10-12）", section_font)
r += 1
header_row(ws, r, ["动作", "内容", "投入"])
items_2c = [
    ("客户看板", "内部案例数据可视化，投标展示用", "¥2-3万"),
    ("白皮书v4.0", "端到端案例+量化数据+行业对比，印刷版+电子版", "¥0.5万"),
    ("私测输出", "1家友好合作事务所私有化部署试点", "¥1-2万"),
    ("行业会议", "1-2个四川省内审计/财政会议，案例分享", "¥1-2万"),
]
for a, b, c in items_2c:
    r += 1
    set_cell(ws, f"A{r}", a, bold_n, None, wrap_align, thin_border)
    set_cell(ws, f"B{r}", b, normal, None, wrap_align, thin_border)
    set_cell(ws, f"C{r}", c, normal, None, wrap_align, thin_border)

r += 2
set_cell(ws, f"A{r}", "▎团队配置 & 总预算", section_font)
r += 1
header_row(ws, r, ["岗位", "人数", "月薪", "年成本"])
team_2 = [
    ("AI全栈工程师(Phase1延续)", "1人", "¥2.0-2.5万", "¥30万"),
    ("数据工程师(新增)", "1人", "¥1.5-2.0万", "¥24万"),
    ("审计业务专家(兼职)", "3人", "各¥0.5-1万补贴", "¥24万"),
    ("合计", "", "", "¥120-140万/年（含基础设施+应用开发）"),
]
for a, b, c, d in team_2:
    r += 1
    set_cell(ws, f"A{r}", a, bold_n if "合计" in a else normal, None, wrap_align, thin_border)
    set_cell(ws, f"B{r}", b, normal, None, wrap_align, thin_border)
    set_cell(ws, f"C{r}", c, normal, None, wrap_align, thin_border)
    set_cell(ws, f"D{r}", d, gold_bold if "合计" in a else normal, None, wrap_align, thin_border)

ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 42
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 18

# ==================== Sheet 4: 三阶段 ====================
ws = wb.create_sheet("三阶段-行业壁垒")
ws.sheet_properties.tabColor = "2C5F2D"

merge_and_set(ws, "A1:D1", "第三阶段：行业壁垒（2028年1月-12月）", title_font)
merge_and_set(ws, "A2:D2", "进入条件：≥3条线采纳率≥50% + ≥20人周用≥5次 + 私测正向 + 标注≥5000条", Font(name="微软雅黑", size=10, color=RED))

set_cell(ws, "A4", "▎阶段3A：数据飞轮加速", section_font)
r = 5
header_row(ws, r, ["动作", "说明", "投入"])
items_3a = [
    ("历史项目脱敏数据集", "融策10年+政府审计项目脱敏，12条业务线标注", "¥15-25万"),
    ("多模态引擎", "PaddleOCR+工程图纸识别+合同字段提取", "¥8-12万"),
    ("知识库自动更新", "法规变动自动触发RAG重建+检测过期chunk", "¥3-5万"),
    ("外部数据合规矩阵", "工商+司法+招标+税务黑名单+舆情", "¥5-8万/年"),
]
for a, b, c in items_3a:
    r += 1
    set_cell(ws, f"A{r}", a, bold_n, None, wrap_align, thin_border)
    set_cell(ws, f"B{r}", b, normal, None, wrap_align, thin_border)
    set_cell(ws, f"C{r}", c, normal, None, wrap_align, thin_border)

r += 2
set_cell(ws, f"A{r}", "▎阶段3B：技术栈升级（条件触发：标注≥5000+采纳率≥60%）", section_font)
r += 1
header_row(ws, r, ["动作", "说明", "投入"])
items_3b = [
    ("追加GPU算力", "RTX 6000 Ada 48GB 或云GPU弹性资源（QLoRA微调7B）", "¥8-15万"),
    ("RAG+微调混合架构", "通用知识走RAG，审计推理走微调模型", "¥3-5万"),
    ("DevOps基础设施", "CI/CD+A/B测试+监控+模型版本管理", "¥3-5万"),
]
for a, b, c in items_3b:
    r += 1
    set_cell(ws, f"A{r}", a, bold_n, None, wrap_align, thin_border)
    set_cell(ws, f"B{r}", b, normal, None, wrap_align, thin_border)
    set_cell(ws, f"C{r}", c, normal, None, wrap_align, thin_border)

r += 2
set_cell(ws, f"A{r}", "▎阶段3C：对外输出（三级递进，不做多租户SaaS）", section_font)
r += 1
header_row(ws, r, ["级别", "模式", "投入"])
outputs = [
    ("Level 1：能力溢价", "投标时以AI辅助审计为差异化卖点，不单独售卖", "¥0"),
    ("Level 2：联合体模式", "与合作事务所组成联合体投标，分成", "¥2-5万"),
    ("Level 3：私有化部署", "¥5-10万/次部署+¥2-3万/年维护，目标3-5家", "¥10-15万"),
]
for a, b, c in outputs:
    r += 1
    set_cell(ws, f"A{r}", a, bold_n, None, wrap_align, thin_border)
    set_cell(ws, f"B{r}", b, normal, None, wrap_align, thin_border)
    set_cell(ws, f"C{r}", c, normal, None, wrap_align, thin_border)

r += 2
set_cell(ws, f"A{r}", "▎团队配置 & 总预算", section_font)
r += 1
header_row(ws, r, ["岗位", "人数", "月薪", "年成本"])
team_3 = [
    ("AI全栈工程师", "2人", "¥2.5-3.5万", ""),
    ("数据工程师", "1人", "¥1.8-2.5万", ""),
    ("审计AI产品经理", "1人", "¥2.0-3.0万", ""),
    ("审计业务专家(兼职)", "5人", "各¥0.5-1万补贴", ""),
    ("合计", "5-7人", "", "¥190-270万/年（含基础设施+应用+输出）"),
]
for a, b, c, d in team_3:
    r += 1
    set_cell(ws, f"A{r}", a, bold_n if "合计" in a else normal, None, wrap_align, thin_border)
    set_cell(ws, f"B{r}", b, normal, None, wrap_align, thin_border)
    set_cell(ws, f"C{r}", c, normal, None, wrap_align, thin_border)
    set_cell(ws, f"D{r}", d, gold_bold if "合计" in a else normal, None, wrap_align, thin_border)

ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 42
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 22

# ==================== Sheet 5: 预算对比 ====================
ws = wb.create_sheet("预算对比")
ws.sheet_properties.tabColor = GOLD

merge_and_set(ws, "A1:F1", "三阶段预算对比：v1.0 → v2.0", title_font)
merge_and_set(ws, "A2:F2", "v2.0省钱逻辑：砍GPU集群→API | 砍微调→RAG | 砍SaaS→私有化部署 | 砍5线→3线", grey_f)

r = 4
header_row(ws, r, ["项目", "v1.0预算", "v2.0修正", "变化", "修正原因"])
items = [
    ("【第一阶段】", "", "", "", ""),
    ("GPU工作站", "¥8-10万", "¥0（推迟至Day90）", "↓100%", "Luna：验证期API够用"),
    ("AI工程师(前3月)", "¥12-18万/年", "¥0（平头哥+OpenClaw）", "↓100%", "苏格拉底Q5"),
    ("AI工程师(后3月)", "—", "¥6-8万(Day90起)", "新增", "月薪¥1.8-2.4万含社保"),
    ("数据平台", "¥3万/年", "¥3万/年", "不变", ""),
    ("API费用", "¥3-5万/年", "¥5-8万/年", "↑60%", "复核token消耗大"),
    ("合规+法务", "¥0", "¥1-2万/年", "新增", "Gemini：Day1建立"),
    ("一阶段小计", "¥40万", "¥15-21万", "↓50%", ""),
    ("", "", "", "", ""),
    ("【第二阶段】", "", "", "", ""),
    ("GPU服务器", "¥20-25万", "¥5-7万(4090工作站)", "↓75%", "不做微调"),
    ("NAS+网络", "含上述", "¥1-2万", "", ""),
    ("数据+标注", "¥13-20万/年", "¥10-15万/年", "↓25%", ""),
    ("应用开发", "¥23-37万/年", "¥20-28万/年", "↓25%", "3线非5线"),
    ("人力", "¥60-80万/年", "¥78万/年", "持平", ""),
    ("品牌+输出", "¥0", "¥5-8万", "新增", ""),
    ("二阶段小计", "¥180万", "¥120-140万", "↓25%", ""),
    ("", "", "", "", ""),
    ("【第三阶段】", "", "", "", ""),
    ("算力集群", "¥100万+", "¥8-15万", "↓90%", "不跑K8s"),
    ("数据+多模态", "含上述", "¥31-50万", "", ""),
    ("应用开发", "含上述", "¥16-25万", "", ""),
    ("SaaS/输出", "含上述", "¥12-20万", "", "私有化部署"),
    ("人力(5-6人)", "含上述", "¥120-160万", "", ""),
    ("三阶段小计", "¥500万+", "¥190-270万", "↓50%", ""),
    ("", "", "", "", ""),
    ("【三年总计】", "¥720万", "¥325-431万", "↓40-55%", ""),
]
for a, b, c, d, e in items:
    r += 1
    is_header = a.startswith("【")
    is_total = "小计" in a or "总计" in a
    set_cell(ws, f"A{r}", a, Font(name="微软雅黑", size=10, bold=is_header or is_total, color=TEAL if is_header else "000000"), None, wrap_align, thin_border)
    set_cell(ws, f"B{r}", b, normal if not is_total else bold_n, None, wrap_align, thin_border)
    set_cell(ws, f"C{r}", c, Font(name="微软雅黑", size=10, bold=is_total, color=GOLD if is_total else "000000"), None, wrap_align, thin_border)
    df = d
    is_down = "↓" in str(d)
    is_up = "↑" in str(d)
    d_color = RED if is_down else GREEN if is_up else None
    set_cell(ws, f"D{r}", d, Font(name="微软雅黑", size=10, color=d_color), None, wrap_align, thin_border)
    set_cell(ws, f"E{r}", e, normal, None, wrap_align, thin_border)

ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 24
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 30

# ==================== Sheet 6: 风险清单 ====================
ws = wb.create_sheet("风险清单")
ws.sheet_properties.tabColor = RED

merge_and_set(ws, "A1:H1", "Top 10 风险清单（按概率×影响排序）", title_font)
merge_and_set(ws, "A2:H2", "每条风险附带：最小验证实验（7天内完成）+ 通过标准 + 所需资源", grey_f)

r = 4
header_row(ws, r, ["#", "风险", "概率", "影响", "等级", "最小验证实验(7天)", "通过标准", "所需资源"])
risk_data = [
    ("1", "你自己放弃——旺季挤压+多线作战", "高", "毁灭", "🔴🔴🔴", "连续7天每天30min打卡", "7天不间断,缺一天标记🚨", "0元"),
    ("2", "AI复核结果不被接受(采纳率<30%)", "高", "高", "🔴🔴🔴", "1份报告AI复核→逐条标注", "≥10条发现+失效分类", "0元"),
    ("3", "被格式多样性打垮(换客户就崩)", "高", "高", "🔴🔴🔴", "≥3个不同来源报告各跑一遍", "记录每种格式失效模式", "0元"),
    ("4", "无量化基线即推进", "中", "高", "🔴🔴", "定下通过标准并书面确认", "标准已写在方案里", "0元"),
    ("5", "合规信息黑洞(等保一问三不知)", "中", "高", "🔴🔴", "查等保费用/周期+数据出本地", "拿到3个明确答案", "30分钟"),
    ("6", "原始数据质量极差", "中", "中", "🟡", "列10个数据质量问题", "10条具体脏数据清单", "0元"),
    ("7", "Prompt调优陷入死循环", "中", "中", "🟡", "每次改prompt记录预期vs实际", "≥2轮有记录的对比", "0元"),
    ("8", "范围蔓延(又想加功能)", "低", "高", "🟡", "写下\"审盾一期不做X\"贴桌上", "一句话已完成", "0元"),
    ("9", "DeepSeek API突然宕机", "低", "中", "🟡", "确认备援模型可用", "备援API已就绪", "已有key"),
    ("10", "客户知道AI参与后要求降价", "低", "低", "🟢", "暂不验证,标记为观察项", "—", "—"),
]
for n, risk, prob, impact, level, exp, std, res in risk_data:
    r += 1
    set_cell(ws, f"A{r}", n, normal, None, wrap_align, thin_border)
    set_cell(ws, f"B{r}", risk, normal, None, wrap_align, thin_border)
    set_cell(ws, f"C{r}", prob, normal, None, wrap_align, thin_border)
    set_cell(ws, f"D{r}", impact, normal, None, wrap_align, thin_border)
    set_cell(ws, f"E{r}", level, Font(name="微软雅黑", size=14), None, wrap_align, thin_border)
    set_cell(ws, f"F{r}", exp, normal, None, wrap_align, thin_border)
    set_cell(ws, f"G{r}", std, normal, None, wrap_align, thin_border)
    set_cell(ws, f"H{r}", res, normal, None, wrap_align, thin_border)

ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 6
ws.column_dimensions['D'].width = 6
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 32
ws.column_dimensions['G'].width = 24
ws.column_dimensions['H'].width = 12

# ==================== Sheet 7: 第1周计划 ====================
ws = wb.create_sheet("第1周计划")
ws.sheet_properties.tabColor = "FF6600"

merge_and_set(ws, "A1:F1", "审盾一期·第1周行动计划（2026年7月22日-28日）", title_font)
merge_and_set(ws, "A2:F2", "总耗时：约8小时 | 总预算：¥0 | 不需要新硬件/新人员", grey_f)
merge_and_set(ws, "A3:F3", "每日硬性规则：早8点发\"审盾 Day X，今天做____\"，晚8点发\"审盾 Day X，完成了____\"", red_f)

r = 5
header_row(ws, r, ["星期", "日期", "任务", "产出物", "耗时", "完成✓"])
plan = [
    ("周二", "7/22", "①收集校园餐+医保全部原始资料 ②合规三问调研", "📁原始资料包+📝合规调研笔记", "1.5h"),
    ("周三", "7/23", "手工标注第1份校园餐报告：记录所有人工检查点", "📝人工复核Checklist v1.0", "1h"),
    ("周四", "7/24", "OpenClaw跑AI复核 vs 人工对比，统计采纳率", "🔀人机对比表#1", "1h"),
    ("周五", "7/25", "换医保资金报告→人机对比#2", "🔀人机对比表#2", "1h"),
    ("周六", "7/26", "第3份不同来源/年份报告→人机对比#3", "🔀人机对比表#3", "1h"),
    ("周日", "7/27", "整理周报：采纳率+失效模式+脏数据清单", "📄第1周验证周报", "1.5h"),
    ("周一", "7/28", "定稿量化通过标准+白皮书大纲", "✅标准文档+📘白皮书大纲", "1h"),
]
for day, date, task, output, time in plan:
    r += 1
    set_cell(ws, f"A{r}", day, normal, None, wrap_align, thin_border)
    set_cell(ws, f"B{r}", date, normal, None, wrap_align, thin_border)
    set_cell(ws, f"C{r}", task, normal, None, wrap_align, thin_border)
    set_cell(ws, f"D{r}", output, normal, None, wrap_align, thin_border)
    set_cell(ws, f"E{r}", time, Font(name="微软雅黑", size=10, bold=True, color=TEAL), None, wrap_align, thin_border)
    set_cell(ws, f"F{r}", "☐", normal, None, wrap_align, thin_border)

ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 45
ws.column_dimensions['D'].width = 32
ws.column_dimensions['E'].width = 8
ws.column_dimensions['F'].width = 6

# ==================== Sheet 8: 决策门 ====================
ws = wb.create_sheet("决策门")
ws.sheet_properties.tabColor = "7030A0"

merge_and_set(ws, "A1:E1", "审盾一期·决策门（Go/No-Go Gates）", title_font)
merge_and_set(ws, "A2:E2", "没有止损线 = 没有真正的验证。每次决策门不达标，必须执行对应的降级动作。", red_f)

r = 4
header_row(ws, r, ["节点", "检查标准", "达标 →", "不达标 →", "责任人"])
gates = [
    ("Day 14 (M0)", "≥5份报告人机对比完成\n量化标准已书面确认\n合规三问有答案", "→ 进入阶段1B\n（规则引擎验证）", "→ 延长1周补做\n暂不进入1B", "平头哥"),
    ("Day 60 (M1)", "规则引擎准确率≥95%\n误报率≤10%\n≥200条标注数据", "→ 进入阶段1C\n（语义复核验证）", "→ 延长至Day90重测\n不招人", "平头哥"),
    ("Day 90 (M2)", "三指标同时达标：\n检出率≥80%\n误报率≤20%\n采纳率≥40%", "→ 招第1个AI工程师\n进入阶段1D", "→ 部分达标:延长30天\n→ 严重不达标:降级为\n  AI辅助建议模式\n  不招人、不买GPU", "平头哥"),
    ("Day 180 (M3)", "3条线AI效率≥30%\n≥10人每周使用≥3次\n工程师可独立维护", "→ 进入第二阶段\n追加GPU+扩团队", "→ 不进入Phase 2\n维持1人+API\n作为AI辅助工具集\n低成本运作", "平头哥"),
    ("Phase 2 第3月", "每条线采纳率趋势上升\n核心用户留存>80%", "→ 继续推广", "→ 缩减为2条线\n砍掉效果最差的那条", "平头哥"),
    ("Phase 2 第12月", "≥3条线采纳率≥50%\n核心用户≥20人\n私测合作方反馈正向\n标注数据集≥5000条", "→ 进入第三阶段", "→ 不进入Phase 3\n仅内部使用\n放弃对外输出", "平头哥"),
]
for node, check, go, nogo, owner in gates:
    r += 1
    set_cell(ws, f"A{r}", node, bold_n, None, wrap_align, thin_border)
    set_cell(ws, f"B{r}", check, normal, None, wrap_align, thin_border)
    set_cell(ws, f"C{r}", go, green_f, None, wrap_align, thin_border)
    set_cell(ws, f"D{r}", nogo, red_f, None, wrap_align, thin_border)
    set_cell(ws, f"E{r}", owner, normal, None, wrap_align, thin_border)

ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 32
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 24
ws.column_dimensions['E'].width = 8

# ==================== Save ====================
outpath = r"C:\Users\scrccpa\Desktop\审盾-三位一体方案-v2.0.xlsx"
wb.save(outpath)
print(f"✅ Excel已保存: {outpath}")
print(f"   8个工作表: 总览 | 一阶段 | 二阶段 | 三阶段 | 预算对比 | 风险清单 | 第1周计划 | 决策门")
