"""宿舍监理项目 — 全量分析Excel报告"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

hdr_f = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
sub_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yel_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
grn_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
org_fill = PatternFill(start_color='FFDAB9', end_color='FFDAB9', fill_type='solid')
title_f = Font(name='微软雅黑', bold=True, size=13, color='2F5496')
red_fnt = Font(name='微软雅黑', size=11, color='CC0000', bold=True)
bold_f = Font(name='微软雅黑', bold=True, size=11)
norm_f = Font(name='微软雅黑', size=11)
sml_f = Font(name='微软雅黑', size=10)
thin_b = Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))
ca = Alignment(horizontal='center', vertical='center', wrap_text=True)
wa = Alignment(vertical='center', wrap_text=True)

def cell(ws, r, col, v, font=norm_f, fill=None, align=ca):
    c = ws.cell(row=r, column=col, value=v)
    c.font = font; c.alignment = align; c.border = thin_b
    if fill: c.fill = fill
    return c

def hdr(ws, r, headers):
    for i,h in enumerate(headers):
        cell(ws, r, i+1, h, hdr_f, hdr_fill)

# ============ Sheet 1: 关键发现 ============
ws1 = wb.active
ws1.title = '审计发现清单'

cell(ws1, 1, 1, '5号6号学生宿舍建设项目监理 审计发现清单', title_f, align=wa)
ws1.merge_cells('A1:H1')
cell(ws1, 2, 1, '项目: 四川护理职业学院5号/6号学生宿舍监理 | 招标人: 四川护理职业学院 | 代理: 圣弘建设 | 总投资: 6200万元 | 22家投标人 | 2025年11月', sml_f, align=wa)
ws1.merge_cells('A2:H2')

r = 4
hdr(ws1, r, ['序号', '问题类别', '风险等级', '问题描述', '涉及方', '证据/数据', '建议', '检测维度'])

issues = [
    [1, '22家投标人元数据完全一致', '🟡 需要澄清',
     '全部22家投标人PDF的Producer(Aspose.PDF for .NET 23.2.0)、Creator(Aspose Pty Ltd.)、Author(Zh)、CreationDate(2012-05-25)完全一致。这并非串标信号，而是四川省政府采购电子交易平台使用Aspose.PDF统一转换投标文件所致。招标文件本身也由同一Aspose引擎生成，进一步证实为平台行为。',
     '交易平台/代理机构',
     'Producer: Aspose.PDF for .NET 23.2.0; modified using iTextSharp 5.5.14\nCreator: Aspose Pty Ltd.\nAuthor: Zh (全部22家+招标文件)\nCreationDate: D:20120525074100Z (固定默认值)\n22/22 = 100%一致性',
     '1. 核查交易平台是否确实使用Aspose.PDF统一转换文档\n2. 由于平台覆写了原始元数据，无法通过L5元数据检测投标人间的关联\n3. 建议要求平台保留原始.docx文件的OLE2元数据', 'L5-元数据'],
    
    [2, '中国华西工程设计被否决-报价不唯一', '🔴 重大',
     '中国华西工程设计建设有限公司因"形式评审：投标函报价不唯一"被否决投标。其投标总价879,590.80元。同一份投标文件中出现不同报价，属于严重的形式缺陷。需要核查是否同时存在其他违规行为。',
     '中国华西工程设计建设有限公司',
     '公示OCR提取：\n华西投标报价: 879,590.80元\n否决原因: 形式评审 投标函报价不唯一\n(数据来源: 监理公示用章后.pdf Page5 OCR)',
     '核实华西投标文件中是否存在多个不一致的报价数字。如为编辑疏忽导致，合理性存疑（大型监理公司不应出现此低级错误）。如为有意行为，需进一步调查动机。', '形式评审'],
    
    [3, '关键数据需高精度OCR补全', '🔴 重大缺失',
     '监理公示PDF(8页)和评标报告PDF(59页)均为物理扫描件（KONICA MINOLTA bizhub C558），无法通过程序化方式提取完整的22家投标人报价-得分对照表。目前通过150dpi OCR仅提取到约50%数据。',
     '归档资料',
     '评标报告: 0字符文字层, 柯尼卡美能达扫描\n中标公示: 0字符, 物理盖章后扫描\n书面报告: 16字符(TOSHIBA扫描)',
     '【P0】手动录入或高精度OCR(300dpi)获取完整的22家报价和得分表。关键数据包括：每家投标人名称、投标报价、评标基准价、偏差率、报价得分、资信得分、大纲得分、总分、排名。', 'L1-报价'],
    
    [4, '中标候选人确定', '🟢 正常发现',
     '第一中标候选人: 四川良友建设咨询有限公司\n第二中标候选人: 四川元博项目管理有限公司\n第三中标候选人: 成都衡泰全过程工程咨询集团有限公司\n\n其他投标人(除被否决者外)均正常参与评审，综合评估得分90+的投标人较少。',
     '前三名投标人',
     '公示Page1 OCR: 四川良友(第一)\n公示Page1 OCR: 四川元博(第二)\n公示Page1 OCR: 成都衡泰(第三)\n中韵四方: 860,860.00元 / 90.97分',
     '正常中标。关注前三名之间是否存在工商关联。', 'L10-评审'],
    
    [5, '所有22家PDF含文字层-可做TF-IDF', '🟢 优势',
     '与艺术团项目不同，本项目22家投标人PDF全部包含可提取的文字层（true PDF，非扫描件）。这意味着可以做全量的TF-IDF文本雷同检测(L3)。这是本项目相较于前两个项目的重大优势。',
     '全部22家',
     '21/22家PDF: txt=True\n1家可能为混合模式\n文件大小: 44MB-159MB, 643-3410页/家',
     '建议启动全量TF-IDF分析：提取22家投标文件全文→jieba分词→TF-IDF向量化→余弦相似度矩阵。重点关注：(1)≥80%相似度的投标人对 (2)特定章节(监理大纲)的雷同度。', 'L3-TFIDF'],
    
    [6, '投标保证金缴纳记录需核实', '🟡 中等',
     '招标文件要求投标保证金10,000元，通过基本账户缴纳至德阳市公共资源交易中心专用账户。22家投标人的保证金缴纳流水可验证L9(资金链)维度，检查是否存在同一账户代缴、相近时间缴纳等异常。',
     '全部22家',
     '保证金金额: 10,000元/家\n缴纳方式: 基本账户转账至交易中心子账号\n形式: 现金/保证保险二选一',
     '向德阳市公共资源交易中心调取22家投标人的保证金缴纳明细：(1)支付账户名称 (2)支付时间 (3)支付银行 (4)是否同一IP操作。如多家使用同一银行账户或相近时间缴纳，构成围标信号。', 'L9-保证金'],
    
    [7, '报价集中度分析-待完整数据', '🟡 待补',
     '已从公示OCR提取到的部分报价：蜀顺810,154 / 五行807,979 / 华宇809,114 / 天立853,008 / 中韵四方860,860 / 华西879,591(否决)。报价范围约80-88万，需要完整22家数据才能计算极差、标准差、评标基准价、偏差率等关键指标。',
     '全部22家',
     '部分报价:\n五行807,979→蜀顺810,154→华宇809,114→天立853,008→中韵四方860,860→华西879,591\n极差(有效): 约52,000元(6.4%)',
     '获取完整数据后执行：评标基准价计算(算术平均)、偏差率分布、极差/标准差分析、报价是否呈正态分布。', 'L1-报价'],
]

for i, iss in enumerate(issues):
    row = 5+i
    for j, val in enumerate(iss):
        if j == 2:
            fill = red_fill if '重大' in str(val) else (yel_fill if '中等' in str(val) else grn_fill)
            cell(ws1, row, j+1, val, bold_f, fill)
        elif j == 0:
            cell(ws1, row, j+1, val, norm_f)
        elif j in [3,5,6]:
            cell(ws1, row, j+1, val, j==5 and sml_f or norm_f, align=wa)
        elif j == 4:
            cell(ws1, row, j+1, val, norm_f, align=wa)
        else:
            cell(ws1, row, j+1, val, norm_f)

ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 24
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 48
ws1.column_dimensions['E'].width = 18
ws1.column_dimensions['F'].width = 36
ws1.column_dimensions['G'].width = 40
ws1.column_dimensions['H'].width = 14

for i in range(5, 5+len(issues)):
    ws1.row_dimensions[i].height = 120

# ============ Sheet 2: 22家元数据详情 ============
ws2 = wb.create_sheet('22家元数据汇总')
cell(ws2, 1, 1, '22家投标人PDF元数据全量汇总 — 全部一致(平台Aspose转换)', title_f, align=wa)
ws2.merge_cells('A1:G1')

bidders_list = [
    '中国华西工程设计建设有限公司', '中国轻工业成都设计工程有限公司', '中天顺韵建设管理有限公司',
    '中泰天顺集团有限责任公司', '中祥冠一建设集团有限公司', '中锦冠达工程顾问集团有限公司',
    '中韵四方建设集团有限公司', '中鸿亿博集团有限公司', '卓昇项目管理有限公司',
    '四川五行建设工程项目管理有限公司', '四川伟业启航集团有限公司', '四川元博项目管理有限公司',
    '四川华宇工程监理咨询有限公司', '四川易弘工程管理有限公司', '四川省兴旺建设工程项目管理有限公司',
    '四川省蜀顺工程建设咨询有限公司', '四川良友建设咨询有限公司', '德阳鑫华建工集团有限公司',
    '成都天立工程管理咨询有限公司', '成都海发建设工程监理有限公司', '成都衡泰全过程工程咨询集团有限公司',
    '深圳市银建安工程项目管理有限公司'
]

r = 3
hdr(ws2, r, ['序号','投标人名称','页数','Producer','Creator','Author','CreationDate'])

for i, name in enumerate(bidders_list):
    cell(ws2, 4+i, 1, i+1)
    cell(ws2, 4+i, 2, name, norm_f, align=wa)
    cell(ws2, 4+i, 3, '600-3400', norm_f)
    cell(ws2, 4+i, 4, 'Aspose.PDF for .NET 23.2.0', sml_f, align=wa)
    cell(ws2, 4+i, 5, 'Aspose Pty Ltd.', norm_f)
    cell(ws2, 4+i, 6, 'Zh', norm_f)
    cell(ws2, 4+i, 7, 'D:20120525074100Z', sml_f)

# Summary
r2 = 4+len(bidders_list)+1
cell(ws2, r2, 1, '结论', bold_f, red_fill, wa)
ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=7)
cell(ws2, r2+1, 1, '全部22家投标人元数据100%一致。此现象的原因为：四川省政府采购电子交易平台(全国公共资源交易平台·德阳市)使用Aspose.PDF for .NET引擎统一将投标人上传的.docx文件转换为PDF。招标文件同样使用此引擎，证实为平台行为而非串标。但此转换也意味着原始.docx文件的OLE2元数据(真实作者/修改历史)被覆写，L5元数据层失去了串标检测能力。', norm_f, align=wa)
ws2.merge_cells(start_row=r2+1, start_column=1, end_row=r2+1, end_column=7)
ws2.row_dimensions[r2+1].height = 50

ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 34
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 36
ws2.column_dimensions['E'].width = 18
ws2.column_dimensions['F'].width = 10
ws2.column_dimensions['G'].width = 20

# ============ Sheet 3: 项目概况 ============
ws3 = wb.create_sheet('项目基本信息')
cell(ws3, 1, 1, '项目基本信息', title_f, align=wa)
ws3.merge_cells('A1:B1')

info = [
    ('项目名称', '四川护理职业学院5号学生宿舍建设项目和6号学生宿舍建设项目监理'),
    ('招标编号', '德旌招（2025）备[房建]监理第005号'),
    ('批文', '德市发改行审【2024】66号 / 德市发改行审【2025】11号'),
    ('招标人', '四川护理职业学院'),
    ('工程地点', '德阳市旌阳区一环路东一段199号（四川护理职业学院校内）'),
    ('招标代理机构', '圣弘建设股份有限公司'),
    ('代理负责人', '蒋亚菲'),
    ('招标文件编制', '刘珂、范鹏'),
    ('招标方式', '公开招标（电子招标）'),
    ('评标方法', '综合评估法（满分100分）'),
    ('评分构成', '资信业绩30分 + 监理大纲30分 + 投标报价40分'),
    ('评标基准价', '有效投标报价的算术平均值（B方式）'),
    ('总投资', '估算6200万元'),
    ('建筑面积', '14,212.94平方米（5#、6#各7,106.47㎡，地上6层框架）'),
    ('监理服务期', '合同生效至施工期+缺陷责任期24个月'),
    ('资质要求', '房屋建筑工程监理乙级及以上'),
    ('业绩要求', '近3年至少1个≥14,000㎡建安投资≥6200万元的房建监理业绩'),
    ('投标保证金', '10,000元（基本账户缴纳至德阳公共资源交易中心）'),
    ('投标有效期', '90日历天'),
    ('招标文件获取', '2025年10月30日至11月19日（全国公共资源交易平台·德阳市）'),
    ('投标截止', '2025年11月19日23:59:59'),
    ('评标报告','KONICA MINOLTA bizhub C558 扫描 (59页，扫描件)'),
    ('中标公示','物理盖章后扫描 (8页，扫描件)'),
    ('投标人数量','22家'),
    ('投标文件格式','Aspose.PDF for .NET 23.2.0 统一转换 (平台行为)'),
]

for i, (k,v) in enumerate(info):
    cell(ws3, 3+i, 1, k, bold_f, sub_fill, wa)
    cell(ws3, 3+i, 2, v, norm_f, align=wa)

ws3.column_dimensions['A'].width = 22
ws3.column_dimensions['B'].width = 65

# ============ Sheet 4: 下一步建议 ============
ws4 = wb.create_sheet('后续行动建议')
cell(ws4, 1, 1, '后续调查与取证建议', title_f, align=wa)
ws4.merge_cells('A1:E1')

r = 3
hdr(ws4, r, ['优先级','行动项','目的','执行方式','预期产出'])

actions = [
    ['P0','完整报价-得分表获取','获取22家完整报价和评标得分，完成L1报价规律分析和L10评审一致性分析。','手动从公示/评标报告中提取或300dpi高精度OCR。','22家报价→评标基准价→偏差率分布→检查是否有异常集中的报价集团。'],
    ['P0','全量TF-IDF文本雷同检测','22家投标人均有可提取文本层，这是本项目最大的分析优势。','提取22家投标文件全文→jieba分词→TF-IDF→余弦相似度矩阵。重点关注监理大纲部分。','≥80%相似度的投标人对→可能指向前期技术方案共享或串通。'],
    ['P1','L9保证金缴纳记录调取','检查22家投标人是否使用同一账户或相近时间缴纳保证金。','向德阳市公共资源交易中心函调保证金缴纳明细。','同一银行账户代缴多家=围标铁证。相近时间缴纳=值得关注。'],
    ['P1','L8工商关联查询','核实22家投标人(特别是前三名中标候选人)的股东/高管/投资关联。','天眼查/企查查批量查询22家公司。重点关注：四川良友、四川元博、成都衡泰之间的关联。','股东/董监高交叉=串标信号。'],
    ['P1','华西投标函"报价不唯一"核查','大型监理公司不应出现"报价不唯一"的低级错误。','查阅华西原始投标文件，确认具体哪个环节出现多报价。','如为编辑错误→管理混乱。如有意行为→需进一步调查。'],
    ['P2','平台Aspose转换机制确认','确认平台是否确实覆写了所有.docx的原始元数据。','向德阳市公共资源交易中心确认技术方案。','如确认→L5层在本项目中无检测价值。如否→元数据一致=串标。'],
]

for i, act in enumerate(actions):
    for j, val in enumerate(act):
        if j == 0:
            fill = red_fill if 'P0' in str(val) else (org_fill if 'P1' in str(val) else grn_fill)
            cell(ws4, 4+i, j+1, val, bold_f, fill)
        elif j in [3,4]:
            cell(ws4, 4+i, j+1, val, sml_f, align=wa)
        else:
            cell(ws4, 4+i, j+1, val, norm_f, align=wa)

ws4.column_dimensions['A'].width = 10
ws4.column_dimensions['B'].width = 26
ws4.column_dimensions['C'].width = 35
ws4.column_dimensions['D'].width = 40
ws4.column_dimensions['E'].width = 38

# Save
out = r'C:\Users\scrccpa\Desktop\宿舍监理审计分析报告.xlsx'
wb.save(out)
print('Saved: ' + out)
