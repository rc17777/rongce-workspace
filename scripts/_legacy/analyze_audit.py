import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

PATH = r'C:\Users\scrccpa\Desktop\融策审计过程记录系统=项目经理版(6).xlsx'
wb = openpyxl.load_workbook(PATH, data_only=True)

# Sheet 2: 审计过程 - full data
ws = wb['2-审计过程']
print(f'=== 2-审计过程: {ws.max_row} rows, {ws.max_column} cols ===')
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
    vals = [str(v) if v is not None else '' for v in row]
    print(f'R{i}: ' + ' | '.join(vals))
