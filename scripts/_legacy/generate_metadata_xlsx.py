#!/usr/bin/env python3
"""医工设备采购 - 元数据/扫描仪/硬件信息深度分析Excel"""
import sys, io, os, re, base64, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def sf(c):
    return PatternFill(patternType='solid', fgColor=c)

RED_FILL = sf('FFD7D7')
YELLOW_FILL = sf('FFF3CD')
GREEN_FILL = sf('D4EDDA')
HEADER_FILL = sf('1A3A6E')
LIGHT_BLUE = sf('E8F0FE')
LIGHT_GRAY = sf('F5F5F5')
H_FONT = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
T_FONT = Font(name='Microsoft YaHei', size=14, bold=True, color='1A3A6E')
SUB_FONT = Font(name='Microsoft YaHei', size=12, bold=True, color='1A3A6E')
N = Font(name='Microsoft YaHei', size=10)
B = Font(name='Microsoft YaHei', size=10, bold=True)
BR = Font(name='Microsoft YaHei', size=10, color='CC0000', bold=True)
I = Font(name='Microsoft YaHei', size=9, italic=True, color='888888')
TH = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
C = Alignment(horizontal='center', vertical='center', wrap_text=True)
L = Alignment(horizontal='left', vertical='center', wrap_text=True)


def hdr(ws, row, headers):
    for c, h in enumerate(headers, 1):
        cl = ws.cell(row=row, column=c, value=h)
        cl.font = H_FONT
        cl.fill = HEADER_FILL
        cl.alignment = C
        cl.border = TH


def cl(ws, r, c, val, font=N, align=C, fill=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = font
    cell.alignment = align
    cell.border = TH
    if fill: cell.fill = fill


wb = Workbook()

# ===== Sheet 1: PDF元数据全景 =====
ws1 = wb.active
ws1.title = 'PDF元数据全景'
ws1.merge_cells('A1:N1')
ws1['A1'] = '投标文件PDF - 扫描设备/图像/元数据全景分析'
ws1['A1'].font = T_FONT
ws1.merge_cells('A2:N2')
ws1[
    'A2'] = f'项目: HTGJ-CS(2025)-97号 四川护理职业学院医工教研实训设备采购 | 分析时间: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
ws1['A2'].font = I

hdr(ws1, 4,
    ['投标单位', '文件类型', '页数', '大小(MB)', '扫描设备(Creator)', '扫描驱动(Producer)', '扫描时间', '修改时间', 'PDF版本',
     '扫描DPI', '图像尺寸', '色彩深度', '色彩空间', 'MD5(前16位)'])

meta_data = [
    ['百安智能', '资格性', '39', '11.3', 'RICOH MP 9003', 'RICOH MP 9003',
     '2025-06-03 13:29 (EDT)', '2025-06-03 10:18 (CST)', 'PDF 1.6', '424×212 (A4 Portrait)', '3508×2480px', '8 bpc',
     'DeviceRGB', 'ee6b7844e3e2a790'],
    ['百安智能', '其他', '121', '42.2', 'RICOH MP 9003', 'RICOH MP 9003', '2025-06-03 13:29 (EDT)',
     '2025-06-03 10:17 (CST)', 'PDF 1.6', '424×212 (A4 Portrait)', '3508×2480px', '8 bpc', 'DeviceRGB', '9aabec3b8876ed96'],
    ['成都逐声科技', '资格性', '15', '12.0', '(无) — 非扫描仪生成', '(无)', '—', '2025-06-02 11:03 (CST)', 'PDF 1.7',
     '192×192 (等比例)', '1587×2245px', '8 bpc', 'DeviceRGB', 'a5ce2a8ae80bd741'],
    ['成都逐声科技', '其他', '69', '76.0', '(无) — 非扫描仪生成', '(无)', '—', '2025-06-02 11:04 (CST)', 'PDF 1.7',
     '192×192 (等比例)', '1587×2245px', '8 bpc', 'DeviceRGB', '1304e51e551c5191'],
    ['中科兴蓉科技', '资格性', '32', '10.1', 'RICOH MP 9003', 'RICOH MP 9003', '2025-06-03 15:22 (EDT)',
     '2025-06-03 15:53 (CST)', 'PDF 1.6', '424×212 (A4 Portrait)', '3508×2480px', '8 bpc', 'DeviceRGB', '57d3d77827889551'],
    ['中科兴蓉科技', '其他', '129', '46.9', 'RICOH MP 9003', 'RICOH MP 9003', '2025-06-03 15:25 (EDT)',
     '2025-06-03 15:53 (CST)', 'PDF 1.6', '424×212 (A4 Portrait)', '3508×2480px', '8 bpc', 'DeviceRGB', 'b7ca306f938b6331'],
]

for r, row in enumerate(meta_data, 5):
    for c, val in enumerate(row, 1):
        fill = None
        if c == 5:
            fill = YELLOW_FILL if 'RICOH' in str(val) else GREEN_FILL
        elif c == 10:
            fill = YELLOW_FILL if '192' in str(val) else GREEN_FILL
        elif c == 9:
            fill = YELLOW_FILL if '1.7' in str(val) else None
        cl(ws1, r, c, val, B if c <= 2 else N, C if c <= 4 else L, fill)

# Analysis block
r1 = 12
ws1.merge_cells(f'A{r1}:N{r1}')
ws1[f'A{r1}'] = '🔍 关键发现'
ws1[f'A{r1}'].font = Font(name='Microsoft YaHei', size=13, bold=True, color='CC0000')

analysis_lines = [
    '1. [扫描设备] 百安智能 & 中科兴蓉: 均使用 RICOH MP 9003 (理光高端A3黑白数码复合机，上市约2013-2015年)。逐声科技: 无creator/producer字段，可能使用非扫描仪软件(如Adobe/Word直接导出)生成PDF。',
    '2. [DPI差异] 百安/中科: 424×212 DPI (A4纵向扫描，水平424dpi为主扫描方向，垂直212dpi为副扫描方向，典型"标准"质量设定)。逐声: 192×192 DPI (等比例，非标准A4比例，扫描质量明显较低)。',
    '3. [PDF版本] 百安/中科: PDF 1.6 (RICOH标准输出)。逐声: PDF 1.7 (更新版本，暗示使用不同的PDF生成软件/驱动)。',
    '4. [时间线] 逐声: 6月2日(最早，开标前2天) → 百安: 6月3日13:29 → 中科: 6月3日15:22-25。开标: 6月4日。三家时间分散在两天内，时空关系不紧密。',
    '5. [时区异常] 百安和中科的创建时间时区标记为-04(美国东部夏令时)，修改时间为+08(北京)。这是扫描仪/打印驱动时区配置错误，常见于RICOH设备的默认出厂设置。',
    '6. [结论] 逐声科技的PDF特征(无scanner metadata、PDF 1.7、低DPI等比例扫描)与百安/中科(RICOH扫描、PDF 1.6、424dpi A4)完全不同。百安和中科共用RICOH MP 9003属于中等信号，但不构成串标铁证。'
]
for i, line in enumerate(analysis_lines):
    ws1.merge_cells(f'A{r1 + 1 + i}:N{r1 + 1 + i}')
    ws1[f'A{r1 + 1 + i}'] = line
    ws1[f'A{r1 + 1 + i}'].font = N
    ws1[f'A{r1 + 1 + i}'].alignment = L

ws1.column_dimensions['A'].width = 16
ws1.column_dimensions['B'].width = 10
ws1.column_dimensions['C'].width = 6
ws1.column_dimensions['D'].width = 10
ws1.column_dimensions['E'].width = 24
ws1.column_dimensions['F'].width = 22
ws1.column_dimensions['G'].width = 24
ws1.column_dimensions['H'].width = 24
ws1.column_dimensions['I'].width = 12
ws1.column_dimensions['J'].width = 20
ws1.column_dimensions['K'].width = 14
ws1.column_dimensions['L'].width = 10
ws1.column_dimensions['M'].width = 14
ws1.column_dimensions['N'].width = 16

# ===== Sheet 2: 招标文件DOCX属性 =====
ws2 = wb.create_sheet('招标文件DOCX属性')
ws2.merge_cells('A1:D1')
ws2['A1'] = '招标采购文件 (.docx) 深层属性提取'
ws2['A1'].font = T_FONT
ws2.merge_cells('A2:D2')
ws2['A2'] = '⚠️ DOCX文件包含可直接关联到具体人的元数据 — 创建者姓名/WPS账号/硬件ID'
ws2['A2'].font = Font(name='Microsoft YaHei', size=10, color='CC0000', italic=True)

hdr(ws2, 4, ['属性类别', '属性名称', '属性值', '含义/注释'])

# Decode custom.xml base64
custom_b64 = 'eyJoZGlkIjoiYzI1ZGZmMjQ0ZjI2ZmZjODllNDg1ZjJkNTg4ZmM4MGUiLCJ1c2VySWQiOiIxMzg5OTkyNDYzIn0='
try:
    custom_json = json.loads(base64.b64decode(custom_b64).decode('utf-8'))
except:
    custom_json = {'hdid': 'decode error', 'userId': 'decode error'}

docx_meta = [
    ['身份信息', '创建者(Creator)', '张鹏举', '原始文档创建者姓名 — 可直接追溯到个人'],
    ['身份信息', '最后修改者', '代理机构', '最终编辑者标记为"代理机构"'],
    ['身份信息', 'WPS用户ID', custom_json.get('userId', ''), 'WPS云账号ID: 1389992463 — 可关联WPS账号'],
    ['设备信息', 'WPS硬件ID(HDID)', custom_json.get('hdid', ''), 'WPS Office硬件设备指纹 — 唯一标识安装WPS的计算机'],
    ['设备信息', 'WPS版本号', '11.1.0.14309_F1E327BC-269C-435d-A152-05C5408002CA', 'GUID后缀F1E327BC...为安装包标识，同版本WPS共享同一后缀'],
    ['时间信息', '创建时间', '2020-04-26 01:34', '原始模板创建时间(约5年前) — 非本次项目创建时间'],
    ['时间信息', '最后修改时间', '2025-05-19 14:00', '本次项目实际修改时间 — 开标(6月4日)前约2周'],
    ['时间信息', '最后打印时间', '2023-11-21 06:59', '上一次打印时间(约1.5年前) — 模板被复用'],
    ['时间信息', '编辑时长(TotalTime)', '33分钟', '累计编辑33分钟 — 说明以模板为基础修改'],
    ['文档信息', '修订次数(Revision)', '557次', '557次修订 — 说明该模板经历多次编辑迭代'],
    ['文档信息', '页数/字数', '70页 / 38,798字', '完整的招标文件编制'],
    ['软件信息', '应用程序', 'WPS Office 11.1.0.14309', '使用WPS Office编辑 — 非Microsoft Word'],
    ['软件信息', '模板', 'Normal', '使用Normal默认模板'],
    ['软件信息', 'Company字段', 'Microsoft', 'WPS Office默认填充"Microsoft"(兼容Word行为) — 非实际公司名'],
    ['安全信息', 'DocSecurity', '0', '文档未加密/未限制编辑'],
]

for r, row in enumerate(docx_meta, 5):
    cat = row[0]
    fill = None
    if '身份' in cat or '设备' in cat:
        fill = RED_FILL
    elif '时间' in cat:
        fill = LIGHT_BLUE
    elif '软件' in cat:
        fill = LIGHT_GRAY
    for c, val in enumerate(row, 1):
        cl(ws2, r, c, val, B if c <= 2 else N, L if c >= 3 else C, fill if c <= 1 else None)

r2 = 21
ws2.merge_cells(f'A{r2}:D{r2}')
ws2[f'A{r2}'] = '⚠️ 注意: DOCX元数据包含创建者"张鹏举"的WPS账号ID(1389992463)和硬件设备指纹。该信息可唯一追溯到制作招标文件的个人和计算机。'
ws2[f'A{r2}'].font = BR
ws2[f'A{r2}'].alignment = L
ws2.row_dimensions[r2].height = 30

ws2.column_dimensions['A'].width = 14
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 50
ws2.column_dimensions['D'].width = 45

# ===== Sheet 3: 设备信息对比 =====
ws3 = wb.create_sheet('设备信息对比')
ws3.merge_cells('A1:H1')
ws3['A1'] = '三家公司投标文件生成设备对比'
ws3['A1'].font = T_FONT

hdr(ws3, 3, ['对比维度', '百安智能', '成都逐声科技', '中科兴蓉科技', '比对结果', '风险等级', '技术说明'])

device_compare = [
    ['扫描设备型号', 'RICOH MP 9003', '(无metadata)', 'RICOH MP 9003', '2/3相同(百安+中科)', 'YELLOW',
     '理光MP 9003是2013年发布的A3黑白复合机，政府/企业办公常见设备'],
    ['扫描DPI(主扫描)', '424 dpi', '192 dpi', '424 dpi', '百安=中科 ≠ 逐声', 'YELLOW', 'DPI相同暗示相同扫描设定，或同一设备相同默认参数'],
    ['扫描DPI(副扫描)', '212 dpi', '192 dpi', '212 dpi', '百安=中科 ≠ 逐声', 'YELLOW', 'A4 Portrait扫描的典型纵向DPI'],
    ['图像尺寸', '3508×2480px', '1587×2245px', '3508×2480px', '百安=中科 ≠ 逐声', 'YELLOW', '像素尺寸由扫描DPI和纸张大小决定'],
    ['PDF版本', '1.6', '1.7', '1.6', '百安=中科(1.6) ≠ 逐声(1.7)', 'GREEN', 'PDF 1.7通常由较新软件生成(Adobe 8+/较新WPS/drivers)'],
    ['色彩空间', 'DeviceRGB', 'DeviceRGB', 'DeviceRGB', '全部相同', 'GREEN', '标准彩色扫描，无特殊处理'],
    ['色彩深度', '8 bpc', '8 bpc', '8 bpc', '全部相同', 'GREEN', '24位真彩色标准'],
    ['嵌入字体', '无(纯图像)', '无(纯图像)', '无(纯图像)', '全部相同', 'GREEN', '扫描件无嵌入字体，证明全为纸质扫描'],
    ['Creator字段', '有(RICOH MP 9003)', '无', '有(RICOH MP 9003)', '部分相同', 'YELLOW', '无creator说明PDF非扫描仪直接生成，可能是软件合成'],
    ['扫描时间', '2025-06-03 13:29', '不适用(非扫描)', '2025-06-03 15:22/25', '百安和中科相差约2小时', 'GREEN', '时间分散，非同一时刻'],
    ['资格:其他时间差', '相同13:29(同1分钟)', '—', '15:22 vs 15:25(差3分钟)', '均在合理范围', 'GREEN', '同一批次扫描多份文件的正常时间间隔'],
    ['文件大小效率', '0.29MB/页(资格) 0.35MB/页(其他)', '0.80MB/页(资格) 1.10MB/页(其他)', '0.31MB/页(资格) 0.36MB/页(其他)',
     '逐声文件密度大3-4倍', 'GREEN', '逐声每页文件更大，可能是高分辨率图像或不同压缩设置'],
]

for r, row in enumerate(device_compare, 4):
    for c, val in enumerate(row, 1):
        fill = RED_FILL if val == 'RED' else (YELLOW_FILL if val == 'YELLOW' else None)
        cl(ws3, r, c, val, B if c == 1 else N, L, fill)

r3 = 17
ws3.merge_cells(f'A{r3}:H{r3}')
ws3[f'A{r3}'] = '核心结论: 百安智能和中科兴蓉极可能使用同一台RICOH MP 9003扫描仪。但逐声科技的PDF特征完全不同(无扫描仪metadata、低DPI、PDF 1.7)，三家并非"统一制作"。这与上个项目(3家全部Fuji Xerox D125)有本质区别。'
ws3[f'A{r3}'].font = BR
ws3[f'A{r3}'].alignment = L
ws3.row_dimensions[r3].height = 35

ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 22
ws3.column_dimensions['C'].width = 22
ws3.column_dimensions['D'].width = 22
ws3.column_dimensions['E'].width = 26
ws3.column_dimensions['F'].width = 14
ws3.column_dimensions['G'].width = 55

# ===== Sheet 4: 硬盘/CPU/软件指纹 =====
ws4 = wb.create_sheet('硬件与软件指纹')
ws4.merge_cells('A1:G1')
ws4['A1'] = '投标文件可提取的硬件/软件指纹信息'
ws4['A1'].font = T_FONT
ws4.merge_cells('A2:G2')
ws4[
    'A2'] = '⚠️ 说明: PDF扫描件不包含源电脑的硬盘/CPU/主板信息。以下为WPS Office .docx文件中可提取的设备指纹。投标PDF无法获取此类信息。'
ws4['A2'].font = Font(name='Microsoft YaHei', size=9, color='CC0000', italic=True)

hdr(ws4, 4, ['信息类别', '存在形式', '本项目中可提取的值', '可追溯性', '隐私/安全风险', '来源文件', '备注'])

hw_info = [
    ['WPS用户ID', 'WPS云账号数字ID', '1389992463', '⭐可追溯到WPS云账号', '高 — 关联云文档、个人信息', '招标文件.docx (custom.xml)',
     '编码在custom.xml中，Base64解码获取'],
    ['WPS硬件设备ID (HDID)', '设备指纹Hash', 'c25dff244f26ffc89e485f2d588fc80e', '⭐可唯一标识安装WPS的计算机', '高 — 精确到单台计算机',
     '招标文件.docx (custom.xml)', 'WPS Office为每台安装设备生成唯一标识'],
    ['WPS版本号', '安装包GUID', 'F1E327BC-269C-435d-A152-05C5408002CA', '只能追溯安装包版本', '低 — 同版本共享',
     '招标文件.docx (app.xml)', '该GUID后缀与上个项目(培训资料采购)的WPS版本一致'],
    ['操作系统用户名', '创建者字段', '张鹏举', '⭐可追溯到制作招标文件的个人', '高 — 直接暴露个人姓名', '招标文件.docx (core.xml)',
     '通常是Windows登录用户名或WPS账号显示名'],
    ['电脑名称/主机名', '不可获取', 'N/A — PDF不包含此信息', '—', '—', '—', 'PDF扫描件无此元数据字段'],
    ['硬盘序列号', '不可获取', 'N/A — PDF不包含此信息', '—', '—', '—', '扫描PDF为纯图像，不包含源电脑硬件信息'],
    ['CPU型号', '不可获取', 'N/A — PDF不包含此信息', '—', '—', '—', '扫描PDF为纯图像，不包含源电脑硬件信息'],
    ['MAC地址', '不可获取', 'N/A — 纸质投标无网络信息', '—', '—', '—', '纸质投标流程无电子网络日志'],
    ['打印机/扫描仪序列号', '不可获取', 'N/A — metadata仅含型号不含序列号', '—', '—', '—',
     'RICOH MP 9003的metadata不包含设备序列号。需从代理机构获取扫描仪日志'],
    ['文档编辑软件', 'Application字段', 'WPS Office 11.1.0.14309', '可追溯软件版本', '低', '招标文件.docx (app.xml)',
     '与投标文件无关，仅招标文件编制者使用'],
]

for r, row in enumerate(hw_info, 5):
    for c, val in enumerate(row, 1):
        fill = RED_FILL if '高' in str(val) and c == 5 else (
            YELLOW_FILL if '不可获取' in str(val) and c == 3 else None)
        cl(ws4, r, c, val, B if c == 1 else N, L, fill)

r4 = 16
ws4.merge_cells(f'A{r4}:G{r4}')
ws4[f'A{r4}'] = ('重要说明:\n'
                 '1. 投标文件PDF均为纸质文件扫描件，元数据仅包含扫描仪信息，不包含投标人源电脑的任何硬件信息(硬盘/CPU/主板等)。\n'
                 '2. 唯一可提取的硬件指纹来自【招标文件.docx】，即代理机构/采购方制作招标文件的计算机，与三家投标供应商无关。\n'
                 '3. WPS Office版本GUID (F1E327BC...050C5408002CA)与上个培训资料采购项目中的WPS版本一致，说明同一WPS版本在本地广泛使用。\n'
                 '4. 如需获取投标人的硬件信息(电脑/MAC/IP)，需调取投标文件制作时的电子日志或电子投标系统后台记录。纸质投标无法提供此类证据。')
ws4[f'A{r4}'].font = N
ws4[f'A{r4}'].alignment = L
ws4.row_dimensions[r4].height = 90

ws4.column_dimensions['A'].width = 20
ws4.column_dimensions['B'].width = 18
ws4.column_dimensions['C'].width = 40
ws4.column_dimensions['D'].width = 22
ws4.column_dimensions['E'].width = 28
ws4.column_dimensions['F'].width = 26
ws4.column_dimensions['G'].width = 42

# ===== Sheet 5: 与前项目对比 =====
ws5 = wb.create_sheet('与培训资料项目对比')
ws5.merge_cells('A1:E1')
ws5['A1'] = '两个项目技术指标对比'
ws5['A1'].font = T_FONT

hdr(ws5, 3, ['对比维度', '培训资料采购项目', '医工设备采购项目', '差异分析', '信号强度变化'])

projects_compare = [
    ['扫描仪一致性', '3/3 全部 Fuji Xerox D125', '2/3 RICOH MP 9003, 1/3 无metadata', '从全同→部分相同', 'RED→YELLOW ↓↓'],
    ['PDF版本分布', '1.3(江楼) / 1.6(建韬+拓奇)', '1.6(百安+中科) / 1.7(逐声)', '医工项目差异更大', '—'],
    ['DPI分布', '未提取(上次分析未做)', '424×212(百安+中科) / 192×192(逐声)', '首次分析DPI维度', 'NEW'],
    ['TF-IDF文本相似度', '86.6%-89.8%(全量) 75.6%-80.9%(技术方案)', '62.7%-95.0%(注意:逐声+中科仅16页)', '培训项目更高但受模板影响', 'RED→YELLOW ↓'],
    ['文本体量(其他文件)', '92 / 58 / 327页', '160 / 84 / 161页', '培训项目体量悬殊5.6倍', '—'],
    ['地理集中度', '3家同在武侯区,建韬+拓奇玉林片区500m', '待查', '—', '待查'],
    ['WPS版本一致', '4/4一致(F1E327BC...)', '仅招标文件使用该WPS版本', '培训项目投标文件也含WPS', 'RED→GREEN ↓↓'],
    ['中标价vs预算', '均低于预算', '中标价=预算上限(35.9万)', '医工项目更极端', 'NEW: YELLOW'],
    ['综合风险评级', '★★★★☆ 中高风险', '★★☆☆☆ 中低风险', '医工项目信号明显更弱', '—'],
]

for r, row in enumerate(projects_compare, 4):
    for c, val in enumerate(row, 1):
        fill = None
        if 'RED' in str(val) and c == 5:
            fill = RED_FILL
        elif 'YELLOW' in str(val) and c == 5:
            fill = YELLOW_FILL
        elif 'GREEN' in str(val) and c == 5:
            fill = GREEN_FILL
        cl(ws5, r, c, val, B if c == 1 else N, L, fill)

r5 = 14
ws5.merge_cells(f'A{r5}:E{r5}')
ws5[f'A{r5}'] = ('结论: 医工设备采购项目的串标信号全面弱于培训资料采购项目。\n'
                 '关键差异: (1) 扫描仪从3/3降至2/3 (2) PDF版本多样性增加 (3) 无WPS版本一致性证据 (4) 逐声科技全部独立特征。\n'
                 '本项目不构成串标围标怀疑。')
ws5[f'A{r5}'].font = Font(name='Microsoft YaHei', size=11, bold=True)
ws5[f'A{r5}'].alignment = L
ws5.row_dimensions[r5].height = 55

ws5.column_dimensions['A'].width = 22
ws5.column_dimensions['B'].width = 30
ws5.column_dimensions['C'].width = 36
ws5.column_dimensions['D'].width = 28
ws5.column_dimensions['E'].width = 20

# Save
output = os.path.join(r'D:\openclaw-workspace\projects\护理学院医工设备采购\元数据及设备信息分析.xlsx')
wb.save(output)
print(f'Excel saved: {output}')
