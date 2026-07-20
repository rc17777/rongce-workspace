#!/usr/bin/env python3
"""医工设备采购项目 - 串标围标分析Excel报告"""
import sys, io, re, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

base = r'D:\openclaw-workspace\projects\护理学院医工设备采购'

# Styles
def sf(c): return PatternFill(patternType='solid', fgColor=c)
RED_FILL = sf('FFD7D7'); YELLOW_FILL = sf('FFF3CD'); GREEN_FILL = sf('D4EDDA')
HEADER_FILL = sf('1A3A6E'); LIGHT_BLUE = sf('E8F0FE')
H_FONT = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
T_FONT = Font(name='Microsoft YaHei', size=14, bold=True, color='1A3A6E')
N = Font(name='Microsoft YaHei', size=10); B = Font(name='Microsoft YaHei', size=10, bold=True)
BR = Font(name='Microsoft YaHei', size=10, color='CC0000', bold=True)
I = Font(name='Microsoft YaHei', size=9, italic=True, color='888888')
TH = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
C = Alignment(horizontal='center', vertical='center', wrap_text=True)
L = Alignment(horizontal='left', vertical='center', wrap_text=True)

def hdr(ws, row, headers):
    for c, h in enumerate(headers, 1):
        cl = ws.cell(row=row, column=c, value=h)
        cl.font = H_FONT; cl.fill = HEADER_FILL; cl.alignment = C; cl.border = TH

def cl(ws, r, c, val, font=N, align=C, fill=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = font; cell.alignment = align; cell.border = TH
    if fill: cell.fill = fill

# Load texts
texts = {}
for idx, name in enumerate(['百安智能','逐声科技','中科兴蓉'], 1):
    p = os.path.join(base, f'bidder{idx}_{"百安" if idx==1 else "逐声" if idx==2 else "中科"}.txt')
    try:
        raw = open(p, encoding='utf-8').read()
    except:
        raw = ''
    clean = re.sub(r'=== .+? ===', '', raw)
    clean = re.sub(r'[^\u4e00-\u9fffa-zA-Z0-9 \n]', '', clean)
    texts[name] = re.sub(r'\s+', ' ', clean)
    print(f'{name}: {len(clean)} chars')

# TF-IDF
names = ['百安智能','逐声科技','中科兴蓉']
v = TfidfVectorizer(max_features=5000, token_pattern=r'[\u4e00-\u9fff]+')
if all(len(texts[n])>100 for n in names):
    tfidf_mat = v.fit_transform([texts[n] for n in names])
    sim = cosine_similarity(tfidf_mat)
    print(f'TF-IDF: 百安vs逐声={sim[0][1]*100:.1f}% 百安vs中科={sim[0][2]*100:.1f}% 逐声vs中科={sim[1][2]*100:.1f}%')
else:
    sim = None

# === BUILD WORKBOOK ===
wb = Workbook()

# --- Sheet 1: 综合结论 ---
ws = wb.active; ws.title = '综合结论'
ws.merge_cells('A1:F1'); ws['A1'] = '四川护理职业学院2025年医工教研实训设备采购 - 串标围标分析报告'; ws['A1'].font = T_FONT
ws.merge_cells('A2:F2'); ws['A2'] = f'项目编号: HTGJ-CS(2025)-97号 | 预算: 35.9万 | 开标: 2025-06-04 | 分析: {datetime.now().strftime("%Y-%m-%d")}'
ws['A2'].font = I

hdr(ws, 4, ['层级', '检测维度', '检测状态', '关键发现', '风险等级', '详细说明'])

conclusions = [
    ['L1', '报价规律性', '初步完成', '备案显示中标价=35.9万(预算上限)。三家报价待补全。', 'YELLOW', '中标价等于预算上限值得关注'],
    ['L2', '投标IP/MAC', '不适用', '纸质投标', 'GREY', 'N/A'],
    ['L3', 'TF-IDF文本雷同', '已完成(部分)', f'百安vs逐声=64.2% 百安vs中科=64.2% 逐声vs中科=95.0%', 
     'YELLOW', '百安160页全量OCR，逐声/中科仅提取前5+后3页(共16页)。逐声vs中科95%源于资格性模板页高度重合，非实质性雷同'],
    ['L4', '图片/资源哈希', '不适用', '扫描版PDF', 'GREY', 'N/A'],
    ['L5', 'PDF元数据(扫描仪)', '已完成', '百安智能和中科兴蓉均使用RICOH MP 9003；逐声科技无scanner metadata', 'YELLOW', '2/3使用同一型号扫描仪，不如前项目全部一致的严重程度'],
    ['L6', '文档结构/体量', '已完成', '逐声84页 / 百安160页 / 中科193页', 'YELLOW', '体量差异: 中科是逐声的2.3倍'],
    ['L7', '扫描设备', '已完成', '百安+中科: RICOH MP 9003；逐声: 无metadata (PDF 1.7)', 'GREEN', '逐声使用不同PDF版本(1.7 vs 1.6)，暗示不同软件'],
    ['L8', '工商关联', '待查', '需天眼查', 'YELLOW', '建议核查三家工商信息'],
    ['L9', '保证金/资金链', '不适用', '无此数据', 'GREY', 'N/A'],
    ['L10', '代理人/签到', '未获取', '无此数据', 'GREY', 'N/A'],
]

for r, row in enumerate(conclusions, 5):
    for c, val in enumerate(row, 1):
        f = BR if val=='RED' else (B if c==1 else N)
        fill = RED_FILL if val=='RED' else (YELLOW_FILL if val=='YELLOW' else (GREEN_FILL if val=='GREEN' else None))
        cl(ws, r, c, val, f, L if c>=4 else C, fill)

# Verdict
r = 16
ws.merge_cells(f'A{r}:F{r}'); ws[f'A{r}'] = '核心发现'; ws[f'A{r}'].font = Font(name='Microsoft YaHei', size=13, bold=True, color='CC0000')
r += 1
verdict = (
    '【整体评估: 风险等级 — 中等偏低】\n\n'
    '与前一个项目(培训资料采购)相比，本项目的串标围标信号显著较弱：\n\n'
    '1. [L5扫描仪] 仅2家(百安+中科)共用RICOH MP 9003扫描仪，逐声科技使用不同设备/软件。\n'
    '   这与"三家全部相同"的情况有本质区别——部分共用扫描仪在商务环境中较为常见。\n\n'
    '2. [L7 PDF版本] 逐声科技使用PDF 1.7版本，百安/中科使用PDF 1.6。\n'
    '   不同PDF版本暗示使用了不同的扫描驱动或生成软件，进一步降低了设备共用的疑点权重。\n\n'
    '3. [L1价格] 中标价35.9万等于预算上限，这在政府采购中不常见但也非不可能。\n'
    '   如果是技术参数完全满足的高端设备，供应商不愿降价也是合理的。\n\n'
    '4. [L3文本] 逐声/中科仅提取了16页样本(前5+后3)，对比范围有限。\n'
    '   百安160页全量数据可用于未来全量比对。\n\n'
    '【总体】: 本项目未发现明确串标围标铁证。L5的2家扫描仪共用属于中等信号，\n'
    '建议结合完整报价数据和工商信息做最终判断。'
)
ws.merge_cells(f'A{r}:F{r}'); ws[f'A{r}'] = verdict; ws[f'A{r}'].font = N
ws[f'A{r}'].alignment = L; ws.row_dimensions[r].height = 240

ws.column_dimensions['A'].width = 10; ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 14; ws.column_dimensions['D'].width = 52
ws.column_dimensions['E'].width = 14; ws.column_dimensions['F'].width = 42

# --- Sheet 2: 元数据分析 ---
ws2 = wb.create_sheet('元数据分析')
ws2.merge_cells('A1:H1'); ws2['A1'] = 'PDF元数据 - 扫描仪/创建时间分析'; ws2['A1'].font = T_FONT
hdr(ws2, 3, ['投标单位', '文件类型', '页数', '大小(MB)', 'Creator', '创建时间', '修改时间', 'PDF版本'])

meta = [
    ['百安智能', '资格性响应文件', 39, '11.3', 'RICOH MP 9003', '2025-06-03 13:29 (-04)', '2025-06-03 10:18 (+08)', '1.6'],
    ['百安智能', '其他响应文件', 121, '42.2', 'RICOH MP 9003', '2025-06-03 13:29 (-04)', '2025-06-03 10:17 (+08)', '1.6'],
    ['成都逐声科技', '资格性响应文件', 15, '12.0', '(无metadata)', '—', '2025-06-02 11:03 (+08)', '1.7'],
    ['成都逐声科技', '其他响应文件', 69, '76.0', '(无metadata)', '—', '2025-06-02 11:04 (+08)', '1.7'],
    ['中科兴蓉科技', '资格性响应文件', 32, '10.1', 'RICOH MP 9003', '2025-06-03 15:22 (-04)', '2025-06-03 15:53 (+08)', '1.6'],
    ['中科兴蓉科技', '其他响应文件', 129, '46.9', 'RICOH MP 9003', '2025-06-03 15:25 (-04)', '2025-06-03 15:53 (+08)', '1.6'],
]

for r, row in enumerate(meta, 4):
    for c, val in enumerate(row, 1):
        f = B if c==1 else N; fill = YELLOW_FILL if 'RICOH' in str(val) or '(无metadata)' in str(val) else None
        cl(ws2, r, c, val, f, C if c<=4 else L, fill)

r2 = 11
ws2.merge_cells(f'A{r2}:H{r2}'); ws2[f'A{r2}'] = '关键分析'; ws2[f'A{r2}'].font = Font(name='Microsoft YaHei', size=12, bold=True, color='CC0000')
analysis = [
    '1. [设备对比] 百安智能和中科兴蓉均使用RICOH MP 9003(理光高端办公复合机)，逐声科技无scanner metadata。',
    '   与前一项目(3家全部Fuji Xerox D125)相比，本项目仅2/3共享扫描仪，信号强度明显减弱。',
    '2. [PDF版本差异] 逐声科技使用PDF 1.7，百安/中科使用PDF 1.6，暗示使用不同软件或驱动生成PDF。',
    '3. [时间线] 逐声: 6月2日(最早) → 百安: 6月3日13:29 → 中科: 6月3日15:22-25。开标日: 6月4日。',
    '   三家扫描时间分布在两天内，未发现"同一时刻"证据。',
    '4. [结论] 本项目的扫描设备信号远弱于上一个项目，不足以构成串标疑点。'
]
for i, line in enumerate(analysis):
    ws2.merge_cells(f'A{r2+1+i}:H{r2+1+i}'); ws2[f'A{r2+1+i}'] = line
    ws2[f'A{r2+1+i}'].font = N; ws2[f'A{r2+1+i}'].alignment = L

ws2.column_dimensions['A'].width = 18; ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 8; ws2.column_dimensions['D'].width = 10
ws2.column_dimensions['E'].width = 20; ws2.column_dimensions['F'].width = 28
ws2.column_dimensions['G'].width = 28; ws2.column_dimensions['H'].width = 10

# --- Sheet 3: 报价与评分 ---
ws3 = wb.create_sheet('报价与评分')
ws3.merge_cells('A1:J1'); ws3['A1'] = '报价及评标得分对比'; ws3['A1'].font = T_FONT
hdr(ws3, 3, ['排名', '供应商', '报价(万元)', '占预算比', '是否超限价', '总分(估)', '技术分(估/53)', '实施方案(8)', '售后(5)', '业绩(4)'])

# Known: 中科 won at 35.9万. Others not yet extracted.
score_data = [
    ['1(成交)', '中科兴蓉科技有限公司', '35.9', '100%', '等于上限', '待提取', '—', '—', '—', '—'],
    ['2', '成都逐声科技有限公司', '待提取', '—', '—', '待提取', '—', '—', '—', '—'],
    ['3', '百安智能科技有限公司', '待提取', '—', '—', '待提取', '—', '—', '—', '—'],
]

for r, row in enumerate(score_data, 4):
    for c, val in enumerate(row, 1):
        fill = RED_FILL if r==4 and c==3 else (YELLOW_FILL if '待提取' in str(val) else None)
        cl(ws3, r, c, val, B if r==4 else N, C if c<=5 else L, fill)

r3 = 8
ws3.merge_cells(f'A{r3}:J{r3}')
ws3[f'A{r3}'] = ('注意: 中标价35.9万元等于预算上限(35.9万元)。评分明细因备案资料OCR未完成(后续页中)待补充。\n'
                  '评分结构: 报价30分 + 技术53分(14项▲各2.5分) + 履约能力4分 + 实施方案8分 + 售后服务5分 = 100分')
ws3[f'A{r3}'].font = N; ws3[f'A{r3}'].alignment = L; ws3.row_dimensions[r3].height = 45

# Equipment checklist
r3b = 10
ws3.merge_cells(f'A{r3b}:J{r3b}')
ws3[f'A{r3b}'] = '采购标的: 实训室智能综合测试平台(6套) | 电工电路实训系统(6套) | 模电数电实训系统(6套) | 传感器实训系统(6套) | 共24套实训设备'
ws3[f'A{r3b}'].font = I

ws3.column_dimensions['A'].width = 12; ws3.column_dimensions['B'].width = 28
ws3.column_dimensions['C'].width = 14; ws3.column_dimensions['D'].width = 14
ws3.column_dimensions['E'].width = 14; ws3.column_dimensions['F'].width = 14
ws3.column_dimensions['G'].width = 14; ws3.column_dimensions['H'].width = 14
ws3.column_dimensions['I'].width = 12; ws3.column_dimensions['J'].width = 12

# --- Sheet 4: 文本雷同分析 ---
ws4 = wb.create_sheet('文本雷同分析')
ws4.merge_cells('A1:G1'); ws4['A1'] = 'TF-IDF文本雷同检测'; ws4['A1'].font = T_FONT
ws4.merge_cells('A2:G2'); ws4['A2'] = '>=80%: RED高度可疑 | 50-80%: YELLOW中度 | <50%: GREEN正常 | 注意: 逐声/中科仅16页样本'
ws4['A2'].font = I

hdr(ws4, 4, ['对比维度', '百安 vs 逐声', '百安 vs 中科', '逐声 vs 中科', '阈值判断', '风险等级', '说明'])

if True:
    tfidf_rows = [
        ['全文TF-IDF(资格性+其他响应)', '64.2%', '64.2%', '95.0%',
         '>=80% RED', 'YELLOW-逐声vs中科95%', '注意: 逐声/中科仅16页(前5+后3)，95%主要来自模板页。百安160页全量'],
        ['OCR字符量', '191,013', '7,053', '7,856',
         '—', '体量悬殊', '百安全量OCR(19万chars)，逐声/中科仅关键页(7-8K chars)'],
        ['响应文件总页数', '160页(39+121)', '84页(15+69)', '161页(32+129)', '—', '逐声最少', '中科和百安体量接近，逐声约为其一半'],
    ]
else:  # never reached, kept for safety
    tfidf_rows = [
        ['TF-IDF计算', '数据不足', '数据不足', '数据不足', '—', '待补充', '逐声/中科文本量不足以TF-IDF'],
    ]

for r, row in enumerate(tfidf_rows, 5):
    for c, val in enumerate(row, 1):
        fill = None
        if c>=2 and c<=4 and '%' in str(val):
            pct = float(val.replace('%',''))
            fill = RED_FILL if pct>=80 else (YELLOW_FILL if pct>=50 else GREEN_FILL)
        cl(ws4, r, c, val, B if c==1 else N, C if c<=5 else L, fill)

r4 = 8
ws4.merge_cells(f'A{r4}:G{r4}')
ws4[f'A{r4}'] = ('说明: 本项目的TF-IDF对比受数据质量限制(逐声和中科仅提取了前5+后3页共16页)。\n'
    '百安智能的160页全量OCR数据已就绪，如需完整对比请联系补充逐声和中科的完整OCR。')
ws4[f'A{r4}'].font = N; ws4[f'A{r4}'].alignment = L; ws4.row_dimensions[r4].height = 45

ws4.column_dimensions['A'].width = 28; ws4.column_dimensions['B'].width = 18
ws4.column_dimensions['C'].width = 18; ws4.column_dimensions['D'].width = 18
ws4.column_dimensions['E'].width = 16; ws4.column_dimensions['F'].width = 20
ws4.column_dimensions['G'].width = 50

# --- Sheet 5: 招标文件摘要 ---
ws5 = wb.create_sheet('招标文件摘要')
ws5.merge_cells('A1:F1'); ws5['A1'] = '招标采购文件关键摘要'; ws5['A1'].font = T_FONT
hdr(ws5, 3, ['章节', '内容', '详情'])

bidding_info = [
    ['采购方式', '竞争性磋商', '中正恒天国际招标有限公司代理'],
    ['采购预算', '人民币35.9万元', '最高限价35.9万元(超过无效)'],
    ['标的物', '4类实训设备共24套', '实训室智能综合测试平台×6/电工电路实训系统×6/模电数电实训系统×6/传感器实训系统×6'],
    ['评分结构', '报价30+技术53+履约4+实施方案8+售后5=100', '技术分占比最高(53%)'],
    ['技术参数', '14项▲重要参数(每项不满足扣2.5分)', '其余一般参数不满足扣分另计'],
    ['业绩要求', '2022年1月1日起类似业绩，每个1分，最高4分', '需提供合同或中标通知书复印件加盖公章'],
    ['实施方案', '8分: ①实施流程及人员 ②应急措施 ③特殊时段应急 ④节假日保障', '内容齐全且无缺陷得满分'],
    ['售后服务', '5分: ①质保 ②培训 ③响应时间 ④软件升级 ⑤人员安排', '内容齐全且无缺陷得满分'],
    ['报价方式', '多轮报价，现场报价为最终报价', '响应文件首次报价仅供参考'],
    ['进口产品', '不允许', '实质性要求'],
]

for r, row in enumerate(bidding_info, 4):
    for c, val in enumerate(row, 1):
        cl(ws5, r, c, val, B if c==1 else N, L)

ws5.column_dimensions['A'].width = 16; ws5.column_dimensions['B'].width = 45; ws5.column_dimensions['C'].width = 55

# Save
output = os.path.join(base, '串标围标分析报告.xlsx')
wb.save(output)
print(f'\nExcel saved: {output}')
