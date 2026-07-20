#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""阿坝发展控股集团 - 审计评估收费测算表（最终版）"""
import sys
sys.stdout.reconfigure(encoding='utf-8-sig')
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DARK_BLUE='0A1F3F';TEAL='1A5C6E';GOLD='C5955C';WARM_GRAY='F5F2EC';LIGHT_GOLD='F5E6C8'
tb=Border(left=Side(style='thin',color='999999'),right=Side(style='thin',color='999999'),top=Side(style='thin',color='999999'),bottom=Side(style='thin',color='999999'))
tf=Font(name='微软雅黑',size=14,bold=True,color=DARK_BLUE)
hf=Font(name='微软雅黑',size=10,bold=True,color='FFFFFF')
sf=Font(name='微软雅黑',size=10,bold=True,color=DARK_BLUE)
nf=Font(name='微软雅黑',size=10,color='333333')
mf=Font(name='微软雅黑',size=9,color='666666')
rf=Font(name='微软雅黑',size=11,bold=True,color=DARK_BLUE)
gf=Font(name='微软雅黑',size=10,bold=True,color=GOLD)
rrf=Font(name='微软雅黑',size=12,bold=True,color='CC0000')
hfl=PatternFill(start_color=DARK_BLUE,end_color=DARK_BLUE,fill_type='solid')
tfl=PatternFill(start_color=TEAL,end_color=TEAL,fill_type='solid')
wfl=PatternFill(start_color=WARM_GRAY,end_color=WARM_GRAY,fill_type='solid')
lgfl=PatternFill(start_color=LIGHT_GOLD,end_color=LIGHT_GOLD,fill_type='solid')
ca=Alignment(horizontal='center',vertical='center',wrap_text=True)
ra=Alignment(horizontal='right',vertical='center',wrap_text=True)
def setw(ws,w):
    for i,v in enumerate(w,1):ws.column_dimensions[get_column_letter(i)].width=v
def bdr(ws,r1,r2,c1,c2):
    for r in range(r1,r2+1):
        for c in range(c1,c2+1):ws.cell(row=r,column=c).border=tb

# 审计费率（‰千分比，按征求意见文件）
AR = [('1','0-100万元',2.5,100),('2','100-500万元',1.5,400),('3','500-1,000万元',0.8,500),
      ('4','1,000-5,000万元',0.4,4000),('5','5,000万-1亿',0.3,5000),
      ('6','1亿-5亿',0.2,40000),('7','5亿-10亿',0.15,50000),('8','>10亿',0.1,None)]
def calc_audit(w):
    dets,tot,rem=[],0.0,w*10000
    for i,n,r,s in AR:
        s2=s if s is None else min(max(rem,0),s*10000)
        if s is None:s2=max(rem,0)
        fee=s2*r/1000;dets.append((i,n,r,s2,fee));tot+=fee;rem-=s2
    if tot<2000:tot=2000
    return dets,tot

# 评估费率（川评协[2017]23号）
AP = [('1','100以下',15.0,100),('2','100-1000',6.25,900),('3','1000-5000',2.0,4000),
      ('4','5000-10000',1.2,5000),('5','10000-100000',0.25,90000),('6','100000以上',0.15,None)]
def calc_appr(w):
    dets,tot,rem=[],0.0,w*10000
    for i,n,r,s in AP:
        s2=s if s is None else min(max(rem,0),s*10000)
        if s is None:s2=max(rem,0)
        fee=s2*r/1000;dets.append((i,n,r,s2,fee));tot+=fee;rem-=s2
    if tot<2000:tot=2000
    return dets,tot

# 读取数据
src = r'C:\Users\scrccpa\Desktop\阿坝州-汇总面积(1).xlsx'
ws = load_workbook(src).active
props = []
for row in ws.iter_rows(min_row=2, max_row=11, values_only=True):
    o,n,a,b,_ = row
    props.append((o or '', n or '', a or 0, b or 0))

ta = sum(p[2] for p in props)
tbv = sum(p[3] for p in props)
pf = [(o,n,a,b,calc_appr(b)[1]) for o,n,a,b in props]
taf = sum(f for _,_,_,_,f in pf)

# 审计费
total_assets = 12091490505.09
net_assets = 8274567855.20
revenue = 940055809.15
wan = lambda v: v/10000
d_ta,b_ta=calc_audit(wan(total_assets))
d_na,b_na=calc_audit(wan(net_assets))
d_re,b_re=calc_audit(wan(revenue))

# 生成Excel
owb = Workbook()

# === Sheet 1: 项目概况 ===
ws1=owb.active;ws1.title='项目概况';setw(ws1,[5,55,30])
lines = [
    ('', '阿坝发展控股集团有限公司'),
    ('', '审计、评估项目收费测算表'),
    ('', ''),
    ('!', '一、项目基本信息'),
    ('', '  项目名称：阿坝发展控股集团有限公司2025年资产评估项目'),
    ('', '          及2025年度与2026年指定一期财务报表审计项目'),
    ('', '  服务内容：控制价专项审核服务'),
    ('', '  甲方：阿坝发展控股集团有限公司'),
    ('', '  乙方：四川融策会计师事务所有限公司'),
    ('', ''),
    ('!', '二、审计收费依据'),
    ('', '  川发改价格[2013]901号 年报审计'),
    ('', ''),
    ('!', '三、评估收费依据'),
    ('', '  川评协[2017]23号'),
    ('', ''),
    ('!', '四、审计收费测算基础数据（2025年初）'),
    ('', '  总资产：12,091,490,505.09元（约120.91亿元）'),
    ('', '  净资产：8,274,567,855.20元（约82.75亿元）'),
    ('', '  收入：940,055,809.15元（约9.40亿元）'),
    ('', ''),
    ('!', '五、评估收费测算基础数据'),
    ('', '  评估基础资料：阿坝州-汇总面积(1).xlsx'),
    ('', f'  各项目账面价值合计：{tbv:,.2f}万元（约{tbv/10000:.2f}亿元）'),
    ('', ''),
    ('', '编制单位：四川融策会计师事务所有限公司'),
    ('', '编制日期：2026年7月'),
]
for i,(t,text) in enumerate(lines,1):
    c=ws1.cell(row=i,column=2,value=text)
    if i==2:c.font=tf
    elif t=='!':c.font=sf
    else:c.font=nf

# === Sheet 2: 审计收费计算 ===
ws2=owb.create_sheet('审计收费计算');setw(ws2,[5,18,16,14,14,14,14,16,14])
ws2.merge_cells('A1:I1');ws2.cell(1,1,'阿坝发展控股集团 - 审计收费测算表（川发改价格[2013]901号）').font=tf;ws2.cell(1,1).alignment=ca
ws2.cell(3,1,'项目名称：阿坝发展控股集团有限公司2025年度及2026年指定一期财务报表审计项目').font=sf
ws2.cell(4,1,'收费依据：川发改价格[2013]901号（‰千分比费率）').font=nf
ws2.cell(5,1,'计算方法：差额定率累进法').font=nf
ws2.cell(6,1,f'总资产：{total_assets:,.2f}元（≈{wan(total_assets):,.2f}万元）').font=nf

# 三指标对比
r=8;ws2.merge_cells(f'A{r}:I{r}');ws2.cell(r,1,'一、三指标收费对比').font=sf
r=9;hs=['序号','计费指标','计费基数(万元)','标准收费(元)','1.5倍(经责/合并)','2.0倍(经济案件)','1.2倍(其他)','大写金额','备注']
for c,h in enumerate(hs,1):c2=ws2.cell(r,c,h);c2.font=hf;c2.fill=hfl;c2.alignment=ca

metrics=[('总资产',wan(total_assets),b_ta,'年报审计常用'),('净资产',wan(net_assets),b_na,'参考对比'),('收入',wan(revenue),b_re,'参考对比')]
for i,(nm,wanv,b,note) in enumerate(metrics):
    rr=r+1+i
    ws2.cell(rr,1,i+1).font=nf;ws2.cell(rr,1).alignment=ca
    ws2.cell(rr,2,nm).font=sf;ws2.cell(rr,2).alignment=ca
    ws2.cell(rr,3,wanv).font=nf;ws2.cell(rr,3).alignment=ra;ws2.cell(rr,3).number_format='#,##0.00'
    ws2.cell(rr,4,b).font=rf;ws2.cell(rr,4).alignment=ra;ws2.cell(rr,4).number_format='#,##0.00'
    ws2.cell(rr,5,b*1.5).font=nf;ws2.cell(rr,5).alignment=ra;ws2.cell(rr,5).number_format='#,##0.00'
    ws2.cell(rr,6,b*2.0).font=nf;ws2.cell(rr,6).alignment=ra;ws2.cell(rr,6).number_format='#,##0.00'
    ws2.cell(rr,7,b*1.2).font=nf;ws2.cell(rr,7).alignment=ra;ws2.cell(rr,7).number_format='#,##0.00'
    ws2.cell(rr,8,f'{b:,.2f}元').font=mf;ws2.cell(rr,9,note).font=mf
    if i==0:
        for c in range(1,10):ws2.cell(rr,c).fill=PatternFill(start_color='FFE0E0',end_color='FFE0E0',fill_type='solid')
    elif i%2==0:
        for c in range(1,10):ws2.cell(rr,c).fill=wfl
bdr(ws2,9,r+2,1,9)

# 各指标明细
r2=r+4
for nm,wanv,det in [('总资产',wan(total_assets),d_ta),('净资产',wan(net_assets),d_na),('收入',wan(revenue),d_re)]:
    b = sum(f for _,_,_,_,f in det)
    r2+=1;ws2.merge_cells(f'A{r2}:I{r2}');ws2.cell(r2,1,f'二、按{nm}差额定率累进计算').font=sf
    r2+=1;ws2.cell(r2,1,f'计费基数：{wanv:,.2f}万元').font=nf
    r2+=1;hs=['档次','计费区间（万元）','费率(‰)','计费额(万元)','费率(%)','收费额(万元)','累计收费(万元)','收费额(元)','计算公式']
    for c,h in enumerate(hs,1):c2=ws2.cell(r2,c,h);c2.font=hf;c2.fill=tfl;c2.alignment=ca
    cum=0
    for i,(idx,n2,rate,seg,fee) in enumerate(det):
        rr=r2+1+i
        ws2.cell(rr,1,idx).font=nf;ws2.cell(rr,1).alignment=ca;ws2.cell(rr,2,n2).font=nf
        ws2.cell(rr,3,f'{rate:.2f}').font=nf;ws2.cell(rr,3).alignment=ca
        ws2.cell(rr,4,seg/10000).font=nf;ws2.cell(rr,4).alignment=ra;ws2.cell(rr,4).number_format='#,##0'
        ws2.cell(rr,5,rate/1000 if seg>0 else 0).font=nf;ws2.cell(rr,5).alignment=ra;ws2.cell(rr,5).number_format='0.0000%'
        ws2.cell(rr,6,fee/10000).font=nf;ws2.cell(rr,6).alignment=ra;ws2.cell(rr,6).number_format='#,##0.0000'
        cum+=fee;ws2.cell(rr,7,cum/10000).font=nf;ws2.cell(rr,7).alignment=ra;ws2.cell(rr,7).number_format='#,##0.0000'
        ws2.cell(rr,8,fee).font=nf;ws2.cell(rr,8).alignment=ra;ws2.cell(rr,8).number_format='#,##0.00'
        ws2.cell(rr,9,f'{seg/10000:,.2f}万×{rate}‰').font=Font(name='微软雅黑',size=8,color='666666')
        if i%2==0:
            for c in range(1,10):ws2.cell(rr,c).fill=wfl
    tr=r2+1+len(AR)
    ws2.merge_cells(f'A{tr}:D{tr}');ws2.cell(tr,1,'合计').font=gf;ws2.cell(tr,1).alignment=ra
    ws2.cell(tr,6,b/10000).font=gf;ws2.cell(tr,6).alignment=ra;ws2.cell(tr,6).number_format='#,##0.0000'
    ws2.cell(tr,8,b).font=rrf;ws2.cell(tr,8).alignment=ra;ws2.cell(tr,8).number_format='#,##0.00'
    ws2.cell(tr,9,'标准收费').font=mf
    for c in range(1,10):ws2.cell(tr,c).fill=lgfl
    bdr(ws2,r2,tr,1,9);r2=tr+1

# === Sheet 3: 评估收费计算 ===
ws3=owb.create_sheet('评估收费计算');setw(ws3,[5,22,14,14,14,14,14,18])
ws3.merge_cells('A1:H1');ws3.cell(1,1,'阿坝发展控股集团 - 评估收费测算表（川评协[2017]23号）').font=tf;ws3.cell(1,1).alignment=ca
ws3.cell(2,1,f'测算基础：阿坝州-汇总面积(1).xlsx（账面价值合计{tbv:,.2f}万元）').font=sf

r=4;ws3.merge_cells(f'A{r}:H{r}');ws3.cell(r,1,'一、差额定率累进计费标准（川评协[2017]23号）').font=sf
r=5;hs=['档次','计费额度（万元）','差额计费率（千分比）','计费额(万元)','费率(%)','收费额(万元)','累计收费(万元)','备注']
for c,h in enumerate(hs,1):c2=ws3.cell(r,c,h);c2.font=hf;c2.fill=hfl;c2.alignment=ca
d_ap,b_ap=calc_appr(tbv);cum=0
for i,(idx,name,rate,seg,fee) in enumerate(d_ap):
    rr=r+1+i
    ws3.cell(rr,1,idx).font=nf;ws3.cell(rr,1).alignment=ca;ws3.cell(rr,2,name).font=nf
    ws3.cell(rr,3,f'{rate:.2f}‰').font=nf;ws3.cell(rr,3).alignment=ca
    ws3.cell(rr,4,seg/10000).font=nf;ws3.cell(rr,4).alignment=ra;ws3.cell(rr,4).number_format='#,##0'
    ws3.cell(rr,5,rate/1000 if seg>0 else 0).font=nf;ws3.cell(rr,5).alignment=ra;ws3.cell(rr,5).number_format='0.0000%'
    ws3.cell(rr,6,fee/10000).font=nf;ws3.cell(rr,6).alignment=ra;ws3.cell(rr,6).number_format='#,##0.0000'
    cum+=fee;ws3.cell(rr,7,cum/10000).font=nf;ws3.cell(rr,7).alignment=ra;ws3.cell(rr,7).number_format='#,##0.0000'
    ws3.cell(rr,8,'').font=mf
    if i%2==0:
        for c in range(1,9):ws3.cell(rr,c).fill=wfl
tr=r+1+len(AP);ws3.merge_cells(f'A{tr}:D{tr}')
ws3.cell(tr,1,'合计（标准收费/万元）').font=gf;ws3.cell(tr,1).alignment=ra
ws3.cell(tr,6,b_ap/10000).font=rf;ws3.cell(tr,6).alignment=ra;ws3.cell(tr,6).number_format='#,##0.0000'
for c in range(1,9):ws3.cell(tr,c).fill=lgfl
bdr(ws3,r,tr,1,8)

rr=tr+2;ws3.merge_cells(f'A{rr}:H{rr}');ws3.cell(rr,1,'二、评估项目明细（按账面价值计算）').font=sf
rr+=1;hs=['权利人','项目名称','面积(m2)','账面价值(万元)','评估收费(元)','备注']
for c,h in enumerate(hs,1):c2=ws3.cell(rr,c,h);c2.font=hf;c2.fill=hfl;c2.alignment=ca
po=''
for i,(o,n,a,b,f) in enumerate(pf):
    rr2=rr+1+i
    dn=o if o and o!=po else ''
    ws3.cell(rr2,1,dn).font=nf;ws3.cell(rr2,2,n).font=nf
    ws3.cell(rr2,3,a).font=nf;ws3.cell(rr2,3).alignment=ra;ws3.cell(rr2,3).number_format='#,##0.00'
    ws3.cell(rr2,4,b).font=nf;ws3.cell(rr2,4).alignment=ra;ws3.cell(rr2,4).number_format='#,##0.00'
    ws3.cell(rr2,5,f).font=nf;ws3.cell(rr2,5).alignment=ra;ws3.cell(rr2,5).number_format='#,##0.00'
    ws3.cell(rr2,6,'').font=mf;po=o
tr=rr+1+len(pf);ws3.merge_cells(f'A{tr}:B{tr}')
ws3.cell(tr,1,'合计').font=gf;ws3.cell(tr,1).alignment=ra
ws3.cell(tr,3,ta).font=gf;ws3.cell(tr,3).alignment=ra;ws3.cell(tr,3).number_format='#,##0.00'
ws3.cell(tr,4,tbv).font=gf;ws3.cell(tr,4).alignment=ra;ws3.cell(tr,4).number_format='#,##0.00'
ws3.cell(tr,5,taf).font=gf;ws3.cell(tr,5).alignment=ra;ws3.cell(tr,5).number_format='#,##0.00'
ws3.cell(tr,6,'').font=mf
for c in range(1,7):ws3.cell(tr,c).fill=lgfl
bdr(ws3,rr,tr,1,6)

# === Sheet 4: 费用汇总 ===
ws4=owb.create_sheet('费用汇总');setw(ws4,[5,22,16,18,18,18,18,18,18])
ws4.merge_cells('A1:I1');ws4.cell(1,1,'阿坝发展控股集团 - 审计评估收费汇总表').font=tf;ws4.cell(1,1).alignment=ca
r=3;hs=['项目','计费指标','计费基数(万元)','标准收费(元)','1.5倍(经责/合并)','2.0倍(经济案件)','1.2倍(其他)','大写金额','备注']
for c,h in enumerate(hs,1):c2=ws4.cell(r,c,h);c2.font=hf;c2.fill=hfl;c2.alignment=ca

for i,(nm,wanv,b,note) in enumerate(metrics):
    rr=r+1+i
    ws4.cell(rr,1,'审计收费').font=sf
    ws4.cell(rr,2,nm).font=nf;ws4.cell(rr,2).alignment=ca
    ws4.cell(rr,3,wanv).font=nf;ws4.cell(rr,3).alignment=ra;ws4.cell(rr,3).number_format='#,##0.00'
    ws4.cell(rr,4,b).font=rf;ws4.cell(rr,4).alignment=ra;ws4.cell(rr,4).number_format='#,##0.00'
    ws4.cell(rr,5,b*1.5).font=nf;ws4.cell(rr,5).alignment=ra;ws4.cell(rr,5).number_format='#,##0.00'
    ws4.cell(rr,6,b*2.0).font=nf;ws4.cell(rr,6).alignment=ra;ws4.cell(rr,6).number_format='#,##0.00'
    ws4.cell(rr,7,b*1.2).font=nf;ws4.cell(rr,7).alignment=ra;ws4.cell(rr,7).number_format='#,##0.00'
    ws4.cell(rr,8,f'{b:,.2f}元').font=mf;ws4.cell(rr,9,note).font=mf
    if i%2==0:
        for c in range(1,10):ws4.cell(rr,c).fill=wfl

# 评估行
rr=r+1+len(metrics)
ws4.cell(rr,1,'评估收费').font=sf
ws4.cell(rr,2,'账面价值').font=nf;ws4.cell(rr,2).alignment=ca
ws4.cell(rr,3,tbv).font=nf;ws4.cell(rr,3).alignment=ra;ws4.cell(rr,3).number_format='#,##0.00'
ws4.cell(rr,4,taf).font=rf;ws4.cell(rr,4).alignment=ra;ws4.cell(rr,4).number_format='#,##0.00'
ws4.cell(rr,5,taf*1.5).font=nf;ws4.cell(rr,5).alignment=ra;ws4.cell(rr,5).number_format='#,##0.00'
ws4.cell(rr,6,taf*2.0).font=nf;ws4.cell(rr,6).alignment=ra;ws4.cell(rr,6).number_format='#,##0.00'
ws4.cell(rr,7,taf*1.2).font=nf;ws4.cell(rr,7).alignment=ra;ws4.cell(rr,7).number_format='#,##0.00'
ws4.cell(rr,8,f'{taf:,.2f}元').font=mf;ws4.cell(rr,9,'川评协[2017]23号').font=mf
bdr(ws4,3,rr,1,9)

# 合计行
rr2=rr+1;ws4.cell(rr2,1,'合计（总资产+评估）').font=sf
ws4.cell(rr2,4,b_ta+taf).font=rrf;ws4.cell(rr2,4).alignment=ra;ws4.cell(rr2,4).number_format='#,##0.00'
ws4.cell(rr2,9,'审计+评估').font=mf
for c in range(1,10):ws4.cell(rr2,c).fill=lgfl

r=rr2+2;ws4.cell(r,1,'补充说明：').font=sf
notes=['1. 审计收费按总资产12,091,490,505.09元差额定率累进计算。','2. 评估收费按各项目账面价值合计，依据川评协[2017]23号计算。','3. 以上为标准收费，实际收费可上下浮动20%。','4. 最低收费2000元/项目，所有项目均超过此标准。','5. 异地项目额外收取交通费、住宿费。']
for i,note in enumerate(notes):ws4.cell(row=r+1+i,column=1,value=note).font=nf

# 保存
out = r'C:\Users\scrccpa\Desktop\阿坝发展控股集团-审计评估收费测算表_v3.xlsx'
owb.save(out)
print(f'OK: {out}')
print()
print('=== 审计收费 ===')
print(f'总资产: {wan(total_assets):,.2f}万元 -> 标准收费 {b_ta:,.2f}元')
print(f'净资产: {wan(net_assets):,.2f}万元 -> 标准收费 {b_na:,.2f}元')
print(f'收入:   {wan(revenue):,.2f}万元 -> 标准收费 {b_re:,.2f}元')
print()
print('=== 评估收费 ===')
print(f'账面价值合计: {tbv:,.2f}万元')
print(f'评估标准收费: {taf:,.2f}元')
print(f'审计+评估合计: {b_ta+taf:,.2f}元')
print()
print('=== 各项目明细 ===')
for o,n,a,b,f in pf:
    print(f'{n:20s}  面积{a:>8.2f}m2  账面{b:>10.2f}万  收费{f:>10.2f}元')