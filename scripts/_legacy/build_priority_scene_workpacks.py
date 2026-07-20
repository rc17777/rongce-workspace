#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Create workpacks for priority SOP scenes: data request letter, Excel working paper, report issue templates."""
from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit("openpyxl is required to build xlsx workpaper templates") from exc

V2_ROOT = Path(r"C:\Users\scrccpa\Documents\Obsidian Vault\审计案例库-OCR\融策标准作业体系 v2.0")
OUT_DIR = V2_ROOT / "08-重点场景作业包"

SCENES = {
    "绩效评价": {
        "slug": "01-绩效评价",
        "org": "项目主管部门/实施单位",
        "scope": "财政重点项目、部门整体支出、专项资金、政府购买服务、公共服务项目、成本预算绩效和事前绩效评估",
        "data_items": [
            ("政策制度", "项目相关政策文件、管理办法、实施方案、会议纪要、审批文件", "PDF/Word/扫描件", "确认目标依据、管理要求和职责边界"),
            ("立项预算", "项目申报书、可研/实施方案、绩效目标申报表、预算批复", "Excel/PDF", "核查目标设置、预算测算和立项论证"),
            ("资金支付", "资金下达文件、预算指标表、国库支付明细、银行流水、凭证附件", "Excel+PDF", "核查资金到位、支出进度和支出合规"),
            ("合同验收", "合同、采购文件、验收报告、服务记录、成果清单、资产移交资料", "Excel/PDF/图片", "核查产出真实性和合同履约"),
            ("绩效结果", "自评报告、监控记录、以前年度评价报告、整改台账", "Word/Excel", "核查结果应用和整改闭环"),
            ("外部反馈", "服务对象清单、满意度原始问卷、投诉举报、第三方监测数据", "Excel/原始记录", "核查效益和满意度真实性"),
        ],
        "sheets": {
            "绩效指标底稿": ["指标名称", "一级指标", "二级指标", "评价标准", "权重", "数据来源", "证据编号", "核查结果", "扣分", "偏差金额/数量", "原因分析", "责任主体", "整改建议"],
            "资金支付核验": ["项目名称", "预算金额", "指标下达日期", "支付日期", "支付对象", "支付金额", "合同/凭证编号", "验收节点", "是否异常", "异常说明", "证据编号"],
            "产出效益核验": ["项目名称", "承诺产出", "实际产出", "质量标准", "现场/资料核验结果", "服务对象反馈", "是否达标", "影响程度", "证据编号"],
            "满意度样本": ["样本编号", "服务对象类型", "联系方式/脱敏ID", "抽样方式", "调查日期", "问卷得分", "异常标记", "备注"],
            "问题汇总": ["问题类别", "问题事实", "涉及金额/数量", "定性依据", "证据索引", "责任主体", "整改建议", "报告引用段落"],
        },
        "issue_templates": [
            ("绩效目标设置不科学", "XX项目绩效目标设置不够科学，部分指标仅表述为“提升、完善、加强”等定性内容，缺少数量、质量、时效、成本等可衡量标准，导致后续绩效监控和评价缺乏明确标尺。", "建议主管部门按投入、产出、效益、满意度等维度重构绩效目标，建立可量化、可核验、可追踪的指标体系。"),
            ("预算安排与绩效目标脱节", "XX项目预算安排与绩效目标、任务量或成本标准匹配不足，预算测算依据不充分，存在资金安排偏高/偏低或项目任务无法支撑预算规模的情况。", "建议建立成本测算和同类项目比价机制，将绩效目标、预算安排和成本标准同步审核。"),
            ("资金支付与项目产出不匹配", "XX项目资金支付进度快于实际产出或验收进度，部分支出缺少对应服务记录、验收资料或成果证明，存在支付成果错配风险。", "建议严格按合同节点、验收结果和实际产出办理支付，完善支付前绩效监控。"),
            ("绩效评价结果应用不足", "XX项目以前年度评价发现的问题未有效整改，评价结果未与预算安排、项目调整、责任追究挂钩，存在评价和预算管理“两张皮”现象。", "建议建立评价结果与预算安排、项目退出、整改问责的联动机制。"),
        ],
    },
    "经济责任审计": {
        "slug": "02-经济责任审计",
        "org": "被审计单位/组织人事部门/主管部门",
        "scope": "党政机关、事业单位、国企、医院学校等单位主要领导干部任中、离任或专项经济责任审计",
        "data_items": [
            ("干部任职", "任免文件、任期起止、职责分工、领导班子分工、权责清单", "PDF/Word", "确认审计期间和责任边界"),
            ("议事决策", "党委会/党组会/办公会/董事会会议纪要、重大事项决策资料", "PDF/Word", "核查重大决策程序"),
            ("财政财务", "财务报表、预算决算、科目余额、序时账、银行流水、往来款明细", "Excel/PDF", "核查财政财务管理责任"),
            ("项目资产", "项目台账、采购合同、工程资料、资产资源台账、国资监管资料", "Excel/PDF", "核查资产资源和项目管理"),
            ("监督整改", "以前年度审计报告、巡视巡察反馈、整改台账、信访举报和问题线索", "Word/Excel", "核查整改责任和历史遗留"),
        ],
        "sheets": {
            "任期事项清单": ["事项名称", "事项类别", "发生期间", "涉及金额", "决策主体", "执行主体", "当前状态", "是否跨任期", "资料来源", "备注"],
            "重大决策核验": ["决策事项", "决策日期", "会议类型", "是否集体研究", "前置论证资料", "审批文件", "执行结果", "异常说明", "证据编号"],
            "财政财务问题": ["问题类别", "科目/账户", "发生日期", "金额", "摘要", "责任期间", "责任主体", "违反依据", "证据编号"],
            "责任界定底稿": ["问题事项", "事实摘要", "发生期间", "领导分工", "决策/审批行为", "责任类型", "责任依据", "证据链", "建议表述"],
            "整改跟踪": ["历史问题", "来源", "整改责任人", "整改期限", "整改措施", "核验结果", "是否销号", "未整改原因"],
        },
        "issue_templates": [
            ("重大事项决策程序不规范", "XX单位在XX事项决策过程中，未见充分的前置论证、集体研究或风险评估资料，相关决策程序不够规范，影响重大事项决策的科学性和合规性。", "建议完善重大事项决策程序，严格落实集体研究、合法合规审查和风险评估要求。"),
            ("历史遗留问题整改不到位", "XX问题在本任期内持续存在或进一步扩大，相关整改措施停留在台账层面，未形成实质性整改效果，反映整改责任压实不够。", "建议明确整改责任部门、责任人和时限，建立整改销号和回头看机制。"),
            ("财政财务管理责任落实不足", "XX单位在预算执行、资金支付、往来款清理或资产管理方面存在管理不规范问题，相关事项发生于被审计领导干部任期内，反映财政财务管理责任落实不足。", "建议加强预算、资金、资产全流程管理，定期开展风险排查和清理。"),
            ("责任边界划分不清", "XX问题跨多个任期或多个责任主体，单位未建立清晰的问题形成、扩大、整改时间线，导致责任界定依据不足。", "建议按事项建立时间线和责任链条，分清历史遗留、任期新增、整改不力等责任类型。"),
        ],
    },
    "医保资金审计": {
        "slug": "03-医保资金审计",
        "org": "医保部门/医保经办机构/定点医疗机构",
        "scope": "医保基金、定点医疗机构、药品耗材采购、医保支付方式改革、医疗服务收费和骗保套保专项审计",
        "data_items": [
            ("医保结算", "医保结算明细：参保人脱敏ID、机构、科室、医生、诊断、结算日期、基金支付、个人支付", "Excel/CSV", "核查基金支付和结算异常"),
            ("诊疗项目", "诊疗项目明细：项目编码、项目名称、数量、单价、金额、执行科室、执行时间", "Excel/CSV", "筛查重复收费、串换项目、过度诊疗"),
            ("药品耗材", "药品耗材明细：医保编码、规格、数量、单价、采购价、使用人、使用时间", "Excel/CSV", "核查药品耗材采购、使用、收费一致性"),
            ("病历资料", "病案首页、入出院记录、医嘱、检查检验报告、手术麻醉记录、护理记录", "系统导出/PDF", "对疑点样本进行病历回源"),
            ("监管规则", "定点协议、医保目录、收费标准、DRG/DIP分组结果、稽核处罚和投诉举报资料", "Excel/PDF", "确认违规依据和处理口径"),
        ],
        "sheets": {
            "医保结算疑点": ["疑点类型", "机构", "科室", "医生", "患者脱敏ID", "结算日期", "诊断", "基金支付", "个人支付", "规则命中原因", "证据编号"],
            "诊疗收费核验": ["患者脱敏ID", "就诊/住院号", "项目编码", "项目名称", "数量", "单价", "金额", "执行时间", "医嘱/报告是否匹配", "异常说明"],
            "药品耗材核验": ["机构", "药品/耗材编码", "名称", "规格", "采购数量", "出库数量", "收费数量", "差异数量", "差异金额", "异常说明"],
            "病历回源底稿": ["疑点编号", "患者脱敏ID", "病历资料", "医嘱核验", "报告核验", "护理记录核验", "专家/医保复核意见", "是否确认违规"],
            "违规金额测算": ["疑点类型", "机构", "项目/药品编码", "涉及人次", "涉及数量", "疑点金额", "确认违规金额", "拟追回金额", "处理建议"],
        },
        "issue_templates": [
            ("重复收费", "经筛查，XX医疗机构存在同一患者同日同项目重复收费或互斥项目同时收费的情况，涉及XX人次、疑点金额XX元。经病历和收费明细核验，部分收费缺少相应诊疗依据。", "建议医保部门按规定核实追回违规基金，并督促医疗机构完善收费审核规则。"),
            ("分解住院", "XX医疗机构部分患者短期内多次出入院，诊断相近、住院间隔异常短，存在分解住院、规避医保支付规则的疑点。", "建议结合病案首页、入出院记录和临床专家意见进一步核实，对确认违规部分依法依规处理。"),
            ("串换诊疗项目", "XX医疗机构存在实际执行项目与医保结算项目不一致的情况，部分低价项目疑似按高价医保项目结算，涉及医保基金支付风险。", "建议建立诊疗项目、医嘱、检查报告、收费编码的自动校验规则。"),
            ("药品耗材账实不符", "XX医疗机构药品耗材采购入库、库存出库、患者收费数量之间存在差异，部分高值耗材收费缺少出库或使用记录支撑。", "建议医保、卫健和医疗机构联合核查耗材流向，完善采购、库存、收费联动监管。"),
        ],
    },
}


def write_request_letter(scene: str, cfg: dict, out_dir: Path) -> Path:
    now = datetime.now().strftime("%Y年%m月%d日")
    lines = [
        f"# {scene}项目资料调取函模板",
        "",
        f"> 适用范围：{cfg['scope']}  ",
        f"> 接收单位：{cfg['org']}  ",
        f"> 生成日期：{now}",
        "",
        "## 一、调取目的",
        "",
        f"为开展{scene}相关审计/评价工作，核实项目管理、资金使用、业务执行和整改落实情况，请贵单位按本函要求提供真实、完整、准确的资料和数据。",
        "",
        "## 二、资料清单",
        "",
        "| 序号 | 资料类别 | 具体内容 | 建议格式 | 用途说明 | 是否提供 | 备注 |",
        "|---:|---|---|---|---|---|---|",
    ]
    for idx, item in enumerate(cfg["data_items"], start=1):
        category, content, fmt, purpose = item
        lines.append(f"| {idx} | {category} | {content} | {fmt} | {purpose} |  |  |")
    lines.extend([
        "",
        "## 三、数据要求",
        "",
        "- Excel/CSV 数据请保留原始字段名，不要合并单元格，不要只提供截图。",
        "- 涉及个人隐私或敏感信息的，应按审计需要提供脱敏字段，并保留可回源核验的映射机制。",
        "- 所有资料请注明来源系统、导出时间、经办人员和数据口径。",
        "- 如部分资料无法提供，请说明原因、替代资料和责任部门。",
        "",
        "## 四、提交方式",
        "",
        "请按资料类别建立文件夹，并以“序号-资料类别-单位名称”命名；电子资料通过安全介质或双方确认的安全方式移交。",
        "",
        "## 五、联系人",
        "",
        "- 审计组联系人：__________",
        "- 联系电话：__________",
        "- 资料提交截止时间：____年__月__日",
        "",
    ])
    path = out_dir / f"{cfg['slug']}-取数函模板.md"
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="0A1F3F")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="C5955C")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = max(len(str(ws.cell(row=row, column=col).value or "")) for row in range(1, min(ws.max_row, 20) + 1))
        ws.column_dimensions[letter].width = min(max(max_len + 4, 14), 35)


def write_workbook(scene: str, cfg: dict, out_dir: Path) -> Path:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for sheet_name, headers in cfg["sheets"].items():
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for _ in range(10):
            ws.append([""] * len(headers))
        style_sheet(ws)
    readme = wb.create_sheet("使用说明", 0)
    readme.append(["项目", "说明"])
    readme.append(["适用场景", scene])
    readme.append(["使用方法", "进场后按工作表逐项登记资料、疑点、证据和整改建议；不要删除字段，可按项目新增列。"])
    readme.append(["质量要求", "每个报告问题必须能在底稿中追溯到证据编号、金额/数量、责任主体和整改建议。"])
    style_sheet(readme)
    path = out_dir / f"{cfg['slug']}-Excel底稿字段表.xlsx"
    wb.save(path)
    return path


def write_issue_template(scene: str, cfg: dict, out_dir: Path) -> Path:
    lines = [
        f"# {scene}报告问题表述模板",
        "",
        f"> 适用范围：{cfg['scope']}",
        "",
        "## 使用说明",
        "",
        "- 以下表述是报告问题段落骨架，使用时必须替换单位名称、项目名称、金额、数量、期间和依据。",
        "- 每条问题必须配套底稿证据，不得直接复制模板上报告。",
        "- 建议采用“事实—影响—依据—建议”的四段式表达。",
        "",
    ]
    for idx, (title, fact, suggestion) in enumerate(cfg["issue_templates"], start=1):
        lines.extend([
            f"## {idx}. {title}",
            "",
            "### 问题表述",
            "",
            fact,
            "",
            "### 影响后果",
            "",
            "上述问题影响资金使用绩效、管理规范性和监管有效性，存在资金损失、低效使用或管理风险。具体影响金额/范围应根据底稿核实结果填列。",
            "",
            "### 整改建议",
            "",
            suggestion,
            "",
            "### 底稿支撑",
            "",
            "- 证据编号：__________",
            "- 涉及金额/数量：__________",
            "- 责任主体：__________",
            "- 定性依据：__________",
            "",
        ])
    path = out_dir / f"{cfg['slug']}-报告问题表述模板.md"
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    outputs = []
    for scene, cfg in SCENES.items():
        scene_dir = OUT_DIR / cfg["slug"]
        scene_dir.mkdir(parents=True, exist_ok=True)
        outputs.append(write_request_letter(scene, cfg, scene_dir))
        outputs.append(write_workbook(scene, cfg, scene_dir))
        outputs.append(write_issue_template(scene, cfg, scene_dir))
    index_lines = [
        "# 重点场景作业包索引",
        "",
        f"- 生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "- 覆盖场景：绩效评价、经济责任审计、医保资金审计",
        "- 每个场景包含：取数函模板、Excel底稿字段表、报告问题表述模板",
        "",
        "## 文件清单",
        "",
    ]
    for path in outputs:
        index_lines.append(f"- `{path}`")
    index = OUT_DIR / "00-重点场景作业包索引.md"
    index.write_text("\n".join(index_lines), encoding="utf-8")
    outputs.append(index)
    print(f"OUT_DIR {OUT_DIR}")
    print(f"FILES {len(outputs)}")
    for path in outputs:
        print(path)


if __name__ == "__main__":
    main()
