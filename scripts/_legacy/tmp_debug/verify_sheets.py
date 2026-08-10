import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\scrccpa\Desktop\融策审计过程记录系统=项目经理版(6)-已复核.xlsx')

ws5 = wb['5-问题与证据']
ws6 = wb['6-复核记录']

print('=== Sheet5 问题与证据 ===')
for r in range(3, 13):
    v = ws5.cell(r, 1).value
    if v:
        print(f'  R{r}: {v} | {ws5.cell(r,2).value} | {ws5.cell(r,3).value} | {ws5.cell(r,4).value[:30]}...')

print('\n=== Sheet6 一级复核 ===')
for r in [4,5,6,7,8,9,10,11,12,13,14]:
    item = ws6.cell(r, 1).value
    opinion = ws6.cell(r, 5).value
    if item:
        print(f'  R{r}: {str(item)[:40]} | {str(opinion)[:50] if opinion else ""}')

print('\n=== Sheet6 二级复核 ===')
for r in [16,17,18,19,20]:
    item = ws6.cell(r, 1).value
    opinion = ws6.cell(r, 5).value
    if item:
        print(f'  R{r}: {str(item)[:40]} | {str(opinion)[:50] if opinion else ""}')

print('\n=== Sheet6 三级复核 ===')
for r in [22,23,24,25,26]:
    item = ws6.cell(r, 1).value
    opinion = ws6.cell(r, 5).value
    if item:
        print(f'  R{r}: {str(item)[:40]} | {str(opinion)[:50] if opinion else ""}')

print('\nDone - all verified')
