# -*- coding: utf-8 -*-
"""Add compliance analysis sheet to the cost Excel"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

src_path = r'D:\openclaw-workspace\output\健康照护师-成本构成测算-v2-2026.05.21.xlsx'
wb = openpyxl.load_workbook(src_path)

# ============ Styles ============
title_font = Font(name='微软雅黑', size=16, bold=True)
h2_font = Font(name='微软雅黑', size=12, bold=True)
h3_font = Font(name='微软雅黑', size=11, bold=True)
normal_font = Font(name='微软雅黑', size=10)
bold_font = Font(name='微软雅黑', size=10, bold=True)
green_font = Font(name='微软雅黑', size=10, color='006100')
red_font = Font(name='微软雅黑', size=10, bold=True, color='CC0000')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
sub_header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header_row(ws, row, max_col):
    for c in range(1, max_col+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def sc(ws, row, col, align='center'):
    cell = ws.cell(row=row, column=col)
    cell.font = normal_font
    cell.alignment = center_align if align == 'center' else left_align
    cell.border = thin_border
    return cell

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# =====================================================
# Sheet 6: 收费合规性分析
# =====================================================
ws6 = wb.create_sheet('收费合规性分析')

ws6.merge_cells('A1:H1')
ws6['A1'].value = '健康照护师（长期照护师）职业技能等级认定 — 收费合规性分析'
ws6['A1'].font = title_font
ws6['A1'].alignment = center_align

# ---- Section 1: Per-item compliance ----
ws6.merge_cells('A3:H3')
ws6['A3'].value = '一、各项费用收取合规性逐项分析'
ws6['A3'].font = h2_font

row = 5
hdrs = ['序号', '费用项目', '测算金额\n(以高级为例)', '能否收取', '收费依据', '依据文号/条款', '关键限制', '风险提示']
for c, h in enumerate(hdrs, 1):
    ws6.cell(row=row, column=c, value=h)
style_header_row(ws6, row, len(hdrs))

items = [
    ['1', '理论考试费', '40元/人', '✅ 能收取',
     '政府定价的行政事业性收费，评价机构按此标准向考生收取',
     '川发改价格〔2017〕472号',
     '不得超过政府定价标准\nB类：初级30/中级35/高级40',
     '无风险，严格按定价执行即可'],
    ['2', '操作技能考核费', '240元/人', '✅ 能收取',
     '政府定价的行政事业性收费，评价机构按此标准向考生收取',
     '川发改价格〔2017〕472号',
     '不得超过政府定价标准\nB类：初级140/中级190/高级240',
     '无风险，严格按定价执行即可'],
    ['3', '考务平台费', '20元/人', '✅ 能收取',
     '评价机构可根据实际工作成本收取。考务平台费属信息化服务成本，是等级认定的必要支出',
     '川人社规〔2022〕9号\n第五条第(五)款',
     '按实际平台服务合同价格分摊\n不得以营利为目的加价',
     '需保留平台服务合同作为\n成本依据备查'],
    ['4', '人工费用\n(含SP模特)', '329.80元/人', '✅ 能收取',
     '考评员、监考、SP模特等人员劳务费是等级认定的核心工作成本。\n费率标准依据学院院长办公会纪要',
     '川人社规〔2022〕9号\n第五条第(五)款\n+ 院长办公会纪要\n(第23期)2022.12.12',
     '考评员100元/h、其他60元/h\nSP模特：老师60/h、学生22/h\n按实际工作时长据实测算',
     '人员配置和工时需有\n考试安排表等佐证材料'],
    ['5', '设施设备使用费', '190元/人', '✅ 能收取',
     '设施设备折旧/磨损是等级认定的必要成本，可按磨损费或折旧模型分摊至单人',
     '川人社规〔2022〕9号\n第五条第(五)款',
     '非营利原则：按实际磨损或折旧测算\n不得虚增设备价格或缩短折旧年限',
     '建议逐步转为年限平均\n折旧法，增强说服力'],
    ['6', '耗材成本', '280.49元/人', '✅ 能收取',
     '实操考核耗材是等级认定的直接成本，逐项按市场价×用量计算',
     '川人社规〔2022〕9号\n第五条第(五)款',
     '按实际采购价和用量测算\n健康照护师8考题，耗材量远超普通职业',
     '需保留采购发票/市场\n询价记录作为依据'],
    ['7', '场地水电费', '106.67元/人', '✅ 能收取',
     '场地使用费覆盖考场、候考室、休息室等区域，含场地管理、设施维护、水电使用',
     '川人社规〔2022〕9号\n第五条第(五)款',
     '按实际场地成本÷考核人次分摊\n3200元/天÷30人',
     '场地费标准需有租赁\n合同或内部成本核算\n作为依据'],
    ['8', '视频监控费', '待补充', '✅ 能收取\n(有依据)',
     '川人社规〔2022〕9号明确要求评价过程\"全程留痕\"、视频材料保管≥5年。\n监控是合规要求，其成本属工作成本',
     '川人社规〔2022〕9号\n第四条第(三)款\n+第五条第(五)款',
     '按设备折旧÷年限÷年人次+存储成本测算',
     '数据未到，暂不计入'],
    ['9', '证书制作费', '待补充', '✅ 能收取\n(有依据)',
     '川人社规〔2022〕9号要求统一编码规则和证书样式，制作职业技能等级证书。\n证书工本费属工作成本',
     '川人社规〔2022〕9号\n第四条第(四)款\n+第五条第(五)款',
     '按实际证书采购价测算\n不得高于工本费加合理损耗',
     '数据未到，暂不计入'],
    ['10', '文印广告费', '待补充', '✅ 能收取\n(有依据)',
     '试卷印刷、标识标牌等是等级认定的必要物料成本',
     '川人社规〔2022〕9号\n第五条第(五)款',
     '按实际印刷费和制作费÷人次分摊',
     '数据未到，暂不计入'],
]

for i, item in enumerate(items):
    r = row + 1 + i
    for ci, v in enumerate(item, 1):
        cell = sc(ws6, r, ci, 'left' if ci in [4, 5, 6, 7, 8] else 'center')
        cell.value = v
    # Green fill for "能收取"
    if '能收取' in str(item[3]):
        ws6.cell(row=r, column=4).fill = green_fill
        ws6.cell(row=r, column=4).font = green_font

set_widths(ws6, [6, 18, 16, 16, 42, 28, 32, 28])

# ---- Section 2: Regulatory framework ----
next_row = row + 1 + len(items) + 2
ws6.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=8)
ws6['A' + str(next_row)].value = '二、核心法规依据详解'
ws6['A' + str(next_row)].font = h2_font
next_row += 1

reg_text = [
    ('上位法依据', ''),
    ('川人社规〔2022〕9号', '四川省人力资源和社会保障厅《关于全面推进职业技能等级认定工作的通知》'),
    ('', ''),
    ('核心条款（第五条第(五)款）', ''),
    ('原文', '「评价机构应坚持把社会效益放在首位，不以评价为营利目的，为职业技能等级认定提供稳定的经费保障。评价机构应根据行业规范、地区实际、工作成本等因素综合测算，并向社会公开发布备案职业（工种）收费标准。」'),
    ('', ''),
    ('条款解读', ''),
    ('① 收费主体', '评价机构（经人社部门备案的企业/院校/社会评价组织）有权制定收费标准'),
    ('② 定价原则', '非营利 — 不以评价为营利目的，收费仅覆盖成本'),
    ('③ 定价依据', '行业规范 + 地区实际 + 工作成本 — 三因素综合测算'),
    ('④ 程序要求', '向社会公开发布 — 收费标准必须公开透明'),
    ('⑤ 成本范围', '"工作成本"包括但不限于：考评人员劳务费、场地费、设施设备折旧、耗材费、考务平台费、证书制作费、视频监控费等所有与等级认定直接相关的支出'),
    ('', ''),
    ('辅助条款', ''),
    ('第四条第(三)款', '评价过程"全程留痕"、视频材料保管≥5年 → 视频监控费的法律依据'),
    ('第四条第(四)款', '统一编码规则和证书样式，制作职业技能等级证书 → 证书制作费的法律依据'),
    ('', ''),
    ('政府定价文件的角色', ''),
    ('川发改价格〔2017〕472号', '规定了理论考试费和操作技能考核费的政府定价标准。在职业技能等级认定制度下，该文作为评价机构制定收费标准的参照基准，评价机构参照该标准核定理论考试费和操作技能考核费，并在此基础上根据实际工作成本测算其他费用。'),
    ('', ''),
    ('三、关键合规要点', ''),
    ('1. 非营利底线', '总收费不得超过实际工作成本+合理结余。不得以等级认定为营利手段。建议每年进行一次成本核算并向社会公示。'),
    ('2. 成本可追溯', '各项费用的测算依据（采购发票、询价记录、人员安排表、设备清单等）须完整保存，确保每项成本有据可查。'),
    ('3. 公开透明', '收费标准应在学院官网或认定公告中向社会公开发布，接受社会监督。'),
    ('4. 动态调整', '收费标准应根据耗材价格变动、人工费率调整等因素定期复核更新。'),
    ('5. 备案合规', '在向人社部门申请评价机构备案时，应一并提交收费标准及成本测算依据。'),
]

for i, (label, content) in enumerate(reg_text):
    r = next_row + 1 + i
    if label:
        cell = ws6.cell(row=r, column=1, value=label)
        if label.startswith('原文') or label.startswith('核心条款') or label.startswith('三、'):
            cell.font = h3_font
        elif any(label.startswith(x) for x in ['上位法依据', '辅助条款', '政府定价']):
            cell.font = h3_font
        else:
            cell.font = bold_font if label.startswith('①') or label.startswith('②') or label.startswith('③') or label.startswith('④') or label.startswith('⑤') or label[0].isdigit() else normal_font
    if content:
        c2 = ws6.cell(row=r, column=2, value=content)
        c2.font = normal_font
        c2.alignment = Alignment(wrap_text=True, vertical='center')
    if label == '三、关键合规要点':
        ws6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws6.cell(row=r, column=1).font = h2_font
        ws6.cell(row=r, column=1).fill = sub_header_fill
        for ci in range(1, 9):
            ws6.cell(row=r, column=ci).fill = sub_header_fill

# Adjust column widths for section 2 area
# Already set above

output_path = r'D:\openclaw-workspace\output\健康照护师-成本构成测算-v3-2026.05.21.xlsx'
wb.save(output_path)
print(f'OK: {output_path}')
print('Added: 收费合规性分析 sheet')
