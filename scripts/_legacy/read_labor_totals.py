import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\scrccpa\Desktop\新建文件夹\长期照护师认定人员成本费用测算.xlsx', data_only=True)

for sn in ['25年11月30日实际发放', '26年1月9日实际']:
    ws = wb[sn]
    print(f'\n=== {sn} (max_row={ws.max_row}) ===')
    # Print last 10 rows with all columns
    for r in range(max(1, ws.max_row-10), ws.max_row+1):
        vals = []
        for c in range(1, min(ws.max_column+1, 19)):
            vals.append(str(ws.cell(row=r, column=c).value))
        print(f'  Row {r}: {vals}')
    
    # Sum column 18 (小计)
    total = 0
    for r in range(2, ws.max_row+1):
        v = ws.cell(row=r, column=18).value
        if isinstance(v, (int, float)):
            total += v
    print(f'  SUM col18 (小计): {total}')
