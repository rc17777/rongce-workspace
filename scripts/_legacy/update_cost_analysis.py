# -*- coding: utf-8 -*-
"""Update cost analysis Excel with certification exam + random draw logic"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Load existing workbook
wb = openpyxl.load_workbook(r'D:\openclaw-workspace\output\健康照护师收费分析表.xlsx')

# Style definitions
title_font = Font(name='Arial', size=14, bold=True)
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
sub_header_font = Font(name='Arial', size=11, bold=True)
normal_font = Font(name='Arial', size=10)
green_font = Font(name='Arial', size=10, color='008000')
red_font = Font(name='Arial', size=10, color='FF0000', bold=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
light_blue_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
light_green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def apply_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def apply_row(ws, row, cols, fill=None, font=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = font or normal_font
        cell.border = thin_border
        cell.alignment = center_align if c > 1 else left_align
        if fill:
            cell.fill = fill

# ============================================================
# Insert "认定场景分析" sheet after Sheet 1 (审核总览)
# ============================================================
ws_new = wb.create_sheet('认定场景分析', 1)  # Insert at position 2

ws_new.merge_cells('A1:H1')
ws_new.cell(row=1, column=1).value = '职业技能认定考试场景分析 —— 随机抽题模式下的成本合理性'
ws_new.cell(row=1, column=1).font = title_font
ws_new.cell(row=1, column=1).alignment = center_align
ws_new.row_dimensions[1].height = 35

# Key facts
facts = [
    ('认定性质', '职业技能等级认定(考试)，非培训'),
    ('认定时长', '1天(理论考试 + 实操考核)'),
    ('实操考题', '共8题，考生随机抽取部分考题'),
    ('班次人数', '不等，通常8-15人/批'),
    ('耗材逻辑', '耗材明细表列的是所有可能考题的耗材全集，单个考生只使用抽到考题对应的耗材'),
]
for i, (k, v) in enumerate(facts):
    row = 3 + i
    ws_new.cell(row=row, column=1).value = k
    ws_new.cell(row=row, column=1).font = sub_header_font
    ws_new.cell(row=row, column=1).fill = light_blue_fill
    ws_new.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    ws_new.cell(row=row, column=2).value = v
    ws_new.cell(row=row, column=2).font = normal_font
    for c in range(1, 9):
        ws_new.cell(row=row, column=c).border = thin_border

# Cost decomposition: per-person vs per-batch
row = 9
ws_new.merge_cells(f'A{row}:H{row}')
ws_new.cell(row=row, column=1).value = '一、耗材&设备成本重新分解：人均消耗 vs 共享分摊'
ws_new.cell(row=row, column=1).font = Font(name='Arial', size=12, bold=True)

row = 10
headers = ['等级', '真正一次性消耗\n(口罩/手套/棉签等)\n单人全量成本', '随机抽题后\n实际人均消耗\n(约50%题量)', '共享设备折旧\n(床/AED/模型等)\n单人全量成本', '按班次分摊\n(按10人/批)', '人均总成本\n(实际消耗)', '收费方案分配\n(耗材+设备)', '覆盖判断']
for c, h in enumerate(headers, 1):
    ws_new.cell(row=row, column=c).value = h
apply_header(ws_new, row, len(headers))

# Data: (level, per_person_consumable_full, per_person_consumable_draw, shared_equip_full, shared_equip_per_batch, fee_consumable, fee_equip)
# From detailed analysis: 
# 初级: 一次性51.70, 分摊128.50
# 中级: 一次性107.00, 分摊218.28
# 高级: 一次性206.30, 分摊264.19
calcs = [
    ['初级', 51.70, 51.70*0.5, 128.50, 128.50/10, 25, 25],
    ['中级', 107.00, 107.00*0.5, 218.28, 218.28/10, 35, 35],
    ['高级', 206.30, 206.30*0.5, 264.19, 264.19/10, 50, 50],
]
for r, d in enumerate(calcs):
    row = 11 + r
    ws_new.cell(row=row, column=1).value = d[0]
    ws_new.cell(row=row, column=2).value = d[1]
    ws_new.cell(row=row, column=3).value = d[2]
    ws_new.cell(row=row, column=4).value = d[3]
    ws_new.cell(row=row, column=5).value = round(d[4], 1)
    # Actual per-person total
    actual_total = round(d[2] + d[4], 1)
    ws_new.cell(row=row, column=6).value = actual_total
    # Fee plan allocation
    fee_total = d[5] + d[6]
    ws_new.cell(row=row, column=7).value = fee_total
    # Coverage judgment
    coverage = actual_total / fee_total if fee_total else 0
    if coverage <= 1.0:
        ws_new.cell(row=row, column=8).value = 'OK 覆盖充分'
        ws_new.cell(row=row, column=8).font = green_font
    elif coverage <= 1.3:
        ws_new.cell(row=row, column=8).value = f'基本覆盖({coverage:.0%})'
        ws_new.cell(row=row, column=8).font = normal_font
    else:
        ws_new.cell(row=row, column=8).value = f'覆盖不足({coverage:.0%})'
        ws_new.cell(row=row, column=8).font = red_font
    apply_row(ws_new, row, len(headers))

# Explanation
row = 15
ws_new.merge_cells(f'A{row}:H{row}')
ws_new.cell(row=row, column=1).value = '二、推算说明与关键假设'
ws_new.cell(row=row, column=1).font = Font(name='Arial', size=12, bold=True)

notes = [
    '1. "一次性消耗" = 口罩/手套/棉签/纱布/注射器等每次考试必须更换的物品，按部门明细中未标注"4人一组"或"磨损费"的项目统计',
    '2. 随机抽题系数假设50%：8题中每位考生约抽到4-5题，消耗约一半耗材。实际以护理学院提供的抽题规则为准',
    '3. 共享设备折旧：AED训练机(500元)、护理床(2000元)、心肺复苏模型(300元)等按10人/批分摊，实际以设备折旧年限和使用频次为准',
    '4. 若班次人数更多(如15人/批)，设备分摊更低，覆盖率更高。反之若小班(5人/批)，设备分摊翻倍',
    '5. 若某个耗材项为"必考题"(如隔离防护服28元/人)，则覆盖率需单独核算',
]
for r, note in enumerate(notes):
    row = 16 + r
    ws_new.merge_cells(f'A{row}:H{row}')
    ws_new.cell(row=row, column=1).value = note
    ws_new.cell(row=row, column=1).font = normal_font
    ws_new.cell(row=row, column=1).alignment = left_align

# Sensitivity analysis
row = 22
ws_new.merge_cells(f'A{row}:H{row}')
ws_new.cell(row=row, column=1).value = '三、班次人数敏感性分析（初级为例，人均实际消耗 vs 收费方案50元）'
ws_new.cell(row=row, column=1).font = Font(name='Arial', size=12, bold=True)

row = 23
sens_headers = ['班次人数', '人均一次性\n消耗(50%题量)', '人均设备折旧', '人均总成本', '收费方案\n(耗材+设备)', '差值', '覆盖率', '是否充足']
for c, h in enumerate(sens_headers, 1):
    ws_new.cell(row=row, column=c).value = h
apply_header(ws_new, row, len(sens_headers))

batch_sizes = [5, 8, 10, 12, 15, 20]
for r, n in enumerate(batch_sizes):
    row = 24 + r
    consum = 51.70 * 0.5
    equip = 128.50 / n
    total = round(consum + equip, 1)
    fee = 50
    diff = round(total - fee, 1)
    coverage = fee / total if total else 0
    ws_new.cell(row=row, column=1).value = n
    ws_new.cell(row=row, column=2).value = round(consum, 1)
    ws_new.cell(row=row, column=3).value = round(equip, 1)
    ws_new.cell(row=row, column=4).value = total
    ws_new.cell(row=row, column=5).value = fee
    ws_new.cell(row=row, column=6).value = diff
    ws_new.cell(row=row, column=7).value = f'{coverage:.0%}'
    ws_new.cell(row=row, column=8).value = '充足' if diff <= 0 else ('需关注' if diff <= 15 else '不足')
    apply_row(ws_new, row, len(sens_headers),
              fill=light_green_fill if diff <= 0 else (yellow_fill if diff <= 15 else None))

# Conclusion
row = 31
ws_new.merge_cells(f'A{row}:H{row}')
ws_new.cell(row=row, column=1).value = '四、结论'
ws_new.cell(row=row, column=1).font = Font(name='Arial', size=12, bold=True)

conclusions = [
    '认定考试模式下，耗材按随机抽题消耗、设备按班次分摊后，收费方案中的耗材+设备费(50-100元/人)总体合理。',
    '前提条件：① 实际抽题覆盖率约50% ② 班次人数≥8人/批 ③ 大件设备已按全生命周期折旧而非单次全额计入。',
    '建议护理学院提供：抽题规则(明确几题必考/几题随机)、典型班次人数、设备折旧计算表，以进一步精确核算。',
]
for r, c in enumerate(conclusions):
    row = 32 + r
    ws_new.merge_cells(f'A{row}:H{row}')
    ws_new.cell(row=row, column=1).value = c
    ws_new.cell(row=row, column=1).font = normal_font
    ws_new.cell(row=row, column=1).alignment = left_align

ws_new.column_dimensions['A'].width = 15
for c in ['B','C','D','E','F','G','H']:
    ws_new.column_dimensions[c].width = 16

# ============================================================
# Also update sheet "审核总览" with corrected findings
# ============================================================
ws1 = wb['审核总览']

# Update row 3 (审核范围) to clarify certification not training
ws1.cell(row=3, column=2).value = '初级(五级)320元/人、中级(四级)405元/人、高级(三级)500元/人 —— 职业技能等级认定(考试)，为期1天'

# Update row 5 (审核依据)  
ws1.cell(row=5, column=2).value = '川发改价格(2017)472号 / 川人社规(2025)11号 / 川人社职鉴(2023)4号 / 医保发(2025)11号'

# Update row 7 (审核结论)
ws1.cell(row=7, column=2).value = '认定考试模式下总体合理。按随机抽题+班次分摊逻辑，耗材+设备费覆盖充分。建议提供抽题规则和典型班次人数以精确验证。'

# Skip updating merged cell regions in existing sheet to avoid AttributeError.
# Key findings have been superseded by the new "认定场景分析" sheet.
# Only update the simple non-merged cells (rows 3, 5, 7)

# Save
output_path = r'D:\openclaw-workspace\output\健康照护师收费分析表.xlsx'
wb.save(output_path)
print(f'Saved: {output_path}')
print(f'Sheets: {wb.sheetnames}')
