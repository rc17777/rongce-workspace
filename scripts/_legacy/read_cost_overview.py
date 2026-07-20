import openpyxl, json

wb = openpyxl.load_workbook(r'C:\Users\scrccpa\Desktop\健康照护师-成本构成测算-v3-2026.05.21.xlsx')
ws = wb['成本构成总览']
print(f'rows={ws.max_row}, cols={ws.max_column}')

data = {}
for r in range(1, ws.max_row+1):
    row_vals = []
    for c in range(1, ws.max_column+1):
        v = ws.cell(row=r, column=c).value
        row_vals.append(v)
    data[r] = row_vals

with open(r'D:\openclaw-workspace\temp_cost_overview.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, default=str)

print("Done")
