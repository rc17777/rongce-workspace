import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.Workbook()

# ========== Style definitions ==========
title_font = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
body_font = Font(name='微软雅黑', size=10)
bold_font = Font(name='微软雅黑', size=10, bold=True)
red_font = Font(name='微软雅黑', size=10, bold=True, color='CC0000')
small_font = Font(name='微软雅黑', size=9, color='666666')

title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
phase_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
red_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
light_gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
wrap_align = Alignment(vertical='top', wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0')
)

def style_sheet(ws, col_widths, header_row=1):
    """Apply standard styling to a sheet."""
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_title_row(ws, row, text, col_count, fill=title_fill):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = title_font
    cell.fill = fill
    cell.alignment = center_align
    for c in range(1, col_count+1):
        ws.cell(row=row, column=c).fill = fill
        ws.cell(row=row, column=c).border = thin_border

def write_headers(ws, row, headers, fill=header_fill):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = center_align
        cell.border = thin_border

def write_row(ws, row, values, fonts=None):
    for i, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.font = fonts[i-1] if fonts else body_font
        cell.alignment = wrap_align if '\n' in str(v) else left_align
        cell.border = thin_border

def write_phase_row(ws, row, text, col_count):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name='微软雅黑', size=11, bold=True, color='1F4E79')
    cell.fill = phase_fill
    cell.alignment = left_align
    for c in range(1, col_count+1):
        ws.cell(row=row, column=c).fill = phase_fill
        ws.cell(row=row, column=c).border = thin_border

# ==========================================
# Sheet 1: 使用说明
# ==========================================
ws0 = wb.active
ws0.title = '使用说明'
col_widths = [20, 60]
style_sheet(ws0, col_widths)

write_title_row(ws0, 1, '查阅资料法标准化操作清单（SOP）', 2)

info = [
    ('版本', 'v1.0 | 2026-05-15'),
    ('适用范围', '财务审计、绩效评价、资产清查、专项检查、工程审计等全部审计项目'),
    ('核心理念', '不是在"阅读"，而是在"找反常" —— 从一堆"正常"里，看出"不正常"'),
    ('', ''),
    ('操作节奏', '进场前建基线 → 进场后四维初筛 → 交叉锁定 → 定向深挖'),
    ('', ''),
    ('工作表导航', ''),
    ('Sheet 2', '第一阶段-进场前：建异常基线'),
    ('Sheet 3-6', '第二阶段-四维初筛：时间/数据/流程/人员异常检查'),
    ('Sheet 7', '第三阶段-交叉锁定：四维交叉矩阵'),
    ('Sheet 8', '第四阶段-定向深挖：五问追因+证据固定'),
    ('Sheet 9', '工具速查：分析工具与技能包对照'),
    ('', ''),
    ('执行口诀', '先建基线画时间轴，四维扫描标异常，两维命中即为靶，五问追因定性质。'),
]

for i, (k, v) in enumerate(info, 2):
    c1 = ws0.cell(row=i, column=1, value=k)
    c2 = ws0.cell(row=i, column=2, value=v)
    c1.font = bold_font
    c2.font = body_font
    c1.alignment = left_align
    c2.alignment = wrap_align
    c1.border = thin_border
    c2.border = thin_border

# ==========================================
# Sheet 2: 第一阶段-进场前
# ==========================================
ws1 = wb.create_sheet('第一阶段-进场前')
col_widths = [8, 22, 22, 28, 22]
style_sheet(ws1, col_widths)

write_title_row(ws1, 1, '第一阶段：进场前 —— 建"异常基线"（1-2天）', 5)

row = 3
write_phase_row(ws1, row, '1.1 历史数据拉取', 5)
row += 1
write_headers(ws1, row, ['序号', '操作', '工具/来源', '产出', '完成'])
items_1_1 = [
    ['1', '拉取被审计单位近3-5年财务数据\n（序时账、科目余额表）', '财政一体化系统/\n单位财务系统', 'CSV/Excel数据集', '□'],
    ['2', '拉取近3-5年采购/工程/项目台账', '采购平台/\n发改立项系统', '项目清单', '□'],
    ['3', '拉取近3-5年决算报表', '财政决算系统', '财务报表', '□'],
    ['4', '获取"三定"方案、内控制度、议事规则', '单位提供', '制度汇编', '□'],
]
for item in items_1_1:
    row += 1
    write_row(ws1, row, item)

row += 2
write_phase_row(ws1, row, '1.2 关键时间节点标注（制作时间轴，标注以下节点 ±30天为"敏感窗口"）', 5)
row += 1
write_headers(ws1, row, ['序号', '节点类型', '具体时间', '是否标注', '备注'])
items_1_2 = [
    ['1', '换届时间', '', '□', ''],
    ['2', '上次审计时间', '', '□', ''],
    ['3', '上次巡察时间', '', '□', ''],
    ['4', '主要领导调整时间', '', '□', '含分管领导'],
    ['5', '关键岗位人员退休时间', '', '□', '财务/采购/审批负责人'],
]
for item in items_1_2:
    row += 1
    write_row(ws1, row, item)

row += 2
write_phase_row(ws1, row, '1.3 调取前期问题清单', 5)
row += 1
write_headers(ws1, row, ['序号', '来源', '重点关注', '问题数量', '完成'])
items_1_3 = [
    ['1', '上次审计报告', '未整改到位的问题', '', '□'],
    ['2', '巡察报告', '移交线索、整改情况', '', '□'],
    ['3', '内审报告', '反复出现的问题', '', '□'],
    ['4', '纪检/信访', '已有举报线索', '', '□'],
]
for item in items_1_3:
    row += 1
    write_row(ws1, row, item)

# ==========================================
# Sheet 3: 时间异常
# ==========================================
ws_t = wb.create_sheet('第二阶段-时间异常')
col_widths = [5, 16, 32, 26, 26, 10, 16]
style_sheet(ws_t, col_widths)

write_title_row(ws_t, 1, '第二阶段：四维初筛 —— 维一：时间异常（盯"不该快的时候快了"）', 7)
row = 2
write_headers(ws_t, row, ['序号', '检查项', '操作方法', '正常基准', '异常标准', '结果', '备注'])

items_t = [
    ['T1', '审批速度异常', '统计同类事项审批间隔天数，\n取中位数，标出<中位数50%的', '同类事项中位天数', '审批天数<中位数50%', '□正常 □异常', ''],
    ['T2', '决策时序反常', '核对会议纪要与合同/付款\n日期的先后顺序', '会议→决策→执行', '执行日期早于决策日期', '□正常 □异常', ''],
    ['T3', '敏感窗口突击花钱', '换届/审计/巡察/退休前30天\n的大额支出密集度', '月度支出波动≤30%', '敏感窗口月支出>均值200%', '□正常 □异常', ''],
    ['T4', '年底突击花钱', 'Q4支出占全年比例', 'Q4≤35%', 'Q4>50%', '□正常 □异常', ''],
    ['T5', '节假日前后异常', '春节/国庆前5个工作日\n的支付集中度', '无集中规律', '假期前3天支付占月>40%', '□正常 □异常', ''],
    ['T6', '集中报销', '同一部门/同一人短期内\n报销频次', '月度人均≤3笔', '月度人均>8笔', '□正常 □异常', ''],
    ['T7', '非工作时间操作', '审批/支付发生在\n非工作时段的比例', '工作时段操作', '非工作时段操作占比>10%', '□正常 □异常', ''],
]
for item in items_t:
    row += 1
    write_row(ws_t, row, item)

row += 2
ws_t.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
c = ws_t.cell(row=row, column=1, value='时间异常总计：______ 项（其中高风险 ______ 项）    ⚠️ 高风险标准：同时命中数据/流程/人员任一维度')
c.font = Font(name='微软雅黑', size=11, bold=True, color='CC0000')
c.alignment = left_align

# ==========================================
# Sheet 4: 数据异常
# ==========================================
ws_d = wb.create_sheet('第二阶段-数据异常')
col_widths = [5, 18, 32, 24, 24, 10, 16]
style_sheet(ws_d, col_widths)

write_title_row(ws_d, 1, '第二阶段：四维初筛 —— 维二：数据异常（盯"数字不讲逻辑"）', 7)
row = 2
write_headers(ws_d, row, ['序号', '检查项', '操作方法', '正常基准', '异常标准', '结果', '备注'])

items_d = [
    ['D1', '单项费用月度突变', '月度同比/环比单项费用\n（招待费/培训费/咨询费/维修费等）', '月度波动≤50%', '单月增幅>100%\n且无合理解释', '□正常 □异常', ''],
    ['D2', '付款与进度不匹配', '工程付款比例 vs\n监理确认的工程进度', '偏差≤10%', '偏差>20%', '□正常 □异常', ''],
    ['D3', '费用结构畸高', '同类单位横向对比\n单项费用占比', '同类均值±1个标准差', '偏离>2个标准差', '□正常 □异常', '统计学方法'],
    ['D4', '金额临界点', '略低于招标/审批阈值的\n合同金额分布', '金额自然分布', '集中在阈值的\n95%-99%区间', '□正常 □异常', '规避招标信号'],
    ['D5', '整数金额异常', '大额支付为整数\n（万元/十万元整）的比例', '正常业务含零头', '大额整数支付\n占比>30%', '□正常 □异常', '资金套现信号'],
    ['D6', '专项资金结余突变', '历年专项资金\n结余变化趋势', '平稳波动', '某年突然清零\n或突然暴增', '□正常 □异常', ''],
    ['D7', '人均费用异常', '按部门/人头分摊后\n的人均费用', '同类均值的±30%', '偏离>100%', '□正常 □异常', ''],
    ['D8', '供应商报价规律', '同一项目多家\n供应商报价分布', '自然离散分布', '报价集中在\n某区间±2%', '□正常 □异常', '围标串标信号'],
]
for item in items_d:
    row += 1
    write_row(ws_d, row, item)

row += 2
ws_d.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
c = ws_d.cell(row=row, column=1, value='数据异常总计：______ 项（其中高风险 ______ 项）')
c.font = Font(name='微软雅黑', size=11, bold=True, color='CC0000')
c.alignment = left_align

# ==========================================
# Sheet 5: 流程异常
# ==========================================
ws_p = wb.create_sheet('第二阶段-流程异常')
col_widths = [5, 20, 30, 22, 22, 10, 16]
style_sheet(ws_p, col_widths)

write_title_row(ws_p, 1, '第二阶段：四维初筛 —— 维三：流程异常（盯"该走的路没走"）', 7)
row = 2
write_headers(ws_p, row, ['序号', '检查项', '操作方法', '法规依据', '异常标准', '结果', '备注'])

items_p = [
    ['P1', '应上会未上会', '对照"三重一大"清单\n逐项核查会议纪要', '单位议事规则', '无会议记录的\n重大事项', '□正常 □异常', ''],
    ['P2', '集体决策变个人签字', '检查决策文件\n签字链完整性', '内控制度', '仅有主要领导签字\n无集体讨论记录', '□正常 □异常', ''],
    ['P3', '招标变直采/单一来源', '应公开招标项目\n实际采购方式比对', '《政府采购法》\n《招标投标法》', '规避公开招标', '□正常 □异常', ''],
    ['P4', '传签代替会议', '检查传签文件\n时间跨度和签字日期', '内控制度', '签字日期高度集中\n或跨度过大', '□正常 □异常', '补签信号'],
    ['P5', '先执行后补程序', '合同签订日期 vs\n实际履约日期', '合同管理规定', '合同日期\n晚于履约日期', '□正常 □异常', ''],
    ['P6', '拆分合同规避招标', '同一供应商短期内的\n多个小额合同', '《招标投标法》', '总金额超阈值\n但拆分为多笔', '□正常 □异常', ''],
    ['P7', '超范围/超标准审批', '审批金额/范围\n对比授权制度', '单位授权制度', '超出审批权限', '□正常 □异常', ''],
]
for item in items_p:
    row += 1
    write_row(ws_p, row, item)

row += 2
ws_p.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
c = ws_p.cell(row=row, column=1, value='流程异常总计：______ 项（其中高风险 ______ 项）')
c.font = Font(name='微软雅黑', size=11, bold=True, color='CC0000')
c.alignment = left_align

# ==========================================
# Sheet 6: 人员异常
# ==========================================
ws_r = wb.create_sheet('第二阶段-人员异常')
col_widths = [5, 22, 30, 20, 24, 10, 16]
style_sheet(ws_r, col_widths)

write_title_row(ws_r, 1, '第二阶段：四维初筛 —— 维四：人员异常（盯"人和事的绑定"）', 7)
row = 2
write_headers(ws_r, row, ['序号', '检查项', '操作方法', '正常基准', '异常标准', '结果', '备注'])

items_r = [
    ['R1', '供应商集中度', '同一领导分管期间\n项目流向分布统计', '企业自然分散', '≥60%流向同一企业\n或其关联方', '□正常 □异常', ''],
    ['R2', '新注册/变更供应商中标', '中标供应商成立时间\nvs 中标时间', '成立>3年', '成立<6个月即中标\n或中标前刚变更', '□正常 □异常', '围标信号'],
    ['R3', '关联关系未回避', '供应商法人/股东/高管\n与采购决策者交叉比对', '无关联关系', '存在关联但\n未书面回避', '□正常 □异常', '利益输送信号'],
    ['R4', '关键岗位超期未轮岗', '财务/采购/审批负责人\n任职年限', '≤5年', '>5年未轮换', '□正常 □异常', ''],
    ['R5', '人员变动后风格突变', '岗位人员变动前后的\n审批/支出模式对比', '平稳过渡', '突变幅度>50%', '□正常 □异常', ''],
    ['R6', '亲友/同乡/同学关联', '供应商关键人员与\n被审计单位人员关系', '无特殊关联', '同籍贯/同学/\n前同事等', '□正常 □异常', ''],
]
for item in items_r:
    row += 1
    write_row(ws_r, row, item)

row += 2
ws_r.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
c = ws_r.cell(row=row, column=1, value='人员异常总计：______ 项（其中高风险 ______ 项）')
c.font = Font(name='微软雅黑', size=11, bold=True, color='CC0000')
c.alignment = left_align

# ==========================================
# Sheet 7: 第三阶段-交叉锁定
# ==========================================
ws_x = wb.create_sheet('第三阶段-交叉锁定')
col_widths = [10, 14, 35, 12, 12, 12, 12, 12, 14]
style_sheet(ws_x, col_widths)

write_title_row(ws_x, 1, '第三阶段：交叉锁定 —— 找出"问题入口"（1天）', 9)

row = 3
write_phase_row(ws_x, row, '3.1 四维交叉矩阵（将四维异常点填入下表，命中≥2维 = 高概率问题入口）', 9)

row += 1
write_headers(ws_x, row, ['异常编号', '异常简述', '涉及资金/事项', '时间维', '数据维', '流程维', '人员维', '命中维度', '优先级'])

# Example rows
examples = [
    ['A001', '（示例）Q4招待费集中爆发', 'X月招待费XX万', '✓', '✓', '', '', '2', '⚠️ 中'],
    ['A002', '（示例）某工程付款超进度+未招标', 'XX项目XX万', '✓', '✓', '✓', '', '3', '🔴 高'],
    ['', '', '', '', '', '', '', '', ''],
    ['', '', '', '', '', '', '', '', ''],
    ['', '', '', '', '', '', '', '', ''],
    ['', '', '', '', '', '', '', '', ''],
    ['', '', '', '', '', '', '', '', ''],
    ['', '', '', '', '', '', '', '', ''],
    ['', '', '', '', '', '', '', '', ''],
    ['', '', '', '', '', '', '', '', ''],
    ['', '', '', '', '', '', '', '', ''],
    ['', '', '', '', '', '', '', '', ''],
]
for item in examples:
    row += 1
    fonts = [body_font] * 9
    write_row(ws_x, row, item)

row += 2
write_phase_row(ws_x, row, '3.2 优先级规则', 9)
row += 1
write_headers(ws_x, row, ['优先级', '条件', '处理方式', '', '', '', '', '', ''])
priorities = [
    ['🔴 必查', '命中≥3个维度\n或同时命中"数据+流程"', '第一时间调取全部相关资料\n安排专人定向深挖', '', '', '', '', '', ''],
    ['⚠️ 优先', '命中2个维度', '列入重点核查清单\n按资金量排序处理', '', '', '', '', '', ''],
    ['💡 关注', '命中1个维度\n但异常程度显著', '注册为观察项\n持续收集信息后决定', '', '', '', '', '', ''],
]
for item in priorities:
    row += 1
    fonts = [bold_font if i == 1 else body_font for i in range(1, 10)]
    write_row(ws_x, row, item)

# ==========================================
# Sheet 8: 第四阶段-定向深挖
# ==========================================
ws_s = wb.create_sheet('第四阶段-定向深挖')
col_widths = [8, 40, 50]
style_sheet(ws_s, col_widths)

write_title_row(ws_s, 1, '第四阶段：定向深挖（视项目周期）', 3)

row = 3
write_phase_row(ws_s, row, '4.1 五问追因法（对每个🔴必查项执行）', 3)
row += 1
write_headers(ws_s, row, ['序号', '追问', '记录'])
questions = [
    ['1', '谁主导的？→ 追到具体决策人', ''],
    ['2', '谁受益的？→ 追到最终资金去向', ''],
    ['3', '为什么这样做？→ 排除合理商业理由', ''],
    ['4', '还有谁知道？→ 扩大谈话/取证范围', ''],
    ['5', '有没有关联事件？→ 横向扩展同一主体/同一类型', ''],
]
for item in questions:
    row += 1
    write_row(ws_s, row, item)

row += 2
write_phase_row(ws_s, row, '4.2 证据固定清单', 3)
row += 1
write_headers(ws_s, row, ['序号', '证据类型', '完成'])
evidences = [
    ['1', '截取异常数据的时间序列截图/导出异常数据表', '□'],
    ['2', '复印/扫描关键审批文件、合同、票据', '□'],
    ['3', '记录谈话/问询的旁证（谈话记录签字版）', '□'],
    ['4', '整理对应的法规依据（逐条标注条款编号）', '□'],
    ['5', '编制问题底稿（事实+证据+法规+结论）', '□'],
]
for item in evidences:
    row += 1
    write_row(ws_s, row, item)

# ==========================================
# Sheet 9: 工具速查
# ==========================================
ws_z = wb.create_sheet('工具速查')
col_widths = [8, 24, 26, 28]
style_sheet(ws_z, col_widths)

write_title_row(ws_z, 1, '常用分析工具速查', 4)
row = 2
write_headers(ws_z, row, ['序号', '异常类型', '分析工具/方法', '对应技能包'])

tools = [
    ['1', '数据波动监测', '描述性统计（均值/中位数/标准差）\n+ 箱线图离群值检测', 'audit-data-analysis-methods'],
    ['2', '数字分布异常', 'Benford定律首数字分布检验', 'financial-fraud-detection'],
    ['3', '时间序列趋势', '时间序列分析 + 移动平均\n+ 同比/环比增长率', 'audit-data-analysis-methods'],
    ['4', '同类单位横向对比', '聚类分析 + 多维对标', 'audit-data-analysis-methods'],
    ['5', '供应商关系网络', '关联图谱分析（法人/股东/高管交叉）', 'procurement-audit-models'],
    ['6', '围标串标识别', '报价规律分析 + 中标概率模型\n+ 关系网络叠加', 'procurement-audit-models'],
    ['7', '文本异常检测', '合同/票据关键词检索\n+ NLP相似度检测', '需加载NLP技能包'],
    ['8', '资金流向追踪', '银行流水穿透分析\n+ 资金闭环检测', 'audit-data-analysis-methods'],
]
for item in tools:
    row += 1
    write_row(ws_z, row, item)

row += 2
ws_z.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
c = ws_z.cell(row=row, column=1, value='执行口诀：先建基线画时间轴，四维扫描标异常，两维命中即为靶，五问追因定性质。')
c.font = Font(name='微软雅黑', size=12, bold=True, color='1F4E79')
c.alignment = center_align
for cc in range(1, 5):
    ws_z.cell(row=row, column=cc).fill = phase_fill
    ws_z.cell(row=row, column=cc).border = thin_border

# ==========================================
# Final: Freeze panes, print settings
# ==========================================
for ws in wb.worksheets:
    ws.sheet_properties.tabColor = '1F4E79'
    # freeze top rows
    if ws.title in ['第二阶段-时间异常', '第二阶段-数据异常', '第二阶段-流程异常', '第二阶段-人员异常']:
        ws.freeze_panes = 'A3'
    elif ws.title in ['第三阶段-交叉锁定', '第四阶段-定向深挖', '工具速查']:
        ws.freeze_panes = 'A3'
    elif ws.title == '第一阶段-进场前':
        ws.freeze_panes = 'A2'

# Set first sheet active
wb.active = 0

output_path = 'D:/openclaw-workspace/查阅资料法标准化操作清单.xlsx'
wb.save(output_path)
print(f'Saved: {output_path}')
