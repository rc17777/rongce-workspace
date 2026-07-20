# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Style definitions
title_font = Font(name='Arial', size=14, bold=True)
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
sub_header_font = Font(name='Arial', size=11, bold=True)
normal_font = Font(name='Arial', size=10)
red_font = Font(name='Arial', size=10, color='FF0000', bold=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
light_blue_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
light_red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def apply_header_style(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def apply_cell_style(ws, row, cols, fill=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = center_align if c > 1 else left_align
        if fill:
            cell.fill = fill

# ============================================================
# Sheet 1: Audit Summary
# ============================================================
ws1 = wb.active
ws1.title = '审核总览'

ws1.merge_cells('A1:H1')
ws1.cell(row=1, column=1).value = '健康照护师(长期照护师)职业技能等级认定收费标准审核报告'
ws1.cell(row=1, column=1).font = title_font
ws1.cell(row=1, column=1).alignment = center_align
ws1.row_dimensions[1].height = 35

info_data = [
    ('审核对象', '健康照护师(长期照护师)职业技能等级认定收费标准'),
    ('审核范围', '初级(五级)320元/人、中级(四级)405元/人、高级(三级)500元/人'),
    ('审核依据', '川发改价格(2017)472号 / 川人社规(2025)11号 / 川人社职鉴(2023)4号等'),
    ('数据来源', '护理学院成本测算汇总表 + 长期照护师耗材成本明细表 + 收费方案5稿'),
    ('审核日期', '2026-05-20'),
    ('审核结论', '总体合理，建议关注耗材与场地设备费用分配的充分性'),
]
for i, (k, v) in enumerate(info_data):
    row = 3 + i
    ws1.cell(row=row, column=1).value = k
    ws1.cell(row=row, column=1).font = sub_header_font
    ws1.cell(row=row, column=1).fill = light_blue_fill
    ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    ws1.cell(row=row, column=2).value = v
    ws1.cell(row=row, column=2).font = normal_font
    for c in range(1, 9):
        ws1.cell(row=row, column=c).border = thin_border

# Section 1: Fee Structure
ws1.merge_cells('A10:H10')
ws1.cell(row=10, column=1).value = '一、收费结构合理性分析'
ws1.cell(row=10, column=1).font = Font(name='Arial', size=12, bold=True)
ws1.row_dimensions[10].height = 25

headers = ['等级', '理论考试费\n(政策定价)', '实操考核费\n(政策定价)', '考务平台费', '人工费用', '场地设备\n使用费', '耗材成本', '合计']
row = 11
for c, h in enumerate(headers, 1):
    ws1.cell(row=row, column=c).value = h
apply_header_style(ws1, row, len(headers))

fee_data = [
    ['初级(五级)', 30, 140, 30, 70, 25, 25, 320],
    ['中级(四级)', 35, 190, 30, 80, 35, 35, 405],
    ['高级(三级)', 40, 240, 30, 90, 50, 50, 500],
]
for r, d in enumerate(fee_data):
    row = 12 + r
    for c, v in enumerate(d, 1):
        ws1.cell(row=row, column=c).value = v
    apply_cell_style(ws1, row, len(headers))

# Section 2: Policy vs Other Costs
ws1.merge_cells('A16:H16')
ws1.cell(row=16, column=1).value = '二、政策定价占比分析'
ws1.cell(row=16, column=1).font = Font(name='Arial', size=12, bold=True)

headers2 = ['等级', '政策定价合计\n(理论+实操)', '其他成本合计\n(考务+人工+场地+耗材)', '总收费', '政策定价占比', '其他成本占比', '是否在\n文件标准内', '备注']
row = 17
for c, h in enumerate(headers2, 1):
    ws1.cell(row=row, column=c).value = h
apply_header_style(ws1, row, len(headers2))

analysis_data = [
    ['初级', 170, 150, 320, '53.1%', '46.9%', '是', '政策定价170元占主导，其他150元覆盖运营成本'],
    ['中级', 225, 180, 405, '55.6%', '44.4%', '是', '政策定价225元，增幅合理'],
    ['高级', 280, 220, 500, '56.0%', '44.0%', '是', '政策定价280元，高级实操更复杂'],
]
for r, d in enumerate(analysis_data):
    row = 18 + r
    for c, v in enumerate(d, 1):
        ws1.cell(row=row, column=c).value = v
    apply_cell_style(ws1, row, len(headers2))

# Section 3: Key Findings
ws1.merge_cells('A22:H22')
ws1.cell(row=22, column=1).value = '三、关键发现与问题'
ws1.cell(row=22, column=1).font = Font(name='Arial', size=12, bold=True)

issues = [
    ('发现1: 耗材成本预算偏低',
     '收费方案中耗材仅25-50元/人，但部门详细测算初级87元、中级167元、高级280元。'
     '实际耗材成本远超预算分配，存在成本覆盖不足风险。需关注：实际耗材中部分为4人一组磨损摊销，'
     '收费方案中的耗材分配是否已充分考虑了全部耗材项目。'),
    ('发现2: 场地设备费覆盖不足',
     '收费方案中场地设备费25-50元/人，部门测算初级93元、中级158元、高级190元。差距较大。'
     '可能原因：(1)设备按全生命周期折旧 vs 收费方案按单次使用摊销 '
     '(2)部门测算含全部设备而收费方案可能只覆盖增量成本'),
    ('发现3: 人工费用偏低',
     '健康照护师共8个实操考题(其他职业4-6题)，但人工费仅比同类职业多10元(70 vs 60)。'
     '8题需要更多考评员和考评时间，人工成本可能被低估。'),
    ('发现4: 与同类职业对比合理',
     '与保育师、养老护理员等B类职业相比，健康照护师各等级多10-15元'
     '(初级多10、中级多10、高级多10)，增幅主要来自人工费用增加，结构合理。'),
]
for r, (title, desc) in enumerate(issues):
    row = 23 + r
    ws1.cell(row=row, column=1).value = title
    ws1.cell(row=row, column=1).font = Font(name='Arial', size=10, bold=True, color='FF0000')
    ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    ws1.cell(row=row, column=2).value = desc
    ws1.cell(row=row, column=2).font = normal_font
    ws1.cell(row=row, column=2).alignment = left_align
    for c in range(1, 9):
        ws1.cell(row=row, column=c).border = thin_border
    ws1.row_dimensions[row].height = 60

# Section 4: Conclusion
ws1.merge_cells('A28:H28')
ws1.cell(row=28, column=1).value = '四、审核结论'
ws1.cell(row=28, column=1).font = Font(name='Arial', size=12, bold=True)

conclusions = [
    ('合规性', '合规',
     '理论考试费和操作技能考核费均按川发改价格(2017)472号标准执行，未超出文件规定的收费标准。'
     '其他成本项目(考务平台费、人工费、场地设备费、耗材)均为实际运营成本，政策允许在文件标准之外合理收取。'),
    ('合理性', '基本合理，有关注点',
     '收费总水平与同类B类职业一致(初级310-320元区间)。但因健康照护师实操考题多(8题)、'
     '全国联考要求高，建议核实：(1)耗材是否全部纳入收费方案 (2)人工费是否覆盖8题考评需求 '
     '(3)是否已申请政府培训补贴以弥补成本缺口。'),
    ('成本覆盖', '存在缺口风险',
     '部门测算的耗材+设备成本(180-470元/人)远超收费方案中对应分配(50-100元/人)。'
     '若部门测算准确，则存在较大成本缺口。建议要求部门重新核实各项成本的分摊方式'
     '(如区分一次性消耗与可复用设备折旧)，并提供测算依据的支撑材料。'),
]
for r, (title, verdict, desc) in enumerate(conclusions):
    row = 29 + r
    ws1.cell(row=row, column=1).value = title
    ws1.cell(row=row, column=1).font = sub_header_font
    ws1.cell(row=row, column=1).fill = light_blue_fill
    ws1.cell(row=row, column=2).value = verdict
    ws1.cell(row=row, column=2).font = Font(name='Arial', size=10, bold=True)
    ws1.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
    ws1.cell(row=row, column=3).value = desc
    ws1.cell(row=row, column=3).font = normal_font
    ws1.cell(row=row, column=3).alignment = left_align
    for c in range(1, 9):
        ws1.cell(row=row, column=c).border = thin_border
    ws1.row_dimensions[row].height = 70

ws1.column_dimensions['A'].width = 15
ws1.column_dimensions['B'].width = 16
ws1.column_dimensions['C'].width = 16
ws1.column_dimensions['D'].width = 14
ws1.column_dimensions['E'].width = 14
ws1.column_dimensions['F'].width = 14
ws1.column_dimensions['G'].width = 14
ws1.column_dimensions['H'].width = 60

# ============================================================
# Sheet 2: Cost Detail vs Fee Plan
# ============================================================
ws2 = wb.create_sheet('成本明细对比')

ws2.merge_cells('A1:I1')
ws2.cell(row=1, column=1).value = '健康照护师(长期照护师) 成本明细 vs 收费方案 对比表'
ws2.cell(row=1, column=1).font = title_font
ws2.cell(row=1, column=1).alignment = center_align
ws2.row_dimensions[1].height = 35

ws2.merge_cells('A2:I2')
ws2.cell(row=2, column=1).value = '对比说明: 部门提交的耗材成本明细表(每个等级数十项明细)与收费方案5稿中对应费用的逐项对比'
ws2.cell(row=2, column=1).font = Font(name='Arial', size=9, color='666666')

row = 4
headers3 = ['等级', '收费方案\n耗材费(元)', '部门测算\n耗材成本(元)', '差异\n(方案-实际)', '差异率', 
            '收费方案\n场地设备费(元)', '部门测算\n设备费(元)', '差异\n(方案-实际)', '差异率']
for c, h in enumerate(headers3, 1):
    ws2.cell(row=row, column=c).value = h
apply_header_style(ws2, row, len(headers3))

comp_data = [
    ['初级(五级)', 25, 87.2, 25-87.2, (25-87.2)/87.2, 25, 93, 25-93, (25-93)/93],
    ['中级(四级)', 35, 167.28, 35-167.28, (35-167.28)/167.28, 35, 158, 35-158, (35-158)/158],
    ['高级(三级)', 50, 280.49, 50-280.49, (50-280.49)/280.49, 50, 190, 50-190, (50-190)/190],
]
for r, d in enumerate(comp_data):
    row = 5 + r
    for c, v in enumerate(d, 1):
        ws2.cell(row=row, column=c).value = v
    for c in [4, 5, 8, 9]:
        ws2.cell(row=row, column=c).font = red_font
    if c in [5, 9]:
        ws2.cell(row=row, column=c).number_format = '0.0%'
    apply_cell_style(ws2, row, len(headers3))

# Key observation
row = 9
ws2.merge_cells(f'A{row}:I{row}')
ws2.cell(row=row, column=1).value = 'WARNING: 部门测算的耗材+设备成本远超收费方案中对应分配。可能原因分析:'
ws2.cell(row=row, column=1).font = Font(name='Arial', size=10, bold=True)

reasons = [
    '1. 部门测算包含全部耗材项目(初级46项/中级53项/高级64项)，但部分项目为4人一组共享(如病号服/大浴巾标注4人一组磨损费)，实际人均成本已按4人分摊',
    '2. 设备折旧计算方式不同: 部门按设备全生命周期折旧(如护理床2000元折40元/人)，收费方案可能仅计入单次直接消耗',
    '3. 收费方案的耗材费可能仅覆盖考核用一次性消耗品，设备使用费可能仅覆盖场地水电等运营成本',
    '4. 建议要求护理学院提供: (a)各项耗材的分类 一次性消耗 vs 共享分摊 vs 设备折旧 (b)设备折旧的计算依据 折旧年限/使用次数',
]
for r, reason in enumerate(reasons):
    row = 10 + r
    ws2.merge_cells(f'A{row}:I{row}')
    ws2.cell(row=row, column=1).value = reason
    ws2.cell(row=row, column=1).font = normal_font
    ws2.cell(row=row, column=1).alignment = left_align

for c in range(1, 10):
    ws2.column_dimensions[get_column_letter(c)].width = 16

# ============================================================
# Sheet 3: Cross-occupation Comparison
# ============================================================
ws3 = wb.create_sheet('同级职业横向对比')

ws3.merge_cells('A1:L1')
ws3.cell(row=1, column=1).value = 'B类/C类职业收费横向对比 - 健康照护师定位分析'
ws3.cell(row=1, column=1).font = title_font
ws3.cell(row=1, column=1).alignment = center_align
ws3.row_dimensions[1].height = 35

row = 3
headers4 = ['类别', '职业', '等级', '理论', '实操', '考务平台', '人工', '场地设备', '耗材', '合计', '实操题数', '备注']
for c, h in enumerate(headers4, 1):
    ws3.cell(row=row, column=c).value = h
apply_header_style(ws3, row, len(headers4))

cross_data = [
    ['B类', '保育师/养老护理员/婴幼儿发展引导员/眼镜定配工/眼镜验光员/医药商品购销员', '5级', 30, 140, 30, 60, 25, 25, 310, '4-6题', 'B类标准职业'],
    ['B类', '健康照护师(长期照护师)', '5级', 30, 140, 30, 70, 25, 25, 320, '8题', '>> 本次审核对象'],
    ['B类', '健康照护师(长期照护师)', '4级', 35, 190, 30, 80, 35, 35, 405, '8题', '>> 本次审核对象'],
    ['B类', '健康照护师(长期照护师)', '3级', 40, 240, 30, 90, 50, 50, 500, '8题', '>> 本次审核对象'],
    ['B类', '保育师/养老护理员/婴幼儿发展引导员', '4级', 35, 190, 30, 70, 35, 35, 395, '4-6题', '对比基准'],
    ['B类', '保育师/养老护理员/婴幼儿发展引导员', '3级', 40, 240, 30, 80, 50, 50, 490, '4-6题', '对比基准'],
    ['B类', '助听器验配师', '3级', 40, 240, 30, 90, 310, 70, 780, '-', '设备昂贵不可比'],
    ['C类', '健康管理师', '3级', 40, 220, 30, 80, 10, 20, 400, '-', 'C类实操费标准较低'],
    ['C类', '保健按摩师/美甲师/美容师/中药炮制工', '5级', 30, 120, 30, 60, 25, 25, 290, '-', 'C类基准'],
]
for r, d in enumerate(cross_data):
    row = 4 + r
    for c, v in enumerate(d, 1):
        ws3.cell(row=row, column=c).value = v
    is_target = '>>' in str(d[-1])
    apply_cell_style(ws3, row, len(headers4), fill=yellow_fill if is_target else None)

row = 14
ws3.merge_cells(f'A{row}:L{row}')
ws3.cell(row=row, column=1).value = '横向分析:'
ws3.cell(row=row, column=1).font = sub_header_font

analysis_rows = [
    '1. 健康照护师与保育师/养老护理员同为B类职业，理论+实操考试费完全一致(按政策定价)',
    '2. 健康照护师比其他B类职业多10-15元人工费用(初级70 vs 60/中级80 vs 70/高级90 vs 80)，合理解释: 实操8题(其他4-6题)，考评工作量更大',
    '3. 健康照护师考务平台费/场地设备费/耗材费与同类B类职业相同(初级均为30+25+25)，但实际耗材和设备需求显著更高，建议核实是否低估',
    '4. 与C类职业相比，B类(含健康照护师)实操考核费标准显著更高(初级多20元/中级多20元/高级多20元)，体现B类职业对设备/材料要求更高的定位',
    '5. 横向对比结论: 健康照护师定价处于B类职业合理区间，但人工费和耗材费的充分性需进一步核实',
]
for r, a in enumerate(analysis_rows):
    row = 15 + r
    ws3.merge_cells(f'A{row}:L{row}')
    ws3.cell(row=row, column=1).value = a
    ws3.cell(row=row, column=1).font = normal_font
    ws3.cell(row=row, column=1).alignment = left_align

for c in range(1, 13):
    ws3.column_dimensions[get_column_letter(c)].width = 16
ws3.column_dimensions['B'].width = 32
ws3.column_dimensions['L'].width = 22

# ============================================================
# Sheet 4: Policy Basis
# ============================================================
ws4 = wb.create_sheet('政策依据')

ws4.merge_cells('A1:E1')
ws4.cell(row=1, column=1).value = '政策依据摘要'
ws4.cell(row=1, column=1).font = title_font
ws4.cell(row=1, column=1).alignment = center_align

row = 3
headers5 = ['序号', '政策文件', '关键条款', '与本次审核的关联', '适用判断']
for c, h in enumerate(headers5, 1):
    ws4.cell(row=row, column=c).value = h
apply_header_style(ws4, row, len(headers5))

policy_refs = [
    ['1', 
     '川发改价格(2017)472号\n关于重新公布全省人力资源社会保障部门行政事业性收费的通知',
     '职业技能鉴定考试费标准:\nB类 初级:理论30+实操140\nB类 中级:理论35+实操190\nB类 高级:理论40+实操240\nC类 初级:理论30+实操120\nC类 中级:理论35+实操170\nC类 高级:理论40+实操220',
     '健康照护师按B类标准收取理论+实操=170/225/280元。此部分为政策硬性标准，不可突破',
     '已合规'],
    ['2',
     '川人社规(2025)11号\n四川省补贴性职业技能培训管理办法',
     '规范补贴性培训的补贴标准/申领流程/监督管理。培训补贴可用于补贴培训机构的成本支出',
     '若健康照护师培训属于补贴性培训范围，可申请政府补贴弥补部分成本(待核实是否已申请)',
     '待核实'],
    ['3',
     '川人社职鉴(2023)4号\n职业技能等级认定实施工作流程',
     '规定等级认定的流程/考核标准/考评员配备要求。实操考核需配备足够考评员确保考核质量',
     '健康照护师8题实操需更多考评员和更长时间，人工费增加有据可依',
     '已参考'],
    ['4',
     '川人社办发(2025)22号\n关于进一步提升职业技能培训质效实施技能照亮前程行动',
     '提升培训质效/规范收费行为/加强资金监管。鼓励合理确定培训收费标准',
     '作为收费合理性判断的上位政策依据，要求收费公开透明',
     '已参考'],
    ['5',
     '医保发(2025)11号\n关于做好当前长期照护师培养培训工作的通知',
     '推进长期照护师培养/明确培训标准和考核要求。长期照护师为全国联考职业',
     '长期照护师为全国联考要求更高，成本增加合理。8题实操有政策依据',
     '已参考'],
    ['6',
     '德市财规(2019)3号\n德阳市就业创业补助资金管理使用办法',
     '规定德阳市就业补助资金的使用范围/标准和审批流程',
     '德阳地区的培训可能享受市级就业补助资金(需核实属地适用性)',
     '待核实'],
]
for r, d in enumerate(policy_refs):
    row = 4 + r
    for c, v in enumerate(d, 1):
        ws4.cell(row=row, column=c).value = v
    apply_cell_style(ws4, row, len(headers5))
    ws4.row_dimensions[row].height = 80

ws4.column_dimensions['A'].width = 6
ws4.column_dimensions['B'].width = 30
ws4.column_dimensions['C'].width = 38
ws4.column_dimensions['D'].width = 32
ws4.column_dimensions['E'].width = 14

# ============================================================
# Sheet 5: Consumable Detail
# ============================================================
ws5 = wb.create_sheet('耗材明细汇总')

ws5.merge_cells('A1:I1')
ws5.cell(row=1, column=1).value = '健康照护师 初级(五级)耗材成本明细汇总 - 来自护理学院原始数据'
ws5.cell(row=1, column=1).font = title_font
ws5.cell(row=1, column=1).alignment = center_align

row = 3
headers6 = ['序号', '耗材名称', '规格', '单价(元)', '单人用量', '单人成本(元)', '类型', '是否4人分组', '备注']
for c, h in enumerate(headers6, 1):
    ws5.cell(row=row, column=c).value = h
apply_header_style(ws5, row, len(headers6))

consum_items = [
    ['一', '耗材合计(46项)', '', '', '', 87.2, '', '', '详见原始Excel表格'],
    ['二', '设施设备合计(11项)', '', '', '', 93, '', '', '详见原始Excel表格'],
    ['三', '合计', '', '', '', 180.2, '', '', '部门提交的耗材+设备原始成本合计'],
]
for r, d in enumerate(consum_items):
    row = 4 + r
    for c, v in enumerate(d, 1):
        ws5.cell(row=row, column=c).value = v
    apply_cell_style(ws5, row, len(headers6), fill=yellow_fill if r == 2 else None)

ws5.merge_cells('A8:I8')
ws5.cell(row=8, column=1).value = '关键高值项目(影响审核判断的主要项目)'
ws5.cell(row=8, column=1).font = sub_header_font

key_items = [
    ['1', '隔离防护服(3M)', '套', 28.23, '1', 28.23, '耗材(一次性)', '否', '单人最高耗材项'],
    ['2', '病号服(磨损费)', '件', 60, '1/4', 15, '设备(分摊)', '4人一组', '4人共享分摊'],
    ['3', '大浴巾(磨损费)', '块', 48.5, '1/4', 12, '设备(分摊)', '4人一组', '4人共享分摊'],
    ['4', '治疗车(设备折旧)', '辆', 800, '3', 20, '设备(折旧)', '4人一组', '单价最高设备折旧分摊'],
    ['5', '轮椅(设备折旧)', '台', 504.4, '-', 10, '设备(折旧)', '4人一组', '大型设备折旧'],
    ['6', '坐便椅(设备折旧)', '个', 120, '3', 20, '设备(折旧)', '4人一组', ''],
    ['7', '电子血压计(折旧)', '台', 185, '1', 10, '设备(折旧)', '4人一组', ''],
    ['8', '血糖仪+试纸', '台+盒', '-', '-', 5.93, '设备+耗材', '-', '设备磨损5元+耗材0.93元'],
]
for r, d in enumerate(key_items):
    row = 9 + r
    for c, v in enumerate(d, 1):
        ws5.cell(row=row, column=c).value = v
    apply_cell_style(ws5, row, len(headers6))

for c in range(1, 10):
    ws5.column_dimensions[get_column_letter(c)].width = 16
ws5.column_dimensions['B'].width = 25
ws5.column_dimensions['C'].width = 12

# ============================================================
# Sheet 6: Cost Comparison by Level
# ============================================================
ws6 = wb.create_sheet('各等级成本对比')

ws6.merge_cells('A1:I1')
ws6.cell(row=1, column=1).value = '健康照护师(长期照护师) 初/中/高级 成本总览对比'
ws6.cell(row=1, column=1).font = title_font
ws6.cell(row=1, column=1).alignment = center_align

row = 3
for c, h in enumerate(['项目', '初级(五级)', '中级(四级)', '高级(三级)', '初级 vs 中级 增幅', '中级 vs 高级 增幅', '趋势判断', '说明'], 1):
    ws6.cell(row=row, column=c).value = h
apply_header_style(ws6, row, 8)

level_comp = [
    ['耗材项目数', 46, 53, 64, '+15.2%', '+20.8%', '递增合理', '等级越高护理操作越复杂'],
    ['部门测算耗材成本', 87.20, 167.28, 280.49, '+91.9%', '+67.7%', '递增合理', '高级含胰岛素笔/腹部模型等高值项目'],
    ['部门测算设备成本', 93, 158, 190, '+69.9%', '+20.3%', '递增合理', '高级含护理床/海姆利克马甲等高值设备'],
    ['部门测算合计', 180.2, 325.28, 470.49, '+80.5%', '+44.6%', '递增合理', ''],
    ['收费方案耗材', 25, 35, 50, '+40.0%', '+42.9%', '递增合理', '但绝对值远低于部门测算'],
    ['收费方案设备', 25, 35, 50, '+40.0%', '+42.9%', '递增合理', '但绝对值远低于部门测算'],
    ['收费方案人工', 70, 80, 90, '+14.3%', '+12.5%', '递增偏低', '高级8题 vs 初级8题，人工费仅增20元'],
    ['收费方案合计', 320, 405, 500, '+26.6%', '+23.5%', '整体合理', ''],
]
for r, d in enumerate(level_comp):
    row = 4 + r
    for c, v in enumerate(d, 1):
        ws6.cell(row=row, column=c).value = v
    apply_cell_style(ws6, row, 8)

for c in range(1, 9):
    ws6.column_dimensions[get_column_letter(c)].width = 18
ws6.column_dimensions['A'].width = 22
ws6.column_dimensions['H'].width = 36

# Save
output_path = r'D:\openclaw-workspace\output\健康照护师收费审核报告.xlsx'
wb.save(output_path)
print(f'Saved to {output_path}')
