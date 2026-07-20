#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
阿坝发展控股集团 - 审计评估收费测算表（三指标对比版）
审计：年报审计，按总资产/净资产/收入三指标对比测算
评估：川评协[2017]23号
"""
import sys
sys.stdout.reconfigure(encoding='utf-8-sig')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DARK_BLUE='0A1F3F';TEAL='1A5C6E';GOLD='C5955C';WARM_GRAY='F5F2EC';LIGHT_GOLD='F5E6C8'
tb=Border(left=Side(style='thin',color='999999'),right=Side(style='thin',color='999999'),top=Side(style='thin',color='999999'),bottom=Side(style='thin',color='999999'))
tf=Font(name='微软雅黑',size=14,bold=True,color=DARK_BLUE)
hf=Font(name='微软雅黑',size=10,bold=True,color='FFFFFF')
sf=Font(name='微软雅黑',size=10,bold=True,color=DARK_BLUE)
nf=Font(name='微软雅黑',size=10,color='333333')
mf=Font(name='微软雅黑',size=9,color='666666')
rf=Font(name='微软雅黑',size=11,bold=True,color=DARK_BLUE)
gf=Font(name='微软雅黑',size=10,bold=True,color=GOLD)
rrf=Font(name='微软雅黑',size=12,bold=True,color='CC0000')
hfl=PatternFill(start_color=DARK_BLUE,end_color=DARK_BLUE,fill_type='solid')
tfl=PatternFill(start_color=TEAL,end_color=TEAL,fill_type='solid')
wfl=PatternFill(start_color=WARM_GRAY,end_color=WARM_GRAY,fill_type='solid')
lgfl=PatternFill(start_color=LIGHT_GOLD,end_color=LIGHT_GOLD,fill_type='solid')
ca=Alignment(horizontal='center',vertical='center',wrap_text=True)
ra=Alignment(horizontal='right',vertical='center',wrap_text=True)
def setw(ws,wids):
    for i,w in enumerate(wids,1):ws.column_dimensions[get_column_letter(i)].width=w
def bdr(ws,r1,r2,c1,c2):
    for r in range(r1,r2+1):
        for c in range(c1,c2+1):ws.cell(row=r,column=c).border=tb

AR=[('1','100万元以下',0.500,100),('2','100-500万元',0.150,400),('3','500-1000万元',0.080,500),('4','1000-5000万元',0.040,4000),('5','5000万-1亿',0.030,5000),('6','1亿-5亿',0.022,40000),('7','5亿-10亿',0.015,50000),('8','10亿-100亿',0.010,900000),('9','100亿以上',0.006,None)]
AB=[('财务报表审计',1.0),('合并/分立/清算审计',1.5),('经责/改制/收支审计',1.5),('经济案件/基建竣工审核',2.0),('其他审计',1.2)]
AP=[('1','100以下',15.0,100),('2','100-1000',6.25,900),('3','1000-5000',2.0,4000),('4','5000-10000',1.2,5000),('5','10000-100000',0.25,90000),('6','100000以上',0.15,None)]

def caudit(w):
    dets,tot,rem=[],0.0,w*10000
    for i,n,r,s in AR:
        if s is None:s2=max(rem,0)
        else:s2=min(max(rem,0),s*10000)
        fee=s2*r/100;dets.append((i,n,r,s2,fee));tot+=fee;rem-=s2
    if tot<2000:tot=2000
    return dets,tot

def capr(w):
    dets,tot,rem=[],0.0,w*10000
    for i,n,r,s in AP:
        if s is None:s2=max(rem,0)
        else:s2=min(max(rem,0),s*10000)
        fee=s2*r/1000;dets.append((i,n,r,s2,fee));tot+=fee;rem-=s2
    if tot<2000:tot=2000
    return dets,tot

# 数据
tota=12091490505.09;neta=8274567855.20;reva=940055809.15
metrics=[('总资产',tota),( '净资产',neta),( '收入',reva)]
areas=[('阿坝州国有资产投资管理有限公司','马尔康市周转房二期商业',6699.21),('','马尔康市周转房四期商业',1951.55),('','马尔康市团结广场',108257.77),('','都江堰市光明街、丽水青城',1518.02),('','广汉中山小区住宅',967.23),('','晶金隆商业',25.86),('成都达勤商业管理有限公司','金牛区抚琴西北街、金鱼街',11879.08),('成都瑞久企业管理有限公司','金堂工业厂房',42950.76),('阿坝州国鑫投资发展有限公司','北海市别墅',2569.30),('','红原宾馆',3240.73)]
total_area=sum(a[2] for a in areas)
est_val_per_sqm=5000

# 预计算
ares={}
for nm,val in metrics:
    wan=val/10000;d,b=caudit(wan);ares[nm]=(wan,d,b)

wb=Workbook()

# Sheet 1: 项目概况
ws1=wb.active;ws1.title='项目概况';setw(ws1,[5,55,30])
txts=[('','阿坝发展控股集团有限公司'),('','审计、评估项目收费测算表'),('',''),('!','一、项目基本信息'),('','  项目名称：阿坝发展控股集团有限公司2025年资产评估项目'),('','          及2025年度与2026年指定一期财务报表审计项目'),('','  服务内容：控制价专项审核服务'),('','  甲方：阿坝发展控股集团有限公司'),('','  乙方：四川融策会计师事务所有限公司'),('',''),('!','二、审计收费依据'),('','  川发改价格[2013]901号 年报审计'),('',''),('!','三、评估收费依据'),('','  川评协[2017]23号'),('',''),('!','四、审计收费测算基础数据（2025年初）'),('','  总资产：12,091,490,505.09元'),('','  净资产：8,274,567,855.20元'),('','  收入：940,055,809.15元'),('','  按三指标分别测算并对比'),('',''),('!','五、评估收费测算基础数据'),('','  评估基础资料：阿坝州-汇总面积.xlsx（总18.01万m2）'),('','  注：需补充各项目评估价值（账面原值）'),('',''),('','编制单位：四川融策会计师事务所有限公司'),('','编制日期：2026年7月')]
for i,(t,text) in enumerate(txts,1):
    c=ws1.cell(row=i,column=2,value=text)
    if i==2:c.font=tf
    elif t=='!':c.font=sf
    else:c.font=nf

# Sheet 2: 审计收费对比
ws2=wb.create_sheet('审计收费对比');setw(ws2,[5,18,16,14,14,14,14,16,14])
ws2.merge_cells('A1:I1');ws2.cell(1,1,'阿坝发展控股集团 - 审计收费三指标对比测算（川发改价格[2013]901号）').font=tf;ws2.cell(1,1).alignment=ca
ws2.cell(2,1,'年报审计 - 按总资产/净资产/收入分别测算').font=sf

r=4;ws2.merge_cells(f'A{r}:I{r}');ws2.cell(r,1,'一、三指标收费对比').font=sf
r=5;hs=['序号','计费指标','计费基数(万元)','标准收费(元)','经责审计(1.5倍/元)','合并审计(1.5倍/元)','经济案件(2倍/元)','其他审计(1.2倍/元)','备注']
for c,h in enumerate(hs,1):
    cell=ws2.cell(r,c,h);cell.font=hf;cell.fill=hfl;cell.alignment=ca

max_val=max(metrics,key=lambda x:x[1])[1]
for i,(nm,val) in enumerate(metrics):
    rr=r+1+i;wan,d,b=ares[nm]
    ws2.cell(rr,1,i+1).font=nf;ws2.cell(rr,1).alignment=ca
    ws2.cell(rr,2,nm).font=sf;ws2.cell(rr,2).alignment=ca
    ws2.cell(rr,3,wan).font=nf;ws2.cell(rr,3).alignment=ra;ws2.cell(rr,3).number_format='#,##0.00'
    ws2.cell(rr,4,b).font=rf;ws2.cell(rr,4).alignment=ra;ws2.cell(rr,4).number_format='#,##0.00'
    ws2.cell(rr,5,b*1.5).font=nf;ws2.cell(rr,5).alignment=ra;ws2.cell(rr,5).number_format='#,##0.00'
    ws2.cell(rr,6,b*1.5).font=nf;ws2.cell(rr,6).alignment=ra;ws2.cell(rr,6).number_format='#,##0.00'
    ws2.cell(rr,7,b*2.0).font=nf;ws2.cell(rr,7).alignment=ra;ws2.cell(rr,7).number_format='#,##0.00'
    ws2.cell(rr,8,b*1.2).font=nf;ws2.cell(rr,8).alignment=ra;ws2.cell(rr,8).number_format='#,##0.00'
    ws2.cell(rr,9,'年报审计常用' if nm=='总资产' else '').font=mf
    if val==max_val:
        for c in range(1,10):ws2.cell(rr,c).fill=PatternFill(start_color='FFE0E0',end_color='FFE0E0',fill_type='solid')
    elif i%2==0:
        for c in range(1,10):ws2.cell(rr,c).fill=wfl
bdr(ws2,r,rr,1,9)

r2=rr+2
for nm,val in metrics:
    wan,d,b=ares[nm]
    r2+=1;ws2.merge_cells(f'A{r2}:I{r2}');ws2.cell(r2,1,f'二、{nm}收费明细（{wan:,.2f}万元）').font=sf
    r2+=1;hs=['档次','计费区间','差额计费率','计费额(万元)','费率(%)','收费额(万元)','累计收费(万元)','收费额(元)','备注']
    for c,h in enumerate(hs,1):
        cell=ws2.cell(r2,c,h);cell.font=hf;cell.fill=tfl;cell.alignment=ca
    cum=0
    for i,(idx,n2,rate,seg,fee) in enumerate(d):
        rr=r2+1+i
        ws2.cell(rr,1,idx).font=nf;ws2.cell(rr,1).alignment=ca;ws2.cell(rr,2,n2).font=nf
        ws2.cell(rr,3,f'{rate:.3f}%').font=nf;ws2.cell(rr,3).alignment=ca
        ws2.cell(rr,4,seg/10000).font=nf;ws2.cell(rr,4).alignment=ra;ws2.cell(rr,4).number_format='#,##0'
        ws2.cell(rr,5,rate/100 if seg>0 else 0).font=nf;ws2.cell(rr,5).alignment=ra;ws2.cell(rr,5).number_format='0.0000%'
        ws2.cell(rr,6,fee/10000).font=nf;ws2.cell(rr,6).alignment=ra;ws2.cell(rr,6).number_format='#,##0.0000'
        cum+=fee;ws2.cell(rr,7,cum/10000).font=nf;ws2.cell(rr,7).alignment=ra;ws2.cell(rr,7).number_format='#,##0.0000'
        ws2.cell(rr,8,fee).font=nf;ws2.cell(rr,8).alignment=ra;ws2.cell(rr,8).number_format='#,##0.00'
        ws2.cell(rr,9,'').font=mf
        if i%2==0:
            for c in range(1,10):ws2.cell(rr,c).fill=wfl
    tr=r2+1+len(AR)
    ws2.merge_cells(f'A{tr}:D{tr}');ws2.cell(tr,1,'合计').font=gf;ws2.cell(tr,1).alignment=ra
    ws2.cell(tr,6,b/10000).font=gf;ws2.cell(tr,6).alignment=ra;ws2.cell(tr,6).number_format='#,##0.0000'
    ws2.cell(tr,8,b).font=rrf;ws2.cell(tr,8).alignment=ra;ws2.cell(tr,8).number_format='#,##0.00'
    ws2.cell(tr,9,'标准收费').font=mf
    for c in range(1,10):ws2.cell(tr,c).fill=lgfl
    bdr(ws2,r2,tr,1,9);r2=tr

r2+=2;ws2.cell(r2,1,'注：1. 年报审计通常以总资产为计费基数；2. 最低收费2000元；3. 基准费率可上下浮动20%').font=mf

# Sheet 3: 评估收费测算
ws3=wb.create_sheet('评估收费测算');setw(ws3,[5,22,16,14,14,14,14,18])
ws3.merge_cells('A1:H1');ws3.cell(1,1,'阿坝发展控股集团 - 评估收费测算表（川评协[2017]23号）').font=tf;ws3.cell(1,1).alignment=ca
ws3.cell(2,1,'测算基础：阿坝州-汇总面积.xlsx（18.01万m2）').font=sf
ws3.cell(3,1,'注意：评估收费需以各项目账面原值（评估值）为计费基数，面积仅作参考').font=Font(name='微软雅黑',size=9,bold=True,color='CC0000')

r=5;ws3.merge_cells(f'A{r}:H{r}');ws3.cell(r,1,'一、差额定率累进计费标准').font=sf
r=6;hs=['档次','计费额度（万元）','差额计费率（千分比）','计费额(万元)','费率(%)','收费额(万元)','累计收费(万元)','备注']
for c,h in enumerate(hs,1):
    cell=ws3.cell(r,c,h);cell.font=hf;cell.fill=hfl;cell.alignment=ca
est_total_val=total_area*est_val_per_sqm/10000
d_ap,b_ap=capr(est_total_val);cum=0
for i,(idx,name,rate,seg,fee) in enumerate(d_ap):
    rr=r+1+i
    ws3.cell(rr,1,idx).font=nf;ws3.cell(rr,1).alignment=ca;ws3.cell(rr,2,name).font=nf
    ws3.cell(rr,3,f'{rate:.2f}‰').font=nf;ws3.cell(rr,3).alignment=ca
    ws3.cell(rr,4,seg/10000).font=nf;ws3.cell(rr,4).alignment=ra;ws3.cell(rr,4).number_format='#,##0'
    ws3.cell(rr,5,rate/1000 if seg>0 else 0).font=nf;ws3.cell(rr,5).alignment=ra;ws3.cell(rr,5).number_format='0.0000%'
    ws3.cell(rr,6,fee/10000).font=nf;ws3.cell(rr,6).alignment=ra;ws3.cell(rr,6).number_format='#,##0.0000'
    cum+=fee;ws3.cell(rr,7,cum/10000).font=nf;ws3.cell(rr,7).alignment=ra;ws3.cell(rr,7).number_format='#,##0.0000'
    ws3.cell(rr,8,'').font=mf
    if i%2==0:
        for c in range(1,9):ws3.cell(rr,c).fill=wfl
tr=r+1+len(AP);ws3.merge_cells(f'A{tr}:D{tr}')
ws3.cell(tr,1,'合计').font=gf;ws3.cell(tr,1).alignment=ra
ws3.cell(tr,6,b_ap/10000).font=rf;ws3.cell(tr,6).alignment=ra;ws3.cell(tr,6).number_format='#,##0.0000'
for c in range(1,9):ws3.cell(tr,c).fill=lgfl
bdr(ws3,r,tr,1,8)
rr=tr+2;ws3.cell(rr,1,'注：1. 最低收费2000元；2. 计费额度通常采用被评估资产账面原值').font=mf

rr=tr+4;ws3.merge_cells(f'A{rr}:H{rr}');ws3.cell(rr,1,'二、评估项目明细（需补充各项目评估价值）').font=sf
rr+=1;hs=['权利人','项目名称','面积(m2)','估算价值(万元)','评估标准收费(元)','备注']
for c,h in enumerate(hs,1):
    cell=ws3.cell(rr,c,h);cell.font=hf;cell.fill=hfl;cell.alignment=ca
tev=0;tef=0;po=''
for i,(owner,name,area) in enumerate(areas):
    rr2=rr+1+i;ev=area*est_val_per_sqm/10000;_,ef=capr(ev);tev+=ev;tef+=ef
    dn=owner if owner and owner!=po else ''
    ws3.cell(rr2,1,dn).font=nf;ws3.cell(rr2,2,name).font=nf
    ws3.cell(rr2,3,area).font=nf;ws3.cell(rr2,3).alignment=ra;ws3.cell(rr2,3).number_format='#,##0.00'
    ws3.cell(rr2,4,ev).font=Font(name='微软雅黑',size=10,color='999999')
    ws3.cell(rr2,4).alignment=ra;ws3.cell(rr2,4).number_format='#,##0.00'
    ws3.cell(rr2,5,ef).font=Font(name='微软雅黑',size=10,color='999999')
    ws3.cell(rr2,5).alignment=ra;ws3.cell(rr2,5).number_format='#,##0.00'
    ws3.cell(rr2,6,'需补充账面原值').font=Font(name='微软雅黑',size=9,color='CC0000')
    po=owner
tr=rr+1+len(areas);ws3.merge_cells(f'A{tr}:B{tr}')
ws3.cell(tr,1,'合计').font=gf;ws3.cell(tr,1).alignment=ra
ws3.cell(tr,3,total_area).font=gf;ws3.cell(tr,3).alignment=ra;ws3.cell(tr,3).number_format='#,##0.00'
ws3.cell(tr,4,tev).font=gf;ws3.cell(tr,4).alignment=ra;ws3.cell(tr,4).number_format='#,##0.00'
ws3.cell(tr,5,tef).font=gf;ws3.cell(tr,5).alignment=ra;ws3.cell(tr,5).number_format='#,##0.00'
ws3.cell(tr,6,'').font=mf
for c in range(1,7):ws3.cell(tr,c).fill=lgfl
bdr(ws3,rr,tr,1,6)

# Sheet 4: 费用汇总
ws4=wb.create_sheet('费用汇总');setw(ws4,[5,22,16,18,18,18,18,18,18])
ws4.merge_cells('A1:I1');ws4.cell(1,1,'阿坝发展控股集团 - 审计评估收费汇总表').font=tf;ws4.cell(1,1).alignment=ca
r=3;hs=['项目','计费指标','计费基数(万元)','标准收费(元)','经责审计(1.5倍/元)','合并审计(1.5倍/元)','经济案件(2倍/元)','其他审计(1.2倍/元)','备注']
for c,h in enumerate(hs,1):
    cell=ws4.cell(r,c,h);cell.font=hf;cell.fill=hfl;cell.alignment=ca
for i,(nm,val) in enumerate(metrics):
    rr=r+1+i;wan,d,b=ares[nm]
    ws4.cell(rr,1,'审计收费').font=sf
    ws4.cell(rr,2,nm).font=nf;ws4.cell(rr,2).alignment=ca
    ws4.cell(rr,3,wan).font=nf;ws4.cell(rr,3).alignment=ra;ws4.cell(rr,3).number_format='#,##0.00'
    ws4.cell(rr,4,b).font=rf;ws4.cell(rr,4).alignment=ra;ws4.cell(rr,4).number_format='#,##0.00'
    ws4.cell(rr,5,b*1.5).font=nf;ws4.cell(rr,5).alignment=ra;ws4.cell(rr,5).number_format='#,##0.00'
    ws4.cell(rr,6,b*1.5).font=nf;ws4.cell(rr,6).alignment=ra;ws4.cell(rr,6).number_format='#,##0.00'
    ws4.cell(rr,7,b*2.0).font=nf;ws4.cell(rr,7).alignment=ra;ws4.cell(rr,7).number_format='#,##0.00'
    ws4.cell(rr,8,b*1.2).font=nf;ws4.cell(rr,8).alignment=ra;ws4.cell(rr,8).number_format='#,##0.00'
    ws4.cell(rr,9,'年报审计常用' if nm=='总资产' else '').font=mf
    if i%2==0:
        for c in range(1,10):ws4.cell(rr,c).fill=wfl
bdr(ws4,3,3+len(metrics),1,9)

r=3+len(metrics)+2;ws4.cell(r,1,'补充说明：').font=sf
notes=['1. 年报审计按总资产、净资产、收入三指标分别测算，供对比参考。','2. 年报审计通常以总资产为计费基数，收费约1,234,489.43元。','3. 评估收费需补充各项目账面原值后方可计算。','4. 以上为标准收费，实际收费可上下浮动20%。','5. 最低收费2000元/项目，所有项目均超过此标准。','6. 异地项目额外收取交通费、住宿费。']
for i,note in enumerate(notes):
    ws4.cell(row=9+i,column=1,value=note).font=nf

out=r'C:\Users\scrccpa\Desktop\阿坝发展控股集团-审计评估收费测算表_v2.xlsx'
wb.save(out)
print(f'OK: {out}')
print()
print('=== 三指标对比 ===')
for nm,val in metrics:
    wan,d,b=ares[nm]
    print(f'{nm}: {val:,.2f}元 = {wan:,.2f}万元 -> 标准收费 {b:,.2f}元')
print()
print(f'评估(按5000元/m2估算): 估值{tev:,.2f}万元, 收费约{tef:,.2f}元')