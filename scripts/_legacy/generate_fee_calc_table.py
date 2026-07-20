#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成审计、评估收费测算表（Excel）
审计依据：川发改价格[2013]901号
评估依据：川评协[2017]23号
"""

import sys
import os

sys.stdout.reconfigure(encoding='utf-8-sig')

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ============================================================
# 样式定义
# ============================================================
DARK_BLUE = "0A1F3F"
TEAL = "1A5C6E"
GOLD = "C5955C"
WARM_GRAY = "F5F2EC"
WHITE = "FFFFFF"
LIGHT_BLUE = "D6E4F0"
LIGHT_GOLD = "F5E6C8"

thin_border = Border(
    left=Side(style='thin', color='999999'),
    right=Side(style='thin', color='999999'),
    top=Side(style='thin', color='999999'),
    bottom=Side(style='thin', color='999999'),
)

title_font = Font(name='微软雅黑', size=14, bold=True, color=DARK_BLUE)
header_font = Font(name='微软雅黑', size=10, bold=True, color=WHITE)
sub_header_font = Font(name='微软雅黑', size=10, bold=True, color=DARK_BLUE)
normal_font = Font(name='微软雅黑', size=10, color='333333')
small_font = Font(name='微软雅黑', size=9, color='666666')
result_font = Font(name='微软雅黑', size=11, bold=True, color=DARK_BLUE)
gold_font = Font(name='微软雅黑', size=10, bold=True, color=GOLD)

header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
teal_fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type='solid')
warm_fill = PatternFill(start_color=WARM_GRAY, end_color=WARM_GRAY, fill_type='solid')
light_blue_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type='solid')
light_gold_fill = PatternFill(start_color=LIGHT_GOLD, end_color=LIGHT_GOLD, fill_type='solid')
gold_fill = PatternFill(start_color=GOLD, end_color=GOLD, fill_type='solid')

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)


def apply_border_range(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = thin_border


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# 审计收费数据
# ============================================================
AUDIT_RATES = [
    ("1", "100万元（含）以下", 0.500, 100),
    ("2", "100万元-500万元（含）", 0.150, 400),
    ("3", "500万元-1000万元（含）", 0.080, 500),
    ("4", "1000万元-5000万元（含）", 0.040, 4000),
    ("5", "5000万元-1亿元（含）", 0.030, 5000),
    ("6", "1亿元-5亿元（含）", 0.022, 40000),
    ("7", "5亿元-10亿元（含）", 0.015, 50000),
    ("8", "10亿元-100亿元（含）", 0.010, 900000),
    ("9", "100亿元以上", 0.006, None),
]

AUDIT_TYPES = [
    ("财务报表审计", 1.0, "标准收费"),
    ("合并审计、分立审计、清算审计", 1.5, "三年及以内按财务报表审计收费标准的150%计收"),
    ("经济责任审计、改制审计、财务收支审计", 1.5, "三年及以内按财务报表审计收费标准的150%计收"),
    ("涉及经济案件审计、基本建设工程竣工决算审核", 2.0, "三年及以内按财务报表审计收费标准的200%计收"),
    ("法律、行政法规规定的其他审计业务", 1.2, "按财务报表审计收费标准的120%计收"),
]

# ============================================================
# 评估收费数据（川评协[2017]23号）
# ============================================================
APPRAISAL_RATES = [
    ("1", "100以下（含100）", 15.0, 100),
    ("2", "100以上-1000（含1000）", 6.25, 900),
    ("3", "1000以上-5000（含5000）", 2.0, 4000),
    ("4", "5000以上-10000（含10000）", 1.2, 5000),
    ("5", "10000以上-100000（含100000）", 0.25, 90000),
    ("6", "100000以上", 0.15, None),
]


def calc_cumulative(rates, total_wan):
    """计算差额定率累进，返回(各档明细, 合计)"""
    details = []
    total_fee = 0.0
    remaining = total_wan * 10000  # 转为元

    for idx, name, rate, segment in rates:
        if segment is None:
            # 最后一档，无限额
            fee = remaining * rate / 1000
            actual_segment = remaining
        else:
            segment_wan = segment
            segment_yuan = segment_wan * 10000
            if remaining <= 0:
                fee = 0
                actual_segment = 0
            elif remaining >= segment_yuan:
                fee = segment_yuan * rate / 1000
                actual_segment = segment_yuan
            else:
                fee = remaining * rate / 1000
                actual_segment = remaining

        details.append((idx, name, rate, actual_segment, fee))
        total_fee += fee
        if segment is not None:
            remaining -= segment * 10000
        else:
            remaining = 0

    return details, total_fee


def calc_audit_fee(total_wan, biz_type_idx=0, float_pct=0):
    """计算审计费用"""
    rates = AUDIT_RATES
    details, base_fee = calc_cumulative(rates, total_wan)
    # 最低收费2000元
    if base_fee < 2000:
        base_fee = 2000

    biz_coeff = AUDIT_TYPES[biz_type_idx][1]
    adjusted_fee = base_fee * biz_coeff * (1 + float_pct / 100)
    return details, base_fee, adjusted_fee


def calc_appraisal_fee(total_wan, float_pct=0, security_multiplier=1):
    """计算评估费用"""
    rates = APPRAISAL_RATES
    details, base_fee = calc_cumulative(rates, total_wan)
    # 最低收费2000元
    if base_fee < 2000:
        base_fee = 2000

    adjusted_fee = base_fee * (1 + float_pct / 100) * security_multiplier
    return details, base_fee, adjusted_fee


def format_wan(val):
    """格式化万元"""
    if val is None or val == 0:
        return "-"
    return f"{val:,.2f}"


def format_yuan(val):
    """格式化元"""
    if val is None or val == 0:
        return "-"
    return f"{val:,.2f}"


# ============================================================
# 创建 Excel
# ============================================================
def create_workbook():
    wb = Workbook()

    # ========== Sheet 1: 说明 ==========
    ws1 = wb.active
    ws1.title = "说明"
    set_col_widths(ws1, [5, 50, 30])

    info = [
        ("", "", ""),
        ("", "审计、评估收费测算表", ""),
        ("", "", ""),
        ("", "一、收费依据文件", ""),
        ("", "  1. 审计收费：川发改价格[2013]901号", ""),
        ("", "     《四川省物价局、四川省财政厅关于印发〈四川省会计师事务所服务收费管理办法〉的通知》", ""),
        ("", "  2. 评估收费：川评协[2017]23号", ""),
        ("", "     《四川省资产评估协会关于资产评估机构报送资产评估服务收费标准的通知》", ""),
        ("", "", ""),
        ("", "二、审计收费说明", ""),
        ("", "  1. 采用差额定率累进计费方式", ""),
        ("", "  2. 基准费率上下浮动不得超过20%", ""),
        ("", "  3. 证券、期货相关业务按2-4倍上浮", ""),
        ("", "  4. 最低收费2000元", ""),
        ("", "  5. 业务类型调整系数：", ""),
        ("", "     - 财务报表审计：100%（基准）", ""),
        ("", "     - 合并审计、分立审计、清算审计：150%", ""),
        ("", "     - 经济责任审计、改制审计、财务收支审计：150%", ""),
        ("", "     - 涉及经济案件审计、基建竣工决算审核：200%", ""),
        ("", "     - 法律、行政法规规定的其他审计业务：120%", ""),
        ("", "", ""),
        ("", "三、评估收费说明", ""),
        ("", "  1. 采用差额定率累进计费方式", ""),
        ("", "  2. 计费额度通常采用被评估资产账面原值", ""),
        ("", "  3. 最低收费2000元", ""),
        ("", "  4. 证券期货业务、新兴业务可按2-4倍上浮", ""),
        ("", "  5. 异地业务额外收取交通费、住宿费", ""),
        ("", "", ""),
        ("", "四、使用说明", ""),
        ("", "  1. 在「审计收费测算」sheet的橙色单元格输入计费基数", ""),
        ("", "  2. 在「评估收费测算」sheet的橙色单元格输入计费基数", ""),
        ("", "  3. 在「批量计算」sheet可批量计算多个项目的费用", ""),
        ("", "", ""),
        ("", "编制单位：四川融策会计师事务所", ""),
        ("", "编制日期：2026年7月", ""),
    ]

    for r, (a, b, c) in enumerate(info, 1):
        cell_a = ws1.cell(row=r, column=1, value=a)
        cell_b = ws1.cell(row=r, column=2, value=b)
        cell_c = ws1.cell(row=r, column=3, value=c)
        cell_a.font = normal_font
        cell_b.font = normal_font
        cell_c.font = normal_font

    # 标题
    ws1.cell(row=2, column=2).font = title_font
    ws1.cell(row=2, column=2).alignment = center_align

    # Section headers
    for r in [4, 10, 23, 30]:
        ws1.cell(row=r, column=2).font = sub_header_font

    # ========== Sheet 2: 审计收费测算 ==========
    ws2 = wb.create_sheet("审计收费测算")
    set_col_widths(ws2, [5, 22, 16, 14, 14, 14, 14, 14])

    # 标题
    ws2.merge_cells('A1:H1')
    ws2.cell(row=1, column=1, value="审计收费测算表（川发改价格[2013]901号）").font = title_font
    ws2.cell(row=1, column=1).alignment = center_align

    # 差额定率累进表
    row = 3
    ws2.merge_cells(f'A{row}:H{row}')
    ws2.cell(row=row, column=1, value="一、差额定率累进计费标准").font = sub_header_font

    row = 4
    headers = ["档次", "计费额度", "差额计费率", "计费额(万元)", "费率(%)", "收费额(万元)", "累计收费(万元)", "备注"]
    for c, h in enumerate(headers, 1):
        cell = ws2.cell(row=row, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # 计算示例
    example_total = 30000  # 万元
    ex_details, ex_base, ex_adjusted = calc_audit_fee(example_total, 0, 0)

    cum = 0
    for i, (idx, name, rate, seg, fee) in enumerate(ex_details):
        r = row + 1 + i
        ws2.cell(row=r, column=1, value=idx).font = normal_font
        ws2.cell(row=r, column=1).alignment = center_align
        ws2.cell(row=r, column=2, value=name).font = normal_font
        ws2.cell(row=r, column=3, value=f"{rate:.3f}%").font = normal_font
        ws2.cell(row=r, column=3).alignment = center_align
        ws2.cell(row=r, column=4, value=seg / 10000 if seg > 0 else 0).font = normal_font
        ws2.cell(row=r, column=4).alignment = right_align
        ws2.cell(row=r, column=4).number_format = '#,##0'
        ws2.cell(row=r, column=5, value=rate / 10 if seg > 0 else 0).font = normal_font
        ws2.cell(row=r, column=5).alignment = right_align
        ws2.cell(row=r, column=5).number_format = '#,##0.0000'
        ws2.cell(row=r, column=6, value=fee / 10000).font = normal_font
        ws2.cell(row=r, column=6).alignment = right_align
        ws2.cell(row=r, column=6).number_format = '#,##0.0000'
        cum += fee / 10000
        ws2.cell(row=r, column=7, value=cum).font = normal_font
        ws2.cell(row=r, column=7).alignment = right_align
        ws2.cell(row=r, column=7).number_format = '#,##0.0000'
        ws2.cell(row=r, column=8, value="").font = small_font
        if i % 2 == 0:
            for c in range(1, 9):
                ws2.cell(row=r, column=c).fill = warm_fill

    # 合计行
    total_row = row + 1 + len(AUDIT_RATES)
    ws2.merge_cells(f'A{total_row}:D{total_row}')
    ws2.cell(row=total_row, column=1, value="合计（收费标准）").font = gold_font
    ws2.cell(row=total_row, column=1).alignment = right_align
    ws2.cell(row=total_row, column=5, value="").font = normal_font
    ws2.cell(row=total_row, column=6, value=ex_base / 10000).font = result_font
    ws2.cell(row=total_row, column=6).alignment = right_align
    ws2.cell(row=total_row, column=6).number_format = '#,##0.0000'
    # 铜金填充
    for c in range(1, 9):
        ws2.cell(row=total_row, column=c).fill = light_gold_fill

    apply_border_range(ws2, row, total_row, 1, 8)

    # 备注
    r = total_row + 2
    ws2.cell(row=r, column=1, value="注：1. 最低收费2000元；2. 基准费率可上下浮动20%；3. 证券期货业务按2-4倍上浮").font = small_font

    # 计算区
    r = total_row + 4
    ws2.merge_cells(f'A{r}:H{r}')
    ws2.cell(row=r, column=1, value="二、费用计算（请输入计费基数）").font = sub_header_font

    # 输入
    r += 1
    ws2.cell(row=r, column=1, value="计费基数（万元）：").font = normal_font
    ws2.merge_cells(f'B{r}:C{r}')
    input_cell = ws2.cell(row=r, column=2, value=30000)
    input_cell.font = Font(name='微软雅黑', size=12, bold=True, color='CC6600')
    input_cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    input_cell.alignment = right_align
    input_cell.number_format = '#,##0'

    # 业务类型
    r += 1
    ws2.cell(row=r, column=1, value="业务类型：").font = normal_font
    ws2.merge_cells(f'B{r}:C{r}')
    ws2.cell(row=r, column=2, value=AUDIT_TYPES[0][0]).font = normal_font

    # 浮动比例
    r += 1
    ws2.cell(row=r, column=1, value="上浮比例（%）：").font = normal_font
    ws2.merge_cells(f'B{r}:C{r}')
    float_cell = ws2.cell(row=r, column=2, value=0)
    float_cell.font = Font(name='微软雅黑', size=12, bold=True, color='CC6600')
    float_cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    float_cell.alignment = right_align

    # 结果
    r += 2
    ws2.merge_cells(f'A{r}:E{r}')
    ws2.cell(row=r, column=1, value="【计算结果】").font = Font(name='微软雅黑', size=12, bold=True, color=DARK_BLUE)

    _, base_fee_30, adj_fee_30 = calc_audit_fee(30000, 0, 0)
    r += 1
    ws2.cell(row=r, column=1, value="标准收费（元）：").font = normal_font
    ws2.merge_cells(f'B{r}:D{r}')
    ws2.cell(row=r, column=2, value=base_fee_30).font = result_font
    ws2.cell(row=r, column=2).alignment = right_align
    ws2.cell(row=r, column=2).number_format = '#,##0.00'

    # 业务类型调整说明
    r += 1
    ws2.cell(row=r, column=1, value="业务类型调整系数：").font = normal_font
    ws2.merge_cells(f'B{r}:D{r}')
    ws2.cell(row=r, column=2, value=f"{AUDIT_TYPES[0][1]}").font = normal_font

    r += 1
    ws2.cell(row=r, column=1, value="调整后收费（元）：").font = normal_font
    ws2.merge_cells(f'B{r}:D{r}')
    ws2.cell(row=r, column=2, value=adj_fee_30).font = Font(name='微软雅黑', size=12, bold=True, color='CC0000')
    ws2.cell(row=r, column=2).alignment = right_align
    ws2.cell(row=r, column=2).number_format = '#,##0.00'

    # 各档明细
    r += 2
    ws2.cell(row=r, column=1, value="三、收费明细计算").font = sub_header_font

    r += 1
    detail_headers = ["档次", "计费区间", "差额计费率", "计费额(万元)", "收费额(万元)"]
    for c, h in enumerate(detail_headers, 1):
        cell = ws2.cell(row=r, column=c, value=h)
        cell.font = header_font
        cell.fill = teal_fill
        cell.alignment = center_align

    cum_total = 0
    for i, (idx, name, rate, seg, fee) in enumerate(ex_details):
        r2 = r + 1 + i
        ws2.cell(row=r2, column=1, value=idx).font = normal_font
        ws2.cell(row=r2, column=1).alignment = center_align
        ws2.cell(row=r2, column=2, value=name).font = normal_font
        ws2.cell(row=r2, column=3, value=f"{rate:.3f}%").font = normal_font
        ws2.cell(row=r2, column=3).alignment = center_align
        ws2.cell(row=r2, column=4, value=seg / 10000 if seg > 0 else 0).font = normal_font
        ws2.cell(row=r2, column=4).alignment = right_align
        ws2.cell(row=r2, column=4).number_format = '#,##0'
        ws2.cell(row=r2, column=5, value=fee / 10000).font = normal_font
        ws2.cell(row=r2, column=5).alignment = right_align
        ws2.cell(row=r2, column=5).number_format = '#,##0.0000'
        cum_total += fee / 10000
        if i % 2 == 0:
            for c in range(1, 6):
                ws2.cell(row=r2, column=c).fill = warm_fill

    apply_border_range(ws2, r, r + len(AUDIT_RATES), 1, 5)

    # ========== Sheet 3: 评估收费测算 ==========
    ws3 = wb.create_sheet("评估收费测算")
    set_col_widths(ws3, [5, 22, 16, 14, 14, 14, 14, 14])

    # 标题
    ws3.merge_cells('A1:H1')
    ws3.cell(row=1, column=1, value="评估收费测算表（川评协[2017]23号）").font = title_font
    ws3.cell(row=1, column=1).alignment = center_align

    # 差额定率累进表
    row = 3
    ws3.merge_cells(f'A{row}:H{row}')
    ws3.cell(row=row, column=1, value="一、差额定率累进计费标准").font = sub_header_font

    row = 4
    headers = ["档次", "计费额度（万元）", "差额计费率（‰）", "计费额(万元)", "费率(%)", "收费额(万元)", "累计收费(万元)", "备注"]
    for c, h in enumerate(headers, 1):
        cell = ws3.cell(row=row, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # 示例计算
    ex_total_ap = 30000  # 万元
    ap_details, ap_base, ap_adj = calc_appraisal_fee(ex_total_ap, 0, 1)

    cum = 0
    for i, (idx, name, rate, seg, fee) in enumerate(ap_details):
        r = row + 1 + i
        ws3.cell(row=r, column=1, value=idx).font = normal_font
        ws3.cell(row=r, column=1).alignment = center_align
        ws3.cell(row=r, column=2, value=name).font = normal_font
        ws3.cell(row=r, column=3, value=f"{rate:.2f}‰").font = normal_font
        ws3.cell(row=r, column=3).alignment = center_align
        ws3.cell(row=r, column=4, value=seg / 10000 if seg > 0 else 0).font = normal_font
        ws3.cell(row=r, column=4).alignment = right_align
        ws3.cell(row=r, column=4).number_format = '#,##0'
        ws3.cell(row=r, column=5, value=rate / 10 if seg > 0 else 0).font = normal_font
        ws3.cell(row=r, column=5).alignment = right_align
        ws3.cell(row=r, column=5).number_format = '#,##0.0000'
        ws3.cell(row=r, column=6, value=fee / 10000).font = normal_font
        ws3.cell(row=r, column=6).alignment = right_align
        ws3.cell(row=r, column=6).number_format = '#,##0.0000'
        cum += fee / 10000
        ws3.cell(row=r, column=7, value=cum).font = normal_font
        ws3.cell(row=r, column=7).alignment = right_align
        ws3.cell(row=r, column=7).number_format = '#,##0.0000'
        ws3.cell(row=r, column=8, value="").font = small_font
        if i % 2 == 0:
            for c in range(1, 9):
                ws3.cell(row=r, column=c).fill = warm_fill

    # 合计行
    total_row = row + 1 + len(APPRAISAL_RATES)
    ws3.merge_cells(f'A{total_row}:D{total_row}')
    ws3.cell(row=total_row, column=1, value="合计（收费标准）").font = gold_font
    ws3.cell(row=total_row, column=1).alignment = right_align
    ws3.cell(row=total_row, column=5, value="").font = normal_font
    ws3.cell(row=total_row, column=6, value=ap_base / 10000).font = result_font
    ws3.cell(row=total_row, column=6).alignment = right_align
    ws3.cell(row=total_row, column=6).number_format = '#,##0.0000'
    for c in range(1, 9):
        ws3.cell(row=total_row, column=c).fill = light_gold_fill

    apply_border_range(ws3, row, total_row, 1, 8)

    # 备注
    r = total_row + 2
    ws3.cell(row=r, column=1, value="注：1. 最低收费2000元；2. 计费额度通常采用被评估资产账面原值；3. 证券期货业务可按2-4倍上浮").font = small_font

    # 计算区
    r = total_row + 4
    ws3.merge_cells(f'A{r}:H{r}')
    ws3.cell(row=r, column=1, value="二、费用计算（请输入计费基数）").font = sub_header_font

    # 输入
    r += 1
    ws3.cell(row=r, column=1, value="计费基数（万元）：").font = normal_font
    ws3.merge_cells(f'B{r}:C{r}')
    input_cell = ws3.cell(row=r, column=2, value=30000)
    input_cell.font = Font(name='微软雅黑', size=12, bold=True, color='CC6600')
    input_cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    input_cell.alignment = right_align
    input_cell.number_format = '#,##0'

    # 浮动比例
    r += 1
    ws3.cell(row=r, column=1, value="上浮比例（%）：").font = normal_font
    ws3.merge_cells(f'B{r}:C{r}')
    float_cell = ws3.cell(row=r, column=2, value=0)
    float_cell.font = Font(name='微软雅黑', size=12, bold=True, color='CC6600')
    float_cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    float_cell.alignment = right_align

    # 证券期货倍数
    r += 1
    ws3.cell(row=r, column=1, value="证券期货上浮倍数：").font = normal_font
    ws3.merge_cells(f'B{r}:C{r}')
    ws3.cell(row=r, column=2, value=1).font = normal_font

    # 结果
    r += 2
    ws3.merge_cells(f'A{r}:E{r}')
    ws3.cell(row=r, column=1, value="【计算结果】").font = Font(name='微软雅黑', size=12, bold=True, color=DARK_BLUE)

    r += 1
    ws3.cell(row=r, column=1, value="标准收费（元）：").font = normal_font
    ws3.merge_cells(f'B{r}:D{r}')
    ws3.cell(row=r, column=2, value=ap_base).font = result_font
    ws3.cell(row=r, column=2).alignment = right_align
    ws3.cell(row=r, column=2).number_format = '#,##0.00'

    r += 1
    ws3.cell(row=r, column=1, value="调整后收费（元）：").font = normal_font
    ws3.merge_cells(f'B{r}:D{r}')
    ws3.cell(row=r, column=2, value=ap_adj).font = Font(name='微软雅黑', size=12, bold=True, color='CC0000')
    ws3.cell(row=r, column=2).alignment = right_align
    ws3.cell(row=r, column=2).number_format = '#,##0.00'

    # 收费明细
    r += 2
    ws3.cell(row=r, column=1, value="三、收费明细计算").font = sub_header_font

    r += 1
    detail_headers = ["档次", "计费区间（万元）", "差额计费率（‰）", "计费额(万元)", "收费额(万元)"]
    for c, h in enumerate(detail_headers, 1):
        cell = ws3.cell(row=r, column=c, value=h)
        cell.font = header_font
        cell.fill = teal_fill
        cell.alignment = center_align

    cum_total = 0
    for i, (idx, name, rate, seg, fee) in enumerate(ap_details):
        r2 = r + 1 + i
        ws3.cell(row=r2, column=1, value=idx).font = normal_font
        ws3.cell(row=r2, column=1).alignment = center_align
        ws3.cell(row=r2, column=2, value=name).font = normal_font
        ws3.cell(row=r2, column=3, value=f"{rate:.2f}‰").font = normal_font
        ws3.cell(row=r2, column=3).alignment = center_align
        ws3.cell(row=r2, column=4, value=seg / 10000 if seg > 0 else 0).font = normal_font
        ws3.cell(row=r2, column=4).alignment = right_align
        ws3.cell(row=r2, column=4).number_format = '#,##0'
        ws3.cell(row=r2, column=5, value=fee / 10000).font = normal_font
        ws3.cell(row=r2, column=5).alignment = right_align
        ws3.cell(row=r2, column=5).number_format = '#,##0.0000'
        cum_total += fee / 10000
        if i % 2 == 0:
            for c in range(1, 6):
                ws3.cell(row=r2, column=c).fill = warm_fill

    apply_border_range(ws3, r, r + len(APPRAISAL_RATES), 1, 5)

    # ========== Sheet 4: 批量计算 ==========
    ws4 = wb.create_sheet("批量计算")
    set_col_widths(ws4, [5, 25, 15, 15, 15, 15, 15, 15, 15, 15])

    ws4.merge_cells('A1:J1')
    ws4.cell(row=1, column=1, value="审计、评估收费批量计算表").font = title_font
    ws4.cell(row=1, column=1).alignment = center_align

    # 表头
    row = 3
    batch_headers = ["序号", "项目名称", "资产总额(万元)", "审计标准收费(元)", "审计调整后收费(元)",
                     "评估标准收费(元)", "评估调整后收费(元)", "审计+评估合计(元)", "备注"]
    for c, h in enumerate(batch_headers, 1):
        cell = ws4.cell(row=row, column=c,