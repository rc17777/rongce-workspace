# -*- coding: utf-8 -*-
import os
import glob
import openpyxl

base = "C:/Users/scrccpa/Desktop/成都轨道资源资料"

# Read ledger
ledger_path = glob.glob(base + "/**/资源公司资产台账信息（固定资产）.xlsx", recursive=True)[0]
print("Ledger path:", ledger_path)

wb = openpyxl.load_workbook(ledger_path, data_only=True, read_only=True)
ws = wb.active
headers = [cell.value for cell in ws[1]]
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    rows.append(row)
wb.close()

# Check for 天府 assets specifically
print("\n=== 天府广场相关资产详细分析 ===")
print("\n位置1=天府广场的资产:")
tf_assets = []
for r in rows:
    if "天府" in str(r[23] or ""):  # loc1 = col 23
        tf_assets.append(r)
        print("  名称=%s | 分类=%s/%s/%s | 原值=%s | 净值=%s | 状态=%s | 位置2=%s | 合同号=%s" % (
            str(r[12] or "")[:30], str(r[4] or ""), str(r[5] or ""), str(r[6] or ""),
            str(r[31] or ""), str(r[41] or ""), str(r[30] or ""), str(r[24] or ""), str(r[28] or "")
        ))

print("\n总计天府广场相关资产: %d条" % len(tf_assets))

# Re-run elevator analysis with proper UTF-8 output
print("\n=== 电梯/扶梯详细分析 ===")
elevators = []
for r in rows:
    name = str(r[12] or "")
    cat2 = str(r[5] or "")
    cat3 = str(r[6] or "")
    full = name + cat2 + cat3
    for kw in ["电梯", "扶梯", "电扶梯", "自动扶梯", "升降"]:
        if kw in full:
            elevators.append(r)
            break

print("电梯扶梯资产条目数: %d" % len(elevators))
count_by_name = {}
for r in elevators:
    name = str(r[12] or "")
    if name not in count_by_name:
        count_by_name[name] = 0
    count_by_name[name] += 1

print("按名称分类:")
for k, v in sorted(count_by_name.items(), key=lambda x: -x[1]):
    print("  %s: %d条" % (k, v))

# Check if there are straight escalators vs elevators
print("\n分类统计:")
elevator_types = {}
for r in elevators:
    cat3 = str(r[6] or "")
    if cat3 not in elevator_types:
        elevator_types[cat3] = {"count": 0, "qty": 0}
    elevator_types[cat3]["count"] += 1
    elevator_types[cat3]["qty"] += r[18] or 1

for k, v in sorted(elevator_types.items(), key=lambda x: -x[1]["count"]):
    print("  %s: %d条, 数量合计=%d" % (k, v["count"], v["qty"]))

# Check for straight elevator/auto lift names
print("\n电梯/自动扶梯/载货电梯:");
for r in elevators:
    print("  %s | %s | %s | 数量=%s | 原值=%s" % (
        str(r[12] or "")[:25], str(r[5] or "")[:15], str(r[6] or "")[:15], 
        str(r[18] or ""), str(r[31] or "")
    ))
