#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
阿坝发展控股集团 - 审计收费 & 资产评估收费计算
收费依据：
  1. 审计：川发改价格〔2013〕901号
  2. 评估：四川省资产评估收费标准（川发改价格〔2012〕827号）
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# ========== 公共样式 ==========
FONT_TITLE = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
FONT_HEADER = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
FONT_NORMAL = Font(name='微软雅黑', size=10)
FONT_BOLD = Font(name='微软雅黑', size=10, bold=True)
FONT_SMALL = Font(name='微软雅黑', size=9, color='666666')
FONT_SUBTITLE = Font(name='微软雅黑', size=11, bold=True, color='0A1F3F')
FONT_SUBTITLE_WHITE = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')

FILL_DARK_BLUE = PatternFill(start_color='0A1F3F', end_color='0A1F3F', fill_type='solid')
FILL_MID_BLUE = PatternFill(start_color='1A5C6E', end_color='1A5C6E', fill_type='solid')
FILL_LIGHT = PatternFill(start_color='F5F2EC', end_color='F5F2EC', fill_type='solid')
FILL_GOLD = PatternFill(start_color='C5955C', end_color='C5955C', fill_type='solid')
FILL_WHITE = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
FILL_TOTAL = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center', wrap_text=True)

THIN_BORDER = Border(
    left=Side(style='thin', color='0A1F3F'),
    right=Side(style='thin', color='0A1F3F'),
    top=Side(style='thin', color='0A1F3F'),
    bottom=Side(style='thin', color='0A1F3F'),
)

def apply_cell(ws, row, col, value, font=FONT_NORMAL, fill=FILL_WHITE, alignment=ALIGN_CENTER, border=THIN_BORDER, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = border
    if number_format:
        cell.number_format = number_format
    return cell

def apply_range(ws, start_row, end_row, start_col, end_col, font=None, fill=None, alignment=None, border=None):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            if font: cell.font = font
            if fill: cell.fill = fill
            if alignment: cell.alignment = alignment
            if border: cell.border = border

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# =============================================
# 一、审计收费计算
# =============================================
def calc_audit_fee():
    """
    川发改价格〔2013〕901号
    年度财务报表审计 差额定率累进（按资产总额）
    
    费率表（‰）:
    | 档次 | 资产总额（万元）    | 费率(‰) |
    | 1    | ≤100             | 2.5    |
    | 2    | 100-500          | 1.5    |
    | 3    | 500-1000         | 0.8    |
    | 4    | 1000-5000        | 0.4    |
    | 5    | 5000-10000       | 0.3    |
    | 6    | 10000-50000      | 0.2    |
    | 7    | 50000-100000     | 0.15   |
    | 8    | >100000          | 0.1    |
    """
    
    # 收费基数
    total_assets = 12_091_544_905.09       # 总资产
    net_assets   =  8_274_567_855.20       # 净资产
    revenue      =    940_055_809.15       # 收入
    
    # 差额定率累进分档
    brackets = [
        (0,           100,        2.5),    # 万
        (100,         500,        1.5),
        (500,         1000,       0.8),
        (1000,        5000,       0.4),
        (5000,        10000,      0.3),
        (10000,       50000,      0.2),
        (50000,       100000,     0.15),
        (100000,      float('inf'), 0.1),
    ]
    
    def tiered_calc(base_amount, brackets):
        """差额定率累进计算"""
        base_wan = base_amount / 10000  # 转为万元
        details = []
        total_fee = 0
        remaining = base_wan
        
        for i, (lo, hi, rate) in enumerate(brackets):
            tier_amount = min(remaining, hi - lo)
            if tier_amount <= 0:
                details.append((i+1, f"{lo:,.0f}-{hi:,.0f}" if hi != float('inf') else f">{lo:,.0f}",
                               tier_amount, rate, 0))
                continue
            
            tier_fee = tier_amount * rate  # 万元×‰ = 元/万×每千 → 万元×rate = 千元
            # 实际上：万元 × rate(‰) = 万元 × rate/1000 = rate*万元/1000 千元
            # 更准确：tier_amount万元 = tier_amount × 10000 元
            # fee = 基数元 × rate/1000
            fee_yuan = tier_amount * 10000 * rate / 1000
            total_fee += fee_yuan
            details.append((i+1, f"{lo:,.0f}-{hi:,.0f}" if hi != float('inf') else f">{lo:,.0f}",
                           tier_amount, rate, fee_yuan))
            remaining -= tier_amount
            if remaining <= 0:
                break
        
        return details, total_fee
    
    # ===== 按总资产计算 =====
    details_asset, fee_asset = tiered_calc(total_assets, brackets)
    
    # ===== 按净资产计算（参考） =====
    details_net, fee_net = tiered_calc(net_assets, brackets)
    
    # ===== 按收入计算（参考）- 收入有单独费率表 =====
    # 川发改901号：收入基础审计（如经济效益审计等）使用不同费率
    # 这里按合同性质（年度报表审计控制价审核），主要使用资产总额
    rev_brackets = [
        (0,           100,        3.0),
        (100,         500,        2.0),
        (500,         1000,       1.0),
        (1000,        5000,      0.5),
        (5000,        10000,     0.3),
        (10000,       50000,     0.2),
        (50000,       100000,    0.1),
        (100000,     float('inf'), 0.05),
    ]
    details_rev, fee_rev = tiered_calc(revenue, rev_brackets)
    
    return {
        'total_assets': total_assets,
        'net_assets': net_assets,
        'revenue': revenue,
        'details_asset': details_asset,
        'fee_asset': fee_asset,
        'details_net': details_net,
        'fee_net': fee_net,
        'details_rev': details_rev,
        'fee_rev': fee_rev,
    }


def write_audit_sheet(wb, data):
    ws = wb.create_sheet("一、审计收费计算")
    set_col_widths(ws, [5, 18, 20, 22, 12, 22, 28])
    
    r = 1
    # 标题
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    apply_cell(ws, r, 1, "审计收费计算——川发改价格〔2013〕901号", FONT_TITLE, FILL_DARK_BLUE, ALIGN_CENTER)
    for c in range(2,8): apply_cell(ws, r, c, None, FONT_TITLE, FILL_DARK_BLUE, ALIGN_CENTER)
    
    r += 2
    # 基本信息
    info = [
        ("项目名称", "阿坝发展控股集团有限公司2025年度及2026年指定一期财务报表审计项目"),
        ("收费依据", "川发改价格〔2013〕901号《四川省会计师事务所服务收费标准》"),
        ("计算方法", "差额定率累进法（年度财务报表审计，按资产总额）"),
        ("总资产", f"{data['total_assets']:,.2f} 元（≈{data['total_assets']/100000000:,.2f}亿元）"),
        ("净资产", f"{data['net_assets']:,.2f} 元（≈{data['net_assets']/100000000:,.2f}亿元）"),
        ("营业收入", f"{data['revenue']:,.2f} 元（≈{data['revenue']/100000000:,.2f}亿元）"),
    ]
    for i, (k, v) in enumerate(info):
        apply_cell(ws, r, 1, k, FONT_BOLD, FILL_LIGHT, ALIGN_LEFT)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        apply_cell(ws, r, 2, v, FONT_NORMAL, FILL_LIGHT, ALIGN_LEFT)
        for c in range(3,8): apply_cell(ws, r, c, None, FONT_NORMAL, FILL_LIGHT, ALIGN_LEFT)
        r += 1
    
    r += 1
    
    # === 一、按资产总额计算 ===
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    apply_cell(ws, r, 1, "一、按资产总额差额定率累进计算（主方法）", FONT_SUBTITLE_WHITE, FILL_MID_BLUE, ALIGN_LEFT)
    for c in range(2,8): apply_cell(ws, r, c, None, FONT_SUBTITLE_WHITE, FILL_MID_BLUE, ALIGN_LEFT)
    r += 1
    
    headers = ["档次", "资产总额区间(万元)", "区间金额(万元)", "费率(‰)", "区间收费(元)", "计算公式"]
    for i, h in enumerate(headers, 1):
        apply_cell(ws, r, i, h, FONT_HEADER, FILL_DARK_BLUE, ALIGN_CENTER)
    r += 1
    
    for d in data['details_asset']:
        tier, bracket_str, amount, rate, fee = d
        apply_cell(ws, r, 1, tier, FONT_NORMAL, FILL_WHITE, ALIGN_CENTER)
        apply_cell(ws, r, 2, bracket_str, FONT_NORMAL, FILL_WHITE, ALIGN_CENTER)
        apply_cell(ws, r, 3, f"{amount:,.2f}", FONT_NORMAL, FILL_WHITE, ALIGN_RIGHT)
        apply_cell(ws, r, 4, f"{rate}", FONT_NORMAL, FILL_WHITE, ALIGN_CENTER)
        apply_cell(ws, r, 5, f"{fee:,.2f}", FONT_NORMAL, FILL_WHITE, ALIGN_RIGHT, number_format='#,##0.00')
        formula = f"{amount:,.2f}万 × 10000 × {rate}‰" if fee > 0 else "区间已过，不计"
        apply_cell(ws, r, 6, formula, FONT_SMALL, FILL_WHITE, ALIGN_LEFT)
        r += 1
    
    # 合计行
    apply_cell(ws, r, 1, "", FONT_BOLD, FILL_TOTAL)
    apply_cell(ws, r, 2, "合 计", FONT_BOLD, FILL_TOTAL, ALIGN_CENTER)
    apply_cell(ws, r, 3, f"{data['total_assets']/10000:,.2f}", FONT_BOLD, FILL_TOTAL, ALIGN_RIGHT)
    apply_cell(ws, r, 4, "—", FONT_BOLD, FILL_TOTAL, ALIGN_CENTER)
    apply_cell(ws, r, 5, f"{data['fee_asset']:,.2f}", FONT_BOLD, FILL_TOTAL, ALIGN_RIGHT, number_format='#,##0.00')
    apply_cell(ws, r, 6, f"大写：{num_to_chinese(data['fee_asset'])}", FONT_BOLD, FILL_TOTAL, ALIGN_LEFT)
    
    r += 2
    # === 二、按净资产计算（参考对比）===
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    apply_cell(ws, r, 1, "二、按净资产差额定率累进计算（参考对比）", FONT_SUBTITLE_WHITE, FILL_MID_BLUE, ALIGN_LEFT)
    for c in range(2,8): apply_cell(ws, r, c, None, FONT_SUBTITLE_WHITE, FILL_MID_BLUE, ALIGN_LEFT)
    r += 1
    
    for i, h in enumerate(headers, 1):
        apply_cell(ws, r, i, h, FONT_HEADER, FILL_DARK_BLUE, ALIGN_CENTER)
    r += 1
    
    for d in data['details_net']:
        tier, bracket_str, amount, rate, fee = d
        apply_cell(ws, r, 1, tier, FONT_NORMAL, FILL_WHITE, ALIGN_CENTER)
        apply_cell(ws, r, 2, bracket_str, FONT_NORMAL, FILL_WHITE, ALIGN_CENTER)
        apply_cell(ws, r, 3, f"{amount:,.2f}", FONT_NORMAL, FILL_WHITE, ALIGN_RIGHT)
        apply_cell(ws, r, 4, f"{rate}", FONT_NORMAL, FILL_WHITE, ALIGN_CENTER)
        apply_cell(ws, r, 5, f"{fee:,.2f}", FONT_NORMAL, FILL_WHITE, ALIGN_RIGHT, number_format='#,##0.00')
        apply_cell(ws, r, 6, f"{amount:,.2f}万 × 10000 × {rate}‰" if fee > 0 else "—", FONT_SMALL, FILL_WHITE, ALIGN_LEFT)
        r += 1
    
    apply_cell(ws, r, 1, "", FONT_BOLD, FILL_TOTAL)
    apply_cell(ws, r, 2, "合 计（净资产基准）", FONT_BOLD, FILL_TOTAL, ALIGN_CENTER)
    apply_cell(ws, r, 3, f"{data['net_assets']/10000:,.2f}", FONT_BOLD, FILL_TOTAL, ALIGN_RIGHT)
    apply_cell(ws, r, 4, "—", FONT_BOLD, FILL_TOTAL, ALIGN_CENTER)
    apply_cell(ws, r, 5, f"{data['fee_net']:,.2f}", FONT_BOLD, FILL_TOTAL, ALIGN_RIGHT, number_format='#,##0.00')
    apply_cell(ws, r, 6, f"大写：{num_to_chinese(data['fee_net'])}", FONT_BOLD, FILL_TOTAL, ALIGN_LEFT)
    
    r += 2
    # === 三、按收入计算（参考对比）===
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    apply_cell(ws, r, 1, "三、按收入差额定率累进计算（参考对比，经济效益审计类费率）", FONT_SUBTITLE_WHITE, FILL_MID_BLUE, ALIGN_LEFT)
    for c in range(2,8): apply_cell(ws, r, c, None, FONT_SUBTITLE_WHITE, FILL_MID_BLUE, ALIGN_LEFT)
    r += 1
    
    # 收入费率表头
    rev_headers = ["档次", "收入区间(万元)", "区间金额(万元)", "费率(‰)", "区间收费(元)", "计算公式"]
    for i, h in enumerate(rev_headers, 1):
        apply_cell(ws, r, i, h, FONT_HEADER, FILL_DARK_BLUE, ALIGN_CENTER)
    r += 1
    
    for d in data['details_rev']:
        tier, bracket_str, amount, rate, fee = d
        apply_cell(ws, r, 1, tier, FONT_NORMAL, FILL_WHITE, ALIGN_CENTER)
        apply_cell(ws, r, 2, bracket_str, FONT_NORMAL, FILL_WHITE, ALIGN_CENTER)
        apply_cell(ws, r, 3, f"{amount:,.2f}", FONT_NORMAL, FILL_WHITE, ALIGN_RIGHT)
        apply_cell(ws, r, 4, f"{rate}", FONT_NORMAL, FILL_WHITE, ALIGN_CENTER)
        apply_cell(ws, r, 5, f"{fee:,.2f}", FONT_NORMAL, FILL_WHITE, ALIGN_RIGHT, number_format='#,##0.00')
        apply_cell(ws, r, 6, f"{amount:,.2f}万 × 10000 × {rate}‰" if fee > 0 else "—", FONT_SMALL, FILL_WHITE, ALIGN_LEFT)
        r += 1
    
    apply_cell(ws, r, 1, "", FONT_BOLD, FILL_TOTAL)
    apply_cell(ws, r, 2, "合 计（收入基准）", FONT_BOLD, FILL_TOTAL, ALIGN_CENTER)
    apply_cell(ws, r, 3, f"{data['revenue']/10000:,.2f}", FONT_BOLD, FILL_TOTAL, ALIGN_RIGHT)
    apply_cell(ws, r, 4, "—", FONT_BOLD, FILL_TOTAL, ALIGN_CENTER)
    apply_cell(ws, r, 5, f"{data['fee_rev']:,.2f}", FONT_BOLD, FILL_TOTAL, ALIGN_RIGHT, number_format='#,##0.00')
    apply_cell(ws, r, 6, f"大写：{num_to_chinese(data['fee_rev'])}", FONT_BOLD, FILL_TOTAL, ALIGN_LEFT)
    
    r += 2
    # === 结论与建议 ===
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    apply_cell(ws, r, 1, "四、审计收费结论", FONT_BOLD, FILL_GOLD, ALIGN_LEFT)
    for c in range(2,8): apply_cell(ws, r, c, None, FONT_BOLD, FILL_GOLD, ALIGN_LEFT)
    r += 1
    
    conclusions = [
        f"1. 年度财务报表审计（按资产总额）标准收费：{data['fee_asset']:,.2f} 元（{num_to_chinese(data['fee_asset'])}）",
        f"2. 按净资产计算参考值：{data['fee_net']:,.2f} 元",
        f"3. 按收入计算参考值（经济效益审计费率）：{data['fee_rev']:,.2f} 元",
        "",
        "说明：",
        "• 川发改〔2013〕901号规定年度财务报表审计以「资产总额」作为主要计费基数",
        "• 净资产和收入计算仅作为参考对比",
        "• 以上为标准收费上限，实际收费可在标准基础上根据项目复杂度、风险等因素协商确定",
        "• 本合同为「控制价专项审核服务」而非审计业务本身，合同总价7,000元为控制价审核服务费",
    ]
    for line in conclusions:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        apply_cell(ws, r, 1, line, FONT_NORMAL if not line.startswith("•") else FONT_SMALL, FILL_WHITE, ALIGN_LEFT)
        for c in range(2,8): apply_cell(ws, r, c, None, FONT_NORMAL, FILL_WHITE, ALIGN_LEFT)
        r += 1
    
    return ws


# =============================================
# 二、资产评估收费计算
# =============================================
def calc_evaluation_fee():
    """
    四川省资产评估收费 —— 川发改价格〔2012〕827号
    差额定率累进
    
    资产评估收费费率表（按评估值）：
    | 档次 | 计费额度(万元) | 费率(‰) |
    | 1 | ≤100          | 8.0   |
    | 2 | 100-1000      | 3.75   |
    | 3 | 1000-5000     | 1.2    |
    | 4 | 5000-10000    | 0.75   |
    | 5 | 10000-100000  | 0.15   |
    | 6 | 100000以上    | 0.1    |
    
    注：房地产评估通常按评估值计算。用户提供了面积数据，需先估算价值或直接使用面积。
    由于面积汇总表中不含评估值，本表按：
    (1) 先列出各物业面积汇总
    (2) 按假设的保守评估单价估算总价值
    (3) 按差额定率累进计算评估费用
    """
    
    # 物业面积数据（来自阿坝州-汇总面积.xlsx）
    properties = [
        ("阿坝州国有资产投资管理有限责任公司", [
            (1, "马尔康市日瓦坝片区企业", 6699.21),
            (2, "马尔康市日瓦坝片区企业", 1951.55),
            (3, "马尔康市嘉绒广场", 108257.77),
            (4, "马尔康市公安局、交警队", 1518.02),
            (5, "罗汉山小区住宅", 967.23),
            (6, "禹隆商业", 25.86),
        ]),
        ("成都金牛区企业经营管理有限责任公司", [
            (7, "金牛区营门口片区、抚琴", 11879.08),
        ]),
        ("成都嘉银企业经营管理有限责任公司", [
            (8, "通用工业厂房", 42950.76),
        ]),
        ("阿坝州国有资产投资发展有限责任公司", [
            (9, "阿坝县库房", 2569.30),
            (10, "高原厂房", 3240.73),
        ]),
    ]
    
    total_area = 180059.51  # 总面积（㎡）
    
    # 没有评估值时的估算方法：面积法
    # 由于用户只提供了面积，合同中包含的评估要求是针对"投资性房地产"以公允价值计量
    # 典型单价估算（保守）：
    price_estimates = {
        "scenario_min": {"马尔康商业/办公": 3000, "马尔康住宅": 2000, "成都金牛商业": 8000, "成都工业厂房": 3000, "阿坝县库房": 1500, "高原厂房": 2000},
        "scenario_mid": {"马尔康商业/办公": 5000, "马尔康住宅": 3000, "成都金牛商业": 12000, "成都工业厂房": 4000, "阿坝县库房": 2000, "高原厂房": 2500},
        "scenario_max": {"马尔康商业/办公": 8000, "马尔康住宅": 5000, "成都金牛商业": 18000, "成都工业厂房": 5000, "阿坝县库房": 2500, "高原厂房": 3000},
    }
    
    # 按场景估算评估值
    def estimate_value(area_m2, scenario='mid'):
        """根据位置和类型估算评估价值"""
        prices = price_estimates[scenario]
        estimates = []
        
        for company, items in properties:
            for idx, name, area in items:
                # 判断物业类型
                if "嘉绒广场" in name:
                    price = prices["马尔康商业/办公"]
                elif "住宅" in name:
                    price = prices["马尔康住宅"]
                elif "金牛" in name or "营门口" in name or "抚琴" in name:
                    price = prices["成都金牛商业"]
                elif "工业" in name or "厂房" in name:
                    price = prices["成都工业厂房"]
                elif "库房" in name:
                    price = prices["阿坝县库房"]
                elif "高原" in name:
                    price = prices["高原厂房"]
                elif "商业" in name:
                    price = prices["马尔康商业/办公"]
                else:
                    price = prices["马尔康商业/办公"]
                
                value = area * price
                estimates.append((company, idx, name, area, price, value))
        
        return estimates
    
    # 评估费用费率（川发改2012-827号）
    eval_brackets = [
        (0,           100,        8.0),
        (100,         1000,       3.75),
        (1000,        5000,       1.2),
        (5000,        10000,      0.75),
        (10000,       100000,     0.15),
        (100000,      float('inf'), 0.1),
    ]
    
    def tiered_eval(base_amount, brackets):
        """评估费差额定率累进（base_amount单位：元）"""
        base_wan = base_amount / 10000
        details = []
        total_fee = 0
        remaining = base_wan
        
        for i, (lo, hi, rate) in enumerate(brackets):
            tier_amount = min(remaining, hi - lo)
            if tier_amount <= 0:
                details.append((i+1, f"{lo:,.0f}-{hi:,.0f}" if hi != float('inf') else f">{lo:,.0f}",
                               tier_amount, rate, 0))
                continue
            fee_yuan = tier_amount * 10000 * rate / 1000
            total_fee += fee_yuan
            details.append((i+1, f"{lo:,.0f}-{hi:,.0f}" if hi != float('inf') else f">{lo:,.0f}",
                           tier_amount, rate, fee_yuan))
            remaining -= tier_amount
            if remaining <= 0:
                break
        
        return details, total_fee
    
    # 三个场景
    scenarios = {}
    for key in ['scenario_min', 'scenario_mid', 'scenario_max']:
        est = estimate_value(0, key)
        total_val = sum(e[5] for e in est)
        details, fee = tiered_eval(total_val, eval_brackets)
        scenarios[key] = {
            'name': {'scenario_min': '保守估算', 'scenario_mid': '中性估算', 'scenario_max': '乐观估算'}[key],
            'estimates': est,
            'total_value': total_val,
            'details': details,
            'fee': fee,
        }
    
    return {
        'properties': properties,
        'total_area': total_area,
        'scenarios': scenarios,
        'price_estimates': price_estimates,
    }


def write_eval_sheet(wb, data):
    ws = wb.create_sheet("二、资产评估收费计算")
    set_col_widths(ws, [5, 20, 26, 14, 16, 18, 18])
    
    r = 1
    # 标题
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    apply_cell(ws, r, 1, "资产评估收费计算——川发改价格〔2012〕827号", FONT_TITLE, FILL_DARK_BLUE, ALIGN_CENTER)
    for c in range(2,8): apply_cell(ws, r, c, None, FONT_TITLE, FILL_DARK_BLUE, ALIGN_CENTER)
    
    r += 2
    info = [
        ("项目名称", "阿坝发展控股集团有限公司2025年资产评估项目"),
        ("收费依据", "川发改价格〔2012〕827号《四川省资产评估收费标准》"),
        ("计算方法", "差额定率累进法（按评估值）"),
        ("总面积", f"{data['total_area']:,.2f} 平方米"),
        ("涉及主体", "4家子公司，10处物业"),
    ]
    for i, (k, v) in enumerate(info):
        apply_cell(ws, r, 1, k, FONT_BOLD, FILL_LIGHT, ALIGN_LEFT)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        apply_cell(ws, r, 2, v, FONT_NORMAL, FILL_LIGHT, ALIGN_LEFT)
        for c in range(3,8): apply_cell(ws, r, c, None, FONT_NORMAL, FILL_LIGHT, ALIGN_LEFT)
        r += 1
    
    r += 1
    # === 物业面积汇总 ===
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    apply_cell(ws, r, 1, "一、资产评估物业面积汇总表", FONT_SUBTITLE_WHITE, FILL_MID_BLUE, ALIGN_LEFT)
    for c in range(2,8): apply_cell(ws, r, c, None, FONT_SUBTITLE_WHITE, FILL_MID_BLUE, ALIGN_LEFT)
    r += 1
    
    area_headers = ["序号", "产权单位", "项目名称", "面积(㎡)"]
    for i, h in enumerate(area_headers, 1):
        apply_cell(ws, r, i, h, FONT_HEADER, FILL_DARK_BLUE, ALIGN_CENTER)
    r += 1
    
    idx = 0
    for company, items in data['properties']:
        for num, name, area in items:
            idx += 1
            apply_cell(ws, r, 1, idx, FONT_NORMAL, FILL_WHITE, ALIGN_CENTER)
            apply_cell(ws, r, 2, company, FONT_NORMAL, FILL_WHITE, ALIGN_LEFT)
            apply_cell(ws, r, 3, name, FONT_NORMAL, FILL_WHITE, ALIGN_LEFT)
            apply_cell(ws, r, 4, f"{area:,.2f}", FONT_NORMAL, FILL_WHITE, ALIGN_RIGHT, number_format='#,##0.00')
            r += 1
    
    # 小计
    for c in range(1,5):
        fill = FILL_TOTAL if c != 1 else FILL_TOTAL
    apply_cell(ws, r, 1, "", FONT_BOLD, FILL_TOTAL)
    apply_cell(ws, r, 2, "合 计", FONT_BOLD, FILL_TOTAL, ALIGN_CENTER)
    apply_cell(ws, r, 3, f"共{idx}处物业", FONT_BOLD, FILL_TOTAL, ALIGN_CENTER)
    apply_cell(ws, r, 4, f"{data['total_area']:,.2f}", FONT_BOLD, FILL_TOTAL, ALIGN_RIGHT, number_format='#,##0.00')
    
    r += 2
    # === 评估费估算 ===
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    apply_cell(ws, r, 1, "二、资产评估费计算（差额定率累进法）", FONT_SUBTITLE_WHITE, FILL_MID_BLUE, ALIGN_LEFT)
    for c in range(2,8): apply_cell(ws, r, c, None, FONT_SUBTITLE_WHITE, FILL_MID_BLUE, ALIGN_LEFT)
    r += 1
    
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    apply_cell(ws, r, 1, "⚠ 注意：资产评估收费以「评估价值」为基数。因原始数据仅含面积、无评估值，下表按三种场景估算评估值后计算。实际计算请代入专业评估值。", 
              Font(name='微软雅黑', size=9, color='CC0000', bold=True), FILL_LIGHT, ALIGN_LEFT)
    for c in range(2,8): apply_cell(ws, r, c, None, FONT_SMALL, FILL_LIGHT, ALIGN_LEFT)
    r += 1
    
    # 费率表
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    apply_cell(ws, r, 1, "资产评估差额定率累进费率表（川发改〔2012〕827号）", FONT_BOLD, FILL_LIGHT, ALIGN_LEFT)
    for c in range(2,8): apply_cell(ws, r, c, None, FONT_BOLD, FILL_LIGHT, ALIGN_LEFT)
    r += 1
    
    rate_headers = ["档次", "计费额度(万元)", "差额计费率(‰)"]
    for i, h in enumerate(rate_headers, 1):
        apply_cell(ws, r, i, h, FONT_HEADER, FILL_DARK_BLUE, ALIGN_CENTER)
    r += 1
    
    eval_brackets = [
        (1, "≤100", "8.0"),
        (2, "100-1,000", "3.75"),
        (3, "1,000-5,000", "1.2"),
        (4, "5,000-10,000", "0.75"),
        (5, "10,000-100,000", "0.15"),
        (6, ">100,000", "0.1"),
    ]
    for tier, bstr, rate in eval_brackets:
        apply_cell(ws, r, 1, tier, FONT_NORMAL, FILL_WHITE, ALIGN_CENTER)
        apply_cell(ws, r, 2, bstr, FONT_NORMAL, FILL_WHITE, ALIGN_CENTER)
        apply_cell(ws, r, 3, rate, FONT_NORMAL, FILL_WHITE, ALIGN_CENTER)
        r += 1
    
    r += 1
    
    # 三种场景
    for sc_key, sc_data in data['scenarios'].items():
        r += 1
        sc_name = sc_data['name']
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        apply_cell(ws, r, 1, f"场景{sc_key[-3:]}：{sc_name}评估费计算", FONT_SUBTITLE, FILL_GOLD, ALIGN_LEFT)
        for c in range(2,8): apply_cell(ws, r, c, None, FONT_SUBTITLE, FILL_GOLD, ALIGN_LEFT)
        r += 1
        
        # 评估值明细
        est_headers = ["序号", "产权单位", "项目名称", "面积(㎡)", "估算单价(元/㎡)", "估算评估值(元)"]
        for i, h in enumerate(est_headers, 1):
            apply_cell(ws, r, i, h, FONT_HEADER, FILL_DARK_BLUE, ALIGN_CENTER)
        r += 1
        
        for idx_est, (company, num, name, area, price, value) in enumerate(sc_data['estimates'], 1):
            apply_cell(ws, r, 1, idx_est, FONT_SMALL, FILL_WHITE, ALIGN_CENTER)
            apply_cell(ws, r, 2, company, FONT_SMALL, FILL_WHITE, ALIGN_LEFT)
            apply_cell(ws, r, 3, name, FONT_SMALL, FILL_WHITE, ALIGN_LEFT)
            apply_cell(ws, r, 4, f"{area:,.2f}", FONT_SMALL, FILL_WHITE, ALIGN_RIGHT)
            apply_cell(ws, r, 5, f"{price:,.0f}", FONT_SMALL, FILL_WHITE, ALIGN_RIGHT)
            apply_cell(ws, r, 6, f"{value:,.2f}", FONT_SMALL, FILL_WHITE, ALIGN_RIGHT)  # number_format not supported in apply_cell custom args
            r += 1
        
        # 评估值合计
        apply_cell(ws, r, 1, "", FONT_BOLD, FILL_TOTAL)
        apply_cell(ws, r, 2, "评估值合计", FONT_BOLD, FILL_TOTAL, ALIGN_CENTER)
        apply_cell(ws, r, 3, f"≈ {sc_data['total_value']/100000000:,.2f} 亿元", FONT_BOLD, FILL_TOTAL, ALIGN_CENTER)
        apply_cell(ws, r, 4, f"{data['total_area']:,.2f}", FONT_BOLD, FILL_TOTAL, ALIGN_RIGHT)
        apply_cell(ws, r, 5, "—", FONT_BOLD, FILL_TOTAL, ALIGN_CENTER)
        apply_cell(ws, r, 6, f"{sc_data['total_value']:,.2f}", FONT_BOLD, FILL_TOTAL, ALIGN_RIGHT, number_format='#,##0.00')
        
        r += 1
        # 收费计算过程
        fee_headers = ["档次", "计费额度(万元)", "区间差额(万元)", "费率(‰)", "区间收费(元)", "计算公式"]
        for i, h in enumerate(fee_headers, 1):
            apply_cell(ws, r, i, h, FONT_HEADER, FILL_MID_BLUE, ALIGN_CENTER)
        r += 1
        
        for d in sc_data['details']:
            tier, bracket_str, amount, rate, fee = d
            apply_cell(ws, r, 1, tier, FONT_NORMAL, FILL_WHITE, ALIGN_CENTER)
            apply_cell(ws, r, 2, bracket_str, FONT_NORMAL, FILL_WHITE, ALIGN_CENTER)
            apply_cell(ws, r, 3, f"{amount:,.2f}", FONT_NORMAL, FILL_WHITE, ALIGN_RIGHT)
            apply_cell(ws, r, 4, f"{rate}", FONT_NORMAL, FILL_WHITE, ALIGN_CENTER)
            apply_cell(ws, r, 5, f"{fee:,.2f}", FONT_NORMAL, FILL_WHITE, ALIGN_RIGHT, number_format='#,##0.00')
            apply_cell(ws, r, 6, f"{amount:,.2f}万 × 10000 × {rate}‰" if fee > 0 else "—", FONT_SMALL, FILL_WHITE, ALIGN_LEFT)
            r += 1
        
        # 合计
        apply_cell(ws, r, 1, "", FONT_BOLD, FILL_TOTAL)
        apply_cell(ws, r, 2, "评估费合计", FONT_BOLD, FILL_TOTAL, ALIGN_CENTER)
        apply_cell(ws, r, 3, "", FONT_BOLD, FILL_TOTAL)
        apply_cell(ws, r, 4, "—", FONT_BOLD, FILL_TOTAL, ALIGN_CENTER)
        apply_cell(ws, r, 5, f"{sc_data['fee']:,.2f}", FONT_BOLD, FILL_TOTAL, ALIGN_RIGHT, number_format='#,##0.00')
        apply_cell(ws, r, 6, f"大写：{num_to_chinese(sc_data['fee'])}", FONT_BOLD, FILL_TOTAL, ALIGN_LEFT)
        r += 1
    
    r += 1
    # === 结论 ===
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    apply_cell(ws, r, 1, "三、资产评估收费结论与说明", FONT_SUBTITLE_WHITE, FILL_DARK_BLUE, ALIGN_LEFT)
    for c in range(2,8): apply_cell(ws, r, c, None, FONT_SUBTITLE_WHITE, FILL_DARK_BLUE, ALIGN_LEFT)
    r += 1
    
    conclusions = [
        f"• 保守估算（单价低）评估费：{data['scenarios']['scenario_min']['fee']:,.2f} 元",
        f"• 中性估算（单价中）评估费：{data['scenarios']['scenario_mid']['fee']:,.2f} 元",
        f"• 乐观估算（单价高）评估费：{data['scenarios']['scenario_max']['fee']:,.2f} 元",
        "",
        "⚠ 重要提示：",
        "1. 以上为基于面积的估值估算，实际评估费应以专业评估机构出具的评估值为基数计算",
        "2. 物业类型包含：商业广场、办公楼、住宅、工业厂房、库房等，不同物业估值差异大",
        "3. 成都金牛区物业及马尔康嘉绒广场（108,257㎡）为主要价值载体",
        "4. 评估费标准可上下浮动20%，最终以合同约定为准",
        "5. 本合同为「控制价专项审核服务」，合同总价7,000元系审核服务费，非评估服务费",
    ]
    for line in conclusions:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        font = FONT_SMALL if line.startswith("•") or line.startswith("⚠") else FONT_NORMAL
        apply_cell(ws, r, 1, line, font, FILL_WHITE, ALIGN_LEFT)
        for c in range(2,8): apply_cell(ws, r, c, None, font, FILL_WHITE, ALIGN_LEFT)
        r += 1
    
    return ws


# =============================================
# 辅助函数：数字转中文大写
# =============================================
def num_to_chinese(num):
    """数字转中文大写"""
    if abs(num) < 0.005:
        return "零元整"
    
    units = ['', '拾', '佰', '仟', '万', '拾', '佰', '仟', '亿', '拾', '佰', '仟', '万']
    digits = '零壹贰叁肆伍陆柒捌玖'
    
    yuan = int(num)
    jiao = int(round(num - yuan, 2) * 100) % 100 // 10
    fen = int(round(num - yuan, 2) * 100) % 10
    
    result = ""
    
    if yuan == 0:
        result = "零"
    else:
        yuan_str = str(yuan)
        n = len(yuan_str)
        zero_flag = False
        
        for i, ch in enumerate(yuan_str):
            digit = int(ch)
            unit_idx = n - i - 1
            unit = units[unit_idx]
            
            if digit == 0:
                zero_flag = True
                if unit in ['万', '亿']:
                    result += unit
                    zero_flag = False
            else:
                if zero_flag:
                    result += '零'
                    zero_flag = False
                result += digits[digit] + unit
        
        if yuan % 10000 < 1000 and yuan >= 10000:
            # handle edge case where 万 is followed by less than 1000
            pass
    
    result += '元'
    
    if jiao == 0 and fen == 0:
        result += '整'
    else:
        if jiao > 0:
            result += digits[jiao] + '角'
        if fen > 0:
            result += digits[fen] + '分'
    
    return result


def main():
    output_path = r"C:\Users\scrccpa\Desktop\阿坝发展控股_审计及评估收费计算.xlsx"
    
    wb = openpyxl.Workbook()
    # 删除默认sheet
    wb.remove(wb.active)
    
    # ===== Sheet 1: 审计收费 =====
    audit_data = calc_audit_fee()
    write_audit_sheet(wb, audit_data)
    
    # ===== Sheet 2: 评估收费 =====
    eval_data = calc_evaluation_fee()
    write_eval_sheet(wb, eval_data)
    
    # ===== 保存 =====
    wb.save(output_path)
    print(f"File saved to: {output_path}")
    print(f"\n=== Audit Fee Summary ===")
    print(f"Total Assets: {audit_data['total_assets']:,.2f} ({audit_data['total_assets']/1e8:,.2f}B)")
    print(f"Fee by Assets: {audit_data['fee_asset']:,.2f} CNY ({num_to_chinese(audit_data['fee_asset'])})")
    print(f"Fee by Net Assets ref: {audit_data['fee_net']:,.2f} CNY")
    print(f"Fee by Revenue ref: {audit_data['fee_rev']:,.2f} CNY")
    print(f"\n=== Valuation Fee Summary ===")
    print(f"Total Area: {eval_data['total_area']:,.2f} sqm")
    for key, sc in eval_data['scenarios'].items():
        print(f"{sc['name']}: Value ~{sc['total_value']/1e8:,.2f}B -> Fee {sc['fee']:,.2f} CNY")


if __name__ == '__main__':
    main()
