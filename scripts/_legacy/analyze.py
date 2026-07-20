import sys, os
sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

import openpyxl
from pathlib import Path

base = Path(r'C:\Users\scrccpa\Desktop\成都轨道资源资料')

# Core data files we need
data_files = []

# 1. 天府广场商户台账 (folder 1)
d1 = base / '业务资料' / '天府广场项目2026年专项审计 项目资料清单' / '1'
if d1.exists():
    for f in d1.iterdir():
        if f.suffix in ('.xlsx', '.xls'):
            data_files.append(('商户台账', str(f)))

# 2. 设备设施台账 (folder 2)
d2 = base / '业务资料' / '天府广场项目2026年专项审计 项目资料清单' / '2'
if d2.exists():
    for f in d2.iterdir():
        if f.suffix in ('.xlsx', '.xls'):
            data_files.append(('设施设备台账', str(f)))

# 3. 能耗数据 (folder 16, 17, 18)
for folder_idx in ['16', '17', '18']:
    d = base / '业务资料' / '天府广场项目2026年专项审计 项目资料清单' / folder_idx
    if d.exists():
        for f in d.iterdir():
            if f.suffix in ('.xlsx', '.xls'):
                data_files.append((f'能耗-文件夹{folder_idx}', str(f)))

# 4. 停车场收费数据 (folder 34)
d34 = base / '业务资料' / '天府广场项目2026年专项审计 项目资料清单' / '34'
if d34.exists():
    for f in d34.iterdir():
        if f.suffix in ('.xlsx', '.xls'):
            data_files.append(('停车场台账', str(f)))

# 5. 停车收费记录 (folder 35)
d35 = base / '业务资料' / '天府广场项目2026年专项审计 项目资料清单' / '35'
if d35.exists():
    for f in d35.iterdir():
        if f.suffix in ('.xlsx', '.xls'):
            data_files.append(('停车收费记录', str(f)))

# 6. 商家经营信息 (folder 28)
d28 = base / '业务资料' / '天府广场项目2026年专项审计 项目资料清单' / '28'
if d28.exists():
    for f in d28.iterdir():
        if f.suffix in ('.xlsx', '.xls'):
            data_files.append(('商家经营信息', str(f)))

# 7. 设施设备清单 ZIP contents (after extraction)
d_equip_zip = base / '业务资料' / '天府广场项目2026年专项审计 项目资料清单' / '2'
if d_equip_zip.exists():
    for f in d_equip_zip.iterdir():
        if f.suffix == '.zip':
            data_files.append(('设备设施清单(ZIP)', str(f)))

# Print all found data files
print('=== Data Files Found ===')
for category, path in data_files:
    size = os.path.getsize(path) // 1024
    print(f'[{category}] {Path(path).name} ({size}KB)')

# Now read the key Excel files
print('\n\n=== READING KEY FILES ===')

for category, path in data_files:
    if '停车收费记录' in category or '停车场' in category:
        print(f'\n\n### {category}: {Path(path).name} ###')
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            print(f'Sheets: {wb.sheetnames}')
            for sn in wb.sheetnames:
                ws = wb[sn]
                print(f'  {sn}: {ws.max_row} rows x {ws.max_column} cols')
                for r in range(1, min(ws.max_row+1, 25)):
                    row_data = []
                    for c in range(1, min(ws.max_column+1, 15)):
                        v = ws.cell(r, c).value
                        if v is not None:
                            row_data.append(str(v)[:50])
                    if row_data:
                        print(f'    R{r}: {" | ".join(row_data)}')
        except Exception as e:
            print(f'  ERROR: {e}')

for category, path in data_files:
    if '商户台账' in category or '经营信息' in category:
        print(f'\n\n### {category}: {Path(path).name} ###')
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            print(f'Sheets: {wb.sheetnames}')
            for sn in wb.sheetnames:
                ws = wb[sn]
                print(f'  {sn}: {ws.max_row} rows x {ws.max_column} cols')
                for r in range(1, min(ws.max_row+1, 30)):
                    row_data = []
                    for c in range(1, min(ws.max_column+1, 15)):
                        v = ws.cell(r, c).value
                        if v is not None:
                            row_data.append(str(v)[:50])
                    if row_data:
                        print(f'    R{r}: {" | ".join(row_data)}')
        except Exception as e:
            print(f'  ERROR: {e}')
