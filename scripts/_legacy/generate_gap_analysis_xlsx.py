#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成23篇文章技能差距分析Excel"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ==================== Sheet1: 技能差距总览 ====================
ws1 = wb.active
ws1.title = "技能差距总览"

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 标题
ws1.merge_cells("A1:F1")
ws1["A1"] = "23篇智能化审计文章 vs 融策现有技能差距分析"
ws1["A1"].font = Font(size=16, bold=True)
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 30

# 表头
headers = ["优先级", "缺失技能", "来源文章", "业务价值", "技术难度", "建议时间"]
for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=3, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align

# 数据
rows = [
    ["P0", "穿透式审计", "\"人工智能+穿透式\"医保审计实施路径研究", "高", "中", "1周"],
    ["P0", "专项债券审计", "政府专项债券领域常见问题分析及数智化审计路径", "高", "低", "1周"],
    ["P0", "BIM工程审计", "数字化转型背景下工程结算审计智能化应用研究", "高", "高", "2周"],
    ["P1", "动态审计预警", "基于大语言模型的动态审计预警机制研究", "高", "高", "1月"],
    ["P1", "风险画像", "基于审计风险\"画像\"和AI大模型的保险公司智能审计监督体系", "高", "中", "1月"],
    ["P1", "审计整改标准化", "构建审计整改\"3个1\"标准化体系", "中", "低", "2周"],
    ["P1", "RPA自动化", "流程和数智双轮驱动的交通投资企业内部审计数字化建设", "中", "中", "2月"],
    ["P1", "审计立项精准化", "数智技术赋能内部审计立项精准化的运行机制探讨", "中", "中", "1月"],
    ["P1", "反腐治乱审计", "新技术赋能审计监督推进反腐治乱的思考", "中", "中", "3月"],
    ["P2", "领导驾驶舱", "流程和数智双轮驱动的交通投资企业内部审计数字化建设", "中", "中", "3月"],
    ["P2", "国际审计标准", "欧洲审计院AI战略的审计智能化转型", "低", "低", "6月"],
    ["P2", "移动端审计", "多家企业案例", "低", "高", "6月"],
    ["P2", "人才培养体系", "审计数智化转型与人才培养适应性变革研究", "低", "低", "6月"],
]

priority_colors = {
    "P0": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
    "P1": PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid"),
    "P2": PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid"),
}

for i, row in enumerate(rows, 4):
    for j, val in enumerate(row, 1):
        cell = ws1.cell(row=i, column=j, value=val)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if j == 1:
            cell.fill = priority_colors.get(val, PatternFill())
            cell.font = Font(bold=True)

ws1.column_dimensions["A"].width = 8
ws1.column_dimensions["B"].width = 20
ws1.column_dimensions["C"].width = 50
ws1.column_dimensions["D"].width = 10
ws1.column_dimensions["E"].width = 10
ws1.column_dimensions["F"].width = 12

# ==================== Sheet2: 已有能力对照 ====================
ws2 = wb.create_sheet("已有能力对照")
ws2["A1"] = "融策已有技能/能力 vs 文章覆盖情况"
ws2["A1"].font = Font(size=14, bold=True)
ws2.merge_cells("A1:D1")

ws2_headers = ["技能名称", "对应文章", "覆盖状态", "备注"]
for col, h in enumerate(ws2_headers, 1):
    cell = ws2.cell(row=3, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align

covered_rows = [
    ["audit-report-review", "多篇提及报告质量", "✅ 已覆盖", "15维检查法领先"],
    ["procurement-audit-models", "招投标审计大数据技术应用", "✅ 已覆盖", "11层检测领先"],
    ["data-analyst-cn", "多篇文章数据分析", "✅ 已覆盖", "7大方法"],
    ["audit-jingze", "经济责任审计相关", "✅ 已覆盖", "量化评价v2.0"],
    ["financial-fraud-detection", "反腐治乱相关", "✅ 已覆盖", "Benford定律"],
    ["bid-document", "标书撰写", "✅ 已覆盖", "模板生成"],
    ["drawio", "流程图绘制", "✅ 已覆盖", "架构图"],
    ["deepseek-charting", "可视化", "✅ 已覆盖", "零代码图表"],
    ["audit-knowledge-graph", "知识图谱", "✅ 概念覆盖", "代码已生成"],
    ["unstructured-audit-data", "非结构化数据", "✅ 已覆盖", "文本挖掘"],
    ["RAG知识库", "知识管理", "✅ 已覆盖", "13,706 chunks"],
    ["智析v2.0-报告复核", "报告质量", "✅ 已覆盖", "规则+LLM"],
    ["智析v2.0-串标L8", "工商关联", "✅ 已覆盖", "本地+API"],
    ["audit-meeting-review", "会议记录", "✅ 已覆盖", "望闻问切"],
    ["regulatory-audit-response", "监管审计", "✅ 已覆盖", "接审报"],
]

for i, row in enumerate(covered_rows, 4):
    for j, val in enumerate(row, 1):
        cell = ws2.cell(row=i, column=j, value=val)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 30
ws2.column_dimensions["C"].width = 15
ws2.column_dimensions["D"].width = 25

# ==================== Sheet3: 技术维度对比 ====================
ws3 = wb.create_sheet("技术维度对比")
ws3["A1"] = "技术维度对比分析"
ws3["A1"].font = Font(size=14, bold=True)
ws3.merge_cells("A1:E1")

ws3_headers = ["技术维度", "已有能力", "缺失能力", "紧急度", "说明"]
for col, h in enumerate(ws3_headers, 1):
    cell = ws3.cell(row=3, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align

tech_rows = [
    ["LLM应用", "报告复核、知识问答", "动态预警、政策追踪", "🔴 高", "需数据中台支撑"],
    ["知识图谱", "概念技能", "实际图谱系统", "🟡 中", "代码已生成待验证"],
    ["机器学习", "部分数据分析", "立项预测、风险评分", "🟡 中", "需训练数据"],
    ["RPA", "无", "数据采集/底稿/报告机器人", "🟡 中", "需引入UiPath/影刀"],
    ["BIM", "无", "工程量自动算量/图纸比对", "🔴 高", "工程业务急需"],
    ["可视化", "基础图表", "驾驶舱、实时看板", "🟡 中", "管理决策支持"],
    ["移动端", "无", "现场审计APP", "🟢 低", "长期规划"],
    ["数据中台", "文件系统", "多源融合、自动ETL", "🔴 高", "全部智能化基础"],
]

for i, row in enumerate(tech_rows, 4):
    for j, val in enumerate(row, 1):
        cell = ws3.cell(row=i, column=j, value=val)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

ws3.column_dimensions["A"].width = 15
ws3.column_dimensions["B"].width = 25
ws3.column_dimensions["C"].width = 25
ws3.column_dimensions["D"].width = 10
ws3.column_dimensions["E"].width = 30

# ==================== Sheet4: 业务线分析 ====================
ws4 = wb.create_sheet("业务线分析")
ws4["A1"] = "按业务线分析缺失情况"
ws4["A1"].font = Font(size=14, bold=True)
ws4.merge_cells("A1:E1")

ws4_headers = ["业务线", "已有技能", "缺失技能", "影响程度", "行动建议"]
for col, h in enumerate(ws4_headers, 1):
    cell = ws4.cell(row=3, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align

biz_rows = [
    ["政府审计", "经责、绩效、专项", "穿透式、专项债全周期、整改标准化", "🔴 高", "立即创建3个技能"],
    ["工程审计", "标书、流程图", "BIM、工程量自动复核", "🔴 高", "对接工程咨询业务"],
    ["招投标审计", "11层检测", "工商关联(已部分)", "🟡 中", "配置天眼查API"],
    ["企业内审", "多Agent平台", "立项精准化、RPA、驾驶舱", "🟡 中", "找第一个企业客户"],
    ["数据分析", "7大方法、可视化", "动态预警、风险画像", "🟡 中", "需数据中台支撑"],
    ["反腐审计", "财务造假检测", "反腐治乱专项方法", "🟡 中", "聚焦权力运行监督"],
    ["知识管理", "RAG、知识图谱", "国际标准、人才培养", "🟢 低", "长期规划"],
]

for i, row in enumerate(biz_rows, 4):
    for j, val in enumerate(row, 1):
        cell = ws4.cell(row=i, column=j, value=val)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

ws4.column_dimensions["A"].width = 15
ws4.column_dimensions["B"].width = 25
ws4.column_dimensions["C"].width = 30
ws4.column_dimensions["D"].width = 10
ws4.column_dimensions["E"].width = 25

# ==================== Sheet5: 行动路线图 ====================
ws5 = wb.create_sheet("行动路线图")
ws5["A1"] = "技能创建行动路线图"
ws5["A1"].font = Font(size=14, bold=True)
ws5.merge_cells("A1:F1")

ws5_headers = ["阶段", "技能名称", "核心功能", "输入", "输出", "预计工时"]
for col, h in enumerate(ws5_headers, 1):
    cell = ws5.cell(row=3, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align

roadmap_rows = [
    ["立即(本周)", "penetrating-audit", "资金/项目/供应链全链条追踪", "审计项目类型", "穿透路径图+风险清单", "4h"],
    ["立即(本周)", "special-bond-audit", "四环节风险检查", "专项债项目资料", "检查清单+常见问题", "3h"],
    ["立即(本周)", "bim-engineering-audit", "BIM解析+工程量比对", "BIM模型+结算书", "差异分析+变更评估", "6h"],
    ["短期(1月)", "audit-risk-portrait", "风险画像+量化评分", "客户历史数据", "风险评分+标签+关注清单", "8h"],
    ["短期(1月)", "dynamic-audit-alert", "动态监测+自动预警", "审计数据+政策文件", "预警信号+异常报告", "10h"],
    ["短期(1月)", "audit-rectification", "整改标准化管理", "审计发现问题", "整改方案+跟踪计划", "4h"],
    ["中期(2月)", "rpa-audit-automation", "数据采集/底稿/报告机器人", "审计任务", "自动执行结果", "16h"],
    ["中期(3月)", "audit-dashboard", "领导驾驶舱可视化", "项目数据", "实时看板+决策支持", "12h"],
    ["长期(6月)", "anti-corruption-audit", "反腐治乱专项方法", "权力运行数据", "腐败线索+风险报告", "8h"],
]

for i, row in enumerate(roadmap_rows, 4):
    for j, val in enumerate(row, 1):
        cell = ws5.cell(row=i, column=j, value=val)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

ws5.column_dimensions["A"].width = 15
ws5.column_dimensions["B"].width = 25
ws5.column_dimensions["C"].width = 25
ws5.column_dimensions["D"].width = 20
ws5.column_dimensions["E"].width = 25
ws5.column_dimensions["F"].width = 12

# 保存
output_path = r"E:\2026\审计方法\智能化\analysis_output\23篇文章技能差距分析.xlsx"
wb.save(output_path)
print(f"Excel已保存: {output_path}")
