import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from copy import copy

PATH = r'C:\Users\scrccpa\Desktop\融策审计过程记录系统=项目经理版(6).xlsx'
wb = openpyxl.load_workbook(PATH)

# ============================================================
# PART 1: Sheet5 问题与证据
# ============================================================
ws5 = wb['5-问题与证据']

# Styles
header_font = Font(name='微软雅黑', bold=True, size=10)
body_font = Font(name='微软雅黑', size=9)
wrap = Alignment(wrap_text=True, vertical='top')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
red_fill = PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFDE0', end_color='FFFDE0', fill_type='solid')

# Issues data - consolidated from all 25 findings across projects
issues = [
    {
        'id': 'RC-2025-001',
        'category': '合同履约',
        'severity': '严重',
        'desc': '多个科创课程/科技服务项目合同履约严重不足。涉及4个项目，无人机/机器人课程实际服务次数均未达到合同约定标准：2024科学院开放日完成率仅33%（合同15次实5次）、2025上半年科技服务完成率60%（合同47次实28次）、2025下半年科技服务完成率56%（合同25次实14次）、5街学生科创课程完成率52%（合同25次实13次）。',
        'ev1': '合同原件：明确约定无人机/机器人服务次数',
        'ev2': '教师签到花名册及学生签到表：实际开展次数远低于合同约定',
        'ev3': '财务支付凭证：款项已按合同全额支付，但服务未足额履行',
        'law': '《政府购买服务管理办法》第二十六条、《政府采购法》第五十条',
        'suggestion': '1.逐项核实各项目实际服务次数与合同差异；2.要求供应商书面说明未履约原因；3.根据实际服务量重新核算费用，追回多付资金；4.将供应商履约不良记录纳入后续采购评价。',
        'dept': '被审计单位（成都教科院附中）',
        'status': '待整改'
    },
    {
        'id': 'RC-2025-002',
        'category': '采购程序',
        'severity': '严重',
        'desc': '各年度科创课程/科技服务采购项目普遍未按规范开展比选/比价需求论证，前期邀请比价环节缺少完整流程佐证材料，询价比价全过程未规范留痕。涉及2021-2025年共计7个以上子项目，属于系统性问题。',
        'ev1': '各项目财务凭证附件：均未见比选需求论证文件',
        'ev2': '现场核查记录：询价/比价环节流程佐证材料普遍缺失',
        'ev3': '3家报价单：均未标注报价签署时间，资料要件不全',
        'law': '《行政事业单位内部控制规范（试行）》第三十六条、《政府采购非招标采购方式管理办法》',
        'suggestion': '1.建立比选/比价标准化操作流程和模板；2.明确比选需求论证为采购前置必经环节；3.制定过程留痕清单（邀约记录、报价接收登记、评审记录、结果确认），逐项归档。',
        'dept': '被审计单位（成都教科院附中）',
        'status': '待整改'
    },
    {
        'id': 'RC-2025-003',
        'category': '验收手续',
        'severity': '较重',
        'desc': '多个项目验收单供应商未加盖公章确认，验收必备手续不全；拨款审批单上会计、出纳均未签字。涉及：2025下半年科创课程服务、2025年科技服务项目、2025上半年科技服务项目、科学院开放日等多个项目。',
        'ev1': '各项目验收单原件：供应商盖章栏均为空白',
        'ev2': '财务拨款单及附件：会计、出纳签字栏空白',
        'ev3': '现场核查记录：经办人和领导已签字，但财务审核缺失',
        'law': '《会计法》第十四条、《行政事业单位内部控制规范》第三十条',
        'suggestion': '1.立即补全验收单供应商盖章；2.完善拨款审批会签流程，确保会计、出纳逐级审核签字；3.将验收盖章和财务签字作为付款前置必要条件。',
        'dept': '被审计单位（成都教科院附中）',
        'status': '待整改'
    },
    {
        'id': 'RC-2025-004',
        'category': '资料归档',
        'severity': '较重',
        'desc': '供应商培训实施方案、课程教案、教师资质证明、学生签到花名册、教师签到册等配套服务资料普遍缺失。竞争性磋商响应文件及汇编材料未归档。涉及科创课程历年项目、科学院开放日等多个项目。',
        'ev1': '各项目档案卷宗：培训实施方案无一留存',
        'ev2': '现场核查：课程教案、教师资质、签到册均未归档',
        'ev3': '竞争性磋商项目：响应文件和汇编材料缺失',
        'law': '《政府采购法》第四十二条、《档案法》第十三条',
        'suggestion': '1.限期要求供应商补交培训实施方案和教师资质证明；2.建立服务类采购资料归档清单（合同、方案、教案、资质、签到册、验收报告、影像资料）；3.将资料归档完整性纳入合同验收条款。',
        'dept': '被审计单位（成都教科院附中）',
        'status': '待整改'
    },
    {
        'id': 'RC-2025-005',
        'category': '信息真实性',
        'severity': '严重',
        'desc': '科创课程（科创环保、火星移民、航天创客营）上课时间真实性存疑。2025年5月22日-31日周期内涉及周末及节假日的课程共15次（5月24、25、31日及6月1、2日）。2022年航天创客营课程11月5、12、19日也均在周末。需核实学校是否在周末及节假日正常行课。',
        'ev1': '课程签到表：签到日期含大量周末及节假日',
        'ev2': '学校校历/放假通知：待获取核实',
        'ev3': '相关教师访谈：待执行',
        'law': '《财政违法行为处罚处分条例》第六条',
        'suggestion': '1.获取学校2022-2025年完整校历及放假安排；2.访谈相关教师及学生，核实周末是否实际开课；3.如确认未上课但已付款，追回对应款项；4.建议公安机关介入如涉及虚假材料骗取财政资金。',
        'dept': '被审计单位（成都教科院附中）',
        'status': '待核实'
    },
    {
        'id': 'RC-2025-006',
        'category': '采购程序',
        'severity': '较重',
        'desc': '2022年科创课程火箭模型采购中，询价登记时间（2022年3月15日）发生在学校行政办公会议审定费用（2022年5月13日）之前，时间逻辑倒挂。未审先采，程序违规。金额：33,600元。',
        'ev1': '询价登记表：日期2022年3月15日',
        'ev2': '学校会议记录（教科附纪〔2022〕14号）：审定日期2022年5月13日',
        'ev3': '采购合同及付款凭证：合同签订于2022年3月21日',
        'law': '《行政事业单位内部控制规范》关于"先审批后执行"的规定',
        'suggestion': '1.核实是否存在"先采后批"问题；2.如确认程序倒挂，相关审批人员应说明原因；3.对同类项目排查是否存在类似程序倒挂。',
        'dept': '被审计单位（成都教科院附中）',
        'status': '待整改'
    },
    {
        'id': 'RC-2025-007',
        'category': '采购程序',
        'severity': '较重',
        'desc': '2025年科技服务项目中，项目立项名称、合同时段标注、供应商报价单内容三者互不一致：①上半年科技服务立项为"2025年上半年科创课程采购项目"，实际服务内容对应下半年；②正式合同时段标注为"2024-2025学年上期"，与本次采购不符；③供应商报价单内容错配，出现2025-2026学年内容。信息混乱，采购管理失控。',
        'ev1': '立项文件：项目名称为"2025年上半年科创课程采购项目"',
        'ev2': '正式合同：服务规格标注为"2024-2025学年上期"',
        'ev3': '供应商报价单：成都蓝色时代报价单规格栏出现"2025-2026学年上期"内容',
        'law': '《政府采购法》第四十六条、《政府采购货物和服务招标投标管理办法》',
        'suggestion': '1.核实项目真实采购需求和服务时段；2.修正合同及报价文件中的时段错误；3.排查是否存在不同项目混用或串用文件的情况；4.完善采购文件审核机制，杜绝名称、时段错配。',
        'dept': '被审计单位（成都教科院附中）',
        'status': '待整改'
    },
    {
        'id': 'RC-2025-008',
        'category': '采购程序',
        'severity': '一般',
        'desc': '2025年12月科技服务项目中，合同签订（9月24日）、询价、验收程序的时间均早于正式活动通知（10月17日），活动于10月24日举行，通知到活动仅7天。采购程序与活动组织节奏不匹配，程序合理性存疑。',
        'ev1': '服务采购合同：签订日期2025年9月24日',
        'ev2': '活动通知：成都高新区教育发展中心2025年10月17日发文',
        'ev3': '验收表：验收日期2025年10月23日（活动前一天即验收）',
        'law': '《政府采购法》关于采购程序的时序要求',
        'suggestion': '1.核实9月24日签订合同时是否已确定举办活动；2.确认活动前一日验收的合理性；3.排查是否存在"先定供应商再走流程"的情况。',
        'dept': '被审计单位（成都教科院附中）',
        'status': '待核实'
    },
    {
        'id': 'RC-2025-009',
        'category': '合同履约',
        'severity': '一般',
        'desc': '科创课程方案要求3个班共计30次课程，但实际服务内容仅对应2个班。方案与执行不一致，采购需求未落实。',
        'ev1': '询价方案/需求文件：标注3个班30次课程',
        'ev2': '实际课程签到表：仅有2个班记录',
        'ev3': '合同文件：需核实约定的班级数量',
        'law': '《政府购买服务管理办法》关于按合同约定履行的规定',
        'suggestion': '1.核实合同约定的班级数量与方案是否一致；2.确认实际服务的2个班是否为合同约定；3.如有差额，按比例调整支付金额。',
        'dept': '被审计单位（成都教科院附中）',
        'status': '待核实'
    },
]

# Write to Sheet5
# Start from row 3 (row 1 is title, row 2 is header)
start_row = 3
for i, issue in enumerate(issues):
    row = start_row + i
    ws5.cell(row=row, column=1, value=issue['id']).font = body_font
    ws5.cell(row=row, column=2, value=issue['category']).font = body_font
    ws5.cell(row=row, column=3, value=issue['severity']).font = body_font
    ws5.cell(row=row, column=4, value=issue['desc']).font = body_font
    ws5.cell(row=row, column=5, value=issue['ev1']).font = body_font
    ws5.cell(row=row, column=6, value=issue['ev2']).font = body_font
    ws5.cell(row=row, column=7, value=issue['ev3']).font = body_font
    ws5.cell(row=row, column=8, value=issue['law']).font = body_font
    ws5.cell(row=row, column=9, value=issue['suggestion']).font = body_font
    ws5.cell(row=row, column=10, value=issue['dept']).font = body_font
    ws5.cell(row=row, column=11, value=issue['status']).font = body_font
    
    # Apply formatting
    for col in range(1, 12):
        cell = ws5.cell(row=row, column=col)
        cell.alignment = wrap
        cell.border = thin_border
        # Color severity
        if issue['severity'] == '严重':
            cell.fill = red_fill
        elif issue['severity'] == '较重':
            cell.fill = yellow_fill

# Set column widths for Sheet5
col_widths_5 = [14, 12, 10, 50, 35, 35, 35, 30, 40, 22, 10]
for i, w in enumerate(col_widths_5, 1):
    ws5.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# ============================================================
# PART 2: Sheet6 复核记录
# ============================================================
ws6 = wb['6-复核记录']

# Unmerge all merged cells in Sheet6 before writing
for merge_range in list(ws6.merged_cells.ranges):
    ws6.unmerge_cells(str(merge_range))

# 一级复核 data
review_items = {
    4:  ('底稿完整性（索引连续）', '否', '有效记录127条中有实质内容的约120条，但序号存在断号（从中途起记录无序号），资料台账(Sheet3)完全空白，勾稽关系未建立'),
    5:  ('审计程序执行与方案一致性', '否', '仅有资料审阅和资料查阅两种程序，未见访谈、函证、实地考察、数据分析等其他审计程序。方案中的访谈记录(Sheet4)完全空白'),
    6:  ('证据充分性与适当性', '否', '发现25条问题，但"是否异常"字段0条标记，"审计判断/结论"字段100%空白。问题未按三证据原则整理，证据分散在过程记录中未汇总'),
    7:  ('发现与证据的勾稽关系', '否', '审计发现仅记录在"审计发现/核查情况"列，未与问题清单(Sheet5)建立勾稽。缺少从发现→证据→结论的完整链条'),
    8:  ('结论恰当性', '否', '所有记录的"审计判断/结论"列均为空。发现的问题没有给出定性结论（合规/不合规/存疑/待核实），无法判断结论的恰当性'),
    9:  ('建议可行性', '否', '仅现场核查记录中有零散建议性描述，无系统的整改建议。问题清单(Sheet5)原为空'),
    10: ('签字齐全性', '否', '执行人字段100%为空，无法追溯工作责任人。拨款审批单等原始资料本身存在会计出纳未签字的情况'),
    11: ('异常发现深入度', '否', '现场核查（5.15日记录）质量明显高于早期记录，发现更结构化。但异常标记字段形同虚设，0条标记异常，也未对异常发现追加深入程序'),
    12: ('资料记录完整性', '否', '日期格式不统一（有的写完整日期，有的简写月日），大量记录缺日期。资料台账(Sheet3)空白。课程教案、教师资质、签到册等关键资料均未归档'),
    13: ('访谈记录完整性', '否', '访谈记录(Sheet4)完全空白，审计过程中未见任何访谈活动的记录'),
    14: ('数据分析过程可复现', '否', '全为资料审阅/查阅，未见数据分析程序。无可复现的数据分析过程记录'),
}

# Fill 一级复核
for row_num, (item, result, opinion) in review_items.items():
    ws6.cell(row=row_num, column=1, value=item).font = body_font
    if result == '否':
        ws6.cell(row=row_num, column=2, value='').font = body_font
        ws6.cell(row=row_num, column=3, value='否').font = body_font
        ws6.cell(row=row_num, column=4, value='').font = body_font
    ws6.cell(row=row_num, column=5, value=opinion).font = body_font
    for col in range(1, 9):
        c = ws6.cell(row=row_num, column=col)
        c.alignment = wrap
        c.border = thin_border
        if result == '否' and col in [5]:
            c.fill = yellow_fill

# 二级复核
l2_items = {
    16: ('审计结论依据充分性', '否', '审计判断/结论列全部空白，无法评价结论依据是否充分。建议补全结论后再进行二级复核'),
    17: ('法规引用准确性', '否', '过程记录和发现中未引用任何法规条文，无法评价法规引用准确性'),
    18: ('问题定性恰当性', '否', '问题未经正式定性，分散在过程记录中。已整理至Sheet5的问题清单尚待二级复核确认定性'),
    19: ('整改建议具体可行性', '否', '原记录中缺少系统整改建议。已整理至Sheet5的建议方案待二级复核确认可行性'),
    20: ('一级复核问题整改闭环', '否', '一级复核完成确认后，需逐项跟踪整改闭环'),
}
for row_num, (item, result, opinion) in l2_items.items():
    ws6.cell(row=row_num, column=1, value=item).font = body_font
    if result == '否':
        ws6.cell(row=row_num, column=2, value='').font = body_font
        ws6.cell(row=row_num, column=3, value='否').font = body_font
        ws6.cell(row=row_num, column=4, value='').font = body_font
    ws6.cell(row=row_num, column=5, value=opinion).font = body_font
    for col in range(1, 9):
        c = ws6.cell(row=row_num, column=col)
        c.alignment = wrap
        c.border = thin_border
        if result == '否' and col in [5]:
            c.fill = yellow_fill

# 三级复核
l3_items = {
    22: ('整体审计质量', '否', '审计程序单一（仅有资料审阅），关键字段（结论、异常标记、执行人）空白，问题未系统汇总，访谈/数据分析/实地考察均缺失。整体审计深度和规范性有待大幅提升'),
    23: ('审计报告质量', 'N/A', '审计报告尚未出具，待一二复核完成后再进行报告质量评审'),
    24: ('重大发现处理', '待定', '合同履约严重不足（完成率33%-60%）、课程周末上课真实性存疑、采购程序系统性不规范，上述重大发现需进一步核实后确定处理方案'),
    25: ('一二复核闭环', '否', '一二复核均尚未实质性开展，待完成后方可确认闭环'),
    26: ('是否同意出具报告', '否', '当前阶段不具备出报告条件。需完成：①补全审计判断/结论；②整理问题清单三证据；③完成一二复核；④补做访谈和数据分析程序'),
}
for row_num, (item, result, opinion) in l3_items.items():
    ws6.cell(row=row_num, column=1, value=item).font = body_font
    if result == '否':
        ws6.cell(row=row_num, column=2, value='').font = body_font
        ws6.cell(row=row_num, column=3, value='否').font = body_font
        ws6.cell(row=row_num, column=4, value='').font = body_font
    elif result == 'N/A':
        ws6.cell(row=row_num, column=2, value='').font = body_font
        ws6.cell(row=row_num, column=3, value='').font = body_font
        ws6.cell(row=row_num, column=4, value='N/A').font = body_font
    elif result == '待定':
        ws6.cell(row=row_num, column=2, value='').font = body_font
        ws6.cell(row=row_num, column=3, value='').font = body_font
        ws6.cell(row=row_num, column=4, value='').font = body_font
    ws6.cell(row=row_num, column=5, value=opinion).font = body_font
    for col in range(1, 9):
        c = ws6.cell(row=row_num, column=col)
        c.alignment = wrap
        c.border = thin_border
        if result in ['否', '待定'] and col in [5]:
            c.fill = yellow_fill if result == '否' else PatternFill()

# Set column widths for Sheet6
col_widths_6 = [35, 8, 8, 8, 55, 12, 12, 25]
for i, w in enumerate(col_widths_6, 1):
    ws6.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# ============================================================
# Save
# ============================================================
out_path = r'C:\Users\scrccpa\Desktop\融策审计过程记录系统=项目经理版(6)-已复核.xlsx'
wb.save(out_path)
print(f"✅ 完成！")
print(f"  Sheet5 问题与证据: {len(issues)}条问题已填入")
print(f"  Sheet6 复核记录: 一级复核11项、二级复核5项、三级复核5项已填入")
