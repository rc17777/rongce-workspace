# -*- coding: utf-8 -*-
"""生成融策数据资产化差距分析 Excel 版"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = openpyxl.Workbook()

# ====== 样式 ======
header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='0A1F3F', end_color='0A1F3F', fill_type='solid')
sub_fill = PatternFill(start_color='1A5C6E', end_color='1A5C6E', fill_type='solid')
gold_fill = PatternFill(start_color='C5955C', end_color='C5955C', fill_type='solid')
warm_fill = PatternFill(start_color='F5F2EC', end_color='F5F2EC', fill_type='solid')
red_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
green_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
orange_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
blue_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')

title_font = Font(name='微软雅黑', size=16, bold=True, color='0A1F3F')
subtitle_font = Font(name='微软雅黑', size=11, color='666666')
normal_font = Font(name='微软雅黑', size=10)
bold_font = Font(name='微软雅黑', size=10, bold=True)
big_bold = Font(name='微软雅黑', size=12, bold=True, color='0A1F3F')
red_font = Font(name='微软雅黑', size=10, color='CC0000', bold=True)
green_font = Font(name='微软雅黑', size=10, color='2E7D32', bold=True)

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_top = Alignment(wrap_text=True, vertical='top')
wrap_center = Alignment(wrap_text=True, vertical='center', horizontal='center')
center_align = Alignment(horizontal='center', vertical='center')

def style_header(ws, row, cols, fill=header_fill):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = wrap_center
        cell.border = thin_border

def dc(ws, row, col, val, font=normal_font, fill=None, align=None):
    """data cell helper"""
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = font
    cell.alignment = align or wrap_top
    cell.border = thin_border
    if fill:
        cell.fill = fill
    return cell

def title_row(ws, row, title, subtitle=''):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value=title).font = title_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 40
    if subtitle:
        ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=8)
        ws.cell(row=row+1, column=1, value=subtitle).font = subtitle_font
        ws.cell(row=row+1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row+1].height = 22
        return row + 2
    return row + 1

# ==========================================
# Sheet 1: 总览
# ==========================================
ws = wb.active
ws.title = '总览'
sr = title_row(ws, 1, '融策 vs 审计数据资产化 · 差距分析', '基于微信公众号《审计数据资产化：从归档到闭环》文章对标')

# 核心区别
sr += 1
ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=8)
dc(ws, sr, 1, '⚠️ 核心区别：知识管理 vs 经验闭环', big_bold, red_fill, wrap_top)
ws.row_dimensions[sr].height = 30

items_overview = [
    ('', '你以为的数据资产化', '文章说的数据资产化', '融策现状'),
    ('定位', '把底稿/报告/法规存整齐 → 建知识库 → 可检索', '把每次项目实践中的经验结构化 → 跨项目关联 → 下次项目自动提醒', '前者已做到，后者刚起步'),
    ('类比', '图书馆：把书整理好，方便找', '教练：记录每次训练数据，指导下一次训练', '有很好的图书馆，还没有教练'),
    ('融策匹配度', 'RAG知识库13,977 chunks ✅\nPARA分类体系 ✅\nObsidian知识管理 ✅', '复核意见结构化记录 ❌\n调整分录跨项目分析 ❌\n资料缺失历史追踪 ❌', '基础设施超前，操作习惯滞后'),
]
for i, (a, b, c, d) in enumerate(items_overview):
    row = sr + 1 + i
    dc(ws, row, 1, a, bold_font, warm_fill if i==0 else None)
    dc(ws, row, 2, b, red_font if i==0 else normal_font, green_fill if i==0 else None)
    dc(ws, row, 3, c, red_font if i==0 else normal_font, red_fill if i==0 else None)
    dc(ws, row, 4, d, red_font if i==0 else normal_font, blue_fill if i==0 else None)
    ws.row_dimensions[row].height = 45 if i > 0 else 30

# 三个闭环总览
sr2 = row + 2
dc(ws, sr2, 1, '三大闭环差距总览', big_bold, gold_fill)
ws.merge_cells(start_row=sr2, start_column=1, end_row=sr2, end_column=8)
ws.row_dimensions[sr2].height = 30

headers = ['闭环', '核心问题', '文章要求', '融策现状', '差距等级', '启动难度', '最近一步', '预计见效']
for c, h in enumerate(headers, 1):
    ws.cell(row=sr2+1, column=c, value=h)
style_header(ws, sr2+1, len(headers))

overview_data = [
    ('闭环一\n复核→模板', '底稿老被打回，\n是老问题还是模板问题？',
     '1.结构化复核意见\n2.统计高频问题\n3.反向优化模板',
     '有复核行为，但意见散落在邮件/微信/口头，未集中记录，未分类，未统计',
     '⭐⭐ 中', '⭐ 低', '打印6个标签贴屏幕旁\n从下一份底稿开始贴', '1个月见统计\n3个月改模板'),
    ('闭环二\n调整→风险', '同类项目历史上有\n哪些典型调整？',
     '1.标注调整类型+行业\n2.跨项目规律挖掘\n3.新项目启动时提醒',
     '调整分录在底稿/报告中，但分散在各项目文件夹，未标注类型，未跨项目关联',
     '⭐⭐⭐ 高', '⭐⭐ 中', 'Excel表填调整分录\n(项目名+行业+类型+科目+金额)', '3个月攒数据\n6个月见规律'),
    ('闭环三\n清单→提醒', '客户每年同样的资料\n追三遍？',
     '1.记录资料缺失历史\n2.动态更新资料清单\n3.新项目自动提醒',
     '靠项目经理/老员工记忆，资料清单可能是固定模板',
     '⭐⭐ 中', '⭐ 低', '资料清单模板加一列\n"往年缺失记录"', '1个月见记录\n3个月见提醒'),
]
for i, row_data in enumerate(overview_data):
    row = sr2 + 2 + i
    for c, val in enumerate(row_data, 1):
        fill = None
        font = normal_font
        if c == 5:
            fill = red_fill if '高' in str(val) else orange_fill
            font = bold_font
        elif c == 6:
            fill = green_fill if '低' in str(val) else orange_fill
        elif c == 1:
            font = bold_font
        dc(ws, row, c, val, font, fill)
    ws.row_dimensions[row].height = 80

# 融策优势清单
sr3 = row + 2
dc(ws, sr3, 1, '融策已领先的基础设施（绝大多数事务所没有）', big_bold, green_fill)
ws.merge_cells(start_row=sr3, start_column=1, end_row=sr3, end_column=8)
ws.row_dimensions[sr3].height = 30

advantages = [
    'RAG审计知识库（13,977 chunks + DeepSeek API）—— 一般事务所：零',
    '审计黑板多Agent平台（7 Agent + 调度中枢 + 项目工作区标准化）—— 一般事务所：零',
    '79个技能 + 场景路由 + 执行追踪 —— 一般事务所：零',
    'PARA四层知识分类体系 + Obsidian集成 —— 一般事务所：零',
    'Token预算 + 费用守卫 + API限流 —— 一般事务所：零',
    '模型路由v4.0（错误代价六级路由 + 12模型池 + 双签制）—— 一般事务所：零',
]
for i, adv in enumerate(advantages):
    row = sr3 + 1 + i
    dc(ws, row, 1, adv, green_font, green_fill)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.row_dimensions[row].height = 25

# 一句话总结
sr4 = row + 2
dc(ws, sr4, 1, '💡 一句话总结', big_bold, gold_fill)
ws.merge_cells(start_row=sr4, start_column=1, end_row=sr4, end_column=8)
ws.row_dimensions[sr4].height = 30

dc(ws, sr4+1, 1, '融策和"审计数据资产化"之间的距离，不是技术距离，不是资金距离，甚至不是能力距离——是习惯距离。\n三个闭环需要的所有操作，都可以在一个Excel文件里完成。唯一需要的是：从下一个项目开始，每次审底稿多花15秒贴一个标签，每次项目结束多花15分钟填一张表。', bold_font, warm_fill)
ws.merge_cells(start_row=sr4+1, start_column=1, end_row=sr4+1, end_column=8)
ws.row_dimensions[sr4+1].height = 50

ws.column_dimensions['A'].width = 16
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 30
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 24
ws.column_dimensions['H'].width = 18

# ==========================================
# Sheet 2: 闭环一
# ==========================================
ws2 = wb.create_sheet('闭环一-复核→模板')
sr = title_row(ws2, 1, '闭环一：复核意见 → 问题分类 → 模板优化', '如果某个底稿每年都被打回，可能不是新人的问题，是模板的问题')

headers2 = ['维度', '文章要求', '融策现状', '差距描述', '根本原因', '补救难度', '启动措施', '预期产出']
sr += 1
for c, h in enumerate(headers2, 1):
    ws2.cell(row=sr, column=c, value=h)
style_header(ws2, sr, len(headers2))

data2 = [
    ('复核意见\n是否有记录', '每次审核底稿的意见被保留', '有复核过程，但意见分散在\n邮件/微信/口头沟通中，\n未集中存储', '有行为无记录——\n意见用完就丢了', '复核是"沟通"而非"记录"', '⭐ 低', '建一个共享Excel/在线表格\n每次复核随手记录', '1个月后有\n第一批结构化数据'),
    ('复核意见\n是否结构化', '每条意见标注问题类型\n（金额不一致/依据不充分/\n说明不清楚/勾稽未检查/\n结论太虚/格式问题）', '未分类标注', '完全没有分类——\n不知道哪种问题最多', '没想到要分类;\n没有分类标准', '⭐ 低', '打印6个标签贴在屏幕旁\n每次审底稿选一个贴上', '即时可见\n马上有分类意识'),
    ('是否跨项目\n统计分析', '按底稿类型×问题类型\n×频次统计\n找出最常被打回的底稿', '未做', '不知道哪些底稿\n是系统性有问题\nvs.个别人犯错', '缺乏跨项目视角;\n没有集中数据源', '⭐ 低\n(需要先攒数据)', '每季度汇总一次:\n底稿类型×问题类型×次数\n→改最严重的', '3个月后出\n第一份统计'),
    ('是否反向\n优化模板', '统计结果用来改进底稿模板:\n增加校验规则/填写提示/\n红色标注', '模板可能多年未更新，\n或凭感觉改', '有改进无依据——\n不知道改哪里、\n改了有没有效果', '没数据支撑改进决策;\n模板改进缺乏责任人', '⭐⭐ 中', '改打回率最高的Top3底稿\n加自动勾稽公式\n加"请填写XX"提示', '3个月后\n打回率下降\n可量化'),
    ('模板是否有\n校验规则', '模板中嵌入自动\n勾稽检查/自动提醒/\n必填项校验', '审计黑板有一些，\n但底稿模板本身可能\n没有系统化的校验规则', '部分有但不系统——\n防错靠人而非靠模板', '模板创建的思维定式:\n"填什么"而非\n"怎么防止填错"', '⭐⭐ 中', '在Excel模板加入:\n=IF(A1<>B1,"⚠️不匹配","")\n条件格式自动标红', '新项目开始\n就能用'),
]

for i, row_data in enumerate(data2):
    row = sr + 1 + i
    for c, val in enumerate(row_data, 1):
        fill = None
        font = normal_font
        if c == 6:
            fill = green_fill if '低' in str(val) else orange_fill
        elif c == 1:
            font = bold_font
        dc(ws2, row, c, val, font, fill)
    ws2.row_dimensions[row].height = 75

sr2 = row + 2
dc(ws2, sr2, 1, '🔑 闭环一实施步骤', big_bold, gold_fill)
ws2.merge_cells(start_row=sr2, start_column=1, end_row=sr2, end_column=8)

steps = [
    ('Week 1', '打印6个标签（金额不一致/依据不充分/说明不清楚/勾稽未检查/结论太虚/格式问题），贴在屏幕旁'),
    ('Week 1-4', '每次审底稿，在Excel表里记录：日期+项目名+底稿名+标签+备注（一句话说哪里不对）'),
    ('Month 3', '第一次季度统计：按底稿类型×标签×次数排序 → 找出打回率最高的Top3'),
    ('Month 3', '改Top3底稿的模板：加自动勾稽公式、加填写提示、常见错误处标红'),
    ('Month 4-6', '对比优化前后的打回率 → 发邮件/开会告知全员"XX底稿模板已优化"'),
    ('Month 6+', '持续循环：每季度统计→改最差的模板→验证效果'),
]
for i, (time, step) in enumerate(steps):
    row = sr2 + 1 + i
    dc(ws2, row, 1, time, bold_font, warm_fill)
    dc(ws2, row, 2, step)
    ws2.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    ws2.row_dimensions[row].height = 28

ws2.column_dimensions['A'].width = 16
ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 24
ws2.column_dimensions['D'].width = 22
ws2.column_dimensions['E'].width = 20
ws2.column_dimensions['F'].width = 12
ws2.column_dimensions['G'].width = 26
ws2.column_dimensions['H'].width = 20

# ==========================================
# Sheet 3: 闭环二
# ==========================================
ws3 = wb.create_sheet('闭环二-调整→风险')
sr = title_row(ws3, 1, '闭环二：调整分录 → 调整类型 → 风险提示', '跨项目看调整分录，才能发现规律："这个行业历史上哪些科目最容易调整？"')

sr += 1
for c, h in enumerate(headers2, 1):
    ws3.cell(row=sr, column=c, value=h)
style_header(ws3, sr, len(headers2))

data3 = [
    ('调整分录\n是否有记录', '每笔调整分录被保留', '在底稿/报告中有，\n但分散在各项目文件夹', '有但分散——\n不能跨项目检索', '项目制思维:\n归档即结束', '⭐ 低', '建一个统一的Excel表:\n调整分录登记表', '即刻开始积累'),
    ('是否标注\n调整类型', '收入跨期/费用重分类/\n坏账准备/存货跌价/\n递延所得税/其他', '未系统标注', '完全没有分类——\n不知道哪种调整最常出现', '缺分类标准;\n没人想到要标注', '⭐ 低', '每笔调整分录选择分类:\n6大类+自定义', '录入即标注'),
    ('是否关联\n行业和客户', '调整分录标注所属行业\n+客户类型+项目类型', '未标注', '完全没有——\n不能按行业/客户\n分析调整规律', '没意识到行业维度\n对调整规律的价值', '⭐ 低', 'Excel表增加三列:\n行业/客户类型/项目类型', '录入即关联'),
    ('是否跨项目\n规律挖掘', '20+项目积累后\n自动发现:\n某行业→收入跨期高频\n某客户→费用重分类高频', '未做', '金矿在但没挖——\n调整数据全在，\n但没人串起来看', '没攒够数据;\n没人做这件事', '⭐⭐ 中', '攒到15+项目后\n做一次交叉分析:\n行业×调整类型×频次', '6个月后\n出第一份\n行业-调整图谱'),
    ('新项目启动时\n是否提醒', '同类项目历史高频调整\n自动推送', '靠审计师个人记忆\n或老员工口头传承', '完全靠人——\n老员工离职=经验消失', '没有系统化推送机制', '⭐⭐⭐ 高', '手工版:\n项目启动时查Excel表\n→同类项目的调整记录\n→打印一张风险提示卡', '每个新项目\n都能用'),
]

for i, row_data in enumerate(data3):
    row = sr + 1 + i
    for c, val in enumerate(row_data, 1):
        fill = None
        font = normal_font
        if c == 6:
            fill = green_fill if '低' in str(val) else (orange_fill if '中' in str(val) else red_fill)
        elif c == 1:
            font = bold_font
        dc(ws3, row, c, val, font, fill)
    ws3.row_dimensions[row].height = 75

sr2 = row + 2
dc(ws3, sr2, 1, '🔑 闭环二实施步骤', big_bold, gold_fill)
ws3.merge_cells(start_row=sr2, start_column=1, end_row=sr2, end_column=8)

steps3 = [
    ('Week 1', '建立"调整分录登记表"Excel，字段：项目名称/行业/客户类型/调整分录/调整类型/涉及科目/金额/备注'),
    ('每项目结束', '项目结束时花15分钟，把本项目所有调整分录录入登记表'),
    ('Month 3', '如果已有历史项目数据，回溯录入——至少回溯3-5个代表性项目'),
    ('Month 6', '攒到15+项目后，做交叉分析：行业×调整类型 透视表 → 发现规律（如"交通类项目收入跨期高频"）'),
    ('Month 6+', '新项目启动时：查登记表→筛选同行业/同客户类型→打印"历史高频调整提示卡"→交给审计组'),
    ('持续', '每次新项目做完，更新登记表→规律越来越准确'),
]
for i, (time, step) in enumerate(steps3):
    row = sr2 + 1 + i
    dc(ws3, row, 1, time, bold_font, warm_fill)
    dc(ws3, row, 2, step)
    ws3.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    ws3.row_dimensions[row].height = 28

for c, w in enumerate([16, 22, 24, 22, 20, 12, 26, 20], 1):
    ws3.column_dimensions[get_column_letter(c)].width = w

# ==========================================
# Sheet 4: 闭环三
# ==========================================
ws4 = wb.create_sheet('闭环三-清单→提醒')
sr = title_row(ws4, 1, '闭环三：资料清单 → 缺失记录 → 下年提醒', '资料清单不是静态模板，而是随项目经验不断进化的')

sr += 1
for c, h in enumerate(headers2, 1):
    ws4.cell(row=sr, column=c, value=h)
style_header(ws4, sr, len(headers2))

data4 = [
    ('是否有资料\n缺失记录', '记录每年每个客户\n缺了什么资料', '大概率靠审计师个人记忆\n或项目经理口头传承', '完全靠人——\n新人不知道去年\n什么资料追了很久', '没有记录习惯;\n觉得"每年不一样\n记了也没用"', '⭐ 低', '资料清单Excel加一列:\n"往年缺失记录"', '下次做同一客户\n就能用'),
    ('资料清单\n是否动态更新', '每年根据上年缺失情况\n更新清单内容', '可能用的是固定模板清单\n多年不变', '静态模板——\n同样的资料每年追', '模板思维:\n"清单是固定的"', '⭐ 低', '每年更新前，\n先看去年缺失记录\n→优先标注', '1年见效果'),
    ('格式问题\n是否有说明', '经常出错的资料\n附格式要求+模板', '口头沟通/微信上说\n"你这个格式不对"', '每次都要重新说——\n消耗沟通成本', '没把格式要求\n写成文档', '⭐⭐ 中', '在资料清单中\n对格式要求加备注\n附模板下载链接', '新客户/新人\n都能直接用'),
    ('新项目是否\n自动提醒', '启动时自动推送\n客户历史资料问题', '靠老员工提醒新人\n或自己想起来', '完全靠人——\n忘了就忘了', '没有系统', '⭐⭐ 中', '手工版:\n项目启动会前\n看一眼资料缺失记录\n→会议第一条讲', '每个项目\n启动就能用'),
]

for i, row_data in enumerate(data4):
    row = sr + 1 + i
    for c, val in enumerate(row_data, 1):
        fill = None
        font = normal_font
        if c == 6:
            fill = green_fill if '低' in str(val) else orange_fill
        elif c == 1:
            font = bold_font
        dc(ws4, row, c, val, font, fill)
    ws4.row_dimensions[row].height = 75

sr2 = row + 2
dc(ws4, sr2, 1, '🔑 闭环三实施步骤', big_bold, gold_fill)
ws4.merge_cells(start_row=sr2, start_column=1, end_row=sr2, end_column=8)

steps4 = [
    ('Week 1', '找到现有的资料清单模板，增加两列："往年缺失记录""格式要求"'),
    ('每项目结束', '勾选：哪些资料追了3次以上才拿到 → 写入"往年缺失记录"列'),
    ('每项目结束', '勾选：哪些资料格式有问题 → 在"格式要求"列写清楚正确格式+模板链接'),
    ('下个项目', '同客户：启动前打开清单 → 缺失记录列自动提醒"⚠️ 往年此项资料延迟XX天"'),
    ('下个项目', '新客户同类项目：参考同行业客户的缺失记录，提前准备'),
    ('持续', '每做一次项目，清单就进化一次 → 3年后每个客户都有一份精准的专属清单'),
]
for i, (time, step) in enumerate(steps4):
    row = sr2 + 1 + i
    dc(ws4, row, 1, time, bold_font, warm_fill)
    dc(ws4, row, 2, step)
    ws4.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    ws4.row_dimensions[row].height = 28

for c, w in enumerate([16, 22, 24, 22, 20, 12, 26, 20], 1):
    ws4.column_dimensions[get_column_letter(c)].width = w

# ==========================================
# Sheet 5: 根因分析
# ==========================================
ws5 = wb.create_sheet('根因分析')
sr = title_row(ws5, 1, '为什么这些差距会产生？——三个根因')

headers5 = ['根因', '具体表现', '在融策的体现', '为什么之前没做', '如何突破']
sr += 1
for c, h in enumerate(headers5, 1):
    ws5.cell(row=sr, column=c, value=h)
style_header(ws5, sr, len(headers5))

causes = [
    ('根因一\n"一次性思维"', 
     '审计行业默认模式是"项目制"——\n开始→做完→归档→结束。\n经验跟着项目一起"结束"。\n资产化要求"连续性思维"——\n项目不因归档结束，经验活在下一个项目。',
     '融策的审计黑板已做了项目工作区标准化\n（raw_data→findings→collision），\n这是连续性思维的基础。\n但复核意见和调整分录\n没有被纳入标准化体系。',
     '工作流设计时，\n"项目结束"被定义为物理归档，\n而非"经验提取"。\n归档=任务的终点。',
     '在项目结束checklist中\n增加一项：\n"经验提取完成？"\n（复核意见录入/调整分录录入/\n资料缺失记录更新）'),
    ('根因二\n数据"麻烦程度"\n不同',
     '容易存：知识型数据\n（法规/政策/案例/方法论文）\n  → 融策已做 ✅\n\n麻烦一点：操作型数据\n（复核意见/调整分录/资料缺失）\n  → 融策未做 ❌\n  → 需在日常工作中顺手记录\n\n最难存：隐性知识\n（判断逻辑/经验直觉/行业感觉）\n  → 全行业难题，先放一放',
     '融策先啃了最难的知识体系化\n（RAG + PARA + Obsidian），\n但跳过了中间层（操作型数据）。\n现在回头补中间层，难度更低。',
     '知识型数据的价值是即时可见的\n（搜法规→立刻找到），\n操作型数据的价值是滞后的\n（攒三个月才能看出规律）。\n人天然倾向做反馈快的。',
     '降低操作型数据的录入成本：\n1. 贴标签（15秒）比写意见（2分钟）快\n2. 下拉选择比手动输入快\n3. 季度批量统计而非每日关注'),
    ('根因三\n"没人负责"',
     '知识管理有明确的负责人\n（"你来管知识库"），\n但数据资产化没有——\n复核意见是谁的责任？\n调整分录归谁录入？\n资料缺失谁来追踪？',
     '融策有质控流程，\n但质控的角色是"把关"，\n不是"提取经验"。\n把关完了，意见也丢了。',
     '组织结构中\n没有"经验管理"这个角色。\n每个人的KPI都是做项目，\n不是让下一个项目更好做。',
     '指定一个"经验管理"负责人：\n- 每季度汇总三个闭环数据\n- 输出模板优化建议\n- 输出行业-调整关联分析\n- 输出客户资料缺失年度报告\n这个人不需要全职，\n每月多花半天即可'),
]

for i, row_data in enumerate(causes):
    row = sr + 1 + i
    for c, val in enumerate(row_data, 1):
        font = bold_font if c == 1 else normal_font
        fill = None
        if c == 1:
            fill = red_fill
        dc(ws5, row, c, val, font, fill)
    ws5.row_dimensions[row].height = 150

for c, w in enumerate([18, 38, 32, 30, 30], 1):
    ws5.column_dimensions[get_column_letter(c)].width = w

# ==========================================
# Sheet 6: 路线图
# ==========================================
ws6 = wb.create_sheet('关闭路线图')
sr = title_row(ws6, 1, '差距关闭路线图：从今天到12个月')

headers6 = ['时间节点', '闭环一\n复核→模板', '闭环二\n调整→风险', '闭环三\n清单→提醒', '里程碑事件', '所需资源']
sr += 1
for c, h in enumerate(headers6, 1):
    ws6.cell(row=sr, column=c, value=h)
style_header(ws6, sr, len(headers6))

roadmap = [
    ('现在\n（本周）', 
     '✅ 打印6个标签贴屏幕旁\n✅ 开始打标签记录',
     '✅ 建立调整分录登记表Excel\n（空表准备好）',
     '✅ 资料清单模板加两列\n"往年缺失""格式要求"',
     '🔵 三个闭环基础设施\n全部就位（零成本）',
     '1个Excel文件\n1张贴纸\n30分钟'),
    ('1个月', 
     '📊 第一批标签数据就绪\n约20-50条复核意见记录',
     '📊 1-2个项目调整分录录入',
     '📊 1-2个项目资料缺失记录',
     '🔵 第一批结构化操作数据\n开始积累',
     '每项目结束\n多花15分钟'),
    ('3个月', 
     '📈 第一次季度统计\n找出打回率Top3底稿\n→ 改模板',
     '📊 3-5个项目调整数据',
     '📊 1-3个客户历史缺失记录\n→ 下个项目用上',
     '🟡 闭环一产生第一次\n模板优化（量化效果）\n闭环三第一次提醒生效',
     '每季度\n1小时汇总'),
    ('6个月', 
     '📈 第二次季度统计\n→ 对比优化前后的打回率\n→ 持续迭代',
     '📈 10-15个项目调整数据\n→ 第一次行业-调整分析\n→ 第一次项目风险提示卡',
     '📊 持续更新\n→ 客户专属清单初具雏形',
     '🟢 闭环二第一次规律浮现\n三个闭环全部运转\n飞轮开始加速',
     '每季度1小时汇总\n每新项目10分钟\n查历史提示'),
    ('12个月', 
     '🔄 持续循环优化\n模板打回率下降30%+',
     '🔄 行业-调整关联图谱\n成熟可用\n→ 可喂给AI审计智能体',
     '🔄 每个客户都有\n精准的专属资料清单\n新人直接用',
     '🔵 数据积累足够\n→ 启动AI审计智能体试点\n（预算执行审计场景优先）',
     '数据够了\n可以接智能体'),
]

for i, row_data in enumerate(roadmap):
    row = sr + 1 + i
    for c, val in enumerate(row_data, 1):
        fill = None
        font = normal_font
        if c == 1:
            font = bold_font
            if i == 0:
                fill = red_fill
            elif i <= 2:
                fill = orange_fill
            elif i == 3:
                fill = green_fill
            else:
                fill = blue_fill
        dc(ws6, row, c, val, font, fill)
    ws6.row_dimensions[row].height = 110

for c, w in enumerate([14, 28, 28, 28, 30, 22], 1):
    ws6.column_dimensions[get_column_letter(c)].width = w

# ==========================================
# Sheet 7: 行动清单 - 本周启动
# ==========================================
ws7 = wb.create_sheet('🔴本周行动清单')
sr = title_row(ws7, 1, '🔴 本周就能启动的行动清单', '不需要新系统，不需要预算，只需要习惯改变')

headers7 = ['序号', '行动', '具体操作', '耗时', '负责', 'Deadline', '产出']
sr += 1
for c, h in enumerate(headers7, 1):
    ws7.cell(row=sr, column=c, value=h)
style_header(ws7, sr, len(headers7))

actions = [
    ('1', '打印复核意见标签', '用A4纸打印6个标签：\n①金额不一致 ②依据不充分\n③说明不清楚 ④勾稽未检查\n⑤结论太虚 ⑥格式问题\n贴在办公桌屏幕旁边', '10分钟', '', '今天', '6个可视化标签'),
    ('2', '准备复核意见记录模板', '新建Excel：\n列=日期/项目名/底稿名/标签/备注\n放在共享文件夹/在线文档', '15分钟', '', '今天', '复核意见记录表\n（空模板）'),
    ('3', '建立调整分录登记表', '新建Excel：\n列=项目名/行业/客户类型/调整分录/调整类型/涉及科目/金额\n调整类型下拉：收入跨期/费用重分类/坏账准备/存货跌价/递延所得税/其他', '20分钟', '', '本周', '调整分录登记表\n（空模板）'),
    ('4', '更新资料清单模板', '找到现有的资料清单模板\n增加两列：\n"往年缺失记录""格式要求"', '15分钟', '', '本周', '更新后的资料清单\n（含历史记录列）'),
    ('5', '选1个试点项目', '选一个近期启动或正在进行的项目\n全流程使用本套工具\n（贴标签+登记调整+记录缺失）', '5分钟', '', '本周', '确定试点项目'),
    ('6', '团队告知', '开一个10分钟短会/发微信群：\n"从本周开始，每次审底稿时\n顺手贴一个标签。\n年底我们就能知道\n哪些底稿最需要优化。"', '10分钟', '', '本周', '全员知晓\n减少执行阻力'),
]

for i, row_data in enumerate(actions):
    row = sr + 1 + i
    for c, val in enumerate(row_data, 1):
        fill = None
        if c == 1:
            fill = red_fill
        elif c == 5:
            fill = warm_fill
        dc(ws7, row, c, val, normal_font, fill)
    ws7.row_dimensions[row].height = 70

for c, w in enumerate([8, 22, 38, 12, 10, 12, 22], 1):
    ws7.column_dimensions[get_column_letter(c)].width = w

# ====== 保存到桌面 ======
desktop = os.path.join(os.environ['USERPROFILE'], 'Desktop')
outpath = os.path.join(desktop, '融策数据资产化差距分析.xlsx')
wb.save(outpath)
print(f'Saved to desktop: {outpath}')
print(f'Sheets: {wb.sheetnames}')
