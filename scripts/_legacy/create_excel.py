import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# Create the Excel file
output_path = r"D:\openclaw-workspace\output\融策OpenClaw技能功能与业务场景映射表.xlsx"

# Ensure output directory exists
os.makedirs(os.path.dirname(output_path), exist_ok=True)

# Create a workbook
wb = openpyxl.Workbook()

# Remove default sheet
wb.remove(wb.active)

# Define styles
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
available_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
pending_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
new_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ==================== Sheet 1: 为什么不能用 ====================
ws1 = wb.create_sheet("为什么不能用")
ws1.append(["模块名称", "当前状态", "原因说明", "解决步骤", "预计时间"])
for cell in ws1[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

data_platform_issues = [
    ["数据中台", "代码就绪，不可用", 
     "PostgreSQL数据库未安装，所有ETL脚本依赖数据库连接。配置文件中DB_HOST=localhost，但本地未安装PostgreSQL服务，运行时会报错Connection refused",
     "1. 安装PostgreSQL（官网下载安装包）\n2. 创建数据库rongce_data_platform\n3. 创建用户rongce/密码rongce123\n4. 运行scripts/init_database.py建表\n5. 验证连接",
     "1-2小时"],
    ["知识图谱", "代码就绪，不可用",
     "Neo4j图数据库未安装，GraphBuilder默认尝试连接Neo4j失败后回退到NetworkX（内存图）。NetworkX无法持久化存储，重启即丢失数据。",
     "1. 安装Neo4j Community Edition\n2. 配置连接参数（bolt://localhost:7687）\n3. 初始化图数据库schema\n4. 验证实体抽取+关系抽取+图构建全流程\n5. 可选：继续用NetworkX做开发测试",
     "1-2小时"],
    ["天眼查API", "未配置",
     "串标检测L8（工商关联分析）需要天眼查API Key。当前代码中API Key为空，调用时会跳过外部数据查询。",
     "1. 注册天眼查开放平台账号\n2. 申请API Key\n3. 配置到环境变量或config文件\n4. 测试企业关联查询",
     "1-3天（申请审批）"],
    ["Redis缓存", "未配置",
     "数据中台配置中Redis用于缓存，但本地未安装。不影响基础功能，仅影响性能。",
     "1. 安装Redis（或改用内存缓存）\n2. 配置连接参数",
     "30分钟"],
    ["MinIO对象存储", "未配置",
     "用于存储非结构化文件（PDF/扫描件）。当前用本地文件系统替代，不影响功能。",
     "1. 安装MinIO\n2. 配置endpoint和access_key\n3. 创建bucket",
     "30分钟"],
]

for row in data_platform_issues:
    ws1.append(row)

for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=5):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(vertical='top', wrap_text=True)
    row[1].fill = pending_fill

ws1.column_dimensions['A'].width = 15
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 50
ws1.column_dimensions['D'].width = 45
ws1.column_dimensions['E'].width = 12
ws1.row_dimensions[2].height = 80
ws1.row_dimensions[3].height = 80
ws1.row_dimensions[4].height = 60
ws1.row_dimensions[5].height = 40
ws1.row_dimensions[6].height = 40

wb.save(output_path)
print(f"Excel created: {output_path}")
