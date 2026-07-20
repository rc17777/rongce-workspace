"""
Create comparison workbook between old (v3-2026.05.21) and new (v3-2026.06.05) versions
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

old_path = r'C:\Users\scrccpa\Desktop\健康照护师-成本构成测算-v3-2026.05.21.xlsx'
new_path = r'C:\Users\scrccpa\Desktop\健康照护师-成本构成测算-v3-2026.06.05.xlsx'
out_path = r'C:\Users\scrccpa\Desktop\成本测算-版本差异对比-2026.06.05.xlsx'

wb_old = openpyxl.load_workbook(old_path)
wb_new = openpyxl.load_workbook(new_path)

# Styles
hdr_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
hdr_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
sec_fill = PatternFill(start_color='FFD6E4F0', end_color='FFD6E4F0', fill_type='solid')
chg_fill = PatternFill(start_color='FFFFEB9C', end_color='FFFFEB9C', fill_type='solid')
up_fill  = PatternFill(start_color='FFC6EFCE', end_color='FFC6EFCE', fill_type='solid')
down_fill= PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid')
nf = Font(name='微软雅黑', size=10)
bf = Font(name='微软雅黑', size=10, bold=True)
red_f = Font(name='微软雅黑', size=10, bold=True, color='FF0000')
ca = Alignment(horizontal='center', vertical='center', wrap_text=True)
la = Alignment(horizontal='left', vertical='center', wrap_text=True)
thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

def write_section(ws, row, title):
    ws.merge_cells(f'A{row}:I{row}')
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(name='微软雅黑', size=10, bold=True); c.fill = sec_fill; c.border = thin

def write_header_row(ws, row, hdrs):
    for i, h in enumerate(hdrs):
        c = ws.cell(row=row, column=i+1, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = ca; c.border = thin

def write_data_row(ws, row, seq, name, old_vals, new_vals, note='', changed=False):
    ws.cell(row=row, column=1, value=seq).font = nf
    ws.cell(row=row, column=1).alignment = ca; ws.cell(row=row, column=1).border = thin
    ws.cell(row=row, column=2, value=name).font = Font(name='微软雅黑', size=10, bold=True) if changed else nf
    ws.cell(row=row, column=2).alignment = la; ws.cell(row=row, column=2).border = thin
    for j, v in enumerate(old_vals):
        c = ws.cell(row=row, column=3+j, value=v)
        c.font = nf; c.alignment = ca; c.border = thin
    for j, v in enumerate(new_vals):
        c = ws.cell(row=row, column=6+j, value=v)
        c.font = nf; c.alignment = ca; c.border = thin
        if changed and isinstance(v, (int, float)) and isinstance(old_vals[j], (int, float)):
            if v != old_vals[j]:
                c.fill = down_fill if v > old_vals[j] else up_fill
    ws.cell(row=row, column=9, value=note).font = nf
    ws.cell(row=row, column=9).alignment = la; ws.cell(row=row, column=9).border = thin

# Read new values
ws_n = wb_new['成本构成总览']
new_13 = [ws_n.cell(row=13, column=c).value for c in range(4,7)]
new_14 = [ws_n.cell(row=14, column=c).value for c in range(4,7)]
new_16 = [ws_n.cell(row=16, column=c).value for c in range(4,7)]
new_21 = [ws_n.cell(row=21, column=c).value for c in range(4,7)]
new_24 = [ws_n.cell(row=24, column=c).value for c in range(4,7)]
new_29 = [ws_n.cell(row=29, column=c).value for c in range(4,7)]

# ======================
# Sheet 1: 成本构成总览对比
# ======================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = '成本构成总览对比'

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 28
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 14
ws.column_dimensions['G'].width = 14
ws.column_dimensions['H'].width = 14
ws.column_dimensions['I'].width = 40

ws.merge_cells('A1:I1')
ws.cell(row=1, column=1, value='健康照护师（长期照护师）成本构成测算 — 版本差异对比（v3-2026.05.21 vs v3-2026.06.05）').font = Font(name='微软雅黑', bold=True, size=14)
ws.merge_cells('A2:I2')
ws.cell(row=2, column=1, value='主要变更：1.人工费用由30人/天改为120人/天,基于两次实际认定数据  2.耗材三级标黄部分调整').font = Font(name='微软雅黑', size=10, italic=True)

row = 4
write_header_row(ws, row, ['序号','成本项目','旧版-五级','旧版-四级','旧版-三级','新版-五级','新版-四级','新版-三级','差异说明'])
row += 1

# Assumptions change
write_section(ws, row, '测算假设变更')
row += 1
ws.merge_cells(f'A{row}:I{row}')
ws.cell(row=row, column=1, value='旧版: 每班次30人/2考站 → 新版: 每班次120人/4考场, 基于2025.11.30及2026.01.09实际认定数据').font = Font(name='微软雅黑', size=10, italic=True)
row += 2

# Data rows
write_data_row(ws, row, '1', '理论考试费', [30,35,40], [30,35,40], '无变化'); row += 1
write_data_row(ws, row, '2', '操作技能考核费', [140,190,240], [140,190,240], '无变化'); row += 1
write_data_row(ws, row, '', '政策收费小计', [170,225,280], [170,225,280], '无变化'); row += 1
write_data_row(ws, row, '3', '考务平台费', [20,20,20], [20,20,20], '无变化'); row += 1

# 4. 人工费用 CHANGED
old_13 = [329.8, 329.8, 329.8]
write_data_row(ws, row, '4', '人工费用 ★', old_13, new_13,
    f'329.80 -> {new_13[0]} (-{(329.8-new_13[0]):.1f}, -{(329.8-new_13[0])/329.8*100:.0f}%)', True); row += 1

# 4a. SP模特 CHANGED
old_14 = [20.8, 20.8, 20.8]
write_data_row(ws, row, '4a', '  其中:SP模特 ★', old_14, new_14,
    f'20.80 -> {new_14[0]} (计算口径变更: 2考站x6h -> 4-9考场均值x7h/120人)', True); row += 1

# 5. 设施设备
write_data_row(ws, row, '5', '设施设备使用费', [93,158,190], [93,158,190], '无变化'); row += 1

# 6. 耗材成本 CHANGED for 三级
old_16 = [87.2, 167.28, 280.49]
write_data_row(ws, row, '6', '耗材成本 ★', old_16, new_16,
    f'三级: 280.49 -> {new_16[2]} (-{280.49-new_16[2]:.1f}, -{(280.49-new_16[2])/280.49*100:.1f}%, 标黄调整)', True); row += 1

# 7. 场地
write_data_row(ws, row, '7', '场地水电费', [106.67,106.67,106.67], [106.67,106.67,106.67], '无变化'); row += 1

# 8-10 待补充
write_data_row(ws, row, '8-10', '监控/证书/文印', ['待补充','待补充','待补充'], ['待补充','待补充','待补充'], '待补充'); row += 1

# 实际成本小计 CHANGED
old_21 = [636.67, 781.75, 926.96]
write_data_row(ws, row, '', '实际成本小计（已测算）★', old_21, new_21,
    f'五级: {636.67}->{new_21[0]} | 四级: {781.75}->{new_21[1]} | 三级: {926.96}->{new_21[2]}', True); row += 2

# 全口径成本 CHANGED
old_24 = [806.67, 1006.75, 1206.96]
write_data_row(ws, row, '', '全口径成本（7项已测算）★', old_24, new_24,
    f'五级: {806.67}->{new_24[0]} | 四级: {1006.75}->{new_24[1]} | 三级: {1206.96}->{new_24[2]}', True); row += 1

write_data_row(ws, row, '', '现行收费标准（方案5稿）', [320,405,500], [320,405,500], '无变化'); row += 1
write_data_row(ws, row, '', '已测算7项成本', old_24, new_24, '同全口径成本', True); row += 1

old_29 = [-486.67, -601.75, -706.96]
write_data_row(ws, row, '', '差额（收费-成本）★', old_29, new_29,
    f'缺口缩小: 五级|{abs(old_29[0])}|->|{abs(new_29[0])}| 四级|{abs(old_29[1])}|->|{abs(new_29[1])}| 三级|{abs(old_29[2])}|->|{abs(new_29[2])}|', True); row += 2

# Summary
write_section(ws, row, '差异总结')
row += 1
summaries = [
    '1. 人工费用为主因：计算基础从"30人/天"改为"120人/天"（实际数据），规模效应使人均成本大幅下降',
    f'2. 人工费降幅：329.80 -> {new_13[0]} 降{(329.8-new_13[0]):.0f}元/人 ({(329.8-new_13[0])/329.8*100:.0f}%)',
    f'3. 三级耗材降幅：280.49 -> {new_16[2]} 降{(280.49-new_16[2]):.0f}元/人 ({(280.49-new_16[2])/280.49*100:.1f}%)',
    f'4. 全口径(三级): 1206.96 -> {new_24[2]} 降{1206.96-new_24[2]:.0f}元/人',
    f'5. 差额(三级): -706.96 -> {new_29[2]} 缩小{abs(706.96-new_29[2]):.0f}元/人, 但价差仍存',
]
for s in summaries:
    ws.merge_cells(f'A{row}:I{row}')
    ws.cell(row=row, column=1, value=s).font = nf
    row += 1

# ======================
# Sheet 2: 耗材三级对比
# ======================
ws2 = wb.create_sheet('耗材三级标黄对比')
ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 26
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 12
ws2.column_dimensions['E'].width = 14
ws2.column_dimensions['F'].width = 12
ws2.column_dimensions['G'].width = 12
ws2.column_dimensions['H'].width = 14
ws2.column_dimensions['I'].width = 35

ws2.merge_cells('A1:I1')
ws2.cell(row=1, column=1, value='耗材成本明细 — 三级/高级工 标黄调整对比').font = Font(name='微软雅黑', bold=True, size=14)

row = 3
write_header_row(ws2, row, ['序号','耗材名称','旧-单价','旧-用量','旧-单人成本','新-单价','新-用量','新-单人成本','调整说明'])
row += 1

changed_items = [
    (4, '隔离防护服 -> 隔离衣(布类)', 28.23, '1', 28.23, 60, '1', 10, '标黄调整: 60元/套,磨损费/6次=10元'),
    (33, '瞳孔笔、儿童晨检笔 -> 瞳孔笔', 7.76, '1/2', 3.88, 7.76, '1/2', 3.88, '仅简化名称,数值不变'),
    (40, '毛巾', 20, '1/4', 5, 20, '1/4', 2, '标黄调整: 磨损费/10次,5->2'),
    (44, '袜子 -> 一次性袜子', 12, '1/4', 3, 1, '1', 1, '标黄调整: 改为一次性,1元/双'),
]

for seq, name, op, ou, oc, np_, nu, nc, note in changed_items:
    c = ws2.cell(row=row, column=1, value=seq)
    c.font = bf; c.border = thin; c.alignment = ca
    ws2.cell(row=row, column=2, value=name).font = bf
    ws2.cell(row=row, column=2).border = thin; ws2.cell(row=row, column=2).alignment = la
    for j, v in enumerate([op, ou, oc]):
        c = ws2.cell(row=row, column=3+j, value=v)
        c.font = nf; c.border = thin; c.alignment = ca
    for j, v in enumerate([np_, nu, nc]):
        c = ws2.cell(row=row, column=6+j, value=v)
        c.font = nf; c.border = thin; c.alignment = ca
        if v != [op, ou, oc][j]:
            c.fill = chg_fill
    ws2.cell(row=row, column=9, value=note).font = nf
    ws2.cell(row=row, column=9).border = thin; ws2.cell(row=row, column=9).alignment = la
    row += 1

row += 1
write_section(ws2, row, '合计对比')
row += 1
ws2.cell(row=row, column=1).value = '合计'
ws2.cell(row=row, column=1).font = Font(name='微软雅黑', bold=True, size=11)
ws2.cell(row=row, column=5).value = 280.49
ws2.cell(row=row, column=5).font = Font(name='微软雅黑', bold=True, size=11)
ws2.cell(row=row, column=8).value = 257.26
ws2.cell(row=row, column=8).font = Font(name='微软雅黑', bold=True, size=11, color='FF0000')
ws2.cell(row=row, column=9).value = '降23.23元/人 (-8.3%)'
ws2.cell(row=row, column=9).font = Font(name='微软雅黑', bold=True, size=11, color='FF0000')
for c_idx in range(1,10):
    ws2.cell(row=row, column=c_idx).border = thin; ws2.cell(row=row, column=c_idx).alignment = ca

wb.save(out_path)
print(f"Saved: {out_path}")
