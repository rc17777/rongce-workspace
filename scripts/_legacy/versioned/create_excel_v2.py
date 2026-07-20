import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

output_path = r"D:\openclaw-workspace\output\rongce_skills_business_mapping.xlsx"
os.makedirs(os.path.dirname(output_path), exist_ok=True)

wb = openpyxl.Workbook()
wb.remove(wb.active)

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
available_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
pending_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
new_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def style_header(ws):
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

def style_rows(ws, start_row=2, status_col=None):
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        if status_col:
            val = str(row[status_col-1].value or '')
            if '可用' in val or 'Available' in val:
                row[status_col-1].fill = available_fill
            elif '就绪' in val or 'Pending' in val:
                row[status_col-1].fill = pending_fill
            elif '新增' in val or 'New' in val:
                row[status_col-1].fill = new_fill

# Sheet 1
ws1 = wb.create_sheet("为什么不能用")
ws1.append(["模块名称", "当前状态", "原因说明", "解决步骤", "预计时间"])
style_header(ws1)
issues = [
    ["数据中台", "代码就绪，不可用", "PostgreSQL数据库未安装，ETL脚本依赖数据库连接。配置文件DB_HOST=localhost，但本地未安装PostgreSQL服务，运行时报Connection refused", "1.安装PostgreSQL 2.创建数据库rongce_data_platform 3.创建用户rongce/密码rongce123 4.运行init_database.py建表 5.验证连接", "1-2小时"],
    ["知识图谱", "代码就绪，不可用", "Neo4j图数据库未安装。GraphBuilder尝试连接Neo4j失败后回退到NetworkX（内存图）。NetworkX无法持久化，重启丢失数据。", "1.安装Neo4j Community 2.配置bolt://localhost:7687 3.初始化schema 4.验证全流程 5.可选：继续用NetworkX开发测试", "1-2小时"],
    ["天眼查API", "未配置", "串标检测L8（工商关联分析）需要天眼查API Key。当前API Key为空，调用时跳过外部数据查询。", "1.注册天眼查开放平台 2.申请API Key 3.配置到环境变量 4.测试企业关联查询", "1-3天"],
    ["Redis缓存", "未配置", "数据中台Redis用于缓存，本地未安装。不影响基础功能，仅影响性能。", "1.安装Redis（或改用内存缓存）2.配置连接参数", "30分钟"],
    ["MinIO对象存储", "未配置", "存储非结构化文件（PDF/扫描件）。当前用本地文件系统替代，不影响功能。", "1.安装MinIO 2.配置endpoint和access_key 3.创建bucket", "30分钟"],
]
for row in issues:
    ws1.append(row)
style_rows(ws1, status_col=2)
ws1.column_dimensions['A'].width = 15
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 50
ws1.column_dimensions['D'].width = 45
ws1.column_dimensions['E'].width = 12
for i in range(2, 7):
    ws1.row_dimensions[i].height = 80

wb.save(output_path)
print(f"Excel created: {output_path}")
