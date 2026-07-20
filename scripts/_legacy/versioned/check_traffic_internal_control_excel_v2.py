from openpyxl import load_workbook
p = r'outputs\交通厅内控项目\四川省交通运输厅2026年内部控制评价指标体系及监督检查模板_财会2025-24校准版.xlsx'
wb = load_workbook(p, data_only=False)
print('sheets:', wb.sheetnames)
for s in ['2厅本级指标评分细则','3厅属单位指标评分细则']:
    ws = wb[s]
    vals = [ws.cell(r,5).value for r in range(4, ws.max_row) if isinstance(ws.cell(r,5).value, (int,float))]
    print(s, '指标数', len(vals), '分值合计', sum(vals))
ws = wb['评分汇总表']
print('评分汇总表表头:', [ws.cell(3,c).value for c in range(1, ws.max_column+1)])
print(p)
