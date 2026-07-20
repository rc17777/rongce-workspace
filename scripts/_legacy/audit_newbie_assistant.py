#!/usr/bin/env python3
"""
审计新手项目助手 — 第一性原理驱动的一键式审计任务分解

用法:
  python audit_newbie_assistant.py                     # 交互式向导
  python audit_newbie_assistant.py --type 招投标审计     # 指定类型
  python audit_newbie_assistant.py --type 财务审计 --o 项目计划.xlsx

输出: 项目执行计划书 (Excel) + 分工清单
"""
import sys, io, argparse, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 样式 ──
GRN = PatternFill(patternType='solid', fgColor='D4EDDA')
YEL = PatternFill(patternType='solid', fgColor='FFF3CD')
RED = PatternFill(patternType='solid', fgColor='FFD7D7')
BLUE = PatternFill(patternType='solid', fgColor='D6E4F0')
HEADER = PatternFill(patternType='solid', fgColor='1A3A6E')
TITLE = Font(name='Microsoft YaHei', size=18, bold=True, color='1A3A6E')
SUBTITLE = Font(name='Microsoft YaHei', size=11, color='666666')
H = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
N = Font(name='Microsoft YaHei', size=10)
B = Font(name='Microsoft YaHei', size=10, bold=True)
BR = Font(name='Microsoft YaHei', size=10, bold=True, color='CC0000')
TH = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
C = Alignment(horizontal='center', vertical='center', wrap_text=True)
L = Alignment(horizontal='left', vertical='center', wrap_text=True)
W = Alignment(horizontal='left', vertical='top', wrap_text=True)


# ═══════════════════════════════════════════════
#  审计任务知识库（面向审计新手优化）
# ═══════════════════════════════════════════════
KNOWLEDGE = {
    '招投标采购审计': {
        'alias': ['招投标', '采购', '招标', '投标', '围标', '串标'],
        'brief': '审查政府采购/工程招标的合规性，识别围标串标等违法行为',
        'duration': '5-10个工作日（3-5家投标人）',
        'steps': [
            {
                'step': 1, 'phase': '项目准备',
                'task': '资料收集与整理',
                'who': '🤖 AI',
                'detail': '批量解压投标文件→定位关键文件→提取PDF元数据→OCR非文字层文件',
                'tool': 'audit_pipeline.py --skip apriori,savings,entity',
                'output': 'pdf_元数据.xlsx + extracted/ 目录',
                'time': '10分钟',
            },
            {
                'step': 2, 'phase': '项目准备',
                'task': '招标文件合规审查',
                'who': '🤖 AI',
                'detail': '扫描招标文件中的资格限制（品牌/地域/规模）、评分标准合理性',
                'tool': 'audit_pipeline.py (step2 Word扫描)',
                'output': 'word_关键词扫描.xlsx',
                'time': '2分钟',
            },
            {
                'step': 3, 'phase': '数据分析',
                'task': '报价规律性分析',
                'who': '🤖 AI',
                'detail': '检查报价是否呈等差数列/阶梯分布/接近预算上限',
                'tool': 'procurement-audit-models L1(报价规律)',
                'output': '报价疑点清单',
                'time': '5分钟',
            },
            {
                'step': 4, 'phase': '数据分析',
                'task': '投标文件文本雷同检测',
                'who': '🤖 AI',
                'detail': 'TF-IDF全文本相似度+跨投标人段落级匹配',
                'tool': 'python scripts/03_tfidf_similarity.py',
                'output': 'pdf_相似度.xlsx',
                'time': '3分钟',
            },
            {
                'step': 5, 'phase': '数据分析',
                'task': '数字取证深度分析',
                'who': '🤖 AI',
                'detail': 'PDF元数据同源+JPEG量化表指纹+WPS签名残留+图片哈希',
                'tool': 'forensic_analysis.py + procurement L4/L5/D1-D5',
                'output': '深度取证报告.xlsx',
                'time': '15分钟',
            },
            {
                'step': 6, 'phase': '数据分析',
                'task': '跨项目围标检测',
                'who': '🤖 AI',
                'detail': 'Apriori关联规则(共现)+节资率箱线图+陪标专业户识别',
                'tool': 'apriori_analysis.py + 12_savings_rate.py + 13_entity_anomalies.py',
                'output': '跨项目围标疑点.xlsx',
                'time': '5分钟',
            },
            {
                'step': 7, 'phase': '人工核实',
                'task': '复核AI发现的疑点',
                'who': '👤 审计员',
                'detail': '逐项核对AI标注的疑点：确认报价异常原因/文本雷同是否因模板/元数据相同是否合理',
                'tool': '人工判断 + 打开AI输出文件逐项复核',
                'output': '疑点复核记录（标注在底稿中）',
                'time': '2-4小时',
            },
            {
                'step': 8, 'phase': '人工核实',
                'task': '工商关联查询',
                'who': '👤 审计员',
                'detail': '天眼查/企查查查询投标人的股东/高管/法人交叉关系',
                'tool': '天眼查/企查查网页',
                'output': '企业关联关系图',
                'time': '1-2小时',
            },
            {
                'step': 9, 'phase': '报告编制',
                'task': '问题定性与法规匹配',
                'who': '🤖 AI',
                'detail': '输入发现→自动输出：问题定性+法规条款+整改建议+佐证材料清单',
                'tool': 'audit_finding_processor.py',
                'output': '审计发现处理底稿.xlsx',
                'time': '2分钟',
            },
            {
                'step': 10, 'phase': '报告编制',
                'task': '撰写审计报告',
                'who': '👤👉🤖 审计员+AI',
                'detail': '审计员撰写结论和意见 → AI辅助生成报告框架/格式/法规引用',
                'tool': 'analysis-report + audit_finding_processor',
                'output': '审计报告（初稿）',
                'time': '2-3小时',
            },
        ],
        'checklist': [
            '□ 招标公告是否公开发布（≥5个工作日）',
            '□ 投标人数量是否≥3家',
            '□ 投标人报价是否在预算范围内',
            '□ 报价偏差率是否符合常规',
            '□ 投标人之间是否存在工商关联',
            '□ 投标文件签章/日期是否齐全',
            '□ 评审打分是否规范（≥5人单数）',
            '□ 中标公示期是否≥3个工作日',
            '□ 合同是否与中标结果一致',
            '□ 是否存在投诉/质疑未处理',
        ],
    },

    '财务收支审计': {
        'alias': ['财务', '收支', '预算', '费用', '报销', '账务'],
        'brief': '审查财务收支真实性/合规性/完整性',
        'duration': '7-15个工作日',
        'steps': [
            {
                'step': 1, 'phase': '项目准备',
                'task': '资料收集',
                'who': '🤖 AI',
                'detail': '采集科目余额表/序时账/凭证扫描件/银行对账单',
                'tool': 'data-analyst-cn + OCR',
                'output': '结构化数据表',
                'time': '15分钟',
            },
            {
                'step': 2, 'phase': '数据分析',
                'task': '舞弊风险扫描',
                'who': '🤖 AI',
                'detail': 'Benford定律首位数分布+异常大额整数交易+关联交易穿透',
                'tool': 'financial-fraud-detection + anomaly-detection',
                'output': '舞弊风险疑点.xlsx',
                'time': '5分钟',
            },
            {
                'step': 3, 'phase': '数据分析',
                'task': '费用报销异常检测',
                'who': '🤖 AI',
                'detail': '连号发票/同一商户高频消费/节假日集中报销/临近审批阈值',
                'tool': 'apriori_analysis.py（frequent模式）',
                'output': '报销异常疑点.xlsx',
                'time': '5分钟',
            },
            {
                'step': 4, 'phase': '人工核实',
                'task': '抽查凭证',
                'who': '👤 审计员',
                'detail': '针对AI标记的疑点，抽查原始凭证→核实业务真实性',
                'tool': '原始凭证/合同/审批单',
                'output': '凭证抽查记录',
                'time': '3-5小时',
            },
            {
                'step': 5, 'phase': '人工核实',
                'task': '访谈关键人员',
                'who': '👤 审计员',
                'detail': '访谈财务负责人/出纳/采购经办人→核实异常交易背景',
                'tool': '访谈提纲 + 录音/纪要',
                'output': '访谈记录',
                'time': '2-3小时',
            },
            {
                'step': 6, 'phase': '报告编制',
                'task': '问题定性与报告',
                'who': '👤👉🤖 审计员+AI',
                'detail': '审计员定性 → AI辅助匹配法规+生成整改建议+输出底稿',
                'tool': 'audit_finding_processor.py + analysis-report',
                'output': '审计报告初稿 + 底稿',
                'time': '2-4小时',
            },
        ],
        'checklist': [
            '□ 银行对账单与账面余额是否一致',
            '□ 大额支出是否有审批/合同/验收',
            '□ 费用报销附件是否齐全',
            '□ 是否存在坐收坐支',
            '□ 专项资金是否专款专用',
            '□ 往来款项是否及时清理',
            '□ 是否存在账外资金',
        ],
    },

    '绩效评价': {
        'alias': ['绩效', '评价', '考核', '目标'],
        'brief': '评价财政资金使用绩效，包括产出/效益/满意度',
        'duration': '10-20个工作日',
        'steps': [
            {'step': 1, 'phase': '项目准备', 'task': '收集绩效目标与自评材料', 'who': '🤖 AI',
             'detail': '采集绩效目标申报表/自评报告/项目验收材料', 'tool': 'OCR + data-analyst-cn',
             'output': '结构化绩效数据表', 'time': '15分钟'},
            {'step': 2, 'phase': '数据分析', 'task': '绩效指标完成度分析', 'who': '🤖 AI',
             'detail': '目标值vs实际完成值对比+同类项目横向对比', 'tool': 'data-analyst-cn',
             'output': '绩效差异分析表', 'time': '10分钟'},
            {'step': 3, 'phase': '数据分析', 'task': '资金使用效率分析', 'who': '🤖 AI',
             'detail': '预算执行率/成本节约率/投入产出比趋势', 'tool': 'forecast-simulation + anomaly-detection',
             'output': '资金效率分析.xlsx', 'time': '5分钟'},
            {'step': 4, 'phase': '人工核实', 'task': '现场勘查', 'who': '👤 审计员',
             'detail': '实地查看项目成果→核实自评报告真实性→访谈受益人', 'tool': '实地走访+拍照取证',
             'output': '现场勘查记录', 'time': '1-2天'},
            {'step': 5, 'phase': '人工核实', 'task': '满意度调查核验', 'who': '👤 审计员',
             'detail': '抽查满意度问卷/回访被调查人→验证数据真实性', 'tool': '电话回访+问卷抽查',
             'output': '满意度核验记录', 'time': '3-5小时'},
            {'step': 6, 'phase': '报告编制', 'task': '评价结论与报告', 'who': '👤👉🤖 审计员+AI',
             'detail': '审计员判断绩效等级 → AI辅助生成报告+建议', 'tool': 'analysis-report + audit_risk_navigator',
             'output': '绩效评价报告', 'time': '3-5小时'},
        ],
        'checklist': [
            '□ 绩效目标是否SMART（具体/可衡量/可达成/相关/有时限）',
            '□ 自评报告数据是否有佐证材料',
            '□ 资金使用是否按预算执行',
            '□ 产出数量/质量是否达到目标',
            '□ 效益指标是否有量化证据',
            '□ 满意度调查样本是否充分',
            '□ 是否存在"重投入轻产出"现象',
        ],
    },

    '资产清查': {
        'alias': ['资产', '清查', '盘点', '固定资产', '折旧'],
        'brief': '核实资产账实相符性，发现盘盈/盘亏/闲置/报废',
        'duration': '5-15个工作日',
        'steps': [
            {'step': 1, 'phase': '项目准备', 'task': '采集资产数据', 'who': '🤖 AI',
             'detail': '采集资产卡片/台账/折旧明细/盘点表', 'tool': 'data-analyst-cn',
             'output': '资产数据汇总表', 'time': '10分钟'},
            {'step': 2, 'phase': '数据分析', 'task': '异常资产识别', 'who': '🤖 AI',
             'detail': '长期未使用资产/账实差异/折旧异常/报废异常', 'tool': 'anomaly-detection + data-analyst-cn',
             'output': '资产异常疑点.xlsx', 'time': '5分钟'},
            {'step': 3, 'phase': '人工核实', 'task': '实地盘查', 'who': '👤 审计员',
             'detail': '根据AI疑点清单→重点抽查实物→核对标签/型号/使用状态',
             'tool': '盘点表 + 拍照 + 资产标签核对',
             'output': '实地盘点记录', 'time': '1-3天'},
            {'step': 4, 'phase': '人工核实', 'task': '权属核实', 'who': '👤 审计员',
             'detail': '核查不动产权证/车辆登记证/无形资产权属证明',
             'tool': '权属证明文件',
             'output': '权属核实记录', 'time': '2-4小时'},
            {'step': 5, 'phase': '报告编制', 'task': '清查结论与报告', 'who': '👤👉🤖 审计员+AI',
             'detail': '审计员判断盘盈盘亏原因 → AI辅助生成处理建议',
             'tool': 'audit_finding_processor.py + analysis-report',
             'output': '资产清查报告', 'time': '2-4小时'},
        ],
        'checklist': [
            '□ 资产卡片信息是否与实物一致',
            '□ 折旧计提是否正确（方法/年限/残值率）',
            '□ 已报废资产是否已销账',
            '□ 新购资产是否及时入账',
            '□ 资产出租出借是否履行审批',
            '□ 资产处置是否经过评估',
        ],
    },
}


def generate_plan(project_type: str, output_dir: str):
    """生成审计项目执行计划书"""

    # Fuzzy match
    matched = project_type
    if project_type not in KNOWLEDGE:
        for name, info in KNOWLEDGE.items():
            for alias in info.get('alias', []):
                if alias in project_type:
                    matched = name
                    break
            if matched != project_type:
                break

    if matched not in KNOWLEDGE:
        print(f"❌ 未找到匹配的审计类型: {project_type}")
        print(f"   支持: {', '.join(KNOWLEDGE.keys())}")
        return

    k = KNOWLEDGE[matched]
    os.makedirs(output_dir, exist_ok=True)

    # ── Excel 主计划表 ──
    wb = Workbook()

    # === Sheet 1: 项目概览 ===
    ws = wb.active
    ws.title = '项目概览'

    ws.merge_cells('A1:H1')
    ws['A1'] = f'📋 {matched} — 项目执行计划书'
    ws['A1'].font = TITLE
    ws.row_dimensions[1].height = 40

    ws.merge_cells('A2:H2')
    ws['A2'] = f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M")}    预计工期: {k["duration"]}    简介: {k["brief"]}'
    ws['A2'].font = SUBTITLE

    # Summary stats
    steps = k['steps']
    ai_steps = [s for s in steps if 'AI' in s['who']]
    human_steps = [s for s in steps if '审计员' in s['who']]

    ws.merge_cells('A4:H4')
    ws['A4'] = f'📊 概览:  共{len(steps)}个步骤 | 🤖 AI执行: {len(ai_steps)}步 | 👤 审计员执行: {len(human_steps)}步 | ⏱️ AI节省时间占比: {len(ai_steps)*100//len(steps)}%'
    ws['A4'].font = B
    ws.row_dimensions[4].height = 30

    # 分工饼图（文字版）
    ws.merge_cells('A6:B10')
    ws['A6'] = f'🤖 AI负责\n═══════\n{len(ai_steps)} 个步骤\n\n• 数据采集与清洗\n• 异常模式扫描\n• 文件对比分析\n• 底稿与报告生成\n\n→ 释放审计员{len(ai_steps)*100//len(steps)}%的\n  重复性工作时间'
    ws['A6'].font = N; ws['A6'].alignment = W; ws['A6'].fill = GRN

    ws.merge_cells('C6:D10')
    ws['C6'] = f'👤 审计员负责\n══════════\n{len(human_steps)} 个步骤\n\n• 复核AI发现的疑点\n• 工商关联查询\n• 访谈关键人员\n• 职业判断与结论\n\n→ 聚焦高价值专业工作'
    ws['C6'].font = N; ws['C6'].alignment = W; ws['C6'].fill = BLUE

    ws.merge_cells('E6:H10')
    ws['E6'] = f'📦 产出清单\n══════════\n\n' + '\n'.join([f'• {s["output"]}' for s in steps])
    ws['E6'].font = N; ws['E6'].alignment = W; ws['E6'].fill = YEL

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18

    # === Sheet 2: 执行步骤 ===
    ws2 = wb.create_sheet('执行步骤')

    ws2.merge_cells('A1:H1')
    ws2['A1'] = f'📝 详细执行步骤 — {matched}'
    ws2['A1'].font = Font(name='Microsoft YaHei', size=14, bold=True, color='1A3A6E')

    headers = ['步骤', '阶段', '任务', '执行者', '具体操作', '使用工具/命令', '预期产出', '预估时间']
    for c, h in enumerate(headers, 1):
        cl = ws2.cell(row=3, column=c, value=h)
        cl.font = H; cl.fill = HEADER; cl.alignment = C; cl.border = TH

    for i, s in enumerate(steps):
        r = i + 4
        vals = [s['step'], s['phase'], s['task'], s['who'], s['detail'], s['tool'], s['output'], s['time']]
        for c, val in enumerate(vals, 1):
            cl = ws2.cell(row=r, column=c, value=val)
            cl.font = N; cl.alignment = W if c >= 4 else C; cl.border = TH
            if c == 4:
                cl.fill = GRN if 'AI' in str(val) else BLUE; cl.alignment = C
                cl.font = B

        ws2.row_dimensions[r].height = 50 if len(s['detail']) > 40 else 30

    ws2.column_dimensions['A'].width = 6
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 22
    ws2.column_dimensions['D'].width = 14
    ws2.column_dimensions['E'].width = 50
    ws2.column_dimensions['F'].width = 40
    ws2.column_dimensions['G'].width = 30
    ws2.column_dimensions['H'].width = 12

    # === Sheet 3: 合规检查清单 ===
    ws3 = wb.create_sheet('合规检查清单')

    ws3.merge_cells('A1:C1')
    ws3['A1'] = f'✅ 审计合规检查清单 — {matched}'
    ws3['A1'].font = Font(name='Microsoft YaHei', size=14, bold=True, color='1A3A6E')

    ws3.merge_cells('A2:C2')
    ws3['A2'] = '说明: 逐项检查，已完成打√，发现问题在备注中记录'
    ws3['A2'].font = SUBTITLE

    for c, h in enumerate(['检查项', '状态', '备注/发现问题'], 1):
        cl = ws3.cell(row=4, column=c, value=h)
        cl.font = H; cl.fill = HEADER; cl.alignment = C; cl.border = TH

    for i, item in enumerate(k.get('checklist', [])):
        r = i + 5
        ws3.cell(row=r, column=1, value=item).font = N
        ws3.cell(row=r, column=1).alignment = L; ws3.cell(row=r, column=1).border = TH
        ws3.cell(row=r, column=2, value='☐ 待检查').font = N
        ws3.cell(row=r, column=2).alignment = C; ws3.cell(row=r, column=2).border = TH
        ws3.cell(row=r, column=2).fill = YEL
        ws3.cell(row=r, column=3).border = TH

    ws3.column_dimensions['A'].width = 45
    ws3.column_dimensions['B'].width = 14
    ws3.column_dimensions['C'].width = 40

    # === Sheet 4: AI工具速查 ===
    ws4 = wb.create_sheet('AI工具速查')

    ws4.merge_cells('A1:D1')
    ws4['A1'] = '🛠️ 本项目可用AI工具速查'
    ws4['A1'].font = Font(name='Microsoft YaHei', size=14, bold=True, color='1A3A6E')

    tools_used = list(set(s['tool'] for s in steps if 'AI' in s['who']))
    h4 = ['工具', '用途', '命令示例', '需要什么数据']
    for c, h in enumerate(h4, 1):
        cl = ws4.cell(row=3, column=c, value=h)
        cl.font = H; cl.fill = HEADER; cl.alignment = C; cl.border = TH

    # Map tools to commands
    tool_map = {
        'audit_pipeline.py': ('审计全流程自动化', 'python audit_pipeline.py -p "项目目录" -o "输出目录"', '项目文件夹（含PDF/DOCX/XLSX）'),
        'audit_finding_processor.py': ('审计发现智能处理', 'python audit_finding_processor.py "发现描述" --o 底稿.xlsx', '审计发现描述文本'),
        'audit_risk_navigator.py': ('审计风险导航', 'python audit_risk_navigator.py "招投标采购审计"', '审计类型名称'),
        'apriori_analysis.py': ('Apriori关联规则', 'python apriori_analysis.py --i 交易数据.xlsx --mode frequent', '事务ID+项 两列数据'),
        'financial-fraud-detection': ('舞弊风险扫描', '触发: "检查财务数据有没有异常"', '财务数据（科目余额表等）'),
        'anomaly-detection': ('时序异常检测', '触发: "检查这个指标是否正常"', '时序数据'),
        'data-analyst-cn': ('通用数据分析', '触发: "分析这个数据表"', 'Excel/CSV数据'),
        'analysis-report': ('自动生成报告', '触发: "出一份审计报告"', '分析步骤已完成'),
    }

    for i, tool_name in enumerate(tools_used):
        r = i + 4
        base = tool_name.split('.')[0]
        info = tool_map.get(base, tool_map.get(tool_name, (tool_name, '', '')))
        for c, val in enumerate([info[0], info[1], info[2]], 1):
            ws4.cell(row=r, column=1, value=tool_name).font = B
            ws4.cell(row=r, column=1).border = TH; ws4.cell(row=r, column=1).alignment = L
            ws4.cell(row=r, column=c+1, value=val).font = N
            ws4.cell(row=r, column=c+1).border = TH; ws4.cell(row=r, column=c+1).alignment = W

    ws4.column_dimensions['A'].width = 32
    ws4.column_dimensions['B'].width = 22
    ws4.column_dimensions['C'].width = 50
    ws4.column_dimensions['D'].width = 30

    # Save
    plan_path = Path(output_dir) / f'{matched}_项目执行计划书.xlsx'
    wb.save(plan_path)
    print(f"\n✅ 计划书: {plan_path}")

    # ── 命令行摘要 ──
    print(f"\n{'='*60}")
    print(f"📋 {matched} — 快速指引")
    print(f"{'='*60}")
    print(f"\n⏱️ 预计工期: {k['duration']}")
    print(f"🤖 AI负责: {len(ai_steps)}/{len(steps)} 步骤")
    print(f"👤 你负责: {len(human_steps)}/{len(steps)} 步骤")
    print(f"\n📝 你的核心任务:")
    for s in human_steps:
        print(f"   {s['step']}. {s['task']} ({s['time']})")
    print(f"\n💡 新手提示:")
    print(f"   1. 先让AI跑完所有自动分析，再开始人工核查")
    print(f"   2. AI标注的疑点都是'可能有问题'，需要你判断'是否真有问题'")
    print(f"   3. 合规检查清单在Excel的Sheet3，逐项打勾不漏项")
    print(f"   4. 遇到问题就问AI助手，比如'这个围标证据够不够充分'")
    print(f"{'='*60}")

    return plan_path


def interactive():
    """交互式向导"""
    print("╔══════════════════════════════════════════╗")
    print("║     🎯 审计新手项目助手 v1.0            ║")
    print("║     第一性原理驱动 · 拿来就用          ║")
    print("╚══════════════════════════════════════════╝\n")

    print("支持的项目类型:")
    for i, (name, info) in enumerate(KNOWLEDGE.items(), 1):
        print(f"  {i}. {name} → {info['brief']}")

    print(f"\n  也可以输入关键词匹配，如 '招标'、'财务'、'绩效'、'资产'\n")

    choice = input("▶ 选择项目类型 (输入编号或名称) > ").strip()

    try:
        idx = int(choice) - 1
        keys = list(KNOWLEDGE.keys())
        if 0 <= idx < len(keys):
            proj_type = keys[idx]
        else:
            proj_type = choice
    except ValueError:
        proj_type = choice

    output = input("▶ 输出目录 (回车默认: ./审计项目计划) > ").strip()
    output = output or './审计项目计划'

    project_name = input("▶ 项目名称/编号 (可选) > ").strip()
    if project_name:
        output = f"{output}/{project_name}"

    print()
    generate_plan(proj_type, output)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='审计新手项目助手 — 一键生成执行计划书')
    parser.add_argument('--type', '-t', help='项目类型: 招投标采购审计/财务收支审计/绩效评价/资产清查')
    parser.add_argument('--o', '--output', dest='output', default='./审计项目计划',
                        help='输出目录 (默认: ./审计项目计划)')
    parser.add_argument('--list', action='store_true', help='列出所有支持的项目类型')
    args = parser.parse_args()

    if args.list:
        print("支持的项目类型:")
        for name, info in KNOWLEDGE.items():
            print(f"  {name}: {info['brief']}")
    elif args.type:
        generate_plan(args.type, args.output)
    else:
        interactive()
