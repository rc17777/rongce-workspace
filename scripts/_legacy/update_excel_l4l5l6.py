"""Add L4/L5/L6 sheets to the existing Excel report"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.load_workbook(r'D:\openclaw-workspace\output\急救实训室_extracted\招投标审计分析报告.xlsx')

# Common styles
header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
sub_header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
red_font = Font(name='微软雅黑', size=11, color='FF0000', bold=True)
bold_font = Font(name='微软雅黑', bold=True, size=11)
normal_font = Font(name='微软雅黑', size=11)
small_font = Font(name='微软雅黑', size=10, color='666666')
title_font = Font(name='微软雅黑', bold=True, size=14, color='2F5496')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
wrap_align = Alignment(vertical='center', wrap_text=True)

def apply_style(ws, row, col, value, font=normal_font, fill=None, alignment=center_align, border=thin_border):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if fill: cell.fill = fill
    cell.alignment = alignment
    if border: cell.border = border
    return cell

def write_header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        apply_style(ws, row, start_col+i, h, font=header_font, fill=header_fill)

# ============ Sheet: L4图片哈希 ============
ws_l4 = wb.create_sheet('L4-图片哈希比对')

apply_style(ws_l4, 1, 1, 'L4: PDF嵌入图片哈希比对', font=title_font, fill=None, border=None)
apply_style(ws_l4, 2, 1, '方法：使用PyMuPDF提取PDF中所有嵌入图片 → SHA256哈希 → 跨投标人比对', font=small_font, fill=None, border=None)

r = 4
write_header_row(ws_l4, r, ['投标人', '嵌入图片数', '独立图片SHA256', '跨公司重复', '说明'])
data = [
    ['四川省好医助医疗器械有限公司', 39, 39, 0, '均为独立图片，无跨公司重复'],
    ['成都易可天地科技有限公司', 208, 208, 0, '图片量远超另两家（208 vs 39/32），可能是证明材料截图多'],
    ['江西正好医疗器械有限公司', 32, 32, 0, '均为独立图片，无跨公司重复'],
]
for i, row_data in enumerate(data):
    for j, val in enumerate(row_data):
        fill = green_fill
        apply_style(ws_l4, 5+i, j+1, val, font=normal_font, fill=fill, 
                    alignment=center_align if j in [1,2,3] else wrap_align)

r = 10
apply_style(ws_l4, r, 1, '结论', font=bold_font, fill=sub_header_fill, alignment=wrap_align)
ws_l4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
conclusion_l4 = '共提取279个图片实例，276个独立SHA256值，3个实例为同投标人内不同PDF间重复。0张图片跨公司重复。排除通过共享扫描件/印章/资质图片实现围标的可能。'
apply_style(ws_l4, 11, 1, conclusion_l4, font=normal_font, fill=green_fill, alignment=wrap_align)
ws_l4.merge_cells(start_row=11, start_column=1, end_row=11, end_column=5)

ws_l4.column_dimensions['A'].width = 30
ws_l4.column_dimensions['B'].width = 14
ws_l4.column_dimensions['C'].width = 18
ws_l4.column_dimensions['D'].width = 14
ws_l4.column_dimensions['E'].width = 50

# ============ Sheet: L5元数据 ============
ws_l5 = wb.create_sheet('L5-PDF元数据')

apply_style(ws_l5, 1, 1, 'L5: PDF元数据交叉分析（关键证据）', font=title_font, fill=None, border=None)
apply_style(ws_l5, 2, 1, '方法：使用PyMuPDF提取PDF元数据字段（Producer/Creator/Author/CreationDate）', font=small_font, fill=None, border=None)

# Summary table
r = 4
apply_style(ws_l5, r, 1, 'A. 文档作者（.docx → PDF转换的WPS文档）', font=bold_font, fill=sub_header_fill, alignment=wrap_align)
ws_l5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)

r = 5
write_header_row(ws_l5, r, ['PDF文件', '好医助 Author', '易可天地 Author', '江西正好 Author', 'Creator软件', 'CreationDate(好医助)', 'CreationDate(易可天地)'])

doc_data = [
    ['其他资料.docx.pdf', '何天真', 'ZXM +++', 'ZXM +++', 'WPS文字(全部)', '05-23 12:37', '05-22 18:07'],
    ['技术要求应答表.docx.pdf', '何天真', 'WPS_1654344806 +++', 'WPS_1654344806 +++', 'WPS文字(全部)', '05-23 12:37', '05-22 15:05'],
    ['投标（响应）函.docx.pdf', '何天真', '汪 +++', '汪 +++', 'WPS文字(全部)', '05-23 12:37', '05-22 15:05'],
    ['服务内容要求、商务要求应答表.docx.pdf', '何天真', 'WPS_1654344806 +++', 'WPS_1654344806 +++', 'WPS文字(全部)', '05-23 12:37', '05-22 15:04'],
    ['中小企业声明函.pdf', '(空)', '(空)', '(空)', 'Qt 5.15.2(全部)', '05-23 12:34', '05-22 15:04'],
    ['残疾人福利性单位声明函.pdf', '(空)', '(空)', '(空)', 'Qt 5.15.2(全部)', '05-23 12:34', '05-22 15:04'],
    ['监狱企业的证明文件.pdf', '(空)', '(空)', '(空)', 'Qt 5.15.2(全部)', '05-23 12:34', '05-22 15:04'],
    ['报价表.pdf', '(空)', '(空)', '(空)', 'Chromium/Skia(全部)', '05-23 05:12', '05-26 01:59'],
]

for i, row_data in enumerate(doc_data):
    for j, val in enumerate(row_data):
        # Highlight shared authors
        is_shared = '+++' in str(val)
        clean_val = str(val).replace(' +++', '')
        fill = red_fill if is_shared else None
        font_style = red_font if is_shared else normal_font
        apply_style(ws_l5, 6+i, j+1, clean_val, font=font_style, fill=fill, 
                    alignment=center_align if j > 0 else wrap_align)

# Software summary
r = 15
apply_style(ws_l5, r, 1, 'B. 软件环境汇总', font=bold_font, fill=sub_header_fill, alignment=wrap_align)
ws_l5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

r = 16
write_header_row(ws_l5, r, ['软件类型', '好医助', '易可天地', '江西正好', '一致性'])
sw_data = [
    ['文档编辑(含docx的PDF)', 'WPS文字 (全部一致)', 'WPS文字 (全部一致)', 'WPS文字 (全部一致)', '三方完全一致'],
    ['报价表(含docx的PDF)', 'Chromium/Skia', 'Chromium/Skia', 'Chromium/Skia', '三方完全一致(平台生成)'],
    ['模板PDF', 'Qt 5.15.2', 'Qt 5.15.2', 'Qt 5.15.2', '三方完全一致(模板)'],
    ['关键.doc作者', '何天真(唯一)', 'ZXM/WPS_1654344806/汪', 'ZXM/WPS_1654344806/汪', '易可天地=江西正好!!'],
]
for i, row_data in enumerate(sw_data):
    for j, val in enumerate(row_data):
        is_red = '!!' in str(val)
        clean = str(val).replace('!!', '')
        fill = red_fill if is_red else None
        font_style = red_font if is_red else normal_font
        apply_style(ws_l5, 17+i, j+1, clean, font=font_style, fill=fill,
                    alignment=center_align if j > 0 else wrap_align)

# Key findings
r = 22
apply_style(ws_l5, r, 1, 'C. 关键发现', font=bold_font, fill=red_fill, alignment=wrap_align)
ws_l5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)

findings_l5 = [
    '1. [铁证级别] 易可天地和江西正好的.docx文档作者完全一致: ZXM / WPS_1654344806 / 汪',
    '   - 其他资料.docx.pdf: 两家的Author都是"ZXM"',
    '   - 技术要求应答表.docx.pdf: 两家的Author都是"WPS_1654344806"(WPS安装时间戳)',
    '   - 投标（响应）函.docx.pdf: 两家的Author都是"汪"',
    '   - 服务内容要求、商务要求应答表.docx.pdf: 两家的Author都是"WPS_1654344806"',
    '2. WPS_1654344806 = Unix时间戳 → 2022-06-04 17:53 (WPS安装时间)，两台机器的WPS安装timestamp不可能相同',
    '3. 好医助所有docx文档Author统一为"何天真"，与另两家完全不同',
    '4. 三家均使用WPS文字编辑.docx文档，再转换为PDF上传',
    '5. 模板PDF（声明函等）三家均用Qt 5.15.2生成，来源一致（招标平台统一模板）',
    '6. 报价表PDF三家均用Chromium/Skia生成（电子交易平台网页导出）',
    '',
    '>>> 结论: 易可天地和江西正好的投标文件由同一人/同一台电脑制作，构成串标围标的强证据 <<<',
]
for i, f_text in enumerate(findings_l5):
    is_conclusion = '>>>' in f_text
    font_style = red_font if is_conclusion else normal_font
    fill = red_fill if is_conclusion else None
    apply_style(ws_l5, 23+i, 1, f_text, font=font_style, fill=fill, alignment=wrap_align)
    ws_l5.merge_cells(start_row=23+i, start_column=1, end_row=23+i, end_column=7)

ws_l5.column_dimensions['A'].width = 35
for c in range(2, 8):
    ws_l5.column_dimensions[chr(64+c)].width = 24

# ============ Sheet: L6字体结构 ============
ws_l6 = wb.create_sheet('L6-字体文档结构')

apply_style(ws_l6, 1, 1, 'L6: PDF字体使用与文档结构分析', font=title_font, fill=None, border=None)
apply_style(ws_l6, 2, 1, '方法：使用PyMuPDF提取每页文本span的字体属性，统计字体使用频次', font=small_font, fill=None, border=None)

r = 4
apply_style(ws_l6, r, 1, 'A. 字体使用频次对比', font=bold_font, fill=sub_header_fill, alignment=wrap_align)
ws_l6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

r = 5
write_header_row(ws_l6, r, ['字体', '好医助(8169次/6种)', '易可天地(11461次/10种)', '江西正好(5168次/14种)', '跨公司共用'])

font_data = [
    ['FangSong', '6207 (76.0%)', '9444 (82.4%)', '13 (0.3%)', '三方共用'],
    ['MicrosoftYaHei', '893 (10.9%)', '905 (7.9%)', '904 (17.5%)', '三方共用'],
    ['SimSun', '747 (9.1%)', '128 (1.1%)', '3965 (76.7%)', '三方共用，江西以宋体为主'],
    ['SimHei', '274 (3.4%)', '690 (6.0%)', '214 (4.1%)', '三方共用'],
    ['MicrosoftYaHei-Bold', '21 (0.3%)', '21 (0.2%)', '—', '好医助+易可天地'],
    ['Calibri', '—', '234 (2.0%)', '—', '仅易可天地'],
    ['SegoePrint-Bold', '—', '8 (0.1%)', '—', '仅易可天地'],
    ['FangSong_GB2312', '—', '2 (0.02%)', '—', '仅易可天地'],
    ['ArialMT', '—', '2 (0.02%)', '—', '仅易可天地'],
    ['Type3字体(嵌入)', '—', '—', '51 (1.0%)', '仅江西正好(报价表)'],
]
for i, row_data in enumerate(font_data):
    for j, val in enumerate(row_data):
        apply_style(ws_l6, 6+i, j+1, val, font=normal_font, 
                    alignment=center_align if j > 0 else wrap_align)

r = 17
apply_style(ws_l6, r, 1, 'B. 结构特征总结', font=bold_font, fill=sub_header_fill, alignment=wrap_align)
ws_l6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

struct_findings = [
    '1. 好医助: 字体使用简洁(6种)，以方正仿宋为主(76%)，微软雅黑为辅(11%)',
    '2. 易可天地: 字体最丰富(10种)，以方正仿宋为主(82%)，另有Calibri/Arial等西文字体',
    '3. 江西正好: 字体类别最多(14种)但以宋体为主(77%)，使用嵌入式Type3字体(约占1%)',
    '4. 好医助的6种字体全部被易可天地的10种字体集合包含',
    '5. 江西正好的主导字体(SimSun 77%)与另两家(FangSong 76%/82%)完全不同',
    '6. 所有三家共用4种基础中文字体: FangSong, MicrosoftYaHei, SimSun, SimHei',
    '',
    '结论: 字体层面仅能反映文档制作习惯差异（好医助/易可天地偏仿宋，江西正好偏宋体），',
    '        未构成独立的串标证据。但好医助字体完全被易可天地包含，值得关注。'
]
for i, f_text in enumerate(struct_findings):
    apply_style(ws_l6, 18+i, 1, f_text, font=normal_font, alignment=wrap_align)
    ws_l6.merge_cells(start_row=18+i, start_column=1, end_row=18+i, end_column=5)

ws_l6.column_dimensions['A'].width = 28
ws_l6.column_dimensions['B'].width = 28
ws_l6.column_dimensions['C'].width = 28
ws_l6.column_dimensions['D'].width = 28
ws_l6.column_dimensions['E'].width = 28

# ============ Update Sheet 1 (综合结论) with new findings ============
ws1 = wb['综合审计结论']

# Update the conclusion
r_update = 20
apply_style(ws1, r_update, 1, '三、核心风险信号（已更新L4/L5/L6数据）', font=bold_font, fill=red_fill, alignment=wrap_align)
ws1.merge_cells(start_row=r_update, start_column=1, end_row=r_update, end_column=2)

updated_risks = [
    '1. [L5铁证] 易可天地 & 江西正好 共享同一文档作者(ZXM/WPS_1654344806/汪) — 串标强证据',
    '2. [L1] 报价极差仅1.21%（校服项目7.8%），离散度异常',
    '3. [L1] 所有38项分项均紧贴限价97-99.9%，不同品牌独立定价不应如此一致',
    '4. [追加] 三家全部放弃中小企业10%价格扣除',
    '5. [追加] 评审得分4/5评委完全一致（精确到0.01分），唯一不同评委恰好各少2分',
    '6. [新增] 好医助字体(6种)完全被易可天地字体(10种)包含'
]
for i, risk in enumerate(updated_risks):
    f = red_font if '[L5铁证]' in risk or '[追加]' in risk else normal_font
    fill_c = red_fill if '铁证' in risk else None
    apply_style(ws1, 21+i, 1, risk, font=f, fill=fill_c, alignment=wrap_align)
    ws1.merge_cells(start_row=21+i, start_column=1, end_row=21+i, end_column=2)

# Update final judgment
r_judge = 35
ws1.cell(row=r_judge, column=1).value = None
apply_style(ws1, r_judge, 1, '六、综合判定（基于L5新证据更新）', font=bold_font, fill=red_fill, alignment=wrap_align)
ws1.merge_cells(start_row=r_judge, start_column=1, end_row=r_judge, end_column=2)
judgment = (
    'L5元数据证据确认：易可天地与江西正好的.docx文档由同一人/同一台电脑制作（Author字段完全一致）。'
    '结合报价集中度(1.21%)和评审得分一致性异常，串标围标可能性从"存在疑点"升级为"高度可疑"。'
    '核心判断：易可天地与江西正好之间存在明确的文档同源性，构成串标围标的强证据。'
)
apply_style(ws1, r_judge+1, 1, judgment, font=red_font, fill=red_fill, alignment=wrap_align)
ws1.merge_cells(start_row=r_judge+1, start_column=1, end_row=r_judge+1, end_column=2)

wb.save(r'D:\openclaw-workspace\output\急救实训室_extracted\招投标审计分析报告_v2.xlsx')
print('Saved: v2.xlsx')
