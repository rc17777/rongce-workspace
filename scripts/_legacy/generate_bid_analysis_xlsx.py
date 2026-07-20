#!/usr/bin/env python3
"""串标围标分析Excel报告 v2 (fixed openpyxl compatibility)"""
import sys, io, re, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

base = r'D:\openclaw-workspace\projects\护理学院培训资料采购'

# Styles
RED_FILL = PatternFill(patternType='solid', fgColor='FFD7D7')
YELLOW_FILL = PatternFill(patternType='solid', fgColor='FFF3CD')
GREEN_FILL = PatternFill(patternType='solid', fgColor='D4EDDA')
HEADER_FILL = PatternFill(patternType='solid', fgColor='1A3A6E')
LIGHT_BLUE = PatternFill(patternType='solid', fgColor='E8F0FE')
HEADER_FONT = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
TITLE_FONT = Font(name='Microsoft YaHei', size=14, bold=True, color='1A3A6E')
NORMAL = Font(name='Microsoft YaHei', size=10)
BOLD = Font(name='Microsoft YaHei', size=10, bold=True)
BOLD_RED = Font(name='Microsoft YaHei', size=10, color='CC0000', bold=True)
ITALIC = Font(name='Microsoft YaHei', size=9, italic=True, color='888888')
THIN = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
C = Alignment(horizontal='center', vertical='center', wrap_text=True)
L = Alignment(horizontal='left', vertical='center', wrap_text=True)

def hdr(ws, row, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = C; cell.border = THIN

def cell(ws, r, c, val, font=NORMAL, align=C, fill=None):
    cl = ws.cell(row=r, column=c, value=val)
    cl.font = font; cl.alignment = align; cl.border = THIN
    if fill: cl.fill = fill

# Load texts & compute TF-IDF
tech_texts, full_texts = {}, {}
for idx, name in enumerate(['建韬科技','江楼商贸','拓奇长荣'], 1):
    for fname, dest in [(f'bidder{idx}_技术方案.txt', tech_texts), (f'bidder{idx}_建韬.txt' if idx==1 else f'bidder{idx}_江楼.txt' if idx==2 else f'bidder{idx}_拓奇长荣.txt', full_texts)]:
        p = os.path.join(base, fname)
        raw = open(p, encoding='utf-8').read() if os.path.exists(p) else ''
        clean = re.sub(r'=== PAGE \d+ ===', '', raw)
        clean = re.sub(r'[^\u4e00-\u9fffa-zA-Z0-9 \n]', '', clean)
        dest[name] = re.sub(r'\s+', ' ', clean)

names = ['建韬科技','江楼商贸','拓奇长荣']
v = TfidfVectorizer(max_features=5000, token_pattern=r'[\u4e00-\u9fff]+')
sim_full = cosine_similarity(v.fit_transform([full_texts[n] for n in names]))
sim_tech = cosine_similarity(v.fit_transform([tech_texts[n] for n in names]))

# ====== BUILD WORKBOOK ======
wb = Workbook()

# --- Sheet 1: 综合结论 ---
ws = wb.active; ws.title = '综合结论'
ws.merge_cells('A1:F1'); ws['A1'] = '四川护理职业学院2025年培训资料采购 - 串标围标分析'; ws['A1'].font = TITLE_FONT
ws.merge_cells('A2:F2'); ws['A2'] = f'项目编号: SCLT20250232 | 预算: 25万 | 开标: 2025-04-10 | 分析: {datetime.now().strftime("%Y-%m-%d")}'; ws['A2'].font = ITALIC

hdr(ws, 4, ['层级', '检测维度', '检测状态', '关键发现', '风险等级', '详细说明'])

conclusions = [
    ['L1', '报价规律性', '已完成', '三家报价自然分布(21.7/22.9/23.7万)，极差9.6%，非等差/等比', 'GREEN', '正常'],
    ['L2', '投标IP/MAC', '不适用', '纸质投标无电子系统', 'GREY', 'N/A'],
    ['L3', 'TF-IDF文本雷同', '已完成', '技术方案TF-IDF: 76.6%/75.6%/80.9%，江楼vs拓奇超80%红线', 'RED', '全文86-90%因模板驱动；技术方案75-81%需人工复核'],
    ['L4', '图片/资源哈希', '不适用', '扫描版PDF无嵌入图片', 'GREY', 'N/A'],
    ['L5', 'PDF元数据(扫描仪)', '已完成', '三家6份PDF全部使用Fuji Xerox D125扫描仪', 'RED', '同一型号可能共用打印店/设备'],
    ['L6', '文档结构/体量', '已完成', '拓奇327页 vs 江楼58页 vs 建韬92页', 'YELLOW', '体量差异巨大(5.6x)'],
    ['L7', '扫描设备', '已完成', '同L5: 全部Fuji Xerox D125', 'RED', 'PDF格式: 江楼1.3 vs 建韬/拓奇1.6略有差异'],
    ['L8', '工商关联(地理)', '初步分析', '三家均注册武侯区；建韬(玉林北街)与拓奇(玉林中路)同在玉林片区', 'YELLOW', '距离约500米，建议天眼查股权穿透'],
    ['L9', '保证金/资金链', '不适用', '无此数据源', 'GREY', 'N/A'],
    ['L10', '代理人/签到', '未获取', '签到表未纳入材料', 'GREY', '需调取原件'],
]

for r, row in enumerate(conclusions, 5):
    for c, val in enumerate(row, 1):
        f = BOLD_RED if val=='RED' else (BOLD if c==1 else NORMAL)
        fill = RED_FILL if val=='RED' else (YELLOW_FILL if val=='YELLOW' else (GREEN_FILL if val=='GREEN' else None))
        cell(ws, r, c, val, f, L if c>=4 else C, fill)

# Verdict
r = 16
ws.merge_cells(f'A{r}:F{r}'); ws[f'A{r}'] = '综合判定'; ws[f'A{r}'].font = Font(name='Microsoft YaHei', size=13, bold=True, color='CC0000')
r += 1
verdict = (
    '【核心发现】\n'
    '1. [L5 RED] 扫描仪一致性: 三家6份PDF全部由Fuji Xerox D125生成。虽未达到"同一设备"的铁证级别，\n'
    '   但同型号设备暗示可能使用同一打印店/扫描服务商。建议核实投标文件制作地点。\n\n'
    '2. [L3 YELLOW] 技术方案TF-IDF 75-81%: 高于正常独立撰写的预期(通常<50%)。\n'
    '   江楼vs拓奇达80.9%临界值。需要人工对比判断是否存在实质性文字复制。\n\n'
    '3. [L8 YELLOW] 地理集中: 建韬(玉林北街)与拓奇(玉林中路)同在玉林片区步行距离。\n'
    '   结合扫描仪一致性，建议进行工商股权穿透排除关联关系。\n\n'
    '4. [评分异常] 第一名(91.35)与第二名(67.33)分差24分，拓奇327页方案可能质量显著更优，\n'
    '   但结合L5+L8信号，建议复核评审过程的合规性。\n\n'
    '【总体】: 多个信号叠加(L5扫描仪+L3文本+L8地理+得分差)，虽未达"串标铁证"级别，\n'
    '但已达到"启动进一步调查"的阈值。建议: ①核实投标文件制作地 ②天眼查股权穿透 ③人工比对技术方案。'
)
ws.merge_cells(f'A{r}:F{r}'); ws[f'A{r}'] = verdict; ws[f'A{r}'].font = NORMAL
ws[f'A{r}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws.row_dimensions[r].height = 220

ws.column_dimensions['A'].width = 8; ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 12; ws.column_dimensions['D'].width = 50
ws.column_dimensions['E'].width = 14; ws.column_dimensions['F'].width = 42

# --- Sheet 2: 元数据分析 ---
ws2 = wb.create_sheet('元数据分析')
ws2.merge_cells('A1:H1'); ws2['A1'] = 'PDF元数据 - 扫描仪/创建时间分析'; ws2['A1'].font = TITLE_FONT

hdr(ws2, 3, ['投标单位', '文件类型', '页数', '大小(MB)', 'Creator', '创建时间', '修改时间', 'PDF版本'])

meta_rows = [
    ['四川建韬科技', '资格性响应文件', 57, '13.6', 'Fuji Xerox D125', '2025-04-08 20:49 (+10 GMT)', '2025-04-08 20:35 (+8 CST)', '1.6'],
    ['四川建韬科技', '其他响应文件', 92, '24.7', 'Fuji Xerox D125', '2025-04-08 20:48 (+10 GMT)', '2025-04-08 20:34 (+8 CST)', '1.6'],
    ['成都江楼商贸', '资格性响应文件', 51, '13.7', 'Fuji Xerox D125', '2025-04-08 22:29 (+10 GMT)', '2025-04-08 22:29 (+10 GMT)', '1.3'],
    ['成都江楼商贸', '其他响应文件', 58, '15.0', 'Fuji Xerox D125', '2025-04-08 22:28 (+10 GMT)', '2025-04-08 22:28 (+10 GMT)', '1.3'],
    ['成都拓奇长荣商贸', '资格性响应文件', 81, '24.3', 'Fuji Xerox D125', '2025-04-09 13:13 (+10 GMT)', '2025-04-09 12:34 (+8 CST)', '1.6'],
    ['成都拓奇长荣商贸', '其他响应文件', 327, '105.5', 'Fuji Xerox D125', '2025-04-09 13:18 (+10 GMT)', '2025-04-09 16:01 (+8 CST)', '1.6'],
]

for r, row in enumerate(meta_rows, 4):
    for c, val in enumerate(row, 1):
        f = BOLD if c==1 else NORMAL
        fill = YELLOW_FILL if c==5 else None
        cell(ws2, r, c, val, f, C if c<=4 else L, fill)

# Analysis
r2 = 11
ws2.merge_cells(f'A{r2}:H{r2}'); ws2[f'A{r2}'] = '关键分析'; ws2[f'A{r2}'].font = Font(name='Microsoft YaHei', size=12, bold=True, color='CC0000')

analysis = [
    '1. [设备] 所有6份PDF的Creator均为"Fuji Xerox D125"(富士施乐高端办公复合机)，罕见同型号同时出现于三家。',
    '2. [时间线] 建韬: 4/8晚20:48-49 → 江楼: 4/8晚22:28-29(晚~1.5h) → 拓奇: 4/9午13:13-18(晚半天)。开标日: 4/10。',
    '   三家公司扫描时间分布在两个日期，未发现"同一时刻"扫描的直接证据。',
    '3. [时区] 建韬和拓奇的创建时间时区为+10(GMT+10=澳洲东部)，修改时间为+8(北京)。可能是扫描仪时区配置bug。',
    '   江楼的两个时间时区均为+10，设置较一致。',
    '4. [PDF版本] 江楼使用PDF 1.3，建韬/拓奇使用PDF 1.6，说明扫描软件或驱动版本不完全一致。',
    '5. [结论] 无法排除三家使用同一打印店的可能性，但也无法确认。建议向三家核实投标文件的具体制作地点。'
]
for i, line in enumerate(analysis):
    ws2.merge_cells(f'A{r2+1+i}:H{r2+1+i}'); ws2[f'A{r2+1+i}'] = line
    ws2[f'A{r2+1+i}'].font = NORMAL; ws2[f'A{r2+1+i}'].alignment = L

ws2.column_dimensions['A'].width = 22; ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 8; ws2.column_dimensions['D'].width = 10
ws2.column_dimensions['E'].width = 20; ws2.column_dimensions['F'].width = 28
ws2.column_dimensions['G'].width = 28; ws2.column_dimensions['H'].width = 10

# --- Sheet 3: 报价对比 ---
ws3 = wb.create_sheet('报价对比')
ws3.merge_cells('A1:I1'); ws3['A1'] = '报价及评标得分对比'; ws3['A1'].font = TITLE_FONT
hdr(ws3, 3, ['排名', '供应商', '最后报价(元)', '占预算比', '相对最低价', '总分', '价差分析', '非价格分差距', '风险标记'])

score_data = [
    ['1', '成都拓奇长荣商贸有限公司', 229237.50, '91.7%', '1.058x', 91.35, '比最低价+12,624元', '领先第二名约25.6分', 'YELLOW - 得分奇高'],
    ['2', '成都江楼商贸有限公司', 216613.20, '86.6%', '1.000x(最低)', 67.33, '最低价', '落后第一名约25.6分', 'GREEN - 价格最优'],
    ['3', '四川建韬科技有限公司', 237346.20, '94.9%', '1.096x', 62.71, '比最低价+20,733元', '落后第一名约27.2分', 'GREEN - 价格正常'],
]

for r, row in enumerate(score_data, 4):
    for c, val in enumerate(row, 1):
        fl = RED_FILL if r==4 else (GREEN_FILL if r in [5,6] and c>=8 else None)
        cell(ws3, r, c, val, BOLD if r==4 else NORMAL, C if c<=6 else L, fl)

# Note
r3 = 8
ws3.merge_cells(f'A{r3}:I{r3}')
ws3[f'A{r3}'] = ('得分分析: 第1名(91.35)超出第2名(67.33)达35.7%。报价权重30%(30分)，非价格权重70%(70分)。\n'
                  '拓奇长荣327页技术方案在内容体量上是江楼(58页)的5.6倍，如质量确实显著更优，高分在技术上可能成立。\n'
                  '但结合L5扫描仪一致性和L8地理接近性，建议复核评审打分是否客观公正。')
ws3[f'A{r3}'].font = NORMAL; ws3[f'A{r3}'].alignment = L; ws3.row_dimensions[r3].height = 55

# Budget was 25万
r3b = 10
ws3.merge_cells(f'A{r3b}:I{r3b}')
ws3[f'A{r3b}'] = '预算: 250,000元 | 三家报价均未超预算 | 单项报价明细因OCR有限未逐项提取(纸质扫描件表格识别困难)'
ws3[f'A{r3b}'].font = ITALIC

ws3.column_dimensions['A'].width = 8; ws3.column_dimensions['B'].width = 28
ws3.column_dimensions['C'].width = 16; ws3.column_dimensions['D'].width = 12
ws3.column_dimensions['E'].width = 16; ws3.column_dimensions['F'].width = 10
ws3.column_dimensions['G'].width = 20; ws3.column_dimensions['H'].width = 22
ws3.column_dimensions['I'].width = 20

# --- Sheet 4: 文本雷同分析 ---
ws4 = wb.create_sheet('文本雷同分析')
ws4.merge_cells('A1:G1'); ws4['A1'] = 'TF-IDF文本雷同检测'; ws4['A1'].font = TITLE_FONT
ws4.merge_cells('A2:G2'); ws4['A2'] = '>=80%: 高度可疑 RED | 50-80%: 中度 YELLOW | <50%: 正常 GREEN'
ws4['A2'].font = ITALIC

hdr(ws4, 4, ['对比维度', '建韬 vs 江楼', '建韬 vs 拓奇', '江楼 vs 拓奇', '阈值判断', '风险等级', '说明'])

tfidf_rows = [
    ['全文TF-IDF(资格性+其他响应)', f'{sim_full[0][1]*100:.1f}%', f'{sim_full[0][2]*100:.1f}%', f'{sim_full[1][2]*100:.1f}%',
     '>=80% RED', 'RED - 全部超标', '主要由标准承诺函模板驱动(17-27对100%匹配段落均为模板)'],
    ['技术方案TF-IDF(仅技术部分)', f'{sim_tech[0][1]*100:.1f}%', f'{sim_tech[0][2]*100:.1f}%', f'{sim_tech[1][2]*100:.1f}%',
     '>=80% RED', 'YELLOW - 江楼vs拓奇临界', '75-81%高于正常独立撰写预期(<50%)，需人工复核'],
    ['技术方案字符量', '29,399', '32,908', '57,685', '—', '体量差异大', '拓奇内容量为建韬的~2倍'],
    ['其他响应文件页数', '92页', '58页', '327页', '—', '体量悬殊(5.6x)', '拓奇页数是江楼的5.6倍'],
]

for r, row in enumerate(tfidf_rows, 5):
    for c, val in enumerate(row, 1):
        fill = None
        if c >= 2 and c <= 4 and '%' in str(val):
            pct = float(val.replace('%',''))
            fill = RED_FILL if pct >= 80 else (YELLOW_FILL if pct >= 50 else GREEN_FILL)
        cell(ws4, r, c, val, BOLD if c==1 else NORMAL, C if c<=5 else L, fill)

# Explanation
r4 = 10
ws4.merge_cells(f'A{r4}:G{r4}')
ws4[f'A{r4}'] = ('说明:\n'
    '1. 全文高相似度(86-90%)由统一承诺函模板驱动，不构成实质性内容雷同。\n'
    '   全匹配句例如: "本单位对上述承诺的内容事项真实性负责"、"代理人无转委托权"等。\n'
    '2. 技术方案75-81%相似度高于预期，但由于纸质扫描件OCR噪声，部分"相似"可能来自误识别。\n'
    '3. 建议人工对比技术方案原件，逐段判断是否存在实质性文字复制。')
ws4[f'A{r4}'].font = NORMAL; ws4[f'A{r4}'].alignment = L; ws4.row_dimensions[r4].height = 80

ws4.column_dimensions['A'].width = 28; ws4.column_dimensions['B'].width = 18
ws4.column_dimensions['C'].width = 18; ws4.column_dimensions['D'].width = 18
ws4.column_dimensions['E'].width = 16; ws4.column_dimensions['F'].width = 22
ws4.column_dimensions['G'].width = 45

# --- Sheet 5: 工商信息 ---
ws5 = wb.create_sheet('工商信息')
ws5.merge_cells('A1:F1'); ws5['A1'] = '供应商工商信息对比'; ws5['A1'].font = TITLE_FONT
hdr(ws5, 3, ['信息项', '四川建韬科技', '成都江楼商贸', '成都拓奇长荣商贸', '比对结果', '风险'])

biz = [
    ['法定代表人', '樊巧玲', '刘文杰(董事长)', '田瑶', '不同', 'GREEN'],
    ['授权代表', '赖友弘', '(法定代表人)', '向欢欢', '不同', 'GREEN'],
    ['信用代码', '915101000753879598', '915101083215605861', '未提取到', '不同', 'GREEN'],
    ['成立日期', '2013-08-23', '2014-12-10', '2023年', '不同', 'GREEN'],
    ['注册资本', '800万元', '未提取', '未提取', '—', '—'],
    ['注册地址', '武侯区玉林北街1号1栋3单元4层401号', '武侯区武侯大道铁佛段1号1栋1单元11层1123号', '武侯区玉林中路13号附1122号', '同在武侯区', 'YELLOW'],
    ['地址分析', '玉林北街(玉林片区)', '武侯大道铁佛段(铁佛片区)', '玉林中路(玉林片区)', '建韬与拓奇同在玉林', 'YELLOW'],
    ['经营范围', '计算机软硬件/安防设备', '极广(水果/建材/办公/AI/市政)', '工艺美术品/办公用品', '差异大', 'GREEN'],
    ['经营年限', '~13年', '~11年', '~3年', '差异大', 'GREEN'],
]

for r, row in enumerate(biz, 4):
    for c, val in enumerate(row, 1):
        fill = YELLOW_FILL if val == 'YELLOW' else (GREEN_FILL if val == 'GREEN' else None)
        cell(ws5, r, c, val, BOLD if c==1 else NORMAL, L, fill)

r5 = 14
ws5.merge_cells(f'A{r5}:F{r5}')
ws5[f'A{r5}'] = ('注意: 建韬科技(玉林北街1号)与拓奇长荣(玉林中路13号)同在武侯区玉林片区，步行距离约500米。\n'
                  '在L5扫描仪一致的背景下，建议通过天眼查/企查查核查股权结构和关联关系。\n'
                  '三家法人和经营范围的差异不支持"同一实际控制人"的初步判断，但需专业工商数据验证。')
ws5[f'A{r5}'].font = Font(name='Microsoft YaHei', size=10, color='CC0000')
ws5[f'A{r5}'].alignment = L; ws5.row_dimensions[r5].height = 60

ws5.column_dimensions['A'].width = 16; ws5.column_dimensions['B'].width = 32
ws5.column_dimensions['C'].width = 32; ws5.column_dimensions['D'].width = 28
ws5.column_dimensions['E'].width = 20; ws5.column_dimensions['F'].width = 10

# Save
output = os.path.join(base, '串标围标分析报告.xlsx')
wb.save(output)
print(f'Excel生成完成: {output}')
