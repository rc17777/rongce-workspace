# -*- coding: utf-8 -*-
"""Build expanded audit Excel with national standard framework + system-implementation gap analysis"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ===== Styles =====
header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
sub_header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
section_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
title_font = Font(name='微软雅黑', bold=True, size=14, color='1F4E79')
section_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
normal_font = Font(name='微软雅黑', size=10)
bold_font = Font(name='微软雅黑', bold=True, size=10)
red_font = Font(name='微软雅黑', bold=True, size=10, color='9C0006')
green_font = Font(name='微软雅黑', bold=True, size=10, color='006100')
border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                top=Side(style='thin'), bottom=Side(style='thin'))
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header(ws, row, cols, fill=header_fill, font=header_font):
    for c in range(1, cols+1):
        cell = ws.cell(row, c)
        cell.fill = fill
        cell.font = font
        cell.alignment = center_align
        cell.border = border

def style_row(ws, row, cols, risk=None, font_override=None):
    for c in range(1, cols+1):
        cell = ws.cell(row, c)
        cell.font = font_override if font_override else normal_font
        cell.alignment = left_align if c > 2 else center_align
        cell.border = border
    if risk == 'red':
        for c in range(1, cols+1):
            ws.cell(row, c).fill = red_fill
    elif risk == 'yellow':
        for c in range(1, cols+1):
            ws.cell(row, c).fill = yellow_fill
    elif risk == 'green':
        for c in range(1, cols+1):
            ws.cell(row, c).fill = green_fill

# ==========================================
# Sheet 0: 考核指标框架
# ==========================================
ws0 = wb.active
ws0.title = '考核指标框架'
ws0.merge_cells('A1:H1')
ws0['A1'] = '高校中层领导干部经济责任审计 — 考核指标框架（依据：中办发〔2019〕45号）'
ws0['A1'].font = title_font
ws0['A1'].alignment = center_align

ws0.cell(3, 1, '法规依据').font = bold_font
ws0.merge_cells('A3:B3')
laws = [
    '1. 《党政主要领导干部和国有企事业单位主要领导人员经济责任审计规定》（中办发〔2019〕45号）',
    '2. 《审计署关于内部审计工作的规定》（审计署令第11号，2018）',
    '3. 《教育部关于加强直属高校内部审计工作的意见》（教财〔2015〕46号）',
    '4. 《教育系统内部审计工作规定》（教育部令第47号，2020）',
    '5. 《行政事业性国有资产管理条例》（国务院令第738号，2021）',
    '6. 《四川省省属高等学校国有资产管理办法》（川教〔2024〕112号）',
    '7. 《中华人民共和国政府采购法》及其实施条例',
]
for i, law in enumerate(laws, 4):
    ws0.merge_cells(f'A{i}:H{i}')
    ws0.cell(i, 1, law).font = normal_font

# Evaluation dimensions
headers0 = ['序号', '审计维度\n（中办发〔2019〕45号）', '高校国资处适配', '权重建议', '本审计覆盖', '评分\n(1-10)', '关键证据来源', '评价等级']
for j, h in enumerate(headers0, 1):
    ws0.cell(12, j, h)
style_header(ws0, 12, len(headers0))

dimensions = [
    [1, '贯彻执行重大经济方针政策\n和决策部署情况', '国家/省/学院资产管理、采购管理\n重大政策落实情况', '10%', '✅ 全覆盖', 8, '述职报告/制度文件/会议纪要\n政策学习记录', '🟢 良好'],
    [2, '本部门重要发展规划和政策\n措施的制定、执行和效果', '国资处"十四五"规划执行\n制度建设规划与落地', '10%', '✅ 全覆盖', 7, '制度文件30份/五年述职', '🟡 基本达标'],
    [3, '重大经济事项的决策、执行\n和效果情况', '三重一大决策、大额采购审批\n资产处置决策', '20%', '✅ 全覆盖', 5, '采购台账/会议纪要/党委会纪要\n报废处置请示批复', '🔴 偏差明显'],
    [4, '财政财务管理和经济风险\n防范情况', '预算执行/经费支出规范\n往来资金/专项资金管理', '10%', '✅ 覆盖', 7, '支出问题#14-16\n财务制度11份', '🟡 基本合规'],
    [5, '国有资产管理情况', '资产配置/使用/处置/收益\n闲置资产/无形资产/在建工程', '20%', '✅ 全覆盖', 5, '资产台账/盘点报告/报废文件\n简阳房产/无形资产报废', '🔴 偏差明显'],
    [6, '政府采购管理情况', '采购程序合规/效率/效益\n招标文件/评标/归档', '15%', '✅ 全覆盖', 5, '采购台账/问题#1-9\n采购制度5份', '🔴 问题集中'],
    [7, '在经济活动中落实党风廉政\n建设和廉洁从业规定', '廉洁从业/一岗双责\n廉政风险防控/作风建设', '10%', '✅ 全覆盖', 8, '述责述廉报告/承诺书\n谈话记录/培训记录', '🟢 良好'],
    [8, '以往审计发现问题的整改', '2022年经责审计整改闭环\n本审计前自查自纠', '5%', '⚠️ 部分', 6, '2022年审计底稿\n后续整改材料', '🟡 需追踪'],
]
for i, row in enumerate(dimensions, 13):
    for j, val in enumerate(row):
        ws0.cell(i, j+1, val)
    risk = 'red' if '🔴' in str(row[7]) else ('yellow' if '🟡' in str(row[7]) else 'green')
    style_row(ws0, i, len(headers0), risk)

# Weighted score
total_row = 21
ws0.cell(total_row, 1, '').font = bold_font
ws0.cell(total_row, 2, '加权综合得分').font = bold_font
ws0.cell(total_row, 4, '100%').font = bold_font
ws0.cell(total_row, 6, '6.15/10').font = red_font
for c in range(1,9):
    ws0.cell(total_row, c).border = border

# Score interpretation
ws0.cell(23, 1, '评分标准：').font = bold_font
score_guide = [
    '9-10分（🟢优秀）= 制度健全且执行到位，无重大偏差',
    '7-8分（🟢良好）= 制度基本健全，执行偶有偏差但无实质影响',
    '5-6分（🟡基本达标）= 制度有缺失或执行存在明显偏差',
    '3-4分（🟠不足）= 制度严重缺失或执行存在重大漏洞',
    '1-2分（🔴严重）= 存在系统性风险或重大违规',
]
for i, s in enumerate(score_guide, 24):
    ws0.merge_cells(f'A{i}:H{i}')
    ws0.cell(i, 1, s).font = normal_font

ws0.column_dimensions['A'].width = 6
ws0.column_dimensions['B'].width = 28
ws0.column_dimensions['C'].width = 28
ws0.column_dimensions['D'].width = 10
ws0.column_dimensions['E'].width = 12
ws0.column_dimensions['F'].width = 10
ws0.column_dimensions['G'].width = 32
ws0.column_dimensions['H'].width = 14

# ==========================================
# Sheet 1: 制度与执行对照分析（核心）
# ==========================================
ws1 = wb.create_sheet('制度执行对照')
ws1.merge_cells('A1:J1')
ws1['A1'] = '国资处制度规定与实际执行——逐项对照分析'
ws1['A1'].font = title_font
ws1['A1'].alignment = center_align

headers1 = ['维度', '子项', '制度规定（应然）', '制度文号', '实际执行（实然）', '证据', '偏差程度', '偏差分析', '风险等级', '改进建议']
for j, h in enumerate(headers1, 1):
    ws1.cell(3, j, h)
style_header(ws1, 3, len(headers1))

gap_analysis = [
    # === 维度1: 国有资产管理 ===
    ['一、国有资产管理', '资产配置', '科学论证、预算控制、\n按标准配置', '国资办法(2025)', '配置总体规范\n但199项采购中是否有超标准配置待核实', '采购台账', '🟡 轻微偏差', '2025年采购量大增(1.28亿)\n配置论证压力加大', '🟡', '完善配置论证模板\n建立超标准预警机制'],
    ['', '资产使用管理', '三级管理体系\n归口管理、责任到人\n定期清查盘点', '国资办法(2025)\n第七条', '资产台账存在(15.6MB)\n2021-2024年每年盘点\n但清查发现的问题闭环待查', '资产台账\n盘点总结', '🟡 轻微偏差', '制度框架完整\n但问题发现→整改→销号\n闭环链条不清晰', '🟡', '建立资产问题整改\n跟踪台账'],
    ['', '闲置资产处置', '及时处置低效运转、\n长期闲置资产', '省属高校国资办法\n(2024)第37条', '简阳12处房屋+7门面+1门卫室\n长期闲置(估值855万)\n仅拍卖尝试，两次流拍后无下文', '拍卖记录\n问题#13', '🔴 严重偏差', '制度要求"及时处置"\n实际闲置数年未有效利用\n违反了国资办法核心条款', '🔴', '立即启动闲置资产\n分类处置方案\n报学院办公会决议'],
    ['', '资产报废处置', '按规定权限报批\n金额一致、流程合规', '国资办法(2022)\n第七条', '2023年报废请示71万\n→机管局同意54.7万\n→实际处置仍按71万\n金额不一致,原因不明', '问题#11', '🔴 严重偏差', '请示→批复→执行三者\n金额不一致\n涉嫌擅自扩大处置范围', '🔴', '核实金额差异原因\n如属程序错误需追责'],
    ['', '无形资产管理', '及时登记、评估、\n报废处置', '国资办法(2018/2022)', '2023年党委会决定报废\n→2026年3月才请示\n间隔超2年\n2021-2023年盘点已发现应报废', '问题#12\n党委会纪要', '🔴 严重偏差', '从发现问题到启动处置\n延误2-3年\n报废处置链条断裂', '🔴', '建立资产处置时效\n督办机制\n≤6个月完成流程'],
    ['', '防疫物资管理', '建立库房管理制度\n出入库手续完备', '低耗品办法(2017)\n第五条', '2021年防疫物资出库\n仅有财务处盖章说明\n无国资处出库手续', '问题#17', '🟡 轻微偏差', '个别物资出库流程\n不完整\n可能为疫情防控特殊时期', '🟡', '补齐出库手续\n完善应急物资管理制度'],

    # === 维度2: 政府采购管理 ===
    ['二、政府采购管理', '招标文件编制', '国有资产处负责编制\n组织采购文件确认', '采购办法(2025)\n第七条', '物管项目招标文件瑕疵\n→两次流标(650万项目)\n2023年云桌面中小企业政策\n评标现场临时修改', '问题#1\n问题#7', '🔴 严重偏差', '招标文件质量不稳定\n核心条款在评标现场修改\n削弱采购公信力', '🔴', '建立招标文件\n多级审核制度\n引入法务复核'],
    ['', '采购人代表管理', '按制度选派监督代表\n履行现场监督职责', '采购人代表办法\n(2025)', '2025年家具采购\n监督代表肖梁颖未见签到\n或监督报告签名', '问题#2', '🔴 严重偏差', '监督代表是否实际到场\n成疑\n若监督缺位将影响公正性', '🔴', '核实监督代表实际\n履职情况\n如未履职需追责'],
    ['', '采购评审与质疑', '严格评审、\n配合质疑和投诉处理', '内控办法(2021)\n第六条', '2023年投影仪采购\n4个质疑中3个成立\n→合格供应商不足3家\n→重新采购', '问题#8', '🔴 严重偏差', '1次采购4个质疑3个成立\n质疑成立率75%\n反映评审质量堪忧', '🔴', '分析质疑成立原因\n强化评审专家管理\n建立质疑预防机制'],
    ['', '采购文档归档', '负责采购文档整理\n归档立卷', '采购办法(2025)\n第七条', '2025年电力改造缺评标报告\n2022年医学设备缺磋商报告\n归档不完整', '问题#3\n问题#9', '🔴 严重偏差', '2个年度2个项目\n均缺少核心评标文档\n不是偶发,是系统性问题', '🔴', '建立归档清单\n采购完成后3个工作日内\n归档完毕并交叉检查'],

    # === 维度3: 合同管理 ===
    ['三、合同管理', '履约保证金管理', '签订合同前收取\n履约保证金或保函', '合同条款(图书\n采购)', '2024年图书采购\n未收7.5万履约保证金\n即签订合同', '问题#5', '🔴 严重偏差', '合同条款明文规定\n实际未执行\n形成合同风险敞口', '🔴', '建立签合同前置审查\n缺失履约保证金\n冻结合同签章流程'],
    ['', '合同签订时序', '采购完成后签订合同', '采购办法(2023)\n第31条', '2024年公务车采购5.23签合同\n采购确认5.27\n采购未完成即签合同', '问题#6', '🔴 严重偏差', '合同签订在前\n采购确认在后\n程序倒置,合同效力存疑', '🔴', '严格执行先确认后\n签合同流程\n系统控制时间顺序'],
    ['', '履约考核', '按合同考核标准\n逐项评分、依据充分', '维护服务合同\n第六条', '耗材维护考核表打分\n与标准不一致\n扣分原因跨指标乱填', '问题#4', '🟡 轻微偏差', '考核流于形式\n打分随意\n建议系统化考核模板', '🟡', '建立标准化考核表\n锁定指标计算公式\n审核人二次复核'],

    # === 维度4: 支出管理 ===
    ['四、支出管理', '报销依据完整性', '原始凭证完整\n审批手续齐备', '经费支出办法\n(2022)第10条', '客座教授协议缺失\n专家劳务费无计算过程\n发票出具方与培训方不一致', '问题#14-16', '🟡 轻微偏差', '3笔报销存在依据缺失\n涉及金额不大(1.28万)\n但反映审核不严', '🟡', '强化报账初审\n建立附件清单制度\n缺失依据不予受理'],

    # === 维度5: 内部控制 ===
    ['五、内部控制', '内控制度建设', '建立健全内控体系\n编制内控手册', '内控规范(2021)\n内控办法(2021)', '2022年7月启动内控手册编制\n制度建设30项\n但三项核心制度至今缺失', '会议纪要\n制度清单', '🟡 轻微偏差', '内控框架已建立\n但资产处置/出租/盘点\n三项操作细则缺位', '🟡', '2026年内补齐三项\n操作细则'],
    ['', '不相容岗位分离', '采购/验收/登记\n不相容岗位分离', '内控办法(2021)', '2023年已实施定责定岗\n采购/验收/资产登记分离\n分工明确', '人员分工文件\n述职报告', '🟢 基本一致', '岗位分离执行到位\n制度与实际一致', '🟢', '继续保持\n定期轮岗'], 

    # === 维度6: 安全管理 ===
    ['六、安全管理', '消防安全管理', '部门负责人为消防\n安全第一责任人', '消防安全办法\n第六条', '2024年初发生消防事故\n2021-2022纸质档案全部损毁\n2023年部分损毁', '问题#19\n档案损毁说明', '🔴 严重偏差', '档案管理安全管理\n存在重大漏洞\n损失不可逆', '🔴', '档案数字化备份\n库房消防改造\n定期安全检查制度化'],

    # === 维度7: 任期目标 ===
    ['七、任期目标', '重点工作完成', '年度重点工作计划\n全部完成', '2025年重点工作\n计划', '"采购大数据分析"任务\n2025年计划列出\n2025年总结未见执行', '问题#18', '🟡 轻微偏差', '个别创新性工作任务\n未按计划推进', '🟡', '建立季度重点工作\n督办机制'],
]

for i, row in enumerate(gap_analysis, 4):
    for j, val in enumerate(row):
        ws1.cell(i, j+1, val)
    risk = 'red' if '🔴' in str(row[8]) else ('yellow' if '🟡' in str(row[8]) else 'green')
    style_row(ws1, i, len(headers1), risk)

# Dimensions merge
# Find where each dimension starts and merge
for dim_num, dim_name in enumerate(['一、国有资产管理', '二、政府采购管理', '三、合同管理', '四、支出管理', '五、内部控制', '六、安全管理', '七、任期目标']):
    first = None
    for r in range(4, 4+len(gap_analysis)):
        if ws1.cell(r, 1).value == dim_name:
            first = r
            break
    if first:
        last = first
        for r in range(first+1, 4+len(gap_analysis)):
            val = ws1.cell(r, 1).value
            if val and val != '':
                break
            last = r
        if last > first:
            ws1.merge_cells(f'A{first}:A{last}')
            ws1.cell(first, 1).font = bold_font

# Summary row
summary_row = 4 + len(gap_analysis) + 1
ws1.cell(summary_row, 2, '合计').font = bold_font
red_count = sum(1 for row in gap_analysis if '🔴' in str(row[8]))
yellow_count = sum(1 for row in gap_analysis if '🟡' in str(row[8]))
green_count = sum(1 for row in gap_analysis if '🟢' in str(row[8]))
ws1.cell(summary_row, 7, f'🔴{red_count}项 🟡{yellow_count}项 🟢{green_count}项').font = bold_font
for c in range(1, 11):
    ws1.cell(summary_row, c).border = border

ws1.column_dimensions['A'].width = 16
ws1.column_dimensions['B'].width = 16
ws1.column_dimensions['C'].width = 28
ws1.column_dimensions['D'].width = 18
ws1.column_dimensions['E'].width = 30
ws1.column_dimensions['F'].width = 18
ws1.column_dimensions['G'].width = 14
ws1.column_dimensions['H'].width = 28
ws1.column_dimensions['I'].width = 10
ws1.column_dimensions['J'].width = 28

# ==========================================
# Sheet 2: 问题清单汇总
# ==========================================
ws2 = wb.create_sheet('问题清单汇总')
ws2.merge_cells('A1:I1')
ws2['A1'] = '经济责任审计问题清单（依据中办发〔2019〕45号分类）'
ws2['A1'].font = title_font
ws2['A1'].alignment = center_align

headers2 = ['序号', '审计维度', '问题类别', '问题定性', '问题描述', '涉及金额(元)', '违反法规简称', '责任认定', '整改要求\n立行立改/限期整改']
for j, h in enumerate(headers2, 1):
    ws2.cell(3, j, h)
style_header(ws2, 3, len(headers2))

problems = [
    [1, '政府采购管理', '采购管理', '招标文件不严谨致两次流标', '物管服务项目(650万)两次招标均因招标文件瑕疵流标', 6500000, '采购办法(2025)\n第七条', '领导责任', '立行立改'],
    [2, '政府采购管理', '采购管理', '采购人监督代表未签到', '家具采购监督代表肖梁颖未见签到或监督报告签名', 1116800, '采购办法(2025)\n第七条', '领导责任', '立行立改'],
    [3, '政府采购管理', '采购管理', '采购文档归档不齐', '电力改造项目缺少评标报告/审查表/评分表', 1806419, '采购办法(2025)\n第七条', '领导责任', '限期整改'],
    [4, '政府采购管理', '采购管理', '评审不严致重新采购', '投影仪采购4个质疑3个成立→重新采购', 889500, '内控办法(2021)\n第六条', '领导责任', '立行立改'],
    [5, '政府采购管理', '采购管理', '招标文件审核不严', '云桌面项目中小企业政策现场修改\n演示分值设10分4家均0分', 2290000, '采购办法(2023)/内控办法(2021)', '领导责任', '立行立改'],
    [6, '政府采购管理', '采购管理', '采购文档归档不齐', '医学设备采购缺少磋商报告/评审表等', 1725000, '内控办法(2021)\n第六条', '领导责任', '限期整改'],
    [7, '政府采购管理', '合同管理', '合同签订晚于执行时间', '绿化工程8.11签合同约定7.26开工', 1100000, '内控规范(2021)\n第五条', '领导责任', '立行立改'],
    [8, '政府采购管理', '合同管理', '采购未完成即签合同', '公务车采购确认5.27,合同签5.23', 360000, '采购办法(2023)\n第31条', '领导责任', '立行立改'],
    [9, '合同管理', '合同管理', '履约考核不到位', '耗材维护考核打分与标准不一致', 19830, '维护服务合同\n第六条', '领导责任', '立行立改'],
    [10, '合同管理', '合同管理', '未按合同收取履约保证金', '图书采购未收7.5万保函即签合同', 75000, '图书采购合同\n第五条', '领导责任', '立行立改'],
    [11, '国有资产管理', '资产管理', '报废处置金额不一致', '请示71.1万→批复54.7万→实际仍71.1万', 711388, '国资办法(2022)\n第七条', '领导责任', '立行立改'],
    [12, '国有资产管理', '资产管理', '无形资产报废处置不及时', '2023年决定报废→2026年才请示(间隔2年+)', 845900, '国资办法(2018)\n第九条', '领导责任', '立行立改'],
    [13, '国有资产管理', '资产管理', '闲置资产长期未处置', '简阳12处房屋+7门面+1门卫室闲置(855万)', 8551533, '省属高校办法(2024)/国资办法(2022)', '领导责任', '限期整改'],
    [14, '国有资产管理', '资产管理', '防疫物资出库手续不全', '出库仅有财务处说明,无国资处手续', 10067, '低耗品办法(2017)\n第五条', '领导责任', '限期整改'],
    [15, '财政财务管理', '支出管理', '报销依据不齐备', '客座教授底薪报销无客座教授协议', 5952, '会计基础规范/经费支出办法', '领导责任', '限期整改'],
    [16, '财政财务管理', '支出管理', '原始凭证不规范', '培训费发票方≠通知培训方,无说明', 4400, '《会计法》\n第十四条', '领导责任', '立行立改'],
    [17, '财政财务管理', '支出管理', '费用报销无计算过程', '专家劳务费论证时间2.42h,签到专家不一致', 2400, '省卫计委专家\n劳务费标准', '领导责任', '限期整改'],
    [18, '任期目标', '任期目标', '未完成大数据分析任务', '2025年计划列"采购大数据分析",总结未见执行', 0, '', '领导责任', '立行立改'],
    [19, '安全管理', '安全管理', '消防事故致档案损毁', '2024年消防事故,2021-2022纸质档案全部损毁', 0, '消防安全办法\n第六条', '领导责任', '立行立改'],
]

for i, row in enumerate(problems, 4):
    for j, val in enumerate(row):
        ws2.cell(i, j+1, val)
    risk = 'red'
    style_row(ws2, i, len(headers2), risk)

# Summary
r = 4 + len(problems)
total_amt = sum(p[5] for p in problems)
ws2.cell(r, 1, '').font = bold_font
ws2.cell(r, 2, '合计19项').font = bold_font
ws2.cell(r, 6, total_amt).font = bold_font
ws2.cell(r, 6).number_format = '#,##0.00'
for c in range(1, 10):
    ws2.cell(r, c).border = border

ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 14
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 28
ws2.column_dimensions['E'].width = 45
ws2.column_dimensions['F'].width = 14
ws2.column_dimensions['G'].width = 24
ws2.column_dimensions['H'].width = 10
ws2.column_dimensions['I'].width = 12

# ==========================================
# Sheet 3: 履职评分矩阵
# ==========================================
ws3 = wb.create_sheet('履职评分矩阵')
ws3.merge_cells('A1:N1')
ws3['A1'] = '李欣任中经济责任审计 — 多维度履职评分矩阵'
ws3['A1'].font = title_font
ws3['A1'].alignment = center_align

headers3 = ['审计维度\n(中办发45号)', '权重', '评价指标', '标准分', '制度分\n(1-10)', '执行分\n(1-10)', '加权分', '偏差项\n数量', '🔴','🟡','🟢', '综合等级', '主要发现', '建议']
for j, h in enumerate(headers3, 1):
    ws3.cell(3, j, h)
style_header(ws3, 3, len(headers3))

score_matrix = [
    ['一、重大政策\n贯彻执行', '10%', '政策学习/传达/落实\n制度对标上级法规', 100, 9, 7, 8.0, 0, 0, 0, 0, '🟢', '政策学习记录完整\n制度对标上位法规\n29号文落实到位', '保持政策敏感性\n及时更新对标'],
    ['二、本部门\n发展规划', '10%', '十四五规划执行\n制度建设规划落地', 100, 8, 6, 7.0, 1, 0, 0, 1, '🟢', '30项制度体系完整\n但核心操作细则\n（3项）长期缺失', '补齐三项空白制度\n明确时间节点'],
    ['三、重大经济\n事项决策', '20%', '三重一大执行\n大额采购/资产处置', 100, 7, 4, 5.5, 11, 7, 3, 1, '🔴', '报废金额不一致\n闲置资产不处置\n采购程序倒置\n招标文件质量堪忧', '需重点整改\n是所有维度中\n问题最多的一项'],
    ['四、财政财务\n管理', '10%', '预算执行/经费支出\n资金安全管理', 100, 8, 6, 7.0, 3, 0, 3, 0, '🟡', '3笔报销依据缺失\n金额不大(1.28万)\n反映审核不严', '强化报账初审\n标准化附件清单'],
    ['五、国有资产\n管理', '20%', '配置/使用/处置\n闲置/无形/在建', 100, 7, 3, 5.0, 4, 3, 1, 0, '🔴', '闲置855万数年不动\n无形资产报废延误2年+\n报废处置金额不一致', '立即分类处置闲置\n建立时效督办机制'],
    ['六、政府采购\n管理', '15%', '程序合规/效率/效益\n归档/监督/质疑', 100, 7, 3, 5.0, 8, 6, 1, 1, '🔴', '问题数量最多(8个)\n涉及金额最大(1,917万)\n招标/评审/归档全面薄弱', '系统整改采购流程\n建立多级审核机制'],
    ['七、党风廉政\n建设', '10%', '一岗双责/廉洁从业\n风险防控/教育', 100, 9, 8, 8.5, 0, 0, 0, 0, '🟢', '防控逐年升级\n2025年零廉政事件\n谈话11次/培训26次', '保持不放松\n突破"两张皮"'],
    ['八、以往审计\n整改', '5%', '2022年经责审计整改\n举一反三', 100, 6, 5, 5.5, 1, 0, 1, 0, '🟡', '2022年整改结果\n本次审计中需交叉\n验证闭环情况', '建立整改台账\n逐项销号确认'],
    ['九、安全管理\n（加分项）', '—', '消防安全/档案安全\n资产安全', 100, 5, 2, 3.5, 1, 1, 0, 0, '🔴', '消防事故不可逆损失\n档案灭失', '数字化备份/消防改造\n定期安全巡检'],
]

for i, row in enumerate(score_matrix, 4):
    for j, val in enumerate(row):
        ws3.cell(i, j+1, val)
    risk = 'red' if '🔴' in str(row[11]) else ('yellow' if '🟡' in str(row[11]) else 'green')
    style_row(ws3, i, len(headers3), risk)

# Weighted average
r = 4 + len(score_matrix)
ws3.cell(r, 2, '加权综合').font = bold_font
ws3.cell(r, 3, '(不含安全加分项)').font = bold_font
# Calculate: (8.0*0.1 + 7.0*0.1 + 5.5*0.2 + 7.0*0.1 + 5.0*0.2 + 5.0*0.15 + 8.5*0.1 + 5.5*0.05)
weighted = 8.0*0.1 + 7.0*0.1 + 5.5*0.2 + 7.0*0.1 + 5.0*0.2 + 5.0*0.15 + 8.5*0.1 + 5.5*0.05
ws3.cell(r, 7, round(weighted, 2)).font = red_font
ws3.cell(r, 11, '🔴').font = bold_font
for c in range(1, 15):
    ws3.cell(r, c).border = border

ws3.column_dimensions['A'].width = 14
ws3.column_dimensions['B'].width = 8
ws3.column_dimensions['C'].width = 24
ws3.column_dimensions['D'].width = 8
ws3.column_dimensions['E'].width = 8
ws3.column_dimensions['F'].width = 8
ws3.column_dimensions['G'].width = 8
ws3.column_dimensions['H'].width = 8
ws3.column_dimensions['I'].width = 5
ws3.column_dimensions['J'].width = 5
ws3.column_dimensions['K'].width = 5
ws3.column_dimensions['L'].width = 8
ws3.column_dimensions['M'].width = 28
ws3.column_dimensions['N'].width = 22

# ==========================================
# Sheet 4: 改进路线图
# ==========================================
ws4 = wb.create_sheet('改进路线图')
ws4.merge_cells('A1:G1')
ws4['A1'] = '审计整改建议路线图'
ws4['A1'].font = title_font
ws4['A1'].alignment = center_align

headers4 = ['优先级', '整改事项', '责任部门', '整改措施', '整改时限', '验收标准', '对应问题编号']
for j, h in enumerate(headers4, 1):
    ws4.cell(3, j, h)
style_header(ws4, 3, len(headers4))

roadmap = [
    ['🔴 立即\n(1个月内)', '核实报废金额不一致原因', '国资处', '逐笔核对2023年报废请示/批复/处置确认书\n查明金额差异原因\n如有违规追责', '2026年6月底', '出具书面说明\n金额差异逐笔销账', '#11'],
    ['🔴 立即\n(1个月内)', '启动简阳闲置资产处置', '国资处+院办', '分类评估:12处房屋+7门面\n拍卖/调剂/公益使用多路径\n报学院办公会决议', '2026年6月底', '形成处置方案\n提交院长办公会', '#13'],
    ['🔴 立即\n(1个月内)', '补齐缺失归档资料', '国资处', '逐项目核对2022-2025年采购档案\n补全评标报告/审查表/评分表', '2026年6月底', '归档完整率100%', '#3,#9'],
    ['🔴 立即\n(1个月内)', '档案数字化备份', '国资处', '现存纸质档案全部扫描\n建立电子档案服务器\n定期备份制度', '2026年6月底', '电子档案可检索\n异地备份就绪', '#19'],
    ['🔴 尽快\n(3个月内)', '无形资产报废流程推进', '国资处', '跟进省卫健委/机管局审批进度\n完成2026年请示的审批闭环', '2026年8月底', '获得报废批复\n完成处置确认', '#12'],
    ['🟡 近期\n(3个月内)', '建立招标文件多级审核', '国资处', '招标文件→经办人→科长→处长三级审核\n重要项目增加法务复核', '2026年8月底', '2026下半年招标文件\n零现场修改', '#1,#7'],
    ['🟡 近期\n(3个月内)', '合同管理前置审核', '国资处+财务处', '签合同前系统校验:\n采购确认→保证金→签约\n流程卡控,顺序不对无法盖章', '2026年8月底', '合同倒签率归零', '#5,#6'],
    ['🟡 近期\n(3个月内)', '采购评审质量提升', '国资处', '评审专家库更新+培训\n评分表预审制度\n质疑分析月度报告', '2026年8月底', '质疑成立率<20%', '#8'],
    ['🟡 中期\n(6个月内)', '补齐三项制度空白', '国资处', '制定:①资产处置管理办法\n②出租出借管理办法\n③清查盘点实施细则', '2026年11月底', '三项制度正式发文\n培训全覆盖', '制度缺失'],
    ['🟡 中期\n(6个月内)', '强化报账审核', '财务处+国资处', '建立标准化附件清单\n报账材料不齐系统自动退回', '2026年11月底', '报销退回率显著下降', '#14-16'],
    ['🟢 持续\n(年度)', '突破党建"两张皮"', '国资处支部', '党建与业务联合考核\n每月1次党建+业务融合学习', '2026年全年', '2026述职中"两张皮"\n不再出现', '任期评价'],
    ['🟢 持续\n(年度)', '采购大数据分析', '国资处', '利用2025年已积累数据\n建立采购效率/节约率/质疑率\n三维分析看板', '2026年12月', '大数据分析看板上线\n月度更新', '#18'],
]

for i, row in enumerate(roadmap, 4):
    for j, val in enumerate(row):
        ws4.cell(i, j+1, val)
    risk = 'red' if '🔴' in str(row[0]) else ('yellow' if '🟡' in str(row[0]) else 'green')
    style_row(ws4, i, len(headers4), risk)

ws4.column_dimensions['A'].width = 12
ws4.column_dimensions['B'].width = 28
ws4.column_dimensions['C'].width = 16
ws4.column_dimensions['D'].width = 42
ws4.column_dimensions['E'].width = 16
ws4.column_dimensions['F'].width = 26
ws4.column_dimensions['G'].width = 14

# ==========================================
# Sheet 5: 补充资料清单
# ==========================================
ws5 = wb.create_sheet('补充资料清单')
ws5.merge_cells('A1:F1')
ws5['A1'] = '后续补充资料文件清单及审计关注点'
ws5['A1'].font = title_font
ws5['A1'].alignment = center_align

headers5 = ['序号', '文件类别', '文件名称', '格式', '审计关注点', '优先级']
for j, h in enumerate(headers5, 1):
    ws5.cell(3, j, h)
style_header(ws5, 3, len(headers5))

supp = [
    [1, '采购台账', '2021-2025年采购项目工作台账（5个）', 'xlsx/xls', '需汇总各年度实际合同金额与预算对比', '🔴'],
    [2, '资产台账', '资产台账20260525.xlsx + 全院资产.xlsx', 'xlsx', '核对资产实物与账面,关注闲置资产分布', '🔴'],
    [3, '土地房屋', '两校区土地、房屋资产明细表.xlsx', 'xlsx', '简阳闲置房产(12处+7门面+1门卫室)', '🔴'],
    [4, '资产盘点', '2021-2024年资产盘点工作总结（4个）', 'docx', '盘点差异处理、问题持续跟踪闭环', '🟡'],
    [5, '资产报废', '2021-2026年报废处置请示/批复/确认书（20+）', 'pdf/docx', '核对问题#11金额不一致原因逐笔追溯', '🔴'],
    [6, '资产清查', '2024年清查报告+省卫健委通知+无形资产清查', 'pdf/xlsx', '无形资产报废滞后问题+清查问题跟踪', '🟡'],
    [7, '采购执行', '2025年1-6月采购执行通报+排名+明细', 'doc/xls', '2025年采购预算执行进度分析', '🟡'],
    [8, '人员分工', '国资处人员分工定稿.docx', 'docx', '核实监督代表肖梁颖职责定位', '🔴'],
    [9, '简阳房产', '拍卖成交确认书+流拍报告+再拍通知', 'pdf', '闲置资产处置尝试:为何两次流拍?', '🔴'],
    [10, '汽车调剂', '省财政厅同意调剂特种专业技术用车复函', 'pdf', '资产调剂合规性核查', '🟢'],
    [11, '报废图片', '报废资产实物照片（24张）', 'jpg', '核实报废资产实物存在性,防止虚假报废', '🟡'],
    [12, '2025采购通知', '关于2025年采购计划下达及执行通知', 'docx', '采购管理规范性,制度落地情况', '🟢'],
    [13, '消防事故', '消防事故调查报告/纪委处理结论', '—', '事故原因、责任认定、问责结果', '🔴'],
    [14, '2022审计', '2022年学院主要领导经责审计报告', '—', '本次审计发现与上次是否重复?整改闭环?', '🔴'],
]
for i, row in enumerate(supp, 4):
    for j, val in enumerate(row):
        ws5.cell(i, j+1, val)
    risk = 'red' if '🔴' in str(row[5]) else ('yellow' if '🟡' in str(row[5]) else 'green')
    style_row(ws5, i, len(headers5), risk)

ws5.column_dimensions['A'].width = 6
ws5.column_dimensions['B'].width = 14
ws5.column_dimensions['C'].width = 48
ws5.column_dimensions['D'].width = 14
ws5.column_dimensions['E'].width = 42
ws5.column_dimensions['F'].width = 8

# ===== Save =====
out_path = r'D:\openclaw-workspace\projects\护理学院任中经责审计\护理学院任中经责审计_制度执行对照分析.xlsx'
wb.save(out_path)
print(f'\n✅ 报告已保存: {out_path}')
print(f'   共 {len(wb.sheetnames)} 个Sheet:')
for s in wb.sheetnames:
    ws = wb[s]
    print(f'   📊 {s} ({ws.max_row}行 × {ws.max_column}列)')
