# -*- coding: utf-8 -*-
"""Create comprehensive audit analysis Excel report"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = openpyxl.Workbook()

# ===== Style definitions =====
header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
sub_header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
title_font = Font(name='微软雅黑', bold=True, size=14, color='1F4E79')
normal_font = Font(name='微软雅黑', size=10)
bold_font = Font(name='微软雅黑', bold=True, size=10)
red_font = Font(name='微软雅黑', bold=True, size=10, color='9C0006')
border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                top=Side(style='thin'), bottom=Side(style='thin'))
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header(ws, row, cols, fill=header_fill, font=header_font):
    for c in range(1, cols+1):
        cell = ws.cell(row, c)
        cell.fill = fill
        cell.font = font
        cell.alignment = center_align
        cell.border = border

def style_data_row(ws, row, cols, risk=None):
    for c in range(1, cols+1):
        cell = ws.cell(row, c)
        cell.font = normal_font
        cell.alignment = left_align if c > 1 else center_align
        cell.border = border
    if risk == 'red':
        for c in range(1, cols+1):
            ws.cell(row, c).fill = red_fill
            if c == 1:
                ws.cell(row, c).font = red_font
    elif risk == 'yellow':
        for c in range(1, cols+1):
            ws.cell(row, c).fill = yellow_fill
    elif risk == 'green':
        for c in range(1, cols+1):
            ws.cell(row, c).fill = green_fill

# ==========================================
# Sheet 1: 综合总览
# ==========================================
ws1 = wb.active
ws1.title = '综合总览'
ws1.merge_cells('A1:G1')
ws1['A1'] = '四川护理职业学院国资处领导（李欣）任中经济责任审计 — 综合分析报告'
ws1['A1'].font = title_font
ws1['A1'].alignment = center_align

# Basic info
info_data = [
    ['', '审计对象', '李欣', '审计期间', '2021-2025年（任中）', '报告日期', '2026年5月26日'],
    ['', '所在部门', '国有资产处', '职务', '副处长（主持工作）→处长/支部书记', '', ''],
]
for i, row in enumerate(info_data, 3):
    for j, val in enumerate(row):
        ws1.cell(i, j+1, val).font = bold_font if j in [1,3,5] else normal_font

headers1 = ['序号', '分析维度', '核心发现', '风险等级', '问题数量', '涉及金额（万元）', '建议']
for j, h in enumerate(headers1, 1):
    ws1.cell(6, j, h)
style_header(ws1, 6, len(headers1))

summary = [
    [1, '采购管理', '招标文件瑕疵致两次流标、采购人监督缺位、归档资料不齐、评审不严等', '🔴 高', 8, '1,916.77', '强化招标文件审核、完善归档制度'],
    [2, '合同管理', '履约考核不到位、未收取履约保证金、先签合同后完成采购', '🔴 高', 3, '45.48', '严格合同签订前审查流程'],
    [3, '国有资产管理', '资产报废处置金额不一致、无形/闲置资产处置严重滞后', '🔴 高', 4, '1,008.31', '建立资产处置台账和时效督办'],
    [4, '支出管理', '报销依据不齐、原始凭证不规范、专家劳务费无计算过程', '🟡 中', 3, '1.28', '严格报销审核、补充凭证'],
    [5, '安全管理', '消防事故致2021-2022纸质档案全部损毁', '🔴 高', 1, '—', '档案数字化+防火设施改造'],
    [6, '制度建设', '资产处置/出租出借/清查盘点三项制度缺失', '🟡 中', 3, '—', '补齐制度短板'],
    [7, '任职履职', '"重业务轻党建"持续5年未根治、2025年"大数据分析"任务未完成', '🟡 中', 2, '—', '加强党建业务融合考核'],
]
for i, row in enumerate(summary, 7):
    for j, val in enumerate(row):
        ws1.cell(i, j+1, val)
    risk = 'red' if '🔴' in str(row[3]) else 'yellow'
    style_data_row(ws1, i, len(headers1), risk)

# Totals
ws1.cell(14, 1, '').font = bold_font
ws1.cell(14, 2, '合计').font = bold_font
ws1.cell(14, 3, '19个问题（含制度缺失3项）').font = bold_font
ws1.cell(14, 5, '19+3').font = bold_font
ws1.cell(14, 6, '2,971.84').font = bold_font
for c in range(1, 8):
    ws1.cell(14, c).border = border

# Column widths
ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 16
ws1.column_dimensions['C'].width = 50
ws1.column_dimensions['D'].width = 12
ws1.column_dimensions['E'].width = 12
ws1.column_dimensions['F'].width = 18
ws1.column_dimensions['G'].width = 35

# ==========================================
# Sheet 2: 问题清单汇总
# ==========================================
ws2 = wb.create_sheet('问题清单汇总')
ws2.merge_cells('A1:I1')
ws2['A1'] = '经济责任审计问题清单汇总表'
ws2['A1'].font = title_font
ws2['A1'].alignment = center_align

headers2 = ['序号', '凭证号', '问题类别', '问题定性', '问题描述', '涉及金额（元）', '违反条例（简称）', '责任认定', '整改要求']
for j, h in enumerate(headers2, 1):
    ws2.cell(2, j, h)
style_header(ws2, 2, len(headers2))

problems = [
    [1, '', '采购管理', '招标文件编制不严谨，导致两次流标', '2026-2028年两校区物业管理服务采购项目因招标文件瑕疵两次流标', 6500000, '采购管理办法(2025)第七条', '领导责任', '立行立改'],
    [2, '', '采购管理', '采购人监督代表未签到', '2025年家具采购项目监督代表肖梁颖未见签到记录', 1116800, '采购管理办法(2025)第七条', '领导责任', '立行立改'],
    [3, '', '采购管理', '采购文档归档资料不齐备', '2025年电力改造项目缺少评标报告、审查表、评分表等', 1806418.82, '采购管理办法(2025)第七条', '领导责任', '限期整改'],
    [4, '2024.12#560', '合同管理', '履约考核不到位', '耗材维护费考核表打分与标准不一致，备注扣分原因跨指标', 19830, '耗材维护服务合同第六条', '领导责任', '立行立改'],
    [5, '', '合同管理', '未按合同收取履约保证金', '2024年图书采购项目未收7.5万履约保证金即签合同', 75000, '图书采购合同第五条', '领导责任', '立行立改'],
    [6, '', '采购管理', '采购未完成即签合同', '2024年公务车采购确认5.27，合同签5.23', 360000, '采购管理办法(2023)第31条', '领导责任', '立行立改'],
    [7, '', '采购管理', '招标文件审核不严，评标现场澄清', '2023年云桌面项目中小企业政策现场修改，演示分值设置不合理', 2290000, '采购管理办法(2023)+内控办法(2021)', '领导责任', '立行立改'],
    [8, '', '采购管理', '政府采购评审不严导致重新采购', '2023年投影仪采购4个质疑3个成立，合格供应商不足3家', 889500, '政府采购内控办法(2021)', '领导责任', '立行立改'],
    [9, '', '采购管理', '采购文档归档资料不齐备', '2022年德阳校区医学设备采购缺少磋商报告、评审表等', 1725000, '政府采购内控办法(2021)', '领导责任', '限期整改'],
    [10, '', '合同管理', '合同签订及公告晚于执行时间', '2021年德阳绿化工程8.11签合同，约定开工7.26', 1100000, '政府采购内控规范(2021)', '领导责任', '立行立改'],
    [11, '', '国有资产管理', '报废处置请示金额与批复/实际不一致', '2023年报废请示71.1万→机管局同意54.7万→实际处置仍71.1万', 711388.02, '国资管理办法(2022)第七条', '领导责任', '立行立改'],
    [12, '', '国有资产管理', '无形资产报废处置不及时', '2023年党委会决定报废→2026.3才请示，间隔超2年', 845900, '国资管理办法(2018)第九条', '领导责任', '立行立改'],
    [13, '', '国有资产管理', '闲置资产长期未处置', '简阳12处房屋+7门面+1门卫室长期闲置(估值855万)', 8551533.37, '省属高校国资办法(2024)+国资办法(2022)', '领导责任', '限期整改'],
    [14, '2022.7#148', '支出管理', '费用报销依据不齐备', '报销客座教授底薪5952元，无客座教授协议', 5952.38, '会计基础工作规范+经费支出办法', '领导责任', '限期整改'],
    [15, '2022.3#109', '支出管理', '支付费用原始凭证不规范', '培训费发票出具方与培训通知单位不一致，无说明', 4400, '会计法第14条', '领导责任', '立行立改'],
    [16, '2021.12#53', '支出管理', '费用报销无计算过程与依据', '专家劳务费论证时间2.42h，签到专家与发放表不一致', 2400, '省卫计委专家劳务费标准', '领导责任', '限期整改'],
    [17, '2021.8#4', '国有资产管理', '防疫物资出库管理不严', '防疫物资出库仅财务处盖章说明，无国资处出库手续', 10066.5, '低值易耗品暂行管理办法(2017)', '领导责任', '限期整改'],
    [18, '', '任期目标', '未完成"采购大数据分析"任务', '2025年重点工作计划列出但工作总结未见执行', 0, '', '领导责任', '立行立改'],
    [19, '', '安全管理', '消防事故致档案及资产损毁', '2024年初消防事故，2021-2022纸质档案全部损毁', 0, '消防安全管理暂行办法第6条', '领导责任', '立行立改'],
]

for i, row in enumerate(problems, 3):
    for j, val in enumerate(row):
        ws2.cell(i, j+1, val)
    risk = 'red' if i-3 in [0,1,2,3,4,5,6,7,8,9,10,11,12,18] else 'yellow'
    style_data_row(ws2, i, len(headers2), risk)

# Totals
total_amount = sum(p[5] for p in problems)
ws2.cell(22, 5, '合计').font = bold_font
ws2.cell(22, 6, total_amount).font = bold_font
ws2.cell(22, 6).number_format = '#,##0.00'
for c in [5,6]:
    ws2.cell(22, c).border = border

ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 14
ws2.column_dimensions['C'].width = 14
ws2.column_dimensions['D'].width = 30
ws2.column_dimensions['E'].width = 55
ws2.column_dimensions['F'].width = 16
ws2.column_dimensions['G'].width = 30
ws2.column_dimensions['H'].width = 10
ws2.column_dimensions['I'].width = 12

# ==========================================
# Sheet 3: 制度分析
# ==========================================
ws3 = wb.create_sheet('制度分析')
ws3.merge_cells('A1:F1')
ws3['A1'] = '国资处制度建设分析'
ws3['A1'].font = title_font
ws3['A1'].alignment = center_align

# 3a: Overall assessment
headers3a = ['类别', '文件数', '覆盖面', '评价', '关键问题', '风险等级']
for j, h in enumerate(headers3a, 1):
    ws3.cell(3, j, h)
style_header(ws3, 3, len(headers3a))

policy_summary = [
    ['现行制度(2023-2026)', 10, '资产/采购/工程/验收/绩效/低耗/代理', '🟢 框架完整', '2025年集中出台8项，执行落地待观察', '🟡'],
    ['以往制度(2018-2023)', 9, '资产/采购/履约/内控', '🟢 已更新替代', '旧版制度需明确清理废止', '🟢'],
    ['财务制度(2021-2022)', 11, '预算/收费/往来/货币/专项资金/基建等', '🟢 覆盖全面', '部分2021年试行版未更新', '🟡'],
    ['制度空白', 3, '资产处置/出租出借/清查盘点', '🔴 缺失', '三项核心操作制度至今空白', '🔴'],
]
for i, row in enumerate(policy_summary, 4):
    for j, val in enumerate(row):
        ws3.cell(i, j+1, val)
    risk = 'red' if '🔴' in str(row[5]) else ('yellow' if '🟡' in str(row[5]) else 'green')
    style_data_row(ws3, i, len(headers3a), risk)

# 3b: Evolution timeline
ws3.cell(9, 1, '制度演进时间线').font = bold_font
ws3.merge_cells('A9:F9')
timeline = [
    ['2018', '初始建立', '川护职院发〔2018〕183号 采购管理办法(初版)\n川护职院发〔2018〕184号 国有资产管理办法(初版)', ''],
    ['2021', '内控补强', '川护职院发〔2021〕66号 政府采购内控规范\n川护职院发〔2021〕179号 政府采购内控管理办法', '李欣任财务/国资副处长'],
    ['2022', '大规模建制', '10项制度集中出台（财务9项+国资1项核心）\n川护职院发〔2022〕56号 国有资产管理办法', '制度建设最密集年'],
    ['2023', '完善采购', '川护职院发〔2023〕150号 采购管理办法\n川护职院发〔2023〕144号 招标代理及供应商管理', '国资处独立办公'],
    ['2025', '体系升级', '8项制度集中出台：资产/采购/工程/验收/绩效/低耗等\n川护职院发〔2025〕78号 国资管理办法（全生命周期）', '制度体系基本完备'],
    ['2026', '补漏', '川护职院发〔2026〕7号 捐赠资产管理实施细则', ''],
]
headers3b = ['年份', '阶段', '主要制度文件', '与任职对应']
for j, h in enumerate(headers3b, 1):
    ws3.cell(10, j, h)
style_header(ws3, 10, len(headers3b))
for i, row in enumerate(timeline, 11):
    for j, val in enumerate(row):
        ws3.cell(i, j+1, val)
    style_data_row(ws3, i, len(headers3b))

# 3c: Key gaps
ws3.cell(18, 1, '制度空白分析').font = bold_font
ws3.merge_cells('A18:F18')
gaps = [
    ['资产处置管理办法', '🔴', '报废处置流程缺乏细则，导致问题#11（金额不一致）', '紧急立项'],
    ['出租出借管理办法', '🔴', '简阳闲置资产(855万)长期未处置的制度根源', '紧急立项'],
    ['清查盘点实施细则', '🔴', '资产清查发现的问题缺乏制度化的处理流程', '2026年内'],
    ['档案数字化管理', '🟡', '消防事故后未建立电子档案备份制度', '2026年内'],
    ['无形资产管理制度', '🟡', '仅2026年出台捐赠资产管理，软件/专利等未覆盖', '2026年内'],
]
for j, h in enumerate(['缺失制度', '紧急度', '影响分析', '建议'], 1):
    ws3.cell(19, j, h)
style_header(ws3, 19, 4)
for i, row in enumerate(gaps, 20):
    for j, val in enumerate(row):
        ws3.cell(i, j+1, val)
    risk = 'red' if '🔴' in str(row[1]) else 'yellow'
    style_data_row(ws3, i, 4, risk)

ws3.column_dimensions['A'].width = 28
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 55
ws3.column_dimensions['D'].width = 30
ws3.column_dimensions['E'].width = 18
ws3.column_dimensions['F'].width = 12

# ==========================================
# Sheet 4: 任职情况分析
# ==========================================
ws4 = wb.create_sheet('任职情况分析')
ws4.merge_cells('A1:H1')
ws4['A1'] = '李欣任职情况全量分析（2021-2025）'
ws4['A1'].font = title_font
ws4['A1'].alignment = center_align

# 4a: Post evolution
headers4a = ['年份', '职务', '政治面貌', '管辖范围', '制度建设', '采购规模', '重大事件', '年度评价']
for j, h in enumerate(headers4a, 1):
    ws4.cell(3, j, h)
style_header(ws4, 3, len(headers4a))

career = [
    ['2021', '财务处/国资处\n副处长(主持工作)', '预备党员', '财务+国资\n两个处室', '10项制度\n征求意见稿', '约95个项目', '组织全院资产清查\n清理2012年以来农行账户', '🟢 开局有力'],
    ['2022', '财务处/国资处\n副处长(主持工作)', '青年党员', '财务+国资\n两个处室', '10项制度\n正式发布', '约60个项目', '配合主要领导经责审计\n德阳校区房屋招租100.73万', '🟢 制度年'],
    ['2023', '国资处\n(独立办公)', '党员/支部书记', '国资+采购\n独立运作', '流程标准化\n风控措施', '约91个项目', '国资处独立办公\n编制采购全流程图\n编制廉政风控手册', '🟢 独立成军'],
    ['2024', '国有资产处\n负责人/支部书记', '党员/支部书记', '国资+采购\n+党建', '完善采购制度\n风控手册', '约143个项目', '全院采购廉政培训\n与10家代理签廉洁承诺\n⚠️年初消防事故', '🟡 有憾（消防事故）'],
    ['2025', '国资处\n处长', '党员/支部书记', '国资+采购\n+党建', '5项国资制度\n全生命周期', '199项/1.28亿', '德阳宿舍6296万\n实训基地813万\n零廉政事件/零安全事故', '🟢 规模跃升'],
]
for i, row in enumerate(career, 4):
    for j, val in enumerate(row):
        ws4.cell(i, j+1, val)
    risk = 'yellow' if '⚠️' in str(row[7]) else 'green'
    style_data_row(ws4, i, len(headers4a), risk)

# 4b: Problem tracking
ws4.cell(10, 1, '"老问题"跨年度持续跟踪').font = bold_font
ws4.merge_cells('A10:H10')
track = [
    ['"重业务轻党建"', '已承认', '已承认', '未提及\n(已专任国资)', '未提及', '变体回归："党建与业务两张皮"', '🔴 5年未根治'],
    ['业财融合/专业服务', '深度不够', '能力需加强', '未提及', '未提及', '未再出现\n(因不再管财务)', '🟢 自然消失'],
    ['敢抓敢管力度', '离组织要求有差距', '未提及', '未提及', '已解决', '已解决\n(专项培训+谈话机制)', '🟢 持续改善'],
    ['资产管理精细化', '未提及', '未提及', '未提及', '未提及', '"数字化水平待提高"', '🟡 新提出'],
    ['档案管理安全', '未提及', '未提及', '未提及', '未提及\n(消防事故当年)', '未提及\n(可能已改善)', '🔴 重大缺陷'],
]
headers4b = ['问题', '2021', '2022', '2023', '2024', '2025', '趋势判断']
for j, h in enumerate(headers4b, 1):
    ws4.cell(11, j, h)
style_header(ws4, 11, len(headers4b))
for i, row in enumerate(track, 12):
    for j, val in enumerate(row):
        ws4.cell(i, j+1, val)
    risk = 'red' if '🔴' in str(row[6]) else ('yellow' if '🟡' in str(row[6]) else 'green')
    style_data_row(ws4, i, len(headers4b), risk)

# 4c: Comprehensive evaluation
ws4.cell(18, 1, '综合履职评价').font = bold_font
ws4.merge_cells('A18:H18')
eval_data = [
    ['制度建设', '9/10', '2022年+2025年两轮大规模建制，30项制度覆盖', '🟢'],
    ['采购管理', '6/10', '流程规范但执行漏洞多（8个问题），归档不严', '🟡'],
    ['资产管理', '5/10', '闲置资产处置严重滞后，报废流程不规范，制度空白', '🔴'],
    ['廉政防控', '8/10', '逐年升级，2025年零事件，措施到位', '🟢'],
    ['团队建设', '7/10', '分工明确但教育管理方式单一', '🟢'],
    ['党建融合', '5/10', '"两张皮"5年未解，理论转化不足', '🔴'],
    ['安全管理', '3/10', '消防事故致档案灭失，不可逆损失', '🔴'],
    ['综合', '6.1/10', '制度建设优秀，执行管理有待加强，安全/党建是短板', '🟡'],
]
for j, h in enumerate(['评价维度', '评分', '说明', '等级'], 1):
    ws4.cell(19, j, h)
style_header(ws4, 19, 4)
for i, row in enumerate(eval_data, 20):
    for j, val in enumerate(row):
        ws4.cell(i, j+1, val)
    risk = 'red' if '🔴' in str(row[3]) else ('yellow' if '🟡' in str(row[3]) else 'green')
    style_data_row(ws4, i, 4, risk)

ws4.column_dimensions['A'].width = 18
ws4.column_dimensions['B'].width = 18
ws4.column_dimensions['C'].width = 14
ws4.column_dimensions['D'].width = 18
ws4.column_dimensions['E'].width = 24
ws4.column_dimensions['F'].width = 20
ws4.column_dimensions['G'].width = 30
ws4.column_dimensions['H'].width = 14

# ==========================================
# Sheet 5: 采购台账汇总
# ==========================================
ws5 = wb.create_sheet('采购台账汇总')
ws5.merge_cells('A1:G1')
ws5['A1'] = '采购数据年度汇总（2021-2025）'
ws5['A1'].font = title_font
ws5['A1'].alignment = center_align

headers5 = ['年度', '政府采购项目数', '院内采购项目数', '其他项目数', '合计项目数', '年度总预算(万元)', '备注']
for j, h in enumerate(headers5, 1):
    ws5.cell(3, j, h)
style_header(ws5, 3, len(headers5))

procurement = [
    ['2021', 45, 45, '发改委5', '约95', '—', '台账金额单位为万元，总量待汇总'],
    ['2022', 27, 28, '急办2', '约57', '—', '含德阳校区招租100.73万'],
    ['2023', '—', '—', '—', '约91\n(一表合并)', '—', '台账格式变化，政府采购/院内合并统计'],
    ['2024', '—', '—', '—', '约143', '—', '全年合并统计，含非政府采购'],
    ['2025', '—', '—', '—', '199', '12,800', '述职报告披露：管辖采购总预算约1.28亿'],
]
for i, row in enumerate(procurement, 4):
    for j, val in enumerate(row):
        ws5.cell(i, j+1, val)
    style_data_row(ws5, i, len(headers5))

# Notes
ws5.cell(10, 1, '说明：').font = bold_font
ws5.merge_cells('A10:G10')
ws5.cell(11, 1, '1. 2021-2024年采购总金额待从台账逐项汇总计算（述职报告未披露）').font = normal_font
ws5.merge_cells('A11:G11')
ws5.cell(12, 1, '2. 2025年采购金额来源于述职报告"全年规范推进采购项目约199项、总预算约1.28亿元"').font = normal_font
ws5.merge_cells('A12:G12')
ws5.cell(13, 1, '3. 2023年起台账格式由"政府采购/院内采购/其他"三表改为统一汇总表').font = normal_font
ws5.merge_cells('A13:G13')
ws5.cell(14, 1, '4. 🔴 建议：应从台账中按年度汇总实际签订合同金额，与预算对比分析节约率').font = red_font
ws5.merge_cells('A14:G14')

ws5.column_dimensions['A'].width = 10
ws5.column_dimensions['B'].width = 18
ws5.column_dimensions['C'].width = 18
ws5.column_dimensions['D'].width = 14
ws5.column_dimensions['E'].width = 14
ws5.column_dimensions['F'].width = 20
ws5.column_dimensions['G'].width = 45

# ==========================================
# Sheet 6: 补充资料清单
# ==========================================
ws6 = wb.create_sheet('补充资料清单')
ws6.merge_cells('A1:E1')
ws6['A1'] = '后续补充资料文件清单'
ws6['A1'].font = title_font
ws6['A1'].alignment = center_align

headers6 = ['序号', '文件类别', '文件名称', '格式', '审计关注点']
for j, h in enumerate(headers6, 1):
    ws6.cell(3, j, h)
style_header(ws6, 3, len(headers6))

supp_files = [
    [1, '采购台账', '2021-2025年采购项目工作台账（5个文件）', 'xlsx/xls', '需汇总各年度实际采购金额'],
    [2, '资产台账', '资产台账20260525.xlsx (15.6MB)\n全院资产.xlsx (7.6MB)', 'xlsx', '核对资产实物与账面，关注闲置资产'],
    [3, '土地房屋', '两校区土地、房屋资产明细表.xlsx', 'xlsx', '简阳闲置房产(12处+7门面)'],
    [4, '资产盘点', '2021-2024年资产盘点工作总结(4个docx)', 'docx', '盘点差异处理、问题持续跟踪'],
    [5, '资产报废', '2021-2026年报废处置请示/批复/确认书(20+个)', 'pdf/docx', '核对问题#11金额不一致原因'],
    [6, '资产清查', '2024年国资清查报告+省卫健委通知+行政事业性报告+无形资产清查', 'pdf/xlsx', '无形资产报废滞后问题'],
    [7, '采购执行', '2025年1-6月采购执行通报+排名表+明细表', 'doc/xls', '2025年采购预算执行进度'],
    [8, '人员分工', '国资处人员分工定稿.docx', 'docx', '核实采购监督代表肖梁颖具体职责'],
    [9, '简阳房产', '拍卖成交确认书+流拍报告+再拍通知', 'pdf', '闲置资产处置尝试与失败记录'],
    [10, '汽车调剂', '省财政厅关于同意调剂特种专业技术用车的复函', 'pdf', '资产调剂合规性'],
    [11, '报废图片', '报废资产实物照片(24张)', 'jpg', '核实报废资产实物存在性'],
    [12, '2025采购通知', '关于2025年采购计划下达及执行相关事宜的通知', 'docx', '采购管理规范性'],
]
for i, row in enumerate(supp_files, 4):
    for j, val in enumerate(row):
        ws6.cell(i, j+1, val)
    style_data_row(ws6, i, len(headers6))

ws6.column_dimensions['A'].width = 6
ws6.column_dimensions['B'].width = 14
ws6.column_dimensions['C'].width = 50
ws6.column_dimensions['D'].width = 14
ws6.column_dimensions['E'].width = 45

# ===== Save =====
out_path = r'D:\openclaw-workspace\projects\护理学院任中经责审计\护理学院任中经责审计_综合分析报告.xlsx'
wb.save(out_path)
print(f'\n✅ 报告已保存: {out_path}')
print(f'   共 {len(wb.sheetnames)} 个Sheet: {", ".join(wb.sheetnames)}')
