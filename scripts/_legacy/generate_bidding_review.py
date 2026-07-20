#!/usr/bin/env python3
"""招标文件合规分析 - 条款公平性/倾向性/规范性检查"""
import sys, io, os, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def sf(c): return PatternFill(patternType='solid', fgColor=c)
RED_FILL=sf('FFD7D7'); YELLOW_FILL=sf('FFF3CD'); GREEN_FILL=sf('D4EDDA')
HEADER_FILL=sf('1A3A6E'); LIGHT_GRAY=sf('F5F5F5'); LIGHT_BLUE=sf('E8F0FE')
H=Font(name='Microsoft YaHei',size=11,bold=True,color='FFFFFF')
T=Font(name='Microsoft YaHei',size=14,bold=True,color='1A3A6E')
SUB=Font(name='Microsoft YaHei',size=12,bold=True,color='CC0000')
N=Font(name='Microsoft YaHei',size=10); B=Font(name='Microsoft YaHei',size=10,bold=True)
BR=Font(name='Microsoft YaHei',size=10,color='CC0000',bold=True)
I=Font(name='Microsoft YaHei',size=9,italic=True,color='888888')
TH=Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))
C=Alignment(horizontal='center',vertical='center',wrap_text=True)
L=Alignment(horizontal='left',vertical='center',wrap_text=True)

def hdr(ws,row,headers):
    for c,h in enumerate(headers,1):
        cl=ws.cell(row=row,column=c,value=h)
        cl.font=H;cl.fill=HEADER_FILL;cl.alignment=C;cl.border=TH
def cell(ws,r,c,val,font=N,align=C,fill=None):
    cl=ws.cell(row=r,column=c,value=val)
    cl.font=font;cl.alignment=align;cl.border=TH
    if fill:cl.fill=fill

wb = Workbook()

# ===== Sheet 1: 招标文件综合合规评审 =====
ws = wb.active; ws.title='综合合规评审'
ws.merge_cells('A1:G1'); ws['A1']='招标文件合规性分析 - 医工教研室实训设备采购项目'; ws['A1'].font=T
ws.merge_cells('A2:G2'); ws['A2']=f'项目编号: HTGJ-CS(2025)-97号 | 分析: {datetime.now().strftime("%Y-%m-%d")}'
ws['A2'].font=I

hdr(ws,4,['检查类别','检查项','检查结果','风险等级','依据','具体问题','建议'])

checks = [
    ['采购方式','是否属于政府采购','本项目明确标注"不属于政府采购项目"','YELLOW',
     '采购公告称竞争性磋商，但声明非政府采购','采用政府采购的程序但不纳入政府采购监管，存在监管盲区',
     '查明采购资金来源，确认是否应纳入政府采购监管'],
    ['资格条件','特定资格要求','无(仅《政府采购法》第22条基本条件)','GREEN',
     '第三章明确"本项目的特定资格要求：无"','门槛低，有利于充分竞争','—'],
    ['资格条件','是否存在地域限制','未发现','GREEN','无注册地/纳税地限制','—','—'],
    ['资格条件','是否存在规模/业绩门槛','业绩仅作为加分项(履约能力4分)，不作资格条件','GREEN',
     '无注册资本/营业额/人员规模门槛','—','—'],
    ['技术参数','14项▲重要参数合理性','存在定向风险','YELLOW',
     '▲参数分布在5类设备中，每项扣2.5分(共35分)','实训室智能综合测试平台"BS架构"是核心产品▲项；单片机"工业组态软件"可能指向特定品牌',
     '核实市场上有≥3家供应商可满足各项▲参数'],
    ['技术参数','72项一般参数','参数极为详细，可能存在倾向性','YELLOW',
     '每项不满足扣0.25分(共18分)','如电工电路系统的电压/功率/尺寸精确到具体数值，可能与某个品牌完全匹配',
     '逐项做市场比对：是否≥3家厂商可完全满足'],
    ['技术参数','佐证材料要求','▲项须提供"公开发行的产品彩页资料"','YELLOW',
     '要求提供厂商官方彩页加盖公章','对新进入市场的供应商不友好——可能拿不到厂商授权彩页盖章',
     '可增加"或提供技术白皮书/检测报告"作为替代'],
    ['评分结构','技术分权重(53%)','技术分过高，报价仅30%','YELLOW',
     '技术53%+实施方案8%+售后5%=66%主观评分','报价分影响力有限，中标结果高度依赖技术参数响应程度',
     '建议适当降低技术分至≤45%'],
    ['核心产品','单一核心产品','仅"实训室智能综合测试平台"','YELLOW',
     '第二章5.3:核心产品为实训室智能综合测试平台','若核心产品只有1家满足，则属变相单一来源',
     '核实市场上核心产品的供应商数量'],
    ['程序规范','公告发布','中国招标投标公共服务平台','GREEN','非政府采购的正规公共平台','—','—'],
    ['程序规范','文件获取期限','日期为空白模板(未填写具体日期)','YELLOW',
     '文件显示"2025年05月 日至2025年05月 日"','模板填写不完整——实际发布时应已填写','确认实际公告中的日期'],
    ['程序规范','磋商保证金','不收取','GREEN','明确"本项目不收取磋商保证金"','降低供应商参与成本','—'],
    ['程序规范','联合体/分包/转包','均不允许','GREEN','符合小额采购的常规做法','—','—'],
    ['实质要求','供应商须知附表','空白表(模板未填写)','YELLOW',
     '第二章第一项"供应商须知附表"内容为空','关键信息(采购预算/最高限价/小微企业折扣等)在附表中','确认正本文件中附表已填写完整'],
    ['采购人','联系人','郭老师 028-84879013','INFO',
     '备案资料显示中标单位(中科兴蓉)联系人也是"郭"','同为"郭"姓，可能是巧合','建议核实采购人项目负责人与中标方联系人是否为同一人'],
]

for r,row in enumerate(checks,5):
    for c,val in enumerate(row,1):
        fill=RED_FILL if val=='RED' else (YELLOW_FILL if val=='YELLOW' else (GREEN_FILL if val=='GREEN' else None))
        fn=B if c<=2 else (BR if val=='YELLOW' and c==4 else N)
        cell(ws,r,c,val,fn,(L if c>=5 else C),fill)

r=21; ws.merge_cells(f'A{r}:G{r}')
ws[f'A{r}']='【综合评级: ⚠️ 中等风险 — 技术参数需做市场比对验证】'; ws[f'A{r}'].font=SUB
r+=1; ws.merge_cells(f'A{r}:G{r}')
ws[f'A{r}']=('1. 招标文件本身存在的主要问题: 技术参数极其详细(14▲+72一般=86项)，在缺少前期市场调研证据的情况下，有"对号入座"风险。\n'
    '2. "本项目不属于政府采购项目"的声明与采用竞争性磋商程序之间存在矛盾，资金来源需明确。\n'
    '3. 核心产品"实训室智能综合测试平台"的BS架构描述可能排除CS架构供应商，需验证市场竞争充分性。\n'
    '4. 采购人联系人(郭老师)与中标方(中科兴蓉)联系人(郭)同姓，建议核查是否同一人。')
ws[f'A{r}'].font=N; ws[f'A{r}'].alignment=L; ws.row_dimensions[r].height=75

ws.column_dimensions['A'].width=14; ws.column_dimensions['B'].width=22
ws.column_dimensions['C'].width=40; ws.column_dimensions['D'].width=14
ws.column_dimensions['E'].width=40; ws.column_dimensions['F'].width=40
ws.column_dimensions['G'].width=35

# ===== Sheet 2: 技术参数倾向性分析 =====
ws2=wb.create_sheet('技术参数倾向性分析')
ws2.merge_cells('A1:F1'); ws2['A1']='技术参数逐项分析 - 是否具有排他性/倾向性'; ws2['A1'].font=T
ws2.merge_cells('A2:F2'); ws2['A2']='▲=14项重要参数(每项不满足扣2.5分) | 一般=72项(每项不满足扣0.25分)'
ws2['A2'].font=I

hdr(ws2,4,['设备类别','参数类型','关键参数概述','潜在倾向性','排斥风险','分析说明'])

tech_analysis = [
    ['实训室智能综合测试平台(核心产品)','▲1','BS架构，所有数据在服务器端，浏览器访问','YELLOW-中等',
     'YELLOW','如确定为唯一核心产品，BS架构可能排除CS架构供应商。需确认市场≥3家BS架构供应商。'],
    ['实训室智能综合测试平台','▲2','学生实验流程控制，预习/实验/习题','GREEN-低','GREEN','教学平台的常见功能，市场竞争充分'],
    ['实训室智能综合测试平台','▲3','实验防抄袭系统','GREEN-低','GREEN','防抄袭为教学平台常见功能'],
    ['电工电路实训系统','▲4','交流输入三相四线380V±10%','YELLOW-中等','YELLOW','参数精确到±10%容差，需确认是否为国标要求'],
    ['电工电路实训系统','▲5','三相灯组负载≥15W灯泡','GREEN-低','GREEN','标准教学设备参数'],
    ['模电数电实训系统','▲6','DC ±12V、+5V多路稳压源，最大1A，短路保护','YELLOW-中等','YELLOW',
     '具体电压值和短路保护要求，需确认市场多品牌满足'],
    ['模电数电实训系统','▲7','实验箱主板提供高稳定直流±15V、±2V~±10V五档','YELLOW-中等','YELLOW','五档可调设计可能是某品牌独有，需验证'],
    ['传感器实训系统','▲8','数据采集卡及处理软件','GREEN-低','GREEN','传感器实验标准配置'],
    ['传感器实训系统','▲9','主板面板装电压/电流表','GREEN-低','GREEN','通用要求'],
    ['单片机实训系统','▲10','工业组态软件应用于单片机控制系统','RED-高','RED',
     '"工业组态软件"高度限制——国内主流为组态王/力控/MCGS等少数品牌，若目标明确指向某品牌则存在排他性'],
    ['单片机实训系统','▲11','支持MCU至少包含51/AVR/PIC/MSP430/STM32五种','GREEN-低','GREEN',
     '要求覆盖主流MCU类型，有利于扩大竞争'],
    ['单片机实训系统','▲12','各功能模块完全独立，互不干扰','GREEN-低','GREEN','合理的教学设备要求'],
    ['单片机实训系统','▲13','模块间采用排线连接','YELLOW-中等','YELLOW','"排线连接"可能指向某品牌设计，需验证是否为行业通用方式'],
    ['网络工程','▲14','根据现场需求定制','GREEN-低','GREEN','网络布线定制为必需项，无排他性'],
    ['全部设备','一般参数(72项)','72项具体技术规格','YELLOW-中等','YELLOW',
     '72项参数极为详细，若基于特定品牌产品编制，等于变相锁定供应商。建议随机抽取20项做市场验证。'],
]

for r,row in enumerate(tech_analysis,5):
    for c,val in enumerate(row,1):
        fill=RED_FILL if 'RED' in str(val) else (YELLOW_FILL if 'YELLOW' in str(val) else (GREEN_FILL if 'GREEN' in str(val) else None))
        cell(ws2,r,c,val,B if c==1 else N,L,fill)

r2=21; ws2.merge_cells(f'A{r2}:F{r2}')
ws2[f'A{r2}']='核心风险: ▲10(工业组态软件)是最强排他信号。如需确定是否存在"对号入座"，应对比中标方(中科兴蓉)的技术响应是否与招标参数完美匹配。'
ws2[f'A{r2}'].font=BR; ws2[f'A{r2}'].alignment=L; ws2.row_dimensions[r2].height=30

ws2.column_dimensions['A'].width=28; ws2.column_dimensions['B'].width=14
ws2.column_dimensions['C'].width=45; ws2.column_dimensions['D'].width=16
ws2.column_dimensions['E'].width=14; ws2.column_dimensions['F'].width=50

# ===== Sheet 3: 评分标准分析 =====
ws3=wb.create_sheet('评分标准分析')
ws3.merge_cells('A1:F1'); ws3['A1']='评分标准公平性分析'; ws3['A1'].font=T

hdr(ws3,3,['评分项','分值','占比','评分方式','公平性评估','风险分析'])

scoring = [
    ['报价','30','30%','客观分: (基准价/投标报价)×30','GREEN-客观公正','最低价中标激励不强(30分权重有限)'],
    ['技术参数(▲项)','35','35%','准客观: 逐项扣2.5分，需彩页佐证','YELLOW-主观空间','14项▲若指向特定品牌，则该品牌可直接拿满35分'],
    ['技术参数(一般项)','18','18%','准客观: 逐项扣0.25分','YELLOW-项数多','72项参数若基于某品牌编制，其他投标人不可避免丢分'],
    ['技术参数小计','53','53%','—','YELLOW-权重过高','技术分占比过半，事实上削弱了价格竞争的作用'],
    ['履约能力(业绩)','4','4%','准客观: 每个类似业绩1分/最高4分','GREEN-合理','4分差距不会大幅改变排名'],
    ['实施方案','8','8%','主观: "无缺陷"判断有自由裁量空间','YELLOW-主观','评委对"缺陷"定义的标准不一，存在暗箱评分风险'],
    ['售后服务方案','5','5%','主观: 同实施方案','YELLOW-主观','同上'],
    ['合计','100','100%','客观分: 30% | 准客观: 57% | 主观: 13%','总体: YELLOW','13%纯主观+57%准客观(技术参数来源不明)=存在操控空间'],
]

for r,row in enumerate(scoring,4):
    for c,val in enumerate(row,1):
        fill=RED_FILL if 'RED' in str(val) else (YELLOW_FILL if 'YELLOW' in str(val) else (GREEN_FILL if 'GREEN' in str(val) else None))
        cell(ws3,r,c,val,B if c==1 else N,L,fill)

r3=13; ws3.merge_cells(f'A{r3}:F{r3}')
ws3[f'A{r3}']=('评分结构问题: 53%技术分+8%实施方案+5%售后=66%的得分与技术参数响应相关，而技术参数来源不明。\n'
    '若技术参数基于(或倾向于)某特定品牌编制 → 该品牌直接获得66%的天然优势 → 招标失去实质竞争意义。\n'
    '建议: ①降低技术分至≤45% ②公开技术参数编制依据(是否做了市场调研) ③增加价格分权重至≥40%')
ws3[f'A{r3}'].font=BR; ws3[f'A{r3}'].alignment=L; ws3.row_dimensions[r3].height=65

ws3.column_dimensions['A'].width=20; ws3.column_dimensions['B'].width=10
ws3.column_dimensions['C'].width=12; ws3.column_dimensions['D'].width=36
ws3.column_dimensions['E'].width=20; ws3.column_dimensions['F'].width=50

# ===== Sheet 4: 同姓联系人分析 =====
ws4=wb.create_sheet('招标采购人分析')
ws4.merge_cells('A1:E1'); ws4['A1']='采购方/代理方/中标方 关联分析'; ws4['A1'].font=T
hdr(ws4,3,['主体','角色','关键信息','数据来源','风险标记'])

stakeholders = [
    ['四川护理职业学院','采购人','联系人: 郭老师 | 电话: 028-84879013 | 地址: 成都龙泉驿区龙都南路173号','招标文件第一章','—'],
    ['中正恒天国际招标有限公司','采购代理机构','联系人: 邓老师 | 电话: 028-81058218 | 邮箱: 2598729619@qq.com | 地址: 金牛区金周路595号','招标文件第一章','—'],
    ['中科兴蓉科技有限公司','中标(成交)供应商','联系人: 郭 | 电话: 13258321367 | 中标金额: 35.9万元','备案资料第3页','⚠️ 采购人联系人也姓"郭"'],
    ['百安智能科技有限公司','投标供应商(未中标)','(信息待OCR提取)','投标文件','—'],
    ['成都逐声科技有限公司','投标供应商(未中标)','(信息待OCR提取)','投标文件','—'],
    ['张鹏举','招标文件创建者','WPS账号: 1389992463 | 硬件ID已提取','招标文件.docx core.xml','招标文件编制者身份已暴露'],
]

for r,row in enumerate(stakeholders,4):
    for c,val in enumerate(row,1):
        fill=YELLOW_FILL if '⚠️' in str(val) else (RED_FILL if c==5 and '已暴露' in str(val) else None)
        cell(ws4,r,c,val,B if c==1 else N,L,fill)

r4=11; ws4.merge_cells(f'A{r4}:E{r4}')
ws4[f'A{r4}']=('⚠️ 采购人联系人"郭老师"与中标方(中科兴蓉)"郭"同为郭姓。\n'
    '这可能是纯粹巧合(郭是常见姓氏)，也可能指向同一人或关联人员。\n'
    '建议: ①核实二人是否为同一人 ②核查中科兴蓉的法定代表人/股东是否含"郭"姓 ③确认采购人项目负责人是否与供应商存在亲属/利益关系')
ws4[f'A{r4}'].font=BR; ws4[f'A{r4}'].alignment=L; ws4.row_dimensions[r4].height=60

ws4.column_dimensions['A'].width=26; ws4.column_dimensions['B'].width=16
ws4.column_dimensions['C'].width=50; ws4.column_dimensions['D'].width=26
ws4.column_dimensions['E'].width=32

# ===== Sheet 5: 招标文件规范性审查 =====
ws5=wb.create_sheet('规范性审查')
ws5.merge_cells('A1:D1'); ws5['A1']='招标文件形式规范性审查'; ws5['A1'].font=T
hdr(ws5,3,['检查项','标准要求','实际情况','状态'])

norm_check = [
    ['文件封面信息完整','应含项目编号/名称/采购人/代理机构/日期','项目编号+名称+编制单位+2025年05月','✅ 完整'],
    ['目录与页码对应','目录页码应与正文一致','目录页码标注完整','✅ 完整'],
    ['采购预算/最高限价','应在供应商须知附表中明确','附表为空(模板未填)','❌ 缺漏'],
    ['小微企业价格扣除','应在须知附表中说明','附表为空','❌ 缺漏'],
    ['采购标的清单','应含数量/单位/是否进口/节能环保','第二章有采购清单表','✅ 完整'],
    ['技术参数规范','不应使用"知名品牌""优质"等模糊表述','未发现模糊表述，参数极为具体','✅ 具体(但过于具体可能是问题)'],
    ['资格条件','不得设置与项目无关的资格条件','仅《政府采购法》第22条+无特定资格要求','✅ 合规'],
    ['评审方法','综合评分法，应公开评分细则','第八章完整列出评分细则','✅ 完整'],
    ['合同草案','应提供政府采购合同草案','第九章提供了合同草案','✅ 完整'],
    ['文件获取方式','应说明获取方式/地点/费用','现场或网络获取，300元/份','✅ 完整(但日期空白)'],
    ['是否要求现场考察','—','未要求','—'],
    ['保证金','投标/履约保证金','不收取投标保证金。履约保证金金额未明确','⚠️ 履约保证金数额未填写'],
]

for r,row in enumerate(norm_check,4):
    for c,val in enumerate(row,1):
        fill=RED_FILL if '❌' in str(val) else (YELLOW_FILL if '⚠️' in str(val) else None)
        cell(ws5,r,c,val,B if c==1 else N,L,fill)

r5=17; ws5.merge_cells(f'A{r5}:D{r5}')
ws5[f'A{r5}']='形式审查结论: 招标文件形式基本规范，但关键信息(预算金额、最高限价、小微企业政策)因供应商须知附表为空模板而缺失。需查阅正本文件确认。'
ws5[f'A{r5}'].font=BR; ws5[f'A{r5}'].alignment=L

ws5.column_dimensions['A'].width=22; ws5.column_dimensions['B'].width=40
ws5.column_dimensions['C'].width=45; ws5.column_dimensions['D'].width=14

# ===== Sheet 6: 结论与建议 =====
ws6=wb.create_sheet('结论与建议')
ws6.merge_cells('A1:C1'); ws6['A1']='招标文件分析结论与追查建议'; ws6['A1'].font=T
hdr(ws6,3,['序号','发现/问题','建议措施'])

actions = [
    ['1','▲10"工业组态软件"是最强排他信号','对比中标方(中科兴蓉)的技术响应文件，核实其产品恰好完全匹配该▲参数 → 如匹配则高度可疑'],
    ['2','72项一般技术参数是否基于某品牌编制','随机抽取20项一般参数，通过公开渠道查询是否有≥3家厂商能完全满足'],
    ['3','采购人"郭老师"与中标方"郭"同姓','通过天眼查核查中科兴蓉的法定代表人/股东中有无"郭"姓，确认为同一人则触发《采购法》回避条款'],
    ['4','"不属于政府采购项目"但使用采购程序','查明采购资金来源(财政资金 vs 自筹资金)，确认是否应受政府采购法监管'],
    ['5','技术分53%权重过高','对比同类采购项目的技术分权重，如明显偏高则存在主观操控空间'],
    ['6','供应商须知附表空白','向代理机构(中正恒天国际招标)调取完整版招标文件，核实预算/限价/政策优惠是否已公开'],
    ['7','招标文件创建者"张鹏举"','如需要，可通过WPS账号ID(1389992463)追溯文件制作源头，核实是否有利益冲突'],
    ['8','中标价=预算上限(35.9万)','调取完整的报价记录和评分表，确认是否存在异常打分或报价'],
]

for r,row in enumerate(actions,4):
    for c,val in enumerate(row,1):
        cell(ws6,r,c,val,B if c==1 else N,L)

ws6.column_dimensions['A'].width=8; ws6.column_dimensions['B'].width=60; ws6.column_dimensions['C'].width=60

output = r'D:\openclaw-workspace\projects\护理学院医工设备采购\招标文件合规分析.xlsx'
wb.save(output)
print(f'Saved: {output}')
