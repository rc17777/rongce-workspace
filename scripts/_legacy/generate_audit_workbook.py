"""
政府审计通用底稿模板生成器 v1.0
融策会计师事务所 | 融策工程咨询公司
适用：12大政府审计业务线
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

OUTPUT_DIR = r"D:\openclaw-workspace\output"
os.makedirs(OUTPUT_DIR, exist_ok=True)
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "融策政府审计通用底稿模板_v1.0.xlsx")

# ============ 样式定义 ============
def create_styles(wb):
    title_style = NamedStyle(name="title_style")
    title_style.font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
    title_style.fill = PatternFill(start_color="0A1F3F", end_color="0A1F3F", fill_type="solid")
    title_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wb.add_named_style(title_style)

    subtitle_style = NamedStyle(name="subtitle_style")
    subtitle_style.font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
    subtitle_style.fill = PatternFill(start_color="1A5C6E", end_color="1A5C6E", fill_type="solid")
    subtitle_style.alignment = Alignment(horizontal="left", vertical="center")
    wb.add_named_style(subtitle_style)

    header_style = NamedStyle(name="header_style")
    header_style.font = Font(name="微软雅黑", size=10, bold=True, color="C5955C")
    header_style.fill = PatternFill(start_color="0A1F3F", end_color="0A1F3F", fill_type="solid")
    header_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_style.border = Border(
        left=Side(style="thin", color="C5955C"),
        right=Side(style="thin", color="C5955C"),
        top=Side(style="thin", color="C5955C"),
        bottom=Side(style="thin", color="C5955C")
    )
    wb.add_named_style(header_style)

    body_style = NamedStyle(name="body_style")
    body_style.font = Font(name="宋体", size=10, color="000000")
    body_style.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    body_style.border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )
    wb.add_named_style(body_style)

    alt_fill = PatternFill(start_color="F5F2EC", end_color="F5F2EC", fill_type="solid")
    return alt_fill


# ============ Sheet 1: 封面 ============
def build_cover(wb, styles):
    ws = wb.active
    ws.title = "① 项目封面"

    ws.merge_cells("A1:H1")
    ws["A1"] = "政府审计项目工作底稿"
    ws["A1"].style = "title_style"
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:H2")
    ws["A2"] = "四川融策会计师事务所 / 四川融策工程咨询公司"
    ws["A2"].style = "subtitle_style"
    ws.row_dimensions[2].height = 28

    fields = [
        ("项目名称", ""),
        ("审计类型", ""),
        ("被审计单位", ""),
        ("被审计期间", ""),
        ("审计组长", ""),
        ("主审", ""),
        ("审计组成员", ""),
        ("进场日期", ""),
        ("预计完成日期", ""),
        ("项目编号", ""),
    ]

    for idx, (label, value) in enumerate(fields, 4):
        ws[f"A{idx}"] = label
        ws[f"A{idx}"].style = "header_style"
        ws.merge_cells(f"A{idx}:B{idx}")
        ws[f"C{idx}"] = value
        ws.merge_cells(f"C{idx}:H{idx}")
        ws[f"C{idx}"].style = "body_style"
        ws.row_dimensions[idx].height = 28

    dv = DataValidation(type="list", formula1='"经济责任审计,收支审计,预算执行审计,专项资金审计,往来款清理,招投标审计,国企审计,成本效益审计,能源审计,工程竣工决算审计,预算绩效管理,政府补贴审计"', allow_blank=True)
    dv.add(f"C5")
    ws.add_data_validation(dv)

    ws["A15"] = "12大业务线速查"
    ws["A15"].style = "subtitle_style"
    ws.merge_cells("A15:H15")

    biz_lines = [
        ("1", "经济责任审计", "任中/离任/自然资源", "干部履职+决策合规+廉洁"),
        ("2", "收支审计", "", "收入完整+支出合规+预算执行"),
        ("3", "预算执行审计", "", "预算编制+执行偏差+绩效目标"),
        ("4", "专项资金审计", "社保/营养餐/扶贫等", "专款专用+拨付时效+使用效益"),
        ("5", "往来款清理", "", "长期挂账+坏账风险+资金盘活"),
        ("6", "招投标审计", "", "程序合规+围标串标+控制价虚高"),
        ("7", "国企审计", "", "资产质量+内控执行+三重一大"),
        ("8", "成本效益审计", "", "成本真实性+效益量化+投入产出"),
        ("9", "能源审计", "含碳中和", "能耗数据+节能措施+碳排放"),
        ("10", "工程竣工决算审计", "", "投资控制+变更签证+结算审减"),
        ("11", "预算绩效管理", "目标/评估/监控/评价", "绩效目标+指标体系+评价结论"),
        ("12", "政府补贴审计", "", "申报合规+使用效益+退出机制"),
    ]

    headers = ["编号", "业务线", "子类型", "关注重点"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=16, column=col, value=h)
        cell.style = "header_style"

    for row_idx, (num, name, sub, focus) in enumerate(biz_lines, 17):
        ws.cell(row=row_idx, column=1, value=num).style = "body_style"
        ws.cell(row=row_idx, column=2, value=name).style = "body_style"
        ws.cell(row=row_idx, column=3, value=sub).style = "body_style"
        ws.cell(row=row_idx, column=4, value=focus).style = "body_style"
        if row_idx % 2 == 0:
            for c in range(1, 5):
                ws.cell(row=row_idx, column=c).fill = styles

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 35
    for c in "EFGH":
        ws.column_dimensions[c].width = 12

    return ws


# ============ Sheet 2: 审计事项清单 ============
def build_item_list(wb, styles):
    ws = wb.create_sheet("② 审计事项清单")

    ws.merge_cells("A1:L1")
    ws["A1"] = "审计事项清单（总览表）"
    ws["A1"].style = "title_style"
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:L2")
    ws["A2"] = "说明：每个审计事项对应一张取证单+一张工作底稿，索引号格式：业务线编号-序号"
    ws["A2"].font = Font(name="宋体", size=9, italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 25

    headers = [
        "索引号", "审计事项", "涉及金额（元）", "风险等级",
        "审计程序", "取证单编号", "工作底稿编号", "问题编号",
        "审计人员", "完成状态", "复核状态", "备注"
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.style = "header_style"

    examples = [
        ("1-01", "三重一大决策程序合规性", "", "高", "查阅会议纪要+对比制度", "QZ-001", "DZ-001", "WT-001", "张三", "已完成", "待复核", ""),
        ("1-02", "招投标程序合规性", "5000000", "高", "开标记录+评标报告+合同台账", "QZ-002", "DZ-002", "WT-002", "李四", "进行中", "", ""),
        ("1-03", "工程变更签证规范性", "1200000", "中", "变更签证单+审批流程+造价复核", "QZ-003", "DZ-003", "", "王五", "待开始", "", ""),
    ]

    for row_idx, ex in enumerate(examples, 4):
        for col, val in enumerate(ex, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.style = "body_style"
            if row_idx % 2 == 0:
                cell.fill = styles

    dv_risk = DataValidation(type="list", formula1='"高,中,低"', allow_blank=True)
    dv_risk.add(f"E4:E100")
    ws.add_data_validation(dv_risk)

    dv_status = DataValidation(type="list", formula1='"待开始,进行中,已完成"', allow_blank=True)
    dv_status.add(f"J4:J100")
    ws.add_data_validation(dv_status)

    dv_review = DataValidation(type="list", formula1='",待复核,已复核,退回修改"', allow_blank=True)
    dv_review.add(f"K4:K100")
    ws.add_data_validation(dv_review)

    col_widths = [10, 28, 16, 10, 30, 12, 14, 10, 10, 10, 10, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"
    return ws


# ============ Sheet 3: 取证单模板 ============
def build_evidence_sheet(wb, styles):
    ws = wb.create_sheet("③ 取证单模板")

    ws.merge_cells("A1:H1")
    ws["A1"] = "审计取证单"
    ws["A1"].style = "title_style"
    ws.row_dimensions[1].height = 35

    fields = [
        ("项目名称", ""),
        ("被审计单位", ""),
        ("审计事项", ""),
        ("取证单编号", ""),
        ("索引号", ""),
        ("编制日期", ""),
        ("编制人", ""),
        ("复核人", ""),
    ]

    for idx, (label, value) in enumerate(fields, 3):
        ws[f"A{idx}"] = label
        ws[f"A{idx}"].style = "header_style"
        ws.merge_cells(f"A{idx}:B{idx}")
        ws[f"C{idx}"] = value
        ws.merge_cells(f"C{idx}:H{idx}")
        ws[f"C{idx}"].style = "body_style"
        ws.row_dimensions[idx].height = 28

    ws.merge_cells("A12:H12")
    ws["A12"] = "取证内容"
    ws["A12"].style = "subtitle_style"
    ws.row_dimensions[12].height = 28

    ws.merge_cells("A13:H20")
    for r in range(13, 21):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.style = "body_style"
            if r % 2 == 0:
                cell.fill = styles

    ws.merge_cells("A21:H21")
    ws["A21"] = "取证依据（法规/制度/文件）"
    ws["A21"].style = "subtitle_style"
    ws.row_dimensions[21].height = 28

    ws.merge_cells("A22:H25")
    for r in range(22, 26):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.style = "body_style"
            if r % 2 == 0:
                cell.fill = styles

    ws.merge_cells("A26:H26")
    ws["A26"] = "附件清单"
    ws["A26"].style = "subtitle_style"
    ws.row_dimensions[26].height = 28

    attach_headers = ["序号", "附件名称", "页数/数量", "备注"]
    for col, h in enumerate(attach_headers, 1):
        cell = ws.cell(row=27, column=col, value=h)
        cell.style = "header_style"

    for r in range(28, 33):
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            cell.style = "body_style"
            if r % 2 == 0:
                cell.fill = styles

    ws.merge_cells("A33:H33")
    ws["A33"] = "被审计单位意见"
    ws["A33"].style = "subtitle_style"

    ws.merge_cells("A34:H38")
    for r in range(34, 39):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.style = "body_style"

    ws["A39"] = "被审计单位（盖章）"
    ws["A39"].style = "header_style"
    ws.merge_cells("A39:B39")
    ws["C39"] = ""
    ws.merge_cells("C39:D39")
    ws["C39"].style = "body_style"

    ws["E39"] = "经办人签字"
    ws["E39"].style = "header_style"
    ws.merge_cells("E39:F39")
    ws["G39"] = ""
    ws.merge_cells("G39:H39")
    ws["G39"].style = "body_style"

    ws["A40"] = "日期"
    ws["A40"].style = "header_style"
    ws.merge_cells("A40:B40")
    ws["C40"] = ""
    ws.merge_cells("C40:H40")
    ws["C40"].style = "body_style"

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    for c in "CDEFGH":
        ws.column_dimensions[c].width = 16

    return ws


# ============ Sheet 4: 工作底稿模板 ============
def build_workpaper(wb, styles):
    ws = wb.create_sheet("④ 工作底稿模板")

    ws.merge_cells("A1:I1")
    ws["A1"] = "审计工作底稿"
    ws["A1"].style = "title_style"
    ws.row_dimensions[1].height = 35

    fields = [
        ("项目名称", ""),
        ("被审计单位", ""),
        ("审计事项", ""),
        ("工作底稿编号", ""),
        ("索引号", ""),
        ("审计期间", ""),
        ("编制人/日期", ""),
        ("复核人/日期", ""),
    ]

    for idx, (label, value) in enumerate(fields, 3):
        ws[f"A{idx}"] = label
        ws[f"A{idx}"].style = "header_style"
        ws.merge_cells(f"A{idx}:B{idx}")
        ws[f"C{idx}"] = value
        ws.merge_cells(f"C{idx}:I{idx}")
        ws[f"C{idx}"].style = "body_style"
        ws.row_dimensions[idx].height = 28

    ws.merge_cells("A12:I12")
    ws["A12"] = "一、审计目标"
    ws["A12"].style = "subtitle_style"
    ws.row_dimensions[12].height = 28

    ws.merge_cells("A13:I15")
    for r in range(13, 16):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.style = "body_style"
            if r % 2 == 0:
                cell.fill = styles

    ws.merge_cells("A16:I16")
    ws["A16"] = "二、审计程序"
    ws["A16"].style = "subtitle_style"
    ws.row_dimensions[16].height = 28

    ws.merge_cells("A17:I20")
    for r in range(17, 21):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.style = "body_style"
            if r % 2 == 0:
                cell.fill = styles

    ws.merge_cells("A21:I21")
    ws["A21"] = "三、审计过程记录"
    ws["A21"].style = "subtitle_style"
    ws.row_dimensions[21].height = 28

    proc_headers = ["日期", "审计步骤", "查阅资料", "数据来源", "核对结果", "异常/疑点", "处理措施", "责任人", "备注"]
    for col, h in enumerate(proc_headers, 1):
        cell = ws.cell(row=22, column=col, value=h)
        cell.style = "header_style"

    for r in range(23, 32):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.style = "body_style"
            if r % 2 == 0:
                cell.fill = styles

    ws.merge_cells("A32:I32")
    ws["A32"] = "四、审计结论"
    ws["A32"].style = "subtitle_style"
    ws.row_dimensions[32].height = 28

    ws.merge_cells("A33:I36")
    for r in range(33, 37):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.style = "body_style"
            if r % 2 == 0:
                cell.fill = styles

    ws.merge_cells("A37:I37")
    ws["A37"] = "五、问题发现（如适用）"
    ws["A37"].style = "subtitle_style"
    ws.row_dimensions[37].height = 28

    issue_headers = ["问题编号", "问题描述", "涉及金额", "定性依据", "责任认定", "整改建议", "附件索引", "重要性", "备注"]
    for col, h in enumerate(issue_headers, 1):
        cell = ws.cell(row=38, column=col, value=h)
        cell.style = "header_style"

    for r in range(39, 44):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.style = "body_style"
            if r % 2 == 0:
                cell.fill = styles

    dv_resp = DataValidation(type="list", formula1='"直接责任,主管责任,领导责任"', allow_blank=True)
    dv_resp.add(f"E39:E100")
    ws.add_data_validation(dv_resp)

    dv_imp = DataValidation(type="list", formula1='"重大,重要,一般"', allow_blank=True)
    dv_imp.add(f"H39:H100")
    ws.add_data_validation(dv_imp)

    col_widths = [10, 25, 14, 20, 12, 20, 12, 10, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return ws


# ============ Sheet 5: 问题汇总表 ============
def build_issue_summary(wb, styles):
    ws = wb.create_sheet("⑤ 问题汇总表")

    ws.merge_cells("A1:M1")
    ws["A1"] = "审计发现问题汇总表"
    ws["A1"].style = "title_style"
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:M2")
    ws["A2"] = "说明：本表自动汇总各工作底稿中的问题，用于报告撰写和整改跟踪"
    ws["A2"].font = Font(name="宋体", size=9, italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 25

    headers = [
        "问题编号", "审计事项索引", "问题描述", "涉及金额（元）",
        "问题类型", "定性法规依据", "责任认定", "责任主体",
        "整改建议", "整改期限", "整改状态", "验证结果", "备注"
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.style = "header_style"

    dv_type = DataValidation(type="list", formula1='"政策执行,重大决策,财政财务,资产管理,内控缺陷,廉洁风险,工程管理,采购招投标,绩效目标,其他"', allow_blank=True)
    dv_type.add(f"E4:E200")
    ws.add_data_validation(dv_type)

    dv_resp = DataValidation(type="list", formula1='"直接责任,主管责任,领导责任"', allow_blank=True)
    dv_resp.add(f"G4:G200")
    ws.add_data_validation(dv_resp)

    dv_status = DataValidation(type="list", formula1='"未整改,整改中,已整改,无法整改,无需整改"', allow_blank=True)
    dv_status.add(f"K4:K200")
    ws.add_data_validation(dv_status)

    examples = [
        ("WT-001", "1-01", "三重一大决策未履行集体决策程序，涉及金额500万元", "5000000",
         "重大决策", "《关于进一步推进国有企业贯彻落实三重一大决策制度的意见》", "直接责任", "张某",
         "完善决策程序，补全会议纪要", "2026-08-30", "整改中", "", ""),
    ]

    for row_idx, ex in enumerate(examples, 4):
        for col, val in enumerate(ex, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.style = "body_style"
            if row_idx % 2 == 0:
                cell.fill = styles

    col_widths = [10, 12, 30, 14, 12, 25, 10, 10, 20, 12, 10, 10, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"
    return ws


# ============ Sheet 6: 法规对照表 ============
def build_law_reference(wb, styles):
    ws = wb.create_sheet("⑥ 法规对照表")

    ws.merge_cells("A1:G1")
    ws["A1"] = "审计法规对照表"
    ws["A1"].style = "title_style"
    ws.row_dimensions[1].height = 35

    headers = ["序号", "问题类型", "适用法规", "具体条款", "定性表述", "处罚/处理依据", "备注"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.style = "header_style"

    laws = [
        ("1", "预算执行", "《中华人民共和国预算法》", "第X条", "预算执行不到位", "", ""),
        ("2", "采购招投标", "《中华人民共和国招标投标法》", "第X条", "规避招标/围标串标", "", ""),
        ("3", "资产管理", "《行政单位国有资产管理暂行办法》", "第X条", "资产处置未审批", "", ""),
        ("4", "内控缺陷", "《行政事业单位内部控制规范（试行）》", "第X条", "内控制度缺失/执行不到位", "", ""),
        ("5", "廉洁风险", "《中国共产党纪律处分条例》", "第X条", "违反廉洁纪律", "", ""),
        ("6", "工程管理", "《政府投资条例》", "第X条", "未批先建/超概算", "", ""),
        ("7", "绩效目标", "《项目支出绩效评价管理办法》", "第X条", "绩效目标未达成", "", ""),
    ]

    for row_idx, law in enumerate(laws, 3):
        for col, val in enumerate(law, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.style = "body_style"
            if row_idx % 2 == 0:
                cell.fill = styles

    col_widths = [8, 12, 28, 15, 25, 25, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return ws


# ============ Sheet 7: 资料索引表 ============
def build_index(wb, styles):
    ws = wb.create_sheet("⑦ 资料索引表")

    ws.merge_cells("A1:H1")
    ws["A1"] = "审计资料索引表"
    ws["A1"].style = "title_style"
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = "说明：索引号规则——1开头=政策制度类 / 2开头=财务资料类 / 3开头=合同招投标类 / 4开头=会议纪要类 / 5开头=现场取证类 / 9开头=其他"
    ws["A2"].font = Font(name="宋体", size=9, italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 25

    headers = ["索引号", "资料名称", "资料类型", "来源单位", "取得日期", "页数/份数", "存放位置", "关联审计事项"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.style = "header_style"

    dv_type = DataValidation(type="list", formula1='"政策制度,财务资料,合同招投标,会议纪要,现场取证,函证回执,影像资料,其他"', allow_blank=True)
    dv_type.add(f"C4:C500")
    ws.add_data_validation(dv_type)

    col_widths = [10, 28, 12, 15, 12, 12, 15, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"
    return ws


# ============ Sheet 8: 使用说明 ============
def build_instructions(wb):
    ws = wb.create_sheet("⑧ 使用说明")

    ws.merge_cells("A1:H1")
    ws["A1"] = "底稿模板使用说明"
    ws["A1"].style = "title_style"
    ws.row_dimensions[1].height = 35

    instructions = [
        ("", ""),
        ("一、模板结构", ""),
        ("本模板包含8个Sheet，按审计流程组织：", ""),
        ("① 项目封面", "填写项目基本信息，选择审计类型，查看12大业务线速查"),
        ("② 审计事项清单", "按审计事项逐条列出，每个事项对应一张取证单+一张工作底稿"),
        ("③ 取证单模板", "用于向被审计单位取证，需盖章签字确认"),
        ("④ 工作底稿模板", "记录审计过程、结论和问题发现"),
        ("⑤ 问题汇总表", "汇总所有审计发现的问题，用于报告撰写和整改跟踪"),
        ("⑥ 法规对照表", "预设常用法规，按问题类型快速引用"),
        ("⑦ 资料索引表", "统一管理所有审计资料，建立索引和关联"),
        ("", ""),
        ("二、索引号规则", ""),
        ("审计事项索引号", "业务线编号-序号，如 1-01=经责审计第1个事项"),
        ("取证单编号", "QZ-XXX，如 QZ-001"),
        ("工作底稿编号", "DZ-XXX，如 DZ-001"),
        ("问题编号", "WT-XXX，如 WT-001"),
        ("资料索引号", "类别号-序号，如 1-01=政策制度类第1份资料"),
        ("", ""),
        ("三、使用流程", ""),
        ("Step 1", "在「项目封面」填写基本信息，选择审计类型"),
        ("Step 2", "在「审计事项清单」列出所有审计事项，分配人员和风险等级"),
        ("Step 3", "每个审计事项：复制「取证单模板」和「工作底稿模板」各一张，填写编号和索引"),
        ("Step 4", "审计过程中，在「资料索引表」登记所有取得的资料"),
        ("Step 5", "发现问题时，在「问题汇总表」登记，关联审计事项索引"),
        ("Step 6", "审计结束前，在「法规对照表」补充具体条款和定性依据"),
        ("Step 7", "撰写报告时，从「问题汇总表」提取问题，从「法规对照表」引用依据"),
    ]

    for row_idx, (col_a, col_b) in enumerate(instructions, 3):
        ws.cell(row=row_idx, column=1, value=col_a).style = "body_style"
        ws.cell(row=row_idx, column=2, value=col_b).style = "body_style"
        ws.merge_cells(f"B{row_idx}:H{row_idx}")
        if col_a in ["一、模板结构", "二、索引号规则", "三、使用流程"]:
            ws.cell(row=row_idx, column=1).font = Font(name="微软雅黑", size=11, bold=True, color="0A1F3F")
        if row_idx % 2 == 0:
            for c in range(1, 9):
                ws.cell(row=row_idx, column=c).fill = PatternFill(start_color="F5F2EC", end_color="F5F2EC", fill_type="solid")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 60
    for c in "CDEFGH":
        ws.column_dimensions[c].width = 12

    return ws


# ============ 主函数 ============
def main():
    wb = openpyxl.Workbook()
    styles = create_styles(wb)

    build_cover(wb, styles)
    build_item_list(wb, styles)
    build_evidence_sheet(wb, styles)
    build_workpaper(wb, styles)
    build_issue_summary(wb, styles)
    build_law_reference(wb, styles)
    build_index(wb, styles)
    build_instructions(wb)

    wb.save(OUTPUT_FILE)
    print(f"底稿模板已生成: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
