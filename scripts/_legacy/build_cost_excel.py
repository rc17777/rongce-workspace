# -*- coding: utf-8 -*-
"""生成健康照护师成本构成测算Excel"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============ Styles ============
title_font = Font(name='微软雅黑', size=16, bold=True)
h1_font = Font(name='微软雅黑', size=14, bold=True)
h2_font = Font(name='微软雅黑', size=12, bold=True)
h3_font = Font(name='微软雅黑', size=11, bold=True)
normal_font = Font(name='微软雅黑', size=10)
bold_font = Font(name='微软雅黑', size=10, bold=True)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
sub_header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
total_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
red_font = Font(name='微软雅黑', size=10, bold=True, color='CC0000')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
money_fmt = '#,##0.00'

def style_header_row(ws, row, max_col):
    for c in range(1, max_col+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def sc(ws, row, col, align='center'):
    """Style a data cell and return it"""
    cell = ws.cell(row=row, column=col)
    cell.font = normal_font
    cell.alignment = center_align if align == 'center' else left_align
    cell.border = thin_border
    return cell

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# =====================================================
# Sheet 1: 成本构成总览
# =====================================================
ws1 = wb.active
ws1.title = '成本构成总览'

ws1.merge_cells('A1:I1')
c = ws1['A1']
c.value = '健康照护师（长期照护师）职业技能等级认定 — 单人成本构成测算总览'
c.font = title_font
c.alignment = center_align

ws1.merge_cells('A2:I2')
c = ws1['A2']
c.value = '数据来源：智护学院-长期照护师-耗材成本表-2026.05.14 | 测算基准日：2026年5月 | 编制单位：四川融策会计师事务所'
c.font = Font(name='微软雅黑', size=9, color='666666')
c.alignment = center_align

# ---- Section A: 成本汇总 ----
ws1.merge_cells('A4:I4')
ws1['A4'].value = '一、单人成本构成汇总（单位：元/人）'
ws1['A4'].font = h1_font

row = 6
headers = ['序号', '成本项目', '测算依据/方法', '五级/初级', '四级/中级', '三级/高级',
           '三级vs初级\n增幅%', '是否已测算', '备注']
for c, h in enumerate(headers, 1):
    ws1.cell(row=row, column=c, value=h)
style_header_row(ws1, row, len(headers))

overview_data = [
    # (seq, name, basis, p5, p4, p3, delta, measured, note)
    ['', '一、政策规定收费', '', '', '', '', '', '', ''],
    ['1', '理论考试费', '川发改价格〔2017〕472号', 30, 35, 40, 33.3, '✓', '政府定价上限'],
    ['2', '操作技能考核费', '川发改价格〔2017〕472号', 140, 190, 240, 71.4, '✓', '政府定价上限'],
    ['', '  政策收费小计', '', 170, 225, 280, 64.7, '', ''],
    ['', '', '', '', '', '', '', '', ''],
    ['', '二、实际发生成本', '', '', '', '', '', '', ''],
    ['3', '考务平台费', '平台服务年费÷年考核人次', 30, 30, 30, 0, '✓', '参考现行收费方案'],
    ['4', '人工成本', '考评员劳务费×人数÷单场考生数', '', '', '', '', '待补充', '需向护理学院索取'],
    ['5', '设施设备使用费', '逐项分摊，详见"设施设备成本明细"表', 93, 158, 190, 104.3, '✓', '磨损费/组均摊模型'],
    ['6', '耗材成本', '逐项计量，详见"耗材成本明细"表', 87.20, 167.28, 280.49, 221.7, '✓', '实操8题全量耗材'],
    ['7', '全程视频监控费', '监控系统投入÷年限÷年人次', '', '', '', '', '待补充', '需向护理学院索取'],
    ['8', '证书制作费', '证书工本费单价', '', '', '', '', '待补充', '需向护理学院索取'],
    ['9', '文印广告费', '试卷+标识标牌÷考核人次', '', '', '', '', '待补充', '需向护理学院索取'],
    ['10', '场地水电费', '(年租金+水电)÷年考核天数÷日人次', '', '', '', '', '待补充', '需向护理学院索取'],
    ['', '  已测算实际成本小计', '', 210.20, 355.28, 500.49, 138.1, '', '仅含考务+设备+耗材'],
    ['', '', '', '', '', '', '', '', ''],
    ['', '三、全口径成本测算（含政策收费）', '', '', '', '', '', '', ''],
    ['', '  全口径成本（已测算部分）', '', 380.20, 580.28, 780.49, 105.3, '', '政策收费+已测算实际成本'],
    ['', '', '', '', '', '', '', '', ''],
    ['', '四、与现行收费方案对比', '', '', '', '', '', '', ''],
    ['', '现行收费标准（方案5稿）', '', 320, 405, 500, 56.3, '', 'B类-健康照护师'],
    ['', '已测算成本覆盖', '', 380.20, 580.28, 780.49, 105.3, '', '不含人工等6项待补充成本'],
    ['', '差额（收费-已测算成本）', '', -60.20, -175.28, -280.49, '', '', '⚠ 已测算成本已超过现行收费'],
]

for i, d in enumerate(overview_data):
    r = row + 1 + i
    for ci, v in enumerate(d, 1):
        cell = sc(ws1, r, ci, 'left' if ci in [2, 3] else 'center')
        cell.value = v
        if ci in [4, 5, 6] and isinstance(v, (int, float)):
            cell.number_format = money_fmt

    # Section header styling
    first_val = str(d[1]) if d[1] else str(d[0])
    if any(first_val.startswith(x) for x in ['一、', '二、', '三、', '四、']):
        ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        ws1.cell(row=r, column=1).font = h3_font
        for ci in range(1, 10):
            ws1.cell(row=r, column=ci).fill = sub_header_fill
            ws1.cell(row=r, column=ci).border = thin_border
    # Subtotal rows
    if '小计' in str(d[1]) or '全口径' in str(d[1]) or '差额' in str(d[1]) or '现行收费' in str(d[1]) or '已测算成本覆盖' in str(d[1]):
        for ci in range(1, 10):
            ws1.cell(row=r, column=ci).font = bold_font
            ws1.cell(row=r, column=ci).fill = total_fill
        if '差额' in str(d[1]):
            for ci in [4, 5, 6]:
                ws1.cell(row=r, column=ci).font = red_font

set_widths(ws1, [6, 26, 32, 14, 14, 14, 12, 12, 24])

# =====================================================
# Sheet 2: 耗材成本明细 (all 3 levels)
# =====================================================
ws2 = wb.create_sheet('耗材成本明细')

ws2.merge_cells('A1:J1')
ws2['A1'].value = '健康照护师（长期照护师）等级认定 — 耗材成本明细测算'
ws2['A1'].font = title_font
ws2['A1'].alignment = center_align

ws2.merge_cells('A2:J2')
ws2['A2'].value = '说明：耗材成本=单价×单人使用量；4人一组项目=单价÷4；磨损费项目=单价÷预估使用次数'
ws2['A2'].font = Font(name='微软雅黑', size=9, color='666666')
ws2['A2'].alignment = center_align

src = openpyxl.load_workbook(
    r'C:\Users\scrccpa\Desktop\成本测算=护理学院\各职业耗材成本明细\智护学院-长期照护师-耗材成本表-2026.05.14.xlsx',
    data_only=True
)

current_row = 4
for level_idx, (level_name, sheet_name) in enumerate([
    ('五级/初级工', '五级（初级）'),
    ('四级/中级工', '四级（中级）'),
    ('三级/高级工', '三级（高级）')
]):
    src_ws = src[sheet_name]

    # Section title
    ws2.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
    ws2.cell(row=current_row, column=1).value = f'{"一二三"[level_idx]}、{level_name} — 耗材成本明细'
    ws2.cell(row=current_row, column=1).font = h2_font
    for ci in range(1, 11):
        ws2.cell(row=current_row, column=ci).fill = sub_header_fill
    current_row += 1

    # Headers
    hdrs = ['序号', '耗材名称', '规格', '参考品牌', '单位', '单价（元）',
            '单人使用量', '单人成本（元）', '计算过程', '备注']
    for ci, h in enumerate(hdrs, 1):
        ws2.cell(row=current_row, column=ci).value = h
    style_header_row(ws2, current_row, 10)
    current_row += 1

    # Parse consumables section
    in_section = False
    seq = 0
    total_cost = 0.0
    for src_row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row, values_only=True):
        r0 = str(src_row[0]) if src_row[0] else ''
        if '耗材成本' in r0 and '等级认定' in r0:
            in_section = True
            continue
        if '设施设备' in r0:
            in_section = False
            continue
        if '合计' in r0 and in_section:
            for ci in range(1, 11):
                ws2.cell(row=current_row, column=ci).fill = total_fill
                ws2.cell(row=current_row, column=ci).font = bold_font
                ws2.cell(row=current_row, column=ci).border = thin_border
                ws2.cell(row=current_row, column=ci).alignment = center_align
            ws2.cell(row=current_row, column=1).value = '合计'
            ws2.cell(row=current_row, column=8).value = round(total_cost, 2)
            ws2.cell(row=current_row, column=8).number_format = money_fmt
            current_row += 1
            continue
        if in_section and r0.strip().isdigit():
            seq += 1
            unit_price = float(src_row[5]) if src_row[5] else 0
            person_cost = float(src_row[7]) if src_row[7] else 0
            total_cost += person_cost
            qty_str = str(src_row[6]).strip() if src_row[6] else ''
            remark = str(src_row[8]).strip() if src_row[8] else ''

            # Build calculation description
            calc = ''
            if '/' in qty_str:
                parts = qty_str.strip().split('/')
                if len(parts) == 2 and parts[0].strip() and parts[1].strip():
                    try:
                        denom = float(parts[1].strip())
                        calc = f'{unit_price:.2f} ÷ {denom:.0f} = {unit_price/denom:.2f}'
                    except:
                        calc = f'{unit_price:.2f} ÷ {parts[1].strip()}'
            elif qty_str:
                try:
                    qty = float(qty_str)
                    calc = f'{unit_price:.2f} × {qty} = {unit_price*qty:.2f}'
                except:
                    calc = qty_str

            row_data = [seq, src_row[1] or '', src_row[2] or '', src_row[3] or '',
                       src_row[4] or '', unit_price, qty_str, person_cost, calc, remark]
            for ci, v in enumerate(row_data, 1):
                cell = sc(ws2, current_row, ci, 'left' if ci in [2, 3, 4, 9, 10] else 'center')
                cell.value = v
                if ci in [6, 8]:
                    cell.number_format = money_fmt
            current_row += 1

    current_row += 2  # spacing

set_widths(ws2, [6, 24, 28, 18, 8, 12, 14, 14, 26, 24])

# =====================================================
# Sheet 3: 设施设备成本明细
# =====================================================
ws3 = wb.create_sheet('设施设备成本明细')

ws3.merge_cells('A1:J1')
ws3['A1'].value = '健康照护师（长期照护师）等级认定 — 设施设备成本明细测算'
ws3['A1'].font = title_font
ws3['A1'].alignment = center_align

ws3.merge_cells('A2:J2')
ws3['A2'].value = '说明：单人设施设备成本=(设备采购价÷预计服务总人次)×单次使用数量；当前采用磨损费简化模型'
ws3['A2'].font = Font(name='微软雅黑', size=9, color='666666')
ws3['A2'].alignment = center_align

current_row = 4
for level_idx, (level_name, sheet_name) in enumerate([
    ('五级/初级工', '五级（初级）'),
    ('四级/中级工', '四级（中级）'),
    ('三级/高级工', '三级（高级）')
]):
    src_ws = src[sheet_name]

    ws3.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
    ws3.cell(row=current_row, column=1).value = f'{"一二三"[level_idx]}、{level_name} — 设施设备成本明细'
    ws3.cell(row=current_row, column=1).font = h2_font
    for ci in range(1, 11):
        ws3.cell(row=current_row, column=ci).fill = sub_header_fill
    current_row += 1

    hdrs = ['序号', '设备名称', '规格', '参考品牌', '单位', '单价（元）',
            '单人分摊量', '单人成本（元）', '分摊方式', '备注']
    for ci, h in enumerate(hdrs, 1):
        ws3.cell(row=current_row, column=ci).value = h
    style_header_row(ws3, current_row, 10)
    current_row += 1

    in_section = False
    seq = 0
    total_cost = 0.0
    for src_row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row, values_only=True):
        r0 = str(src_row[0]) if src_row[0] else ''
        if '设施设备' in r0:
            in_section = True
            continue
        if '合计' in r0 and in_section:
            for ci in range(1, 11):
                ws3.cell(row=current_row, column=ci).fill = total_fill
                ws3.cell(row=current_row, column=ci).font = bold_font
                ws3.cell(row=current_row, column=ci).border = thin_border
                ws3.cell(row=current_row, column=ci).alignment = center_align
            ws3.cell(row=current_row, column=1).value = '合计'
            ws3.cell(row=current_row, column=8).value = round(total_cost, 2)
            ws3.cell(row=current_row, column=8).number_format = money_fmt
            current_row += 1
            continue
        if in_section and r0.strip().isdigit():
            seq += 1
            unit_price = float(src_row[5]) if src_row[5] else 0
            person_cost = float(src_row[7]) if src_row[7] else 0
            total_cost += person_cost
            qty_str = str(src_row[6]).strip() if src_row[6] else ''
            remark = str(src_row[8]).strip() if src_row[8] else ''

            alloc = ''
            if '磨损费' in remark:
                alloc = '磨损费分摊'
            elif '4人一组' in remark:
                alloc = '4人一组均摊'
            else:
                alloc = remark

            row_data = [seq, src_row[1] or '', src_row[2] or '', src_row[3] or '',
                       src_row[4] or '', unit_price, qty_str, person_cost, alloc, remark]
            for ci, v in enumerate(row_data, 1):
                cell = sc(ws3, current_row, ci, 'left' if ci in [2, 3, 4, 9, 10] else 'center')
                cell.value = v
                if ci in [6, 8]:
                    cell.number_format = money_fmt
            current_row += 1

    current_row += 2

set_widths(ws3, [6, 24, 28, 18, 8, 12, 14, 14, 20, 28])

# =====================================================
# Sheet 4: 测算依据与过程说明
# =====================================================
ws4 = wb.create_sheet('测算依据与过程说明')

ws4.merge_cells('A1:C1')
ws4['A1'].value = '测算依据与过程说明'
ws4['A1'].font = title_font
ws4['A1'].alignment = center_align

notes = [
    ('一、测算目的', ''),
    ('', '本表对健康照护师（长期照护师）职业技能等级认定（五级/四级/三级）的单人考核成本进行系统测算，为收费定价提供成本依据。测算覆盖耗材成本、设施设备使用费两大核心成本项，并列出其他成本项的测算方法和所需基础数据。'),
    ('', ''),
    ('二、数据来源', ''),
    ('', '1. 智护学院-长期照护师-耗材成本表-2026.05.14.xlsx —— 含五级/四级/三级逐项耗材与设施设备明细（品牌、规格、单价、用量）'),
    ('', '2. 各职业系部成本测算汇总.xlsx —— 系部提交的各职业各等级成本汇总（含健康照护师三级数据）'),
    ('', '3. 职业技能等级认定收费方案（5稿）.docx —— 现行收费方案及职业分类（B类/C类）'),
    ('', '4. 川发改价格〔2017〕472号 关于重新公布全省人力资源社会保障部门行政事业性收费的通知 —— 理论考试费+操作考核费政府定价标准'),
    ('', ''),
    ('三、测算方法与过程', ''),
    ('', ''),
    ('（一）耗材成本测算方法', ''),
    ('', '步骤1：逐项列明实操考核所需全部耗材（覆盖8个考题），含名称、规格、参考品牌、单位、单价'),
    ('', '步骤2：确定单人使用量：'),
    ('', '  - 一次性消耗品：按实际每人消耗计算（如口罩1个、注射器1支）'),
    ('', '  - 可分摊耗材：按4人一组均摊（如餐巾纸1包4人用，单人=1/4包）'),
    ('', '  - 磨损品：按预估使用寿命分摊（如病号服、浴巾按磨损费计入）'),
    ('', '步骤3：单人成本 = 单价 × 单人使用量（或 单价 ÷ 分摊人数）'),
    ('', '步骤4：所有耗材单人成本累加 = 该等级单人总耗材成本'),
    ('', ''),
    ('（二）设施设备成本测算方法', ''),
    ('', '步骤1：逐项列明实操考核所需全部设施设备，含名称、规格、采购单价'),
    ('', '步骤2：确定单人分摊方式：'),
    ('', '  - 高频共用设备：按4人一组均摊（如治疗车、计时器）'),
    ('', '  - 耐用设备：按磨损费模型（采购价 ÷ 预估服务总人次），如护理床2000元转40元/人'),
    ('', '  - 单人次使用设备：按实际使用计入（如血糖仪试纸配套）'),
    ('', '步骤3：单人成本 = 设备采购价 ÷ 预估服务人次 × 本次使用数量（或直接按磨损费标准）'),
    ('', '步骤4：所有设备单人成本累加 = 该等级单人设施设备总成本'),
    ('', ''),
    ('（三）其他成本项测算方法（待补充数据）', ''),
    ('', '人工成本 = (考评员日劳务费 × 考评员数量 + 督导员日劳务费 + 工作人员日劳务费) ÷ 单日考核人数'),
    ('', '考务平台费 = 平台年服务费 ÷ 年考核总人次'),
    ('', '视频监控费 = (监控设备购置费 ÷ 折旧年限 + 年存储成本) ÷ 年考核总人次'),
    ('', '证书制作费 = 证书工本费单价（按实际采购价）'),
    ('', '文印广告费 = (试卷印刷单价 × 考核人数 + 标识标牌制作费) ÷ 考核人数'),
    ('', '场地水电费 = (场地年租金 + 年水电费) ÷ 年使用天数 ÷ 日均考核人数'),
    ('', ''),
    ('四、测算假设与限制条件', ''),
    ('', '1. 耗材价格以2026年5月智护学院提供的参考品牌市场价为准，实际采购价可能因批量采购、品牌替换等因素浮动'),
    ('', '2. 设施设备成本当前采用"磨损费"简化模型，未采用直线折旧法（年限平均法）；如转化为会计准则下的折旧模型，需补充设备预计使用年限、残值率等参数'),
    ('', '3. 4人一组分摊模式基于智护学院现行教学编组，如实际考核编组人数变化需重新测算'),
    ('', '4. 健康照护师实操考核共8个考题，多于其他职业（4-6题），耗材种类和数量均高于同类职业'),
    ('', '5. 当前仅完成"设施设备使用费"和"耗材成本"两项的逐项测算，其余6项成本待护理学院补充基础数据后完成'),
    ('', '6. 本测算未考虑间接费用分摊（如管理费用、财务费用等）'),
    ('', ''),
    ('五、政策依据清单', ''),
    ('', '1. 川发改价格〔2017〕472号《关于重新公布全省人力资源社会保障部门行政事业性收费的通知》'),
    ('', '   — 理论考试费：初级30元/人·次、中级35元/人·次、高级40元/人·次'),
    ('', '   — 操作技能考核费：初级140元/人·次、中级190元/人·次、高级240元/人·次'),
    ('', '2. 川人社规〔2022〕9号《关于全面推进职业技能等级认定工作的通知》'),
    ('', '3. 医保发〔2025〕11号《关于做好当前长期照护师培养培训工作的通知》'),
    ('', '4. 川人社办发〔2025〕22号《关于进一步提升职业技能培训质效实施"技能照亮前程"行动的通知》'),
    ('', '5. 川人社规〔2025〕11号《四川省补贴性职业技能培训管理办法》'),
    ('', ''),
    ('六、待补充事项（需向护理学院索取）', ''),
    ('', '以下基础数据为完成全口径成本测算所必需，建议尽快收集：'),
    ('', '  ① 考评员、督导员、工作人员人数及日劳务费标准'),
    ('', '  ② 单场考核考生人数、年度考核总人次及场次安排'),
    ('', '  ③ 考务平台服务合同（年费或按次计费标准）'),
    ('', '  ④ 视频监控系统投入金额及预计使用年限'),
    ('', '  ⑤ 证书工本费实际采购价'),
    ('', '  ⑥ 试卷印刷单价及标识标牌制作费'),
    ('', '  ⑦ 场地年租金（或自有场地折旧）、年水电费、年使用天数'),
    ('', '  ⑧ 如采用折旧模型，需各设备预计使用年限及残值率'),
    ('', ''),
    ('七、关键结论与建议', ''),
    ('', ''),
    ('(1) 已测算成本汇总：', ''),
    ('', '  ┌──────────┬──────────┬──────────┬──────────┐'),
    ('', '  │  成本项   │ 五级/初级 │ 四级/中级 │ 三级/高级 │'),
    ('', '  ├──────────┼──────────┼──────────┼──────────┤'),
    ('', '  │ 政策收费   │   170.00  │   225.00  │   280.00  │'),
    ('', '  │ 考务平台   │    30.00  │    30.00  │    30.00  │'),
    ('', '  │ 设施设备   │    93.00  │   158.00  │   190.00  │'),
    ('', '  │ 耗材成本   │    87.20  │   167.28  │   280.49  │'),
    ('', '  ├──────────┼──────────┼──────────┼──────────┤'),
    ('', '  │ 已测算合计  │   380.20  │   580.28  │   780.49  │'),
    ('', '  └──────────┴──────────┴──────────┴──────────┘'),
    ('', ''),
    ('(2) 现行收费方案对比：', ''),
    ('', '  现行方案（5稿）收费标准：320 / 405 / 500 元/人（五级/四级/三级）'),
    ('', '  已测算成本（不含6项待补充成本）已达：380.20 / 580.28 / 780.49 元/人'),
    ('', '  ⚠ 差额：-60.20 / -175.28 / -280.49 元/人 —— 已测算成本已显著超过现行收费标准'),
    ('', '  若计入待补充的6项成本（人工、监控、证书、文印、场地水电），全口径成本将进一步上升'),
    ('', ''),
    ('(3) 建议：', ''),
    ('', '  a) 尽快向护理学院收集6项待补充成本的基础数据，完成全口径成本测算'),
    ('', '  b) 基于全口径成本测算结果，重新核定收费方案，确保收费覆盖合理成本'),
    ('', '  c) 对于长期照护师这一新职业，建议在收费方案中单独列明成本构成，增强定价透明度'),
    ('', '  d) 设施设备成本可进一步细化为折旧模型（年限平均法），提高测算精度和说服力'),
    ('', '  e) 建议建立成本定期复核机制（如每年更新一次耗材价格和人工标准）'),
]

for i, (label, content) in enumerate(notes):
    r = i + 4
    if label:
        cell = ws4.cell(row=r, column=1, value=label)
        if any(label.startswith(x) for x in ['一、', '二、', '三、', '四、', '五、', '六、', '七、']):
            cell.font = h2_font
        elif label.startswith('（'):
            cell.font = h3_font
        else:
            cell.font = normal_font
    if content:
        c2 = ws4.cell(row=r, column=2, value=content)
        c2.font = normal_font
        c2.alignment = Alignment(wrap_text=True, vertical='center')

set_widths(ws4, [30, 90, 15])

# =====================================================
# Save
# =====================================================
output_path = r'D:\openclaw-workspace\output\健康照护师-成本构成测算-2026.05.21.xlsx'
wb.save(output_path)
print(f'OK: {output_path}')
