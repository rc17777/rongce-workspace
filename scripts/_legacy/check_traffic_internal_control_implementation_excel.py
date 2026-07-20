from openpyxl import load_workbook
p = r'outputs\交通厅内控项目\四川省交通运输厅内控评价监督检查实施方案及落地资料清单.xlsx'
wb = load_workbook(p, data_only=False)
print('sheets:', wb.sheetnames)
for s in wb.sheetnames:
    ws = wb[s]
    print(s, ws.max_row, ws.max_column)
print(p)
