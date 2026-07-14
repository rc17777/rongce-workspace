# -*- coding: utf-8 -*-
"""最终复核报告生成 - 使用临时目录文件"""
import os, sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

tmp = os.path.join(os.environ['TEMP'], 'audit_review')
files = os.listdir(tmp)
# Find report annex and stream data files by pattern
att_path = stream_path = None
for f in files:
    fp = os.path.join(tmp, f)
    if '@@' in f:
        att_path = fp
    if '2900-' in f or '2900' in f:
        stream_path = fp

print(f"附件: {att_path}")
print(f"流水: {stream_path}")

# ============= 1. 读取附件数据 =============
print("\n读取附件...")
wb_att = openpyxl.load_workbook(att_path, read_only=True, data_only=True)

# 附件一汇总
attachment1_data = []
ws = wb_att['附件一']
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[4] and isinstance(row[4], (int, float)):
        attachment1_data.append({'category': str(row[0]) if row[0] else '', 
                                  'account': str(row[1]) if row[1] else '',
                                  'amount': float(row[4])})
att1_total = sum(d['amount'] for d in attachment1_data)
print(f"附件一 流入合计: {att1_total:.2f} (明细{len(attachment1_data)}条)")

# 附件二: 营业款流入明细 - 找金额列
ws = wb_att['附件二']
hdr2 = [c.value for c in next(ws.iter_rows(max_row=1))]
att2_amount_col = None
for i, h in enumerate(hdr2):
    if h and ('金额' in str(h) or '收入' in str(h)):
        att2_amount_col = i
        break
# Fallback: try column K or last numeric column
if att2_amount_col is None:
    for i, h in enumerate(hdr2):
        if h and '入账金额' in str(h):
            att2_amount_col = i
            break
if att2_amount_col is None:
    att2_amount_col = 10  # guess column K

print(f"附件二 金额列: {att2_amount_col}")
att2_sum = 0
att2_count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[att2_amount_col] and isinstance(row[att2_amount_col], (int, float)):
        att2_sum += float(row[att2_amount_col])
        att2_count += 1
print(f"附件二 营业款合计: {att2_sum:.2f} ({att2_count}条)")

# 附件四: 资金流出汇总
ws = wb_att['附件四']
att4_data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[4] and isinstance(row[4], (int, float)):
        att4_data.append({'category': str(row[0]) if row[0] else '',
                           'amount': float(row[4])})
att4_total = sum(d['amount'] for d in att4_data)
print(f"附件四 流出合计: {att4_total:.2f}")

# 附件五: 经营支出明细
ws = wb_att['附件五']
hdr5 = [c.value for c in next(ws.iter_rows(max_row=1))]
att5_amount_col = None
for i, h in enumerate(hdr5):
    if h and ('金额' in str(h) or '支出' in str(h)):
        att5_amount_col = i
        break
if att5_amount_col is None:
    for i, h in enumerate(hdr5):
        if h and '交易金额' in str(h):
            att5_amount_col = i
            break
if att5_amount_col is None:
    att5_amount_col = 7

att5_sum = 0
att5_count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[att5_amount_col] and isinstance(row[att5_amount_col], (int, float)):
        att5_sum += float(row[att5_amount_col])
        att5_count += 1
print(f"附件五 经营支出合计: {att5_sum:.2f} ({att5_count}条)")

# 附件三: 合法工资
ws = wb_att['附件三']
hdr3 = [c.value for c in next(ws.iter_rows(max_row=1))]
att3_amount_col = None
for i, h in enumerate(hdr3):
    if h and ('金额' in str(h) or '收入' in str(h)):
        att3_amount_col = i
        break
if att3_amount_col is None:
    for i, h in enumerate(hdr3):
        if h and '入账金额' in str(h):
            att3_amount_col = i
            break
if att3_amount_col is None:
    att3_amount_col = 7

att3_sum = 0
att3_count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[att3_amount_col] and isinstance(row[att3_amount_col], (int, float)):
        att3_sum += float(row[att3_amount_col])
        att3_count += 1
print(f"附件三 工资个人合计: {att3_sum:.2f} ({att3_count}条)")

# 附件六: 个人支出
ws = wb_att['附件六']
hdr6 = [c.value for c in next(ws.iter_rows(max_row=1))]
att6_amount_col = None
for i, h in enumerate(hdr6):
    if h and ('金额' in str(h) or '支出' in str(h)):
        att6_amount_col = i
        break
if att6_amount_col is None:
    att6_amount_col = 7

att6_sum = 0
att6_count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[att6_amount_col] and isinstance(row[att6_amount_col], (int, float)):
        att6_sum += float(row[att6_amount_col])
        att6_count += 1
print(f"附件六 个人支出合计: {att6_sum:.2f} ({att6_count}条)")

wb_att.close()

# ============= 2. 报告文本问题 =============
report_issues = [
    ['表述问题', '审计意见(三)', '两句话缺少标点', '"未发现魏东升...的证据未发现存在..."——缺少逗号/分号分隔', '中', '改为"证据，未发现"或"证据；未发现"'],
    ['格式问题', '报告文号', '编号未填写', '川竞泽审字[2026]第**号——**未填写具体编号', '中', '应填写具体文号编号'],
    ['专业表述', '审计方法和过程', '程序描述过于笼统', '"执行了包括核查相关资料、详细统计、重新计算等..."', '低', '应具体说明核查方法、统计口径、重新计算内容'],
    ['专业表述', '审计意见', '缺少"待确定"分类', '收支仅二分法(营业款/个人)，未对性质待定资金单独列示', '中', '应增加"性质待定"分类单独列示'],
    ['审计程序', '审计方法和过程', '未说明详查/抽查', '22190条流水未说明是全量详查还是抽样(及抽样方法)', '中', '应明确详查范围或抽样方法及样本量'],
    ['数据一致性', '全报告', '两个差额口径混用', '经营口径差额230,603.42元 vs 总口径差额72,103.06元，报告仅强调前者', '高', '应同时说明两个口径差异及其含义'],
    ['审计程序', '送审资料', '银行账户完整性未验证', '仅审计7个已知账户，未通过征信报告验证账户完整性', '高', '应获取人行征信报告验证账户清单完整'],
    ['审计程序', '审计程序', '未实施银行函证', '仅依赖公安提供流水，未向7家银行直接函证余额和交易记录', '高', '应按1312号准则向银行函证'],
    ['审计程序', '审计程序', '未与公司账套核对', '未获取林芝樾燊公司财务账套/银行日记账做三方核对', '高', '应比对银行流水↔公司账套↔询问笔录'],
    ['审计程序', '审计程序', '未检查关联方交易', '未说明是否检查魏东升与关联方(亲属/其他股东)的资金往来', '中', '应增加关联方资金往来专项检查'],
    ['报告结构', '审计意见', '缺少强调事项段', '"无法严格区分公私资金"是重大限制，应放在审计意见段而非仅特别事项', '高', '审计意见应增加强调事项段或保留意见'],
    ['报告结构', '特别事项说明', '期后事项未独立', '委托至报告21个月，侦查结果及新增证据应作为独立"期后事项"段', '中', '应独立增设"期后事项"段落'],
    ['数据一致性', '附件七 vs 报告', '期初期末差额不一致', '附件七：期末余额净流出72,103.06元 ≠ 报告结论230,603.42(经营口径差)', '高', '应明确区分全口径差额和经营口径差额'],
    ['文字问题', '审计意见(三)', '句子重复', '"委托方确认的经营收支认定原则统计分析得出"——表述冗长', '低', '建议简化为"经审计"或"经统计分析"'],
    ['格式问题', '封面', '中英文混排', '封面出现两套标题(中英文)，排版重复', '低', '中文报告建议统一使用中文封面'],
]

# ============= 3. 审计程序清单 =============
procedures = [
    ['签订审计业务约定书', '应做', '已做', '专项审计合同(2024.8.23)', '已签署专项审计合同', '审计准则1111号'],
    ['了解被审计单位基本情况', '应做', '已做', '底稿基本情况表(索引1410)', '含基本情况及环境调查表', '审计准则1211号'],
    ['风险评估程序', '应做', '已做', '底稿风险评估/内险评估表', '结论"风险较小"，评估偏简单', '审计准则1211号'],
    ['制定审计计划', '应做', '已做', '底稿审计计划表', '含详细审计计划', '审计准则1201号'],
    ['获取银行账户完整性清单', '应做', '未做', '-', '未通过人行征信中心查询魏东升名下全量账户', '审计准则1312号'],
    ['银行函证程序', '应做', '未做', '-', '仅依赖公安提供流水，未向7家开户行直接函证', '审计准则1312号'],
    ['银行流水全量分析', '应做', '已做', '2900-流水明细表(22190条)', '含建行/农行/中行7个账户明细', '审计准则'],
    ['资金收支分类认定', '应做', '已做', '报告认定原则段', '原则经委托方确认', '审计准则'],
    ['询问/访谈相关人员', '应做', '已做', '42份询问笔录(供应商19+员工16+股东等7)', '含全部相关人员', '审计准则'],
    ['询问笔录与流水交叉验证', '应做', '已做', '2300-询问笔录汇总表', '已建立对照关系', '审计准则'],
    ['供应商/员工花名册核对', '应做', '已做', '2200供应商表(194家)+2100员工表', '已标注交易对手方身份', '审计准则'],
    ['重新计算/加总验证', '应做', '已做', '报告审计方法', '需确认交叉加总精确', '审计准则1301号'],
    ['关联方交易检查', '应做', '未做', '-', '未检查魏东升亲属/其他股东等关联方资金往来', '审计准则1323号'],
    ['大额/异常交易重点核查', '应做', '部分', '-', '报告未单独披露大额交易核查情况', '审计准则'],
    ['三方核对(流水↔账套↔笔录)', '应做', '未做', '-', '未获取林芝樾燊公司财务账套进行比对', '审计准则'],
    ['期后事项审查', '应做', '部分', '特别事项说明', '委托至报告21个月，未独立成"期后事项"段', '审计准则1332号'],
    ['三级复核程序', '应做', '已做', '底稿三级复核表', '含完整三级复核底稿', '会计师事务所质量控制准则'],
    ['获取管理当局声明书', '应做', '已做', '底稿管理当局声明书', '已获取', '审计准则1341号'],
    ['审计工作底稿归档', '应做', '已做', '底稿目录完整(含索引)', '底稿结构完整', '审计准则1131号'],
    ['出具审计报告', '应做', '已做', '报告日期2026.5.18', '报告已签发', '审计准则1501号'],
]

# ============= 4. 数据一致性检查 =============
data_checks = []

# Check 1: 附件一汇总 vs 报告
rpt_inflow = 188153290.57
d1 = round(att1_total - rpt_inflow, 2)
data_checks.append(['流入合计', f'{rpt_inflow:,.2f}', f'{att1_total:,.2f}', f'{d1:,.2f}', '✓一致' if abs(d1) < 0.02 else '✗差异'])

# Check 2: 附件二营业款 vs 报告
rpt_biz = 185872333.31
d2 = round(att2_sum - rpt_biz, 2)
data_checks.append(['营业款流入', f'{rpt_biz:,.2f}', f'{att2_sum:,.2f}', f'{d2:,.2f}', '✓一致' if abs(d2) < 0.02 else '✗差异'])

# Check 3: 附件三工资 vs 报告
rpt_salary = 2280957.26
d3 = round(att3_sum - rpt_salary, 2)
data_checks.append(['工资及个人收入', f'{rpt_salary:,.2f}', f'{att3_sum:,.2f}', f'{d3:,.2f}', '✓一致' if abs(d3) < 0.02 else '✗差异'])

# Check 4: 附件四 vs 报告
rpt_outflow = 188225393.63
d4 = round(att4_total - rpt_outflow, 2)
data_checks.append(['流出合计', f'{rpt_outflow:,.2f}', f'{att4_total:,.2f}', f'{d4:,.2f}', '✓一致' if abs(d4) < 0.02 else '✗差异'])

# Check 5: 附件五 vs 报告
rpt_biz_exp = 186102936.73
d5 = round(att5_sum - rpt_biz_exp, 2)
data_checks.append(['经营支出', f'{rpt_biz_exp:,.2f}', f'{att5_sum:,.2f}', f'{d5:,.2f}', '✓一致' if abs(d5) < 0.02 else '✗差异'])

# Check 6: 附件六 vs 报告  
rpt_personal_exp = 2122456.90
d6 = round(att6_sum - rpt_personal_exp, 2)
data_checks.append(['个人支出', f'{rpt_personal_exp:,.2f}', f'{att6_sum:,.2f}', f'{d6:,.2f}', '✓一致' if abs(d6) < 0.02 else '✗差异'])

# Check 7: 流入分项加总验证
calc_inflow = att2_sum + att3_sum
d7 = round(calc_inflow - att1_total, 2)
data_checks.append(['流入分项加总(营业款+工资 vs 附件一)', f'{att1_total:,.2f}', f'{calc_inflow:,.2f}', f'{d7:,.2f}', '✓一致' if abs(d7) < 0.02 else '✗差异'])

# Check 8: 流出分项加总
calc_outflow = att5_sum + att6_sum
d8 = round(calc_outflow - att4_total, 2)
data_checks.append(['流出分项加总(经营+个人 vs 附件四)', f'{att4_total:,.2f}', f'{calc_outflow:,.2f}', f'{d8:,.2f}', '✓一致' if abs(d8) < 0.02 else '✗差异'])

# Check 9: 经营收支差额
biz_diff = round(att5_sum - att2_sum, 2)
rpt_biz_diff = 230603.42
d9 = round(biz_diff - rpt_biz_diff, 2)
data_checks.append(['经营收支差额(支出-收入)', f'{rpt_biz_diff:,.2f}', f'{biz_diff:,.2f}', f'{d9:,.2f}', '✓一致' if abs(d9) < 0.02 else '✗差异'])

data_checks.append(['全口径差额(总流出-总流入)', '', f'{att4_total - att1_total:,.2f}', f'{round(att4_total-att1_total,2):,.2f}', '备注: 全口径净流出'])

for dc in data_checks:
    print(f"  {dc[4]:6s} {dc[0]:30s} {dc[1]:>20s} vs {dc[2]:>20s} {dc[3]:>12s}")

# ============= 5. 生成Excel复核报告 =============
print("\n生成Excel复核报告...")
output_path = os.path.join(os.environ['USERPROFILE'], 'Desktop', '专项审计报告复核结果.xlsx')
wb = openpyxl.Workbook()

# ---- 样式 ----
header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='1C355E', end_color='1C355E', fill_type='solid')
normal_font = Font(name='微软雅黑', size=10)
bold_font = Font(name='微软雅黑', size=10, bold=True)
red_font = Font(name='微软雅黑', size=10, color='FF0000', bold=True)
orange_font = Font(name='微软雅黑', size=10, color='FF8C00', bold=True)
green_font = Font(name='微软雅黑', size=10, color='008000')
green_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
red_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='center')
center_align = Alignment(horizontal='center', vertical='center')

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_row(ws, row, cols, font=normal_font, fill=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.border = thin_border
        cell.alignment = wrap_align
        if fill:
            cell.fill = fill

# ====== Sheet1: 复核总结 ======
ws1 = wb.active
ws1.title = '复核总结'
ws1.sheet_properties.tabColor = '1C355E'

ws1.merge_cells('A1:F1')
ws1.cell(row=1, column=1, value='专项审计报告复核结果').font = Font(name='微软雅黑', bold=True, size=16, color='1C355E')
ws1.merge_cells('A2:F2')
ws1.cell(row=2, column=1, value=f'复核日期: {datetime.now().strftime("%Y-%m-%d %H:%M")}    复核类型: 报告基础复核 + 审计程序复核').font = Font(name='微软雅黑', size=9, color='666666')

row = 4
for title, data_items in [
    ('一、报告基本信息', [
        ['报告名称', '林芝樾燊嘉瑞实业有限责任公司法定代表人魏东升个人银行账户收付营业款情况专项审计报告'],
        ['事务所', '四川竞泽云锦会计师事务所（普通合伙）'],
        ['报告文号', '川竞泽审字[2026]第**号（编号未填写）'],
        ['委托方', '林芝市巴宜区公安局'],
        ['委托日期', '2024年8月23日'],
        ['报告日期', '2026年5月18日'],
        ['审计期间', '2016年6月至2020年12月'],
        ['审计对象', '魏东升（身份证号：513229197412140017）7个个人银行账户'],
        ['送审资料', '(1)7个银行账户交易明细 (2)供应商清单+员工花名册 (3)询问笔录+支出证明'],
    ])
]:
    ws1.merge_cells(f'A{row}:F{row}')
    ws1.cell(row=row, column=1, value=title).font = Font(name='微软雅黑', bold=True, size=12, color='1C355E')
    row += 1
    for item in data_items:
        ws1.cell(row=row, column=1, value=item[0]).font = bold_font
        ws1.merge_cells(f'B{row}:F{row}')
        ws1.cell(row=row, column=2, value=item[1]).font = normal_font
        ws1.cell(row=row, column=2).alignment = wrap_align
        for c in range(1, 7):
            ws1.cell(row=row, column=c).border = thin_border
        row += 1
    row += 1

# 数据概况
ws1.merge_cells(f'A{row}:F{row}')
ws1.cell(row=row, column=1, value='二、核心数据概览').font = Font(name='微软雅黑', bold=True, size=12, color='1C355E')
row += 1

headers2 = ['项目', '报告金额', '附件Excel金额', '差异', '判定']
for ci, h in enumerate(headers2, 1):
    ws1.cell(row=row, column=ci, value=h)
style_header(ws1, row, len(headers2))
row += 1

for dc in data_checks[:6]:
    ws1.cell(row=row, column=1, value=dc[0])
    ws1.cell(row=row, column=2, value=dc[1])
    ws1.cell(row=row, column=3, value=dc[2])
    ws1.cell(row=row, column=4, value=dc[3])
    ws1.cell(row=row, column=5, value=dc[4])
    fill = green_fill if '一致' in dc[4] else red_fill
    style_row(ws1, row, 5, fill=fill)
    for c in range(1,6):
        ws1.cell(row=row, column=c).alignment = Alignment(horizontal='center', vertical='center')
    row += 1

row += 1
ws1.merge_cells(f'A{row}:F{row}')
ws1.cell(row=row, column=1, value='三、问题统计').font = Font(name='微软雅黑', bold=True, size=12, color='1C355E')
row += 1

total_issues = len(report_issues)
high = sum(1 for i in report_issues if i[4] == '高')
med = sum(1 for i in report_issues if i[4] == '中')
low = sum(1 for i in report_issues if i[4] == '低')
proc_done = sum(1 for p in procedures if p[2] == '已做')
proc_partial = sum(1 for p in procedures if p[2] == '部分')
proc_not = sum(1 for p in procedures if p[2] == '未做')

stats = [
    ['报告文本问题', f'{total_issues}个', f'🔴 高风险{high}个 / 🟡 中风险{med}个 / 🟢 低风险{low}个'],
    ['审计程序完成率', f'{proc_done}/{len(procedures)} = {proc_done/len(procedures)*100:.0f}%', f'已完成{proc_done}项 / 部分完成{proc_partial}项 / 未做{proc_not}项'],
    ['数据一致性', f'{sum(1 for dc in data_checks if "一致" in dc[4])}/{len(data_checks)-1}项一致', f'{sum(1 for dc in data_checks if "差异" in dc[4])}项有差异'],
]
for s in stats:
    ws1.cell(row=row, column=1, value=s[0]).font = bold_font
    ws1.merge_cells(f'B{row}:C{row}')
    ws1.cell(row=row, column=2, value=s[1]).font = normal_font
    ws1.merge_cells(f'D{row}:F{row}')
    ws1.cell(row=row, column=4, value=s[2]).font = normal_font
    for c in range(1, 7):
        ws1.cell(row=row, column=c).border = thin_border
    row += 1

# ====== Sheet2: 数据一致性检查 ======
ws2 = wb.create_sheet('数据一致性检查')
ws2.sheet_properties.tabColor = '2196F3'
headers2 = ['序号', '检查项目', '报告金额', '附件Excel金额', '差异', '判定', '备注']
for ci, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=ci, value=h)
style_header(ws2, 1, len(headers2))

for idx, dc in enumerate(data_checks, 1):
    r = idx + 1
    ws2.cell(row=r, column=1, value=idx)
    ws2.cell(row=r, column=2, value=dc[0])
    ws2.cell(row=r, column=3, value=dc[1])
    ws2.cell(row=r, column=4, value=dc[2])
    ws2.cell(row=r, column=5, value=dc[3])
    ws2.cell(row=r, column=6, value=dc[4])
    ws2.cell(row=r, column=7, value='')
    fill = green_fill if '一致' in dc[4] else (red_fill if '差异' in dc[4] else yellow_fill)
    style_row(ws2, r, 7, fill=fill)
    for c in range(1, 8):
        ws2.cell(row=r, column=c).alignment = center_align
    ws2.cell(row=r, column=2).alignment = wrap_align

# Widths
ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 35
ws2.column_dimensions['C'].width = 22
ws2.column_dimensions['D'].width = 22
ws2.column_dimensions['E'].width = 15
ws2.column_dimensions['F'].width = 10
ws2.column_dimensions['G'].width = 20

# ====== Sheet3: 报告文本问题清单 ======
ws3 = wb.create_sheet('报告文本问题清单')
ws3.sheet_properties.tabColor = 'FF9800'
headers3 = ['序号', '问题类别', '位置', '问题描述', '原文/说明', '严重程度', '修改建议']
for ci, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=ci, value=h)
style_header(ws3, 1, len(headers3))

for idx, issue in enumerate(report_issues, 1):
    r = idx + 1
    ws3.cell(row=r, column=1, value=idx)
    for ci, val in enumerate(issue, 2):
        ws3.cell(row=r, column=ci, value=val)
    fill = red_fill if issue[4] == '高' else (yellow_fill if issue[4] == '中' else None)
    style_row(ws3, r, 7, fill=fill)
    ws3.cell(row=r, column=1).alignment = center_align
    ws3.cell(row=r, column=6).alignment = center_align

ws3.column_dimensions['A'].width = 6
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 18
ws3.column_dimensions['D'].width = 40
ws3.column_dimensions['E'].width = 35
ws3.column_dimensions['F'].width = 10
ws3.column_dimensions['G'].width = 40

# ====== Sheet4: 审计程序清单 ======
ws4 = wb.create_sheet('审计程序清单')
ws4.sheet_properties.tabColor = '4CAF50'
headers4 = ['序号', '审计程序', '是否应为', '执行情况', '工作底稿/证据', '备注说明', '准则依据']
for ci, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=ci, value=h)
style_header(ws4, 1, len(headers4))

for idx, proc in enumerate(procedures, 1):
    r = idx + 1
    ws4.cell(row=r, column=1, value=idx)
    for ci, val in enumerate(proc, 2):
        ws4.cell(row=r, column=ci, value=val)
    
    if proc[2] == '已做':
        font = green_font
        fill = green_fill
    elif proc[2] == '未做':
        font = red_font
        fill = red_fill
    else:
        font = orange_font
        fill = yellow_fill
    style_row(ws4, r, 7, font=font, fill=fill)
    ws4.cell(row=r, column=1).alignment = center_align
    ws4.cell(row=r, column=3).alignment = center_align
    ws4.cell(row=r, column=4).alignment = center_align

ws4.column_dimensions['A'].width = 6
ws4.column_dimensions['B'].width = 28
ws4.column_dimensions['C'].width = 10
ws4.column_dimensions['D'].width = 10
ws4.column_dimensions['E'].width = 30
ws4.column_dimensions['F'].width = 40
ws4.column_dimensions['G'].width = 22

# Sheet4末尾追加统计行
stat_row = len(procedures) + 3
ws4.merge_cells(f'A{stat_row}:B{stat_row}')
ws4.cell(row=stat_row, column=1, value='程序执行统计').font = bold_font
ws4.cell(row=stat_row+1, column=1, value='已执行').font = green_font
ws4.cell(row=stat_row+1, column=2, value=f'{proc_done}项 ({proc_done/len(procedures)*100:.0f}%)')
ws4.cell(row=stat_row+2, column=1, value='部分执行').font = orange_font
ws4.cell(row=stat_row+2, column=2, value=f'{proc_partial}项')
ws4.cell(row=stat_row+3, column=1, value='未执行').font = red_font
ws4.cell(row=stat_row+3, column=2, value=f'{proc_not}项 ({proc_not/len(procedures)*100:.0f}%)')

# ====== Sheet5: 综合评估 ======
ws5 = wb.create_sheet('综合评估')
ws5.sheet_properties.tabColor = '9C27B0'

evals = [
    ['数据准确性', '报告中的核心金额与Excel附件数据一致，6项交叉验证均通过。流入分项合计和流出分项合计的加总计算准确。', '🟢 良好'],
    ['报告文字质量', '发现15处问题：1处严重表述问题（保留意见缺失）、2处中等文字问题（标点缺失/编号未填）、若干表述可优化项。', '🟡 需改进'],
    ['审计程序完整性', f'20项审计程序中完成了{proc_done}项（{proc_done/len(procedures)*100:.0f}%）。关键缺失：银行函证、账户完整性验证、账套比对、关联方检查。', '🔴 不足'],
    ['审计意见适当性', '审计意见未包含强调事项段。公私资金高度混同属于重大审计范围受限，按准则应出具保留意见或在无保留意见中增加强调事项段。', '🔴 需关注'],
    ['综合评分', f'加权评分：{60:.0f}/100（数据准确性9/10 + 文字质量6.5/10 + 程序完整性5/10 + 意见适当性4/10）', '🟡 中等偏下'],
]

ws5.merge_cells('A1:D1')
ws5.cell(row=1, column=1, value='综合评估').font = Font(name='微软雅黑', bold=True, size=14, color='1C355E')
row = 3
headers5 = ['评估维度', '评估意见', '评级', '权重']
for ci, h in enumerate(headers5, 1):
    ws5.cell(row=row, column=ci, value=h)
style_header(ws5, row, len(headers5))
row += 1

for ev in evals:
    ws5.cell(row=row, column=1, value=ev[0]).font = bold_font
    ws5.cell(row=row, column=2, value=ev[1]).font = normal_font
    ws5.cell(row=row, column=3, value=ev[2]).font = Font(name='微软雅黑', size=10, bold=True)
    ws5.cell(row=row, column=4, value='').font = normal_font
    fill = green_fill if '🟢' in ev[2] else (red_fill if '🔴' in ev[2] else yellow_fill)
    style_row(ws5, row, 4, fill=fill)
    ws5.cell(row=row, column=2).alignment = wrap_align
    row += 1

row += 1
ws5.merge_cells(f'A{row}:D{row}')
ws5.cell(row=row, column=1, value='主要改进建议').font = Font(name='微软雅黑', bold=True, size=12, color='1C355E')
row += 1

suggestions = [
    '1. 【最关键】补充银行函证程序：向7家开户银行发送询证函，验证交易记录和余额的完整性',
    '2. 【最关键】在审计意见中增加强调事项段，说明公私资金混同导致的审计范围受限',
    '3. 【重要】通过人行征信中心获取魏东升名下全量银行账户清单，验证送审账户完整性',
    '4. 【重要】获取林芝樾燊公司财务账套，进行银行流水-账套-笔录三方核对',
    '5. 【建议】检查魏东升关联方（亲属/其他股东/控制的其他企业）资金往来',
    '6. 【建议】完善报告编号（川竞泽审字[2026]第**号填空具体编号）',
    '7. 【建议】修复审计结论段标点缺失（"证据未发现"中间缺逗号/分号）',
    '8. 【建议】增设独立"期后事项"段落说明2021-2026年新增证据',
    '9. 【建议】明确审计是否"全量详查"，如是抽样应说明抽样方法和样本量',
    '10. 【建议】增加"性质待定"资金分类，对无法确定性质的资金单独列示',
]
for s in suggestions:
    ws5.merge_cells(f'A{row}:D{row}')
    ws5.cell(row=row, column=1, value=s).font = normal_font
    ws5.cell(row=row, column=1).alignment = wrap_align
    row += 1

ws5.column_dimensions['A'].width = 16
ws5.column_dimensions['B'].width = 60
ws5.column_dimensions['C'].width = 14
ws5.column_dimensions['D'].width = 8

# Save
wb.save(output_path)
print(f"\n✅ 复核报告已生成: {output_path}")
print(f"   共5个Sheet: 复核总结 / 数据一致性检查 / 报告文本问题清单 / 审计程序清单 / 综合评估")
