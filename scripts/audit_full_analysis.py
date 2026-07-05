import openpyxl
import json
from collections import Counter, defaultdict

PATH = r'C:\Users\scrccpa\Desktop\融策审计过程记录系统=项目经理版(6).xlsx'
wb = openpyxl.load_workbook(PATH, data_only=True)

ws = wb['2-审计过程']
headers = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]

# Collect all data
rows = []
for row in ws.iter_rows(min_row=3, values_only=True):
    vals = [str(v) if v is not None else '' for v in row]
    rows.append(vals)

# Convert date from serial to readable
def serial_to_date(s):
    try:
        from datetime import datetime, timedelta
        serial = int(s)
        base = datetime(1899, 12, 30)
        dt = base + timedelta(days=serial)
        return dt.strftime('%Y-%m-%d')
    except:
        return s

# Basic stats
total_records = len(rows)
dates = [serial_to_date(r[1]) for r in rows if r[1].strip()]
subprojects = [r[3] for r in rows if r[3].strip()]
procedures = [r[2] for r in rows if r[2].strip()]
anomalies = [r[8] for r in rows if r[8].strip() == '是']

with open(r'D:\openclaw-workspace\output\audit_analysis.txt', 'w', encoding='utf-8') as f:
    f.write(f"=== 融策审计过程记录系统 分析报告 ===\n\n")
    f.write(f"总记录数: {total_records}\n")
    f.write(f"日期跨度: {dates[0] if dates else 'N/A'} ~ {dates[-1] if dates else 'N/A'}\n\n")
    
    f.write("--- 子项目统计 ---\n")
    sp_count = Counter(subprojects)
    for sp, cnt in sp_count.most_common():
        f.write(f"  {sp}: {cnt}条\n")
    
    f.write(f"\n--- 审计程序分布 ---\n")
    proc_count = Counter(procedures)
    for p, cnt in proc_count.most_common():
        f.write(f"  {p}: {cnt}条\n")
    
    f.write(f"\n--- 异常标记统计 ---\n")
    f.write(f"  标记为异常的记录: {len(anomalies)}条\n")
    
    # Records with findings
    findings = [r for r in rows if r[6].strip()]
    f.write(f"  有审计发现的记录: {len(findings)}条\n")
    
    # Records with conclusions
    conclusions = [r for r in rows if r[7].strip()]
    f.write(f"  有审计判断/结论的记录: {len(conclusions)}条\n\n")
    
    f.write("--- 有审计发现的记录详情 ---\n")
    for i, r in enumerate(findings):
        f.write(f"\n[记录{r[0]}] {serial_to_date(r[1])} | {r[3]} | {r[2]}\n")
        f.write(f"  发现: {r[6]}\n")
        if r[7]:
            f.write(f"  结论: {r[7]}\n")
    
    f.write(f"\n\n--- 所有记录列表 ---\n")
    for r in rows:
        f.write(f"[{r[0]}] {serial_to_date(r[1])} | {r[3]} | {r[2]} | 异常={r[8]}\n")

print(f"分析完成，共{total_records}条记录，报告已写入 output/audit_analysis.txt")
