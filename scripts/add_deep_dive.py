# -*- coding: utf-8 -*-
"""Add deep-dive dimension sheets to existing workbook"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

src_path = r'D:\openclaw-workspace\projects\护理学院任中经责审计\护理学院任中经责审计_制度执行对照分析.xlsx'
wb = openpyxl.load_workbook(src_path)

# ===== Styles =====
header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
sub_header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
section_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
orange_fill = PatternFill(start_color='F4B183', end_color='F4B183', fill_type='solid')
title_font = Font(name='微软雅黑', bold=True, size=14, color='1F4E79')
section_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
normal_font = Font(name='微软雅黑', size=10)
bold_font = Font(name='微软雅黑', bold=True, size=10)
red_font = Font(name='微软雅黑', bold=True, size=10, color='9C0006')
green_font = Font(name='微软雅黑', bold=True, size=10, color='006100')
border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                top=Side(style='thin'), bottom=Side(style='thin'))
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header(ws, row, cols, fill=header_fill, font=header_font):
    for c in range(1, cols+1):
        cell = ws.cell(row, c)
        cell.fill = fill
        cell.font = font
        cell.alignment = center_align
        cell.border = border

def style_row(ws, row, cols, risk=None):
    for c in range(1, cols+1):
        cell = ws.cell(row, c)
        cell.font = normal_font
        cell.alignment = left_align if c > 2 else center_align
        cell.border = border
    if risk == 'red':
        for c in range(1, cols+1):
            ws.cell(row, c).fill = red_fill
    elif risk == 'yellow':
        for c in range(1, cols+1):
            ws.cell(row, c).fill = yellow_fill
    elif risk == 'green':
        for c in range(1, cols+1):
            ws.cell(row, c).fill = green_fill
    elif risk == 'orange':
        for c in range(1, cols+1):
            ws.cell(row, c).fill = orange_fill

# ==========================================
# Sheet A: 政府采购管理深度分析
# ==========================================
ws_a = wb.create_sheet('深析-政府采购管理')
ws_a.merge_cells('A1:L1')
ws_a['A1'] = '维度深析：政府采购管理 — 制度链 × 执行链 × 风险链 全链路分析'
ws_a['A1'].font = title_font
ws_a['A1'].alignment = center_align

# A1: 全生命周期节点分析
ws_a.merge_cells('A3:L3')
ws_a['A3'] = '一、采购全生命周期节点——制度规定 vs 实际执行（8个问题逐节点定位）'
ws_a['A3'].font = bold_font
ws_a['A3'].fill = sub_header_fill

headers_a = ['生命周期\n节点', '子环节', '制度依据', '应然标准', '实然表现', '偏差\n类型', '涉及\n问题号', '涉及金额\n(万元)', '根因分析', '风险传导', '严重度', '改进措施']
for j, h in enumerate(headers_a, 1):
    ws_a.cell(4, j, h)
style_header(ws_a, 4, len(headers_a))

lifecycle = [
    # 预算与计划
    ['1.预算\n与计划', '采购预算编制', '预算管理办法\n采购办法', '应编尽编\n无预算不采购', '基本合规\n但199项/1.28亿规模下\n需求论证充分性待查', '🟡', '—', '—', '规模激增→论证压力\n未发现超预算采购但\n需求合理性审查不足', '配置超标\n预算浪费', '🟡', '强化需求论证模板\n建立超预算预警'],
    
    # 招标文件
    ['2.招标\n文件编制', '文件编制', '采购办法(2025)\n第七条(三)', '编制招标文件\n组织采购文件确认', '物管项目(650万)两次流标\n云桌面项目中小企业政策\n评标现场临时修改', '🔴', '#1\n#7', '879', '招标文件审核流程薄弱\n缺少多级复核机制\n核心条款质量不稳定', '流标→延误工期\n现场修改→公信力受损\n质疑增加', '🔴', '建立三级审核制度\n经办人→科长→处长\n重要项目加法务复核'],
    ['', '评分标准', '采购办法(2025)', '评分标准合理可操作\n分值设置科学', '云桌面项目演示分10分\n4家投标人均0分\n评分项形同虚设', '🔴', '#7', '229', '评分标准脱离市场实际\n未做投标人能力摸底\n评分项无区分度', '评审无法有效\n择优选择', '🔴', '评分表预审制度\n投标人能力预判\n区分度不足的评分项\n降低分值或取消'],
    ['', '政策适用', '内控办法(2021)', '中小企业政策\n在招标文件中\n事先明确', '云桌面项目中小企业政策\n评标现场才修改\n从"专门面向"改为"非专门"', '🔴', '#7', '229', '编制阶段政策适用错误\n评标现场仓促修改\n影响投标策略公平性', '投标人无法\n及时调整策略\n可能引发投诉', '🔴', '政策适用审查加入\n招标文件审核清单\n开标前24h锁定'],
    
    # 采购人代表
    ['3.采购人\n代表', '代表选派', '采购人代表\n管理办法(2025)', '按规定选派\n到现场监督', '家具采购(112万)\n监督代表肖梁颖\n未见签到/监督报告签名', '🔴', '#2', '111.68', '监督代表是否实际到场成疑\n可能是:①未到场②到场未签\n③签到表遗失\n均属管理漏洞', '监督缺位→\n评审公正性存疑\n→采购结果效力\n可能受影响', '🔴', '现场签到电子化\n监督报告当日提交\n迟到/缺席即时替补'],
    
    # 评审
    ['4.评审\n与质疑', '评审质量', '内控办法(2021)\n第六条', '严格评审\n合规高效', '投影仪采购(89万)\n4个质疑中3个成立\n→合格供应商不足3家\n→重新采购', '🔴', '#8', '88.95', '质疑成立率75%远超正常\n评审专家专业能力\n或独立性存疑\n代理机构选择可能不当', '重新采购→\n时间+费用损失\n采购公信力下降', '🔴', '评审专家动态考核\n高质疑率代理机构\n暂停委托资格'],
    ['', '质疑处理', '内控办法(2021)', '配合质疑处理\n及时回应', '4个质疑3个成立\n承认了评审疏漏', '🟡', '#8', '88.95', '承认质疑说明有改进意愿\n但"承认≠改进"\n同类问题需预防', '如不改进→\n持续质疑→\n重复采购', '🟡', '质疑处理月度分析\n发现共性问题→\n制度性预防'],
    
    # 合同签订
    ['5.合同\n签订', '签订时序', '采购办法(2023)\n第31条', '采购完成→\n签订合同', '公务车采购(36万)\n5.23签合同→5.27确认\n程序倒置4天', '🔴', '#6', '36', '经办人可能先上车后补票\n内部协调时间差\n但制度不允许此种操作', '合同效力争议\n→履约纠纷风险', '🔴', '系统控制:\n采购确认未完成\n→合同章不可用'],
    ['', '履约保证金', '各项目合同\n专用条款', '签合同前收取\n保证金/保函', '图书采购(7.5万保函)\n未收取即签合同', '🔴', '#5', '7.5', '审批流中缺少保证金\n校验节点\n签章前无强制核对', '供应商违约\n→无担保追索', '🔴', '保证金状态接入\n合同审批流\n未缴→流程阻断'],
    ['', '合同倒签', '采购办法(2023)\n第31条', '先采购后签约', '绿化工程(110万)\n8.11签合同约定7.26开工\n倒签16天', '🔴', '#7(绿化)', '110', '工期紧迫→先开工再补手续\n典型施工类项目通病', '合同效力瑕疵\n→付款纠纷\n→验收依据动摇', '🔴', '紧急项目事先报备\n分管院领导特批\n→事后限期补正'],
    
    # 归档
    ['6.文档\n归档', '归档完整性', '采购办法(2025)\n第七条(十)', '采购文档\n完整归档立卷', '电力改造(181万)缺评标报告\n医学设备(173万)缺磋商报告\n2个不同年度2个项目', '🔴', '#3\n#9', '353.5', '归档不是偶发失误\n是系统性缺乏归档清单\n缺乏归档→交接→核查流程', '问题追溯无依据\n→审计无法还原\n→法律风险', '🔴', '归档清单标准化\n采购完成3日内归档\n交叉检查签字确认'],
    
    # 执行通报
    ['7.执行\n与监督', '执行通报', '2025年采购通知', '定期通报执行进度\n大数据分析', '2025年1-6月有通报\n但全年大数据分析\n任务未完成', '🟡', '#18', '—', '通报机制已建立\n大数据分析属创新要求\n可能缺数据分析人才', '无数据驱动→\n管理决策凭经验\n→效率难以提升', '🟡', '借用校内信息技术\n人才资源\n或委托第三方'],
]

for i, row in enumerate(lifecycle, 5):
    for j, val in enumerate(row):
        ws_a.cell(i, j+1, val)
    risk = 'red' if '🔴' in str(row[10]) else ('yellow' if '🟡' in str(row[10]) else 'green')
    style_row(ws_a, i, len(headers_a), risk)

# A2: 根因总结
r = 5 + len(lifecycle) + 1
ws_a.cell(r, 1, '根因总结：').font = bold_font
ws_a.merge_cells(f'A{r}:L{r}')
root_causes = [
    ['1', '流程审核层级不足', '招标文件/合同签订均依赖单一审核人,缺少多级复核,导致错误直接从编制环节进入执行环节'],
    ['2', '系统卡控缺失', '采购确认→签约→保证金三个环节在线下串联,缺少系统强制校验,给"先上车"留出空间'],
    ['3', '归档文化薄弱', '归档被当作"事后补"而非"流程必备",缺少归档→交接→核查的刚性流程'],
    ['4', '评审专家管理松散', '专家库更新不及时,评审质量差(质疑成立率75%),缺少动态考核退出机制'],
    ['5', '规模增长与管理能力不匹配', '2025年199项/1.28亿规模下,原有管理方式已不适应,需要系统化升级'],
]
headers_root = ['序号', '根因', '详细分析']
for j, h in enumerate(headers_root, 1):
    ws_a.cell(r+1, j, h)
style_header(ws_a, r+1, 3)
for i, row in enumerate(root_causes, r+2):
    for j, val in enumerate(row):
        ws_a.cell(i, j+1, val)
    style_row(ws_a, i, 3, 'red')

ws_a.column_dimensions['A'].width = 12
ws_a.column_dimensions['B'].width = 14
ws_a.column_dimensions['C'].width = 20
ws_a.column_dimensions['D'].width = 22
ws_a.column_dimensions['E'].width = 28
ws_a.column_dimensions['F'].width = 8
ws_a.column_dimensions['G'].width = 10
ws_a.column_dimensions['H'].width = 12
ws_a.column_dimensions['I'].width = 26
ws_a.column_dimensions['J'].width = 20
ws_a.column_dimensions['K'].width = 8
ws_a.column_dimensions['L'].width = 22

# ==========================================
# Sheet B: 国有资产管理深度分析
# ==========================================
ws_b = wb.create_sheet('深析-国有资产管理')
ws_b.merge_cells('A1:J1')
ws_b['A1'] = '维度深析：国有资产管理 — 全生命周期 × 制度空白 × 执行偏差'
ws_b['A1'].font = title_font
ws_b['A1'].alignment = center_align

headers_b = ['资产生命周期', '管理制度\n覆盖情况', '制度文号', '应然要求', '实然表现', '偏差\n等级', '问题号', '关键数据', '根因诊断', '改进路径']
for j, h in enumerate(headers_b, 1):
    ws_b.cell(3, j, h)
style_header(ws_b, 3, len(headers_b))

asset_lifecycle = [
    ['1.资产配置\n(入口)', '✅ 已覆盖', '国资办法(2025)\n预算办法', '科学论证\n按标准配置\n与预算衔接', '2025年199项采购(1.28亿)\n配置论证在采购流程中\n未发现超标准,但量大面广\n', '🟢', '—', '199项/年', '配置端管控较好\n但体量快速增长下\n论证深度待加强', '建立配置标准库\n超标准自动预警'],
    
    ['2.资产登记\n入库', '✅ 已覆盖', '国资办法(2025)\n在建工程办法\n(2025)', '及时登记\n建卡入账\n账实相符', '资产台账15.6MB\n在建工程产权登记\n2025年新建制度', '🟢', '—', '台账15.6MB\n约数万条', '登记入库环节\n制度与执行匹配度较高', '持续'],
    
    ['3.资产使用\n(过程)', '⚠️ 部分覆盖\n缺出租出借办法', '国资办法(2025)\n低耗品办法(2025)', '合理使用\n定期盘点\n出租出借需审批', '2021-2024年每年盘点\n但盘点问题闭环不明\n简阳房屋长期闲置', '🟡', '#17', '简阳855万\n+防疫物资', '使用管理有日常操作\n但对闲置/低效资产的\n主动识别和处置能力弱', '每季度生成闲置\n资产报告\n自动标注低效资产'],
    
    ['4.资产盘点\n清查', '⚠️ 部分覆盖\n缺实施细则', '国资办法(2025)\n但无专门\n盘点细则', '定期全面盘点\n账实核对\n差异分析处理', '2021-2024年每年盘点\n有工作总结\n但:①盘点差异处理流程\n不清晰②问题发现到\n整改缺乏闭环', '🟡', '—', '4年盘点总结\n2024年全面清查', '盘点沦为"例行公事"\n发现问题→报告→\n没有整改→明年再发现\n缺乏闭环驱动力', '建立"发现→\n责任认定→整改\n→销号"闭环\n纳入绩效考核'],
    
    ['5.资产处置\n(出口-实物)', '🔴 严重缺失\n缺处置管理办法', '仅国资办法\n第七章粗线条', '按权限报批\n金额一致\n流程合规', '报废请示71万→\n批复54.7万→\n实际处置仍71万\n金额不一致', '🔴', '#11', '差额16.4万\n原因不明', '处置流程关键节点\n(请示/批复/执行)\n三个金额缺乏比对\n校验机制\n→中间可插入操作', '处置前锁定清单\n批复差异→\n二次请示确认\n处置后比对销账'],
    
    ['6.闲置资产\n处置', '🔴 严重缺失\n缺处置办法', '省属高校办法\n(2024)第37条\n国资办法第7条', '及时处置低效\n闲置资产', '简阳12处房屋+7门面\n+1门卫室 长期闲置\n仅尝试拍卖(两次流拍)\n后无进一步动作', '🔴', '#13', '闲置≥3年\n估值855万\n年机会成本\n约40-50万', '流拍=资产本体存在问题\n(位置/产权/市场)\n应切换策略:\n调剂/公益/置换\n而非等待', '①产权问题调查\n②分类处置方案\n(不能一刀切)\n③设定处置deadline'],
    
    ['7.无形资产\n管理', '⚠️ 仅框架\n缺专门办法', '国资办法(2025)\n捐赠资产办法\n(2026)', '及时登记\n评估\n报废处置', '2023年党委会决定报废\n→2026年3月才请示\n间隔2年+\n2021年盘点已发现应报废', '🔴', '#12', '报废延误\n2-3年\n涉及84.6万', '无形资产"看不见"\n→重视度低\n→报废优先级低\n→拖延成习惯', '建立资产到期\n自动提醒\n到期未处置→\n升级报告至院领导'],
    
    ['8.资产绩效\n评价', '✅ 新覆盖', '绩效评价办法\n(2025)', '建立绩效评价\n提高使用效益', '2025年刚出台\n尚无执行数据', '🟢', '—', '2025年新制度', '评价办法已建立\n需尽快落地执行\n否则成为"纸面制度"', '2026年开展首次\n绩效评价试点\n选3-5类资产'],
]

for i, row in enumerate(asset_lifecycle, 4):
    for j, val in enumerate(row):
        ws_b.cell(i, j+1, val)
    risk = 'red' if '🔴' in str(row[5]) else ('yellow' if '🟡' in str(row[5]) else ('orange' if '⚠️' in str(row[1]) else 'green'))
    style_row(ws_b, i, len(headers_b), risk)

# Summary
r = 4 + len(asset_lifecycle) + 1
ws_b.cell(r, 1, '制度空白统计：').font = bold_font
ws_b.merge_cells(f'A{r}:E{r}')
ws_b.cell(r, 6, '🔴 3项缺失').font = red_font
ws_b.cell(r, 7, '处置办法/出租出借/盘点细则').font = red_font
r2 = r + 1
ws_b.cell(r2, 1, '执行偏差统计：').font = bold_font
ws_b.merge_cells(f'A{r2}:E{r2}')
ws_b.cell(r2, 6, '🔴 3项严重').font = red_font
ws_b.cell(r2, 7, '报废不一致/闲置不处置/无形资产拖延').font = red_font
ws_b.cell(r2+1, 6, '🟡 2项轻微').font = normal_font
ws_b.cell(r2+1, 7, '盘点闭环/使用监管').font = normal_font

ws_b.column_dimensions['A'].width = 16
ws_b.column_dimensions['B'].width = 16
ws_b.column_dimensions['C'].width = 18
ws_b.column_dimensions['D'].width = 22
ws_b.column_dimensions['E'].width = 30
ws_b.column_dimensions['F'].width = 8
ws_b.column_dimensions['G'].width = 8
ws_b.column_dimensions['H'].width = 16
ws_b.column_dimensions['I'].width = 28
ws_b.column_dimensions['J'].width = 24

# ==========================================
# Sheet C: 跨年度趋势分析
# ==========================================
ws_c = wb.create_sheet('深析-跨年度趋势')
ws_c.merge_cells('A1:K1')
ws_c['A1'] = '维度深析：关键指标跨年度变化趋势（2021-2025）'
ws_c['A1'].font = title_font
ws_c['A1'].alignment = center_align

headers_c = ['指标类别', '具体指标', '2021', '2022', '2023', '2024', '2025', '趋势', '判断', '备注', '数据来源']
for j, h in enumerate(headers_c, 1):
    ws_c.cell(3, j, h)
style_header(ws_c, 3, len(headers_c))

trends = [
    # 制度建设
    ['制度建设', '年度新增/修订制度数', '10(征求意见)', '10(正式发文)', '2', '1', '8', '↗↘↗', '🟡 脉冲式', '两轮密集(2022/2025)\n中间放缓', '述职+制度清单'],
    ['制度建设', '制度总数(累计)', '~5', '15', '17', '18', '30', '↗', '🟢 持续增长', '2025年补齐大量空白', '制度文件清单'],
    ['制度建设', '制度空白项', '—', '—', '—', '—', '3', '—', '🔴 仍有缺失', '资产处置/出租出借/盘点', '制度Gap分析'],
    
    # 采购规模
    ['采购规模', '项目数(个/年)', '~95', '~60', '~91', '~143', '199', '↘↗↗', '🟡 2025激增', '2024→2025增39%', '采购台账'],
    ['采购规模', '总预算(万元)', '—', '—', '—', '—', '12,800', '—', '🔴 数据缺失', '2021-2024金额未量化', '述职仅披露2025'],
    ['采购规模', '人均管理项目数', '高(兼两处)', '高(兼两处)', '中(专任)', '中(专任)', '高(规模增)', '↗', '🟡 压力回升', '2025年规模增长超人员增速', '项目数/人员数'],
    
    # 问题数量
    ['问题发生', '审计发现问题数', '1', '2', '3', '3', '10', '↗', '🔴 逐年上升', '2025年问题集中暴露\n(可能因规模扩大)', '问题清单'],
    ['问题发生', '采购类问题', '1(程序)', '1(归档)', '2(评审)', '2(合同)', '4(多环节)', '↗', '🔴 持续恶化', '采购问题覆盖面扩大', '问题清单'],
    ['问题发生', '资产类问题', '1(物资)', '0', '1(报废)', '0', '2(处置/无形)', '—', '🟡 偶发', '资产老问题未根除', '问题清单'],
    
    # 廉政建设
    ['廉政建设', '廉政培训/学习次数', '4', '—', '—', '1(全院)', '26', '↗', '🟢 大幅提升', '2025年显著加强', '述职报告'],
    ['廉政建设', '谈话次数', '—', '—', '—', '—', '11+', '—', '🟢 新建立', '2025年新建谈话机制', '述职报告'],
    ['廉政建设', '廉政风险事件', '0', '0', '0', '0', '0', '→', '🟢 零事件', '五年保持零廉政事件', '述职报告'],
    
    # 党建
    ['党建', '政治面貌', '预备党员', '正式党员', '党员/支书', '党员/支书', '党员/支书', '↗', '🟢 正常进步', '2021预备→2022转正', '述职报告'],
    ['党建', '"重业务轻党建"', '✅承认', '✅承认', '—', '—', '✅变体回归', '→→↗', '🔴 5年未根治', '"两张皮"从直接表述\n变为间接表述,实质未变', '述职报告'],
    
    # 安全
    ['安全管理', '安全事故', '0', '0', '0', '⚠️消防事故', '0', '—', '🔴 重大事件', '2024年档案灭失\n不可逆损失', '述职/问题#19'],
]

for i, row in enumerate(trends, 4):
    for j, val in enumerate(row):
        ws_c.cell(i, j+1, val)
    risk = 'red' if '🔴' in str(row[8]) else ('yellow' if '🟡' in str(row[8]) else 'green')
    style_row(ws_c, i, len(headers_c), risk)

ws_c.column_dimensions['A'].width = 14
ws_c.column_dimensions['B'].width = 26
ws_c.column_dimensions['C'].width = 16
ws_c.column_dimensions['D'].width = 16
ws_c.column_dimensions['E'].width = 16
ws_c.column_dimensions['F'].width = 16
ws_c.column_dimensions['G'].width = 16
ws_c.column_dimensions['H'].width = 8
ws_c.column_dimensions['I'].width = 18
ws_c.column_dimensions['J'].width = 28
ws_c.column_dimensions['K'].width = 18

# ==========================================
# Sheet D: 风险传导地图
# ==========================================
ws_d = wb.create_sheet('深析-风险传导地图')
ws_d.merge_cells('A1:I1')
ws_d['A1'] = '维度深析：风险传导地图 — 从"表面问题"到"深层后果"的全链追踪'
ws_d['A1'].font = title_font
ws_d['A1'].alignment = center_align

headers_d = ['风险源', '表面问题', '直接后果', '传导路径\n(→→→)', '最终风险', '影响对象', '可能性', '严重度', '风险值']
for j, h in enumerate(headers_d, 1):
    ws_d.cell(3, j, h)
style_header(ws_d, 3, len(headers_d))

risk_map = [
    ['招标文件\n质量失控', '650万物管项目\n两次流标', '采购延误≥3个月\n服务衔接空档', '流标→重新招标→\n供应商信心下降→\n投标减少→竞争不充分\n→价格偏高', '服务断档影响\n全校师生\n采购成本上升', '全校师生\n学院声誉', '高', '🔴 严重', '🔴'],
    ['采购人监督\n缺位', '监督代表未见签到\n家具采购(112万)', '评审过程无\n内部监督\n公正性存疑', '监督缺位→评审不受控→\n可能出现倾向性评分→\n采购结果不公→供应商质疑', '采购结果效力\n受质疑\n→法律风险', '学院\n中标供应商', '中', '🔴 严重', '🔴'],
    ['评审质量\n滑坡', '4个质疑3个成立\n(质疑率75%)', '重新采购\n时间+费用浪费', '评审不严→质疑→\n废标→重新采购→\n上级关注→专项检查\n→更多问题暴露', '采购效率低下\n监管部门问责\n→审计移送', '学院\n代理机构', '高', '🟡 中等', '🔴'],
    ['闲置资产\n不处置', '简阳855万房产\n闲置≥3年', '年机会成本\n约40-50万\n资产贬值', '闲置→持续贬值→\n被上级通报→\n强制处置→被动接受\n不利条件', '国有资产流失\n→领导追责\n→绩效考核扣分', '学院\n李欣个人', '极高', '🔴 严重', '🔴'],
    ['报废处置\n金额不一致', '请示71万→批复54.7万\n→实际71万', '程序违规已发生\n金额差异16.4万', '金额不一致→触发\n纪检监察关注→\n定性为违规处置→\n个人追责+组织处理', '个人纪律风险\n→行政处分\n→影响晋升', '李欣个人\n国资处', '中', '🔴 严重', '🔴'],
    ['无形资产\n报废拖延', '2023年决定→\n2026年才请示', '2-3年延迟\n84.6万资产\n账实不符', '拖延→资产报表失真→\n审计发现→被定性为\n管理失职→整改\n→影响单位考核', '财务报告失真\n→审计意见类型\n受影响', '学院\n李欣', '极高', '🟡 中等', '🟡'],
    ['合同倒签\n/保证金缺失', '先签合同后采购(36万)\n无保证金签合同(7.5万)', '合同效力瑕疵\n担保缺失', '合同瑕疵→供应商违约\n→无担保追索→\n学院损失→\n追究审批人责任', '学院经济损失\n→个人经济责任', '学院\n李欣', '中', '🟡 中等', '🟡'],
    ['档案灭失', '消防事故→\n2021-2022纸质档案\n全部损毁', '制度建设过程\n不可追溯\n问责依据丢失', '档案灭失→历史问题\n无法查证→形成审计\n盲区→其他问题可能\n被掩盖', '审计无法还原\n→增加个人嫌疑\n→信任受损', '审计组\n李欣个人\n学院', '已发生', '🔴 严重', '🔴'],
    ['制度空白\n3项', '资产处置/出租出借\n/盘点无细则', '操作无章可循\n自由裁量空间大', '制度空白→操作不规范\n→审计发现问题→\n系统性问题归因于\n制度建设不完善', '审计评价降级\n→影响学院整体\n考核', '学院', '高', '🟡 中等', '🟡'],
    ['党建"两张皮"', '5年未根治\n2025年变体回归', '党建考核可能\n受影响', '两张皮→上级巡视\n→指出政治意识不强\n→影响个人政治评价', '个人政治前途\n→晋升受限', '李欣个人', '中', '🟡 中等', '🟡'],
]

for i, row in enumerate(risk_map, 4):
    for j, val in enumerate(row):
        ws_d.cell(i, j+1, val)
    risk = 'red' if '🔴' in str(row[8]) else 'yellow'
    style_row(ws_d, i, len(headers_d), risk)

ws_d.column_dimensions['A'].width = 14
ws_d.column_dimensions['B'].width = 24
ws_d.column_dimensions['C'].width = 22
ws_d.column_dimensions['D'].width = 28
ws_d.column_dimensions['E'].width = 24
ws_d.column_dimensions['F'].width = 14
ws_d.column_dimensions['G'].width = 8
ws_d.column_dimensions['H'].width = 10
ws_d.column_dimensions['I'].width = 8

# ===== Save =====
out_path = src_path  # Overwrite same file
wb.save(out_path)
print(f'✅ 已更新: {out_path}')
print(f'   共 {len(wb.sheetnames)} 个Sheet:')
for s in wb.sheetnames:
    ws = wb[s]
    print(f'   📊 {s} ({ws.max_row}行 × {ws.max_column}列)')
