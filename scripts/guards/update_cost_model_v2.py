"""
Update 健康照护师-成本构成测算 FINAL - using user-specified numbers from screenshots
人工成本 = (23670/120 + 23155/120) / 2 = 195.10
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
import shutil

src = r'C:\Users\scrccpa\Desktop\健康照护师-成本构成测算-v3-2026.05.21.xlsx'
dst = r'C:\Users\scrccpa\Desktop\健康照护师-成本构成测算-v3-2026.06.05.xlsx'
shutil.copy2(src, dst)
print(f"Backup: {dst}")

wb = openpyxl.load_workbook(dst)

# ============================================================
# PART 1: 耗材成本明细 - 三级/高级工 标黄调整
# ============================================================
ws_hc = wb['耗材成本明细']

# Item 4 (row 81): 隔离防护服 -> 隔离衣（布类）
ws_hc.cell(row=81, column=2).value = '隔离衣（布类）'
ws_hc.cell(row=81, column=3).value = '175cm（±5%）'
ws_hc.cell(row=81, column=4).value = '/'
ws_hc.cell(row=81, column=6).value = 60
ws_hc.cell(row=81, column=8).value = 10
ws_hc.cell(row=81, column=9).value = '60.00(磨损费,预估÷6次) = 10.00'
ws_hc.cell(row=81, column=10).value = '磨损费'

# Item 33 (row 110): 瞳孔笔、儿童晨检笔 -> 瞳孔笔
ws_hc.cell(row=110, column=2).value = '瞳孔笔'

# Item 40 (row 117): 毛巾 cost 5->2
ws_hc.cell(row=117, column=8).value = 2
ws_hc.cell(row=117, column=9).value = '20.00(磨损费,预估÷10次) = 2.00'

# Item 44 (row 121): 袜子 -> 一次性袜子
ws_hc.cell(row=121, column=2).value = '一次性袜子'
ws_hc.cell(row=121, column=3).value = '纯棉，中筒'
ws_hc.cell(row=121, column=6).value = 1
ws_hc.cell(row=121, column=7).value = 1
ws_hc.cell(row=121, column=8).value = 1
ws_hc.cell(row=121, column=9).value = '1.00 x 1.0 = 1.00'
ws_hc.cell(row=121, column=10).value = '一次性消耗品'

# Recalculate 三级/高级 合计
total_3 = 0
for r in range(78, 122):
    v = ws_hc.cell(row=r, column=8).value
    if isinstance(v, (int, float)): total_3 += v
total_3 = round(total_3, 2)
ws_hc.cell(row=122, column=8).value = total_3
print(f"三级（高级）新合计: {total_3}")

# ============================================================
# PART 2: 人工费用明细 - 重构
# ============================================================
ws_ld = wb['人工费用明细']

# Unmerge all first
for mr in list(ws_ld.merged_cells.ranges):
    ws_ld.unmerge_cells(str(mr))

for r in range(1, 31):
    for c in range(1, 9):
        ws_ld.cell(row=r, column=c).value = None

# Styles
hfill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
hfont = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
sfill = PatternFill(start_color='FFD6E4F0', end_color='FFD6E4F0', fill_type='solid')
tfill = PatternFill(start_color='FFFFF2CC', end_color='FFFFF2CC', fill_type='solid')
rfill = PatternFill(start_color='FFE2EFDA', end_color='FFE2EFDA', fill_type='solid')
rf = Font(name='微软雅黑', size=11)
bf = Font(name='微软雅黑', size=11, bold=True)

def hs(ws, r, fill=None, font=None):
    for c in range(1,9):
        cell = ws.cell(row=r, column=c)
        if fill: cell.fill = fill
        if font: cell.font = font

# Title
ws_ld.cell(row=1, column=1).value = '健康照护师（长期照护师）职业技能等级认定 — 人工费用测算明细'
ws_ld.merge_cells('A1:H1')
ws_ld.cell(row=1, column=1).font = Font(name='微软雅黑', bold=True, size=14)

ws_ld.cell(row=2, column=1).value = '数据来源：2025年11月30日及2026年1月9日两次认定实际数据（每次120人）| 执行标准：院长办公会纪要（第23期）2022.12.12'
ws_ld.merge_cells('A2:H2')

ws_ld.cell(row=3, column=1).value = '测算依据：考评员100元/小时，其他工作人员60元/小时，学生22元/小时 | 健康照护师全国联考'
ws_ld.merge_cells('A3:H3')

# === 2025.11.30 实际数据 ===
r = 5
ws_ld.cell(row=r, column=1).value = '一、2025年11月30日认定数据（4个实操考场，认定120人）'
ws_ld.merge_cells(f'A{r}:H{r}')
hs(ws_ld, r, sfill, bf)
r += 1

headers = ['序号', '人员类别', '人数', '费率', '工作时长', '日费用(元)', '单人成本(元/人)', '计算说明']
for i, h in enumerate(headers):
    c = ws_ld.cell(row=r, column=i+1)
    c.value = h; c.font = hfont; c.fill = hfill
r += 1

# 2025.11.30 items
items_1130 = [
    [1, '实操考评员', 12, '100元/h', '7h', 8400, 70.00, '4考场x3人x7h'],
    [2, 'SP模特-老师', 4, '60元/h', '7h', 1680, 14.00, '4考场x1人x7h'],
    [3, 'SP模特-学生', 8, '22元/h', '7h', 1232, 10.27, '4考场x2人x7h'],
    [4, '实操监考', 12, '60元/h', '7h', 5040, 42.00, '4考场x2人+候考室4人x7h'],
    [5, '主考/副主考/考务/督导', 6, '60元/h', '8h', 2880, 24.00, '主考1+副主考1+考务2+督导1+综合1'],
    [6, '理论监考+考场布置', 18, '60元/h', '2.5h', 2700, 22.50, '4考场x2人x1.5h监考+布置'],
    [7, '技术保障+安保+志愿者', 32, '—', '—', 1738, 14.48, '技术1人180+安保1人150+志愿者30人x15+布置/用物510'],
]
for item in items_1130:
    for i, val in enumerate(item):
        ws_ld.cell(row=r, column=i+1).value = val
    r += 1

total_1130 = 23670
ws_ld.cell(row=r, column=2).value = '2025.11.30 合计'
ws_ld.cell(row=r, column=6).value = total_1130
ws_ld.cell(row=r, column=7).value = round(total_1130/120, 2)
ws_ld.cell(row=r, column=8).value = f'{total_1130}÷120 = {round(total_1130/120,2)}'
hs(ws_ld, r, rfill, bf)
print(f"2025.11.30: total={total_1130}, per_person={round(total_1130/120,2)}")
r += 2

# === 2026.01.09 实际数据 ===
ws_ld.cell(row=r, column=1).value = '二、2026年1月9日认定数据（4个实操考场，认定120人）'
ws_ld.merge_cells(f'A{r}:H{r}')
hs(ws_ld, r, sfill, bf)
r += 1

for i, h in enumerate(headers):
    c = ws_ld.cell(row=r, column=i+1)
    c.value = h; c.font = hfont; c.fill = hfill
r += 1

items_0109 = [
    [1, '实操考评员', 12, '100元/h', '7h', 8400, 70.00, '4考场x3人x7h'],
    [2, 'SP模特-老师', 4, '60元/h', '7h', 1680, 14.00, '4考场x1人x7h'],
    [3, 'SP模特-学生', 8, '22元/h', '7h', 1232, 10.27, '4考场x2人x7h'],
    [4, '实操监考', 12, '60元/h', '7h', 5040, 42.00, '4考场x2人+候考室4人x7h'],
    [5, '主考/副主考/考务/督导', 6, '60元/h', '8h', 2880, 24.00, '主考1+副主考1+考务2+督导1+综合1'],
    [6, '理论监考+考场布置', 18, '60元/h', '2.5h', 2700, 22.50, '4考场x2人x1.5h监考+布置'],
    [7, '技术保障+安保+志愿者', 32, '—', '—', 1223, 10.19, '技术1人+安保1人+志愿者+考场用物准备'],
]
for item in items_0109:
    for i, val in enumerate(item):
        ws_ld.cell(row=r, column=i+1).value = val
    r += 1

total_0109 = 23155
ws_ld.cell(row=r, column=2).value = '2026.01.09 合计'
ws_ld.cell(row=r, column=6).value = total_0109
ws_ld.cell(row=r, column=7).value = round(total_0109/120, 2)
ws_ld.cell(row=r, column=8).value = f'{total_0109}÷120 = {round(total_0109/120,2)}'
hs(ws_ld, r, rfill, bf)
print(f"2026.01.09: total={total_0109}, per_person={round(total_0109/120,2)}")
r += 2

# === 9考场理论测算 ===
ws_ld.cell(row=r, column=1).value = '三、9个实操考场理论测算（参考，认定120人）'
ws_ld.merge_cells(f'A{r}:H{r}')
hs(ws_ld, r, sfill, bf)
r += 1

for i, h in enumerate(headers):
    c = ws_ld.cell(row=r, column=i+1)
    c.value = h; c.font = hfont; c.fill = hfill
r += 1

items_9kc = [
    [1, '实操考评员', 27, '100元/h', '7h', 18900, 157.50, '9考场x3人x7h'],
    [2, 'SP模特-老师', 9, '60元/h', '7h', 3780, 31.50, '9考场x1人x7h'],
    [3, 'SP模特-学生', 18, '22元/h', '7h', 2772, 23.10, '9考场x2人x7h'],
    [4, '实操监考', 22, '60元/h', '7h', 9240, 77.00, '9考场x2人+候考室4人x7h'],
    [5, '主考/副主考/考务/督导', 6, '60元/h', '8h', 2880, 24.00, '主考1+副主考1+考务2+督导1+综合1'],
    [6, '理论监考+考场布置', 38, '60元/h', '2.5h', 5700, 47.50, '9考场x2人x1.5h监考+布置'],
    [7, '技术保障+安保+志愿者', 56, '—', '—', 2920, 24.33, '技术+安保+志愿者50人x15+用物准备'],
]
for item in items_9kc:
    for i, val in enumerate(item):
        ws_ld.cell(row=r, column=i+1).value = val
    r += 1

total_9kc = 46192
ws_ld.cell(row=r, column=2).value = '9考场理论测算合计'
ws_ld.cell(row=r, column=6).value = total_9kc
ws_ld.cell(row=r, column=7).value = round(total_9kc/120, 2)
ws_ld.cell(row=r, column=8).value = f'(理论参考) {total_9kc}÷120 = {round(total_9kc/120,2)}'
hs(ws_ld, r, tfill, bf)
r += 2

# === 人工成本建议 ===
ws_ld.cell(row=r, column=1).value = '四、人工成本建议'
ws_ld.merge_cells(f'A{r}:H{r}')
ws_ld.cell(row=r, column=1).font = Font(name='微软雅黑', bold=True, size=13, color='FF0000')
ws_ld.cell(row=r, column=1).fill = rfill
r += 1

p_1130 = round(total_1130/120, 2)
p_0109 = round(total_0109/120, 2)
avg_labor = round((p_1130 + p_0109) / 2, 2)

ws_ld.cell(row=r, column=2).value = '2025.11.30 单人成本'
ws_ld.cell(row=r, column=7).value = p_1130
ws_ld.cell(row=r, column=8).value = f'= {total_1130} / 120'
r += 1
ws_ld.cell(row=r, column=2).value = '2026.01.09 单人成本'
ws_ld.cell(row=r, column=7).value = p_0109
ws_ld.cell(row=r, column=8).value = f'= {total_0109} / 120'
r += 1

ws_ld.cell(row=r, column=2).value = '★★ 建议人均人工成本（两次认定平均值）'
ws_ld.cell(row=r, column=7).value = avg_labor
ws_ld.cell(row=r, column=8).value = f'= ({total_1130}/120 + {total_0109}/120) / 2 = {avg_labor}'
for c in range(1,9):
    ws_ld.cell(row=r, column=c).font = Font(name='微软雅黑', bold=True, size=12, color='FF0000')
    ws_ld.cell(row=r, column=c).fill = rfill

# ============================================================
# PART 3: 成本构成总览
# ============================================================
ws_zl = wb['成本构成总览']

# Row 13: 人工费用 -> 195.10
ws_zl.cell(row=13, column=4).value = avg_labor
ws_zl.cell(row=13, column=5).value = avg_labor
ws_zl.cell(row=13, column=6).value = avg_labor

# Row 14: SP模特 component
sp_avg = round((4*(60*7+44*7)/120 + 9*(60*7+44*7)/120) / 2, 2)
ws_zl.cell(row=14, column=3).value = '1师60/h+2生22/hx7h/120人(4-9考场均值)'
ws_zl.cell(row=14, column=4).value = sp_avg
ws_zl.cell(row=14, column=5).value = sp_avg
ws_zl.cell(row=14, column=6).value = sp_avg

# Row 16: 耗材成本 三级 -> 257.26
ws_zl.cell(row=16, column=6).value = total_3

# Row 21: 实际成本小计
five_actual = round(20 + avg_labor + 93 + 87.2 + 106.67, 2)
four_actual = round(20 + avg_labor + 158 + 167.28 + 106.67, 2)
three_actual = round(20 + avg_labor + 190 + total_3 + 106.67, 2)
ws_zl.cell(row=21, column=4).value = five_actual
ws_zl.cell(row=21, column=5).value = four_actual
ws_zl.cell(row=21, column=6).value = three_actual

# Row 24: 全口径成本
five_full = round(170 + five_actual, 2)
four_full = round(225 + four_actual, 2)
three_full = round(280 + three_actual, 2)
ws_zl.cell(row=24, column=4).value = five_full
ws_zl.cell(row=24, column=5).value = four_full
ws_zl.cell(row=24, column=6).value = three_full

# Row 28: 已测算7项成本
ws_zl.cell(row=28, column=4).value = five_full
ws_zl.cell(row=28, column=5).value = four_full
ws_zl.cell(row=28, column=6).value = three_full

# Row 29: 差额
ws_zl.cell(row=29, column=4).value = round(320 - five_full, 2)
ws_zl.cell(row=29, column=5).value = round(405 - four_full, 2)
ws_zl.cell(row=29, column=6).value = round(500 - three_full, 2)

# Update row 3
ws_zl.cell(row=3, column=1).value = '测算假设：每班次120人(4实操考场) | 实操7小时/天 | SP模特(1老师60/h+2学生22/h) | 场地3200元/天 | 基于2025.11.30及2026.1.9两次认定实际数据'

wb.save(dst)

print(f"\n=== FINAL SUMMARY ===")
print(f"耗材成本三级/高级: {total_3} (原280.49)")
print(f"2025.11.30: {total_1130}, 单人={p_1130}")
print(f"2026.01.09: {total_0109}, 单人={p_0109}")
print(f"建议人均人工成本: {avg_labor} = ({total_1130}/120 + {total_0109}/120) / 2")
print(f"实际成本小计: 五级={five_actual}, 四级={four_actual}, 三级={three_actual}")
print(f"全口径成本: 五级={five_full}, 四级={four_full}, 三级={three_full}")
print(f"差额: 五级={round(320-five_full,2)}, 四级={round(405-four_full,2)}, 三级={round(500-three_full,2)}")
print(f"输出: {dst}")
