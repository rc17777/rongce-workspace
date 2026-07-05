# -*- coding: utf-8 -*-
"""Update analysis: ALL 8 questions per person, 30 people per batch"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(r'D:\openclaw-workspace\output\健康照护师收费分析表.xlsx')

# Styles
title_font = Font(name='Arial', size=14, bold=True)
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
sub_header_font = Font(name='Arial', size=11, bold=True)
normal_font = Font(name='Arial', size=10)
red_font = Font(name='Arial', size=10, color='FF0000', bold=True)
green_font = Font(name='Arial', size=10, color='008000')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color='C0504D', end_color='C0504D', fill_type='solid')
light_blue_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
light_red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
light_green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def ah(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center_align; cell.border = thin_border

def ar(ws, row, cols, fill=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = normal_font; cell.border = thin_border
        cell.alignment = center_align if c > 1 else left_align
        if fill: cell.fill = fill

# ============================================================
# Replace "认定场景分析" sheet with corrected version
# ============================================================
# Remove old sheet and recreate
idx = wb.sheetnames.index('认定场景分析')
wb.remove(wb['认定场景分析'])
ws = wb.create_sheet('认定场景分析(全8题×30人)', idx)

ws.merge_cells('A1:H1')
ws.cell(row=1, column=1).value = '职业技能认定考试成本重算 —— 全8题 × 班次30人'
ws.cell(row=1, column=1).font = title_font
ws.cell(row=1, column=1).alignment = center_align
ws.row_dimensions[1].height = 35

# Scenario description
scenarios = [
    ('认定性质', '职业技能等级认定(考试)，非培训'),
    ('实操考题', '共8题，每个考生全部考完(非抽题)'),
    ('班次人数', '约30人/批'),
    ('认定时长', '1天'),
    ('数据来源', '护理学院长期照护师耗材成本明细表 2026.05.14 + 收费方案5稿'),
]
for i, (k, v) in enumerate(scenarios):
    row = 3 + i
    ws.cell(row=row, column=1).value = k
    ws.cell(row=row, column=1).font = sub_header_font
    ws.cell(row=row, column=1).fill = light_blue_fill
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    ws.cell(row=row, column=2).value = v
    ws.cell(row=row, column=2).font = normal_font
    for c in range(1, 9):
        ws.cell(row=row, column=c).border = thin_border

# Section 1: Cost breakdown
row = 9
ws.merge_cells(f'A{row}:H{row}')
ws.cell(row=row, column=1).value = '一、实际人均成本 vs 收费方案分配'
ws.cell(row=row, column=1).font = Font(name='Arial', size=12, bold=True)

row = 10
h = ['等级', '真正一次性\n耗材(元/人)', '可分摊耗材\n(4人组/磨损)\n(元/人)', '设备折旧\n(30人分摊后)\n(元/人)', '实际人均\n耗材+设备\n合计(元)', '收费方案\n耗材+设备\n(元)', '差额\n(元)', '缺口率']
for c, hv in enumerate(h, 1):
    ws.cell(row=row, column=c).value = hv
ah(ws, row, len(h))

data = [
    ['初级', 51.7, 35.5, 44.6, 131.8, 50, -81.8, '-163.6%'],
    ['中级', 107.0, 60.3, 75.8, 243.1, 70, -173.1, '-247.3%'],
    ['高级', 206.3, 74.2, 91.2, 371.7, 100, -271.7, '-271.7%'],
]
for r, d in enumerate(data):
    row = 11 + r
    for c, v in enumerate(d, 1):
        ws.cell(row=row, column=c).value = v
    ar(ws, row, len(h), fill=light_red_fill)

# Section 2: Key high-cost items driving the gap
row = 15
ws.merge_cells(f'A{row}:H{row}')
ws.cell(row=row, column=1).value = '二、驱动成本差距的关键高值项目（初级为例）'
ws.cell(row=row, column=1).font = Font(name='Arial', size=12, bold=True)

row = 16
h2 = ['序号', '项目', '单人成本(元)', '占一次性耗材比', '可否压降', '替代方案建议']
for c, hv in enumerate(h2, 1):
    ws.cell(row=row, column=c).value = hv
ah(ws, row, len(h2))

key_items = [
    ['1', '隔离防护服(3M) 28.23元', '28.23', '54.6%', '是', '国产替代品牌约15-18元/套，可省10-13元'],
    ['2', '口腔护理包 4.85元', '4.85', '9.4%', '是', '批量采购可谈至3-4元'],
    ['3', '纸尿裤(成人) 3.3元', '3.3', '6.4%', '是', '考核可用小包装/单品'],
    ['4', '治疗巾(一次性) 2.5元×5', '2.5', '4.8%', '部分', '核实是否每题都需5张'],
    ['5', '病号服(磨损) 15元', '15', '—', '待核实', '60元/件÷4人，若可用30次则降至2元'],
    ['6', '治疗车(折旧) 20元', '20', '—', '待核实', '800元/辆，折旧年限和次数需确认'],
    ['7', '坐便椅(折旧) 20元', '20', '—', '待核实', '120元/个，折旧次数需确认'],
]
for r, d in enumerate(key_items):
    row = 17 + r
    for c, v in enumerate(d, 1):
        ws.cell(row=row, column=c).value = v
    ar(ws, row, len(h2))

# Section 3: Cost verification checklist
row = 25
ws.merge_cells(f'A{row}:H{row}')
ws.cell(row=row, column=1).value = '三、需护理学院逐一核实的成本项'
ws.cell(row=row, column=1).font = Font(name='Arial', size=12, bold=True)

checks = [
    '1. 隔离防护服28.23元/人——是否必须用3M品牌？国产替代品(15-18元)能否满足考核要求？',
    '2. 治疗车800元折旧每人20元——治疗车折旧几年？每年使用多少次？800/(年数×次数)=?',
    '3. 坐便椅120元折旧每人20元——折旧计算方式同上，20元/人的依据？',
    '4. 轮椅504元折旧每人10元——同上，折旧计算是否合理？',
    '5. 血糖仪82元折旧每人5元——同上',
    '6. 电子血压计185元折旧每人10元——同上',
    '7. 病号服60元按4人分每人15元——实际使用多少次后报废？若可用20次则÷20',
    '8. 大浴巾48.5元按4人分每人12元——同上',
    '9.  所有"4人一组磨损费"项目——实际使用频次和报废周期',
    '10. 是否考虑批量采购折扣？（30人×多批次=大量采购，单价应低于零售）',
]
for r, c in enumerate(checks):
    row = 26 + r
    ws.merge_cells(f'A{row}:H{row}')
    ws.cell(row=row, column=1).value = c
    ws.cell(row=row, column=1).font = normal_font
    ws.cell(row=row, column=1).alignment = left_align

# Section 4: Overall assessment
row = 37
ws.merge_cells(f'A{row}:H{row}')
ws.cell(row=row, column=1).value = '四、综合判断'
ws.cell(row=row, column=1).font = Font(name='Arial', size=12, bold=True)

judgments = [
    '按全8题+30人/批，部门测算的人均耗材+设备成本(132~372元)远超收费方案分配(50~100元)。',
    '但部门测算中多个高值项目存在压降空间：国产替代品、折旧重算、批量采购折扣。',
    '核心待核实问题：设备折旧年限和使用次数。若折旧计算合理，则收费方案确实偏低；若折旧虚高，则缺口可缩小。',
    '建议：要求护理学院提供①每项设备的折旧计算明细(原值/折旧年限/年使用次数) ②高值耗材的替代品牌对比 ③批量采购报价。',
    '核对完毕后，按实际合理成本重新确定耗材+设备费在收费方案中的分配比例。',
]
for r, j in enumerate(judgments):
    row = 38 + r
    ws.merge_cells(f'A{row}:H{row}')
    ws.cell(row=row, column=1).value = j
    ws.cell(row=row, column=1).font = normal_font
    ws.cell(row=row, column=1).alignment = left_align

ws.column_dimensions['A'].width = 20
for c in ['B','C','D','E','F','G','H']:
    ws.column_dimensions[c].width = 18

# Save
wb.save(r'D:\openclaw-workspace\output\健康照护师收费分析表.xlsx')
print('Done. Sheets:', wb.sheetnames)
