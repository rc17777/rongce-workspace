"""
Update 健康照护师-成本构成测算-v3-2026.05.21.xlsx
Changes:
1. 耗材成本明细 - 三级（高级）标黄部分调整
2. 人工费用明细 - 重构基于实际认定数据
3. 成本构成总览 - 更新汇总数字
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import shutil

# Backup
src = r'C:\Users\scrccpa\Desktop\健康照护师-成本构成测算-v3-2026.05.21.xlsx'
dst = r'C:\Users\scrccpa\Desktop\健康照护师-成本构成测算-v3-2026.06.05.xlsx'
shutil.copy2(src, dst)
print(f"Backed up to: {dst}")

wb = openpyxl.load_workbook(dst)

# ============================================================
# PART 1: 耗材成本明细 - 三级/高级工 标黄调整
# ============================================================
ws_hc = wb['耗材成本明细']

# Item 4 (row 81): 隔离防护服 → 隔离衣（布类）
ws_hc.cell(row=81, column=2).value = '隔离衣（布类）'
ws_hc.cell(row=81, column=3).value = '175cm（\u00b15%）'
ws_hc.cell(row=81, column=4).value = '/'
ws_hc.cell(row=81, column=6).value = 60
ws_hc.cell(row=81, column=8).value = 10
ws_hc.cell(row=81, column=9).value = '60.00(磨损费,预估÷6次) = 10.00'
ws_hc.cell(row=81, column=10).value = '磨损费'

# Item 33 (row 110): 瞳孔笔、儿童晨检笔 → 瞳孔笔
ws_hc.cell(row=110, column=2).value = '瞳孔笔'

# Item 40 (row 117): 毛巾 cost 5→2
ws_hc.cell(row=117, column=8).value = 2
ws_hc.cell(row=117, column=9).value = '20.00(磨损费,预估÷10次) = 2.00'

# Item 44 (row 121): 袜子 → 一次性袜子
ws_hc.cell(row=121, column=2).value = '一次性袜子'
ws_hc.cell(row=121, column=3).value = '纯棉，中筒'
ws_hc.cell(row=121, column=6).value = 1
ws_hc.cell(row=121, column=7).value = 1
ws_hc.cell(row=121, column=8).value = 1
ws_hc.cell(row=121, column=9).value = '1.00 x 1.0 = 1.00'
ws_hc.cell(row=121, column=10).value = '一次性消耗品'

# Recalculate 三级/高级 合计 (row 122)
total_3 = 0
for r in range(78, 122):
    cost_val = ws_hc.cell(row=r, column=8).value
    if isinstance(cost_val, (int, float)):
        total_3 += cost_val
total_3 = round(total_3, 2)
ws_hc.cell(row=122, column=8).value = total_3
print(f"三级（高级）新合计: {total_3}")

# Recalculate 五级/初级 合计
for r in range(1, 40):
    if ws_hc.cell(row=r, column=1).value and '五级' in str(ws_hc.cell(row=r, column=1).value):
        t5 = 0
        for rr in range(r+2, 40):
            if ws_hc.cell(row=rr, column=1).value == '合计':
                t5 = round(t5, 2)
                ws_hc.cell(row=rr, column=8).value = t5
                print(f"五级（初级）合计: {t5}")
                break
            v = ws_hc.cell(row=rr, column=8).value
            if isinstance(v, (int, float)): t5 += v
        break

# Recalculate 四级/中级 合计
for r in range(40, 76):
    if ws_hc.cell(row=r, column=1).value and '四级' in str(ws_hc.cell(row=r, column=1).value):
        t4 = 0
        for rr in range(r+2, 76):
            if ws_hc.cell(row=rr, column=1).value == '合计':
                t4 = round(t4, 2)
                ws_hc.cell(row=rr, column=8).value = t4
                print(f"四级（中级）合计: {t4}")
                break
            v = ws_hc.cell(row=rr, column=8).value
            if isinstance(v, (int, float)): t4 += v
        break

# ============================================================
# PART 2: 人工费用明细 - 重构
# ============================================================
ws_ld = wb['人工费用明细']

# Remove all existing merged cells first
merged_ranges = list(ws_ld.merged_cells.ranges)
for mr in merged_ranges:
    ws_ld.unmerge_cells(str(mr))

# Now clear all content
for r in range(1, 31):
    for c in range(1, 9):
        ws_ld.cell(row=r, column=c).value = None

# Styles
header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
section_fill = PatternFill(start_color='FFD6E4F0', end_color='FFD6E4F0', fill_type='solid')
subtotal_fill = PatternFill(start_color='FFFFF2CC', end_color='FFFFF2CC', fill_type='solid')
result_fill = PatternFill(start_color='FFE2EFDA', end_color='FFE2EFDA', fill_type='solid')
nf = Font(name='微软雅黑', size=11)
bf = Font(name='微软雅黑', size=11, bold=True)

# Title
ws_ld.cell(row=1, column=1).value = '健康照护师（长期照护师）职业技能等级认定 — 人工费用测算明细'
ws_ld.merge_cells('A1:H1')
ws_ld.cell(row=1, column=1).font = Font(name='微软雅黑', bold=True, size=14)

ws_ld.cell(row=2, column=1).value = '数据来源：2025年11月30日及2026年1月9日实际认定数据 | 执行标准：院长办公会纪要（第23期）2022.12.12'
ws_ld.merge_cells('A2:H2')

ws_ld.cell(row=3, column=1).value = '测算假设：每日1班次120人 | 实操7小时 | 考评员100元/h，其他60元/h，学生22元/h | 健康照护师全国联考'
ws_ld.merge_cells('A3:H3')

def write_header(ws, row):
    headers = ['序号', '人员类别', '人数', '费率(元/h)', '工作时长(h)', '日费用(元)', '单人成本(元/人)', '计算说明']
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=i+1)
        c.value = h
        c.font = header_font
        c.fill = header_fill

def write_section(ws, row, title):
    ws.cell(row=row, column=1).value = title
    ws.merge_cells(f'A{row}:H{row}')
    ws.cell(row=row, column=1).font = bf
    ws.cell(row=row, column=1).fill = section_fill

def write_items(ws, row, items):
    for item in items:
        for i, val in enumerate(item):
            ws.cell(row=row, column=i+1).value = val
        row += 1
    return row

def write_subtotal(ws, row, label, total, per_person):
    ws.cell(row=row, column=2).value = label
    ws.cell(row=row, column=6).value = total
    ws.cell(row=row, column=7).value = per_person
    for c in range(1, 9):
        ws.cell(row=row, column=c).font = bf
        ws.cell(row=row, column=c).fill = subtotal_fill

def write_result(ws, row, label, total=None, per_person=None, note=None):
    ws.cell(row=row, column=2).value = label
    if total: ws.cell(row=row, column=6).value = total
    if per_person: ws.cell(row=row, column=7).value = per_person
    if note: ws.cell(row=row, column=8).value = note
    for c in range(1, 9):
        ws.cell(row=row, column=c).font = bf
        ws.cell(row=row, column=c).fill = result_fill

# === 4考场 ===
r = 5
write_section(ws_ld, r, '一、4个实操考场人工费用明细')
r += 1
write_header(ws_ld, r)
r += 1

items_4_ops = [
    [1, '实操考评员', 12, 100, 7, 8400, round(8400/120, 2), '4考场x3人x7h'],
    [2, 'SP模特-老师', 4, 60, 7, 1680, round(1680/120, 2), '4考场x1人x7h'],
    [3, 'SP模特-学生', 8, 22, 7, 1232, round(1232/120, 2), '4考场x2人x7h'],
    [4, '实操监考', 12, 60, 7, 5040, round(5040/120, 2), '4考场x2人+候考室4人x7h'],
]
r = write_items(ws_ld, r, items_4_ops)
sub_4_ops = 8400+1680+1232+5040
write_subtotal(ws_ld, r, '实操部分小计', sub_4_ops, round(sub_4_ops/120, 2))
r += 2

items_4_mgmt = [
    [5, '主考', 1, 60, 8, 480, round(480/120, 2), '全程负责'],
    [6, '副主考', 1, 60, 8, 480, round(480/120, 2), ''],
    [7, '考务人员', 2, 60, 8, 960, round(960/120, 2), '2人'],
    [8, '督导员', 1, 60, 8, 480, round(480/120, 2), ''],
    [9, '综合管理', 1, 60, 8, 480, round(480/120, 2), '考前统筹+当天协调'],
]
r = write_items(ws_ld, r, items_4_mgmt)
sub_mgmt = 480*5
write_subtotal(ws_ld, r, '管理部分小计', sub_mgmt, round(sub_mgmt/120, 2))
r += 2

items_4_theory = [
    [10, '理论监考', 8, 60, 1.5, 720, round(720/120, 2), '4考场x2人x1.5h'],
    [11, '考场布置', 10, 60, 1, 600, round(600/120, 2), '理论4+实操4+用物准备'],
    [12, '技术保障', 1, 60, 3, 180, round(180/120, 2), '考前部署+考试保障'],
    [13, '安保', 1, 150, 1, 150, round(150/120, 2), '1天'],
    [14, '志愿者午餐', 30, 15, 1, 450, round(450/120, 2), '30人x15元'],
]
r = write_items(ws_ld, r, items_4_theory)
sub_4_other = 720+600+180+150+450
write_subtotal(ws_ld, r, '理论/其他小计', sub_4_other, round(sub_4_other/120, 2))
r += 2

total_4 = sub_4_ops + sub_mgmt + sub_4_other
write_result(ws_ld, r, '4考场每日人工费总计', total_4)
r += 1
write_result(ws_ld, r, '4考场单人人工成本', None, round(total_4/120, 2), f'{total_4}/120={round(total_4/120,2)}')
print(f"4考场: total={total_4}, per_person={round(total_4/120, 2)}")
r += 3

# === 9考场 ===
write_section(ws_ld, r, '二、9个实操考场人工费用明细')
r += 1
write_header(ws_ld, r)
r += 1

items_9_ops = [
    [1, '实操考评员', 27, 100, 7, 18900, round(18900/120, 2), '9考场x3人x7h'],
    [2, 'SP模特-老师', 9, 60, 7, 3780, round(3780/120, 2), '9考场x1人x7h'],
    [3, 'SP模特-学生', 18, 22, 7, 2772, round(2772/120, 2), '9考场x2人x7h'],
    [4, '实操监考', 22, 60, 7, 9240, round(9240/120, 2), '9考场x2人+候考室4人x7h'],
]
r = write_items(ws_ld, r, items_9_ops)
sub_9_ops = 18900+3780+2772+9240
write_subtotal(ws_ld, r, '实操部分小计', sub_9_ops, round(sub_9_ops/120, 2))
r += 2

r = write_items(ws_ld, r, items_4_mgmt)  # same management
write_subtotal(ws_ld, r, '管理部分小计', sub_mgmt, round(sub_mgmt/120, 2))
r += 2

items_9_theory = [
    [10, '理论监考', 18, 60, 1.5, 1620, round(1620/120, 2), '9考场x2人x1.5h'],
    [11, '考场布置', 20, 60, 1, 1200, round(1200/120, 2), '理论9+实操9+用物准备'],
    [12, '技术保障', 1, 60, 4, 240, round(240/120, 2), '考前部署+考试保障'],
    [13, '安保', 1, 150, 1, 150, round(150/120, 2), '1天'],
    [14, '志愿者午餐', 50, 15, 1, 750, round(750/120, 2), '50人x15元'],
]
r = write_items(ws_ld, r, items_9_theory)
sub_9_other = 1620+1200+240+150+750
write_subtotal(ws_ld, r, '理论/其他小计', sub_9_other, round(sub_9_other/120, 2))
r += 2

total_9 = sub_9_ops + sub_mgmt + sub_9_other
write_result(ws_ld, r, '9考场每日人工费总计', total_9)
r += 1
write_result(ws_ld, r, '9考场单人人工成本', None, round(total_9/120, 2), f'{total_9}/120={round(total_9/120,2)}')
print(f"9考场: total={total_9}, per_person={round(total_9/120, 2)}")
r += 3

# === 建议 ===
ws_ld.cell(row=r, column=1).value = '三、人工成本建议'
ws_ld.merge_cells(f'A{r}:H{r}')
ws_ld.cell(row=r, column=1).font = Font(name='微软雅黑', bold=True, size=13, color='FF0000')
ws_ld.cell(row=r, column=1).fill = result_fill
r += 1

p_4 = round(total_4/120, 2)
p_9 = round(total_9/120, 2)
avg_labor = round((p_4 + p_9) / 2, 2)

write_result(ws_ld, r, '4考场单人成本', None, p_4, f'= {total_4} / 120')
r += 1
write_result(ws_ld, r, '9考场单人成本', None, p_9, f'= {total_9} / 120')
r += 1
ws_ld.cell(row=r, column=2).value = '建议人均人工成本（平均值）'
ws_ld.cell(row=r, column=7).value = avg_labor
ws_ld.cell(row=r, column=8).value = f'= ({total_4}/120 + {total_9}/120) / 2 = {avg_labor}'
for c in range(1, 9):
    ws_ld.cell(row=r, column=c).font = Font(name='微软雅黑', bold=True, size=12, color='FF0000')
    ws_ld.cell(row=r, column=c).fill = result_fill
print(f"建议人均人工成本: {avg_labor}")

# ============================================================
# PART 3: 成本构成总览 - 更新
# ============================================================
ws_zl = wb['成本构成总览']

# Row 13: 人工费用
ws_zl.cell(row=13, column=4).value = avg_labor
ws_zl.cell(row=13, column=5).value = avg_labor
ws_zl.cell(row=13, column=6).value = avg_labor

# Row 14: SP模特 (component of labor)
sp_avg = round((4*(60+44)*7/120 + 9*(60+44)*7/120) / 2, 2)
ws_zl.cell(row=14, column=3).value = '1师60/h+2生22/hx7h/120人(4-9考场均值)'
ws_zl.cell(row=14, column=4).value = sp_avg
ws_zl.cell(row=14, column=5).value = sp_avg
ws_zl.cell(row=14, column=6).value = sp_avg

# Row 16: 耗材成本 三级
ws_zl.cell(row=16, column=6).value = total_3

# Row 21: 实际成本小计
five_actual = round(20 + avg_labor + 93 + 87.2 + 106.67, 2)
four_actual = round(20 + avg_labor + 158 + 167.28 + 106.67, 2)
three_actual = round(20 + avg_labor + 190 + total_3 + 106.67, 2)
ws_zl.cell(row=21, column=4).value = five_actual
ws_zl.cell(row=21, column=5).value = four_actual
ws_zl.cell(row=21, column=6).value = three_actual

# Row 24: 全口径成本
five_full = round(170 + five_actual, 2)
four_full = round(225 + four_actual, 2)
three_full = round(280 + three_actual, 2)
ws_zl.cell(row=24, column=4).value = five_full
ws_zl.cell(row=24, column=5).value = four_full
ws_zl.cell(row=24, column=6).value = three_full

# Row 28: 已测算7项成本
ws_zl.cell(row=28, column=4).value = five_full
ws_zl.cell(row=28, column=5).value = four_full
ws_zl.cell(row=28, column=6).value = three_full

# Row 29: 差额
ws_zl.cell(row=29, column=4).value = round(320 - five_full, 2)
ws_zl.cell(row=29, column=5).value = round(405 - four_full, 2)
ws_zl.cell(row=29, column=6).value = round(500 - three_full, 2)

# Update row 3 assumptions
ws_zl.cell(row=3, column=1).value = '测算假设：每班次120人(4-9实操考场) | 实操7小时/天 | SP模特(1老师60/h+2学生22/h) | 场地3200元/天 | 基于2025.11.30及2026.1.9实际数据'

# Save
wb.save(dst)

print(f"\n=== SUMMARY ===")
print(f"耗材成本三级/高级: {total_3}")
print(f"4考场人工费总计: {total_4}, 单人: {p_4}")
print(f"9考场人工费总计: {total_9}, 单人: {p_9}")
print(f"建议人均人工成本: {avg_labor}")
print(f"实际成本小计: 五级={five_actual}, 四级={four_actual}, 三级={three_actual}")
print(f"全口径成本: 五级={five_full}, 四级={four_full}, 三级={three_full}")
print(f"差额: 五级={round(320-five_full,2)}, 四级={round(405-four_full,2)}, 三级={round(500-three_full,2)}")
print(f"Saved: {dst}")
print("=== DONE ===")
