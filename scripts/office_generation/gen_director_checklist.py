# -*- coding: utf-8 -*-
"""生成成都市轨道资源公司-董事监事履职专项审计检查清单Excel"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()

# === Styles ===
hdr_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
title_font = Font(name='Microsoft YaHei', size=16, bold=True, color='FFFFFF')
sec_font = Font(name='Microsoft YaHei', size=11, bold=True, color='8B0000')
body_font = Font(name='Microsoft YaHei', size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

red_fill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
red_hdr = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
blue_fill = PatternFill(start_color='2E86C1', end_color='2E86C1', fill_type='solid')
blue_hdr = PatternFill(start_color='3498DB', end_color='3498DB', fill_type='solid')
green_fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
green_hdr = PatternFill(start_color='2ECC71', end_color='2ECC71', fill_type='solid')
dark_fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
dark_hdr = PatternFill(start_color='5D6D7E', end_color='5D6D7E', fill_type='solid')
light_gray = PatternFill(start_color='F2F3F4', end_color='F2F3F4', fill_type='solid')
light_red = PatternFill(start_color='FDEDEC', end_color='FDEDEC', fill_type='solid')
light_blue = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')
light_green = PatternFill(start_color='EAFAF1', end_color='EAFAF1', fill_type='solid')
light_dark = PatternFill(start_color='EBEDEF', end_color='EBEDEF', fill_type='solid')

def setup_sheet(ws, title, fill, subtitle='被审计单位：成都市轨道资源经营管理有限公司'):
    ws.merge_cells('A1:G1')
    c = ws['A1']
    c.value = title
    c.font = title_font
    c.fill = fill
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 38
    ws['A2'] = subtitle
    ws['A2'].font = Font(name='Microsoft YaHei', size=10, italic=True)
    ws.merge_cells('A2:G2')

def write_headers(ws, row, headers, fill):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = hdr_font
        c.fill = fill
        c.alignment = center_align
        c.border = thin_border

def write_section(ws, row, text, fill):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    c = ws.cell(row=row, column=1, value=text)
    c.font = sec_font
    c.fill = fill
    c.alignment = Alignment(vertical='center')
    c.border = thin_border
    for col in range(2, 8):
        ws.cell(row=row, column=col).fill = fill
        ws.cell(row=row, column=col).border = thin_border

def write_row(ws, row, data, zebra_fill):
    for col, val in enumerate(data, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = body_font
        c.alignment = wrap_align
        c.border = thin_border
        if row % 2 == 0:
            c.fill = zebra_fill

def set_col_widths(ws):
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 48
    ws.column_dimensions['D'].width = 48
    ws.column_dimensions['E'].width = 38
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 32

# =============================================
# Sheet 1: 董事履职合规检查清单（18条）
# =============================================
ws1 = wb.active
ws1.title = '董事履职合规检查'

setup_sheet(ws1, '董事履职合规性检查清单（18条）', red_fill)
headers = ['序号', '检查维度', '检查项', '检查方法', '法规依据', '风险等级', '审计证据']
write_headers(ws1, 3, headers, red_hdr)

directors = [
    ('一、人员选聘与资格（第1-5项）', light_red),
    (1, '选聘与资格', '外部董事人数是否占董事会成员的半数以上？',
     '查阅董事会名册、公司章程、工商登记信息，统计内外部董事比例',
     '新《公司法》第173条：国有独资公司董事会成员中应当过半数为外部董事', '\U0001F534 高', '董事会名册、公司章程、工商登记信息'),
    (2, '选聘与资格', '是否建立了外部董事人才库？选聘来源是否单一？',
     '查阅外部董事遴选制度文件、人才库名录，分析选聘渠道分布',
     '国企改革三年行动方案、国资委关于加强外部董事选聘管理的通知', '\U0001F7E1 中', '外部董事遴选制度、人才库名单、选聘记录'),
    (3, '选聘与资格', '外派董事是否有明确的岗位说明书（含职责权限义务表述）？',
     '查阅董事岗位说明书、聘任协议，核查职责条款完整性',
     '公司治理准则（G20/OECD）第二章第VI条', '\U0001F7E1 中', '岗位说明书、聘任协议'),
    (4, '选聘与资格', '是否存在董事同时在3家以上企业兼职的情况？',
     '通过董事个人信息调查表+天眼查/企查查交叉比对，统计兼职数量',
     '新《公司法》第184条：董事忠实义务', '\U0001F534 高', '兼职声明、调查表、工商查询记录'),
    (5, '选聘与资格', '董事上岗前是否经过系统性培训？有无持证上岗制度？',
     '查阅培训档案、考核记录、证书，核实岗前培训完成情况',
     '国资委关于董事会建设相关指导意见', '\U0001F7E2 低', '培训档案、考核成绩、培训证书'),

    ('二、决策机制（第6-9项）', light_red),
    (6, '决策机制', '是否建立了重大事项分类决策机制？非重大事项董事能否自主表决？',
     '查阅董事会议事规则、"三重一大"制度，检查决策权限划分表',
     '新《公司法》第67条：董事会职权、国务院国资委"三重一大"决策制度', '\U0001F534 高', '董事会议事规则、三重一大制度、决策权限表'),
    (7, '决策机制', '董事会表决是否存在"先内部审批、后形式上开会"的流程？',
     '时间链比对法：抽取10项重大决策，比对股东审批日期与董事会会议日期',
     '新《公司法》第125条：董事会决议须经全体董事过半数通过', '\U0001F534 高', '股东审批记录、董事会会议记录、表决票'),
    (8, '决策机制', '会议通知是否提前发送？议案材料是否完整规范？',
     '抽取最近10次董事会会议通知和议案材料，检查通知时间与章程规定的符合性',
     '新《公司法》第124条：会议通知义务', '\U0001F7E1 中', '会议通知邮件/签收单、议案材料清单'),
    (9, '决策机制', '董事是否就表决事项进行了充分的会前调研？',
     '查阅董事调研报告、工作日志，核对其覆盖的议题数量和质量',
     '国资委关于加强国有企业董事会建设的意见', '\U0001F7E1 中', '调研报告、工作日志、出差记录'),

    ('三、履职记录（第10-12项）', light_red),
    (10, '履职记录', '董事是否有完整的履职工作记录（出席情况、表决意见、调研报告等）？',
     '查阅董事履职台账，统计出席率、独立表决率、调研天数',
     '国企改革三年行动方案关于董事会考核的要求', '\U0001F534 高', '履职台账、出席签到表、表决票、调研报告'),
    (11, '履职记录', '是否建立了董事年度述职制度？述职报告是否存档？',
     '查阅述职制度文件和近三年述职报告存档情况',
     '国资委关于外部董事管理办法的相关要求', '\U0001F7E1 中', '述职制度、述职报告、存档记录'),
    (12, '履职记录', '是否建立了董事履职台账（参会次数、发言内容、表决记录）？',
     '抽取5次重大会议检验台账与会议记录一致性',
     '国企公司治理示范企业评价标准', '\U0001F534 高', '履职台账、会议记录、表决票'),

    ('四、考核与激励（第13-16项）', light_red),
    (13, '考核与激励', '是否建立了董事考核评价机制？评价维度是否包含忠实勤勉、专业性、独立性？',
     '查阅考核制度文件，抽取近两年考核档案评估执行情况',
     '新《公司法》第147条：董事忠实勤勉义务、国资委外部董事考核办法', '\U0001F7E1 中', '考核制度、考核档案、评分表'),
    (14, '考核与激励', '考核结果是否与薪酬、续聘挂钩？',
     '抽取近两年薪酬发放记录与考核结果进行匹配分析',
     '国企薪酬改革相关文件', '\U0001F7E1 中', '薪酬发放记录、考核结果、聘任文件'),
    (15, '考核与激励', '外派董事是否在任职企业领取报酬？是否存在"兼职不兼薪、有责无利"？',
     '查阅任职企业薪酬台账和派出单位工资单，交叉比对收入来源',
     '新《公司法》第118条：董事薪酬由股东会决定', '\U0001F534 高', '薪酬台账、派出单位工资单、任职合同'),
    (16, '考核与激励', '是否有容错纠错机制？对非主观性错误的免责标准是否明确？',
     '查阅容错纠错制度文件，评估免责条件、程序和适用范围',
     '国务院国资委关于建立容错纠错机制的指导意见', '\U0001F7E2 低', '容错纠错制度、免责申请记录'),

    ('五、法律合规（第17-18项）', light_red),
    (17, '法律合规', '公司是否投保了董事责任保险？',
     '查阅保险单、保费支付凭证、股东会审议记录',
     '新《公司法》第193条：公司可为董事投保责任保险', '\U0001F7E1 中', '保险单、保费凭证、股东会决议'),
    (18, '法律合规', '股东单位是否越权直接干预子公司经营管理？',
     '反向追踪法：选取近一年10项重大经营决策，追踪决策发起源头',
     '新《公司法》第59条：股东会职权、第67条：董事会职权', '\U0001F534 高', '重大经营决策档案、章程职权条款、决策流程文件'),
]

row = 4
for item in directors:
    if isinstance(item[0], str):
        write_section(ws1, row, item[0], item[1])
    else:
        write_row(ws1, row, item, light_gray)
    ws1.row_dimensions[row].height = 50 if isinstance(item[0], str) else 55
    row += 1

set_col_widths(ws1)

# =============================================
# Sheet 2: 监事履职检查清单（12条）
# =============================================
ws2 = wb.create_sheet('监事履职检查')

setup_sheet(ws2, '监事履职检查清单（12条）', blue_fill)
write_headers(ws2, 3, headers, blue_hdr)

supervisors = [
    ('一、任职资格与独立性（第1-3项）', light_blue),
    (1, '任职资格', '监事是否由董事、高管兼任？',
     '查阅监事名册和任职文件，与董事/高管名单交叉比对',
     '新《公司法》第117条：董事、高级管理人员不得兼任监事', '\U0001F534 高', '监事名册、任职文件、董事/高管名单'),
    (2, '任职资格', '外派监事是否有财务、法务或审计专业背景？',
     '查阅监事个人简历、资格证书，评估专业匹配度',
     '国资委关于加强监事会建设的指导意见', '\U0001F7E1 中', '监事简历、资格证书、任职文件'),
    (3, '任职资格', '外派监事是否经过岗前培训？是否清楚自身责任风险？',
     '查阅培训档案，抽取3名监事进行访谈测试',
     '国资委关于外派监事管理办法', '\U0001F7E1 中', '培训档案、访谈记录'),

    ('二、履职保障（第4-7项）', light_blue),
    (4, '履职保障', '监事会是否制定了议事规则？条款是否具体可操作？',
     '查阅监事会议事规则，核查议事程序和权限条款的明确性',
     '新《公司法》第78条：监事会职权和议事规则', '\U0001F7E1 中', '监事会议事规则'),
    (5, '履职保障', '是否有专门对接监事工作的部门或人员？',
     '查阅组织机构设置和岗位说明书',
     '国企公司治理示范企业评价标准', '\U0001F7E2 低', '组织机构图、岗位说明书'),
    (6, '履职保障', '履职企业是否按规定及时向监事提供财务、决策和经营资料？',
     '抽取近一年监事收到的文件清单，检验资料完整性和时效性',
     '新《公司法》第79条：公司应提供必要履职条件', '\U0001F534 高', '文件接收清单、资料签收记录'),
    (7, '履职保障', '涉及"三重一大"的会议是否提前通知监事参加并送达文件？',
     '抽取近一年10次重大会议，检查监事通知记录',
     '国企"三重一大"决策制度', '\U0001F534 高', '会议通知、文件送达记录'),

    ('三、履职记录与成效（第8-10项）', light_blue),
    (8, '履职记录', '监事是否有完整的履职工作底稿？',
     '查阅监事工作底稿，包含查阅资料记录、约谈记录、调研报告、书面质询等',
     '国资委外派监事管理办法', '\U0001F534 高', '工作底稿、查阅记录、约谈记录、调研报告'),
    (9, '履职记录', '监事是否在证券发行文件/定期报告上签署了书面确认意见？',
     '查阅近年年报/发债文件，核查监事签字页',
     '新《证券法》第82条：监事签署书面确认意见', '\U0001F534 高', '定期报告、监事签字页'),
    (10, '履职成效', '监事发现的问题是否形成书面报告并提交股东？有无整改跟踪？',
     '查阅监事工作报告、质询函、整改通知书，追踪问题闭环情况',
     '新《公司法》第78条：监事会监督职责', '\U0001F534 高', '监事工作报告、质询函、整改通知、跟踪记录'),

    ('四、制度与追责（第11-12项）', light_blue),
    (11, '制度保障', '是否建立了外派监事职责清单？清单内容是否可量化、可操作？',
     '查阅外派监事管理办法和职责清单文件',
     '国资委关于完善外派监事制度的指导意见', '\U0001F7E1 中', '外派监事管理办法、职责清单'),
    (12, '追责机制', '是否建立了监事尽职合规免责机制？认定程序是否清晰？',
     '查阅免责机制文件，评估免责条件的明确性和程序完备性',
     '国资委关于建立容错纠错机制的指导意见', '\U0001F7E1 中', '免责制度文件、免责申请案例'),
]

row = 4
for item in supervisors:
    if isinstance(item[0], str):
        write_section(ws2, row, item[0], item[1])
    else:
        write_row(ws2, row, item, light_gray)
    ws2.row_dimensions[row].height = 50 if isinstance(item[0], str) else 55
    row += 1

set_col_widths(ws2)

# =============================================
# Sheet 3: 三会运作规范性检查（10条）
# =============================================
ws3 = wb.create_sheet('三会运作规范性检查')

setup_sheet(ws3, '三会运作规范性检查清单（10条）', green_fill)
write_headers(ws3, 3, ['序号', '检查维度', '检查项', '检查方法', '法规依据', '风险等级', '审计证据'], green_hdr)

three_hui = [
    (1, '制度基础', '股东会、董事会、监事会是否有明确的议事规则且归档备查？',
     '查阅三会议事规则文件，检验完整性、时效性和备案情况',
     '新《公司法》第25条（章程必备条款）、各级国资委公司治理指引', '\U0001F534 高', '三会议事规则、章程、备案文件'),
    (2, '会议频次', '三会是否有年度会议计划？会议频次是否达标？',
     '查阅会议计划和近三年实际召开记录，比对法规要求',
     '新《公司法》第62条（股东会每年至少一次）、第123条（董事会每年至少两次）', '\U0001F7E1 中', '年度会议计划、近三年会议记录'),
    (3, '通知合规', '会议通知时间是否符合章程规定？议案资料是否提前送达？',
     '抽取近一年三会各2次，比对通知日期与章程规定天数',
     '新《公司法》第124条、章程关于通知时限的条款', '\U0001F7E1 中', '会议通知、送达签收记录'),
    (4, '记录完整', '会议记录是否完整（出席情况、表决过程、异议意见、决议内容）？',
     '抽取近一年三会各2次会议记录，逐项检查完整性',
     '新《公司法》第128条（会议记录要求）、第126条（异议记录）', '\U0001F534 高', '会议记录、签到表、表决票、异议声明'),
    (5, '决议跟踪', '决议执行是否有跟踪机制？落实情况是否定期向董事会/监事会反馈？',
     '抽取5项重大决议，追踪从决议日到执行完成的全链条',
     '国资委关于国有企业决策执行监督的要求', '\U0001F534 高', '决议文件、执行跟踪台账、反馈报告'),
    (6, '形式合规', '是否存在"未开会但有决议"或"传签替代会议"且无合理记录的情况？',
     '抽取年度全部决议事项，逐一核实对应的会议记录',
     '新《公司法》关于董事会/监事会会议形式的规定', '\U0001F534 高', '全部决议文件、对应的会议记录、传签文件'),
    (7, '权责边界', '股东会/董事会/监事会之间是否存在越权决策？',
     '反向追踪法：抽取近一年10项重大决策，逐项比对三会职权清单',
     '新《公司法》第59条（股东会职权）、第67条（董事会职权）、第78条（监事会职权）', '\U0001F534 高', '重大决策档案、三会职权清单、章程'),
    (8, '专委会运作', '是否设立了董事会专门委员会（审计、薪酬、提名等）？运作是否正常？',
     '查阅专委会设置文件、会议记录、工作报告',
     '上市公司治理准则、国资委关于专委会建设的指导意见', '\U0001F7E1 中', '专委会制度、会议记录、工作报告'),
    (9, '质效匹配', '会议频率与质量是否匹配企业规模和业务复杂度？',
     '对比同规模同行业企业的三会运作数据+评估年度议题覆盖率',
     '公司治理最佳实践标准', '\U0001F7E2 低', '行业对标数据、议题清单、年度报告'),
    (10, '预置条款', '对国有参股企业，是否在投资协议和章程中预先约定了董监事席位和行权方式？',
     '查阅投资协议和参股企业章程中的董监事条款',
     '国资委关于加强参股企业管理的通知', '\U0001F7E1 中', '投资协议、参股企业章程'),
]

row = 4
for item in three_hui:
    write_row(ws3, row, item, light_gray)
    ws3.row_dimensions[row].height = 55
    row += 1

set_col_widths(ws3)

# =============================================
# Sheet 4: 审计取证技巧
# =============================================
ws4 = wb.create_sheet('审计取证技巧')

setup_sheet(ws4, '审计取证技巧（3大核心方法+取证路径对照表）', dark_fill)
ws4['A1'].fill = dark_fill

# Part A: 三大取证技巧
ws4.merge_cells('A3:G3')
c = ws4['A3']
c.value = '一、三大核心取证技巧'
c.font = Font(name='Microsoft YaHei', size=12, bold=True, color='2C3E50')
c.fill = light_dark
c.border = thin_border
for col in range(2, 8):
    ws4.cell(row=3, column=col).fill = light_dark
    ws4.cell(row=3, column=col).border = thin_border

skill_headers = ['技巧', '原理说明', '适用场景', '操作步骤', '关键证据', '输出物', '注意事项']
write_headers(ws4, 4, skill_headers, dark_hdr)

skills = [
    ('反向追踪法',
     '先锁定子公司实际发生的重大决策事项，倒查这些事项是否经过了规定的董事会/监事会决策程序。从结果反推过程合规性。',
     '经责审计中对被审计单位重大投资、资产处置、大额资金使用等事项的合规性检查',
     '1.锁定近一年重大决策事项清单\n2.逐项检查是否有对应的董事会议案\n3.检查议案是否经过充分论证\n4.比对决议内容与实际执行的一致性',
     '重大决策清单、董事会议案、论证报告、会议决议',
     '重大决策合规性去向标表',
     '注意排除章程中已授权管理层直接决策的事项'),
    ('时间链比对法',
     '将股东内部审批日期与董事会会议日期进行比对——审批在前、会议在后且结果完全一致，是表决流于形式的强证据。',
     '外派董事的独立性检查、判断董事会是否为"橡皮图章"',
     '1.收集近一年重大决策的股东方审批文件（含审批日期）\n2.收集对应董事会会议记录（含会议日期）\n3.制表比对审批日期与会议日期的先后\n4.对"先批后议"事项逐项核实原因',
     '股东审批文件（含日期）、董事会会议通知/记录/决议',
     '审批-会议日期比对表',
     '审批日期与会议日期接近但会议在后属正常；审批日期显著早于会议日期+决议内容完全一致=强证据'),
    ('信息落差测试',
     '同时访谈外派董监事和内部高管，对同一重大事项的了解程度进行比对——差距悬殊则证明信息壁垒存在，外派人员无法实质性履职。',
     '判断外派董监事履职是否具备"知情"前提、参股企业外派人员的信息获取充分性检查',
     '1.选取3-5个近期重大事项\n2.分别访谈外派董事/监事和内部高管\n3.从"事项知晓时间、决策细节、后续进展"三个维度比对\n4.评分量化信息落差等级',
     '访谈记录（外派董事版+内部高管版）、重大事项清单',
     '信息落差评估对比表',
     '访谈应独立逐一进行，避免相互串联；差距分级：小(1-2周延迟)、中(缺失关键细节)、大(完全不知情)'),
]

row = 5
for item in skills:
    write_row(ws4, row, item, light_gray)
    ws4.row_dimensions[row].height = 110
    row += 1

# Part B: 取证路径对照表
row += 1
ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
c = ws4.cell(row=row, column=1, value='二、审计取证路径对照表（10条）')
c.font = Font(name='Microsoft YaHei', size=12, bold=True, color='2C3E50')
c.fill = light_dark
c.border = thin_border
for col in range(2, 8):
    ws4.cell(row=row, column=col).fill = light_dark
    ws4.cell(row=row, column=col).border = thin_border
row += 1

path_headers = ['序号', '审计目标', '取证方法', '关键证据材料', '取证技巧提示', '预期发现', '适用清单']
write_headers(ws4, row, path_headers, dark_hdr)
row += 1

paths = [
    (1, '董事任职合规', '查阅+比对', '公司章程、董事名册、聘任文件、外部董事人才库名单',
     '用企查查交叉验证工商注册董事与实际任职董事的一致性', '外部董事未过半、无人才库、兼职超限', '董事清单1-5项'),
    (2, '董事会运作规范性', '抽样+穿行测试', '董事会会议通知、签到表、会议记录、决议文件、表决票',
     '按重大决策类型（投资/资产处置/人事）各抽3项穿行', '先批后议、会议记录缺失、传签替代开会', '董事清单6-9项'),
    (3, '外部董事独立性', '访谈+问卷', '外部董事访谈记录、利益冲突声明、任职企业关联方清单',
     '访谈时注意"外派董事是否主动提及独立意见"', '董事附同性投票、不能独立表决', '董事清单7、15项'),
    (4, '信息传递有效性', '时间比对', '会议通知日期vs实际送达日期、议案材料完整性',
     '抽取紧急议案检查通知提前量', '议案未提前送达、材料不完整', '董事清单8项、监事清单6-7项'),
    (5, '履职勤勉程度', '量化分析', '履职台账（出席率、发言次数、调研天数）、年度述职报告',
     '出席率<75%即异常，发言记录为零即预警', '空壳董事、挂名监事', '董事清单10-12项、监事清单8-9项'),
    (6, '考核激励机制', '制度审查', '考核办法、薪酬发放记录、奖惩文件',
     '核对考核结果与薪酬的联动比例', '干好干坏一个样、无差异化激励', '董事清单13-16项'),
    (7, '监事监督实效', '正向追踪+反向验证', '监事工作报告、质询函、整改通知书',
     '选3个月监事报告中的问题，查是否整改闭环', '监事发现问题但未督促整改', '监事清单8-10项'),
    (8, '三会权责边界', '制度比对', '三会议事规则、章程职权条款、实际决策事项清单与权限对照',
     '选10项决策事项逐项标注决策主体', '股东越权、监事会空转', '三会清单7项'),
    (9, '外派人员权责对等', '制度审查+访谈', '外派管理办法、内部审批流程记录、表决指令文件',
     '比对表决票内容与股东审批单内容', '表决票内容与股东审批完全一致', '董事清单15、18项、监事清单11项'),
    (10, '责任保险覆盖', '查阅+确认', '董事责任保险保单、保费支付记录、股东会审议记录',
     '核对投保范围是否覆盖全体董事、保额是否合理', '未投保或投保范围不完整', '董事清单17项'),
]

for item in paths:
    write_row(ws4, row, item, light_dark)
    ws4.row_dimensions[row].height = 65
    row += 1

ws4.column_dimensions['A'].width = 6
ws4.column_dimensions['B'].width = 18
ws4.column_dimensions['C'].width = 30
ws4.column_dimensions['D'].width = 35
ws4.column_dimensions['E'].width = 30
ws4.column_dimensions['F'].width = 22
ws4.column_dimensions['G'].width = 20

# =============================================
# Save
# =============================================
out = r'D:\openclaw-workspace\output\成都市轨道资源公司-董事监事履职专项审计检查清单.xlsx'
wb.save(out)
print(f'Saved: {out}')
print('Sheets: 董事履职合规检查 / 监事履职检查 / 三会运作规范性检查 / 审计取证技巧')
