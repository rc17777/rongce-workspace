# -*- coding: utf-8 -*-
"""校服采购项目全量分析 Excel 生成脚本"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
import os

out = r'D:\openclaw-workspace\output\校服分析\校服2采购分析.xlsx'
wb = openpyxl.Workbook()

# ── Style definitions ──
hdr_font = Font(name='微软雅黑', bold=True, size=11)
hdr_fill = PatternFill('solid', fgColor='1F4E79')
hdr_font_w = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
title_font = Font(name='微软雅黑', bold=True, size=14, color='1F4E79')
sub_font = Font(name='微软雅黑', bold=True, size=11, color='1F4E79')
normal_font = Font(name='微软雅黑', size=10)
red_font = Font(name='微软雅黑', size=10, color='FF0000', bold=True)
orange_font = Font(name='微软雅黑', size=10, color='ED7D31', bold=True)
red_fill = PatternFill('solid', fgColor='FFF2CC')
green_fill = PatternFill('solid', fgColor='E2EFDA')
yellow_fill = PatternFill('solid', fgColor='FCE4D6')
wrap = Alignment(wrap_text=True, vertical='center')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, row, cols, fill=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font_w
        cell.fill = fill or hdr_fill
        cell.alignment = center
        cell.border = thin_border

def style_row(ws, row, cols, fonts=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = fonts or normal_font
        cell.alignment = wrap
        cell.border = thin_border

def auto_width(ws, min_w=10, max_w=45):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                lines = str(cell.value).split('\n')
                for line in lines:
                    # rough CJK width: CJK chars ~2x
                    length = sum(2 if ord(c)>127 else 1 for c in line)
                    max_len = max(max_len, length)
        ws.column_dimensions[col_letter].width = min(max(max_len+2, min_w), max_w)

# ═══════════════════════════════════════════
# Sheet 1: 报价分析
# ═══════════════════════════════════════════
ws = wb.active
ws.title = '报价分析'

data = [
    ['成都市教育科学研究院附属中学学生校服采购项目 - 报价全量分析', None],
    ['项目编号: TQ-CG-(2025)093 | 最高限价: 795元/全套 | 投标截止: 2025-10-12 10:00', None],
    [None, None],
    ['一、报价总览', None],
    ['排名','投标人','全套报价(元)','与限价比','折扣率','是否小微企业','价格扣除后(元)','备注'],
    ['1','四川乐吉玛帝诺服饰有限公司','645','81.1%','18.9%','是(小微企业)','580.50',''],
    ['2','四川牧森服饰有限公司','685','86.2%','13.8%','是(小微企业)','616.50',''],
    ['3','江苏苏美达伊顿纪德品牌管理有限公司','695','87.4%','12.6%','否(央企)','695.00','不享受价格扣除'],
    ['—','成都顺华服装有限公司','空白⚠️','—','—','是(小型企业)','—','报价栏全部空白→应认定无效投标'],
    ['—','弘博士服饰集团有限公司','待提取❓','—','—','待确认','—','.doc格式无法程序提取'],
    [None, None],
    ['二、单项报价对比（元）', None],
    ['项目','招标限价','乐吉玛帝诺','牧森','苏美达伊顿纪德','最大偏离'],
    ['春秋校服-长袖外套','112','80(-28.6%)','90(-19.6%)','105(-6.2%)','乐吉玛帝诺-28.6%'],
    ['春秋校服-长裤','76','70(-7.9%)','70(-7.9%)','58(-23.7%)','苏美达-23.7%'],
    ['夏季校服-T恤','65','60(-7.7%)','55(-15.4%)','52(-20.0%)','苏美达-20.0%'],
    ['夏季校服-夏长裤','54','50(-7.4%)','48(-11.1%)','52(-3.7%)','牧森-11.1%'],
    ['夏季校服-齐膝短裤','48','45(-6.2%)','—','45(-6.2%)','乐/苏-6.2%'],
    ['冬季校服-冲锋衣外套','220','180(含羽绒内胆)','—','195(-11.4%)','—'],
    ['冬季校服-内胆','140','含在冲锋衣中','—','108(-22.9%)','—'],
    ['冬季校服-冬裤','80','—','—','80(0.0%)','—'],
    [None, None],
    ['三、报价规律检测', None],
    ['检测项','结果','判断', None],
    ['报价区间','645 - 695元','正常', None],
    ['报价均值','675.00元','', None],
    ['极差','50元(7.8%)','正常', None],
    ['等差数列检测','差额序列[40, 10]','非等差数列，无操纵迹象', None],
    ['与限价比范围','81.1% - 87.4%','正常竞争区间', None],
    ['价格分最大差异','2.16分(645元vs695元)','价格竞争对总分影响极小', None],
]

for i, row in enumerate(data):
    for j, val in enumerate(row):
        if val is not None:
            ws.cell(row=i+1, column=j+1, value=val)

ws.merge_cells('A1:H1')
ws.merge_cells('A2:H2')
ws.merge_cells('A4:H4')
ws.merge_cells('A12:H12')
ws.merge_cells('A24:H24')

ws['A1'].font = title_font
ws['A2'].font = Font(name='微软雅黑', size=9, color='666666')
for r in [4, 12, 24]:
    ws.cell(row=r, column=1).font = sub_font

style_header(ws, 5, 8)
for r in [6,7,8,9,10,13,14,15,16,17,18,19,20,21,25,26,27,28,29,30]:
    style_row(ws, r, 8)

for r in [6,7,8,9,10]:
    for c in range(1, 9):
        ws.cell(row=r, column=c).alignment = center
for r in [13,14,15,16,17,18,19,20,21]:
    for c in range(1, 7):
        ws.cell(row=r, column=c).alignment = center
for r in [25,26,27,28,29,30]:
    for c in range(1, 4):
        ws.cell(row=r, column=c).alignment = center

# highlight
ws['C9'].font = red_font
ws['C9'].fill = red_fill
ws['D9'].font = red_font
ws['C10'].font = orange_font
ws['E14'].fill = yellow_fill
ws['F14'].font = red_font
ws['E17'].fill = yellow_fill
ws['F17'].font = red_font

auto_width(ws)

# ═══════════════════════════════════════════
# Sheet 2: 元数据分析
# ═══════════════════════════════════════════
ws2 = wb.create_sheet('元数据分析')

data2 = [
    ['投标文件元数据分析', None],
    ['招标文件发布: 2025-09-15 | 投标截止: 2025-10-12 10:00 | 红色=异常发现', None],
    ['投标人','文件','大小(MB)','文档作者','最后修改者','文档创建时间','文档最后修改(CST)','创建应用','异常标记'],
    ['乐吉玛帝诺','商务标.docx','89.0','1⚠️','┞iベLytton','2025-10-03','2025-10-12 03:32🔴','WPS 12.1.0.22529','作者名"1";凌晨3:32修改'],
    ['乐吉玛帝诺','资格标.docx','19.3','1⚠️','┞iベLytton','2025-10-03','2025-10-12 00:06','WPS 12.1.0.22529','作者名"1"'],
    ['牧森','商务标.docx','61.4','Administrator','旧时光','2024-12-10🔴','2025-10-11 11:27','WPS 12.1.0.23125','创建于招标前10个月→旧模板复用'],
    ['牧森','资格标.docx','10.1','Administrator','旧时光','2024-12-10🔴','2025-10-10 10:20','WPS 12.1.0.22529','资格/商务用不同WPS版本'],
    ['顺华','商务标.doc(ZIP)','110.1','张磊','徐奕','2025-08-05🟡','2025-10-11 23:16🔴','WPS 12.1.0.22529','招标前40天;报价空白;投标前夜修改'],
    ['顺华','资格标.doc(ZIP)','18.6','张磊','徐奕','2025-08-05🟡','2025-10-11 23:14🔴','WPS 12.1.0.22529','招标前40天;投标前夜修改'],
    ['苏美达伊顿纪德','商务标.docx','166.0','浮生六记','浮生六记','2025-09-23','2025-10-10 14:31','WPS 12.1.0.22529','招标发布后8天→正常'],
    ['苏美达伊顿纪德','资格标.docx','87.2','浮生六记','浮生六记','2025-09-23','2025-10-10 10:38','WPS 12.1.0.22529','正常'],
    ['弘博士','商务标.doc(OLE2)','112.5','—','—','—','2025-10-11 17:00(FS)','—','真.doc格式→元数据不可读'],
    ['弘博士','资格标.doc(OLE2)','11.6','—','—','—','2025-10-11 16:58(FS)','—','真.doc格式→元数据不可读'],
    [None, None],
    ['关键发现汇总', None],
    ['#','发现','风险等级','详情'],
    ['1','牧森文档创建于招标前10个月','🔴高','创建2024-12-10→招标2025-09-15→旧模板复用→可能存在旧项目信息残留'],
    ['2','顺华文档创建于招标前40天','🟡中','创建2025-08-05→提前一个多月→模板复用+报价空白→极可能未完成版本'],
    ['3','乐吉玛帝诺作者名为"1"','🟡中','非正常人名→WPS默认/未配置用户名→可能批量模板或有意隐藏身份'],
    ['4','乐吉玛帝诺凌晨3:32最后修改','🟡中','投标当天凌晨仍修改→距截止仅6.5小时→645元最低价可能为临时紧急降价'],
    ['5','4家公司使用相同WPS版本build号','🟡中','乐吉玛帝诺/牧森(资格)/顺华/苏美达均用12.1.0.22529(GUID相同)→4/4=100%'],
    ['6','顺华投标前夜23:16仍在修改','🟡中','距投标截止不足11小时→且报价栏最终为空白→极可能提交了错误版本'],
    ['7','牧森资格/商务标使用不同WPS版本','🟢低','资格标12.1.0.22529 vs 商务标12.1.0.23125→可能不同电脑制作→需关注'],
]

for i, row in enumerate(data2):
    for j, val in enumerate(row):
        if val is not None:
            ws2.cell(row=i+1, column=j+1, value=val)

ws2.merge_cells('A1:I1')
ws2.merge_cells('A2:I2')
ws2['A1'].font = title_font
ws2['A2'].font = Font(name='微软雅黑', size=9, color='666666')
style_header(ws2, 3, 9)
for r in range(4, 14):
    style_row(ws2, r, 9)
    for c in range(1, 10):
        ws2.cell(row=r, column=c).alignment = center

ws2.merge_cells('A14:I14')
ws2.cell(row=14, column=1).font = sub_font
style_header(ws2, 15, 4, fill=PatternFill('solid', fgColor='C00000'))

for r in range(16, 23):
    style_row(ws2, r, 4)
    ws2.cell(row=r, column=3).alignment = center

auto_width(ws2)

# ═══════════════════════════════════════════
# Sheet 3: 招标条件与评分分析
# ═══════════════════════════════════════════
ws3 = wb.create_sheet('招标条件与评分')

data3 = [
    ['招标条件设置合理性、合规性分析', None],
    ['采购方式: 公开招标(综合评分法) | 资金来源: 学生自有资金(非财政性) | 限价: 795元/全套 | 服务期: 一采三年', None],
    [None, None],
    ['一、合规性审查', None],
    ['审查项','合规状态','依据','说明'],
    ['适用法律','✅合规','非财政资金→适用《招标投标法》而非《政府采购法》','学生自有资金，不涉及财政性资金，采用公开招标方式合法'],
    ['采购方式','✅合规','《招标投标法》规定公开招标为首选方式','公开招标→符合"公开透明公平诚信"原则'],
    ['代理机构资质','✅合规','四川天启建设项目管理有限公司已在官网公示','具备招标代理资格→在自有系统(scbidding.com)发布'],
    ['招标公告发布','⚠️需关注','公告发布于代理机构官网(非法定平台)','公告在代理机构官网发布，未强制要求在法定公告平台同时发布→可能影响公告覆盖面'],
    ['招标文件售价','✅合规','300元/份→成本价售卖','收费合理，未禁止'],
    ['联合体投标','✅合规','明确不接受联合体','校服采购通常不接受联合体→合理'],
    ['备选方案','✅合规','不接受备选方案和多个报价','符合招标惯例'],
    ['资格审查方式','✅合规','资格后审','开标后评标阶段审查→符合常规流程'],
    ['投诉救济渠道','⚠️需关注','未明确设定独立投诉渠道','仅列联系方式→未说明异议/投诉的具体处理程序和时间要求'],
    [None, None],
    ['二、资格条件合理性分析', None],
    ['条件项','要求','合理性评估','风险说明'],
    ['基础资格','符合政府采购法22条','✅基本要求','无特殊行业资质要求→任何服装企业均可参与'],
    ['特殊条件','无','⚠️门槛偏低','未要求校服生产经验/检测报告/生产能力证明→低门槛可能导致不合格供应商进入'],
    ['注册资本','未要求','🟢合理(减负)','不设注册资本门槛符合营商环境优化方向'],
    ['业绩要求','列入评分(10分)而非资格条件','🟢合理','业绩作为加分项而非资格门槛→降低新企业进入壁垒'],
    ['信用要求','未被列入失信/重大税收违法','✅合理','标准信用审查条款'],
    ['行贿记录','投标人及法定代表人均不得有','✅合理','标准廉洁要求'],
    [None, None],
    ['三、评分体系合理性分析', None],
    ['评分项','权重','分值','评分类型','评分方式','自由裁量度','合理性评价'],
    ['报价','30%','30分','客观','公式计算(基准价/报价×30)','无','价格权重偏低'],
    ['实施方案','10%','10分','主观','评委评审(5项×2分)','高("缺陷"定义模糊)','评分标准不够细化'],
    ['售后服务方案','10%','10分','主观','评委评审(5项×2分)','高("缺陷"定义模糊)','缺乏量化指标'],
    ['样品评审','20%','20分','主观','盲样+评委感官评审','高(体验性指标)','主观性最强→评委个人偏好影响大'],
    ['现场述标','20%','20分','主观','评委自行打分(5项×4分)','极高(完全主观)','🔴争议最大→述标表现替代标书质量评估'],
    ['类似项目业绩','10%','10分','客观','数量计分(1个4分+每多1个3分)','无','无业绩上限→大企业碾压小企业'],
    [None, None],
    ['四、评分结构比重', None],
    ['类型','分值合计','占比','评价'],
    ['客观分(报价+业绩)','40分','40%','偏低(行业实践通常40-60%)'],
    ['主观分(方案+售后+样品+述标)','60分','60%','🔴偏高→评委自由裁量权过大'],
    [None, None],
    ['五、报价分模拟', None],
    ['投标人','报价(元)','小微企业扣除后','价格分','与最高分差'],
    ['乐吉玛帝诺','645','580.50','30.00','0'],
    ['牧森','685','616.50','28.25','1.75'],
    ['苏美达','695','695.00','27.84','2.16'],
    ['结论:价格分最大差异仅2.16分→样品20+述标20(共40分)才是决胜关键→价格竞争被边缘化'],
    [None, None],
    ['六、★实质性条款分析', None],
    ['类别','数量','主要内容','门槛评价','风险关注点'],
    ['安全质量标准','12项','甲醛/PH/色牢度/异味/纤维含量等','✅到位','校服安全底线→合理且必要'],
    ['服务承诺','8项','IP转让/线上禁售/统一招标/供货保障等','✅到位','保护学校品牌和家长权益→合理'],
    ['知识产权','2项','版权归采购人/中标人原创','✅到位','与业界实践一致'],
    ['其他商务','3项','合同一年一签/据实结算/不得分包','✅到位','一采三年的合理管理模式'],
    [None, None],
    ['七、条件设置风险总评', None],
    ['#','问题','风险等级','说明'],
    ['1','主观分占比60%过高','🔴高','述标20+样品20+方案20→评委个人判断占比过大→可能偏离"物有所值"原则'],
    ['2','述标不盲→投标人可见','🔴高','述标时投标人身份暴露→可能受人情关系影响→存在围猎评委风险'],
    ['3','价格权重仅30%','🟡中','国际上对"物"的采购通常40-60%→30%削弱价格竞争→不利于资金使用效率'],
    ['4','资格门槛过低','🟡中','无特殊行业资质要求→任何服装企业都可参与→围标串标门槛降低'],
    ['5','业绩无上限','🟡中','"4+3×n"无上限→大企业可用大量业绩碾压→中小企业无法竞争'],
    ['6','无履约保证金','🟢低','减轻企业负担但对违约约束力不足→一采三年模式下风险增大'],
    ['7','公告范围可能偏窄','🟡中','仅在代理机构自有平台公告→可能影响充分竞争(只招到5家)'],
]

for i, row in enumerate(data3):
    for j, val in enumerate(row):
        if val is not None:
            ws3.cell(row=i+1, column=j+1, value=val)

ws3.merge_cells('A1:G1')
ws3.merge_cells('A2:G2')
ws3['A1'].font = title_font
ws3['A2'].font = Font(name='微软雅黑', size=9, color='666666')

ws3.merge_cells('A4:G4'); ws3.cell(row=4, column=1).font = sub_font
style_header(ws3, 5, 5)
for r in range(6, 14):
    style_row(ws3, r, 5)
    ws3.cell(row=r, column=2).alignment = center

ws3.merge_cells('A15:G15'); ws3.cell(row=15, column=1).font = sub_font
style_header(ws3, 16, 5)
for r in range(17, 23):
    style_row(ws3, r, 5)

ws3.merge_cells('A24:G24'); ws3.cell(row=24, column=1).font = sub_font
style_header(ws3, 25, 7)
for r in range(26, 31):
    style_row(ws3, r, 7)

ws3.merge_cells('A32:G32'); ws3.cell(row=32, column=1).font = sub_font
style_header(ws3, 33, 4)
for r in range(34, 35):
    style_row(ws3, r, 4)

ws3.merge_cells('A36:G36'); ws3.cell(row=36, column=1).font = sub_font
style_header(ws3, 37, 5)
for r in range(38, 42):
    style_row(ws3, r, 5)

ws3.merge_cells('A43:G43'); ws3.cell(row=43, column=1).font = sub_font

ws3.merge_cells('A44:G44'); ws3.cell(row=44, column=1).font = sub_font
style_header(ws3, 45, 5)
for r in range(46, 49):
    style_row(ws3, r, 5)

ws3.merge_cells('A50:G50'); ws3.cell(row=50, column=1).font = sub_font
style_header(ws3, 51, 4)
for r in range(52, 59):
    style_row(ws3, r, 4)

auto_width(ws3)

# ═══════════════════════════════════════════
# Sheet 4: 面料偏离
# ═══════════════════════════════════════════
ws4 = wb.create_sheet('面料偏离')

data4 = [
    ['面料参数与招标要求偏离分析', None],
    ['★=实质性条款(不满足即废标) | 🔴=严重偏离 | 🟡=轻微偏离 | ±5%允差已考虑', None],
    ['投标人','品类','招标要求','实际参数','偏离幅度','严重程度','是否★条款','风险说明'],
    ['苏美达伊顿纪德','春秋校服克重','≥300g/m²','290g/m²','-3.3%','🟡轻微','否','在±5%允差范围内'],
    ['苏美达伊顿纪德','春秋面料棉含量','50%±5%(45%-55%)','45%','-5%(允差边界)','🟡轻微','否','刚好处在允差下限'],
    ['苏美达伊顿纪德','夏季长裤棉含量','65%±5%(60%-70%)','45%','-30.8%','🔴严重','★(技术参数)','远超±5%允差→可能触发无效投标'],
    ['苏美达伊顿纪德','冲锋衣面料克重','≥200g/m²','140g/m²','-30.0%','🔴严重','★(技术参数)','仅为招标要求的70%→严重缩水'],
    ['苏美达伊顿纪德','冲锋衣里料克重','≥140g/m²','60g/m²','-57.1%','🔴严重','★(技术参数)','仅为招标要求的43%→严重缩水'],
    ['苏美达伊顿纪德','冬裤克重','≥350g/m²','280g/m²','-20.0%','🔴严重','★(技术参数)','仅为招标要求的80%→大概率触发无效投标'],
    ['乐吉玛帝诺','冬季内胆材质','新雪丽(100%聚酯)','90鸭绒(可拆卸)','完全替代','🟡中等','★(技术参数)','材质从化纤棉→羽绒→需确认是否经采购人认可'],
    ['牧森','夏季长裤材质','65%棉35%聚酯','94%棉6%氨纶','配方完全不同','🟡中等','否(可调整)','招标允许调整面料成分→但差异需在分项报价表明确'],
    [None, None],
    ['苏美达伊顿纪德偏离汇总:', None],
    ['共4项严重实质性偏离→春秋克重-3.3%(在允差内)+夏裤棉-30.8%+冲锋衣面料-30%+里料-57%+冬裤-20%', None],
    ['判断:多处偏离远超±5%允差范围→建议评委会重点审查→存在被判定无效投标的风险', None],
]

for i, row in enumerate(data4):
    for j, val in enumerate(row):
        if val is not None:
            ws4.cell(row=i+1, column=j+1, value=val)

ws4.merge_cells('A1:H1')
ws4.merge_cells('A2:H2')
ws4['A1'].font = title_font
ws4['A2'].font = Font(name='微软雅黑', size=9, color='666666')

style_header(ws4, 3, 8)
for r in range(4, 12):
    style_row(ws4, r, 8)
    ws4.cell(row=r, column=6).alignment = center
    ws4.cell(row=r, column=7).alignment = center

ws4.merge_cells('A13:H13'); ws4.cell(row=13, column=1).font = sub_font
ws4.merge_cells('A14:H14'); ws4.cell(row=14, column=1).font = red_font
ws4.merge_cells('A15:H15'); ws4.cell(row=15, column=1).font = red_font

# highlight severe deviations
for r in [6,7,8,9]:
    ws4.cell(row=r, column=6).fill = red_fill

auto_width(ws4)

# ═══════════════════════════════════════════
# Sheet 5: 串标围标分析
# ═══════════════════════════════════════════
ws5 = wb.create_sheet('串标围标分析')

data5 = [
    ['串标围标专项分析', None],
    ['基于7维证据链评估 + 投标文件元数据交叉比对 + 报价规律检测', None],
    [None, None],
    ['一、7维度串标围标证据链评估', None],
    ['检测维度','检测内容','证据强度','判断','说明'],
    ['1.报价规律性','645/685/695分布','无证据','✅正常','差额[40,10,?]→无等差数列→无报价操纵特征'],
    ['2.文档作者交叉','各公司独立作者','无证据','✅正常','浮生六记/Administrator/1/张磊→各公司独立→无同一人制作多家标书迹象'],
    ['3.WPS版本一致性','4/4公司同一build号','🟡弱证据','⚠️需关注','WPS 12.1.0.22529(GUID相同)→100%一致→可能是广泛分发也可能协同一体'],
    ['4.文档创建时间','牧森提前10月/顺华提前40天','🟡弱证据','⚠️模板复用','非围标证据→但表明旧模板复用→需检查旧项目信息残留'],
    ['5.深夜修改','乐吉3:32AM/顺华23:16','极弱证据','✅赶工迹象','不构成围标证据→但体现异常投标节奏'],
    ['6.业绩交叉','牧森12+顺华14项目','无证据','✅正常','无同一学校→无联合围标嫌疑'],
    ['7.地域集中','3家成都企业','正常','✅正常市场分布','3/5为成都企业→无明显异常'],
    [None, None],
    ['二、串标围标综合判断', None],
    ['结论:基于现有可获取数据，未发现明确串标围标证据。', None],
    ['注意:以下数据缺口可能隐藏围标线索→建议补充后重新评估', None],
    [None, None],
    ['三、需要补充的关键数据', None],
    ['缺口项','重要性','建议获取途径','如发现异常的串标含义'],
    ['弘博士报价+元数据','🔴高','人工查看.doc文件中开标一览表','补充后可完成5家全量报价规律分析'],
    ['投标IP/签到信息','🔴高','向代理机构(天启建设)调取投标登记记录','同IP=围标铁证;连续签到=协商迹象'],
    ['工商关联信息','🟡中','天眼查/企查查查5家法人/股东/高管交叉','同一实控人=围标明确证据'],
    ['代理人信息','🟡中','调取开标现场签到表比对法定代表人/代理人','同一代理人=围标可疑证据'],
    ['保证金账户','🟡中','调取投标保证金汇款账户','同一账户=围标确认证据'],
    ['标书全文相似度','🟡中','商务标全量文本转换后TF-IDF比对','相似度>60%=协商制作可能'],
    ['顺华最终报价','🔴高','确认提交的纸质标书是否含报价','报价空白为无效投标;如有报价需排除版本问题'],
    ['牧森/苏美达PDF扫描件','🟡中','OCR提取PDF内容与docx对比','验证提交版本与电子版一致性'],
]

for i, row in enumerate(data5):
    for j, val in enumerate(row):
        if val is not None:
            ws5.cell(row=i+1, column=j+1, value=val)

ws5.merge_cells('A1:F1')
ws5.merge_cells('A2:F2')
ws5['A1'].font = title_font
ws5['A2'].font = Font(name='微软雅黑', size=9, color='666666')

ws5.merge_cells('A4:F4'); ws5.cell(row=4, column=1).font = sub_font
style_header(ws5, 5, 5)
for r in range(6, 13):
    style_row(ws5, r, 5)

ws5.merge_cells('A14:F14'); ws5.cell(row=14, column=1).font = sub_font
ws5.merge_cells('A15:F15'); ws5.cell(row=15, column=1).font = Font(name='微软雅黑', size=11, color='1F4E79')
ws5.merge_cells('A16:F16'); ws5.cell(row=16, column=1).font = orange_font

ws5.merge_cells('A18:F18'); ws5.cell(row=18, column=1).font = sub_font
style_header(ws5, 19, 5)
for r in range(20, 28):
    style_row(ws5, r, 5)

auto_width(ws5)

# ═══════════════════════════════════════════
# Sheet 6: 综合结论
# ═══════════════════════════════════════════
ws6 = wb.create_sheet('综合结论')

data6 = [
    ['校服采购项目全量分析 — 综合结论', None],
    ['分析日期: 2025-05-25 | 数据源: 5家投标文件+83页招标文件 | 分析维度: 报价+元数据+面料+关联+评分+合规', None],
    [None, None],
    ['一、投标有效性判定', None],
    ['投标人','有效性判定','判定依据'],
    ['成都顺华服装有限公司','🔴应认定无效投标','报价栏全部空白→招标文件第二章1条实质性要求+3.3.3(四)条'],
    ['江苏苏美达伊顿纪德品牌管理有限公司','🟡需评委会判定','面料5项实质性偏离(夏裤棉-30.8%/冲锋衣面料-30%/里料-57%/冬裤-20%)→远超±5%允差'],
    ['四川乐吉玛帝诺服饰有限公司','🟢形式有效','报价645元(最低)→羽绒替代新雪丽需采购人确认→其余响应完整'],
    ['四川牧森服饰有限公司','🟢形式有效','报价685元→文档创建于招标前10月(旧模板)→注意检查旧项目信息'],
    ['弘博士服饰集团有限公司','❓数据不足','.doc格式无法程序提取→需人工补充报价及元数据'],
    [None, None],
    ['二、招标文件条件评价', None],
    ['评价维度','评级','说明'],
    ['法规合规性','✅通过','适用《招标投标法》、招标方式/审查程序/公告/拒联合体均合规'],
    ['条件设置合理性','⚠️存疑','资格门槛偏低(无行业特殊要求)+主观分60%偏高→影响评审科学性'],
    ['评分公正性','🔴需关注','述标20分完全主观+不盲→存在围猎评委风险;价格仅30%→削弱竞争'],
    ['信息透明度','⚠️存疑','仅代理机构平台公告→覆盖面可能不足→5家投标人中3家成都本地'],
    [None, None],
    ['三、串标围标判断', None],
    ['检测维度','结论','依据'],
    ['报价规律','✅无异常','645-695元正常分布→无等差数列/异常集中'],
    ['文档作者','✅无异常','各公司作者独立→无交叉'],
    ['WPS版本','⚠️关注点','4/4相同build号→记录备查'],
    ['创建时间','⚠️关注点','模板复用→非围标证据'],
    ['最终判断','未发现明确串标围标证据','建议补充投标IP/工商关联/保证金核查后方可确认'],
    [None, None],
    ['四、中标概率排序(基于当前已知数据)', None],
    ['排名','投标人','有利因素','不利因素','综合评估'],
    ['1','乐吉玛帝诺','最低价+小微企业扣除+640有效价','羽绒替代★条款需确认','最具价格竞争力→但需过样品+述标关'],
    ['2','牧森','小微企业扣除后616.5+业绩好','旧模板复用需核查','价格有优势→述标+样品决定'],
    ['3','苏美达伊顿纪德','央企品牌+无价格扣除后=695','面料5项实质性偏离→可能被废标','评委裁决→若面料过关则有品牌优势→否则资格出局'],
    ['—','顺华','—','报价空白→应认定无效投标','建议从有效投标人中排除'],
    ['—','弘博士','—','数据缺失','无法评估'],
]

for i, row in enumerate(data6):
    for j, val in enumerate(row):
        if val is not None:
            ws6.cell(row=i+1, column=j+1, value=val)

ws6.merge_cells('A1:E1')
ws6.merge_cells('A2:E2')
ws6['A1'].font = title_font
ws6['A2'].font = Font(name='微软雅黑', size=9, color='666666')

ws6.merge_cells('A4:E4'); ws6.cell(row=4, column=1).font = sub_font
style_header(ws6, 5, 3)
for r in range(6, 11):
    style_row(ws6, r, 3)
    ws6.cell(row=r, column=2).alignment = center

ws6.merge_cells('A12:E12'); ws6.cell(row=12, column=1).font = sub_font
style_header(ws6, 13, 3, fill=PatternFill('solid', fgColor='4472C4'))
for r in range(14, 18):
    style_row(ws6, r, 3)

ws6.merge_cells('A19:E19'); ws6.cell(row=19, column=1).font = sub_font
style_header(ws6, 20, 3)
for r in range(21, 27):
    style_row(ws6, r, 3)

ws6.merge_cells('A28:E28'); ws6.cell(row=28, column=1).font = sub_font
style_header(ws6, 29, 5, fill=PatternFill('solid', fgColor='2C5F2D'))
for r in range(30, 35):
    style_row(ws6, r, 5)

auto_width(ws6)

# ── Set tab colors ──
ws.sheet_properties.tabColor = '2C5F2D'
ws2.sheet_properties.tabColor = 'C00000'
ws3.sheet_properties.tabColor = '4472C4'
ws4.sheet_properties.tabColor = 'ED7D31'
ws5.sheet_properties.tabColor = 'FF0000'
ws6.sheet_properties.tabColor = '7030A0'

# ── Save ──
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print(f'Done! Saved to: {out}')
print(f'Sheets: {wb.sheetnames}')
print(f'Size: {os.path.getsize(out)/1024:.1f} KB')
