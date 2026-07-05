#!/usr/bin/env python3
"""智析智能体 vs 审计厅供应商要求·对标分析+功能完善方案"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# 样式
# ============================================================
DARK = "1B3A5C"
MED = "2E6B9E"
LIGHT = "D6E4F0"
WHITE = "FFFFFF"
GRAY = "F5F5F5"
RED = "C0392B"
ORANGE = "E67E22"
GREEN = "27AE60"
YELLOW = "FFF8E1"

hfont = Font(name="微软雅黑", bold=True, color=WHITE, size=11)
tfont = Font(name="微软雅黑", bold=True, color=DARK, size=14)
sfont = Font(name="微软雅黑", bold=True, color=MED, size=12)
bfont = Font(name="微软雅黑", size=10)
bold = Font(name="微软雅黑", bold=True, size=10)
rfont = Font(name="微软雅黑", bold=True, color=RED, size=10)
gfont = Font(name="微软雅黑", bold=True, color=GREEN, size=10)
ofont = Font(name="微软雅黑", bold=True, color=ORANGE, size=10)
sfont9 = Font(name="微软雅黑", size=9, color="888888")

hfill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
lfill = PatternFill(start_color=LIGHT, end_color=LIGHT, fill_type="solid")
gfill = PatternFill(start_color=GRAY, end_color=GRAY, fill_type="solid")
yfill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
rfill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
gfill2 = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
wrap = Alignment(wrap_text=True, vertical="top")

thin = Border(
    left=Side(style="thin", color="00CCCCCC"),
    right=Side(style="thin", color="00CCCCCC"),
    top=Side(style="thin", color="00CCCCCC"),
    bottom=Side(style="thin", color="00CCCCCC"),
)

def header(ws, r, n):
    for c in range(1, n+1):
        cell = ws.cell(row=r, column=c)
        cell.font = hfont; cell.fill = hfill; cell.alignment = center; cell.border = thin

def data_row(ws, r, n, alt=False):
    for c in range(1, n+1):
        cell = ws.cell(row=r, column=c)
        cell.font = bfont; cell.alignment = left if c > 1 else center
        cell.border = thin
        if alt: cell.fill = gfill

def widths(ws, w):
    for i, v in enumerate(w, 1):
        ws.column_dimensions[get_column_letter(i)].width = v

def title(ws, r, c, txt, merge=None):
    cell = ws.cell(row=r, column=c, value=txt)
    cell.font = tfont
    if merge: ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=merge)

# ============================================================
# Sheet 1: 对标分析总览
# ============================================================
ws1 = wb.active
ws1.title = "1-对标分析总览"
widths(ws1, [5, 28, 22, 22, 32])

title(ws1, 1, 1, "智析智能体 vs 审计厅数据分析供应商要求 · 对标总览", 5)
ws1.cell(row=2, column=1, value="基于审计厅《数据分析服务要求》+ 智析智能体现有功能架构").font = sfont9

row = 4
ws1.cell(row=row, column=1, value="七大要求领域 · 覆盖度总览").font = sfont
row += 1

data = [
    ["序号", "审计厅要求", "智析现有能力", "覆盖度", "缺口说明"],
    ["1", "数据采集（行业+项目，40T/年，多格式）", "行业数据定期采集（zhixi_collector.exe）", "🟡 部分", "缺少项目数据采集、政务共享平台交换、网络数据采集"],
    ["2", "数据校验、清洗", "基本数据恢复", "🔴 极弱", "缺少完整性/总量/业务规则校验；无标准化清洗流程"],
    ["3", "数据迁移、标准化", "无", "🔴 缺失", "无多数据库迁移能力，无审计数据标准库建设"],
    ["4", "非结构化数据处理（OCR/关键要素提取）", "无", "🔴 缺失", "无OCR识别、无合同/招投标/会议纪要关键要素提取"],
    ["5", "数据维护（可视化/资源目录/备份/财务账套）", "Web界面+可视化", "🟡 部分", "缺少财务账套采集软件维护、数据资源目录、离线备份机制"],
    ["6", "审计模型构建", "内置分析逻辑", "🟡 部分", "无模型构建协作流程、无建模技术文档"],
    ["7", "大数据技术服务（挖掘/图谱/OCR/文本/培训）", "基础数据分析", "🔴 极弱", "无图谱挖掘、文本挖掘、现场技术支持、技术培训"],
]
for i, d in enumerate(data):
    for j, val in enumerate(d):
        ws1.cell(row=row, column=j+1, value=val)
    row += 1
header(ws1, row - len(data), 5)
for r in range(row - len(data), row):
    c3 = ws1.cell(row=r, column=4)
    if "缺失" in str(c3.value) or "极弱" in str(c3.value):
        c3.font = rfont
    elif "部分" in str(c3.value):
        c3.font = ofont
    else:
        c3.font = gfont

row += 1
ws1.cell(row=row, column=1, value="结论：智析在审计厅7大要求中，仅1项（数据采集）有基础能力，3项部分覆盖，3项几乎空白。需要系统性的功能补强。").font = bold

# ============================================================
# Sheet 2: 逐项对标与完善方案
# ============================================================
ws2 = wb.create_sheet("2-逐项完善方案")
widths(ws2, [5, 20, 22, 25, 30, 20])

title(ws2, 1, 1, "七大领域 · 逐项对标 + 功能完善方案", 6)

row = 3
areas = [
    # [序号, 审计厅要求, 智析现状, 差距, 完善方案, 优先级]
    ["1", "数据采集", "", "", "", ""],
    ["", "①行业数据定期采集\n②项目相关数据采集\n③多格式：Oracle/SQL Server/MySQL/DB2/Sybase/国产库+XML+非结构化\n④40T/年\n⑤多种采集方式", "✅ 行业数据采集器(zhixi_collector)\n❌ 无项目数据采集\n❌ 仅限部分数据库格式\n❌ 无国产数据库支持\n❌ 无政务共享平台对接\n❌ 无网络数据采集", "差距较大", "1.增加「项目数据采集」模块：支持按审计项目组织数据采集任务\n2.扩展数据库适配器：达梦/神通/人大金仓等国产库驱动\n3.增加数据源连接器：政务共享平台API对接、网络爬虫\n4.采集任务调度：支持定时、增量采集\n5.采集进度监控与日志", "🔴 P0"],
    ["2", "数据校验、清洗", "", "", "", ""],
    ["", "①数据恢复到指定数据库\n②完整性校验\n③总量/变量/业务规则校验\n④按审计规则清洗", "✅ 基本数据恢复\n❌ 无结构化校验规则\n❌ 无业务规则校验\n❌ 清洗流程不标准", "差距大", "1.构建「数据校验引擎」：完整性检查 / 总量核对 / 字段级校验 / 业务规则验证\n2.构建「数据清洗工作台」：去重/格式化/缺失值处理/异常值标记\n3.预设审计常用校验规则模板（资金平衡、科目对应、期间匹配等）\n4.输出校验报告+清洗日志", "🔴 P0"],
    ["3", "数据迁移、标准化", "", "", "", ""],
    ["", "①多数据库间迁移（→SQL Server/Oracle/国产库/大数据中心）\n②按审计行业数据规划标准化\n③构建审计数据标准库\n④工具不得另行收费", "❌ 完全缺失", "空白", "1.「数据迁移工具」：源库→目标库，支持全量/增量迁移\n2.「数据标准化引擎」：字段映射/编码统一/格式规范/审计行业标准模板\n3.内置国家标准：GB/T 数据元标准、审计署数据规划\n4.迁移+标准化一体化操作面板\n5.审计标准库管理：表结构/字典/元数据", "🔴 P0"],
    ["4", "非结构化数据处理", "", "", "", ""],
    ["", "①OCR识别：图片、扫描件→可编辑文档\n②关键要素提取：会议纪要/合同/招投标文件→结构化数据库表\n③标准化后形成可关联分析数据", "❌ 完全缺失", "空白", "1.「OCR识别模块」：集成PaddleOCR/Tesseract，支持扫描件、PDF图片\n2.「文档关键要素提取」：\n   - 合同：甲乙方/金额/标的/履约期限/违约责任\n   - 招投标：中标单位/金额/评标方法/评委\n   - 会议纪要：议题/决议/责任人/时间节点\n3.提取结果→标准化数据库表\n4.与结构化数据关联分析", "🟠 P1"],
    ["5", "数据维护", "", "", "", ""],
    ["", "①可视化工具维护数据资源\n②维护/提供财务账套采集软件\n③原始数据维护+离线备份\n④建立数据资源目录（按行业/项目/专题）\n⑤标准化数据维护+权限+备份", "✅ Web可视化界面\n❌ 无财务账套采集软件\n❌ 无数据资源目录\n❌ 无备份机制\n❌ 无权限管理", "差距较大", "1.「财务账套采集模块」：支持主流财务软件（用友/金蝶/浪潮）账套导入解析\n2.「数据资源目录」：按行业/项目/专题三维分类，可视化浏览\n3.「备份管理」：定时离线备份+备份状态监控+恢复验证\n4.「权限管理」：基于角色的数据访问控制\n5.升级现有Web可视化：数据血缘、表关系图", "🟠 P1"],
    ["6", "审计模型构建", "", "", "", ""],
    ["", "①协助构建审计数据分析模型\n②完善数据建模技术文档", "✅ 内置分析逻辑\n❌ 无模型构建协作流程\n❌ 无建模技术文档\n❌ 无模型管理/版本", "差距中", "1.「审计模型工作台」：SQL/Python双模式建模\n2.「模型库管理」：分类/版本/共享/复用\n3.内置审计常用模型模板（参照bid-document SQL模型库31个模型）\n4.模型文档自动生成\n5.模型运行+结果可视化+疑点标注", "🟡 P2"],
    ["7", "大数据技术服务", "", "", "", ""],
    ["", "①数据挖掘+图谱关系挖掘\n②OCR+文本挖掘+可视化\n③现场技术支持\n④数据模型核实修正\n⑤大数据技术/课题研究\n⑥技术培训", "✅ 基础数据可视化\n❌ 无图谱分析\n❌ 无文本挖掘\n❌ 无现场支持流程\n❌ 无培训体系", "差距大", "1.「关系图谱分析」：资金流向/股权穿透/供应商关系/人员关联\n2.「文本挖掘」：词频/共现/情感/热点分析\n3.「现场支持工具包」：便携版+离线运行能力\n4.「培训课程体系」：分层培训（基础/进阶/专题）\n5.技术文档+课题研究模板", "🟡 P2"],
]

section_row = None
alt = False
for i, d in enumerate(areas):
    if d[0] and d[0] != "":
        # Section header
        if section_row:
            ws2.merge_cells(start_row=section_row, start_column=1, end_row=row-1, end_column=1)
            ws2.cell(row=section_row, column=1).font = bold
            ws2.cell(row=section_row, column=1).alignment = center
            for c in range(1, 7):
                ws2.cell(row=section_row, column=c).fill = lfill

        section_row = row
        for j, val in enumerate(d):
            ws2.cell(row=row, column=j+1, value=val)
        row += 1
        # Detail row
        for j, val in enumerate(areas[i+1]):
            ws2.cell(row=row, column=j+1, value=val)
        row += 1
        alt = not alt
    elif d[0] == "":
        # Already handled above as detail row
        continue

# Handle last section
if section_row:
    ws2.merge_cells(start_row=section_row, start_column=1, end_row=row-1, end_column=1)
    ws2.cell(row=section_row, column=1).font = bold
    ws2.cell(row=section_row, column=1).alignment = center
    for c in range(1, 7):
        ws2.cell(row=section_row, column=c).fill = lfill

# Style detail rows
for r in range(4, row + 1):
    for c in range(1, 7):
        ws2.cell(row=r, column=c).font = bfont
        ws2.cell(row=r, column=c).alignment = wrap
        ws2.cell(row=r, column=c).border = thin
        if c == 6 and ws2.cell(row=r, column=c).value:
            val = str(ws2.cell(row=r, column=c).value)
            if "P0" in val:
                ws2.cell(row=r, column=c).font = rfont
            elif "P1" in val:
                ws2.cell(row=r, column=c).font = ofont
            elif "P2" in val:
                ws2.cell(row=r, column=c).font = gfont

# ============================================================
# Sheet 3: 功能完善路线图
# ============================================================
ws3 = wb.create_sheet("3-实施路线图")
widths(ws3, [5, 12, 20, 35, 15, 18, 18])

title(ws3, 1, 1, "功能完善 · 分阶段实施路线图", 7)

row = 3
roadmap = [
    ["阶段", "时间", "目标", "具体任务", "优先级", "所需资源", "产出物"],
    ["Phase 1\n基础补齐", "1-2个月", "满足审计厅最基础的\n数据采集+校验+迁移需求", "1.扩展数据库适配器（国产库）\n2.数据校验引擎（完整性+规则+业务）\n3.数据迁移工具（多库互迁）\n4.数据采集增加项目模式和API对接", "P0\n🔴紧急", "2人（1后端+1数据分析）\n开源组件零成本", "1.多数据库适配器\n2.数据校验引擎V1\n3.数据迁移工具V1"],
    ["Phase 2\n能力提升", "2-4个月", "补齐非结构化处理+数据维护\n达到供应商入围基本线", "1.OCR识别模块\n2.文档关键要素提取（合同/招投标/纪要）\n3.财务账套采集解析\n4.数据资源目录+备份+权限", "P1\n🟠重要", "3人（+1前端）\nPaddleOCR(免费)\nNLP模型(可接入大模型API)", "1.OCR识别模块\n2.文档提取引擎V1\n3.财务账套解析器\n4.资源目录+备份系统"],
    ["Phase 3\n智能升级", "4-6个月", "形成差异化竞争力\n具备投标资格", "1.审计模型工作台（SQL+Python双模）\n2.内置模型库（31+审计模型模板）\n3.关系图谱分析\n4.文本挖掘+可视化增强\n5.现场支持工具包", "P2\n🟡增强", "3-4人\nECharts/G6免费\nNetworkX图分析\nSQL模型库已有（bid-document）", "1.审计模型工作台\n2.模型库（31+模板）\n3.图谱分析模块\n4.现场工具包"],
    ["Phase 4\n持续运营", "持续", "驻场服务+培训+迭代\n形成长期合同", "1.驻场技术团队组建\n2.培训课程体系开发\n3.数据模型持续迭代\n4.课题研究参与\n5.知识库沉淀", "-", "驻场团队2-3人\n利用现有审计方法论+AI技能", "1.培训课程\n2.课题成果\n3.知识库"],
]
for i, d in enumerate(roadmap):
    for j, val in enumerate(d):
        ws3.cell(row=row, column=j+1, value=val)
    row += 1
header(ws3, row - len(roadmap), 7)
for r in range(row - len(roadmap), row):
    for c in range(1, 8):
        ws3.cell(row=r, column=c).alignment = wrap
        ws3.cell(row=r, column=c).border = thin
        ws3.cell(row=r, column=c).font = bfont
    # color priority col
    pri = str(ws3.cell(row=r, column=5).value or "")
    if "P0" in pri:
        ws3.cell(row=r, column=5).fill = rfill
        ws3.cell(row=r, column=5).font = rfont
    elif "P1" in pri:
        ws3.cell(row=r, column=5).fill = yfill
        ws3.cell(row=r, column=5).font = ofont
    elif "P2" in pri:
        ws3.cell(row=r, column=5).fill = gfill2
        ws3.cell(row=r, column=5).font = gfont

# ============================================================
# Sheet 4: 功能模块架构
# ============================================================
ws4 = wb.create_sheet("4-目标功能架构")
widths(ws4, [5, 22, 35, 25, 18])

title(ws4, 1, 1, "智析智能体 · 目标功能架构 + 模块清单", 5)

row = 3
arch = [
    ["层级", "模块名称", "功能点", "对接审计厅要求", "开发状态"],
    ["数据接入层", "多源数据采集", "行业数据/项目数据/API对接/网络爬虫/政务平台", "1.数据采集", "🟡 升级中"],
    ["", "国产数据库适配", "达梦/神通/人大金仓/GaussDB/OceanBase", "1.数据采集", "🔴 待开发"],
    ["", "财务账套解析", "用友/金蝶/浪潮/新中大/sap", "5.数据维护", "🔴 待开发"],
    ["数据处理层", "数据校验引擎", "完整性/总量/字段/业务规则/统计抽样", "2.数据校验", "🔴 待开发"],
    ["", "数据清洗工作台", "去重/格式化/缺失处理/异常标记/审计规则清洗", "2.数据清洗", "🔴 待开发"],
    ["", "数据迁移工具", "全量/增量迁移/多库互迁/进度监控", "3.数据迁移", "🔴 待开发"],
    ["", "数据标准化引擎", "字段映射/编码统一/审计数据规划/元数据管理", "3.数据标准化", "🔴 待开发"],
    ["智能分析层", "审计模型工作台", "SQL/Python双模/模型库/模板/版本/文档", "6.审计模型构建", "🟡 升级中"],
    ["", "关系图谱分析", "资金流向/股权穿透/供应商关联/人员关系", "7.图谱挖掘", "🔴 待开发"],
    ["", "文本挖掘引擎", "词频/共现/热点/异常文本检测", "7.文本挖掘", "🔴 待开发"],
    ["", "文档要素提取", "合同条款/招投标信息/会议决议/关键词提取", "4.非结构化处理", "🔴 待开发"],
    ["", "OCR识别模块", "扫描件/图片→文本/表格结构化", "4.非结构化处理", "🔴 待开发"],
    ["基础服务层", "数据资源目录", "行业/项目/专题三维目录", "5.数据维护", "🔴 待开发"],
    ["", "备份与安全", "离线备份/状态监控/权限管理/审计日志", "5.数据维护", "🔴 待开发"],
    ["", "可视化仪表盘", "数据血缘/表关系/资源地图/监控大盘", "5.可视化", "🟡 升级中"],
    ["", "培训与支持", "在线培训/现场支持工具包/知识库", "7.技术培训", "🔴 待开发"],
]
for i, d in enumerate(arch):
    for j, val in enumerate(d):
        ws4.cell(row=row, column=j+1, value=val)
    row += 1
header(ws4, row - len(arch), 5)
for r in range(row - len(arch), row):
    for c in range(1, 6):
        ws4.cell(row=r, column=c).border = thin
        ws4.cell(row=r, column=c).font = bfont
        ws4.cell(row=r, column=c).alignment = wrap
    dev = str(ws4.cell(row=r, column=5).value or "")
    if "待开发" in dev:
        ws4.cell(row=r, column=5).fill = rfill
    elif "升级" in dev:
        ws4.cell(row=r, column=5).fill = yfill

# ============================================================
# Sheet 5: 与融策现有资产联动
# ============================================================
ws5 = wb.create_sheet("5-融策资产联动")
widths(ws5, [5, 22, 22, 35])

title(ws5, 1, 1, "融策现有AI资产 × 智析功能完善 · 联动复用矩阵", 4)

row = 3
linkage = [
    ["融策资产", "可直接赋能智析的功能", "复用方式"],
    ["SQL审计模型库（bid-document 31个模型）", "审计模型工作台→模型模板库", "直接内嵌为智析模型模板，覆盖预算执行/采购/社保/经责等7大类"],
    ["cot-capture（审计思维链沉淀）", "审计规则库→数据校验业务规则", "将专家判断经验转化为数据校验规则引擎的If/Then条件"],
    ["prompt-librarian（提示词资产库）", "文档要素提取的Prompt模板", "合同/招投标/纪要的提取提示词直接复用"],
    ["agent-data-standard（12项检查）", "数据标准化引擎→数据质量检查", "12项标准内嵌为数据标准化后的质检规则"],
    ["procurement-audit-models（采购审计模型）", "图谱分析→供应商关联分析", "围标串标检测模型转化为图谱分析的关系规则"],
    ["audit-data-analysis-methods（7大方法）", "审计模型工作台→分析方法模板", "描述性/相关性/聚类/异常检测等方法内嵌为分析模板"],
    ["audit-knowledge-graph（知识图谱）", "关系图谱分析模块", "Neo4j+Cypher的图分析方案直接参考"],
    ["digital-audit-methodology（数字化审计方法论）", "培训课程体系+操作手册", "10大框架作为培训内容骨架"],
    ["workflow-embedder（6种嵌入模式）", "驻场服务→流程嵌入方案", "帮助审计厅设计数据服务嵌入审计流程的方式"],
]
for i, d in enumerate(linkage):
    for j, val in enumerate(d):
        ws5.cell(row=row, column=j+1, value=val)
    row += 1
header(ws5, row - len(linkage), 4)

# ============================================================
# 保存
# ============================================================
out = r"D:\openclaw-workspace\output\智析智能体-功能完善方案.xlsx"
wb.save(out)
print(f"Saved: {out}")
