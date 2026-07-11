# -*- coding: utf-8 -*-
import os, sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

base = r'C:\Users\scrccpa\Desktop\林芝樾~1'
for root, dirs, files in os.walk(base):
    for f in files:
        if f.startswith('~$'):
            continue
        if f.endswith('.xlsx'):
            fp = os.path.join(root, f)
            fs = os.path.getsize(fp)
            print(f'FILE: {f} ({fs:,} bytes)')
            try:
                wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
                for sn in wb.sheetnames:
                    ws = wb[sn]
                    r1 = [c.value for c in next(ws.iter_rows(max_row=1))]
                    last_row = [c.value for c in next(ws.iter_rows(min_row=ws.max_row))]
                    print(f'  [{sn}] {ws.max_row}R x {ws.max_column}C')
                    h = [str(v) for v in r1 if v]
                    l = [str(v) for v in last_row if v]
                    print(f'    Hdr: {h[:8]}')
                    print(f'    Last: {l[:8]}')
                wb.close()
            except Exception as e:
                print(f'  ERR: {e}')
            print()
