# -*- coding: utf-8 -*-
"""生成健康照护师成本构成测算Excel v2 — 补充SP模特、更新场地/考务/人工"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============ Styles ============
title_font = Font(name='微软雅黑', size=16, bold=True)
h1_font = Font(name='微软雅黑', size=14, bold=True)
h2_font = Font(name='微软雅黑', size=12, bold=True)
h3_font = Font(name='微软雅黑', size=11, bold=True)
normal_font = Font(name='微软雅黑', size=10)
bold_font = Font(name='微软雅黑', size=10, bold=True)
red_font = Font(name='微软雅黑', size=10, bold=True, color='CC0000')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
sub_header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
total_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
money_fmt = '#,##0.00'

def style_header_row(ws, row, max_col):
    for c in range(1, max_col+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def sc(ws, row, col, align='center'):
    cell = ws.cell(row=row, column=col)
    cell.font = normal_font
    cell.alignment = center_align if align == 'center' else left_align
    cell.border = thin_border
    return cell

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ============ Data ============
# Cost components from source data
equip_cost = {'初级': 93, '中级': 158, '高级': 190}
consumable_cost = {'初级': 87.20, '中级': 167.28, '高级': 280.49}
policy_theory = {'初级': 30, '中级': 35, '高级': 40}
policy_practical = {'初级': 140, '中级': 190, '高级': 240}

# Updated parameters
platform_fee = 20  # 考务平台费 (was 30)
venue_daily = 3200  # 场地费/天
venue_per_person = round(3200 / 30, 2)  # 106.67

# SP模特成本
sp_teacher_rate = 60  # 老师/小时
sp_student_rate = 22  # 学生/小时
sp_count_teacher = 1
sp_count_student = 2
sp_hours = 6  # 实操约6小时
sp_daily_cost = sp_count_teacher * sp_teacher_rate * sp_hours + sp_count_student * sp_student_rate * sp_hours
sp_per_person = round(sp_daily_cost / 30, 2)  # 20.80

# 人工费用 (院长办公会纪要第23期 2022.12.12 费率)
# 考评员: 100元/小时, 其他: 60元/小时
# 健康照护师8考题, 估计2-3考站, 30人/天
rate_examiner = 100  # 考评员
rate_staff = 60      # 其他工作人员

# --- 理论部分 (2.5h) ---
theory_invigilators = 4  # 2考场×2人
theory_hours = 2.5
theory_tech = 1
theory_tech_hours = 3.5
theory_daily = theory_invigilators * theory_hours * rate_staff + theory_tech * theory_tech_hours * rate_staff

# --- 实操部分 (6h) ---
prac_stations = 2  # 考站数(8考题轮转,2站并行)
prac_examiners = prac_stations * 3  # 每站3名考评员
prac_examiner_hours = 6
prac_invigilators = prac_stations * 2 + 2  # 每站2监考+候考室2
prac_invigilator_hours = 6
prac_daily = (prac_examiners * prac_examiner_hours * rate_examiner +
              prac_invigilators * prac_invigilator_hours * rate_staff)

# --- SP模特 (已单独计算) ---
# sp_daily_cost already computed

# --- 管理协调 ---
mgmt_staff = 5  # 主考/考务×2/督导/综合(30人小班次可精简)
mgmt_hours = 8
mgmt_daily = mgmt_staff * mgmt_hours * rate_staff

# --- 考场布置 ---
setup_staff = 5
setup_hours = 1
setup_daily = setup_staff * setup_hours * rate_staff

# --- 合计 ---
daily_labor_total = theory_daily + prac_daily + sp_daily_cost + mgmt_daily + setup_daily
labor_per_person = round(daily_labor_total / 30, 2)

# Print for verification
print(f'理论部分: {theory_daily:.0f}')
print(f'实操考评员: {prac_daily:.0f}')
print(f'SP模特: {sp_daily_cost:.0f}')
print(f'管理协调: {mgmt_daily:.0f}')
print(f'考场布置: {setup_daily:.0f}')
print(f'每日人工合计: {daily_labor_total:.0f}')
print(f'单人人工成本: {labor_per_person:.2f}')
print(f'场地单人: {venue_per_person}')
print(f'SP单人: {sp_per_person}')

# =====================================================
# Sheet 1: 成本构成总览
# =====================================================
ws1 = wb.active
ws1.title = '成本构成总览'

ws1.merge_cells('A1:J1')
ws1['A1'].value = '健康照护师（长期照护师）职业技能等级认定 — 单人成本构成测算'
ws1['A1'].font = title_font
ws1['A1'].alignment = center_align

ws1.merge_cells('A2:J2')
ws1['A2'].value = '测算依据：院长办公会纪要(第23期)2022.12.12 | 智护学院耗材成本表2026.05.14 | 川发改价格〔2017〕472号 | 编制：四川融策会计师事务所'
ws1['A2'].font = Font(name='微软雅黑', size=9, color='666666')
ws1['A2'].alignment = center_align

ws1.merge_cells('A3:J3')
ws1['A3'].value = '测算假设：每班次30人 | 实操6小时/2考站并行(8考题轮转) | SP模特3人(1老师+2学生) | 场地3200元/天'
ws1['A3'].font = Font(name='微软雅黑', size=9, color='666666')
ws1['A3'].alignment = center_align

row = 5
headers = ['序号', '成本项目', '测算依据/方法', '五级/初级\n(元/人)', '四级/中级\n(元/人)', '三级/高级\n(元/人)',
           '三级vs初级\n增幅%', '数据状态', '备注']
for c, h in enumerate(headers, 1):
    ws1.cell(row=row, column=c, value=h)
style_header_row(ws1, row, len(headers))

# Build overview data
p5_cost = policy_theory['初级'] + policy_practical['初级']
p4_cost = policy_theory['中级'] + policy_practical['中级']
p3_cost = policy_theory['高级'] + policy_practical['高级']

subtotal_measured_5 = p5_cost + platform_fee + labor_per_person + equip_cost['初级'] + consumable_cost['初级'] + venue_per_person
subtotal_measured_4 = p4_cost + platform_fee + labor_per_person + equip_cost['中级'] + consumable_cost['中级'] + venue_per_person
subtotal_measured_3 = p3_cost + platform_fee + labor_per_person + equip_cost['高级'] + consumable_cost['高级'] + venue_per_person

overview_data = [
    ['', '一、政策规定收费（政府定价）', '', '', '', '', '', '', ''],
    ['1', '理论考试费', '川发改价格〔2017〕472号', 30, 35, 40, 33.3, '政府定价', ''],
    ['2', '操作技能考核费', '川发改价格〔2017〕472号', 140, 190, 240, 71.4, '政府定价', ''],
    ['', '  政策收费小计', '', 170, 225, 280, 64.7, '', ''],
    ['', '', '', '', '', '', '', '', ''],
    ['', '二、实际发生成本', '', '', '', '', '', '', ''],
    ['3', '考务平台费', '固定20元/人', 20, 20, 20, 0, '✓已测算', ''],
    ['4', '人工费用', '院长办公会纪要(第23期)\n考评员100/h 其他60/h', labor_per_person, labor_per_person, labor_per_person, 0, '✓已测算', f'详见表"人工费用明细"'],
    ['4a', '  其中：SP模特', '1老师60/h+2学生22/h×6h÷30人', sp_per_person, sp_per_person, sp_per_person, 0, '✓已测算', '新增项'],
    ['5', '设施设备使用费', '磨损费/组均摊，详见表"设施设备成本明细"', 93, 158, 190, 104.3, '✓已测算', ''],
    ['6', '耗材成本', '逐项计量，详见表"耗材成本明细"', 87.20, 167.28, 280.49, 221.7, '✓已测算', '实操8题全量耗材'],
    ['7', '场地水电费', '3200元/天÷30人/班次', venue_per_person, venue_per_person, venue_per_person, 0, '✓已测算', ''],
    ['8', '全程视频监控费', '监控系统投入÷年限÷年人次', '', '', '', '', '待补充', '需学院提供设备折旧数据'],
    ['9', '证书制作费', '证书工本费单价', '', '', '', '', '待补充', '需学院提供'],
    ['10', '文印广告费', '试卷+标识标牌÷考核人次', '', '', '', '', '待补充', '需学院提供'],
    ['', '  实际成本小计（已测算）', '', '', '', '', '', '', ''],
    ['', '', '', '', '', '', '', '', ''],
    ['', '三、全口径成本汇总', '', '', '', '', '', '', ''],
    ['', '全口径成本（7项已测算）', '', subtotal_measured_5, subtotal_measured_4, subtotal_measured_3, '', '', '=政策收费+实际成本(已测算)'],
    ['', '', '', '', '', '', '', '', ''],
    ['', '四、与现行收费方案对比', '', '', '', '', '', '', ''],
    ['', '现行收费标准（方案5稿）', '', 320, 405, 500, 56.3, '', 'B类-健康照护师'],
    ['', '已测算7项成本', '', subtotal_measured_5, subtotal_measured_4, subtotal_measured_3, '', '', ''],
    ['', '差额（收费-已测算成本）', '', round(320 - subtotal_measured_5, 2), round(405 - subtotal_measured_4, 2), round(500 - subtotal_measured_3, 2), '', '', ''],
]

for i, d in enumerate(overview_data):
    r = row + 1 + i
    for ci, v in enumerate(d, 1):
        cell = sc(ws1, r, ci, 'left' if ci in [2, 3] else 'center')
        cell.value = v
        if ci in [4, 5, 6] and isinstance(v, (int, float)):
            cell.number_format = money_fmt

    first_val = str(d[1]) if d[1] else str(d[0])
    if any(first_val.startswith(x) for x in ['一、', '二、', '三、', '四、']):
        ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        for ci in range(1, 10):
            ws1.cell(row=r, column=ci).font = h3_font
            ws1.cell(row=r, column=ci).fill = sub_header_fill
            ws1.cell(row=r, column=ci).border = thin_border

    if '小计' in str(d[1]) or '全口径' in str(d[1]) or '差额' in str(d[1]) or '现行收费' in str(d[1]) or '已测算7项' in str(d[1]):
        for ci in range(1, 10):
            ws1.cell(row=r, column=ci).font = bold_font
            ws1.cell(row=r, column=ci).fill = total_fill
        if '差额' in str(d[1]):
            for ci in [4, 5, 6]:
                ws1.cell(row=r, column=ci).font = red_font

# Fill in calculated subtotals
for i, d in enumerate(overview_data):
    r = row + 1 + i
    if '实际成本小计' in str(d[1]):
        for ci, lvl in enumerate(['初级', '中级', '高级'], 4):
            total = platform_fee + labor_per_person + equip_cost[lvl] + consumable_cost[lvl] + venue_per_person
            ws1.cell(row=r, column=ci).value = total
            ws1.cell(row=r, column=ci).number_format = money_fmt

set_widths(ws1, [6, 28, 32, 14, 14, 14, 12, 12, 28])

# =====================================================
# Sheet 2: 人工费用明细
# =====================================================
ws2 = wb.create_sheet('人工费用明细')

ws2.merge_cells('A1:H1')
ws2['A1'].value = '健康照护师（长期照护师）职业技能等级认定 — 人工费用测算明细'
ws2['A1'].font = title_font
ws2['A1'].alignment = center_align

ws2.merge_cells('A2:H2')
ws2['A2'].value = '执行依据：四川护理职业学院院长办公会纪要（第23期）2022年12月12日 — 考评员100元/小时，其他工作人员60元/小时'
ws2['A2'].font = Font(name='微软雅黑', size=9, color='666666')
ws2['A2'].alignment = center_align

ws2.merge_cells('A3:H3')
ws2['A3'].value = '测算假设：每日1班次30人 | 实操6小时(含准备)2考站并行(8考题轮转) | SP模特3人(1老师+2学生)全程 | 健康照护师全国联考'
ws2['A3'].font = Font(name='微软雅黑', size=9, color='666666')
ws2['A3'].alignment = center_align

row = 5
hdrs2 = ['序号', '人员类别', '人数', '费率(元/小时)', '工作时长(小时)', '日费用(元)', '单人成本(元/人)', '计算说明']
for c, h in enumerate(hdrs2, 1):
    ws2.cell(row=row, column=c, value=h)
style_header_row(ws2, row, len(hdrs2))

# Build labor detail
labor_items = [
    # (seq, category, count, rate, hours, note)
    ['', '一、理论考试部分', '', '', '', '', '', ''],
    ['1', '理论监考', theory_invigilators, rate_staff, theory_hours, '', '', f'{theory_invigilators}人×{theory_hours}h (2考场×2人)'],
    ['2', '技术保障', theory_tech, rate_staff, theory_tech_hours, '', '', f'{theory_tech}人×{theory_tech_hours}h (考前部署+考试保障)'],
    ['', '  理论部分小计', '', '', '', '', '', ''],
    ['', '', '', '', '', '', '', ''],
    ['', '二、实操考试部分', '', '', '', '', '', ''],
    ['3', '实操考评员', prac_examiners, rate_examiner, prac_examiner_hours, '', '', f'{prac_examiners}人×{prac_examiner_hours}h ({prac_stations}考站×3人, 8考题轮转)'],
    ['4', 'SP模特-老师', sp_count_teacher, sp_teacher_rate, sp_hours, '', '', f'{sp_count_teacher}人×{sp_hours}h (标准化病人)'],
    ['5', 'SP模特-学生', sp_count_student, sp_student_rate, sp_hours, '', '', f'{sp_count_student}人×{sp_hours}h (标准化病人)'],
    ['6', '实操监考', prac_invigilators, rate_staff, prac_invigilator_hours, '', '', f'{prac_invigilators}人×{prac_invigilator_hours}h ({prac_stations}站×2人+候考室2人)'],
    ['', '  实操部分小计', '', '', '', '', '', ''],
    ['', '', '', '', '', '', '', ''],
    ['', '三、管理与保障', '', '', '', '', '', ''],
    ['7', '主考', 1, rate_staff, mgmt_hours, '', '', '全程负责'],
    ['8', '考务人员', 2, rate_staff, mgmt_hours, '', '', '2人'],
    ['9', '督导员', 1, rate_staff, mgmt_hours, '', '', ''],
    ['10', '综合管理', 1, rate_staff, mgmt_hours, '', '', '考前统筹+当天协调'],
    ['', '  管理部分小计', '', '', '', '', '', ''],
    ['', '', '', '', '', '', '', ''],
    ['', '四、考场布置', '', '', '', '', '', ''],
    ['11', '考场布置人员', setup_staff, rate_staff, setup_hours, '', '', f'{setup_staff}人×{setup_hours}h (考前一天+考后还原)'],
    ['', '', '', '', '', '', '', ''],
    ['', '五、合计', '', '', '', '', '', ''],
    ['', '每日人工费总计', '', '', '', daily_labor_total, '', ''],
    ['', '单人人工成本', '', '', '', '', labor_per_person, f'{daily_labor_total:.0f}÷30人'],
]

labor_totals = {
    '理论部分小计': theory_daily,
    '实操部分小计': prac_daily + sp_daily_cost,
    '管理部分小计': mgmt_daily,
}

for i, item in enumerate(labor_items):
    r = row + 1 + i
    seq, cat, count, rate, hours, daily, per_person, note = item[0], item[1], item[2], item[3], item[4], item[5] if len(item) > 5 else '', item[6] if len(item) > 6 else '', item[7] if len(item) > 7 else ''

    for ci, v in enumerate(item, 1):
        cell = sc(ws2, r, ci, 'left' if ci in [2, 8] else 'center')
        cell.value = v if v != '' else None

    # Calculate daily and per-person for data rows
    if isinstance(count, (int, float)) and count > 0:
        daily_val = count * rate * hours
        per_person_val = round(daily_val / 30, 2)
        ws2.cell(row=r, column=6).value = daily_val
        ws2.cell(row=r, column=6).number_format = money_fmt
        ws2.cell(row=r, column=7).value = per_person_val
        ws2.cell(row=r, column=7).number_format = money_fmt

    # Section headers
    if cat and any(cat.startswith(x) for x in ['一、', '二、', '三、', '四、', '五、']):
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        for ci in range(1, 9):
            ws2.cell(row=r, column=ci).font = h3_font
            ws2.cell(row=r, column=ci).fill = sub_header_fill
            ws2.cell(row=r, column=ci).border = thin_border

    # Subtotals / Totals
    if '小计' in str(cat):
        for ci in range(1, 9):
            ws2.cell(row=r, column=ci).font = bold_font
            ws2.cell(row=r, column=ci).fill = total_fill
        section_total = labor_totals.get(cat.replace('  ', ''), 0)
        if section_total > 0:
            ws2.cell(row=r, column=6).value = section_total
            ws2.cell(row=r, column=6).number_format = money_fmt
            ws2.cell(row=r, column=7).value = round(section_total / 30, 2)
            ws2.cell(row=r, column=7).number_format = money_fmt

    if '每日人工费总计' in str(cat):
        for ci in range(1, 9):
            ws2.cell(row=r, column=ci).font = bold_font
            ws2.cell(row=r, column=ci).fill = total_fill

    if '单人人工成本' in str(cat):
        for ci in range(1, 9):
            ws2.cell(row=r, column=ci).font = bold_font
            ws2.cell(row=r, column=ci).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

set_widths(ws2, [6, 22, 8, 14, 16, 14, 16, 40])

# =====================================================
# Sheet 3: 耗材成本明细
# =====================================================
ws3 = wb.create_sheet('耗材成本明细')

ws3.merge_cells('A1:J1')
ws3['A1'].value = '健康照护师（长期照护师）等级认定 — 耗材成本明细测算'
ws3['A1'].font = title_font
ws3['A1'].alignment = center_align

ws3.merge_cells('A2:J2')
ws3['A2'].value = '数据来源：智护学院-长期照护师-耗材成本表-2026.05.14 | 单价×单人使用量(或÷小组人数) | 4人一组项目÷4 | 磨损费=单价÷预估使用次数'
ws3['A2'].font = Font(name='微软雅黑', size=9, color='666666')
ws3['A2'].alignment = center_align

src = openpyxl.load_workbook(
    r'C:\Users\scrccpa\Desktop\成本测算=护理学院\各职业耗材成本明细\智护学院-长期照护师-耗材成本表-2026.05.14.xlsx',
    data_only=True
)

current_row = 4
for level_idx, (level_name, sheet_name) in enumerate([
    ('五级/初级工', '五级（初级）'),
    ('四级/中级工', '四级（中级）'),
    ('三级/高级工', '三级（高级）')
]):
    src_ws = src[sheet_name]

    ws3.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
    ws3.cell(row=current_row, column=1).value = f'{"一二三"[level_idx]}、{level_name} — 耗材成本明细'
    ws3.cell(row=current_row, column=1).font = h2_font
    for ci in range(1, 11):
        ws3.cell(row=current_row, column=ci).fill = sub_header_fill
    current_row += 1

    hdrs = ['序号', '耗材名称', '规格', '参考品牌', '单位', '单价（元）',
            '单人使用量', '单人成本（元）', '计算过程', '备注']
    for ci, h in enumerate(hdrs, 1):
        ws3.cell(row=current_row, column=ci).value = h
    style_header_row(ws3, current_row, 10)
    current_row += 1

    in_section = False
    seq = 0
    total_cost = 0.0
    for src_row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row, values_only=True):
        r0 = str(src_row[0]) if src_row[0] else ''
        if '耗材成本' in r0 and '等级认定' in r0:
            in_section = True
            continue
        if '设施设备' in r0:
            in_section = False
            continue
        if '合计' in r0 and in_section:
            for ci in range(1, 11):
                ws3.cell(row=current_row, column=ci).fill = total_fill
                ws3.cell(row=current_row, column=ci).font = bold_font
                ws3.cell(row=current_row, column=ci).border = thin_border
                ws3.cell(row=current_row, column=ci).alignment = center_align
            ws3.cell(row=current_row, column=1).value = '合计'
            ws3.cell(row=current_row, column=8).value = round(total_cost, 2)
            ws3.cell(row=current_row, column=8).number_format = money_fmt
            current_row += 1
            continue
        if in_section and r0.strip().isdigit():
            seq += 1
            unit_price = float(src_row[5]) if src_row[5] else 0
            person_cost = float(src_row[7]) if src_row[7] else 0
            total_cost += person_cost
            qty_str = str(src_row[6]).strip() if src_row[6] else ''
            remark = str(src_row[8]).strip() if src_row[8] else ''

            calc = ''
            if '/' in qty_str:
                parts = qty_str.strip().split('/')
                if len(parts) == 2 and parts[0].strip() and parts[1].strip():
                    try:
                        denom = float(parts[1].strip())
                        calc = f'{unit_price:.2f} ÷ {denom:.0f} = {unit_price/denom:.2f}'
                    except:
                        calc = f'{unit_price:.2f} ÷ {parts[1].strip()}'
            elif qty_str:
                try:
                    qty = float(qty_str)
                    calc = f'{unit_price:.2f} × {qty} = {unit_price*qty:.2f}'
                except:
                    calc = qty_str

            row_data = [seq, src_row[1] or '', src_row[2] or '', src_row[3] or '',
                       src_row[4] or '', unit_price, qty_str, person_cost, calc, remark]
            for ci, v in enumerate(row_data, 1):
                cell = sc(ws3, current_row, ci, 'left' if ci in [2, 3, 4, 9, 10] else 'center')
                cell.value = v
                if ci in [6, 8]:
                    cell.number_format = money_fmt
            current_row += 1

    current_row += 2

set_widths(ws3, [6, 24, 28, 18, 8, 12, 14, 14, 26, 24])

# =====================================================
# Sheet 4: 设施设备成本明细
# =====================================================
ws4 = wb.create_sheet('设施设备成本明细')

ws4.merge_cells('A1:J1')
ws4['A1'].value = '健康照护师（长期照护师）等级认定 — 设施设备成本明细测算'
ws4['A1'].font = title_font
ws4['A1'].alignment = center_align

ws4.merge_cells('A2:J2')
ws4['A2'].value = '说明：单人设施设备成本=(设备采购价÷预计服务总人次)×单次使用数量 | 当前采用磨损费简化模型 | 4人一组均摊'
ws4['A2'].font = Font(name='微软雅黑', size=9, color='666666')
ws4['A2'].alignment = center_align

current_row = 4
for level_idx, (level_name, sheet_name) in enumerate([
    ('五级/初级工', '五级（初级）'),
    ('四级/中级工', '四级（中级）'),
    ('三级/高级工', '三级（高级）')
]):
    src_ws = src[sheet_name]

    ws4.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
    ws4.cell(row=current_row, column=1).value = f'{"一二三"[level_idx]}、{level_name} — 设施设备成本明细'
    ws4.cell(row=current_row, column=1).font = h2_font
    for ci in range(1, 11):
        ws4.cell(row=current_row, column=ci).fill = sub_header_fill
    current_row += 1

    hdrs = ['序号', '设备名称', '规格', '参考品牌', '单位', '单价（元）',
            '单人分摊量', '单人成本（元）', '分摊方式', '备注']
    for ci, h in enumerate(hdrs, 1):
        ws4.cell(row=current_row, column=ci).value = h
    style_header_row(ws4, current_row, 10)
    current_row += 1

    in_section = False
    seq = 0
    total_cost = 0.0
    for src_row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row, values_only=True):
        r0 = str(src_row[0]) if src_row[0] else ''
        if '设施设备' in r0:
            in_section = True
            continue
        if '合计' in r0 and in_section:
            for ci in range(1, 11):
                ws4.cell(row=current_row, column=ci).fill = total_fill
                ws4.cell(row=current_row, column=ci).font = bold_font
                ws4.cell(row=current_row, column=ci).border = thin_border
                ws4.cell(row=current_row, column=ci).alignment = center_align
            ws4.cell(row=current_row, column=1).value = '合计'
            ws4.cell(row=current_row, column=8).value = round(total_cost, 2)
            ws4.cell(row=current_row, column=8).number_format = money_fmt
            current_row += 1
            continue
        if in_section and r0.strip().isdigit():
            seq += 1
            unit_price = float(src_row[5]) if src_row[5] else 0
            person_cost = float(src_row[7]) if src_row[7] else 0
            total_cost += person_cost
            qty_str = str(src_row[6]).strip() if src_row[6] else ''
            remark = str(src_row[8]).strip() if src_row[8] else ''

            alloc = ''
            if '磨损费' in remark:
                alloc = '磨损费分摊'
            elif '4人一组' in remark:
                alloc = '4人一组均摊'
            else:
                alloc = remark

            row_data = [seq, src_row[1] or '', src_row[2] or '', src_row[3] or '',
                       src_row[4] or '', unit_price, qty_str, person_cost, alloc, remark]
            for ci, v in enumerate(row_data, 1):
                cell = sc(ws4, current_row, ci, 'left' if ci in [2, 3, 4, 9, 10] else 'center')
                cell.value = v
                if ci in [6, 8]:
                    cell.number_format = money_fmt
            current_row += 1

    current_row += 2

set_widths(ws4, [6, 24, 28, 18, 8, 12, 14, 14, 20, 28])

# =====================================================
# Sheet 5: 测算依据与过程
# =====================================================
ws5 = wb.create_sheet('测算依据与过程')

ws5.merge_cells('A1:C1')
ws5['A1'].value = '测算依据与过程说明'
ws5['A1'].font = title_font
ws5['A1'].alignment = center_align

notes = [
    ('一、测算目的', '本表对健康照护师（长期照护师）职业技能等级认定（五级/四级/三级）的单人考核成本进行系统测算，为收费定价提供成本依据。'),
    ('', ''),
    ('二、数据来源', ''),
    ('1', '智护学院-长期照护师-耗材成本表-2026.05.14.xlsx — 含五级/四级/三级逐项耗材与设施设备明细'),
    ('2', '各职业系部成本测算汇总.xlsx — 系部提交的各职业各等级成本汇总数据'),
    ('3', '职业技能等级认定收费方案（5稿）.docx — 现行收费方案及职业分类'),
    ('4', '川发改价格〔2017〕472号 — 人社部门行政事业性收费标准（理论+操作考核费）'),
    ('5', '四川护理职业学院院长办公会纪要（第23期）2022年12月12日 — 人工费用发放标准'),
    ('', ''),
    ('三、测算方法与过程', ''),
    ('', ''),
    ('（一）政策规定收费', '理论考试费和操作技能考核费按川发改价格〔2017〕472号规定的B类标准执行，为政府定价上限，不可调整。'),
    ('', ''),
    ('（二）考务平台费', '按学院实际支付的考务平台服务费计算，固定20元/人。'),
    ('', ''),
    ('（三）人工费用', '按院长办公会纪要（第23期）2022.12.12议定标准：考评员100元/小时，其他工作人员60元/小时。'),
    ('', '测算步骤：'),
    ('', '  1) 确定每日考试班次30人、实操6小时含准备'),
    ('', '  2) 理论部分：监考4人×2.5h+技术保障1人×3.5h，按60元/h计'),
    ('', '  3) 实操考评员：3考站×3人×6h，按100元/h计（健康照护师8考题/全国联考，比普通职业多设考站）'),
    ('', '  4) SP模特（标准化病人）：1名老师×60元/h + 2名学生×22元/h，全程6h'),
    ('', '  5) 实操监考：3考站×2人+候考室2人，共8人×6h×60元/h'),
    ('', '  6) 管理协调：主考、副主考、督导、考务×2、安保、综合管理共7人×7.5h×60元/h'),
    ('', '  7) 考场布置：5人×1h×60元/h'),
    ('', '  8) 各项汇总 ÷ 30人 = 单人人工成本'),
    ('', ''),
    ('（四）设施设备使用费', ''),
    ('', '  1) 逐项列明考核所需设施设备名称、规格、采购单价'),
    ('', '  2) 共用设备按4人一组均摊或磨损费模型分摊至单人'),
    ('', '  3) 设备折旧当前采用磨损费简化模型；如转化为会计准则下的直线折旧法，需补充预计使用年限和残值率'),
    ('', ''),
    ('（五）耗材成本', ''),
    ('', '  1) 逐项列明8个考题所需全部耗材（名称、规格、品牌、单价、用量）'),
    ('', '  2) 一次性消耗品：单价×单人使用量'),
    ('', '  3) 可分摊耗材：单价÷分摊人数（4人一组）'),
    ('', '  4) 磨损品（病号服、浴巾等）：按磨损费计入'),
    ('', ''),
    ('（六）场地水电费', '场地费3,200元/天 ÷ 30人/班次 = 106.67元/人。含考场、候考室、休息室等区域使用及水电。'),
    ('', ''),
    ('四、测算假设与限制条件', ''),
    ('', '1. 每班次30人、实操6小时（含考前准备）、3个考站并行 —— 实际安排可能因考生数量调整'),
    ('', '2. SP模特3人（1老师+2学生），全程参与实操考核 —— 如考题编制变化，SP需求可能调整'),
    ('', '3. 三级考试人工成本与五级/四级相同 —— 主要差异在耗材和设施设备'),
    ('', '4. 设施设备成本采用"磨损费"简化模型，未采用直线折旧法'),
    ('', '5. 视频监控费、证书制作费、文印广告费尚未取得基础数据，暂未计入'),
    ('', '6. 本测算未包含间接费用（管理费用分摊、财务费用等）'),
    ('', ''),
    ('五、关键结论', ''),
    ('', ''),
    (f'已测算7项成本合计：五级/初级{subtotal_measured_5:.2f}元/人 | 四级/中级{subtotal_measured_4:.2f}元/人 | 三级/高级{subtotal_measured_3:.2f}元/人', ''),
    ('', ''),
    (f'现行收费方案（5稿）：320 / 405 / 500 元/人', ''),
    (f'差额（收费-成本）：{320-subtotal_measured_5:.2f} / {405-subtotal_measured_4:.2f} / {500-subtotal_measured_3:.2f} 元/人', ''),
    ('', ''),
    ('⚠ 已测算7项成本已超过现行收费标准。若计入待补充的3项成本（视频监控/证书/文印），全口径成本将进一步上升。', ''),
    ('建议：(1)尽快收集待补充成本数据完成全口径测算；(2)基于全口径成本重新核定收费方案。', ''),
]

for i, (label, content) in enumerate(notes):
    r = i + 4
    if label:
        cell = ws5.cell(row=r, column=1, value=label)
        if any(label.startswith(x) for x in ['一、', '二、', '三、', '四、', '五、']):
            cell.font = h2_font
        elif label.startswith('（'):
            cell.font = h3_font
        else:
            cell.font = normal_font
    if content:
        c2 = ws5.cell(row=r, column=2, value=content)
        c2.font = normal_font
        c2.alignment = Alignment(wrap_text=True, vertical='center')

set_widths(ws5, [32, 90, 15])

# =====================================================
# Save
# =====================================================
output_path = r'D:\openclaw-workspace\output\健康照护师-成本构成测算-v2-2026.05.21.xlsx'
wb.save(output_path)
print(f'\nOK: {output_path}')
print(f'\n=== 成本汇总 ===')
print(f'五级/初级: {subtotal_measured_5:.2f} 元/人')
print(f'四级/中级: {subtotal_measured_4:.2f} 元/人')
print(f'三级/高级: {subtotal_measured_3:.2f} 元/人')
print(f'\n=== 差额（收费-成本） ===')
print(f'五级/初级: {320-subtotal_measured_5:.2f} 元/人')
print(f'四级/中级: {405-subtotal_measured_4:.2f} 元/人')
print(f'三级/高级: {500-subtotal_measured_3:.2f} 元/人')
print(f'\nSP模特单人成本: {sp_per_person} 元/人')
print(f'场地单人成本: {venue_per_person} 元/人')
print(f'人工费用单人: {labor_per_person} 元/人')
