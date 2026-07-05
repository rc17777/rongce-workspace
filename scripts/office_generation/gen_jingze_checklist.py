# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.Workbook()

header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
title_font = Font(name='微软雅黑', size=14, bold=True)
sub_font = Font(name='微软雅黑', size=10, bold=True)
body_font = Font(name='微软雅黑', size=10)
bold_red = Font(name='微软雅黑', size=10, bold=True, color='C00000')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
section_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap = Alignment(wrap_text=True, vertical='center')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

# ============ Sheet 1 ============
ws = wb.active
ws.title = '取数函清单'
ws.sheet_properties.tabColor = '2F5496'

ws.merge_cells('A1:G1')
c = ws['A1']; c.value = '审计资料需求清单 - 经济责任审计'; c.font = title_font
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 30

ws.merge_cells('A2:G2')
c = ws['A2']; c.value = '被审计单位：          | 审计期间：          | 填发日期：'
c.font = body_font; ws.row_dimensions[2].height = 22

headers = ['序号', '大类', '资料名称', '数据格式', '详细要求', '优先级', '备注']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=h)
    cell.font = header_font; cell.fill = header_fill; cell.alignment = center; cell.border = thin_border
ws.row_dimensions[3].height = 25

data = [
    (True, '一、财务资料', '', '', '', '', ''),
    ('', '序时账', '.xlsx / .csv', '末级科目，含摘要和辅助核算（供应商/客户/项目/部门/个人五维度）。任期全周期+前后各推1年', 'RED', '缺摘要则改要明细账'),
    ('', '科目余额表', '.xlsx', '逐月！末级科目，含辅助核算维度。任期全周期+前后各1年', 'RED', '只给年末数则退回重取'),
    ('', '总账', '.xlsx', '一级科目汇总，任期内每年', 'YELLOW', '可用科目余额表替代'),
    ('', '明细账', '.xlsx', '末级科目逐笔。序时账无摘要时必须提供', 'YELLOW', '序时账有摘要可不取'),
    ('', '资产负债表、收入费用表、净资产变动表', '.xlsx / .pdf', '任期内每年，全口径（含合并报表如适用）', 'RED', ''),
    ('', '预算执行情况表', '.xlsx', '任期内每年，含预算批复数与实际执行数', 'RED', ''),
    ('', '三公经费预算批复及执行明细', '.xlsx', '逐月逐笔，因公出国/公务用车/公务接待分开，含预算数和执行数', 'RED', '另加会议费+培训费'),
    ('', '固定资产台账', '.xlsx', '任期起始日和终止日两个时点的全量台账（含入账日期/原值/折旧/净值/使用部门/存放地点）', 'RED', '两个时点才能算保值增值'),
    ('', '往来款明细表', '.xlsx', '应收账款/应付账款/其他应收款/其他应付款/预付账款，任期末全量+账龄分析，按对方单位展开含挂账日期+事由', 'RED', '含3年以上长账龄标记'),
    ('', '银行对账单', '.xlsx / .pdf', '任期内全部银行账户逐月对账单', 'YELLOW', '发现现金异常时补取'),
    ('', '现金日记账 / 银行日记账', '.xlsx', '任期内全量', 'YELLOW', '辅助参考'),
    (True, '二、文本资料', '', '', '', '', ''),
    ('', '内控制度汇编', '.pdf / .docx', '含财务管理制度、采购管理办法、资产管理制度、三重一大决策制度。文字版优先，扫描件需OCR', 'RED', ''),
    ('', '合同台账', '.xlsx', '合同编号|合同名称|对方单位|签订日期|合同金额|已付金额|标的|履约状态|验收日期', 'RED', '合同编号必须唯一'),
    ('', '合同原件电子版', '.pdf', '任期内全量合同扫描件', 'RED', '与合同台账一一对应'),
    ('', '党组/党委会议纪要', '.pdf / .docx', '任期内全量', 'RED', '判断该上会是否上会的唯一依据'),
    ('', '局长（主任）办公会议纪要', '.pdf / .docx', '任期内全量', 'RED', ''),
    ('', '三重一大上会清单', '.xlsx / .pdf', '含上会日期、议题、表决结果', 'RED', ''),
    ('', '上级审计/巡视/督查问题及整改情况', '.pdf / .docx', '任期内上级监督检查发现的所有问题+整改报告', 'RED', '查整改闭环'),
    ('', '任职文件', '.pdf', '含任免日期，精确到日', 'RED', '确定审计起止时间'),
    ('', '单位三定方案', '.pdf', '现行有效版本', 'RED', '确定职责范围'),
    ('', '下属单位清单', '.xlsx / .pdf', '含自收自支事业单位、企业化管理单位、协会学会', 'RED', ''),
    (True, '三、招投标资料（如有工程建设/大额设备/服务外包）', '', '', '', '', ''),
    ('', '任期内采购项目清单', '.xlsx', '含项目名称、采购方式、预算金额、中标金额、中标单位', 'RED', '有大额采购时必取'),
    ('', '招标文件', '.pdf', '每个限额以上项目的招标文件全套', 'RED', '提取评标办法/限价/资格条件'),
    ('', '投标文件（各投标人）', '.docx优先', '每个项目全部投标人的投标文件。必须.docx格式！PDF扫描件只能做L1报价，L3-L6文本雷同不可用', 'RED', '围标检测核心数据源'),
    ('', '开标一览表', '.xlsx', '含所有投标人名称和报价', 'RED', 'L1报价规律分析数据源'),
    ('', '评标报告', '.pdf', '每个项目', 'RED', ''),
    ('', '中标通知书', '.pdf', '每个中标项目', 'YELLOW', '中标金额与合同金额对照'),
    ('', '投标系统登录日志', '代理机构提供', '每个项目各投标人的IP地址、登录时间、MAC地址', 'YELLOW', '详见Sheet2攻略'),
    (True, '四、专项资金（任期内每笔专项资金逐笔取）', '', '', '', '', ''),
    ('', '专项资金收支明细表', '.xlsx', '专项资金名称|资金来源|下达金额|下达时间|支出金额|结余', 'RED', '先汇总再逐笔深入'),
    ('', '资金下达文件', '.pdf', '上级拨款的正式文件，含文号', 'RED', ''),
    ('', '专项资金管理办法', '.pdf', '该专项的使用规则和限制', 'RED', ''),
    ('', '项目实施方案', '.pdf', '钱花在哪、怎么花', 'RED', ''),
    ('', '验收报告', '.pdf', '项目完成情况', 'YELLOW', '项目完结时取'),
    ('', '支出发票和凭证（大额抽查）', '.pdf / 图片', '专项资金大额支出的原始凭证', 'YELLOW', '关注异常供应商/异常时间'),
    (True, '五、国企附加（仅国有企业适用）', '', '', '', '', ''),
    ('', '董事会/监事会决议+会议记录', '.pdf', '任期内全量', 'RED', ''),
    ('', '经营业绩考核文件+指标完成数据', '.pdf / .xlsx', '任期各年度', 'RED', ''),
    ('', '关联方清单+关联交易台账', '.xlsx', '任期内全量，含交易金额+定价依据', 'RED', ''),
    ('', '高管薪酬方案+实际发放明细', '.xlsx', '任期各年度', 'RED', ''),
    ('', '对外投资台账', '.xlsx', '投了哪些公司/持股比例/分红情况', 'RED', ''),
    ('', '企业产权登记表', '.pdf / .xlsx', '任期末全量', 'YELLOW', ''),
]

row = 4; seq = 0
for item in data:
    if item[0] is True:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cell = ws.cell(row=row, column=1, value=item[1])
        cell.font = Font(name='微软雅黑', size=11, bold=True, color='2F5496')
        cell.fill = section_fill; cell.alignment = Alignment(vertical='center'); cell.border = thin_border
        for c in range(2, 8):
            ws.cell(row=row, column=c).fill = section_fill
            ws.cell(row=row, column=c).border = thin_border
        ws.row_dimensions[row].height = 24; row += 1; seq = 0
    else:
        seq += 1
        ws.cell(row=row, column=1, value=seq).font = body_font
        ws.cell(row=row, column=1).alignment = center
        for ci, val in enumerate(item[1:], 2):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font = bold_red if (ci == 5 and val == 'RED') else body_font
            cell.alignment = wrap
            if ci == 5:
                if val == 'RED': cell.value = '最优先'
                elif val == 'YELLOW': cell.value = '辅助/条件'
                else: cell.value = val
        for c in range(1, 8):
            ws.cell(row=row, column=c).border = thin_border
        ws.row_dimensions[row].height = 36; row += 1

widths = [6, 22, 28, 14, 48, 13, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.auto_filter.ref = f'A3:G{row-1}'
ws.freeze_panes = 'A4'

# ============ Sheet 2 ============
ws2 = wb.create_sheet('代理机构不给IP日志怎么破')
ws2.sheet_properties.tabColor = 'FF6B6B'
ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 18
ws2.column_dimensions['D'].width = 55
ws2.column_dimensions['E'].width = 38

ws2.merge_cells('A1:E1')
c = ws2['A1']; c.value = '审计实战：代理机构不给投标IP/MAC日志——7层破解法'
c.font = title_font; c.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 30

ws2.merge_cells('A2:E2')
c = ws2['A2']; c.value = '典型场景：打电话要投标系统登录日志→代理机构回复：没有/系统坏了/数据已删除/没权限给。怎么办？'
c.font = Font(name='微软雅黑', size=10, color='C00000'); c.alignment = wrap
ws2.row_dimensions[2].height = 24

for col, h in enumerate(['序号', '层级', '应对方法', '具体操作', '效力评估'], 1):
    cell = ws2.cell(row=3, column=col, value=h)
    cell.font = header_font; cell.fill = header_fill; cell.alignment = center; cell.border = thin_border
ws2.row_dimensions[3].height = 22

tactics = [
    ('1', '硬要\n（第一步）', '书面发函\n+法规施压',
     '向代理机构发正式函：\n"根据《招标投标法》及实施条例，你单位作为招标代理机构，有义务保存和提供招标过程记录。请你单位在X个工作日内提供：\n1.各投标人投标文件上传IP地址\n2.投标系统登录日志（含时间/IP/MAC）\n3.开标签到表"',
     '直接有效。约60%代理机构收到正式函件后会提供。'),
    ('2', '投诉施压\n（第二步）', '向监管部门投诉\n代理机构不配合',
     '向财政局/公共资源交易监管局投诉：代理机构拒不提供招标过程数据，妨碍审计取证。监管部门有权责令其提供并处罚。',
     '强力。代理机构怕监管部门的处罚比怕审计多。'),
    ('3', '曲线取证', '从投标人端\n反推IP',
     '投标文件是电子上传的→PDF的/Producer字段有时含创建IP（部分代理系统有）。投标文件是纸质送达的→查送达签收记录（同一人同一天送多家标书=旁证）。',
     '间接证据，可支撑七成结论。'),
    ('4', '旁证替代\n（核心策略）', '用其他L层级发现\n替代L2',
     'L2（IP/MAC）只是11层围标检测中的1层。即使没有IP数据，以下发现依然是铁证：\nL3 文本雷同>=89%（TF-IDF）\nL4 图片MD5完全一致（共享同一扫描件）\nL5 WPS版本GUID完全一致（同机编制）\nL7 打印机型号相同\n→三层以上同时命中=基本确认围标，不依赖IP数据。',
     '核心策略！L3+L4+L5三杀即可定案，法院采信。'),
    ('5', '资金溯源', '查保证金\n汇款账户',
     '向银行调取投标保证金汇款凭证：同一账户汇出多家保证金=围标铁证。同一时间段同一柜台汇款=强信号。注意：代理机构管不了银行，向银行调取不需要经过代理机构。',
     '铁证。银行流水比IP日志更难伪造，法院采信度最高之一。'),
    ('6', '工商穿透', '查工商关联\n替代IP关联',
     '用天眼查/企查查查所有投标人：同一实控人控制两家以上投标公司=围标。法人/监事/高管/财务负责人交叉任职。注册地址/联系电话相同。→法院采信度最高的围标证据类型之一。',
     '铁证级。L8工商关联发现即定案。'),
    ('7', '倒逼承认', '把现有证据\n甩代理机构面前',
     '"你单位拒不提供IP数据，但我们已通过其他渠道发现：A和C投标文件雷同度89%；A/B/C三家共享同一张认证证书扫描件（MD5完全一致）。你单位作为招标代理是否知情？现在你单位是配合调查还是等待移送监管部门？"→代理机构意识到事情比他们想象的大→主动配合。',
     '心理战。当代理机构发现你在别处已有铁证时，隐瞒成本急剧上升，通常会主动配合。'),
]

row = 4
for t in tactics:
    for ci, val in enumerate(t, 1):
        cell = ws2.cell(row=row, column=ci, value=val)
        cell.font = sub_font if ci == 1 else body_font
        cell.alignment = center if ci == 1 else wrap
        cell.border = thin_border
    if '核心策略' in t[4] or '铁证' in t[4]:
        for c in range(1, 6):
            ws2.cell(row=row, column=c).fill = green_fill
    elif '强力' in t[4]:
        for c in range(1, 6):
            ws2.cell(row=row, column=c).fill = yellow_fill
    ws2.row_dimensions[row].height = 120
    row += 1

row += 1
ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
c = ws2.cell(row=row, column=1)
c.value = '核心原则：L2（IP/MAC）丢了不可怕。L3+L4+L5同时三杀的直接证据强度>单一L2的IP证据。11层体系中任何三层以上同时命中即可形成闭合证据链，不依赖代理机构配合。'
c.font = Font(name='微软雅黑', size=11, bold=True, color='2F5496'); c.alignment = wrap
ws2.row_dimensions[row].height = 40

ws2.freeze_panes = 'A4'

outpath = r'C:\Users\scrccpa\Desktop\审计资料需求清单-经济责任审计.xlsx'
wb.save(outpath)
print(f'Done: {outpath}')
