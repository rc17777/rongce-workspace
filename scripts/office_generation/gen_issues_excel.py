"""生成审计发现问题清单Excel — 放到桌面"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============ Styles ============
header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
sub_header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
orange_fill = PatternFill(start_color='FFDAB9', end_color='FFDAB9', fill_type='solid')
title_font = Font(name='微软雅黑', bold=True, size=13, color='2F5496')
red_font = Font(name='微软雅黑', size=11, color='CC0000', bold=True)
bold_font = Font(name='微软雅黑', bold=True, size=11)
normal_font = Font(name='微软雅黑', size=11)
s_font = Font(name='微软雅黑', size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
c_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
w_align = Alignment(vertical='center', wrap_text=True)

def cell(ws, r, col, v, font=normal_font, fill=None, align=c_align):
    c = ws.cell(row=r, column=col, value=v)
    c.font = font; c.alignment = align; c.border = thin_border
    if fill: c.fill = fill
    return c

def hdr(ws, r, headers):
    for i, h in enumerate(headers):
        cell(ws, r, i+1, h, header_font, header_fill, c_align)

# ============ Sheet 1: 问题清单 ============
ws = wb.active
ws.title = '审计发现问题清单'

# Title
cell(ws, 1, 1, '2024年多功能急救实训室建设项目 审计发现问题清单', title_font, align=w_align)
ws.merge_cells('A1:H1')
cell(ws, 2, 1, '项目编号: N5100012025000628 | 采购人: 四川护理职业学院 | 代理机构: 联投项目管理(集团)有限公司 | 中标人: 成都易可天地科技有限公司 | 中标金额: 1,566,000元', s_font, align=w_align)
ws.merge_cells('A2:H2')

r = 4
hdr(ws, r, ['序号', '问题类别', '风险等级', '问题描述', '涉及投标人', '证据/数据', '审计建议', '对应检测层'])

issues = [
    # L5 铁证
    [1, '文档同源性 — 串标铁证', '🔴 重大',
     '易可天地与江西正好的4份核心投标文档（其他资料、技术应答表、投标函、商务应答表）的.docx原始文件Author字段完全一致：ZXM / WPS_1654344806 / 汪。WPS_1654344806为WPS安装时间戳，两台不同电脑的安装时间戳不可能相同，证明两份投标文件由同一人/同一台电脑制作。',
     '成都易可天地科技有限公司\n江西正好医疗器械有限公司',
     'PDF元数据提取(PyMuPDF)：\n• 其他资料.docx.pdf Author=ZXM (两家一致)\n• 技术应答表.docx.pdf Author=WPS_1654344806 (两家一致)\n• 投标函.docx.pdf Author=汪 (两家一致)\n• 商务应答表.docx.pdf Author=WPS_1654344806 (两家一致)',
     '构成《政府采购法实施条例》第74条第(七)项"供应商之间协商报价、技术方案等投标文件实质性内容"的串通投标行为。建议移送财政部门立案调查。', 'L5-元数据'],
    
    # 报价集中
    [2, '报价高度集中', '🔴 重大',
     '三家投标报价极差仅19,000元(1.21%)，全部紧贴最高限价98.0%-99.2%区间。对比同类项目（校服采购5家极差7.8%），离散度异常偏低。所有38项分项报价均紧贴各自限价97%-99.9%。',
     '三家全部',
     '好医助: 1,585,000(99.2%)\n江西正好: 1,574,022(98.5%)\n易可天地: 1,566,000(98.0%)\n极差: 19,000元(1.21%)\n标准差: 9,538元',
     '向代理机构调取电子交易系统投标IP记录。同一IP或同一IP段即构成围标铁证。', 'L1-报价规律'],
    
    # 评审得分异常
    [3, '评审得分一致性异常', '🔴 重大',
     '5位评委中4位（郑雁、杨京儒、王文霞、肖晓辉）对三家投标人的评分完全一致（精确到小数点后2位）。含10分主观分的评审中，4/5评委的评分完全一致在统计学上概率极低。唯一不同的评委罗征洪，恰好给每个投标人各少打2分。',
     '评审委员会',
     '评分汇总表(P106)：\n易可天地: 评委4/5打93.50分\n江西正好: 评委4/5打69.18分\n好医助: 评委4/5打65.97分\n唯一差异评委罗征洪: 92.50/67.18/63.97\n(恰好各少2分)',
     '核查评审委员会成员是否存在利益关联。调取原始评分底稿，核实是否存在"预填"或"统一口径"情况。', '追加-评审'],
    
    # 技术得分断崖
    [4, '非价格得分断崖式差距', '🔴 重大',
     '第1名与第2名非价格得分(满分70分)差距达24.37分，第1名是第2名的1.6倍。而第2名与第3名仅差3分。三家投标人的技术实力不应存在如此巨大的阶梯差，提示招标条件可能为中标人"量身定制"或第2/3名为"陪标"。',
     '三家全部',
     '非价格得分(满分70)：\n易可天地: 63.30 (90.4%)\n江西正好: 38.93 (55.6%)\n好医助: 35.93 (51.3%)\n第1→2名差距: 24.37分',
     '复核招标技术参数是否具有指向性。核实中标人所投品牌/型号在市场中的独占性。比对第2/3名技术响应的扣分项是否合理。', '追加-技术'],
    
    # 中小企业
    [5, '中小企业声明一致性', '🟡 中等',
     '三家投标人均在中小企业声明函中标注"不响应此评审点内容"，即全部放弃小微企业10%价格扣除优惠。若确有符合小微标准的企业主动放弃10%价格优势，不符合商业逻辑。好医助与江西正好声明日期相同(2025-05-23)。',
     '三家全部',
     '中小企业声明函：\n好医助: "不响应此评审点内容" 日期5/23\n易可天地: "不响应此评审点内容" 日期5/22\n江西正好: "不响应此评审点内容" 日期5/23',
     '通过国家企业信用信息公示系统或工信部中小企业认定平台核实三家实际企业规模。如存在符合小微标准却主动放弃的情况，需查明原因。', '追加-中小企业'],
    
    # 文件结构
    [6, '投标文件结构完全一致', '🟡 中等',
     '三家投标人均提交8个PDF文件，文件名完全一致：中小企业声明函.pdf、其他资料.docx.pdf、技术要求应答表.docx.pdf、投标（响应）函.docx.pdf、报价表.pdf、服务内容要求商务要求应答表.docx.pdf、残疾人福利性单位声明函.pdf、监狱企业的证明文件.pdf。均使用WPS文字编辑.docx后转为PDF。',
     '三家全部',
     '文件列表对比：三家8个PDF文件名100%一致。软件环境: 全部WPS文字+Qt 5.15.2+Chromium/Skia。',
     '文件命名一致性本身可能源于招标平台模板要求，但结合L5作者同源性，进一步强化串标嫌疑。', 'L6-文档结构'],
    
    # 字体包含
    [7, '好医助字体完全被易可天地包含', '🟡 中等',
     '好医助使用的6种字体(FangSong/MicrosoftYaHei/SimSun/SimHei等)100%被易可天地的10种字体集合包含。虽然字体本身为基础中文字体，但完全包含关系值得关注。',
     '好医助 + 易可天地',
     '好医助: 6种字体\n易可天地: 10种字体\n好医助6种⊂易可天地10种',
     '配合L5证据使用。同城两家投标人(成都)字体集合的包含关系，结合文档作者的独立性(L5好医助≠易可天地)，暂时不构成独立证据。', 'L6-字体'],
    
    # 图片量差异
    [8, '易可天地图片量异常庞大', '🟢 关注',
     '易可天地嵌入208张图片，远超好医助(39张)和江西正好(32张)。可能是提供了更详尽的证明材料和技术截图，但也可能是刻意增加文档复杂度。',
     '易可天地',
     'PDF嵌入图片数：\n好医助: 39张\n易可天地: 208张\n江西正好: 32张\n0张跨公司重复(SHA256)',
     '关注中标后实际交付产品是否与投标时提供的技术证明材料一致。', 'L4-图片'],
    
    # 喉镜地域集中
    [9, '可视化喉镜供应商地域集中', '🟢 关注',
     '三家可视化喉镜均采购自江苏省泰州/泰兴地区（中国医疗器械产业聚集区），但为三家不同厂家（斯美特/辉春/永乐）。属于产业聚集现象，不构成独立串标证据。',
     '三家全部',
     '好医助: 泰兴市斯美特\n易可天地: 泰州市辉春\n江西正好: 江苏永乐(泰兴)',
     '无需单独行动。如后续获取L2(IP)证据，可作为辅助佐证。', 'L1-品牌'],
    
    # L2
    [10, '投标IP/MAC记录缺失', '🔴 重大缺失',
     '最关键的第2层检测（投标IP/MAC）数据缺失。电子交易系统后台可查询三家投标文件上传IP地址和时间。同一IP即构成围标不可辩驳的铁证。',
     '代理机构联投项目管理公司',
     '数据源: 四川省政府采购一体化平台电子交易系统日志。需正式向代理机构发函调取。',
     '【最高优先级】立即向代理机构（联投项目管理公司）调取投标IP/MAC记录。同时核实投标文件上传时间戳是否异常接近。', 'L2-IP'],
    
    # L8
    [11, '工商关联关系待查', '🔴 重大缺失',
     '三家公司的股东/高管/对外投资是否存在交叉持股或关联关系，尚未查询。天眼查/企查查数据可直接验证。',
     '三家全部',
     '待查平台: 天眼查(www.tianyancha.com) 或 企查查(www.qcc.com)',
     '立即查询三家公司的工商信息: (1)法定代表人 (2)股东及持股比例 (3)董监高人员 (4)对外投资 (5)历史股东变更。如存在人员交叉或投资关联，构成围标辅助证据。', 'L8-工商'],
    
    # L7
    [12, 'PDF生成设备信息可提取', '🟢 可补查',
     'PDF Producer/Creator字段已提取。报价表(Chromium/Skia)为平台统一生成，模板PDF(Qt 5.15.2)为统一模板。但可进一步分析PDF的XMP元数据和字体嵌入特征。',
     '三家全部',
     'Producer: Qt 5.15.2(模板PDF) / Chromium+Skia(报价表)\nCreator: WPS文字(docx类)',
     '获取原始.docx文件后提取OLE2流中的作者信息，与PDF元数据交叉验证。', 'L7-设备'],
]

for i, issue in enumerate(issues):
    row = 5 + i
    for j, val in enumerate(issue):
        if j == 2:  # 风险等级
            fill = red_fill if '重大' in str(val) else (yellow_fill if '中等' in str(val) else green_fill)
            cell(ws, row, j+1, val, bold_font, fill)
        elif j == 0:  # 序号
            cell(ws, row, j+1, val, normal_font, align=c_align)
        elif j == 5:  # 证据
            cell(ws, row, j+1, val, s_font, align=w_align)
        elif j in [3, 6]:  # 描述/建议
            cell(ws, row, j+1, val, normal_font, align=w_align)
        else:
            cell(ws, row, j+1, val, normal_font, align=c_align if j != 4 else w_align)

# Row heights
for i in range(5, 5+len(issues)):
    ws.row_dimensions[i].height = 120

ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 50
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 35
ws.column_dimensions['G'].width = 40
ws.column_dimensions['H'].width = 14

# ============ Sheet 2: 评审得分明细 ============
ws2 = wb.create_sheet('评审得分明细')

cell(ws2, 1, 1, '评审得分明细表', title_font, align=w_align)
ws2.merge_cells('A1:J1')
cell(ws2, 2, 1, '数据来源: 归档资料 P104(价格评审表) + P106(评分汇总表) | 评审日期: 2025-05-26 | 专家: 杨京儒(采购人代表)、罗征洪、郑雁、王文霞、肖晓辉', s_font, align=w_align)
ws2.merge_cells('A2:J2')

# Part A: Score summary
r = 4
cell(ws2, r, 1, 'A. 各评委打分及总分', bold_font, sub_header_fill, w_align)
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)

r = 5
hdr(ws2, r, ['排名', '供应商', '郑雁', '杨京儒', '罗征洪', '王文霞', '肖晓辉', '总分合计', '平均得分', '价格得分'])

scores = [
    [1, '成都易可天地科技有限公司', 93.50, 93.50, 92.50, 93.50, 93.50, 466.50, 93.30, 30.00],
    [2, '江西正好医疗器械有限公司', 69.18, 69.18, 67.18, 69.18, 69.18, 343.90, 68.78, 29.85],
    [3, '四川省好医助医疗器械有限公司', 65.97, 65.97, 63.97, 65.97, 65.97, 327.85, 65.57, 29.64],
]
for i, row_data in enumerate(scores):
    for j, val in enumerate(row_data):
        f = green_fill if i == 0 else None
        font_style = bold_font if j == 7 else normal_font
        cell(ws2, 6+i, j+1, val, font_style, f)
        if isinstance(val, float):
            ws2.cell(row=6+i, column=j+1).number_format = '0.00'

# Highlight consistency issue
r = 10
cell(ws2, r, 1, '异常分析', red_font, red_fill, w_align)
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
notes = [
    '4/5评委(郑雁/杨京儒/王文霞/肖晓辉)评分完全一致(精确到0.01分)，含10分主观分的评审中出现此等一致性概率极低。',
    '唯一不同的评委罗征洪恰好给每个投标人各少2分，差异模式过于工整。'
]
for i, n in enumerate(notes):
    cell(ws2, 11+i, 1, n, normal_font, align=w_align)
    ws2.merge_cells(start_row=11+i, start_column=1, end_row=11+i, end_column=10)

# Part B: Non-price score gap
r = 14
cell(ws2, r, 1, 'B. 非价格得分(满分70=技术45+演示12+服务10+业绩2+环境1)差距分析', bold_font, sub_header_fill, w_align)
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)

r = 15
hdr(ws2, r, ['排名', '供应商', '总得分', '价格得分', '非价格得分', '非价格得分率', '与第1名差距', '差距倍数'])

np_scores = [
    [1, '成都易可天地科技有限公司', 93.30, 30.00, 63.30, '90.4%', '—', '—'],
    [2, '江西正好医疗器械有限公司', 68.78, 29.85, 38.93, '55.6%', -24.37, '1.63倍'],
    [3, '四川省好医助医疗器械有限公司', 65.57, 29.64, 35.93, '51.3%', -27.37, '1.76倍'],
]
for i, row_data in enumerate(np_scores):
    for j, val in enumerate(row_data):
        f = green_fill if i == 0 else None
        font_style = red_font if (j == 6 and i > 0) else normal_font
        cell(ws2, 16+i, j+1, val, font_style, f)
        if isinstance(val, float):
            ws2.cell(row=16+i, column=j+1).number_format = '0.00'

r = 20
cell(ws2, r, 1, '断崖结论', red_font, red_fill, w_align)
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
cell(ws2, 21, 1, '第1名非价格得分是第2名的1.63倍、第3名的1.76倍。排除价格因素后，技术+演示+服务的综合实力差距呈断崖式分布，与正常竞争分布严重不符。提示招标技术参数可能为中标人"量身定制"，或第2/3名为"陪标"。', normal_font, align=w_align)
ws2.merge_cells(start_row=21, start_column=1, end_row=21, end_column=10)

# Part C: Price scores
r = 23
cell(ws2, r, 1, 'C. 价格评分明细(满分30)', bold_font, sub_header_fill, w_align)
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)

r = 24
hdr(ws2, r, ['排名', '供应商', '投标报价', '评标报价', '占限价', '价格得分', '与最低价分差', '计算公式'])

price_data = [
    [1, '成都易可天地科技有限公司', 1566000, 1566000, '98.0%', 30.00, '—', '30×(1,566,000/1,566,000)'],
    [2, '江西正好医疗器械有限公司', 1574022, 1574022, '98.5%', 29.85, -0.15, '30×(1,566,000/1,574,022)'],
    [3, '四川省好医助医疗器械有限公司', 1585000, 1585000, '99.2%', 29.64, -0.36, '30×(1,566,000/1,585,000)'],
]
for i, row_data in enumerate(price_data):
    for j, val in enumerate(row_data):
        f = green_fill if i == 0 else None
        cell(ws2, 25+i, j+1, val, normal_font, f, c_align if j != 7 else w_align)
        if j in [2, 3]:
            ws2.cell(row=25+i, column=j+1).number_format = '#,##0'

ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 28
for col_letter in ['C','D','E','F','G','H','I','J']:
    ws2.column_dimensions[col_letter].width = 13

# ============ Sheet 3: 元数据铁证 ============
ws3 = wb.create_sheet('L5-元数据铁证')

cell(ws3, 1, 1, 'L5 元数据交叉分析 — 易可天地与江西正好文档同源铁证', title_font, align=w_align)
ws3.merge_cells('A1:G1')
cell(ws3, 2, 1, '方法: PyMuPDF提取PDF元数据(Producer/Creator/Author/CreationDate) | 软件: 全部使用WPS文字编辑.docx后转PDF', s_font, align=w_align)
ws3.merge_cells('A2:G2')

r = 4
hdr(ws3, r, ['文件', '好医助 Author', '易可天地 Author', '江西正好 Author', 'Creator', '核查结论'])

data = [
    ['其他资料.docx.pdf', '何天真', 'ZXM', 'ZXM', 'WPS文字', '易可天地=江西正好'],
    ['技术要求应答表.docx.pdf', '何天真', 'WPS_1654344806', 'WPS_1654344806', 'WPS文字', '易可天地=江西正好'],
    ['投标(响应)函.docx.pdf', '何天真', '汪', '汪', 'WPS文字', '易可天地=江西正好'],
    ['服务/商务应答表.docx.pdf', '何天真', 'WPS_1654344806', 'WPS_1654344806', 'WPS文字', '易可天地=江西正好'],
    ['中小企业声明函.pdf', '(空)', '(空)', '(空)', 'Qt 5.15.2', '三方一致(模板)'],
    ['残疾人福利性单位声明函.pdf', '(空)', '(空)', '(空)', 'Qt 5.15.2', '三方一致(模板)'],
    ['监狱企业证明文件.pdf', '(空)', '(空)', '(空)', 'Qt 5.15.2', '三方一致(模板)'],
    ['报价表.pdf', '(空)', '(空)', '(空)', 'Chromium/Skia', '三方一致(平台生成)'],
]

for i, row_data in enumerate(data):
    for j, val in enumerate(row_data):
        is_shared = (j == 6 and '易可天地=江西正好' in str(val)) or (j in [2,3] and j == 3 and row_data[2] == row_data[3])
        if j == 6 and '易可天地=江西正好' in str(val):
            cell(ws3, 5+i, j+1, val, red_font, red_fill)
        elif j in [2,3] and i < 4:
            cell(ws3, 5+i, j+1, val, red_font, red_fill)
        else:
            cell(ws3, 5+i, j+1, val, normal_font, align=w_align if j == 0 else c_align)

# Explanation
r = 14
cell(ws3, r, 1, '证据解析', bold_font, red_fill, w_align)
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)

exp = [
    '1. WPS_1654344806: WPS Office安装时生成的用户标识。其中1654344806为Unix时间戳 → 转换为北京时间 2022-06-04 17:53:26。两台不同电脑的WPS不可能有相同的安装时间戳 → 证明两个文档在同一个WPS环境下创建。',
    '2. ZXM: 自定义用户名，在两个公司的"其他资料.docx"中同时作为Author出现。排除姓名拼音巧合(同日、同项目、同文档类型)的概率。',
    '3. 汪: 另一个自定义用户名，同时出现在两个公司的投标函中。',
    '4. 核心推论: A公司(易可天地)和B公司(江西正好)的投标文件编排由同一人/同一台电脑/同一WPS安装完成。两家注册地为成都和江西樟树，法人/股东不同，按常理不可能"恰好"在同一台电脑上编辑投标文件。',
    '5. 法律适用: 构成《政府采购法实施条例》第74条第(七)项"供应商之间协商报价、技术方案等投标文件实质性内容"的串通投标行为。'
]
for i, e_text in enumerate(exp):
    cell(ws3, 15+i, 1, e_text, normal_font, align=w_align)
    ws3.merge_cells(start_row=15+i, start_column=1, end_row=15+i, end_column=7)

ws3.column_dimensions['A'].width = 35
for c in range(2, 8):
    ws3.column_dimensions[get_column_letter(c)].width = 22

# ============ Sheet 4: 十层检测汇总 ============
ws4 = wb.create_sheet('十层检测全量结果')

cell(ws4, 1, 1, '串标围标十层检测体系 — 本项目完整结果', title_font, align=w_align)
ws4.merge_cells('A1:G1')

r = 3
hdr(ws4, r, ['层级', '检测维度', '检测方法', '本项目结果', '风险等级', '数据来源', '说明'])

matrix = [
    ['L1', '报价规律性', '报价极差/标准差/等差数列检测/分项限价比', '极差1.21%，全部紧贴限价97-99.9%', '🔴 高', '投标报价表.pdf', '异常集中。校服项目参考极差7.8%。'],
    ['L2', '投标IP/MAC', '电子交易系统登录日志', '❓ 缺数据', '— 待查', '需向代理机构调取', '同一IP=围标不可辩驳的铁证。'],
    ['L3', '文本雷同(TF-IDF)', 'jieba分词→TF-IDF→余弦相似度', '词级3-8%(正常)/char级89-94%(模板)', '🟢 低', '投标文件全量文本', '高char相似均对应标准模板文本。'],
    ['L4', '图片/资源哈希', 'PyMuPDF提取嵌入图片→SHA256哈希', '0跨公司重复(276独立/279实例)', '🟢 低', '投标PDF嵌入图片', '易可天地208张图片远超另两家(39/32)。'],
    ['L5', '元数据交叉', 'PyMuPDF提取PDF元数据(Author/Creator/Producer)', '易可天地=江西正好 Author完全一致', '🔴 高', 'PDF元数据', '铁证。同一人/同一电脑制作。详见Sheet3。'],
    ['L6', '文档结构/字体', 'PyMuPDF提取页面字体span统计', '好医助字体(6种)⊂易可天地(10种)', '🟡 中', 'PDF字体分析', '完全包含关系。非独立铁证。'],
    ['L7', '生成设备信息', 'PDF Producer/Creator字段', '三方软件一致(WPS/Qt/Chromium)', '🟢 低', 'PDF Producer字段', '均为常见软件，非独立证据。'],
    ['L8', '工商关联', '天眼查/企查查', '❓ 待查(网络超时)', '— 待查', '天眼查/企查查', '需查询股东/高管/对外投资交叉关系。'],
    ['L9', '保证金/资金链', '银行汇款凭证', 'N/A(本项目不收取投标保证金)', '— N/A', '招标文件', '无投标保证金，不适用。'],
    ['L10', '代理/签到/得分', '开标记录/评分表/授权委托书', '评分4/5一致+非价格断崖', '🔴 高', '归档资料P104/P106', '评审得分异常+非价格差距断崖。'],
]

for i, row_data in enumerate(matrix):
    for j, val in enumerate(row_data):
        if j == 4:
            fill = red_fill if '高' in str(val) else (yellow_fill if '中' in str(val) else (green_fill if '低' in str(val) else None))
            cell(ws4, 4+i, j+1, val, bold_font, fill)
        elif j in [3, 6]:
            cell(ws4, 4+i, j+1, val, s_font, align=w_align)
        else:
            cell(ws4, 4+i, j+1, val, normal_font, align=c_align if j != 3 else w_align)

ws4.column_dimensions['A'].width = 8
ws4.column_dimensions['B'].width = 16
ws4.column_dimensions['C'].width = 35
ws4.column_dimensions['D'].width = 35
ws4.column_dimensions['E'].width = 10
ws4.column_dimensions['F'].width = 22
ws4.column_dimensions['G'].width = 40

# ============ Sheet 5: 档案信息 ============
ws5 = wb.create_sheet('项目基本信息')

cell(ws5, 1, 1, '项目基本信息与档案', title_font, align=w_align)
ws5.merge_cells('A1:B1')

info = [
    ('项目名称', '2024年多功能急救实训室建设项目'),
    ('采购编号', 'N5100012025000628'),
    ('备案编号', '51000024210200046723'),
    ('采购人', '四川护理职业学院'),
    ('采购人地址', '成都市龙泉驿区龙都南路173号'),
    ('采购人联系人', '梅老师 028-63955482'),
    ('代理机构', '联投项目管理(集团)有限公司'),
    ('代理机构地址', '成都市高新区天府大道北段1700号'),
    ('采购预算', '1,751,000.00元'),
    ('最高限价', '1,598,400.00元'),
    ('采购方式', '公开招标(综合评分法，电子化采购)'),
    ('评审方法', '技术45+演示12+服务10+业绩2+环境1+价格30=100分'),
    ('核心产品', '急救训练系统(3套)/虚实结合系统(1套)，共38项'),
    ('中标供应商', '成都易可天地科技有限公司'),
    ('中标金额', '1,566,000.00元'),
    ('代理服务费', '16,980元(1.698万元)'),
    ('评审日期', '2025年5月26日 10:37-16:00'),
    ('评审专家', '杨京儒(采购人代表)、罗征洪、郑雁、王文霞、肖晓辉'),
    ('复核人员', '张星梅、韩茂宇'),
    ('监督投诉', '四川省财政厅 028-86723581/028-86723539/028-86723553'),
    ('投标人1', '四川省好医助医疗器械有限公司(成都) | 1,585,000元 | 总分65.57'),
    ('投标人2', '成都易可天地科技有限公司(成都) | 1,566,000元 | 总分93.30 | 中标'),
    ('投标人3', '江西正好医疗器械有限公司(樟树) | 1,574,022元 | 总分68.78'),
]
for i, (k, v) in enumerate(info):
    cell(ws5, 3+i, 1, k, bold_font, sub_header_fill, w_align)
    cell(ws5, 3+i, 2, v, normal_font, align=w_align)

ws5.column_dimensions['A'].width = 20
ws5.column_dimensions['B'].width = 65

# Save
out = r'C:\Users\scrccpa\Desktop\急救实训室审计发现问题清单.xlsx'
wb.save(out)
print('Saved to desktop: 急救实训室审计发现问题清单.xlsx')
