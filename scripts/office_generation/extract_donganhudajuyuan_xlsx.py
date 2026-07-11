import json
import sys
from pathlib import Path

import openpyxl
from docx import Document

sys.stdout.reconfigure(encoding='utf-8')

BASE = Path(r'C:\Users\scrccpa\Desktop\新建文件夹')
XLSX = BASE / '年末补贴申请资料' / '佐证资料' / '成都东安湖大剧院运营项目补贴年度考核兑付评分表（第一年度）.xlsx'
OUT = Path(r'C:\Users\scrccpa\.openclaw\workspace\outputs\donganhudajuyuan_review')
OUT.mkdir(parents=True, exist_ok=True)

wb = openpyxl.load_workbook(XLSX, data_only=False)
rows_out = []
for ws in wb.worksheets:
    for row in ws.iter_rows():
        vals = []
        has = False
        for c in row:
            v = c.value
            if v is not None:
                has = True
            vals.append(v)
        if has:
            rows_out.append({'sheet': ws.title, 'row': row[0].row, 'values': vals})

(OUT / 'scoring_table_rows.json').write_text(json.dumps(rows_out, ensure_ascii=False, indent=2, default=str), encoding='utf-8')

lines = ['# 评分表抽取', '']
for r in rows_out:
    lines.append(f"## {r['sheet']}!{r['row']}")
    lines.append(' | '.join('' if v is None else str(v) for v in r['values']))
    lines.append('')
(OUT / 'scoring_table_rows.md').write_text('\n'.join(lines), encoding='utf-8')
print(json.dumps({'xlsx': str(XLSX), 'sheets': wb.sheetnames, 'non_empty_rows': len(rows_out)}, ensure_ascii=False, indent=2))
