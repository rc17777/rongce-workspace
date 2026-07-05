import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

paths = [
    Path(r'C:\Users\scrccpa\.openclaw\workspace\outputs\donganhudajuyuan_review\东安湖大剧院运营补贴专项审计报告复核结果.xlsx'),
    Path(r'C:\Users\scrccpa\Desktop\东安湖大剧院运营补贴专项审计报告复核结果.xlsx'),
]

rows = [
    ['F-001', '原P1重大', '报告前文合格档表述与考核办法不一致，是否影响319.60万元结论？', '报告审计结果载明年终考核91分；考核办法载明良好80分（含）以上按100%兑付；金额计算1700-850-530.40=319.60。', '91分已达到良好档，报告后文金额计算依据成立；错误仅限前文合格档文字引用。', 'P2一般', '是', '由“重大问题”调整为“文字引用瑕疵”，不影响本次金额结论。'],
    ['F-002', 'P1重大', '底稿第四/第五年600万元是否影响报告结论？', '底稿风险初步评价写600万元；报告和考核办法均为1200万元。', '本项目为第一年度，核心计算使用前三年1700万元，不直接影响319.60万元；但底稿与依据明显不一致。', 'P1重大', '否', '涉及协议核心补贴档位，影响底稿归档一致性，仍需修改。'],
    ['F-003', 'P1重大', '底稿年度考核周期6个月是否影响复核判断？', '底稿风险初步评价写年度考核周期6个月；报告和考核办法均为12个月。', '报告正文已写12个月，金额计算使用年终考核91分；错误主要在底稿。', 'P1重大', '否', '考核周期属于兑付核心条件，底稿必须修正。'],
    ['F-004', 'P2一般', '统一社会信用代码底稿错误是否影响报告结论？', '报告和取证单为91510112MAE39T7D4F；底稿局部少“12”。', '主体名称、报告正文正确，属于底稿基础信息错误。', 'P2一般', '否', '不影响金额，但需修正底稿主体信息。'],
    ['F-005', 'P2一般', '第二笔530.40万元来源说明不足是否影响金额？', '报告列示第一笔850、第二笔530.40、第三笔319.60；算术关系成立。', '第二笔为已支付金额，当前报告未展开形成依据，但不导致第三笔算术错误。', 'P2一般', '否', '建议补充说明或附计算表。'],
]

for path in paths:
    wb = openpyxl.load_workbook(path)
    if '结论自检表' in wb.sheetnames:
        del wb['结论自检表']
    ws = wb.create_sheet('结论自检表', 2)
    headers = ['原编号', '原风险等级', '自检问题', '支持证据', '反证/缓释因素', '校准后等级', '是否调整', '调整理由']
    ws.append(headers)
    for row in rows:
        ws.append(row)
    thin = Side(style='thin', color='A6A6A6')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = border
    for cell in ws[1]:
        cell.fill = PatternFill('solid', fgColor='1F4E78')
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    widths = [12, 14, 40, 58, 52, 14, 12, 42]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 72
        if ws.cell(row, 7).value == '是':
            ws.cell(row, 7).fill = PatternFill('solid', fgColor='FFF2CC')
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:H{ws.max_row}'
    ws.sheet_view.showGridLines = False
    wb.save(path)
    print(f'updated: {path}')
