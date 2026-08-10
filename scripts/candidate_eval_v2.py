# -*- coding: utf-8 -*-
"""融策审计助理招聘 — 候选人综合评分表 v2（含陈卓懿实操报告评分）"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ============ 品牌色 ============
DARK_BLUE = "0A1F3F"
TEAL = "1A5C6E"
COPPER = "C5955C"
WARM_GRAY = "F5F2EC"
WHITE = "FFFFFF"
LIGHT_BLUE = "D6E4F0"
LIGHT_GREEN = "E2EFDA"
LIGHT_RED = "FCE4D6"
HEADER_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
COPPER_FILL = PatternFill(start_color=COPPER, end_color=COPPER, fill_type="solid")
WARM_FILL = PatternFill(start_color=WARM_GRAY, end_color=WARM_GRAY, fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
LIGHT_RED_FILL = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_font = Font(name='微软雅黑', size=11, bold=True, color=WHITE)
title_font = Font(name='微软雅黑', size=14, bold=True, color=DARK_BLUE)
normal_font = Font(name='微软雅黑', size=10)
bold_font = Font(name='微软雅黑', size=10, bold=True)
copper_font = Font(name='微软雅黑', size=10, bold=True, color=COPPER)
small_font = Font(name='微软雅黑', size=9, color="666666")
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def apply_header_style(ws, row, col_range):
    for col in col_range:
        c = ws.cell(row=row, column=col)
        c.font = header_font
        c.fill = HEADER_FILL
        c.alignment = center_align
        c.border = thin_border

def apply_cell_style(cell, font=normal_font, fill=None, align=center_align):
    cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = align
    cell.border = thin_border

# ==========================================
# Sheet 1: 综合评分对比
# ==========================================
ws1 = wb.active
ws1.title = "综合评分对比"

ws1.merge_cells('A1:H1')
ws1['A1'] = '融策审计助理招聘 — 候选人综合评分对比'
ws1['A1'].font = title_font
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 36

ws1.merge_cells('A2:H2')
ws1['A2'] = '评估日期：2026年8月6日 | 陈卓懿试卷类型：实操报告（竣工财务决算审核报告）| 刘洋/杜巧丽：应届生笔试'
ws1['A2'].font = small_font
ws1['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[2].height = 22

headers = ['评分维度', '权重', '陈卓懿\n(社招·男·25岁)', '得分率', '刘洋\n(校招·应届·22岁)', '得分率', '杜巧丽\n(校招·应届·24岁)', '得分率']
for i, h in enumerate(headers, 1):
    ws1.cell(row=3, column=i, value=h)
apply_header_style(ws1, 3, range(1, 9))
ws1.row_dimensions[3].height = 36

dimensions = [
    ("一、硬性条件匹配度", 25,
     ["本科(会计学) ✓ | 初级会计证 ✓ | 专业对口 ✓", 23],
     ["本科(财务管理·前5%) ✓ | 初级会计证+初级审计证 ✓ | 专业对口 ✓", 24],
     ["本科(审计学) ✓ | 初级会计证 ✓ | 专业对口 ✓", 23]),
    ("  学历（全日制本科/财会专业）", 10,
     ["南京理工大学泰州科技学院·会计学·本科", 9],
     ["西南交通大学希望学院·财务管理·本科·GPA 3.52·前5%", 9],
     ["西南财经大学天府学院·审计学·本科", 9]),
    ("  证书（初级/中级会计证等）", 10,
     ["初级会计证", 8],
     ["初级会计证 + 初级审计证 + 英语四级 + 计算机二级", 9],
     ["初级会计证(2025年) + 计算机二级", 8]),
    ("  专业对口度", 5,
     ["会计学 → 审计岗位 ★★★★", 5],
     ["财务管理 → 审计岗位 ★★★★", 5],
     ["审计学 → 审计岗位 ★★★★★", 5]),
    ("二、工作经验与项目匹配度", 25,
     ["2.5年事务所审计助理 | 多类型政府审计项目 | 竣工决算报告编制", 25],
     ["碧桂园助理会计+亿企薪福政务岗+财政处实习 | 固定资产盘点经验", 19],
     ["事务所审计实习6个月+外勤会计2个月 | 全流程审计参与", 20]),
    ("  审计/事务所经验年限", 8,
     ["2024.02-至今(2.5年)·四川中御会计师事务所", 8],
     ["实习经历为主·无正式事务所工作经验", 4],
     ["2025.07-2026.01(6个月)·四川新智会计师事务所", 6]),
    ("  政府审计项目经验", 12,
     ["高标准农田·校园餐·帮扶资产清查·预决算审计·经责审计·竣工决算", 12],
     ["亿企薪福：6家事业单位固定资产盘点 | 与政府审计有一定关联", 7],
     ["事务所实习偏社会审计·政府项目经验较少", 5]),
    ("  项目参与深度与独立性", 5,
     ["独立编制竣工决算审核报告·独立完成年报·多类项目独立承担", 5],
     ["实习生角色参与·执行层工作", 3],
     ["从凭证抽查到底稿编制全流程参与·独立性较好", 4]),
    ("三、专业能力与技能", 20,
     ["WPS+Excel公式·底稿编制·报告撰写·数据透视·多类型项目覆盖", 19],
     ["用友U8+金蝶·竞赛获奖(国家级/省级)·学习能力强·底稿经验欠缺", 15],
     ["底稿编制·审计程序执行·函证·监盘·办公软件", 16]),
    ("四、综合素质", 15,
     ["跨方沟通强·事务所高压适应·社招意向明确", 14],
     ["学生干部·婚礼主持人(沟通突出)·多项奖学金·竞赛获奖", 13],
     ["吃苦耐劳·适应力强·态度积极·踏实肯干", 12]),
    ("五、试卷/实操成绩", 15,
     ["实操报告83分→折算12/15 | 竣工决算审核报告·结构完整·审计发现到位", 12],
     ["单选10/10 | 多选5/5 | 简答12/15 | 案例分析25/35 | 总分87→折算13/15", 13],
     ["单选~9/10 | 多选5/5 | 简答13/15 | 案例分析28/35 | 总分88→折算13/15", 13]),
]

row = 4
for dim_name, weight, czj, ly, dql in dimensions:
    ws1.cell(row=row, column=1, value=dim_name)
    ws1.cell(row=row, column=2, value=weight)
    ws1.cell(row=row, column=3, value=f"{czj[0]}\n得分：{czj[1]}/{weight}")
    ws1.cell(row=row, column=4, value=f"{czj[1]/weight*100:.0f}%")
    ws1.cell(row=row, column=5, value=f"{ly[0]}\n得分：{ly[1]}/{weight}")
    ws1.cell(row=row, column=6, value=f"{ly[1]/weight*100:.0f}%")
    ws1.cell(row=row, column=7, value=f"{dql[0]}\n得分：{dql[1]}/{weight}")
    ws1.cell(row=row, column=8, value=f"{dql[1]/weight*100:.0f}%")

    for c in range(1, 9):
        apply_cell_style(ws1.cell(row=row, column=c), align=left_align if c in [1,3,5,7] else center_align)

    if dim_name.startswith("  "):
        ws1.cell(row=row, column=1).font = Font(name='微软雅黑', size=9, color="555555")
        ws1.row_dimensions[row].height = 28
    elif dim_name.startswith("一") or dim_name.startswith("二") or dim_name.startswith("三") or dim_name.startswith("四") or dim_name.startswith("五"):
        for c in range(1, 9):
            ws1.cell(row=row, column=c).fill = LIGHT_BLUE_FILL
            ws1.cell(row=row, column=c).font = bold_font
        ws1.row_dimensions[row].height = 36
    else:
        ws1.row_dimensions[row].height = 64

    row += 1

# Totals
czj_main = [23, 25, 19, 14, 12]  # Updated with new scores
ly_main = [24, 19, 15, 13, 13]
dql_main = [23, 20, 16, 12, 13]

czj_total = sum(czj_main)
ly_total = sum(ly_main)
dql_total = sum(dql_main)

row_total = row
ws1.merge_cells(start_row=row_total, start_column=1, end_row=row_total, end_column=2)
ws1.cell(row=row_total, column=1, value='综合总分（满分100）')
ws1.cell(row=row_total, column=3, value=czj_total)
ws1.cell(row=row_total, column=4, value=f'{czj_total}%')
ws1.cell(row=row_total, column=5, value=ly_total)
ws1.cell(row=row_total, column=6, value=f'{ly_total}%')
ws1.cell(row=row_total, column=7, value=dql_total)
ws1.cell(row=row_total, column=8, value=f'{dql_total}%')
for c in range(1, 9):
    apply_cell_style(ws1.cell(row=row_total, column=c), font=Font(name='微软雅黑', size=11, bold=True, color=WHITE))
    ws1.cell(row=row_total, column=c).fill = COPPER_FILL
ws1.row_dimensions[row_total].height = 30
row += 1

# Ranking
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
ws1.cell(row=row, column=1, value='综合排名')
ws1.cell(row=row, column=3, value='🥇 第1名（93分）')
ws1.cell(row=row, column=5, value='🥈 第2名（84分）')
ws1.cell(row=row, column=7, value='🥈 第2名（84分）')
for c in range(1, 9):
    apply_cell_style(ws1.cell(row=row, column=c), font=bold_font, fill=WARM_FILL)
    if c in [4, 6, 8]:
        ws1.cell(row=row, column=c).value = ''
ws1.row_dimensions[row].height = 28
row += 1

# Recommendation
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
ws1.cell(row=row, column=1, value='录用建议')
ws1.cell(row=row, column=3, value='✅ 强烈推荐录用\n政府审计实战经验突出，竣工决算报告独立编制能力已达标')
ws1.cell(row=row, column=5, value='⭐ 推荐录用\n综合素质出众，双证+竞赛，可培养潜力大')
ws1.cell(row=row, column=7, value='⭐ 推荐录用\n审计专业基础扎实，绩效评价体系设计能力突出')
for c in range(1, 9):
    apply_cell_style(ws1.cell(row=row, column=c), font=bold_font)
    if c in [4, 6, 8]:
        ws1.cell(row=row, column=c).value = ''
ws1.row_dimensions[row].height = 48
row += 1

# Old score comparison row
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
ws1.cell(row=row, column=1, value='📝 说明')
ws1.cell(row=row, column=3, value='原评86分（无试卷折算）\n→更新为93分（含实操报告）')
ws1.cell(row=row, column=5, value='')
ws1.cell(row=row, column=7, value='')
for c in range(1, 9):
    apply_cell_style(ws1.cell(row=row, column=c), font=small_font, fill=YELLOW_FILL)
ws1.row_dimensions[row].height = 30
row += 2

# Summary text
ws1.merge_cells(start_row=row, start_column=1, end_row=row+6, end_column=8)
summary_text = (
    "【评估说明】\n"
    "1. 评分采用百分制，五大维度加权：硬性条件25% + 工作经验25% + 专业能力20% + 综合素质15% + 试卷/实操15%\n"
    "2. 陈卓懿以实操成果（竣工财务决算审核报告）代替笔试，报告评分83/100折合12/15分\n"
    "   — 报告亮点：结构完整规范，法规引用准确，审计发现（未按合同支付进度款）有具体金额和条款支撑\n"
    "   — 报告瑕疵：P5行存在复制粘贴错误（'四三七处'应为'马尔康市卫健局'），Excel附表'方案编制费'页标题误写为'审查费'\n"
    "3. 刘洋与杜巧丽均为应届生笔试，客观题均接近满分；杜巧丽Q1（政府审计主体）选错，但绩效评价方案更系统\n"
    "4. 三人均可匹配「审计助理（持初级会计证）」岗位，陈卓懿可考虑按高级助理/准项目经理方向培养\n"
    "5. 建议全部安排面试，面试重点：陈卓懿—核实报告独立编制程度；刘洋—考察审计底稿实操；杜巧丽—补测基础概念"
)
ws1.cell(row=row, column=1, value=summary_text)
ws1.cell(row=row, column=1).font = Font(name='微软雅黑', size=9, color="444444")
ws1.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

col_widths = [28, 6, 32, 7, 32, 7, 32, 7]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.sheet_properties.tabColor = DARK_BLUE

# ==========================================
# Sheet 2: 评分明细
# ==========================================
ws2 = wb.create_sheet("评分明细")

ws2.merge_cells('A1:G1')
ws2['A1'] = '融策审计助理招聘 — 评分明细表'
ws2['A1'].font = title_font
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 36

detail_headers = ['评估维度', '子项', '满分', '陈卓懿\n得分', '刘洋\n得分', '杜巧丽\n得分', '评分说明']
for i, h in enumerate(detail_headers, 1):
    ws2.cell(row=3, column=i, value=h)
apply_header_style(ws2, 3, range(1, 8))
ws2.row_dimensions[3].height = 32

details = [
    ("一、硬性条件", "学历层次", 5, 4, 4, 4, "均为全日制本科，非985/211统一给4分"),
    ("", "专业匹配度", 5, 5, 4, 5, "审计学>会计学>财务管理；杜巧丽审计学最优"),
    ("", "初级会计证", 5, 5, 5, 5, "三人均持初级会计证"),
    ("", "附加证书", 5, 1, 4, 2, "刘洋：+初级审计证/+英语四级/+计算机二级；杜巧丽：+计算机二级"),
    ("", "学业表现", 5, 3, 5, 3, "刘洋：GPA 3.52/前5%/多项奖学金；其余两人未突出"),
    ("二、工作经验", "事务所/审计经验", 10, 10, 3, 6, "陈2.5年全职事务所→满分；杜6个月事务所实习；刘仅有非审计实习"),
    ("", "政府审计项目", 10, 10, 6, 4, "陈：高标准农田/校园餐/帮扶资产/预决算/经责/竣工决算→满分"),
    ("", "项目角色深度", 5, 5, 3, 4, "陈独立出具年报+竣工决算报告；杜协助底稿全流程；刘实习角色执行层"),
    ("三、专业能力", "底稿编制能力", 5, 5, 2, 4, "陈独立编制底稿+Excel附表公式联动；杜在指导下编制；刘无明确经验"),
    ("", "报告撰写能力", 5, 5, 2, 3, "陈独立编制竣工决算审核报告（实操证明）；杜辅助性参与；刘无经验"),
    ("", "软件工具掌握", 5, 4, 4, 3, "陈：WPS+Excel公式/数据透视；刘：用友U8/金蝶；杜：办公软件"),
    ("", "项目类型覆盖", 5, 5, 2, 3, "陈覆盖预决算/经责/专项资金/竣工决算等多类型→满分"),
    ("四、综合素质", "沟通表达能力", 4, 3, 4, 3, "刘洋婚礼主持人经验突出；陈跨方沟通强；杜待人友善"),
    ("", "抗压与责任心", 4, 4, 3, 3, "陈事务所高压环境2年+/融策报告产出；刘学生干部+竞赛；杜吃苦耐劳"),
    ("", "学习与成长性", 4, 3, 4, 3, "刘：前5%+多项奖学金+竞赛奖；陈和杜均一般"),
    ("", "稳定性与意向", 3, 3, 2, 3, "陈社招意向明确；杜审计专业意向清晰；刘校招初次就业"),
    ("五、试卷/实操", "单选题(30分) / 实操-结构完整", 5, "—", 5, 4, "刘10/10满分；杜约9/10(Q1错)"),
    ("", "多选题(20分) / 实操-法规引用", 4, "—", 4, 4, "两人均5/5满分"),
    ("", "简答题(15分) / 实操-审计发现", 3, "—", 2, 2, "两人要点覆盖但表述有提升空间"),
    ("", "案例分析(35分) / 实操-专业表达", 3, "—", 2, 3, "杜绩效评价体系更完整"),
    ("", "【陈卓懿专项】实操报告综合评分", 15, 12, "—", "—",
     "报告评分83/100→折合12/15。\n"
     "✓ 结构完整25/25 ✓ 法规引用15/15 ✓ 数据底稿18/20\n"
     "✓ 审计发现15/15 ✓ 专业表达15/15\n"
     "✗ 复制粘贴错误-5(四三七处≠卫健局) ✗ Excel标题错误-2"),
]

row = 4
current_section = None
for section, sub, max_score, s1, s2, s3, note in details:
    if section and section != current_section:
        current_section = section
        ws2.cell(row=row, column=1, value=section)
        ws2.cell(row=row, column=1).font = bold_font
        for c in range(1, 8):
            ws2.cell(row=row, column=c).fill = LIGHT_BLUE_FILL
            apply_cell_style(ws2.cell(row=row, column=c), font=bold_font)
        ws2.row_dimensions[row].height = 24
        row += 1

    ws2.cell(row=row, column=1, value='')
    ws2.cell(row=row, column=2, value=sub)
    ws2.cell(row=row, column=3, value=max_score)
    ws2.cell(row=row, column=4, value=s1)
    ws2.cell(row=row, column=5, value=s2)
    ws2.cell(row=row, column=6, value=s3)
    ws2.cell(row=row, column=7, value=note)
    for c in range(1, 8):
        apply_cell_style(ws2.cell(row=row, column=c), align=left_align if c in [2,7] else center_align)
    # Highlight the new row
    if "陈卓懿专项" in sub:
        ws2.row_dimensions[row].height = 72
        for c in range(1, 8):
            ws2.cell(row=row, column=c).fill = YELLOW_FILL
    else:
        ws2.row_dimensions[row].height = 26
    row += 1

ws2.column_dimensions['A'].width = 14
ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 6
ws2.column_dimensions['D'].width = 12
ws2.column_dimensions['E'].width = 12
ws2.column_dimensions['F'].width = 12
ws2.column_dimensions['G'].width = 56
ws2.sheet_properties.tabColor = TEAL

# ==========================================
# Sheet 3: 试卷分析（陈卓懿实操报告 + 刘洋/杜巧丽笔试）
# ==========================================
ws3 = wb.create_sheet("试卷分析")

ws3.merge_cells('A1:G1')
ws3['A1'] = '试卷/实操成果分析'
ws3['A1'].font = title_font
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 36

# ====== Part 0: 陈卓懿实操报告评分 ======
ws3.merge_cells('A3:G3')
ws3['A3'] = '【陈卓懿】实操成果：马尔康市脚木足乡卫生院灾后维修加固项目 竣工财务决算审核报告'
ws3['A3'].font = Font(name='微软雅黑', size=10, bold=True, color=DARK_BLUE)
ws3['A3'].fill = LIGHT_BLUE_FILL
ws3.row_dimensions[3].height = 24

report_headers = ['评分维度', '满分', '得分', '评价要点', '', '', '']
for i, h in enumerate(report_headers, 1):
    if h:
        ws3.cell(row=4, column=i, value=h)
apply_header_style(ws3, 4, [1,2,3,4])

report_data = [
    ("报告结构完整性", 25, 25, "标准竣工财务决算审核报告格式：项目基本情况→审核依据→审核程序方法→审核结果（概算/合同/结算/资金/工期/效益）→发现问题→审核结论→附件。完整规范，符合财建〔2016〕503号要求。"),
    ("法规引用准确性", 15, 15, "正确引用14项审核依据：财建〔2016〕503号、财政部令第81号、财建〔2016〕504号、财协〔1999〕103号、《中国注册会计师审计准则》、招标投标法、政府采购法等。覆盖全面，层级清晰。"),
    ("数据分析与底稿质量", 20, 18, "各费用科目（设计/施工/监理/造价咨询/鉴定/审查/方案编制）合同金额、结算价、审定金额、支付金额逐项列示。Excel附表含公式联动（SUM/跨表引用）。扣2分：方案编制费页标题误写为'审查费'。"),
    ("审计发现质量", 15, 15, "发现'未按照合同约定支付进度款'问题：合同约定验收后付至75%→实际付75%，但完成审计后应付至97%未执行。问题有具体金额、合同条款支撑、整改建议。精准到位。"),
    ("专业表达与格式", 15, 15, "专业用语规范，逻辑清晰，报告层次分明。含'征求意见书'格式。签章页完整（双注+事务所盖章+日期）。"),
    ("扣分：复制粘贴错误", 0, -5, "P5行：'是四三七处的责任'应改为'是马尔康市卫生健康局的责任'。属于模板复制遗留错误，影响专业性。"),
    ("扣分：附表标题错误", 0, -2, "Excel附表'方案编制费'页标题行写的是'审查费'而非'方案编制费'。细节检查不到位。"),
]

for i, (dim, max_s, score, note) in enumerate(report_data):
    r = 5 + i
    ws3.cell(row=r, column=1, value=dim)
    ws3.cell(row=r, column=2, value=max_s)
    ws3.cell(row=r, column=3, value=score)
    ws3.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
    ws3.cell(row=r, column=4, value=note)
    for c in [1,2,3]:
        apply_cell_style(ws3.cell(row=r, column=c))
    apply_cell_style(ws3.cell(row=r, column=4), align=left_align)
    if score < 0:
        ws3.cell(row=r, column=3).fill = LIGHT_RED_FILL
    ws3.row_dimensions[r].height = 36

r_total = 12
ws3.merge_cells(start_row=r_total, start_column=1, end_row=r_total, end_column=2)
ws3.cell(row=r_total, column=1, value='报告总分（满分100）')
ws3.cell(row=r_total, column=3, value=83)
ws3.cell(row=r_total, column=4, value='良好。报告整体质量达到了执业水平，具备独立编制竣工决算审核报告能力。主要扣分在细节检查（模板复制残留+Excel标题错误），属于经验积累问题。')
for c in [1,2,3,4]:
    apply_cell_style(ws3.cell(row=r_total, column=c), font=Font(name='微软雅黑', size=11, bold=True, color=WHITE))
ws3.cell(row=r_total, column=1).fill = COPPER_FILL
ws3.cell(row=r_total, column=2).fill = COPPER_FILL
ws3.cell(row=r_total, column=3).fill = COPPER_FILL
ws3.cell(row=r_total, column=4).fill = COPPER_FILL
ws3.row_dimensions[r_total].height = 30

# ====== Part 1: 刘洋/杜巧丽 笔试 ======
r_start = 15
ws3.merge_cells(start_row=r_start, start_column=1, end_row=r_start, end_column=7)
ws3.cell(row=r_start, column=1, value='【刘洋 / 杜巧丽】应届生招聘笔试 — 客观题答案对照')
ws3.cell(row=r_start, column=1).font = Font(name='微软雅黑', size=10, bold=True, color=DARK_BLUE)
ws3.cell(row=r_start, column=1).fill = LIGHT_BLUE_FILL
ws3.row_dimensions[r_start].height = 24

mc_headers = ['题号', '题目（摘要）', '正确答案', '刘洋答案', '刘洋结果', '杜巧丽答案', '杜巧丽结果']
r = r_start + 1
for i, h in enumerate(mc_headers, 1):
    ws3.cell(row=r, column=i, value=h)
apply_header_style(ws3, r, range(1, 8))

mc_data = [
    (1, '我国政府审计的主体', 'B.审计机关', 'B', '✓', 'C', '✗'),
    (2, '政府审计的核心依据不包括', 'C.企业会计准则', 'C', '✓', 'D(OCR存疑)', '?'),
    (3, '不属于政府审计重点对象', 'B.国企经营财报', 'B', '✓', 'B', '✓'),
    (4, '三公经费不包含', 'C.公务员工资', 'C', '✓', 'C', '✓'),
    (5, '现场审计首要基础工作', 'B.编制审计实施方案', 'B', '✓', 'B', '✓'),
    (6, '专项资金最常见违规', 'B.资金挤占挪用滞留', 'B', '✓', 'B', '✓'),
    (7, '政府审计与社会审计区别', 'B.强制性vs委托自愿', 'B', '✓', 'B', '✓'),
    (8, '预算执行审计期间', 'B.政府财政预算年度', 'B', '✓', 'B', '✓'),
    (9, '审计证据质量要求不包括', 'C.随意性', 'C', '✓', 'C', '✓'),
    (10, '政府投资项目审计重点', 'B.工程造价·资金使用·履约', 'B', '✓', 'B', '✓'),
]
for i, (num, topic, answer, ly_ans, ly_r, dql_ans, dql_r) in enumerate(mc_data):
    r2 = r + 1 + i
    ws3.cell(row=r2, column=1, value=num)
    ws3.cell(row=r2, column=2, value=topic)
    ws3.cell(row=r2, column=3, value=answer)
    ws3.cell(row=r2, column=4, value=ly_ans)
    ws3.cell(row=r2, column=5, value=ly_r)
    ws3.cell(row=r2, column=6, value=dql_ans)
    ws3.cell(row=r2, column=7, value=dql_r)
    for c in range(1, 8):
        apply_cell_style(ws3.cell(row=r2, column=c), align=left_align if c == 2 else center_align)
    ws3.cell(row=r2, column=5).fill = LIGHT_GREEN_FILL if '✓' in str(ly_r) else LIGHT_RED_FILL
    ws3.cell(row=r2, column=7).fill = LIGHT_GREEN_FILL if '✓' in str(dql_r) else LIGHT_RED_FILL
    ws3.row_dimensions[r2].height = 24

r2 = r + 1 + len(mc_data)
ws3.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=3)
ws3.cell(row=r2, column=1, value='单选题小计（满分30）')
ws3.cell(row=r2, column=4, value='30')
ws3.cell(row=r2, column=5, value='✓ 满分')
ws3.cell(row=r2, column=6, value='~27')
ws3.cell(row=r2, column=7, value='⚠ 第1题错·第2题OCR存疑')
for c in range(1, 8):
    apply_cell_style(ws3.cell(row=r2, column=c), font=bold_font)
ws3.row_dimensions[r2].height = 24

# 多选题
r3 = r2 + 2
ws3.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=7)
ws3.cell(row=r3, column=1, value='多项选择题（共5题，每题4分，共20分）')
ws3.cell(row=r3, column=1).font = bold_font
ws3.cell(row=r3, column=1).fill = WARM_FILL
ws3.row_dimensions[r3].height = 24

r3 += 1
for i, h in enumerate(mc_headers, 1):
    ws3.cell(row=r3, column=i, value=h)
apply_header_style(ws3, r3, range(1, 8))

ms_data = [
    (1, '政府审计常见业务类型', 'ABCD', 'ABCD', '✓', 'ABCD', '✓'),
    (2, '常用审计方法', 'ABCD', 'ABCD', '✓', 'ABCD', '✓'),
    (3, '经责审计对象', 'AB', 'AB', '✓', 'AB', '✓'),
    (4, '财政资金违规使用', 'ABC', 'ABC', '✓', 'ABC', '✓'),
    (5, '完整底稿应包含要素', 'ABCD', 'ABCD', '✓', 'ABCD', '✓'),
]
for i, (num, topic, answer, ly_ans, ly_r, dql_ans, dql_r) in enumerate(ms_data):
    r4 = r3 + 1 + i
    ws3.cell(row=r4, column=1, value=num)
    ws3.cell(row=r4, column=2, value=topic)
    ws3.cell(row=r4, column=3, value=answer)
    ws3.cell(row=r4, column=4, value=ly_ans)
    ws3.cell(row=r4, column=5, value=ly_r)
    ws3.cell(row=r4, column=6, value=dql_ans)
    ws3.cell(row=r4, column=7, value=dql_r)
    for c in range(1, 8):
        apply_cell_style(ws3.cell(row=r4, column=c), align=left_align if c == 2 else center_align)
    ws3.cell(row=r4, column=5).fill = LIGHT_GREEN_FILL
    ws3.cell(row=r4, column=7).fill = LIGHT_GREEN_FILL
    ws3.row_dimensions[r4].height = 24

r4 = r3 + 1 + len(ms_data)
ws3.merge_cells(start_row=r4, start_column=1, end_row=r4, end_column=3)
ws3.cell(row=r4, column=1, value='多选题小计（满分20）')
ws3.cell(row=r4, column=4, value='20')
ws3.cell(row=r4, column=5, value='✓ 满分')
ws3.cell(row=r4, column=6, value='20')
ws3.cell(row=r4, column=7, value='✓ 满分')
for c in range(1, 8):
    apply_cell_style(ws3.cell(row=r4, column=c), font=bold_font)

# 简答+案例
r5 = r4 + 2
ws3.merge_cells(start_row=r5, start_column=1, end_row=r5, end_column=7)
ws3.cell(row=r5, column=1, value='简答题（15分）+ 案例分析题（35分）— 主观题评分对比')
ws3.cell(row=r5, column=1).font = bold_font
ws3.cell(row=r5, column=1).fill = WARM_FILL
ws3.row_dimensions[r5].height = 24

r5 += 1
essay_headers = ['题型', '题目', '满分', '刘洋得分', '刘洋评价', '杜巧丽得分', '杜巧丽评价']
for i, h in enumerate(essay_headers, 1):
    ws3.cell(row=r5, column=i, value=h)
apply_header_style(ws3, r5, range(1, 8))

essay_data = [
    ('简答Q1', '如何判断被审计单位存在"小金库"', 5, 4, '要点覆盖（核对收支/突击盘点/访谈），但表述不够系统条理', 4, '要点覆盖（账外资金/资产/账簿核对/账外信息），较为完整'),
    ('简答Q2', '设置项目支出绩效评价一二级指标', 10, 8, '结构正确（投入/过程/产出/效益/满意度），二级指标有待细化', 9, '决策/过程/产出/效益/满意度，四级分解更完整，术语更规范'),
    ('案例分析Q1', '针对虚列/挪用/不公开提出审计建议(5分)', 5, 4, '建议方向正确（追回资金/完善制度/信息公开），可更具体', 4, '建议全面（完善内控/预算管理/资金管理/追回/公开），较完整'),
    ('案例分析Q2', '如何进一步调查取证确认问题(5分)', 5, 4, '方法到位（原始凭证/实物盘点/访谈/核对/查询），较全面', 4, '证据链完整（凭证/银行流水/访谈/材料缺失核查/专款核对）'),
    ('案例分析Q3', '退耕还林绩效评价工作方案+指标体系(20分)', 20, 14, '框架完整（目的/范围/方法/指标体系），指标覆盖较全', 17, '结构完整（目的/内容/步骤/指标/结论），绩效指标更系统'),
]
for i, (etype, topic, max_s, ly_s, ly_eval, dql_s, dql_eval) in enumerate(essay_data):
    r6 = r5 + 1 + i
    ws3.cell(row=r6, column=1, value=etype)
    ws3.cell(row=r6, column=2, value=topic)
    ws3.cell(row=r6, column=3, value=max_s)
    ws3.cell(row=r6, column=4, value=ly_s)
    ws3.cell(row=r6, column=5, value=ly_eval)
    ws3.cell(row=r6, column=6, value=dql_s)
    ws3.cell(row=r6, column=7, value=dql_eval)
    for c in range(1, 8):
        apply_cell_style(ws3.cell(row=r6, column=c), align=left_align if c in [2,5,7] else center_align)
    ws3.row_dimensions[r6].height = 48

r6 = r5 + 1 + len(essay_data)
ws3.merge_cells(start_row=r6, start_column=1, end_row=r6, end_column=3)
ws3.cell(row=r6, column=1, value='试卷总分（满分100）')
ws3.cell(row=r6, column=4, value='87')
ws3.cell(row=r6, column=5, value='优秀（87%）')
ws3.cell(row=r6, column=6, value='88')
ws3.cell(row=r6, column=7, value='优秀（88%）')
for c in range(1, 8):
    apply_cell_style(ws3.cell(row=r6, column=c), font=Font(name='微软雅黑', size=11, bold=True, color=WHITE))
    ws3.cell(row=r6, column=c).fill = COPPER_FILL

# Note
r7 = r6 + 2
ws3.merge_cells(start_row=r7, start_column=1, end_row=r7+4, end_column=7)
ws3.cell(row=r7, column=1, value=(
    "【判卷说明】\n"
    "1. 陈卓懿以实操报告代替笔试，评分侧重报告完整性、法规引用、审计发现质量和专业表达，扣分项为细节疏忽。\n"
    "2. 单选题第1题：我国政府审计的主体是国家审计机关（B），非企事业单位内审部门（C）。杜巧丽此题答错。\n"
    "3. 单选题第2题：杜巧丽OCR识别存疑，按标记处理；建议查看原件确认。\n"
    "4. 主观题因手写体OCR存在误差，评分基于可识别内容判断，建议面试时口头复核关键知识点。\n"
    "5. 综合来看：陈卓懿实操能力已达执业水平；刘洋基础理论扎实；杜巧丽绩效评价体系设计能力突出。"
))
ws3.cell(row=r7, column=1).font = Font(name='微软雅黑', size=9, color="444444")
ws3.cell(row=r7, column=1).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

ws3.column_dimensions['A'].width = 14
ws3.column_dimensions['B'].width = 34
ws3.column_dimensions['C'].width = 12
ws3.column_dimensions['D'].width = 12
ws3.column_dimensions['E'].width = 38
ws3.column_dimensions['F'].width = 12
ws3.column_dimensions['G'].width = 38
ws3.sheet_properties.tabColor = COPPER

# ==========================================
# Sheet 4: 陈卓懿实操报告评审
# ==========================================
ws4 = wb.create_sheet("陈卓懿实操报告评审")

ws4.merge_cells('A1:F1')
ws4['A1'] = '陈卓懿实操报告 — 竣工财务决算审核报告 详细评审'
ws4['A1'].font = title_font
ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[1].height = 36

# Report basic info
info_data = [
    ('报告名称', '关于马尔康市卫生健康局"马尔康市脚木足乡卫生院灾后维修加固项目"竣工财务决算审核报告'),
    ('报告日期', '2026年8月6日'),
    ('审计机构', '四川融策会计师事务所有限公司'),
    ('被审计单位', '马尔康市卫生健康局'),
    ('项目概算', '50万元（省预算+地方配套）'),
    ('审定投资', '494,652.68元（节约概算5,347.32元，节概率1.07%）'),
    ('报告类型', '竣工财务决算审核报告（含征求意见书）'),
    ('报告结构', '八章+附件：基本情况/审核依据/程序方法/审核结果/发现问题/审核结论'),
    ('附件质量', 'Excel附表含汇总+7个费用明细页，公式联动（SUM/跨表引用）'),
]

for i, (label, value) in enumerate(info_data):
    r = 3 + i
    ws4.cell(row=r, column=1, value=label)
    ws4.cell(row=r, column=1).font = bold_font
    ws4.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    ws4.cell(row=r, column=2, value=value)
    apply_cell_style(ws4.cell(row=r, column=1), font=bold_font)
    apply_cell_style(ws4.cell(row=r, column=2), align=left_align)
    ws4.row_dimensions[r].height = 22

ws4.column_dimensions['A'].width = 14
ws4.column_dimensions['B'].width = 18
ws4.column_dimensions['C'].width = 18
ws4.column_dimensions['D'].width = 18
ws4.column_dimensions['E'].width = 18
ws4.column_dimensions['F'].width = 18
ws4.sheet_properties.tabColor = "4472C4"

# save
output_path = r"C:\Users\scrccpa\Desktop\融策审计助理招聘_候选人综合评分.xlsx"
wb.save(output_path)
print(f"✅ 评分表（v2）已更新保存至：{output_path}")
print(f"陈卓懿总分：93（原86→+7分，因实操报告代替笔试大幅提升）")
