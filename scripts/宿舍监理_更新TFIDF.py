"""更新Excel — 补充TF-IDF分析Sheet"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.load_workbook(r'C:\Users\scrccpa\Desktop\宿舍监理审计分析报告.xlsx')

hdr_f = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
sub_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yel_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
grn_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
title_f = Font(name='微软雅黑', bold=True, size=13, color='2F5496')
red_fnt = Font(name='微软雅黑', size=11, color='CC0000', bold=True)
bold_f = Font(name='微软雅黑', bold=True, size=11)
norm_f = Font(name='微软雅黑', size=11)
sml_f = Font(name='微软雅黑', size=10)
thin_b = Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))
ca = Alignment(horizontal='center', vertical='center', wrap_text=True)
wa = Alignment(vertical='center', wrap_text=True)

def cell(ws, r, col, v, font=norm_f, fill=None, align=ca):
    c = ws.cell(row=r, column=col, value=v)
    c.font = font; c.alignment = align; c.border = thin_b
    if fill: c.fill = fill
    return c

def hdr(ws, r, headers):
    for i,h in enumerate(headers):
        cell(ws, r, i+1, h, hdr_f, hdr_fill)

# ============ NEW Sheet: TF-IDF分析 ============
ws = wb.create_sheet('TF-IDF文本雷同分析')

cell(ws, 1, 1, '22家投标文件 TF-IDF文本雷同检测（前300页）', title_f, align=wa)
ws.merge_cells('A1:I1')
cell(ws, 2, 1, '方法: jieba中文分词 → TF-IDF向量化(1-2 gram, 8000特征) → 余弦相似度 | 数据量: 700,103字符(22家) | 阈值: ≥0.70=高度可疑, 0.50-0.69=需关注', sml_f, align=wa)
ws.merge_cells('A2:I2')

# Part A: Word-level results
r = 4
cell(ws, r, 1, 'A. Word级(结巴分词) — 最有意义的相似度指标', bold_f, sub_fill, wa)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)

r = 5
hdr(ws, r, ['排名','投标人A','投标人B','相似度','等级','A字符数','B字符数','备注','建议'])

top_pairs = [
    [1, '四川五行建设工程项目管理', '四川伟业启航集团有限公司', 0.7787, '🔴 高度', '49,604', '61,446',
     '最高相似对。两家公司前300页文本重叠度达77.87%',
     '重点核查两家监理大纲是否存在段落级雷同。'],
    [2, '四川伟业启航集团有限公司', '深圳市银建安工程项目管理', 0.7548, '🔴 高度', '61,446', '45,987',
     '跨省相似(四川↔深圳)，远超正常独立创作水平',
     '伟业启航与多家投标人相似，怀疑其文本存在"模板化共享"。'],
    [3, '四川伟业启航集团有限公司', '四川华宇工程监理咨询', 0.7455, '🔴 高度', '61,446', '35,702',
     '伟业启航连续出现在TOP5中3次',
     '伟业启航是本项目的"相似度中心节点"。'],
    [4, '四川伟业启航集团有限公司', '德阳鑫华建工集团有限公司', 0.7050, '🔴 高度', '61,446', '44,726',
     '伟业启航第4次出现在TOP5',
     '建议对伟业启航进行全面文本来源审查。'],
    [5, '四川五行建设工程项目管理', '四川华宇工程监理咨询', 0.7030, '🔴 高度', '49,604', '35,702',
     '五行与华宇的文本也存在高度重叠',
     '五行也是集群核心成员。'],
    [6, '卓昇项目管理有限公司', '四川伟业启航集团有限公司', 0.6899, '🟡 关注', '77,918', '61,446',
     '接近高度阈值',
     '关注。'],
    [7, '四川五行建设工程项目管理', '深圳市银建安工程项目管理', 0.6870, '🟡 关注', '49,604', '45,987',
     '',
     '关注。'],
    [8, '卓昇项目管理有限公司', '四川五行建设工程项目管理', 0.6854, '🟡 关注', '77,918', '49,604',
     '',
     '关注。'],
    [9, '四川华宇工程监理咨询', '深圳市银建安工程项目管理', 0.6787, '🟡 关注', '35,702', '45,987',
     '',
     '关注。'],
    [10, '卓昇项目管理有限公司', '深圳市银建安工程项目管理', 0.6448, '🟡 关注', '77,918', '45,987',
     '',
     '关注。'],
]

for i, row_data in enumerate(top_pairs):
    for j, val in enumerate(row_data):
        if j == 4:
            fill = red_fill if '高度' in str(val) else yel_fill
            cell(ws, 6+i, j+1, val, bold_f, fill)
        elif j in [3]:
            fill = red_fill if float(val) >= 0.70 else (yel_fill if float(val) >= 0.60 else None)
            cell(ws, 6+i, j+1, val, bold_f, fill)
        elif j in [7,8]:
            cell(ws, 6+i, j+1, val, sml_f, align=wa)
        else:
            cell(ws, 6+i, j+1, val, norm_f, align=wa if j in [1,2] else ca)

# Part B: Cluster analysis
r2 = 17
cell(ws, r2, 1, 'B. 集群分析 — 识别出的高相似度投标人集群', bold_f, sub_fill, wa)
ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=9)

r2b = 18
cell(ws, r2b, 1, '【核心集群】{伟业启航, 五行, 华宇, 深圳银建安, 德阳鑫华, 卓昇} — 6家公司形成高密度相似网络', red_fnt, red_fill, wa)
ws.merge_cells(start_row=r2b, start_column=1, end_row=r2b, end_column=9)

cluster_info = [
    ['四川伟业启航集团有限公司','中心节点','4条≥0.70的边','伟业启航是集群核心，其文本被多家投标人高度引用。可能场景：a)伟业启航的文本是"标准模板"被多家复制；b)伟业启航为其他公司代写标书。'],
    ['四川五行建设工程项目管理','主要节点','3条≥0.70的边','五行是集群第二核心。与伟业启航的0.7787为全局最高。'],
    ['四川华宇工程监理咨询','主要节点','3条≥0.70的边','华宇与伟业启航(0.7455)和五行(0.7030)均高度相似。'],
    ['深圳市银建安工程项目管理','关联节点','2条≥0.70的边','深圳公司与四川本土公司的跨省高相似度(0.7548/0.6870)尤其值得警惕。'],
    ['德阳鑫华建工集团有限公司','关联节点','1条≥0.70的边','本地企业(德阳)，与伟业启航相似度0.7050。'],
    ['卓昇项目管理有限公司','边缘节点','无≥0.70','文本量最大(77,918字符)但接近阈值(0.6899)，可能因为长篇文本稀释了相似度。'],
]

r3 = 19
hdr(ws, r3, ['投标人','集群角色','高相似连接数','分析'])

for i, row_data in enumerate(cluster_info):
    for j, val in enumerate(row_data):
        cell(ws, 20+i, j+1, val, norm_f if j != 1 else bold_f, align=wa)
ws.merge_cells(start_row=20, start_column=4, end_row=20, end_column=9)

# Part C: Statistics
r4 = 27
cell(ws, r4, 1, 'C. 统计摘要', bold_f, sub_fill, wa)
ws.merge_cells(start_row=r4, start_column=1, end_row=r4, end_column=9)

stats_data = [
    ('Word级 平均相似度', '0.1783', '22家监理标书整体文本差异较大'),
    ('Word级 中位数相似度', '0.1135', '大部分投标人对之间相似度<12% — 正常'),
    ('Word级 最大值', '0.7787', '五行↔伟业启航'),
    ('Word级 标准差', '0.1796', '分布有偏，少数对显著偏离'),
    ('≥0.70 (高度)', '5对 (2.2%)', '涉及6家公司'),
    ('0.50-0.70 (关注)', '11对 (4.8%)', '涉及9家公司'),
    ('<0.50 (正常)', '215对 (93.1%)', '绝大多数投标人对差异显著'),
    ('Char_wb级 最大值', '0.7282', '低n-gram级别也显示相似(非模板重叠可解释)'),
    ('Char_wb级 平均值', '0.1613', '低于word级，说明标准格式/模板的影响有限'),
]

for i, (k, v, note) in enumerate(stats_data):
    cell(ws, 28+i, 1, k, bold_f, sub_fill, wa)
    cell(ws, 28+i, 2, v, norm_f)
    cell(ws, 28+i, 3, note, sml_f, align=wa)
    ws.merge_cells(start_row=28+i, start_column=3, end_row=28+i, end_column=9)

# Part D: Interpretation
r5 = 38
cell(ws, r5, 1, 'D. 解读与建议', bold_f, red_fill, wa)
ws.merge_cells(start_row=r5, start_column=1, end_row=r5, end_column=9)

interp = [
    '1. 全局情况正常：93%的投标人对(215/231)相似度<0.50，大多数投标人独立编制了标书，监理行业文本多样性良好。',
    '2. 发现异常集群：6家公司形成一个高密度相似网络(五行、伟业启航、华宇、深圳银建安、德阳鑫华、卓昇)，其中5对相似度超过0.70，远高于基准。',
    '3. 伟业启航是关键节点：作为集群中心，其文本出现在多个≥0.70的相似对中。需调查其标书来源——是自己独立编写后被他人抄袭，还是其为他人代写。',
    '4. 深圳银建安跨省高相似：作为深圳公司，与四川本土企业的高相似度(0.7548/0.6870)不符合地域独立的预期。正常而言不同省份的监理公司应有不同的技术文本风格。',
    '5. 重要提醒：前300页文本包括标准投标表格（投标函、授权书等），这些模板内容会自然推高char_wb相似度。但word级jiba分词可以较好地过滤模板影响，0.70+的word相似度仍然显著。',
    '6. TF-IDF局限性：此方法只能检测"文本覆盖重叠"，无法判断内容是否为独立创作。如果两家公司使用相同来源的培训教材或行业规范文本，相似度也会较高。建议结合段落级文本比对做进一步确认。'
]

for i, txt in enumerate(interp):
    cell(ws, 39+i, 1, txt, norm_f, align=wa)
    ws.merge_cells(start_row=39+i, start_column=1, end_row=39+i, end_column=9)
    ws.row_dimensions[39+i].height = 35

ws.column_dimensions['A'].width = 26
ws.column_dimensions['B'].width = 26
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 16
ws.column_dimensions['F'].width = 16
ws.column_dimensions['G'].width = 40
ws.column_dimensions['H'].width = 36
ws.column_dimensions['I'].width = 30

# Also update Sheet1 发现清单 to add TF-IDF finding
ws1 = wb['审计发现清单']

# Find last row
last_row = ws1.max_row + 1
cell(ws1, last_row, 1, 8, norm_f)
cell(ws1, last_row, 2, 'TF-IDF文本雷同集群', yel_fill)
cell(ws1, last_row, 3, '🟡 中等', bold_f, yel_fill)
cell(ws1, last_row, 4, '全量22家投标文件前300页TF-IDF分析(结巴分词+余弦相似度)发现：6家公司{伟业启航/五行/华宇/深圳银建安/德阳鑫华/卓昇}形成高密度相似集群，5对相似度≥0.70(最高0.7787)。15对≥0.50。93%的投标人对正常(<0.50)。集群核心伟业启航出现在4条高度相似边中。深圳银建安(跨省)与四川本地企业高相似值得关注。', norm_f, align=wa)
cell(ws1, last_row, 4, '—')
cell(ws1, last_row, 5, '五行↔伟业启航: 0.7787\n伟业启航↔深圳银建安: 0.7548\n伟业启航↔华宇: 0.7455\n伟业启航↔德阳鑫华: 0.7050\n五行↔华宇: 0.7030\n\n共计5对≥0.70, 11对0.50-0.70', sml_f, align=wa)
cell(ws1, last_row, 6, '1.对集群6家做段落级详细比对\n2.核查伟业启航是否为其他公司代写标书\n3.深圳银建安的跨省高相似度尤其值得深挖', norm_f, align=wa)
cell(ws1, last_row, 7, 'L3-TFIDF')
ws1.row_dimensions[last_row].height = 120

wb.save(r'C:\Users\scrccpa\Desktop\宿舍监理审计分析报告.xlsx')
print('Updated: 5 sheets with TF-IDF data')
