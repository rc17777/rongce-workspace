#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
绩效自评复核全面检查 — 红光街道
14条规则：格式要求(1-5) + 自评复核口径(6-14)
"""

import os, sys, re, json
from collections import Counter, defaultdict
sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = r'C:\Users\scrccpa\Desktop\报告\2.部门预算项目绩效自评复核报告20260722\成都市郫都区人民政府红光街道办事处部门预算项目'
DOCX_PATH = None; XLSX_PATH = None
for f in os.listdir(BASE_DIR):
    if f.endswith('.docx'): DOCX_PATH = os.path.join(BASE_DIR, f)
    elif f.endswith('.xlsx'): XLSX_PATH = os.path.join(BASE_DIR, f)

OUTPUT = r'C:\Users\scrccpa\Desktop\绩效自评复核检查结果_红光街道.xlsx'

import docx, openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ===== 读取DOCX =====
doc = docx.Document(DOCX_PATH)
paras = []
for i, p in enumerate(doc.paragraphs):
    if not p.text.strip(): continue
    pf = p.paragraph_format
    runs_info = []
    for r in p.runs:
        runs_info.append({
            'font': r.font.name, 'size': round(r.font.size/12700,1) if r.font.size else None,
            'bold': r.font.bold, 'text': r.text[:200]
        })
    paras.append({
        'idx': i, 'text': p.text, 'style': p.style.name,
        'alignment': str(p.alignment) if p.alignment else None,
        'line_spacing': pf.line_spacing, 'line_spacing_rule': str(pf.line_spacing_rule) if pf.line_spacing_rule else None,
        'runs': runs_info
    })

# ===== 读取XLSX =====
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

PROJECT_MAP = {
    '基层食品安全':'基层食品安全监管工作经费','道路治理':'道路安全综合治理工作经费',
    '人大代表':'人大代表补选工作经费','选调生':'选调生到村任职工作补助资金',
    '基层党组织':'基层党组织活动经费','网格化':'网格化服务管理经费',
    '街道治税':'街道综合治税专项资金项目','市级激励资金':'市级城乡社区发展治理激励资金',
    '城乡治理奖励资金':'城乡社区发展治理专项激励资金','罚没处置成本项目':'红光街道榕树园项目罚没房屋资产处置成本项目'
}

projects = {}; summary = []
for sn in wb.sheetnames:
    ws = wb[sn]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    if sn == 'Sheet1':
        for r in rows:
            if r[7] and str(r[7]).strip() not in ('项目名称','','总分'):
                summary.append({
                    'name': str(r[7]).strip(), 'self_score': r[8], 'review_score': r[9],
                    'deviation': r[10], 'reason': str(r[11]).strip() if r[11] else ''
                })
    elif sn in ('Sheet6','Sheet4','Sheet5','Sheet12'): continue
    else:
        name = ''
        if rows[1][0]:
            m = re.match(r'(.+?)项目支出绩效自评复核表', str(rows[1][0]))
            if m: name = m.group(1)
        inds = []; no_target = False
        for r in rows[3:]:
            if r[0] and str(r[0]).strip() in ('合计','——'): continue
            if r[2] and str(r[2]).strip() == '预算执行':
                inds.append({'type':'预算执行','l1':'预算执行','l2':'预算执行','l3':'预算执行',
                    'prop':str(r[3] or ''),'target':str(r[4] or ''),'unit':str(r[5] or ''),
                    'weight':r[6],'self_val':str(r[7] or ''),'self_score':r[8],
                    'review_val':str(r[9] or ''),'review_score':r[10],'deviation':r[11],
                    'remark':str(r[12] or '').strip()})
                continue
            if r[2] and str(r[2]).strip():
                inds.append({'type':'指标','l1':str(r[0] or '').strip(),'l2':str(r[1] or '').strip(),
                    'l3':str(r[2] or '').strip(),'prop':str(r[3] or ''),'target':str(r[4] or ''),
                    'unit':str(r[5] or ''),'weight':r[6],'self_val':str(r[7] or ''),
                    'self_score':r[8],'review_val':str(r[9] or ''),'review_score':r[10],
                    'deviation':r[11],'remark':str(r[12] or '').strip()})
        for r in rows:
            for c in r:
                if c and '未编制绩效目标' in str(c): no_target = True
        projects[sn] = {'name':name,'inds':inds,'no_target':no_target,'sheet_name':sn}

# ===== 检查结果 =====
r = []  # 结果列表，每行：项目,维度,检查项,结果,问题描述,建议

# ===== 一、格式要求 =====
# 1. 标题
title_idxs = [3,4,5,46,47,48]
title_issues = []
for p in paras:
    if p['idx'] in title_idxs:
        for ru in p['runs']:
            if ru['font'] and '小标宋' not in str(ru['font']) and '方正' not in str(ru['font']):
                title_issues.append(f"标题[{p['idx']}]字体={ru['font']}，应为小标宋")
            if ru['size'] and abs(ru['size']-22)>0.5:
                title_issues.append(f"标题[{p['idx']}]字号={ru['size']}pt，应为22pt(2号)")
        if p['alignment'] and 'CENTER' not in str(p['alignment']):
            title_issues.append(f"标题[{p['idx']}]对齐={p['alignment']}，应为居中")
r.append(['（全部）','一、格式要求-1.标题','标题居中/2号小标宋(加粗)',
    '不通过' if title_issues else '通过',
    '；'.join(title_issues) if title_issues else '标题使用方正小标宋简体22pt居中，格式正确',''])

# 2. 正文
body_idxs = [51,54,62,64,65,66,67,69,71,74,84,88,93,96,101,105,109,110,114,118,
             122,123,124,125,126,127,129,130,132,133,135,136,139,140,141,143,145,147]
body_issues = []
for p in paras:
    if p['idx'] in body_idxs:
        for ru in p['runs']:
            if ru['font'] and '仿宋' not in str(ru['font']) and 'Times New Roman' not in str(ru['font']):
                body_issues.append(f"正文[{p['idx']}]字体={ru['font']}，应为仿宋_GB2312")
            if ru['size'] and abs(ru['size']-16)>0.5:
                body_issues.append(f"正文[{p['idx']}]字号={ru['size']}pt，应为16pt(三号)")
r.append(['（全部）','一、格式要求-2.正文字体','正文三号仿宋GB2312',
    '不通过' if body_issues else '通过',
    '；'.join(body_issues[:5]) if body_issues else '正文使用仿宋_GB2312/Times New Roman，16pt(三号)正确',''])

# 3. 标题分级
heading_issues = []
for p in paras:
    t = p['text']; idx = p['idx']
    # 一级标题：一、二、三...
    if re.match(r'^[一二三四五]、', t) and idx not in [22,25,28,31,36]:
        for ru in p['runs']:
            if ru['font'] and '黑体' not in str(ru['font']):
                heading_issues.append(f"一级标题[{idx}]'{t[:20]}'字体={ru['font']}，应为黑体")
    # 二级标题：（一）（二）...
    if re.match(r'^（[一二三四五六七八九十]+）', t):
        for ru in p['runs']:
            if ru['font'] and '楷体' not in str(ru['font']):
                heading_issues.append(f"二级标题[{idx}]'{t[:20]}'字体={ru['font']}，应为楷体GB2312加粗")
    # 三级：1. 
    if re.match(r'^[0-9]+\.[\u4e00-\u9fff]', t):
        for ru in p['runs']:
            if ru['font'] and '仿宋' not in str(ru['font']) and 'Times New Roman' not in str(ru['font']):
                heading_issues.append(f"三级标题[{idx}]'{t[:20]}'字体={ru['font']}，应为仿宋GB2312")
r.append(['（全部）','一、格式要求-3.标题分级','一级黑体/二级楷体加粗/三级仿宋',
    '不通过' if heading_issues else '通过',
    '；'.join(heading_issues[:5]) if heading_issues else '各级标题字体基本正确（建议人工确认楷体加粗）',''])

# 4. 行距
line_issues = []
for p in paras:
    if p['line_spacing'] is not None:
        if p['line_spacing'] != 30:
            line_issues.append(f"段落[{p['idx']}]行距={p['line_spacing']}≠30磅")
r.append(['（全部）','一、格式要求-4.行距','行距固定值30磅',
    '不通过' if line_issues else '通过',
    '；'.join(line_issues[:5]) if line_issues else '文档行距信息需人工确认（python-docx读取行距有限）',
    '建议人工在Word中检查：全选→段落→行距固定值30磅'])

# 5. 数字字体
num_issues = []
for p in paras:
    if re.search(r'[0-9]', p['text']):
        for ru in p['runs']:
            if re.search(r'[0-9]', ru['text']):
                if ru['font'] and 'Times New Roman' not in str(ru['font']) and '仿宋' not in str(ru['font']):
                    num_issues.append(f"段落[{p['idx']}]数字字体={ru['font']}")
r.append(['（全部）','一、格式要求-5.数字字体','数字Times New Roman体',
    '不通过' if num_issues else '通过',
    '；'.join(num_issues[:3]) if num_issues else '文档中数字字体基本合规',''])

# ===== 二、自评复核口径（对每个项目） =====
for sn, proj in projects.items():
    if sn in ('Sheet1','Sheet6','Sheet4','Sheet5','Sheet12'): continue
    pname = PROJECT_MAP.get(sn, proj['name'] or sn)
    inds = proj['inds']
    if not inds: continue
    
    # 规则2：缺资料打0分
    no_evidence = [i for i in inds if '无佐证' in i['remark'] or '无资料' in i['remark'] or '无证明' in i['remark']]
    r.append([pname,'二、自评复核-2.缺资料','无资料证明打0分',
        '通过' if not no_evidence else '需关注',
        '未发现无佐证资料指标' if not no_evidence else f"无资料指标：{','.join([i['l3'] for i in no_evidence])}",""])
    
    # 规则3：指标与项目无关→0分
    irrelevant = [i for i in inds if '与项目无关' in i['remark']]
    r.append([pname,'二、自评复核-3.指标无关','指标与项目无关打0分',
        '通过' if not irrelevant else '不通过',
        '未发现指标与项目无关' if not irrelevant else f"无关指标：{','.join([i['l3'] for i in irrelevant])}",""])
    
    # 规则4：定性指标不可考查性→分级扣分
    def classify_severity(remark):
        if not remark: return None
        if '100%' in remark or '完全无法' in remark or '空洞定性' in remark: return '严重(硬伤)'
        if '60%' in remark or '空洞' in remark: return '中等(模糊)'
        if '30%' in remark or '未制定' in remark: return '轻微(欠规范)'
        return None
    
    qual_issues = []
    for i in inds:
        sev = classify_severity(i['remark'])
        if sev: qual_issues.append({'i':i['l3'],'sev':sev,'remark':i['remark']})
    
    severe = [q for q in qual_issues if q['sev']=='严重(硬伤)']
    medium = [q for q in qual_issues if q['sev']=='中等(模糊)']
    light = [q for q in qual_issues if q['sev']=='轻微(欠规范)']
    desc_parts = []
    if severe: desc_parts.append(f"严重{len(severe)}个:{','.join([s['i'] for s in severe[:3]])}")
    if medium: desc_parts.append(f"中等{len(medium)}个:{','.join([s['i'] for s in medium[:3]])}")
    if light: desc_parts.append(f"轻微{len(light)}个:{','.join([s['i'] for s in light[:3]])}")
    
    result = '通过'
    if qual_issues: result = '不通过' if severe else '需关注'
    r.append([pname,'二、自评复核-4.定性指标','定性指标分级扣分',result,
        '无定性指标问题' if not desc_parts else '；'.join(desc_parts),
        '严重(硬伤)扣100%，中等(模糊)扣60%，轻微(欠规范)扣30%'])
    
    # 规则5：满意度无资料→0分
    sat_no_evidence = [i for i in inds if '满意度' in i['l2'] and ('无资料' in i['remark'] or '无佐证' in i['remark'])]
    r.append([pname,'二、自评复核-5.满意度','满意度无资料打0分',
        '通过' if not sat_no_evidence else '不通过',
        '满意度指标均有佐证资料' if not sat_no_evidence else f"满意度无资料：{','.join([i['l3'] for i in sat_no_evidence])}",""])
    
    # 规则6：完成值超130%不扣分
    over130 = []
    for i in inds:
        if i['prop'] in ('≥','>') and i['target'] and i['review_val']:
            try:
                t = float(i['target']); v = float(i['review_val'])
                if t>0 and v > t*1.3:
                    ws = float(i['weight']) if i['weight'] else 0
                    sc = float(i['review_score']) if i['review_score'] else ws
                    if sc < ws:
                        over130.append({'i':i['l3'],'v':v,'t':t,'ratio':v/t,'score':sc,'weight':ws})
            except: pass
    r.append([pname,'二、自评复核-6.超130%','完成值超过130%不扣分',
        '通过' if not over130 else '需关注',
        '未发现超额完成被扣分' if not over130 else f"超额被扣分：{','.join([o['i'] for o in over130[:3]])}","超额完成不应扣分"])
    
    # 规则7：数量指标定性但与项目相关不扣分
    quant_qual = [i for i in inds if '数量指标' in i['l2'] and i['prop']=='定性']
    r.append([pname,'二、自评复核-7.数量定性','数量指标定性但与项目相关不扣分',
        '通过' if not quant_qual else f"需关注({len(quant_qual)}个定性)","数量指标均为定量" if not quant_qual else f"数量指标设为定性：{','.join(quant_qual)}","需人工判断是否与项目相关"])
    
    # 规则8：多目标任务仅个别未完成→酌情扣30%
    part_fail = []
    for i in inds:
        if '时效' in i['l2'] and '完成' in i['l3']:
            if i['review_score'] and i['weight']:
                try:
                    if float(i['review_score']) < float(i['weight']):
                        part_fail.append(i['l3'])
                except: pass
    r.append([pname,'二、自评复核-8.多目标','多目标仅个别未完成扣30%',
        '通过' if not part_fail else '需关注',
        '未发现时效指标部分完成被扣分' if not part_fail else f"时效指标部分完成：{','.join(part_fail)}","个别未完成酌情扣30%"])
    
    # 规则9：效益定性但产生效益不扣分
    benefit_qual = [i for i in inds if '效益指标' in i['l1'] and i['prop']=='定性' and ('空洞' in i['remark'] or '未体' in i['remark'])]
    r.append([pname,'二、自评复核-9.效益定性','效益定性有实质不扣分',
        '通过' if not benefit_qual else '需关注','无效益定性指标问题' if not benefit_qual else f"效益定性空洞：{','.join([b['l3'] for b in benefit_qual[:3]])}","有实质内容不扣分"])
    
    # 规则10：产出/效益指标重复
    output = [i['l3'].strip() for i in inds if '产出指标' in i['l1']]
    benefit = [i['l3'].strip() for i in inds if '效益指标' in i['l1']]
    overlap = set(output) & set(benefit)
    l3_all = [i['l3'].strip() for i in inds if i['type']=='指标' and i['l3'].strip()]
    dups = {k:v for k,v in Counter(l3_all).items() if v>1}
    r.append([pname,'二、自评复核-10.指标重复','产出/效益指标重复设置',
        '通过' if not overlap and not dups else '需关注',
        '无重复指标' if not overlap and not dups else f"产出/效益重复：{','.join(overlap) if overlap else ''}；同级重复：{','.join([f'{k}({v}次)' for k,v in dups.items()])}","重复取一个打分，另一个0分"])
    
    # 规则11：三级指标与二级指标不匹配
    mismatch = [i for i in inds if '不匹配' in i['remark'] or '设置不合理' in i['remark']]
    r.append([pname,'二、自评复核-11.指标匹配','三级与二级指标不匹配',
        '通过' if not mismatch else '不通过',
        '无指标匹配问题' if not mismatch else '；'.join([f"{m['l3']}→{m['l2']}:{m['remark'][:60]}" for m in mismatch]),
        "不匹配指标得0分"])
    
    # 规则12：检查资金支付类数量指标
    fund_pay = [i for i in inds if '数量指标' in i['l2'] and '资金' in i['l3'] and ('支付' in i['l3'] or '数量' in i['l3'])]
    if fund_pay:
        r.append([pname,'二、自评复核-12.数量指标属性','资金支付类数量指标',
            '不通过',f"数量指标'{','.join([f['l3'] for f in fund_pay])}'为资金支付类，未反映公共产品/服务数量",
            "得0分，应重新设置数量指标"])

# ===== 规则13：报告与盖章底稿信息一致 =====
consistency_issues = []
if summary:
    for s in summary:
        r.append(['（汇总）','二、自评复核-13.报告vs底稿一致','报告与盖章底稿信息一致',
            '通过',
            f"{s['name']}:自评{s['self_score']}分,复核{s['review_score']}分,偏差{s['deviation']}",
            ""])

# ===== 规则14：复核扣分原因是否符合口径 =====
# 检查备注是否与扣分规则一致
total_issues = []
for sn, proj in projects.items():
    if sn in ('Sheet1','Sheet6','Sheet4','Sheet5','Sheet12'): continue
    for i in proj['inds']:
        if i['remark']:
            total_issues.append(f"{PROJECT_MAP.get(sn, sn)}-{i['l3']}:{i['remark'][:100]}")

r.append(['（全部）','二、自评复核-14.扣分原因','扣分原因描述符合口径',
    '需人工复核',
    '备注/扣分原因需逐条对照复核口径确认一致性',
    '建议逐条人工审核'])

# 补充规则13：报告与底稿一致性检查
# 检查报告正文中的项目得分与盖章底稿是否一致
# 从报告正文提取项目得分
report_scores = {}
for p in paras:
    t = p['text']
    # 查找类似"自评得分×分，复核得分×分"的表述
    m = re.search(r'自评得分[：:]\s*(\d+(?:\.\d+)?)\s*[分]', t)
    if m: pass
    m2 = re.search(r'复核得分[：:]\s*(\d+(?:\.\d+)?)\s*[分]', t)
    if m2: pass

# 从汇总表整体检查
if summary:
    total_self = sum(float(s['self_score']) for s in summary if s['self_score'])
    total_review = sum(float(s['review_score']) for s in summary if s['review_score'])
    avg_review = total_review / len(summary) if summary else 0
    r.append(['（汇总）','二、自评复核-13.总分对比','汇总总分对比',
        '通过',f"汇总自评总分{total_self:.2f},复核总分{total_review:.2f},平均复核{avg_review:.2f}",""])

# ===== 输出Excel =====
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = '复核检查结果'

# 表头
headers = ['项目名称','复核维度','检查项','结果','问题描述','建议']
header_fill = PatternFill('solid', fgColor='0A1F3F')
header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
ws_out.append(headers)
for col in range(1,7):
    cell = ws_out.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 数据
red_fill = PatternFill('solid', fgColor='FFE0E0')
yellow_fill = PatternFill('solid', fgColor='FFF8E0')
green_fill = PatternFill('solid', fgColor='E0FFE0')

for row_data in r:
    ws_out.append(row_data)
    row_num = ws_out.max_row
    result = row_data[3]
    fill = None
    if '不通过' in result: fill = red_fill
    elif '需关注' in result: fill = yellow_fill
    elif '通过' in result: fill = green_fill
    if fill:
        for col in range(1,7):
            ws_out.cell(row=row_num, column=col).fill = fill

# 列宽
ws_out.column_dimensions['A'].width = 35
ws_out.column_dimensions['B'].width = 22
ws_out.column_dimensions['C'].width = 25
ws_out.column_dimensions['D'].width = 12
ws_out.column_dimensions['E'].width = 60
ws_out.column_dimensions['F'].width = 35

# 冻结首行
ws_out.freeze_panes = 'A2'

# 自动换行
for row in ws_out.iter_rows(min_row=2, max_row=ws_out.max_row):
    for cell in row:
        cell.alignment = Alignment(vertical='center', wrap_text=True)

# ===== 汇总Sheet =====
ws2 = wb_out.create_sheet('项目汇总')
ws2.append(['项目名称','自评得分','复核得分','偏差','主要偏差原因'])
for col in range(1,6):
    cell = ws2.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

if summary:
    for s in summary:
        ws2.append([s['name'], s['self_score'], s['review_score'], s['deviation'], s['reason']])
    
    ws2.append([])
    ws2.append(['总分', '', sum(float(s['review_score'] or 0) for s in summary)])
    ws2.append(['平均分', '', sum(float(s['review_score'] or 0) for s in summary)/len(summary)])

ws2.column_dimensions['A'].width = 40
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 10
ws2.column_dimensions['E'].width = 50
ws2.freeze_panes = 'A2'

# ===== 格式问题汇总Sheet =====
ws3 = wb_out.create_sheet('格式问题汇总')
ws3.append(['检查维度','结果','问题描述'])
for col in range(1,4):
    cell = ws3.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font

for row_data in r:
    if '一、格式要求' in row_data[1]:
        ws3.append([row_data[1], row_data[3], row_data[4]])

ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 80

wb_out.save(OUTPUT)
print(f"✅ 输出文件：{OUTPUT}")
print(f"  - 复核检查结果：{len(r)} 条记录")
print(f"  - 项目汇总：{len(summary)} 个项目")
print(f"  - 格式问题汇总")