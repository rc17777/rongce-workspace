#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
天府广场停车场出入记录全量分析脚本
使用openpyxl和csv模块（避免numpy问题）
"""
import os, csv, sys
from openpyxl import load_workbook
from collections import defaultdict, Counter
from datetime import datetime, timedelta
import json

# ============================================================
# STEP 0: Locate files
# ============================================================
base = r'C:\Users\scrccpa\Desktop\成都轨道资源资料'

# Find directory 35 (parking records)
def find_dir_by_name(base_dir, target_name):
    for root, dirs, files in os.walk(base_dir):
        for d in dirs:
            if d == target_name:
                return os.path.join(root, d)
    return None

parking_dir = find_dir_by_name(base, '35')
protocol_dir = find_dir_by_name(base, '34')
asset_dir1 = find_dir_by_name(base, '1')   # 资产台账
asset_dir3 = find_dir_by_name(base, '3')   # 资产台账 also
policy_dir = find_dir_by_name(base, '33')  # 政策文件
drawing_dir = find_dir_by_name(base, '37')  # 车位图

print("=== 文件定位结果 ===")
print(f"停车记录目录(35): {parking_dir}")
print(f"协议台账目录(34): {protocol_dir}")
print(f"资产台账目录(1):  {asset_dir1}")
print(f"资产台账目录(3):  {asset_dir3}")
print(f"政策文件目录(33): {policy_dir}")
print(f"车位图目录(37):  {drawing_dir}")

# ============================================================
# STEP 1: Read parking record data
# ============================================================
def read_csv_file(filepath):
    """Read CSV with encoding detection"""
    for enc in ['utf-8', 'gbk', 'gb2312', 'gb18030', 'utf-16']:
        try:
            with open(filepath, 'r', encoding=enc) as f:
                reader = csv.reader(f)
                header = next(reader)
                rows = list(reader)
            print(f"  CSV OK: enc={enc}, header={header}, rows={len(rows)}")
            return header, rows
        except Exception as e:
            continue
    return None, None

def read_xlsx_file(filepath):
    """Read xlsx file using openpyxl"""
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = list(rows[0])
            data = rows[1:]
            print(f"  XLSX OK: header={header}, rows={len(data)}")
            return header, data
        return None, None
    except Exception as e:
        print(f"  XLSX ERROR: {e}")
        return None, None

def read_xls_file(filepath):
    """Try reading .xls file (older format)"""
    try:
        # Try openpyxl first (may work if it's really xlsx)
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = list(rows[0])
            data = rows[1:]
            print(f"  XLS OK (as xlsx): header={header}, rows={len(data)}")
            return header, data
    except:
        pass
    try:
        # Try xlrd
        import xlrd
        wb = xlrd.open_workbook(filepath)
        ws = wb.sheet_by_index(0)
        header = [ws.cell_value(0, c) for c in range(ws.ncols)]
        data = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(1, ws.nrows)]
        print(f"  XLS OK (xlrd): header={header}, rows={len(data)}")
        return header, data
    except Exception as e:
        print(f"  XLS ERROR: {e}")
        return None, None

all_parking_data = {}  # {period: (header, rows)}

print("\n\n=== 步骤1: 读取停车记录文件 ===")
for fname in sorted(os.listdir(parking_dir)):
    fpath = os.path.join(parking_dir, fname)
    ext = os.path.splitext(fname)[1].lower()
    
    # Map file to period name
    period = fname  # Use filename as key
    
    print(f"\n读取: {repr(fname)}")
    
    if ext == '.csv':
        header, rows = read_csv_file(fpath)
    elif ext == '.xlsx':
        header, rows = read_xlsx_file(fpath)
    elif ext == '.xls':
        header, rows = read_xls_file(fpath)
    else:
        print(f"  跳过: 不支持的文件类型 {ext}")
        continue
    
    if header and rows:
        all_parking_data[period] = (header, rows)

print(f"\n\n成功读取 {len(all_parking_data)} 个数据文件")

# ============================================================
# STEP 2: Standardize column names
# ============================================================
print("\n\n=== 步骤2: 检查各文件列名并标准化 ===")

for period, (header, data) in all_parking_data.items():
    print(f"\n{repr(period)}:")
    for i, h in enumerate(header):
        print(f"  [{i}] {repr(h)}")

# ============================================================
# STEP 3: Read protocol car ledger (协议车位台账)
# ============================================================
print("\n\n=== 步骤3: 读取协议车位台账 ===")
protocol_data = None
if protocol_dir:
    for fname in os.listdir(protocol_dir):
        fpath = os.path.join(protocol_dir, fname)
        print(f"\n读取协议台账: {repr(fname)}")
        header, rows = read_xls_file(fpath)
        if header and rows:
            protocol_data = (header, rows)
            for i, h in enumerate(header):
                print(f"  [{i}] {repr(h)}")
            # Print first 5 rows
            for r in rows[:5]:
                print(f"  {r}")
            break

# ============================================================
# STEP 4: Read asset ledger (资产台账)
# ============================================================
print("\n\n=== 步骤4: 读取资产台账 ===")
for asset_dir in [asset_dir1, asset_dir3]:
    if asset_dir and os.path.exists(asset_dir):
        for fname in os.listdir(asset_dir):
            fpath = os.path.join(asset_dir, fname)
            ext = os.path.splitext(fname)[1].lower()
            if ext in ['.xlsx', '.xls']:
                print(f"\n读取资产台账: {repr(fname)}")
                header, rows = read_xlsx_file(fpath) if ext == '.xlsx' else read_xls_file(fpath)
                if header:
                    for i, h in enumerate(header):
                        print(f"  [{i}] {repr(h)}")
                    print(f"  总行数: {len(rows)}")
                    # Print first 3 rows
                    for r in rows[:3]:
                        print(f"  {r}")

# ============================================================
# Main parking data analysis
# ============================================================
print("\n\n" + "="*80)
print("开始停车数据分析...")
print("="*80)

# Now let's deeply analyze each period
for period, (header, rows) in sorted(all_parking_data.items()):
    print(f"\n{'='*60}")
    print(f"分析期间: {repr(period)}")
    print(f"列名: {header}")
    print(f"记录数: {len(rows)}")
    
    # Analyze columns for content
    col_types = {}
    for col_idx in range(len(header)):
        col_name = header[col_idx]
        vals = []
        for r in rows[:100]:
            if col_idx < len(r):
                v = r[col_idx]
                if v is not None and str(v).strip():
                    vals.append(str(v).strip())
        if vals:
            # Sample some values
            sample = list(set(vals))[:5]
            col_types[col_idx] = {
                'name': col_name,
                'sample': sample,
                'non_null': len(vals),
                'type_guess': 'text' if all(len(v) > 0 for v in sample) else 'mixed'
            }
    
    # Detect car plate columns, time columns, etc.
    print(f"\n  列内容分析:")
    for ci, info in col_types.items():
        print(f"    [{ci}] {repr(info['name'])}: {info['sample']}")

print("\n\n=== 分析完成（初步）===")

