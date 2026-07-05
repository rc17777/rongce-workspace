#!/usr/bin/env python3
"""
审计发现智能处理 — 基于审计大模型框架的知识引擎

参考: 柳絮/李欣潼《审计大模型》场景2+4
功能: 输入审计发现描述 → 输出: 问题定性 + 法规依据 + 整改措施 + 佐证材料清单
"""
import sys, io, argparse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

RED = PatternFill(patternType='solid', fgColor='FFD7D7')
HEADER = PatternFill(patternType='solid', fgColor='1A3A6E')
H = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
N = Font(name='Microsoft YaHei', size=10)
TH = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
C = Alignment(horizontal='center', vertical='center', wrap_text=True)
L = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ── 审计知识库 (基于已归档的审计方法论) ──
AUDIT_KNOWLEDGE = {
    # 采购/招投标领域
    '围标串标': {
        '定性': '投标人之间横向串通，采取轮流坐庄、价格联盟或陪伴补偿方式操纵中标结果',
        '法规': [
            '《中华人民共和国招标投标法》第三十二条（投标人不得相互串通投标）',
            '《中华人民共和国招标投标法实施条例》第三十九条（属于投标人相互串通投标的情形）',
            '《中华人民共和国招标投标法实施条例》第四十条（视为投标人相互串通投标的情形）',
        ],
        '整改': [
            '对涉事企业依法处以中标项目金额5‰-10‰罚款',
            '将涉事企业列入不良行为记录名单，1-3年内禁止参加依法必须招标项目',
            '完善电子招投标系统异常行为自动监测功能',
            '建立投标人关联关系数据库，事前拦截关联企业同时投标',
        ],
        '佐证材料': [
            '各投标人投标文件全文',
            '投标保证金缴纳记录（银行回单/电子保函）',
            '投标人联系方式/授权委托书',
            '投标IP/MAC/硬盘特征码记录',
            '工商关联关系查询结果（天眼查/企查查）',
            '评审打分汇总表',
        ],
    },
    '招标文件设限': {
        '定性': '招标文件设置不合理条件限制或排斥潜在投标人',
        '法规': [
            '《中华人民共和国招标投标法》第十八条（不得以不合理条件限制或排斥潜在投标人）',
            '《中华人民共和国招标投标法实施条例》第三十二条（属于不合理条件限制的情形）',
            '《政府采购法实施条例》第二十条（属于差别待遇或歧视待遇的情形）',
        ],
        '整改': [
            '删除招标文件中涉及特定品牌/地域/规模等不合理条件',
            '重新发布招标公告，延长投标截止时间',
            '对招标文件编制人员进行法规培训',
        ],
        '佐证材料': [
            '招标文件全文',
            '招标文件编制人及审批人信息',
            '潜在投标人质疑/投诉记录',
        ],
    },
    '节资率异常': {
        '定性': '招标节资率明显低于同类项目平均水平，竞争不充分',
        '法规': [
            '《中华人民共和国招标投标法》第五条（招标投标活动应遵循公开、公平、公正原则）',
            '《政府投资条例》第十二条（政府投资项目应遵循科学决策、规范管理原则）',
        ],
        '整改': [
            '扩大招标公告发布范围，增加潜在投标人数量',
            '引入省外/行业外供应商参与竞争',
            '建立节资率异常预警机制，低于阈值自动触发复核',
        ],
        '佐证材料': [
            '招标控制价编制依据',
            '同类项目节资率对比分析表',
            '投标人来源分析（本地/外地/行业）',
        ],
    },
    # 财务审计领域
    '费用报销异常': {
        '定性': '费用报销存在审批不严、凭证不全、超标准报销等问题',
        '法规': [
            '《会计法》第十四条（会计凭证的要求）',
            '《企业内部控制应用指引第6号——资金活动》',
            '国务院《关于改进工作作风、密切联系群众的八项规定》',
        ],
        '整改': [
            '追回违规报销款项，追究相关责任人责任',
            '完善费用报销审批权限和流程',
            '建立费用报销预警机制（单笔上限/单人累计/敏感类别）',
        ],
        '佐证材料': [
            '费用报销单及附件（发票/审批单/出差审批）',
            '报销人银行卡收款记录',
            '单位差旅费/招待费管理制度',
        ],
    },
    '资产账实不符': {
        '定性': '固定资产账面记录与实际使用状态不符，存在盘亏/盘盈/闲置等',
        '法规': [
            '《会计法》第十七条（定期进行财产清查）',
            '《企业会计准则第4号——固定资产》',
            '《行政事业单位国有资产管理办法》',
        ],
        '整改': [
            '开展全面资产清查，调整账务至与实物一致',
            '对盘亏资产查明原因并追究责任',
            '建立资产动态管理台账，每年至少清查一次',
        ],
        '佐证材料': [
            '资产卡片/台账/标签',
            '资产盘点表（签字确认）',
            '盘盈/盘亏审批文件',
            '资产实物照片',
        ],
    },
    # 工程领域
    '工程款超付': {
        '定性': '工程款支付超过合同约定或超出实际完成进度',
        '法规': [
            '《中华人民共和国民法典》第七百八十八条（建设工程合同）',
            '《建设工程价款结算暂行办法》',
            '《保障农民工工资支付条例》',
        ],
        '整改': [
            '追缴超付工程款项',
            '完善工程款支付审批流程，实行计量支付',
            '追究超付审批相关人员责任',
        ],
        '佐证材料': [
            '工程合同及补充协议',
            '工程计量/计价文件',
            '工程款支付审批单',
            '银行付款凭证',
            '工程进度确认表（监理/甲方签字）',
        ],
    },
    '工程签证变更不规范': {
        '定性': '工程设计变更/现场签证未经规范审批，变更依据不充分',
        '法规': [
            '《建设工程质量管理条例》',
            '《建设工程监理规范》（GB/T50319）',
        ],
        '整改': [
            '补全签证审批手续',
            '对不合理签证进行重新核定工程造价',
            '建立工程变更分级审批制度',
        ],
        '佐证材料': [
            '签证变更单及审批记录',
            '设计变更通知单',
            '变更前后对比照片/图纸',
            '相关会议纪要',
        ],
    },
    # 资金/绩效领域
    '专项资金挪用': {
        '定性': '专项资金未按规定用途使用，存在截留/挪用/虚列支出',
        '法规': [
            '《预算法》第五十七条（专项资金专款专用）',
            '《财政违法行为处罚处分条例》第六条',
        ],
        '整改': [
            '追回被挪用资金',
            '完善专项资金专账核算',
            '追究相关责任人行政/刑事责任',
        ],
        '佐证材料': [
            '专项资金预算批复文件',
            '资金拨付凭证',
            '支付凭证及附件',
            '项目验收/结算报告',
        ],
    },
    '绩效目标未完成': {
        '定性': '项目设定的绩效目标未达到，存在管理不善或投入不足',
        '法规': [
            '《中共中央 国务院关于全面实施预算绩效管理的意见》',
            '《项目支出绩效评价管理办法》',
        ],
        '整改': [
            '分析原因，制定绩效改进计划',
            '调整后续年度预算安排',
            '对责任单位进行绩效问责',
        ],
        '佐证材料': [
            '项目绩效目标申报表',
            '绩效自评报告',
            '项目产出/效益证明材料',
        ],
    },
}


def process_finding(finding_desc: str, category: str = None):
    """处理审计发现，匹配知识库 → 结构化输出"""

    # Auto-detect category if not specified
    if not category:
        for kw, info in AUDIT_KNOWLEDGE.items():
            if kw in finding_desc:
                category = kw
                break

    if category and category in AUDIT_KNOWLEDGE:
        matched = AUDIT_KNOWLEDGE[category]
        print(f"\n📌 匹配知识库: {category}")
    else:
        print(f"\n⚠️ 未匹配到知识库类别，请手动录入")
        matched = None

    return {
        'finding': finding_desc,
        'category': category,
        'knowledge': matched,
    }


def generate_workpaper(findings: list, output: str):
    """生成审计底稿格式的Excel"""
    wb = Workbook()

    # Sheet 1: 审计发现处理表
    ws = wb.active
    ws.title = '审计发现处理表'

    ws.merge_cells('A1:F1')
    ws['A1'] = '审计发现智能处理底稿'
    ws['A1'].font = Font(name='Microsoft YaHei', size=16, bold=True, color='1A3A6E')

    headers = ['序号', '审计发现描述', '问题定性', '适用法规', '整改建议', '佐证材料清单']
    for c, h in enumerate(headers, 1):
        cl = ws.cell(row=3, column=c, value=h)
        cl.font = H; cl.fill = HEADER; cl.alignment = C; cl.border = TH

    for i, finding in enumerate(findings):
        r = i + 4
        kw = finding.get('knowledge')

        ws.cell(row=r, column=1, value=i+1).font = N
        ws.cell(row=r, column=1).alignment = C; ws.cell(row=r, column=1).border = TH

        ws.cell(row=r, column=2, value=finding['finding'])
        ws.cell(row=r, column=2).font = N; ws.cell(row=r, column=2).alignment = L
        ws.cell(row=r, column=2).border = TH

        if kw:
            ws.cell(row=r, column=3, value=kw['定性'])
            ws.cell(row=r, column=4, value='\n'.join(kw['法规']))
            ws.cell(row=r, column=5, value='\n'.join(kw['整改']))
            ws.cell(row=r, column=6, value='\n'.join(kw['佐证材料']))
        else:
            for c in [3,4,5,6]:
                ws.cell(row=r, column=c, value='[需人工填写]')

        for c in range(1, 7):
            ws.cell(row=r, column=c).font = N
            ws.cell(row=r, column=c).alignment = L
            ws.cell(row=r, column=c).border = TH

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 35

    wb.save(output)
    print(f"\n✅ 底稿: {output}")
    return output


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='审计发现智能处理 — 问题定性+法规+整改+佐证材料',
        epilog='''
示例:
  python audit_finding_processor.py "XX项目存在围标串标行为"
  python audit_finding_processor.py --file 发现清单.xlsx --col 审计发现 --o 底稿.xlsx
        '''
    )
    parser.add_argument('finding', nargs='?', help='审计发现描述文本')
    parser.add_argument('--file', '-f', help='批量处理: 审计发现清单xlsx')
    parser.add_argument('--col', default='审计发现', help='发现描述列名 (默认: 审计发现)')
    parser.add_argument('--o', '--output', dest='output', default='审计发现处理底稿.xlsx')
    args = parser.parse_args()

    findings = []

    if args.file:
        import pandas as pd
        df = pd.read_excel(args.file)
        for _, row in df.iterrows():
            desc = str(row.get(args.col, ''))
            if desc and desc != 'nan':
                result = process_finding(desc)
                findings.append(result)
    elif args.finding:
        result = process_finding(args.finding)
        findings.append(result)
    else:
        # Interactive mode
        print("审计发现智能处理 — 输入审计发现，自动匹配知识库")
        print("(输入空行完成)\n")

        # Show available categories
        print("已加载知识库类别:")
        for i, kw in enumerate(AUDIT_KNOWLEDGE.keys(), 1):
            print(f"  {i}. {kw}")
        print()

        while True:
            desc = input("审计发现 > ").strip()
            if not desc:
                break
            result = process_finding(desc)
            if result['knowledge']:
                kw = result['knowledge']
                print(f"  定性: {kw['定性']}")
                print(f"  法规: {kw['法规'][0]}...")
                print(f"  整改: {len(kw['整改'])} 项")
                print(f"  佐证: {len(kw['佐证材料'])} 项")
            findings.append(result)

    if findings:
        generate_workpaper(findings, args.output)

    # Print summary
    if findings:
        matched = sum(1 for f in findings if f['knowledge'])
        print(f"\n📊 处理: {len(findings)} 条, 知识库匹配: {matched}, 需人工: {len(findings)-matched}")
