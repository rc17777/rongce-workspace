import openpyxl
import json

path = r"C:\Users\scrccpa\Desktop\（吴博医保基金）融策审计过程记录系统=项目经理版(2).xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

print(f"Sheet names: {wb.sheetnames}")

for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n{'='*80}")
    print(f"Sheet: {name}  (rows={ws.max_row}, cols={ws.max_column})")
    print(f"{'='*80}")
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200), values_only=False):
        vals = []
        for cell in row:
            v = cell.value
            if v is not None:
                vals.append(str(v))
            else:
                vals.append("")
        line = "\t".join(vals)
        if line.strip():
            print(line)
    print(f"... (total rows: {ws.max_row})")
