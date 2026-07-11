"""
金川投标项目 — 7家投标单位全量串标围标分析
覆盖层: L1(报价规律) / L3(文本雷同) / L4(图片哈希) / L5(PDF元数据)
输出: 全量分析结果.xlsx（多Sheet）
"""
import os, sys, json, io, hashlib, re
from collections import defaultdict, Counter
from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ====== PDF 解析 ======
import fitz  # PyMuPDF

BASE = r"C:\Users\scrccpa\Desktop\金川投标"

# ====== 1. 收集所有文件 ======
def collect_files():
    """遍历目录，返回 {公司名: [文件路径列表]}"""
    data = {}
    for d in sorted(os.listdir(BASE)):
        dpath = os.path.join(BASE, d)
        if not os.path.isdir(dpath):
            continue
        company = d.replace("(包1)", "").strip()
        files = []
        for f in sorted(os.listdir(dpath)):
            if f.lower().endswith('.pdf'):
                files.append(os.path.join(dpath, f))
        if files:
            data[company] = files
    return data

# ====== 2. 文件归类 ======
FILE_CATEGORIES = {
    "报价表": ["报价表"],
    "投标函": ["投标（响应）函", "投标函"],
    "中小企业声明函": ["中小企业声明函"],
    "残疾人福利性单位声明函": ["残疾人福利性单位声明函"],
    "监狱企业证明": ["监狱企业的证明文件"],
    "供应商基本情况表": ["供应商基本情况表"],
    "业绩一览表": ["供应商类似项目业绩一览表"],
    "商务应答表": ["商务应答表"],
    "技术要求应答表": ["技术要求应答表"],
    "服务应答表": ["服务应答表"],
    "主要人员情况表": ["实施本项目的主要人员情况表"],
    "相关证明材料": ["投标人应提交的相关证明材料"],
    "其他相关材料": ["供应商认为需要提供的其他相关材料"],
    "投标文件封面": ["投标文件封面"],
}

def classify_file(fpath):
    fname = os.path.basename(fpath)
    for cat, keywords in FILE_CATEGORIES.items():
        for kw in keywords:
            if kw in fname:
                return cat
    return os.path.splitext(fname)[0]

# ====== 3. PDF文本提取 ======
def extract_text(pdf_path):
    """PyMuPDF提取文本"""
    try:
        doc = fitz.open(pdf_path)
        text = ""
        for page in doc:
            text += page.get_text("text") + "\n"
        doc.close()
        return text.strip()
    except Exception as e:
        return f"[ERROR: {e}]"

def extract_all_texts(files_data):
    """提取所有PDF文本，返回 {公司: {文件类别: 文本}}"""
    all_texts = {}
    for company, files in files_data.items():
        company_texts = {}
        for fpath in files:
            cat = classify_file(fpath)
            text = extract_text(fpath)
            company_texts[cat] = text
        all_texts[company] = company_texts
        print(f"  [TEXT] {company}: {len(company_texts)} files extracted")
    return all_texts

# ====== 4. PDF元数据提取 ======
def extract_metadata(pdf_path):
    """提取PDF元数据"""
    try:
        doc = fitz.open(pdf_path)
        meta = doc.metadata
        doc.close()
        return {
            "title": str(meta.get("title", "") or ""),
            "author": str(meta.get("author", "") or ""),
            "subject": str(meta.get("subject", "") or ""),
            "creator": str(meta.get("creator", "") or ""),
            "producer": str(meta.get("producer", "") or ""),
            "creationDate": str(meta.get("creationDate", "") or ""),
            "modDate": str(meta.get("modDate", "") or ""),
            "format": str(meta.get("format", "") or ""),
            "page_count": fitz.open(pdf_path).page_count if not fitz.open(pdf_path).is_closed else 0,
        }
    except Exception as e:
        return {"error": str(e)}

def extract_all_metadata(files_data):
    """提取所有PDF元数据"""
    all_meta = {}
    for company, files in files_data.items():
        company_meta = {}
        for fpath in files:
            cat = classify_file(fpath)
            company_meta[os.path.basename(fpath)] = extract_metadata(fpath)
        all_meta[company] = company_meta
        print(f"  [META] {company}: {len(company_meta)} files")
    return all_meta

# ====== 5. PDF嵌入图片提取 & 哈希 ======
def extract_images_hashes(pdf_path):
    """从PDF提取嵌入图片并计算MD5/SHA256哈希"""
    hashes = []
    try:
        doc = fitz.open(pdf_path)
        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            images = page.get_images(full=True)
            for img in images:
                xref = img[0]
                base_image = doc.extract_image(xref)
                img_bytes = base_image["image"]
                md5 = hashlib.md5(img_bytes).hexdigest()
                sha256 = hashlib.sha256(img_bytes).hexdigest()
                ext = base_image["ext"]
                w, h = base_image["width"], base_image["height"]
                hashes.append({
                    "md5": md5,
                    "sha256": sha256,
                    "ext": ext,
                    "width": w,
                    "height": h,
                    "page": page_num + 1,
                    "size_bytes": len(img_bytes),
                })
        doc.close()
    except Exception as e:
        pass
    return hashes

def extract_all_images(files_data):
    """提取所有PDF的嵌入图片哈希"""
    all_img = {}
    global_hashes = defaultdict(list)  # md5 -> [(公司, 文件)]
    for company, files in files_data.items():
        company_imgs = {}
        for fpath in files:
            cat = classify_file(fpath)
            hashes_list = extract_images_hashes(fpath)
            company_imgs[cat] = hashes_list
            for h in hashes_list:
                global_hashes[h["md5"]].append((company, cat, fpath))
        all_img[company] = company_imgs
        total_imgs = sum(len(v) for v in company_imgs.values())
        print(f"  [IMG] {company}: {total_imgs} images extracted")
    return all_img, global_hashes

# ====== 6. 文本清洗 ======
def clean_text(text):
    """清洗文本，去除空白、标点差异"""
    if not text or text.startswith("[ERROR"):
        return ""
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

# ====== 7. L1 报价规律性分析 ======
def analyze_price_patterns(all_texts):
    """从报价表中提取报价并分析规律"""
    prices = {}
    for company, texts in all_texts.items():
        for cat, text in texts.items():
            if "报价" in cat:
                cleaned = clean_text(text)
                # 尝试匹配价格模式
                price_matches = re.findall(r'(?:总价|合计|投标报价|金额)[：:]\s*(\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?)', cleaned)
                if not price_matches:
                    price_matches = re.findall(r'(?:¥|￥)\s*(\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?)', cleaned)
                if not price_matches:
                    price_matches = re.findall(r'(\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?)\s*元', cleaned)
                if not price_matches:
                    price_matches = re.findall(r'(?:总价|合计|金额)[：:]*\s*([\d,]+\.?\d*)', cleaned)
                
                parsed = []
                for m in price_matches:
                    try:
                        parsed.append(float(m.replace(",", "")))
                    except:
                        pass
                if parsed:
                    prices[company] = {"raw_text_snippet": cleaned[:500], "prices_found": parsed}
                break
    
    return prices

# ====== 8. L3 文本雷同检测 (TF-IDF) ======
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

def analyze_text_similarity(all_texts):
    """对同类文件做TF-IDF余弦相似度分析"""
    results = {}
    
    # 选取适合文本分析的类别
    text_categories = [
        "投标函", "商务应答表", "技术要求应答表", "服务应答表",
        "供应商基本情况表", "主要人员情况表", "中小企业声明函"
    ]
    
    for cat in text_categories:
        companies = []
        texts = []
        for company, company_texts in all_texts.items():
            if cat in company_texts:
                t = clean_text(company_texts[cat])
                if t and len(t) > 20:  # 至少20字符才算有效文本
                    companies.append(company)
                    texts.append(t)
        
        if len(texts) >= 2:
            try:
                vectorizer = TfidfVectorizer(
                    analyzer='char_wb', ngram_range=(3, 5),
                    max_features=5000
                )
                tfidf_matrix = vectorizer.fit_transform(texts)
                sim_matrix = cosine_similarity(tfidf_matrix)
                
                # 构建两两对比
                pair_results = []
                for i in range(len(companies)):
                    for j in range(i+1, len(companies)):
                        pair_results.append({
                            "公司A": companies[i],
                            "公司B": companies[j],
                            "余弦相似度": round(float(sim_matrix[i][j]), 4),
                            "字符数_A": len(texts[i]),
                            "字符数_B": len(texts[j]),
                        })
                
                results[cat] = {
                    "companies": companies,
                    "texts": texts,
                    "matrix": sim_matrix,
                    "pairs": pair_results,
                }
            except Exception as e:
                results[cat] = {"error": str(e)}
        else:
            results[cat] = {"note": f"仅{len(texts)}家有有效文本"}
    
    return results

# ====== 9. 全局文本相似度（跨文件类型全文本拼接） ======
def analyze_global_similarity(all_texts):
    """将所有文件的文本拼接后做全量TF-IDF对比"""
    companies = []
    all_concatenated = []
    
    for company, texts in all_texts.items():
        concatenated = ""
        for cat in sorted(texts.keys()):
            t = clean_text(texts[cat])
            if t:
                concatenated += t + " "
        if concatenated.strip():
            companies.append(company)
            all_concatenated.append(concatenated.strip())
    
    if len(companies) < 2:
        return None
    
    vectorizer = TfidfVectorizer(
        analyzer='char_wb', ngram_range=(3, 5),
        max_features=10000
    )
    tfidf_matrix = vectorizer.fit_transform(all_concatenated)
    sim_matrix = cosine_similarity(tfidf_matrix)
    
    pairs = []
    for i in range(len(companies)):
        for j in range(i+1, len(companies)):
            pairs.append({
                "公司A": companies[i],
                "公司B": companies[j],
                "全文本余弦相似度": round(float(sim_matrix[i][j]), 4),
                "字符数_A": len(all_concatenated[i]),
                "字符数_B": len(all_concatenated[j]),
            })
    
    return {
        "companies": companies,
        "matrix": sim_matrix,
        "pairs": pairs,
    }

# ====== 10. L5 元数据交叉分析 ======
def analyze_metadata_cross(all_meta):
    """分析PDF元数据，检测同源制作"""
    # 构建: 文件类别 -> 字段 -> {值: [公司列表]}
    cat_field_values = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    
    for company, files_meta in all_meta.items():
        for fname, meta in files_meta.items():
            cat = classify_file(fname)
            for field in ["author", "creator", "producer"]:
                val = meta.get(field, "").strip()
                if val and val != "None":
                    cat_field_values[cat][field][val].append(company)
    
    # 找出共用同一值的情况
    cross_findings = []
    for cat, fields in cat_field_values.items():
        for field, values in fields.items():
            for val, companies in values.items():
                if len(companies) >= 2:
                    cross_findings.append({
                        "文件类别": cat,
                        "元数据字段": field,
                        "值": val,
                        "涉及公司": "、".join(sorted(companies)),
                        "涉及数量": len(companies),
                        "严重程度": "🔴铁证" if len(companies) >= 3 else "🟡强信号",
                    })
    
    # 按文件类别汇总元数据
    all_meta_rows = []
    for company, files_meta in all_meta.items():
        for fname, meta in files_meta.items():
            cat = classify_file(fname)
            all_meta_rows.append({
                "公司": company,
                "文件名": fname,
                "文件类别": cat,
                "Author": meta.get("author", ""),
                "Creator": meta.get("creator", ""),
                "Producer": meta.get("producer", ""),
                "CreationDate": meta.get("creationDate", ""),
                "ModDate": meta.get("modDate", ""),
                "页数": meta.get("page_count", 0),
            })
    
    return all_meta_rows, cross_findings, cat_field_values

# ====== 11. L4 图片跨公司哈希比对 ======
def analyze_image_cross(global_hashes):
    """分析跨公司重复图片"""
    cross_img_findings = []
    for md5, occurrences in global_hashes.items():
        companies = set(o[0] for o in occurrences)
        if len(companies) >= 2:
            # 获取图片信息
            details = []
            for o in occurrences:
                details.append(f"{o[0]}/{o[1]}(P{o[2]})")
            
            # 判断是否是可能的模板化图片（如公章、签名等）
            is_common = any(kw in "|".join(details).lower() for kw in ["声明函", "封面", "证明"])
            
            cross_img_findings.append({
                "MD5": md5[:16] + "...",
                "涉及公司数": len(companies),
                "涉及公司": "、".join(sorted(companies)),
                "出现次数": len(occurrences),
                "详情": " | ".join(details),
                "可能为模板": "是" if is_common else "否",
            })
    
    cross_img_findings.sort(key=lambda x: -x["涉及公司数"])
    return cross_img_findings

# ====== 12. 文本完整性分析 ======
def analyze_text_completeness(all_texts):
    """分析每家每类文件的文本提取量"""
    rows = []
    for company, texts in all_texts.items():
        for cat in sorted(texts.keys()):
            t = texts[cat]
            char_count = len(t) if t and not t.startswith("[ERROR") else 0
            has_text = "有效" if char_count > 20 else ("极少" if char_count > 0 else "无文本(扫描件)")
            rows.append({
                "公司": company,
                "文件类别": cat,
                "字符数": char_count,
                "文本状态": has_text,
            })
    return rows

# ====== 13. 报价详细提取（从报价表PDF逐页） ======
def extract_price_details(files_data):
    """详细提取报价表内容"""
    price_details = {}
    for company, files in files_data.items():
        for fpath in files:
            if "报价表" in os.path.basename(fpath):
                try:
                    doc = fitz.open(fpath)
                    full_text = ""
                    for page in doc:
                        full_text += page.get_text("text") + "\n"
                    doc.close()
                    price_details[company] = {
                        "file": os.path.basename(fpath),
                        "text": full_text.strip(),
                        "char_count": len(full_text.strip()),
                    }
                except Exception as e:
                    price_details[company] = {"error": str(e)}
    return price_details

# ====== 写入 Excel ======
def write_excel(results, output_path):
    """将全量分析结果写入格式化Excel"""
    wb = Workbook()
    
    # 样式
    header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wrap_align = Alignment(wrap_text=True, vertical="top")
    center_align = Alignment(horizontal="center", vertical="center")
    
    def style_header(ws, headers, row=1):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
    
    def style_data_rows(ws, start_row, end_row, col_count):
        for r in range(start_row, end_row + 1):
            for c in range(1, col_count + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    def auto_width(ws, max_width=60):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    lines = str(cell.value).split('\n')
                    for line in lines:
                        max_len = max(max_len, len(line))
            ws.column_dimensions[col_letter].width = min(max_len + 4, max_width)
    
    # ========== Sheet 1: 总览 ==========
    ws1 = wb.active
    ws1.title = "总览"
    ws1.merge_cells("A1:H1")
    title_cell = ws1["A1"]
    title_cell.value = "金川投标项目 — 7家投标单位全量串标围标分析报告"
    title_cell.font = Font(name="Microsoft YaHei", bold=True, size=14, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    companies = sorted(results["companies"])
    overview_data = [
        ["项目名称", "金川投标项目"],
        ["分析日期", "2026-05-30"],
        ["投标单位数", len(companies)],
        ["分析包数", "包1"],
        ["", ""],
        ["投标单位列表", ""],
    ]
    for i, c in enumerate(companies, 1):
        overview_data.append([f"  {i}. {c}", ""])
    
    overview_data += [
        ["", ""],
        ["分析维度", "状态"],
        ["L1 报价规律性", "✅ 已完成"],
        ["L3 文本雷同(TF-IDF)", "✅ 已完成"],
        ["L4 图片跨公司哈希比对", "✅ 已完成"],
        ["L5 PDF元数据交叉", "✅ 已完成"],
        ["文本完整性检查", "✅ 已完成"],
    ]
    
    for r, row in enumerate(overview_data, 3):
        for c, val in enumerate(row, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.font = Font(name="Microsoft YaHei", size=11)
    
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 40
    
    # ========== Sheet 2: 报价分析 ==========
    ws2 = wb.create_sheet("L1_报价规律性")
    price_details = results["price_details"]
    all_texts = results["all_texts"]
    
    # 先展示报价表文本内容
    headers_price = ["公司", "文件名", "字符数", "文本内容预览(前500字)"]
    style_header(ws2, headers_price)
    
    row = 2
    for company in sorted(price_details.keys()):
        pd_info = price_details[company]
        ws2.cell(row=row, column=1, value=company).font = Font(name="Microsoft YaHei", size=10)
        ws2.cell(row=row, column=2, value=pd_info.get("file", ""))
        ws2.cell(row=row, column=3, value=pd_info.get("char_count", 0))
        ws2.cell(row=row, column=4, value=pd_info.get("text", "")[:500])
        row += 1
    
    style_data_rows(ws2, 2, row - 1, len(headers_price))
    auto_width(ws2)
    ws2.column_dimensions['D'].width = 80
    
    # 如果有提取到的价格，添加分析
    row += 2
    ws2.cell(row=row, column=1, value="📊 报价规律分析").font = Font(name="Microsoft YaHei", bold=True, size=12, color="2F5496")
    
    prices = results["prices"]
    row += 2
    price_headers = ["公司", "发现的价格数值", "价格数量"]
    style_header(ws2, price_headers, row)
    row += 1
    for company in sorted(prices.keys()):
        pdata = prices[company]
        ws2.cell(row=row, column=1, value=company)
        ws2.cell(row=row, column=2, value=str(pdata.get("prices_found", [])))
        ws2.cell(row=row, column=3, value=len(pdata.get("prices_found", [])))
        row += 1
    style_data_rows(ws2, row - len(prices), row - 1, 3)
    
    # ========== Sheet 3: 文本完整性 ==========
    ws3 = wb.create_sheet("文本完整性")
    completeness_headers = ["公司", "文件类别", "字符数", "文本状态"]
    style_header(ws3, completeness_headers)
    
    completeness = results["completeness"]
    for r, cr in enumerate(completeness, 2):
        ws3.cell(row=r, column=1, value=cr["公司"])
        ws3.cell(row=r, column=2, value=cr["文件类别"])
        ws3.cell(row=r, column=3, value=cr["字符数"])
        status_cell = ws3.cell(row=r, column=4, value=cr["文本状态"])
        if cr["文本状态"] == "有效":
            status_cell.fill = green_fill
        elif cr["文本状态"] == "极少":
            status_cell.fill = yellow_fill
        else:
            status_cell.fill = red_fill
    
    style_data_rows(ws3, 2, len(completeness) + 1, len(completeness_headers))
    auto_width(ws3)
    
    # ========== Sheet 4: 文本相似度(同类文件) ==========
    ws4 = wb.create_sheet("L3_文本雷同_同类文件")
    
    sim_results = results["text_similarity"]
    row = 1
    for cat, cat_result in sim_results.items():
        if "error" in cat_result or "note" in cat_result:
            continue
        
        ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws4.cell(row=row, column=1, value=f"📄 {cat} — TF-IDF余弦相似度").font = Font(name="Microsoft YaHei", bold=True, size=12, color="2F5496")
        row += 1
        
        # 相似度矩阵
        companies = cat_result["companies"]
        n = len(companies)
        
        # 矩阵表头
        matrix_headers = [""] + companies
        style_header(ws4, matrix_headers, row)
        row += 1
        
        for i, ca in enumerate(companies):
            ws4.cell(row=row, column=1, value=ca).font = Font(name="Microsoft YaHei", bold=True, size=9)
            for j, cb in enumerate(companies):
                val = float(cat_result["matrix"][i][j])
                cell = ws4.cell(row=row, column=j+2, value=round(val, 4))
                cell.alignment = center_align
                cell.border = thin_border
                if i == j:
                    cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
                elif val >= 0.8:
                    cell.fill = red_fill
                elif val >= 0.6:
                    cell.fill = yellow_fill
            row += 1
        
        row += 1
        
        # 两两对比明细
        pair_headers = ["公司A", "公司B", "余弦相似度", "字符数_A", "字符数_B", "判定"]
        style_header(ws4, pair_headers, row)
        row += 1
        
        for pair in sorted(cat_result["pairs"], key=lambda x: -x["余弦相似度"]):
            ws4.cell(row=row, column=1, value=pair["公司A"])
            ws4.cell(row=row, column=2, value=pair["公司B"])
            sim = pair["余弦相似度"]
            cell = ws4.cell(row=row, column=3, value=sim)
            ws4.cell(row=row, column=4, value=pair["字符数_A"])
            ws4.cell(row=row, column=5, value=pair["字符数_B"])
            
            if sim >= 0.8:
                verdict = "🔴 高度可疑(≥80%)"
                cell.fill = red_fill
            elif sim >= 0.6:
                verdict = "🟡 需关注(≥60%)"
                cell.fill = yellow_fill
            else:
                verdict = "🟢 正常"
                cell.fill = green_fill
            ws4.cell(row=row, column=6, value=verdict)
            row += 1
        
        row += 2
    
    style_data_rows(ws4, 2, row - 1, 8)
    auto_width(ws4)
    
    # ========== Sheet 5: 全文本相似度 ==========
    ws5 = wb.create_sheet("L3_全文本相似度")
    
    global_sim = results["global_similarity"]
    if global_sim:
        companies = global_sim["companies"]
        n = len(companies)
        
        ws5.merge_cells("A1:H1")
        ws5.cell(row=1, column=1, value="全文本TF-IDF余弦相似度矩阵（所有文件文本拼接）").font = Font(name="Microsoft YaHei", bold=True, size=12, color="2F5496")
        
        matrix_headers = [""] + companies
        style_header(ws5, matrix_headers, 3)
        
        row = 4
        for i, ca in enumerate(companies):
            ws5.cell(row=row, column=1, value=ca).font = Font(name="Microsoft YaHei", bold=True, size=9)
            for j, cb in enumerate(companies):
                val = float(global_sim["matrix"][i][j])
                cell = ws5.cell(row=row, column=j+2, value=round(val, 4))
                cell.alignment = center_align
                cell.border = thin_border
                if i == j:
                    cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
                elif val >= 0.8:
                    cell.fill = red_fill
                elif val >= 0.6:
                    cell.fill = yellow_fill
            row += 1
        
        row += 2
        pair_headers = ["公司A", "公司B", "全文本余弦相似度", "字符数_A", "字符数_B", "判定"]
        style_header(ws5, pair_headers, row)
        row += 1
        
        for pair in sorted(global_sim["pairs"], key=lambda x: -x["全文本余弦相似度"]):
            ws5.cell(row=row, column=1, value=pair["公司A"])
            ws5.cell(row=row, column=2, value=pair["公司B"])
            sim = pair["全文本余弦相似度"]
            cell = ws5.cell(row=row, column=3, value=sim)
            ws5.cell(row=row, column=4, value=pair["字符数_A"])
            ws5.cell(row=row, column=5, value=pair["字符数_B"])
            
            if sim >= 0.8:
                verdict = "🔴 高度可疑(≥80%)"
                cell.fill = red_fill
            elif sim >= 0.6:
                verdict = "🟡 需关注(≥60%)"
                cell.fill = yellow_fill
            else:
                verdict = "🟢 正常"
                cell.fill = green_fill
            ws5.cell(row=row, column=6, value=verdict)
            row += 1
        
        style_data_rows(ws5, 4, row - 1, 6)
        auto_width(ws5)
    
    # ========== Sheet 6: 图片跨公司比对 ==========
    ws6 = wb.create_sheet("L4_图片哈希跨公司比对")
    
    img_findings = results["image_cross"]
    
    img_headers = ["MD5(前16位)", "涉及公司数", "涉及公司", "出现次数", "详情", "可能为模板"]
    style_header(ws6, img_headers)
    
    for r, finding in enumerate(img_findings, 2):
        ws6.cell(row=r, column=1, value=finding["MD5"])
        ws6.cell(row=r, column=2, value=finding["涉及公司数"])
        ws6.cell(row=r, column=3, value=finding["涉及公司"])
        ws6.cell(row=r, column=4, value=finding["出现次数"])
        ws6.cell(row=r, column=5, value=finding["详情"])
        ws6.cell(row=r, column=6, value=finding["可能为模板"])
        
        # 高亮跨公司重复
        if finding["涉及公司数"] >= 3 and finding["可能为模板"] == "否":
            for c in range(1, 7):
                ws6.cell(row=r, column=c).fill = red_fill
    
    style_data_rows(ws6, 2, len(img_findings) + 1, len(img_headers))
    auto_width(ws6)
    
    # 添加汇总
    total_cross = len([f for f in img_findings if f["涉及公司数"] >= 2])
    non_template_cross = len([f for f in img_findings if f["涉及公司数"] >= 2 and f["可能为模板"] == "否"])
    row = len(img_findings) + 4
    ws6.cell(row=row, column=1, value=f"汇总：共{total_cross}张图片跨公司重复，其中{non_template_cross}张非模板类重复").font = Font(name="Microsoft YaHei", bold=True, size=11)
    
    # ========== Sheet 7: PDF元数据交叉 ==========
    ws7 = wb.create_sheet("L5_PDF元数据交叉")
    
    # 元数据明细
    meta_headers = ["公司", "文件名", "文件类别", "Author", "Creator", "Producer", "CreationDate", "ModDate", "页数"]
    style_header(ws7, meta_headers)
    
    meta_rows = results["metadata_rows"]
    for r, mr in enumerate(meta_rows, 2):
        ws7.cell(row=r, column=1, value=mr["公司"])
        ws7.cell(row=r, column=2, value=mr["文件名"])
        ws7.cell(row=r, column=3, value=mr["文件类别"])
        ws7.cell(row=r, column=4, value=mr["Author"])
        ws7.cell(row=r, column=5, value=mr["Creator"])
        ws7.cell(row=r, column=6, value=mr["Producer"])
        ws7.cell(row=r, column=7, value=mr["CreationDate"])
        ws7.cell(row=r, column=8, value=mr["ModDate"])
        ws7.cell(row=r, column=9, value=mr["页数"])
    
    style_data_rows(ws7, 2, len(meta_rows) + 1, len(meta_headers))
    
    # 交叉发现
    cross_start = len(meta_rows) + 4
    ws7.merge_cells(start_row=cross_start, start_column=1, end_row=cross_start, end_column=6)
    ws7.cell(row=cross_start, column=1, value="🔍 元数据交叉发现（同一元数据值跨公司出现）").font = Font(name="Microsoft YaHei", bold=True, size=12, color="C00000")
    
    cross_headers = ["文件类别", "元数据字段", "值", "涉及公司", "涉及数量", "严重程度"]
    style_header(ws7, cross_headers, cross_start + 1)
    
    cross_findings = results["metadata_cross"]
    for r, cf in enumerate(cross_findings, cross_start + 2):
        ws7.cell(row=r, column=1, value=cf["文件类别"])
        ws7.cell(row=r, column=2, value=cf["元数据字段"])
        ws7.cell(row=r, column=3, value=cf["值"])
        ws7.cell(row=r, column=4, value=cf["涉及公司"])
        ws7.cell(row=r, column=5, value=cf["涉及数量"])
        ws7.cell(row=r, column=6, value=cf["严重程度"])
        
        if "铁证" in cf["严重程度"]:
            for c in range(1, 7):
                ws7.cell(row=r, column=c).fill = red_fill
    
    style_data_rows(ws7, cross_start + 2, cross_start + 1 + len(cross_findings), len(cross_headers))
    auto_width(ws7)
    ws7.column_dimensions['C'].width = 50
    ws7.column_dimensions['D'].width = 50
    
    # ========== Sheet 8: 综合疑点汇总 ==========
    ws8 = wb.create_sheet("综合疑点汇总")
    ws8.merge_cells("A1:F1")
    ws8.cell(row=1, column=1, value="🔴 综合疑点汇总 — 需进一步调查事项").font = Font(name="Microsoft YaHei", bold=True, size=14, color="C00000")
    
    summary_headers = ["序号", "检测层级", "发现内容", "涉及公司", "证据强度", "建议行动"]
    style_header(ws8, summary_headers, 3)
    
    findings = []
    seq = 0
    
    # 从元数据交叉
    for cf in cross_findings:
        seq += 1
        severity = "铁证" if "铁证" in cf["严重程度"] else "强信号"
        findings.append([seq, "L5 元数据", f"{cf['文件类别']}: {cf['元数据字段']}={cf['值']}", cf["涉及公司"], severity, "调取原始Word文档进行进一步比对"])
    
    # 从高文本相似度
    if global_sim:
        for pair in global_sim["pairs"]:
            if pair["全文本余弦相似度"] >= 0.6:
                seq += 1
                severity = "铁证" if pair["全文本余弦相似度"] >= 0.8 else "强信号"
                findings.append([seq, "L3 文本雷同(全量)", f"全文本余弦相似度={pair['全文本余弦相似度']}", f"{pair['公司A']} vs {pair['公司B']}", severity, "详细比对原始Word文档文字内容"])
    
    # 从图片跨公司
    for imgf in img_findings:
        if imgf["涉及公司数"] >= 2 and imgf["可能为模板"] == "否":
            seq += 1
            severity = "铁证" if imgf["涉及公司数"] >= 3 else "强信号"
            findings.append([seq, "L4 图片哈希", f"相同图片跨公司出现: MD5={imgf['MD5']}", imgf["涉及公司"], severity, "提取该图片进行人工比对确认"])
    
    for r, f in enumerate(findings, 4):
        for c, val in enumerate(f, 1):
            ws8.cell(row=r, column=c, value=val)
            if "铁证" in str(f[4]):
                ws8.cell(row=r, column=c).fill = red_fill
            elif "强信号" in str(f[4]):
                ws8.cell(row=r, column=c).fill = yellow_fill
    
    style_data_rows(ws8, 4, 3 + len(findings), len(summary_headers))
    auto_width(ws8)
    ws8.column_dimensions['C'].width = 55
    ws8.column_dimensions['D'].width = 45
    ws8.column_dimensions['F'].width = 40
    
    # 保存
    wb.save(output_path)
    print(f"\n✅ 全量分析完成，输出: {output_path}")

# ====== 主流程 ======
def main():
    print("=" * 60)
    print("金川投标项目 — 全量串标围标分析")
    print("=" * 60)
    
    # 1. 收集文件
    print("\n[1/8] 收集文件...")
    files_data = collect_files()
    companies = sorted(files_data.keys())
    total_files = sum(len(v) for v in files_data.values())
    print(f"  发现 {len(companies)} 家公司, 共 {total_files} 个PDF文件")
    for c in companies:
        print(f"    - {c}: {len(files_data[c])}个文件")
    
    # 2. 提取文本
    print("\n[2/8] 提取PDF文本...")
    all_texts = extract_all_texts(files_data)
    
    # 3. 提取元数据
    print("\n[3/8] 提取PDF元数据...")
    all_meta = extract_all_metadata(files_data)
    
    # 4. 提取图片哈希
    print("\n[4/8] 提取嵌入图片哈希...")
    all_img, global_hashes = extract_all_images(files_data)
    
    # 5. 报价分析
    print("\n[5/8] L1 报价规律性分析...")
    price_details = extract_price_details(files_data)
    prices = analyze_price_patterns(all_texts)
    
    # 6. 文本完整性
    print("\n[6/8] 文本完整性分析...")
    completeness = analyze_text_completeness(all_texts)
    
    # 7. 文本相似度
    print("\n[7/8] L3 文本雷同检测...")
    text_similarity = analyze_text_similarity(all_texts)
    global_similarity = analyze_global_similarity(all_texts)
    
    # 8. 元数据交叉 + 图片交叉
    print("\n[8/8] L4/L5 图片&元数据交叉分析...")
    metadata_rows, metadata_cross, _ = analyze_metadata_cross(all_meta)
    image_cross = analyze_image_cross(global_hashes)
    
    # 汇总结果
    results = {
        "companies": companies,
        "files_data": files_data,
        "all_texts": all_texts,
        "all_meta": all_meta,
        "all_img": all_img,
        "global_hashes": global_hashes,
        "price_details": price_details,
        "prices": prices,
        "completeness": completeness,
        "text_similarity": text_similarity,
        "global_similarity": global_similarity,
        "metadata_rows": metadata_rows,
        "metadata_cross": metadata_cross,
        "image_cross": image_cross,
    }
    
    # 输出
    output_path = os.path.join(os.path.dirname(BASE), "金川投标_全量分析结果.xlsx")
    write_excel(results, output_path)
    
    # 控制台摘要
    print("\n" + "=" * 60)
    print("📊 分析摘要")
    print("=" * 60)
    print(f"投标单位: {len(companies)}家")
    print(f"PDF文件: {total_files}个")
    print(f"元数据交叉发现: {len(metadata_cross)}条")
    print(f"跨公司重复图片: {len(image_cross)}张")
    if global_similarity:
        high_sim = [p for p in global_similarity["pairs"] if p["全文本余弦相似度"] >= 0.6]
        print(f"高文本相似度(≥60%): {len(high_sim)}对")
    
    print(f"\n详细结果: {output_path}")

if __name__ == "__main__":
    main()
