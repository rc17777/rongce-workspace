#!/usr/bin/env python3
"""Generate comprehensive Excel analysis report for 护理学院采购项目"""
import sys, io, re, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime

base = r'D:\openclaw-workspace\projects\护理学院培训资料采购'

# ===== Color definitions =====
def safe_fill(color):
    from openpyxl.styles.fills import GradientFill
    return PatternFill(patternType='solid', fgColor=color)

RED_FILL = safe_fill('FFD7D7')
YELLOW_FILL = safe_fill('FFF3CD')
GREEN_FILL = safe_fill('D4EDDA')
HEADER_FILL = safe_fill('1A3A6E')
HEADER_FONT = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
TITLE_FONT = Font(name='微软雅黑', size=14, bold=True, color='1A3A6E')
NORMAL_FONT = Font(name='微软雅黑', size=10)
BOLD_FONT = Font(name='微软雅黑', size=10, bold=True)
RED_FONT = Font(name='微软雅黑', size=10, color='CC0000', bold=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ===== Load data =====
# TF-IDF on tech proposals
tech_texts = {}
for idx, name in enumerate(['建韬科技', '江楼商贸', '拓奇长荣'], 1):
    fpath = os.path.join(base, f'bidder{idx}_技术方案.txt')
    raw = ''
    if os.path.exists(fpath):
        with open(fpath, 'r', encoding='utf-8') as f:
            raw = f.read()
    clean = re.sub(r'=== PAGE \d+ ===', '', raw)
    clean = re.sub(r'[^\u4e00-\u9fffa-zA-Z0-9 \n]', '', clean)
    clean = re.sub(r'\s+', ' ', clean)
    tech_texts[name] = clean

# Full text (from earlier extraction)
full_texts = {}
for idx, name in enumerate(['建韬科技', '江楼商贸', '拓奇长荣'], 1):
    fpath = os.path.join(base, f'bidder{idx}_建韬.txt' if idx==1 else f'bidder{idx}_江楼.txt' if idx==2 else f'bidder{idx}_拓奇长荣.txt')
    raw = ''
    if os.path.exists(fpath):
        with open(fpath, 'r', encoding='utf-8') as f:
            raw = f.read()
    clean = re.sub(r'=== PAGE \d+ ===', '', raw)
    clean = re.sub(r'[^\u4e00-\u9fffa-zA-Z0-9 \n]', '', clean)
    clean = re.sub(r'\s+', ' ', clean)
    full_texts[name] = clean

# TF-IDF calculations
vectorizer_full = TfidfVectorizer(max_features=5000, token_pattern=r'[\u4e00-\u9fff]+')
names_list = ['建韬科技', '江楼商贸', '拓奇长荣']
tfidf_full = vectorizer_full.fit_transform([full_texts[n] for n in names_list])
sim_full = cosine_similarity(tfidf_full)

vectorizer_tech = TfidfVectorizer(max_features=5000, token_pattern=r'[\u4e00-\u9fff]+')
tfidf_tech = vectorizer_tech.fit_transform([tech_texts[n] for n in names_list])
sim_tech = cosine_similarity(tfidf_tech)

# ===== Create workbook =====
wb = openpyxl.Workbook()

# ---- Sheet 1: 综合结论 ----
ws1 = wb.active
ws1.title = '综合结论'
ws1.sheet_properties.tabColor = '1A3A6E'

# Title
ws1.merge_cells('A1:G1')
ws1['A1'] = '四川护理职业学院2025年培训教学资料委托制作服务采购项目 — 串标围标分析报告'
ws1['A1'].font = TITLE_FONT
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 30

ws1.merge_cells('A2:G2')
ws1['A2'] = f'项目编号: SCLT20250232 | 采购方式: 竞争性磋商 | 预算: 25万元 | 分析日期: {datetime.now().strftime("%Y-%m-%d")}'
ws1['A2'].font = Font(name='微软雅黑', size=10, color='666666')
ws1['A2'].alignment = Alignment(horizontal='center')

# Ten-layer conclusion table
headers_conclusion = ['层级', '检测维度', '检测状态', '关键发现', '风险等级', '详细说明']
for c, h in enumerate(headers_conclusion, 1):
    cell = ws1.cell(row=4, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

conclusions = [
    ['L1', '报价规律性', '已完成', '三家报价自然分布(21.6万/22.9万/23.7万)，极差9.6%，非等差数列', '🟢 正常', '报价无机械阶梯或等比分布痕迹'],
    ['L2', '投标IP/MAC', '不适用', '本项目为纸质投标，无电子投标系统', '—', '纸质投标无法获取IP/MAC数据'],
    ['L3', 'TF-IDF文本雷同', '已完成', '全文86.6%-89.8%高度相似；技术方案75.6%-80.9%中高相似', '🔴 关注', '全文相似主要来自标准承诺函模板(17-27对段落100%一致均为模板)；技术方案75-81%相似度偏高，需进一步排除行业通稿因素'],
    ['L4', '图片/资源哈希', '不适用', '所有投标文件为扫描版PDF，无嵌入图片', '—', '纸质盖章后扫描，无法提取图片哈希'],
    ['L5', 'PDF元数据', '已完成', '🚨 三家6份PDF全部使用同一型号扫描仪: Fuji Xerox D125', '🔴 高度关注', '扫描时间: 建韬4/8晚20:48-49, 江楼4/8晚22:28-29, 拓奇4/9午13:13-18。同型号设备可能是同一打印店'],
    ['L6', '文档结构/体量', '已完成', '拓奇技术方案327页 vs 江楼58页 vs 建韬92页，差异极大', '🟡 关注', '拓奇327页含极详细方案(8大部分47章)，远超过竞争对手'],
    ['L7', '打印机/扫描仪', '已完成', '同L5: 全部Fuji Xerox D125扫描', '🔴 关注', '同一型号扫描仪。江楼PDF格式1.3 vs 建韬/拓奇1.6，扫描设置略有不同'],
    ['L8', '工商关联', '初步分析', '三家注册地均在武侯区；建韬(玉林北街)与拓奇(玉林中路)同在玉林片区', '🟡 关注', '建韬: 武侯区玉林北街1号; 拓奇: 武侯区玉林中路13号; 江楼: 武侯区武侯大道铁佛段。建议天眼查股权穿透'],
    ['L9', '保证金/资金链', '不适用', '纸质投标无此数据', '—', '—'],
    ['L10', '代理人/签到', '未获取', '签到表未包含在分析材料中', '—', '需调取原件'],
]

for r, row_data in enumerate(conclusions, 5):
    for c, val in enumerate(row_data, 1):
        cell = ws1.cell(row=r, column=c, value=val)
        cell.font = NORMAL_FONT
        cell.alignment = LEFT_WRAP
        cell.border = THIN_BORDER
        if '🔴' in str(val):
            cell.fill = RED_FILL
        elif '🟡' in str(val):
            cell.fill = YELLOW_FILL
        elif '🟢' in str(val):
            cell.fill = GREEN_FILL

# Overall assessment
r = 5 + len(conclusions) + 1
ws1.merge_cells(f'A{r}:G{r}')
ws1[f'A{r}'] = '综合判定'
ws1[f'A{r}'].font = Font(name='微软雅黑', size=12, bold=True, color='CC0000')
r += 1
summary_text = (
    '【核心发现】\n'
    '1. [L5扫描仪] 三家6份PDF全部使用Fuji Xerox D125扫描，这是最值得关注的技术线索。\n'
    '   可能存在共用同一打印店/扫描设备的情况，建议核实投标文件制作地点。\n'
    '2. [L3文本雷同] 技术方案TF-IDF相似度75-81%，虽低于全文(86-90%)，但仍偏高。\n'
    '   需要人工对比具体段落判断是否存在实质性内容雷同。\n'
    '3. [L8工商] 三家均在武侯区注册，建韬与拓奇同在玉林片区(相距约500米)。\n'
    '   建议通过天眼查/企查查做股权穿透，排除关联关系。\n'
    '4. [得分] 第一名(91.35)与第二名(67.33)分差24分，虽然拓奇327页方案可能质量显著更优，\n'
    '   但结合扫描仪一致性和地理接近性，建议对评审过程做合规性复查。\n\n'
    '【总体评估】: 存在多个需要关注的信号(L5扫描仪+ L3文本雷同+ L8地理接近+ 得分极差)，\n'
    '虽未达到"铁证"级别，但足以启动进一步调查程序。'
)
ws1.merge_cells(f'A{r}:G{r}')
ws1[f'A{r}'] = summary_text
ws1[f'A{r}'].font = NORMAL_FONT
ws1[f'A{r}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws1.row_dimensions[r].height = 180

# Column widths
ws1.column_dimensions['A'].width = 10
ws1.column_dimensions['B'].width = 16
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 55
ws1.column_dimensions['E'].width = 14
ws1.column_dimensions['F'].width = 50
ws1.column_dimensions['G'].width = 2

# ---- Sheet 2: 元数据分析 ----
ws2 = wb.create_sheet('元数据分析')
ws2.sheet_properties.tabColor = 'C0392B'

ws2.merge_cells('A1:H1')
ws2['A1'] = 'PDF元数据分析'
ws2['A1'].font = TITLE_FONT

metadata_headers = ['投标单位', '文件类型', '页数', '文件大小(MB)', 'Creator(扫描仪)', 'Producer', '创建日期', '修改日期']
for c, h in enumerate(metadata_headers, 1):
    cell = ws2.cell(row=3, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

metadata_rows = [
    ['四川建韬科技', '资格性响应文件', 57, '13.6 MB', 'Fuji Xerox D125', 'Fuji Xerox D125', '2025-04-08 20:49 (+10)', '2025-04-08 20:35 (+8)'],
    ['四川建韬科技', '其他响应文件', 92, '24.7 MB', 'Fuji Xerox D125', 'Fuji Xerox D125', '2025-04-08 20:48 (+10)', '2025-04-08 20:34 (+8)'],
    ['成都江楼商贸', '资格性响应文件', 51, '13.7 MB', 'Fuji Xerox D125', 'Fuji Xerox D125', '2025-04-08 22:29 (+10)', '2025-04-08 22:29 (+10)'],
    ['成都江楼商贸', '其他响应文件', 58, '15.0 MB', 'Fuji Xerox D125', 'Fuji Xerox D125', '2025-04-08 22:28 (+10)', '2025-04-08 22:28 (+10)'],
    ['成都拓奇长荣商贸', '资格性响应文件', 81, '24.3 MB', 'Fuji Xerox D125', 'Fuji Xerox D125', '2025-04-09 13:13 (+10)', '2025-04-09 12:34 (+8)'],
    ['成都拓奇长荣商贸', '其他响应文件', 327, '105.5 MB', 'Fuji Xerox D125', 'Fuji Xerox D125', '2025-04-09 13:18 (+10)', '2025-04-09 16:01 (+8)'],
]

for r, row_data in enumerate(metadata_rows, 4):
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.font = NORMAL_FONT
        cell.alignment = CENTER if c <= 4 else LEFT_WRAP
        cell.border = THIN_BORDER
        # Highlight scan time differences
        if c == 7 and '22:2' in str(val):
            cell.fill = YELLOW_FILL

# Analysis box
r_analysis = 4 + len(metadata_rows) + 1
ws2.merge_cells(f'A{r_analysis}:H{r_analysis}')
ws2[f'A{r_analysis}'] = '▎关键发现'
ws2[f'A{r_analysis}'].font = Font(name='微软雅黑', size=12, bold=True, color='CC0000')

analysis_lines = [
    '1. 【设备一致性】 所有6份PDF的Creator/Producer均为"Fuji Xerox D125"，表明三家公司使用了同一型号的扫描/复印设备。',
    '   - Fuji Xerox D125 是一款高端办公复合机(打印/复印/扫描一体)，常见于专业打印店或大型企业。',
    '   - 如果三家分别在不同地点扫描，同时使用同型号设备的概率较低，暗示可能在同一打印店完成投标文件制作。',
    '',
    '2. 【时间线分析】',
    '   - 建韬科技: 4月8日晚上20:48-20:49扫描(两个文件几乎同时完成)',
    '   - 江楼商贸: 4月8日晚上22:28-22:29扫描(比建韬晚约1.5小时)',
    '   - 拓奇长荣: 4月9日下午13:13-13:18扫描(晚一天)',
    '   - 投标截止/开标日期: 4月10日',
    '   - 三家扫描时间分散在两天内，不构成"同一时刻"的证据。',
    '',
    '3. 【时区异常】 建韬和拓奇的创建时间时区为+10(澳洲东部)，修改时间为+8(北京时间)，可能是扫描仪时区配置问题。',
    '   江楼的两个时间均为+10，设置较一致。',
    '',
    '4. 【PDF版本差异】 江楼的PDF格式为1.3，建韬/拓奇为1.6，表明扫描软件或设置存在差异，不完全一致。',
    '',
    '5. 【建议】 向三家供应商核实投标文件的制作地点(自行扫描 vs 委托打印店)，交叉验证。'
]
for i, line in enumerate(analysis_lines):
    ws2.merge_cells(f'A{r_analysis+1+i}:H{r_analysis+1+i}')
    ws2[f'A{r_analysis+1+i}'] = line
    ws2[f'A{r_analysis+1+i}'].font = NORMAL_FONT
    ws2[f'A{r_analysis+1+i}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 14
ws2.column_dimensions['E'].width = 22
ws2.column_dimensions['F'].width = 22
ws2.column_dimensions['G'].width = 26
ws2.column_dimensions['H'].width = 26

# ---- Sheet 3: 报价对比 ----
ws3 = wb.create_sheet('报价对比')
ws3.sheet_properties.tabColor = '2C3E50'

ws3.merge_cells('A1:J1')
ws3['A1'] = '报价明细对比分析'
ws3['A1'].font = TITLE_FONT

price_headers = ['序号', '标的名称', '单项限价(元)', '建韬报价(元)', '江楼报价(元)', '拓奇报价(元)', '最低报价', '最高报价', '极差(%)', '判定']
for c, h in enumerate(price_headers, 1):
    cell = ws3.cell(row=3, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

# Estimated pricing data (from filing doc final quotes + bidding doc limits)
# Total: 建韬 237,346 | 江楼 216,613 | 拓奇 229,237 | Budget 250,000
# We don't have per-item breakdowns, so use totals + item descriptions
items = [
    ['学员袋1', 40, '—', '—', '—', '—', '—', '—', '见备注'],
    ['学员袋2', 35, '—', '—', '—', '—', '—', '—', '见备注'],
    ['学员袋3', 30, '—', '—', '—', '—', '—', '—', '见备注'],
    ['学员证1', 10, '—', '—', '—', '—', '—', '—', '见备注'],
    ['学员证2', 6, '—', '—', '—', '—', '—', '—', '见备注'],
    ['笔1', 3.8, '—', '—', '—', '—', '—', '—', '见备注'],
    ['笔2', 2, '—', '—', '—', '—', '—', '—', '见备注'],
    ['笔记本1', 20, '—', '—', '—', '—', '—', '—', '见备注'],
    ['笔记本2', 16, '—', '—', '—', '—', '—', '—', '见备注'],
    ['笔记本3', 14, '—', '—', '—', '—', '—', '—', '见备注'],
    ['笔记本4', 8, '—', '—', '—', '—', '—', '—', '见备注'],
    ['学员手册', 10, '—', '—', '—', '—', '—', '—', '见备注'],
    ['结业证书1', 13, '—', '—', '—', '—', '—', '—', '见备注'],
    ['结业证书2', 12, '—', '—', '—', '—', '—', '—', '见备注'],
    ['结业证书3', 8, '—', '—', '—', '—', '—', '—', '见备注'],
]

for r, item in enumerate(items, 4):
    for c, val in enumerate(item, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        cell.font = NORMAL_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        if c == 1:
            cell.value = r - 3

# Total row
total_row = 4 + len(items)
totals = ['', '合计', '—', '237,346.20', '216,613.20', '229,237.50', '216,613.20', '237,346.20', '9.57%', '🟢 正常']
for c, val in enumerate(totals, 1):
    cell = ws3.cell(row=total_row, column=c, value=val)
    cell.font = BOLD_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    if c >= 7:
        cell.fill = GREEN_FILL if '🟢' in str(val) else None

# Note
note_row = total_row + 2
ws3.merge_cells(f'A{note_row}:J{note_row}')
ws3[f'A{note_row}'] = '说明: 单项报价明细因OCR限于纸质扫描件表格识别能力无法逐项提取。上表仅汇总最后报价总金额。'
ws3[f'A{note_row}'].font = Font(name='微软雅黑', size=10, italic=True, color='888888')

# Score analysis
score_row = note_row + 2
ws3.merge_cells(f'A{score_row}:J{score_row}')
ws3[f'A{score_row}'] = '▎评标得分对比'
ws3[f'A{score_row}'].font = Font(name='微软雅黑', size=12, bold=True, color='CC0000')

score_headers = ['排名', '供应商', '总分', '最后报价', '报价排名(低→高)', '报价分(估30%)', '非价格分(70%)']
for c, h in enumerate(score_headers, 1):
    cell = ws3.cell(row=score_row+1, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

score_data = [
    ['1', '成都拓奇长荣商贸有限公司', '91.35', '229,237.50', '第2(中)', '28.42(估)', '62.93(领先)'],
    ['2', '成都江楼商贸有限公司', '67.33', '216,613.20', '第1(低)', '30.00(估)', '37.33(落后)'],
    ['3', '四川建韬科技有限公司', '62.71', '237,346.20', '第3(高)', '27.38(估)', '35.33(落后)'],
]
for r, row_data in enumerate(score_data, score_row+2):
    for c, val in enumerate(row_data, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        cell.font = BOLD_FONT if c == 2 else NORMAL_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        if r == score_row+2:  # Winner
            cell.font = RED_FONT

score_note = score_row + 6
ws3.merge_cells(f'A{score_note}:J{score_note}')
ws3[f'A{score_note}'] = ('关键: 第1名与第2名分差24.02分，但报价仅差5.5%。非价格因素(实施方案/样品/设计/业绩)差距高达25.6分，'
                          '表明评委会对拓奇长荣的技术方案和样品认可度远高于其他两家。')
ws3[f'A{score_note}'].font = Font(name='微软雅黑', size=10, color='CC0000')

ws3.column_dimensions['A'].width = 8
ws3.column_dimensions['B'].width = 30
ws3.column_dimensions['C'].width = 16
ws3.column_dimensions['D'].width = 16
ws3.column_dimensions['E'].width = 16
ws3.column_dimensions['F'].width = 16
ws3.column_dimensions['G'].width = 16
ws3.column_dimensions['H'].width = 16
ws3.column_dimensions['I'].width = 12
ws3.column_dimensions['J'].width = 14

# ---- Sheet 4: 文本雷同分析 ----
ws4 = wb.create_sheet('文本雷同分析')
ws4.sheet_properties.tabColor = 'F39C12'

ws4.merge_cells('A1:G1')
ws4['A1'] = 'TF-IDF文本雷同分析'
ws4['A1'].font = TITLE_FONT

ws4.merge_cells('A2:G2')
ws4['A2'] = '说明: 85%+ 高度可疑 | 50-85% 中度相似需结合内容判断 | <50% 正常'
ws4['A2'].font = Font(name='微软雅黑', size=10, color='888888')

tfidf_headers = ['对比维度', '建韬 vs 江楼', '建韬 vs 拓奇', '江楼 vs 拓奇', '阈值', '判定', '备注']
for c, h in enumerate(tfidf_headers, 1):
    cell = ws4.cell(row=4, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

# TF-IDF data
sim_pairs = [(0,1), (0,2), (1,2)]
tfidf_data = [
    ['全文TF-IDF(资格性+其他响应文件)', 
     f'{sim_full[sim_pairs[0]]*100:.1f}%', 
     f'{sim_full[sim_pairs[1]]*100:.1f}%', 
     f'{sim_full[sim_pairs[2]]*100:.1f}%',
     '>=80% 可疑', '🔴 全部超标', '高度相似主要来自标准承诺函模板'],
    ['技术方案TF-IDF(其他响应文件技术部分)', 
     f'{sim_tech[sim_pairs[0]]*100:.1f}%', 
     f'{sim_tech[sim_pairs[1]]*100:.1f}%', 
     f'{sim_tech[sim_pairs[2]]*100:.1f}%',
     '>=80% 可疑', '🟡 江楼vs拓奇超标', '建韬vs江楼/拓奇在80%以下'],
    ['技术方案字符量', '29,399', '32,908', '57,685', '—', '差距大', '拓奇内容量约为建韬的2倍'],
    ['其他响应文件总页数', '92页', '58页', '327页', '—', '体量悬殊', '拓奇页数是江楼的5.6倍'],
    ['技术关键词命中(20词)', '0词', '2词', '5词', '—', '—', 'OCR噪声可能影响命中统计'],
]

for r, row_data in enumerate(tfidf_data, 5):
    for c, val in enumerate(row_data, 1):
        cell = ws4.cell(row=r, column=c, value=val)
        cell.font = NORMAL_FONT
        cell.alignment = CENTER if c <= 5 else LEFT_WRAP
        cell.border = THIN_BORDER
        # Color code
        if c >= 2 and c <= 4 and '%' in str(val):
            pct = float(val.replace('%',''))
            if pct >= 80:
                cell.fill = RED_FILL
            elif pct >= 50:
                cell.fill = YELLOW_FILL
            else:
                cell.fill = GREEN_FILL

# Analysis
r2 = 5 + len(tfidf_data) + 1
ws4.merge_cells(f'A{r2}:G{r2}')
ws4[f'A{r2}'] = '▎分析说明'
ws4[f'A{r2}'].font = Font(name='微软雅黑', size=12, bold=True, color='1A3A6E')

tfidf_explanations = [
    '1. 全文TF-IDF(86.6%-89.8%): 严重超标，但经段落级验证，17-27对100%匹配段落均为采购文件要求的统一承诺函/声明模板。',
    '   典型模板句: "本单位对上述承诺的内容事项真实性负责"、"代理人无转委托权"、"不存在同时是采购代理机构工作人员的情形"',
    '   结论: 全文高相似度主要由格式性承诺函文本驱动，不构成实质性内容雷同证据。',
    '',
    '2. 技术方案TF-IDF(75.6%-80.9%): 江楼vs拓奇达到80.9%超过红线，处于需关注的临界区间。',
    '   由于OCR识别精度有限(扫描件噪声)，部分"相似度"可能来自OCR误识别产生的随机字符组合。',
    '   建议: 人工对比技术方案原件，逐段检查是否存在实质性的文字复制。',
    '',
    '3. 文档体量(92/58/327页): 拓奇的327页技术方案是江楼58页的5.6倍，内容量级差异巨大，',
    '   从侧面说明三家技术方案的原创程度存在本质差异。'
]
for i, line in enumerate(tfidf_explanations):
    ws4.merge_cells(f'A{r2+1+i}:G{r2+1+i}')
    ws4[f'A{r2+1+i}'] = line
    ws4[f'A{r2+1+i}'].font = NORMAL_FONT
    ws4[f'A{r2+1+i}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

ws4.column_dimensions['A'].width = 32
ws4.column_dimensions['B'].width = 18
ws4.column_dimensions['C'].width = 18
ws4.column_dimensions['D'].width = 18
ws4.column_dimensions['E'].width = 14
ws4.column_dimensions['F'].width = 20
ws4.column_dimensions['G'].width = 35

# ---- Sheet 5: 工商信息对比 ----
ws5 = wb.create_sheet('工商信息对比')
ws5.sheet_properties.tabColor = '27AE60'

ws5.merge_cells('A1:G1')
ws5['A1'] = '供应商工商信息对比'
ws5['A1'].font = TITLE_FONT

biz_headers = ['信息项', '四川建韬科技', '成都江楼商贸', '成都拓奇长荣商贸', '比对结果', '风险标记']
for c, h in enumerate(biz_headers, 1):
    cell = ws5.cell(row=3, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

biz_data = [
    ['法定代表人', '樊巧玲', '刘文杰', '田瑶', '各不相同', '🟢'],
    ['授权代表', '赖友弘', '(法定代表人直接参与)', '向欢欢', '各不相同', '🟢'],
    ['统一社会信用代码', '915101000753879598', '915101083215605861', '未提取到', '不同', '🟢'],
    ['成立日期', '2013年8月23日', '2014年12月10日', '2023年(具体月日未提取)', '不同', '🟢'],
    ['注册资本', '捌佰万元(800万)', '未提取', '未提取', '—', '🟢'],
    ['注册地址', '成都市武侯区玉林北街1号1栋3单元4层401号', '成都市武侯区武侯大道铁佛段1号1栋1单元11层1123号', '成都市武侯区玉林中路13号附1122号', '同在武侯区', '🟡'],
    ['地址细分', '玉林北街', '武侯大道铁佛段', '玉林中路', '建韬与拓奇同在玉林片区', '🟡'],
    ['经营范围', '计算机软硬件/安防设备', '百货/水果/建材/办公/物业/AI/市政(极广)', '工艺美术品/办公用品/日用百货', '差异大', '🟢'],
    ['经营年限', '约13年', '约11年', '约3年', '差异大', '🟢'],
]

for r, row_data in enumerate(biz_data, 4):
    for c, val in enumerate(row_data, 1):
        cell = ws5.cell(row=r, column=c, value=val)
        cell.font = BOLD_FONT if c == 1 else NORMAL_FONT
        cell.alignment = LEFT_WRAP
        cell.border = THIN_BORDER
        if '🟡' in str(val):
            cell.fill = YELLOW_FILL
        elif '🟢' in str(val):
            cell.fill = GREEN_FILL

# Note
r_biz_note = 4 + len(biz_data) + 1
ws5.merge_cells(f'A{r_biz_note}:F{r_biz_note}')
ws5[f'A{r_biz_note}'] = ('注意: 建韬科技与拓奇长荣同在武侯区玉林片区(玉林北街 vs 玉林中路)，步行距离约500米。'
                          '虽不构成关联证据，但在扫描仪一致(L5)的背景下，建议做股权穿透验证。')
ws5[f'A{r_biz_note}'].font = Font(name='微软雅黑', size=10, color='CC0000')

ws5.column_dimensions['A'].width = 18
ws5.column_dimensions['B'].width = 30
ws5.column_dimensions['C'].width = 30
ws5.column_dimensions['D'].width = 30
ws5.column_dimensions['E'].width = 22
ws5.column_dimensions['F'].width = 14

# Save
output_path = os.path.join(base, '串标围标分析报告.xlsx')
wb.save(output_path)
print(f'Excel saved: {output_path}')
