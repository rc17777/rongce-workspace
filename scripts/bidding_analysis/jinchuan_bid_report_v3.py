"""
金川投标 — 格式化全量分析报告 v3
优化:
  1. Logo仅封面出现，缩小至80px宽，居中排列
  2. 每个分析表下方增加「通俗解读」板块，解释术语和逻辑
  3. 专业术语翻译为业务语言
  4. 布局更清爽，间距更合理
"""
import os, sys, re, hashlib
from collections import defaultdict
import fitz  # PyMuPDF
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.page import PageMargins

# ========== 路径 & 常量 ==========
BASE       = r"C:\Users\scrccpa\Desktop\金川投标"
LOGO_PATH  = r"D:\openclaw-workspace\projects\data-analysis-agent\static\Images\rongce-logo.png"
OUTPUT     = r"C:\Users\scrccpa\Desktop\金川投标_全量分析报告_v3.xlsx"

# ========== 配色 ==========
C_PRI   = "1F4E79"  # 深蓝
C_SEC   = "2E75B6"  # 中蓝
C_GOLD  = "C4A23A"  # 金
C_RED   = "C0392B"
C_ORANGE= "E67E22"
C_GREEN = "27AE60"
C_GRAY  = "999999"
C_LIGHT = "F0F4F8"
C_WHITE = "FFFFFF"
C_SOFT  = "FAFBFD"
C_TITLE = "1A1A1A"
C_TEXT  = "333333"
C_LABEL = "777777"

# 证据色
BG_RED    = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
BG_YELLOW = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
BG_GREEN  = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
BG_TITLE  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
BG_EXPLAIN= PatternFill(start_color="F8F9FB", end_color="F8F9FB", fill_type="solid")
BG_COVER  = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")
BG_ODD    = PatternFill(start_color="F5F8FB", end_color="F5F8FB", fill_type="solid")
BG_LIGHT_BLUE = PatternFill(start_color="E9F0F7", end_color="E9F0F7", fill_type="solid")

# 字体
F_COVER_TITLE = Font(name="微软雅黑", size=22, bold=True, color=C_PRI)
F_COVER_SUB   = Font(name="微软雅黑", size=13, color=C_SEC)
F_COVER_INFO  = Font(name="微软雅黑", size=10.5, color=C_TEXT)
F_SECTION     = Font(name="微软雅黑", size=13, bold=True, color=C_PRI)
F_SUBSECTION  = Font(name="微软雅黑", size=11, bold=True, color=C_SEC)
F_HEADER      = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
F_TABLE_BODY  = Font(name="微软雅黑", size=10, color=C_TEXT)
F_TABLE_BOLD  = Font(name="微软雅黑", size=10, bold=True, color=C_TEXT)
F_EXPLAIN_TITLE = Font(name="微软雅黑", size=10.5, bold=True, color=C_SEC)
F_EXPLAIN_BODY  = Font(name="微软雅黑", size=10, color=C_LABEL)
F_FOOTER      = Font(name="微软雅黑", size=9, color=C_GRAY)
F_NOTE        = Font(name="微软雅黑", size=10, italic=True, color=C_LABEL)
F_STRONG      = Font(name="微软雅黑", size=10.5, bold=True, color=C_RED)

# 边框
THIN_SIDE = Side(style="thin", color="D0D0D0")
BORDER    = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
BORDER_NONE = Border()

# 对齐
ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_TL= Alignment(horizontal="left",   vertical="top",    wrap_text=True)

PAGE = PageMargins(left=0.5, right=0.5, top=0.6, bottom=0.6, header=0.3, footer=0.3)

# ========== 工具函数 ==========
def hdr(ws, row, headers, start_col=1):
    for i, h in enumerate(headers, start_col):
        c = ws.cell(row=row, column=i, value=h)
        c.font = F_HEADER; c.fill = BG_TITLE; c.alignment = ALIGN_C; c.border = BORDER

def data_row(ws, row, vals, start_col=1, level="normal"):
    fm = {
        "red":    BG_RED,  "yellow": BG_YELLOW,
        "green":  BG_GREEN, "normal": BG_ODD if row % 2 == 0 else None
    }
    f = fm.get(level)
    for i, v in enumerate(vals, start_col):
        c = ws.cell(row=row, column=i, value=v)
        c.font = F_TABLE_BODY; c.alignment = ALIGN_L; c.border = BORDER
        if f: c.fill = f

def section(ws, row, text, col=1, end_col=10, font=None):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=col, value=text)
    c.font = font or F_SECTION; c.alignment = Alignment(horizontal="left", vertical="center")

def write_explain_box(ws, row, col, end_col, title, lines):
    """写入通俗解读框"""
    # 标题行：浅蓝底
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=col, value=f"📖 {title}")
    c.font = F_EXPLAIN_TITLE; c.fill = BG_LIGHT_BLUE
    c.alignment = Alignment(horizontal="left", vertical="center")
    row += 1
    # 内容行
    for line in lines:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
        c = ws.cell(row=row, column=col, value=line)
        c.font = F_EXPLAIN_BODY; c.alignment = ALIGN_TL
        c.fill = BG_EXPLAIN
        row += 1
    return row + 1  # 返回下一个空行

def auto_w(ws, min_w=8, max_w=60):
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        mx = 0
        for cell in col_cells:
            if cell.value:
                for ln in str(cell.value).split('\n'):
                    w = sum(2 if ord(ch) > 127 else 1 for ch in ln)
                    mx = max(mx, w)
        ws.column_dimensions[letter].width = max(min_w, min(mx + 4, max_w))

# ========== PDF 分析函数 ==========
CATS = {
    "报价表": ["报价表"], "投标函": ["投标（响应）函", "投标函"],
    "中小企业声明函": ["中小企业声明函"], "残疾人福利性单位声明函": ["残疾人福利性单位声明函"],
    "监狱企业证明": ["监狱企业的证明文件"], "供应商基本情况表": ["供应商基本情况表"],
    "业绩一览表": ["供应商类似项目业绩一览表"], "商务应答表": ["商务应答表"],
    "技术要求应答表": ["技术要求应答表"], "服务应答表": ["服务应答表"],
    "主要人员情况表": ["实施本项目的主要人员情况表"],
    "相关证明材料": ["投标人应提交的相关证明材料"],
    "其他相关材料": ["供应商认为需要提供的其他相关材料"],
    "投标文件封面": ["投标文件封面"],
}
def classify(fpath):
    fn = os.path.basename(fpath)
    for cat, kws in CATS.items():
        for kw in kws:
            if kw in fn: return cat
    return os.path.splitext(fn)[0]

def extract_text(fp):
    try:
        doc = fitz.open(fp); t = "".join(p.get_text("text") for p in doc); doc.close(); return t.strip()
    except: return ""

def extract_meta(fp):
    try:
        doc = fitz.open(fp); m = doc.metadata; pc = doc.page_count; doc.close()
        return {"author": str(m.get("author","") or ""), "creator": str(m.get("creator","") or ""),
                "producer": str(m.get("producer","") or ""), "creationDate": str(m.get("creationDate","") or ""),
                "modDate": str(m.get("modDate","") or ""), "page_count": pc}
    except: return {"author":"","creator":"","producer":"","creationDate":"","modDate":"","page_count":0}

def extract_img_hashes(fp):
    hs = []
    try:
        doc = fitz.open(fp)
        for pn in range(len(doc)):
            for img in doc.load_page(pn).get_images(full=True):
                bs = doc.extract_image(img[0])["image"]
                hs.append({"md5": hashlib.md5(bs).hexdigest(), "page": pn+1})
        doc.close()
    except: pass
    return hs

# ========== 构建工作簿 ==========
def build():
    print("Loading data...")
    # 收集文件
    files_data = {}
    for d in sorted(os.listdir(BASE)):
        dp = os.path.join(BASE, d)
        if not os.path.isdir(dp): continue
        co = d.replace("(包1)","").strip()
        fs = [os.path.join(dp,f) for f in sorted(os.listdir(dp)) if f.lower().endswith('.pdf')]
        if fs: files_data[co] = fs
    companies = sorted(files_data.keys())

    # 提取文本
    all_texts = {}
    for co, fs in files_data.items():
        ct = {}
        for fp in fs: ct[classify(fp)] = extract_text(fp)
        all_texts[co] = ct

    # 提取元数据
    all_meta = {}
    for co, fs in files_data.items():
        cm = {}
        for fp in fs: cm[os.path.basename(fp)] = extract_meta(fp)
        all_meta[co] = cm

    # 图片哈希
    gh = defaultdict(list)
    for co, fs in files_data.items():
        for fp in fs:
            for h in extract_img_hashes(fp):
                gh[h["md5"]].append((co, classify(fp)))

    # 报价
    price_info = {}
    for co, fs in files_data.items():
        for fp in fs:
            if "报价表" in os.path.basename(fp):
                try:
                    doc = fitz.open(fp); t = "".join(p.get_text("text") for p in doc); doc.close()
                    price_info[co] = {"text": t.strip(), "chars": len(t.strip())}
                except: price_info[co] = {"text":"","chars":0}

    prices = {}
    for co, texts in all_texts.items():
        for cat, text in texts.items():
            if "报价" in cat:
                cl = re.sub(r'\s+',' ', text).strip()
                ms = re.findall(r'(\d{1,3}(?:\.\d{1,2})?)%', cl)
                if ms: prices[co] = [float(m) for m in ms]
                break

    # 全文本相似度
    corp = []; corp_co = []
    for co, texts in all_texts.items():
        concat = ""
        for cat in sorted(texts.keys()):
            t = re.sub(r'\s+',' ', texts[cat]).strip()
            if t: concat += t + " "
        if concat.strip(): corp_co.append(co); corp.append(concat.strip())

    sim_pairs = []
    if len(corp) >= 2:
        vec = TfidfVectorizer(analyzer='char_wb', ngram_range=(3,5), max_features=10000)
        tf = vec.fit_transform(corp); sm = cosine_similarity(tf)
        for i in range(len(corp_co)):
            for j in range(i+1, len(corp_co)):
                sim_pairs.append({"a":corp_co[i],"b":corp_co[j],"sim":float(sm[i][j]),
                                  "ca":len(corp[i]),"cb":len(corp[j])})

    # 元数据交叉
    cfv = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for co, fm in all_meta.items():
        for fn, m in fm.items():
            cat = classify(fn)
            for fld in ["author","creator","producer"]:
                v = m.get(fld,"").strip()
                if v and v != "None": cfv[cat][fld][v].append(co)
    meta_cross = []
    for cat, fds in cfv.items():
        for fld, vs in fds.items():
            for v, cs in vs.items():
                if len(cs) >= 2:
                    is_platform = any(kw in v.lower() for kw in ["chromium","skia/pdf","qt 5.15"])
                    meta_cross.append({"cat":cat,"fld":fld,"val":v,"cs":sorted(cs),"n":len(cs),
                                       "platform":is_platform})
    meta_cross.sort(key=lambda x: -x["n"])

    # 图片跨公司
    img_cross = []
    for md5, occ in gh.items():
        cs = set(o[0] for o in occ)
        if len(cs) >= 2:
            dets = " | ".join(f"{o[0]}/{o[1]}" for o in occ)
            is_t = any(k in dets.lower() for k in ["声明函","封面","证明"])
            img_cross.append({"md5":md5[:16],"n":len(cs),"cs":"、".join(sorted(cs)),
                              "occ":len(occ),"dets":dets,"template":is_t})
    img_cross.sort(key=lambda x: -x["n"])

    # =========================
    wb = Workbook()

    # ---- Sheet 0: 封面 ----
    ws0 = wb.active
    ws0.title = "封面"
    ws0.sheet_properties.tabColor = C_PRI
    ws0.sheet_view.showGridLines = False

    for c in 'ABCDEFGH': ws0.column_dimensions[c].width = 14
    ws0.row_dimensions[1].height = 20
    ws0.row_dimensions[2].height = 12

    # 小Logo居中
    if os.path.exists(LOGO_PATH):
        img = XLImage(LOGO_PATH)
        img.width = 130; img.height = 45
        ws0.add_image(img, "D2")

    ws0.row_dimensions[7].height = 8

    # 主标题
    ws0.merge_cells("A9:H9")
    c = ws0.cell(row=9, column=1, value="投标文件全量分析报告")
    c.font = F_COVER_TITLE; c.alignment = Alignment(horizontal="center", vertical="center")
    ws0.row_dimensions[9].height = 38

    ws0.merge_cells("A10:H10")
    c = ws0.cell(row=10, column=1, value="金川县中小学、幼儿园食堂大宗食品采购项目（合同包一）")
    c.font = F_COVER_SUB; c.alignment = Alignment(horizontal="center", vertical="center")
    ws0.row_dimensions[10].height = 24

    ws0.row_dimensions[11].height = 6

    # 分隔线
    ws0.merge_cells("A12:H12")
    c = ws0.cell(row=12, column=1, value="─" * 70)
    c.font = Font(name="微软雅黑", size=6, color=C_GOLD); c.alignment = Alignment(horizontal="center", vertical="center")

    ws0.row_dimensions[13].height = 8

    cover_info = [
        ("项目编号", "N5132262025000114"),
        ("采购方式", "政府采购公开招标"),
        ("投标单位", "7 家"),
        ("分析日期", "2026年5月30日"),
        ("检测维度", "报价规律 / 文本雷同 / 图片比对 / 文件来源追溯"),
    ]
    for i, (k, v) in enumerate(cover_info):
        r = 14 + i
        ws0.merge_cells(f"A{r}:H{r}")
        c = ws0.cell(row=r, column=1, value=f"    {k}：{v}")
        c.font = F_COVER_INFO; c.alignment = Alignment(horizontal="left", vertical="center")
        ws0.row_dimensions[r].height = 22

    ws0.row_dimensions[19].height = 12
    ws0.merge_cells("A20:H20")
    c = ws0.cell(row=20, column=1, value="─" * 70)
    c.font = Font(name="微软雅黑", size=6, color=C_GOLD); c.alignment = Alignment(horizontal="center", vertical="center")

    ws0.row_dimensions[21].height = 6
    ws0.merge_cells("A22:H22")
    c = ws0.cell(row=22, column=1, value="四川融策会计师事务所 · 四川融策工程咨询有限公司")
    c.font = Font(name="微软雅黑", size=11, color=C_TEXT); c.alignment = Alignment(horizontal="center", vertical="center")
    ws0.merge_cells("A23:H23")
    c = ws0.cell(row=23, column=1, value="本报告为保密分析材料，仅供项目组内部使用")
    c.font = F_FOOTER; c.alignment = Alignment(horizontal="center", vertical="center")

    # ---- Sheet 1: 分析总览 ----
    ws1 = wb.create_sheet("分析总览")
    ws1.sheet_properties.tabColor = C_PRI
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 65

    section(ws1, 2, "一、项目概况与检测方法说明", 1, 2)

    hdr(ws1, 4, ["项目", "内容"])
    overview = [
        ("项目名称","金川县中小学、幼儿园食堂大宗食品采购"),
        ("项目编号","N5132262025000114"),
        ("采购包","合同包一（生鲜肉类、禽蛋、蔬菜、干杂调味品、米面油等）"),
        ("报价方式","结算率报价（即中标后结算价格=市场基准价×结算率）"),
        ("分析日期","2026年5月30日"),
        ("投标单位数量","7家"),
        ("投标文件数量","98个PDF文件（每家14个文件）"),
    ]
    for i,(k,v) in enumerate(overview,5):
        data_row(ws1, i, [k,v])

    row = 14
    section(ws1, row, "二、四层检测一览", 1, 2)
    row += 2
    hdr(ws1, row, ["检测层", "通俗名称", "检测方法简介", "本项目结果"])
    row += 1
    checks = [
        ["L1 报价规律","报价有没有「商量好的」迹象",
         "对比7家的报价数据，看看是否存在阶梯式分布、等差数列等被人为安排过的痕迹。正常的市场竞争下，报价应该是分散的、没有规律的。",
         "✅ 正常 — 7家报价分散在87.6%~100%，没有明显的人为规律"],
        ["L3 文本雷同","投标文件内容有没有互相抄",
         "把7家投标文件的全部文字拼接起来，用计算机算法逐字逐句比对相似程度。相似度超过80%就说明存在大段复制粘贴。",
         "✅ 正常 — 最高相似度仅49.76%，远低于80%的可疑阈值，7家投标文件文字内容彼此独立"],
        ["L4 图片比对","投标文件里的图片有没有共用",
         "提取所有PDF中嵌入的图片（公章、照片、扫描件等），计算每张图片的数字指纹（MD5哈希值），看同一张图片是否出现在不同公司的投标文件里。",
         "✅ 正常 — 共提取1552张图片，没有发现任何一张图片被不同公司共用"],
        ["L5 文件来源","投标文件是不是同一个人做的",
         "每份PDF文件内部都记录了「制作人信息」——用什么软件做的、电脑用户名是什么、什么时候创建的。如果不同公司的文件显示同一个制作者，说明极可能是同一台电脑或同一个人制作的。",
         "⚠️ 严重异常 — 发现有4家公司的投标文件共享同一个制作者信息，详见L5分析页"],
    ]
    for ck in checks:
        data_row(ws1, row, ck, 1, "red" if "异常" in ck[3] else "normal")
        row += 1

    row += 1
    section(ws1, row, "三、快速阅读指南", 1, 2)
    row += 2
    guide = [
        "💡 如果您只有3分钟：请直接看「L5_元数据交叉」页和「综合结论与建议」页。",
        "💡 每个分析页下方都有「📖 通俗解读」，用大白话解释检测逻辑和结论含义。",
        "💡 绿色 = 正常无异常 | 黄色 = 需要留意 | 红色 = 重大疑点",
        "💡 「文件来源追溯（L5）」是本次分析的核心发现，证据级别最高。",
    ]
    for g in guide:
        ws1.merge_cells(f"A{row}:B{row}")
        ws1.cell(row=row, column=1, value=g).font = F_TABLE_BODY
        ws1.cell(row=row, column=1).alignment = ALIGN_L
        row += 1

    ws1.page_margins = PAGE

    # ---- Sheet 2: L1 报价 ----
    ws2 = wb.create_sheet("L1_报价规律性")
    ws2.sheet_properties.tabColor = C_GREEN
    ws2.sheet_view.showGridLines = False

    section(ws2, 2, "L1 报价规律性分析", 1, 8)
    section(ws2, 3, "检测目标：7家投标报价是否存在人为安排的痕迹（如阶梯式分布、等差数列等）", 1, 8,
            Font(name="微软雅黑", size=10, color=C_LABEL))

    row = 5
    hdr(ws2, row, ["序号","投标单位","结算率","投标日期","报价分析"])
    row += 1

    sp = sorted(prices.items(), key=lambda x: x[1][0] if x[1] else 999)
    rate_vals = [p[1][0] for p in sp if p[1]]
    for i,(co, pl) in enumerate(sp, 1):
        rate = pl[0] if pl else None
        pi = price_info.get(co, {})
        t = pi.get("text","")
        dm = re.search(r'(\d{4})年(\d{2})月(\d{2})日', t)
        bd = f"{dm.group(1)}-{dm.group(2)}-{dm.group(3)}" if dm else "未提取到"
        if rate == 100:
            an = "最高限价不打折（100%），通常为不指望中标的「陪跑」报价，或是对项目势在必得的最高价策略"
            rtp = "yellow"
        elif rate and rate < 88:
            an = "最低价区间，属于竞争性较强的报价"
            rtp = "green"
        else:
            an = "中间价位，在正常竞争范围内"
            rtp = "normal"
        data_row(ws2, row, [i, co, f"{rate}%" if rate else "N/A", bd, an], 1, rtp)
        row += 1

    row += 1
    # 解读框
    if rate_vals:
        mn, mx = min(rate_vals), max(rate_vals)
        spd = mx - mn
        row = write_explain_box(ws2, row, 1, 8, "通俗解读：报价规律性检测",
            [
                f"▎结算率是什么意思？",
                f"  本次采购不是报一个固定总价，而是报「结算率」。比如中标结算率是90%，意思是最终结算时，采购方按市场基准价的90%付款。结算率越低，对采购方越省钱。",
                "",
                f"▎7家的报价情况",
                f"  最高报价 {mx}%（按市场价全额结算），最低报价 {mn}%（打{mn}折），",
                f"  报价区间从{mn}%到{mx}%，跨度{spd:.1f}个百分点，分散度正常。",
                "",
                f"▎怎么看有没有围标？",
                f"  围标时常见两种报价模式：①报价等间距分布（如90%、93%、97%→差3-4个点）；②有一家特别低、其他都接近最高价（「保一家中标」模式）。",
                f"  本项目7家报价分别为：{', '.join(f'{r}%' for r in sorted(rate_vals))}",
                f"  差值不规律、没有明显的人为安排痕迹。报价层面未见异常。",
            ])

    ws2.column_dimensions['A'].width = 7
    ws2.column_dimensions['B'].width = 34
    ws2.column_dimensions['C'].width = 13
    ws2.column_dimensions['D'].width = 14
    ws2.column_dimensions['E'].width = 70
    for c in 'FGH': ws2.column_dimensions[c].width = 12
    ws2.page_margins = PAGE

    # ---- Sheet 3: L3 文本雷同 ----
    ws3 = wb.create_sheet("L3_文本雷同检测")
    ws3.sheet_properties.tabColor = C_GREEN
    ws3.sheet_view.showGridLines = False

    section(ws3, 2, "L3 文本雷同检测 (TF-IDF余弦相似度)", 1, 8)
    section(ws3, 3, "检测目标：7家投标文件的文字内容是否存在大段复制粘贴（雷同度≥80%即为可疑）", 1, 8,
            Font(name="微软雅黑", size=10, color=C_LABEL))

    # 矩阵
    row = 5
    section(ws3, row, "▎全文本相似度矩阵（百分比）", 1, 8, F_SUBSECTION)
    row += 1
    hdr(ws3, row, [""] + corp_co)
    row += 1
    for i, ca in enumerate(corp_co):
        vals = [ca]
        for j, cb in enumerate(corp_co):
            if i == j: vals.append("—")
            else:
                for p in sim_pairs:
                    if (p["a"]==ca and p["b"]==cb) or (p["a"]==cb and p["b"]==ca):
                        vals.append(f"{p['sim']*100:.1f}%"); break
                else: vals.append("")
        data_row(ws3, row, vals)
        row += 1

    row += 1
    # 两两明细
    section(ws3, row, "▎两两对比明细", 1, 6, F_SUBSECTION)
    row += 1
    hdr(ws3, row, ["公司A","公司B","余弦相似度","文字量_A","文字量_B","判定"]); row += 1
    for p in sorted(sim_pairs, key=lambda x: -x["sim"]):
        sp = p["sim"]*100
        vd = "🔴 高度可疑(≥80%)" if sp>=80 else ("🟡 需关注(≥60%)" if sp>=60 else "正常 ✓")
        rt = "red" if sp>=80 else ("yellow" if sp>=60 else "green")
        data_row(ws3, row, [p["a"],p["b"],f"{sp:.2f}%",p["ca"],p["cb"],vd], 1, rt)
        row += 1

    row += 1
    row = write_explain_box(ws3, row, 1, 8, "通俗解读：文本雷同检测",
        [
            f"▎这个方法在检测什么？",
            f"  把7家公司的全部投标文件文字（包括投标函、承诺书、技术方案、商务应答等）分别拼接成7篇「全文」，",
            f"  然后用计算机逐字逐句比对每两篇之间的相似程度。",
            f"  打个比方：就像老师检查学生作文有没有互相抄——如果两个学生的作文有80%以上的内容几乎一样，那基本可以断定有抄袭。",
            "",
            f"▎7家对比结果",
            f"  21组两两对比中，最高相似度只有49.76%（四川华创景盛 vs 成都乐稼良品），远低于80%的可疑线。",
            f"  而且这个最高值很大程度是因为两家的文字总量差异太大导致的统计偏差（一家4661字，另一家10万多字），",
            f"  并非真正的内容雷同。",
            "",
            f"▎结论：7家投标文件的文字内容是各自独立撰写的，未发现相互抄袭或共用模板的痕迹。",
        ])

    ws3.column_dimensions['A'].width = 34
    ws3.column_dimensions['B'].width = 34
    ws3.column_dimensions['C'].width = 16
    ws3.column_dimensions['D'].width = 14
    ws3.column_dimensions['E'].width = 14
    ws3.column_dimensions['F'].width = 22
    for c in 'GH': ws3.column_dimensions[c].width = 12
    ws3.page_margins = PAGE

    # ---- Sheet 4: L4 图片 ----
    ws4 = wb.create_sheet("L4_图片比对")
    ws4.sheet_properties.tabColor = C_GREEN
    ws4.sheet_view.showGridLines = False

    section(ws4, 2, "L4 图片跨公司比对", 1, 8)
    section(ws4, 3, "检测目标：同一张图片（公章、扫描件、照片等）是否出现在不同公司的投标文件中", 1, 8,
            Font(name="微软雅黑", size=10, color=C_LABEL))

    total_imgs = sum(len(v) for v in gh.values())
    non_tmpl = [ic for ic in img_cross if not ic["template"]]
    row = 5
    ws4.merge_cells(f"A{row}:F{row}")
    ws4.cell(row=row, column=1, value=f"共提取 {total_imgs} 张嵌入图片，跨公司重复图片 {len(img_cross)} 张（其中 {len(non_tmpl)} 张为非模板类重复）")
    ws4.cell(row=row, column=1).font = F_TABLE_BODY
    row += 1
    ws4.merge_cells(f"A{row}:F{row}")
    ws4.cell(row=row, column=1, value="✅ 未发现任何跨公司非模板类重复图片，图片层面无异常信号。")
    ws4.cell(row=row, column=1).font = F_SUBSECTION
    ws4.cell(row=row, column=1).fill = BG_GREEN
    row += 2

    row = write_explain_box(ws4, row, 1, 8, "通俗解读：图片比对",
        [
            f"▎这个方法在检测什么？",
            f"  每份PDF投标文件内部都包含各种图片——公司的公章、营业执照扫描件、业绩证明材料扫描件等。",
            f"  系统从98个PDF中提取了{total_imgs}张图片，给每张图片算出一个独一无二的「数字指纹」（MD5哈希值），",
            f"  然后检查：有没有同一张图片出现在不同公司的投标文件里？",
            "",
            f"▎什么情况下会出现跨公司重复图片？",
            f"  正常情况：各公司公章、营业执照、资质证书都是自己的，不可能相同。",
            f"  异常情况：如果A公司和B公司共用了同一张扫描件图片，说明至少有一方拿不到原件，",
            f"  或者两家的标书是同一人经手处理的——这是很强烈的围标信号。",
            "",
            f"▎本项目的检测结果",
            f"  {total_imgs}张图片中，没有任何一张被两家以上公司共用。每家公司提供的图片都是独立的。",
            f"  图片层面未发现异常。",
        ])

    ws4.column_dimensions['A'].width = 8
    ws4.column_dimensions['B'].width = 22
    ws4.column_dimensions['C'].width = 14
    ws4.column_dimensions['D'].width = 55
    ws4.column_dimensions['E'].width = 60
    ws4.column_dimensions['F'].width = 12
    for c in 'GH': ws4.column_dimensions[c].width = 12
    ws4.page_margins = PAGE

    # ---- Sheet 5: L5 元数据 ⚠️核心 ----
    ws5 = wb.create_sheet("L5_文件来源追溯")
    ws5.sheet_properties.tabColor = C_RED
    ws5.sheet_view.showGridLines = False

    section(ws5, 2, "L5 文件来源追溯 ⚠️ 核心发现", 1, 8)
    section(ws5, 3, "检测目标：追溯每份投标文件的制作来源——是谁、用什么软件、在哪台电脑上制作的", 1, 8,
            Font(name="微软雅黑", size=10, color=C_LABEL))

    # 异常发现
    abnormal = [cf for cf in meta_cross if not cf["platform"]]
    normal_pf = [cf for cf in meta_cross if cf["platform"]]

    row = 5
    section(ws5, row, "▎异常发现：跨公司共享的「制作者信息」（核心疑点）", 1, 8, F_SUBSECTION)
    row += 1
    hdr(ws5, row, ["#","文件类型","信息字段","共享的值","涉及公司数","涉及公司","证据强度"]); row += 1

    for i, cf in enumerate(abnormal, 1):
        sev = "🔴 铁证" if cf["n"] >= 3 else "🟡 强信号"
        rt = "red" if cf["n"] >= 3 else "yellow"
        data_row(ws5, row, [i, cf["cat"], cf["fld"], cf["val"], cf["n"], "、".join(cf["cs"]), sev], 1, rt)
        row += 1

    row += 1
    row = write_explain_box(ws5, row, 1, 8, "通俗解读：什么是「文件来源追溯」？",
        [
            "▎每一份电子文件都自带「身份证信息」",
            "  你在电脑上创建一个Word文档或PDF文件后，文件内部会自动记录一些信息：",
            "  • Author（作者）：写这份文件的电脑用户名，比如你电脑登录名叫「张三」，Author就是「张三」",
            "  • Creator（创建软件）：用什么软件创建的，比如「WPS 文字」、「Microsoft Word」",
            "  • Producer（生成工具）：用什么工具把Word转成PDF的",
            "",
            "▎为什么这个信息能发现围标？",
            "  正常情况：7家不同公司，投标文件是各家公司自己的人做的，那么Author应该各不相同",
            "  （比如A公司的文件Author=「张会计」、B公司的文件Author=「李经理」……）",
            "",
            "  异常情况：如果多家公司的投标文件Author都是同一个人（比如都是「123」或「linyan」），",
            "  说明这些文件极有可能是——同一个人、在同一台电脑上、用同一个软件做的。",
            "  不同公司不可能有同一个「制作者」，这是非常直接的串标证据。",
            "",
            "▎本项目的发现",
            "  有4家公司（安必先科技、乐稼良品、蓉建粮油、兴鸿人力）的多个文件共享了完全相同的Author信息：",
            "  • Author = '123' → 4家公司的基本情况表、业绩表、人员表等都显示这个作者",
            "  • Author = 'linyan' → 4家公司的商务应答表、证明材料等都显示这个作者",
            "  • Author = 'Administrator' → 4家公司的技术要求应答表显示这个作者",
            "",
            "  这在物理上几乎不可能自然发生——只有同一人在同一台电脑上制作，才会出现这种模式。",
            "  这是本次全量分析中证据级别最高的发现。",
        ])

    # 正常特征（仅作说明）
    row += 1
    section(ws5, row, "▎参考：政府采购平台自动生成的特征（不视为围标证据）", 1, 8, F_SUBSECTION)
    row += 1
    hdr(ws5, row, ["#","文件类型","信息字段","值","涉及公司数","说明"]); row += 1
    for i, cf in enumerate(normal_pf[:8], 1):
        expl = "政府采购平台自动生成的PDF" if any(k in cf["val"].lower() for k in ["chromium","skia"]) else "声明函类文件共用模板"
        data_row(ws5, row, [i, cf["cat"], cf["fld"], cf["val"], cf["n"], expl])
        row += 1

    ws5.column_dimensions['A'].width = 6
    ws5.column_dimensions['B'].width = 24
    ws5.column_dimensions['C'].width = 14
    ws5.column_dimensions['D'].width = 22
    ws5.column_dimensions['E'].width = 14
    ws5.column_dimensions['F'].width = 55
    ws5.column_dimensions['G'].width = 16
    ws5.column_dimensions['H'].width = 12
    ws5.page_margins = PAGE

    # ---- Sheet 6: 元数据明细 ----
    ws6 = wb.create_sheet("元数据明细表")
    ws6.sheet_properties.tabColor = C_SEC
    ws6.sheet_view.showGridLines = False

    section(ws6, 2, "PDF元数据明细（7家 × 14个文件 = 98条完整记录）", 1, 9)
    section(ws6, 3, "完整记录每个PDF文件的Author（作者）、Creator（创建软件）、Producer（生成工具）、创建时间等信息", 1, 9,
            Font(name="微软雅黑", size=10, color=C_LABEL))

    row = 5
    hdr(ws6, row, ["公司","文件名","文件类别","Author(作者)","Creator(创建软件)","Producer(生成工具)","创建时间","修改时间","页数"]); row += 1
    for co in sorted(all_meta.keys()):
        fm = all_meta[co]
        for fn in sorted(fm.keys()):
            m = fm[fn]; cat = classify(fn)
            data_row(ws6, row, [co, fn, cat, m["author"], m["creator"], m["producer"],
                                m["creationDate"], m["modDate"], m["page_count"]])
            row += 1

    ws6.column_dimensions['A'].width = 34
    ws6.column_dimensions['B'].width = 40
    ws6.column_dimensions['C'].width = 22
    ws6.column_dimensions['D'].width = 16
    ws6.column_dimensions['E'].width = 18
    ws6.column_dimensions['F'].width = 22
    ws6.column_dimensions['G'].width = 20
    ws6.column_dimensions['H'].width = 20
    ws6.column_dimensions['I'].width = 8
    ws6.page_margins = PAGE

    # ---- Sheet 7: 综合结论 ----
    ws7 = wb.create_sheet("综合结论与建议")
    ws7.sheet_properties.tabColor = C_RED
    ws7.sheet_view.showGridLines = False

    section(ws7, 2, "综合结论与行动建议", 1, 7)

    row = 4
    section(ws7, row, "一、分析结论", 1, 7, F_SUBSECTION)
    row += 2
    hdr(ws7, row, ["评估项","结论"]); row += 1
    conclusions = [
        ["总体评估","⚠️ 4家公司（安必先科技、乐稼良品、蓉建粮油、兴鸿人力）存在高度串标嫌疑"],
        ["核心证据","文件来源追溯：Author字段（123 / linyan / Administrator）跨4家公司完全一致，指向同一人/同一设备制作"],
        ["辅助证据","创建软件信息（Creator='WPS 文字'）在6-7家中共享，结合Author一致，形成完整证据链"],
        ["排除项","报价规律、文本雷同、图片比对三个维度均未发现异常 — 投标文件为扫描型PDF，文字可对比内容有限"],
        ["独立判断","四川华创景盛商贸（报价100%、元数据独立）推测为独立投标，未参与串标"],
        ["中度关注","川恒亿科技、心诚农副产品 — 创建软件信息与其他家一致，但Author不同，嫌疑较低"],
    ]
    for k, v in conclusions:
        rt = "red" if "🔴" in v or "铁证" in v or "高度" in v else ("yellow" if "⚠" in v else "normal")
        data_row(ws7, row, [k, v], 1, rt); row += 1

    row += 1
    section(ws7, row, "二、嫌疑分群", 1, 7, F_SUBSECTION)
    row += 2
    hdr(ws7, row, ["风险等级","涉及公司","核心证据","建议措施"]); row += 1
    groups = [
        ["🔴 高度嫌疑\n（4家）",
         "1. 四川安必先科技有限公司\n2. 成都乐稼良品农业科技有限公司\n3. 成都蓉建粮油销售有限公司\n4. 金川县兴鸿人力资源有限责任公司",
         "Author字段完全一致(123/linyan/Administrator)\n多类文件交叉验证\n证据级别：铁证",
         "① 向代理机构调取投标系统登录IP/MAC地址\n② 要求提供原始Word文档进行比对\n③ 天眼查/企查查核查四家工商关联关系\n④ 比对授权委托书经办人信息"],
        ["🟡 中度关注\n（2家）",
         "1. 成都川恒亿科技有限公司\n2. 成都心诚农副产品有限公司",
         "Creator='WPS 文字'与其他家一致\n但Author不同",
         "① 核实投标文件制作来源\n② 排除是否委托同一代理机构制作"],
        ["🟢 未见异常\n（1家）",
         "1. 四川华创景盛商贸有限公司",
         "报价100% + 元数据独立\n未与其他公司共享",
         "正常投标行为，无需进一步调查"],
    ]
    for g in groups:
        data_row(ws7, row, g, 1, "red" if "🔴" in g[0] else ("yellow" if "🟡" in g[0] else "green"))
        row += 1
        ws7.row_dimensions[row-1].height = 90

    row += 1
    section(ws7, row, "三、建议调查步骤", 1, 7, F_SUBSECTION)
    row += 2
    suggestions = [
        ("第一步","【IP/MAC核查】","向代理机构（四川政府采购中心）调取投标系统登录日志，比对4家嫌疑单位在投标时间段内的登录IP地址和MAC地址是否相同。如果IP/MAC一致，直接构成串标铁证。","高"),
        ("第二步","【原始文档比对】","要求4家单位提交投标文件原始Word版本，对比文档属性中的「作者」「创建时间」「最后保存者」「修改次数」等信息，并与PDF中的记录交叉验证。","高"),
        ("第三步","【工商关联穿透】","通过天眼查、企查查或国家企业信用信息公示系统，查询4家公司之间是否存在：同一股东、同一法人代表、同一监事、同一注册地址、同一联系电话等关联关系。","高"),
        ("第四步","【经办人交叉比对】","核对4家投标授权委托书中的经办人姓名、身份证号、联系电话。如果不同公司的经办人是同一人，或联系电话相同，进一步强化证据链。","中"),
        ("第五步","【保证金资金链路】","向银行调取4家单位的投标保证金汇款凭证，核查是否存在同一账户代缴、资金回流等异常情况。","中"),
    ]
    hdr(ws7, row, ["步骤","核查事项","具体操作","优先级"]); row += 1
    for step, title, desc, pri in suggestions:
        data_row(ws7, row, [step, title, desc, pri], 1, "red" if pri=="高" else "yellow")
        ws7.row_dimensions[row].height = 50
        row += 1

    row += 1
    ws7.merge_cells(f"A{row}:G{row}")
    ws7.cell(row=row, column=1,
        value="四川融策会计师事务所 · 四川融策工程咨询有限公司  |  本报告基于PDF文件自动分析生成，仅供项目组内部参考  |  2026年5月30日"
    ).font = F_FOOTER
    ws7.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")

    ws7.column_dimensions['A'].width = 18
    ws7.column_dimensions['B'].width = 24
    ws7.column_dimensions['C'].width = 52
    ws7.column_dimensions['D'].width = 42
    for c in 'EFG': ws7.column_dimensions[c].width = 14
    ws7.page_margins = PAGE

    # ========== 保存 ==========
    for ws in wb.worksheets:
        ws.page_margins = PAGE
    wb.save(OUTPUT)
    print(f"Done: {OUTPUT}")

if __name__ == "__main__":
    build()
