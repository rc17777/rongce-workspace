# -*- coding: utf-8 -*-
"""重新精确提取五份复核底稿的全部列(含复核得分/偏差/备注等)"""
import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(r"C:\Users\scrccpa\Desktop\报告\2.部门预算项目绩效自评复核报告20260722")

def vb(val): return "" if val is None else str(val).strip()

for d in sorted(ROOT.iterdir()):
    if not d.is_dir(): continue
    name = d.name[:8]
    xls = list(d.glob("*.xlsx"))
    if not xls: continue
    print(f"\n{'='*60}")
    print(f"=== {d.name[:50]} ===")
    wb = load_workbook(xls[0], data_only=True)
    for s in wb.worksheets:
        st = s.title
        # Skip non-project sheets
        if st in ("Sheet4","Sheet5","Sheet6","Sheet1","Sheet2","Sheet3","目标完成，偏离度","自评复核报告表"): continue
        # Print ALL columns for rows with content
        rows = [[vb(c.value) for c in row] for row in s.iter_rows(min_row=1, max_row=min(s.max_row, 20))]
        # Print header row fully
        header = rows[0] if rows else []
        # Find rows with real content (not just dashes or empty)
        content_rows = [r for r in rows if any(v not in ("","","——","——","附件1","附件2","附件3","附件4","附件5","附件6","附件7","附件8","附件9","附件10") for v in r) and not r[0].startswith("附件")]
        for i,r in enumerate(content_rows):
            # Print all columns for each row
            non_empty = [(j,v) for j,v in enumerate(r) if v]
            if non_empty:
                col_str = " | ".join(f"[{j}]{v}" for j,v in non_empty)
                print(f"  {st[:25]:>25} R{i+1}: {col_str}")
            else:
                print(f"  {st[:25]:>25} R{i+1}: (empty)")
        print()
