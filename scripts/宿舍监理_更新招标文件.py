"""更新Excel — 补充招标文件分析Sheet"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.load_workbook(r'C:\Users\scrccpa\Desktop\宿舍监理审计分析报告.xlsx')

hdr_f = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
sub_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yel_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
grn_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
title_f = Font(name='微软雅黑', bold=True, size=13, color='2F5496')
bold_f = Font(name='微软雅黑', bold=True, size=11)
norm_f = Font(name='微软雅黑', size=11)
sml_f = Font(name='微软雅黑', size=10)
thin_b = Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))
ca = Alignment(horizontal='center', vertical='center', wrap_text=True)
wa = Alignment(vertical='center', wrap_text=True)

def cell(ws, r, col, v, font=norm_f, fill=None, align=ca):
    c = ws.cell(row=r, column=col, value=v)
    c.font = font; c.alignment = align; c.border = thin_b
    if fill: c.fill = fill
    return c

def hdr(ws, r, headers):
    for i,h in enumerate(headers):
        cell(ws, r, i+1, h, hdr_f, hdr_fill)

# ============ Sheet: 招标文件分析 ============
ws = wb.create_sheet('招标文件审查')

cell(ws, 1, 1, '宿舍监理项目 招标文件关键条款审查', title_f, align=wa)
ws.merge_cells('A1:G1')
cell(ws, 2, 1, '模板: 四川省2021版标准监理招标文件 | 152页 | 编制: 刘珂/范鹏 | 圣弘建设 | 2025年10月28日 | 最高限价946,000元 | PDF 1.7 (Aspose.PDF)', sml_f, align=wa)
ws.merge_cells('A2:G2')

# Part A: Basic Info
r = 4
cell(ws, r, 1, 'A. 评分体系全览（满分100分）', bold_f, sub_fill, wa)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)

r = 5
hdr(ws, r, ['评分板块','分值','评分因素','分项分值','评分标准摘要','是否为四川省标准模板','审查结论'])

scoring = [
    ['资信业绩','30分','企业综合实力/诚信','3分',
     '诚信分100-95=3分, 95-90=2分, 90-85=1分, 85以下=0分\n来源: 四川省住建厅网站, 投标截止前10日内打印',
     '✅ 标准','🟢 合理'],
    ['','','企业类似业绩','8分',
     '除资格业绩外, 近3年每1个类似项目(≥14000㎡+≥6200万元房建监理)得4分, 最多8分\n(即最多2个加分业绩 + 1个资格业绩 = 3个)',
     '✅ 标准','🟢 合理 (≥本项目规模, 不是<<小于>>)'],
    ['','','总监理工程师','7分',
     '中级职称2分, 高级及以上4分\n已完成1个类似业绩得3分, 最多3分',
     '✅ 标准','🟢 合理'],
    ['','','现场监理人员','12分',
     '房建专监: 注册证2分+中级+1/高级+2→最多4分\n安装专监: 同上, 最多4分\n造价专监: 一级造价师2分\n安全专监: 注册监理工程师2分\n以上4岗不可兼任',
     '✅ 标准','🟢 合理'],
    ['监理大纲','30分','监理依据/目标','2分','目标与项目建设总体目标的符合性/完整性(1-2分)','✅ 标准','🟢 主观评分项(弹性)'],
    ['','','机构设置/岗位职责','2分','机构设置合理/职责分工明确(1-2分)','✅ 标准','🟢 主观评分项'],
    ['','','工作程序/方法/制度','3分','程序严谨/方法合理/制度可操作(1-3分)','✅ 标准','🟢 主观评分项'],
    ['','','质量/进度/造价/安全/环保措施','8分','与项目实际相符/针对性强/可行(4-8分)','✅ 标准','🟡 权重最高, 主观空间大'],
    ['','','合同/信息管理方案','4分','合理/可行/创新(2-4分)','✅ 标准','🟢 主观评分项'],
    ['','','组织协调内容/措施','4分','完整/合理/可行(2-4分)','✅ 标准','🟢 主观评分项'],
    ['','','工作重点/难点分析','4分','重点难点准确/分析清晰/理解透彻(2-4分)','✅ 标准','🟢 主观评分项'],
    ['','','合理化建议','3分','认识项目特点/建议可行(1-3分)','✅ 标准','🟢 主观评分项'],
    ['投标报价','40分','偏差率','40分',
     '等于评标基准价=满分\n每高1%扣1分\n每低1%扣0.5分\n中间值线性内插, 保留两位小数',
     '✅ 标准','🟢 扣分不对称(高扣多/低扣少), 鼓励低价'],
]

for i, row_data in enumerate(scoring):
    for j, val in enumerate(row_data):
        if j == 0:
            cell(ws, 6+i, j+1, val, bold_f if val else norm_f)
        elif j in [5,6]:
            cell(ws, 6+i, j+1, val, norm_f, fill=grn_fill if '合理' in str(val) else yel_fill if '主观' in str(val) else None)
        elif j == 4:
            cell(ws, 6+i, j+1, val, sml_f, align=wa)
        elif j == 1:
            cell(ws, 6+i, j+1, val, bold_f)
        else:
            cell(ws, 6+i, j+1, val, norm_f)

# Part B: Review Points
r2 = 20
cell(ws, r2, 1, 'B. 招标文件合规性审查要点', bold_f, sub_fill, wa)
ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=7)

r2 = 21
hdr(ws, r2, ['审查维度','审查要点','招标文件规定','是否合规','风险等级','分析','建议'])

compliance = [
    ['资质设置','是否设置过高门槛','房屋建筑工程监理乙级及以上\n(最低乙级, 无甲级要求)','✅ 合规','🟢',
     '乙级即可承接本工程监理(投资6200万, 非超高/超大项目)。乙级资质覆盖面广, 不会限制竞争。',
     '—'],
    ['业绩门槛','是否≥本项目规模','资格: 1个≥14000㎡+≥6200万元房建监理\n加分: 最多2个同类(与资格相同标准)','✅ 合规','🟢',
     '业绩门槛与项目本身规模一致, ≤而非>项目规模, 符合"不得高于"标准模板规定。最多3个业绩(含资格)在允许范围内。',
     '—'],
    ['奖项/认证','是否存在特定奖项加分','无任何奖项加分条款\n其他评分因素=0分','✅ 合规','🟢',
     '采用四川省标准模板, 未设置特定行业/区域奖项。诚信分来自省住建厅标准化系统, 非个别奖项。',
     '—'],
    ['特定行业/区域限制','是否排斥外省企业','企业注册地不在四川省→需提供入川信息网页截图\n(标准要求, 非排斥)','✅ 合规','🟢',
     '入川截图是住建部强制性信息公示要求, 不是限制条件。未设置本地办公面积/纳税/社保等规模条件。',
     '—'],
    ['评标基准价','是否存在操纵空间','B方式: 有效投标报价算术平均\nS=(a1+...+an)/n\n无剔除极值机制','⚠️ 理论风险','🟡',
     '算术平均法在投标人少时易被操纵。如多家联手报高价, 可拉高基准价→高价者反而得分高。但22家投标人的大样本降低了此风险。',
     '关注是否有投标人集群集体偏离基准价。'],
    ['报价扣分机制','是否对称','每高1%扣1分\n每低1%扣0.5分\n(高扣多, 低扣少)','✅ 合规','🟢',
     '不对称扣分鼓励低价竞争, 符合公共财政节约原则。低报价策略更优.但可能导致恶意低价。',
     '关注异常低价(低于成本评审)'],
    ['监理大纲评分','主观分是否过大','30分上限, 8项分散评分\n最低得分=各分项下限之和=18分\n最高=各分项上限=30分\n评委可浮动范围=12分','⚠️ 需关注','🟡',
     '8项评分因素占30分, 主观分数较大(12分可浮动空间)。评委自由裁量权可用于"定向打分"。但22家投标人、至少5位评委、去掉最高最低→降低了定向操作空间。',
     '调取各评委原始打分表, 检查是否有评委给特定投标人每项都打上限/下限。'],
    ['否决投标条款','是否存在模糊条款','形式评审: 报价唯一/签字盖章/格式\n资格评审: 资质/财务/业绩/信誉/人员\n响应性: 报价/内容/工期/质量/保证金等','✅ 明确','🟢',
     '否决条款全部为明确的客观标准, 无模糊/主观否决条款。',
     '—'],
    ['保证金/资金','是否限制缴纳方式','10000元, 三种方式: 现金(基本账户)/保证保险/银行保函\n全部通过德阳交易平台在线缴纳','✅ 合规','🟢',
     '必须从基本账户缴纳, 可追溯缴纳来源。三种方式灵活, 不限制投标人选择。',
     'L9保证金缴纳记录可向交易中心调取'],
    ['总监在建项目','是否不合理限制','拟任总监在同一时间可在其他≤1个项目中担任总监\n(即可同时监理2个项目)','✅ 合理','🟢',
     '允许总监有1个在建项目, 不要求"无在建", 比很多项目更宽松。增加了投标人池。',
     '—'],
]

for i, row_data in enumerate(compliance):
    for j, val in enumerate(row_data):
        if j == 4:
            fill = yel_fill if '理论' in str(val) or '需关注' in str(val) else grn_fill
            cell(ws, 22+i, j+1, val, bold_f, fill)
        elif j in [5,6]:
            cell(ws, 22+i, j+1, val, sml_f, align=wa)
        elif j == 3:
            cell(ws, 22+i, j+1, val, norm_f, fill=grn_fill)
        elif j == 1:
            cell(ws, 22+i, j+1, val, bold_f, align=wa)
        elif j == 0:
            cell(ws, 22+i, j+1, val, bold_f)
        else:
            cell(ws, 22+i, j+1, val, norm_f, align=wa if j==2 else ca)

# Part C: Summary
r3 = 33
cell(ws, r3, 1, 'C. 审查结论', bold_f, sub_fill, wa)
ws.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=7)

conclusions = [
    '总体评价: 招标文件使用四川省2021版标准监理招标文件模板, 整体合规, 未发现明显倾向性或排斥性条款。',
    '优势: (1)奖项/特定行业业绩/地区限制均未设置;(2)投标报价权重40%合理,扣分机制鼓励低价;(3)总监可同时有1个在建项目,投标人池充足。',
    '⚠️ 需关注的潜在风险:',
    '  ① 评标基准价B方式(算术平均)在投标人样本小时可能被操纵。本项目22家降低了此风险。',
    '  ② 监理大纲30分全部为弹性分值结构,评委自由裁量空间最大12分。需调取原始评委打分表检查分布。',
    '  ③ 报价扣分不对称(高扣多/低扣少),鼓励低价策略→可能导致恶意低价抢标→低于成本评审的重要性增加。',
    '结论: 招标文件本身无明显为特定投标人"量身定制"的痕迹。审计重点应放在评审执行过程(评委打分一致性)和投标人关联关系(L8/L9),加上已完成的TF-IDF分析(L3)。'
]

for i, txt in enumerate(conclusions):
    cell(ws, 34+i, 1, txt, norm_f, align=wa)
    ws.merge_cells(start_row=34+i, start_column=1, end_row=34+i, end_column=7)
    ws.row_dimensions[34+i].height = 28

ws.column_dimensions['A'].width = 16
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 40
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 30
ws.column_dimensions['F'].width = 36
ws.column_dimensions['G'].width = 30

for i in range(22, 22+len(compliance)):
    ws.row_dimensions[i].height = 85

# Also update 发现清单 to add bidding doc finding
ws1 = wb['审计发现清单']
last_row = ws1.max_row + 1

cell(ws1, last_row, 1, ws1.cell(row=last_row-1, column=1).value + 1 if isinstance(ws1.cell(row=last_row-1, column=1).value, int) else 9, norm_f)
cell(ws1, last_row, 2, '招标文件合规性审查', grn_fill)
cell(ws1, last_row, 3, '🟢 正常', bold_f, grn_fill)
cell(ws1, last_row, 4, '招标文件使用四川省2021版标准监理招标文件模板。评分体系(资信30/大纲30/报价40)为标准配置。未设置特定奖项/行业业绩/地区限制。资格要求(乙级+1个同类业绩)合理, 未≥项目规模。报价评标基准价为B方式(有效投标算术平均),扣分不对称(高扣多低扣少), 鼓励低价竞争。监理大纲30分中12分弹性空间需关注评委是否定向打分。', norm_f, align=wa)
cell(ws1, last_row, 5, '圣弘建设(代理)/四川护理职业学院(招标人)')
cell(ws1, last_row, 6, '招标文件152页, 四川省2021版标准模板\n评分: 资信30+大纲30+报价40\n限价: 946,000元\n基准: 算术平均(B方式)\n偏差: +1%扣1, -1%扣0.5', sml_f, align=wa)
cell(ws1, last_row, 7, '调取各评委原始打分表;关注监理大纲评分一致性;核查低于成本投标', norm_f, align=wa)
cell(ws1, last_row, 8, '招标文件')
ws1.row_dimensions[last_row].height = 100

wb.save(r'C:\Users\scrccpa\Desktop\宿舍监理审计分析报告.xlsx')
print('Updated: 6 sheets with bidding document analysis')
