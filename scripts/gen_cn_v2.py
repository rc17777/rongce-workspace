# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
red_fill = PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid')
yellow_fill = PatternFill(start_color='FFF8E0', end_color='FFF8E0', fill_type='solid')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
normal_font = Font(name='Arial', size=10)
title_font = Font(name='Arial', bold=True, size=14)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
wrap_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(wrap_text=True, vertical='center', horizontal='center')

def sc(ws, r, c, font=normal_font, fill=white_fill, align=wrap_align, border=thin_border):
    cell = ws.cell(row=r, column=c)
    cell.font = font; cell.fill = fill; cell.alignment = align; cell.border = border

print('Script loaded OK')
