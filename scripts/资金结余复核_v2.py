#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
融策审计 - 资金来源与预算执行指标复核
基于已提取数据
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

# ========== 从附表1(财务决算表)和附表2(审核汇总表)提取的资金数据 ==========

data = {
    'S220安羌镇至茸安乡段灾害修复整治工程': {
        # 附表1 财务决算表 02表
        '基建拨款': 26900000.00,    # C1=26,900,000 元
        '应付款合计': 289662.53,    # C1=289,662.53
        '资金来源合计': 27189662.53, # C1=27,189,662.53
        
        '基本建设支出(待核销)': 24917712.92, # C3=24,917,712.92 元（已支出）
        '货币资金(结余)': 2271949.61,      # C3=2,271,949.61 元
        '资金占用合计': 27189662.53,       # C3=27,189,662.53
        
        # 附表2 审核汇总表 01表
        '批准概算': 39922900.00,    # C3=39,922,900 元（注意：含征地拆迁）
        '送审投资': 24917712.92,   # C5=24,917,712.92
        '审定投资': 24917712.92,   # C7=24,917,712.92
        
        # 审核报告数据（已提取）
        '送审金额_报告': 29788228.07,  # 29,788,228.07元
        '审定金额_报告': 24841922.02,  # 24,841,922.02元
        '已支付': 24628050.39,        # 24,628,050.39元
        '概算总投资': 20260000.00,    # 2,026万（报告口径，不含征地拆迁？）
        
        # 二类费用明细（元）
        '二类费用合计': 2445684.33,
    },
    'S452垮沙乡至柯河乡段灾害修复整治工程': {
        '基建拨款': 61400000.00,
        '应付款合计': 0.00,          # 02表中为0
        '资金来源合计': 61400000.00,
        
        '基本建设支出(待核销)': 49122735.12,
        '货币资金(结余)': 12277264.88,
        '资金占用合计': 61400000.00,
        
        '批准概算': 1010000.00,      # 101万？这个数字存疑（审核汇总表R4的C3=1,010,000）
        '送审投资': 49122735.12,
        '审定投资': 49122735.12,
        
        '送审金额_报告': 56535896.00,  # 56,535,896元
        '审定金额_报告': 50177672.00,  # 50,177,672元
        '已支付': 49122735.12,         # 49,122,735.12元
        '概算总投资': 20260000.00,     # 2,026万
        
        '二类费用合计': 3792683.20,
    }
}

print("="*70)
print("  资金来源与预算执行指标复核")
print(f"  {'─'*50}")
print("  复核基准日：2026-06-24")
print("="*70)

for proj, d in data.items():
    print(f"\n{'#'*70}")
    print(f"#  {proj}")
    print(f"{'#'*70}")
    
    # ===== 指标1：资金来源构成 =====
    print(f"\n【1. 资金来源构成】")
    print(f"{'─'*50}")
    
    zj = d.get('基建拨款', 0) / 10000
    yfk = d.get('应付款合计', 0) / 10000
    total_src = d['资金来源合计'] / 10000
    
    print(f"  基建拨款: {zj:>10.2f}万元 ({zj/total_src*100:.1f}%)")
    print(f"  应付款:   {yfk:>10.2f}万元 ({yfk/total_src*100:.1f}%)")
    print(f"  {'─'*50}")
    print(f"  资金来源合计: {total_src:>10.2f}万元")
    
    # ===== 指标2：资金去向 =====
    print(f"\n【2. 资金去向（资金占用）】")
    print(f"{'─'*50}")
    
    zc = d['基本建设支出(待核销)'] / 10000
    jy = d['货币资金(结余)'] / 10000
    total_use = d['资金占用合计'] / 10000
    yf = d['已支付'] / 10000
    
    print(f"  基本建设支出(已支付): {zc:>10.2f}万元 ({zc/total_use*100:.1f}%)")
    print(f"  货币资金(结余):       {jy:>10.2f}万元 ({jy/total_use*100:.1f}%)")
    print(f"  {'─'*50}")
    print(f"  资金占用合计:         {total_use:>10.2f}万元")
    
    # ===== 指标3：资金来源 vs 资金占用（平衡性） =====
    print(f"\n【3. 资金来源 vs 资金占用 平衡性】")
    print(f"{'─'*50}")
    balance = total_src - total_use
    print(f"  来源合计: {total_src:.2f}万元")
    print(f"  占用合计: {total_use:.2f}万元")
    if abs(balance) < 0.01:
        print(f"  ✅ 来源=占用，平衡")
    else:
        print(f"  ⚠️ 差额: {balance:.4f}万元 （{'来源>占用' if balance > 0 else '占用>来源'})")
    
    # ===== 指标4：资金结余率 =====
    print(f"\n【4. 资金结余分析】")
    print(f"{'─'*50}")
    balance_rate = jy / total_src * 100
    print(f"  资金结余(货币资金): {jy:.2f}万元 ({balance_rate:.2f}%)")
    if balance_rate < 2:
        print(f"  ✅ 结余率<2%，资金使用充分")
    elif balance_rate < 5:
        print(f"  ⚠️ 结余率{balance_rate:.1f}%，需说明结余原因")
    elif balance_rate < 10:
        print(f"  ⚠️ 结余率{balance_rate:.1f}%偏高，需核实是否应缴回财政")
    else:
        print(f"  ❌ 结余率{balance_rate:.1f}%过高，应缴回财政")
    
    # ===== 指标5：概算执行率 =====
    print(f"\n【5. 概算/预算执行率】")
    print(f"{'─'*50}")
    
    gs = d['概算总投资']
    ss = d['送审金额_报告']
    ds = d['审定金额_报告']
    yf_amt = d['已支付']
    
    # 执行率 = 审定/概算
    exec_rate = ds / gs * 100
    print(f"  概算总投资: {gs/10000:.2f}万元")
    print(f"  送审金额:   {ss/10000:.2f}万元")
    print(f"  审定金额:   {ds/10000:.2f}万元")
    print(f"  {'─'*50}")
    print(f"  预算执行率(审定/概算): {exec_rate:.1f}%")
    if exec_rate <= 100:
        print(f"  ✅ 在概算范围内")
    else:
        print(f"  ⚠️ 超概{exec_rate-100:.1f}%，需说明")
    
    # 超概分析
    over = ss - gs
    over_rate = over / gs * 100
    
    # 注意：概算可能不含征地拆迁
    # 从数据看：S220报告说概算2,026万，但批准概算表上是3,992万
    # 差异 = 3,992 - 2,026 = 1,966万 ≈ 征地拆迁
    print(f"\n  【超概分析】")
    print(f"  报告口径概算: {gs/10000:.2f}万元")
    print(f"  送审超概:     {ss/10000:.2f} - {gs/10000:.2f} = {over/10000:.2f}万元")
    
    # 如果审核汇总表有批准概算（合计行C3），则用那个
    if '批准概算' in d:
        gs_full = d['批准概算']
        print(f"  审核汇总表口径概算(含征地拆迁?): {gs_full/10000:.2f}万元")
        exec_rate_full = ds / gs_full * 100
        print(f"  预算执行率(审定/含征地拆迁): {exec_rate_full:.1f}%")
        if exec_rate_full <= 100:
            print(f"  ✅ 包含征地拆迁后，未超概")
        else:
            print(f"  ⚠️ 仍超概{exec_rate_full-100:.1f}%")
    
    # ===== 指标6：资金支付率 =====
    print(f"\n【6. 资金支付率（应付尽付情况）】")
    print(f"{'─'*50}")
    pay_rate = yf_amt / ds * 100
    unpaid = ds - yf_amt
    print(f"  审定金额:  {ds/10000:.2f}万元")
    print(f"  已支付:    {yf_amt/10000:.2f}万元")
    print(f"  {'─'*50}")
    print(f"  支付率:    {pay_rate:.1f}%")
    print(f"  未支付:    {unpaid/10000:.2f}万元")
    
    if pay_rate >= 99:
        print(f"  ✅ 基本支付完毕")
    elif pay_rate >= 90:
        print(f"  ⚠️ 尚有{100-pay_rate:.0f}%未支付")
    else:
        print(f"  ❌ 支付率偏低，需核实")
    
    # ===== 指标7：结余资金应退财政 =====
    print(f"\n【7. 结余资金（应退还财政）】")
    print(f"{'─'*50}")
    # 结余 = 资金来源 - 已支付（或占用-已支付？）
    # 通常结余 = 拨款 - 支付
    if '基建拨款' in d:
        cb = d['基建拨款']
        surplus = cb - yf_amt
        print(f"  财政拨款:      {cb/10000:.2f}万元")
        print(f"  已支付(审计): {yf_amt/10000:.2f}万元")
        print(f"  {'─'*50}")
        print(f"  应退财政结余:  {surplus/10000:.2f}万元")
        if surplus > 0:
            print(f"  ⚠️ 结余{surplus/10000:.2f}万元应退还财政部门")
        elif surplus < 0:
            print(f"  ⚠️ 超付{abs(surplus)/10000:.2f}万元")
        else:
            print(f"  ✅ 收支平衡")
        
        # 加上应付款的影响
        yfk = d.get('应付款合计', 0)
        net_surplus = surplus - yfk
        print(f"\n  考虑应付款后净结余: {net_surplus/10000:.2f}万元")
        if net_surplus > 0 and yfk > 0:
            print(f"  ⚠️ 有应付款{yfk/10000:.2f}万元未付，扣除后仍有结余{net_surplus/10000:.2f}万需退财政")
        elif net_surplus > 0:
            print(f"  ⚠️ 净结余{net_surplus/10000:.2f}万元应退财政")
        elif abs(net_surplus) < 0.01:
            print(f"  ✅ 收支平衡（含应付款后）")
    
    print(f"\n{'#'*70}")

# ========== 输出Excel ==========
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '预算执行指标复核'
    
    thin_border = Border(
        left=Side(style='thin', color='BDBDBD'),
        right=Side(style='thin', color='BDBDBD'),
        top=Side(style='thin', color='BDBDBD'),
        bottom=Side(style='thin', color='BDBDBD'),
    )
    header_fill = PatternFill(start_color='1A237E', fill_type='solid')
    green_fill = PatternFill(start_color='E8F5E9', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFF3E0', fill_type='solid')
    red_fill = PatternFill(start_color='FFEBEE', fill_type='solid')
    light_blue = PatternFill(start_color='E3F2FD', fill_type='solid')
    
    ws.merge_cells('A1:G1')
    ws['A1'] = '四川融策会计师事务所 - 资金到位与预算执行指标复核表'
    ws['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='1A237E')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 35
    
    # Sheet1: 综合指标表
    r = 3
    for proj, d in data.items():
        ws.merge_cells(f'A{r}:G{r}')
        ws.cell(row=r, column=1, value=proj).font = Font(size=11, bold=True, color='1A237E')
        ws.cell(row=r, column=1).fill = light_blue
        r += 1
        
        gs = d['概算总投资'] / 10000
        ss = d['送审金额_报告'] / 10000
        ds = d['审定金额_报告'] / 10000
        yf = d['已支付'] / 10000
        zj = d.get('基建拨款', 0) / 10000
        jy = d['货币资金(结余)'] / 10000
        total_src = d['资金来源合计'] / 10000
        
        # 指标清单
        indicators = [
            ('1. 资金来源总额', f'{total_src:.2f}万元', '', ''),
            ('  其中：基建拨款', f'{zj:.2f}万元', f'{zj/total_src*100:.1f}%', ''),
            ('  其中：应付款', f'{d.get("应付款合计",0)/10000:.2f}万元', '', ''),
            ('2. 资金占用总额', f'{d["资金占用合计"]/10000:.2f}万元', '', ''),
            ('  其中：基本建设支出', f'{d["基本建设支出(待核销)"]/10000:.2f}万元', '', ''),
            ('  其中：货币资金(结余)', f'{jy:.2f}万元', f'{jy/total_src*100:.1f}%', 
             '✅正常' if jy/total_src*100 < 2 else '⚠️偏高' if jy/total_src*100 < 5 else '❌过高应退'),
            ('3. 资金平衡', f'{total_src - d["资金占用合计"]/10000:.2f}万元', '', '✅平衡' if abs(total_src - d["资金占用合计"]/10000) < 0.01 else '⚠️不平衡'),
            ('4. 概算总投资(报告口径)', f'{gs:.2f}万元', '', ''),
            ('5. 送审金额', f'{ss:.2f}万元', '', ''),
            ('6. 审定金额', f'{ds:.2f}万元', '', ''),
            ('7. 预算执行率(审定/概算)', '', f'{ds/gs*100:.1f}%', '✅在概算内' if ds <= gs else f'超概{ds/gs*100-100:.1f}%'),
            ('8. 已支付金额', f'{yf:.2f}万元', '', ''),
            ('9. 支付率(已付/审定)', '', f'{yf/ds*100:.1f}%', '✅基本支付' if yf/ds*100 >= 99 else f'尚欠{(ds-yf):.2f}万'),
            ('10. 结余应退财政', f'{zj - yf:.2f}万元', '', 
             '⚠️应退财政' if zj - yf > 0 else '✅已用完' if abs(zj-yf) < 0.01 else '⚠️超付'),
        ]
        
        for ind in indicators:
            ws.cell(row=r, column=1, value=ind[0]).border = thin_border
            ws.cell(row=r, column=2, value=ind[1]).border = thin_border
            ws.cell(row=r, column=3, value=ind[2]).border = thin_border
            ws.cell(row=r, column=4, value=ind[3]).border = thin_border
            
            if '⚠️' in ind[3] or '❌' in ind[3]:
                fill = red_fill if '❌' in ind[3] else yellow_fill
                ws.cell(row=r, column=4).fill = fill
            elif '✅' in ind[3]:
                ws.cell(row=r, column=4).fill = green_fill
            
            for c in range(1, 5):
                ws.cell(row=r, column=c).font = Font(name='微软雅黑', size=9)
                ws.cell(row=r, column=c).alignment = Alignment(vertical='center', wrap_text=True)
            r += 1
        r += 1  # 空行
    
    for i, w in enumerate([35, 18, 14, 22], 1):
        ws.column_dimensions[chr(64+i)].width = w
    
    # Sheet2: 原始数据明细
    ws2 = wb.create_sheet('明细数据')
    ws2.merge_cells('A1:E1')
    ws2['A1'] = '附表1(财务决算表)提取的原始资金数据'
    ws2['A1'].font = Font(size=12, bold=True)
    
    ws2.append(['项目', '指标', '金额(元)', '金额(万元)', '来源'])
    for c in range(1, 6):
        ws2.cell(row=2, column=c).font = Font(bold=True)
        ws2.cell(row=2, column=c).fill = PatternFill(start_color='E0E0E0', fill_type='solid')
    
    r = 3
    for proj, d in data.items():
        ws2.cell(row=r, column=1, value=proj).font = Font(bold=True)
        ws2.cell(row=r, column=1).fill = light_blue
        ws2.merge_cells(f'A{r}:E{r}')
        r += 1
        
        detail_items = [
            ('基建拨款', d.get('基建拨款', 0), '附表1-02表'),
            ('应付款合计', d.get('应付款合计', 0), '附表1-02表'),
            ('资金来源合计', d['资金来源合计'], '附表1-02表'),
            ('基本建设支出(待核销)', d['基本建设支出(待核销)'], '附表1-02表'),
            ('货币资金(结余)', d['货币资金(结余)'], '附表1-02表'),
            ('资金占用合计', d['资金占用合计'], '附表1-02表'),
            ('批准概算', d.get('批准概算', 0), '附表2-01表'),
            ('送审投资', d.get('送审投资', 0), '附表2-01表'),
            ('审定投资', d.get('审定投资', 0), '附表2-01表'),
            ('送审金额(报告)', d.get('送审金额_报告', 0), '审核报告'),
            ('审定金额(报告)', d.get('审定金额_报告', 0), '审核报告'),
            ('已支付', d.get('已支付', 0), '审核报告'),
        ]
        
        for name, val, src in detail_items:
            ws2.cell(row=r, column=1, value=name).border = thin_border
            ws2.cell(row=r, column=2, value=val).border = thin_border
            ws2.cell(row=r, column=3, value=round(val/10000, 4)).border = thin_border
            ws2.cell(row=r, column=4, value=src).border = thin_border
            for c in range(1, 6):
                ws2.cell(row=r, column=c).font = Font(name='微软雅黑', size=9)
            r += 1
        r += 1
    
    for i, w in enumerate([6, 30, 16, 16, 20], 1):
        ws2.column_dimensions[chr(64+i)].width = w
    
    output_path = r'C:\Users\scrccpa\Desktop\融策审计资金预算执行复核表.xlsx'
    wb.save(output_path)
    print(f"\n✅ 资金预算执行复核表已保存至: {output_path}")
    print(f"   Sheet1: 预算执行指标复核")
    print(f"   Sheet2: 明细数据")

except ImportError:
    print("\n⚠️ openpyxl未安装，跳过Excel输出")

print("\n复核完毕！")
