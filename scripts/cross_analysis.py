import sys, os
sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

base = r'C:\Users\scrccpa\Desktop\成都轨道资源资料'

# ====================================
# STEP 0: POLICY CONTEXT SUMMARY
# (PDFs are scanned images, extracted from filenames + audit plan context)
# ====================================
print("=" * 70)
print("一、政策文件概要")
print("=" * 70)
print("""
文件1: [2020]109号 成都市属国企资产招商管理办法 (14页扫描件)
  - 规范市属国企资产招商（招租/招商）管理流程
  - 涉及：评估定价、公开招商、合同管理、收益上缴

文件2: SW-2025-1554 关于进一步规范市属国有企业资产租赁管理 (8页扫描件)
  - 2025年新规，进一步收紧资产租赁管理要求
  - 涉及：租赁审批权限、租期限制、租金评估、公示程序

文件3: 2025年制度建设(2).zip (37MB)
文件4: 资产/资产经营/资金监督(1).zip (23MB)
  
审计依据：上述政策 + 资源公司内部制度 + 租赁合同约定
审计重点合规项：
  1. 资产出租是否经评估/公开招商
  2. 租金标准是否符合政策要求
  3. 合同是否到期续签、有无违规转租
  4. 租金/停车费是否及时足额上缴
""")

# ====================================
# STEP 1: 商户台账分析
# ====================================
print("\n" + "=" * 70)
print("二、商户台账数据交叉分析")
print("=" * 70)

import openpyxl

# Try to find the actual file paths by scanning
def find_data_files():
    found = []
    for root, dirs, files in os.walk(base):
        for f in files:
            if f.endswith('.xlsx') or f.endswith('.xls'):
                fp = os.path.join(root, f)
                fn_lower = f.lower()
                cat = '其他'
                if '商户' in fn_lower or '商家' in fn_lower:
                    cat = '商户台账'
                elif '车位' in fn_lower or '停车' in fn_lower or '收费' in fn_lower or '协议' in fn_lower:
                    cat = '停车场'
                elif '资产' in fn_lower:
                    cat = '资产台账'
                elif '设备' in fn_lower or '设施' in fn_lower:
                    cat = '设施设备'
                elif '水费' in fn_lower or '能耗' in fn_lower or '分摊' in fn_lower:
                    cat = '能耗'
                elif '转租' in fn_lower or '转供' in fn_lower:
                    cat = '转供/转租'
                found.append((cat, fp))
    return found

all_files = find_data_files()
print(f"\n找到 {len(all_files)} 个数据文件:\n")
for cat, path in all_files:
    size = os.path.getsize(path) // 1024
    print(f"  [{cat}] {os.path.basename(path)} ({size}KB)")

# Read merchant/商家 data files
print("\n\n--- 商户台账/商家经营信息 ---")
merchant_files = [(c, p) for c, p in all_files if '商户' in c or '商家' in c]
for cat, path in merchant_files:
    print(f"\n文件: {os.path.basename(path)}")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        print(f"  Sheet页: {wb.sheetnames}")
        for sn in wb.sheetnames:
            ws = wb[sn]
            print(f"  --- {sn} ({ws.max_row}行 x {ws.max_column}列) ---")
            for r in range(1, min(ws.max_row+1, 35)):
                row_data = []
                for c in range(1, min(ws.max_column+1, 15)):
                    v = ws.cell(r, c).value
                    if v is not None:
                        s = str(v).strip()
                        if len(s) > 60:
                            s = s[:57] + '...'
                        row_data.append(s)
                if row_data:
                    print(f"    R{r}: {' | '.join(row_data)}")
    except Exception as e:
        print(f"  ERROR: {e}")

# Read parking data files
print("\n\n--- 停车场/收费数据 ---")
parking_files = [(c, p) for c, p in all_files if c == '停车场']
for cat, path in parking_files:
    print(f"\n文件: {os.path.basename(path)}")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        print(f"  Sheet页: {wb.sheetnames}")
        for sn in wb.sheetnames:
            ws = wb[sn]
            print(f"  --- {sn} ({ws.max_row}行 x {ws.max_column}列) ---")
            for r in range(1, min(ws.max_row+1, 35)):
                row_data = []
                for c in range(1, min(ws.max_column+1, 15)):
                    v = ws.cell(r, c).value
                    if v is not None:
                        s = str(v).strip()
                        if len(s) > 60:
                            s = s[:57] + '...'
                        row_data.append(s)
                if row_data:
                    print(f"    R{r}: {' | '.join(row_data)}")
    except Exception as e:
        print(f"  ERROR: {e}")
