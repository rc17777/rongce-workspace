import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ====== 样式定义 ======
header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
sub_header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
sub_header_font = Font(name='微软雅黑', bold=True, size=10, color='2F5496')
section_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
section_font = Font(name='微软雅黑', bold=True, size=10, color='BF8F00')
normal_font = Font(name='微软雅黑', size=10)
title_font = Font(name='微软雅黑', bold=True, size=14, color='2F5496')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='center', horizontal='left')
center_align = Alignment(wrap_text=True, vertical='center', horizontal='center')

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_row(ws, row, cols, font=None, fill=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = font or normal_font
        cell.alignment = wrap_align
        cell.border = thin_border
        if fill:
            cell.fill = fill

def auto_width(ws, min_w=12, max_w=55):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = min_w
        for cell in col:
            if cell.value:
                lines = str(cell.value).split('\n')
                for line in lines:
                    # 中文字符算2个宽度
                    char_len = sum(2 if ord(c) > 127 else 1 for c in line)
                    max_len = max(max_len, char_len)
        ws.column_dimensions[col_letter].width = min(max_len + 4, max_w)

# ====== Sheet 1: 制度总览 ======
ws1 = wb.active
ws1.title = '制度总览'

ws1.merge_cells('A1:F1')
ws1['A1'] = '成都市教育科学研究院附属中学 — 政府采购管理制度（节选）要点分析'
ws1['A1'].font = title_font
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 30

headers1 = ['制度类型', '适用范围', '核心原则', '决策机构', '归口管理部门', '监督部门']
for i, h in enumerate(headers1, 1):
    ws1.cell(row=3, column=i, value=h)
style_header(ws1, 3, 6)

data1 = [
    ['政府采购管理制度\n（12章）',
     '使用财政性资金采购集中采购目录以内或采购限额标准以上的货物、工程和服务',
     '公开透明、公平竞争、公正、诚实信用；\n"分事行权、分岗设权、分级授权"',
     '校长办公会/党组织委员会',
     '总务处',
     '纪检小组'],
    ['一般采购管理制度\n（10章+7附件）',
     '集中采购目录以外、限额标准以下的：\n· 货物/服务类：50万元以下\n· 工程类：100万元以下',
     '权责分离原则；\n三重一大议事规则',
     '校长办公会/党组织委员会',
     '总务处',
     '纪检小组'],
]
for r, row in enumerate(data1, 4):
    for c, val in enumerate(row, 1):
        ws1.cell(row=r, column=c, value=val)
    style_row(ws1, r, 6)

# 金额速查表
ws1.merge_cells('A7:F7')
ws1['A7'] = '政府采购金额标准速查'
ws1['A7'].font = Font(name='微软雅黑', bold=True, size=12, color='2F5496')
style_row(ws1, 7, 6)

headers_p = ['项目类型', '金额区间', '采购方式', '评审小组要求', '备注']
for i, h in enumerate(headers_p, 1):
    ws1.cell(row=8, column=i, value=h)
style_header(ws1, 8, 5)

amount_data = [
    ['货物/服务', '1万元以下', '自行采购', '—', '报业务处室负责人、分管副校长、校长审批'],
    ['货物/服务', '1万元（含）—5万元', '询价/比选/磋商', '3人采购小组\n（业务处室组成）', '单位负责人审批'],
    ['货物/服务/工程', '5万元（含）—10万元', '公开比选\n可不委托代理机构', '≥7人评审小组\n（总务处牵头）', '需校长办公会/党组织委员会审议采购文件'],
    ['货物/服务/工程', '10万元—50万元\n（货物/服务限额以下）', '公开比选\n须委托采购代理机构', '≥7人评审小组\n（代理机构比选）', '非政府采购项目'],
    ['货物/服务', '50万元以上\n（达政府采购限额）', '政府采购\n（公开招标为主）', '≥5人单数\n评审专家≥2/3', '执行政府采购制度\n编制采购实施计划报备'],
    ['工程', '100—400万元', '政府采购\n（原则上专门面向中小企业）', '≥5人单数\n评审专家≥2/3', '预留份额面向中小企业'],
    ['工程', '400万元以上\n（达政府采购限额）', '政府采购\n（公开招标为主）', '≥5人单数\n评审专家≥2/3', '可预留30%以上份额给中小企业'],
]
for r, row in enumerate(amount_data, 9):
    for c, val in enumerate(row, 1):
        ws1.cell(row=r, column=c, value=val)
    style_row(ws1, r, 5)

auto_width(ws1, max_w=45)
ws1.column_dimensions['C'].width = 28
ws1.column_dimensions['D'].width = 28
ws1.column_dimensions['E'].width = 35

# ====== Sheet 2: 管理机构与职责 ======
ws2 = wb.create_sheet('管理机构与职责')

ws2.merge_cells('A1:E1')
ws2['A1'] = '管理机构设置与职责分工'
ws2['A1'].font = title_font
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 30

headers2 = ['机构角色', '部门/岗位', '序号', '主要职责', '备注']
for i, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=i, value=h)
style_header(ws2, 3, 5)

org_data = [
    # 政府采购管理制度
    ['决策机构', '校长办公会/\n党组织委员会', '1', '审定政府采购内部管理制度', '政府采购'],
    ['', '', '2', '研究决定重大政府采购事项，审定政府采购预算和计划', ''],
    ['', '', '3', '督促按制度办理政府采购业务，协调解决重大问题', ''],
    ['', '', '4', '负责采购管理其他事项的决策', ''],
    ['归口管理部门', '总务处', '1', '拟定政府采购内部管理制度，及时更新工作细则', '政府采购'],
    ['', '', '2', '汇总各业务部门提交的采购预算数据、计划、申请', ''],
    ['', '', '3', '确定政府采购组织形式和采购方式', ''],
    ['', '', '4', '加强单位自行组织采购活动的管理', ''],
    ['', '', '5', '指导和督促各业务部门归档政府采购合同', ''],
    ['', '', '6', '组织实施政府采购验收', ''],
    ['', '', '7', '组织处理政府采购纠纷调处', ''],
    ['', '', '8', '保管政府采购相关资料，移交合同/验收报告给财务', ''],
    ['', '', '9', '采购需求处室确认采购文件，异议报行政会审议', ''],
    ['执行部门', '各业务处室', '1', '申报本处室的政府采购预算建议数', '政府采购'],
    ['', '', '2', '编制政府采购计划，进行需求登记，提出采购申请', ''],
    ['', '', '3', '确认政府采购文件，有异议的进行调整修改', ''],
    ['', '', '4', '确认公开招标预中标结果，领取中标通知书，参与合同签订', ''],
    ['', '', '5', '对政府采购合同和相关文件进行备案', ''],
    ['', '', '6', '提出政府采购资金支付申请', ''],
    ['监督部门', '纪检小组', '1', '监督检查执行政府采购法律法规和相关规定的情况', '政府采购'],
    ['', '', '2', '参与政府采购业务投诉答复的处理', ''],
    # 一般采购管理制度
    ['决策机构', '校长办公会/\n党组织委员会', '1', '审议并批复一般采购管理制度和流程', '一般采购'],
    ['', '', '2', '审议并批复采购管理内部控制机构设置情况', ''],
    ['', '', '3', '审议并批复一般采购项目（5万以上），5万以下授权领导审批', ''],
    ['', '', '4', '审议其他有必要经会议决策的一般采购事项', ''],
    ['归口管理部门', '总务处', '1', '拟定一般采购制度及流程', '一般采购'],
    ['', '', '2', '审定单位采购项目的采购（招标）文件', ''],
    ['', '', '3', '会同各处室审核采购合同', ''],
    ['', '', '4', '组织参与采购需求评审及采购验收工作', ''],
    ['', '', '5', '组建一般采购评标、验收小组（3万元以上），组织评审验收', ''],
    ['', '', '6', '负责采购资料的归档与备案工作', ''],
    ['', '', '7', '指导采购法律法规贯彻执行，加强业务培训', ''],
    ['', '', '8', '负责会议交办的其他采购管理工作', ''],
    ['执行部门', '各业务处室', '1', '提出采购需求并按流程报批', '一般采购'],
    ['', '', '2', '编制采购计划并按程序报批', ''],
    ['', '', '3', '执行各一般采购项目，拟定和确认采购文件', ''],
    ['', '', '4', '5万元以上拟定比选文件，明确采购项目、预算、比选程序等', ''],
    ['', '', '5', '协助受理采购质疑与投诉事项', ''],
    ['', '', '6', '对一般采购合同相关文件进行备案', ''],
    ['', '', '7', '五万以上经费报销需由经费牵头部门申请过会（三重一大）', ''],
    ['', '', '8', '经费报销时项目验收由经费牵头部门验收', ''],
    ['临时机构', '评审/验收小组', '—', '3万元以上由总务处组建；3万元以下由采购处室组建', '一般采购'],
    ['监督部门', '纪检小组', '1', '建立健全一般采购管理内部控制监督机制', '一般采购'],
    ['', '', '2', '监督检查一般采购制度流程是否建立健全并有效执行', ''],
    ['', '', '—', '不得从事采购执行事务', ''],
]

for r, row in enumerate(org_data, 4):
    for c, val in enumerate(row, 1):
        ws2.cell(row=r, column=c, value=val)
    fill = None
    if row[0]:
        fill = section_fill
    style_row(ws2, r, 5, fill=fill if row[0] else None)

auto_width(ws2)
ws2.column_dimensions['A'].width = 16
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['D'].width = 55
ws2.column_dimensions['E'].width = 14

# ====== Sheet 3: 采购方式与条件 ======
ws3 = wb.create_sheet('采购方式与适用条件')

ws3.merge_cells('A1:D1')
ws3['A1'] = '政府采购方式及适用条件'
ws3['A1'].font = title_font
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 30

headers3 = ['采购方式', '适用条件', '适用制度', '关键要求']
for i, h in enumerate(headers3, 1):
    ws3.cell(row=3, column=i, value=h)
style_header(ws3, 3, 4)

method_data = [
    ['公开招标\n（主要采购方式）', '达到公开招标数额标准的货物/服务/工程\n（执行高新区相关标准）', '政府采购制度', '不得化整为零规避公开招标'],
    ['邀请招标', '①具有特殊性，只能从有限范围供应商处采购\n②公开招标费用占项目总价值比例过大', '政府采购制度', '—'],
    ['竞争性谈判', '①招标后无供应商投标/无合格标/重新招标未成立\n②技术复杂或性质特殊，不能确定详细规格\n③招标时间不能满足紧急需要\n④不能事先计算价格总额', '政府采购制度', '谈判文件须经采购人书面同意'],
    ['单一来源', '①只能从唯一供应商处采购\n②不可预见的紧急情况\n③保证一致性和配套要求，添购金额≤原合同10%', '政府采购制度', '需严格审核'],
    ['询价', '采购货物规格标准统一、现货充足且价格变化幅度小', '政府采购制度', '适用于标准化货物'],
    ['网上竞价', '经常性零星小额采购：\n· 计算机、打印设备、传真机、复印机等\n· 单项或批量采购预算2万以上软件/碎纸机/复印纸\n· 5万以上家具类\n年累计单项≤50万元（公务车除外）', '政府采购制度', '如对竞价结果不满意可改用其他方式'],
    ['公开比选\n（5-10万）', '非政府采购项目，金额5万元（含）-10万元（不含）', '一般采购制度', '可不委托代理机构\n≥7人评审小组'],
    ['公开比选\n（10万-限额以下）', '非政府采购项目，金额10万元（含）-50万元（货物/服务）\n10万元-100万元（工程）', '一般采购制度', '须委托采购代理机构\n公开比选方式'],
    ['询价/比选/磋商\n（1-5万）', '非政府采购项目，金额1万元（含）-5万元', '一般采购制度', '3人采购小组\n单位负责人审批'],
    ['自行采购\n（1万以下）', '非政府采购项目，金额1万元以下', '一般采购制度', '报业务处室负责人、分管副校长、校长审批'],
    ['议定方式', '①实施2次比选流程后，供应商仍不足3家\n②涉及国家安全、保密等要求\n③特定/唯一、国家省市要求的定点供应商', '一般采购制度', '校长办公会/党组织委员会研究议定'],
]
for r, row in enumerate(method_data, 4):
    for c, val in enumerate(row, 1):
        ws3.cell(row=r, column=c, value=val)
    style_row(ws3, r, 4)

auto_width(ws3)
ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 52
ws3.column_dimensions['C'].width = 18
ws3.column_dimensions['D'].width = 32

# ====== Sheet 4: 流程控制要点 ======
ws4 = wb.create_sheet('流程控制要点')

ws4.merge_cells('A1:E1')
ws4['A1'] = '政府采购关键流程与控制要点'
ws4['A1'].font = title_font
ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[1].height = 30

headers4 = ['流程环节', '控制要点', '负责部门/人', '制度依据（章节）', '风险提示']
for i, h in enumerate(headers4, 1):
    ws4.cell(row=3, column=i, value=h)
style_header(ws4, 3, 5)

flow_data = [
    ['预算编制', '硬化预算约束，细化编制；未编制预算不得实施采购', '各处室→总务处预算岗→区教育文体局', '第三章', '避免重复购置、浪费资金'],
    ['需求调查', '面向≥3个代表性市场主体开展调查；\n1000万以上货物/服务、3000万以上工程必须调查', '各处室', '第五章\n（财库〔2021〕22号）', '涉及公共利益/技术复杂项目必须调查'],
    ['需求编制', '包括技术、商务等要求；可自行或委托第三方编制', '业务处室', '第五章', '—'],
    ['实施计划编制', '明确类别、名称、标的、预算、数量、组织形式、采购方式等', '业务处室', '第五章', '必须报上级部门备案'],
    ['需求与计划审查', '组建≥3人审查工作组；编制专家和第三方不得参与审查', '总务处牵头', '第五章', '利益相关方回避'],
    ['意向公开', '公开采购项目名称、需求概况、预算金额、预计采购时间；\n不晚于采购活动开始前30日', '—', '第四章', '除小额零星采购外均需公开'],
    ['代理机构选择', '在年初确定的代理机构库中随机抽取', '采购人', '第六章', '任何单位和个人不得指定代理机构'],
    ['采购方式变更', '填写《政府采购方式变更申请表》，书面申请', '采购需求处室/代理机构', '第六章', '严格按成高财发〔2019〕139号执行'],
    ['评审委员会组建', '公开招标：≥5人单数，评审专家≥2/3\n竞争性谈判/询价：≥3人单数，评审专家≥2/3\n采购人/代理机构人员不得作为评审专家', '—', '第六章', '评审专家不得少于成员总数2/3'],
    ['合同签订', '中标通知书发出之日起30日内签订\n自签订之日起7个工作日报备\n自签订之日起2个工作日内公告', '各业务处室', '第九章', '改变中标结果须承担法律责任'],
    ['补充合同', '追加与合同标的相同的货物/工程/服务\n所有补充合同金额≤原合同金额10%', '—', '第九章', '不得擅自变更、中止或终止合同'],
    ['资金支付', '收到供应商发票后15日内支付\n中小企业：收到发票后10个工作日内支付', '采购人', '第九章', '—'],
    ['履约验收', '组织对供应商履约验收；大型/复杂项目邀请质量检测机构参加；验收方签字并承担法律责任', '采购人/代理机构', '第九章', '公共服务项目验收应邀请服务对象参与'],
    ['档案管理', '采购专员收集整理（一式两份）移交总务处归档\n保存期限≥15年（可用电子档案）', '总务处', '第九章', '包括会议纪要、实施计划表、招投标资料、合同、验收报告等'],
    ['质疑处理', '询问：3个工作日内答复\n书面质疑：7个工作日内书面答复', '采购人', '第十章', '答复内容不得涉及商业秘密'],
    ['中小企业预留', '200万以下货物/服务、400万以下工程原则上专门面向中小企业；\n超过的预留≥30%份额', '各处室', '第三章', '对照《四川省政府采购面向中小企业采购指导目录》确定'],
]
for r, row in enumerate(flow_data, 4):
    for c, val in enumerate(row, 1):
        ws4.cell(row=r, column=c, value=val)
    style_row(ws4, r, 5)

auto_width(ws4)
ws4.column_dimensions['A'].width = 16
ws4.column_dimensions['B'].width = 48
ws4.column_dimensions['C'].width = 24
ws4.column_dimensions['D'].width = 20
ws4.column_dimensions['E'].width = 32

# ====== Sheet 5: 关键时间节点 ======
ws5 = wb.create_sheet('关键时间节点')

ws5.merge_cells('A1:D1')
ws5['A1'] = '政府采购关键时间节点汇总'
ws5['A1'].font = title_font
ws5['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws5.row_dimensions[1].height = 30

headers5 = ['序号', '事项', '时间要求', '适用制度']
for i, h in enumerate(headers5, 1):
    ws5.cell(row=3, column=i, value=h)
style_header(ws5, 3, 4)

time_data = [
    ['1', '采购意向公开', '原则上不晚于采购活动开始前30日', '政府采购'],
    ['2', '比选文件发售时间', '不少于3个工作日', '一般采购'],
    ['3', '比选文件发售至响应提交截止', '不少于5个工作日', '一般采购'],
    ['4', '邀请比选 → 响应提交截止', '不少于3个工作日', '一般采购'],
    ['5', '合同签订时限', '中标通知书发出之日起30日内', '政府采购'],
    ['6', '比选合同签订时限', '中选通知书发出之日起3日后、10日内', '一般采购'],
    ['7', '合同备案', '签订之日起7个工作日内', '政府采购'],
    ['8', '合同公告', '签订之日起2个工作日内', '政府采购'],
    ['9', '评审报告送交采购人', '评审结束之日起2个工作日内', '政府采购'],
    ['10', '确定中标/成交供应商', '收到评审报告之日起5个工作日内', '政府采购'],
    ['11', '发出中标/成交通知书', '确定之日起2个工作日内', '政府采购'],
    ['12', '资金支付（一般）', '收到供应商发票后15日内', '政府采购'],
    ['13', '资金支付（中小企业）', '收到供应商发票后10个工作日内', '政府采购'],
    ['14', '询问答复', '供应商提出后3个工作日内', '政府采购'],
    ['15', '书面质疑答复', '收到书面质疑后7个工作日内', '政府采购'],
    ['16', '一般采购质疑答复', '收到质疑后7个工作日内', '一般采购'],
    ['17', '投诉答复（纪检小组）', '收到书面投诉后7个工作日内', '一般采购'],
    ['18', '中选结果公示期', '不少于2个工作日', '一般采购'],
    ['19', '档案保存期限', '从采购结束之日起至少15年', '政府采购'],
    ['20', '验收时限', '收到供应商验收申请后15个工作日内确定验收时间', '一般采购'],
    ['21', '追加采购限额', '所有补充合同金额≤原合同金额10%', '政府采购'],
    ['22', '服务合同期限上限', '不超过3年履行期限（需年度预算保障+事前载明）', '政府采购/政府购买服务'],
]
for r, row in enumerate(time_data, 4):
    for c, val in enumerate(row, 1):
        ws5.cell(row=r, column=c, value=val)
    style_row(ws5, r, 4)

auto_width(ws5)
ws5.column_dimensions['A'].width = 8
ws5.column_dimensions['B'].width = 32
ws5.column_dimensions['C'].width = 38
ws5.column_dimensions['D'].width = 18

# ====== Sheet 6: 进口产品与政府购买服务 ======
ws6 = wb.create_sheet('进口产品与购买服务')

ws6.merge_cells('A1:D1')
ws6['A1'] = '进口产品采购 & 政府购买服务管理要点'
ws6['A1'].font = title_font
ws6['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws6.row_dimensions[1].height = 30

headers6 = ['类别', '控制目标/要求', '具体内容', '注意事项']
for i, h in enumerate(headers6, 1):
    ws6.cell(row=3, column=i, value=h)
style_header(ws6, 3, 4)

special_data = [
    ['进口产品采购\n（第八章）', '建立内部会商决策机制\n需求制定与专家论证分离', '①政府采购应当采购本国产品\n②确需采购进口产品实行审核管理\n③严格厉行节约，不得申请采购禁止进口产品\n④未纳入统一论证范围的须报财政部门审核', '全面了解产品信息，科学合理确定需求'],
    ['政府购买服务\n（第七章）', '将自身职责范围且适合市场化方式提供的服务事项，交由符合条件的供应商承担', '①认真开展采购需求调研和咨询论证\n②预算批复后及时向财政部门申报采购计划\n③通过四川政府采购网等媒体公布服务项目、内容、对承接主体要求和绩效评价标准\n④委托采购代理机构组织实施\n⑤确定合适采购方式（公开招标/邀请招标/竞争性谈判/单一来源/询价）\n⑥签订采购合同，期限不超过3年', '需事前在采购文件中载明可签多年合同'],
    ['政府购买服务\n预算编制', '结合项目特点合理测算', '应发挥行业组织、专业咨询评估机构、专家等优势\n综合物价、工资、税费等因素测算所需预算', '—'],
    ['预留份额\n面向中小企业', '落实中小企业采购政策\n（财库〔2016〕46号）', '①200万以下货物/服务 → 原则上专门面向中小企业\n②400万以下工程 → 原则上专门面向中小企业\n③超过以上标准的 → 预留≥30%份额\n④对照《四川省政府采购面向中小企业采购指导目录（2021年版）》确定', '除《办法》规定可不专门面向中小企业的情形外'],
]
for r, row in enumerate(special_data, 4):
    for c, val in enumerate(row, 1):
        ws6.cell(row=r, column=c, value=val)
    style_row(ws6, r, 4)

auto_width(ws6)
ws6.column_dimensions['A'].width = 18
ws6.column_dimensions['B'].width = 35
ws6.column_dimensions['C'].width = 55
ws6.column_dimensions['D'].width = 35

# ====== Sheet 7: 供应商资质与验收 ======
ws7 = wb.create_sheet('供应商资质与验收标准')

ws7.merge_cells('A1:D1')
ws7['A1'] = '供应商资质要求 & 验收管理标准'
ws7['A1'].font = title_font
ws7['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws7.row_dimensions[1].height = 30

headers7 = ['类别', '序号', '内容要求', '说明']
for i, h in enumerate(headers7, 1):
    ws7.cell(row=3, column=i, value=h)
style_header(ws7, 3, 4)

qual_data = [
    ['供应商基本资质', '1', '依法设立，具有独立承担法律责任的能力；提供合法完整有效的营业执照；治理结构健全，内部管理和监督制度完善', ''],
    ['', '2', '具备所必需的设施和专业技术能力等资源条件；有专业资质要求的应具备相应要求', ''],
    ['', '3', '具有独立健全的财务管理、会计核算和资产管理制度；依法缴纳税收和社会保险费', ''],
    ['', '4', '参与采购前3年内无重大违法违纪记录；信用状况良好，未被列入经营异常名录或严重违法企业名单', '成立不足3年则自成立之日起算'],
    ['', '5', '法律、法规规定以及购买项目要求的其他条件', ''],
    ['验收内容', '1', '合同期限情况', ''],
    ['', '2', '核对供货清单：品牌、型号规格、品质材质、货物数量等', '与合同完全相符'],
    ['', '3', '检查质量与技术水平：进行检验、实验或测试', '达到合同规定标准'],
    ['', '4', '服务质量：核查服务内容、培训方案、服务响应时间等', '按合同承诺提供相应服务'],
    ['', '5', '其它与合同有关的内容', ''],
    ['验收不合格情形', '1', '与采购合同规定的内容不相符（型号、规格、品牌等）', ''],
    ['', '2', '产品质量、系统技术水平不达标', '无法达到国家标准及合同规定'],
    ['', '3', '无正当理由延期完工或不能按期供货造成学校损失', ''],
    ['', '4', '验收小组成员经投票确认不合格', ''],
    ['验收归档资料\n（比选采购）', '—', '签到表、密封情况检查记录、报价表、技术参数情况表、评审表、比选决议、中选通知书、评审报告、比选文件、供应商响应文件、验收评估结论表', '由总务处采购专员归档'],
    ['验收归档资料\n（询价采购）', '—', '询价函、询价评审报告、验收评估结论表、供应商响应文件', '由总务处采购专员归档'],
]
for r, row in enumerate(qual_data, 4):
    for c, val in enumerate(row, 1):
        ws7.cell(row=r, column=c, value=val)
    fill = section_fill if row[0] else None
    style_row(ws7, r, 4, fill=fill)

auto_width(ws7)
ws7.column_dimensions['A'].width = 22
ws7.column_dimensions['C'].width = 60

# ====== Sheet 8: 比选评审标准 ======
ws8 = wb.create_sheet('比选评审标准')

ws8.merge_cells('A1:C1')
ws8['A1'] = '比选评审标准（供参考）'
ws8['A1'].font = title_font
ws8['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws8.row_dimensions[1].height = 30

headers8 = ['评审项', '评审标准', '分值范围/说明']
for i, h in enumerate(headers8, 1):
    ws8.cell(row=3, column=i, value=h)
style_header(ws8, 3, 3)

eval_data = [
    ['团队实力', '业绩、获奖等', '20分'],
    ['项目要求', '满足项目要求的方案', '50分'],
    ['报价', '报价得分 = (评标基准价 / 投标报价) × 分值\n基准价确定方式：①最低报价为基准价 ②所有报价取平均值', '满分30分\n服务类：30≥报价分值≥10\n货物类：60≥报价分值≥30'],
    ['合计', '综合评分法', '100分'],
    ['', '', ''],
    ['评审规则补充说明', '', ''],
    ['· 价格分权重', '服务类项目', '报价分值范围：10-30分'],
    ['', '货物类项目', '报价分值范围：30-60分'],
    ['· 基准价确认', '方式一', '以最低报价为基准价'],
    ['', '方式二', '以所有报价取平均值为基准价'],
    ['· 评审项可调整', '按照项目要求可增加评审项，分值按需求自行分配', ''],
]
for r, row in enumerate(eval_data, 4):
    for c, val in enumerate(row, 1):
        ws8.cell(row=r, column=c, value=val)
    fill = section_fill if '补充说明' in str(row[0]) else None
    style_row(ws8, r, 3, fill=fill)

auto_width(ws8)
ws8.column_dimensions['A'].width = 22
ws8.column_dimensions['B'].width = 52
ws8.column_dimensions['C'].width = 35

# 保存
output_path = r'D:\openclaw-workspace\政府采购管理制度要点分析.xlsx'
wb.save(output_path)
print(f'已保存: {output_path}')
