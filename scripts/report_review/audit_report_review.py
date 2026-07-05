# -*- coding: utf-8 -*-
"""
专项审计报告复核脚本
1. 基础复核：报告数据 vs Excel表格数据一致性 / 错别字 / 表述不精准
2. 审计程序复核：应有的 vs 已做的程序清单
输出：Excel格式复核结果
"""
import os, sys, re, json
from pathlib import Path
from collections import defaultdict

BASE = r'C:\Users\scrccpa\Desktop\林芝樾~1'

# ========================
# 1. 读取报告文本
# ========================
report_text = r"""林芝樾燊嘉瑞实业有限责任公司
法定代表人魏东升个人银行账户收付营业款情况
专项审计报告
四川竞泽云锦会计师事务所（普通合伙）
林芝樾燊嘉瑞实业有限责任公司
法定代表人魏东升个人银行账户收付营业款情况
专项审计报告
川竞泽审字[2026]第**号
林芝市巴宜区公安局：
我们接受贵单位的委托，对林芝樾燊嘉瑞实业有限责任公司（以下简称"公司"）法定代表人、股东（持股比例32.1%）、执行董事兼总经理魏东升（身份证号：513229197412140017）在2016年6月至2020年12月期间，以其个人银行账户收付公司营业款的情况进行了专项审计。送审案件材料的真实性、完整性、合法性是委托方的责任，我们的责任是在实施审计工作的基础上，对魏东升以个人银行账户收付营业款的情况发表审计意见。
我们的审计是依据《中国注册会计师审计准则》、《中国注册会计师其他鉴证业务准则第3101号—历史财务信息审计或审阅以外的鉴证业务》等相关准则的规定进行的。审计工作包括详查案件材料、分析、重新计算等我们认为必要的审计程序。我们相信，我们获取的审计证据是充分、适当的，为发表审计意见提供了基础。
委托方及委托日期：1、委托方：林芝市巴宜区公安局。2、委托日期：2024年8月23日。
委托审计事项：依据林芝市巴宜区公安局与我所签订的《专项审计合同》，我所受托对送审的魏东升2016年6月至2020年12月期间个人银行账户交易明细进行统计分析，核实魏东升以个人银行账户收支归属于公司的营业款项，并对收支金额发表审计意见。
送审资料：
1、林芝市巴宜区公安局提供的魏东升个人银行账户交易明细包括：中国建设银行3个账户，中国农业银行3个账户，中国银行1个账户
2、林芝市巴宜区公安局提供的林芝樾燊嘉瑞实业有限责任公司供应商清单、员工花名册；
3、案件涉及的相关人员询问笔录、支出证明材料。
审计依据...
审计原则...
审计方法和过程...
审计意见：
（一）剔除魏东升个人银行账户之间转账交易、交易失败退回等不影响资金净流入的银行交易明细，魏东升个人银行账户2016年6月至2020年12月资金流入合计188,153,290.57元。分类统计结果如下：
1、认定为林芝樾燊嘉瑞实业有限责任公司营业款的资金流入185,872,333.31元；
2、认定为魏东升合法工资及个人合法收入的资金流入2,280,957.26元。
（二）资金支出及去向情况
魏东升个人银行账户2016年6月至2020年12月资金流出合计188,225,393.63元。分类统计结果如下：
1、认定为公司经营支出的资金流出186,102,936.73元，主要包括支付货款、房租、水电、员工工资等与经营相关支出。
2、认定为个人支出等非公司经营支出的资金流出2,122,456.90元，主要包括生活消费、购买车辆、私人借贷等。
（三）审计结论
根据与委托方确认的经营收支认定原则统计分析得出，在2016年6月至2020年12月期间，魏东升个人银行账户中归属于公司的资金流入共计185,872,333.31元，用于公司经营的资金流出共计186,102,936.73元，送审个人银行账户中用于公司经营的资金流出金额大于归属于公司的资金流入230,603.42元，未发现魏东升通过个人银行账户侵占公司营业款的证据未发现存在涉嫌职务侵占单位资金的情形。"""

print("=" * 60)
print("报告文本已加载，长度:", len(report_text), "字符")

# ========================
# 2. 提取报告中的核心数据
# ========================
REPORT_NUMBERS = {
    '资金流入合计': 188153290.57,
    '营业款资金流入': 185872333.31,
    '合法工资及个人收入': 2280957.26,
    '资金流出合计': 188225393.63,
    '公司经营支出': 186102936.73,
    '个人支出等非经营支出': 2122456.90,
    '流出减流入差额': 230603.42,  # 186102936.73 - 185872333.31
}

# 验证算数
def verify_calcs():
    results = []
    obj = REPORT_NUMBERS
    
    # 流入分项合计
    inflow_sum = obj['营业款资金流入'] + obj['合法工资及个人收入']
    diff_inflow = round(inflow_sum - obj['资金流入合计'], 2)
    results.append(('流入分项合计', obj['资金流入合计'],
                    inflow_sum, diff_inflow,
                    '✅' if abs(diff_inflow) < 0.02 else '❌'))
    
    # 流出分项合计
    outflow_sum = obj['公司经营支出'] + obj['个人支出等非经营支出']
    diff_outflow = round(outflow_sum - obj['资金流出合计'], 2)
    results.append(('流出分项合计', obj['资金流出合计'],
                    outflow_sum, diff_outflow,
                    '✅' if abs(diff_outflow) < 0.02 else '❌'))
    
    # 差额验证
    calc_diff = round(obj['公司经营支出'] - obj['营业款资金流入'], 2)
    results.append(('经营收支差额', obj['流出减流入差额'],
                    calc_diff, round(calc_diff - obj['流出减流入差额'], 2),
                    '✅' if abs(calc_diff - obj['流出减流入差额']) < 0.02 else '❌'))
    
    # 总流入 vs 总流出
    total_diff = round(obj['资金流出合计'] - obj['资金流入合计'], 2)
    calc_total_diff = round(obj['个人支出等非经营支出'] - obj['合法工资及个人收入'], 2)
    results.append(('总差额(流出-流入)', total_diff,
                    calc_total_diff, round(total_diff - calc_total_diff, 2),
                    '✅' if abs(total_diff - calc_total_diff) < 0.02 else '❌'))
    
    return results

calc_results = verify_calcs()
print("\n===== 报告数据自校验 =====")
for item, expect, actual, diff, status in calc_results:
    print(f"  {status} {item}: 报告={expect}, 计算={actual}, 差异={diff}")

# ========================
# 3. 读取Excel附件数据进行交叉比对
# ========================
print("\n===== Excel附件数据读取 =====")

excel_report = {}
# Find Excel files
for root, dirs, files in os.walk(BASE):
    for f in files:
        if f.startswith('~$'):
            continue
        fp = os.path.join(root, f)
        if f.endswith('.xlsx'):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(fp, data_only=True)
                print(f"\n  📊 {f} ({os.path.getsize(fp)} bytes)")
                print(f"     Sheets: {wb.sheetnames}")
                
                for sn in wb.sheetnames:
                    ws = wb[sn]
                    print(f"     [{sn}] {ws.max_row} rows × {ws.max_column} cols")
                    
                    # Read first 3 rows for context
                    for row_idx, row in enumerate(ws.iter_rows(max_row=min(5, ws.max_row), values_only=True), 1):
                        vals = [str(v)[:50] if v else '' for v in row]
                        print(f"       R{row_idx}: {vals[:5]}")
                    
                    # Try to find totals/summary rows
                    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                        vals = [str(v) if v else '' for v in row]
                        if any('合计' in v or '总计' in v or '小计' in v for v in vals):
                            print(f"       🔸 R{row_idx} (含合计): {[v[:80] for v in vals if v][:8]}")
                
                wb.close()
            except Exception as e:
                print(f"  ❌ Error reading {f}: {e}")

print("\n" + "=" * 60)
print("数据读取完成")
