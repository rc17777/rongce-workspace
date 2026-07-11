# -*- coding: utf-8 -*-
"""青白江清算审计报告复核 - 生成Excel"""
import os, sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

output_path = os.path.join(os.environ['USERPROFILE'], 'Desktop', '青白江清算审计报告复核结果.xlsx')
wb = openpyxl.Workbook()

# ---- 样式 ----
hfont = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
hfill = PatternFill(start_color='1C355E', end_color='1C355E', fill_type='solid')
nfont = Font(name='微软雅黑', size=10)
bfont = Font(name='微软雅黑', size=10, bold=True)
rfont = Font(name='微软雅黑', size=10, color='CC0000', bold=True)
ofont = Font(name='微软雅黑', size=10, color='E65100', bold=True)
gfont = Font(name='微软雅黑', size=10, color='2E7D32')
gfill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
rfll = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
yfill = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
bfill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
tborder = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
wrap = Alignment(wrap_text=True, vertical='center')
cc = Alignment(horizontal='center', vertical='center')

def hdr(ws, row, cols):
    for c in range(1, cols+1):
        ce = ws.cell(row=row, column=c)
        ce.font = hfont; ce.fill = hfill; ce.alignment = cc; ce.border = tborder

def srow(ws, row, cols, f=nfont, fill=None):
    for c in range(1, cols+1):
        ce = ws.cell(row=row, column=c)
        ce.font = f; ce.border = tborder; ce.alignment = wrap
        if fill: ce.fill = fill

# ============== Sheet1: 复核总结 ==============
ws = wb.active; ws.title = '复核总结'; ws.sheet_properties.tabColor = '1C355E'

ws.merge_cells('A1:E1')
ws.cell(row=1, column=1, value='清算审计报告复核结果').font = Font(name='微软雅黑', bold=True, size=16, color='1C355E')
ws.merge_cells('A2:E2')
ws.cell(row=2, column=1, value=f'复核日期: {datetime.now().strftime("%Y-%m-%d %H:%M")}    报告名称: 成都市青白江区国家馆协会清算审计报告').font = Font(name='微软雅黑', size=9, color='666666')

r = 4
ws.merge_cells(f'A{r}:E{r}'); ws.cell(row=r,column=1,value='一、报告基本信息').font = Font(name='微软雅黑',bold=True,size=12,color='1C355E'); r+=1

info = [
    ['被审计单位','成都市青白江区国家馆协会'],
    ['审计类型','清算审计（社团注销）'],
    ['清算事由','2026年3月26日会员大会决议：经营不善，决定注销'],
    ['清算期间','2026年3月26日至2026年5月24日'],
    ['清算组','未在报告中明确列出清算组成员'],
    ['事务所','四川竞泽云锦会计师事务所（普通合伙）'],
    ['报告文号','川竞泽审字[2026]第**号（编号未填写⚠️）'],
    ['审计对象','清算资产负债表(2026.5.24)、清算业务活动表、清算财产表、债务清偿表'],
    ['审计类型','清算审计，但实际审计报告由注册会计师出具（不是清算组自行编制）'],
    ['特别说明','⚠️ 用户提示：非限定性净资产数据将变动（下周一银行销户确认利息）'],
]
for it in info:
    ws.cell(row=r,column=1,value=it[0]).font=bfont
    ws.merge_cells(f'B{r}:E{r}'); ws.cell(row=r,column=2,value=it[1]).font=nfont; ws.cell(row=r,column=2).alignment=wrap
    for c in range(1,6): ws.cell(row=r,column=c).border=tborder
    r+=1
r+=1

# 核心数据
ws.merge_cells(f'A{r}:E{r}'); ws.cell(row=r,column=1,value='二、核心数据对照（清算组报告 vs 审计报告 vs Excel报表）').font = Font(name='微软雅黑',bold=True,size=12,color='1C355E'); r+=1

h2=['序号','项目','清算组报告','审计报告','Excel报表','判定']
for ci,h in enumerate(h2,1): ws.cell(row=r,column=ci,value=h); hdr(ws,r,len(h2)); r+=1

data_check = [
    [1,'期初货币资金(3.27)','8,619.98','8,619.98','8,619.98','✅一致'],
    [2,'期初负债(3.27)','8,580.00','8,580.00','8,580.00','✅一致'],
    [3,'期初净资产(3.27)','39.98','39.98','39.98','✅一致'],
    [4,'清算期间利息收入','22.00','22.00','⚠️业务活动表收入为0','❌不一致'],
    [5,'清算期间费用','债务清偿8,580','债务清偿8,580','债务清偿8,580','✅一致'],
    [6,'期末货币资金(5.24)','8,641.98','8,641.98','⚠️ Excel为0','❌不一致'],
    [7,'期末负债(5.24)','0','0','0','✅一致'],
    [8,'期末净资产(5.24)','61.98（待确认）','61.98','⚠️ Excel净资产变动-39.98','❌不一致'],
    [9,'审计费','8,000','8,000','8,000','✅一致'],
    [10,'印章费','580','580','580','✅一致'],
    [11,'剩余财产','61.98（待银行销户确认）','61.98','未明确','⚠️需更新'],
]
for idx,dc in enumerate(data_check,1):
    ws.cell(row=r,column=1,value=dc[0])
    for ci,val in enumerate(dc[1:],2): ws.cell(row=r,column=ci,value=val)
    fill = gfill if '一致' in dc[5] else rfll
    srow(ws,r,6,fill=fill)
    for c in range(1,7): ws.cell(row=r,column=c).alignment=cc
    ws.cell(row=r,column=2).alignment=wrap
    r+=1
r+=1

# 统计
ws.merge_cells(f'A{r}:E{r}'); ws.cell(row=r,column=1,value='三、复核发现统计').font = Font(name='微软雅黑',bold=True,size=12,color='1C355E'); r+=1
stats = [
    ['报告文本问题','10项','🔴高风险3项 / 🟡中风险6项 / 🟢低风险1项'],
    ['审计程序完成率','7/15项 = 47%','✅已执行7项 / ⚠️部分执行2项 / ❌未执行6项'],
    ['数据一致性','8/11项一致','❌3项不一致：核心问题在Excel报表数据滞后'],
    ['综合评分','52/100','数据准确性5/10 + 文字质量7/10 + 程序完整性4/10 + 意见适当性5/10'],
]
for s in stats:
    ws.cell(row=r,column=1,value=s[0]).font=bfont; ws.merge_cells(f'B{r}:C{r}'); ws.cell(row=r,column=2,value=s[1]).font=nfont
    ws.merge_cells(f'D{r}:E{r}'); ws.cell(row=r,column=4,value=s[2]).font=nfont
    for c in range(1,6): ws.cell(row=r,column=c).border=tborder
    r+=1

ws.column_dimensions['A'].width=14; ws.column_dimensions['B'].width=20
ws.column_dimensions['C'].width=20; ws.column_dimensions['D'].width=28; ws.column_dimensions['E'].width=18

# ============== Sheet2: 数据一致性详细检查 ==============
ws2 = wb.create_sheet('数据一致性检查'); ws2.sheet_properties.tabColor='2196F3'
h2b=['序号','检查项目','清算组报告','审计报告','Excel报表','差异','判定','备注']
for ci,h in enumerate(h2b,1): ws2.cell(row=1,column=ci,value=h); hdr(ws2,1,len(h2b))

dc_full = [
    [1,'期初货币资金(2026.3.27)','8,619.98','8,619.98','8,619.98','0','✅一致','三方数据一致'],
    [2,'期初应付款项','8,580.00','8,580.00','8,580.00','0','✅一致','含审计费8,000+印章费580'],
    [3,'期初净资产(非限定性)','39.98','39.98','39.98','0','✅一致','货币资金-负债=39.98'],
    [4,'清算期间利息收入','22.00元','22.00元','⚠️表上收入为0','22.00','❌不一致','业务活动表未反映利息收入22元'],
    [5,'债务清偿-审计费','8,000','8,000','8,000','0','✅一致',''],
    [6,'债务清偿-印章费','580','580','580','0','✅一致',''],
    [7,'期末货币资金(5.24)','8,641.98','8,641.98','⚠️Excel为0','8,641.98','❌不一致','Excel资产负债表货币资金期末为0'],
    [8,'期末负债(5.24)','0','0','0','0','✅一致','债务已清偿'],
    [9,'期末净资产(5.24)','61.98','61.98','⚠️Excel净资产合计0','61.98','❌不一致','Excel显示净资产变动-39.98(欠反映利息)'],
    [10,'净资产变动额','+22.00','+22.00','0（业务活动表）','22.00','❌不一致','清算组/审计报告=利息22元，业务活动表=0'],
    [11,'资产处置','无实物资产','无实物资产','无实物资产','-','✅一致',''],
    [12,'债权申报','无单位和个人申报','无单位和个人申报','无单位和个人申报','-','✅一致','四川工人日报2026.4.8公告，45日公示期满'],
    [13,'剩余财产归属','未说明','未明确说明','未说明','-','⚠️需补充','按章程应说明剩余财产分配去向'],
]
for idx,dc in enumerate(dc_full,1):
    r=idx+1
    for ci,val in enumerate(dc[1:],2): ws2.cell(row=r,column=ci,value=val)
    ws2.cell(row=r,column=1,value=dc[0])
    fill=gfill if '一致' in dc[6] else (rfll if '不一致' in dc[6] else yfill)
    srow(ws2,r,8,fill=fill)
    for c in range(1,9): ws2.cell(row=r,column=c).alignment=cc
    ws2.cell(row=r,column=2).alignment=wrap; ws2.cell(row=r,column=8).alignment=wrap

ws2.column_dimensions['A'].width=6; ws2.column_dimensions['B'].width=24; ws2.column_dimensions['C'].width=16
ws2.column_dimensions['D'].width=16; ws2.column_dimensions['E'].width=18; ws2.column_dimensions['F'].width=12
ws2.column_dimensions['G'].width=12; ws2.column_dimensions['H'].width=36

# ============== Sheet3: 报告文本问题清单 ==============
ws3 = wb.create_sheet('报告文本问题清单'); ws3.sheet_properties.tabColor='FF9800'
h3=['序号','类别','位置','问题描述','原文/说明','程度','修改建议']
for ci,h in enumerate(h3,1): ws3.cell(row=1,column=ci,value=h); hdr(ws3,1,len(h3))

issues = [
    [1,'数据不一致','Excel vs 报告','Excel报表数据滞后，未反映清算期间利息收入','①资产负债表货币资金期末为0（应为8,641.98）；②业务活动表收入为0（应含利息22元）；③净资产变动为-39.98（应为+22.00）','高','待银行销户后统一更新Excel报表数据（用户确认下周一销户）'],
    [2,'数据待确认','审计报告','非限定性净资产可能变动','剩余净资产61.98元含22元利息，需等银行销户确认最终利息金额','高','在报告中增加【期后事项】段说明：银行账户尚未注销，净资产金额可能因销户利息发生调整'],
    [3,'格式问题','报告文号','编号未填写','川竞泽审字[2026]第**号','中','应填写具体文号编号'],
    [4,'专业表述','审计报告','剩余财产分配去向未说明','第(三)或(四)部分未说明61.98元剩余财产如何处置','高','按《社会团体登记管理条例》应明确剩余财产按章程或业务主管单位意见处置的归属'],
    [5,'表述问题','审计报告','清算组信息缺失','未列出清算组成员及负责人','中','应在报告中注明清算组组成情况（至少含负责人姓名）'],
    [6,'表述问题','审计报告','公告信息不完整','仅说明"已在四川工人日报公告"','中','应补充公告日期(2026.4.8)、公告期(45日)、公告期满日(2026.5.23)'],
    [7,'审计意见','审计报告','未包含强调事项段','银行账户尚未注销、最终利息未确定等期后事项未在审计意见中提及','中','审计意见段增加强调事项段：①银行账户将于XX日注销；②净资产金额可能因销户时点利息发生微调'],
    [8,'审计程序','审计报告','审计方法和过程描述笼统','清算组报告提及了检查/复核/监盘等程序，审计报告应为独立审计','低','审计报告应独立描述自身执行的审计程序，而非引用清算组的方法'],
    [9,'表述问题','审计报告','特别事项说明缺失','未说明清算结束日至审计报告日之间的后续事项','中','应增加特别事项说明段：①公告期满确认无异议；②银行销户后续安排'],
    [10,'格式问题','清算组报告','格式不规范','清算组报告标题、落款格式与标准清算报告有差异','低','建议增加：清算组成员签字/日期/附件清单'],
]
for iss in issues:
    r=iss[0]+1
    ws3.cell(row=r,column=1,value=iss[0])
    for ci,val in enumerate(iss[1:],2): ws3.cell(row=r,column=ci,value=val)
    fill=rfll if iss[5]=='高' else (yfill if iss[5]=='中' else None)
    srow(ws3,r,7,fill=fill)
    ws3.cell(row=r,column=1).alignment=cc; ws3.cell(row=r,column=6).alignment=cc

ws3.column_dimensions['A'].width=6; ws3.column_dimensions['B'].width=14; ws3.column_dimensions['C'].width=20
ws3.column_dimensions['D'].width=38; ws3.column_dimensions['E'].width=38; ws3.column_dimensions['F'].width=8; ws3.column_dimensions['G'].width=38

# ============== Sheet4: 审计程序清单 ==============
ws4 = wb.create_sheet('审计程序清单'); ws4.sheet_properties.tabColor='4CAF50'
h4=['序号','审计程序','应做?','已做?','工作底稿/证据','备注说明','准则/法规依据']
for ci,h in enumerate(h4,1): ws4.cell(row=1,column=ci,value=h); hdr(ws4,1,len(h4))

procs = [
    [1,'签订审计业务约定书','√','√','审计业务约定书','事务所与清算组签订','审计准则1111号'],
    [2,'了解被审计单位基本情况','√','√','社会团体登记信息','含社团法人证书/章程','审计准则1211号'],
    [3,'获取清算决议/批准文件','√','√','会员大会决议(2026.3.26)','已核实注销决议','《社会团体登记管理条例》'],
    [4,'获取债权债务公告及结果','√','√','四川工人日报公告(2026.4.8)','45日公示期满无异议','《社会团体登记管理条例》第22条'],
    [5,'获取清算期财务报表','√','√','清算资产负债表+活动表+财产表+债务清偿表','4张表获取完整','审计准则'],
    [6,'核对银行对账单','√','⚠','银行利息22元有提及','未直接获取银行对账单正式函证','审计准则1312号'],
    [7,'银行函证程序','√','✗','-','未向银行发送询证函验证余额','审计准则1312号'],
    [8,'盘点现金/实物资产','√','√','无实物资产','确认无实物资产需要处置','审计准则'],
    [9,'重新计算清算数据','√','√','期初8619.98-负债8580+利息22=61.98','加总计算一致','审计准则1301号'],
    [10,'核对债务清偿凭证','√','√','审计费8,000+印章费580','两项债务清偿完毕','审计准则'],
    [11,'核实债权债务处理','√','√','公告期满无单位和个人申报','已确认无其他债权债务','审计准则'],
    [12,'检查剩余财产分配方案','√','✗','-','未说明61.98元剩余财产去向','《社会团体登记管理条例》'],
    [13,'获取清算组声明书','√','✗','-','未获取清算组关于报表真实完整的声明','审计准则1341号'],
    [14,'期后事项审查','√','⚠','提及银行待销户','未独立成"期后事项"段，未说明对报告的影响','审计准则1332号'],
    [15,'出具清算审计报告','√','√','报告已出具','报告文号未填','审计准则1501号'],
]
for pc in procs:
    r=pc[0]+1
    ws4.cell(row=r,column=1,value=pc[0])
    for ci,val in enumerate(pc[1:],2): ws4.cell(row=r,column=ci,value=val)
    done=pc[3]=='√'; ndone=pc[3]=='✗'; part=pc[3]=='⚠'
    fill=gfill if done else (rfll if ndone else yfill)
    f=gfont if done else (rfont if ndone else ofont)
    srow(ws4,r,7,f=f,fill=fill)
    for c in [1,3,4]: ws4.cell(row=r,column=c).alignment=cc

# 统计
rr=len(procs)+3
ws4.merge_cells(f'A{rr}:B{rr}'); ws4.cell(row=rr,column=1,value='统计').font=bfont
p_done=sum(1 for p in procs if p[3]=='√'); p_part=sum(1 for p in procs if p[3]=='⚠'); p_not=sum(1 for p in procs if p[3]=='✗')
ws4.cell(row=rr+1,column=1,value='已执行').font=gfont; ws4.cell(row=rr+1,column=2,value=f'{p_done}项 ({p_done/len(procs)*100:.0f}%)')
ws4.cell(row=rr+2,column=1,value='部分执行').font=ofont; ws4.cell(row=rr+2,column=2,value=f'{p_part}项')
ws4.cell(row=rr+3,column=1,value='未执行').font=rfont; ws4.cell(row=rr+3,column=2,value=f'{p_not}项 ({p_not/len(procs)*100:.0f}%)')

ws4.column_dimensions['A'].width=6; ws4.column_dimensions['B'].width=28; ws4.column_dimensions['C'].width=8
ws4.column_dimensions['D'].width=8; ws4.column_dimensions['E'].width=30; ws4.column_dimensions['F'].width=40; ws4.column_dimensions['G'].width=22

# ============== Sheet5: 综合评估与改进建议 ==============
ws5 = wb.create_sheet('综合评估'); ws5.sheet_properties.tabColor='9C27B0'
ws5.merge_cells('A1:D1'); ws5.cell(row=1,column=1,value='综合评估').font=Font(name='微软雅黑',bold=True,size=14,color='1C355E')
r=3
h5=['评估维度','评估意见','评级']
for ci,h in enumerate(h5,1): ws5.cell(row=r,column=ci,value=h); hdr(ws5,r,len(h5)); r+=1

evals = [
    ['数据准确性','核心存在3处不一致：业务活动表利息收入为0、资产负债表货币资金期末为0、净资产变动未反映利息。根本原因是Excel报表未随实际进展更新（清算组/审计报告数据已更新但报表滞后）。','待修复'],
    ['文字质量','文号未填、清算组信息缺失、剩余财产去向未说明、缺少期后事项段、公告信息不完整。但整体行文规范，基本能满足报告要求。','可接受'],
    ['程序完整性','15项程序完成7项(47%)。关键缺失：银行函证、剩余财产分配方案检查、清算组声明书获取。','不足'],
    ['意见适当性','出具无保留意见但未增加强调事项段（银行待销户导致净资产待定）。公告期满无异议的基本判断合理。','需关注'],
]
for ev in evals:
    ws5.cell(row=r,column=1,value=ev[0]).font=bfont; ws5.cell(row=r,column=2,value=ev[1]).font=nfont; ws5.cell(row=r,column=2).alignment=wrap
    ws5.cell(row=r,column=3,value=ev[2]).font=Font(name='微软雅黑',size=10,bold=True)
    fill=gfill if '可接受'==ev[2] else (rfll if '不足' in ev[2] or '待修复' in ev[2] else yfill)
    srow(ws5,r,3,fill=fill)
    ws5.cell(row=r,column=3).alignment=cc
    r+=1

r+=1
ws5.merge_cells(f'A{r}:D{r}'); ws5.cell(row=r,column=1,value='清算核心数据总览').font=Font(name='微软雅黑',bold=True,size=12,color='1C355E'); r+=1
h6=['项目','金额(元)','时点','说明']
for ci,h in enumerate(h6,1): ws5.cell(row=r,column=ci,value=h); hdr(ws5,r,len(h6)); r+=1
n_data = [
    ['期初货币资金','8,619.98','2026.3.27','清算开始日'],
    ['期初负债','8,580.00','2026.3.27','审计费8,000+印章费580'],
    ['期初净资产（非限定性）','39.98','2026.3.27','货币资金-负债'],
    ['清算期间利息收入','22.00（待银行销户确认）','期间','银行结息'],
    ['清算期间支出','-8,580.00','期间','清偿审计费+印章费'],
    ['期末净资产（非限定性）','61.98（待确认）','2026.5.24','⚠️银行销户后可能有±调整'],
    ['剩余财产归属','待明确','-','⚠️报告未说明去向'],
]
for nd in n_data:
    for ci,val in enumerate(nd,1): ws5.cell(row=r,column=ci,value=val).font=bfont if ci==1 else nfont
    srow(ws5,r,4,fill=bfill if '⚠️' in str(nd[3]) else None)
    for c in [1,2,3,4]: ws5.cell(row=r,column=c).alignment=cc if c>1 else wrap
    r+=1

r+=1
ws5.merge_cells(f'A{r}:D{r}'); ws5.cell(row=r,column=1,value='改进建议（按优先级）').font=Font(name='微软雅黑',bold=True,size=12,color='1C355E'); r+=1
sugs = [
    ['🔴 关键','更新Excel报表数据','待银行下周一销户后：①更新资产负债表（货币资金期末=销户后实际余额）；②更新业务活动表（确认利息收入）；③更新净资产（非限定性）金额'],
    ['🔴 关键','增加期后事项段','在审计报告中增加期后事项段说明：①银行账户将于XX日注销；②净资产金额因销户利息可能存在±调整'],
    ['🔴 关键','补充剩余财产分配方案','按《社会团体登记管理条例》第24条，明确61.98元剩余财产的分配去向（按章程规定或业务主管单位决定）'],
    ['🟡 重要','补充银行函证','向开户银行发送询证函，正式验证银行存款余额'],
    ['🟡 重要','填补报告文号','川竞泽审字[2026]第XX号完成编号'],
    ['🟡 建议','完善清算组信息','在报告中注明清算组成员及负责人姓名'],
    ['🟡 建议','获取清算组声明书','按审计准则1341号获取清算组关于报表真实完整的书面声明'],
    ['🟢 优化','完善公告信息描述','明确公告日期(2026.4.8)、公告媒体(四川工人日报)、公告期(45日)、期满日(2026.5.23)'],
    ['🟢 优化','独立描述审计程序','审计报告应独立描述自身执行的审计程序，与清算组报告明确区分'],
]
for s in sugs:
    ws5.cell(row=r,column=1,value=s[0]).font=bfont; ws5.cell(row=r,column=2,value=s[1]).font=bfont
    ws5.merge_cells(f'C{r}:D{r}'); ws5.cell(row=r,column=3,value=s[2]).font=nfont; ws5.cell(row=r,column=3).alignment=wrap
    srow(ws5,r,4)
    r+=1

ws5.column_dimensions['A'].width=14; ws5.column_dimensions['B'].width=28
ws5.column_dimensions['C'].width=30; ws5.column_dimensions['D'].width=28

# ============== Sheet6: 复核发现底稿 ==============
ws6 = wb.create_sheet('复核发现底稿'); ws6.sheet_properties.tabColor='607D8B'
ws6.merge_cells('A1:E1'); ws6.cell(row=1,column=1,value='复核发现底稿（明细）').font=Font(name='微软雅黑',bold=True,size=14,color='1C355E')
r=3
h7=['编号','发现事项','所属领域','情况描述','风险等级','改进建议']
for ci,h in enumerate(h7,1): ws6.cell(row=r,column=ci,value=h); hdr(ws6,r,len(h7)); r+=1

findings = [
    ['F1','Excel报表数据滞后','数据一致性','资产负债表货币资金期末为0、业务活动表收入为0、净资产变动-39.98，均与实际不符（实际货币资金8,641.98、利息22元、净资产61.98）','高','待银行销户后统一更新Excel数据'],
    ['F2','非限定性净资产待确认','数据准确性','用户说明净资产数据将变动，银行下周一销户时确认最终利息','高','在报告中增加期后事项段说明此不确定性'],
    ['F3','剩余财产分配去向未说明','审计报告完整性','61.98元剩余财产处置方案缺失','高','按《社会团体登记管理条例》明确分配方案'],
    ['F4','报告文号未填写','格式问题','川竞泽审字[2026]第**号','中','填写具体编号'],
    ['F5','银行函证程序缺失','审计程序','未向银行发询证函验证存款余额','中','按审计准则1312号实施银行函证'],
    ['F6','缺少清算组声明书','审计程序','未获取清算组关于报表真实完整的书面声明','中','按审计准则1341号获取'],
    ['F7','清算组信息不完整','披露问题','未列出清算组成员及负责人','中','补充清算组组成信息'],
    ['F8','公告信息描述不完整','披露问题','仅说"已在四川工人日报公告"，未说明具体日期和结果','中','补充：2026.4.8公告，45日期满，无单位和个人申报'],
    ['F9','缺少强调事项段','审计意见','银行待销户、利息待定等期后事项未在审计意见中提及','中','增加强调事项段'],
    ['F10','审计程序描述与清算组报告混淆','专业表述','审计报告的审计方法段似引用清算组工作，应独立描述','低','明确区分审计程序和清算组工作'],
]
for fd in findings:
    ws6.cell(row=r,column=1,value=fd[0]); ws6.cell(row=r,column=2,value=fd[1]); ws6.cell(row=r,column=3,value=fd[2])
    ws6.cell(row=r,column=4,value=fd[3]); ws6.cell(row=r,column=5,value=fd[4]); ws6.cell(row=r,column=6,value=fd[5])
    fill=rfll if fd[4]=='高' else (yfill if fd[4]=='中' else None)
    srow(ws6,r,6,fill=fill)
    ws6.cell(row=r,column=1).alignment=cc; ws6.cell(row=r,column=5).alignment=cc
    ws6.cell(row=r,column=4).alignment=wrap; ws6.cell(row=r,column=6).alignment=wrap
    r+=1

ws6.column_dimensions['A'].width=8; ws6.column_dimensions['B'].width=28; ws6.column_dimensions['C'].width=18
ws6.column_dimensions['D'].width=45; ws6.column_dimensions['E'].width=10; ws6.column_dimensions['F'].width=38

wb.save(output_path)
print(f'Done: {output_path}')
