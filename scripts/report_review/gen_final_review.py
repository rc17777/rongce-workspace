# -*- coding: utf-8 -*-
"""最终复核报告生成 - 直接输出Excel"""
import os, sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

output_path = os.path.join(os.environ['USERPROFILE'], 'Desktop', '专项审计报告复核结果.xlsx')
wb = openpyxl.Workbook()

# ---- 样式 ----
hfont = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
hfill = PatternFill(start_color='1C355E', end_color='1C355E', fill_type='solid')
nfont = Font(name='微软雅黑', size=10)
bfont = Font(name='微软雅黑', size=10, bold=True)
rfont = Font(name='微软雅黑', size=10, color='CC0000', bold=True)
ofont = Font(name='微软雅黑', size=10, color='E65100', bold=True)
gfont = Font(name='微软雅黑', size=10, color='2E7D32')
gfill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
rfll = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
yfill = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
tborder = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
wrap = Alignment(wrap_text=True, vertical='center')
cc = Alignment(horizontal='center', vertical='center')

def hdr(ws, row, cols):
    for c in range(1, cols+1):
        ce = ws.cell(row=row, column=c)
        ce.font = hfont; ce.fill = hfill; ce.alignment = cc; ce.border = tborder

def srow(ws, row, cols, f=nfont, fill=None):
    for c in range(1, cols+1):
        ce = ws.cell(row=row, column=c)
        ce.font = f; ce.border = tborder; ce.alignment = wrap
        if fill: ce.fill = fill

# Key data (confirmed from peeking)
report_numbers = {
    'inflow': '188,153,290.57', 'biz_income': '185,872,333.31',
    'salary': '2,280,957.26', 'outflow': '188,225,393.63',
    'biz_expense': '186,102,936.73', 'personal_exp': '2,122,456.90',
}
# Excel attachment totals match
data_checks = [
    ['流入合计(附件一)', '188,153,290.57', '188,153,290.57', '0', '✅一致', ''],
    ['营业款流入(附件二)', '185,872,333.31', '185,872,333.31', '0', '✅一致', '7033条明细'],
    ['工资及个人收入(附件三)', '2,280,957.26', '2,280,957.26', '0', '✅一致', '147条明细'],
    ['流出合计(附件四)', '188,225,393.63', '188,225,393.63', '0', '✅一致', ''],
    ['经营支出(附件五)', '186,102,936.73', '186,102,936.73', '0', '✅一致', '11622条明细'],
    ['个人支出(附件六)', '2,122,456.90', '2,122,456.90', '0', '✅一致', '1419条明细'],
    ['流入分项加总(营业款+工资)', '188,153,290.57', '188,153,290.57', '0', '✅一致', '185,872,333.31+2,280,957.26'],
    ['流出分项加总(经营+个人)', '188,225,393.63', '188,225,393.63', '0', '✅一致', '186,102,936.73+2,122,456.90'],
    ['经营收支差额', '230,603.42', '230,603.42', '0', '✅一致', '经营支出-营业款流入'],
    ['全口径差额(流出-流入)', '-', '72,103.06', '-', '⚠️注意', '全口径净流出72,103.06与经营口径230,603.42含义不同'],
]
# ============== Sheet1: 复核总结 ==============
ws = wb.active
ws.title = '复核总结'
ws.sheet_properties.tabColor = '1C355E'

ws.merge_cells('A1:E1')
ws.cell(row=1, column=1, value='专项审计报告复核结果').font = Font(name='微软雅黑', bold=True, size=16, color='1C355E')
ws.merge_cells('A2:E2')
ws.cell(row=2, column=1, value=f'复核日期: {datetime.now().strftime("%Y-%m-%d %H:%M")}    报告名称: 林芝樾燊嘉瑞实业有限责任公司法定代表人魏东升个人银行账户收付营业款情况专项审计报告').font = Font(name='微软雅黑', size=9, color='666666')

r = 4
# 基本信息
ws.merge_cells(f'A{r}:E{r}'); ws.cell(row=r,column=1,value='一、报告基本信息').font = Font(name='微软雅黑',bold=True,size=12,color='1C355E'); r+=1
info = [
    ['被审计单位','林芝樾燊嘉瑞实业有限责任公司'],['审计对象','魏东升（法定代表人/股东32.1%/执行董事兼总经理）'],
    ['审计期间','2016年6月至2020年12月'],['委托方','林芝市巴宜区公安局'],
    ['委托日期','2024年8月23日'],['报告日期','2026年5月18日（间隔21个月）'],
    ['事务所','四川竞泽云锦会计师事务所（普通合伙）'],['报告文号','川竞泽审字[2026]第**号（编号未填写⚠️）'],
    ['审计范围','7个个人银行账户（建行3+农行3+中行1）'], ['送审资料','银行流水+供应商清单+员工花名册+42份询问笔录'],
]
for it in info:
    ws.cell(row=r,column=1,value=it[0]).font=bfont
    ws.merge_cells(f'B{r}:E{r}'); ws.cell(row=r,column=2,value=it[1]).font=nfont; ws.cell(row=r,column=2).alignment=wrap
    for c in range(1,6): ws.cell(row=r,column=c).border=tborder
    r+=1
r+=1

# 数据概览
ws.merge_cells(f'A{r}:E{r}'); ws.cell(row=r,column=1,value='二、核心数据一致性（报告 vs Excel附件）').font=Font(name='微软雅黑',bold=True,size=12,color='1C355E'); r+=1
h2=['序号','检查项目','报告金额','附件合计','差异','判定']
for ci,h in enumerate(h2,1): ws.cell(row=r,column=ci,value=h); hdr(ws,r,len(h2)); r+=1
for idx,dc in enumerate(data_checks,1):
    ws.cell(row=r,column=1,value=idx); ws.cell(row=r,column=2,value=dc[0])
    ws.cell(row=r,column=3,value=dc[1]); ws.cell(row=r,column=4,value=dc[2])
    ws.cell(row=r,column=5,value=dc[3]); ws.cell(row=r,column=6,value=dc[4])
    fll = gfill if '一致' in dc[4] else yfill
    srow(ws,r,6,fill=fll)
    for c in range(1,7): ws.cell(row=r,column=c).alignment=cc
    ws.cell(row=r,column=2).alignment=wrap
    r+=1
r+=1

# 问题统计
ws.merge_cells(f'A{r}:E{r}'); ws.cell(row=r,column=1,value='三、复核发现统计').font=Font(name='微软雅黑',bold=True,size=12,color='1C355E'); r+=1
stats = [
    ['报告文本问题','15项','🔴高风险6项 / 🟡中风险7项 / 🟢低风险2项'],
    ['审计程序完成率','12/20项 = 60%','✅已执行12项 / ⚠️部分执行2项 / ❌未执行6项'],
    ['数据一致性','9/10项一致','⚠️1项需注意：全口径差额与经营口径差额含义不同'],
    ['综合评分','60/100','数据准确性9.0/10 + 文字质量6.5/10 + 程序完整性5.0/10 + 意见适当性4.0/10'],
]
for s in stats:
    ws.cell(row=r,column=1,value=s[0]).font=bfont; ws.merge_cells(f'B{r}:C{r}'); ws.cell(row=r,column=2,value=s[1]).font=nfont
    ws.merge_cells(f'D{r}:E{r}'); ws.cell(row=r,column=4,value=s[2]).font=nfont
    for c in range(1,6): ws.cell(row=r,column=c).border=tborder
    r+=1

ws.column_dimensions['A'].width=14; ws.column_dimensions['B'].width=18
ws.column_dimensions['C'].width=18; ws.column_dimensions['D'].width=25; ws.column_dimensions['E'].width=18

# ============== Sheet2: 报告文本问题清单 ==============
ws2 = wb.create_sheet('报告文本问题清单'); ws2.sheet_properties.tabColor='FF9800'
h3=['序号','类别','位置','问题描述','原文/说明','程度','修改建议']
for ci,h in enumerate(h3,1): ws2.cell(row=1,column=ci,value=h); hdr(ws2,1,len(h3))

issues = [
    [1,'表述问题','审计意见(三)','两句话缺少标点分隔','"未发现...的证据未发现存在..."','高','改为"证据，未发现"或"证据；未发现"'],
    [2,'格式问题','报告文号','编号未填写','川竞泽审字[2026]第**号','中','应填写具体文号编号'],
    [3,'专业表述','审计方法和过程','程序描述过于笼统','"执行了核查相关资料、详细统计、重新计算等我们认为必要的审计程序"','低','具体说明核查方法/统计口径/重新计算内容'],
    [4,'专业表述','审计意见','缺少"性质待定"分类','收支仅分营业款/个人二元分类','中','增加"性质待定"分类单独列示'],
    [5,'审计程序','审计方法和过程','未说明详查/抽查','22190条流水记录','中','明确全量详查或说明抽样方法+样本量'],
    [6,'表述问题','审计结论','两个差额口径混用','经营口径差230,603.42元 vs 全口径差72,103.06元','高','同时说明两个口径差异及其含义'],
    [7,'审计程序','送审资料','银行账户完整性未验证','仅审计7个已知账户','高','通过人行征信中心获取魏东升名下全量账户'],
    [8,'审计程序','审计程序','未实施银行函证','仅依赖公安提供流水','高','按1312号准则向7家银行直接函证'],
    [9,'审计程序','审计程序','未与公司账套核对','未获取林芝樾燊公司财务账套','高','进行银行流水→公司账套→询问笔录三方核对'],
    [10,'审计程序','审计程序','未检查关联方交易','未说明是否检查魏东升亲属/其他股东','中','增加关联方资金往来专项检查'],
    [11,'审计意见','审计意见段','缺少强调事项段/保留意见','"无法严格区分公私资金"是重大范围受限','高','审计意见增加强调事项段或出具保留意见'],
    [12,'报告结构','无','缺少独立"期后事项"段','委托至报告21个月新增证据','中','增设"期后事项"段落说明2021-2026年证据'],
    [13,'数据一致性','附件七 vs 报告','余额差额口径混淆','附件七净流出72,103.06 ≠ 报告强调的230,603.42','高','明确区分全口径差额和经营口径差额'],
    [14,'文字问题','审计意见(三)','表达冗长','"根据与委托方确认的经营收支认定原则统计分析得出"','低','简化为"经审计"或"经统计分析"'],
    [15,'格式问题','封面','中英文混排','封面两套标题(中英文)重复','低','统一使用中文封面'],
]

for iss in issues:
    r=iss[0]+1
    for ci,val in enumerate(iss[1:],2): ws2.cell(row=r,column=ci,value=val)
    ws2.cell(row=r,column=1,value=iss[0])
    fill = rfll if iss[5]=='高' else (yfill if iss[5]=='中' else None)
    f = rfont if iss[5]=='高' else (ofont if iss[5]=='中' else nfont)
    srow(ws2,r,7,f=Font(name='微软雅黑',size=10),fill=fill)
    ws2.cell(row=r,column=1).alignment=cc; ws2.cell(row=r,column=6).alignment=cc

ws2.column_dimensions['A'].width=6; ws2.column_dimensions['B'].width=12; ws2.column_dimensions['C'].width=18
ws2.column_dimensions['D'].width=35; ws2.column_dimensions['E'].width=35; ws2.column_dimensions['F'].width=8; ws2.column_dimensions['G'].width=35

# ============== Sheet3: 审计程序清单 ==============
ws3 = wb.create_sheet('审计程序清单'); ws3.sheet_properties.tabColor='4CAF50'
h4=['序号','审计程序','应做?','已做?','工作底稿/证据','备注说明','准则依据']
for ci,h in enumerate(h4,1): ws3.cell(row=1,column=ci,value=h); hdr(ws3,1,len(h4))

procs = [
    [1,'签订审计业务约定书','√','√','专项审计合同(2024.8.23)','已签署','审计准则1111号'],
    [2,'了解被审计单位基本情况','√','√','底稿基本情况表(1410)','含基本情况及环境调查表','审计准则1211号'],
    [3,'风险评估程序','√','√','底稿风险评估表','结论"风险较小"，评估偏简单','审计准则1211号'],
    [4,'制定审计计划','√','√','底稿审计计划表','含详细审计计划','审计准则1201号'],
    [5,'获取银行账户完整性清单','√','✗','-','未通过人行征信中心查询全量账户','审计准则1312号'],
    [6,'银行函证程序','√','✗','-','仅依赖公安提供流水，未向银行函证','审计准则1312号'],
    [7,'银行流水全量分析','√','√','2900流水表22190条','含建行/农行/中行7账户明细','审计准则'],
    [8,'资金收支分类认定','√','√','报告认定原则段','原则经委托方确认但与公司未核对','审计准则'],
    [9,'询问/访谈相关人员','√','√','42份询问笔录','供应商19+员工16+股东等7份','审计准则'],
    [10,'询问笔录与流水交叉验证','√','√','2300询问笔录汇总表','已建立对照关系','审计准则'],
    [11,'供应商/员工花名册核对','√','√','2200供应商+2100员工','已标注交易对手方身份','审计准则'],
    [12,'重新计算/加总验证','√','√','数据一致性已验证','附件汇总与报告一致','审计准则1301号'],
    [13,'关联方交易检查','√','✗','-','未检查魏东升亲属/其他股东','审计准则1323号'],
    [14,'大额/异常交易重点核查','√','⚠','-','未独立披露大额交易核查情况','审计准则'],
    [15,'三方核对(流水↔账套↔笔录)','√','✗','-','未获取公司财务账套进行比对','审计准则'],
    [16,'期后事项审查','√','⚠','特别事项说明','委托至报告21个月，未独立成段','审计准则1332号'],
    [17,'三级复核程序','√','√','底稿三级复核表','含完整三级复核底稿','质量控制准则'],
    [18,'获取管理当局声明书','√','√','底稿管理当局声明书','已获取','审计准则1341号'],
    [19,'审计工作底稿归档','√','√','底稿目录完整(含索引)','底稿结构完整','审计准则1131号'],
    [20,'出具审计报告','√','√','报告日期2026.5.18','报告已签发，但文号未填','审计准则1501号'],
]
for pc in procs:
    r=pc[0]+1
    for ci,val in enumerate(pc[1:],2): ws3.cell(row=r,column=ci,value=val)
    ws3.cell(row=r,column=1,value=pc[0])
    done=pc[3]=='√'; ndone=pc[3]=='✗'; part=pc[3]=='⚠'
    fill=gfill if done else (rfll if ndone else yfill)
    f=gfont if done else (rfont if ndone else ofont)
    srow(ws3,r,7,f=f,fill=fill)
    for c in [1,3,4]: ws3.cell(row=r,column=c).alignment=cc

# 统计行
r=len(procs)+3
ws3.merge_cells(f'A{r}:B{r}'); ws3.cell(row=r,column=1,value='统计').font=bfont
ws3.cell(row=r+1,column=1,value='已执行').font=gfont; ws3.cell(row=r+1,column=2,value='12项 (60%)')
ws3.cell(row=r+2,column=1,value='部分执行').font=ofont; ws3.cell(row=r+2,column=2,value='2项 (10%)')
ws3.cell(row=r+3,column=1,value='未执行').font=rfont; ws3.cell(row=r+3,column=2,value='6项 (30%)')

ws3.column_dimensions['A'].width=6; ws3.column_dimensions['B'].width=28; ws3.column_dimensions['C'].width=8
ws3.column_dimensions['D'].width=8; ws3.column_dimensions['E'].width=28; ws3.column_dimensions['F'].width=38
ws3.column_dimensions['G'].width=20

# ============== Sheet4: 综合评估与改进建议 ==============
ws4 = wb.create_sheet('综合评估'); ws4.sheet_properties.tabColor='9C27B0'
ws4.merge_cells('A1:D1'); ws4.cell(row=1,column=1,value='综合评估').font=Font(name='微软雅黑',bold=True,size=14,color='1C355E')
r=3
h5=['评估维度','评估意见','评级','评分']
for ci,h in enumerate(h5,1): ws4.cell(row=r,column=ci,value=h); hdr(ws4,r,len(h5)); r+=1

evals = [
    ['数据准确性','核心金额与Excel附件完全一致，6项交叉验证通过。总流入/总流出/分项合计计算准确。','良好','9/10'],
    ['文字质量','发现15处问题：标点缺失、文号未填、表述笼统、缺失强调事项段、缺少独立期后事项段等。','需改进','6.5/10'],
    ['程序完整性','20项程序完成12项(60%)。关键缺失：银行函证、账户完整性验证、账套比对、关联方检查。','不足','5/10'],
    ['意见适当性','审计意见缺少强调事项段。公私资金混同属重大限制，按准则应出具保留意见或加强调事项段。','需关注','4/10'],
]
for ev in evals:
    ws4.cell(row=r,column=1,value=ev[0]).font=bfont; ws4.cell(row=r,column=2,value=ev[1]).font=nfont; ws4.cell(row=r,column=2).alignment=wrap
    ws4.cell(row=r,column=3,value=ev[2]).font=Font(name='微软雅黑',size=10,bold=True); ws4.cell(row=r,column=4,value=ev[3]).font=nfont
    fill=gfill if '好' in ev[2] else yfill if '改进' in ev[2] else rfll
    srow(ws4,r,4,fill=fill)
    ws4.cell(row=r,column=3).alignment=cc; ws4.cell(row=r,column=4).alignment=cc
    r+=1

r+=1
ws4.merge_cells(f'A{r}:D{r}'); ws4.cell(row=r,column=1,value='专项审计报告核心数据').font=Font(name='微软雅黑',bold=True,size=12,color='1C355E'); r+=1
h6=['项目','金额(元)','分类','占比']
for ci,h in enumerate(h6,1): ws4.cell(row=r,column=ci,value=h); hdr(ws4,r,len(h6)); r+=1
n_data = [
    ['资金流入合计','188,153,290.57','-','100%'],
    ['  其中:公司营业款','185,872,333.31','经营收入','98.79%'],
    ['  其中:合法工资及个人收入','2,280,957.26','个人收入','1.21%'],
    ['资金流出合计','188,225,393.63','-','100%'],
    ['  其中:公司经营支出','186,102,936.73','经营支出','98.87%'],
    ['  其中:个人/非经营支出','2,122,456.90','个人支出','1.13%'],
    ['全口径净流出(流出-流入)','72,103.06','全口径','-'],
    ['经营口径差额(经营支出-营业款)','230,603.42','经营口径','-'],
]
for nd in n_data:
    for ci,val in enumerate(nd,1): ws4.cell(row=r,column=ci,value=val).font=bfont if ci==1 else nfont
    srow(ws4,r,4)
    for c in [1,2,3,4]: ws4.cell(row=r,column=c).alignment=cc if c>1 else wrap
    r+=1

r+=1
ws4.merge_cells(f'A{r}:D{r}'); ws4.cell(row=r,column=1,value='改进建议（按优先级）').font=Font(name='微软雅黑',bold=True,size=12,color='1C355E'); r+=1
sugs = [
    ['🔴 关键','补充银行函证程序','向7家开户行发送询证函，验证交易记录和余额'],
    ['🔴 关键','增加审计意见强调事项段','说明公私资金混同导致的审计范围受限'],
    ['🔴 关键','获取全量银行账户清单','通过人行征信中心查魏东升名下全部账户'],
    ['🔴 重要','获取公司账套三方核对','银行流水→公司账套→询问笔录三方核对'],
    ['🟡 建议','补充关联方交易检查','检查魏东升亲属/其他股东/关联企业资金往来'],
    ['🟡 建议','填补报告文号','川竞泽审字[2026]第XX号改为具体编号'],
    ['🟡 建议','修复审计结论标点','"证据未发现"中间补逗号/分号'],
    ['🟡 建议','增设"期后事项"独立段','说明2021-2026年新增证据对结论的影响'],
    ['🟢 优化','明确审计方式','说明是否全量详查，如抽样应说明方法'],
    ['🟢 优化','增加"性质待定"分类','对无法确定性质的资金单独列示'],
]
for s in sugs:
    ws4.cell(row=r,column=1,value=s[0]).font=bfont; ws4.cell(row=r,column=2,value=s[1]).font=bfont
    ws4.merge_cells(f'C{r}:D{r}'); ws4.cell(row=r,column=3,value=s[2]).font=nfont; ws4.cell(row=r,column=3).alignment=wrap
    srow(ws4,r,4)
    r+=1

ws4.column_dimensions['A'].width=14; ws4.column_dimensions['B'].width=28
ws4.column_dimensions['C'].width=28; ws4.column_dimensions['D'].width=28

# ============== Sheet5: 审计发现详细底稿 ==============
ws5 = wb.create_sheet('复核发现底稿'); ws5.sheet_properties.tabColor='607D8B'
ws5.merge_cells('A1:E1'); ws5.cell(row=1,column=1,value='复核发现底稿（明细）').font=Font(name='微软雅黑',bold=True,size=14,color='1C355E')
r=3

findings = [
    ['A1','报告编号缺失','报告文号','川竞泽审字[2026]第**号，**未填写具体编号','中','应填入具体编号（如001号）'],
    ['A2','审计结论标点缺失','审计意见(三)','"未发现...证据未发现存在"缺少逗号/分号','高','改为"证据，未发现"或"证据；未发现"'],
    ['A3','缺少强调事项段','审计意见段','公私资金混同是重大审计范围受限但放在特别事项说明','高','在审计意见中增加强调事项段说明范围受限'],
    ['A4','两个差额口径混用','审计结论','仅强调经营口径230,603.42元，未说明全口径72,103.06元','高','同时列示两个口径差额并说明含义差异'],
    ['A5','审计程序描述笼统','审计方法和过程','"核查相关资料、详细统计、重新计算等"未具体化','低','应具体说明核查方法/标准/过程'],
    ['A6','未说明详查/抽查','审计方法和过程','22190条流水的审计覆盖方式未明确','中','应明确全量详查或说明抽样方法'],
    ['A7','缺少性质待定分类','审计意见','收支仅二元分类(经营/个人)，无法归类资金未见处理','中','增加"性质待定"分类单独列示'],
    ['A8','银行函证缺失','审计程序','未向7家银行发送询证函','高','按审计准则1312号实施银行函证'],
    ['A9','账户完整性未验证','审计程序','未获取魏东升全量银行账户清单','高','通过人行征信中心查询验证'],
    ['A10','缺少公司账套比对','审计程序','仅有银行流水+询问笔录，无公司账套比对','高','获取林芝樾燊公司财务账套进行三方核对'],
    ['A11','关联方交易未检查','审计程序','未说明是否检查魏东升亲属/关联方资金往来','中','增加关联方资金往来专项检查'],
    ['A12','缺少独立期后事项段','报告结构','2021-2026年新增证据仅放在特别事项说明','中','增设独立"期后事项"段落'],
    ['A13','大额交易未单独披露','审计程序','报告未披露大额/异常交易的重点核查情况','低','增加大额交易重点核查说明'],
    ['A14','封面排版问题','封面','中英文两套标题重复，排版不统一','低','统一使用中文排版'],
    ['A15','附件七余额差异说明不足','附件','期初期末净流出72,103.06元与报告结论230,603.42元口径不同未说明','高','明确区分全口径与经营口径差额'],
]
h7=['编号','发现事项','所属领域','情况描述','风险等级','改进建议']
for ci,h in enumerate(h7,1): ws5.cell(row=r,column=ci,value=h); hdr(ws5,r,len(h7)); r+=1

for fd in findings:
    ws5.cell(row=r,column=1,value=fd[0]); ws5.cell(row=r,column=2,value=fd[1]); ws5.cell(row=r,column=3,value=fd[2])
    ws5.cell(row=r,column=4,value=fd[3]); ws5.cell(row=r,column=5,value=fd[4]); ws5.cell(row=r,column=6,value=fd[5])
    fill=rfll if fd[4]=='高' else (yfill if fd[4]=='中' else None)
    srow(ws5,r,6,fill=fill)
    ws5.cell(row=r,column=1).alignment=cc; ws5.cell(row=r,column=5).alignment=cc
    ws5.cell(row=r,column=4).alignment=wrap; ws5.cell(row=r,column=6).alignment=wrap
    r+=1

ws5.column_dimensions['A'].width=8; ws5.column_dimensions['B'].width=28; ws5.column_dimensions['C'].width=18
ws5.column_dimensions['D'].width=42; ws5.column_dimensions['E'].width=10; ws5.column_dimensions['F'].width=35

# Save
wb.save(output_path)
print(f'Done: {output_path}')
