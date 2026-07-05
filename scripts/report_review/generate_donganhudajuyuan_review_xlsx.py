import json
import sys
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

sys.stdout.reconfigure(encoding='utf-8')

OUT = Path(r'C:\Users\scrccpa\.openclaw\workspace\outputs\donganhudajuyuan_review')
XLSX_OUT = OUT / '东安湖大剧院运营补贴专项审计报告复核结果.xlsx'
INV = json.loads((OUT / 'file_inventory.json').read_text(encoding='utf-8'))

now = datetime.now().strftime('%Y-%m-%d %H:%M')

findings = [
    {
        '编号': 'F-001', '风险等级': 'P1重大', '模型维度': 'B2问题定性精确度/C2依据对应',
        '报告/底稿位置': '报告“一、单位及项目基本情况-3.考核结果运用”；底稿风险初步评价、取证单相关段落',
        '原文/现象': '报告及部分底稿写：“合格”：年度考核总分为80（包括）分以上，按照全年补贴的100%比例兑付。',
        '复核发现': '与《考核办法》不一致。《考核办法》明确：“良好”为80分（含）以上按100%兑付；“合格”为60分（含）—80分按90%兑付；“不合格”为60分以下按得分比例兑付。报告将“合格”档误写为80分以上100%兑付，属于政策依据表述错误。因本项目得分91分，不影响本次319.60万元结论，但影响报告严谨性。',
        '影响': '政策依据表述错误，若被委托方或监管方追问，容易被认为未准确引用考核办法。',
        '修改建议': '将年终考核兑付比例按《考核办法》原文改为：良好80分（含）以上按100%；合格60分（含）—80分按90%；不合格60分以下按得分比例兑付。',
        '状态': '需修改'
    },
    {
        '编号': 'F-002', '风险等级': 'P1重大', '模型维度': 'C2报告-取证单证据对应/C5全链路金额追踪',
        '报告/底稿位置': '专项审计底稿-风险初步评价',
        '原文/现象': '底稿写：第四年、第五年给予每年不超过600万元。',
        '复核发现': '与报告及《考核办法》不一致。报告和考核办法均为“第四年、第五年给予每年不超过1200万元”。该处属于底稿残留或复制错误。',
        '影响': '底稿与报告、依据文件不一致，削弱证据链一致性；虽不影响第一年度1700万元补贴计算，但影响归档质量。',
        '修改建议': '将底稿相关表述统一修改为“第四年、第五年给予每年不超过1200万元”。',
        '状态': '需修改底稿'
    },
    {
        '编号': 'F-003', '风险等级': 'P1重大', '模型维度': 'A2审计范围边界/B1逻辑一致性',
        '报告/底稿位置': '专项审计底稿-风险初步评价',
        '原文/现象': '底稿写：首次考核周期以剧院进行首次演出之日起，年度考核周期为6个月。',
        '复核发现': '与报告及《考核办法》“年度考核周期为12个月”不一致。',
        '影响': '考核周期是补贴兑付判断的核心条件，底稿错误会影响项目归档和复核判断。',
        '修改建议': '将底稿中“年度考核周期为6个月”修改为“年度考核周期为12个月”；如指半年考核，应明确为“半年考核周期为6个月”。',
        '状态': '需修改底稿'
    },
    {
        '编号': 'F-004', '风险等级': 'P2一般', '模型维度': 'B1逻辑一致性/C2证据对应',
        '报告/底稿位置': '专项审计底稿-风险初步评价；报告单位基本情况',
        '原文/现象': '底稿局部统一社会信用代码为“9151016MAE39T7D4F”；报告为“91510112MAE39T7D4F”。',
        '复核发现': '底稿中统一社会信用代码少“12”，与报告及取证单不一致。',
        '影响': '主体识别信息错误，属于底稿基础信息质量问题。',
        '修改建议': '将底稿统一社会信用代码统一为“91510112MAE39T7D4F”，并与营业执照/证照资料核对。',
        '状态': '需修改底稿'
    },
    {
        '编号': 'F-005', '风险等级': 'P2一般', '模型维度': 'B6审计目标覆盖度/C5金额追踪',
        '报告/底稿位置': '报告“二、审计结果”',
        '原文/现象': '报告写已支付第一笔850万元、第二笔530.40万元，第三笔补贴319.6万元。',
        '复核发现': '算术关系成立：1700-850-530.40=319.60万元；但报告未简要说明第二笔530.40万元的形成依据（例如半年考核得分/兑付比例）。',
        '影响': '第三笔金额虽然可勾稽，但第二笔已支付金额来源说明不足，报告可读性和可追溯性偏弱。',
        '修改建议': '建议在报告中补一句“第二笔补贴金额530.40万元系根据半年考核结果及已兑付资料确认”，或在附件/附表列示第一、二、三笔补贴计算过程。',
        '状态': '建议补充'
    },
    {
        '编号': 'F-006', '风险等级': 'P2一般', '模型维度': 'C2报告-取证单证据对应/C3取证单-附表溯源',
        '报告/底稿位置': '佐证资料目录',
        '原文/现象': '600份评价、场次确认表、安全生产、经营效益、社会评价、协议履约等PDF均为扫描件，当前未做OCR文本抽取。',
        '复核发现': '文件存在性已确认，但无法直接自动核验扫描件中具体场次、评价份数、营业收入、证照审批等明细内容。评分表显示91分及各项得分，但底层佐证细节尚未OCR穿透。',
        '影响': '影响AI自动复核的证据穿透深度；人工出具前应抽样核验关键扫描件。',
        '修改建议': '建议对关键扫描件进行人工抽样或OCR：场次确认表、营业收入情况、100/600份评价、协议履约、安全生产、证照资料。',
        '状态': '需人工/OCR核验'
    },
    {
        '编号': 'F-007', '风险等级': 'P2一般', '模型维度': 'B3整改/报告表达/B8摘要可读性',
        '报告/底稿位置': '报告全文',
        '原文/现象': '报告未附补贴计算明细表、评分表摘要或资料清单。',
        '复核发现': '专项审计报告结果较简洁，核心结论为319.60万元，但缺少可视化计算表和评分表关键摘要。',
        '影响': '委托方阅读报告时需回到底稿或佐证资料才能理解计算链条。',
        '修改建议': '建议增加附件或正文小表：全年补贴上限、年终得分、应兑付比例、应兑付全年金额、已兑付第一/二笔、第三笔应兑付金额。',
        '状态': '建议优化'
    },
]

amount_rows = [
    ['全年补贴上限', 1700.00, '万元', '报告、考核办法', '前三年每年不超过1700万元', '一致'],
    ['年度考核得分', 91.00, '分', '评分表、报告、底稿', '评分表实际得分91分', '一致'],
    ['应兑付比例', 100.00, '%', '考核办法', '80分（含）以上按100%兑付', '一致'],
    ['全年应兑付金额', 1700.00, '万元', '计算', '1700×100%', '一致'],
    ['第一笔已兑付', 850.00, '万元', '报告、底稿', '1700×50%', '一致'],
    ['第二笔已兑付', 530.40, '万元', '报告、底稿', '需补充半年考核依据；530.40/680=78%', '需说明来源'],
    ['第三笔应兑付', 319.60, '万元', '报告、底稿', '1700-850-530.40=319.60', '算术一致'],
]

score_rows = [
    ['不予兑付事项', '7项均为否', '', '评分表第4-10行', '未触发不予兑付条件'],
    ['安全生产', '10', '10', '评分表第12行', '满分'],
    ['演出场次', '30', '30', '评分表第13行', '满分'],
    ['经营效益-收入', '10', '10', '评分表第14行', '满分'],
    ['经营效益-综合目标', '5', '2', '评分表第15行', '②④⑤项不得分'],
    ['经营效益-就业人数', '5', '5', '评分表第16行', '满分'],
    ['协议履约-品牌项目', '5', '4', '评分表第17行', '②项不得分'],
    ['协议履约-培训人才', '5', '0', '评分表第18行', '未得分'],
    ['协议履约-审批运营', '5', '5', '评分表第19行', '满分'],
    ['社会评价-满意度', '10', '10', '评分表第20行', '满分'],
    ['社会评价-影响力', '10', '10', '评分表第21行', '满分'],
    ['社会评价-惠民演出', '5', '5', '评分表第22行', '满分'],
    ['合计', '100', '91', '评分表第23行', '与报告一致'],
]

wb = Workbook()
ws = wb.active
ws.title = '复核结论'

blue = '1F4E78'
gold = 'C5955C'
light = 'F5F2EC'
red = 'F4CCCC'
yellow = 'FFF2CC'
green = 'D9EAD3'
gray = 'E7E6E6'
white = 'FFFFFF'
thin = Side(style='thin', color='A6A6A6')
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_sheet(ws):
    ws.freeze_panes = 'A2'
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = border
    for cell in ws[1]:
        cell.fill = PatternFill('solid', fgColor=blue)
        cell.font = Font(color=white, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.sheet_view.showGridLines = False


def set_widths(ws, widths):
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

ws.append(['项目', '内容'])
summary = [
    ['复核对象', '关于东安湖大剧院运营项目运营补贴专项审计报告'],
    ['复核模型', '融策专项审计报告AI复核模型 v2.0（四层十五维）'],
    ['复核时间', now],
    ['复核结论', '修改后可出具。金额计算链条基本成立，但报告及底稿存在政策档位表述错误、底稿口径残留和证据穿透限制。'],
    ['P0致命问题', '0项'],
    ['P1重大问题', '3项：合格档兑付比例表述错误；底稿第四/第五年补贴金额错误；底稿年度考核周期错误。'],
    ['P2一般问题', '4项：主体信用代码底稿错误、第二笔补贴来源说明不足、扫描件未OCR、报告缺少计算明细附件。'],
    ['金额结论', '1700-850-530.40=319.60万元，第三笔补贴金额算术勾稽一致。'],
    ['证据限制', '多数佐证PDF为扫描件，当前未OCR，AI复核仅确认文件存在和评分表/底稿/报告勾稽，不能替代人工抽样核验。'],
    ['建议处理', '先修改P1问题；补充第三笔补贴计算明细；对关键扫描件进行抽样或OCR后归档。'],
]
for r in summary:
    ws.append(r)
style_sheet(ws)
set_widths(ws, [24, 110])
for r in range(2, ws.max_row+1):
    ws.cell(r,1).fill = PatternFill('solid', fgColor=light)
    ws.cell(r,1).font = Font(bold=True)

ws2 = wb.create_sheet('复核问题清单')
headers = list(findings[0].keys())
ws2.append(headers)
for f in findings:
    ws2.append([f[h] for h in headers])
style_sheet(ws2)
set_widths(ws2, [10, 12, 22, 34, 42, 58, 34, 54, 14])
for row in range(2, ws2.max_row+1):
    level = ws2.cell(row,2).value
    fill = red if 'P1' in level else yellow if 'P2' in level else green
    ws2.cell(row,2).fill = PatternFill('solid', fgColor=fill)

ws3 = wb.create_sheet('金额核对表')
ws3.append(['项目', '金额/数值', '单位', '来源', '复核说明', '结论'])
for r in amount_rows:
    ws3.append(r)
style_sheet(ws3)
set_widths(ws3, [22, 14, 10, 24, 58, 16])
for row in range(2, ws3.max_row+1):
    if '需' in str(ws3.cell(row,6).value):
        ws3.cell(row,6).fill = PatternFill('solid', fgColor=yellow)
    else:
        ws3.cell(row,6).fill = PatternFill('solid', fgColor=green)

ws4 = wb.create_sheet('评分表核对')
ws4.append(['考核项目', '全年计分', '考核得分', '来源', '备注'])
for r in score_rows:
    ws4.append(r)
style_sheet(ws4)
set_widths(ws4, [28, 14, 14, 24, 42])

ws5 = wb.create_sheet('资料清单与可抽取性')
ws5.append(['类别', '文件名', '类型', '页数', '文本字符数', '抽取状态', '复核影响', '路径'])
for rec in INV:
    status = rec.get('status')
    impact = '可文本复核' if status == 'text_extracted' else '扫描件/旧格式/未直接穿透，需人工或OCR核验' if status in ['no_embedded_text_or_scanned','legacy_doc_not_extracted'] else '已通过后续转换抽取' if rec['suffix']=='.doc' else '文件级核验'
    if rec['suffix'] == '.doc':
        impact = '已通过Word转换抽取底稿文本'
        status = 'converted_and_extracted'
    if rec['suffix'] == '.xlsx':
        impact = '已读取评分表关键行'
        status = 'xlsx_extracted'
    ws5.append([rec['category'], rec['name'], rec['suffix'], rec.get('pages'), rec.get('text_chars'), status, impact, rec['path']])
style_sheet(ws5)
set_widths(ws5, [12, 46, 10, 10, 14, 28, 42, 90])
for row in range(2, ws5.max_row+1):
    status = str(ws5.cell(row,6).value)
    if 'scanned' in status or 'OCR' in str(ws5.cell(row,7).value):
        ws5.cell(row,6).fill = PatternFill('solid', fgColor=yellow)
    elif 'extracted' in status:
        ws5.cell(row,6).fill = PatternFill('solid', fgColor=green)

ws6 = wb.create_sheet('建议改写')
ws6.append(['位置', '建议替换/补充文本'])
rewrite_rows = [
    ['考核结果运用段落', '第三笔补贴（年终考核补贴）兑付比例为：“良好”：年度考核总分为80分（含）以上，按照全年补贴的100%比例兑付；“合格”：年度考核总分为60分（含）—80分，按照全年补贴的90%兑付；“不合格”：年度考核总分为60分以下，按照得分比例划拨考核经费。'],
    ['审计结果段落', '根据区文广体旅局组织的考核工作组评分，成都东安湖大剧院运营项目补贴年终考核得分为91分，未触发不予兑付补贴事项，年度考核结果达到“良好”等级，按全年补贴100%比例计算，应兑付全年补贴1,700.00万元。'],
    ['补贴计算补充表述', '因已支付第一笔补贴850.00万元、第二笔补贴530.40万元，本次第三笔应兑付补贴金额为319.60万元（1,700.00-850.00-530.40）。建议同步附补贴计算明细表。'],
    ['底稿修正', '将底稿中“第四年、第五年每年不超过600万元”统一修正为“第四年、第五年每年不超过1200万元”；将“年度考核周期为6个月”修正为“年度考核周期为12个月”。'],
]
for r in rewrite_rows:
    ws6.append(r)
style_sheet(ws6)
set_widths(ws6, [28, 120])

# Add autofilters/tables where reasonable
for wsx in [ws2, ws3, ws4, ws5, ws6]:
    ref = f'A1:{get_column_letter(wsx.max_column)}{wsx.max_row}'
    wsx.auto_filter.ref = ref

# Row heights
for wsx in wb.worksheets:
    for row in range(1, wsx.max_row+1):
        wsx.row_dimensions[row].height = 42 if row > 1 else 28

wb.save(XLSX_OUT)
print(str(XLSX_OUT))
