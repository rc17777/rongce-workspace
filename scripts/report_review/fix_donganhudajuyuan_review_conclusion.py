import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill

sys.stdout.reconfigure(encoding='utf-8')

paths = [
    Path(r'C:\Users\scrccpa\.openclaw\workspace\outputs\donganhudajuyuan_review\东安湖大剧院运营补贴专项审计报告复核结果.xlsx'),
    Path(r'C:\Users\scrccpa\Desktop\东安湖大剧院运营补贴专项审计报告复核结果.xlsx'),
]

for path in paths:
    wb = openpyxl.load_workbook(path)
    ws = wb['复核问题清单']
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, headers['编号']).value == 'F-001':
            ws.cell(row, headers['风险等级']).value = 'P2一般'
            ws.cell(row, headers['模型维度']).value = 'B2问题定性精确度/B1文字口径一致性'
            ws.cell(row, headers['复核发现']).value = '经复核，报告“二、审计结果”所依据的年终考核得分为91分，按《考核办法》“良好：80分（含）以上按100%兑付”计算，319.60万元金额结论成立。需修正的是报告前文“考核结果运用”段落对“合格”档的引用表述：考核办法原文为“合格：60分（含）—80分按90%兑付”，报告误写为“合格：80分（含）以上按100%兑付”。该问题属于政策档位文字引用瑕疵，不影响本项目91分按100%兑付及第三笔补贴金额。'
            ws.cell(row, headers['影响']).value = '不影响本次319.60万元金额结论；影响报告引用依据的严谨性，建议出具前修正。'
            ws.cell(row, headers['修改建议']).value = '保留审计结果中“91分、按100%兑付、第三笔319.60万元”的结论；仅将前文“合格”档改为“60分（含）—80分按全年补贴90%兑付”。'
            ws.cell(row, headers['状态']).value = '建议修改'
            ws.cell(row, headers['风险等级']).fill = PatternFill('solid', fgColor='FFF2CC')
    ws0 = wb['复核结论']
    # B5/B6/B7 based on sheet layout from generator: rows 5-7 contain conclusion/P0/P1/P2-ish fields? Use labels for safety.
    for row in range(1, ws0.max_row + 1):
        label = ws0.cell(row, 1).value
        if label == '复核结论':
            ws0.cell(row, 2).value = '可出具前建议修改。金额计算链条成立，91分支持按100%兑付；需修正报告前文“合格档”文字引用、底稿残留口径和证据穿透限制。'
        elif label == 'P1重大问题':
            ws0.cell(row, 2).value = '2项：底稿第四/第五年补贴金额错误；底稿年度考核周期错误。'
        elif label == 'P2一般问题':
            ws0.cell(row, 2).value = '5项：报告合格档文字引用瑕疵、主体信用代码底稿错误、第二笔补贴来源说明不足、扫描PDF未OCR、报告缺少补贴计算明细附件。'
        elif label == '金额结论':
            ws0.cell(row, 2).value = '年终考核得分91分，按良好档100%兑付；1700-850-530.40=319.60万元，第三笔补贴金额算术勾稽一致。'
    ws6 = wb['建议改写']
    for row in range(2, ws6.max_row + 1):
        if ws6.cell(row, 1).value == '考核结果运用段落':
            ws6.cell(row, 2).value = '第三笔补贴（年终考核补贴）兑付比例为：“良好”：年度考核总分为80分（含）以上，按照全年补贴的100%比例兑付；“合格”：年度考核总分为60分（含）—80分，按照全年补贴的90%兑付；“不合格”：年度考核总分为60分以下，按照得分比例划拨考核经费。'
        elif ws6.cell(row, 1).value == '审计结果段落':
            ws6.cell(row, 2).value = '根据区文广体旅局组织的考核工作组评分，成都东安湖大剧院运营项目补贴年终考核得分为91分，达到“良好”等级，按全年补贴100%比例计算，应兑付全年补贴1,700.00万元。'
    wb.save(path)
    print(f'updated: {path}')
