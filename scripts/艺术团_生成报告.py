"""艺术团采购项目 — 全量分析Excel报告"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============ Styles ============
hdr_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
sub_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yel_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
grn_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
org_fill = PatternFill(start_color='FFDAB9', end_color='FFDAB9', fill_type='solid')
title_f = Font(name='微软雅黑', bold=True, size=13, color='2F5496')
red_f = Font(name='微软雅黑', size=11, color='CC0000', bold=True)
bold_f = Font(name='微软雅黑', bold=True, size=11)
norm_f = Font(name='微软雅黑', size=11)
sml_f = Font(name='微软雅黑', size=10)
thin_b = Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))
ca = Alignment(horizontal='center', vertical='center', wrap_text=True)
wa = Alignment(vertical='center', wrap_text=True)

def cell(ws, r, col, v, font=norm_f, fill=None, align=ca):
    c = ws.cell(row=r, column=col, value=v)
    c.font = font; c.alignment = align; c.border = thin_b
    if fill: c.fill = fill
    return c

def hdr_row(ws, r, headers):
    for i,h in enumerate(headers):
        cell(ws, r, i+1, h, hdr_font, hdr_fill)

# ============ Sheet 1: 问题清单 ============
ws1 = wb.active
ws1.title = '审计发现问题清单'

cell(ws1, 1, 1, '2025年校级艺术团技能培训与迎新晚会编导服务采购项目 审计发现问题清单', title_f, align=wa)
ws1.merge_cells('A1:H1')
cell(ws1, 2, 1, '项目编号: ZHH-F〔2025〕85号 | 采购人: 四川护理职业学院 | 代理机构: 四川正汇恒招标代理 | 采购方式: 竞争性磋商 | 预算: 180,000元', sml_f, align=wa)
ws1.merge_cells('A2:H2')

r=4
hdr_row(ws1, r, ['序号','问题类别','风险等级','问题描述','涉及方','证据/数据','审计建议','检测层'])

issues = [
    [1, '响应文件无效/损坏', '🔴 重大',
     '供应商"立美"提交的响应文件PDF(56.2MB)无法正常打开：PyMuPDF报告0页，pdfplumber报告"无/Root对象"。该文件不是有效的PDF文档，应被视为未提交有效响应文件。根据磋商文件第二章2.2.5条"通过资格审查的供应商不足3家的，终止本次采购活动"。',
     '立美(供应商)',
     'PyMuPDF: pages=0\npdfplumber: No /Root object\n文件大小: 56,193,715 bytes\n胤皓: 153页(正常)\n太格: 51页(正常)',
     '立即核查评审时立美的响应文件是否真实可读。如评审时文件已损坏，则实质性响应供应商仅2家，应终止采购。如评审后文件才损坏，需核实是否存在篡改。', '新增-文件完整性'],
    
    [2, '胤皓投标文件元数据被清除', '🟡 中等',
     '四川胤皓文化传媒有限公司的响应文件PDF(153页，72MB)所有元数据字段(Producer/Creator/Author/CreationDate)全部为空。WPS文字转PDF通常会保留Author/Creator信息，元数据被完全清除可能是人为操作，目的为规避L5元数据交叉检测。',
     '胤皓',
     'PyMuPDF提取：\nProducer: (空)\nCreator: (空)\nAuthor: (空)\nCreationDate: (空)\nModDate: (空)',
     '向胤皓索取原始.docx源文件，提取OLE2流中的作者、创建者、修订历史。元数据人为清除行为本身即可作为可疑信号。', 'L5-元数据'],
    
    [3, '太格投标文件为物理扫描件', '🟡 中等',
     '太格电子文档的响应文件PDF是物理扫描件，Producer=RICOH Pro 8100S(理光生产型打印机/扫描仪)。这意味着太格先将文件打印成纸质件，再用扫描仪扫描成PDF——而非.docx直接导出PDF。此操作流程异常，增大了文件被"二次制作"或"统一制作后分头扫描"的可能性。',
     '太格',
     'Producer: RICOH Pro 8100S\nCreator: RICOH Pro 8100S\n51页, 19MB，0字文字层\n扫描日期: 2025-04-02 03:53:48 (UTC-4)',
     '核查太格是否具备独立制作投标文件的能力。调取其历史投标文件，比对文件制作习惯(是否一贯使用物理扫描方式)。如其他项目均为.docx直接导出PDF，则本项目的扫描行为存在异常。', 'L5-元数据'],
    
    [4, '胤皓投标文件无文字层(全扫描)', '🟡 中等',
     '四川胤皓文化传媒有限公司的153页响应文件PDF全部为扫描图像，无任何可搜索文字层。即使元数据被清空，从数字文档导出的PDF通常会有可提取的文字层。胤皓的文件0字符可提取文本，可能是：1)打印后扫描；2)使用图片合成工具制作；3)刻意转换为图像以规避文本分析。',
     '胤皓',
     'PyMuPDF get_text(): 0 chars\npdfplumber extract_text(): 0 chars\n153页全部为图片，无文字层\n72MB文件大小(图片压缩质量较高)',
     '要求胤皓提供原始.docx投标文件进行比对。如无法提供，则投标文件的真实性存疑。扫描件投标不符合电子化采购的常规操作习惯。', 'L3/L6-文本/字体'],
    
    [5, '项目缺少文本/字体/样式检测基础', '🟡 中等',
     '因三家投标文件均为图片扫描件(胤皓)或物理扫描件(太格)或无效文件(立美)，L3(TF-IDF文本雷同)、L6(字体结构比对)均无法执行。客观上形成了规避文本分析检测的效果，使得串标围标检测缺少了一个关键维度。',
     '系统性问题',
     '胤皓: 0可提取字符\n太格: 0可提取字符\n立美: 0页(文件损坏)\nL3/L6检测: N/A',
     '将此作为制度性风险提示：电子化采购中，应要求供应商提交可检索的PDF文本层文件，禁止用扫描图片替代数字文档。', 'L3/L6-系统'],
    
    [6, '响应报价信息无法提取', '🔴 重大缺失',
     '三家投标文件因全部为扫描件/无效文件，报价金额无法通过程序化方式提取。目前通过72dpi OCR扫描了胤皓前10页和最后5页，未找到报价页。竞争性磋商的报价表可能在文件中间部分或作为单独密封件提交。',
     '三家全部',
     '胤皓: 153页全扫描，报价页位置待定位\n太格: 51页RICOH扫描，报价页待定位\n立美: 文件损坏\n(归档备案资料PDF也已损坏，无法通过归档记录获取)',
     '【最高优先级】通过以下渠道获取报价数据：\n1. 向采购代理机构(四川正汇恒)调取开标记录和最后报价记录\n2. 查询中国招标投标公共服务平台成交公告\n3. 向采购人(四川护理职业学院)调取评审报告', 'L1-报价'],
    
    [7, '归档备案资料PDF损坏', '🔴 重大缺失',
     '备案资料PDF(14.1MB)无法被任何PDF工具打开。PyMuPDF报"no objects found"，pdfplumber报"No /Root object"。该文件应包含评审过程记录、评委打分、中标结果等关键信息。文件的不可用性严重影响了审计的完整性。',
     '代理机构',
     '文件大小: 14,118,969 bytes\nPyMuPDF: FileDataError\npdfplumber: PDFSyntaxError\n两个独立PDF库均无法解析',
     '要求代理机构重新提供归档资料的原始文件。核实该文件是否为正确的PDF，还是文件传输/存储过程中损坏。如原始文件即如此，应追究代理机构归档责任。', '新增-归档'],
    
    [8, '图片层未发现跨公司重复', '🟢 已排除',
     '对胤皓(153张)和太格(51张)的PDF嵌入图片进行SHA256哈希比对，共204张图片实例，0张跨公司重复。立美PDF无法提取图片。本层检测未发现串标证据。',
     '胤皓/太格',
     '胤皓: 153张独立图片\n太格: 51张独立图片\n跨公司重复: 0\n立美: 0张(N/A)',
     '无进一步建议。', 'L4-图片'],
    
    [9, '中小企业和残疾人福利政策响应待核实', '🟢 关注',
     '竞磋文件未明确标注"专门面向中小企业"。需核实三家供应商是否按规定享受了价格扣除优惠，以及其提交的声明是否真实。',
     '三家全部',
     '待提取(扫描件限制)\n预算18万元，属于小额服务采购\n中小企业价格扣除通常为10%',
     '通过国家企业信用信息公示系统核实三家供应商的企业规模。如有应享受而未享受价格扣除的情况，应追溯评审合规性。', '追加-政策'],
    
    [10, '采购方式选择合理性', '🟢 关注',
     '本项目采用竞争性磋商方式，预算18万元，属于小额服务采购。竞争性磋商适用于"技术复杂或性质特殊，不能确定详细规格或具体要求"的情形。艺术培训与晚会编导服务是否符合此条件，需评估。如实际仅有1-2家有效供应商，竞争性不足。',
     '采购人/代理机构',
     '采购方式: 竞争性磋商\n预算: 180,000元\n磋商文件获取期: 3月21-27日(7天)\n递交截止: 4月3日(获取后仅7个工作日)',
     '核实项目是否在规定的采购方式范围内。磋商文件获取期仅7天偏短，可能影响潜在供应商的充分准备。', '追加-采购方式'],
]

for i, iss in enumerate(issues):
    row = 5 + i
    for j, val in enumerate(iss):
        if j == 2:
            fill = red_fill if '重大' in str(val) else (yel_fill if '中等' in str(val) else grn_fill)
            cell(ws1, row, j+1, val, bold_f, fill)
        elif j == 0:
            cell(ws1, row, j+1, val, norm_f)
        elif j in [3, 5, 6]:
            cell(ws1, row, j+1, val, j == 5 and sml_f or norm_f, align=wa)
        elif j == 4:
            cell(ws1, row, j+1, val, norm_f, align=wa)
        else:
            cell(ws1, row, j+1, val, norm_f)

for i in range(5, 5+len(issues)):
    ws1.row_dimensions[i].height = 110

ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 22
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 50
ws1.column_dimensions['E'].width = 16
ws1.column_dimensions['F'].width = 35
ws1.column_dimensions['G'].width = 42
ws1.column_dimensions['H'].width = 14

# ============ Sheet 2: 十层检测结果 ============
ws2 = wb.create_sheet('十层检测全量结果')

cell(ws2, 1, 1, '串标围标十层检测体系 — 本项目完整结果', title_f, align=wa)
ws2.merge_cells('A1:G1')
cell(ws2, 2, 1, '项目: ZHH-F〔2025〕85号 | 类型: 竞争性磋商(服务类) | 预算: 180,000元 | 重要提示: 所有投标文件均为扫描件/无效文件，L3/L6无法执行', sml_f, align=wa)
ws2.merge_cells('A2:G2')

r = 4
hdr_row(ws2, r, ['层级','检测维度','检测方法','本项目结果','风险等级','数据来源','说明'])

matrix = [
    ['L1','报价规律性','报价极差/标准差/预算对比','❓ 报价数据缺失','— 待补','投标响应文件','所有文件为扫描件/损坏，报价值无法程序化提取。需人工获取。'],
    ['L2','投标IP/MAC','电子交易系统登录日志','❓ 缺数据','— 待查','四川省政府采购一体化平台','本项目可能通过线下磋商方式进行。需向代理机构确认是否使用电子交易系统。'],
    ['L3','文本雷同(TF-IDF)','jieba→TF-IDF→余弦相似度','❓ N/A(全扫描件)','— 无法执行','N/A','胤皓0字符、太格0字符、立美0页。客观上形成规避文本分析的效果。'],
    ['L4','图片/资源哈希','PyMuPDF提取嵌入图片→SHA256','0跨公司重复(204实例)','🟢 低','投标PDF嵌入图片','胤皓153张+太格51张，无重复。立美无图片。'],
    ['L5','PDF元数据','PyMuPDF Author/Creator/Producer','胤皓:全部空白(可疑)\n太格:RICOH物理扫描\n立美:文件损坏','🟡 中','PDF元数据字段','胤皓元数据被完全清除(WPS正常会保留)。太格为物理扫描件。元数据可用性低。'],
    ['L6','文档结构/字体','PyMuPDF提取页面字体','❓ N/A(全扫描件)','— 无法执行','N/A','所有文件0文字层，无字体span可提取。'],
    ['L7','生成设备信息','PDF Producer/Creator字段','太格:RICOH Pro 8100S\n胤皓:(全部空白)\n立美:N/A','🟡 中','PDF Producer','太格使用打印→扫描流程而非数字导出，异常。胤皓元数据清空。'],
    ['L8','工商关联','天眼查/企查查','❓ 待查(网络超时)','— 待查','天眼查/企查查','需手动查询三家公司工商信息:股东/高管/对外投资交叉关系。'],
    ['L9','保证金/资金链','银行汇款凭证','本磋商文件未明确要求投标保证金','— N/A','磋商文件','小额服务采购通常不收取投标保证金。'],
    ['L10','代理/签到/得分','开标记录/评审报告','归档资料PDF损坏(14MB)，无法读取','🔴 缺失','归档备案资料','归档文件损坏导致评审过程完全不可见。需向代理机构重新调取。'],
    ['追加','立美文件有效性','PDF解析','立美响应文件.pdf 56MB → 0页','🔴 重大','投标文件','文件完全无效。实质性响应供应商可能仅2家。'],
    ['追加','胤皓文件真实性','元数据+文本层双重检查','153页全部为图片，元数据空白','🟡 中等','胤皓响应文件','数字文档导出应有文字层和元数据。完全空白+纯图片=可疑。'],
]

for i, row_data in enumerate(matrix):
    for j, val in enumerate(row_data):
        if j == 4:
            fill = red_fill if '高' in str(val) or '重大' in str(val) else (yel_fill if '中' in str(val) else (grn_fill if '低' in str(val) else None))
            cell(ws2, 5+i, j+1, val, bold_f, fill)
        elif j in [3, 6]:
            cell(ws2, 5+i, j+1, val, sml_f, align=wa)
        else:
            cell(ws2, 5+i, j+1, val, norm_f, align=wa if j != 0 else ca)

ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 16
ws2.column_dimensions['C'].width = 30
ws2.column_dimensions['D'].width = 35
ws2.column_dimensions['E'].width = 10
ws2.column_dimensions['F'].width = 22
ws2.column_dimensions['G'].width = 42

# ============ Sheet 3: L5元数据详情 ============
ws3 = wb.create_sheet('L5-元数据详情')
cell(ws3, 1, 1, 'L5 PDF元数据交叉分析', title_f, align=wa)
ws3.merge_cells('A1:G1')

r = 3
hdr_row(ws3, r, ['文件来源','文件名','页数','Producer','Creator','Author','创建日期','状态评估'])

meta_data = [
    ['招标文件','招标采购文件-ZHH-F〔2025〕85号磋商文件...','65','(空)','WPS 文字','Administrator','2025-03-20','正常(WPS导出)'],
    ['胤皓','四川胤皓文化传媒有限公司.pdf','153','(空)','(空)','(空)','(空)','🔴 元数据完全清空(可疑)'],
    ['太格','太格电子文档.pdf','51','RICOH Pro 8100S','RICOH Pro 8100S','(空)','2025-04-02 03:53','🟡 物理扫描件(非数字导出)'],
    ['立美','立美响应文件.pdf','0','N/A','N/A','N/A','N/A','🔴 文件损坏(0页/无Root对象)'],
    ['归档资料','备案资料-ZHH-F〔2025〕85号...','N/A','N/A','N/A','N/A','N/A','🔴 文件损坏(无法打开)'],
]

for i, row_data in enumerate(meta_data):
    for j, val in enumerate(row_data):
        if j == 7:
            fill = red_fill if '🔴' in str(val) else (yel_fill if '🟡' in str(val) else None)
            cell(ws3, 4+i, j+1, val, norm_f, fill, wa)
        else:
            cell(ws3, 4+i, j+1, val, norm_f, align=wa if j != 2 else ca)

r2 = 10
cell(ws3, r2, 1, '关键发现与推断', red_f, red_fill, wa)
ws3.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=7)

findings = [
    '1. 胤皓元数据完全清空：正常情况下，WPS文字导出PDF会保留Author(用户名)和CreationDate。元数据被完全清除(包括Producer)，说明有人为干预。目的可能是：a)隐藏原始文档作者信息；b)规避L5元数据交叉检测。此行为本身就构成可疑信号。',
    '2. 太格使用物理扫描：Producer=RICOH Pro 8100S表明文件经过"打印→物理扫描"流程。正常电子化投标应直接.docx→PDF。物理扫描增加了文件被中间人"统一制作后分头扫描"的风险。扫描时间2025-04-02 03:53(UTC-4=北京时间15:53)，在递交截止前约18小时。',
    '3. 立美文件完全无效：56MB的PDF文件0页，表明该文件要么从未被正确创建，要么在传输/存储过程中严重损坏。如果评审时该文件已不可读，则实质上只有2家供应商通过资格性审查，应终止采购活动。',
    '4. 归档资料同步损坏：备案资料PDF同样无法打开。考虑到立美文件也损坏，不排除存储介质/传输通道存在问题。但两个关键文件同时损坏的概率较低。',
    '5. 缺少交叉验证基础：由于胤皓无元数据、太格为扫描件、立美损坏，L5层无法进行Author/Creator/Producer的交叉比对。三家恰好规避了元数据检测的三种方式(清空/物理扫描/损坏)。'
]

for i, f in enumerate(findings):
    cell(ws3, 11+i, 1, f, norm_f, align=wa)
    ws3.merge_cells(start_row=11+i, start_column=1, end_row=11+i, end_column=7)

ws3.column_dimensions['A'].width = 16
ws3.column_dimensions['B'].width = 38
for c in range(3, 8):
    ws3.column_dimensions[get_column_letter(c)].width = 18

# ============ Sheet 4: 项目概况 ============
ws4 = wb.create_sheet('项目基本信息')
cell(ws4, 1, 1, '项目基本信息', title_f, align=wa)
ws4.merge_cells('A1:B1')

info = [
    ('项目名称','四川护理职业学院2025年校级艺术团专业技能培训与迎新晚会编导服务采购项目'),
    ('采购编号','ZHH-F〔2025〕85号'),
    ('采购人','四川护理职业学院'),
    ('采购人地址','成都市龙泉驿区龙都南路173号'),
    ('采购人联系人','赵老师 028-63955482'),
    ('代理机构','四川正汇恒招标代理有限公司'),
    ('代理机构地址','成都市高新区吉泰路666号1栋9层10号(花样年福年广场)'),
    ('代理机构联系人','陈先生 028-67171868'),
    ('代理机构邮箱','service@sc-zhh.com'),
    ('采购方式','竞争性磋商'),
    ('公告平台','中国招标投标公共服务平台'),
    ('采购预算','180,000.00元'),
    ('最高限价','180,000.00元'),
    ('评审方法','综合评分法(100分)'),
    ('评分构成','报价10分+履约能力6分+服务内容30分+实施方案20分+质量保障24分+应急方案10分'),
    ('中小企业政策','磋商文件未明确标注"专门面向中小企业"'),
    ('是否接受联合体','不接受联合体'),
    ('磋商文件获取期','2025年3月21日至3月27日(7天)'),
    ('响应文件递交截止','2025年4月3日10:30'),
    ('磋商地点','成都市高新区吉泰路666号1栋9层8号(花样年福年广场)'),
    ('响应有效期','递交响应文件截止日期之日起120天'),
    ('磋商文件售价','人民币300元/份(售后不退)'),
    ('供应商1','四川胤皓文化传媒有限公司 | 153页(扫描件) | 72MB | 元数据清空'),
    ('供应商2','太格电子文档 | 51页(RICOH扫描) | 19MB | 扫描日期2025-04-02'),
    ('供应商3','立美(全称待确认) | 0页(文件损坏) | 56MB | 无效响应文件'),
    ('归档资料状态','备案资料PDF(14MB) — 损坏，无法打开'),
]

for i, (k,v) in enumerate(info):
    cell(ws4, 3+i, 1, k, bold_f, sub_fill, wa)
    cell(ws4, 3+i, 2, v, norm_f, align=wa)

ws4.column_dimensions['A'].width = 22
ws4.column_dimensions['B'].width = 70

# ============ Sheet 5: 后续行动建议 ============
ws5 = wb.create_sheet('后续行动建议')
cell(ws5, 1, 1, '后续调查与取证建议(按优先级排序)', title_f, align=wa)
ws5.merge_cells('A1:E1')

r = 3
hdr_row(ws5, r, ['优先级','行动项','目的','执行方式','预期产出'])

actions = [
    ['P0-立刻','调取最后报价记录','获取三家供应商的实际报价金额，完成L1报价规律检测。','向代理机构(四川正汇恒)发函调取磋商记录和最后报价表。','报价数据→完成L1检测→判断是否存在价格串通。'],
    ['P0-立刻','调取评审报告','获取评委打分明细、成交结果、评审过程记录。','向采购人(四川护理职业学院)或代理机构调取。','评分数据→检查评分异常→完成L10检测。'],
    ['P0-立刻','核查立美响应文件有效性','确认评审时立美文件是否可读。如当时已损坏，则实质性响应供应商不足3家，应终止采购。','询问代理机构和评审专家：立美文件在评审时能否正常打开和审阅？','判断采购程序合法性。可能触发重新采购。'],
    ['P1-尽快','向胤皓调取原始.docx投标文件','核实投标文件真实性，提取原始作者/修订历史信息。','发函要求胤皓提供投标文件的原始可编辑电子版。','OLE2元数据→完成L5交叉验证→判断文件是否由第三方代制。'],
    ['P1-尽快','获取归档资料原件','备案资料PDF损坏，需获取可读版本。','要求代理机构重新提供归档资料的原始文件或光盘。','归档评审数据→评审一致性分析→完善审计证据链。'],
    ['P1-尽快','工商关联查询','核查三家供应商是否存在股东/高管/投资关联。','天眼查/企查查查询三家公司：(1)法定代表人 (2)股东 (3)董监高 (4)对外投资。','工商关联图→判断是否存在利益关联→L8检测完成。'],
    ['P1-尽快','胤皓资质核实','核实胤皓是否存在，是否具备承办艺术培训/晚会编导的能力。','国家企业信用信息公示系统查询+实地走访(如需)。','企业真实性和履约能力评估。'],
    ['P2-常规','采购方式合规性审查','竞争性磋商方式是否适用于该项目。18万小额服务是否可采用更简单的采购方式。','查阅四川省政府采购限额标准+本项目采购审批文件。','采购方式选择是否存在规避公开招标的嫌疑。'],
    ['P2-常规','磋商文件获取期评估','7天获取期是否充分，是否限制了潜在供应商。','比对同类项目的文件获取期惯例。','是否存在限制竞争的迹象。'],
]

for i, act in enumerate(actions):
    for j, val in enumerate(act):
        if j == 0:
            fill = red_fill if 'P0' in str(val) else (org_fill if 'P1' in str(val) else grn_fill)
            cell(ws5, 4+i, j+1, val, bold_f, fill)
        elif j in [3, 4]:
            cell(ws5, 4+i, j+1, val, sml_f, align=wa)
        else:
            cell(ws5, 4+i, j+1, val, norm_f, align=wa)

ws5.column_dimensions['A'].width = 12
ws5.column_dimensions['B'].width = 30
ws5.column_dimensions['C'].width = 35
ws5.column_dimensions['D'].width = 42
ws5.column_dimensions['E'].width = 38

# Save
out = r'C:\Users\scrccpa\Desktop\艺术团采购审计分析报告.xlsx'
wb.save(out)
print('Saved: ' + out)
