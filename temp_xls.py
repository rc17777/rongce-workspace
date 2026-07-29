# -*- coding: utf-8 -*-
"""测算数据基础表 Excel"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ====== COLOR / STYLE ======
HEADER_FILL = PatternFill(start_color='0A1F3F', end_color='0A1F3F', fill_type='solid')
HEADER_FONT = Font(name='微软雅黑', bold=True, color='FFFFFF', size=10)
SUB_FILL = PatternFill(start_color='C5955C', end_color='C5955C', fill_type='solid')
SUB_FONT = Font(name='宋体', bold=True, size=10)
DATA_FONT = Font(name='宋体', size=10)
BOLD_FONT = Font(name='宋体', bold=True, size=10)
ALT_FILL = PatternFill(start_color='F5F2EC', end_color='F5F2EC', fill_type='solid')
WARN_FONT = Font(name='宋体', bold=True, color='CC0000', size=10)
TITLE_FONT = Font(name='微软雅黑', bold=True, size=14, color='0A1F3F')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))
NUM_FMT = '#,##0.00'
PCT_FMT = '0.00%'

def write_header(ws, row, cols, fill=HEADER_FILL, font=HEADER_FONT):
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=j, value=c)
        cell.font = font; cell.fill = fill; cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def write_row(ws, row, vals, font=DATA_FONT, alt=False, fmt=NUM_FMT):
    for j, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=j, value=v)
        cell.font = font; cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        if isinstance(v, (int, float)):
            cell.number_format = fmt
        if alt:
            cell.fill = ALT_FILL

def write_title(ws, row, text):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)

RATE = 0.0799; K = 0.0799
CAPITAL_AMORT = 80_000_000; CAPITAL_LUMP = 22_298_077.79
A = CAPITAL_AMORT * RATE * (1+RATE)**18 / ((1+RATE)**18 - 1)

# ============================================================
# SHEET 1: 资本金回报测算表
# ============================================================
ws1 = wb.active; ws1.title = '表1-资本金回报'
write_title(ws1, 1, '表1：项目资本金回报测算表（单位：元）')
ws1.merge_cells('A2:G2')
cell = ws1['A2']; cell.value = f'P=80,000,000, k=7.99%, n=18, PMT={A:,.2f}/年; 剩余22,298,077.79第18年一次性支付'
cell.font = Font(name='宋体', size=9); cell.alignment = Alignment(horizontal='left')

headers = ['年', '等额本息本金', '等额本息利息', '等额本息小计', '一次性支付(第18年)', '资本金回报合计', '剩余(等额本息)']
write_header(ws1, 4, headers)

cr = CAPITAL_AMORT
for y in range(18):
    interest = cr * RATE; principal = A - interest
    if y == 17: principal = cr; interest = A - principal
    cr -= principal
    if cr < 0: cr = 0
    lump = CAPITAL_LUMP if y == 17 else 0
    vals = [2023+y, round(principal,2), round(interest,2), round(A,2),
            round(lump,2) if lump else '-', round(A+lump,2), round(cr,2) if cr > 0 else 0]
    write_row(ws1, 5+y, vals, alt=(y%2==1))

# totals
actual_pr_sum = 0; cr2 = CAPITAL_AMORT
for y in range(18):
    int_ = cr2 * RATE; pr_ = A - int_
    if y == 17: pr_ = cr2; int_ = A - pr_
    actual_pr_sum += pr_
    cr2 -= pr_
    if cr2 < 0: cr2 = 0

si_total = A*18 - actual_pr_sum
tot_vals = ['合计', round(actual_pr_sum,2), round(si_total,2), round(A*18,2),
            round(CAPITAL_LUMP,2), round(A*18+CAPITAL_LUMP,2), '-']
write_row(ws1, 23, tot_vals, font=BOLD_FONT)

ws1.column_dimensions['A'].width = 8
for c in range(2,8): ws1.column_dimensions[get_column_letter(c)].width = 18

# ============================================================
# SHEET 2: 银行融资本息
# ============================================================
ws2 = wb.create_sheet('表2-银行融资')
write_title(ws2, 1, '表2：银行融资本息测算表（单位：元）')
cell = ws2['A2']; cell.value = '逐年数据=原征求意见稿表2×0.997695(缩放至银行确认总数710,822,119.32)；非银行原始逐年还款计划'
cell.font = Font(name='宋体', color='CC0000', size=9)

headers2 = ['年', '原征求意见稿值', '缩放系数', '本报告取值', '剩余金额', '适用利率']
write_header(ws2, 4, headers2)

bo = [(31686745.84,381300000),(26809671.03,381300000),(24811756.73,381250000),
      (21351219.32,380250000),(20263854.17,379250000),(20265833.33,378250000),
      (20162465.30,377250000),(20111770.82,376250000),(23022951.39,372250000),
      (22871874.99,368250000),(22617395.82,364250000),(22414618.06,360250000),
      (27148298.64,351250000),(27728124.99,341250000),(35070729.16,323250000),
      (34158229.18,305250000),(311968819.45,0),(0,0)]
sf = 710822119.32 / sum(b[0] for b in bo)
rates = ['6.95%','6.45%','5.5%/5.0%','5.0%','5.0%','5.0%','5.0%','5.0%','5.0%',
         '5.0%','5.0%','5.0%','5.0%','5.0%','5.0%','5.0%','5.0%','-']

for y in range(18):
    orig, rem = bo[y]; scaled = round(orig * sf, 2)
    vals = [2023+y, orig, sf, scaled, rem if rem > 0 else '-', rates[y]]
    write_row(ws2, 5+y, vals, alt=(y%2==1))
    ws2.cell(row=5+y, column=3).number_format = '0.000000'

# totals
orig_sum = sum(b[0] for b in bo)
scaled_sum = sum(round(b[0]*sf,2) for b in bo)
tot2 = ['合计', round(orig_sum,2), '-', round(scaled_sum,2), '-', '—']
write_row(ws2, 23, tot2, font=BOLD_FONT)

ws2.column_dimensions['A'].width = 8
for i,c in enumerate([20,14,18,18,20,14]): ws2.column_dimensions[get_column_letter(i+2)].width = c

# ============================================================
# SHEET 3: 运维成本(含K验证)
# ============================================================
ws3 = wb.create_sheet('表3-运维成本')
write_title(ws3, 1, '表3：运营维护成本明细（含K=7.99%加成验证）')
cell = ws3['A2']; cell.value = '合同公式: 运维成本×(1+K), K=7.99%'
cell.font = Font(name='宋体', size=9)

headers3 = ['年', '底层成本(税前)', 'K系数(1+K)', '运维成本(含K)', '数据来源/验证']
write_header(ws3, 4, headers3)

# 2023: 第一经营年度
raw_2023 = 5341344.03; op_2023 = raw_2023 * (1+K)
vals3a = [2023, raw_2023, 1+K, round(op_2023,2),
          f'第一经营年度{raw_2023:,.2f}×1.0799={op_2023:,.2f} (与原表一致)']
write_row(ws3, 5, vals3a)

# 2024: 第二经营年度+第三经营年度partial
raw_2024_y2 = 5948956.93; raw_2024_y3 = 431456.22; raw_2024 = raw_2024_y2 + raw_2024_y3
op_2024_part = [round(raw_2024_y2*(1+K),2), round(raw_2024_y3*(1+K),2)]
op_2024 = sum(op_2024_part)
vals3b = [2024, round(raw_2024,2), 1+K, round(op_2024,2),
          f'Y2:{raw_2024_y2:,.2f}×1.0799={op_2024_part[0]:,.2f} + Y3:{raw_2024_y3:,.2f}×1.0799={op_2024_part[1]:,.2f}']
write_row(ws3, 6, vals3b)

# 2025: 审计数据
raw_2025 = 5713241.86; op_2025 = raw_2025 * (1+K)
vals3c = [2025, raw_2025, 1+K, round(op_2025,2),
          f'2025审计: 主营成本4,474,027.75+管理费1,239,214.11={raw_2025:,.2f}×1.0799 (独立验证)']
write_row(ws3, 7, vals3c, alt=True)

# 2026-2040: 估算
raw_future = 5710000; op_future = raw_future * (1+K)
for i in range(15):
    vals3f = [2026+i, raw_future, 1+K, round(op_future,2),
              f'按2025税前取整{raw_future:,.0f}×1.0799估算 (第{i+1}/15年)']
    write_row(ws3, 8+i, vals3f, alt=((8+i)%2==1))

# totals
raw_total = raw_2023 + raw_2024 + raw_2025 + raw_future * 15
op_total = sum([round(op_2023,2), round(op_2024,2), round(op_2025,2)] + [round(op_future,2)]*15)
tot3 = ['合计', round(raw_total,2), '-', round(op_total,2), f'税前总计{raw_total:,.2f}×1.0799={round(raw_total*(1+K),2):,.2f}']
write_row(ws3, 23, tot3, font=BOLD_FONT)

for c,w in zip('ABCDE', [8,20,12,20,70]):
    ws3.column_dimensions[c].width = w

# ============================================================
# SHEET 4: 第三方收入
# ============================================================
ws4 = wb.create_sheet('表4-第三方收入')
write_title(ws4, 1, '表4：第三方收入明细（单位：元）')
cell = ws4['A2']; cell.value = '合同公式中第三方收入不乘K系数，直接扣减'
cell.font = Font(name='宋体', size=9)

headers4 = ['年', '第三方收入', '数据来源']
write_header(ws4, 4, headers4)

inc_data = [(2023, 4285070.31, '沿用征求意见稿表3'),
            (2024, 6229093.45, '沿用征求意见稿表3 (含第三经营年度partial 387,927.58)'),
            (2025, 5564017.88, '2025审计: 主营收入4,961,142.62+其他收入602,875.26 (独立验证)')]
for i, (yr, inc, src) in enumerate(inc_data):
    write_row(ws4, 5+i, [yr, inc, src], alt=(i%2==1))
for i in range(15):
    write_row(ws4, 8+i, [2026+i, 5560000.00, '按2025实际取整估算'], alt=((8+i)%2==1))
inc_total = sum([4285070.31, 6229093.45, 5564017.88] + [5560000]*15)
write_row(ws4, 23, ['合计', inc_total, ''], font=BOLD_FONT)

ws4.column_dimensions['A'].width = 8; ws4.column_dimensions['B'].width = 20; ws4.column_dimensions['C'].width = 55

# ============================================================
# SHEET 5: 2025审计数据溯源
# ============================================================
ws5 = wb.create_sheet('表5-2025审计溯源')
write_title(ws5, 1, '表5：2025年度运营审计数据逐季拆解')
cell = ws5['A2']; cell.value = '数据来源: 巴中恩阳医院PPP项目2025年运营审计 XLS明细账 (6册), 取4个标准季度损益结转合计值'
cell.font = Font(name='宋体', size=9)

headers5 = ['科目', 'Q1', 'Q2', 'Q3', 'Q4', '4Q合计', '年末调整项*', '全年合计(含调整)']
write_header(ws5, 4, headers5)

quarters = [
    ('主营业务成本', 1016316.83, 667176.24, 1476847.53, 1313687.15,
     1409346.55+579752.49, '物业费+护工服务费'),
    ('管理费用', 558779.11, 287198.91, 213914.32, 179321.77,
     428178.07+169550.88, '工资/社保/差旅/办公等'),
    ('主营业务收入', 1160281.89, 1210580.37, 1407481.80, 1182798.56,
     1529843.40+1251323.21, '物业收入+护工收入'),
    ('其他业务收入', 166207.80, 29264.39, 173959.48, 233443.59,
     44851.74+283341.09, '停车场/充电宝/场地租赁等'),
]
for i, (name, q1, q2, q3, q4, adj, desc) in enumerate(quarters):
    qsum = q1+q2+q3+q4; total = qsum + adj
    write_row(ws5, 5+i, [name, q1, q2, q3, q4, round(qsum,2), round(adj,2), round(total,2)], alt=(i%2==1))

# summary rows
write_row(ws5, 9, ['运维成本(主营+管理)', '', '', '', '', 4474027.75+1239214.11, 1989099.04, 8300069.89], font=BOLD_FONT)
write_row(ws5, 10, ['第三方收入(主营+其他)', '', '', '', '', 4961142.62+602875.26, 2828016.23, 8392034.11], font=BOLD_FONT)
write_row(ws5, 11, ['本报告采用值(4Q)', '', '', '', '', '', '', ''], font=SUB_FONT)

# annotation
ws5.cell(row=13, column=1, value='*年末调整项包含额外两个损益结转期间, 本报告暂未纳入2025年度标准值, 取4个标准季度合计数')
ws5.cell(row=13, column=1).font = Font(name='宋体', size=9, color='CC0000')
ws5.merge_cells('A13:H13')

for i,w in enumerate([18,15,15,15,15,15,18,18]): ws5.column_dimensions[get_column_letter(i+1)].width = w

# ============================================================
# SHEET 6: K因子验证
# ============================================================
ws6 = wb.create_sheet('表6-K因子验证')
write_title(ws6, 1, '表6：K因子(1+7.99%)验证')
cell = ws6['A2']; cell.value = '合同公式: 运维成本×(1+K), K=7.99%. 验证原报告表3数据是否已包含K'
cell.font = Font(name='宋体', size=9)

headers6 = ['验证项', '底层成本(原文)', '×(1+7.99%)', '计算值', '原表3值', '偏差', '结论']
write_header(ws6, 4, headers6)

checks = [
    ('2023运维成本', 5341344.03, 1.0799, 5341344.03*1.0799, 5768117.42, '原表已含K'),
    ('2024-第二经营年度', 5948956.93, 1.0799, 5948956.93*1.0799, 0, '仅验证不单独对表'),
    ('2024-第三经营年度partial', 431456.22, 1.0799, 431456.22*1.0799, 0, ''),
    ('2024合计', 5948956.93+431456.22, 1.0799, (5948956.93+431456.22)*1.0799, 6890208.16, '原表已含K'),
    ('2025运维成本(本所验证)', 5713241.86, 1.0799, 5713241.86*1.0799, 0, '本报告v5已修正'),
    ('2026+运维成本(估算)', 5710000, 1.0799, 5710000*1.0799, 0, '本报告v5已修正'),
]
for i, (item, raw, kk, calc, orig, note) in enumerate(checks):
    diff = calc - orig if orig > 0 else 0
    result = 'OK' if abs(diff) < 1 else f'偏差{diff:,.2f}'
    vals = [item, round(raw,2), kk, round(calc,2), orig if orig > 0 else '-', result, note]
    write_row(ws6, 5+i, vals, alt=(i%2==1))

for i,w in enumerate([28,22,12,22,22,15,30]): ws6.column_dimensions[get_column_letter(i+1)].width = w

# ============================================================
# SHEET 7: 可用性付费汇总
# ============================================================
ws7 = wb.create_sheet('表7-可用性付费汇总')
write_title(ws7, 1, '表7：可用性付费测算汇总表（单位：元）')
headers7 = ['年', '资本金回报', '实际融资本息', '运维成本(含K)', '第三方收入', '可用性付费']
write_header(ws7, 4, headers7)

for y in range(18):
    pr, interest = 0, 0
    # compute from scratch
    cr_temp = CAPITAL_AMORT
    for yy in range(y):
        i_ = cr_temp * RATE; p_ = A - i_
        if yy == 17: p_ = cr_temp; i_ = A - p_
        cr_temp -= p_
        if cr_temp < 0: cr_temp = 0
    i_y = cr_temp * RATE; p_y = A - i_y
    if y == 17: p_y = cr_temp; i_y = A - p_y
    lump = CAPITAL_LUMP if y == 17 else 0
    cap_ret = A + lump

    # bank
    bk_val = round(bo[y][0] * sf, 2)
    # op
    if y == 0: op_val = raw_2023 * (1+K)
    elif y == 1: op_val = raw_2024 * (1+K)
    elif y == 2: op_val = raw_2025 * (1+K)
    else: op_val = raw_future * (1+K)
    # inc
    if y == 0: inc_val = 4285070.31
    elif y == 1: inc_val = 6229093.45
    elif y == 2: inc_val = 5564017.88
    else: inc_val = 5560000.00

    avail = cap_ret + bk_val + op_val - inc_val
    vals7 = [2023+y, round(cap_ret,2), round(bk_val,2), round(op_val,2), round(inc_val,2), round(avail,2)]
    write_row(ws7, 5+y, vals7, alt=(y%2==1))

# totals
cap_total = A * 18 + CAPITAL_LUMP
bank_total = sum(round(b[0]*sf,2) for b in bo)
op_total = (raw_2023+raw_2024+raw_2025+raw_future*15) * (1+K)
inc_total = 4285070.31 + 6229093.45 + 5564017.88 + 5560000*15
grand = cap_total + bank_total + op_total - inc_total

tot7 = ['合计', round(cap_total,2), round(bank_total,2), round(op_total,2), round(inc_total,2), round(grand,2)]
write_row(ws7, 23, tot7, font=BOLD_FONT)

for i,w in enumerate([8,20,20,20,20,22]): ws7.column_dimensions[get_column_letter(i+1)].width = w

# Grand total row
ws7.cell(row=25, column=1, value=f'可用性付费总额: {grand:,.2f}元 (约{grand/1e8:.2f}亿元)')
ws7.cell(row=25, column=1).font = Font(name='微软雅黑', bold=True, size=14, color='0A1F3F')
ws7.merge_cells('A25:F25')

# ============================================================
# SHEET 8: 差异对照 (v4 vs v5 vs 原报告)
# ============================================================
ws8 = wb.create_sheet('表8-版本对照')
write_title(ws8, 1, '表8：版本差异对照 (征求意见稿 vs v4 vs v5)')
headers8 = ['项目', '原征求意见稿', 'v4(含bug)', 'v5(修正后)', 'v4-v5差异来源']
write_header(ws8, 3, headers8)

v4_op = 104021567.44
v5_op = op_total
v4_total = 891208028.23
comp = [
    ('资本金回报', 177955158.81, 175842523.10, round(cap_total,2), '资本金8000万等额+2229.8万一次付vs原8229.8万等额'),
    ('银行融资本息', 712464358.22, 710822119.33, round(bank_total,2), '银行报告总数710,822,119 vs 原表712,464,358'),
    ('运维成本(含K)', 110108501.58, round(v4_op,2), round(v5_op,2), f'v4未对2025+乘K(差{round(v5_op-v4_op,0):,.0f})'),
    ('第三方收入', 91474163.76, 99478181.64, round(inc_total,2), '2025审计实际556万 vs 原表估算506万/年'),
    ('可用性付费合计', 909053854.85, round(v4_total,2), round(grand,2), ''),
]
for i, (item, a, b, c, note) in enumerate(comp):
    write_row(ws8, 4+i, [item, round(a,2), round(b,2), round(c,2), note], alt=(i%2==1))

write_row(ws8, 9, ['原报告-v5差异', '', '', round(909053854.85-grand,2),
                    f'约{round((909053854.85-grand)/1e4,0):,.0f}万'], font=BOLD_FONT)

for i,w in enumerate([18,22,22,22,55]): ws8.column_dimensions[get_column_letter(i+1)].width = w

# Save
outpath = r'C:\Users\scrccpa\Desktop\恩阳医养园PPP测算数据基础表v5.xlsx'
wb.save(outpath)
print(f'Saved: {outpath} ({os.path.getsize(outpath):,} bytes)')
print(f'Grand total: {grand:,.2f} ({grand/1e8:.2f}yi)')
