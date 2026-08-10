import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\scrccpa\Desktop\算法\政府审计算法资产库_v4.xlsx')

# Count detail cards
pc = wb['☆算法详细卡片']
cards = []
for r in range(1, pc.max_row + 1):
    v = pc.cell(row=r, column=1).value
    if v and str(v).startswith('算法卡：'):
        sn = str(v).split('算法卡：')[1].split(' — ')[0]
        cards.append(sn)

# Count summary rows
ws = wb['☆算法资产库总览']
summary_names = []
for r in range(4, ws.max_row + 1):
    v = ws.cell(row=r, column=2).value
    if v:
        summary_names.append(v)

print(f'Detail cards: {len(cards)}')
print(f'Summary rows: {len(summary_names)}')

# Check for mismatches
# Let me also check ENV-CHECKLIST specifically
for r in range(2, pc.max_row + 1):
    v = pc.cell(row=r, column=1).value
    if v and 'ENV-CHECKLIST' in str(v):
        # Count elements in this card
        elements = 0
        for rr in range(r, min(r+50, pc.max_row + 1)):
            elem = pc.cell(row=rr, column=1).value
            if elem and str(elem).startswith('要素名称'):
                continue
            if elem and elem != 'None' and elem.strip():
                elements += 1
            # stop at next card
            if rr > r and pc.cell(row=rr, column=1).value and str(pc.cell(row=rr, column=1).value).startswith('算法卡：'):
                break
        print(f'\nENV-CHECKLIST-001: ~{elements} rows')
        break

# Check if ENV-CHECKLIST is in summary
found_env = False
for r in range(4, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=r, column=c).value
        if v and 'ENV-CHECKLIST' in str(v):
            found_env = True
            print(f'ENV-CHECKLIST in summary: row {r}, col {c}')

if not found_env:
    print('⚠️  ENV-CHECKLIST-001 NOT FOUND in summary sheet!')
    # The summary should have 40 entries but has 39

# Spot check a few cards for completeness
spot_checks = ['PERF-OUTLIER-001', 'SUPV-POCKET-001', 'SOE-MIDMAN-001', 'SOCIAL-INS-001']
for sn in spot_checks:
    count = 0
    counting = False
    for r in range(1, pc.max_row + 1):
        v = pc.cell(row=r, column=1).value
        if v and f'算法卡：{sn}' in str(v):
            counting = True
            continue
        if counting:
            if v and str(v).startswith('算法卡：'):
                break
            if v and v != 'None':
                count += 1
    print(f'  {sn}: ~{count} fields in card')
