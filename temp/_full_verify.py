import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\scrccpa\Desktop\算法\政府审计算法资产库_v5.xlsx')

# Scene map
ws3 = wb['☆业务场景地图']
lines = {}
for row in ws3.iter_rows(min_row=2, values_only=True):
    if row[0] and row[3]:
        lines[row[0]] = row[3].split(chr(10))
covered = set()
print("=== Scene Map Distribution ===")
for ln in sorted(lines.keys()):
    covered.update(lines[ln])
    print(f"  {ln}: {len(lines[ln])}")
print(f"\nTotal lines: {len(lines)} | Covered: {len(covered)}")

# Overview
ws1 = wb['☆算法资产库总览']
all_sns = [r[1] for r in ws1.iter_rows(min_row=2, values_only=True)]
missing = [s for s in all_sns if s not in covered]
print(f"Missing SNs: {missing if missing else 'NONE'}")
print(f"Total SNs: {len(all_sns)} | Duplicates: {[s for s in set(all_sns) if all_sns.count(s)>1] or 'NONE'}")

# Key card placement
print("\n=== Key Card Placement ===")
targets = ['WHISTLE-FLOW-001','EINV-CROSS-001','ASSET-REVIVE-001','BUDGET-006',
           'ASSET-MATCH-001','TRAVEL-SIGNAL-001','LOSS-PENETRATE-001','BID-DARKMARK-001',
           'BID-ROTATE-001','VENDOR-VERIFY-001','PERF-DEVIATION-001','CHK2-001']
for r in ws1.iter_rows(min_row=2, values_only=True):
    if r[1] in targets:
        # Find which scene line it's in
        loc = "?"
        for ln, sns in lines.items():
            if r[1] in sns:
                loc = ln
                break
        print(f"  {r[1]:25s} → {loc}")

# Status spot-check
print("\n=== Status Spot-check ===")
for r in ws1.iter_rows(min_row=2, values_only=True):
    if r[1] in ('TRAVEL-SIGNAL-001','LOSS-PENETRATE-001','BID-DARKMARK-001','BUDGET-006','BUDGET-001'):
        print(f"  {r[1]:25s} | type={r[3]} | risk={r[5]} | cx={r[6]} | status={str(r[10])[:60]}")

# Sheet2 spot-check
print("\n=== Sheet2 (Detailed Cards) Structure ===")
ws2 = wb['☆算法详细卡片']
# Count flagship vs skeleton sections
flagship_count = 0
skeleton_count = 0
for r in range(1, ws2.max_row+1):
    v = str(ws2.cell(row=r, column=1).value or '')
    if '旗舰卡 v4' in v or '40要素' in v:
        flagship_count += 1
    if '骨架卡 v1' in v or '15要素' in v:
        skeleton_count += 1
print(f"  Flagship sections: {flagship_count} | Skeleton sections: {skeleton_count}")

# Check first skeleton card
for r in range(1, ws2.max_row+1):
    v = str(ws2.cell(row=r, column=1).value or '')
    if '骨架卡 v1' in v:
        print(f"  First skeleton card at row {r}: {v[:80]}")
        # Print next 5 rows
        for i in range(1, 6):
            a = ws2.cell(row=r+i, column=1).value or ''
            c = ws2.cell(row=r+i, column=3).value or ''
            print(f"    {a}: {str(c)[:80]}")
        break

print("\n✅ All checks complete")
