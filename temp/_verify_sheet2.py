import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\scrccpa\Desktop\算法\政府审计算法资产库_v5.xlsx')
ws2 = wb['☆算法详细卡片']
print(f"Sheet2 max_row: {ws2.max_row}, max_col: {ws2.max_column}")

# Sample rows from different parts
print("\n=== Row 1-3 ===")
for r in range(1, 4):
    vals = [ws2.cell(row=r, column=c).value for c in range(1, 5)]
    print(f"  Row {r}: {vals}")

# Find section headers
print("\n=== Section headers ===")
for r in range(1, ws2.max_row+1, 1):
    v = str(ws2.cell(row=r, column=1).value or '')
    if len(v) > 10 and ('旗舰' in v or '骨架' in v or '算法' in v or '=== ' in v):
        print(f"  Row {r}: {v[:100]}")

# Check styling - find a flagship and skeleton row
print("\n=== Styling check ===")
for r in range(1, min(ws2.max_row+1, 50)):
    fill = ws2.cell(row=r, column=1).fill
    if fill and fill.start_color and fill.start_color.rgb:
        rgb = fill.start_color.rgb
        if rgb not in ('00000000', 'None'):
            v = str(ws2.cell(row=r, column=1).value or '')[:50]
            print(f"  Row {r}: fill={rgb} val={v}")

# Check Sheet1 styling
print("\n=== Sheet1 Styling ===")
ws1 = wb['☆算法资产库总览']
for r in range(1, min(ws1.max_row+1, 5)):
    fill = ws1.cell(row=r, column=1).fill
    rgb = fill.start_color.rgb if fill and fill.start_color else 'none'
    v = str(ws1.cell(row=r, column=1).value or '')[:30]
    print(f"  Row {r}: fill={rgb} val={v}")
# Also check row 41 (first skeleton)
fill41 = ws1.cell(row=41, column=1).fill
rgb41 = fill41.start_color.rgb if fill41 and fill41.start_color else 'none'
print(f"  Row 41 (first skeleton): fill={rgb41}")

print("\n✅ Sheet2 verification complete")
