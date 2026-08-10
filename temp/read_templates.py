import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

for fname in ['审计算法要素模板（小白版）_v2.xlsx', '政府审计算法资产库架构_v2.xlsx']:
    wb = openpyxl.load_workbook(f'C:\\Users\\scrccpa\\Desktop\\算法\\{fname}')
    print(f'\n===== {fname} =====')
    print(f'Sheets: {wb.sheetnames}')
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f'\n--- Sheet: {sn} (rows={ws.max_row}, cols={ws.max_column}) ---')
        for r, row in enumerate(ws.iter_rows(min_row=1, max_row=min(80, ws.max_row), values_only=False), 1):
            vals = []
            for c in row:
                v = str(c.value)[:100] if c.value is not None else ''
                vals.append(v)
            print(f'  R{r}: | ' + ' | '.join(vals))
    wb.close()
