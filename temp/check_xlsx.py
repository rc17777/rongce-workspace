# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

path = r'C:\Users\scrccpa\Desktop\算法\政府审计算法资产库_v3.xlsx'
wb = openpyxl.load_workbook(path)
print('文件:', path)
print('Sheet列表:', wb.sheetnames)

# 架构总览
ws = wb['☆算法资产库总览']
alg_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0] and str(r[0]).startswith(('SOE','AGR','FIN','ENG','MED','ENV','BUD')) and '【' not in str(r[0])]
print(f'\n[架构总览] 总行数={ws.max_row}, 新增算法行={len(alg_rows)}')
for r in alg_rows:
    print('  ', r[0], '|', r[1], '| 页码', r[8])

# 详细卡片
ws2 = wb['☆算法详细卡片']
titles = [c.value for row in ws2.iter_rows(min_col=1, max_col=1) for c in row if c.value and '算法卡：' in str(c.value)]
print(f'\n[算法详细卡片] 算法卡总数={len(titles)}')
for t in titles[:3]:
    print('  ', t)
print('   ...')
for t in titles[-4:]:
    print('  ', t)

# 风险矩阵
ws3 = wb['☆风险机制与算法矩阵']
cnt = sum(1 for r in ws3.iter_rows(min_row=2, values_only=True) if r[0])
print(f'\n[风险机制矩阵] 数据行={cnt}')
new3 = [r for r in ws3.iter_rows(min_row=2, values_only=True) if r[2] in ('SOE-MIDMAN-001','AGR-INSFAKE-001','FIN-SHELL-001','FIN-INSFAKE-001','ENG-RATIO-001','MED-BIDRIG-001','ENV-RS-001','BUD-CHECKLIST-001')]
print('  新增算法行数:', len(new3))

# 场景地图
ws4 = wb['☆业务场景地图']
cnt4 = sum(1 for r in ws4.iter_rows(min_row=2, values_only=True) if r[0])
print(f'\n[业务场景地图] 数据行={cnt4}')
new4 = [r for r in ws4.iter_rows(min_row=2, values_only=True) if r[3] and ('SOE-MIDMAN' in str(r[3]) or 'AGR-INSFAKE' in str(r[3]) or 'FIN-SHELL' in str(r[3]) or 'FIN-INSFAKE' in str(r[3]) or 'ENG-RATIO' in str(r[3]) or 'MED-BIDRIG' in str(r[3]) or 'ENV-RS' in str(r[3]) or 'BUD-CHECKLIST' in str(r[3]))]
print('  含新增算法行数:', len(new4))

# 文献来源
ws5 = wb['☆文献来源对照']
cnt5 = sum(1 for r in ws5.iter_rows(min_row=2, values_only=True) if r[0] and r[0] != '——')
print(f'\n[文献来源对照] 数据行={cnt5}')
new5 = [r for r in ws5.iter_rows(min_row=2, values_only=True) if r[0] in ('SOE-MIDMAN-001','AGR-INSFAKE-001','FIN-SHELL-001','FIN-INSFAKE-001','ENG-RATIO-001','MED-BIDRIG-001','ENV-RS-001','BUD-CHECKLIST-001')]
print('  新增来源行数:', len(new5))

# 建设路线图
ws6 = wb['☆建设路线图']
r1 = [c.value for c in ws6[2]]
print(f'\n[建设路线图] 第1阶段行: {r1[1][:60]} | {r1[2]} | {r1[5]}')

# 使用声明
ws7 = wb['☆使用声明与责任边界']
r7 = [c.value for c in ws7[8]]
print(f'\n[使用声明] 第7条(节选): {str(r7[2])[:80]}...')

import os
print(f'\n文件大小: {os.path.getsize(path)/1024:.0f} KB')
