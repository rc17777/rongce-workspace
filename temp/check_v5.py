import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\scrccpa\Desktop\算法\政府审计算法资产库_v5.xlsx')

print('=== Sheets ===')
for s in wb.sheetnames:
    ws = wb[s]
    print(f'{s}: {ws.max_row} rows x {ws.max_column} cols')

print('\n=== Summary (first 5 + last 3) ===')
ws = wb[wb.sheetnames[0]]
# Count algorithm rows
alg_count = 0
for r in range(4, ws.max_row + 1):
    sn = ws.cell(row=r, column=2).value
    if sn and sn.strip():
        alg_count += 1
        if alg_count <= 5:
            name = ws.cell(row=r, column=3).value
            atype = ws.cell(row=r, column=4).value
            print(f'  {alg_count}. {sn} {name[:40] if name else ""} [{atype}]')

# Last 3
last_rows = []
for r in range(max(4, ws.max_row - 5), ws.max_row + 1):
    sn = ws.cell(row=r, column=2).value
    if sn and sn.strip():
        name = ws.cell(row=r, column=3).value
        atype = ws.cell(row=r, column=4).value
        last_rows.append(f'  . {sn} {name[:40] if name else ""} [{atype}]')
for lr in last_rows[-3:]:
    print(lr)

print(f'\nTotal algorithms in summary: {alg_count}')

# Count detail cards
pc = wb[wb.sheetnames[1]]
card_count = 0
for r in range(1, pc.max_row + 1):
    v = pc.cell(row=r, column=1).value
    if v and str(v).startswith('算法卡：'):
        card_count += 1
        if card_count <= 3:
            print(f'  Card: {str(v)[:80]}')

print(f'\nTotal detail cards: {card_count}')
