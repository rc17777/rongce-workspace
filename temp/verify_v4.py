# -*- coding: utf-8 -*-
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

path = r'C:\Users\scrccpa\Desktop\算法\政府审计算法资产库_v4.xlsx'
print('文件大小:', os.path.getsize(path), 'bytes')
wb = openpyxl.load_workbook(path)
print('Sheet列表:', wb.sheetnames)

# 验证算法卡片完整性
ws = wb['☆算法详细卡片']
# 找出所有"算法卡："标题行
titles = []
for row in ws.iter_rows(min_col=1, max_col=1):
    v = row[0].value
    if v and isinstance(v, str) and v.startswith('算法卡：'):
        titles.append(v)
print(f'\n算法卡片总数: {len(titles)}')
for t in titles:
    print(' ', t)

# 验证每张卡片40要素齐全（统计"必须"要素行）
required = ['算法编号','算法名称','版本/编制人/复核人','适用业务场景','审计目标','风险假设','适用范围','不适用范围',
            '法规及业务依据','输入数据表','核心字段','主键与关联键','数据质量检查','计算公式/步骤','参数与阈值',
            '阈值依据','输出字段','疑点解释模板','追加证据','人工核查程序','结论边界','测试案例','历史回测结果',
            '数据就绪度评估','数据质量分级','验证标准（目标象限）','底稿嵌入模板编号','可解释性输出模板',
            '算法退役条件','定期复查周期']
# 检查每个v4新算法在卡片sheet中是否有完整字段（通过检查字典重建：直接重跑算法定义验证）
import ast
src = open(r'C:\Users\scrccpa\.openclaw\workspace\temp\build_algorithm_lib_v4.py', encoding='utf-8').read()
# 提取 algorithms_v4 字面量
tree = ast.parse(src)
algs4 = None
for node in ast.walk(tree):
    if isinstance(node, ast.Assign):
        for t in node.targets:
            if isinstance(t, ast.Name) and t.id == 'algorithms_v4':
                algs4 = ast.literal_eval(node.value)
print(f'\nv4新算法字典数: {len(algs4)}')
field_map_keys = ['sn','name','scene','objective','risk_hypothesis','scope_yes','scope_no','law_basis',
 'data_tables','fields','keys','data_quality','calc_logic','threshold','threshold_basis','output_fields',
 'explain','evidence','check_procedure','conclusion_boundary','test_cases','backtest','risk_score',
 'multi_rule','sensitivity','fpr_fnr','visual','reuse','perf','privacy','data_readiness','data_grade',
 'verify_standard','workpaper_template','explainability','retire_condition','review_cycle','dependency',
 'expected_value','history_output']
print('要素总数要求: 40')
for a in algs4:
    missing = [k for k in field_map_keys if k not in a or not a[k]]
    empty = [k for k in field_map_keys if k in a and (a[k] is None or str(a[k]).strip()=='')]
    print(f'  {a["sn"]} {a["name"]}: 缺字段={missing} 空字段={empty} 字段数={len([k for k in field_map_keys if a.get(k)])}/40')

# 总览sheet行数
ws2 = wb['☆算法资产库总览']
print(f'\n总览sheet行数: {ws2.max_row}')
ws3 = wb['☆风险机制与算法矩阵']
print(f'风险矩阵行数: {ws3.max_row}')
ws4 = wb['☆业务场景地图']
print(f'场景图行数: {ws4.max_row}')
ws5 = wb['☆文献来源对照']
print(f'文献行数: {ws5.max_row}')

# 抽查一个v3算法卡和一个v4算法卡内容是否完整写入
# 找到SOCIAL-INS-001卡片位置
for row in ws.iter_rows(min_col=1, max_col=1):
    v = row[0].value
    if v and 'SOCIAL-INS-001' in str(v):
        print(f'\n抽查: {v} 位于第{row[0].row}行')
        break
