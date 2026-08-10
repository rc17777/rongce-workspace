import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\scrccpa\Desktop\算法\政府审计算法资产库_v5.xlsx')
ws1 = wb['☆算法资产库总览']
# Check fills for rows 2,3,4 (flagship) and 41,42,43 (skeleton)
for r in [1,2,3,40,41,42,135,136]:
    fill = ws1.cell(row=r, column=1).fill
    rgb = fill.start_color.rgb if fill and fill.start_color else 'none'
    v = str(ws1.cell(row=r, column=1).value or '')[:20]
    t = str(ws1.cell(row=r, column=4).value or '')[:10]  # type column
    print(f"Row {r:3d}: fill={rgb} type={t} val={v}")
