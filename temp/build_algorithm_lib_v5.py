# -*- coding: utf-8 -*-
r"""
政府审计算法资产库 v5.0 生成脚本
================================
- 40 张旗舰卡（完整40要素）：直接复用 v4 全部算法（31 v3 + 9 v4）
- 95 张骨架卡（15要素）：4 批扫描提取（batch1=26 + batch2=33 + batch3=15 + batch4=21）
- 合并去重：骨架卡与旗舰卡高度重叠的标记为"补充细化"（保留独立SN），其余按新算法入库
- 输出：C:\Users\scrccpa\Desktop\算法\政府审计算法资产库_v5.xlsx
  * Sheet1 ☆算法资产库总览（135算法，旗舰深蓝行/骨架浅灰行）
  * Sheet2 ☆算法详细卡片（旗舰40要素 + 骨架15要素紧凑格式）
  * Sheet3 ☆业务场景地图（12条业务线+延伸场景）
  * Sheet4 ☆风险机制矩阵（覆盖全部135个算法）
  * Sheet5 ☆建设路线图（v5.0里程碑：135算法完成全量提取）
  * Sheet6 ☆使用声明（更新来源说明）
  * Sheet7 ☆文献来源（合并全部来源）
"""
import json
import sys
sys.stdout.reconfigure(encoding='utf-8')

import re

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

V4_PATH = r'C:\Users\scrccpa\.openclaw\workspace\temp\build_algorithm_lib_v4.py'
OUTPUT_PATH = r'C:\Users\scrccpa\Desktop\算法\政府审计算法资产库_v5.xlsx'
BATCH_PATHS = {
    1: r'C:\Users\scrccpa\.openclaw\workspace\temp\batch_algorithms_batch1.json',
    2: r'C:\Users\scrccpa\.openclaw\workspace\temp\batch_algorithms_batch2.json',
    3: r'C:\Users\scrccpa\.openclaw\workspace\temp\batch_algorithms_batch3.json',
    4: r'C:\Users\scrccpa\.openclaw\workspace\temp\batch_algorithms_batch4.json',
}
VERSION = 'v5.0'

# ========== 1. 提取 v4 全局数据（40旗舰卡 + 各Sheet基础数据） ==========
print('① 提取 v4 旗舰卡数据...')
src = open(V4_PATH, encoding='utf-8').read()
src = src.replace('wb.save(OUTPUT_PATH)', 'pass  # v5 save stripped')
ns = {}
exec(compile(src, 'v4', 'exec'), ns)
algorithms_v4 = ns['algorithms']            # 40 旗舰卡（40要素）
arch_data_v4 = ns['arch_data']              # 总览基础
risk_data_v4 = ns['risk_data']              # 风险矩阵基础
scene_data_v4 = ns['scene_data']            # 场景图基础
road_data_v4 = ns['road_data']              # 路线图基础
decl_data_v4 = ns['decl_data']              # 声明基础
paper_data_v4 = ns['paper_data']            # 文献基础
print(f'   ✅ 旗舰卡 {len(algorithms_v4)} 张；arch={len(arch_data_v4)} risk={len(risk_data_v4)} '
      f'scene={len(scene_data_v4)} road={len(road_data_v4)} decl={len(decl_data_v4)} paper={len(paper_data_v4)}')

# ========== 2. 加载 4 批骨架卡并统一为15要素结构 ==========
def _as_list(v):
    if v is None:
        return []
    if isinstance(v, list):
        return [str(x) for x in v]
    return [str(v)]

def _fmt(v, joiner='\n'):
    items = _as_list(v)
    items = [x.strip() for x in items if str(x).strip() and str(x).strip() != 'None']
    return joiner.join(items)

def load_batch(batch_no):
    d = json.load(open(BATCH_PATHS[batch_no], encoding='utf-8'))
    if isinstance(d, dict):
        return d['algorithms']
    return d

print('② 加载4批骨架卡...')
skeleton_cards = []
for bno in [1, 2, 3, 4]:
    raw = load_batch(bno)
    for a in raw:
        if bno == 1:
            # batch1 字段名：scenario/risk_assumption/core_signals/formula/audit_steps/regulation/agent/duplicate
            card = {
                'sn': a.get('sn'), 'name': a.get('name'),
                'scene': a.get('scenario', ''),
                'objective': '识别并核实：' + str(a.get('risk_assumption', '')),
                'risk_hypothesis': a.get('risk_assumption', ''),
                'signals': a.get('core_signals', []),
                'calc_logic': a.get('formula', ''),
                'threshold': a.get('threshold', ''),
                'evidence': a.get('evidence', ''),
                'check_procedure': a.get('audit_steps', []),
                'law_basis': a.get('regulation', ''),
                'source_file': a.get('source_file', ''),
                'biz_line': a.get('biz_line', ''),
                'complexity': a.get('complexity', ''),
                'agent': a.get('agent', ''),
                'dup_note': a.get('duplicate', '') or '',
            }
        else:
            # batch2-4 字段名：scene/risk_hypothesis/signals/calc_logic/check_procedure/law_basis/agent_map
            card = {
                'sn': a.get('sn'), 'name': a.get('name'),
                'scene': a.get('scene', ''),
                'objective': '在「' + str(a.get('scene', '')) + '」场景中，识别并核实：' + str(a.get('risk_hypothesis', '')),
                'risk_hypothesis': a.get('risk_hypothesis', ''),
                'signals': a.get('signals', []),
                'calc_logic': a.get('calc_logic', ''),
                'threshold': a.get('threshold', ''),
                'evidence': a.get('evidence', ''),
                'check_procedure': a.get('check_procedure', ''),
                'law_basis': a.get('law_basis', ''),
                'source_file': a.get('source_file', ''),
                'biz_line': a.get('biz_line', ''),
                'complexity': a.get('complexity', ''),
                'agent': a.get('agent_map', ''),
                'dup_note': '',
            }
        card['batch'] = bno
        skeleton_cards.append(card)
print(f'   ✅ 骨架卡 {len(skeleton_cards)} 张 (batch1=26, batch2=33, batch3=15, batch4=21)')

# ========== 3. 去重处理 ==========
# 全部135张保留独立SN；与旗舰卡高度重叠的标记为"补充细化"，几乎不重叠的按新算法入库
KNOWN_PAIRS = {
    'TRAVEL-SIGNAL-001': '与SUPV-TRAVEL-001同源（差旅费四信号）',
    'LOSS-PENETRATE-001': '与CHK-LOSS-001同源（亏损六步穿透）',
    'HR-EATEMPTY-001': '与HR-RF-002/FUND-SIPHON-001同源（吃空饷五对照）',
    'VENDOR-VERIFY-001': '与PROC-FAKE-001同源（供应商虚假材料三查）',
    'HOSP-PARAM-001': '与MED-BIDRIG-001同源（医院围标串标）',
    'NATRES-AUDIT-001': '与ENV-CHECKLIST-001同源（自然资源五维清单）',
    'BID-DARKMARK-001': '与招投标猎手检测阶段重叠（暗标隐形记号）',
    'BID-ROTATE-001': '与BID-PATTERN-005重叠（互惠轮庄陪标）',
}

def status_for(card):
    sn = card['sn']
    if sn in KNOWN_PAIRS:
        return '补充细化入库：' + KNOWN_PAIRS[sn]
    note = (card.get('dup_note') or '').strip()
    if note:
        note = note.replace('否，', '').strip()
        if '补充细化' in note:
            return '补充细化入库：' + note
        if '互补' in note:
            return '新算法入库（' + note + '）'
        return '新算法入库（' + note + '）'
    return '新算法入库（15要素骨架卡）'

# SN 唯一性校验
all_sns = [a['sn'] for a in algorithms_v4] + [c['sn'] for c in skeleton_cards]
dup_sns = [s for s in set(all_sns) if all_sns.count(s) > 1]
assert not dup_sns, f'SN重复：{dup_sns}'
print(f'③ 去重校验：{len(algorithms_v4)}+{len(skeleton_cards)}={len(all_sns)} 个SN全部唯一，无重复卡被跳过')

# ========== 4. 派生字段：风险机制 / 复杂度 / Agent / 业务线 ==========
# 4.1 旗舰卡风险机制：从 v4 arch_data 查询（SN → 风险机制）
arch_lookup = {}
for row in arch_data_v4:
    if row and str(row[0]).startswith(('SUPV', 'PERF', 'PROC', 'BID', 'FUND', 'HR-', 'ENG', 'CHK', 'ASSET', 'RULE', 'FIN', 'ENV', 'SOE', 'AGR', 'MED', 'BUD', 'SOCIAL', 'BIGDATA', 'TRANSFER', 'BOND', 'REV')):
        arch_lookup[row[0]] = row
def flagship_risk(sn):
    row = arch_lookup.get(sn)
    return row[3] if row and len(row) > 3 else '真实性'
def flagship_biz(sn):
    row = arch_lookup.get(sn)
    return row[2] if row and len(row) > 2 else '——'
def flagship_cx(sn):
    row = arch_lookup.get(sn)
    return str(row[5]) if row and len(row) > 5 else 'L3'

# 4.2 旗舰卡 Agent 映射（人工映射，与融策22-Agent分工一致）
FLAGSHIP_AGENT = {
    'PERF-OUTLIER-001': '数据侦察兵（异常检测）', 'REV-PREDICT-001': '数据侦察兵（预测建模）',
    'PROC-CONCEN-001': '招投标猎手（集中度分析）', 'BID-PATTERN-005': '招投标猎手（模式检测）',
    'FUND-FRAUD-001': '数据侦察兵（资金分析）', 'HR-RF-001': '数据侦察兵（人员比对）',
    'HR-RF-002': '数据侦察兵（名单比对）', 'ENG-SAMPLE-001': '结算审计师（抽样核算）',
    'ENG-SCORE-001': '结算审计师（评分测算）', 'CHK-RECON-001': '数据侦察兵（勾稽核对）',
    'ASSET-MATCH-001': '数据侦察兵（资产盘点）', 'RULE-MATCH-001': '数据侦察兵（规则引擎）',
    'SUPV-ANOMALY-001': '数据侦察兵（统计检测）', 'SUPV-POCKET-001': '数据侦察兵（资金轨迹）',
    'SUPV-TRAVEL-001': '数据侦察兵（费用信号）', 'FUND-SIPHON-001': '数据侦察兵（资金穿透）',
    'CHK-LOSS-001': '数据侦察兵（损益穿透）', 'PROC-FAKE-001': '合同猎犬（材料真实性）',
    'PROC-RELATED-001': '招投标猎手（关联分析）', 'CHK-RD-001': '数据侦察兵（研发费用核）',
    'ENG-FINAL-001': '结算审计师（结算审核）', 'SUPV-WARNING-001': '数据侦察兵（预警建模）',
    'ENV-CHECKLIST-001': '法规检察官（合规清单）', 'SOE-MIDMAN-001': '数据侦察兵（中间商穿透）',
    'AGR-INSFAKE-001': '数据侦察兵（保险比对）', 'FIN-SHELL-001': '数据侦察兵（壳特征识别）',
    'FIN-INSFAKE-001': '数据侦察兵（理赔交叉）', 'ENG-RATIO-001': '结算审计师（比例测算）',
    'MED-BIDRIG-001': '招投标猎手（围标识别）', 'ENV-RS-001': '数据侦察兵（遥感比对）',
    'BUD-CHECKLIST-001': '预算工程师（清单核查）', 'SOCIAL-INS-001': '数据侦察兵（多源比对）',
    'SOCIAL-MAT-001': '数据侦察兵（多维模型）', 'SOCIAL-WORK-001': '数据侦察兵（信号检测）',
    'SOCIAL-WELFARE-001': '数据侦察兵（双向比对）', 'BIGDATA-SERVICE-001': '数据侦察兵（时空验证）',
    'BIGDATA-SQL-001': '数据侦察兵（SQL范式）', 'PERF-DEVIATION-001': '绩效评价师（偏离度检测）',
    'TRANSFER-TRACE-001': '预算工程师（全链路追踪）', 'BOND-PENETRATE-001': '预算工程师（穿透监测）',
}

# 4.3 骨架卡 Agent 英译中
AGENT_TRANSLATE = {
    'data-scout': '数据侦察兵', 'bid-hunter': '招投标猎手', 'contract-hound': '合同猎犬',
    'settlement-auditor': '结算审计师', 'perf-evaluator': '绩效评价师', 'fiscal-reviewer': '财政评审员',
    'budget-engineer': '预算工程师', 'adjustment-entries': '调整分录师', 'law-prosecutor': '法规检察官',
    'report-writer': '报告笔杆子', 'review-sentinel': '复核哨兵', 'meeting-minutes': '会议纪要分析',
    'ocr-preprocessor': 'OCR预处理', 'data-classifier': '数据分类员', 'data-masker': '数据脱敏',
    'plan-writer': '方案撰写师', 'bid-deviation': '评标偏离度',
}
def translate_agent(s):
    if not s:
        return '数据侦察兵'
    for en, zh in AGENT_TRANSLATE.items():
        if en in s:
            s = s.replace(en, zh)
    return s

# 4.4 骨架卡风险机制分类器
def classify_mechanism(card):
    text = ' '.join([
        str(card.get('name', '')), str(card.get('scene', '')),
        _fmt(card.get('signals', [])), str(card.get('risk_hypothesis', '')),
    ])
    rules = [
        ('真实性', ['伪造', '虚假', '造假', '骗保', '冒领', '虚报', '套取', '套码', '虚增', '虚列', '回流',
                   '闭环', '循环', '画皮', '虚构', '无实物', '白条', '隐匿', '截留', '挪用', '挤占', '假',
                   '过户', '壳', '出险', '兑付', '逃逸', '空饷', '套用', '骗']),
        ('完整性', ['漏收', '少收', '漏征', '应享未享', '应免未免', '未上缴', '未缴', '欠缴', '少计', '流失',
                   '侵占', '足额', '违约金', '完整性']),
        ('合规性', ['违规', '超标准', '超限', '超许可', '规避', '未执行', '违反', '不合规', '围标', '串标',
                   '陪标', '受控', '越权', '无期限', '审批', '超编', '超概算', '红线', '标准', '暗标',
                   '履约', '分拆', '退二进三', '合规', '受控拍卖']),
        ('穿透性', ['穿透', '全链路', '资金轨迹', '资金链', '资金往来', '追征']),
        ('异常性', ['异常', '偏离', '波动', '集中', '离群', '突变', '偏差', '高频', '超常', '重复', '预警',
                   '信号', '突击', '冲突', '顶格', '溢价']),
        ('效率性', ['闲置', '低效', '效益', '盘活', '绩效', '成本', '亏损', '测算', '评估']),
        ('勾稽性', ['比对', '一致性', '勾稽', '交叉', '校验', '核对', '复核', '稽核', '核销', '真实性']),
    ]
    overrides = {
        'WHISTLE-FLOW-001': '有效性', 'CONCESS-FEE-001': '效率性',
        'NATRES-AUDIT-001': '合规性', 'HR-EATEMPTY-001': '真实性',
    }
    if card['sn'] in overrides:
        return overrides[card['sn']]
    for mech, kws in rules:
        if any(k in text for k in kws):
            return mech
    return '真实性'

FAMILY_MAP = {
    '真实性': '交叉验证', '完整性': '勾稽核对', '合规性': '规则筛查', '穿透性': '穿透分析',
    '异常性': '偏离度检测', '效率性': '绩效评价', '勾稽性': '勾稽核对', '有效性': '机制评估',
}

def norm_cx(v):
    s = str(v or '').strip()
    mm = re.search(r'L[234]', s)
    if not mm:
        return s or 'L2'
    extra = s[mm.end():].strip('（）() 算法规则')
    if '机器学习' in s:
        extra = '机器学习'
    return mm.group(0) + (f'（{extra}）' if extra else '')

# ========== 5. 构建总览数据 ==========
overview = []   # 每行: [序号, sn, name, 类型, 场景, 风险机制, 复杂度, 业务线, Agent, 来源批次, 状态]
for i, a in enumerate(algorithms_v4, 1):
    overview.append([
        i, a['sn'], a['name'], '旗舰', a.get('scene', ''),
        flagship_risk(a['sn']), flagship_cx(a['sn']), flagship_biz(a['sn']),
        FLAGSHIP_AGENT.get(a['sn'], '数据侦察兵'), '旗舰卡·v4.0（31 v3 + 9 v4）',
        '正式入库（40要素完整卡）',
    ])
for c in skeleton_cards:
    overview.append([
        len(overview) + 1, c['sn'], c['name'], '骨架', c.get('scene', ''),
        classify_mechanism(c), norm_cx(c.get('complexity', '')), c.get('biz_line', ''),
        translate_agent(c.get('agent', '')), f"骨架卡·batch{c['batch']}",
        status_for(c),
    ])
assert len(overview) == 135, f'总览行数异常：{len(overview)}'
print(f'④ 总览数据构建完成：{len(overview)} 行（旗舰40 + 骨架95）')

# ========== 6. 业务场景地图：12条业务线+延伸场景 ==========
# 关键词按“先具体后通用”排序；全部未命中时按业务线前缀回退；仍无则归入财政审计/预算执行
LINES = [
    ('政府债务/专项债券审计', '债券资金使用真实合规、收益自平衡、化债风险穿透', ['专项债', '债券', '债务', '国债', '化债']),
    ('转移支付/专项资金审计', '资金拨付及时性、使用合规性、绩效真实性、重复申报', ['转移支付', '就业补助', '代征非税', '专项资金']),
    ('国有资产审计/财政资源统筹（延伸）', '资产出租收入、闲置资产盘活、实物资产匹配', ['国有资产', '资产出租', '资产盘活', '资产清查', '资产状态', '账实', '闲置', '校舍', '沉没', '实物资产']),
    ('民生审计（社保/医保/养老/惠民补贴）', '社保基金安全、骗保冒领、补贴应享尽享', ['社保', '医保', '养老保险', '工伤保险', '生育津贴', '惠民', '补贴', '津贴', '殡葬', '低保', '一卡通']),
    ('医疗机构审计（延伸）', '医用耗材结算、DRG套码、器械供应商壳特征、医院采购围标', ['耗材', 'DRG', '病组', '器械', '医院', '医用']),
    ('政府购买服务/养老服务（延伸）', '服务工单真实性、工单配置超限、人员资质', ['养老', '上门服务', '护理', '服务工单', '购买服务', '居家']),
    ('农业农村/乡村振兴审计', '涉农资金真实使用、农机补贴、农业保险、乡村振兴', ['农业', '农村', '农机', '以工代赈', '涉农', '养殖', '乡村振兴']),
    ('粮食安全审计（延伸）', '储备粮轮换运输真实性、损耗率超限', ['储备粮', '粮食', '损耗']),
    ('招投标/政府采购审计', '围标串标识别、虚假材料核验、价格偏离、受控拍卖、暗标记号', ['招投标', '投标', '招标', '采购', '围标', '串标', '评标', '供应商', '卖场', '拍卖', '产权交易', '陪标', '暗标', '竞价']),
    ('投资审计/政府投资项目', '工程量真实性、征拆补偿、资金归集专款专用、规避审批、造价', ['工程', '造价', '土方', '征拆', '拆迁', '签证', '水利', '概算', '投资', '信息化', '软件']),
    ('经济责任审计', '履职尽责、公车管理、吃空饷、收入完整性、离任责任', ['经责', '经济责任', '公车', '公务用车', '吃空饷', '离任', '领导干部', '履职', '未履约', '台账']),
    ('三公经费/八项规定检查（延伸）', '差旅费报销信号、公务车辆费用、通信费用异常', ['三公', '差旅', '公务', '会议', '接待', '车辆', '通信', '八项', '差旅费']),
    ('金融审计', '信贷真实性、资管估值、保险理赔、资金回流、壳特征', ['金融', '银行', '信贷', '资管', '农商行', '快贷', '政策性保险', '理赔']),
    ('企业审计/国有企业审计', '国资流失、中间人绕道、亏损穿透、虚假研发、基金循环', ['国企', '集团', '企业', '基金', '合资', '研发', '工资', '亏损', '中间商', '工商', '员工', '高息', '融资']),
    ('公用事业/收费征管（延伸）', '水费燃气费征收、阶梯水价、污水处理费、违约金', ['水费', '燃气', '污水', '水价', '阶梯', '公用事业', '用水', '违约金', '收费']),
    ('资源环境审计', '矿产资源、土地、林地、生态红线、自然资源责任、公益林', ['资源', '环境', '矿产', '采矿', '矿业', '土地', '林地', '公益林', '自然资源', '遥感', '用地', 'NDVI', '退二进三', '生态']),
    ('财会监督（延伸）', '举报受理闭环、会计监督、发票报销内控', ['财会监督', '举报受理', '会计监督']),
    ('税务审计（延伸）', '虚开发票、网络货运、逃逸式注销、税源追征', ['税务', '税', '虚开', '货运', '逃逸', '注销', '税源']),
    ('大数据审计（延伸）', '多源数据交叉、SQL范式、电子发票稽核、特征码检测', ['大数据', '数据', 'SQL', '范式', '电子发票', '稽核', '特征码', '电子标书']),
    ('绩效评价/预算绩效管理', '绩效指标真实性、成本预算绩效、盘活效益、费用测算', ['绩效评价', '绩效管理', '绩效指标', '成本预算', '特许经营', '四维测算', '六步法', '偏离度', '评估']),
    ('财政审计/预算执行', '预算编制真实性、执行进度、存量资金盘活、支出标准符合性', ['预算', '财政', '存量资金', '公用经费', '经费', '收支', '往来', '决算', '支出标准', '非税收入']),
]

BIZ_FALLBACK = [
    ('经责', '经济责任审计'), ('民生', '民生审计（社保/医保/养老/惠民补贴）'),
    ('农业', '农业农村/乡村振兴审计'), ('农村', '农业农村/乡村振兴审计'),
    ('投资', '投资审计/政府投资项目'), ('工程', '投资审计/政府投资项目'), ('造价', '投资审计/政府投资项目'),
    ('招投标', '招投标/政府采购审计'), ('采购', '招投标/政府采购审计'),
    ('金融', '金融审计'), ('银行', '金融审计'),
    ('企业', '企业审计/国有企业审计'), ('国企', '企业审计/国有企业审计'), ('国资', '企业审计/国有企业审计'),
    ('资源', '资源环境审计'), ('环境', '资源环境审计'),
    ('绩效', '绩效评价/预算绩效管理'), ('评价', '绩效评价/预算绩效管理'),
    ('税务', '税务审计（延伸）'), ('税', '税务审计（延伸）'),
    ('粮食', '粮食安全审计（延伸）'), ('储备', '粮食安全审计（延伸）'),
    ('大数据', '大数据审计（延伸）'), ('数据', '大数据审计（延伸）'),
    ('债务', '政府债务/专项债券审计'), ('债', '政府债务/专项债券审计'),
    ('转移支付', '转移支付/专项资金审计'), ('专项', '转移支付/专项资金审计'),
    ('三公', '三公经费/八项规定检查（延伸）'), ('公务', '三公经费/八项规定检查（延伸）'),
    ('公用事业', '公用事业/收费征管（延伸）'),
    ('养老', '政府购买服务/养老服务（延伸）'), ('服务', '政府购买服务/养老服务（延伸）'),
    ('医疗', '医疗机构审计（延伸）'), ('医保', '医疗机构审计（延伸）'), ('医院', '医疗机构审计（延伸）'),
    ('财会', '财会监督（延伸）'), ('监督', '财会监督（延伸）'),
    ('资产', '国有资产审计/财政资源统筹（延伸）'),
    ('补贴', '民生审计（社保/医保/养老/惠民补贴）'), ('保险', '民生审计（社保/医保/养老/惠民补贴）'),
]

DEFAULT_LINE = '财政审计/预算执行'


def primary_line(card_or_algo):
    text = ' '.join([str(card_or_algo.get('name', '')), str(card_or_algo.get('scene', '')),
                     str(card_or_algo.get('biz_line', ''))])
    for line_name, goal, kws in LINES:
        if any(k in text for k in kws):
            return line_name
    biz = str(card_or_algo.get('biz_line', ''))
    for k, line_name in BIZ_FALLBACK:
        if k in biz:
            return line_name
    return DEFAULT_LINE

scene_extra = []
for line_name, goal, kws in LINES:
    members = [o for o in overview if primary_line({'name': o[2], 'scene': o[4], 'biz_line': o[7]}) == line_name]
    if not members:
        continue
    mechs = list(dict.fromkeys(o[5] for o in members))
    sns = '\n'.join(o[1] for o in members)
    names = '\n'.join(o[2] for o in members)
    scene_extra.append([
        line_name, goal, '/'.join(mechs), sns, names,
        '见各算法卡「输入数据表」', '见各算法卡「数据就绪度」',
    ])
scene_data_v5 = list(scene_data_v4) + scene_extra
print(f'⑤ 场景地图：v4基础{len(scene_data_v4)}行 + 新增{len(scene_extra)}条业务线 = {len(scene_data_v5)}行')

# ========== 7. 风险机制矩阵（覆盖全部135） ==========
risk_extra = []
for o in overview:
    if o[3] == '旗舰':
        continue  # 旗舰卡已在 risk_data_v4 中
    mech = o[5]
    risk_extra.append([
        mech, FAMILY_MAP.get(mech, '规则筛查'), o[1], o[2],
        '骨架卡（15要素），升级40要素后跨场景复用',
        '规则/阈值未随政策更新=中风险',
        '每半年复核（随旗舰卡族）',
    ])
risk_data_v5 = list(risk_data_v4) + risk_extra
assert len(risk_data_v5) >= 135, f'风险矩阵未覆盖135算法：{len(risk_data_v5)}'
print(f'⑥ 风险矩阵：v4基础{len(risk_data_v4)}行 + 骨架{len(risk_extra)}行 = {len(risk_data_v5)}行（覆盖全部135算法）')

# ========== 8. 路线图 / 声明 / 文献 ==========
road_data_v5 = list(road_data_v4)
road_data_v5[0] = [
    '第1阶段：选场景',
    '文献/案例/杂志全量扫描：4批共95张骨架卡 + 40张旗舰卡，135个算法覆盖20条以上业务线（v5.0里程碑：全量提取完成）',
    '135个算法卡（40张40要素旗舰卡 + 95张15要素骨架卡）',
    '每个场景2-3个算法；135个算法全部有唯一SN与场景映射；骨架卡与旗舰卡重叠项已标记"补充细化"',
    '旗舰L3≥85%；骨架L2-L3（待升级）',
    '✅ 已完成（v5.0交付：135算法完成全量提取）',
]
road_data_v5.append([
    '第8阶段：骨架卡升级（v5.1规划）',
    '将95张骨架卡按40要素模板补齐（审计目标/输入数据表/核心字段/回测/验证标准/底稿模板等25个缺失要素）',
    '95张骨架卡全部升级为40要素完整卡',
    '全部135算法40要素齐全，可直接嵌入审计工作流',
    'L3≥90%',
    '⏳ 待开发',
])

decl_data_v5 = list(decl_data_v4)
decl_data_v5[6] = [
    '7',
    '来源说明',
    '本算法库共135个算法：40张旗舰卡（13个从公开学术论文案例中提炼、18个来自融策审计方法论库内部沉淀、9个来自第四批知识源[审计案例2026年第1-5册+财政监督杂志]，40要素齐全）+ 95张骨架卡（4批扫描提取：batch1-26张/预算与专项债，batch2-33张/农业民生金融等，batch3-15张/采购工程税务，batch4-21张/发票稽核医疗招投标等，15要素骨架格式，来源见文献来源Sheet）。骨架卡与旗舰卡高度重叠的8组已标记"补充细化"（同源保留独立SN），其余按新算法入库。算法逻辑经文献/项目实证验证，部分有实测回测数据。具体应用需结合审计项目实际情况调整参数。',
    '——',
]
decl_data_v5.append([
    '8',
    '骨架卡使用定位',
    '骨架卡为"快速检索+场景覆盖"级资产：使用前需由审计人员结合对应旗舰卡交叉验证，或按40要素模板补齐后方可嵌入正式审计工作流；标记"补充细化"的骨架卡与旗舰卡同源，优先复用旗舰卡40要素，骨架卡信号作为细化补充。',
    '融策算法资产管理办法',
])

paper_data_v5 = list(paper_data_v4)
for c in skeleton_cards:
    paper_data_v5.append([
        c['sn'], c['name'], '骨架卡：来源文献批量扫描提取（批次扫描）',
        '融策算法工作组', f"骨架卡batch{c['batch']}", str(c.get('source_file', '')),
    ])
print(f'⑦ 文献来源：{len(paper_data_v5)}条（v4的{len(paper_data_v4)}条 + 骨架{len(skeleton_cards)}条）')

# ========== 9. 生成 Excel ==========
print('⑧ 开始生成政府审计算法资产库 v5.0...')
wb = openpyxl.Workbook()

# 样式定义
DEEP_BLUE = '0A1F3F'
TEAL = '1A5C6E'
COPPER = 'C5955C'
WARM_GREY = 'F5F2EC'
SKELETON_GREY = 'D9D9D9'     # 骨架卡浅灰
SKELETON_ROW_GREY = 'F2F2F2' # 总览骨架行浅灰

header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color=DEEP_BLUE, end_color=DEEP_BLUE, fill_type='solid')
normal_font = Font(name='微软雅黑', size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
wrap_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_header_row(ws, row, max_col, fill=None):
    if fill is None:
        fill = header_fill
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = center_align

def style_data_cell(ws, row, col, bold=False):
    cell = ws.cell(row=row, column=col)
    cell.font = Font(name='微软雅黑', size=10, bold=bold)
    cell.border = thin_border
    cell.alignment = wrap_align
    return cell

# ========== Sheet1: ☆算法资产库总览 ==========
ws_arch = wb.active
ws_arch.title = '☆算法资产库总览'
arch_headers = ['序号', '算法编号', '算法名称', '类型(旗舰/骨架)', '适用场景', '风险机制',
                '复杂度(L2/L3)', '业务线', 'Agent映射', '来源批次', '状态']
for c, h in enumerate(arch_headers, 1):
    ws_arch.cell(row=1, column=c, value=h)
style_header_row(ws_arch, 1, len(arch_headers))

for r, row in enumerate(overview, 2):
    is_flag = (row[3] == '旗舰')
    fill = PatternFill(start_color=DEEP_BLUE, end_color=DEEP_BLUE, fill_type='solid') if is_flag \
        else PatternFill(start_color=SKELETON_ROW_GREY, end_color=SKELETON_ROW_GREY, fill_type='solid')
    for c, val in enumerate(row, 1):
        cell = style_data_cell(ws_arch, r, c)
        cell.value = val
        cell.fill = fill
        if is_flag:
            cell.font = Font(name='微软雅黑', size=10, bold=False, color='FFFFFF')
    # 类型列着色强调
    type_cell = ws_arch.cell(row=r, column=4)
    type_cell.font = Font(name='微软雅黑', size=10, bold=True,
                          color='FFFFFF' if is_flag else DEEP_BLUE)
    type_cell.alignment = center_align
    type_cell.fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type='solid') if is_flag \
        else PatternFill(start_color=COPPER, end_color=COPPER, fill_type='solid')

widths = [6, 20, 36, 12, 42, 10, 13, 24, 22, 16, 34]
for i, w in enumerate(widths, 1):
    ws_arch.column_dimensions[get_column_letter(i)].width = w
ws_arch.freeze_panes = 'A2'

# ========== Sheet2: ☆算法详细卡片 ==========
ws_cards = wb.create_sheet('☆算法详细卡片')
ws_cards.column_dimensions['A'].width = 22
ws_cards.column_dimensions['B'].width = 10
ws_cards.column_dimensions['C'].width = 95

# 图例说明
ws_cards.merge_cells('A1:C1')
legend = ws_cards.cell(row=1, column=1,
    value='v5.0 算法详细卡片：■ 旗舰卡（40要素完整卡，深蓝标题，40张）  □ 骨架卡（15要素紧凑卡，灰色标题，95张，待升级）')
legend.font = Font(name='微软雅黑', size=11, bold=True, color=DEEP_BLUE)
legend.fill = PatternFill(start_color=WARM_GREY, end_color=WARM_GREY, fill_type='solid')
legend.alignment = Alignment(horizontal='center', vertical='center')
for c in range(1, 4):
    ws_cards.cell(row=1, column=c).border = thin_border
    ws_cards.cell(row=1, column=c).fill = PatternFill(start_color=WARM_GREY, end_color=WARM_GREY, fill_type='solid')

field_map_flag = [
    ('算法编号', '必须', 'sn', 1), ('算法名称', '必须', 'name', 1),
    ('版本/编制人/复核人', '必须', None, 1), ('适用业务场景', '必须', 'scene', 1),
    ('审计目标', '必须', 'objective', 1), ('风险假设', '必须', 'risk_hypothesis', 2),
    ('适用范围', '必须', 'scope_yes', 1), ('不适用范围', '必须', 'scope_no', 1),
    ('法规及业务依据', '必须', 'law_basis', 2), ('输入数据表', '必须', 'data_tables', 2),
    ('核心字段', '必须', 'fields', 2), ('主键与关联键', '必须', 'keys', 1),
    ('数据质量检查', '必须', 'data_quality', 2), ('计算公式/步骤', '必须', 'calc_logic', 3),
    ('参数与阈值', '必须', 'threshold', 1), ('阈值依据', '必须', 'threshold_basis', 1),
    ('输出字段', '必须', 'output_fields', 2), ('疑点解释模板', '必须', 'explain', 2),
    ('追加证据', '必须', 'evidence', 2), ('人工核查程序', '必须', 'check_procedure', 2),
    ('结论边界', '必须', 'conclusion_boundary', 2), ('测试案例', '必须', 'test_cases', 1),
    ('历史回测结果', '必须', 'backtest', 2), ('风险评分', '锦上添花', 'risk_score', 1),
    ('多规则组合', '锦上添花', 'multi_rule', 1), ('敏感性分析', '锦上添花', 'sensitivity', 1),
    ('误报率与漏报率', '锦上添花', 'fpr_fnr', 1), ('可视化方案', '锦上添花', 'visual', 1),
    ('跨场景复用', '锦上添花', 'reuse', 1), ('运行与触发条件', '锦上添花', 'perf', 1),
    ('脱敏和权限', '锦上添花', 'privacy', 1), ('数据就绪度评估', '必须', 'data_readiness', 1),
    ('数据质量分级', '必须', 'data_grade', 1), ('验证标准（目标象限）', '必须', 'verify_standard', 1),
    ('底稿嵌入模板编号', '必须', 'workpaper_template', 1), ('可解释性输出模板', '必须', 'explainability', 2),
    ('算法退役条件', '必须', 'retire_condition', 1), ('定期复查周期', '必须', 'review_cycle', 1),
    ('前置依赖算法', '锦上添花', 'dependency', 1), ('预期审计价值', '锦上添花', 'expected_value', 1),
    ('历史产出记录', '锦上添花', 'history_output', 1),
]
field_map_skel = [
    ('算法编号', '骨架', 'sn'), ('算法名称', '骨架', 'name'),
    ('版本/编制人', '骨架', None), ('适用业务场景', '骨架', 'scene'),
    ('审计目标', '骨架', 'objective'), ('风险假设', '骨架', 'risk_hypothesis'),
    ('核心信号', '骨架', 'signals'), ('公式/计算逻辑', '骨架', 'calc_logic'),
    ('参数与阈值', '骨架', 'threshold'), ('证据清单', '骨架', 'evidence'),
    ('人工核查程序', '骨架', 'check_procedure'), ('法规及业务依据', '骨架', 'law_basis'),
    ('来源文献', '骨架', 'source_file'), ('复杂度', '骨架', 'complexity'),
    ('Agent映射', '骨架', 'agent'),
]

def fill_card(ws, algo, start_row, is_flag=True):
    r = start_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    title_cell = ws.cell(row=r, column=1, value=f'算法卡：{algo["sn"]} — {algo["name"]}')
    if is_flag:
        title_cell.font = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color=DEEP_BLUE, end_color=DEEP_BLUE, fill_type='solid')
    else:
        title_cell.font = Font(name='微软雅黑', size=13, bold=True, color=DEEP_BLUE)
        title_fill = PatternFill(start_color=SKELETON_GREY, end_color=SKELETON_GREY, fill_type='solid')
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    for c in range(1, 4):
        ws.cell(row=r, column=c).border = thin_border
        ws.cell(row=r, column=c).fill = title_fill
    r += 1

    for c, h in enumerate(['要素名称', '重要级别', '填写内容'], 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
    r += 1

    fields = field_map_flag if is_flag else field_map_skel
    for name, level, key, *rest in fields:
        row_h = rest[0] if rest else 1
        if key is None:
            if is_flag:
                val = f'{VERSION} / 融策算法工作组 / 融策平头哥'
            else:
                val = f'骨架卡 v1.0（batch{algo.get("batch", "-")}）/ 融策算法工作组'
        else:
            val = algo.get(key, '——')
            if isinstance(val, list):
                val = _fmt(val, '\n')
            if val is None:
                val = '——'
            if key == 'signals' and val:
                items = str(val).split('\n')
                val = '\n'.join(f'信号{i+1}：{x}' for i, x in enumerate(items) if x.strip())
            if key == 'check_procedure' and isinstance(algo.get(key), list):
                val = _fmt(algo.get(key), '\n')
        cell_name = style_data_cell(ws, r, 1)
        cell_name.value = name
        if level == '必须':
            cell_name.font = Font(name='微软雅黑', size=10, bold=True, color='CC0000')
        cell_level = style_data_cell(ws, r, 2)
        cell_level.value = level
        cell_level.alignment = center_align
        if level == '必须':
            cell_level.font = Font(name='微软雅黑', size=10, bold=True, color='CC0000')
        cell_val = style_data_cell(ws, r, 3)
        cell_val.value = str(val)
        if row_h > 1:
            end_row = r + row_h - 1
            for col in (1, 2, 3):
                ws.merge_cells(start_row=r, start_column=col, end_row=end_row, end_column=col)
            for mr in range(r, end_row + 1):
                for mc in range(1, 4):
                    ws.cell(row=mr, column=mc).border = thin_border
                    ws.cell(row=mr, column=mc).alignment = wrap_align
            r = end_row + 1
        else:
            r += 1
    return r

current_row = 2
for i, algo in enumerate(algorithms_v4):
    current_row = fill_card(ws_cards, algo, current_row, is_flag=True)
    current_row += 2
for c in skeleton_cards:
    current_row = fill_card(ws_cards, c, current_row, is_flag=False)
    current_row += 2
ws_cards.freeze_panes = 'A2'

# ========== Sheet3: ☆业务场景地图 ==========
ws_scene = wb.create_sheet('☆业务场景地图')
scene_headers = ['业务场景', '重点审计目标', '典型风险机制', '匹配算法编号', '匹配算法名称', '核心数据对象', '数据就绪度']
for c, h in enumerate(scene_headers, 1):
    ws_scene.cell(row=1, column=c, value=h)
style_header_row(ws_scene, 1, len(scene_headers))
for r, row in enumerate(scene_data_v5, 2):
    for c, val in enumerate(row, 1):
        cell = style_data_cell(ws_scene, r, c)
        cell.value = val
for i, w in enumerate([16, 32, 20, 30, 34, 24, 18], 1):
    ws_scene.column_dimensions[get_column_letter(i)].width = w
ws_scene.freeze_panes = 'A2'

# ========== Sheet4: ☆风险机制矩阵 ==========
ws_risk = wb.create_sheet('☆风险机制矩阵')
risk_headers = ['风险机制', '算法族', '算法编号', '算法名称', '可跨场景复用', '退化风险', '维护周期']
for c, h in enumerate(risk_headers, 1):
    ws_risk.cell(row=1, column=c, value=h)
style_header_row(ws_risk, 1, len(risk_headers))
for r, row in enumerate(risk_data_v5, 2):
    for c, val in enumerate(row, 1):
        cell = style_data_cell(ws_risk, r, c)
        cell.value = val
for i, w in enumerate([12, 12, 22, 40, 30, 26, 18], 1):
    ws_risk.column_dimensions[get_column_letter(i)].width = w
ws_risk.freeze_panes = 'A2'

# ========== Sheet5: ☆建设路线图 ==========
ws_road = wb.create_sheet('☆建设路线图')
road_headers = ['阶段', '重点工作', '建议产出', '验收标准', '成熟度目标', '当前状态']
for c, h in enumerate(road_headers, 1):
    ws_road.cell(row=1, column=c, value=h)
style_header_row(ws_road, 1, len(road_headers))
for r, row in enumerate(road_data_v5, 2):
    for c, val in enumerate(row, 1):
        cell = style_data_cell(ws_road, r, c)
        cell.value = val
for i, w in enumerate([22, 44, 36, 42, 16, 26], 1):
    ws_road.column_dimensions[get_column_letter(i)].width = w
ws_road.freeze_panes = 'A2'

# ========== Sheet6: ☆使用声明 ==========
ws_decl = wb.create_sheet('☆使用声明')
decl_headers = ['序号', '声明事项', '具体内容', '法律/准则依据']
for c, h in enumerate(decl_headers, 1):
    ws_decl.cell(row=1, column=c, value=h)
style_header_row(ws_decl, 1, len(decl_headers))
for r, row in enumerate(decl_data_v5, 2):
    for c, val in enumerate(row, 1):
        cell = style_data_cell(ws_decl, r, c)
        cell.value = val
for i, w in enumerate([8, 22, 80, 50], 1):
    ws_decl.column_dimensions[get_column_letter(i)].width = w
ws_decl.freeze_panes = 'A2'

# ========== Sheet7: ☆文献来源 ==========
ws_papers = wb.create_sheet('☆文献来源')
paper_headers = ['算法编号', '算法名称', '来源论文', '作者', '期刊/年份', '案例场景']
for c, h in enumerate(paper_headers, 1):
    ws_papers.cell(row=1, column=c, value=h)
style_header_row(ws_papers, 1, len(paper_headers))
for r, row in enumerate(paper_data_v5, 2):
    for c, val in enumerate(row, 1):
        cell = style_data_cell(ws_papers, r, c)
        cell.value = val
for i, w in enumerate([22, 36, 55, 22, 26, 50], 1):
    ws_papers.column_dimensions[get_column_letter(i)].width = w
ws_papers.freeze_panes = 'A2'

wb.save(OUTPUT_PATH)
print(f'\n✅ 政府审计算法资产库 v5.0 已生成！')
print(f'📁 输出文件：{OUTPUT_PATH}')
print(f'📊 包含：')
print(f'   - 总览：{len(overview)}个算法（旗舰40 + 骨架95）')
print(f'   - 详细卡片：{len(algorithms_v4)}张40要素旗舰卡 + {len(skeleton_cards)}张15要素骨架卡')
print(f'   - 业务场景地图：{len(scene_data_v5)}行（12条业务线+延伸场景）')
print(f'   - 风险机制矩阵：{len(risk_data_v5)}行（覆盖全部135个算法）')
print(f'   - 建设路线图：{len(road_data_v5)}阶段（v5.0里程碑：135算法完成全量提取）')
print(f'   - 使用声明：{len(decl_data_v5)}条')
print(f'   - 文献来源：{len(paper_data_v5)}条')

# ========== 10. 验证 ==========
print('\n⑨ 验证：')
wb2 = openpyxl.load_workbook(OUTPUT_PATH)
ws1 = wb2['☆算法资产库总览']
n_flag = sum(1 for row in ws1.iter_rows(min_row=2, values_only=True) if row[3] == '旗舰')
n_skel = sum(1 for row in ws1.iter_rows(min_row=2, values_only=True) if row[3] == '骨架')
sns = [row[1] for row in ws1.iter_rows(min_row=2, values_only=True)]
print(f'   - 总览算法数：{ws1.max_row - 1}（旗舰{n_flag} + 骨架{n_skel}）')
print(f'   - SN唯一性：{"✅ 全部唯一" if len(sns) == len(set(sns)) else "❌ 存在重复"}')
ws3 = wb2['☆业务场景地图']
print(f'   - 场景地图行数：{ws3.max_row - 1}')
ws4 = wb2['☆风险机制矩阵']
print(f'   - 风险矩阵行数：{ws4.max_row - 1}（覆盖算法数≥135：{"✅" if ws4.max_row - 1 >= 135 else "❌"}）')
print('   - Sheet列表：', wb2.sheetnames)
print('\n🎉 v5.0 合并完成：40张旗舰卡 + 95张骨架卡 = 135算法，全量提取里程碑达成')
