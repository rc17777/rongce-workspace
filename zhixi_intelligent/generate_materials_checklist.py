"""
生成「审计数据采集·被采集单位资料准备清单」Excel
对应合同第3条第1款(1)项，13个行业的完整资料清单
"""
import sys
sys.path.insert(0, r"D:\openclaw-workspace")
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = Workbook()

# ============================================================
# 样式定义
# ============================================================
header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1a1a2e", end_color="16213e", fill_type="solid")
title_font = Font(name="微软雅黑", size=16, bold=True, color="1a1a2e")
subtitle_font = Font(name="微软雅黑", size=10, color="7f8c8d")
section_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
section_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
normal_font = Font(name="微软雅黑", size=10)
bold_font = Font(name="微软雅黑", size=10, bold=True)
required_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")  # 必需项 浅红
important_fill = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")  # 重要 浅黄
optional_fill = PatternFill(start_color="EAFAF1", end_color="EAFAF1", fill_type="solid")   # 建议 浅绿
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
wrap_align = Alignment(wrap_text=True, vertical="center")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ============================================================
# 13个行业的资料清单数据
# ============================================================
industry_materials = [
    {
        "industry": "工商",
        "unit": "省/市市场监督管理局",
        "data_source": "企业登记注册系统、年报系统、信用信息公示系统",
        "materials": [
            ("数据库备份文件（.dmp / .bak）或直连权限", "必需", "含全部在册企业基本信息、股东、高管、注册资本、经营范围、经营状态等"),
            ("企业年度报告数据", "必需", "近3个完整年度，含资产负债表、利润表、现金流量表"),
            ("企业变更登记台账", "必需", "逐条变更记录，含变更前后内容、变更日期"),
            ("行政处罚和经营异常名录", "必需", "全部记录，含处罚原因、处罚结果、列入/移出日期"),
            ("股权出质/冻结信息", "重要", "用于关联分析企业间控制关系"),
            ("数据字典/数据库设计文档", "必需", "表结构说明，含字段含义、代码表、关联关系"),
            ("系统管理员联系方式", "必需", "用于协调数据导出和技术对接"),
            ("数据安全管理联系人及电话", "必需", "签订保密协议和确认数据用途"),
            ("API接口文档（如有政务共享平台对接）", "建议", "用于政务共享平台自动采集"),
        ]
    },
    {
        "industry": "财政财务",
        "unit": "省/市/区县财政局",
        "data_source": "财政一体化系统、国库集中支付系统、预算编制系统、部门决算系统",
        "materials": [
            ("预算编制数据", "必需", "一般公共预算/政府性基金/国有资本经营/社保基金四本预算完整数据，含支出功能分类到项级"),
            ("国库集中支付数据", "必需", "近3年全量支付明细，含指标/计划/支付/清算全流程，支出经济分类到款级"),
            ("总预算会计账套数据", "必需", "含科目余额表、总账、明细账、辅助账，覆盖所有财政专户"),
            ("部门决算数据", "必需", "全部预算单位决算报表（主表+附表），含年初预算/调整预算/决算对比"),
            ("非税收入征缴数据", "必需", "缴款人/执收单位/收费项目/金额/时间/分成比例"),
            ("政府债务数据", "重要", "一般债/专项债台账，含发行/还本/付息/用途/项目对应"),
            ("政府采购数据", "重要", "采购预算/采购计划/招标公告/中标结果/合同备案全流程"),
            ("财政专户银行对账单（电子版）", "必需", "用于校验支付数据的完整性和准确性"),
            ("往来款/暂付款台账", "重要", "逐笔明细，含借款单位/金额/日期/用途/还款情况"),
            ("转移支付资金分配下达数据", "重要", "中央/省级转移支付指标的分配/下达/拨付/使用全链条"),
            ("数据字典及系统操作手册", "必需", "含各业务表字段说明、代码表（功能分类/经济分类/资金来源/支出用途等）"),
            ("信息中心技术负责人联系方式", "必需", ""),
        ]
    },
    {
        "industry": "民生民政",
        "unit": "省/市/县民政局",
        "data_source": "社会救助系统、低保系统、特困供养系统、残疾人补贴系统、殡葬系统",
        "materials": [
            ("城乡低保在册人员数据", "必需", "含姓名/身份证号/家庭人口/保障金额/保障类别/审批日期"),
            ("特困人员供养数据", "必需", "供养对象信息、供养标准、供养机构、护理等级"),
            ("临时救助台账", "必需", "救助原因/金额/时间/救助对象"),
            ("残疾人两项补贴数据", "必需", "含补贴类型（生活/护理）/金额/发放记录"),
            ("高龄津贴发放数据", "重要", "80岁以上老人津贴发放台账"),
            ("孤儿/事实无人抚养儿童保障数据", "重要", "基本生活保障发放台账"),
            ("殡葬火化数据", "重要", "用于跨行业比对验证生存状态"),
            ("社会组织和慈善机构登记数据", "建议", ""),
            ("数据安全管理及保密协议", "必需", "民生数据涉及大量个人隐私"),
        ]
    },
    {
        "industry": "教科文卫",
        "unit": "省/市/县教育局、科技局、文旅局、卫健委",
        "data_source": "学籍管理系统、学生资助系统、科技项目管理系统、文化场馆管理系统",
        "materials": [
            ("中小学学籍数据", "必需", "在校学生基本信息，含学籍号/姓名/身份证号/学校/年级/班级"),
            ("学生资助发放数据", "必需", "含资助类型（营养餐/助学金/免学费）/金额/时间/受益学生"),
            ("教师编制及工资发放数据", "重要", "用于核查教师待遇落实情况"),
            ("教育经费拨款及使用数据", "必需", "含项目名称/拨款金额/支出明细/结余"),
            ("校舍建设及维修改造项目数据", "重要", "含立项/批复/招标/合同/验收/审计全流程"),
            ("营养餐专项资金管理数据", "必需", "经费拨付/支出/受益学生人数/供餐企业"),
            ("科技项目立项及经费使用数据", "建议", "用于科技专项资金审计"),
            ("公共文化服务资金使用数据", "建议", "三馆一站免费开放经费等"),
            ("卫健委：基层医疗/公共卫生服务数据", "重要", "基本公卫项目资金拨付/服务人次/考核结果"),
        ]
    },
    {
        "industry": "社保医保",
        "unit": "省/市人社局、医保局",
        "data_source": "社保核心系统、医保结算系统、养老保险系统、工伤保险系统",
        "materials": [
            ("养老保险参保及发放数据", "必需", "在职/退休人员基本信息，含缴费基数/缴费年限/退休金/发放记录"),
            ("医疗保险参保及结算数据", "必需", "参保人信息/定点机构/诊疗项目/药品明细/费用/报销比例"),
            ("失业保险参保及发放数据", "必需", "含失业登记/失业金申领/发放金额/发放期限"),
            ("工伤保险认定及待遇数据", "重要", "工伤认定/劳动能力鉴定/待遇支付"),
            ("生育保险待遇支付数据", "重要", ""),
            ("社保基金收支及结余数据", "必需", "会计总账/明细账/银行对账单"),
            ("定点医疗机构/药店结算明细", "必需", "用于大数据分析识别异常结算"),
            ("异地就医结算数据", "重要", ""),
            ("社保卡发放及使用记录", "建议", "用于跨行业数据关联"),
            ("数据字典/系统ER图/接口文档", "必需", ""),
        ]
    },
    {
        "industry": "公积金",
        "unit": "省/市住房公积金管理中心",
        "data_source": "公积金核心业务系统、贷款管理系统",
        "materials": [
            ("住房公积金缴存数据", "必需", "含单位/个人缴存基数/比例/月缴额/累计余额"),
            ("住房公积金提取数据", "必需", "提取原因（购房/租房/还贷/退休等）/金额/时间"),
            ("住房公积金贷款数据", "必需", "贷款申请/审批/发放/还款明细，含贷款金额/利率/期限/抵押物"),
            ("公积金增值收益及分配数据", "重要", "含管理费用支出/风险准备金/廉租房补充资金"),
            ("单位缴存台账", "重要", "用于与工商数据进行比对，识别未依法缴存的单位"),
            ("银行专户对账单（电子版）", "必需", "用于校验数据完整性"),
        ]
    },
    {
        "industry": "企业及金融机构",
        "unit": "省/市国资委、金融监管局、国有企业集团",
        "data_source": "国有企业财务系统、金融监管系统、产权管理系统",
        "materials": [
            ("国有企业年度财务决算数据", "必需", "含合并报表/单体报表/审计报告/附注"),
            ("国有企业预算编制及执行数据", "必需", "年初预算/年中调整/年末实际执行对比"),
            ("国有企业重大投资决策文件", "重要", "含立项/可研/董事会决议/批复/合同"),
            ("国有企业高管薪酬数据", "重要", "含年薪/绩效/津贴/福利/履职待遇"),
            ("国有企业产权登记及变更数据", "必需", "含股权结构/对外投资/产权转让/增资扩股"),
            ('国有企业"三重一大"决策记录', "重要", "重大决策/重要人事任免/重大项目安排/大额资金使用"),
            ("融资平台公司债务数据", "必需", "含融资方式/金额/期限/利率/用途/偿还计划"),
            ("金融机构信贷投放及不良资产数据", "必需", "含贷款企业/金额/利率/行业/担保方式/五级分类"),
            ("国有企业负责人经济责任审计报告", "建议", "已有审计结果可复用"),
            ("信息安全保密协议", "必需", "金融机构数据涉及商业秘密"),
        ]
    },
    {
        "industry": "重大投资项目",
        "unit": "省/市发改委、交通局、住建局、水利局",
        "data_source": "投资项目在线审批监管平台、重大项目管理系统",
        "materials": [
            ("重大项目台账", "必需", "含项目名称/代码/总投资/年度投资/建设内容/工期/责任单位"),
            ("项目立项批复文件", "必需", "含项目建议书/可研报告/初步设计及其批复"),
            ("招投标数据", "必需", "招标公告/招标文件/投标文件/评标报告/中标通知书/合同"),
            ("项目进度及资金拨付数据", "必需", "施工进度/计量支付/资金到位/支出明细/工程变更签证"),
            ("项目竣工验收资料", "必需", "竣工报告/验收报告/结算报告/决算报告/审计报告"),
            ("土地征收及拆迁补偿数据", "重要", "征收方案/补偿标准/补偿款发放台账"),
            ("环评/水保/能评等专项审批文件", "重要", ""),
            ("项目绩效目标及自评报告", "重要", "用于绩效审计"),
            ("项目建设单位/监理单位/设计单位联系方式", "必需", "用于现场核实和延伸调查"),
        ]
    },
    {
        "industry": "公共资源交易",
        "unit": "省/市公共资源交易中心",
        "data_source": "公共资源交易平台、政府采购系统、建设工程交易系统、土地矿权交易系统",
        "materials": [
            ("工程建设项目招投标数据", "必需", "全量交易记录：招标/投标/开标/评标/定标，含投标人/报价/得分/中标"),
            ("政府采购交易数据", "必需", "采购预算/采购方式/供应商/报价/中标/合同"),
            ("土地出让/矿权交易数据", "重要", "出让公告/竞买人/起始价/成交价/出让合同"),
            ("国有产权交易数据", "重要", "产权转让/增资扩股/资产转让"),
            ("投标保证金缴纳及退还数据", "必需", "用于围标串标分析——关键线索来源"),
            ("供应商/投标人基本信息", "必需", "含统一社会信用代码/法人/联系人/注册资本/资质等级"),
            ("评标专家库及抽取记录", "重要", "用于识别评标专家利益冲突"),
            ("交易异常/质疑/投诉处理记录", "重要", ""),
            ("系统数据字典/接口规范", "必需", ""),
        ]
    },
    {
        "industry": "农业",
        "unit": "省/市/县农业农村局",
        "data_source": "惠农补贴系统、农村三资管理系统、高标准农田管理系统",
        "materials": [
            ("耕地地力保护补贴发放数据", "必需", "含农户姓名/身份证号/耕地面积/补贴标准/发放金额/账号"),
            ("农机购置补贴数据", "必需", "含购机者/机具型号/出厂编号/补贴金额/经销商"),
            ("农业保险保费补贴数据", "重要", "投保面积/品种/保费/理赔"),
            ("高标准农田建设项目数据", "必需", "含项目规划/立项/招标/施工/验收/资金拨付"),
            ('农村集体"三资"管理数据', "必需", "资金/资产/资源台账，含村级财务收支/资产登记/资源发包"),
            ("新型农业经营主体数据", "重要", "农业龙头企业/合作社/家庭农场登记信息"),
            ("农村土地承包经营权确权登记数据", "重要", "用于核实耕地面积和补贴对象一致性"),
            ('惠农补贴"一卡通"发放系统数据', "必需", "用于跨行业比对（与低保/社保/公职人员等交叉验证）"),
        ]
    },
    {
        "industry": "高校",
        "unit": "省属/市属高校",
        "data_source": "高校财务系统、教务管理系统、科研管理系统、国有资产管理系统",
        "materials": [
            ("高校财务账套数据", "必需", "含会计科目/凭证/明细账/总账/辅助账，覆盖教育事业费/科研经费/基建经费"),
            ("收费系统数据", "必需", "学费/住宿费/代办费等收费标准和实际收缴明细"),
            ("学生资助资金管理数据", "必需", "奖助学金/助学贷款/勤工助学/学费减免"),
            ("科研项目经费管理数据", "必需", "纵向/横向科研项目立项/经费到账/支出明细/结题结余"),
            ("资产管理系统数据", "必需", "固定资产/无形资产的登记/使用/处置，含房产/设备/车辆/软件"),
            ("采购及招投标数据", "重要", "货物/服务/工程采购的招标/合同/验收/付款"),
            ("基建项目数据", "重要", "新校区建设/大型修缮的立项/概算/招标/合同/结算"),
            ("教职工薪酬管理数据", "重要", "含编制内/编外/劳务派遣人员工资/绩效/津贴"),
            ("学校章程/内控制度文件", "建议", "了解重大经济决策程序和审批权限"),
            ("预算编制及执行情况数据", "必需", "年度预算/调整预算/决算对比"),
        ]
    },
    {
        "industry": "医院",
        "unit": "省/市/县级公立医院",
        "data_source": "HIS系统（医院信息系统）、HRP系统（医院资源规划）、LIS/PACS/RIS等系统",
        "materials": [
            ("医院财务账套数据", "必需", "含医疗收入/药品收入/财政补助/科教项目等分类核算"),
            ("HIS收费系统数据", "必需", "门诊/住院收费明细，含诊疗项目/药品/耗材/检查检验/手术等"),
            ("药品进销存数据", "必需", "药品采购/入库/出库/库存台账，含品名/规格/厂家/批次/进价/售价"),
            ("高值耗材管理数据", "必需", "含骨科/心内/眼科等高值耗材的采购/使用追溯"),
            ("医保结算数据", "必需", "医保患者刷卡结算明细，含统筹支付/个账支付/自费"),
            ("医疗设备采购及管理数据", "重要", "大型设备采购论证/招标/合同/付款/使用效益分析"),
            ("基建/修缮项目数据", "重要", "新院区建设/病区改造项目全流程"),
            ("人员编制及薪酬数据", "重要", "在职/编外/规培人员工资绩效，含夜班费/手术费等"),
            ("医疗服务价格和收费项目标准", "必需", "物价部门批复的收费项目和标准"),
            ("医疗收入和成本核算数据", "重要", "科室全成本核算数据，用于绩效评价"),
            ("病案首页数据", "建议", "DRG/DIP付费改革相关的病案数据"),
        ]
    },
    {
        "industry": "其他",
        "unit": "根据审计需要灵活确定",
        "data_source": "根据具体审计项目的行业特征确定",
        "materials": [
            ("根据具体审计项目需求确定", "必需", "如：自然资源/生态环境/乡村振兴/应急救灾/等专项领域数据"),
            ("政务数据共享交换平台对接凭证", "必需", "用于跨部门数据共享交换"),
            ("其他行业主管部门的法定职能数据", "必需", "根据审计实施方案中确定的审计范围来决定"),
        ]
    },
]

# ============================================================
# Sheet 1: 总览目录
# ============================================================
ws0 = wb.active
ws0.title = "总览目录"

# 标题
ws0.merge_cells("A1:H1")
ws0["A1"] = "四川省审计厅 · 数据采集被采集单位资料准备清单"
ws0["A1"].font = title_font
ws0["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws0.row_dimensions[1].height = 36

ws0.merge_cells("A2:H2")
ws0["A2"] = f"合同编号: N5100012024002699 | 编制单位: 融策会计师事务所 | 编制日期: {datetime.now().strftime('%Y-%m-%d')}"
ws0["A2"].font = subtitle_font
ws0["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws0.row_dimensions[2].height = 22

# 目录表头
dir_headers = ["序号", "行业", "被采集单位", "主要数据来源系统", "必需项数", "重要项数", "建议项数", "合计"]
for col, h in enumerate(dir_headers, 1):
    cell = ws0.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# 填充目录数据
for i, item in enumerate(industry_materials):
    row = 5 + i
    required_count = sum(1 for m in item["materials"] if m[1] == "必需")
    important_count = sum(1 for m in item["materials"] if m[1] == "重要")
    optional_count = sum(1 for m in item["materials"] if m[1] == "建议")

    data = [i+1, item["industry"], item["unit"], item["data_source"],
            required_count, important_count, optional_count, len(item["materials"])]
    for col, val in enumerate(data, 1):
        cell = ws0.cell(row=row, column=col, value=val)
        cell.font = normal_font
        cell.alignment = center_align if col != 4 else Alignment(wrap_text=True, vertical="center")
        cell.border = thin_border
    ws0.row_dimensions[row].height = 32

# 汇总行
total_row = 5 + len(industry_materials)
total_required = sum(sum(1 for m in item["materials"] if m[1] == "必需") for item in industry_materials)
total_important = sum(sum(1 for m in item["materials"] if m[1] == "重要") for item in industry_materials)
total_optional = sum(sum(1 for m in item["materials"] if m[1] == "建议") for item in industry_materials)
total_all = total_required + total_important + total_optional

for col, val in enumerate(["", "合计", "", "", total_required, total_important, total_optional, total_all], 1):
    cell = ws0.cell(row=total_row, column=col, value=val)
    cell.font = bold_font
    cell.border = thin_border
    cell.alignment = center_align
    cell.fill = PatternFill(start_color="D5D8DC", end_color="D5D8DC", fill_type="solid")

# 图例
legend_row = total_row + 2
ws0.cell(row=legend_row, column=1, value="图例：").font = bold_font
ws0.cell(row=legend_row, column=2, value="").fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
ws0.cell(row=legend_row, column=3, value="必需提供（不提供则无法完成采集）").font = normal_font
ws0.cell(row=legend_row+1, column=2, value="").fill = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
ws0.cell(row=legend_row+1, column=3, value="重要材料（建议提供，影响审计深度）").font = normal_font
ws0.cell(row=legend_row+2, column=2, value="").fill = PatternFill(start_color="EAFAF1", end_color="EAFAF1", fill_type="solid")
ws0.cell(row=legend_row+2, column=3, value="建议提供（辅助分析用）").font = normal_font

# 列宽
col_widths = [6, 14, 22, 40, 10, 10, 10, 8]
for i, w in enumerate(col_widths, 1):
    ws0.column_dimensions[get_column_letter(i)].width = w

# 冻结
ws0.freeze_panes = "A5"

# ============================================================
# Sheet 2: 详细清单
# ============================================================
ws1 = wb.create_sheet("详细资料清单")

ws1.merge_cells("A1:G1")
ws1["A1"] = "审计数据采集 · 被采集单位资料准备详细清单"
ws1["A1"].font = title_font
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 36

ws1.merge_cells("A2:G2")
ws1["A2"] = "说明：本清单按合同第3条服务内容编制，被采集单位需在采集前10个工作日内完成资料准备。标红为必需项。"
ws1["A2"].font = Font(name="微软雅黑", size=9, color="E74C3C")
ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")

# 表头
detail_headers = ["序号", "行业", "资料名称", "优先级", "说明/备注", "参考格式", "采集方式"]
for col, h in enumerate(detail_headers, 1):
    cell = ws1.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

priority_fills = {
    "必需": required_fill,
    "重要": important_fill,
    "建议": optional_fill
}

row_idx = 5
seq = 0
for item in industry_materials:
    # 行业标题行
    ws1.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
    title_cell = ws1.cell(row=row_idx, column=1,
                          value=f"{item['industry']} — {item['unit']}  |  主要系统: {item['data_source']}")
    title_cell.font = section_font
    title_cell.fill = section_fill
    title_cell.alignment = Alignment(vertical="center")
    for c in range(1, 8):
        ws1.cell(row=row_idx, column=c).border = thin_border
    ws1.row_dimensions[row_idx].height = 26
    row_idx += 1

    # 资料明细
    for mat in item["materials"]:
        seq += 1
        material_name = mat[0]
        priority = mat[1]
        note = mat[2]

        # 格式推测
        if "数据库" in material_name or "系统" in material_name or "账套" in material_name:
            fmt = "数据库备份(.dmp/.bak)\n或CSV/Excel导出"
            method = "拷贝采集 / 直连采集"
        elif "台账" in material_name or "明细" in material_name or "表" in material_name:
            fmt = "Excel(.xlsx)\n或CSV(.csv)"
            method = "拷贝采集"
        elif "合同" in material_name or "文件" in material_name or "报告" in material_name or "记录" in material_name or "批复" in material_name:
            fmt = "PDF / Word / 扫描件"
            method = "OCR识别 + 要素提取"
        elif "制度" in material_name or "章程" in material_name or "手册" in material_name:
            fmt = "PDF / Word"
            method = "拷贝采集"
        elif "联系方式" in material_name or "联系人" in material_name:
            fmt = "姓名+电话+职务"
            method = "书面确认"
        elif "协议" in material_name or "凭证" in material_name:
            fmt = "PDF / Word / 扫描件"
            method = "OCR识别"
        elif "接口" in material_name or "API" in material_name:
            fmt = "接口规范文档\n(.docx/.pdf)"
            method = "政务共享平台交换"
        elif "对账单" in material_name:
            fmt = "Excel / PDF / 电子流水"
            method = "拷贝采集"
        elif "数据字典" in material_name or "ER图" in material_name:
            fmt = "Excel / Word\n数据库导出"
            method = "拷贝采集"
        else:
            fmt = "Excel / PDF"
            method = "拷贝采集"

        data = [seq, item["industry"], material_name, priority, note, fmt, method]
        for col, val in enumerate(data, 1):
            cell = ws1.cell(row=row_idx, column=col, value=val)
            cell.font = normal_font
            cell.alignment = center_align if col in (1, 4) else wrap_align
            cell.border = thin_border
            # 优先级着色
            if col == 4:
                cell.fill = priority_fills.get(priority, PatternFill())

        ws1.row_dimensions[row_idx].height = 42
        row_idx += 1

# 列宽
detail_col_widths = [6, 12, 38, 8, 45, 18, 16]
for i, w in enumerate(detail_col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.freeze_panes = "A5"

# ============================================================
# Sheet 3: 采集工作流程建议
# ============================================================
ws2 = wb.create_sheet("采集工作流程")

ws2.merge_cells("A1:E1")
ws2["A1"] = "数据采集工作流程与时间安排建议"
ws2["A1"].font = title_font
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 36

flow_headers = ["步骤", "工作内容", "责任方", "时间要求", "输出成果"]
for col, h in enumerate(flow_headers, 1):
    cell = ws2.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

flow_data = [
    ("第1步", "发送资料准备清单给被采集单位", "审计厅数据管理部门", "采集开始前15个工作日", "《资料准备清单》签收回执"),
    ("第2步", "被采集单位准备资料并反馈准备情况", "被采集单位", "收到清单后10个工作日内", "资料准备情况反馈表"),
    ("第3步", "确认资料完整性，签订保密协议", "审计厅 + 被采集单位", "采集开始前5个工作日", "保密协议 + 资料完整性确认单"),
    ("第4步", "现场/远程数据采集", "采集工程师(融策)", "合同约定期限内（按行业分批）", "原始数据库备份文件"),
    ("第5步", "数据恢复至审计大数据中心", "采集工程师(融策)", "采集后3个工作日内", "恢复确认报告"),
    ("第6步", "数据校验（完整性/总量/业务规则）", "智析智能 + 采集工程师复核", "恢复后1个工作日内", "《数据校验报告》"),
    ("第7步", "数据清洗及标准化", "智析智能 + 采集工程师复核", "校验后5个工作日内", "标准库数据 + 处理记录"),
    ("第8步", "元数据扫描 + 数据资源目录建立", "智析智能 · 元数据扫描器", "标准化完成后", "《元数据报告》+ 数据资源目录"),
    ("第9步", "更新数据采集进度看板", "智析智能 · 采集进度看板", "每行业完成后", "最新HTML看板 + JSON报告"),
    ("第10步", "数据移交给分析团队", "采集工程师(融策)", "全行业采集完成后", "数据移交确认单"),
]

for i, (step, content, responsible, timing, output) in enumerate(flow_data):
    row = 4 + i
    for col, val in enumerate([step, content, responsible, timing, output], 1):
        cell = ws2.cell(row=row, column=col, value=val)
        cell.font = normal_font
        cell.alignment = center_align if col in (1, 3, 4) else wrap_align
        cell.border = thin_border
    ws2.row_dimensions[row].height = 36

flow_col_widths = [8, 38, 22, 22, 30]
for i, w in enumerate(flow_col_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.freeze_panes = "A4"

# ============================================================
# ✨ Sheet 4: 给被采集单位的通知函模板
# ============================================================
ws3 = wb.create_sheet("通知函模板")

ws3.merge_cells("A1:C1")
ws3["A1"] = "数据采集通知函模板（供审计厅发送给被采集单位）"
ws3["A1"].font = title_font
ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 36

template_text = """关于配合提供审计电子数据的函

XXXXXXXX单位：

根据《中华人民共和国审计法》及审计署关于金审工程建设相关工作要求，四川省审计厅现开展年度审计电子数据采集工作。
贵单位为本次数据采集的被采集单位之一，请按照以下要求配合提供相关电子数据。

一、数据提供范围
请参照随函附送的《数据采集资料准备清单》中"XXX行业"部分所列资料清单进行准备。

二、数据提供方式
□ 数据库备份文件（.dmp / .bak）提供
□ 数据文件导出（Excel / CSV格式）
□ 数据库直连只读授权（需开通临时只读账号）
□ 政务共享平台数据交换

三、时间要求
请于XXXX年XX月XX日前完成资料准备，并将准备情况反馈表发送至审计厅数据管理部门。

四、保密要求
所有提供的数据将严格按照《中华人民共和国审计法》《中华人民共和国数据安全法》等法律法规进行管理和使用，
仅限用于审计工作目的，审计厅将与被采集单位签订数据安全保密协议。

五、联系方式
数据采集联系人：XXX  电话：XXX-XXXXXXXX
技术对接联系人：XXX  电话：XXX-XXXXXXXX

特此函告。

附件：
1. 数据采集资料准备清单
2. 资料准备情况反馈表
3. 数据安全保密协议（模板）

四川省审计厅
XXXX年XX月XX日"""

ws3.merge_cells("A3:C30")
cell = ws3["A3"]
cell.value = template_text
cell.font = Font(name="楷体", size=10)
cell.alignment = Alignment(wrap_text=True, vertical="top")
ws3.row_dimensions[3].height = 500
ws3.column_dimensions["A"].width = 20
ws3.column_dimensions["B"].width = 50
ws3.column_dimensions["C"].width = 20

# ============================================================
# 保存
# ============================================================
output_path = r"D:\openclaw-workspace\zhixi_intelligent\reports\数据采集资料准备清单.xlsx"
wb.save(output_path)
print(f"已生成: {output_path}")
print(f"包含4个Sheet: 总览目录 | 详细资料清单 | 采集工作流程 | 通知函模板")
print(f"共13个行业、{total_all}项资料需求")
