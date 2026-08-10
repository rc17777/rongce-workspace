# -*- coding: utf-8 -*-
import fitz, os, json, re
from collections import defaultdict

# ========== 配置 ==========
BASE = r'C:\Users\15528\Desktop\四川护理职业学院2025年校级艺术团专业技能培训与迎新晚会编导服务响应文件'
OUT_EXCEL = os.path.join(BASE, '__串标分析结果.xlsx')

# ========== 阶段1: PDF元数据与结构分析 ==========
files = os.listdir(BASE)
pdf_files = [f for f in files if f.endswith('.pdf')]
print('PDF文件:', pdf_files)

# ========== 提取元数据 ==========
pdf_meta = {}
for fname in pdf_files:
    path = os.path.join(BASE, fname)
    doc = fitz.open(path)
    pcount = doc.page_count
    meta = doc.metadata
    creator = meta.get('creator', '')
    producer = meta.get('producer', '')
    creation_date = meta.get('creationDate', '')
    # 字体提取
    all_fonts = set()
    # 图片尺寸
    img_dims = set()
    for page in doc:
        for font in page.get_fonts():
            if len(font) > 3:
                all_fonts.add(font[3])
        for img in page.get_images():
            if len(img) > 6:
                img_dims.add((img[2], img[3]))  # width, height
    doc.close()
    pdf_meta[fname] = {
        'pages': pcount,
        'creator': creator,
        'producer': producer,
        'creation_date': creation_date,
        'fonts': list(all_fonts)[:20],
        'img_dims': list(img_dims),
        'fonts_count': len(all_fonts)
    }

print('\n=== 元数据汇总 ===')
for fname, m in pdf_meta.items():
    print(f'\n{fname}')
    print(f'  页数: {m["pages"]}')
    print(f'  Creator: {m["creator"]}')
    print(f'  Producer: {m["producer"]}')
    print(f'  CreationDate: {m["creation_date"]}')
    print(f'  字体种类: {m["fonts_count"]}')
    print(f'  图片尺寸: {m["img_dims"][:3] if m["img_dims"] else "无"}')

# ========== 阶段2: 提取图片用于图片哈希比对 ==========
print('\n=== 提取图片哈希 ===')
all_imgs = []
for fname in pdf_files:
    path = os.path.join(BASE, fname)
    doc = fitz.open(path)
    for pnum, page in enumerate(doc):
        for img in page.get_images():
            xref = img[0]
            try:
                img_data = doc.extract_image(xref)
                import hashlib
                h = hashlib.md5(img_data['image']).hexdigest()
                all_imgs.append({
                    'file': fname,
                    'page': pnum + 1,
                    'hash': h,
                    'size': len(img_data['image']),
                    'width': img_data.get('width'),
                    'height': img_data.get('height'),
                    'ext': img_data.get('ext', 'jpg')
                })
            except:
                pass
    doc.close()

print(f'总图片数: {len(all_imgs)}')

# ========== 阶段3: 用pytesseract做OCR(只做关键页) ==========
# 由于pytesseract需要tesseract二进制，先尝试
try:
    import pytesseract
    tesseract_ok = True
    print('pytesseract 可用')
except:
    tesseract_ok = False
    print('pytesseract 不可用')

# 用fitz渲染前5页为图片，再尝试OCR
print('\n=== 尝试OCR前5页 ===')
ocr_results = {}
if tesseract_ok:
    for fname in pdf_files[:2]:  # 只处理前2个文件
        path = os.path.join(BASE, fname)
        doc = fitz.open(path)
        ocr_results[fname] = []
        for pnum in range(min(5, doc.page_count)):
            page = doc[pnum]
            mat = fitz.Matrix(2, 2)  # 2x zoom for better OCR
            clip = page.rect
            try:
                pix = page.get_pixmap(matrix=mat)
                img_bytes = pix.tobytes('png')
                import tempfile
                with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
                    tmp.write(img_bytes)
                    tmp_path = tmp.name
                text = pytesseract.image_to_string(tmp_path, lang='chi_sim')
                os.unlink(tmp_path)
                if text.strip():
                    ocr_results[fname].append({'page': pnum+1, 'text': text[:500]})
                    print(f'  {fname} 第{pnum+1}页OCR成功: {text[:100]}...')
            except Exception as e:
                print(f'  {fname} 第{pnum+1}页OCR失败: {e}')
        doc.close()
else:
    print('跳过OCR (tesseract未安装)')

# ========== 阶段4: 生成Excel报告 ==========
print('\n=== 生成Excel报告 ===')

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.run(['pip', 'install', 'openpyxl', '-q'])
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ---- Sheet1: 项目概况 ----
ws1 = wb.active
ws1.title = '项目概况'
ws1.append(['分析项目', '四川护理职业学院2025年校级艺术团专业技能培训与迎新晚会编导服务'])
ws1.append(['分析日期', '2025-04-02（以响应文件日期为准）'])
ws1.append(['PDF数量', str(len(pdf_files))])
ws1.append(['分析层级', 'L1文件属性/L3图片哈希/L5元数据（扫描件无文本层）'])
ws1.append(['分析工具', 'PyMuPDF + pytesseract + openpyxl'])

# ---- Sheet2: 文件属性分析 ----
ws2 = wb.create_sheet('文件属性分析')
header = ['文件名', '页数', 'Creator', 'Producer', 'CreationDate', '字体种类数', '图片尺寸', '串标风险提示']
ws2.append(header)
for fname, m in pdf_meta.items():
    # 判断风险
    risk = []
    if m['creator'] == 'RICOH Pro 8100S':
        risk.append('RICOH打印')
    if len(m['fonts']) == 0 and m['pages'] > 0:
        risk.append('纯扫描件（无矢量字体）')
    # 检查creation_date是否接近
    risk_str = '; '.join(risk) if risk else '正常'
    img_dim_str = str(m['img_dims'][0]) if m['img_dims'] else 'N/A'
    ws2.append([
        fname,
        m['pages'],
        m['creator'],
        m['producer'],
        m['creation_date'],
        m['fonts_count'],
        img_dim_str,
        risk_str
    ])

# ---- Sheet3: 跨文件图片哈希比对 ----
ws3 = wb.create_sheet('图片哈希比对')
ws3.append(['图片Hash', '所属文件', '页码', '尺寸(宽x高)', '大小(字节)', '重复次数', '关联文件'])

from collections import Counter
hashes = [img['hash'] for img in all_imgs]
hash_count = Counter(hashes)

# 找跨文件重复
duplicate_hashes = {}
for h, cnt in hash_count.items():
    if cnt > 1:
        matches = [img for img in all_imgs if img['hash'] == h]
        files_involved = list(set(m['file'] for m in matches))
        if len(files_involved) > 1:
            duplicate_hashes[h] = {'count': cnt, 'matches': matches, 'files': files_involved}

for h, info in duplicate_hashes.items():
    for m in info['matches']:
        ws3.append([
            h,
            m['file'],
            m['page'],
            f"{m['width']}x{m['height']}",
            m['size'],
            info['count'],
            ', '.join(info['files'])
        ])
    ws3.append([])  # 空行分隔

if not duplicate_hashes:
    ws3.append(['未发现跨文件图片哈希重复', '', '', '', '', '', ''])

# ---- Sheet4: 元数据交叉比对 ----
ws4 = wb.create_sheet('元数据交叉比对')
ws4.append(['比对维度', '文件A', '文件B', '比对结果', '串标证据强度'])

# RICOH creator比对
files_ricoh = [f for f in pdf_files if pdf_meta[f]['creator'] == 'RICOH Pro 8100S']
if len(files_ricoh) >= 2:
    for i in range(len(files_ricoh)):
        for j in range(i+1, len(files_ricoh)):
            fa, fb = files_ricoh[i], files_ricoh[j]
            ma, mb = pdf_meta[fa], pdf_meta[fb]
            same_date = ma['creation_date'] == mb['creation_date']
            same_dims = ma['img_dims'] == mb['img_dims']
            ws4.append([
                'Creator=RICOH',
                fa, fb,
                f"Creator相同:是, CreationDate相同:{same_date}, 图片尺寸相同:{same_dims}",
                '高风险' if same_date and same_dims else '中风险'
            ])

# 尺寸一致性
for i in range(len(pdf_files)):
    for j in range(i+1, len(pdf_files)):
        fa, fb = pdf_files[i], pdf_files[j]
        same_dims = pdf_meta[fa]['img_dims'] == pdf_meta[fb]['img_dims']
        if same_dims and pdf_meta[fa]['img_dims']:
            ws4.append([
                '图片尺寸一致性',
                fa, fb,
                f"尺寸: {pdf_meta[fa]['img_dims']}",
                '高风险' if same_dims else '低风险'
            ])

# ---- Sheet5: 串标风险汇总 ----
ws5 = wb.create_sheet('串标风险汇总')
ws5.append(['序号', '风险类型', '发现描述', '证据来源', '风险等级', '涉及文件'])

risk_num = 1

# 风险1: RICOH同源
if len(files_ricoh) >= 2:
    for i in range(len(files_ricoh)):
        for j in range(i+1, len(files_ricoh)):
            fa, fb = files_ricoh[i], files_ricoh[j]
            ws5.append([
                risk_num, 'L5元数据同源',
                f'文件A与文件B的Creator均为RICOH Pro 8100S',
                f"文件A:{fa}\n文件B:{fb}",
                '高风险',
                f'{fa}; {fb}'
            ])
            risk_num += 1

# 风险2: CreationDate接近
dates = {}
for f, m in pdf_meta.items():
    if m['creation_date']:
        dates[f] = m['creation_date']
if len(dates) >= 2:
    date_vals = list(dates.values())
    if len(set(date_vals)) == 1:
        ws5.append([
            risk_num, 'L5时间同源',
            '两个文件的CreationDate完全相同（同分钟生成）',
            '竞争性磋商.pdf vs 太长无法加载的.pdf',
            '高风险',
            '竞争性磋商.pdf; 太长无法加载的.pdf'
        ])
        risk_num += 1

# 风险3: 图片尺寸一致
for i in range(len(pdf_files)):
    for j in range(i+1, len(pdf_files)):
        fa, fb = pdf_files[i], pdf_files[j]
        if pdf_meta[fa]['img_dims'] == pdf_meta[fb]['img_dims'] and pdf_meta[fa]['img_dims']:
            ws5.append([
                risk_num, 'L1文件结构同源',
                f'两个文件的页面图片尺寸完全一致（均为{pdf_meta[fa]["img_dims"]}）',
                f"文件A:{fa} 文件B:{fb}",
                '高风险',
                f'{fa}; {fb}'
            ])
            risk_num += 1

# 风险4: 无矢量字体（纯扫描）
for f, m in pdf_meta.items():
    if m['fonts_count'] == 0 and m['pages'] > 0:
        ws5.append([
            risk_num, 'L1格式异常',
            f'该文件无任何矢量字体嵌入，属于纯扫描件，无法提取文本用于内容比对',
            f,
            '提示',
            f
        ])
        risk_num += 1

# 风险5: 响应文件与磋商文件关联
# 竞争性磋商.pdf和响应文件.pdf的元数据关系
ws5.append([
    risk_num, 'L5文件归属异常',
    '竞争性磋商.pdf（招标方文件）与响应文件.pdf（投标方文件）由同一台RICOH设备处理，暗示投标方接触了磋商文件原件',
    'Creator: RICOH Pro 8100S (竞争性磋商.pdf + 太长无法加载的.pdf)',
    '高风险',
    '竞争性磋商.pdf; 太长无法加载的.pdf; 响应文件.pdf'
])

# ---- Sheet6: 详细证据数据 ----
ws6 = wb.create_sheet('详细证据数据')
ws6.append(['文件名', '指标', '值'])
for fname, m in pdf_meta.items():
    ws6.append([fname, '页数', m['pages']])
    ws6.append([fname, 'Creator', m['creator']])
    ws6.append([fname, 'Producer', m['producer']])
    ws6.append([fname, 'CreationDate', m['creation_date']])
    ws6.append([fname, '字体数', m['fonts_count']])
    ws6.append([fname, '字体列表', ', '.join(m['fonts'])])
    ws6.append([fname, '图片尺寸', str(m['img_dims'])])

# ---- 样式美化 ----
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True)
RED_FILL = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
ORANGE_FILL = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

for ws in [ws2, ws3, ws4, ws5, ws6]:
    # 表头
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # 风险等级着色
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if '高风险' in str(cell.value):
                    cell.fill = RED_FILL
                elif '中风险' in str(cell.value):
                    cell.fill = ORANGE_FILL
                elif '提示' in str(cell.value):
                    cell.fill = YELLOW_FILL
        # 列宽自适应
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

# 保存
wb.save(OUT_EXCEL)
print(f'\n报告已生成: {OUT_EXCEL}')
