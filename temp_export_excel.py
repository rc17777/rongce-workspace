# -*- coding: utf-8 -*-
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ===== STYLES =====
header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='0A1F3F', end_color='0A1F3F', fill_type='solid')
sub_header_fill = PatternFill(start_color='1A5C6E', end_color='1A5C6E', fill_type='solid')
p0_fill = PatternFill(start_color='FFD7D7', end_color='FFD7D7', fill_type='solid')
p1_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
p2_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
total_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
normal_font = Font(name='微软雅黑', size=10)
bold_font = Font(name='微软雅黑', size=10, bold=True)
title_font = Font(name='微软雅黑', size=14, bold=True, color='0A1F3F')
subtitle_font = Font(name='微软雅黑', size=11, bold=True, color='1A5C6E')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header(ws, row, max_col):
    for c in range(1, max_col+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_row(ws, row, max_col, fill=None):
    for c in range(1, max_col+1):
        cell = ws.cell(row=row, column=c)
        cell.font = normal_font
        cell.alignment = left_align if c > 1 else center_align
        cell.border = thin_border
        if fill:
            cell.fill = fill

# ===== CREATE WORKBOOK =====
wb = openpyxl.Workbook()

# ========== SHEET 1: 复核总览 ==========
ws1 = wb.active
ws1.title = '复核总览'

# Title
ws1.merge_cells('A1:H1')
ws1.cell(row=1, column=1, value='马尔康市党坝乡石加友村2026年特色旅居村以工代赈项目').font = title_font
ws1.merge_cells('A2:H2')
ws1.cell(row=2, column=1, value='工程量清单及招标控制价复核报告').font = subtitle_font

ws1.merge_cells('A4:H4')
ws1.cell(row=4, column=1, value='控制价总额：¥7,170,602.52（柒佰壹拾柒万零陆佰零贰元伍角贰分）').font = bold_font
ws1.cell(row=5, column=1, value='复核日期：2026-07-30').font = normal_font

# Summary table
headers1 = ['序号', '费用项目', '金额（元）', '占控制价%', '暂估价（元）', '备注', '', '']
for c, h in enumerate(headers1, 1):
    ws1.cell(row=7, column=c, value=h)
style_header(ws1, 7, len(headers1))

data1 = [
    ['1', '分部分项及单价措施项目', 6878935.65, '95.93%', 568829.30, '含单价措施21,455.71', '', ''],
    ['2', '总价措施项目', 291666.87, '4.07%', 0, '', '', ''],
    ['2.1', '  其中：安全文明施工费', 279284.79, '3.89%', 0, '费率4.06%', '', ''],
    ['3', '其他项目', 0, '-', 0, '', '', ''],
    ['4', '规费', 0, '-', 0, '以工代赈不计取', '', ''],
    ['5', '销项增值税', 0, '-', 0, '以工代赈不计取', '', ''],
    ['', '招标控制价总价', 7170602.52, '100%', 568829.30, '', '', ''],
]
for i, d in enumerate(data1):
    r = 8 + i
    for c, v in enumerate(d, 1):
        ws1.cell(row=r, column=c, value=v)
    fill = total_fill if d[0] == '' else None
    style_row(ws1, r, len(headers1), fill)

# Findings summary
ws1.merge_cells('A16:H16')
ws1.cell(row=16, column=1, value='发现问题汇总').font = subtitle_font

headers_f = ['等级', '数量', '关键问题', '涉及金额（元）', '', '', '', '']
for c, h in enumerate(headers_f, 1):
    ws1.cell(row=17, column=c, value=h)
style_header(ws1, 17, len(headers_f))

findings_summary = [
    ['🔴 P0 必须整改', '2项', '管理费+利润违规列支、排泥阀井vs阀门井单价矛盾', '~627,000', '', '', '', ''],
    ['🟡 P1 建议核实', '4项', '土石比无依据、工程量两套数、路面恢复减薄、消能池单价', '~550,000', '', '', '', ''],
    ['🔵 P2 建议关注', '5项', 'DN50压力等级不一致、道路模板、二次搬运重复、湿井缺井盖、安文费费率', '-', '', '', '', ''],
]
for i, d in enumerate(findings_summary):
    r = 18 + i
    for c, v in enumerate(d, 1):
        ws1.cell(row=r, column=c, value=v)
    fill = [p0_fill, p1_fill, p2_fill][i]
    style_row(ws1, r, len(headers_f), fill)

# Column widths
ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 30
ws1.column_dimensions['C'].width = 18
ws1.column_dimensions['D'].width = 14
ws1.column_dimensions['E'].width = 18
ws1.column_dimensions['F'].width = 40

# ========== SHEET 2: P0问题详情 ==========
ws2 = wb.create_sheet('P0-必须整改')

ws2.merge_cells('A1:G1')
ws2.cell(row=1, column=1, value='P0 必须整改 — 政策合规/重大计价错误').font = title_font

# --- P0-1: 管理费+利润 ---
ws2.merge_cells('A3:G3')
ws2.cell(row=3, column=1, value='P0-1：综合单价含管理费和利润，违反以工代赈政策（川发改赈〔2024〕37号）').font = subtitle_font

ws2.cell(row=5, column=1, value='编制说明第2.10条明确："不应列支…企业管理费、利润及税金"。但全部27项清单综合单价分析表中均包含管理费和利润，合计589,668元，占控制价8.22%。').font = normal_font

h2 = ['序号', '项目编码', '项目名称', '综合单价', '管理费', '利润', '管+利占比', '管+利合计(元)']
for c, h in enumerate(h2, 1):
    ws2.cell(row=7, column=c, value=h)
style_header(ws2, 7, len(h2))

p0_1_data = [
    [1, '040101002001', '挖沟槽土方', 44.39, 2.03, 4.63, '15.0%', 83947],
    [2, '040102002002', '挖沟槽石方', 87.64, 4.02, 9.15, '15.0%', 166004],
    [3, '040101003003', '挖基坑土方', 44.39, 2.03, 4.63, '15.0%', 2827],
    [4, '040102003004', '挖基坑石方', 87.64, 4.02, 9.15, '15.0%', 5590],
    [5, '040103001005', '回填土方', 14.06, 0.66, 1.50, '15.4%', 23507],
    [6, '040103001006', '回填石方', 14.70, 0.69, 1.57, '15.4%', 24596],
    [7, '040103002007', '余方弃置', 46.29, 2.12, 4.83, '15.0%', 29831],
    [8, '040103001008', '砂石基础', 209.01, 5.79, 13.18, '9.1%', 74624],
    [9, '041001008009', '人工拆除边沟', 348.33, 8.55, 19.42, '8.0%', 738],
    [10, '041001001010', '混凝土路面拆除', 54.22, 2.31, 5.27, '14.0%', 12128],
    [11, '040203007011', '混凝土路面恢复', 143.46, 1.94, 4.41, '4.4%', 10160],
    [12, '040503002013', '管道混凝土支墩', 592.79, 16.18, 36.84, '8.9%', 1374],
    [13, '040501017014', '公路混凝土边沟', 191.08, 2.65, 6.06, '4.6%', 523],
    [14, '030801017015', 'DN150钢丝骨架管', 260.02, 3.09, 7.06, '3.9%', 97643],
    [15, '030801017016', 'DN50钢丝骨架管', 81.06, 1.02, 2.33, '4.1%', 14070],
    [16, '030817009017', 'DN150压力试验', 9.34, 0.31, 0.71, '10.9%', 9812],
    [17, '030817009018', 'DN50压力试验', 5.78, 0.21, 0.48, '11.9%', 2898],
    [18, '040504002019', '排气阀井1200x1200', 5375.85, 106.22, 242.05, '6.5%', 5224],
    [19, '040504002020', '排泥阀井1300x1300', 9058.35, 159.27, 363.14, '5.8%', 5747],
    [20, '040504002021', '阀门井1300x1300', 5692.56, 112.19, 255.63, '6.5%', 6989],
    [21, '040504002022', '排泥湿井1100x1100', 4385.15, 80.73, 183.88, '6.0%', 2911],
    [22, '040504002023', '消能池', 25160.86, 491.92, 1120.46, '6.4%', 6450],
    [23, '自编024', '二次搬运', 231.00, 0, 0, '0.0%', 0],
]

for i, d in enumerate(p0_1_data):
    r = 8 + i
    for c, v in enumerate(d, 1):
        ws2.cell(row=r, column=c, value=v)
    style_row(ws2, r, len(h2))

# Total row
r_total = 8 + len(p0_1_data)
for c in range(1, len(h2)+1):
    ws2.cell(row=r_total, column=c).border = thin_border
    ws2.cell(row=r_total, column=c).fill = total_fill
    ws2.cell(row=r_total, column=c).font = bold_font
ws2.cell(row=r_total, column=1, value='')
ws2.cell(row=r_total, column=2, value='合计')
ws2.cell(row=r_total, column=8, value=589668)

r_note = r_total + 2
ws2.merge_cells(f'A{r_note}:H{r_note}')
ws2.cell(row=r_note, column=1, value='建议：所有综合单价重新组价，剔除管理费和利润。仅保留人工费、材料费、机械费。预计控制价应调减约58.9万元。').font = Font(name='微软雅黑', size=10, bold=True, color='CC0000')

# --- P0-2: 排泥阀井vs阀门井 ---
r_start = r_note + 3
ws2.merge_cells(f'A{r_start}:H{r_start}')
ws2.cell(row=r_start, column=1, value='P0-2：排泥阀井 vs 阀门井 — 同图同量不同价').font = subtitle_font

ws2.merge_cells(f'A{r_start+1}:H{r_start+1}')
ws2.cell(row=r_start+1, column=1, value='两者均为1300×1300mm、井深2.17m、C25混凝土、同一标准图集07MS101-2 P66，井.xls中工程量完全相同。').font = normal_font

h3 = ['对比项', '排泥阀井 1300x1300', '阀门井 1300x1300', '差额', '差异率', '', '', '']
for c, h in enumerate(h3, 1):
    ws2.cell(row=r_start+3, column=c, value=h)
style_header(ws2, r_start+3, len(h3))

comp_data = [
    ['综合单价（元/座）', 9058.35, 5692.56, 3365.79, '37.2%', '', '', ''],
    ['人工费', 2088.92, 1114.24, 974.68, '46.7%', '', '', ''],
    ['材料费', 3696.00, 3692.05, 3.95, '0.1%', '', '', ''],
    ['机械费', 408.85, 243.26, 165.59, '40.5%', '', '', ''],
    ['管理费', 159.27, 112.19, 47.08, '29.6%', '', '', ''],
    ['利润', 363.14, 255.63, 107.51, '29.6%', '', '', ''],
    ['数量', 11, 19, '', '', '', '', ''],
    ['合价', 99641.85, 108158.64, '', '', '', '', ''],
]
for i, d in enumerate(comp_data):
    r = r_start + 4 + i
    for c, v in enumerate(d, 1):
        ws2.cell(row=r, column=c, value=v)
    style_row(ws2, r, len(h3))

r_note2 = r_start + 4 + len(comp_data) + 1
ws2.merge_cells(f'A{r_note2}:H{r_note2}')
ws2.cell(row=r_note2, column=1, value='建议：核实两者定额子目选取是否一致。如排泥阀井确含阀体设备费，应在项目特征中单独列明，不得混入井体结构。').font = Font(name='微软雅黑', size=10, bold=True, color='CC0000')

# Column widths
for c, w in enumerate([8, 28, 22, 22, 14, 14, 14, 18], 1):
    ws2.column_dimensions[get_column_letter(c)].width = w

# ========== SHEET 3: P1问题详情 ==========
ws3 = wb.create_sheet('P1-建议核实')

ws3.merge_cells('A1:F1')
ws3.cell(row=1, column=1, value='P1 建议核实 — 工程量/计价依据存疑').font = title_font

p1_items = [
    ('P1-1 沟槽土石方比例50:50无地质依据',
     '挖沟槽土方=挖沟槽石方=12,604.69m³，恰好各半。',
     '土方单价44.39元/m³ vs 石方单价87.64元/m³，此比例影响造价约54.5万元。建议提供地质勘察报告或现场签证确认土石比。'),
    ('P1-2 土石方工程数量两套数据',
     '控制价与工程量表存在系统性偏差。',
     '挖方差-848.93m³，填方差-692.04m³（约3.3%）。建议统一取数来源，以设计图纸核算量为准。'),
    ('P1-3 路面恢复厚度(15cm)低于拆除厚度(20cm)',
     '拆除20cm→恢复15cm，恢复厚度低于原路面。',
     '建议提供恢复厚度15cm的设计依据（交通量等级、路基承载力计算书）。'),
    ('P1-4 消能池综合单价偏高',
     '消能池25,160.86元/座，4座合计100,643元。',
     '混凝土量27.13m³远超普通检查井(1.82m³)，但单位造价仍需核实是否合理。'),
]

headers_p1 = ['序号', '问题', '现状', '建议', '', '']
for c, h in enumerate(headers_p1, 1):
    ws3.cell(row=3, column=c, value=h)
style_header(ws3, 3, len(headers_p1))

for i, (title, desc, suggestion) in enumerate(p1_items):
    r = 4 + i
    ws3.cell(row=r, column=1, value=i+1)
    ws3.cell(row=r, column=2, value=title)
    ws3.cell(row=r, column=3, value=desc)
    ws3.cell(row=r, column=4, value=suggestion)
    style_row(ws3, r, len(headers_p1), p1_fill)
    ws3.row_dimensions[r].height = 45

for c, w in enumerate([6, 42, 50, 50, 10, 10], 1):
    ws3.column_dimensions[get_column_letter(c)].width = w

# ========== SHEET 4: P2问题详情 ==========
ws4 = wb.create_sheet('P2-建议关注')

ws4.merge_cells('A1:F1')
ws4.cell(row=1, column=1, value='P2 建议关注 — 技术细节/一致性').font = title_font

p2_items = [
    ('P2-1 DN50管材压力等级不一致', '工程量表标注2.5MPa，控制价标注2.0MPa', '压力等级影响管材壁厚和单价，建议统一。'),
    ('P2-2 道路模板计量可能偏大', '混凝土道路模板1,600m²×3.72=5,952元', '道路模板按定额应计侧模面积(周长×厚度)，1600m²似为路面面积而非模板面积。'),
    ('P2-3 二次搬运重复列项', 'F.1表以"自编024"列入57,750元，F.4总价措施二次搬运费却为0', '核实是否为重复计取；暂估价形式缺乏计算依据。'),
    ('P2-4 排泥湿井无井盖', '井.xls中排泥湿井(1100×1100)无井盖及支座', '湿井通常有井盖，建议核实是否漏项。'),
    ('P2-5 安全文明施工费费率', '4项合计费率4.06%', '对照川建行规〔2024〕15号确认是否符合市政工程对应费率。'),
]

headers_p2 = ['序号', '问题', '现状', '建议', '', '']
for c, h in enumerate(headers_p2, 1):
    ws4.cell(row=3, column=c, value=h)
style_header(ws4, 3, len(headers_p2))

for i, (title, desc, suggestion) in enumerate(p2_items):
    r = 4 + i
    ws4.cell(row=r, column=1, value=i+1)
    ws4.cell(row=r, column=2, value=title)
    ws4.cell(row=r, column=3, value=desc)
    ws4.cell(row=r, column=4, value=suggestion)
    style_row(ws4, r, len(headers_p2), p2_fill)
    ws4.row_dimensions[r].height = 45

for c, w in enumerate([6, 36, 50, 50, 10, 10], 1):
    ws4.column_dimensions[get_column_letter(c)].width = w

# ========== SHEET 5: 合规性审查 ==========
ws5 = wb.create_sheet('以工代赈合规性审查')

ws5.merge_cells('A1:E1')
ws5.cell(row=1, column=1, value='以工代赈政策合规性审查（川发改赈〔2024〕37号 / 〔2026〕248号）').font = title_font

headers_c = ['序号', '检查项', '政策要求', '现状', '结论']
for c, h in enumerate(headers_c, 1):
    ws5.cell(row=3, column=c, value=h)
style_header(ws5, 3, len(headers_c))

compliance_data = [
    ['1', '计价方式', '定额计价', '定额计价', '✅ 符合'],
    ['2', '人工为主', '能用人工不用机械', '开挖/回填均标注"人工"', '✅ 符合'],
    ['3', '管理费+利润', '不应列支', '全部含管理费+利润 589,668元', '❌ 违规 - P0'],
    ['4', '规费', '不应列支', '0', '✅ 符合'],
    ['5', '增值税', '不应列支（村民自建）', '0', '✅ 符合'],
    ['6', '招投标代理费', '不应列支', '未出现', '✅ 符合'],
    ['7', '劳务报酬标准', '150-300元/天', '已注明', '✅ 符合'],
    ['8', '最低工资', '2,200元/月（阿坝州）', '已注明', '✅ 符合'],
    ['9', '安全文明施工费', '按川建行规〔2024〕15号', '费率4.06%已计取', '✅ 符合（费率待核实）'],
    ['10', '暂列金额', '村民自建可酌减', '0', '✅ 符合'],
]

for i, d in enumerate(compliance_data):
    r = 4 + i
    for c, v in enumerate(d, 1):
        ws5.cell(row=r, column=c, value=v)
    fill = p0_fill if '❌' in str(d[4]) else None
    style_row(ws5, r, len(headers_c), fill)
    ws5.row_dimensions[r].height = 28

for c, w in enumerate([6, 20, 38, 38, 15], 1):
    ws5.column_dimensions[get_column_letter(c)].width = w

# ========== SHEET 6: 综合单价分析 ==========
ws6 = wb.create_sheet('综合单价分析-管理费利润')

ws6.merge_cells('A1:K1')
ws6.cell(row=1, column=1, value='综合单价分析表 — 管理费+利润提取（来源：F.2综合单价分析表）').font = title_font

h6 = ['序号', '项目编码', '项目名称', '单位', '工程量', '综合单价', '人工费', '材料费', '机械费', '管理费', '利润']
for c, h in enumerate(h6, 1):
    ws6.cell(row=3, column=c, value=h)
style_header(ws6, 3, len(h6))

price_data = [
    [1, '040101002001', '挖沟槽土方', 'm3', 12604.69, 44.39, 37.72, 0, 0, 2.03, 4.63],
    [2, '040102002002', '挖沟槽石方', 'm3', 12604.69, 87.64, 74.47, 0, 0, 4.02, 9.15],
    [3, '040101003003', '挖基坑土方', 'm3', 424.46, 44.39, 37.72, 0, 0, 2.03, 4.63],
    [4, '040102003004', '挖基坑石方', 'm3', 424.46, 87.64, 74.47, 0, 0, 4.02, 9.15],
    [5, '040103001005', '回填土方', 'm3', 10883.02, 14.06, 10.25, 0, 1.59, 0.66, 1.50],
    [6, '040103001006', '回填石方', 'm3', 10883.02, 14.70, 10.34, 0, 2.02, 0.69, 1.57],
    [7, '040103002007', '余方弃置', 'm3', 4292.26, 46.29, 39.34, 0, 0, 2.12, 4.83],
    [8, '040103001008', '砂石基础', 'm3', 3933.80, 209.01, 58.19, 129.92, 1.93, 5.79, 13.18],
    [9, '041001008009', '人工拆除边沟', 'm3', 26.40, 348.33, 320.37, 0, 0, 8.55, 19.42],
    [10, '041001001010', '混凝土路面拆除', 'm2', 1600.00, 54.22, 46.63, 0, 0, 2.31, 5.27],
    [11, '040203007011', '混凝土路面恢复', 'm2', 1600.00, 143.46, 22.64, 108.18, 3.97, 1.94, 4.41],
    [12, '040503002013', '管道混凝土支墩', 'm3', 25.92, 592.79, 160.07, 372.17, 7.53, 16.18, 36.84],
    [13, '040501017014', '公路混凝土边沟', 'm', 60.00, 191.08, 44.21, 131.10, 4.49, 2.65, 6.06],
    [14, '030801017015', 'DN150钢丝骨架管', 'm', 9620.00, 260.02, 65.10, 0.30, 3.27, 3.09, 7.06],
    [15, '030801017016', 'DN50钢丝骨架管', 'm', 4200.00, 81.06, 21.88, 0.07, 0.76, 1.02, 2.33],
    [16, '030817009017', 'DN150压力试验', 'm', 9620.00, 9.34, 6.74, 1.19, 0.16, 0.31, 0.71],
    [17, '030817009018', 'DN50压力试验', 'm', 4200.00, 5.78, 4.58, 0.36, 0.13, 0.21, 0.48],
    [18, '040504002019', '排气阀井1200x1200', '座', 15.00, 5375.85, 1060.57, 3457.97, 235.71, 106.22, 242.05],
    [19, '040504002020', '排泥阀井1300x1300', '座', 11.00, 9058.35, 2088.92, 3696.00, 408.85, 159.27, 363.14],
    [20, '040504002021', '阀门井1300x1300', '座', 19.00, 5692.56, 1114.24, 3692.05, 243.26, 112.19, 255.63],
    [21, '040504002022', '排泥湿井1100x1100', '座', 11.00, 4385.15, 730.75, 3274.02, 98.82, 80.73, 183.88],
    [22, '040504002023', '消能池', '座', 4.00, 25160.86, 4649.72, 17839.06, 671.23, 491.92, 1120.46],
    [23, '自编024', '二次搬运', '工日', 250.00, 231.00, 231.00, 0, 0, 0, 0],
]

for i, d in enumerate(price_data):
    r = 4 + i
    for c, v in enumerate(d, 1):
        ws6.cell(row=r, column=c, value=v)
    style_row(ws6, r, len(h6))

# Total
r_t = 4 + len(price_data)
for c in range(1, len(h6)+1):
    ws6.cell(row=r_t, column=c).border = thin_border
    ws6.cell(row=r_t, column=c).fill = total_fill
    ws6.cell(row=r_t, column=c).font = bold_font
ws6.cell(row=r_t, column=1, value='')
ws6.cell(row=r_t, column=3, value='合计')

# Add formulas for totals
ws6.cell(row=r_t, column=10, value=589668)

for c, w in enumerate([6, 18, 22, 6, 12, 10, 10, 10, 10, 10, 10], 1):
    ws6.column_dimensions[get_column_letter(c)].width = w

# ========== SAVE ==========
output_path = r'C:\Users\scrccpa\Desktop\石加友村以工代赈-控制价复核-20260730.xlsx'
wb.save(output_path)
print(f'Saved to: {output_path}')
print('Sheets: 复核总览 | P0-必须整改 | P1-建议核实 | P2-建议关注 | 以工代赈合规性审查 | 综合单价分析-管理费利润')
