"""
Review Report Generator (统一复核报告生成器)

Generates Excel + Markdown dual-format review reports from
filtered check results. Includes page references, original excerpts,
expected/actual/diff, and severity classification.

Output:
  - 复核报告.xlsx:  one sheet per check type
  - 复核报告.md:    readable markdown with cross-references
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from .rule_engine import CheckResult
from .review_filter import ReviewFilter


# ---------------------------------------------------------------------------
# Markdown Report
# ---------------------------------------------------------------------------

_MD_TEMPLATE = """# {title}

**生成时间**：{generated_at}
**检查领域**：{domain}
**源文件**：{source_path}

---

## 检查概要

| 指标 | 数值 |
|------|------|
| 检查总数 | {total_checks} |
| 通过 | {passed} |
| 未通过 | {failed} |
| 确认错误 | {confirmed} |
| 需人工复核 | {needs_review} |
| 误报（已筛除） | {false_positives} |
| 误报率 | {fp_rate:.1%} |

---

## 确认错误

{confirmed_section}

---

## 需人工复核

{needs_review_section}

---

## 误报明细（已自动筛除，供参考）

{fp_section}

---

## 各维度检查明细

{check_type_section}
"""


def _format_result_md(result: CheckResult, index: int) -> str:
    """Format a single check result as a markdown entry."""
    lines = [
        f"### {index}. {result.description}",
        f"",
        f"- **规则ID**：`{result.rule_id}`",
        f"- **严重程度**：{'🔴 错误' if result.severity == 'error' else '🟡 警告' if result.severity == 'warning' else 'ℹ️ 信息'}",
    ]

    if result.expected is not None and result.actual is not None:
        lines.append(f"- **预期值**：{result.expected:,.2f}")
        lines.append(f"- **实际值**：{result.actual:,.2f}")
        if result.diff is not None:
            lines.append(f"- **差异**：{result.diff:,.2f}（容差：{result.tolerance}）")

    if result.page_ref:
        lines.append(f"- **位置**：第 {result.page_ref} 页")
    if result.sheet_context:
        lines.append(f"- **表格**：{result.sheet_context}")
    if result.excerpt:
        lines.append(f"- **原文摘录**：`{result.excerpt}`")
    if result.detail:
        lines.append(f"- **说明**：{result.detail}")
    if result.false_positive:
        lines.append(f"- **误报原因**：{result.false_positive_reason}")

    lines.append("")
    return "\n".join(lines)


def generate_markdown_report(
    classified: dict[str, list[CheckResult]],
    domain: str,
    source_path: str = "",
    title: str = "审计复核报告",
) -> str:
    """Generate a complete Markdown review report."""

    confirmed = classified["confirmed"]
    needs_review = classified["needs_review"]
    false_positives = classified["false_positive"]

    total = len(confirmed) + len(needs_review) + len(false_positives)
    passed = 0  # passed checks aren't included in classified
    failed = len(confirmed) + len(needs_review)
    fp_rate = len(false_positives) / max(1, total)

    # Confirmed errors section
    if confirmed:
        confirmed_section = f"共 **{len(confirmed)}** 项：\n\n"
        for i, r in enumerate(confirmed, 1):
            confirmed_section += _format_result_md(r, i)
    else:
        confirmed_section = "✅ 未发现确认错误。"

    # Needs review section
    if needs_review:
        needs_review_section = f"共 **{len(needs_review)}** 项：\n\n"
        for i, r in enumerate(needs_review, 1):
            needs_review_section += _format_result_md(r, i)
    else:
        needs_review_section = "无需人工复核项。"

    # False positives section
    if false_positives:
        fp_section = f"共 **{len(false_positives)}** 项（已自动筛除）：\n\n"
        for i, r in enumerate(false_positives, 1):
            fp_section += _format_result_md(r, i)
    else:
        fp_section = "无误报。"

    # By check type
    by_type: dict[str, dict] = {}
    all_results = confirmed + needs_review + false_positives
    for r in all_results:
        ct = r.check_type or "other"
        if ct not in by_type:
            by_type[ct] = {"confirmed": 0, "needs_review": 0, "false_positive": 0}
        if r.false_positive:
            by_type[ct]["false_positive"] += 1
        elif r.requires_human_review:
            by_type[ct]["needs_review"] += 1
        else:
            by_type[ct]["confirmed"] += 1

    ct_lines = []
    for ct, counts in sorted(by_type.items()):
        ct_lines.append(
            f"| {ct} | {counts['confirmed']} | {counts['needs_review']} | "
            f"{counts['false_positive']} |"
        )
    check_type_section = (
        "| 检查维度 | 确认错误 | 需复核 | 误报 |\n"
        "|----------|---------|--------|------|\n"
        + "\n".join(ct_lines)
    )

    return _MD_TEMPLATE.format(
        title=title,
        generated_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        domain=domain,
        source_path=source_path,
        total_checks=total,
        passed=passed,
        failed=failed,
        confirmed=len(confirmed),
        needs_review=len(needs_review),
        false_positives=len(false_positives),
        fp_rate=fp_rate,
        confirmed_section=confirmed_section,
        needs_review_section=needs_review_section,
        fp_section=fp_section,
        check_type_section=check_type_section,
    )


# ---------------------------------------------------------------------------
# Excel Report
# ---------------------------------------------------------------------------

def generate_excel_report(
    classified: dict[str, list[CheckResult]],
    output_path: str,
    domain: str = "",
) -> str:
    """
    Generate an Excel review report using openpyxl.
    Returns the output path on success.
    
    Falls back to JSON if openpyxl is not available.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return _generate_json_fallback(classified, output_path, domain)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "复核概要"

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    info_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fp_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def apply_header_style(ws, row, max_col):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    # ---- Summary Sheet ----
    ws.merge_cells("A1:H1")
    ws["A1"] = "审计复核报告"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A3"] = f"检查领域：{domain}"

    summary_data = [
        ("确认错误", len(classified["confirmed"])),
        ("需人工复核", len(classified["needs_review"])),
        ("误报（已筛除）", len(classified["false_positive"])),
    ]
    for i, (label, count) in enumerate(summary_data, 5):
        ws.cell(row=i, column=1, value=label).font = header_font
        ws.cell(row=i, column=2, value=count)

    # ---- Detailed sheets by category ----
    for category, results in [
        ("确认错误", classified["confirmed"]),
        ("需人工复核", classified["needs_review"]),
        ("误报明细", classified["false_positive"]),
    ]:
        if not results:
            continue

        ws2 = wb.create_sheet(title=category[:31])  # Excel sheet name limit
        headers = [
            "规则ID", "检查维度", "描述", "预期值", "实际值", "差异",
            "容差", "严重程度", "页码", "表格", "原文摘录", "说明",
        ]
        for col_idx, header in enumerate(headers, 1):
            ws2.cell(row=1, column=col_idx, value=header)
        apply_header_style(ws2, 1, len(headers))

        for row_idx, r in enumerate(results, 2):
            values = [
                r.rule_id, r.check_type, r.description,
                r.expected, r.actual, r.diff,
                r.tolerance, r.severity, r.page_ref,
                r.sheet_context, r.excerpt, r.detail or r.false_positive_reason,
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border

            # Color row by severity
            if r.false_positive:
                fill = fp_fill
            elif r.severity == "error":
                fill = error_fill
            elif r.severity == "warning":
                fill = warning_fill
            else:
                fill = info_fill

            for col_idx in range(1, len(headers) + 1):
                ws2.cell(row=row_idx, column=col_idx).fill = fill

        # Auto-width
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            ws2.column_dimensions[col_letter].width = 18

    # Save
    if not output_path.endswith(".xlsx"):
        output_path += ".xlsx"
    wb.save(output_path)
    return output_path


def _generate_json_fallback(
    classified: dict[str, list[CheckResult]],
    output_path: str,
    domain: str,
) -> str:
    """Fallback: generate JSON report when openpyxl is unavailable."""
    data = {
        "title": "审计复核报告",
        "generated_at": datetime.now().isoformat(),
        "domain": domain,
        "summary": {
            "confirmed_errors": len(classified["confirmed"]),
            "needs_human_review": len(classified["needs_review"]),
            "false_positives": len(classified["false_positive"]),
        },
        "confirmed": [
            {
                "rule_id": r.rule_id,
                "check_type": r.check_type,
                "description": r.description,
                "expected": r.expected,
                "actual": r.actual,
                "diff": r.diff,
                "tolerance": r.tolerance,
                "severity": r.severity,
                "page_ref": r.page_ref,
                "excerpt": r.excerpt,
                "detail": r.detail,
            }
            for r in classified["confirmed"]
        ],
        "needs_review": [
            {
                "rule_id": r.rule_id,
                "check_type": r.check_type,
                "description": r.description,
                "expected": r.expected,
                "actual": r.actual,
                "diff": r.diff,
                "page_ref": r.page_ref,
                "excerpt": r.excerpt,
                "detail": r.detail,
            }
            for r in classified["needs_review"]
        ],
        "false_positives": [
            {
                "rule_id": r.rule_id,
                "description": r.description,
                "reason": r.false_positive_reason,
            }
            for r in classified["false_positive"]
        ],
    }

    if not output_path.endswith(".json"):
        output_path = output_path.replace(".xlsx", ".json")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return output_path


# ---------------------------------------------------------------------------
# Report Generator (Orchestrator)
# ---------------------------------------------------------------------------

class ReviewReportGenerator:
    """
    Generates dual-format (Excel + Markdown) review reports.
    """

    def __init__(self, classified: dict[str, list[CheckResult]],
                 domain: str = "", source_path: str = ""):
        self.classified = classified
        self.domain = domain
        self.source_path = source_path

    def generate_all(self, output_dir: str, base_name: str = "复核报告") -> dict[str, str]:
        """
        Generate both Excel and Markdown reports.
        
        Returns: {"excel": path, "markdown": path}
        """
        out_dir = Path(output_dir)
        out_dir.mkdir(parents=True, exist_ok=True)

        # Markdown
        md_content = generate_markdown_report(
            self.classified, self.domain, self.source_path
        )
        md_path = out_dir / f"{base_name}.md"
        with open(md_path, "w", encoding="utf-8") as f:
            f.write(md_content)

        # Excel
        xlsx_path = str(out_dir / f"{base_name}.xlsx")
        xlsx_result = generate_excel_report(
            self.classified, xlsx_path, self.domain
        )

        return {
            "excel": xlsx_result,
            "markdown": str(md_path),
        }

    @staticmethod
    def from_filter_results(
        confirmed: list[CheckResult],
        needs_review: list[CheckResult],
        false_positives: list[CheckResult],
        domain: str = "",
        source_path: str = "",
    ) -> "ReviewReportGenerator":
        return ReviewReportGenerator(
            classified={
                "confirmed": confirmed,
                "needs_review": needs_review,
                "false_positive": false_positives,
            },
            domain=domain,
            source_path=source_path,
        )
